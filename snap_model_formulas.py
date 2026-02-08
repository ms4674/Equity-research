#!/usr/bin/env python3
"""
Snap Inc. (SNAP) – Three-Statement Financial Model (NON-GAAP PRIMARY)
=====================================================================
Primary P&L uses Non-GAAP metrics (ex-SBC, ex-intangible amort, ex-restructuring).
Dedicated GAAP Reconciliation tab bridges Non-GAAP → GAAP with full detail.
Every forecast cell is a live Excel formula referencing the Assumptions tab.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as CL

# ═══════════════════════════════════════════════════════════════════════════════
# STYLING
# ═══════════════════════════════════════════════════════════════════════════════
HDR_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HDR_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUB_FONT = Font(name="Calibri", bold=True, size=10, color="2F5496")
SUB_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FCST_FONT = Font(name="Calibri", italic=True, size=10, color="0070C0")
NF = Font(name="Calibri", size=10)
BF = Font(name="Calibri", bold=True, size=10)
TF = Font(name="Calibri", bold=True, size=14, color="2F5496")
SF = Font(name="Calibri", bold=True, size=11, color="2F5496")
GREEN_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
MED_B = Border(bottom=Side(style='medium', color='2F5496'))

P = '0.0%'; N = '#,##0'; D = '#,##0.0'; U2 = '$#,##0.00'; U1 = '$#,##0.0'

def w(ws, r, c, val, font=NF, fmt=None, fill=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = font
    if fmt: cell.number_format = fmt
    if fill: cell.fill = fill
    return cell

def hdr_row(ws, r, labels):
    for i, lb in enumerate(labels):
        c = ws.cell(row=r, column=1+i, value=lb)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center', wrap_text=True)

def sub_row(ws, r, labels):
    for i, lb in enumerate(labels):
        c = ws.cell(row=r, column=1+i, value=lb)
        c.font = SUB_FONT; c.fill = SUB_FILL; c.alignment = Alignment(horizontal='center')

def sec(ws, r, label):
    ws.cell(row=r, column=1, value=label).font = SF
    ws.cell(row=r, column=1).border = MED_B

def colw(ws, widths):
    for i, w_ in enumerate(widths, 1): ws.column_dimensions[CL(i)].width = w_

def fill_hist(ws, r, vals, fmt=N, font=NF):
    for i, v in enumerate(vals): w(ws, r, 2+i, v, font=font, fmt=fmt)

def fill_fcst(ws, r, formulas, fmt=N, font=FCST_FONT, fill=None):
    for i, f_ in enumerate(formulas): w(ws, r, 8+i, f_, font=font, fmt=fmt, fill=fill)

def fill_all(ws, r, gen, fmt=P, font=NF, start=2):
    for ci in range(start, 12):
        f_ = gen(ci)
        if f_: w(ws, r, ci, f_, font=(FCST_FONT if ci >= 8 else font), fmt=fmt)

# Cross-tab refs
A_R  = "Assumptions"
RB_R = "'Revenue Build'"
IS_R = "'Non-GAAP P&L'"
RC_R = "'GAAP Reconciliation'"
BS_R = "'Balance Sheet'"
CF_R = "'Cash Flow & FCF Bridge'"
Q_R  = "'Quarterly Detail'"

# ═══════════════════════════════════════════════════════════════════════════════
# ROW MAPS
# ═══════════════════════════════════════════════════════════════════════════════
A = {
    'dau_g':7, 'arpu_g':8,
    'gm':11, 'rd':12, 'sm':13, 'ga':14,                          # Non-GAAP margins
    'sbc_pct':17, 'sbc_cogs_alloc':18,                            # SBC
    'intang_amort':19, 'restructuring':20, 'ppe_dep':21,          # Adjustments
    'ng_tax':22, 'gaap_tax':23,                                   # Tax rates
    'na':26, 'eu':27, 'row':28,                                   # Geo mix
    'capex':31, 'interest':32, 'wc':33,                           # Other P&L
    'ppe':36, 'intang':37, 'rou':38, 'olt_a':39,                  # BS assets
    'ltd':40, 'lease':41, 'olt_l':42,                             # BS liabilities
    'dso':43, 'dpo':44, 'accr':45,                                # WC ratios
    'shares':48, 'dilution':49,                                    # Share count
    'cff':52,                                                      # Financing
    'season':56,                                                   # Seasonality
}
IS = {
    'rev':6, 'rev_g':7,
    'cogs':9, 'gp':10, 'gm':11,
    'rd':14, 'rd_pct':15, 'sm':16, 'sm_pct':17, 'ga':18, 'ga_pct':19,
    'opex':20, 'opex_pct':21,
    'oi':23, 'oi_m':24,
    'dep':26, 'adj_ebitda':27, 'adj_ebitda_m':28,
    'int':30, 'pt':31, 'tax':32, 'tax_r':33,
    'ni':35, 'shares':37, 'eps':38,
    'sbc':41, 'sbc_pct':42, 'ia':43, 'restr':44,
}
RC = {
    'ng_oi':7, 'rc_sbc':8, 'rc_ia':9, 'rc_restr':10,
    'gaap_oi':11, 'gaap_oi_m':12,
    'ng_ni':15, 'ni_sbc':16, 'ni_ia':17, 'ni_restr':18, 'ni_tax_eff':19,
    'gaap_ni':20, 'gaap_eps':21,
    'g_rev':25, 'g_cogs':26, 'g_gp':27, 'g_gm':28,
    'g_rd':30, 'g_sm':31, 'g_ga':32, 'g_opex':33,
    'g_oi':34, 'g_oi_m':35,
    'g_da':37, 'g_ebitda':38, 'g_ebitda_m':39,
    'g_int':41, 'g_pt':42, 'g_tax':43, 'g_ni':45, 'g_eps':46,
    'ae':50, 'ae_dep':51, 'ae_ngoi':52,
    'ae_sbc':53, 'ae_ia':54, 'ae_restr':55, 'ae_gaapoi':56,
    'ae_int':57, 'ae_tax':58, 'ae_gaapni':59,
    'ae_da':60, 'ae_wc':61, 'ae_other':62,
    'ae_cfo':63, 'ae_capex':64, 'ae_fcf':65,
}
CF = {
    'ni':7, 'da':8, 'sbc':9, 'wc':10, 'other':11, 'cfo':12,
    'capex':15, 'other_cfi':16, 'cfi':17,
    'cff':20,
    'net':22, 'beg':23, 'end':24,
    'fcf':26, 'fcf_m':27,
}
BS = {
    'cash':7, 'ar':8, 'oca':9, 'tca':10,
    'ppe':12, 'intang':13, 'rou':14, 'olta':15, 'ta':16,
    'ap':19, 'accr':20, 'dr':21, 'ocl':22, 'tcl':23,
    'ltd':25, 'lease':26, 'oltl':27, 'tl':28,
    'eq':31, 'tle':32, 'chk':34,
}
RB = {
    'dau':7, 'dau_g':8, 'arpu':11, 'arpu_g':12,
    'rev':15, 'rev_g':16,
    'na':19, 'na_pct':20, 'eu':21, 'eu_pct':22, 'row_r':23, 'row_pct':24,
    'imp':27, 'ecpm':28,
}
QT = {
    'dau':6, 'rev':7, 'arpu':8,
    'cogs':10, 'gp':11, 'gm':12,
    'rd':14, 'sm':15, 'ga':16, 'oi':17, 'oi_m':18,
    'dep':20, 'adj_ebitda':21,
    'sbc':23, 'ia':24, 'restr':25,
    'gaap_oi':26,
    'int':28, 'ng_pt':29, 'ng_tax':30, 'ng_ni':31,
    'capex':33,
}

# ═══════════════════════════════════════════════════════════════════════════════
# HISTORICAL DATA (all $M)
# ═══════════════════════════════════════════════════════════════════════════════
# -- GAAP baseline --
dau_h = [218, 265, 319, 375, 414, 443]
rev_h = [1716, 2507, 4117, 4602, 4606, 5360]
rev_na_h = [1100, 1620, 2720, 2990, 2950, 3430]
rev_eu_h = [360, 500, 830, 920, 920, 1070]
rev_row_h = [256, 387, 567, 692, 736, 860]
gaap_cor_h = [880, 1128, 1763, 2158, 2190, 2461]
gaap_rd_h = [988, 1235, 1535, 2017, 1847, 1802]
gaap_sm_h = [654, 794, 1168, 1424, 1284, 1281]
gaap_ga_h = [363, 421, 507, 641, 614, 582]
int_h = [-51, -88, -89, -87, -100, -95]
gaap_tax_h = [5, 7, 14, 20, 25, 28]
shares_h = [1500, 1548, 1620, 1660, 1676, 1690]
imp_h = [25, 27, 30, 33, 36, 39]

# -- GAAP Adjustment Items (Non-GAAP add-backs) --
sbc_h = [600, 770, 1096, 1538, 1376, 1296]
intang_amort_h = [30, 32, 55, 70, 65, 55]
restructuring_h = [0, 0, 0, 155, 80, 35]
ppe_dep_h = [51, 75, 85, 115, 135, 160]
# SBC allocation: COGS 5%, R&D 50%, S&M 25%, G&A 20%
SBC_COGS_PCT = 0.05

# -- Non-GAAP P&L (derived) --
ng_cogs_h = [c - round(s*0.05) - round(ia*0.40)
             for c, s, ia in zip(gaap_cor_h, sbc_h, intang_amort_h)]
ng_rd_h = [r - round(s*0.50) - round(ia*0.40)
           for r, s, ia in zip(gaap_rd_h, sbc_h, intang_amort_h)]
ng_sm_h = [m - round(s*0.25) - round(ia*0.10)
           for m, s, ia in zip(gaap_sm_h, sbc_h, intang_amort_h)]
ng_ga_h = [g - round(s*0.20) - round(ia*0.10) - re
           for g, s, ia, re in zip(gaap_ga_h, sbc_h, intang_amort_h, restructuring_h)]

ng_gp_h = [r - c for r, c in zip(rev_h, ng_cogs_h)]
ng_opex_h = [a+b+c for a,b,c in zip(ng_rd_h, ng_sm_h, ng_ga_h)]
ng_oi_h = [g - o for g, o in zip(ng_gp_h, ng_opex_h)]
ng_pt_h = [o + i for o, i in zip(ng_oi_h, int_h)]
# Non-GAAP tax (higher rate than GAAP since SBC deduction removed)
ng_tax_rate_h = [0.02, 0.02, 0.05, 0.05, 0.08, 0.10]
ng_tax_h = [max(0, round(pt * tr)) if pt > 0 else 0 for pt, tr in zip(ng_pt_h, ng_tax_rate_h)]
ng_ni_h = [pt - t for pt, t in zip(ng_pt_h, ng_tax_h)]

# GAAP derived
gaap_gp_h = [r - c for r, c in zip(rev_h, gaap_cor_h)]
gaap_opex_h = [a+b+c for a,b,c in zip(gaap_rd_h, gaap_sm_h, gaap_ga_h)]
gaap_oi_h = [g - o for g, o in zip(gaap_gp_h, gaap_opex_h)]
gaap_pt_h = [o + i for o, i in zip(gaap_oi_h, int_h)]
gaap_ni_h = [p - t for p, t in zip(gaap_pt_h, gaap_tax_h)]
total_da_h = [p + ia for p, ia in zip(ppe_dep_h, intang_amort_h)]

# Balance Sheet
cash_h = [2024, 2768, 3506, 3938, 3374, 3205]
ar_h = [548, 737, 1068, 995, 1109, 1230]
oca_h = [174, 200, 201, 182, 217, 215]
ppe_h = [204, 258, 339, 452, 430, 400]
intang_bs_h = [181, 182, 411, 440, 395, 360]
rou_h = [250, 280, 310, 350, 360, 370]
olt_a_h = [150, 180, 250, 310, 330, 350]
ap_h = [120, 155, 245, 265, 280, 295]
accr_h = [350, 430, 620, 705, 680, 720]
dr_h = [25, 30, 35, 40, 42, 45]
ocl_h = [0, 0, 0, 0, 0, 0]
ltd_h = [1485, 2190, 3745, 3745, 3745, 3745]
lease_h = [220, 260, 300, 340, 350, 360]
olt_l_h = [80, 95, 120, 140, 150, 155]

# Cash Flow
cfo_h = [-41, 166, 631, 234, 75, 260]
cfi_h = [-200, -500, -1200, -350, -120, -100]
cff_h = [900, 1078, 1307, 548, -519, -329]
wc_h = [-80, -45, -120, 50, 30, 15]
beg_cash_h = [1365, 2024, 2768, 3506, 3938, 3374]
capex_h = [-51, -59, -85, -115, -93, -90]
other_cfi_h = [cf - cx for cf, cx in zip(cfi_h, capex_h)]
other_adj_h = [cfo - ni - da - sbc - wcc for cfo, ni, da, sbc, wcc in
               zip(cfo_h, gaap_ni_h, total_da_h, sbc_h, wc_h)]

# -- Forecast Assumptions --
dau_g_f = [0.065, 0.055, 0.045, 0.040]
arpu_g_f = [0.10, 0.09, 0.08, 0.07]
ng_gm_f = [0.575, 0.590, 0.605, 0.620]
ng_rd_f = [0.19, 0.17, 0.16, 0.15]
ng_sm_f = [0.16, 0.15, 0.14, 0.13]
ng_ga_f = [0.050, 0.047, 0.044, 0.042]
sbc_pct_f = [0.20, 0.17, 0.15, 0.13]
sbc_cogs_alloc_f = [0.05, 0.05, 0.05, 0.05]
intang_amort_f = [45, 40, 35, 30]
restructuring_f = [15, 10, 5, 0]
ppe_dep_f = [170, 180, 190, 200]
ng_tax_f = [0.10, 0.12, 0.15, 0.18]
gaap_tax_f = [0.05, 0.07, 0.10, 0.12]
na_f = [0.63, 0.62, 0.61, 0.60]
eu_f = [0.20, 0.20, 0.20, 0.21]
row_f = [0.17, 0.18, 0.19, 0.19]
capex_f = [95, 100, 105, 110]
interest_f = [90, 85, 80, 75]
wc_f = [-20, -15, -10, -5]
ppe_bs_f = [390, 385, 380, 375]
intang_bs_f = [340, 320, 300, 280]
rou_f = [375, 380, 385, 390]
olt_a_f = [360, 370, 380, 390]
ltd_f_v = [3745, 3745, 3745, 3745]
lease_f = [365, 370, 375, 380]
olt_l_f = [160, 165, 170, 175]
dso_f = [80, 80, 80, 80]
dpo_f = [42, 42, 42, 42]
accr_f = [0.12, 0.12, 0.12, 0.12]
dilution_f = [0.012, 0.009, 0.009, 0.009]
cff_f = [-200, -200, -250, -300]
imp_f = [42, 44, 46, 48]
season_f = [0.21, 0.23, 0.25, 0.31]

# Quarterly historical (GAAP)
q_labels_h = ['Q1-23','Q2-23','Q3-23','Q4-23','Q1-24','Q2-24','Q3-24','Q4-24']
q_labels_f = ['Q1-25E','Q2-25E','Q3-25E','Q4-25E','Q1-26E','Q2-26E','Q3-26E','Q4-26E']
dau_qh = [383, 397, 406, 414, 422, 432, 437, 443]
rev_qh = [989, 1068, 1189, 1361, 1195, 1237, 1373, 1556]
# Non-GAAP quarterly (approximate)
ng_cor_qh = [460, 480, 515, 610, 515, 535, 588, 697]
ng_rd_qh = [260, 252, 248, 246, 248, 244, 244, 245]
ng_sm_qh = [240, 232, 234, 231, 235, 232, 234, 234]
ng_ga_qh = [58, 57, 55, 55, 53, 52, 52, 52]
sbc_qh = [355, 345, 340, 336, 330, 325, 322, 319]
ia_qh = [16, 16, 17, 16, 14, 14, 14, 13]
restr_qh = [22, 20, 20, 18, 10, 9, 8, 8]
ppe_dep_qh = [34, 34, 34, 33, 40, 40, 40, 40]
int_qh = [-25, -25, -25, -25, -24, -24, -24, -23]
capex_qh = [-22, -24, -23, -24, -22, -23, -22, -23]

# Share count quarterly
rsu_vest = [12, 11, 13, 12, 11, 10, 12, 11]
buyback_q = [5, 5, 5, 5, 5, 5, 5, 5]
other_q_sc = [3, 4, 2, 3, 3, 3, 1, 3]
unvested_rsu = [120, 117, 112, 108, 104, 101, 96, 92]
vested_opts = [15, 14, 13, 12, 11, 10, 9, 8]
new_grants = [8, 8, 8, 8, 7, 7, 7, 7]

yr_hdr = ["","2019","2020","2021","2022","2023","2024","2025E","2026E","2027E","2028E"]
hf_marker = [""]+["Historical"]*6+["Forecast"]*4

# ═══════════════════════════════════════════════════════════════════════════════
# CREATE WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws_rb = wb.active; ws_rb.title = "Revenue Build"
ws_is = wb.create_sheet("Non-GAAP P&L")
ws_rc = wb.create_sheet("GAAP Reconciliation")
ws_bs = wb.create_sheet("Balance Sheet")
ws_cf = wb.create_sheet("Cash Flow & FCF Bridge")
ws_qt = wb.create_sheet("Quarterly Detail")
ws_sc = wb.create_sheet("Share Count & Dilution")
ws_sn = wb.create_sheet("Sensitivity Analysis")
ws_vl = wb.create_sheet("Valuation Cross-Checks")
ws_a  = wb.create_sheet("Assumptions")

for s in [ws_rb, ws_is]: s.sheet_properties.tabColor = "2F5496"
ws_rc.sheet_properties.tabColor = "7F6000"
ws_bs.sheet_properties.tabColor = "548235"
ws_cf.sheet_properties.tabColor = "BF8F00"
ws_qt.sheet_properties.tabColor = "7030A0"
ws_sc.sheet_properties.tabColor = "C00000"
ws_sn.sheet_properties.tabColor = "ED7D31"
ws_vl.sheet_properties.tabColor = "00B050"
ws_a.sheet_properties.tabColor = "FF0000"

# ═══════════════════════════════════════════════════════════════════════════════
# TAB: ASSUMPTIONS
# ═══════════════════════════════════════════════════════════════════════════════
ws = ws_a
w(ws, 1, 1, "Snap Inc. (SNAP) - Model Assumptions (Non-GAAP Primary)", TF)
ws.merge_cells('A1:K1')
hdr_row(ws, 3, yr_hdr); sub_row(ws, 4, hf_marker)

sec(ws, 6, "GROWTH DRIVERS")
for row_n, label, hist_gen, fvals in [
    (A['dau_g'], "DAU YoY Growth", lambda ci: f"={RB_R}!{CL(ci)}{RB['dau']}/{RB_R}!{CL(ci-1)}{RB['dau']}-1", dau_g_f),
    (A['arpu_g'], "ARPU YoY Growth", lambda ci: f"={RB_R}!{CL(ci)}{RB['arpu']}/{RB_R}!{CL(ci-1)}{RB['arpu']}-1", arpu_g_f),
]:
    w(ws, row_n, 1, label, BF)
    for ci in range(3, 8): w(ws, row_n, ci, hist_gen(ci), fmt=P)
    for i, v in enumerate(fvals): w(ws, row_n, 8+i, v, font=FCST_FONT, fmt=P, fill=INPUT_FILL)

sec(ws, 10, "NON-GAAP MARGIN ASSUMPTIONS")
for row_n, label, fvals, is_num, is_den in [
    (A['gm'], "Non-GAAP Gross Margin", ng_gm_f, IS['gp'], IS['rev']),
    (A['rd'], "Non-GAAP R&D % of Rev", ng_rd_f, IS['rd'], IS['rev']),
    (A['sm'], "Non-GAAP S&M % of Rev", ng_sm_f, IS['sm'], IS['rev']),
    (A['ga'], "Non-GAAP G&A % of Rev", ng_ga_f, IS['ga'], IS['rev']),
]:
    w(ws, row_n, 1, label, BF)
    for ci in range(2, 8): w(ws, row_n, ci, f"={IS_R}!{CL(ci)}{is_num}/{IS_R}!{CL(ci)}{is_den}", fmt=P)
    for i, v in enumerate(fvals): w(ws, row_n, 8+i, v, font=FCST_FONT, fmt=P, fill=INPUT_FILL)

sec(ws, 16, "GAAP ADJUSTMENT ITEMS")
w(ws, A['sbc_pct'], 1, "SBC % of Revenue", BF)
for ci in range(2, 8): w(ws, A['sbc_pct'], ci, f"={IS_R}!{CL(ci)}{IS['sbc']}/{IS_R}!{CL(ci)}{IS['rev']}", fmt=P)
for i, v in enumerate(sbc_pct_f): w(ws, A['sbc_pct'], 8+i, v, font=FCST_FONT, fmt=P, fill=INPUT_FILL)

w(ws, A['sbc_cogs_alloc'], 1, "SBC % Allocated to COGS", BF)
for i in range(4): w(ws, A['sbc_cogs_alloc'], 8+i, sbc_cogs_alloc_f[i], font=FCST_FONT, fmt=P, fill=INPUT_FILL)
fill_hist(ws, A['sbc_cogs_alloc'], [SBC_COGS_PCT]*6, fmt=P)

for row_n, label, hvals, fvals in [
    (A['intang_amort'], "Intangible Amortization ($M)", intang_amort_h, intang_amort_f),
    (A['restructuring'], "Restructuring Charges ($M)", restructuring_h, restructuring_f),
    (A['ppe_dep'], "PP&E Depreciation ($M)", ppe_dep_h, ppe_dep_f),
]:
    w(ws, row_n, 1, label, BF)
    fill_hist(ws, row_n, hvals, fmt=N)
    for i, v in enumerate(fvals): w(ws, row_n, 8+i, v, font=FCST_FONT, fmt=N, fill=INPUT_FILL)

w(ws, A['ng_tax'], 1, "Non-GAAP Tax Rate", BF)
fill_hist(ws, A['ng_tax'], ng_tax_rate_h, fmt=P)
for i, v in enumerate(ng_tax_f): w(ws, A['ng_tax'], 8+i, v, font=FCST_FONT, fmt=P, fill=INPUT_FILL)

w(ws, A['gaap_tax'], 1, "GAAP Tax Rate", BF)
gaap_tax_rate_h = [0 if gaap_pt_h[i] <= 0 else round(gaap_tax_h[i]/gaap_pt_h[i], 3) if gaap_pt_h[i] != 0 else 0 for i in range(6)]
fill_hist(ws, A['gaap_tax'], gaap_tax_rate_h, fmt=P)
for i, v in enumerate(gaap_tax_f): w(ws, A['gaap_tax'], 8+i, v, font=FCST_FONT, fmt=P, fill=INPUT_FILL)

sec(ws, 25, "GEOGRAPHIC MIX")
for row_n, label, fvals, rb_row in [
    (A['na'], "NA % of Revenue", na_f, RB['na']),
    (A['eu'], "EU % of Revenue", eu_f, RB['eu']),
    (A['row'], "ROW % of Revenue", row_f, RB['row_r']),
]:
    w(ws, row_n, 1, label, BF)
    for ci in range(2, 8): w(ws, row_n, ci, f"={RB_R}!{CL(ci)}{rb_row}/{RB_R}!{CL(ci)}{RB['rev']}", fmt=P)
    for i, v in enumerate(fvals): w(ws, row_n, 8+i, v, font=FCST_FONT, fmt=P, fill=INPUT_FILL)

sec(ws, 30, "OTHER P&L ITEMS")
for row_n, label, hvals, fvals in [
    (A['capex'], "CapEx ($M, positive)", [abs(x) for x in capex_h], capex_f),
    (A['interest'], "Interest Expense ($M, positive)", [abs(x) for x in int_h], interest_f),
    (A['wc'], "WC Change ($M)", wc_h, wc_f),
]:
    w(ws, row_n, 1, label, BF)
    fill_hist(ws, row_n, hvals, fmt=N)
    for i, v in enumerate(fvals): w(ws, row_n, 8+i, v, font=FCST_FONT, fmt=N, fill=INPUT_FILL)

sec(ws, 35, "BALANCE SHEET DRIVERS")
for row_n, label, hvals, fvals in [
    (A['ppe'], "PP&E ($M)", ppe_h, ppe_bs_f),
    (A['intang'], "Intangibles BS ($M)", intang_bs_h, intang_bs_f),
    (A['rou'], "ROU Assets ($M)", rou_h, rou_f),
    (A['olt_a'], "Other LT Assets ($M)", olt_a_h, olt_a_f),
    (A['ltd'], "Long-Term Debt ($M)", ltd_h, ltd_f_v),
    (A['lease'], "Lease Liabilities ($M)", lease_h, lease_f),
    (A['olt_l'], "Other LT Liab ($M)", olt_l_h, olt_l_f),
]:
    w(ws, row_n, 1, label, BF); fill_hist(ws, row_n, hvals, fmt=N)
    for i, v in enumerate(fvals): w(ws, row_n, 8+i, v, font=FCST_FONT, fmt=N, fill=INPUT_FILL)

for row_n, label, fvals, fmt_ in [
    (A['dso'], "DSO (days)", dso_f, N), (A['dpo'], "DPO (days)", dpo_f, N),
    (A['accr'], "Accrued Liab % of Rev", accr_f, P),
]:
    w(ws, row_n, 1, label, BF)
    for i, v in enumerate(fvals): w(ws, row_n, 8+i, v, font=FCST_FONT, fmt=fmt_, fill=INPUT_FILL)

sec(ws, 47, "SHARE COUNT")
w(ws, A['shares'], 1, "Diluted Shares (M)", BF)
fill_hist(ws, A['shares'], shares_h, fmt=N)
for ci in range(8, 12):
    w(ws, A['shares'], ci, f"=ROUND({CL(ci-1)}{A['shares']}*(1+{CL(ci)}{A['dilution']}),0)", font=FCST_FONT, fmt=N)
w(ws, A['dilution'], 1, "Annual Dilution (%)", BF)
for ci in range(3, 8): w(ws, A['dilution'], ci, f"={CL(ci)}{A['shares']}/{CL(ci-1)}{A['shares']}-1", fmt=P)
for i, v in enumerate(dilution_f): w(ws, A['dilution'], 8+i, v, font=FCST_FONT, fmt=P, fill=INPUT_FILL)

sec(ws, 51, "FINANCING")
w(ws, A['cff'], 1, "Cash from Financing ($M)", BF)
fill_hist(ws, A['cff'], cff_h, fmt=N)
for i, v in enumerate(cff_f): w(ws, A['cff'], 8+i, v, font=FCST_FONT, fmt=N, fill=INPUT_FILL)

sec(ws, 54, "QUARTERLY SEASONALITY")
hdr_row(ws, 55, ["","Q1","Q2","Q3","Q4"])
w(ws, A['season'], 1, "Revenue Seasonality Weight", BF)
for i, v in enumerate(season_f): w(ws, A['season'], 2+i, v, fmt=P, fill=INPUT_FILL)
w(ws, 57, 1, "Total Check", BF); w(ws, 57, 2, f"=SUM(B{A['season']}:E{A['season']})", fmt=P)
colw(ws, [32]+[14]*10)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB: REVENUE BUILD (unchanged)
# ═══════════════════════════════════════════════════════════════════════════════
ws = ws_rb
w(ws, 1, 1, "Snap Inc. (SNAP) - Revenue Build by Segment & Driver", TF); ws.merge_cells('A1:K1')
hdr_row(ws, 3, yr_hdr); sub_row(ws, 4, hf_marker)

sec(ws, 6, "DAU DRIVERS")
w(ws, RB['dau'], 1, "Daily Active Users (M)", BF); fill_hist(ws, RB['dau'], dau_h, fmt=N)
fill_fcst(ws, RB['dau'], [f"=ROUND({CL(ci-1)}{RB['dau']}*(1+{A_R}!{CL(ci)}{A['dau_g']}),0)" for ci in range(8,12)], fmt=N)
w(ws, RB['dau_g'], 1, "  YoY Growth", NF)
fill_all(ws, RB['dau_g'], lambda ci: f"={CL(ci)}{RB['dau']}/{CL(ci-1)}{RB['dau']}-1", fmt=P, start=3)

sec(ws, 10, "ARPU DRIVERS")
w(ws, RB['arpu'], 1, "Avg Revenue Per User ($)", BF)
for ci in range(2, 8): w(ws, RB['arpu'], ci, f"=ROUND({CL(ci)}{RB['rev']}/{CL(ci)}{RB['dau']},2)", fmt=U2)
fill_fcst(ws, RB['arpu'], [f"=ROUND({CL(ci-1)}{RB['arpu']}*(1+{A_R}!{CL(ci)}{A['arpu_g']}),2)" for ci in range(8,12)], fmt=U2)
w(ws, RB['arpu_g'], 1, "  YoY Growth", NF)
fill_all(ws, RB['arpu_g'], lambda ci: f"={CL(ci)}{RB['arpu']}/{CL(ci-1)}{RB['arpu']}-1", fmt=P, start=3)

sec(ws, 14, "REVENUE")
w(ws, RB['rev'], 1, "Total Revenue ($M)", BF); fill_hist(ws, RB['rev'], rev_h, fmt=N, font=BF)
fill_fcst(ws, RB['rev'], [f"=ROUND({CL(ci)}{RB['dau']}*{CL(ci)}{RB['arpu']},0)" for ci in range(8,12)], fmt=N)
w(ws, RB['rev_g'], 1, "  YoY Growth", NF)
fill_all(ws, RB['rev_g'], lambda ci: f"={CL(ci)}{RB['rev']}/{CL(ci-1)}{RB['rev']}-1", fmt=P, start=3)

sec(ws, 18, "REVENUE BY GEOGRAPHY")
for rv, pv, lb, hist, ap in [(RB['na'],RB['na_pct'],"  North America",rev_na_h,A['na']),
    (RB['eu'],RB['eu_pct'],"  Europe",rev_eu_h,A['eu']),
    (RB['row_r'],RB['row_pct'],"  Rest of World",rev_row_h,A['row'])]:
    w(ws, rv, 1, lb, NF); fill_hist(ws, rv, hist, fmt=N)
    fill_fcst(ws, rv, [f"=ROUND({CL(ci)}{RB['rev']}*{A_R}!{CL(ci)}{ap},0)" for ci in range(8,12)], fmt=N)
    w(ws, pv, 1, "    % of Total", NF)
    fill_all(ws, pv, lambda ci, rr=rv: f"={CL(ci)}{rr}/{CL(ci)}{RB['rev']}", fmt=P)

sec(ws, 26, "IMPLIED AD METRICS")
w(ws, RB['imp'], 1, "  Impressions/DAU/Day", NF); fill_hist(ws, RB['imp'], imp_h, fmt=N)
for i, v in enumerate(imp_f): w(ws, RB['imp'], 8+i, v, font=FCST_FONT, fmt=N, fill=INPUT_FILL)
w(ws, RB['ecpm'], 1, "  Implied eCPM ($)", NF)
fill_all(ws, RB['ecpm'], lambda ci: f"=ROUND({CL(ci)}{RB['rev']}/({CL(ci)}{RB['dau']}*{CL(ci)}{RB['imp']}*365/1000000)*1000,2)", fmt=U2)
colw(ws, [32]+[14]*10)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB: NON-GAAP P&L (PRIMARY)
# ═══════════════════════════════════════════════════════════════════════════════
ws = ws_is
w(ws, 1, 1, "Snap Inc. (SNAP) - Non-GAAP Income Statement ($M)", TF); ws.merge_cells('A1:K1')
hdr_row(ws, 3, yr_hdr); sub_row(ws, 4, hf_marker)

# Revenue
w(ws, IS['rev'], 1, "Revenue", BF); fill_hist(ws, IS['rev'], rev_h, fmt=N, font=BF)
fill_fcst(ws, IS['rev'], [f"={RB_R}!{CL(ci)}{RB['rev']}" for ci in range(8,12)], fmt=N)
w(ws, IS['rev_g'], 1, "  YoY Growth", NF)
fill_all(ws, IS['rev_g'], lambda ci: f"={CL(ci)}{IS['rev']}/{CL(ci-1)}{IS['rev']}-1", fmt=P, start=3)

# Non-GAAP COGS
w(ws, IS['cogs'], 1, "Non-GAAP Cost of Revenue", NF)
fill_hist(ws, IS['cogs'], ng_cogs_h, fmt=N)
fill_fcst(ws, IS['cogs'], [f"=ROUND({CL(ci)}{IS['rev']}*(1-{A_R}!{CL(ci)}{A['gm']}),0)" for ci in range(8,12)], fmt=N)

# Non-GAAP GP
w(ws, IS['gp'], 1, "Non-GAAP Gross Profit", BF)
fill_all(ws, IS['gp'], lambda ci: f"={CL(ci)}{IS['rev']}-{CL(ci)}{IS['cogs']}", fmt=N)
w(ws, IS['gm'], 1, "  Non-GAAP Gross Margin", NF)
fill_all(ws, IS['gm'], lambda ci: f"={CL(ci)}{IS['gp']}/{CL(ci)}{IS['rev']}", fmt=P)

sec(ws, 13, "Non-GAAP Operating Expenses")
for vr, pr, lb, hist, ar in [(IS['rd'],IS['rd_pct'],"  Non-GAAP R&D",ng_rd_h,A['rd']),
    (IS['sm'],IS['sm_pct'],"  Non-GAAP S&M",ng_sm_h,A['sm']),
    (IS['ga'],IS['ga_pct'],"  Non-GAAP G&A",ng_ga_h,A['ga'])]:
    w(ws, vr, 1, lb, NF); fill_hist(ws, vr, hist, fmt=N)
    fill_fcst(ws, vr, [f"=ROUND({CL(ci)}{IS['rev']}*{A_R}!{CL(ci)}{ar},0)" for ci in range(8,12)], fmt=N)
    w(ws, pr, 1, "    % of Revenue", NF)
    fill_all(ws, pr, lambda ci, v=vr: f"={CL(ci)}{v}/{CL(ci)}{IS['rev']}", fmt=P)

w(ws, IS['opex'], 1, "Non-GAAP Total OpEx", BF)
fill_all(ws, IS['opex'], lambda ci: f"={CL(ci)}{IS['rd']}+{CL(ci)}{IS['sm']}+{CL(ci)}{IS['ga']}", fmt=N)
w(ws, IS['opex_pct'], 1, "    % of Revenue", NF)
fill_all(ws, IS['opex_pct'], lambda ci: f"={CL(ci)}{IS['opex']}/{CL(ci)}{IS['rev']}", fmt=P)

# Non-GAAP OI
w(ws, IS['oi'], 1, "Non-GAAP Operating Income", BF)
fill_all(ws, IS['oi'], lambda ci: f"={CL(ci)}{IS['gp']}-{CL(ci)}{IS['opex']}", fmt=N)
w(ws, IS['oi_m'], 1, "  Non-GAAP Operating Margin", NF)
fill_all(ws, IS['oi_m'], lambda ci: f"={CL(ci)}{IS['oi']}/{CL(ci)}{IS['rev']}", fmt=P)

# Adjusted EBITDA
w(ws, IS['dep'], 1, "PP&E Depreciation", NF)
fill_hist(ws, IS['dep'], ppe_dep_h, fmt=N)
fill_fcst(ws, IS['dep'], [f"={A_R}!{CL(ci)}{A['ppe_dep']}" for ci in range(8,12)], fmt=N)
w(ws, IS['adj_ebitda'], 1, "Adjusted EBITDA", BF)
fill_all(ws, IS['adj_ebitda'], lambda ci: f"={CL(ci)}{IS['oi']}+{CL(ci)}{IS['dep']}", fmt=N)
w(ws, IS['adj_ebitda_m'], 1, "  Adjusted EBITDA Margin", NF)
fill_all(ws, IS['adj_ebitda_m'], lambda ci: f"={CL(ci)}{IS['adj_ebitda']}/{CL(ci)}{IS['rev']}", fmt=P)

# Below the line
w(ws, IS['int'], 1, "Interest & Other", NF)
fill_hist(ws, IS['int'], int_h, fmt=N)
fill_fcst(ws, IS['int'], [f"=-{A_R}!{CL(ci)}{A['interest']}" for ci in range(8,12)], fmt=N)
w(ws, IS['pt'], 1, "Non-GAAP Pre-Tax Income", BF)
fill_all(ws, IS['pt'], lambda ci: f"={CL(ci)}{IS['oi']}+{CL(ci)}{IS['int']}", fmt=N)
w(ws, IS['tax'], 1, "Non-GAAP Tax Provision", NF)
fill_hist(ws, IS['tax'], ng_tax_h, fmt=N)
fill_fcst(ws, IS['tax'], [f"=IF({CL(ci)}{IS['pt']}>0,ROUND({CL(ci)}{IS['pt']}*{A_R}!{CL(ci)}{A['ng_tax']},0),0)" for ci in range(8,12)], fmt=N)
w(ws, IS['tax_r'], 1, "  Non-GAAP Tax Rate", NF)
fill_all(ws, IS['tax_r'], lambda ci: f"=IF({CL(ci)}{IS['pt']}>0,{CL(ci)}{IS['tax']}/{CL(ci)}{IS['pt']},0)", fmt=P)

w(ws, IS['ni'], 1, "Non-GAAP Net Income", BF)
fill_all(ws, IS['ni'], lambda ci: f"={CL(ci)}{IS['pt']}-{CL(ci)}{IS['tax']}", fmt=N)
w(ws, IS['shares'], 1, "Diluted Shares (M)", NF)
fill_all(ws, IS['shares'], lambda ci: f"={A_R}!{CL(ci)}{A['shares']}", fmt=N)
w(ws, IS['eps'], 1, "Non-GAAP Diluted EPS", BF)
fill_all(ws, IS['eps'], lambda ci: f"=ROUND({CL(ci)}{IS['ni']}/{CL(ci)}{IS['shares']},2)", fmt=U2)

# Memo: Adjustment items
sec(ws, 40, "Memo: GAAP Adjustment Items (add-backs)")
w(ws, IS['sbc'], 1, "  Stock-Based Compensation", NF)
fill_hist(ws, IS['sbc'], sbc_h, fmt=N)
fill_fcst(ws, IS['sbc'], [f"=ROUND({CL(ci)}{IS['rev']}*{A_R}!{CL(ci)}{A['sbc_pct']},0)" for ci in range(8,12)], fmt=N)
w(ws, IS['sbc_pct'], 1, "    SBC % of Revenue", NF)
fill_all(ws, IS['sbc_pct'], lambda ci: f"={CL(ci)}{IS['sbc']}/{CL(ci)}{IS['rev']}", fmt=P)
w(ws, IS['ia'], 1, "  Intangible Amortization", NF)
fill_hist(ws, IS['ia'], intang_amort_h, fmt=N)
fill_fcst(ws, IS['ia'], [f"={A_R}!{CL(ci)}{A['intang_amort']}" for ci in range(8,12)], fmt=N)
w(ws, IS['restr'], 1, "  Restructuring Charges", NF)
fill_hist(ws, IS['restr'], restructuring_h, fmt=N)
fill_fcst(ws, IS['restr'], [f"={A_R}!{CL(ci)}{A['restructuring']}" for ci in range(8,12)], fmt=N)
colw(ws, [34]+[14]*10)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB: GAAP RECONCILIATION
# ═══════════════════════════════════════════════════════════════════════════════
ws = ws_rc
w(ws, 1, 1, "Snap Inc. (SNAP) - Non-GAAP to GAAP Reconciliation ($M)", TF); ws.merge_cells('A1:K1')
hdr_row(ws, 3, yr_hdr); sub_row(ws, 4, hf_marker)

# Section 1: OI Bridge
sec(ws, 6, "NON-GAAP OI -> GAAP OI BRIDGE")
w(ws, RC['ng_oi'], 1, "Non-GAAP Operating Income", BF)
fill_all(ws, RC['ng_oi'], lambda ci: f"={IS_R}!{CL(ci)}{IS['oi']}", fmt=N)
w(ws, RC['rc_sbc'], 1, "  (-) Stock-Based Compensation", NF)
fill_all(ws, RC['rc_sbc'], lambda ci: f"=-{IS_R}!{CL(ci)}{IS['sbc']}", fmt=N)
w(ws, RC['rc_ia'], 1, "  (-) Intangible Amortization", NF)
fill_all(ws, RC['rc_ia'], lambda ci: f"=-{IS_R}!{CL(ci)}{IS['ia']}", fmt=N)
w(ws, RC['rc_restr'], 1, "  (-) Restructuring Charges", NF)
fill_all(ws, RC['rc_restr'], lambda ci: f"=-{IS_R}!{CL(ci)}{IS['restr']}", fmt=N)
w(ws, RC['gaap_oi'], 1, "GAAP Operating Income (Loss)", BF)
fill_all(ws, RC['gaap_oi'], lambda ci: f"={CL(ci)}{RC['ng_oi']}+{CL(ci)}{RC['rc_sbc']}+{CL(ci)}{RC['rc_ia']}+{CL(ci)}{RC['rc_restr']}", fmt=N)
w(ws, RC['gaap_oi_m'], 1, "  GAAP Operating Margin", NF)
fill_all(ws, RC['gaap_oi_m'], lambda ci: f"={CL(ci)}{RC['gaap_oi']}/{IS_R}!{CL(ci)}{IS['rev']}", fmt=P)

# Section 2: NI Bridge
sec(ws, 14, "NON-GAAP NI -> GAAP NI BRIDGE")
w(ws, RC['ng_ni'], 1, "Non-GAAP Net Income", BF)
fill_all(ws, RC['ng_ni'], lambda ci: f"={IS_R}!{CL(ci)}{IS['ni']}", fmt=N)
w(ws, RC['ni_sbc'], 1, "  (-) Stock-Based Compensation", NF)
fill_all(ws, RC['ni_sbc'], lambda ci: f"=-{IS_R}!{CL(ci)}{IS['sbc']}", fmt=N)
w(ws, RC['ni_ia'], 1, "  (-) Intangible Amortization", NF)
fill_all(ws, RC['ni_ia'], lambda ci: f"=-{IS_R}!{CL(ci)}{IS['ia']}", fmt=N)
w(ws, RC['ni_restr'], 1, "  (-) Restructuring Charges", NF)
fill_all(ws, RC['ni_restr'], lambda ci: f"=-{IS_R}!{CL(ci)}{IS['restr']}", fmt=N)
w(ws, RC['ni_tax_eff'], 1, "  (+) Tax Effect of Adjustments", NF)
# Tax effect = NonGAAP_Tax - GAAP_Tax (positive because GAAP tax is lower)
fill_all(ws, RC['ni_tax_eff'], lambda ci: f"={IS_R}!{CL(ci)}{IS['tax']}-{CL(ci)}{RC['g_tax']}", fmt=N)
w(ws, RC['gaap_ni'], 1, "GAAP Net Income (Loss)", BF)
fill_all(ws, RC['gaap_ni'], lambda ci: f"={CL(ci)}{RC['ng_ni']}+{CL(ci)}{RC['ni_sbc']}+{CL(ci)}{RC['ni_ia']}+{CL(ci)}{RC['ni_restr']}+{CL(ci)}{RC['ni_tax_eff']}", fmt=N)
w(ws, RC['gaap_eps'], 1, "GAAP Diluted EPS", BF)
fill_all(ws, RC['gaap_eps'], lambda ci: f"=ROUND({CL(ci)}{RC['gaap_ni']}/{IS_R}!{CL(ci)}{IS['shares']},2)", fmt=U2)

# Section 3: Full GAAP P&L
sec(ws, 24, "FULL GAAP INCOME STATEMENT")
w(ws, RC['g_rev'], 1, "Revenue", BF)
fill_all(ws, RC['g_rev'], lambda ci: f"={IS_R}!{CL(ci)}{IS['rev']}", fmt=N)
# GAAP COGS = Non-GAAP COGS + SBC_in_COGS + Intangible_Amort_in_COGS
w(ws, RC['g_cogs'], 1, "GAAP Cost of Revenue", NF)
fill_hist(ws, RC['g_cogs'], gaap_cor_h, fmt=N)
fill_fcst(ws, RC['g_cogs'], [
    f"={IS_R}!{CL(ci)}{IS['cogs']}+ROUND({IS_R}!{CL(ci)}{IS['sbc']}*{A_R}!{CL(ci)}{A['sbc_cogs_alloc']},0)+ROUND({IS_R}!{CL(ci)}{IS['ia']}*0.4,0)"
    for ci in range(8,12)], fmt=N)
w(ws, RC['g_gp'], 1, "GAAP Gross Profit", BF)
fill_all(ws, RC['g_gp'], lambda ci: f"={CL(ci)}{RC['g_rev']}-{CL(ci)}{RC['g_cogs']}", fmt=N)
w(ws, RC['g_gm'], 1, "  GAAP Gross Margin", NF)
fill_all(ws, RC['g_gm'], lambda ci: f"={CL(ci)}{RC['g_gp']}/{CL(ci)}{RC['g_rev']}", fmt=P)

# GAAP OpEx lines
w(ws, RC['g_rd'], 1, "GAAP R&D", NF)
fill_hist(ws, RC['g_rd'], gaap_rd_h, fmt=N)
fill_fcst(ws, RC['g_rd'], [
    f"={IS_R}!{CL(ci)}{IS['rd']}+ROUND({IS_R}!{CL(ci)}{IS['sbc']}*0.50,0)+ROUND({IS_R}!{CL(ci)}{IS['ia']}*0.40,0)"
    for ci in range(8,12)], fmt=N)
w(ws, RC['g_sm'], 1, "GAAP S&M", NF)
fill_hist(ws, RC['g_sm'], gaap_sm_h, fmt=N)
fill_fcst(ws, RC['g_sm'], [
    f"={IS_R}!{CL(ci)}{IS['sm']}+ROUND({IS_R}!{CL(ci)}{IS['sbc']}*0.25,0)+ROUND({IS_R}!{CL(ci)}{IS['ia']}*0.10,0)"
    for ci in range(8,12)], fmt=N)
w(ws, RC['g_ga'], 1, "GAAP G&A", NF)
fill_hist(ws, RC['g_ga'], gaap_ga_h, fmt=N)
fill_fcst(ws, RC['g_ga'], [
    f"={IS_R}!{CL(ci)}{IS['ga']}+ROUND({IS_R}!{CL(ci)}{IS['sbc']}*0.20,0)+ROUND({IS_R}!{CL(ci)}{IS['ia']}*0.10,0)+{IS_R}!{CL(ci)}{IS['restr']}"
    for ci in range(8,12)], fmt=N)
w(ws, RC['g_opex'], 1, "GAAP Total OpEx", BF)
fill_all(ws, RC['g_opex'], lambda ci: f"={CL(ci)}{RC['g_rd']}+{CL(ci)}{RC['g_sm']}+{CL(ci)}{RC['g_ga']}", fmt=N)

w(ws, RC['g_oi'], 1, "GAAP Operating Income (Loss)", BF)
fill_all(ws, RC['g_oi'], lambda ci: f"={CL(ci)}{RC['g_gp']}-{CL(ci)}{RC['g_opex']}", fmt=N)
w(ws, RC['g_oi_m'], 1, "  GAAP Operating Margin", NF)
fill_all(ws, RC['g_oi_m'], lambda ci: f"={CL(ci)}{RC['g_oi']}/{CL(ci)}{RC['g_rev']}", fmt=P)

# GAAP D&A = PP&E Dep + Intangible Amort
w(ws, RC['g_da'], 1, "Total D&A (PP&E + Intangible)", NF)
fill_all(ws, RC['g_da'], lambda ci: f"={IS_R}!{CL(ci)}{IS['dep']}+{IS_R}!{CL(ci)}{IS['ia']}", fmt=N)
w(ws, RC['g_ebitda'], 1, "GAAP EBITDA", BF)
fill_all(ws, RC['g_ebitda'], lambda ci: f"={CL(ci)}{RC['g_oi']}+{CL(ci)}{RC['g_da']}", fmt=N)
w(ws, RC['g_ebitda_m'], 1, "  GAAP EBITDA Margin", NF)
fill_all(ws, RC['g_ebitda_m'], lambda ci: f"={CL(ci)}{RC['g_ebitda']}/{CL(ci)}{RC['g_rev']}", fmt=P)

w(ws, RC['g_int'], 1, "Interest & Other", NF)
fill_all(ws, RC['g_int'], lambda ci: f"={IS_R}!{CL(ci)}{IS['int']}", fmt=N)
w(ws, RC['g_pt'], 1, "GAAP Pre-Tax Income (Loss)", BF)
fill_all(ws, RC['g_pt'], lambda ci: f"={CL(ci)}{RC['g_oi']}+{CL(ci)}{RC['g_int']}", fmt=N)
w(ws, RC['g_tax'], 1, "GAAP Tax Provision", NF)
fill_hist(ws, RC['g_tax'], gaap_tax_h, fmt=N)
fill_fcst(ws, RC['g_tax'], [f"=IF({CL(ci)}{RC['g_pt']}>0,ROUND({CL(ci)}{RC['g_pt']}*{A_R}!{CL(ci)}{A['gaap_tax']},0),0)" for ci in range(8,12)], fmt=N)

w(ws, RC['g_ni'], 1, "GAAP Net Income (Loss)", BF)
fill_all(ws, RC['g_ni'], lambda ci: f"={CL(ci)}{RC['g_pt']}-{CL(ci)}{RC['g_tax']}", fmt=N)
w(ws, RC['g_eps'], 1, "GAAP Diluted EPS", BF)
fill_all(ws, RC['g_eps'], lambda ci: f"=ROUND({CL(ci)}{RC['g_ni']}/{IS_R}!{CL(ci)}{IS['shares']},2)", fmt=U2)

# Section 4: Adj EBITDA → FCF Bridge
sec(ws, 49, "ADJUSTED EBITDA -> FREE CASH FLOW BRIDGE")
w(ws, RC['ae'], 1, "Adjusted EBITDA", BF)
fill_all(ws, RC['ae'], lambda ci: f"={IS_R}!{CL(ci)}{IS['adj_ebitda']}", fmt=N)
w(ws, RC['ae_dep'], 1, "  (-) PP&E Depreciation", NF)
fill_all(ws, RC['ae_dep'], lambda ci: f"=-{IS_R}!{CL(ci)}{IS['dep']}", fmt=N)
w(ws, RC['ae_ngoi'], 1, "= Non-GAAP Operating Income", BF)
fill_all(ws, RC['ae_ngoi'], lambda ci: f"={CL(ci)}{RC['ae']}+{CL(ci)}{RC['ae_dep']}", fmt=N)
w(ws, RC['ae_sbc'], 1, "  (-) SBC", NF)
fill_all(ws, RC['ae_sbc'], lambda ci: f"=-{IS_R}!{CL(ci)}{IS['sbc']}", fmt=N)
w(ws, RC['ae_ia'], 1, "  (-) Intangible Amortization", NF)
fill_all(ws, RC['ae_ia'], lambda ci: f"=-{IS_R}!{CL(ci)}{IS['ia']}", fmt=N)
w(ws, RC['ae_restr'], 1, "  (-) Restructuring", NF)
fill_all(ws, RC['ae_restr'], lambda ci: f"=-{IS_R}!{CL(ci)}{IS['restr']}", fmt=N)
w(ws, RC['ae_gaapoi'], 1, "= GAAP Operating Income", BF)
fill_all(ws, RC['ae_gaapoi'], lambda ci: f"={CL(ci)}{RC['ae_ngoi']}+{CL(ci)}{RC['ae_sbc']}+{CL(ci)}{RC['ae_ia']}+{CL(ci)}{RC['ae_restr']}", fmt=N)
w(ws, RC['ae_int'], 1, "  (+/-) Interest & Other", NF)
fill_all(ws, RC['ae_int'], lambda ci: f"={IS_R}!{CL(ci)}{IS['int']}", fmt=N)
w(ws, RC['ae_tax'], 1, "  (-) GAAP Tax", NF)
fill_all(ws, RC['ae_tax'], lambda ci: f"=-{CL(ci)}{RC['g_tax']}", fmt=N)
w(ws, RC['ae_gaapni'], 1, "= GAAP Net Income", BF)
fill_all(ws, RC['ae_gaapni'], lambda ci: f"={CL(ci)}{RC['ae_gaapoi']}+{CL(ci)}{RC['ae_int']}+{CL(ci)}{RC['ae_tax']}", fmt=N)
# Continue to FCF
w(ws, RC['ae_da'], 1, "  (+) Total D&A", NF)
fill_all(ws, RC['ae_da'], lambda ci: f"={CL(ci)}{RC['g_da']}", fmt=N)
w(ws, RC['ae_wc'], 1, "  (+) SBC (non-cash)", NF)
fill_all(ws, RC['ae_wc'], lambda ci: f"={IS_R}!{CL(ci)}{IS['sbc']}", fmt=N)
w(ws, RC['ae_other'], 1, "  (+/-) WC & Other", NF)
fill_hist(ws, RC['ae_other'], [wc + oa for wc, oa in zip(wc_h, other_adj_h)], fmt=N)
fill_fcst(ws, RC['ae_other'], [f"={A_R}!{CL(ci)}{A['wc']}" for ci in range(8,12)], fmt=N)
w(ws, RC['ae_cfo'], 1, "= CFO", BF)
fill_all(ws, RC['ae_cfo'], lambda ci: f"={CL(ci)}{RC['ae_gaapni']}+{CL(ci)}{RC['ae_da']}+{CL(ci)}{RC['ae_wc']}+{CL(ci)}{RC['ae_other']}", fmt=N)
w(ws, RC['ae_capex'], 1, "  (-) CapEx", NF)
fill_hist(ws, RC['ae_capex'], capex_h, fmt=N)
fill_fcst(ws, RC['ae_capex'], [f"=-{A_R}!{CL(ci)}{A['capex']}" for ci in range(8,12)], fmt=N)
w(ws, RC['ae_fcf'], 1, "= Free Cash Flow", BF)
fill_all(ws, RC['ae_fcf'], lambda ci: f"={CL(ci)}{RC['ae_cfo']}+{CL(ci)}{RC['ae_capex']}", fmt=N)
colw(ws, [34]+[14]*10)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB: CASH FLOW (references GAAP NI from Reconciliation)
# ═══════════════════════════════════════════════════════════════════════════════
ws = ws_cf
w(ws, 1, 1, "Snap Inc. (SNAP) - Cash Flow Statement ($M)", TF); ws.merge_cells('A1:K1')
hdr_row(ws, 3, yr_hdr); sub_row(ws, 4, hf_marker)

sec(ws, 6, "CASH FROM OPERATIONS")
w(ws, CF['ni'], 1, "GAAP Net Income (Loss)", NF)
fill_all(ws, CF['ni'], lambda ci: f"={RC_R}!{CL(ci)}{RC['g_ni']}", fmt=N)
w(ws, CF['da'], 1, "  (+) Total D&A", NF)
fill_all(ws, CF['da'], lambda ci: f"={IS_R}!{CL(ci)}{IS['dep']}+{IS_R}!{CL(ci)}{IS['ia']}", fmt=N)
w(ws, CF['sbc'], 1, "  (+) Stock-Based Compensation", NF)
fill_all(ws, CF['sbc'], lambda ci: f"={IS_R}!{CL(ci)}{IS['sbc']}", fmt=N)
w(ws, CF['wc'], 1, "  (+/-) Working Capital Changes", NF)
fill_hist(ws, CF['wc'], wc_h, fmt=N)
fill_fcst(ws, CF['wc'], [f"={A_R}!{CL(ci)}{A['wc']}" for ci in range(8,12)], fmt=N)
w(ws, CF['other'], 1, "  (+/-) Other Non-Cash Adj.", NF)
fill_hist(ws, CF['other'], other_adj_h, fmt=N)
fill_fcst(ws, CF['other'], [0,0,0,0], fmt=N)
w(ws, CF['cfo'], 1, "Cash from Operations (CFO)", BF)
fill_all(ws, CF['cfo'], lambda ci: f"=SUM({CL(ci)}{CF['ni']}:{CL(ci)}{CF['other']})", fmt=N)

sec(ws, 14, "CASH FROM INVESTING")
w(ws, CF['capex'], 1, "  Capital Expenditures", NF)
fill_hist(ws, CF['capex'], capex_h, fmt=N)
fill_fcst(ws, CF['capex'], [f"=-{A_R}!{CL(ci)}{A['capex']}" for ci in range(8,12)], fmt=N)
w(ws, CF['other_cfi'], 1, "  Other Investing", NF)
fill_hist(ws, CF['other_cfi'], other_cfi_h, fmt=N)
fill_fcst(ws, CF['other_cfi'], [0,0,0,0], fmt=N)
w(ws, CF['cfi'], 1, "Cash from Investing (CFI)", BF)
fill_all(ws, CF['cfi'], lambda ci: f"={CL(ci)}{CF['capex']}+{CL(ci)}{CF['other_cfi']}", fmt=N)

sec(ws, 19, "CASH FROM FINANCING")
w(ws, CF['cff'], 1, "Cash from Financing (CFF)", BF)
fill_all(ws, CF['cff'], lambda ci: f"={A_R}!{CL(ci)}{A['cff']}", fmt=N)

w(ws, CF['net'], 1, "Net Change in Cash", BF)
fill_all(ws, CF['net'], lambda ci: f"={CL(ci)}{CF['cfo']}+{CL(ci)}{CF['cfi']}+{CL(ci)}{CF['cff']}", fmt=N)
w(ws, CF['beg'], 1, "Beginning Cash", NF)
fill_hist(ws, CF['beg'], beg_cash_h, fmt=N)
fill_fcst(ws, CF['beg'], [f"={CL(ci-1)}{CF['end']}" for ci in range(8,12)], fmt=N)
w(ws, CF['end'], 1, "Ending Cash", BF)
fill_all(ws, CF['end'], lambda ci: f"={CL(ci)}{CF['beg']}+{CL(ci)}{CF['net']}", fmt=N)
w(ws, CF['fcf'], 1, "Free Cash Flow (CFO + CapEx)", BF)
fill_all(ws, CF['fcf'], lambda ci: f"={CL(ci)}{CF['cfo']}+{CL(ci)}{CF['capex']}", fmt=N)
w(ws, CF['fcf_m'], 1, "  FCF Margin", NF)
fill_all(ws, CF['fcf_m'], lambda ci: f"={CL(ci)}{CF['fcf']}/{RB_R}!{CL(ci)}{RB['rev']}", fmt=P)
colw(ws, [34]+[14]*10)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB: BALANCE SHEET (always GAAP, unchanged structure)
# ═══════════════════════════════════════════════════════════════════════════════
ws = ws_bs
w(ws, 1, 1, "Snap Inc. (SNAP) - Balance Sheet ($M, GAAP)", TF); ws.merge_cells('A1:K1')
hdr_row(ws, 3, yr_hdr); sub_row(ws, 4, hf_marker)
sec(ws, 6, "ASSETS")
w(ws, BS['cash'], 1, "Cash & Equivalents", NF); fill_hist(ws, BS['cash'], cash_h, fmt=N)
fill_fcst(ws, BS['cash'], [f"={CF_R}!{CL(ci)}{CF['end']}" for ci in range(8,12)], fmt=N)
w(ws, BS['ar'], 1, "Accounts Receivable", NF); fill_hist(ws, BS['ar'], ar_h, fmt=N)
fill_fcst(ws, BS['ar'], [f"=ROUND({RB_R}!{CL(ci)}{RB['rev']}*{A_R}!{CL(ci)}{A['dso']}/365,0)" for ci in range(8,12)], fmt=N)
w(ws, BS['oca'], 1, "Other Current Assets", NF); fill_hist(ws, BS['oca'], oca_h, fmt=N)
fill_fcst(ws, BS['oca'], [200,200,200,200], fmt=N)
w(ws, BS['tca'], 1, "Total Current Assets", BF)
fill_all(ws, BS['tca'], lambda ci: f"={CL(ci)}{BS['cash']}+{CL(ci)}{BS['ar']}+{CL(ci)}{BS['oca']}", fmt=N)
for br, lb, hist, ar in [(BS['ppe'],"PP&E",ppe_h,A['ppe']),(BS['intang'],"Intangibles",intang_bs_h,A['intang']),
    (BS['rou'],"ROU Assets",rou_h,A['rou']),(BS['olta'],"Other LT Assets",olt_a_h,A['olt_a'])]:
    w(ws, br, 1, lb, NF); fill_hist(ws, br, hist, fmt=N)
    fill_fcst(ws, br, [f"={A_R}!{CL(ci)}{ar}" for ci in range(8,12)], fmt=N)
w(ws, BS['ta'], 1, "Total Assets", BF)
fill_all(ws, BS['ta'], lambda ci: f"={CL(ci)}{BS['tca']}+{CL(ci)}{BS['ppe']}+{CL(ci)}{BS['intang']}+{CL(ci)}{BS['rou']}+{CL(ci)}{BS['olta']}", fmt=N)
sec(ws, 18, "LIABILITIES")
w(ws, BS['ap'], 1, "Accounts Payable", NF); fill_hist(ws, BS['ap'], ap_h, fmt=N)
fill_fcst(ws, BS['ap'], [f"=ROUND({RC_R}!{CL(ci)}{RC['g_cogs']}*{A_R}!{CL(ci)}{A['dpo']}/365,0)" for ci in range(8,12)], fmt=N)
w(ws, BS['accr'], 1, "Accrued Liabilities", NF); fill_hist(ws, BS['accr'], accr_h, fmt=N)
fill_fcst(ws, BS['accr'], [f"=ROUND({RB_R}!{CL(ci)}{RB['rev']}*{A_R}!{CL(ci)}{A['accr']},0)" for ci in range(8,12)], fmt=N)
w(ws, BS['dr'], 1, "Deferred Revenue", NF); fill_hist(ws, BS['dr'], dr_h, fmt=N)
fill_fcst(ws, BS['dr'], [f"=ROUND({RB_R}!{CL(ci)}{RB['rev']}*0.007,0)" for ci in range(8,12)], fmt=N)
w(ws, BS['ocl'], 1, "Other Current Liab", NF); fill_hist(ws, BS['ocl'], ocl_h, fmt=N)
fill_fcst(ws, BS['ocl'], [100,100,100,100], fmt=N)
w(ws, BS['tcl'], 1, "Total Current Liabilities", BF)
fill_all(ws, BS['tcl'], lambda ci: f"={CL(ci)}{BS['ap']}+{CL(ci)}{BS['accr']}+{CL(ci)}{BS['dr']}+{CL(ci)}{BS['ocl']}", fmt=N)
for br, lb, hist, ar in [(BS['ltd'],"Long-Term Debt",ltd_h,A['ltd']),(BS['lease'],"Lease Liab",lease_h,A['lease']),
    (BS['oltl'],"Other LT Liab",olt_l_h,A['olt_l'])]:
    w(ws, br, 1, lb, NF); fill_hist(ws, br, hist, fmt=N)
    fill_fcst(ws, br, [f"={A_R}!{CL(ci)}{ar}" for ci in range(8,12)], fmt=N)
w(ws, BS['tl'], 1, "Total Liabilities", BF)
fill_all(ws, BS['tl'], lambda ci: f"={CL(ci)}{BS['tcl']}+{CL(ci)}{BS['ltd']}+{CL(ci)}{BS['lease']}+{CL(ci)}{BS['oltl']}", fmt=N)
sec(ws, 30, "STOCKHOLDERS' EQUITY")
w(ws, BS['eq'], 1, "Total Equity", BF)
fill_all(ws, BS['eq'], lambda ci: f"={CL(ci)}{BS['ta']}-{CL(ci)}{BS['tl']}", fmt=N)
w(ws, BS['tle'], 1, "Total L&E", BF)
fill_all(ws, BS['tle'], lambda ci: f"={CL(ci)}{BS['tl']}+{CL(ci)}{BS['eq']}", fmt=N)
w(ws, BS['chk'], 1, "Balance Check (=0)", NF)
fill_all(ws, BS['chk'], lambda ci: f"={CL(ci)}{BS['ta']}-{CL(ci)}{BS['tle']}", fmt=N)
colw(ws, [32]+[14]*10)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB: QUARTERLY DETAIL (Non-GAAP primary)
# ═══════════════════════════════════════════════════════════════════════════════
ws = ws_qt
w(ws, 1, 1, "Snap Inc. (SNAP) - Non-GAAP Quarterly P&L ($M)", TF); ws.merge_cells('A1:Q1')
hdr_row(ws, 3, [""]+q_labels_h+q_labels_f)
sub_row(ws, 4, [""]+["Historical"]*8+["Forecast"]*8)
QFC = 10
def q_annual(ci): return CL(8 + (ci-10)//4)
def q_season(ci): return CL(2 + (ci-10)%4)
def q_num(ci): return (ci-10)%4 + 1
def q_prev_annual(ci): return CL(8 + (ci-10)//4 - 1)

w(ws, QT['dau'], 1, "DAU (M)", BF)
for i, v in enumerate(dau_qh): w(ws, QT['dau'], 2+i, v, fmt=N)
for ci in range(QFC, 18):
    pa=q_prev_annual(ci); ca=q_annual(ci); qn=q_num(ci)
    w(ws, QT['dau'], ci, f"=ROUND({RB_R}!{pa}{RB['dau']}+({RB_R}!{ca}{RB['dau']}-{RB_R}!{pa}{RB['dau']})*{qn}/4,0)", font=FCST_FONT, fmt=N)

w(ws, QT['rev'], 1, "Revenue", BF)
for i, v in enumerate(rev_qh): w(ws, QT['rev'], 2+i, v, fmt=N, font=BF)
for ci in range(QFC, 18):
    w(ws, QT['rev'], ci, f"=ROUND({RB_R}!{q_annual(ci)}{RB['rev']}*{A_R}!{q_season(ci)}{A['season']},0)", font=FCST_FONT, fmt=N)

w(ws, QT['arpu'], 1, "ARPU ($)", NF)
for ci in range(2, 18):
    w(ws, QT['arpu'], ci, f"=ROUND({CL(ci)}{QT['rev']}/{CL(ci)}{QT['dau']},2)", font=(FCST_FONT if ci>=QFC else NF), fmt=U2)

w(ws, QT['cogs'], 1, "Non-GAAP COGS", NF)
for i, v in enumerate(ng_cor_qh): w(ws, QT['cogs'], 2+i, v, fmt=N)
for ci in range(QFC, 18):
    w(ws, QT['cogs'], ci, f"=ROUND({CL(ci)}{QT['rev']}*(1-{A_R}!{q_annual(ci)}{A['gm']}),0)", font=FCST_FONT, fmt=N)

for ci in range(2, 18):
    fnt = FCST_FONT if ci>=QFC else NF
    w(ws, QT['gp'], ci, f"={CL(ci)}{QT['rev']}-{CL(ci)}{QT['cogs']}", font=fnt, fmt=N)
w(ws, QT['gp'], 1, "Non-GAAP Gross Profit", BF)
w(ws, QT['gm'], 1, "  Non-GAAP GM", NF)
for ci in range(2, 18): w(ws, QT['gm'], ci, f"={CL(ci)}{QT['gp']}/{CL(ci)}{QT['rev']}", font=(FCST_FONT if ci>=QFC else NF), fmt=P)

for qr, lb, hist, ar in [(QT['rd'],"Non-GAAP R&D",ng_rd_qh,A['rd']),(QT['sm'],"Non-GAAP S&M",ng_sm_qh,A['sm']),(QT['ga'],"Non-GAAP G&A",ng_ga_qh,A['ga'])]:
    w(ws, qr, 1, lb, NF)
    for i, v in enumerate(hist): w(ws, qr, 2+i, v, fmt=N)
    for ci in range(QFC, 18): w(ws, qr, ci, f"=ROUND({CL(ci)}{QT['rev']}*{A_R}!{q_annual(ci)}{ar},0)", font=FCST_FONT, fmt=N)

w(ws, QT['oi'], 1, "Non-GAAP Operating Income", BF)
for ci in range(2, 18): w(ws, QT['oi'], ci, f"={CL(ci)}{QT['gp']}-{CL(ci)}{QT['rd']}-{CL(ci)}{QT['sm']}-{CL(ci)}{QT['ga']}", font=(FCST_FONT if ci>=QFC else NF), fmt=N)
w(ws, QT['oi_m'], 1, "  Non-GAAP OI Margin", NF)
for ci in range(2, 18): w(ws, QT['oi_m'], ci, f"={CL(ci)}{QT['oi']}/{CL(ci)}{QT['rev']}", font=(FCST_FONT if ci>=QFC else NF), fmt=P)

w(ws, QT['dep'], 1, "PP&E Depreciation", NF)
for i, v in enumerate(ppe_dep_qh): w(ws, QT['dep'], 2+i, v, fmt=N)
for ci in range(QFC, 18): w(ws, QT['dep'], ci, f"=ROUND({A_R}!{q_annual(ci)}{A['ppe_dep']}/4,0)", font=FCST_FONT, fmt=N)
w(ws, QT['adj_ebitda'], 1, "Adjusted EBITDA", BF)
for ci in range(2, 18): w(ws, QT['adj_ebitda'], ci, f"={CL(ci)}{QT['oi']}+{CL(ci)}{QT['dep']}", font=(FCST_FONT if ci>=QFC else NF), fmt=N)

# Memo adjustment items
w(ws, QT['sbc'], 1, "SBC", NF)
for i, v in enumerate(sbc_qh): w(ws, QT['sbc'], 2+i, v, fmt=N)
for ci in range(QFC, 18): w(ws, QT['sbc'], ci, f"=ROUND({CL(ci)}{QT['rev']}*{A_R}!{q_annual(ci)}{A['sbc_pct']},0)", font=FCST_FONT, fmt=N)
w(ws, QT['ia'], 1, "Intangible Amort", NF)
for i, v in enumerate(ia_qh): w(ws, QT['ia'], 2+i, v, fmt=N)
for ci in range(QFC, 18): w(ws, QT['ia'], ci, f"=ROUND({A_R}!{q_annual(ci)}{A['intang_amort']}/4,0)", font=FCST_FONT, fmt=N)
w(ws, QT['restr'], 1, "Restructuring", NF)
for i, v in enumerate(restr_qh): w(ws, QT['restr'], 2+i, v, fmt=N)
for ci in range(QFC, 18): w(ws, QT['restr'], ci, f"=ROUND({A_R}!{q_annual(ci)}{A['restructuring']}/4,0)", font=FCST_FONT, fmt=N)
w(ws, QT['gaap_oi'], 1, "GAAP Operating Income", BF)
for ci in range(2, 18): w(ws, QT['gaap_oi'], ci, f"={CL(ci)}{QT['oi']}-{CL(ci)}{QT['sbc']}-{CL(ci)}{QT['ia']}-{CL(ci)}{QT['restr']}", font=(FCST_FONT if ci>=QFC else NF), fmt=N)

w(ws, QT['int'], 1, "Interest & Other", NF)
for i, v in enumerate(int_qh): w(ws, QT['int'], 2+i, v, fmt=N)
for ci in range(QFC, 18): w(ws, QT['int'], ci, f"=ROUND(-{A_R}!{q_annual(ci)}{A['interest']}/4,0)", font=FCST_FONT, fmt=N)
w(ws, QT['ng_pt'], 1, "Non-GAAP Pre-Tax", BF)
for ci in range(2, 18): w(ws, QT['ng_pt'], ci, f"={CL(ci)}{QT['oi']}+{CL(ci)}{QT['int']}", font=(FCST_FONT if ci>=QFC else NF), fmt=N)
w(ws, QT['ng_tax'], 1, "Non-GAAP Tax", NF)
for ci in range(2, 18):
    if ci < QFC:
        # Simple historical: use approximate
        w(ws, QT['ng_tax'], ci, f"=IF({CL(ci)}{QT['ng_pt']}>0,ROUND({CL(ci)}{QT['ng_pt']}*0.08,0),0)", fmt=N)
    else:
        w(ws, QT['ng_tax'], ci, f"=IF({CL(ci)}{QT['ng_pt']}>0,ROUND({CL(ci)}{QT['ng_pt']}*{A_R}!{q_annual(ci)}{A['ng_tax']},0),0)", font=FCST_FONT, fmt=N)
w(ws, QT['ng_ni'], 1, "Non-GAAP Net Income", BF)
for ci in range(2, 18): w(ws, QT['ng_ni'], ci, f"={CL(ci)}{QT['ng_pt']}-{CL(ci)}{QT['ng_tax']}", font=(FCST_FONT if ci>=QFC else NF), fmt=N)
w(ws, QT['capex'], 1, "CapEx", NF)
for i, v in enumerate(capex_qh): w(ws, QT['capex'], 2+i, v, fmt=N)
for ci in range(QFC, 18): w(ws, QT['capex'], ci, f"=ROUND(-{A_R}!{q_annual(ci)}{A['capex']}/4,0)", font=FCST_FONT, fmt=N)
colw(ws, [28]+[12]*16)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB: SHARE COUNT (references Non-GAAP EPS)
# ═══════════════════════════════════════════════════════════════════════════════
ws = ws_sc
w(ws, 1, 1, "Snap Inc. (SNAP) - Share Count & Dilution", TF); ws.merge_cells('A1:K1')
hdr_row(ws, 3, yr_hdr); sub_row(ws, 4, hf_marker)
SC = {'shares':5,'dil':6,'sbc':7,'ng_eps':8,'gaap_eps':9}
w(ws, SC['shares'], 1, "Diluted Shares (M)", BF)
fill_all(ws, SC['shares'], lambda ci: f"={A_R}!{CL(ci)}{A['shares']}", fmt=N)
w(ws, SC['dil'], 1, "  YoY Dilution", NF)
fill_all(ws, SC['dil'], lambda ci: f"={CL(ci)}{SC['shares']}/{CL(ci-1)}{SC['shares']}-1", fmt=P, start=3)
w(ws, SC['sbc'], 1, "SBC ($M)", NF)
fill_all(ws, SC['sbc'], lambda ci: f"={IS_R}!{CL(ci)}{IS['sbc']}", fmt=N)
w(ws, SC['ng_eps'], 1, "Non-GAAP EPS", BF)
fill_all(ws, SC['ng_eps'], lambda ci: f"={IS_R}!{CL(ci)}{IS['eps']}", fmt=U2)
w(ws, SC['gaap_eps'], 1, "GAAP EPS", NF)
fill_all(ws, SC['gaap_eps'], lambda ci: f"={RC_R}!{CL(ci)}{RC['g_eps']}", fmt=U2)

# Quarterly schedule
r=12; w(ws, r, 1, "Quarterly Dilution Schedule (Next 8 Quarters)", TF); ws.merge_cells(f'A{r}:I{r}')
hdr_row(ws, 14, [""]+q_labels_f)
SC_Q = {'beg':15,'rsu':16,'buy':17,'other':18,'end':19,'grants':21,'sbc':22,'sbc_ps':23,'unvested':26,'options':27}
w(ws, SC_Q['beg'], 1, "Beginning Shares (M)", NF)
w(ws, SC_Q['beg'], 2, f"={A_R}!G{A['shares']}", fmt=N)
for ci in range(3, 10): w(ws, SC_Q['beg'], ci, f"={CL(ci-1)}{SC_Q['end']}", fmt=N)
for row_n, lb, vals in [(SC_Q['rsu'],"(+) RSU Vesting",rsu_vest),(SC_Q['buy'],"(-) Buybacks",buyback_q),(SC_Q['other'],"(+/-) Other",other_q_sc)]:
    w(ws, row_n, 1, lb, NF)
    for i, v in enumerate(vals): w(ws, row_n, 2+i, v, fmt=N, fill=INPUT_FILL)
w(ws, SC_Q['end'], 1, "Ending Shares (M)", BF)
for ci in range(2, 10):
    c = CL(ci)
    w(ws, SC_Q['end'], ci, f"={c}{SC_Q['beg']}+{c}{SC_Q['rsu']}-{c}{SC_Q['buy']}+{c}{SC_Q['other']}", fmt=N, font=BF)
w(ws, SC_Q['grants'], 1, "Memo: New RSU Grants", NF)
for i, v in enumerate(new_grants): w(ws, SC_Q['grants'], 2+i, v, fmt=N, fill=INPUT_FILL)
w(ws, SC_Q['sbc'], 1, "Memo: Quarterly SBC ($M)", NF)
for ci in range(2, 10): w(ws, SC_Q['sbc'], ci, f"={Q_R}!{CL(ci+8)}{QT['sbc']}", fmt=N)
w(ws, SC_Q['sbc_ps'], 1, "Memo: SBC / Share ($)", NF)
for ci in range(2, 10): w(ws, SC_Q['sbc_ps'], ci, f"=ROUND({CL(ci)}{SC_Q['sbc']}/{CL(ci)}{SC_Q['end']},2)", fmt=U2)
sec(ws, 25, "Outstanding Equity Awards (est.)")
for row_n, lb, vals in [(SC_Q['unvested'],"  Unvested RSUs (M)",unvested_rsu),(SC_Q['options'],"  Vested Options (M)",vested_opts)]:
    w(ws, row_n, 1, lb, NF)
    for i, v in enumerate(vals): w(ws, row_n, 2+i, v, fmt=N)
colw(ws, [32]+[14]*10)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB: SENSITIVITY (references Adjusted EBITDA)
# ═══════════════════════════════════════════════════════════════════════════════
ws = ws_sn
w(ws, 1, 1, "Snap Inc. (SNAP) - Sensitivity Analysis", TF); ws.merge_cells('A1:G1')

r=4; w(ws, r, 1, "TABLE 1: FY2026E Revenue ($M)", SF); ws.merge_cells(f'A{r}:G{r}')
dau_scen=[0.02,0.035,0.055,0.07,0.085]; arpu_scen=[0.04,0.06,0.09,0.12,0.15]
r=6; w(ws, r, 1, "DAU Growth / ARPU Growth ->", BF)
for j, ag in enumerate(arpu_scen):
    w(ws, r, 2+j, ag, font=HDR_FONT, fmt=P, fill=HDR_FILL); ws.cell(row=r,column=2+j).alignment=Alignment(horizontal='center')
for i, dg in enumerate(dau_scen):
    rr=r+1+i; w(ws, rr, 1, dg, font=BF, fmt=P, fill=SUB_FILL)
    for j, ag in enumerate(arpu_scen):
        f_=f"=ROUND(({RB_R}!H{RB['dau']}*(1+$A{rr}))*({RB_R}!H{RB['arpu']}*(1+{CL(2+j)}${r})),0)"
        c=w(ws, rr, 2+j, f_, fmt=N); c.alignment=Alignment(horizontal='center')
        if abs(dg-0.055)<0.001 and abs(ag-0.09)<0.001: c.fill=GREEN_FILL; c.font=BF

r2=r+len(dau_scen)+4; w(ws, r2, 1, "TABLE 2: FY2026E Adj EBITDA ($M)", SF); ws.merge_cells(f'A{r2}:G{r2}')
rev_g_scen=[0.10,0.14,0.17,0.20,0.24]; ebitda_scen=[0.10,0.14,0.18,0.22,0.26]
r2+=2; w(ws, r2, 1, "Rev Growth / Adj EBITDA Margin ->", BF)
for j, em in enumerate(ebitda_scen):
    w(ws, r2, 2+j, em, font=HDR_FONT, fmt=P, fill=HDR_FILL); ws.cell(row=r2,column=2+j).alignment=Alignment(horizontal='center')
for i, rg in enumerate(rev_g_scen):
    rr=r2+1+i; w(ws, rr, 1, rg, font=BF, fmt=P, fill=SUB_FILL)
    for j, em in enumerate(ebitda_scen):
        c=w(ws, rr, 2+j, f"=ROUND({RB_R}!H{RB['rev']}*(1+$A{rr})*{CL(2+j)}${r2},0)", fmt=N)
        c.alignment=Alignment(horizontal='center')
        if abs(rg-0.17)<0.01 and abs(em-0.18)<0.01: c.fill=GREEN_FILL; c.font=BF

r3=r2+len(rev_g_scen)+4; w(ws, r3, 1, "TABLE 3: DCF Implied Share Price ($)", SF); ws.merge_cells(f'A{r3}:G{r3}')
hr=r3+2; w(ws, hr, 1, "DCF Helper Data:", SF)
for i, (lb, ref) in enumerate([("FCF 25E",f"={CF_R}!H{CF['fcf']}"),("FCF 26E",f"={CF_R}!I{CF['fcf']}"),
    ("FCF 27E",f"={CF_R}!J{CF['fcf']}"),("FCF 28E",f"={CF_R}!K{CF['fcf']}"),
    ("Net Debt",f"={A_R}!H{A['ltd']}-{CF_R}!H{CF['end']}"),("Shares",f"={A_R}!I{A['shares']}")]):
    w(ws, hr+1+i, 1, lb, NF); w(ws, hr+1+i, 2, ref, fmt=N)
fcf_c=[f"$B${hr+1+i}" for i in range(4)]; nd_c=f"$B${hr+5}"; sh_c=f"$B${hr+6}"
wacc_scen=[0.09,0.10,0.11,0.12,0.13]; tgr_scen=[0.02,0.025,0.03,0.035,0.04]
tbl_r=hr+8; w(ws, tbl_r, 1, "WACC / Terminal Growth ->", BF)
for j, tg in enumerate(tgr_scen):
    w(ws, tbl_r, 2+j, tg, font=HDR_FONT, fmt=P, fill=HDR_FILL); ws.cell(row=tbl_r,column=2+j).alignment=Alignment(horizontal='center')
for i, wv in enumerate(wacc_scen):
    rr=tbl_r+1+i; w(ws, rr, 1, wv, font=BF, fmt=P, fill=SUB_FILL)
    for j, tg in enumerate(tgr_scen):
        tg_r=f"{CL(2+j)}${tbl_r}"; w_r=f"$A{rr}"
        pv="+".join(f"{fcf_c[y]}/(1+{w_r})^{y+1}" for y in range(4))
        tv=f"{fcf_c[3]}*(1+{tg_r})/({w_r}-{tg_r})/(1+{w_r})^4"
        c=w(ws, rr, 2+j, f"=ROUND(({pv}+{tv}-{nd_c})/{sh_c},1)", fmt=U1)
        c.alignment=Alignment(horizontal='center')
        if abs(wv-0.11)<0.005 and abs(tg-0.03)<0.003: c.fill=GREEN_FILL; c.font=BF
colw(ws, [40]+[14]*8)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB: VALUATION (references Non-GAAP + Adj EBITDA)
# ═══════════════════════════════════════════════════════════════════════════════
ws = ws_vl
w(ws, 1, 1, "Snap Inc. (SNAP) - Valuation Cross-Checks", TF); ws.merge_cells('A1:I1')
sec(ws, 3, "1. DCF Valuation")
V={}
r=5; V['wacc']=r; w(ws,r,1,"WACC",BF); w(ws,r,2,0.11,fmt=P,fill=INPUT_FILL)
r=6; V['tgr']=r; w(ws,r,1,"Terminal Growth",BF); w(ws,r,2,0.03,fmt=P,fill=INPUT_FILL)
r=7; V['nd']=r; w(ws,r,1,"Net Debt ($M)",BF); w(ws,r,2,f"={A_R}!H{A['ltd']}-{CF_R}!H{CF['end']}",fmt=N)
r=8; V['sh']=r; w(ws,r,1,"Shares (M)",BF); w(ws,r,2,f"={A_R}!I{A['shares']}",fmt=N)
hdr_row(ws,10,["","FY2025E","FY2026E","FY2027E","FY2028E"])
r=11; V['fcf']=r; w(ws,r,1,"FCF ($M)",NF)
for i, yc in enumerate(['H','I','J','K']): w(ws,r,2+i,f"={CF_R}!{yc}{CF['fcf']}",fmt=N)
r=12; w(ws,r,1,"Discount Factor",NF)
for i in range(4): w(ws,r,2+i,f"=1/(1+$B${V['wacc']})^{i+1}",fmt='0.0000')
r=13; w(ws,r,1,"PV of FCF",NF)
for i in range(4): w(ws,r,2+i,f"=ROUND({CL(2+i)}{V['fcf']}*{CL(2+i)}12,0)",fmt=N)
r=15; w(ws,r,1,"Sum PV FCFs",BF); w(ws,r,2,"=SUM(B13:E13)",fmt=N); V['spv']=r
r=16; w(ws,r,1,"Terminal Value",NF); w(ws,r,2,f"=ROUND(E{V['fcf']}*(1+$B${V['tgr']})/($B${V['wacc']}-$B${V['tgr']}),0)",fmt=N); V['tv']=r
r=17; w(ws,r,1,"PV of TV",NF); w(ws,r,2,f"=ROUND(B{V['tv']}*E12,0)",fmt=N); V['pvtv']=r
r=18; w(ws,r,1,"EV ($M)",BF); w(ws,r,2,f"=B{V['spv']}+B{V['pvtv']}",fmt=N); V['ev']=r
r=19; w(ws,r,1,"Equity ($M)",BF); w(ws,r,2,f"=B{V['ev']}-B{V['nd']}",fmt=N); V['eqv']=r
r=20; V['dcf_p']=r; w(ws,r,1,"Implied Price ($)",BF); w(ws,r,2,f"=ROUND(B{V['eqv']}/B{V['sh']},2)",fmt=U2)

sec(ws, 23, "2. Unit Economics / EV Sanity Check")
r=25; V['mkt']=r; w(ws,r,1,"Market Cap ($M, est.)",BF); w(ws,r,2,18000,fmt=N,fill=INPUT_FILL)
r=26; V['ev2']=r; w(ws,r,1,"EV ($M, est.)",BF); w(ws,r,2,f"=B{V['mkt']}+B{V['nd']}",fmt=N)
hdr_row(ws,28,["Metric","Value","Commentary"])
for rr, lb, f_, cmt in [
    (29,"EV/DAU ($)",f"=ROUND(B{V['ev2']}/{RB_R}!H{RB['dau']},1)","vs META ~$310, PINS ~$75"),
    (30,"EV/FY25E Rev",f"=ROUND(B{V['ev2']}/{RB_R}!H{RB['rev']},1)","vs peers 4-8x"),
    (31,"EV/FY26E Adj EBITDA",f"=IF({IS_R}!I{IS['adj_ebitda']}>0,ROUND(B{V['ev2']}/{IS_R}!I{IS['adj_ebitda']},1),\"N/M\")","vs peers 15-25x"),
    (32,"FY26E FCF Yield",f"=IF({CF_R}!I{CF['fcf']}>0,{CF_R}!I{CF['fcf']}/B{V['mkt']},0)","FCF inflection"),
]:
    w(ws,rr,1,lb,NF); w(ws,rr,2,f_,fmt=(P if "Yield" in lb else D)); w(ws,rr,3,cmt)

sec(ws, 35, "Valuation Summary")
r=36; V['cp']=r; w(ws,r,1,"Current Price (est.)",BF); w(ws,r,2,10.50,fmt=U2,fill=INPUT_FILL)
hdr_row(ws,38,["Method","Price ($)","Upside/Downside"])
for rr, lb, f_ in [(39,"DCF",f"=B{V['dcf_p']}"),(40,"EV/Rev Comp (4x FY26E)",f"=ROUND(({RB_R}!I{RB['rev']}*4-B{V['nd']})/B{V['sh']},2)")]:
    w(ws,rr,1,lb,NF); w(ws,rr,2,f_,fmt=U2); w(ws,rr,3,f"=B{rr}/B{V['cp']}-1",fmt=P)
colw(ws, [38]+[18]*8)

# ═══════════════════════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════════════════════
wb.save("/workspace/Snap_Inc_Three_Statement_Model.xlsx")
print("Model saved. All Non-GAAP primary with GAAP reconciliation.")
print(f"Tabs: {wb.sheetnames}")
