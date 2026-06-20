"""Recalculate the workbook with the `formulas` engine and report key outputs
and the integrity checks (balance, cash tie-out, segment NI/EPS tie-out)."""
import formulas
import openpyxl

XL = "SpaceX_Cursor_Pro_Forma_Model.xlsx"
xl = formulas.ExcelModel().loads(XL).finish()
sol = xl.calculate()
vals = {k.upper(): v for k, v in sol.items()}
wb = openpyxl.load_workbook(XL)

YEARS = [2025, 2026, 2027, 2028, 2029, 2030]
C = {2025: "C", 2026: "D", 2027: "E", 2028: "F", 2029: "G", 2030: "H"}
BS = {2025: "C", "Open": "F", 2026: "G", 2027: "H", 2028: "I", 2029: "J", 2030: "K"}


def cell(sheet, a1):
    v = vals.get(f"'[{XL.upper()}]{sheet.upper()}'!{a1.upper()}")
    try:
        return float(v.value[0, 0])
    except Exception:
        try:
            return v.value
        except Exception:
            return v


def rownum(sheet, label):
    s = wb[sheet]
    for r in range(1, s.max_row + 1):
        if s.cell(row=r, column=1).value == label:
            return r


def fmt(x):
    try:
        return f"{x:>11,.0f}"
    except Exception:
        return f"{str(x):>11}"


def fmt2(x):
    try:
        return f"{x:>11,.2f}"
    except Exception:
        return f"{str(x):>11}"


def show(sheet, label, years=YEARS, money=True, key=None):
    r = rownum(sheet, label)
    f = fmt if money else fmt2
    print(f"{(key or label)[:26]:<26} " + "".join(f(cell(sheet, f"{C[y]}{r}")) for y in years))


print("\n=== INTEGRITY CHECKS ===")
print("Year                       " + "".join(f"{y:>11}" for y in YEARS))
show("Checks", "Balance sheet balances (A − L&E)", key="BS balance")
show("Checks", "Cash flow ties to balance sheet cash", key="CF tie")
show("Checks", "Segment net income sums to consolidated NI", key="Seg NI tie")
show("Checks", "Segment EPS contributions sum to diluted EPS", key="Seg EPS tie", money=False)

print("\n=== REVENUE BY SEGMENT (US$M) ===")
print("Year                       " + "".join(f"{y:>11}" for y in YEARS))
show("Income Statement", "  Space")
show("Income Statement", "  Starlink")
show("Income Statement", "  xAI-Cursor")
show("Income Statement", "Total revenue")

print("\n=== EBITDA BY SEGMENT (US$M) ===")
print("Year                       " + "".join(f"{y:>11}" for y in YEARS))
# "  Space"/"  Starlink" appear in revenue, EBITDA and EPS sections; EBITDA is index 1
s = wb["Income Statement"]
rows = [r for r in range(1, s.max_row + 1) if s.cell(row=r, column=1).value == "  Space"]
print("Space EBITDA               " + "".join(fmt(cell("Income Statement", f"{C[y]}{rows[1]}")) for y in YEARS))
rows = [r for r in range(1, s.max_row + 1) if s.cell(row=r, column=1).value == "  Starlink"]
print("Starlink EBITDA            " + "".join(fmt(cell("Income Statement", f"{C[y]}{rows[1]}")) for y in YEARS))
show("Income Statement", "Total EBITDA")

print("\n=== P&L (US$M) ===")
print("Year                       " + "".join(f"{y:>11}" for y in YEARS))
show("Income Statement", "EBIT (operating income)")
show("Income Statement", "Pre-tax income (loss)")
show("Income Statement", "Net income (loss)")
show("Income Statement", "Diluted EPS ($)", money=False)

print("\n=== PP&E FORECAST BY SEGMENT (US$M) ===")
print("Year                       " + "".join(f"{y:>11}" for y in YEARS))
for seg, lbl in [("Space", "Space"), ("Starlink", "Starlink"), ("xAI-Cursor", "xAI-Cursor")]:
    sc = wb["Schedules"]
    # locate the segment block header then its capex/dep/ending rows
    hdr = next(r for r in range(1, sc.max_row + 1) if sc.cell(row=r, column=1).value == f"  {seg}")
    cap = next(r for r in range(hdr, hdr + 6) if "Capital expenditures" in str(sc.cell(row=r, column=1).value))
    dep = next(r for r in range(hdr, hdr + 6) if "Depreciation" in str(sc.cell(row=r, column=1).value))
    end = next(r for r in range(hdr, hdr + 6) if "Ending net PP&E" in str(sc.cell(row=r, column=1).value))
    print(f"{lbl} capex            " [:26] + "".join(fmt(cell("Schedules", f"{C[y]}{cap}")) for y in YEARS))
    print(f"{lbl} depreciation     " [:26] + "".join(fmt(cell("Schedules", f"{C[y]}{dep}")) for y in YEARS))
    print(f"{lbl} ending PP&E      " [:26] + "".join(fmt(cell("Schedules", f"{C[y]}{end}")) for y in YEARS))
show("Schedules", "     (+) Total capital expenditures", key="TOTAL capex")
show("Schedules", "     (−) Total depreciation", key="TOTAL depreciation")
show("Schedules", "     Total ending net PP&E", key="TOTAL ending PP&E")

print("\n=== EPS CONTRIBUTION BY SEGMENT ($) ===")
print("Year                       " + "".join(f"{y:>11}" for y in YEARS))
# EPS contribution rows are in the dedicated section; Space/Starlink/xAI-Cursor appear again
rows = [r for r in range(1, s.max_row + 1) if s.cell(row=r, column=1).value == "  Space"]
print("Space EPS contribution     " + "".join(fmt2(cell("Income Statement", f"{C[y]}{rows[2]}")) for y in YEARS))
rows = [r for r in range(1, s.max_row + 1) if s.cell(row=r, column=1).value == "  Starlink"]
print("Starlink EPS contribution  " + "".join(fmt2(cell("Income Statement", f"{C[y]}{rows[2]}")) for y in YEARS))
# "  xAI-Cursor" appears in the revenue section and the EPS section (EBITDA label differs)
rows = [r for r in range(1, s.max_row + 1) if s.cell(row=r, column=1).value == "  xAI-Cursor"]
print("xAI-Cursor EPS contribution" + "".join(fmt2(cell("Income Statement", f"{C[y]}{rows[-1]}")) for y in YEARS))
show("Income Statement", "  Total (= diluted EPS)", money=False, key="Total EPS (check)")
