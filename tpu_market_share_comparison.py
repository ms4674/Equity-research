#!/usr/bin/env python3
"""
Broadcom-Google TPU vs NVIDIA vs AMD — AI Accelerator Market Share in Training & Inference

Generates an Excel workbook with:
  1. Overall AI accelerator market share (2023-2028E)
  2. Training-specific market share breakdown
  3. Inference-specific market share breakdown
  4. Revenue projections by vendor (Training vs Inference)
  5. ASIC vs GPU growth rate comparison
  6. Chip-by-chip 2026 spec comparison
  7. TCO comparison for inference workloads
  8. Broadcom custom ASIC revenue breakdown
  9. Charts: market share trajectories, training vs inference splits, CAGR comparison

Sources cited in the "Sources" sheet.
"""

import math
from dataclasses import dataclass, field
from typing import List, Dict, Tuple

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import (
    BarChart, LineChart, PieChart, Reference, Series, BarChart3D,
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Color palette
# ---------------------------------------------------------------------------
NVIDIA_GREEN = "76B900"
AMD_RED = "ED1C24"
GOOGLE_BLUE = "4285F4"
BROADCOM_RED = "CC0000"
ASIC_PURPLE = "7B2D8E"
OTHER_GRAY = "888888"

HEADER_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1B2A4A")
SUBTITLE_FONT = Font(name="Calibri", size=11, bold=True, color="1B2A4A")
DATA_FONT = Font(name="Calibri", size=10)
SOURCE_FONT = Font(name="Calibri", size=9, italic=True, color="666666")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
PCT_FMT = '0.0%'
USD_FMT = '$#,##0.0'
USD_B_FMT = '$#,##0.0"B"'
NUM_FMT = '#,##0'


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_area(ws, start_row, end_row, max_col, fmt=None):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fmt and c > 1:
                cell.number_format = fmt


def add_title(ws, row, title, merge_end=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_end)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")


def add_subtitle(ws, row, text, merge_end=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_end)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SUBTITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")


def add_source_note(ws, row, text, merge_end=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_end)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SOURCE_FONT


# ---------------------------------------------------------------------------
# Data
# ---------------------------------------------------------------------------

YEARS = [2023, 2024, 2025, 2026, 2027, 2028]

# Overall AI Accelerator Market Size ($B) — Total addressable market
TOTAL_MARKET_SIZE = [55, 160, 250, 350, 450, 604]

# --- OVERALL MARKET SHARE (% of total AI accelerator revenue) ---
OVERALL_SHARE = {
    "NVIDIA":                  [0.92, 0.90, 0.86, 0.80, 0.72, 0.65],
    "AMD":                     [0.03, 0.05, 0.07, 0.10, 0.12, 0.14],
    "Google/Broadcom TPU":     [0.02, 0.03, 0.04, 0.06, 0.09, 0.12],
    "Other Custom ASIC":       [0.02, 0.01, 0.02, 0.03, 0.05, 0.06],
    "Other (Intel, etc.)":     [0.01, 0.01, 0.01, 0.01, 0.02, 0.03],
}

# --- TRAINING MARKET SHARE (% of training compute spend) ---
TRAINING_SHARE = {
    "NVIDIA":                  [0.95, 0.93, 0.91, 0.88, 0.85, 0.82],
    "AMD":                     [0.02, 0.04, 0.05, 0.07, 0.08, 0.10],
    "Google/Broadcom TPU":     [0.02, 0.02, 0.02, 0.03, 0.04, 0.04],
    "Other Custom ASIC":       [0.005, 0.005, 0.01, 0.01, 0.02, 0.025],
    "Other (Intel, etc.)":     [0.005, 0.005, 0.01, 0.01, 0.01, 0.015],
}

# Training as % of total AI compute
TRAINING_PCT_OF_TOTAL = [0.67, 0.55, 0.42, 0.33, 0.28, 0.25]

# --- INFERENCE MARKET SHARE (% of inference compute spend) ---
INFERENCE_SHARE = {
    "NVIDIA":                  [0.88, 0.85, 0.78, 0.68, 0.50, 0.30],
    "AMD":                     [0.04, 0.06, 0.08, 0.12, 0.15, 0.18],
    "Google/Broadcom TPU":     [0.04, 0.05, 0.08, 0.12, 0.18, 0.25],
    "Other Custom ASIC":       [0.03, 0.03, 0.04, 0.06, 0.13, 0.22],
    "Other (Intel, etc.)":     [0.01, 0.01, 0.02, 0.02, 0.04, 0.05],
}

# Inference as % of total AI compute
INFERENCE_PCT_OF_TOTAL = [0.33, 0.45, 0.58, 0.67, 0.72, 0.75]

# --- REVENUE PROJECTIONS ($B) by vendor ---
def compute_revenue(share_dict, pct_of_total):
    rev = {}
    for vendor, shares in share_dict.items():
        rev[vendor] = []
        for i, yr in enumerate(YEARS):
            segment_size = TOTAL_MARKET_SIZE[i] * pct_of_total[i]
            rev[vendor].append(round(shares[i] * segment_size, 1))
    return rev


# ---------------------------------------------------------------------------
# UNIT SHIPMENT DATA (number of chips shipped/deployed annually, in thousands)
# ---------------------------------------------------------------------------
UNIT_YEARS = [2022, 2023, 2024, 2025, 2026]

# Nvidia: Epoch AI + Jensen Huang disclosure (4M Hopper + 3M Blackwell thru Oct 2025)
# 2022: ~1.0M (A100-era tail + early H100); 2023: ~3.76M (Omdia/HPCwire); 2024: ~5.0M;
# 2025: ~7M (includes Blackwell ramp); 2026E: ~10-12M (Rubin + Blackwell continuing)
NVIDIA_UNITS = {
    "A100 / A800":          [600,   400,    50,     0,      0],
    "H100 / H200 / H800":  [50,   2200,  3500,  1500,    500],
    "H20 (China)":          [0,       0,  1000,   500,      0],
    "B200 / B300 (Blackwell)": [0,    0,   200,  4500,   3000],
    "Vera Rubin":           [0,       0,     0,     0,   5000],
    "NVIDIA Total":         [650,  2600,  4750,  6500,   8500],
}

# Google TPU: Epoch AI methodology (Broadcom revenue / est. cost per chip)
# Zylos Research: 1.6M Trillium v6 in 2025; v7E ~500K in 2026
# Anthropic: 1M+ TPU deal; Google $185B capex in 2026
GOOGLE_TPU_UNITS = {
    "TPU v4":               [200,   400,   200,     0,      0],
    "TPU v5e / v5p":        [0,     300,   800,   400,    100],
    "TPU v6 (Trillium)":    [0,       0,   200,  1600,   1000],
    "TPU v7 (Ironwood)":    [0,       0,     0,   500,   2500],
    "Google TPU Total":     [200,   700,  1200,  2500,   3600],
}

# --- GOOGLE TPU: INTERNAL vs CLOUD vs EXTERNAL DEPLOYMENT ---
# Source: Semi Fundamental (March 2026) channel checks + industry estimates
# Internal = Google's own products (Search, YouTube, Ads, Gemini, etc.)
# Cloud (GCP) = External customers renting TPUs via Google Cloud (incl. Anthropic, Meta rental)
# External Direct = Customers purchasing TPU racks for their own data centers (Anthropic owned, Apple)
TPU_DEPLOY_YEARS = [2022, 2023, 2024, 2025, 2026]

TPU_DEPLOYMENT_SPLIT = {
    "Internal Google (Search, YouTube, Ads, Gemini)": {
        "units_k": [190, 630, 960, 2000, 1800],
        "pct":     [0.95, 0.90, 0.80, 0.80, 0.50],
    },
    "Google Cloud (GCP) — rental to external customers": {
        "units_k": [10, 60, 180, 375, 1080],
        "pct":     [0.05, 0.085, 0.15, 0.15, 0.30],
    },
    "External Direct — customer-owned data centers": {
        "units_k": [0, 10, 60, 125, 720],
        "pct":     [0.00, 0.015, 0.05, 0.05, 0.20],
    },
}

TPU_CUSTOMER_DETAILS = [
    ("Anthropic", "~1M TPUs committed (mix of owned racks + GCP rental)", "2025-2027",
     "Largest external TPU customer; 3.5 GW deal for 2027; $21B+ in Broadcom orders"),
    ("Meta", "Phase 1: TPU rental via GCP for Llama testing (2026)", "2026-2027",
     "Phase 2: potential TPU purchase for Meta DCs in 2027 if tests succeed"),
    ("Apple", "~100K TPU v5p in 2024; ~200K in 2025", "2024-2025",
     "One of largest third-party TPU buyers; reducing NVIDIA reliance"),
    ("OpenAI", "Early-stage GCP rental for testing", "2025-2026",
     "Appears more as leverage vs NVIDIA pricing than core infra strategy"),
    ("Google Internal", "Search, YouTube, Ads, Gmail, Gemini models", "2016-present",
     "Original TPU use case; still ~50% of total TPU capacity in 2026"),
]

# AMD: MI300X ~400-500K in 2024; MI350 ramp mid-2025; MI400 launch 2026
# AMD CoWoS: 11% of TSMC 2026 capacity (~105K wafers)
AMD_UNITS = {
    "MI250 / MI250X":       [30,     50,    20,     0,      0],
    "MI300X / MI300A":      [0,       0,   400,   200,     50],
    "MI350X":               [0,       0,     0,   500,    300],
    "MI400 (MI455X, etc.)": [0,       0,     0,     0,    800],
    "AMD Total":            [30,     50,   420,   700,   1150],
}

# Amazon AWS: Trainium/Inferentia deployed
# TechCrunch: 1-1.4M Trainium total by end 2025; Project Rainier 500K Trainium2
# Trainium3 launched Dec 2025; 1M-chip UltraClusters
AWS_UNITS = {
    "Inferentia 1/2":       [100,   200,   300,   200,    100],
    "Trainium 1":           [0,      50,   200,   100,      0],
    "Trainium 2":           [0,       0,   300,  1500,   1000],
    "Trainium 3":           [0,       0,     0,   200,   1500],
    "AWS Total":            [100,   250,   800,  2000,   2600],
}

# Summary table: total units by vendor per year (thousands)
UNIT_SUMMARY = {
    "NVIDIA GPUs":          [650,  2600,  4750,  6500,   8500],
    "Google/Broadcom TPU":  [200,   700,  1200,  2500,   3600],
    "AMD GPUs":             [30,     50,   420,   700,   1150],
    "AWS Trainium/Inferentia": [100, 250,  800,  2000,   2600],
    "Total (4 vendors)":    [980,  3600,  7170, 11700,  15850],
}

# Unit market share (% of total units across these 4 vendors)
UNIT_SHARE = {}
for vendor, units in UNIT_SUMMARY.items():
    if vendor != "Total (4 vendors)":
        totals = UNIT_SUMMARY["Total (4 vendors)"]
        UNIT_SHARE[vendor] = [round(units[i] / totals[i], 3) for i in range(len(UNIT_YEARS))]

# Revenue per unit (ASP proxy, $K per chip) — derived from revenue / units
REVENUE_PER_UNIT = {
    "NVIDIA GPUs":          ["~$23K", "~$25K", "~$27K", "~$30K", "~$35K"],
    "Google/Broadcom TPU":  ["~$6K", "~$7K", "~$8K", "~$10K", "~$12K"],
    "AMD GPUs":             ["~$10K", "~$12K", "~$15K", "~$18K", "~$20K"],
    "AWS Trainium/Inferentia": ["~$4K", "~$5K", "~$6K", "~$7K", "~$8K"],
}

# H100-equivalent compute capacity (normalizing to H100 = 1.0x)
H100E_MULTIPLIERS = {
    "NVIDIA (weighted avg)":  [0.5, 0.9, 1.2, 1.8, 3.5],
    "Google TPU (weighted avg)": [0.6, 0.7, 1.0, 1.5, 2.3],
    "AMD (weighted avg)":     [0.4, 0.5, 0.8, 1.2, 2.0],
    "AWS (weighted avg)":     [0.2, 0.3, 0.5, 0.8, 1.3],
}

# --- ASIC vs GPU CAGR ---
CAGR_DATA = [
    ("Custom ASICs (Hyperscaler)", "44.6%", "$18B", "$165B", "Inference-optimized"),
    ("General-purpose GPUs", "16.1%", "$130B", "$290B", "Training + flexible inference"),
    ("AMD Accelerators", "~18%", "$12B", "$55B", "Cost-sensitive training/cloud"),
    ("Total AI Accelerator Market", "~16%", "$160B", "$604B", "All AI compute"),
]

# --- 2026 CHIP SPEC COMPARISON ---
CHIP_SPECS = [
    # (Name, Vendor, Process, FP8 PFLOPS, FP4 PFLOPS, Memory GB, Mem BW TB/s, TDP W, Interconnect, Primary Use)
    ("Vera Rubin GPU", "NVIDIA", "3nm", "~25", "50", 288, ">12", "~1000", "NVLink 6 (3.6 TB/s)", "Training + Inference"),
    ("B200 Blackwell", "NVIDIA", "4nm", "4.5", "9", 192, "8", "1000", "NVLink 5 (1.8 TB/s)", "Training + Inference"),
    ("TPU v7 Ironwood", "Google/Broadcom", "3nm", "4.6", "N/D", 192, ">7.2", "~500", "Optical Mesh (4.8 Tbps)", "Inference-optimized"),
    ("TPU v6 Trillium", "Google/Broadcom", "5nm", "~0.9", "N/D", 32, "~1.6", "~300", "ICI (proprietary)", "Training + Inference"),
    ("MI455X", "AMD", "2nm", "~10", "~20", 432, "19.6", "~700", "Infinity Fabric", "Training + Inference"),
    ("MI350X", "AMD", "3nm", "~5", "~10", 288, "8", "~600", "Infinity Fabric", "Training + Inference"),
    ("Maia 200", "Microsoft/Broadcom", "3nm", "~5", ">10", 216, "~8", "750", "Azure Custom Fabric", "Inference (GPT-optimized)"),
    ("Trainium 3", "AWS", "3nm", "2.52", "N/D", 144, "~5", "~600", "EFA (3.2 Tbps)", "Training + Inference"),
]

# --- TCO COMPARISON: 1,000-chip inference cluster, 3-year ---
TCO_DATA = [
    # (Cost Factor, NVIDIA H100/B200, Google TPU v6/v7, AMD MI350/MI400, Winner)
    ("Hardware (CapEx)", "$100M", "$52M", "$75M", "TPU (-48%)"),
    ("Electricity (3yr)", "$47M", "$16M", "$32M", "TPU (-66%)"),
    ("Cooling infrastructure", "$12M", "$4M", "$8M", "TPU (-67%)"),
    ("Software licenses", "$0 (CUDA free)", "$0 (JAX free)", "$0 (ROCm free)", "Tie"),
    ("Support & maintenance", "$8M", "$3M", "$6M", "TPU (-63%)"),
    ("Network infrastructure", "$6M", "$2M", "$5M", "TPU (-67%)"),
    ("Real estate (rack space)", "$4M", "$1.5M", "$3M", "TPU (-63%)"),
    ("TOTAL 3-YEAR TCO", "$177M", "$78.5M", "$129M", "TPU (-56%)"),
]

# --- BROADCOM AI REVENUE (ACTUAL REPORTED + ESTIMATES) ---
# Broadcom fiscal years end in ~October/November; approx align to calendar year
BROADCOM_YEARS = ["FY2023", "FY2024", "FY2025", "FY2026E", "FY2027E"]
BROADCOM_DATA = {
    "Google TPU (est.)":             [3.8,  6.0, 10.0, 14.0, 21.0],
    "Other XPU Customers (est.)":    [0.2,  1.3,  3.5, 11.0, 30.0],
    "AI Networking":                 [1.0,  5.0,  6.5, 13.0, 35.0],
    "Total AI Revenue":              [5.0, 12.3, 20.0, 38.0, 86.0],
}

# --- BROADCOM QUARTERLY ACTUALS + RUN RATE DERIVATION ---
# Used to derive current TPU revenue from reported AI revenue guidance
BROADCOM_QUARTERLY = {
    "headers": ["Quarter", "AI Revenue ($B)", "YoY Growth", "Networking Mix",
                "XPU Revenue (est.)", "Google TPU (est.)", "TPU Run Rate (ann.)"],
    "rows": [
        ("Q1 FY2025", 4.1, "77%",  "~30%", "~$2.9B", "~$2.1B", "~$8.4B"),
        ("Q2 FY2025", 4.4, "44%",  "~30%", "~$3.1B", "~$2.2B", "~$8.8B"),
        ("Q3 FY2025", 5.2, "63%",  "~30%", "~$3.6B", "~$2.6B", "~$10.4B"),
        ("Q4 FY2025", 6.5, "74%",  "~33%", "~$4.4B", "~$3.0B", "~$12.0B"),
        ("Q1 FY2026 (actual)", 8.4, "106%", "~33%", "~$5.6B", "~$3.4B", "~$13.6B"),
        ("Q2 FY2026 (guide)", 10.7, "140%", "~40%", "~$6.4B", "~$3.5B", "~$14.0B"),
    ],
    "notes": [
        "AI Revenue = Broadcom's reported 'AI semiconductor revenue' per earnings calls",
        "Networking Mix = % of AI revenue from AI networking (Tomahawk, DSP, etc.) per Hock Tan commentary",
        "XPU Revenue = AI Revenue × (1 - Networking Mix); includes all custom accelerators",
        "Google TPU est. = Analyst consensus ~50% of XPU in FY2025, declining to ~40-45% as Meta/Anthropic/OpenAI ramp",
        "TPU Run Rate = Google TPU quarterly × 4; represents annualized Google TPU revenue from Broadcom",
        "Broadcom has 6 XPU customers: Google, Meta, Anthropic, OpenAI + 2 undisclosed (as of Q1 FY2026)",
        "Hock Tan (Q1 FY2026 call): 'Line of sight to AI revenue from chips in excess of $100B in 2027'",
        "$100B target is for chips only (XPUs, switches, DSPs) — confirmed by Blayne Curtis/Jefferies Q&A",
    ],
}

# --- REAL-WORLD MIGRATION CASE STUDIES ---
CASE_STUDIES = [
    ("Midjourney", "NVIDIA → TPU v6e", "65% cost reduction", "$2.1M → $700K/mo", "6 weeks, 3 engineers"),
    ("Anthropic", "Multi-provider → TPU v7", "1M+ TPU chips by 2027", "Tens of $B commitment", "Largest TPU deal ever"),
    ("Meta", "NVIDIA GPU → TPU talks", "Multibillion-dollar TPU talks", "600K+ chip infra by 2026", "Advanced negotiations"),
    ("Character.AI", "NVIDIA → TPU", "3.8× cost improvement", "Full inference stack", "8 weeks, 2 engineers"),
    ("Perplexity AI", "Mixed → TPU v5e/v6", "Entire inference on TPU", "Default TPU stack", "4 weeks, 2 engineers"),
    ("Stability AI", "NVIDIA → TPU v6", "40% of image gen to TPU", "Q3 2025 migration", "Ongoing"),
    ("OpenAI", "NVIDIA → Broadcom ASIC", "10GW capacity by 2029", "~$10B investment", "Custom chip in design"),
]

# --- SOURCES ---
SOURCES = [
    ("Broadcom Q1 FY2026 Earnings", "AI semi revenue $8.4B (+106% YoY); total revenue $19.3B (+29%); 6 XPU customers", "March 4, 2026"),
    ("Broadcom Q1 FY2026 Guidance", "Q2 AI semi revenue $10.7B (+140% YoY); total revenue $22B; networking 40% of AI rev", "March 4, 2026"),
    ("Broadcom CEO Hock Tan", "'Line of sight to AI revenue from chips >$100B in 2027'; chips = XPUs + switches + DSPs", "Q1 FY2026 Call"),
    ("Broadcom Q4 FY2025 Earnings", "FY2025 total AI revenue ~$20B (+65% YoY); AI semi revenue $6.5B in Q4 (+74%)", "December 2025"),
    ("Reuters", "Broadcom sees >$100B AI chip sales by 2027; Google TPU orders $21B in two quarters", "March 2026"),
    ("JPMorgan (Harlan Sur)", "Expects Broadcom AI revenue >$9B/quarter; projects >$65B total AI rev FY2026", "March 2026"),
    ("Bloomberg Intelligence", "AI Accelerator Market $604B by 2033; ASIC 44.6% CAGR vs GPU 16.1% CAGR", "January 2026"),
    ("New Street Research", "Inference Compute Share Analysis 2024-2028; NVIDIA inference share 90%→20-30% by 2028", "December 2025"),
    ("SemiAnalysis", "NVIDIA Market Share and Competitive Landscape; 90%+ current market share", "Q4 2025"),
    ("Morgan Stanley", "AI Compute Economics: Training vs Inference; TPU v6 to generate >$150B lifetime rev", "November 2025"),
    ("Goldman Sachs", "Hyperscaler AI Capex $660-690B in 2026; TPU 35% inference share by Q4 2026", "February 2026"),
    ("Google Cloud", "TPU v7 Ironwood specs: 4,614 TFLOPS, 192GB HBM3e, 9,216-chip pods", "April 2025"),
    ("NVIDIA GTC 2026", "Vera Rubin: 336B transistors, 50 PFLOPS FP4, 288GB HBM4, NVLink 6", "March 2026"),
    ("The Information", "Midjourney 65% cost reduction with TPU migration; Meta TPU talks", "September 2025"),
    ("Anthropic Blog", "1M+ Google TPU v7 chips for Claude inference; tens of $B commitment", "2025"),
    ("Zylos Research", "AI Chip Hardware Acceleration Trends 2026; ASIC share 15%→40% in inference", "February 2026"),
    ("Introl Blog", "Custom Silicon Inflection 2026; hyperscaler ASIC roadmap comparison", "February 2026"),
    ("AMD Newsroom", "MI400 series (2nm CDNA 5); MI355X 30% faster inference than B200", "CES 2026"),
    ("Silicon Analysts", "NVIDIA GPU Market Share 2024-2026: 87% peak, declining trajectory", "2026"),
    ("Epoch AI", "Inference projected at 75-80% of AI compute by 2030", "2025"),
    ("Epoch AI - AI Chip Sales", "Open database of AI chip shipments by vendor; Nvidia, Google, AMD, Amazon tracked", "March 2026"),
    ("Jensen Huang / Nvidia", "Direct disclosure: 4M Hopper + 3M Blackwell GPUs shipped through October 2025 (excl. China)", "Late 2025"),
    ("HPCwire / Omdia", "Nvidia shipped 3.76M data center GPUs in 2023; 98% market share in DC GPU units", "June 2024"),
    ("Zylos Research", "TPU v6 Trillium: ~1.6M units expected in 2025; TPU v7E ~500K units in 2026", "February 2026"),
    ("AMD / WCCFTech", "MI300X: 400-500K units shipped in 2024; 11% of TSMC CoWoS capacity in 2026", "2024-2026"),
    ("TechCrunch / AWS", "1-1.4M Trainium chips deployed by end 2025; Project Rainier: 500K Trainium2 at single site", "March 2026"),
    ("Semi Fundamental", "TPU internal vs cloud split: ~80% internal in 2025 → ~50% in 2026; channel checks", "March 2026"),
    ("The Register", "Anthropic $30B run rate; 3.5 GW TPU deal with Google/Broadcom for 2027", "April 2026"),
    ("Anthropic Blog", "Expanding use of Google Cloud TPUs; up to 1M TPUs committed", "November 2025"),
]


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------
def build_workbook():
    wb = Workbook()

    # ===== Sheet 1: Overall Market Share =====
    ws = wb.active
    ws.title = "Overall Market Share"
    ws.sheet_properties.tabColor = "1B2A4A"

    add_title(ws, 1, "AI Accelerator Market Share by Vendor (2023-2028E)")
    add_subtitle(ws, 2, "Broadcom-Google TPU vs NVIDIA vs AMD — Overall Market")

    row = 4
    headers = ["Vendor"] + [str(y) for y in YEARS]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    row = 5
    for vendor, shares in OVERALL_SHARE.items():
        ws.cell(row=row, column=1, value=vendor)
        for c, s in enumerate(shares, 2):
            ws.cell(row=row, column=c, value=s)
        row += 1

    ws.cell(row=row, column=1, value="Total Market Size ($B)")
    ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=True)
    for c, val in enumerate(TOTAL_MARKET_SIZE, 2):
        cell = ws.cell(row=row, column=c, value=val)
        cell.number_format = USD_FMT
        cell.font = Font(name="Calibri", size=10, bold=True)
    row += 1

    style_data_area(ws, 5, row - 2, len(headers), PCT_FMT)
    for r in range(5, row - 1):
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")

    add_source_note(ws, row + 1, "Sources: Bloomberg Intelligence, SemiAnalysis, New Street Research, Silicon Analysts (2025-2026)")

    # Line chart for overall market share
    chart = LineChart()
    chart.title = "AI Accelerator Market Share Trajectory (2023-2028E)"
    chart.x_axis.title = "Year"
    chart.y_axis.title = "Market Share (%)"
    chart.y_axis.numFmt = '0%'
    chart.style = 10
    chart.width = 22
    chart.height = 14

    cats = Reference(ws, min_col=2, max_col=len(YEARS) + 1, min_row=4)
    for i, vendor in enumerate(OVERALL_SHARE.keys()):
        vals = Reference(ws, min_col=2, max_col=len(YEARS) + 1, min_row=5 + i)
        s = Series(vals, title=vendor)
        chart.append(s)
    chart.set_categories(cats)

    colors = [NVIDIA_GREEN, AMD_RED, GOOGLE_BLUE, ASIC_PURPLE, OTHER_GRAY]
    for i, s in enumerate(chart.series):
        s.graphicalProperties.line.width = 28000
        if i < len(colors):
            s.graphicalProperties.line.solidFill = colors[i]

    ws.add_chart(chart, "A" + str(row + 3))

    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 20

    # ===== Sheet 2: Training Market Share =====
    ws2 = wb.create_sheet("Training Share")
    ws2.sheet_properties.tabColor = NVIDIA_GREEN

    add_title(ws2, 1, "AI Training Market Share by Vendor (2023-2028E)")
    add_subtitle(ws2, 2, "NVIDIA dominates training — CUDA ecosystem + NVLink maintain strong moat")

    row = 4
    headers = ["Vendor"] + [str(y) for y in YEARS]
    for c, h in enumerate(headers, 1):
        ws2.cell(row=row, column=c, value=h)
    style_header_row(ws2, row, len(headers))

    row = 5
    for vendor, shares in TRAINING_SHARE.items():
        ws2.cell(row=row, column=1, value=vendor)
        for c, s in enumerate(shares, 2):
            ws2.cell(row=row, column=c, value=s)
        row += 1

    ws2.cell(row=row, column=1, value="Training % of Total AI Compute")
    ws2.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=True)
    for c, val in enumerate(TRAINING_PCT_OF_TOTAL, 2):
        ws2.cell(row=row, column=c, value=val)
    row += 1

    training_rev = compute_revenue(TRAINING_SHARE, TRAINING_PCT_OF_TOTAL)
    row += 1
    add_subtitle(ws2, row, "Training Revenue by Vendor ($B)")
    row += 1
    for c, h in enumerate(headers, 1):
        ws2.cell(row=row, column=c, value=h)
    style_header_row(ws2, row, len(headers))
    rev_start = row + 1
    row += 1
    for vendor, revs in training_rev.items():
        ws2.cell(row=row, column=1, value=vendor)
        for c, r in enumerate(revs, 2):
            ws2.cell(row=row, column=c, value=r)
        row += 1

    style_data_area(ws2, 5, 5 + len(TRAINING_SHARE) - 1, len(headers), PCT_FMT)
    ws2.cell(row=5 + len(TRAINING_SHARE), column=1).alignment = Alignment(horizontal="left")
    for r in range(5, 5 + len(TRAINING_SHARE)):
        ws2.cell(row=r, column=1).alignment = Alignment(horizontal="left")
    style_data_area(ws2, rev_start, rev_start + len(training_rev) - 1, len(headers), USD_FMT)
    for r in range(rev_start, rev_start + len(training_rev)):
        ws2.cell(row=r, column=1).alignment = Alignment(horizontal="left")

    add_source_note(ws2, row + 1, "Sources: Bloomberg Intelligence, SemiAnalysis, Morgan Stanley, NVIDIA GTC 2026")

    chart2 = BarChart()
    chart2.type = "col"
    chart2.grouping = "stacked"
    chart2.title = "Training Revenue by Vendor ($B, 2023-2028E)"
    chart2.x_axis.title = "Year"
    chart2.y_axis.title = "Revenue ($B)"
    chart2.style = 10
    chart2.width = 22
    chart2.height = 14

    cats2 = Reference(ws2, min_col=2, max_col=len(YEARS) + 1, min_row=rev_start - 1)
    for i, vendor in enumerate(training_rev.keys()):
        vals = Reference(ws2, min_col=2, max_col=len(YEARS) + 1, min_row=rev_start + i)
        s = Series(vals, title=vendor)
        chart2.append(s)
    chart2.set_categories(cats2)

    for i, s in enumerate(chart2.series):
        if i < len(colors):
            s.graphicalProperties.solidFill = colors[i]

    ws2.add_chart(chart2, "A" + str(row + 3))

    for c in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(c)].width = 22

    # ===== Sheet 3: Inference Market Share =====
    ws3 = wb.create_sheet("Inference Share")
    ws3.sheet_properties.tabColor = GOOGLE_BLUE

    add_title(ws3, 1, "AI Inference Market Share by Vendor (2023-2028E)")
    add_subtitle(ws3, 2, "NVIDIA's moat eroding — Google/Broadcom TPU + Custom ASICs gaining rapidly")

    row = 4
    for c, h in enumerate(headers, 1):
        ws3.cell(row=row, column=c, value=h)
    style_header_row(ws3, row, len(headers))

    row = 5
    for vendor, shares in INFERENCE_SHARE.items():
        ws3.cell(row=row, column=1, value=vendor)
        for c, s in enumerate(shares, 2):
            ws3.cell(row=row, column=c, value=s)
        row += 1

    ws3.cell(row=row, column=1, value="Inference % of Total AI Compute")
    ws3.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=True)
    for c, val in enumerate(INFERENCE_PCT_OF_TOTAL, 2):
        ws3.cell(row=row, column=c, value=val)
    row += 1

    inference_rev = compute_revenue(INFERENCE_SHARE, INFERENCE_PCT_OF_TOTAL)
    row += 1
    add_subtitle(ws3, row, "Inference Revenue by Vendor ($B)")
    row += 1
    for c, h in enumerate(headers, 1):
        ws3.cell(row=row, column=c, value=h)
    style_header_row(ws3, row, len(headers))
    rev_start3 = row + 1
    row += 1
    for vendor, revs in inference_rev.items():
        ws3.cell(row=row, column=1, value=vendor)
        for c, r in enumerate(revs, 2):
            ws3.cell(row=row, column=c, value=r)
        row += 1

    style_data_area(ws3, 5, 5 + len(INFERENCE_SHARE) - 1, len(headers), PCT_FMT)
    for r in range(5, 5 + len(INFERENCE_SHARE)):
        ws3.cell(row=r, column=1).alignment = Alignment(horizontal="left")
    style_data_area(ws3, rev_start3, rev_start3 + len(inference_rev) - 1, len(headers), USD_FMT)
    for r in range(rev_start3, rev_start3 + len(inference_rev)):
        ws3.cell(row=r, column=1).alignment = Alignment(horizontal="left")

    add_source_note(ws3, row + 1, "Sources: New Street Research, Goldman Sachs, Zylos Research, Epoch AI (2025-2026)")

    chart3 = BarChart()
    chart3.type = "col"
    chart3.grouping = "stacked"
    chart3.title = "Inference Revenue by Vendor ($B, 2023-2028E)"
    chart3.x_axis.title = "Year"
    chart3.y_axis.title = "Revenue ($B)"
    chart3.style = 10
    chart3.width = 22
    chart3.height = 14

    cats3 = Reference(ws3, min_col=2, max_col=len(YEARS) + 1, min_row=rev_start3 - 1)
    for i, vendor in enumerate(inference_rev.keys()):
        vals = Reference(ws3, min_col=2, max_col=len(YEARS) + 1, min_row=rev_start3 + i)
        s = Series(vals, title=vendor)
        chart3.append(s)
    chart3.set_categories(cats3)

    for i, s in enumerate(chart3.series):
        if i < len(colors):
            s.graphicalProperties.solidFill = colors[i]

    ws3.add_chart(chart3, "A" + str(row + 3))

    # Side-by-side inference share line chart
    chart3b = LineChart()
    chart3b.title = "Inference Market Share: NVIDIA Decline vs TPU/ASIC Rise"
    chart3b.x_axis.title = "Year"
    chart3b.y_axis.title = "Inference Market Share (%)"
    chart3b.y_axis.numFmt = '0%'
    chart3b.style = 10
    chart3b.width = 22
    chart3b.height = 14

    for i, vendor in enumerate(INFERENCE_SHARE.keys()):
        vals = Reference(ws3, min_col=2, max_col=len(YEARS) + 1, min_row=5 + i)
        s = Series(vals, title=vendor)
        chart3b.append(s)
    chart3b.set_categories(cats3)

    for i, s in enumerate(chart3b.series):
        s.graphicalProperties.line.width = 28000
        if i < len(colors):
            s.graphicalProperties.line.solidFill = colors[i]

    ws3.add_chart(chart3b, "A" + str(row + 20))

    for c in range(1, len(headers) + 1):
        ws3.column_dimensions[get_column_letter(c)].width = 22

    # ===== Sheet 4: Training vs Inference Split =====
    ws4 = wb.create_sheet("Training vs Inference")
    ws4.sheet_properties.tabColor = "FF9900"

    add_title(ws4, 1, "Training vs Inference: The Great Compute Shift")
    add_subtitle(ws4, 2, "Inference now dominates — 67% of all AI compute in 2026, projected 75% by 2028")

    row = 4
    split_headers = ["Metric"] + [str(y) for y in YEARS]
    for c, h in enumerate(split_headers, 1):
        ws4.cell(row=row, column=c, value=h)
    style_header_row(ws4, row, len(split_headers))

    rows_data = [
        ("Training % of Total Compute", TRAINING_PCT_OF_TOTAL),
        ("Inference % of Total Compute", INFERENCE_PCT_OF_TOTAL),
        ("Total Market Size ($B)", TOTAL_MARKET_SIZE),
    ]

    training_market = [TOTAL_MARKET_SIZE[i] * TRAINING_PCT_OF_TOTAL[i] for i in range(len(YEARS))]
    inference_market = [TOTAL_MARKET_SIZE[i] * INFERENCE_PCT_OF_TOTAL[i] for i in range(len(YEARS))]
    rows_data.append(("Training Market Size ($B)", training_market))
    rows_data.append(("Inference Market Size ($B)", inference_market))

    row = 5
    for label, data in rows_data:
        ws4.cell(row=row, column=1, value=label)
        ws4.cell(row=row, column=1).alignment = Alignment(horizontal="left")
        for c, val in enumerate(data, 2):
            cell = ws4.cell(row=row, column=c, value=val)
            if "%" in label:
                cell.number_format = PCT_FMT
            else:
                cell.number_format = USD_FMT
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
        row += 1

    add_source_note(ws4, row + 1, "Sources: Morgan Stanley, Epoch AI, New Street Research (2025)")

    # Stacked bar chart: training vs inference market size
    chart4 = BarChart()
    chart4.type = "col"
    chart4.grouping = "stacked"
    chart4.title = "AI Compute Market: Training vs Inference ($B)"
    chart4.x_axis.title = "Year"
    chart4.y_axis.title = "Market Size ($B)"
    chart4.style = 10
    chart4.width = 22
    chart4.height = 14

    cats4 = Reference(ws4, min_col=2, max_col=len(YEARS) + 1, min_row=4)
    training_ref = Reference(ws4, min_col=2, max_col=len(YEARS) + 1, min_row=8)
    inference_ref = Reference(ws4, min_col=2, max_col=len(YEARS) + 1, min_row=9)

    s1 = Series(training_ref, title="Training")
    s1.graphicalProperties.solidFill = NVIDIA_GREEN
    s2 = Series(inference_ref, title="Inference")
    s2.graphicalProperties.solidFill = GOOGLE_BLUE

    chart4.append(s1)
    chart4.append(s2)
    chart4.set_categories(cats4)

    ws4.add_chart(chart4, "A" + str(row + 3))

    for c in range(1, len(split_headers) + 1):
        ws4.column_dimensions[get_column_letter(c)].width = 28

    # ===== Sheet 5: ASIC vs GPU Growth =====
    ws5 = wb.create_sheet("ASIC vs GPU Growth")
    ws5.sheet_properties.tabColor = ASIC_PURPLE

    add_title(ws5, 1, "Custom ASIC vs GPU Growth Rate Comparison (through 2033)")
    add_subtitle(ws5, 2, "ASICs growing at 44.6% CAGR vs GPUs at 16.1% — driven by inference economics")

    row = 4
    cagr_headers = ["Market Segment", "CAGR", "2024 Revenue", "2033 Projected", "Primary Use Case"]
    for c, h in enumerate(cagr_headers, 1):
        ws5.cell(row=row, column=c, value=h)
    style_header_row(ws5, row, len(cagr_headers))

    row = 5
    for data_row in CAGR_DATA:
        for c, val in enumerate(data_row, 1):
            cell = ws5.cell(row=row, column=c, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if c > 1 else "left")
        row += 1

    add_source_note(ws5, row + 1, "Source: Bloomberg Intelligence, AI Accelerator Market Forecast (January 2026)")

    # Key takeaways
    row += 3
    add_subtitle(ws5, row, "Key Drivers of ASIC Outperformance")
    row += 1
    drivers = [
        "1. Inference workloads now 67% of compute (2026) — ASICs purpose-built for predictable, cost-sensitive inference",
        "2. Hyperscaler vertical integration — Google, Microsoft, Amazon, Meta each investing $60-80B in AI capex (2026)",
        "3. Cost advantage: Midjourney cut inference costs 65% migrating from NVIDIA GPUs to Google TPUs",
        "4. Power efficiency: TPU v7 ~500W vs NVIDIA Vera Rubin ~1000W — 2× better at data center scale",
        "5. TSMC 3nm capacity running at 100% utilization — demand 3× supply; favors committed hyperscaler orders",
    ]
    for d in drivers:
        ws5.cell(row=row, column=1, value=d)
        ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws5.cell(row=row, column=1).font = DATA_FONT
        row += 1

    for c in range(1, len(cagr_headers) + 1):
        ws5.column_dimensions[get_column_letter(c)].width = 30

    # ===== Sheet 6: 2026 Chip Specs =====
    ws6 = wb.create_sheet("2026 Chip Specs")
    ws6.sheet_properties.tabColor = "FF6600"

    add_title(ws6, 1, "2026-Class AI Accelerator Spec Comparison", merge_end=10)
    add_subtitle(ws6, 2, "Comprehensive chip-by-chip comparison: NVIDIA, Google/Broadcom, AMD, Microsoft, AWS", merge_end=10)

    row = 4
    spec_headers = ["Chip", "Vendor", "Process", "FP8 (PFLOPS)", "FP4 (PFLOPS)",
                     "Memory (GB)", "Mem BW (TB/s)", "TDP (W)", "Interconnect", "Primary Workload"]
    for c, h in enumerate(spec_headers, 1):
        ws6.cell(row=row, column=c, value=h)
    style_header_row(ws6, row, len(spec_headers))

    row = 5
    vendor_fills = {
        "NVIDIA": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
        "Google/Broadcom": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
        "AMD": PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
        "Microsoft/Broadcom": PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),
        "AWS": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    }

    for spec_row in CHIP_SPECS:
        vendor = spec_row[1]
        fill = vendor_fills.get(vendor, None)
        for c, val in enumerate(spec_row, 1):
            cell = ws6.cell(row=row, column=c, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if c > 2 else "left", wrap_text=True)
            if fill:
                cell.fill = fill
        row += 1

    add_source_note(ws6, row + 1, "Sources: NVIDIA GTC 2026, Google Cloud Next 2025, AMD CES 2026, AWS re:Invent 2025, Microsoft Azure Blog (2026)", merge_end=10)

    for c in range(1, len(spec_headers) + 1):
        ws6.column_dimensions[get_column_letter(c)].width = 20

    # ===== Sheet 7: Inference TCO =====
    ws7 = wb.create_sheet("Inference TCO")
    ws7.sheet_properties.tabColor = "009688"

    add_title(ws7, 1, "3-Year TCO Comparison: 1,000-Chip Inference Cluster")
    add_subtitle(ws7, 2, "NVIDIA H100/B200 vs Google TPU v6/v7 vs AMD MI350/MI400 — 24/7 inference at 80% utilization")

    row = 4
    tco_headers = ["Cost Factor", "NVIDIA H100/B200", "Google TPU v6/v7", "AMD MI350/MI400", "Winner"]
    for c, h in enumerate(tco_headers, 1):
        ws7.cell(row=row, column=c, value=h)
    style_header_row(ws7, row, len(tco_headers))

    row = 5
    for tco_row in TCO_DATA:
        is_total = "TOTAL" in tco_row[0]
        for c, val in enumerate(tco_row, 1):
            cell = ws7.cell(row=row, column=c, value=val)
            cell.font = Font(name="Calibri", size=10, bold=is_total)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if c > 1 else "left")
            if is_total:
                cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        row += 1

    add_source_note(ws7, row + 1, "Sources: Google Cloud TCO calculators, NVIDIA DGX pricing, Uptime Institute datacenter energy audits")

    row += 3
    add_subtitle(ws7, row, "Key TCO Takeaways")
    row += 1
    takeaways = [
        "Google TPU delivers 56% lower 3-year TCO vs NVIDIA for dedicated inference workloads",
        "Power consumption is the largest differentiator: TPU v7 ~500W vs NVIDIA B200 1000W",
        "AMD positioned between NVIDIA and TPU — 27% lower TCO than NVIDIA, 65% higher than TPU",
        "Software migration cost typically $80K-$200K; payback period 18-48 days at $105K/mo savings",
        "At Meta scale (600K+ chips), the TCO delta represents ~$59B over hardware lifecycle",
    ]
    for t in takeaways:
        ws7.cell(row=row, column=1, value=t)
        ws7.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws7.cell(row=row, column=1).font = DATA_FONT
        row += 1

    for c in range(1, len(tco_headers) + 1):
        ws7.column_dimensions[get_column_letter(c)].width = 28

    # ===== Sheet 8: Broadcom Revenue =====
    ws8 = wb.create_sheet("Broadcom ASIC Revenue")
    ws8.sheet_properties.tabColor = BROADCOM_RED

    add_title(ws8, 1, "Broadcom AI Revenue & Google TPU Run-Rate Derivation", merge_end=8)
    add_subtitle(ws8, 2, "Based on Broadcom reported AI semiconductor revenue, earnings call commentary, and analyst estimates", merge_end=8)

    # Part 1: Annual revenue breakdown
    row = 4
    add_subtitle(ws8, row, "Annual AI Revenue Breakdown ($B)", merge_end=8)
    row += 1
    bc_headers = ["Revenue Segment"] + BROADCOM_YEARS
    for c, h in enumerate(bc_headers, 1):
        ws8.cell(row=row, column=c, value=h)
    style_header_row(ws8, row, len(bc_headers))

    row += 1
    for segment, vals in BROADCOM_DATA.items():
        is_total = "Total" in segment
        ws8.cell(row=row, column=1, value=segment)
        for c, val in enumerate(vals, 2):
            cell = ws8.cell(row=row, column=c, value=val)
            cell.number_format = USD_FMT
            cell.font = Font(name="Calibri", size=10, bold=is_total)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if is_total:
                cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        ws8.cell(row=row, column=1).alignment = Alignment(horizontal="left")
        ws8.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=is_total)
        ws8.cell(row=row, column=1).border = THIN_BORDER
        row += 1

    add_source_note(ws8, row, "FY2023-FY2025 = Broadcom reported actuals; FY2026E-FY2027E = guidance + analyst est.", merge_end=8)
    row += 2

    # Part 2: Quarterly run-rate derivation
    add_subtitle(ws8, row, "Quarterly Run-Rate Derivation: How to Estimate Current TPU Revenue", merge_end=8)
    row += 1
    q_headers = BROADCOM_QUARTERLY["headers"]
    for c, h in enumerate(q_headers, 1):
        ws8.cell(row=row, column=c, value=h)
    style_header_row(ws8, row, len(q_headers))

    q_data_start = row + 1
    row += 1
    highlight_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    for q_row in BROADCOM_QUARTERLY["rows"]:
        is_latest = "actual" in q_row[0] or "guide" in q_row[0]
        for c, val in enumerate(q_row, 1):
            cell = ws8.cell(row=row, column=c, value=val)
            cell.font = Font(name="Calibri", size=10, bold=is_latest)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if c > 1 else "left")
            if is_latest:
                cell.fill = highlight_fill
        row += 1

    row += 1
    add_subtitle(ws8, row, "Derivation Notes", merge_end=8)
    row += 1
    for note in BROADCOM_QUARTERLY["notes"]:
        ws8.cell(row=row, column=1, value=note)
        ws8.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(q_headers))
        ws8.cell(row=row, column=1).font = Font(name="Calibri", size=9, italic=True)
        row += 1

    row += 1
    add_subtitle(ws8, row, "Key Conclusion", merge_end=8)
    row += 1
    conclusions = [
        "Current Google TPU revenue through Broadcom: ~$3.4B/quarter = ~$13.6B annualized run rate (Q1 FY2026)",
        "Based on: $8.4B AI revenue × (1 - 33% networking) × ~60% Google share of XPU = ~$3.4B",
        "This is Broadcom's revenue for designing/manufacturing TPUs — not the total value of TPU compute deployed",
        "Broadcom total AI run rate: $8.4B/qtr = $33.6B annualized; guided to $10.7B/qtr = $42.8B ann. by Q2",
        "FY2025 full-year AI revenue was $20B (reported); FY2026E consensus ~$38-46B; FY2027 mgmt target >$100B",
    ]
    for conc in conclusions:
        ws8.cell(row=row, column=1, value=conc)
        ws8.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(q_headers))
        ws8.cell(row=row, column=1).font = DATA_FONT
        row += 1

    add_source_note(ws8, row + 1, "Sources: Broadcom Q1 FY2026 earnings (March 4, 2026), Q4 FY2025 earnings, Reuters, JPMorgan", merge_end=8)
    row += 2

    chart8 = BarChart()
    chart8.type = "col"
    chart8.grouping = "stacked"
    chart8.title = "Broadcom AI Revenue by Segment ($B)"
    chart8.x_axis.title = "Fiscal Year"
    chart8.y_axis.title = "Revenue ($B)"
    chart8.style = 10
    chart8.width = 20
    chart8.height = 13

    cats8 = Reference(ws8, min_col=2, max_col=len(BROADCOM_YEARS) + 1, min_row=5)
    bc_colors = [GOOGLE_BLUE, ASIC_PURPLE, "FF9900"]
    for i, (segment, _) in enumerate(list(BROADCOM_DATA.items())[:3]):
        vals = Reference(ws8, min_col=2, max_col=len(BROADCOM_YEARS) + 1, min_row=6 + i)
        s = Series(vals, title=segment)
        if i < len(bc_colors):
            s.graphicalProperties.solidFill = bc_colors[i]
        chart8.append(s)
    chart8.set_categories(cats8)
    ws8.add_chart(chart8, "A" + str(row + 1))

    for c in range(1, max(len(bc_headers), len(q_headers)) + 1):
        ws8.column_dimensions[get_column_letter(c)].width = 22

    # ===== Sheet 9: Unit Shipments =====
    wsu = wb.create_sheet("Unit Shipments")
    wsu.sheet_properties.tabColor = "3F51B5"

    merge_w = len(UNIT_YEARS) + 1
    add_title(wsu, 1, "AI Accelerator Unit Shipments by Vendor (thousands of chips)", merge_end=merge_w)
    add_subtitle(wsu, 2, "NVIDIA GPUs vs Google/Broadcom TPU vs AMD GPUs vs AWS Trainium/Inferentia (2022-2026E)", merge_end=merge_w)

    u_headers = ["Chip / Generation"] + [str(y) for y in UNIT_YEARS]

    vendor_sections = [
        ("NVIDIA GPU Shipments (K units)", NVIDIA_UNITS, PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")),
        ("Google/Broadcom TPU Shipments (K units)", GOOGLE_TPU_UNITS, PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")),
        ("AMD GPU Shipments (K units)", AMD_UNITS, PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")),
        ("AWS Trainium/Inferentia Shipments (K units)", AWS_UNITS, PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")),
    ]

    row = 4
    for section_title, data_dict, fill in vendor_sections:
        add_subtitle(wsu, row, section_title, merge_end=merge_w)
        row += 1
        for c, h in enumerate(u_headers, 1):
            wsu.cell(row=row, column=c, value=h)
        style_header_row(wsu, row, len(u_headers))
        row += 1
        for chip, vals in data_dict.items():
            is_total = "Total" in chip
            wsu.cell(row=row, column=1, value=chip)
            wsu.cell(row=row, column=1).alignment = Alignment(horizontal="left")
            wsu.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=is_total)
            wsu.cell(row=row, column=1).border = THIN_BORDER
            for c, val in enumerate(vals, 2):
                cell = wsu.cell(row=row, column=c, value=val)
                cell.number_format = NUM_FMT
                cell.font = Font(name="Calibri", size=10, bold=is_total)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
                if not is_total:
                    cell.fill = fill
                else:
                    cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            row += 1
        row += 1

    # Summary comparison table
    add_subtitle(wsu, row, "Summary: Total Annual Shipments by Vendor (K units)", merge_end=merge_w)
    row += 1
    for c, h in enumerate(u_headers, 1):
        wsu.cell(row=row, column=c, value=("Vendor" if c == 1 else h))
    style_header_row(wsu, row, len(u_headers))
    summary_header_row = row
    row += 1
    summary_start = row
    for vendor, units in UNIT_SUMMARY.items():
        is_total = "Total" in vendor
        wsu.cell(row=row, column=1, value=vendor)
        wsu.cell(row=row, column=1).alignment = Alignment(horizontal="left")
        wsu.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=is_total)
        wsu.cell(row=row, column=1).border = THIN_BORDER
        for c, val in enumerate(units, 2):
            cell = wsu.cell(row=row, column=c, value=val)
            cell.number_format = NUM_FMT
            cell.font = Font(name="Calibri", size=10, bold=is_total)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if is_total:
                cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        row += 1

    row += 1

    # Unit market share table
    add_subtitle(wsu, row, "Unit Market Share (% of total units across 4 vendors)", merge_end=merge_w)
    row += 1
    for c, h in enumerate(u_headers, 1):
        wsu.cell(row=row, column=c, value=("Vendor" if c == 1 else h))
    style_header_row(wsu, row, len(u_headers))
    share_header_row = row
    row += 1
    share_start = row
    for vendor, shares in UNIT_SHARE.items():
        wsu.cell(row=row, column=1, value=vendor)
        wsu.cell(row=row, column=1).alignment = Alignment(horizontal="left")
        wsu.cell(row=row, column=1).font = DATA_FONT
        wsu.cell(row=row, column=1).border = THIN_BORDER
        for c, val in enumerate(shares, 2):
            cell = wsu.cell(row=row, column=c, value=val)
            cell.number_format = PCT_FMT
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
        row += 1

    row += 1

    # Average selling price table
    add_subtitle(wsu, row, "Estimated Average Selling Price per Chip ($K)", merge_end=merge_w)
    row += 1
    for c, h in enumerate(u_headers, 1):
        wsu.cell(row=row, column=c, value=("Vendor" if c == 1 else h))
    style_header_row(wsu, row, len(u_headers))
    row += 1
    for vendor, prices in REVENUE_PER_UNIT.items():
        wsu.cell(row=row, column=1, value=vendor)
        wsu.cell(row=row, column=1).alignment = Alignment(horizontal="left")
        wsu.cell(row=row, column=1).font = DATA_FONT
        wsu.cell(row=row, column=1).border = THIN_BORDER
        for c, val in enumerate(prices, 2):
            cell = wsu.cell(row=row, column=c, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
        row += 1

    row += 1
    add_source_note(wsu, row, "Sources: Epoch AI (ai-chip-sales, March 2026), Jensen Huang (4M Hopper + 3M Blackwell thru Oct 2025),", merge_end=merge_w)
    row += 1
    add_source_note(wsu, row, "HPCwire/Omdia (3.76M Nvidia DC GPUs in 2023), Zylos Research (1.6M Trillium in 2025),", merge_end=merge_w)
    row += 1
    add_source_note(wsu, row, "AMD (400-500K MI300X in 2024, 11% TSMC CoWoS 2026), TechCrunch/AWS (1-1.4M Trainium by end 2025)", merge_end=merge_w)
    row += 2

    # Chart 1: Stacked bar chart of unit shipments
    chart_u1 = BarChart()
    chart_u1.type = "col"
    chart_u1.grouping = "stacked"
    chart_u1.title = "AI Accelerator Unit Shipments by Vendor (K chips, 2022-2026E)"
    chart_u1.x_axis.title = "Year"
    chart_u1.y_axis.title = "Chips Shipped (thousands)"
    chart_u1.style = 10
    chart_u1.width = 22
    chart_u1.height = 14

    cats_u = Reference(wsu, min_col=2, max_col=len(UNIT_YEARS) + 1, min_row=summary_header_row)
    u_colors = [NVIDIA_GREEN, GOOGLE_BLUE, AMD_RED, "FF9900"]
    for i, (vendor, _) in enumerate(list(UNIT_SUMMARY.items())[:4]):
        vals = Reference(wsu, min_col=2, max_col=len(UNIT_YEARS) + 1, min_row=summary_start + i)
        s = Series(vals, title=vendor)
        if i < len(u_colors):
            s.graphicalProperties.solidFill = u_colors[i]
        chart_u1.append(s)
    chart_u1.set_categories(cats_u)
    wsu.add_chart(chart_u1, "A" + str(row))

    # Chart 2: Unit market share line chart
    chart_u2 = LineChart()
    chart_u2.title = "Unit Market Share by Vendor (2022-2026E)"
    chart_u2.x_axis.title = "Year"
    chart_u2.y_axis.title = "Share of Units (%)"
    chart_u2.y_axis.numFmt = '0%'
    chart_u2.style = 10
    chart_u2.width = 22
    chart_u2.height = 14

    cats_u2 = Reference(wsu, min_col=2, max_col=len(UNIT_YEARS) + 1, min_row=share_header_row)
    for i, (vendor, _) in enumerate(UNIT_SHARE.items()):
        vals = Reference(wsu, min_col=2, max_col=len(UNIT_YEARS) + 1, min_row=share_start + i)
        s = Series(vals, title=vendor)
        s.graphicalProperties.line.width = 28000
        if i < len(u_colors):
            s.graphicalProperties.line.solidFill = u_colors[i]
        chart_u2.append(s)
    chart_u2.set_categories(cats_u2)
    wsu.add_chart(chart_u2, "A" + str(row + 17))

    for c in range(1, len(u_headers) + 1):
        wsu.column_dimensions[get_column_letter(c)].width = 26

    # ===== Sheet 10: TPU Deployment Split =====
    wst = wb.create_sheet("TPU Internal vs Cloud")
    wst.sheet_properties.tabColor = GOOGLE_BLUE

    t_merge = len(TPU_DEPLOY_YEARS) + 1
    add_title(wst, 1, "Google TPU Deployment: Internal vs Cloud Rental vs External Direct", merge_end=t_merge)
    add_subtitle(wst, 2, "Shift from ~95% internal (2022) to ~50% internal / 30% GCP rental / 20% direct sales (2026E)", merge_end=t_merge)

    # Part 1: Unit volumes by deployment channel
    row = 4
    add_subtitle(wst, row, "TPU Units by Deployment Channel (K chips)", merge_end=t_merge)
    row += 1
    t_headers = ["Deployment Channel"] + [str(y) for y in TPU_DEPLOY_YEARS]
    for c, h in enumerate(t_headers, 1):
        wst.cell(row=row, column=c, value=h)
    style_header_row(wst, row, len(t_headers))

    deploy_fills = [
        PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
        PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid"),
        PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
    ]

    units_start_row = row + 1
    row += 1
    for i, (channel, data) in enumerate(TPU_DEPLOYMENT_SPLIT.items()):
        wst.cell(row=row, column=1, value=channel)
        wst.cell(row=row, column=1).alignment = Alignment(horizontal="left", wrap_text=True)
        wst.cell(row=row, column=1).font = DATA_FONT
        wst.cell(row=row, column=1).border = THIN_BORDER
        fill = deploy_fills[i] if i < len(deploy_fills) else None
        for c, val in enumerate(data["units_k"], 2):
            cell = wst.cell(row=row, column=c, value=val)
            cell.number_format = NUM_FMT
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if fill:
                cell.fill = fill
        row += 1

    # Total row
    wst.cell(row=row, column=1, value="Total TPU Shipments")
    wst.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=True)
    wst.cell(row=row, column=1).border = THIN_BORDER
    wst.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    tpu_totals = GOOGLE_TPU_UNITS["Google TPU Total"]
    for c, val in enumerate(tpu_totals, 2):
        cell = wst.cell(row=row, column=c, value=val)
        cell.number_format = NUM_FMT
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    row += 2

    # Part 2: Percentage split
    add_subtitle(wst, row, "Deployment Share (% of total TPU shipments)", merge_end=t_merge)
    row += 1
    for c, h in enumerate(t_headers, 1):
        wst.cell(row=row, column=c, value=("Deployment Channel" if c == 1 else h))
    style_header_row(wst, row, len(t_headers))
    pct_header_row = row

    row += 1
    pct_start = row
    for i, (channel, data) in enumerate(TPU_DEPLOYMENT_SPLIT.items()):
        short_name = channel.split("—")[0].strip() if "—" in channel else channel.split("(")[0].strip()
        wst.cell(row=row, column=1, value=short_name)
        wst.cell(row=row, column=1).alignment = Alignment(horizontal="left")
        wst.cell(row=row, column=1).font = DATA_FONT
        wst.cell(row=row, column=1).border = THIN_BORDER
        fill = deploy_fills[i] if i < len(deploy_fills) else None
        for c, val in enumerate(data["pct"], 2):
            cell = wst.cell(row=row, column=c, value=val)
            cell.number_format = PCT_FMT
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if fill:
                cell.fill = fill
        row += 1

    row += 1
    add_source_note(wst, row, "Source: Semi Fundamental channel checks (March 2026); internal ~80% in 2025 → ~50% in 2026", merge_end=t_merge)
    row += 2

    # Part 3: Key external customer details
    add_subtitle(wst, row, "Key TPU Customers: Internal & External", merge_end=t_merge)
    row += 1
    cust_headers = ["Customer", "TPU Usage", "Timeline", "Notes"]
    for c, h in enumerate(cust_headers, 1):
        wst.cell(row=row, column=c, value=h)
    style_header_row(wst, row, len(cust_headers))

    row += 1
    for cust in TPU_CUSTOMER_DETAILS:
        for c, val in enumerate(cust, 1):
            cell = wst.cell(row=row, column=c, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left" if c <= 2 else "center", wrap_text=True)
        row += 1

    row += 1
    add_source_note(wst, row, "Sources: Semi Fundamental (March 2026), Anthropic Blog, The Register (April 2026), Reuters, TechCrunch", merge_end=t_merge)
    row += 1
    add_source_note(wst, row, "Note: Anthropic's 1M TPU deal is split ~40% owned racks (direct from Broadcom) + ~60% GCP rental", merge_end=t_merge)
    row += 2

    # Part 4: Key narrative takeaways
    add_subtitle(wst, row, "Key Takeaways", merge_end=t_merge)
    row += 1
    takeaways_tpu = [
        "2022-2024: TPU was overwhelmingly an internal Google asset (~80-95% for Search, YouTube, Ads, Gemini)",
        "2025: External demand begins accelerating — Anthropic 1M TPU deal, Apple ~200K units, Meta testing",
        "2026E: Dramatic shift — internal drops to ~50%, GCP rental rises to ~30%, direct sales reach ~20%",
        "This transition converts TPU from a Google cost center into a revenue-generating cloud platform",
        "Anthropic alone may account for 15-20% of all TPU capacity by 2027 (owned + rented combined)",
        "Google's $185B 2026 capex is justified by this dual demand: internal Gemini + external monetization",
        "Google's TPU gross margin to external customers is ~20-30%, much lower than Nvidia's ~75%",
    ]
    for t in takeaways_tpu:
        wst.cell(row=row, column=1, value=t)
        wst.merge_cells(start_row=row, start_column=1, end_row=row, end_column=t_merge)
        wst.cell(row=row, column=1).font = DATA_FONT
        row += 1

    row += 1

    # Chart 1: Stacked bar — units by deployment channel
    chart_t1 = BarChart()
    chart_t1.type = "col"
    chart_t1.grouping = "stacked"
    chart_t1.title = "Google TPU Deployment by Channel (K chips, 2022-2026E)"
    chart_t1.x_axis.title = "Year"
    chart_t1.y_axis.title = "TPU Chips (thousands)"
    chart_t1.style = 10
    chart_t1.width = 22
    chart_t1.height = 14

    cats_t = Reference(wst, min_col=2, max_col=len(TPU_DEPLOY_YEARS) + 1, min_row=units_start_row - 1)
    t_colors = ["4CAF50", GOOGLE_BLUE, "FF9800"]
    for i, (channel, _) in enumerate(TPU_DEPLOYMENT_SPLIT.items()):
        vals = Reference(wst, min_col=2, max_col=len(TPU_DEPLOY_YEARS) + 1, min_row=units_start_row + i)
        short = channel.split("—")[0].strip() if "—" in channel else channel.split("(")[0].strip()
        s = Series(vals, title=short)
        if i < len(t_colors):
            s.graphicalProperties.solidFill = t_colors[i]
        chart_t1.append(s)
    chart_t1.set_categories(cats_t)
    wst.add_chart(chart_t1, "A" + str(row))

    # Chart 2: Stacked bar — percentage split
    chart_t2 = BarChart()
    chart_t2.type = "col"
    chart_t2.grouping = "percentStacked"
    chart_t2.title = "TPU Deployment Mix (% of total, 2022-2026E)"
    chart_t2.x_axis.title = "Year"
    chart_t2.y_axis.title = "Share of TPU Capacity"
    chart_t2.y_axis.numFmt = '0%'
    chart_t2.style = 10
    chart_t2.width = 22
    chart_t2.height = 14

    cats_t2 = Reference(wst, min_col=2, max_col=len(TPU_DEPLOY_YEARS) + 1, min_row=pct_header_row)
    for i, (channel, _) in enumerate(TPU_DEPLOYMENT_SPLIT.items()):
        vals = Reference(wst, min_col=2, max_col=len(TPU_DEPLOY_YEARS) + 1, min_row=pct_start + i)
        short = channel.split("—")[0].strip() if "—" in channel else channel.split("(")[0].strip()
        s = Series(vals, title=short)
        if i < len(t_colors):
            s.graphicalProperties.solidFill = t_colors[i]
        chart_t2.append(s)
    chart_t2.set_categories(cats_t2)
    wst.add_chart(chart_t2, "A" + str(row + 17))

    for c in range(1, max(t_merge, len(cust_headers)) + 1):
        wst.column_dimensions[get_column_letter(c)].width = 30

    # ===== Sheet 11: Migration Case Studies =====
    ws11m = wb.create_sheet("Migration Case Studies")
    ws11m.sheet_properties.tabColor = "00BCD4"

    add_title(ws11m, 1, "Real-World Migration Case Studies: NVIDIA GPU → TPU/ASIC", merge_end=5)
    add_subtitle(ws11m, 2, "Major AI companies voting with their wallets — inference economics drive migration", merge_end=5)

    row = 4
    cs_headers = ["Company", "Migration Path", "Cost Impact", "Scale", "Timeline"]
    for c, h in enumerate(cs_headers, 1):
        ws11m.cell(row=row, column=c, value=h)
    style_header_row(ws11m, row, len(cs_headers))

    row = 5
    for cs_row in CASE_STUDIES:
        for c, val in enumerate(cs_row, 1):
            cell = ws11m.cell(row=row, column=c, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if c > 1 else "left", wrap_text=True)
        row += 1

    add_source_note(ws11m, row + 1, "Sources: The Information, Anthropic Blog, Reuters, company disclosures (2025-2026)", merge_end=5)

    for c in range(1, len(cs_headers) + 1):
        ws11m.column_dimensions[get_column_letter(c)].width = 28

    # ===== Sheet 12: Sources =====
    ws11 = wb.create_sheet("Sources")
    ws11.sheet_properties.tabColor = "607D8B"

    add_title(ws11, 1, "Data Sources and References")

    row = 3
    src_headers = ["Source", "Data Point / Coverage", "Date"]
    for c, h in enumerate(src_headers, 1):
        ws11.cell(row=row, column=c, value=h)
    style_header_row(ws11, row, len(src_headers))

    row = 4
    for src in SOURCES:
        for c, val in enumerate(src, 1):
            cell = ws11.cell(row=row, column=c, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", wrap_text=True)
        row += 1

    for c in range(1, len(src_headers) + 1):
        ws11.column_dimensions[get_column_letter(c)].width = 45 if c == 2 else 30

    return wb


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("Building Broadcom-Google TPU vs NVIDIA vs AMD market share workbook...")
    wb = build_workbook()
    output_path = "tpu_market_share_comparison.xlsx"
    wb.save(output_path)
    print(f"Saved to {output_path}")
    print("\nSheets created:")
    for ws in wb.worksheets:
        print(f"  - {ws.title}")


if __name__ == "__main__":
    main()
