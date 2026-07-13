#!/usr/bin/env python3
"""Build the 'Climate Change Impact on Data Centers' Excel workbook.

Regenerate the spreadsheet with:  python3 build_workbook.py
Output: climate_change_impact_on_data_centers.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ACCENT = "1F4E79"        # dark blue header
ACCENT_LIGHT = "DDEBF7"  # light blue banding
WHITE = "FFFFFF"

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
HEADER_FILL = PatternFill("solid", fgColor=ACCENT)
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=ACCENT)
SUB_FONT = Font(name="Calibri", size=10, italic=True, color="595959")
BODY_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def add_sheet(wb, name, title, subtitle, headers, rows, widths):
    ws = wb.create_sheet(name)
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.cell(row=2, column=1, value=subtitle).font = SUB_FONT

    header_row = 4
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER

    for r, row in enumerate(rows, header_row + 1):
        band = PatternFill("solid", fgColor=ACCENT_LIGHT) if (r - header_row) % 2 == 0 else None
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = BODY_ALIGN
            cell.border = BORDER
            if band:
                cell.fill = band

    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + len(rows)}"
    return ws


wb = Workbook()
wb.remove(wb.active)

# ---------------------------------------------------------------- 1. Overview
add_sheet(
    wb, "Overview",
    "Climate Change Impact on Data Centers - Facet Overview",
    "Aggregated view of all impact facets. Each facet has a dedicated worksheet with detail. Compiled July 2026.",
    ["#", "Facet", "Worksheet", "What it covers", "Overall severity", "Trend", "Headline finding"],
    [
        [1, "Physical climate hazards", "Physical Hazards",
         "Direct damage and disruption from extreme heat, flooding, wildfire, storms, drought, sea-level rise",
         "High", "Worsening",
         "79% of global data center capacity sits in markets exposed to elevated flood, wind or wildfire risk (First Street, 2026)."],
        [2, "Cooling & water stress", "Cooling & Water",
         "Reduced cooling efficiency, free-cooling limits, water scarcity and consumption conflicts",
         "High", "Worsening",
         "Two-thirds of data centers built or in development since 2022 are in water-stressed regions; a large AI facility can use up to 5 million gallons/day."],
        [3, "Energy & grid", "Energy & Grid",
         "Grid strain during heat waves, power-supply outages, demand growth colliding with climate stress",
         "High", "Worsening",
         "Power supply causes ~45% of data center outages (Uptime Institute); extreme heat stresses data centers and the grid simultaneously."],
        [4, "Indirect / interdependency risk", "Physical Hazards",
         "Failure of surrounding infrastructure (grid, telecom, water, transport, supply chains)",
         "Very high", "Worsening",
         "XDI modelling: including indirect risk increases disruption risk to European data center operations roughly ten-fold."],
        [5, "Financial & insurance", "Financial & Insurance",
         "Asset value erosion, business interruption, insurance pricing and availability, financing terms",
         "High", "Worsening",
         "Global data-center insurance premiums projected to rise from US$10.6bn to US$24.2bn by 2030 (Swiss Re); climate exposure could grow from $388bn to $1.0-3.7tn under AI growth scenarios."],
        [6, "Regulatory & policy", "Regulatory & Policy",
         "Carbon costs, disclosure mandates, water/energy permitting, moratoria, community opposition",
         "Medium", "Tightening",
         "EU (EED recast) requires energy/water reporting; US states shifting from tax incentives to grid-upgrade cost-sharing; 7 in 10 Americans oppose local data center construction (Gallup)."],
        [7, "Data centers' own climate footprint", "DC Climate Footprint",
         "The reverse facet: emissions, energy and water consumption of the sector driving climate change",
         "High", "Growing",
         "Global data center power consumption projected to rise ~160% by 2030 (Goldman Sachs); data centers could reach ~14% of US power demand by 2030, up from ~5% in 2025 (S&P)."],
        [8, "Adaptation & resilience", "Adaptation Strategies",
         "Mitigation levers: siting, cooling technology, water reuse, microgrids, hardening, insurance structuring",
         "Opportunity", "Accelerating",
         "Proactive adaptation protects ~$150bn in net asset value (30-39% of exposed value in every scenario tested)."],
        [9, "Regional exposure", "Regional Exposure",
         "How hazard mix and exposure vary by geography and market",
         "Varies", "Diverging",
         "Climate value-erosion exposure: China ~56% and Europe ~53%, roughly double the US (~28%) and APMEA (~28%)."],
    ],
    [4, 26, 20, 52, 14, 13, 70],
)

# ---------------------------------------------------- 2. Physical hazards
add_sheet(
    wb, "Physical Hazards",
    "Facet 1 - Physical Climate Hazards",
    "Direct and indirect physical threats to facilities and operations.",
    ["Hazard", "Type", "Impact mechanism on data centers", "Operational consequences",
     "Exposure indicators / evidence", "Trend under climate change"],
    [
        ["Extreme heat", "Chronic + acute",
         "Higher ambient temperatures reduce cooling efficiency and shrink thermal headroom; equipment derating; staff heat stress during outdoor maintenance",
         "Higher PUE and energy cost, throttling or shutdown of IT load, shortened equipment life, SLA breaches",
         "Cooling is ~40% of energy use at normal temperatures and rises in heat waves; heat assessed separately by XDI because it drives operational continuity rather than physical damage",
         "More frequent, longer, more intense heat waves globally"],
        ["High humidity (with heat)", "Chronic",
         "Co-occurring high temperature and humidity limits air-side economization (free cooling) and evaporative cooling effectiveness",
         "Loss of free-cooling hours forces mechanical cooling, raising energy and water use",
         "Hours limiting direct air cooling have risen significantly over 45 years, especially in the tropics and southeastern US; share of sites losing free cooling for >25% of the year is rising (Sci Reports, 2026)",
         "Continued expansion of constrained hours through mid-century"],
        ["Riverine / surface flooding", "Acute",
         "Inundation of ground-level equipment, substations, fuel storage, access roads",
         "Outages, equipment write-offs, restricted site access, generator fuel-supply failure",
         "Flooding is one of the top three acute hazards in First Street's 79%-of-capacity exposure finding",
         "Heavier precipitation extremes increase flood frequency"],
        ["Coastal inundation / sea-level rise", "Chronic + acute",
         "Seawater flooding from high tides, storm surge, low pressure and waves damages coastal land, buildings and infrastructure",
         "Permanent site viability loss in worst cases; saltwater damage; rising flood-defense capex",
         "Named as a key hazard in XDI 2026 global analysis of planned data centers",
         "Accelerating with sea-level rise"],
        ["Wildfire", "Acute",
         "Direct fire threat, smoke/particulate ingestion into air handling, PSPS (preemptive power shutoffs)",
         "Air filtration failures, forced shutdowns, grid supply interruption, evacuation of staff",
         "One of the three acute hazards driving the First Street 79% exposure figure",
         "Longer fire seasons, larger burned areas"],
        ["Severe convective storms (tornado, hail, high wind)", "Acute",
         "Damage to roofs, exposed HVAC, cooling towers, solar installations; debris impact",
         "Physical damage, cooling loss, long repair lead-times for bespoke equipment",
         "SCS caused ~US$51bn global insured losses in 2025; 64% of US capacity under construction is moving to frontier markets (West Texas, Tennessee, Wisconsin, Ohio) with elevated tornado/hail/wind risk",
         "Insured nat-cat losses rising 5-7%/yr in real terms (Swiss Re sigma)"],
        ["Tropical cyclones / hurricanes", "Acute",
         "Wind and surge damage, prolonged regional grid and telecom outages",
         "Multi-day outages, fuel logistics failure for backup generation",
         "Cyclones listed among key hazards in XDI 2026 planned-facility analysis",
         "Higher intensity storms; slower-moving, wetter systems"],
        ["Drought / water scarcity", "Chronic",
         "Restricts evaporative cooling water; municipal restrictions can curtail supply",
         "Cooling capacity limits, higher operating cost, community conflict and permitting friction",
         "Two-thirds of new data centers since 2022 sit in water-stressed regions; ~80% of potable water used in cooling evaporates",
         "Expanding drought-prone areas"],
        ["Indirect / interdependency failure", "Systemic",
         "A resilient facility still fails if grid, water, telecom, transport or supply chains fail around it",
         "Outages despite on-site hardening; extended recovery times",
         "XDI: disruption risk to a synthetic European portfolio increased ~10x when indirect risk was included; surrounding infrastructure often faces higher heat-disruption risk than the facilities themselves (notably Indonesia, Brazil, Mexico, Spain, India, US, Malaysia)",
         "Grows with hazard frequency and grid stress"],
    ],
    [26, 15, 46, 40, 52, 32],
)

# ---------------------------------------------------- 3. Cooling & water
add_sheet(
    wb, "Cooling & Water",
    "Facet 2 - Cooling & Water Stress",
    "How warming and water scarcity degrade thermal management, and what it costs.",
    ["Issue", "Detail", "Quantified evidence", "Consequence", "Mitigation options"],
    [
        ["Cooling energy share", "Cooling dominates non-IT energy use and grows disproportionately in heat waves",
         "~40% of data center energy use at normal temperatures, higher during extreme heat (Rhizome/CNBC, 2026)",
         "Peak energy costs coincide with peak grid stress and peak tariffs",
         "Raise supply-water temperature (each +1 degC chiller setpoint cuts cooling energy ~4% - Nvidia); liquid cooling at up to 45 degC coolant"],
        ["Free-cooling (air-side economization) limits", "High heat + humidity hours make outside air unusable for direct cooling",
         "45-year significant rise in limiting hours, worst in tropics and southeastern US; growing share of sites constrained >25% of the year; constraints expand through mid-century (Scientific Reports, June 2026)",
         "Loss of the cheapest, most water-efficient cooling mode; retrofit pressure on hybrid designs",
         "Hybrid architectures, indirect evaporative with backup DX, rear-door heat exchangers, immersion cooling"],
        ["Evaporative cooling water demand", "Water use spikes exactly when supplies are shortest",
         "A single large AI data center can consume up to ~5 million gallons (18.9 ML) per day; ~80% of potable cooling water evaporates; many systems do not recirculate",
         "Direct competition with municipal supply during heat waves and droughts; reputational and permitting risk",
         "Closed-loop systems, non-potable/reclaimed water, air-cooled chillers (energy trade-off), water-use effectiveness (WUE) targets"],
        ["Siting in water-stressed regions", "Growth is concentrated where water is already scarce",
         "Two-thirds of all data centers built or in development since 2022 are in regions already facing water scarcity (Al Jazeera, 2026)",
         "Long-term license-to-operate risk; drought curtailment exposure",
         "Climate-adjusted site selection; water-positive commitments; on-site water storage and reuse"],
        ["Hardware thermal tolerance", "Chip and server design is adapting to hotter operations",
         "Nvidia's newer AI servers run cooling liquid at 45 degC, up from previously lower temperatures",
         "Reduces chiller dependence and enables heat reuse, but raises facility design complexity",
         "Warm-water liquid cooling, heat-reuse export to district heating"],
    ],
    [30, 42, 56, 44, 50],
)

# ---------------------------------------------------- 4. Energy & grid
add_sheet(
    wb, "Energy & Grid",
    "Facet 3 - Energy & Grid Impacts",
    "Climate stress on power supply meets explosive demand growth.",
    ["Issue", "Detail", "Quantified evidence", "Consequence"],
    [
        ["Simultaneous peak stress", "Extreme heat raises data center cooling load exactly when air-conditioning demand peaks and thermal generation/transmission capacity derates",
         "\"Data centers need the most energy exactly when the grid has the least available to give\" (Rhizome CEO, 2026); 2026 US heat wave strained grid and water supplies simultaneously",
         "Elevated blackout/brownout risk; exposure to demand-response curtailment; price spikes"],
        ["Power supply as top outage cause", "Grid supply failure remains the dominant driver of downtime",
         "~45% of data center outages are power-related (Uptime Institute Global Data Center Survey)",
         "Business-interruption losses; investment needed in on-site generation and storage"],
        ["Demand growth outpacing infrastructure", "AI build-out is outstripping grid and water expansion",
         "Global data center power consumption projected +160% by 2030 (Goldman Sachs); data centers ~14% of US power demand by 2030 vs ~5% in 2025 (S&P); utilities and regulators warn construction pace exceeds infrastructure expansion",
         "Interconnection queues, delayed energizations, higher tariffs, political backlash"],
        ["Frontier-market grid fragility", "New capacity is shifting to regions with weaker infrastructure",
         "64% of US capacity under construction is outside traditional hubs; many global planned facilities are in markets with less resilient grids (XDI 2026)",
         "Higher indirect-risk exposure; longer restoration after climate events"],
        ["Backup power vulnerabilities", "Generators and fuel logistics are themselves climate-exposed",
         "Flooded fuel storage, blocked transport access, and multi-day regional outages exceed typical 24-72h fuel reserves",
         "Extended outages during compounding events; pressure for microgrids and long-duration storage"],
    ],
    [32, 52, 62, 50],
)

# ---------------------------------------------------- 5. Financial & insurance
add_sheet(
    wb, "Financial & Insurance",
    "Facet 4 - Financial & Insurance Impacts",
    "How physical climate risk is repricing data center assets, insurance and capital.",
    ["Item", "Detail", "Quantified evidence", "Implication"],
    [
        ["Insurance premium growth", "Rapidly expanding, climate-exposed insured base",
         "Global data-center insurance premiums projected to rise from US$10.6bn to US$24.2bn by 2030 (Swiss Re sigma); hyperscale premiums ~$10bn/yr already, double the global aviation market",
         "Premiums increasingly priced on site-level hazard data and resilience specs"],
        ["Asset value concentration", "Single campuses now rival national infrastructure in insured value",
         "Total insurable values for one hyperscale campus can reach US$20-30bn (S&P); insurable asset base of ~11,000 operating data centers exceeds US$2tn; annual investment to surpass US$300bn by 2027",
         "Severe accumulation/aggregation risk for insurers in cat-exposed locations"],
        ["Climate exposure of asset value", "Downtime cost at AI scale multiplies exposure",
         "Physical risk per GW ~2.6x higher for new AI-era facilities than the installed base; total climate exposure could grow from US$388bn to US$1.0-3.7tn under published AI growth scenarios",
         "Climate-adjusted DCF valuation is becoming standard for facility-level underwriting"],
        ["Five core risk channels", "Framework used in climate-adjusted valuation studies",
         "(1) cooling costs, (2) business interruption, (3) physical damage, (4) heat productivity, (5) carbon costs",
         "All five channels feed discounted-cash-flow impacts on asset value"],
        ["Business interruption (BI)", "Downtime is the dominant financial exposure, tied to SLAs and multi-tenant dependencies",
         "BI, loss of rent, and service interruption flagged as critical operational-phase covers (Swiss Re); power supply drives ~45% of outages",
         "BI cover harder to structure than for conventional property; parametric products emerging"],
        ["Underwriting/valuation gap", "Capital still flows to highest-exposure markets without adjusted pricing",
         "First Street: markets equivalent on power cost and connectivity diverge sharply on climate risk; gap showing in NOI stability, insurance availability, debt capacity, refinancing terms and exit valuations",
         "Growing disconnect between current valuations and forward-looking physical risk"],
        ["Rising nat-cat baseline", "The loss environment underneath is deteriorating",
         "Insured losses from natural catastrophes rising 5-7%/yr on average in real terms; severe convective storms alone caused ~US$51bn global losses in 2025 (Swiss Re)",
         "Structural upward pressure on premiums and deductibles for exposed sites"],
    ],
    [30, 44, 64, 48],
)

# ---------------------------------------------------- 6. Regulatory & policy
add_sheet(
    wb, "Regulatory & Policy",
    "Facet 5 - Regulatory, Policy & Social License",
    "Transition risks: disclosure, carbon costs, permitting and community opposition.",
    ["Area", "Jurisdiction / scope", "Requirement or trend", "Impact on data centers"],
    [
        ["Energy & water disclosure", "European Union",
         "EU Energy Efficiency Directive (recast) introduces mandatory reporting of data center energy performance and water usage; broader transparency requirements tightening",
         "Compliance cost; public benchmarking of PUE/WUE; laggards exposed"],
        ["Carbon pricing & costs", "EU ETS, national carbon taxes, corporate net-zero commitments",
         "Carbon costs identified as one of five core climate risk channels in facility-level valuation studies",
         "Raises cost of fossil-backed power and diesel backup; accelerates PPA and clean-firm procurement"],
        ["Grid cost allocation", "US state regulators",
         "Shift from offering tax incentives toward requiring data centers to pay for grid upgrades and demonstrate clean-energy commitments",
         "Higher connection costs; longer negotiation timelines; clean-energy procurement as a de facto permitting condition"],
        ["Water permitting", "Water-stressed municipalities globally",
         "Heightened scrutiny of cooling-water withdrawals; drought-period restrictions",
         "Permit denials/delays; requirements for reclaimed water or closed-loop cooling"],
        ["Moratoria & zoning", "Ireland (Dublin), Netherlands (Amsterdam), Singapore (historic), various US counties",
         "Grid- and resource-driven pauses or conditions on new connections and construction",
         "Constrains growth in established hubs; pushes capacity to frontier markets with higher physical risk"],
        ["Community opposition", "United States (national polling)",
         "7 in 10 Americans opposed data center construction in their local communities (Gallup, 2026); opposition intensified during heat waves that strained shared grid and water",
         "Political/social license risk; project cancellations and delays; local benefit agreements increasingly required"],
        ["Climate risk disclosure", "Investors, rating agencies, insurers",
         "S&P treats hyperscale data centers as an emerging risk pool; insurers demand granular site-selection and resilience data",
         "Financing conditions and project economics now tied to demonstrated climate resilience"],
    ],
    [26, 38, 60, 50],
)

# ---------------------------------------------------- 7. DC climate footprint
add_sheet(
    wb, "DC Climate Footprint",
    "Facet 6 - Data Centers' Own Climate Footprint (the reverse impact)",
    "The sector is both victim and contributor: its energy, water and carbon footprint feeds the problem.",
    ["Dimension", "Current state", "Trajectory", "Key figures"],
    [
        ["Electricity consumption", "Data centers are among the fastest-growing electricity consumers, driven by AI training and inference",
         "Consumption projected to roughly double-to-triple by 2030",
         "Global power consumption +160% by 2030 (Goldman Sachs); ~14% of US power demand by 2030 vs ~5% in 2025 (S&P)"],
        ["Carbon emissions", "Emissions depend heavily on grid mix; rapid build-out is extending fossil generation life in some markets",
         "Voluntary reporting frameworks remain largely non-binding; EU moving to mandatory transparency",
         "Carbon costs are one of the five valuation risk channels; hyperscalers report rising scope 2/3 emissions since AI ramp"],
        ["Water footprint", "Evaporative cooling consumes potable water, concentrated in stressed basins",
         "Rising with AI capacity, unless closed-loop and reclaimed-water designs are adopted",
         "Up to ~5 million gallons/day for a large AI campus; ~80% of potable cooling water evaporates; two-thirds of new builds since 2022 in water-stressed regions"],
        ["Construction & embodied carbon", "Concrete, steel and IT-hardware supply chains carry large embodied emissions",
         "Construction spend climbing steeply, multiplying embodied footprint",
         "US data center construction spend grew from $1.8bn (2014) to $28.3bn (2024); 565 operating US facilities plus 571 in development"],
        ["Grid decarbonization role", "Large flexible loads can help or hinder the energy transition",
         "PPAs, 24/7 carbon-free energy matching, and demand flexibility becoming competitive differentiators",
         "Clean-energy commitments increasingly a regulatory and community expectation for new connections"],
    ],
    [28, 52, 46, 62],
)

# ---------------------------------------------------- 8. Adaptation strategies
add_sheet(
    wb, "Adaptation Strategies",
    "Facet 7 - Adaptation & Resilience Strategies",
    "Levers to protect uptime and asset value, with indicative effectiveness.",
    ["Strategy", "Category", "Description", "Effectiveness / evidence", "Adoption stage"],
    [
        ["Climate-adjusted site selection", "Planning",
         "Screen candidate sites for heat, flood, wind, wildfire, water stress and indirect infrastructure risk across scenarios to 2100 before committing capital",
         "Planned facilities offer a one-time window: siting and design decisions now materially change future insurability and continuity (XDI 2026); avoids the highest-cost retrofits entirely",
         "Becoming standard among hyperscalers; inconsistent in frontier markets"],
        ["Engineering hardening (archetype upgrades)", "Design",
         "Elevated equipment, flood barriers, wind-rated roofs/HVAC, fire-resistant envelopes, hail-protected cooling towers",
         "Resilience archetypes materially reduce modelled disruption in XDI analysis; FM loss-prevention guidance now recommends increased fire and equipment protection for new builds",
         "Mainstream for new tier-IV builds"],
        ["Liquid / warm-water cooling", "Cooling",
         "Direct-to-chip or immersion cooling tolerating coolant up to 45 degC; reduces chiller and water dependence",
         "Each +1 degC chiller setpoint cuts cooling energy ~4% (Nvidia); decouples cooling from ambient heat extremes",
         "Rapid adoption in AI clusters"],
        ["Closed-loop & reclaimed water systems", "Water",
         "Recirculating cooling water, using non-potable/reclaimed sources, WUE targets",
         "Directly addresses the ~80% evaporation loss of once-through potable systems; reduces drought curtailment and permitting risk",
         "Growing; regulatory pressure accelerating it"],
        ["Microgrids, storage & on-site generation", "Energy",
         "Battery storage, long-duration storage, fuel cells, gas/renewable microgrids islanding from stressed grids",
         "Targets the ~45% of outages caused by power supply; named among value-protecting measures in climate-adjusted valuation studies",
         "Early-to-mid adoption; economics improving"],
        ["Power purchase agreements (PPAs) & clean firm energy", "Energy",
         "Long-term contracted renewable/clean-firm supply reduces carbon cost exposure and supports permitting",
         "Listed among measures whose protected asset value is quantifiable (part of the ~$150bn net protected value finding)",
         "Mainstream among hyperscalers"],
        ["Demand flexibility & workload shifting", "Operations",
         "Temporal/geographic shifting of deferrable AI workloads during grid stress and heat peaks",
         "Reduces exposure to curtailment and price spikes; supports grid during simultaneous-peak events",
         "Emerging; strongest with batch training workloads"],
        ["Insurance structuring & parametrics", "Financial",
         "Layered property/BI programs, parametric triggers for heat/outage events, resilience-linked pricing",
         "Insurers already price on granular resilience data; proactive adaptation protects 30-39% of exposed value in every scenario tested (~$150bn net across installed base)",
         "Maturing rapidly as premiums rise toward $24.2bn by 2030"],
        ["Redundancy & geographic distribution", "Architecture",
         "Multi-region replication and failover so a single-site climate event does not interrupt service",
         "Mitigates the 10x indirect-risk multiplier by de-correlating from any one region's infrastructure",
         "Standard for cloud; harder for latency-sensitive and single-campus AI training"],
    ],
    [34, 14, 52, 58, 34],
)

# ---------------------------------------------------- 9. Regional exposure
add_sheet(
    wb, "Regional Exposure",
    "Facet 8 - Regional Exposure Patterns",
    "How climate risk to data centers differs by geography.",
    ["Region / market", "Dominant hazards", "Exposure notes", "Key figures"],
    [
        ["United States - traditional hubs (N. Virginia, etc.)", "Heat waves, grid strain, water stress (localized)",
         "Mature infrastructure but concentrated load; interconnection queues and community pushback growing",
         "US value-erosion exposure ~28% (lower than China/Europe); data centers ~5% of US power demand in 2025, heading to ~14% by 2030"],
        ["United States - frontier markets (West Texas, Tennessee, Wisconsin, Ohio)", "Tornado, hail, high wind, heat",
         "64% of US capacity under construction is outside traditional hubs; exposed HVAC, cooling towers and solar on vast roofs are vulnerable to severe convective storms",
         "SCS caused ~US$51bn global insured losses in 2025"],
        ["Europe", "Extreme heat, drought, river flooding, grid stress",
         "2026 heat waves disrupted power infrastructure; older grids; strong regulatory tightening (EED)",
         "Value-erosion exposure ~53%; XDI found ~10x disruption-risk increase when indirect risk included for a European portfolio"],
        ["China", "Flooding, typhoons, extreme heat",
         "Largest exposure share in climate-adjusted valuation studies; rapid coastal and inland build-out",
         "Value-erosion exposure ~56% - highest of major regions"],
        ["APMEA (Asia-Pacific, Middle East, Africa)", "Extreme heat + humidity, cyclones, water scarcity, fragile grids",
         "Tropics face the fastest loss of free-cooling hours; surrounding-infrastructure heat risk exceeds facility risk in Indonesia, India, Malaysia",
         "Value-erosion exposure ~28% on average, but with extreme within-region variance"],
        ["Latin America", "Heat, drought, flooding, grid fragility",
         "Brazil and Mexico flagged for surrounding-infrastructure disruption risk above facility-level risk",
         "Included in XDI list of markets with elevated indirect heat-disruption risk"],
        ["Global aggregate", "All hazards",
         "Risk is being locked in now: most future capacity is still in planning, where siting and design changes are cheapest",
         "79% of global capacity in markets with elevated flood/wind/wildfire risk (First Street); global capacity projected to nearly double by 2030"],
    ],
    [40, 38, 56, 56],
)

# ---------------------------------------------------- 10. Key stats & sources
add_sheet(
    wb, "Key Stats & Sources",
    "Key Statistics & Source Register",
    "Headline quantitative findings with attribution. Figures compiled July 2026; verify before external use.",
    ["Statistic", "Value", "Source", "Year"],
    [
        ["Global data center capacity in markets with elevated flood/wind/wildfire risk", "79%", "First Street, 18th National Risk Assessment", "2026"],
        ["Increase in disruption risk when indirect (infrastructure) risk is included - European portfolio", "~10x", "XDI Global Data Centres Report", "2026"],
        ["Global data-center insurance premiums, today vs 2030", "US$10.6bn -> US$24.2bn", "Swiss Re Institute, sigma insights 07/2026", "2026"],
        ["Physical climate risk per GW, new AI-era facilities vs installed base", "~2.6x higher", "Climate-adjusted valuation study (reported by The Tech Data)", "2026"],
        ["Potential total climate exposure under AI growth scenarios", "US$388bn -> US$1.0-3.7tn", "Climate-adjusted valuation study", "2026"],
        ["Net asset value protected by proactive adaptation (installed base)", "~US$150bn (30-39% of exposed value)", "Climate-adjusted valuation study", "2026"],
        ["Regional value-erosion exposure", "China 56%, Europe 53%, US 28%, APMEA 28%", "Climate-adjusted valuation study", "2026"],
        ["Share of data center energy used for cooling (normal conditions)", "~40%", "Rhizome, via CNBC", "2026"],
        ["Cooling energy saved per +1 degC chiller setpoint", "~4%", "Nvidia", "2026"],
        ["Max coolant temperature of newest AI servers", "45 degC", "Nvidia", "2026"],
        ["Share of data center outages caused by power supply", "~45%", "Uptime Institute Global Data Center Survey", "2025"],
        ["Projected growth in global data center power consumption by 2030", "+160%", "Goldman Sachs, via Insurance Business", "2026"],
        ["Data centers' share of US power demand", "~5% (2025) -> ~14% (2030)", "S&P Global", "2026"],
        ["Water consumption of a single large AI data center", "Up to 5 million gallons (18.9 ML) per day", "Al Jazeera reporting", "2026"],
        ["Share of potable cooling water that evaporates", "~80%", "Al Jazeera reporting", "2026"],
        ["New data centers (since 2022) located in water-stressed regions", "Two-thirds", "Al Jazeera reporting", "2026"],
        ["US capacity under construction outside traditional hubs", "64%", "Zurich Insurance, via CNBC", "2026"],
        ["Global insured losses from severe convective storms in 2025", "US$51bn", "Swiss Re sigma", "2026"],
        ["Long-term real growth of insured natural catastrophe losses", "5-7% per year", "Swiss Re sigma", "2026"],
        ["Insurable value of a single hyperscale campus", "US$20-30bn", "S&P Global Ratings", "2026"],
        ["Total insurable asset base of ~11,000 operating data centers", ">US$2tn", "Insurance Journal", "2026"],
        ["Annual hyperscale data center investment by 2027", ">US$300bn", "Insurance Journal", "2026"],
        ["US data center construction spending, 2014 vs 2024", "US$1.8bn -> US$28.3bn", "Insurance Business / industry data", "2026"],
        ["Americans opposing data center construction in their community", "7 in 10", "Gallup", "2026"],
        ["Trend in hours limiting air free cooling (45-yr record)", "Significant increase, worst in tropics & SE US", "Scientific Reports (Nature)", "2026"],
    ],
    [66, 34, 46, 8],
)

wb.save("climate_change_impact_on_data_centers.xlsx")
print("Workbook written: climate_change_impact_on_data_centers.xlsx")
print("Sheets:", wb.sheetnames)
