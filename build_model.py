"""
SpaceX / Cursor (Anysphere) pro-forma three-statement model builder.

Generates a fully formula-linked Excel workbook with an integrated Income
Statement, Balance Sheet and Cash Flow Statement for the combined entity
following SpaceX's all-stock acquisition of Cursor (announced 16-Jun-2026,
$60bn equity value, expected close Q3 2026).

The workbook is "live": every projected cell is an Excel formula that
references the Assumptions tab, so a user can flex any driver and the three
statements recalculate and stay in balance.

Run:  python build_model.py
Out:  SpaceX_Cursor_Pro_Forma_Model.xlsx

All figures are in US$ millions unless stated otherwise.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTFILE = "SpaceX_Cursor_Pro_Forma_Model.xlsx"

YEARS = [2025, 2026, 2027, 2028, 2029, 2030]
YEAR_LABEL = {2025: "2025A", 2026: "2026E", 2027: "2027E",
              2028: "2028E", 2029: "2029E", 2030: "2030E"}

# --- column maps -----------------------------------------------------------
# IS / CF / Schedules / Assumptions: years occupy columns C..H
def yl(year):
    return get_column_letter(3 + YEARS.index(year))

# Balance Sheet has a purchase-accounting bridge before the projection years
BS_COL = {"2025A": "C", "Cursor": "D", "PPA": "E", "Open": "F",
          2026: "G", 2027: "H", 2028: "I", 2029: "J", 2030: "K"}
BS_PREV = {2026: "F", 2027: "G", 2028: "H", 2029: "I", 2030: "J"}

# Sheet titles
S_COVER = "Cover"
S_ASSUM = "Assumptions"
S_SEG = "Segments"
S_PPA = "Deal & PPA"
S_SCH = "Schedules"
S_IS = "Income Statement"
S_BS = "Balance Sheet"
S_CF = "Cash Flow"
S_CHK = "Checks"

# --- styling ---------------------------------------------------------------
NAVY = "0B1F3A"
BLUE = "1F4E79"
LIGHT = "DDEBF7"
GREY = "F2F2F2"
GREEN = "E2EFDA"
INPUT_FONT = Font(color="0000CC")            # blue = hard-coded input
LABEL_FONT = Font(color="000000")
BOLD = Font(bold=True)
WHITE_BOLD = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=16, color=NAVY)
SUB_FONT = Font(italic=True, size=9, color="595959")

FMT_USD = '#,##0;(#,##0)'
FMT_USD1 = '#,##0.0;(#,##0.0)'
FMT_PCT = '0.0%'
FMT_X = '#,##0.00'

thin = Side(style="thin", color="BFBFBF")
TOP_BORDER = Border(top=Side(style="thin", color="404040"))
DBL_TOP = Border(top=Side(style="double", color="404040"))


class _Null:
    """Swallows attribute assignment during the layout pass."""
    def __setattr__(self, k, v):
        pass


_NULL = _Null()


def header_fill():
    return PatternFill("solid", fgColor=BLUE)


def section_fill():
    return PatternFill("solid", fgColor=LIGHT)


class ModelBuilder:
    def __init__(self):
        self.wb = Workbook()
        self.ws = {}
        self.row = {}          # (sheet_title, key) -> row number
        self._rowcount = {}    # sheet_title -> last used row
        self.layout_only = False
        # create sheets in display order
        first = self.wb.active
        first.title = S_COVER
        self.ws[S_COVER] = first
        for t in [S_ASSUM, S_SEG, S_PPA, S_SCH, S_IS, S_BS, S_CF, S_CHK]:
            self.ws[t] = self.wb.create_sheet(t)

    # -- reference helpers --------------------------------------------------
    def r(self, sheet, key):
        return self.row[(sheet, key)]

    def ref(self, sheet, key, col):
        """Absolute-row reference like 'IS SpaceX'!C12 (col is a letter)."""
        if (sheet, key) not in self.row:
            if self.layout_only:
                return "PENDING"      # discarded; formulas only written in pass 2
            raise KeyError((sheet, key))
        return f"'{sheet}'!{col}{self.row[(sheet, key)]}"

    def yref(self, sheet, key, year):
        return self.ref(sheet, key, yl(year))

    def aref(self, key, year):
        return self.yref(S_ASSUM, key, year)

    def aref1(self, key):
        """single-value assumption (stored in column C)."""
        return self.ref(S_ASSUM, key, "C")

    # -- low-level writers --------------------------------------------------
    def put(self, sheet, key, label, *, indent=0, bold=False, fmt=FMT_USD,
            section=False, note=None):
        ws = self.ws[sheet]
        if (sheet, key) in self.row:
            row = self.row[(sheet, key)]
        else:
            row = self._rowcount.get(sheet, 0) + 1
            self._rowcount[sheet] = row
            self.row[(sheet, key)] = row
        c = ws.cell(row=row, column=1, value=label)
        c.alignment = Alignment(indent=indent)
        if section:
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.fill = section_fill()
            c.font = Font(bold=True, color=NAVY)
        elif bold:
            c.font = BOLD
        if note:
            ws.cell(row=row, column=12, value=note).font = SUB_FONT
        return row

    def setval(self, sheet, key, col, value, *, fmt=FMT_USD, is_input=False,
               bold=False, border=None):
        if self.layout_only:
            return _NULL
        ws = self.ws[sheet]
        row = self.row[(sheet, key)]
        cell = ws.cell(row=row, column=_col_idx(col), value=value)
        cell.number_format = fmt
        if is_input:
            cell.font = INPUT_FONT
        if bold:
            cell.font = Font(bold=True, color=(INPUT_FONT.color.rgb if is_input else "000000"))
        if border:
            cell.border = border
        return cell

    # ----------------------------------------------------------------------
    def _run_all(self):
        self.build_assumptions()
        self.build_segments()
        self.build_ppa()
        self.build_schedules()
        self.build_income_statement()
        self.build_balance_sheet()
        self.build_cash_flow()
        self.build_checks()
        self.build_cover()

    def build(self):
        # pass 1: register all rows on all sheets (labels only)
        self.layout_only = True
        self._run_all()
        # pass 2: write formulas/values now that every row is known
        self.layout_only = False
        self._run_all()
        self.finalize_layout()
        self.wb.save(OUTFILE)

    # ====================================================================
    # ASSUMPTIONS
    # ====================================================================
    def build_assumptions(self):
        s = S_ASSUM
        ws = self.ws[s]
        self.put(s, "_t", "ASSUMPTIONS & DRIVERS  (US$ millions unless noted; blue = input)")
        ws.cell(row=1, column=1).font = TITLE_FONT
        # year header row
        self.put(s, "_hdr", "")
        for y in YEARS:
            c = self.setval(s, "_hdr", yl(y), YEAR_LABEL[y], fmt='@', bold=True)
            c.fill = header_fill(); c.font = WHITE_BOLD
            c.alignment = Alignment(horizontal="right")

        def yrow(key, label, values, fmt=FMT_PCT, indent=1, note=None):
            self.put(s, key, label, indent=indent, fmt=fmt, note=note)
            for y, v in zip(YEARS, values):
                if v is None:
                    continue
                self.setval(s, key, yl(y), v, fmt=fmt, is_input=True)

        def single(key, label, value, fmt=FMT_PCT, note=None):
            self.put(s, key, label, indent=1, fmt=fmt, note=note)
            self.setval(s, key, "C", value, fmt=fmt, is_input=True)

        # Operating segments ----------------------------------------------
        # The combined entity is modelled as three reporting segments:
        #   Space        = launch services (Falcon/Starship) + Starshield
        #   Starlink     = satellite connectivity
        #   xAI-Cursor   = AI segment (xAI/Grok + acquired Cursor) incl. synergies
        self.put(s, "_seg", "Operating segments — revenue & profitability", section=True)
        self.put(s, "_sp", "  Space (launch services + Starshield)", indent=0, bold=True)
        self.put(s, "space_rev_base", "    Revenue, FY2025 ($M)", indent=1)
        self.setval(s, "space_rev_base", "C", 6000, fmt=FMT_USD, is_input=True)
        yrow("space_g", "    Revenue growth %", [None, 0.28, 0.30, 0.28, 0.24, 0.20])
        yrow("space_ebitda", "    EBITDA margin %", [-0.08, 0.00, 0.08, 0.14, 0.19, 0.24])

        self.put(s, "_sl", "  Starlink (satellite connectivity)", indent=0, bold=True)
        self.put(s, "star_rev_base", "    Revenue, FY2025 ($M)", indent=1)
        self.setval(s, "star_rev_base", "C", 11400, fmt=FMT_USD, is_input=True)
        yrow("star_g", "    Revenue growth %", [None, 0.42, 0.36, 0.30, 0.25, 0.21])
        yrow("star_ebitda", "    EBITDA margin %", [0.45, 0.48, 0.51, 0.53, 0.55, 0.57])

        self.put(s, "_ai", "  xAI-Cursor (AI: xAI/Grok + Cursor)", indent=0, bold=True)
        self.put(s, "ai_rev_base", "    Organic revenue, FY2025 ($M) [xAI ~1,300 + Cursor ~450]", indent=1)
        self.setval(s, "ai_rev_base", "C", 1750, fmt=FMT_USD, is_input=True)
        yrow("ai_g", "    Organic revenue growth %", [None, 0.85, 0.70, 0.50, 0.38, 0.30])
        yrow("ai_ebitda", "    EBITDA margin % (pre-synergies)", [-0.60, -0.25, -0.02, 0.08, 0.16, 0.22])

        # Synergies / deal effects (attributed to the xAI-Cursor segment) ---
        self.put(s, "_syn", "Pro-forma deal effects (within xAI-Cursor segment)", section=True)
        yrow("syn_rev", "  Revenue synergies ($M) [Grok+Cursor cross-sell]", [0, 50, 200, 350, 450, 550], fmt=FMT_USD)
        yrow("syn_cost", "  Cost synergies ($M) [in-source inference on xAI/Starlink]", [0, 100, 350, 550, 700, 850], fmt=FMT_USD)
        yrow("integ", "  Integration & one-time deal costs ($M)", [0, 500, 0, 0, 0, 0], fmt=FMT_USD)

        # Capital / financing ---------------------------------------------
        self.put(s, "_cap", "Capital, financing & other", section=True)
        yrow("capex", "  Capex % of combined revenue", [0.260, 0.230, 0.210, 0.190, 0.175, 0.165])
        single("dep_rate", "  Depreciation % of beginning net PP&E", 0.090)
        # depreciation allocation across segments (capital intensity), sums to 100%
        single("dep_w_space", "  D&A allocation — Space %", 0.30)
        single("dep_w_star", "  D&A allocation — Starlink %", 0.50)
        single("dep_w_ai", "  D&A allocation — xAI-Cursor %", 0.20)
        yrow("sbc", "  Stock-based comp % of revenue", [0.060, 0.060, 0.055, 0.050, 0.045, 0.040])
        yrow("debt_draw", "  Net new long-term debt ($M)", [0, 2000, 2000, 1500, 1000, 1000], fmt=FMT_USD)
        yrow("spec_repay", "  Spectrum obligation repayment ($M)", [0, -2000, -2000, -2000, -2000, -2000], fmt=FMT_USD)
        yrow("equity_raise", "  Primary equity issuance / IPO proceeds ($M)", [0, 50000, 0, 0, 0, 0], fmt=FMT_USD,
             note="SpaceX June-2026 IPO")
        single("int_rate", "  Interest rate on debt & spectrum obligation", 0.065)
        single("int_inc", "  Interest income rate on cash", 0.040)
        single("tax_rate", "  Statutory tax rate", 0.210)
        single("nol_open", "  Opening NOL carryforward, FY2025 ($M)", 26000, fmt=FMT_USD,
               note="SpaceX accumulated + Cursor")
        single("nol_limit", "  NOL usage limit (% of taxable income)", 0.80)
        single("shares_pre", "  SpaceX shares pre-deal (M)", 2000.0, fmt=FMT_USD1)
        single("share_px", "  SpaceX share price, deal pricing ($)", 450.0, fmt=FMT_X)

        # Working-capital ratios (calibrated to FY2025 opening B/S) --------
        # Only recurring operating working capital scales with revenue; large
        # structural balances (LT deferred revenue, other LT items) are held flat.
        self.put(s, "_wc", "Working capital ratios (% of revenue)", section=True)
        single("ar_pct", "  Accounts receivable (% of revenue)", round(2500/18674, 4))
        single("inv_pct", "  Inventory (% of revenue)", round(2800/18674, 4))
        single("prep_pct", "  Prepaid & other current assets (% of revenue)", round(1500/18674, 4))
        single("ap_pct", "  Accounts payable (% of revenue)", round(2800/18674, 4))
        single("accr_pct", "  Accrued liabilities (% of revenue)", round(1700/18674, 4))
        single("defc_pct", "  Deferred revenue, current (% of revenue)", round(2800/18674, 4))
        single("cust_pct", "  Customer deposits & advances (% of revenue)", round(1900/18674, 4))

    # ====================================================================
    # SEGMENTS — revenue, EBITDA and EPS-contribution bridge
    # ====================================================================
    # (key, label, growth_assum, margin_assum, dep_weight_assum)
    SEGMENTS = [
        ("space", "Space", "space_g", "space_ebitda", "dep_w_space"),
        ("star", "Starlink", "star_g", "star_ebitda", "dep_w_star"),
        ("ai", "xAI-Cursor", "ai_g", "ai_ebitda", "dep_w_ai"),
    ]

    def build_segments(self):
        s = S_SEG
        self.put(s, "_t", "OPERATING SEGMENTS — revenue, EBITDA & EPS contribution (US$M)")
        self.ws[s].cell(row=1, column=1).font = TITLE_FONT
        self.put(s, "_n", "xAI-Cursor combines xAI/Grok with the acquired Cursor business and carries all deal synergies, integration costs and acquired-intangible amortization.")
        self.ws[s].cell(row=self.r(s, "_n"), column=1).font = SUB_FONT
        self._year_header(s)

        # ---- Revenue by segment ----
        self.put(s, "_rev", "Revenue by segment", section=True)
        self.put(s, "space_rev", "  Space", indent=1)
        self.put(s, "star_rev", "  Starlink", indent=1)
        self.put(s, "ai_rev_org", "  xAI-Cursor — organic", indent=1)
        self.put(s, "ai_rev_syn", "  xAI-Cursor — revenue synergies", indent=1)
        self.put(s, "ai_rev", "  xAI-Cursor — total", indent=1)
        self.put(s, "rev_tot", "  Total revenue", bold=True)

        # ---- EBITDA by segment ----
        self.put(s, "_ebd", "EBITDA by segment", section=True)
        self.put(s, "space_ebitda", "  Space", indent=1)
        self.put(s, "star_ebitda", "  Starlink", indent=1)
        self.put(s, "ai_ebitda", "  xAI-Cursor (incl. cost synergies less integration)", indent=1)
        self.put(s, "ebitda_tot", "  Total EBITDA", bold=True)
        self.put(s, "ebitda_m", "  Total EBITDA margin %", indent=1, fmt=FMT_PCT)

        # ---- per-segment bridge to net income & EPS ----
        self.put(s, "_eps", "Segment contribution to net income & EPS", section=True)
        self.put(s, "_method", "  D&A by capital-intensity weights; net interest & tax allocated by revenue share.", indent=1)
        self.ws[s].cell(row=self.r(s, "_method"), column=1).font = SUB_FONT
        for k, label, _g, _m, _w in self.SEGMENTS:
            self.put(s, f"{k}_hd", f"  {label}", bold=True)
            self.put(s, f"{k}_b_ebitda", "     EBITDA", indent=1)
            self.put(s, f"{k}_b_dep", "     Less: depreciation", indent=1)
            self.put(s, f"{k}_b_amort", "     Less: intangible amortization", indent=1)
            self.put(s, f"{k}_b_ebit", "     EBIT", indent=1, bold=True)
            self.put(s, f"{k}_b_nint", "     Net interest (alloc.)", indent=1)
            self.put(s, f"{k}_b_pretax", "     Pre-tax income", indent=1)
            self.put(s, f"{k}_b_tax", "     Income tax (alloc.)", indent=1)
            self.put(s, f"{k}_b_ni", "     Net income contribution", indent=1, bold=True)
            self.put(s, f"{k}_b_eps", "     EPS contribution ($)", indent=1, fmt=FMT_X)
        self.put(s, "ni_tot", "  Total net income (check vs IS)", bold=True)
        self.put(s, "eps_tot", "  Total diluted EPS ($)", bold=True, fmt=FMT_X)

        for y in YEARS:
            col = yl(y)
            prevcol = yl(YEARS[YEARS.index(y) - 1]) if y != 2025 else None
            # revenue
            for k in ("space", "star"):
                base = self.aref1(f"{k}_rev_base")
                if y == 2025:
                    self.setval(s, f"{k}_rev", col, f"={base}")
                else:
                    self.setval(s, f"{k}_rev", col, f"='{s}'!{prevcol}{self.r(s, f'{k}_rev')}*(1+{self.aref(f'{k}_g', y)})")
            if y == 2025:
                self.setval(s, "ai_rev_org", col, f"={self.aref1('ai_rev_base')}")
            else:
                self.setval(s, "ai_rev_org", col, f"='{s}'!{prevcol}{self.r(s,'ai_rev_org')}*(1+{self.aref('ai_g', y)})")
            self.setval(s, "ai_rev_syn", col, f"={self.aref('syn_rev', y)}")
            self.setval(s, "ai_rev", col, f"={self.yref(s,'ai_rev_org',y)}+{self.yref(s,'ai_rev_syn',y)}")
            self.setval(s, "rev_tot", col,
                        f"={self.yref(s,'space_rev',y)}+{self.yref(s,'star_rev',y)}+{self.yref(s,'ai_rev',y)}", bold=True)

            # EBITDA
            self.setval(s, "space_ebitda", col, f"={self.yref(s,'space_rev',y)}*{self.aref('space_ebitda', y)}")
            self.setval(s, "star_ebitda", col, f"={self.yref(s,'star_rev',y)}*{self.aref('star_ebitda', y)}")
            self.setval(s, "ai_ebitda", col,
                        f"={self.yref(s,'ai_rev',y)}*{self.aref('ai_ebitda', y)}+{self.aref('syn_cost', y)}-{self.aref('integ', y)}")
            self.setval(s, "ebitda_tot", col,
                        f"={self.yref(s,'space_ebitda',y)}+{self.yref(s,'star_ebitda',y)}+{self.yref(s,'ai_ebitda',y)}", bold=True)
            self.setval(s, "ebitda_m", col, f"={self.yref(s,'ebitda_tot',y)}/{self.yref(s,'rev_tot',y)}", fmt=FMT_PCT)

            # bridge to net income & EPS (per segment)
            net_int = f"({self.yref(S_IS,'int_exp',y)}+{self.yref(S_IS,'int_inc',y)})"
            tot_tax = f"({self.yref(S_IS,'tax_cur',y)}+{self.yref(S_IS,'tax_def',y)})"
            for k, label, _g, _m, w in self.SEGMENTS:
                rev_share = f"({self.yref(s, ('ai_rev' if k=='ai' else f'{k}_rev'), y)}/{self.yref(s,'rev_tot',y)})"
                self.setval(s, f"{k}_b_ebitda", col, f"={self.yref(s, ('ai_ebitda' if k=='ai' else f'{k}_ebitda'), y)}")
                self.setval(s, f"{k}_b_dep", col, f"=-{self.yref(S_IS,'dep',y)}*{self.aref1(w)}")
                # acquired-intangible amortization sits entirely in xAI-Cursor
                if k == "ai":
                    self.setval(s, f"{k}_b_amort", col, f"=-{self.yref(S_IS,'amort',y)}")
                else:
                    self.setval(s, f"{k}_b_amort", col, 0, fmt=FMT_USD)
                self.setval(s, f"{k}_b_ebit", col,
                            f"={self.ref(s, f'{k}_b_ebitda', col)}+{self.ref(s, f'{k}_b_dep', col)}+{self.ref(s, f'{k}_b_amort', col)}", bold=True)
                self.setval(s, f"{k}_b_nint", col, f"={net_int}*{rev_share}")
                self.setval(s, f"{k}_b_pretax", col, f"={self.ref(s, f'{k}_b_ebit', col)}+{self.ref(s, f'{k}_b_nint', col)}")
                self.setval(s, f"{k}_b_tax", col, f"={tot_tax}*{rev_share}")
                self.setval(s, f"{k}_b_ni", col, f"={self.ref(s, f'{k}_b_pretax', col)}+{self.ref(s, f'{k}_b_tax', col)}", bold=True)
                self.setval(s, f"{k}_b_eps", col, f"={self.ref(s, f'{k}_b_ni', col)}/{self.yref(S_IS,'shares',y)}", fmt=FMT_X)
            self.setval(s, "ni_tot", col,
                        "=" + "+".join(self.ref(s, f"{k}_b_ni", col) for k, *_ in self.SEGMENTS), bold=True)
            self.setval(s, "eps_tot", col,
                        "=" + "+".join(self.ref(s, f"{k}_b_eps", col) for k, *_ in self.SEGMENTS), bold=True, fmt=FMT_X)

    # ====================================================================
    # DEAL & PURCHASE PRICE ALLOCATION
    # ====================================================================
    def build_ppa(self):
        s = S_PPA
        self.put(s, "_t", "DEAL STRUCTURE & PURCHASE PRICE ALLOCATION (US$M)")
        self.ws[s].cell(row=1, column=1).font = TITLE_FONT
        self.put(s, "_n1", "All-stock acquisition announced 16-Jun-2026; expected close Q3-2026. Modelled as effective 1-Jan-2026 for full-year pro-forma comparability.")
        self.ws[s].cell(row=self.r(s, "_n1"), column=1).font = SUB_FONT

        self.put(s, "_c", "Consideration", section=True)
        self.put(s, "consid", "  Equity purchase price (100% stock)", indent=1)
        self.setval(s, "consid", "C", 60000, is_input=True)
        self.put(s, "px", "  SpaceX share price at pricing ($)")
        self.setval(s, "px", "C", f"={self.aref1('share_px')}", fmt=FMT_X)
        self.put(s, "newsh", "  New SpaceX shares issued (M)")
        self.setval(s, "newsh", "C", f"={self.ref(s,'consid','C')}/{self.ref(s,'px','C')}", fmt=FMT_USD1)

        self.put(s, "_fv", "Fair value of identifiable net assets acquired", section=True)
        self.put(s, "fv_cash", "  Cash & equivalents", indent=1)
        self.setval(s, "fv_cash", "C", 3000, is_input=True)
        self.put(s, "fv_ar", "  Accounts receivable", indent=1)
        self.setval(s, "fv_ar", "C", 250, is_input=True)
        self.put(s, "fv_prep", "  Prepaid & other current assets", indent=1)
        self.setval(s, "fv_prep", "C", 150, is_input=True)
        self.put(s, "fv_ppe", "  Property & equipment", indent=1)
        self.setval(s, "fv_ppe", "C", 300, is_input=True)
        self.put(s, "fv_liab", "  Less: liabilities assumed", indent=1)
        self.setval(s, "fv_liab", "C", -800, is_input=True)
        self.put(s, "fv_tang", "  Net tangible assets", bold=True)
        self.setval(s, "fv_tang", "C",
                    f"={self.ref(s,'fv_cash','C')}+{self.ref(s,'fv_ar','C')}+{self.ref(s,'fv_prep','C')}+{self.ref(s,'fv_ppe','C')}+{self.ref(s,'fv_liab','C')}",
                    bold=True)

        self.put(s, "_int", "Identifiable intangible assets", section=True)
        self.put(s, "int_dt", "  Developed technology (IDE + agent models)", indent=1)
        self.setval(s, "int_dt", "C", 8000, is_input=True)
        self.put(s, "life_dt", "    Useful life (yrs)", indent=1)
        self.setval(s, "life_dt", "C", 8, fmt='0', is_input=True)
        self.put(s, "int_cr", "  Customer relationships (enterprise contracts)", indent=1)
        self.setval(s, "int_cr", "C", 5000, is_input=True)
        self.put(s, "life_cr", "    Useful life (yrs)", indent=1)
        self.setval(s, "life_cr", "C", 10, fmt='0', is_input=True)
        self.put(s, "int_tn", "  Trade name (Cursor)", indent=1)
        self.setval(s, "int_tn", "C", 2000, is_input=True)
        self.put(s, "life_tn", "    Useful life (yrs)", indent=1)
        self.setval(s, "life_tn", "C", 15, fmt='0', is_input=True)
        self.put(s, "int_tot", "  Total identifiable intangibles", bold=True)
        self.setval(s, "int_tot", "C", f"={self.ref(s,'int_dt','C')}+{self.ref(s,'int_cr','C')}+{self.ref(s,'int_tn','C')}", bold=True)
        self.put(s, "amort", "  Annual intangible amortization", bold=True)
        self.setval(s, "amort", "C",
                    f"={self.ref(s,'int_dt','C')}/{self.ref(s,'life_dt','C')}+{self.ref(s,'int_cr','C')}/{self.ref(s,'life_cr','C')}+{self.ref(s,'int_tn','C')}/{self.ref(s,'life_tn','C')}",
                    bold=True)

        self.put(s, "_g", "Goodwill & deferred tax", section=True)
        self.put(s, "dtl", "  Deferred tax liability on intangibles", indent=1)
        self.setval(s, "dtl", "C", f"={self.aref1('tax_rate')}*{self.ref(s,'int_tot','C')}")
        self.put(s, "fv_net", "  FV of identifiable net assets (incl. intangibles, net of DTL)", bold=True)
        self.setval(s, "fv_net", "C", f"={self.ref(s,'fv_tang','C')}+{self.ref(s,'int_tot','C')}-{self.ref(s,'dtl','C')}", bold=True)
        self.put(s, "gw", "  Goodwill", bold=True)
        self.setval(s, "gw", "C", f"={self.ref(s,'consid','C')}-{self.ref(s,'fv_net','C')}", bold=True, border=TOP_BORDER)

    # ====================================================================
    # SUPPORTING SCHEDULES
    # ====================================================================
    def build_schedules(self):
        s = S_SCH
        self.put(s, "_t", "SUPPORTING SCHEDULES (US$M)")
        self.ws[s].cell(row=1, column=1).font = TITLE_FONT
        self._year_header(s)

        # combined revenue / cost references come from Income Statement
        # PP&E roll
        self.put(s, "_ppe", "Property, plant & equipment", section=True)
        self.put(s, "ppe_beg", "  Beginning net PP&E", indent=1)
        self.put(s, "capex", "  Capital expenditures", indent=1)
        self.put(s, "dep", "  Depreciation", indent=1)
        self.put(s, "ppe_end", "  Ending net PP&E", bold=True)

        # Intangibles & DTL roll
        self.put(s, "_intan", "Acquired intangibles & deferred tax", section=True)
        self.put(s, "int_beg", "  Beginning intangibles, net", indent=1)
        self.put(s, "int_am", "  Amortization", indent=1)
        self.put(s, "int_end", "  Ending intangibles, net", bold=True)
        self.put(s, "dtl_beg", "  Beginning deferred tax liability", indent=1)
        self.put(s, "dtl_chg", "  Change in DTL (deferred tax benefit)", indent=1)
        self.put(s, "dtl_end", "  Ending deferred tax liability", bold=True)

        # Debt & spectrum
        self.put(s, "_debt", "Debt & spectrum obligation", section=True)
        self.put(s, "debt_beg", "  Beginning long-term debt", indent=1)
        self.put(s, "debt_draw", "  Net draws / (repayments)", indent=1)
        self.put(s, "debt_end", "  Ending long-term debt", bold=True)
        self.put(s, "spec_beg", "  Beginning spectrum obligation", indent=1)
        self.put(s, "spec_repay", "  Repayments", indent=1)
        self.put(s, "spec_end", "  Ending spectrum obligation", bold=True)
        self.put(s, "int_exp", "  Interest expense (on beginning balances)", bold=True)

        # NOL / tax
        self.put(s, "_tax", "Tax & NOL carryforward", section=True)
        self.put(s, "pretax", "  Pre-tax income (from IS)", indent=1)
        self.put(s, "taxbase", "  Taxable income (add back non-deductible amort)", indent=1)
        self.put(s, "nol_beg", "  Beginning NOL balance", indent=1)
        self.put(s, "nol_use", "  NOL utilised", indent=1)
        self.put(s, "nol_new", "  NOL generated", indent=1)
        self.put(s, "nol_end", "  Ending NOL balance", bold=True)
        self.put(s, "cur_tax", "  Current income tax", bold=True)

        for y in YEARS:
            col = yl(y)
            prev = yl(YEARS[YEARS.index(y) - 1]) if y != 2025 else None
            rev_is = self.yref(S_IS, "rev_tot", y)
            # ---- PP&E ----
            if y == 2025:
                self.setval(s, "ppe_beg", col, 46000, fmt=FMT_USD)
                self.setval(s, "capex", col, f"={self.aref('capex', y)}*{rev_is}")
                self.setval(s, "dep", col, f"={self.ref(s,'ppe_beg',col)}*{self.aref1('dep_rate')}")
                self.setval(s, "ppe_end", col, 46000, fmt=FMT_USD, bold=True)
            else:
                if y == 2026:
                    self.setval(s, "ppe_beg", col, f"={self.ref(S_BS,'ppe', BS_COL['Open'])}")
                else:
                    self.setval(s, "ppe_beg", col, f"='{s}'!{prev}{self.r(s,'ppe_end')}")
                self.setval(s, "capex", col, f"={self.aref('capex', y)}*{rev_is}")
                self.setval(s, "dep", col, f"={self.ref(s,'ppe_beg',col)}*{self.aref1('dep_rate')}")
                self.setval(s, "ppe_end", col, f"={self.ref(s,'ppe_beg',col)}+{self.ref(s,'capex',col)}-{self.ref(s,'dep',col)}", bold=True)

            # ---- intangibles & DTL (post-deal only) ----
            if y == 2025:
                for k in ["int_beg", "int_am", "int_end", "dtl_beg", "dtl_chg", "dtl_end"]:
                    self.setval(s, k, col, 0, fmt=FMT_USD)
            else:
                if y == 2026:
                    self.setval(s, "int_beg", col, f"={self.ref(S_PPA,'int_tot','C')}")
                    self.setval(s, "dtl_beg", col, f"={self.ref(S_PPA,'dtl','C')}")
                else:
                    self.setval(s, "int_beg", col, f"='{s}'!{prev}{self.r(s,'int_end')}")
                    self.setval(s, "dtl_beg", col, f"='{s}'!{prev}{self.r(s,'dtl_end')}")
                self.setval(s, "int_am", col, f"=MIN({self.ref(s,'int_beg',col)},{self.ref(S_PPA,'amort','C')})")
                self.setval(s, "int_end", col, f"={self.ref(s,'int_beg',col)}-{self.ref(s,'int_am',col)}", bold=True)
                self.setval(s, "dtl_chg", col, f"=-{self.aref1('tax_rate')}*{self.ref(s,'int_am',col)}")
                self.setval(s, "dtl_end", col, f"={self.ref(s,'dtl_beg',col)}+{self.ref(s,'dtl_chg',col)}", bold=True)

            # ---- debt & spectrum ----
            if y == 2025:
                self.setval(s, "debt_beg", col, 12000, fmt=FMT_USD)
                self.setval(s, "debt_draw", col, 0, fmt=FMT_USD)
                self.setval(s, "debt_end", col, 12000, fmt=FMT_USD, bold=True)
                self.setval(s, "spec_beg", col, 19600, fmt=FMT_USD)
                self.setval(s, "spec_repay", col, 0, fmt=FMT_USD)
                self.setval(s, "spec_end", col, 19600, fmt=FMT_USD, bold=True)
            else:
                if y == 2026:
                    self.setval(s, "debt_beg", col, f"={self.ref(S_BS,'debt', BS_COL['Open'])}")
                    self.setval(s, "spec_beg", col, f"={self.ref(S_BS,'spec', BS_COL['Open'])}")
                else:
                    self.setval(s, "debt_beg", col, f"='{s}'!{prev}{self.r(s,'debt_end')}")
                    self.setval(s, "spec_beg", col, f"='{s}'!{prev}{self.r(s,'spec_end')}")
                self.setval(s, "debt_draw", col, f"={self.aref('debt_draw', y)}")
                self.setval(s, "debt_end", col, f"={self.ref(s,'debt_beg',col)}+{self.ref(s,'debt_draw',col)}", bold=True)
                self.setval(s, "spec_repay", col, f"={self.aref('spec_repay', y)}")
                self.setval(s, "spec_end", col, f"={self.ref(s,'spec_beg',col)}+{self.ref(s,'spec_repay',col)}", bold=True)
            self.setval(s, "int_exp", col, f"=({self.ref(s,'debt_beg',col)}+{self.ref(s,'spec_beg',col)})*{self.aref1('int_rate')}", bold=True)

            # ---- tax / NOL ----
            self.setval(s, "pretax", col, f"={self.yref(S_IS,'pretax',y)}")
            self.setval(s, "taxbase", col, f"={self.ref(s,'pretax',col)}+{self.ref(s,'int_am',col)}")
            if y == 2025:
                self.setval(s, "nol_beg", col, f"={self.aref1('nol_open')}")
            elif y == 2026:
                self.setval(s, "nol_beg", col, f"='{s}'!{prev}{self.r(s,'nol_end')}")
            else:
                self.setval(s, "nol_beg", col, f"='{s}'!{prev}{self.r(s,'nol_end')}")
            self.setval(s, "nol_use", col,
                        f"=MIN({self.ref(s,'nol_beg',col)},{self.aref1('nol_limit')}*MAX(0,{self.ref(s,'taxbase',col)}))")
            self.setval(s, "nol_new", col, f"=MAX(0,-{self.ref(s,'taxbase',col)})")
            self.setval(s, "nol_end", col,
                        f"={self.ref(s,'nol_beg',col)}-{self.ref(s,'nol_use',col)}+{self.ref(s,'nol_new',col)}", bold=True)
            self.setval(s, "cur_tax", col,
                        f"={self.aref1('tax_rate')}*MAX(0,{self.ref(s,'taxbase',col)}-{self.ref(s,'nol_use',col)})", bold=True)

    # ====================================================================
    # PRO-FORMA INCOME STATEMENT
    # ====================================================================
    def build_income_statement(self):
        s = S_IS
        self.put(s, "_t", "PRO-FORMA COMBINED INCOME STATEMENT (US$M)")
        self.ws[s].cell(row=1, column=1).font = TITLE_FONT
        self.put(s, "_n", "FY2025 shown pro-forma for comparability; deal effective 1-Jan-2026 in the model.")
        self.ws[s].cell(row=self.r(s, "_n"), column=1).font = SUB_FONT
        self._year_header(s)

        self.put(s, "_revh", "Revenue by segment", section=True)
        self.put(s, "rev_space", "  Space", indent=1)
        self.put(s, "rev_star", "  Starlink", indent=1)
        self.put(s, "rev_ai", "  xAI-Cursor", indent=1)
        self.put(s, "rev_tot", "Total revenue", bold=True)
        self.put(s, "_ebdh", "EBITDA by segment", section=True)
        self.put(s, "ebitda_space", "  Space", indent=1)
        self.put(s, "ebitda_star", "  Starlink", indent=1)
        self.put(s, "ebitda_ai", "  xAI-Cursor (incl. synergies, integration)", indent=1)
        self.put(s, "ebitda", "Total EBITDA", bold=True)
        self.put(s, "ebitda_m", "  EBITDA margin %", indent=1, fmt=FMT_PCT)
        self.put(s, "_pnl", "Consolidated P&L", section=True)
        self.put(s, "dep", "Depreciation", indent=1)
        self.put(s, "amort", "Amortization of acquired intangibles", indent=1)
        self.put(s, "ebit", "EBIT (operating income)", bold=True)
        self.put(s, "ebit_m", "  Operating margin %", indent=1, fmt=FMT_PCT)
        self.put(s, "int_exp", "Interest expense", indent=1)
        self.put(s, "int_inc", "Interest income", indent=1)
        self.put(s, "pretax", "Pre-tax income (loss)", bold=True)
        self.put(s, "tax_cur", "Current income tax", indent=1)
        self.put(s, "tax_def", "Deferred income tax (benefit)", indent=1)
        self.put(s, "ni", "Net income (loss)", bold=True)
        self.put(s, "ni_m", "  Net margin %", indent=1, fmt=FMT_PCT)
        self.put(s, "shares", "Diluted shares outstanding (M)", indent=1, fmt=FMT_USD1)
        self.put(s, "eps", "Diluted EPS ($)", bold=True, fmt=FMT_X)
        self.put(s, "_epsh", "EPS contribution by segment ($)", section=True)
        self.put(s, "eps_space", "  Space", indent=1)
        self.put(s, "eps_star", "  Starlink", indent=1)
        self.put(s, "eps_ai", "  xAI-Cursor", indent=1)
        self.put(s, "eps_chk", "  Total (= diluted EPS)", bold=True, fmt=FMT_X)

        for y in YEARS:
            col = yl(y)
            self.setval(s, "rev_space", col, f"={self.yref(S_SEG,'space_rev',y)}")
            self.setval(s, "rev_star", col, f"={self.yref(S_SEG,'star_rev',y)}")
            self.setval(s, "rev_ai", col, f"={self.yref(S_SEG,'ai_rev',y)}")
            self.setval(s, "rev_tot", col, f"={self.yref(S_SEG,'rev_tot',y)}", bold=True)
            self.setval(s, "ebitda_space", col, f"={self.yref(S_SEG,'space_ebitda',y)}")
            self.setval(s, "ebitda_star", col, f"={self.yref(S_SEG,'star_ebitda',y)}")
            self.setval(s, "ebitda_ai", col, f"={self.yref(S_SEG,'ai_ebitda',y)}")
            self.setval(s, "ebitda", col, f"={self.yref(S_SEG,'ebitda_tot',y)}", bold=True)
            self.setval(s, "ebitda_m", col, f"={self.yref(s,'ebitda',y)}/{self.yref(s,'rev_tot',y)}", fmt=FMT_PCT)
            self.setval(s, "dep", col, f"={self.yref(S_SCH,'dep',y)}")
            self.setval(s, "amort", col, f"={self.yref(S_SCH,'int_am',y)}")
            self.setval(s, "ebit", col, f"={self.yref(s,'ebitda',y)}-{self.yref(s,'dep',y)}-{self.yref(s,'amort',y)}", bold=True)
            self.setval(s, "ebit_m", col, f"={self.yref(s,'ebit',y)}/{self.yref(s,'rev_tot',y)}", fmt=FMT_PCT)
            self.setval(s, "int_exp", col, f"=-{self.yref(S_SCH,'int_exp',y)}")
            # interest income on beginning cash
            if y == 2025:
                begcash = "12000"
            elif y == 2026:
                begcash = self.ref(S_BS, "cash", BS_COL["Open"])
            else:
                begcash = self.ref(S_BS, "cash", BS_PREV[y])
            self.setval(s, "int_inc", col, f"={begcash}*{self.aref1('int_inc')}")
            self.setval(s, "pretax", col, f"={self.yref(s,'ebit',y)}+{self.yref(s,'int_exp',y)}+{self.yref(s,'int_inc',y)}", bold=True)
            self.setval(s, "tax_cur", col, f"=-{self.yref(S_SCH,'cur_tax',y)}")
            # DTL release is a deferred-tax benefit (increases net income)
            self.setval(s, "tax_def", col, f"=-{self.yref(S_SCH,'dtl_chg',y)}")
            self.setval(s, "ni", col, f"={self.yref(s,'pretax',y)}+{self.yref(s,'tax_cur',y)}+{self.yref(s,'tax_def',y)}", bold=True)
            self.setval(s, "ni_m", col, f"={self.yref(s,'ni',y)}/{self.yref(s,'rev_tot',y)}", fmt=FMT_PCT)
            # shares
            if y == 2025:
                self.setval(s, "shares", col, f"={self.aref1('shares_pre')}+{self.ref(S_PPA,'newsh','C')}", fmt=FMT_USD1)
            else:
                self.setval(s, "shares", col, f"={self.aref1('shares_pre')}+{self.ref(S_PPA,'newsh','C')}", fmt=FMT_USD1)
            self.setval(s, "eps", col, f"={self.yref(s,'ni',y)}/{self.yref(s,'shares',y)}", fmt=FMT_X, bold=True)
            # EPS contribution by segment (from Segments tab)
            self.setval(s, "eps_space", col, f"={self.yref(S_SEG,'space_b_eps',y)}", fmt=FMT_X)
            self.setval(s, "eps_star", col, f"={self.yref(S_SEG,'star_b_eps',y)}", fmt=FMT_X)
            self.setval(s, "eps_ai", col, f"={self.yref(S_SEG,'ai_b_eps',y)}", fmt=FMT_X)
            self.setval(s, "eps_chk", col, f"={self.yref(S_SEG,'eps_tot',y)}", fmt=FMT_X, bold=True)

    # ====================================================================
    # BALANCE SHEET (with PPA bridge)
    # ====================================================================
    def build_balance_sheet(self):
        s = S_BS
        self.put(s, "_t", "PRO-FORMA COMBINED BALANCE SHEET (US$M)")
        self.ws[s].cell(row=1, column=1).font = TITLE_FONT
        # header
        self.put(s, "_hdr", "")
        labels = {"2025A": "SpaceX 2025A", "Cursor": "+ Cursor", "PPA": "+ Purch. acctg",
                  "Open": "PF Open 1/1/26", 2026: "2026E", 2027: "2027E",
                  2028: "2028E", 2029: "2029E", 2030: "2030E"}
        for k, colL in BS_COL.items():
            c = self.setval(s, "_hdr", colL, labels[k], fmt='@', bold=True)
            c.fill = header_fill(); c.font = WHITE_BOLD
            c.alignment = Alignment(horizontal="right", wrap_text=True)

        # ---- Assets ----
        self.put(s, "_a", "ASSETS", section=True)
        self.put(s, "cash", "Cash, equivalents & ST investments", indent=1)
        self.put(s, "ar", "Accounts receivable", indent=1)
        self.put(s, "inv", "Inventory", indent=1)
        self.put(s, "prep", "Prepaid & other current assets", indent=1)
        self.put(s, "ca", "Total current assets", bold=True)
        self.put(s, "ppe", "Property & equipment, net", indent=1)
        self.put(s, "intan", "Acquired intangibles, net", indent=1)
        self.put(s, "gw", "Goodwill", indent=1)
        self.put(s, "spec_int", "Spectrum & other intangibles", indent=1)
        self.put(s, "olta", "Other long-term assets", indent=1)
        self.put(s, "ta", "TOTAL ASSETS", bold=True)

        # ---- Liabilities ----
        self.put(s, "_l", "LIABILITIES", section=True)
        self.put(s, "ap", "Accounts payable", indent=1)
        self.put(s, "accr", "Accrued liabilities", indent=1)
        self.put(s, "defc", "Deferred revenue, current", indent=1)
        self.put(s, "cust", "Customer deposits & advances", indent=1)
        self.put(s, "cl", "Total current liabilities", bold=True)
        self.put(s, "debt", "Long-term debt", indent=1)
        self.put(s, "spec", "Spectrum acquisition obligation", indent=1)
        self.put(s, "ltdef", "LT deferred revenue & deposits", indent=1)
        self.put(s, "dtl", "Deferred tax liability", indent=1)
        self.put(s, "oltl", "Other long-term liabilities", indent=1)
        self.put(s, "tl", "TOTAL LIABILITIES", bold=True)

        # ---- Equity ----
        self.put(s, "_e", "EQUITY", section=True)
        self.put(s, "apic", "Common stock & additional paid-in capital", indent=1)
        self.put(s, "re", "Accumulated deficit", indent=1)
        self.put(s, "te", "TOTAL EQUITY", bold=True)
        self.put(s, "tle", "TOTAL LIABILITIES & EQUITY", bold=True)
        self.put(s, "chk", "Balance check (assets − L&E)", indent=1)

        # opening 2025A actual values
        open2025 = {"cash": 12000, "ar": 2500, "inv": 2800, "prep": 1500,
                    "ppe": 46000, "intan": 0, "gw": 0, "spec_int": 19600, "olta": 7700,
                    "ap": 2800, "accr": 1700, "defc": 2800, "cust": 1900,
                    "debt": 12000, "spec": 19600, "ltdef": 28000, "dtl": 0, "oltl": 20700,
                    "apic": 25000, "re": -22400}
        cursor_fv = {"cash": 3000, "ar": 250, "inv": 0, "prep": 150, "ppe": 300,
                     "intan": 0, "gw": 0, "spec_int": 0, "olta": 0,
                     "ap": 150, "accr": 150, "defc": 300, "cust": 0,
                     "debt": 0, "spec": 0, "ltdef": 200, "dtl": 0, "oltl": 0,
                     "apic": 2900, "re": 0}

        # 2025A column (C)
        for k, v in open2025.items():
            self.setval(s, k, "C", v, fmt=FMT_USD)
        # Cursor column (D)
        for k, v in cursor_fv.items():
            self.setval(s, k, "D", v, fmt=FMT_USD)
        # PPA adjustments column (E)
        E = "E"
        self.setval(s, "intan", E, f"={self.ref(S_PPA,'int_tot','C')}", fmt=FMT_USD)
        self.setval(s, "gw", E, f"={self.ref(S_PPA,'gw','C')}", fmt=FMT_USD)
        self.setval(s, "dtl", E, f"={self.ref(S_PPA,'dtl','C')}", fmt=FMT_USD)
        # consideration to equity, eliminate Cursor's carried equity
        self.setval(s, "apic", E, f"={self.ref(S_PPA,'consid','C')}-{self.ref(s,'apic','D')}", fmt=FMT_USD)
        for k in ["cash", "ar", "inv", "prep", "ppe", "spec_int", "olta",
                  "ap", "accr", "defc", "cust", "debt", "spec", "ltdef", "oltl", "re"]:
            self.setval(s, k, E, 0, fmt=FMT_USD)

        # Open column (F) = C+D+E
        for k in open2025.keys():
            self.setval(s, k, "F", f"={self.ref(s,k,'C')}+{self.ref(s,k,'D')}+{self.ref(s,k,'E')}", fmt=FMT_USD)

        # subtotals/total formulas for bridge + projection columns
        for colL in ["C", "D", "E", "F", "G", "H", "I", "J", "K"]:
            self.setval(s, "ca", colL, f"={self.ref(s,'cash',colL)}+{self.ref(s,'ar',colL)}+{self.ref(s,'inv',colL)}+{self.ref(s,'prep',colL)}", bold=True, border=TOP_BORDER)
            self.setval(s, "ta", colL,
                        f"={self.ref(s,'ca',colL)}+{self.ref(s,'ppe',colL)}+{self.ref(s,'intan',colL)}+{self.ref(s,'gw',colL)}+{self.ref(s,'spec_int',colL)}+{self.ref(s,'olta',colL)}",
                        bold=True, border=TOP_BORDER)
            self.setval(s, "cl", colL, f"={self.ref(s,'ap',colL)}+{self.ref(s,'accr',colL)}+{self.ref(s,'defc',colL)}+{self.ref(s,'cust',colL)}", bold=True, border=TOP_BORDER)
            self.setval(s, "tl", colL,
                        f"={self.ref(s,'cl',colL)}+{self.ref(s,'debt',colL)}+{self.ref(s,'spec',colL)}+{self.ref(s,'ltdef',colL)}+{self.ref(s,'dtl',colL)}+{self.ref(s,'oltl',colL)}",
                        bold=True, border=TOP_BORDER)
            self.setval(s, "te", colL, f"={self.ref(s,'apic',colL)}+{self.ref(s,'re',colL)}", bold=True, border=TOP_BORDER)
            self.setval(s, "tle", colL, f"={self.ref(s,'tl',colL)}+{self.ref(s,'te',colL)}", bold=True, border=TOP_BORDER)
            self.setval(s, "chk", colL, f"={self.ref(s,'ta',colL)}-{self.ref(s,'tle',colL)}", fmt=FMT_USD1)

        # projection columns G..K (2026..2030)
        for y in YEARS:
            if y == 2025:
                continue
            col = BS_COL[y]
            prev = BS_PREV[y]
            rev = self.yref(S_IS, "rev_tot", y)
            # assets driven by revenue ratios
            self.setval(s, "cash", col, f"={self.yref(S_CF,'cash_end',y)}")
            self.setval(s, "ar", col, f"={rev}*{self.aref1('ar_pct')}")
            self.setval(s, "inv", col, f"={rev}*{self.aref1('inv_pct')}")
            self.setval(s, "prep", col, f"={rev}*{self.aref1('prep_pct')}")
            self.setval(s, "ppe", col, f"={self.yref(S_SCH,'ppe_end',y)}")
            self.setval(s, "intan", col, f"={self.yref(S_SCH,'int_end',y)}")
            self.setval(s, "gw", col, f"={self.ref(s,'gw',prev)}")          # constant
            self.setval(s, "spec_int", col, f"={self.ref(s,'spec_int',prev)}")  # flat
            self.setval(s, "olta", col, f"={self.ref(s,'olta',prev)}")      # flat
            # liabilities
            self.setval(s, "ap", col, f"={rev}*{self.aref1('ap_pct')}")
            self.setval(s, "accr", col, f"={rev}*{self.aref1('accr_pct')}")
            self.setval(s, "defc", col, f"={rev}*{self.aref1('defc_pct')}")
            self.setval(s, "cust", col, f"={rev}*{self.aref1('cust_pct')}")
            self.setval(s, "debt", col, f"={self.yref(S_SCH,'debt_end',y)}")
            self.setval(s, "spec", col, f"={self.yref(S_SCH,'spec_end',y)}")
            self.setval(s, "ltdef", col, f"={self.ref(s,'ltdef',prev)}")    # flat (structural)
            self.setval(s, "dtl", col, f"={self.yref(S_SCH,'dtl_end',y)}")
            self.setval(s, "oltl", col, f"={self.ref(s,'oltl',prev)}")      # flat
            # equity
            self.setval(s, "apic", col, f"={self.ref(s,'apic',prev)}+{self.yref(S_CF,'sbc',y)}+{self.aref('equity_raise', y)}")
            self.setval(s, "re", col, f"={self.ref(s,'re',prev)}+{self.yref(S_IS,'ni',y)}")

    # ====================================================================
    # CASH FLOW STATEMENT (2026E..2030E)
    # ====================================================================
    def build_cash_flow(self):
        s = S_CF
        self.put(s, "_t", "PRO-FORMA COMBINED CASH FLOW STATEMENT (US$M)")
        self.ws[s].cell(row=1, column=1).font = TITLE_FONT
        self.put(s, "_n", "Post-acquisition years. The all-stock purchase is non-cash and reflected in opening balances (acquired cash included).")
        self.ws[s].cell(row=self.r(s, "_n"), column=1).font = SUB_FONT
        self._year_header(s)

        self.put(s, "_o", "Operating activities", section=True)
        self.put(s, "ni", "Net income (loss)", indent=1)
        self.put(s, "dep", "Depreciation", indent=1)
        self.put(s, "amort", "Amortization of acquired intangibles", indent=1)
        self.put(s, "sbc", "Stock-based compensation", indent=1)
        self.put(s, "deftax", "Deferred income taxes", indent=1)
        self.put(s, "d_ar", "(Increase)/decrease in receivables", indent=1)
        self.put(s, "d_inv", "(Increase)/decrease in inventory", indent=1)
        self.put(s, "d_prep", "(Increase)/decrease in prepaids", indent=1)
        self.put(s, "d_ap", "Increase/(decrease) in payables", indent=1)
        self.put(s, "d_accr", "Increase/(decrease) in accrued liabilities", indent=1)
        self.put(s, "d_defc", "Increase/(decrease) in deferred revenue (cur.)", indent=1)
        self.put(s, "d_cust", "Increase/(decrease) in customer deposits", indent=1)
        self.put(s, "d_ltdef", "Increase/(decrease) in LT deferred revenue", indent=1)
        self.put(s, "cfo", "Cash from operations", bold=True)

        self.put(s, "_i", "Investing activities", section=True)
        self.put(s, "capex", "Capital expenditures", indent=1)
        self.put(s, "cfi", "Cash from investing", bold=True)

        self.put(s, "_f", "Financing activities", section=True)
        self.put(s, "debt", "Net debt draws / (repayments)", indent=1)
        self.put(s, "spec", "Spectrum obligation repayments", indent=1)
        self.put(s, "raise", "Primary equity issuance / IPO", indent=1)
        self.put(s, "cff", "Cash from financing", bold=True)

        self.put(s, "net", "Net change in cash", bold=True)
        self.put(s, "cash_beg", "Cash, beginning of period", indent=1)
        self.put(s, "cash_end", "Cash, end of period", bold=True)

        for y in YEARS:
            if y == 2025:
                continue
            col = yl(y)
            bcol = BS_COL[y]
            bprev = BS_PREV[y]

            def d(line_key):  # change in a BS line vs prior column
                return f"({self.ref(S_BS, line_key, bcol)}-{self.ref(S_BS, line_key, bprev)})"

            self.setval(s, "ni", col, f"={self.yref(S_IS,'ni',y)}")
            self.setval(s, "dep", col, f"={self.yref(S_SCH,'dep',y)}")
            self.setval(s, "amort", col, f"={self.yref(S_SCH,'int_am',y)}")
            self.setval(s, "sbc", col, f"={self.aref('sbc', y)}*{self.yref(S_IS,'rev_tot',y)}")
            self.setval(s, "deftax", col, f"={self.yref(S_SCH,'dtl_chg',y)}")
            self.setval(s, "d_ar", col, f"=-{d('ar')}")
            self.setval(s, "d_inv", col, f"=-{d('inv')}")
            self.setval(s, "d_prep", col, f"=-{d('prep')}")
            self.setval(s, "d_ap", col, f"={d('ap')}")
            self.setval(s, "d_accr", col, f"={d('accr')}")
            self.setval(s, "d_defc", col, f"={d('defc')}")
            self.setval(s, "d_cust", col, f"={d('cust')}")
            self.setval(s, "d_ltdef", col, f"={d('ltdef')}")
            cfo_terms = ["ni", "dep", "amort", "sbc", "deftax", "d_ar", "d_inv",
                         "d_prep", "d_ap", "d_accr", "d_defc", "d_cust", "d_ltdef"]
            self.setval(s, "cfo", col, "=" + "+".join(self.ref(s, k, col) for k in cfo_terms), bold=True)

            self.setval(s, "capex", col, f"=-{self.yref(S_SCH,'capex',y)}")
            self.setval(s, "cfi", col, f"={self.ref(s,'capex',col)}", bold=True)

            self.setval(s, "debt", col, f"={self.yref(S_SCH,'debt_draw',y)}")
            self.setval(s, "spec", col, f"={self.yref(S_SCH,'spec_repay',y)}")
            self.setval(s, "raise", col, f"={self.aref('equity_raise', y)}")
            self.setval(s, "cff", col, f"={self.ref(s,'debt',col)}+{self.ref(s,'spec',col)}+{self.ref(s,'raise',col)}", bold=True)

            self.setval(s, "net", col, f"={self.ref(s,'cfo',col)}+{self.ref(s,'cfi',col)}+{self.ref(s,'cff',col)}", bold=True)
            if y == 2026:
                self.setval(s, "cash_beg", col, f"={self.ref(S_BS,'cash',BS_COL['Open'])}")
            else:
                self.setval(s, "cash_beg", col, f"='{s}'!{yl(YEARS[YEARS.index(y)-1])}{self.r(s,'cash_end')}")
            self.setval(s, "cash_end", col, f"={self.ref(s,'cash_beg',col)}+{self.ref(s,'net',col)}", bold=True, border=TOP_BORDER)

    # ====================================================================
    # CHECKS
    # ====================================================================
    def build_checks(self):
        s = S_CHK
        self.put(s, "_t", "MODEL INTEGRITY CHECKS (should all be ~0)")
        self.ws[s].cell(row=1, column=1).font = TITLE_FONT
        self._year_header(s)
        self.put(s, "bs", "Balance sheet balances (A − L&E)", bold=True)
        self.put(s, "cf", "Cash flow ties to balance sheet cash", bold=True)
        self.put(s, "ppa", "PPA bridge balances (Open: A − L&E)", bold=True)
        self.put(s, "segni", "Segment net income sums to consolidated NI", bold=True)
        self.put(s, "segeps", "Segment EPS contributions sum to diluted EPS", bold=True, fmt=FMT_X)
        for y in YEARS:
            col = yl(y)
            if y == 2025:
                self.setval(s, "bs", col, f"={self.ref(S_BS,'chk','C')}", fmt=FMT_USD1)
                self.setval(s, "cf", col, "n/a", fmt='@')
            else:
                self.setval(s, "bs", col, f"={self.ref(S_BS,'chk',BS_COL[y])}", fmt=FMT_USD1)
                self.setval(s, "cf", col, f"={self.yref(S_CF,'cash_end',y)}-{self.ref(S_BS,'cash',BS_COL[y])}", fmt=FMT_USD1)
            self.setval(s, "segni", col, f"={self.yref(S_SEG,'ni_tot',y)}-{self.yref(S_IS,'ni',y)}", fmt=FMT_USD1)
            self.setval(s, "segeps", col, f"={self.yref(S_SEG,'eps_tot',y)}-{self.yref(S_IS,'eps',y)}", fmt='0.0000')
        self.setval(s, "ppa", "C", f"={self.ref(S_BS,'chk','F')}", fmt=FMT_USD1)

    # ====================================================================
    # COVER
    # ====================================================================
    def build_cover(self):
        s = S_COVER
        ws = self.ws[s]
        ws.cell(row=2, column=2, value="SpaceX  ×  Cursor").font = Font(bold=True, size=24, color=NAVY)
        ws.cell(row=3, column=2, value="Pro-Forma Three-Statement Model — Segmented").font = Font(bold=True, size=16, color=BLUE)
        lines = [
            "",
            "Transaction: SpaceX acquires Cursor (Anysphere, Inc.) — announced 16-Jun-2026",
            "Consideration: ~US$60bn, 100% SpaceX Class A stock; expected close Q3-2026",
            "Strategic rationale: AI / developer-tools capability following the Feb-2026 xAI combination",
            "Segments modelled: Space (launch + Starshield), Starlink, and xAI-Cursor (xAI/Grok + Cursor)",
            "",
            "Contents:",
            "   • Assumptions — all input drivers (blue cells are editable inputs)",
            "   • Segments — Space, Starlink & xAI-Cursor: revenue, EBITDA & EPS contribution",
            "   • Deal & PPA — purchase price allocation, goodwill, intangibles",
            "   • Schedules — PP&E, intangibles/DTL, debt & spectrum, tax/NOL",
            "   • Income Statement — segment-driven P&L with EPS contribution by segment",
            "   • Balance Sheet — purchase-accounting bridge + projections",
            "   • Cash Flow — integrated cash flow statement",
            "   • Checks — balance, cash-flow and segment EPS integrity checks",
            "",
            "All figures in US$ millions unless stated. The workbook is fully formula-linked:",
            "change any blue input on the Assumptions tab and all three statements recalculate.",
            "",
            "DISCLAIMER: Illustrative model for analytical purposes only. Built from public",
            "estimates and reasonable assumptions; not investment advice and not company-verified.",
        ]
        r = 5
        for ln in lines:
            c = ws.cell(row=r, column=2, value=ln)
            if ln.startswith("Contents"):
                c.font = Font(bold=True)
            elif ln.startswith("DISCLAIMER"):
                c.font = Font(italic=True, size=9, color="808080")
            else:
                c.font = Font(size=10)
            r += 1
        ws.sheet_view.showGridLines = False

    # ====================================================================
    # layout helpers
    # ====================================================================
    def _year_header(self, sheet):
        self.put(sheet, "_hdr", "")
        for y in YEARS:
            c = self.setval(sheet, "_hdr", yl(y), YEAR_LABEL[y], fmt='@', bold=True)
            c.fill = header_fill(); c.font = WHITE_BOLD
            c.alignment = Alignment(horizontal="right")

    def finalize_layout(self):
        for title, ws in self.ws.items():
            ws.column_dimensions["A"].width = 46
            ws.column_dimensions["B"].width = 2
            for col in "CDEFGHIJK":
                ws.column_dimensions[col].width = 13
            ws.column_dimensions["L"].width = 22
            ws.freeze_panes = "C3" if title not in (S_COVER,) else None
            ws.sheet_view.showGridLines = False


def _col_idx(col_letter):
    idx = 0
    for ch in col_letter:
        idx = idx * 26 + (ord(ch.upper()) - ord("A") + 1)
    return idx


if __name__ == "__main__":
    mb = ModelBuilder()
    mb.build()
    print(f"Wrote {OUTFILE}")
