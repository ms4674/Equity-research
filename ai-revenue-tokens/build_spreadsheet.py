"""Build an Excel workbook + bar-chart PNGs from the AI revenue / token CSVs.

Outputs:
    ai_revenue_tokens.xlsx   - multi-sheet workbook with raw + pivoted data
                               and embedded bar charts
    charts/revenue_bar.png   - grouped bar chart of ARR by snapshot date
    charts/tokens_bar.png    - grouped bar chart of monthly tokens by date

Requires: pandas, matplotlib, openpyxl (`pip install -r requirements.txt`).
"""
from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment

HERE = Path(__file__).resolve().parent
DATA_DIR = HERE / "data"
CHART_DIR = HERE / "charts"
OUT_XLSX = HERE / "ai_revenue_tokens.xlsx"

REVENUE_SNAPSHOTS = [
    ("2022-12-31", "Dec-22"),
    ("2023-12-31", "Dec-23"),
    ("2024-06-30", "Jun-24"),
    ("2024-12-31", "Dec-24"),
    ("2025-06-30", "Jun-25"),
    ("2025-12-31", "Dec-25"),
    ("2026-04-15", "Apr-26"),
]

TOKEN_SNAPSHOTS = [
    ("2023-10-31", "Oct-23"),
    ("2024-12-31", "Dec-24"),
    ("2025-06-30", "Jun-25"),
    ("2025-10-31", "Oct-25"),
    ("2026-03-31", "Mar-26"),
]

COMPANY_ORDER = [
    "OpenAI",
    "Anthropic",
    "Harvey",
    "Sierra",
    "Decagon",
    "Intercom Fin",
]

TOKEN_COMPANY_ORDER = [
    "OpenAI",
    "Anthropic",
    "Google (reference)",
    "Microsoft Foundry (reference)",
    "Together.ai (reference)",
    "OpenRouter (reference)",
]


def _snapshot_value(df: pd.DataFrame, company: str, as_of: pd.Timestamp, value_col: str) -> float | None:
    """Most-recent observation for *company* on or before *as_of*."""
    sub = df[(df["company"] == company) & (df["as_of_date"] <= as_of)]
    if sub.empty:
        return None
    return float(sub.sort_values("as_of_date").iloc[-1][value_col])


def build_revenue_pivot() -> pd.DataFrame:
    df = pd.read_csv(DATA_DIR / "revenue_arr_timeseries.csv", parse_dates=["as_of_date"])
    rows = []
    for company in COMPANY_ORDER:
        row = {"Company": company}
        for iso_date, label in REVENUE_SNAPSHOTS:
            row[label] = _snapshot_value(df, company, pd.Timestamp(iso_date), "arr_usd_millions")
        rows.append(row)
    return pd.DataFrame(rows)


def build_token_pivot() -> pd.DataFrame:
    df = pd.read_csv(DATA_DIR / "monthly_token_consumption.csv", parse_dates=["as_of_date"])
    rows = []
    for company in TOKEN_COMPANY_ORDER:
        row = {"Provider": company}
        for iso_date, label in TOKEN_SNAPSHOTS:
            row[label] = _snapshot_value(df, company, pd.Timestamp(iso_date), "tokens_per_month_trillions")
        rows.append(row)
    return pd.DataFrame(rows)


def _grouped_bar(
    pivot: pd.DataFrame,
    label_col: str,
    title: str,
    ylabel: str,
    out: Path,
    log: bool = True,
) -> None:
    period_cols = [c for c in pivot.columns if c != label_col]
    n_periods = len(period_cols)
    n_groups = len(pivot)
    width = 0.8 / n_periods
    x = np.arange(n_groups)

    fig, ax = plt.subplots(figsize=(13, 7))
    cmap = plt.colormaps.get_cmap("viridis").resampled(n_periods)
    for i, col in enumerate(period_cols):
        vals = pd.to_numeric(pivot[col], errors="coerce").fillna(0).values
        bars = ax.bar(x + i * width - 0.4 + width / 2, vals, width=width, label=col, color=cmap(i))
        for bar, v in zip(bars, vals):
            if v > 0:
                ax.annotate(
                    f"{v:,.0f}" if v >= 10 else f"{v:,.1f}",
                    xy=(bar.get_x() + bar.get_width() / 2, v),
                    xytext=(0, 2),
                    textcoords="offset points",
                    ha="center",
                    va="bottom",
                    fontsize=7,
                    rotation=90,
                )

    ax.set_xticks(x)
    ax.set_xticklabels(pivot[label_col], rotation=15, ha="right")
    ax.set_ylabel(ylabel)
    ax.set_title(title)
    if log:
        ax.set_yscale("log")
        ax.set_ylim(bottom=0.5)
    ax.grid(axis="y", which="both", alpha=0.3)
    ax.legend(title="Snapshot", loc="upper right", ncol=2, fontsize=9)
    fig.tight_layout()
    fig.savefig(out, dpi=160)
    plt.close(fig)
    print(f"wrote {out}")


def _write_df_to_sheet(ws, df: pd.DataFrame, title: str) -> None:
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    header_row_idx = ws.max_row + 1
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[header_row_idx]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for col_letter in [chr(ord("A") + i) for i in range(len(df.columns))]:
        ws.column_dimensions[col_letter].width = 18
    ws.column_dimensions["A"].width = 28
    for row in ws.iter_rows(min_row=header_row_idx + 1, max_col=len(df.columns)):
        for cell in row[1:]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.0"


def _add_excel_bar_chart(ws, df: pd.DataFrame, title: str, y_axis: str, anchor: str) -> None:
    n_cols = len(df.columns)
    n_rows = len(df)
    header_row = 3
    first_data_row = header_row + 1
    last_data_row = header_row + n_rows
    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.title = title
    chart.y_axis.title = y_axis
    chart.x_axis.title = df.columns[0]
    data_ref = Reference(
        ws,
        min_col=2,
        min_row=header_row,
        max_col=n_cols,
        max_row=last_data_row,
    )
    cats_ref = Reference(ws, min_col=1, min_row=first_data_row, max_row=last_data_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width = 26
    chart.height = 14
    chart.dataLabels = DataLabelList(showVal=False)
    ws.add_chart(chart, anchor)


def main() -> None:
    CHART_DIR.mkdir(exist_ok=True)
    rev_pivot = build_revenue_pivot()
    tok_pivot = build_token_pivot()

    _grouped_bar(
        rev_pivot,
        label_col="Company",
        title="AI revenue / annualized run-rate by company (USD millions, log scale)",
        ylabel="ARR (USD millions, log)",
        out=CHART_DIR / "revenue_bar.png",
        log=True,
    )
    _grouped_bar(
        tok_pivot,
        label_col="Provider",
        title="Monthly API tokens processed by provider (trillions, log scale)",
        ylabel="Tokens per month (trillions, log)",
        out=CHART_DIR / "tokens_bar.png",
        log=True,
    )

    raw_rev = pd.read_csv(DATA_DIR / "revenue_arr_timeseries.csv")
    raw_tok = pd.read_csv(DATA_DIR / "monthly_token_consumption.csv")

    wb = Workbook()
    wb.remove(wb.active)

    ws_rev = wb.create_sheet("Revenue (pivot)")
    _write_df_to_sheet(ws_rev, rev_pivot, "Annualized run-rate revenue (USD millions)")
    _add_excel_bar_chart(
        ws_rev,
        rev_pivot,
        "ARR by company across snapshots",
        "USD millions",
        anchor=f"A{ws_rev.max_row + 3}",
    )

    ws_tok = wb.create_sheet("Tokens (pivot)")
    _write_df_to_sheet(ws_tok, tok_pivot, "Monthly API tokens processed (trillions)")
    _add_excel_bar_chart(
        ws_tok,
        tok_pivot,
        "Monthly tokens by provider across snapshots",
        "Trillions of tokens",
        anchor=f"A{ws_tok.max_row + 3}",
    )

    ws_raw_rev = wb.create_sheet("Revenue (raw)")
    _write_df_to_sheet(ws_raw_rev, raw_rev, "Raw ARR observations")

    ws_raw_tok = wb.create_sheet("Tokens (raw)")
    _write_df_to_sheet(ws_raw_tok, raw_tok, "Raw monthly token observations")

    ws_notes = wb.create_sheet("Notes", 0)
    ws_notes["A1"] = "AI Revenue & Monthly Token Consumption"
    ws_notes["A1"].font = Font(bold=True, size=16)
    notes_lines = [
        "",
        "Workbook contents:",
        "  - Revenue (pivot): companies x snapshot dates, with bar chart.",
        "  - Tokens (pivot):  providers x snapshot dates, with bar chart.",
        "  - Revenue (raw):   every dated ARR observation with source notes.",
        "  - Tokens (raw):    every tokens-per-minute / per-month observation.",
        "",
        "Snapshot rule: each pivoted cell uses the most recent disclosed",
        "observation on or before the snapshot date (forward-fill from",
        "milestone disclosures).",
        "",
        "Caveats:",
        "  - ARR = annualized run-rate (most recent month x 12), NOT TTM.",
        "  - OpenAI books net of cloud share; Anthropic books gross.",
        "  - Sierra / Decagon / Fin / Anthropic token rows include interpolations",
        "    between disclosed milestones (flagged in raw sheets).",
        "  - 'Intercom Fin' is the AI-agent product line within Intercom",
        "    (~$400M parent ARR in Mar 2026).",
        "",
        "Sources: Reuters, The Information, FT, CNBC, Sacra, SaaStr,",
        "Bloomberg/Sherwood, Mostly Metrics, OpenAI DevDay 2025 keynote,",
        "Tomasz Tunguz, Demirer/Fradkin et al. (Dec 2025), Microsoft &",
        "Alphabet earnings releases, company blogs.",
    ]
    for i, line in enumerate(notes_lines, start=2):
        ws_notes.cell(row=i, column=1, value=line)
    ws_notes.column_dimensions["A"].width = 90

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
