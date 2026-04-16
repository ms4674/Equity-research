"""
Generate a multi-sheet Excel workbook comparing different types of AI agent
harnesses / orchestration frameworks.

Data compiled from public sources, GitHub, company disclosures, and analyst
reports as of April 2026.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=11)
BODY_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)


def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_body(ws, start_row, end_row, max_col):
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            if (r - start_row) % 2 == 1:
                cell.fill = alt_fill


def auto_width(ws, max_col, cap=45):
    for col in range(1, max_col + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, max(len(str(l)) for l in str(cell.value).split("\n")))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, cap)


def add_title_row(ws, title, max_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30


# ── Sheet 1: Framework Overview ──────────────────────────────────────────────

def sheet_framework_overview(wb):
    ws = wb.create_sheet("Framework Overview")
    title = "AI Agent Harness / Orchestration Framework Overview"
    headers = [
        "Framework", "Provider / Maintainer", "Harness Category",
        "Architecture Pattern", "Primary Language(s)", "License",
        "GitHub Stars (Apr 2026)", "GitHub Forks", "Contributors",
        "Monthly Downloads", "Latest Stable Release",
        "Production-Ready?", "Key Differentiator",
    ]
    max_col = len(headers)
    add_title_row(ws, title, max_col)

    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, max_col)

    data = [
        ["LangGraph", "LangChain (Sequoia, Benchmark)", "Open-Source Framework",
         "Graph-based state machines", "Python, TypeScript", "MIT",
         "~29,100", "5,003", "280", "~90M (LangChain ecosystem)",
         "v2.0 (Feb 2026)", "Yes",
         "Deterministic control flow; state persistence; explicit branching logic"],
        ["CrewAI", "CrewAI Inc. (Insight Partners)", "Open-Source Framework",
         "Role-based agent crews", "Python", "MIT",
         "~48,000", "6,700+", "300+", "~5M (PyPI)",
         "v4.x (2026)", "Yes",
         "Fastest time-to-first-agent (~25 min); YAML config; lowest entry barrier"],
        ["AutoGen / AG2", "Microsoft (OSS, maintenance mode)", "Open-Source Framework (Legacy)",
         "Conversational multi-agent message passing", "Python", "MIT",
         "~57,000", "8,565", "450+", "Declining (migrating to MS Agent Framework)",
         "v0.7.5 (Sep 2025)", "Maintenance only",
         "Native cyclic loop support; dynamic agent collaboration; research tasks"],
        ["Microsoft Agent Framework", "Microsoft", "Platform SDK",
         "Graph + middleware pipeline", ".NET, Python", "MIT",
         "Merged from AutoGen + Semantic Kernel", "N/A", "N/A", "N/A (v1.0 Apr 2026)",
         "v1.0 (Apr 2026)", "Yes",
         "Unifies AutoGen + Semantic Kernel; enterprise .NET support; Azure integration"],
        ["OpenAI Agents SDK", "OpenAI", "Platform SDK",
         "Handoff-based agent chains", "Python", "MIT",
         "~20,700", "3,390", "240", "N/A",
         "v0.13.6 (Apr 2026)", "Yes",
         "Native OpenAI model integration; handoffs; guardrails; realtime voice"],
        ["Claude Agent SDK", "Anthropic", "Platform SDK",
         "Tool-use agentic loop", "Python, TypeScript", "Proprietary",
         "N/A (embedded in Claude Code)", "N/A", "N/A", "N/A",
         "2026", "Yes",
         "Computer-like interface (files, shell, web); 18 lifecycle hooks; zero tool setup"],
        ["Google ADK", "Google", "Platform SDK",
         "Model-driven orchestration", "Python, Java, Go, TypeScript", "Apache 2.0",
         "N/A", "N/A", "N/A", "7M+ (Nov 2025)",
         "Java 1.0.0 (Mar 2026)", "Yes",
         "A2A protocol; Vertex AI integration; 100+ connectors; model-agnostic"],
        ["AWS Bedrock AgentCore", "Amazon Web Services", "Managed Platform",
         "Framework-agnostic runtime", "Python, Java (Spring AI)", "Proprietary",
         "N/A (managed service)", "N/A", "N/A", "N/A",
         "GA 2026", "Yes",
         "MicroVM session isolation; multi-framework support; AWS-native scaling"],
        ["Strands Agents", "AWS (OSS)", "Open-Source Framework",
         "Model-driven orchestration", "Python, TypeScript", "Apache 2.0",
         "~6,100", "N/A", "120", "N/A",
         "v1.35.0 (Apr 2026)", "Yes",
         "Minimal code; MCP support; native AWS integrations; handoffs + swarms"],
        ["Salesforce Agentforce", "Salesforce", "Enterprise Platform",
         "Declarative agent builder", "Low-code / Apex", "Proprietary",
         "N/A (SaaS)", "N/A", "N/A", "N/A",
         "Continuous (SaaS)", "Yes",
         "$800M ARR (Q4 FY26); 29K deals; CRM-native; 19T tokens processed"],
        ["LlamaIndex Agents", "LlamaIndex (run-llama)", "Open-Source Framework",
         "Pipeline + tool-based agents", "Python, TypeScript", "MIT",
         "~47,800", "7,038", "1,500+", "~25M",
         "v0.12.x (2026)", "Yes",
         "Best-in-class RAG; document ingestion pipelines; composable query engines"],
    ]

    for r, row_data in enumerate(data, 3):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_body(ws, 3, 2 + len(data), max_col)
    auto_width(ws, max_col)
    ws.freeze_panes = "A3"
    return ws


# ── Sheet 2: Harness Category Taxonomy ───────────────────────────────────────

def sheet_harness_taxonomy(wb):
    ws = wb.create_sheet("Harness Category Taxonomy")
    title = "Agent Harness Categories: Taxonomy & Characteristics"
    headers = [
        "Harness Category", "Definition", "Target User",
        "Hosting Model", "Customization Level",
        "Typical Integration", "Examples", "Trade-offs",
    ]
    max_col = len(headers)
    add_title_row(ws, title, max_col)

    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, max_col)

    data = [
        ["Open-Source Framework",
         "Standalone libraries that provide agent primitives (loops, tools, memory) for developers to compose custom agent architectures",
         "Developers, AI engineers, startups",
         "Self-hosted (cloud or on-prem)",
         "High — full control over agent topology, model selection, tool definitions",
         "pip/npm install; integrated via code into existing apps",
         "LangGraph, CrewAI, AutoGen, Strands, LlamaIndex",
         "Maximum flexibility but requires engineering effort; ops burden on user"],
        ["Platform SDK",
         "SDKs tightly coupled with a specific model provider's ecosystem, offering native access to proprietary model features",
         "Developers building on a specific model provider",
         "Cloud-managed by provider; SDK runs client-side",
         "Medium — model-specific features with some customization",
         "API key + SDK; usually tied to provider's billing",
         "OpenAI Agents SDK, Claude Agent SDK, Google ADK, MS Agent Framework",
         "Deep model integration but creates vendor lock-in; fast start"],
        ["Managed Platform / Runtime",
         "Fully managed infrastructure for deploying, scaling, and operating agents built with any framework",
         "Enterprise DevOps, platform engineers",
         "Fully managed by cloud provider",
         "Low-Medium — runtime config, not agent logic",
         "Deploy via CLI/CI; provider handles scaling, isolation, observability",
         "AWS Bedrock AgentCore, Google Vertex Agent Engine",
         "Zero-ops scaling but limited to provider's runtime constraints"],
        ["Enterprise SaaS Platform",
         "End-to-end SaaS products with built-in agent capabilities, typically low-code / no-code",
         "Business users, CRM admins, non-developers",
         "SaaS (multi-tenant cloud)",
         "Low — declarative config, templates, pre-built workflows",
         "Point-and-click builder within existing SaaS product",
         "Salesforce Agentforce, ServiceNow Agent, SAP Joule",
         "Fastest time-to-value but least flexible; data stays in SaaS silo"],
        ["Research / Experimental",
         "Frameworks primarily for academic research, benchmarking, or prototyping novel agent architectures",
         "Researchers, academics, advanced prototypers",
         "Local / notebook",
         "Very High — full source access, novel patterns",
         "pip install; notebook-driven experimentation",
         "OpenAI Swarm (deprecated), CAMEL-AI, AgentBench",
         "Cutting-edge ideas but not production-hardened; may be abandoned"],
    ]

    for r, row_data in enumerate(data, 3):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_body(ws, 3, 2 + len(data), max_col)
    auto_width(ws, max_col)
    ws.freeze_panes = "A3"
    return ws


# ── Sheet 3: Architecture Comparison ─────────────────────────────────────────

def sheet_architecture_comparison(wb):
    ws = wb.create_sheet("Architecture Comparison")
    title = "Agent Harness Architecture Patterns"
    headers = [
        "Architecture Pattern", "How It Works",
        "Frameworks Using This Pattern",
        "Control Flow", "State Management",
        "Multi-Agent Support", "Human-in-the-Loop",
        "Best For", "Limitations",
    ]
    max_col = len(headers)
    add_title_row(ws, title, max_col)

    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, max_col)

    data = [
        ["Graph-Based State Machines",
         "Agents are nodes in a directed graph; edges define transitions; state is passed explicitly between nodes",
         "LangGraph, Microsoft Agent Framework",
         "Deterministic; explicit branching and conditional edges",
         "Built-in state persistence; checkpointing; time-travel debugging",
         "Yes — parallel branches, sub-graphs",
         "Native support (interrupt nodes)",
         "Complex conditional workflows; production reliability; regulatory compliance",
         "Verbose for simple tasks; steep learning curve"],
        ["Role-Based Agent Crews",
         "Each agent is assigned a role (researcher, writer, reviewer); a manager coordinates task delegation",
         "CrewAI",
         "Sequential or hierarchical task delegation",
         "Partial — shared context between crew members",
         "Yes — built-in multi-agent crew orchestration",
         "Native support",
         "Rapid prototyping; business process automation; content pipelines",
         "Limited cyclic graphs; less fine-grained state control"],
        ["Conversational Message Passing",
         "Agents communicate via structured messages in a shared chat; each agent responds to messages from others",
         "AutoGen (legacy), AG2",
         "Dynamic — agents decide when to respond",
         "Conversation history as implicit state",
         "Yes — core design principle",
         "Add-on required",
         "Research tasks; multi-expert deliberation; brainstorming",
         "Higher latency; token overhead; less deterministic"],
        ["Handoff-Based Chains",
         "A primary agent delegates to specialist sub-agents via handoffs; control returns after sub-task completion",
         "OpenAI Agents SDK, Strands Agents",
         "Linear with delegated branches",
         "Per-agent context; handoff transfers relevant state",
         "Yes — via handoff mechanism",
         "Native support",
         "Customer service routing; triage workflows; voice agents",
         "Less suited for deeply collaborative multi-agent scenarios"],
        ["Tool-Use Agentic Loop",
         "A single agent iterates in a loop: reason → select tool → execute → observe → repeat until done",
         "Claude Agent SDK, basic LangChain agents",
         "Model-driven; LLM decides tool calls",
         "Conversation context + tool results",
         "Limited — primarily single-agent",
         "Via tool callbacks",
         "Coding agents; computer use; autonomous task completion",
         "Single-agent focus; orchestration complexity for multi-agent"],
        ["Model-Driven Orchestration",
         "The LLM itself plans and executes the workflow; framework provides tools and memory, model decides sequence",
         "Google ADK, Strands Agents",
         "Model-driven; framework provides guardrails",
         "Sessions + Memory Bank (short & long-term)",
         "Yes — A2A protocol (Google); swarms (Strands)",
         "Native support",
         "Flexible workflows where plan isn't known ahead of time",
         "Less deterministic; model quality directly affects reliability"],
        ["Declarative / Low-Code Builder",
         "Visual or config-driven agent definition; no custom code required",
         "Salesforce Agentforce, ServiceNow Agent",
         "Template-driven; rule-based routing",
         "Managed by SaaS platform",
         "Limited — single-agent per topic",
         "Built-in",
         "Business users; CRM automation; customer service",
         "Least flexible; vendor lock-in; limited to platform's capabilities"],
    ]

    for r, row_data in enumerate(data, 3):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_body(ws, 3, 2 + len(data), max_col)
    auto_width(ws, max_col)
    ws.freeze_panes = "A3"
    return ws


# ── Sheet 4: Performance Benchmarks ──────────────────────────────────────────

def sheet_benchmarks(wb):
    ws = wb.create_sheet("Performance Benchmarks")
    title = "Agent Framework Performance Benchmarks (2026)"
    headers = [
        "Metric", "LangGraph", "CrewAI", "AutoGen (Legacy)",
        "OpenAI Agents SDK", "Strands Agents", "Notes",
    ]
    max_col = len(headers)
    add_title_row(ws, title, max_col)

    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, max_col)

    data = [
        ["Median E2E Latency (research task)", "14.1s", "18.4s", "22.7s",
         "~12s (est.)", "~15s (est.)", "Lower is better; benchmark on standardized research task"],
        ["Cost per 1,000 Tasks", "$41.70", "$48.20", "$67.40",
         "Model-dependent", "Model-dependent", "Token costs vary by model; OSS frameworks add no markup"],
        ["Integration Complexity (1-10)", "6.8", "3.5", "5.9",
         "4.0", "3.0", "Lower is simpler; based on developer surveys"],
        ["Time to First Agent", "~2 hours", "~25 min", "~1 hour",
         "~30 min", "~15 min", "From install to working agent prototype"],
        ["State Persistence", "Native (checkpointing)", "Partial", "Conversation-based",
         "Session-based", "Session-based", "LangGraph most robust for production state"],
        ["Max Concurrent Agents (tested)", "50+", "20+", "30+",
         "20+", "40+", "Depends on infrastructure; not hard limits"],
        ["MCP Protocol Support", "Yes", "Yes", "No",
         "Yes", "Yes", "Model Context Protocol for tool interoperability"],
        ["A2A Protocol Support", "No", "No", "No",
         "No", "No", "Agent-to-Agent protocol (Google-led standard)"],
    ]

    for r, row_data in enumerate(data, 3):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_body(ws, 3, 2 + len(data), max_col)
    auto_width(ws, max_col)
    ws.freeze_panes = "A3"
    return ws


# ── Sheet 5: Market Size & Projections ───────────────────────────────────────

def sheet_market_projections(wb):
    ws = wb.create_sheet("Market Projections")
    title = "Agent Orchestration / Harness Market Projections"
    headers = [
        "Metric / Source", "2024", "2025", "2026E", "2027E",
        "2028E", "2030E", "2034E", "CAGR", "Source",
    ]
    max_col = len(headers)
    add_title_row(ws, title, max_col)

    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, max_col)

    data = [
        ["Agentic AI Enterprise Platform Market (Marqstats)",
         "—", "$4.35B", "$7.03B", "—", "—", "$47.8B", "—", "61.5%", "Marqstats 2026"],
        ["Agent Orchestration Platform Market (MarketIntelo)",
         "—", "$5.8B", "—", "—", "—", "—", "$38.6B", "23.7%", "MarketIntelo 2026"],
        ["Autonomous AI Agent Market (Deloitte)",
         "—", "—", "$8.5B", "—", "—", "$35-45B", "—", "~35%", "Deloitte TMT Predictions 2026"],
        ["Global AI Spending (all categories)",
         "—", "$1.75T", "$2.52T", "—", "—", "—", "—", "44% YoY", "Gartner 2026"],
        ["Salesforce Agentforce ARR",
         "—", "~$300M", "$800M", "—", "—", "—", "—", "169% YoY", "Salesforce Q4 FY26 earnings"],
        ["Enterprise Agentic AI Adoption (% at scale)",
         "—", "2%", "~5% (est.)", "—", "—", "~25% (est.)", "—", "—",
         "Marqstats; Fortune 500 survey"],
        ["Fortune 500 Piloting Multi-Agent Workflows",
         "18%", "62%", "~75% (est.)", "—", "—", "—", "—", "—",
         "MarketIntelo; industry surveys"],
        ["LangChain Ecosystem Revenue",
         "—", "$16M", "—", "—", "—", "—", "—", "—", "CB Insights FY2025"],
        ["CrewAI Agentic Executions (trailing 12mo)",
         "—", "—", "2B+", "—", "—", "—", "—", "—", "CrewAI public disclosure"],
    ]

    for r, row_data in enumerate(data, 3):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_body(ws, 3, 2 + len(data), max_col)
    auto_width(ws, max_col)
    ws.freeze_panes = "A3"
    return ws


# ── Sheet 6: Funding & Valuation ─────────────────────────────────────────────

def sheet_funding(wb):
    ws = wb.create_sheet("Funding & Valuation")
    title = "Agent Framework Companies: Funding & Investors"
    headers = [
        "Company / Framework", "Total Funding", "Latest Round",
        "Lead Investors", "Valuation (est.)",
        "Revenue (Latest)", "Monetization Model",
    ]
    max_col = len(headers)
    add_title_row(ws, title, max_col)

    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, max_col)

    data = [
        ["LangChain / LangGraph", "$160M (6 rounds)", "Series B $125M (Oct 2025)",
         "Institutional Venture Partners, Sequoia, Benchmark, CapitalG",
         "~$500M+", "$16M (FY2025)", "LangSmith SaaS (tracing, evaluation, deployment)"],
        ["CrewAI", "$18M", "Series A (Oct 2024)",
         "Insight Partners, Boldstart Ventures",
         "~$100M+", "Not disclosed", "CrewAI Enterprise (managed platform)"],
        ["LlamaIndex", "$~35M", "Series A (2024)",
         "Greylock, Madrona, a16z",
         "~$200M+", "Not disclosed", "LlamaCloud (managed RAG, parsing)"],
        ["OpenAI (Agents SDK)", "$40B+ (total)", "Latest round $40B (2025)",
         "SoftBank, Microsoft, Thrive Capital",
         "$300B", "~$13B ARR (2026)", "API usage (per-token billing)"],
        ["Anthropic (Claude Agent SDK)", "$~15B+ (total)", "Series E $2B (2025)",
         "Google, Menlo Ventures, Spark, Salesforce",
         "$60B+", "~$4B+ ARR (2026 est.)", "API usage (per-token billing)"],
        ["Microsoft (Agent Framework)", "N/A (public co)", "N/A",
         "N/A", "$3T+ mkt cap",
         "$245B+ (FY2025)", "Azure AI consumption; Copilot subscriptions"],
        ["Google (ADK / Vertex AI)", "N/A (public co)", "N/A",
         "N/A", "$2T+ mkt cap",
         "$350B+ (FY2025)", "Google Cloud consumption; Vertex AI"],
        ["AWS (Bedrock AgentCore)", "N/A (Amazon subsidiary)", "N/A",
         "N/A", "N/A",
         "$115B+ AWS rev (2025)", "Bedrock API consumption; AgentCore runtime"],
        ["Salesforce (Agentforce)", "N/A (public co)", "N/A",
         "N/A", "$280B mkt cap",
         "$41.5B (FY2026); Agentforce $800M ARR", "Per-conversation pricing; seat licenses"],
    ]

    for r, row_data in enumerate(data, 3):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_body(ws, 3, 2 + len(data), max_col)
    auto_width(ws, max_col)
    ws.freeze_panes = "A3"
    return ws


# ── Sheet 7: Token Usage by Harness ──────────────────────────────────────────

def sheet_token_usage(wb):
    ws = wb.create_sheet("Token Usage by Harness")
    title = "Token Consumption & Cost Impact by Agent Harness Type"
    headers = [
        "Framework / Harness", "Token Overhead vs. Raw API",
        "Primary Overhead Source",
        "Cost per 1K Tasks (GPT-4o)", "Multi-Agent Multiplier",
        "Context Mgmt Strategy", "Token Savings from Mgmt",
        "Annual Cost (100K tasks/mo)",
    ]
    max_col = len(headers)
    add_title_row(ws, title, max_col)

    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, max_col)

    data = [
        ["LangGraph", "+9%",
         "State serialization; minimal wrapper",
         "$41.70", "3–8× (sub-graphs most efficient)",
         "Autonomous compression; trimming; summarization; checkpointing",
         "60–80%", "$135K"],
        ["CrewAI", "+18%",
         "Role prompts; delegation instructions; crew coordination",
         "$48.20", "3–5× (manager + workers)",
         "Shared context window; manager summarization between handoffs",
         "20–40%", "$360K"],
        ["AutoGen (legacy)", "+31%",
         "Conversational message passing; bilateral agent-to-agent messages",
         "$67.40", "5–10× (all agents read all messages)",
         "Full conversation replay; no built-in compression",
         "0% (none)", "$750K"],
        ["OpenAI Agents SDK", "~+12% (est.)",
         "Handoff context; guardrail checks; tracing metadata",
         "Model-dependent", "1.5–2× (handoff pairs)",
         "Per-agent context isolation; handoff transfers relevant state",
         "30–50%", "$210K"],
        ["Claude Agent SDK", "~+8–15% (est.)",
         "Tool-use loop; tool results can be large (files, shell output)",
         "Model-dependent", "1× (primarily single-agent)",
         "Conversation + tool results; relies on model context window",
         "0–20%", "Varies"],
        ["Google ADK", "~+10–15% (est.)",
         "Model-driven planning; session management overhead",
         "Model-dependent", "2–4× (A2A protocol)",
         "Sessions + Memory Bank; automatic session compaction",
         "40–60%", "Varies"],
        ["Strands Agents", "~+10% (est.)",
         "Minimal orchestration; model-driven planning",
         "Model-dependent", "2–3× (handoffs + swarms)",
         "Session-based; handoff with condensed context",
         "30–50%", "Varies"],
        ["Salesforce Agentforce", "Abstracted",
         "Token costs hidden; $2/conversation or $0.10/action",
         "$2.00/conversation", "N/A (managed)",
         "Managed by platform; no user control",
         "N/A", "$2.4M (per-conversation)"],
    ]

    for r, row_data in enumerate(data, 3):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_body(ws, 3, 2 + len(data), max_col)

    # Add sub-table: Token anatomy breakdown
    gap_row = 3 + len(data) + 1
    ws.merge_cells(start_row=gap_row, start_column=1, end_row=gap_row, end_column=max_col)
    sub_cell = ws.cell(row=gap_row, column=1,
                       value="Token Anatomy: Where Tokens Go in a Typical 10-Step Agent Task")
    sub_cell.font = Font(name="Calibri", bold=True, size=12, color="2F5496")

    sub_headers = [
        "Token Category", "Tokens per Call", "Calls (10-step)",
        "Total Tokens", "% of Total", "Cost Driver", "", "",
    ]
    sub_row = gap_row + 1
    for c, h in enumerate(sub_headers, 1):
        ws.cell(row=sub_row, column=c, value=h)
    style_header(ws, sub_row, max_col)

    sub_data = [
        ["System prompt", "1,500–4,000", "10", "15,000–40,000", "10–15%",
         "Repeated every call; fixed cost per framework", "", ""],
        ["Tool / function schemas", "2,000–5,000", "10", "20,000–50,000", "12–18%",
         "Grows with number of tools registered", "", ""],
        ["Conversation history", "2,500 → 26,500", "10", "~145,000", "40–50%",
         "Quadratic growth; largest cost component", "", ""],
        ["Tool results", "200–3,000", "10", "2,000–30,000", "5–12%",
         "Highly variable; file reads / web fetches can be massive", "", ""],
        ["Thinking / reasoning tokens", "3–10× output", "10", "30,000–100,000+", "15–30%",
         "Reasoning models only (o3, o4-mini, Opus 4); invisible but billed", "", ""],
        ["Output tokens", "200–800", "10", "2,000–8,000", "3–5%",
         "3–8× more expensive per token than input", "", ""],
    ]

    for r, row_data in enumerate(sub_data, sub_row + 1):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_body(ws, sub_row + 1, sub_row + len(sub_data), max_col)

    # Add sub-table: Multi-agent multipliers
    gap2 = sub_row + len(sub_data) + 2
    ws.merge_cells(start_row=gap2, start_column=1, end_row=gap2, end_column=max_col)
    sub2_cell = ws.cell(row=gap2, column=1,
                        value="Multi-Agent Topology Token Multipliers")
    sub2_cell.font = Font(name="Calibri", bold=True, size=12, color="2F5496")

    topo_headers = [
        "Topology", "Token Multiplier vs. Single Agent", "Explanation",
        "Best Framework Fit", "", "", "", "",
    ]
    topo_row = gap2 + 1
    for c, h in enumerate(topo_headers, 1):
        ws.cell(row=topo_row, column=c, value=h)
    style_header(ws, topo_row, max_col)

    topo_data = [
        ["Single agent (tool loop)", "1× (baseline)", "One context window, one loop",
         "Claude Agent SDK, LangGraph", "", "", "", ""],
        ["Sequential handoff (2 agents)", "1.5–2×",
         "Handoff transfers condensed context; second agent builds new history",
         "OpenAI Agents SDK, Strands", "", "", "", ""],
        ["Parallel fan-out (3 agents)", "2.5–3.5×",
         "Each agent runs independently; results merged by orchestrator",
         "LangGraph, Google ADK", "", "", "", ""],
        ["Crew with manager (4 agents)", "3–5×",
         "Manager delegates + reviews; each worker has full context cycle",
         "CrewAI", "", "", "", ""],
        ["Conversational swarm (4+ agents)", "5–10×",
         "All agents read all messages; bilateral token costs; negotiation rounds",
         "AutoGen (legacy)", "", "", "", ""],
        ["Hierarchical sub-graphs (nested)", "3–8×",
         "Sub-graphs encapsulate; parent only sees summaries; most efficient multi-agent",
         "LangGraph", "", "", "", ""],
    ]

    for r, row_data in enumerate(topo_data, topo_row + 1):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_body(ws, topo_row + 1, topo_row + len(topo_data), max_col)

    auto_width(ws, max_col)
    ws.freeze_panes = "A3"
    return ws


# ── Sheet 8: Sources & Notes ─────────────────────────────────────────────────

def sheet_sources(wb):
    ws = wb.create_sheet("Sources & Notes")
    title = "Sources, Methodology & Disclaimers"
    max_col = 2
    add_title_row(ws, title, max_col)

    headers = ["Topic", "Detail"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, max_col)

    data = [
        ["Data As Of", "April 15, 2026"],
        ["GitHub Stars", "Sourced from GitHub repository pages; snapshot values subject to daily fluctuation"],
        ["Download Counts", "PyPI and npm registry statistics; LangChain 90M/mo per company disclosure"],
        ["Market Size Estimates",
         "Marqstats Agentic AI Enterprise Platform Market Report 2026-2030; "
         "MarketIntelo Agent Orchestration Platform Market 2034; "
         "Deloitte TMT Predictions 2026; Gartner 2026"],
        ["Framework Benchmarks",
         "Multi-Agent Orchestration Frameworks Benchmark (agent-harness.ai); "
         "standardized research task with median latency, cost per 1K tasks, integration complexity"],
        ["Funding Data",
         "CB Insights, Tracxn, Crunchbase, company press releases"],
        ["Revenue Data",
         "Public earnings (Salesforce Q4 FY26, Microsoft FY2025); "
         "CB Insights (LangChain FY2025); analyst estimates where noted"],
        ["Enterprise Adoption",
         "Fortune 500 survey data from Marqstats and MarketIntelo reports; "
         "Camunda State of Agentic Orchestration 2026"],
        ["Token Overhead Benchmarks",
         "Multi-Agent Orchestration Frameworks Benchmark (agent-harness.ai); "
         "TokenMix framework comparison 2026; "
         "AI Cost Check agent cost guide 2026; "
         "Zylos Research token economics report (Feb 2026)"],
        ["Token Anatomy & Multipliers",
         "Context Window Budgeting (gantz.ai); "
         "ContextBudget paper (arXiv 2604.01664); "
         "Augment Code loop cost model; "
         "LangChain autonomous context compression blog; "
         "Salesforce Agentforce pricing page (salesforce.com)"],
        ["Methodology",
         "Frameworks categorized by: (1) open-source vs. proprietary, "
         "(2) standalone vs. platform-embedded, (3) architecture pattern, "
         "(4) target user persona. Categories are non-exclusive. "
         "Token overhead measured vs. equivalent raw API calls on same model. "
         "Annual cost estimates assume 100K tasks/mo on GPT-4o pricing."],
        ["Disclaimers",
         "Estimates marked (est.) are author projections based on available data. "
         "Market size figures vary widely across analysts due to differing category definitions. "
         "GitHub stars ≠ adoption. Download counts may include CI/CD bots. "
         "This is for informational purposes and is not investment advice."],
    ]

    for r, row_data in enumerate(data, 3):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_body(ws, 3, 2 + len(data), max_col)
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 90
    ws.freeze_panes = "A3"
    return ws


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sheet_framework_overview(wb)
    sheet_harness_taxonomy(wb)
    sheet_architecture_comparison(wb)
    sheet_benchmarks(wb)
    sheet_market_projections(wb)
    sheet_funding(wb)
    sheet_token_usage(wb)
    sheet_sources(wb)

    output = "Agent_Harness_Types_Comparison.xlsx"
    wb.save(output)
    print(f"Saved {output}")


if __name__ == "__main__":
    main()
