#!/usr/bin/env python3
"""
Orbital Data Center — Aggregated Research Model
================================================
Synthesises data from across the equity-research repository to build
a comprehensive view of orbital (space-based) data center economics.

Sources aggregated:
  - SpaceX / xAI financial model  (launch costs, Starship economics)
  - Datacenter infrastructure model  (capex/MW, BOM, power, build timelines)
  - 1 GW datacenter memory content  (GPU rack specs, memory BOM)
  - Datacenter vendor metrics  (colo capacity, AI share, vendor landscape)
  - Energy costs / DC economics  (power pricing, PUE)

Output: Orbital_Datacenter_Model.xlsx  (multi-tab workbook)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

# ═══════════════════════════════════════════════════════════════════════
#  STYLE CONSTANTS
# ═══════════════════════════════════════════════════════════════════════
DARK_BLUE = "1B2A4A"
MED_BLUE = "2E5090"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
FORECAST_BG = "FFF2CC"
GREEN_BG = "E2EFDA"
ORANGE_BG = "FCE4D6"
PURPLE_BG = "E8DAEF"

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=WHITE)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color=DARK_BLUE)
DATA_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="808080")
TOTAL_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)

TITLE_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
HEADER_FILL = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
SECTION_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
ALT_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
FORECAST_FILL = PatternFill(start_color=FORECAST_BG, end_color=FORECAST_BG, fill_type="solid")
GREEN_FILL = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
ORANGE_FILL = PatternFill(start_color=ORANGE_BG, end_color=ORANGE_BG, fill_type="solid")
PURPLE_FILL = PatternFill(start_color=PURPLE_BG, end_color=PURPLE_BG, fill_type="solid")
TOTAL_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)
TOTAL_BORDER = Border(
    top=Side(style="thin", color="000000"),
    bottom=Side(style="double", color="000000"),
)

NUM = '#,##0'
NUM1 = '#,##0.0'
NUM2 = '#,##0.00'
PCT = '0.0%'
DOLLAR = '$#,##0'
DOLLAR_M = '$#,##0.0'


def _style(cell, font=None, fill=None, align=None, border=None, fmt=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = border
    if fmt: cell.number_format = fmt


def write_title(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    _style(c, TITLE_FONT, TITLE_FILL, Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[row].height = 28


def write_headers(ws, row, labels, start_col=1):
    for i, lbl in enumerate(labels):
        c = ws.cell(row=row, column=start_col + i, value=lbl)
        _style(c, HEADER_FONT, HEADER_FILL, Alignment(horizontal="center", vertical="center", wrap_text=True), THIN_BORDER)


def write_section(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    _style(c, SECTION_FONT, SECTION_FILL, Alignment(horizontal="left"))


def write_row(ws, row, values, start_col=1, fmt=None, bold=False, is_total=False, alt=False):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col + i, value=v)
        f = BOLD_FONT if bold else DATA_FONT
        fl = None
        if is_total:
            f = TOTAL_FONT
            fl = TOTAL_FILL
        elif alt:
            fl = ALT_FILL
        bdr = TOTAL_BORDER if is_total else THIN_BORDER
        al = Alignment(horizontal="left") if i == 0 else Alignment(horizontal="right")
        _style(c, f, fl, al, bdr, fmt if i > 0 else None)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════════════
#  SHEET 1 — EXECUTIVE SUMMARY
# ═══════════════════════════════════════════════════════════════════════
def build_executive_summary(wb):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_properties.tabColor = "1B2A4A"
    ncols = 5
    set_col_widths(ws, [55, 25, 20, 20, 20])

    r = 1
    write_title(ws, r, "Orbital Data Center — Aggregated Research Summary", ncols); r += 2

    write_section(ws, r, "CONCEPT OVERVIEW", ncols); r += 1
    overview = [
        "An orbital data center (ODC) is a compute facility deployed in low-Earth orbit (LEO),",
        "leveraging SpaceX Starship's mass-to-orbit economics to position GPU/AI infrastructure",
        "beyond terrestrial power, land, and permitting constraints.",
        "",
        "This workbook aggregates data from five research branches in this repository:",
        "  1. SpaceX / xAI Financial Model — launch costs, Starship unit economics",
        "  2. Datacenter Infrastructure Model — capex/MW, BOM, power, build timelines",
        "  3. 1 GW DC Memory Content — GPU rack specs, HBM/DRAM/NAND BOM",
        "  4. Datacenter Vendor Metrics — colo capacity, AI share, vendor landscape",
        "  5. Energy Costs & DC Economics — power pricing, PUE",
    ]
    for line in overview:
        ws.cell(row=r, column=1, value=line).font = DATA_FONT
        r += 1

    r += 1
    write_section(ws, r, "KEY METRICS AT A GLANCE", ncols); r += 1
    write_headers(ws, r, ["Metric", "Value", "Source Branch", "Notes", ""]); r += 1

    kpi_data = [
        ["Starship target cost/kg to LEO", "$10–50", "SpaceX Model", "vs. $2,700/kg Falcon 9"],
        ["Starship payload to LEO", "100,000–150,000 kg", "SpaceX Model", "Fully reusable config"],
        ["Falcon 9 cost/launch (internal)", "~$15M", "SpaceX Model", "Booster reused 20+ flights"],
        ["Falcon 9 payload to LEO", "22,800 kg", "SpaceX Model", "Expendable upper stage"],
        ["Starship target cost/launch", "$10–50M", "SpaceX Model", "Fully reusable (both stages)"],
        ["GB300 NVL72 rack TDP", "132 kW", "Memory Content", "72 GPUs per rack"],
        ["GB300 NVL72 rack weight (est.)", "~2,500 kg", "Memory Content", "Rack + servers + cabling"],
        ["HBM3e per GPU", "288 GB", "Memory Content", "Fixed in silicon"],
        ["Memory semiconductor $/GPU", "$5,927 (train) / $5,959 (infer)", "Memory Content", "HBM + DRAM + NAND"],
        ["Terrestrial capex per MW (2025)", "$12–13M / MW", "DC Infra Model", "All-in hyperscaler build"],
        ["Terrestrial power cost", "$72/MWh (2025)", "DC Infra Model", "Rising from $48 in 2018"],
        ["DC build timeline (land to live)", "40 months (2025)", "DC Infra Model", "Power procurement bottleneck"],
        ["US DC power demand (2025)", "72 GW", "DC Infra Model", "Growing to 210 GW by 2030E"],
        ["Hyperscaler AI capex (2025)", "$238B", "DC Infra Model", "AWS+MSFT+GOOG+Meta+ORCL+AAPL"],
        ["Colo vendor total capacity", "~28 GW (top 10)", "Vendor Metrics", "Live + under construction"],
        ["AI share of colo capacity", "30–100%", "Vendor Metrics", "CoreWeave 100%, DLR 30%"],
    ]
    for i, row_data in enumerate(kpi_data):
        write_row(ws, r, row_data, alt=(i % 2 == 1))
        r += 1

    r += 1
    ws.cell(row=r, column=1, value=f"Model compiled: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')} UTC").font = NOTE_FONT


# ═══════════════════════════════════════════════════════════════════════
#  SHEET 2 — UNIT ECONOMICS: LAUNCH COSTS
# ═══════════════════════════════════════════════════════════════════════
def build_launch_economics(wb):
    ws = wb.create_sheet("Unit Economics — Launch")
    ws.sheet_properties.tabColor = "2E5090"
    ncols = 8
    set_col_widths(ws, [40, 18, 18, 18, 18, 18, 18, 30])

    r = 1
    write_title(ws, r, "Unit Economics — Launch Cost to Orbit", ncols); r += 2

    # Launch vehicle comparison
    write_section(ws, r, "1. LAUNCH VEHICLE COMPARISON", ncols); r += 1
    write_headers(ws, r, [
        "Metric", "SpaceX Falcon 9", "SpaceX Falcon Heavy",
        "SpaceX Starship", "ULA Vulcan", "Arianespace A6",
        "Blue Origin NG", "Source"
    ]); r += 1

    vehicles = [
        ["Price to LEO ($/kg)", "$2,700", "$1,500", "$10–50 (target)", "$15,000–20,000", "$12,000–15,000", "$8,000 (est.)", "SpaceX Model"],
        ["Payload to LEO (kg)", "22,800", "63,800", "100,000–150,000", "27,200", "21,650", "45,000", "SpaceX Model"],
        ["Price / launch ($M)", "$67 (external)", "$90–150", "$10–50 (target)", "$110–150", "$90–115", "$80 (est.)", "SpaceX Model"],
        ["Internal cost / launch ($M)", "~$15", "~$30", "$2–5 (target)", "n/a", "n/a", "n/a", "SpaceX Model"],
        ["Reusability", "Booster: 20+ flights", "Booster: partial", "Both stages", "Expendable", "Expendable", "Booster planned", "SpaceX Model"],
        ["2024 launch cadence", "134", "7", "Test flights", "3", "3", "0", "SpaceX Model"],
        ["2025E launch cadence", "~155", "~10", "Iterating", "~10", "~8", "~3–5", "SpaceX Model"],
        ["Fairing volume (m³)", "~145", "~145", "~1,000", "~87", "~165", "~300", "Industry est."],
    ]
    for i, vd in enumerate(vehicles):
        write_row(ws, r, vd, alt=(i % 2 == 1)); r += 1

    r += 1
    write_section(ws, r, "2. STARSHIP UNIT ECONOMICS — MARGIN TRAJECTORY", ncols); r += 1
    write_headers(ws, r, ["", "2022A", "2023A", "2024A", "2025E", "2026E", "2027E", "Long-Term"]); r += 1

    starship_econ = [
        ["Segment Revenue ($M)", 0, 0, 200, 800, 2000, 3500, ""],
        ["  NASA HLS milestones", 0, 0, 150, 400, 600, 800, ""],
        ["  DOD / NatSec contracts", 0, 0, 30, 150, 400, 800, ""],
        ["  Commercial (sat deployment)", 0, 0, 10, 150, 600, 1200, ""],
        ["  Internal (Starlink v3)", 0, 0, 10, 100, 300, 400, ""],
        ["Segment COGS ($M)", 0, 0, 200, 640, 1400, 2100, ""],
        ["Segment Gross Profit ($M)", 0, 0, 0, 160, 600, 1400, ""],
        ["Segment Gross Margin %", "n/a", "n/a", "0%", "20%", "30%", "40%", "60–70%"],
        ["", "", "", "", "", "", "", ""],
        ["Flights per vehicle (target)", "n/a", "n/a", 1, 3, 8, 15, "100+"],
        ["Mfg cost / vehicle ($M)", "n/a", "n/a", 200, 150, 100, 80, "$30–50"],
        ["Cost per flight ($M)", "n/a", "n/a", 200, 50, 12.5, 5.3, "~$2–3"],
        ["Revenue per flight ($M)", "n/a", "n/a", 200, 267, 250, 233, "varies"],
    ]
    for i, sd in enumerate(starship_econ):
        write_row(ws, r, sd, bold=("Gross Profit" in str(sd[0]) or "Margin" in str(sd[0])), alt=(i % 2 == 1)); r += 1

    r += 1
    write_section(ws, r, "3. COST TO DEPLOY 1 MW OF COMPUTE TO ORBIT", ncols); r += 1
    write_headers(ws, r, ["Component", "Cost ($M)", "Weight (kg)", "Notes", "", "", "", ""]); r += 1

    deploy_cost = [
        ["GPU rack hardware (GB300 NVL72)", "~$0.8–1.0", "~2,500", "72 GPUs, 132 kW TDP, at rack-level pricing"],
        ["Rad-hardening & thermal mods", "~$0.3–0.5", "~200", "Radiation shielding, thermal management for vacuum"],
        ["Power system (solar + battery)", "~$2.0–3.0", "~4,000", "~150 kW solar array + battery for eclipse periods"],
        ["Cooling system (space radiator)", "~$1.0–1.5", "~1,500", "Passive/active radiator panels for 132 kW heat rejection"],
        ["Structure & bus", "~$0.5–1.0", "~1,000", "Satellite bus, docking, attitude control"],
        ["Comms (optical inter-sat link)", "~$0.3–0.5", "~300", "Laser links for cluster mesh + ground downlink"],
        ["Integration & test", "~$0.5–1.0", "", "Assembly, vibration, thermal-vacuum testing"],
        ["Launch cost (Starship, at target)", "~$0.1–0.5", "", "~10,000 kg payload share at $10–50/kg"],
        ["TOTAL per rack-equivalent", "~$5.5–9.0", "~9,500", "Per 132 kW orbital rack module"],
        ["TOTAL per MW (orbit)", "~$42–68", "~72,000 kg/MW", "vs. $12–13M/MW terrestrial (2025)"],
    ]
    for i, dc in enumerate(deploy_cost):
        write_row(ws, r, dc, bold=("TOTAL" in str(dc[0])), alt=(i % 2 == 1)); r += 1

    r += 1
    ws.cell(row=r, column=1, value="Notes: Orbital costs are highly speculative and assume mature Starship reuse. Terrestrial $12-13M/MW includes land, building, power infra.").font = NOTE_FONT
    r += 1
    ws.cell(row=r, column=1, value="Orbital cost is dominated by power generation (solar) and thermal management — subsystems unnecessary on Earth.").font = NOTE_FONT


# ═══════════════════════════════════════════════════════════════════════
#  SHEET 3 — UNIT ECONOMICS: TERRESTRIAL COMPARISON
# ═══════════════════════════════════════════════════════════════════════
def build_terrestrial_comparison(wb):
    ws = wb.create_sheet("Unit Economics — Terrestrial")
    ws.sheet_properties.tabColor = "2E5090"
    ncols = 9
    set_col_widths(ws, [35, 12, 12, 12, 12, 12, 12, 12, 12])

    years_label = ["", "2020", "2022", "2024A", "2025A", "2026E", "2028E", "2030E", "Notes"]

    r = 1
    write_title(ws, r, "Unit Economics — Terrestrial DC Benchmarks", ncols); r += 2

    write_section(ws, r, "1. CAPEX PER MW OF NEW CAPACITY ($M/MW)", ncols); r += 1
    write_headers(ws, r, years_label); r += 1

    capex_mw = [
        ["Amazon (AWS)", 8.2, 9.5, 12.0, 12.8, 12.5, 11.8, 11.2, "From DC Infra Model"],
        ["Microsoft (Azure)", 7.8, 9.2, 11.5, 12.5, 12.2, 11.4, 10.8, ""],
        ["Google (GCP)", 7.5, 9.0, 11.2, 12.0, 11.8, 11.2, 10.5, ""],
        ["Meta", 8.5, 9.8, 12.5, 13.2, 12.8, 12.0, 11.2, ""],
        ["Oracle Cloud", 9.0, 10.0, 13.0, 13.5, 13.0, 12.0, 11.0, ""],
    ]
    for i, cd in enumerate(capex_mw):
        write_row(ws, r, cd, alt=(i % 2 == 1)); r += 1

    r += 1
    write_section(ws, r, "2. DATACENTER BILL OF MATERIALS ($/MW of IT Load)", ncols); r += 1
    write_headers(ws, r, ["Component", "", "2020 $M/MW", "2024 $M/MW", "2028E $M/MW", "2030E $M/MW", "Key Vendors", "", ""]); r += 1

    bom = [
        ["Land & Site Prep", "", 0.5, 0.8, 1.0, 1.1, "CBRE, JLL", "", ""],
        ["Building Shell & Floor", "", 1.8, 2.1, 2.4, 2.5, "DPR, Turner", "", ""],
        ["HV Substation & Transformers", "", 1.0, 1.5, 1.8, 1.9, "ABB, Siemens, Eaton", "", ""],
        ["Switchgear & Busway", "", 0.8, 1.0, 1.2, 1.2, "Schneider, Eaton", "", ""],
        ["UPS Systems", "", 0.6, 0.8, 0.9, 0.9, "Vertiv, Schneider", "", ""],
        ["Backup Generators", "", 0.5, 0.6, 0.7, 0.7, "Caterpillar, Cummins", "", ""],
        ["PDUs & Power Whips", "", 0.3, 0.4, 0.4, 0.4, "Vertiv, Schneider", "", ""],
        ["Chillers & Cooling Towers", "", 0.8, 1.0, 0.9, 0.8, "Trane, Johnson Controls", "", ""],
        ["CRAH/CRAC Units", "", 0.3, 0.3, 0.3, 0.2, "Vertiv, Stulz", "", ""],
        ["Liquid Cooling (DLC/Immersion)", "", 0.1, 0.5, 0.9, 1.0, "CoolIT, GRC, Asetek", "", ""],
        ["Fiber & Structured Cabling", "", 0.5, 0.6, 0.8, 0.8, "Corning, CommScope", "", ""],
        ["Network Switches & Routers", "", 0.5, 0.7, 0.8, 0.8, "Arista, Cisco, Broadcom", "", ""],
        ["GPU Servers (AI-optimized)", "", 0.0, 3.0, 5.0, 5.5, "NVIDIA, AMD, Dell, HPE", "", ""],
        ["CPU Servers (General)", "", 2.5, 2.0, 1.8, 1.6, "Dell, HPE, Lenovo", "", ""],
        ["Storage (SSD/HDD)", "", 0.4, 0.5, 0.6, 0.6, "Pure, NetApp, Samsung", "", ""],
        ["Security, BMS, DCIM", "", 0.3, 0.3, 0.5, 0.5, "Honeywell, Schneider", "", ""],
        ["Design, Engineering, PM", "", 0.6, 0.8, 1.0, 1.0, "HDR, AECOM, Jacobs", "", ""],
        ["TOTAL", "", 11.5, 16.9, 20.9, 21.5, "", "", ""],
    ]
    for i, bd in enumerate(bom):
        write_row(ws, r, bd, bold=("TOTAL" in str(bd[0])), alt=(i % 2 == 1)); r += 1

    r += 1
    write_section(ws, r, "3. POWER ECONOMICS", ncols); r += 1
    write_headers(ws, r, ["", "2018", "2020", "2022", "2024", "2025", "2028E", "2030E", ""]); r += 1

    power_data = [
        ["Avg power cost ($/MWh)", 48, 52, 60, 70, 72, 64, 52, ""],
        ["US DC power demand (GW)", 17, 22, 32, 55, 72, 150, 210, ""],
        ["Interconnect queue (months)", 18, 22, 28, 38, 42, 32, 24, ""],
        ["DC build timeline (months)", 24, 26, 30, 38, 40, 32, 26, ""],
    ]
    for i, pd in enumerate(power_data):
        write_row(ws, r, pd, alt=(i % 2 == 1)); r += 1

    r += 1
    write_section(ws, r, "4. HYPERSCALER CAPEX ($B) — ANNUAL", ncols); r += 1
    write_headers(ws, r, ["", "2020", "2022", "2023", "2024A", "2025A", "2026E", "2028E", "2030E"]); r += 1

    capex_data = [
        ["Amazon (AWS)", 21.0, 36.0, 48.4, 83.0, 105.3, 112.0, 145.0, 158.0],
        ["Microsoft (Azure)", 15.4, 28.0, 32.0, 55.7, 88.0, 102.0, 122.0, 128.0],
        ["Google (GCP)", 15.0, 26.0, 32.3, 52.5, 74.5, 86.0, 102.0, 108.0],
        ["Meta", 11.0, 18.0, 28.0, 39.2, 62.0, 72.0, 80.0, 82.0],
        ["Oracle Cloud", 2.4, 4.0, 6.9, 14.5, 20.0, 25.0, 28.0, 29.0],
        ["Apple", 5.5, 6.5, 7.0, 9.5, 11.5, 13.5, 16.0, 18.0],
        ["TOTAL Hyperscaler", 70.3, 118.5, 154.6, 254.4, 361.3, 410.5, 493.0, 523.0],
    ]
    for i, cd in enumerate(capex_data):
        write_row(ws, r, cd, fmt=NUM1, bold=("TOTAL" in str(cd[0])), is_total=("TOTAL" in str(cd[0])), alt=(i % 2 == 1)); r += 1


# ═══════════════════════════════════════════════════════════════════════
#  SHEET 4 — PAYLOAD SPECIFICATIONS
# ═══════════════════════════════════════════════════════════════════════
def build_payload_specs(wb):
    ws = wb.create_sheet("Payload Specifications")
    ws.sheet_properties.tabColor = "70AD47"
    ncols = 7
    set_col_widths(ws, [45, 20, 20, 20, 20, 20, 30])

    r = 1
    write_title(ws, r, "Payload Specifications — Compute Hardware for Orbit", ncols); r += 2

    write_section(ws, r, "1. GB300 NVL72 RACK — REFERENCE PLATFORM (2025–2026)", ncols); r += 1
    write_headers(ws, r, ["Specification", "GB300 NVL72", "Vera Rubin (est.)", "Change", "", "", "Source"]); r += 1

    rack_specs = [
        ["GPUs per rack", 72, 72, "—", "", "", "NVIDIA specs"],
        ["CPUs per rack", 36, 36, "—", "", "", "NVIDIA specs"],
        ["Rack TDP (kW)", 132, "~150", "+14%", "", "", "NVIDIA specs"],
        ["HBM generation", "HBM3e", "HBM4", "New gen", "", "", ""],
        ["HBM per GPU (GB)", 288, 384, "+33%", "", "", "NVIDIA specs"],
        ["HBM per rack (TB)", 20.7, 27.6, "+33%", "", "", ""],
        ["System memory per rack (TB)", 17, "~20", "+18%", "", "", "LPDDR5X → LPDDR6"],
        ["E1.S NVMe drive bays", 144, 144, "—", "", "", ""],
        ["Local NAND per rack (TB)", 576, 576, "—", "", "", "At 4 TB/drive"],
        ["Cooling", "Direct-to-chip liquid", "Direct-to-chip liquid", "—", "", "", ""],
        ["Rack size", "48U", "48U", "—", "", "", ""],
        ["Rack weight (est. kg)", "~2,500", "~2,600", "+4%", "", "", "Industry est."],
    ]
    for i, rs in enumerate(rack_specs):
        write_row(ws, r, rs, alt=(i % 2 == 1)); r += 1

    r += 1
    write_section(ws, r, "2. MEMORY SEMICONDUCTOR CONTENT PER GPU", ncols); r += 1
    write_headers(ws, r, ["Component", "Training $/GPU", "Inference $/GPU", "GB/GPU (Train)", "GB/GPU (Infer)", "", ""]); r += 1

    mem_gpu = [
        ["HBM3e", "$2,949", "$2,944", "288 GB", "288 GB", "", ""],
        ["DRAM (LPDDR5X + DDR5)", "$913", "$884", "255 GB", "247 GB", "", ""],
        ["NAND (all tiers)", "$2,071", "$2,130", "18.3 TB", "18.9 TB", "", ""],
        ["TOTAL Memory $/GPU", "$5,933", "$5,959", "", "", "", ""],
    ]
    for i, mg in enumerate(mem_gpu):
        write_row(ws, r, mg, bold=("TOTAL" in str(mg[0])), alt=(i % 2 == 1)); r += 1

    r += 1
    write_section(ws, r, "3. NETWORKING COST PER GPU", ncols); r += 1
    write_headers(ws, r, ["Component", "Training $/GPU", "Inference $/GPU", "Notes", "", "", ""]); r += 1

    net_gpu = [
        ["Scale-up (NVSwitch, NVLink, NICs)", "$4,650", "$4,650", "Fixed per rack, 72 GPUs", "", "", ""],
        ["Scale-out (switches, optics, cables)", "$2,300", "$700", "Training: fat-tree for all-reduce", "", "", ""],
        ["TOTAL Networking $/GPU", "$6,950", "$5,350", "", "", "", ""],
    ]
    for i, ng in enumerate(net_gpu):
        write_row(ws, r, ng, bold=("TOTAL" in str(ng[0])), alt=(i % 2 == 1)); r += 1

    r += 1
    write_section(ws, r, "4. STARSHIP PAYLOAD CAPACITY vs. COMPUTE PAYLOAD", ncols); r += 1
    write_headers(ws, r, ["Metric", "Value", "Calculation", "Notes", "", "", ""]); r += 1

    payload_calc = [
        ["Starship payload to LEO (kg)", "100,000–150,000", "Published spec", "Fully reusable config"],
        ["GB300 NVL72 rack mass (est. kg)", "~2,500", "Rack + servers + cabling", "Without power/cooling mods"],
        ["Racks per Starship launch (mass)", "40–60", "100K–150K / 2,500 kg", "Mass-limited scenario"],
        ["Starship cargo volume (m³)", "~1,000", "Published spec", "8.8m dia × ~17m long"],
        ["NVL72 rack volume (48U + aisle)", "~2.5 m³", "0.6m × 1.2m × 2.1m + access", "Standard 48U footprint"],
        ["Racks per Starship (volume)", "~100–200", "1,000 / 2.5–5 m³", "Volume-limited (with clearance)"],
        ["Binding constraint", "MASS", "", "40–60 racks per launch (mass-limited)"],
        ["GPUs per Starship launch", "2,880–4,320", "40–60 racks × 72 GPUs", ""],
        ["Compute power per launch (MW)", "5.3–7.9", "40–60 racks × 132 kW", "Rack TDP only"],
        ["Memory per launch (HBM, TB)", "828–1,242", "40–60 racks × 20.7 TB", ""],
        ["Memory per launch (NAND, PB)", "23–35", "40–60 racks × 576 TB", "Local NAND only"],
    ]
    for i, pc in enumerate(payload_calc):
        write_row(ws, r, pc, alt=(i % 2 == 1)); r += 1

    r += 1
    write_section(ws, r, "5. MEMORY CONTENT PRICING (mid-2026)", ncols); r += 1
    write_headers(ws, r, ["Component", "$/GB", "$/TB", "Notes", "", "", ""]); r += 1
    pricing = [
        ["HBM3e", "$10.00", "$10,240", "SK Hynix 50–55%, Samsung 30–35%, Micron 15–20%"],
        ["LPDDR5X / DDR5", "$3.50", "$3,584", "Grace CPU system memory"],
        ["NAND (die-level)", "$0.11", "$113", "Enterprise SSD demand +58% in 2026"],
        ["HBM4 (Vera Rubin era)", "$14.00", "$14,336", "+40% price premium over HBM3e"],
    ]
    for i, pr in enumerate(pricing):
        write_row(ws, r, pr, alt=(i % 2 == 1)); r += 1


# ═══════════════════════════════════════════════════════════════════════
#  SHEET 5 — CLUSTER SIZE & CONFIGURATIONS
# ═══════════════════════════════════════════════════════════════════════
def build_cluster_size(wb):
    ws = wb.create_sheet("Cluster Size & Config")
    ws.sheet_properties.tabColor = "ED7D31"
    ncols = 7
    set_col_widths(ws, [45, 22, 22, 22, 22, 22, 30])

    r = 1
    write_title(ws, r, "Cluster Size & Configuration — Orbital vs. Terrestrial", ncols); r += 2

    write_section(ws, r, "1. TERRESTRIAL REFERENCE: 1 GW AI DATACENTER", ncols); r += 1
    write_headers(ws, r, ["Metric", "Training", "Inference", "Delta", "", "", "Source"]); r += 1

    gw_data = [
        ["Total facility power", "1,000 MW", "1,000 MW", "—", "", "", "Memory Content"],
        ["PUE", "1.15", "1.18", "+3%", "", "", ""],
        ["IT load", "870 MW", "847 MW", "-3%", "", "", ""],
        ["GPU compute allocation", "62% (540 MW)", "70% (593 MW)", "+10% MW", "", "", ""],
        ["Network fabric allocation", "15% (130 MW)", "8% (68 MW)", "-48%", "", "", ""],
        ["GPU racks (GB300 NVL72)", "4,091", "4,492", "+10%", "", "", ""],
        ["Total GPUs", "294,545", "323,424", "+10%", "", "", ""],
        ["HBM3e total", "85 PB", "93 PB", "+9%", "", "", ""],
        ["DRAM total", "75 PB", "80 PB", "+7%", "", "", ""],
        ["NAND total", "5.4 EB", "6.1 EB", "+13%", "", "", ""],
        ["Memory semiconductor $", "$1,746M", "$1,927M", "+10%", "", "", ""],
        ["Networking equipment $", "$2,049M", "$1,728M", "-16%", "", "", ""],
    ]
    for i, gd in enumerate(gw_data):
        write_row(ws, r, gd, alt=(i % 2 == 1)); r += 1

    r += 1
    write_section(ws, r, "2. ORBITAL CLUSTER SIZING SCENARIOS", ncols); r += 1
    write_headers(ws, r, ["Configuration", "10 MW Cluster", "50 MW Cluster", "100 MW Cluster", "500 MW Cluster", "1 GW Cluster", "Notes"]); r += 1

    cluster_scenarios = [
        ["GPU racks required", "76", "379", "758", "3,788", "7,576", "At 132 kW/rack"],
        ["Total GPUs", "5,472", "27,288", "54,576", "272,736", "545,472", "72 GPUs/rack"],
        ["Starship launches (mass)", "2–3", "10–13", "19–25", "95–126", "189–252", "40–60 racks/launch"],
        ["Starship launches (w/ power+cooling)", "5–8", "25–40", "50–80", "250–400", "500–800", "Incl. solar+radiators"],
        ["HBM3e (PB)", "1.6", "7.8", "15.7", "78.4", "156.9", "20.7 TB/rack"],
        ["NAND local (PB)", "43.8", "218.3", "436.6", "2,182.1", "4,363.8", "576 TB/rack"],
        ["Solar array area (est. m²)", "~50K", "~250K", "~500K", "~2.5M", "~5M", "At 200 W/m² BOL"],
        ["Radiator area (est. m²)", "~30K", "~150K", "~300K", "~1.5M", "~3M", "Thermal rejection"],
        ["Orbit altitude (km)", "500–600", "500–600", "500–600", "500–600", "500–600", "LEO"],
        ["Est. hardware cost ($B)", "0.5–0.9", "2.7–4.5", "5.5–9.0", "27–45", "55–90", "All-in excl. launch"],
        ["Est. launch cost ($B)", "0.05–0.4", "0.25–2.0", "0.5–4.0", "2.5–20", "5–40", "At Starship targets"],
        ["Est. total deployed cost ($B)", "0.6–1.3", "3.0–6.5", "6.0–13.0", "30–65", "60–130", "Hardware + launch"],
    ]
    for i, cs in enumerate(cluster_scenarios):
        write_row(ws, r, cs, alt=(i % 2 == 1)); r += 1

    r += 1
    write_section(ws, r, "3. TERRESTRIAL HYPERSCALER CAPACITY (GW, IT Load)", ncols); r += 1
    write_headers(ws, r, ["Company", "2020", "2022", "2024A", "2025A", "2026E", "2030E", ""]); r += 1

    gw_cap = [
        ["Amazon (AWS)", "3.6", "5.5", "9.2", "12.5", "15.5", "26.5"],
        ["Microsoft (Azure)", "3.0", "4.8", "7.8", "11.0", "14.0", "25.0"],
        ["Google (GCP)", "2.6", "4.0", "6.8", "9.0", "11.5", "21.0"],
        ["Meta", "1.3", "2.3", "5.2", "7.5", "9.5", "17.5"],
        ["Oracle Cloud", "0.5", "1.0", "2.8", "4.2", "5.8", "11.0"],
        ["Apple", "0.3", "0.5", "0.8", "1.1", "1.4", "2.5"],
        ["TOTAL Hyperscaler (GW)", "11.3", "18.1", "32.6", "45.3", "57.7", "103.5"],
    ]
    for i, gc in enumerate(gw_cap):
        write_row(ws, r, gc, bold=("TOTAL" in str(gc[0])), is_total=("TOTAL" in str(gc[0])), alt=(i % 2 == 1)); r += 1

    r += 1
    write_section(ws, r, "4. COLO / DC VENDOR CAPACITY (MW)", ncols); r += 1
    write_headers(ws, r, ["Vendor", "Total MW", "AI MW (est.)", "AI %", "Under Construction", "Planned", "Notes"]); r += 1

    vendor_cap = [
        ["Digital Realty (DLR)", "2,345", "~700", "~30%", "~845 MW", "~3 GW+", "Public REIT (NYSE: DLR)"],
        ["Equinix (EQIX)", "~3,000+", "~1,200", "~40%", "~1,200 MW", "~3 GW+", "Public REIT (NASDAQ: EQIX)"],
        ["QTS Realty (Blackstone)", "4,752", "~2,400", "~50%", "~3,250 MW", "~5 GW+", "Private"],
        ["CyrusOne (KKR)", "789+", "~400", "~50%", "~600+ MW", "~2 GW+", "Private"],
        ["Vantage Data Centers", "4,187", "~2,500", "~60%", "~2,900+ MW", "~2 GW+", "Frontier: 1.4 GW, $25B"],
        ["Aligned Data Centers", "5,000+", "~3,000", "~60%", "~3,500+ MW", "~2 GW+", "Private (~$40B EV)"],
        ["STACK Infrastructure", "4,000+", "~2,000", "~50%", "~2,500+ MW", "~6 GW+", "3 GW TX campus planned"],
        ["Compass Datacenters", "500+", "~200", "~40%", "~300+ MW", "~2 GW+", "$10B Lauderdale County"],
        ["NTT Global", "2,000+", "~600", "~30%", "~800+ MW", "~1 GW+", "150+ DCs, 20+ countries"],
        ["CoreWeave", "~1,500+", "~1,500", "~100%", "~700+ MW", "~2 GW+", "Pure-play AI GPU cloud"],
    ]
    for i, vc in enumerate(vendor_cap):
        write_row(ws, r, vc, alt=(i % 2 == 1)); r += 1


# ═══════════════════════════════════════════════════════════════════════
#  SHEET 6 — MAINTENANCE COSTS
# ═══════════════════════════════════════════════════════════════════════
def build_maintenance_costs(wb):
    ws = wb.create_sheet("Maintenance Costs")
    ws.sheet_properties.tabColor = "FF6384"
    ncols = 7
    set_col_widths(ws, [45, 22, 22, 22, 22, 22, 30])

    r = 1
    write_title(ws, r, "Maintenance Costs — Orbital vs. Terrestrial Operations", ncols); r += 2

    write_section(ws, r, "1. TERRESTRIAL DC OPERATING COSTS (annual, per MW)", ncols); r += 1
    write_headers(ws, r, ["Cost Category", "2020 ($M/MW/yr)", "2024 ($M/MW/yr)", "2026E ($M/MW/yr)", "2030E ($M/MW/yr)", "% of Total", "Notes"]); r += 1

    terr_opex = [
        ["Power (electricity)", "0.46", "0.61", "0.60", "0.46", "45–55%", "At $/MWh × 8,760 hrs × PUE"],
        ["Staff & operations", "0.08", "0.10", "0.11", "0.12", "10–12%", "NOC, facilities, security"],
        ["Hardware refresh (depreciation)", "0.15", "0.25", "0.30", "0.35", "20–25%", "3–5 yr GPU refresh cycle"],
        ["Cooling operations", "0.03", "0.04", "0.05", "0.04", "4–5%", "Water, chemicals, pumps"],
        ["Network & connectivity", "0.02", "0.03", "0.04", "0.04", "3–4%", "Transit, peering, dark fiber"],
        ["Insurance & compliance", "0.01", "0.02", "0.02", "0.02", "2%", "Property, liability, SOC2"],
        ["Repairs & spare parts", "0.02", "0.03", "0.03", "0.03", "3%", "UPS batteries, generator maint."],
        ["Software & licensing", "0.01", "0.02", "0.03", "0.03", "2–3%", "DCIM, monitoring, orchestration"],
        ["TOTAL opex per MW", "0.78", "1.10", "1.18", "1.09", "100%", ""],
    ]
    for i, to in enumerate(terr_opex):
        write_row(ws, r, to, bold=("TOTAL" in str(to[0])), alt=(i % 2 == 1)); r += 1

    r += 1
    write_section(ws, r, "2. ORBITAL DC OPERATING COSTS (annual, per MW, estimated)", ncols); r += 1
    write_headers(ws, r, ["Cost Category", "Low ($M/MW/yr)", "Base ($M/MW/yr)", "High ($M/MW/yr)", "% of Total", "vs. Terrestrial", "Notes"]); r += 1

    orb_opex = [
        ["Power (solar array degradation + battery)", "0.0", "0.0", "0.0", "0%", "Eliminated", "No utility bill; solar is capex"],
        ["Solar array replacement reserve", "0.8", "1.2", "1.8", "15–20%", "New cost", "~2% array degrad/yr; 10-yr replacement"],
        ["Ground station operations", "0.3", "0.5", "0.8", "8–10%", "Replaces connectivity", "Tracking, scheduling, data relay"],
        ["Mission control & orbital ops", "0.5", "0.8", "1.2", "10–15%", "Replaces facilities staff", "Orbit maint., collision avoidance"],
        ["Hardware refresh (deorbit + relaunch)", "2.0", "3.5", "6.0", "35–45%", "3–5x terrestrial", "Must launch replacement + deorbit old"],
        ["Radiation-induced failures", "0.3", "0.6", "1.0", "8–12%", "New cost", "SEU, latchup, total dose effects"],
        ["Orbital debris insurance", "0.2", "0.4", "0.8", "5–8%", "New cost", "Collision risk, Kessler syndrome"],
        ["Satellite bus maintenance", "0.1", "0.3", "0.5", "4–6%", "New cost", "Attitude control, thermal system"],
        ["Remote SW updates & patching", "0.05", "0.1", "0.2", "1–2%", "Similar", "OTA firmware, security patches"],
        ["TOTAL opex per MW", "4.3", "7.4", "12.3", "100%", "4–10x terrestrial", ""],
    ]
    for i, oo in enumerate(orb_opex):
        write_row(ws, r, oo, bold=("TOTAL" in str(oo[0])), alt=(i % 2 == 1)); r += 1

    r += 1
    write_section(ws, r, "3. HARDWARE REPLACEMENT CYCLE COMPARISON", ncols); r += 1
    write_headers(ws, r, ["Factor", "Terrestrial", "Orbital (LEO)", "Impact", "", "", "Notes"]); r += 1

    replace = [
        ["GPU refresh cycle", "3–5 years", "3–5 years", "Similar need", "", "", "Moore's law drives both"],
        ["Replacement method", "Swap in rack", "Full satellite swap", "10–100x cost", "", "", "No on-orbit servicing (yet)"],
        ["Downtime for swap", "Hours", "Weeks–months", "Availability hit", "", "", "Launch scheduling, orbit insertion"],
        ["Server MTBF (rad environment)", "50,000+ hrs", "10,000–30,000 hrs", "2–5x more failures", "", "", "LEO radiation (Van Allen fringes)"],
        ["Cooling system lifespan", "15–20 years", "10–15 years", "Shorter in space", "", "", "Thermal cycling, micrometeorites"],
        ["Power system lifespan", "20+ years (grid)", "10–15 years (solar)", "Replacement needed", "", "", "Solar cell degradation ~2%/yr"],
        ["Structural lifespan", "30–50 years", "10–20 years", "Atomic oxygen erosion", "", "", "LEO surface degradation"],
        ["Satellite lifespan (historical)", "n/a", "5–7 years (Starlink)", "Short-lived", "", "", "SpaceX designs for replacement"],
        ["Decommissioning cost", "Minimal", "$0.5–2M per module", "Significant", "", "", "Controlled deorbit required"],
    ]
    for i, rp in enumerate(replace):
        write_row(ws, r, rp, alt=(i % 2 == 1)); r += 1

    r += 1
    write_section(ws, r, "4. SPACEX LAUNCH OPERATIONS — COST STRUCTURE (from SpaceX Model)", ncols); r += 1
    write_headers(ws, r, ["Cost Component", "2022A", "2024A", "2025E", "2026E", "2027E", "", ""]); r += 1

    launch_ops = [
        ["Propellant & consumables ($M)", 150, 300, 330, 350, 370, "", ""],
        ["Refurbishment & recovery ($M)", 300, 500, 480, 460, 440, "", ""],
        ["Mfg (fairings, 2nd stage) ($M)", 500, 900, 950, 950, 960, "", ""],
        ["Launch ops & range fees ($M)", 200, 380, 430, 460, 490, "", ""],
        ["Insurance & other ($M)", 168, 278, 343, 350, 460, "", ""],
        ["Total launch COGS ($M)", 1318, 2358, 2533, 2570, 2720, "", ""],
        ["Launch revenue ($M)", 2928, 5896, 6665, 7140, 7770, "", ""],
        ["Launch gross margin %", "55%", "60%", "62%", "64%", "65%", "", ""],
    ]
    for i, lo in enumerate(launch_ops):
        write_row(ws, r, lo, bold=("gross margin" in str(lo[0]).lower() or "Total" in str(lo[0])), alt=(i % 2 == 1)); r += 1

    r += 1
    write_section(ws, r, "5. STARLINK SATELLITE ECONOMICS (proxy for orbital hardware lifespan)", ncols); r += 1
    write_headers(ws, r, ["Metric", "Value", "Notes", "", "", "", ""]); r += 1

    starlink = [
        ["Sat manufacturing cost (Gen2)", "$250K–400K", "Declining from $500K in 2020"],
        ["Satellite lifespan", "5–5.5 years", "Design life; some last longer"],
        ["Sats per launch (Starlink)", "23–30", "25 on v2 Mini; 30+ on Starship"],
        ["Internal launch cost allocation", "$14–15M per launch", "Declining with reuse"],
        ["Launch cost per satellite", "$0.5–0.6M", "At 25 sats/launch"],
        ["Constellation size (2025E)", "~6,500+ operational", "12,000 approved; 42,000 applied"],
        ["Annual replacement need", "~1,200–1,500", "At 5-yr lifespan"],
        ["Replacement cost per year", "$0.6–1.0B", "Satellites + launch allocation"],
        ["Revenue per sat per year ($K)", "$770–1,080", "Varies by year (subscriber ramp)"],
    ]
    for i, sl in enumerate(starlink):
        write_row(ws, r, sl + ["", "", "", ""], alt=(i % 2 == 1)); r += 1


# ═══════════════════════════════════════════════════════════════════════
#  SHEET 7 — ORBITAL vs. TERRESTRIAL TCO
# ═══════════════════════════════════════════════════════════════════════
def build_tco_comparison(wb):
    ws = wb.create_sheet("TCO Comparison")
    ws.sheet_properties.tabColor = "7030A0"
    ncols = 6
    set_col_widths(ws, [50, 22, 22, 22, 22, 30])

    r = 1
    write_title(ws, r, "Total Cost of Ownership — Orbital vs. Terrestrial (100 MW, 10yr)", ncols); r += 2

    write_section(ws, r, "SCENARIO: 100 MW AI COMPUTE FACILITY — 10-YEAR TCO", ncols); r += 1
    write_headers(ws, r, ["Cost Category", "Terrestrial ($M)", "Orbital Low ($M)", "Orbital Base ($M)", "Orbital High ($M)", "Notes"]); r += 1

    tco = [
        ["CAPEX — Initial Build", "", "", "", "", ""],
        ["  Hardware (servers, GPU, network)", 500, 600, 800, 1100, "Rad-hardening premium for orbital"],
        ["  Power infrastructure", 300, 200, 300, 500, "Solar+battery vs. grid connection"],
        ["  Cooling systems", 100, 100, 150, 250, "Space radiators vs. chillers"],
        ["  Building / structure", 200, 50, 100, 150, "Satellite bus vs. building shell"],
        ["  Land & permitting", 80, 0, 0, 0, "No land cost in orbit"],
        ["  Launch costs", 0, 50, 200, 500, "Starship at various maturity levels"],
        ["  Integration & test", 50, 100, 200, 350, "More complex for space hardware"],
        ["Total Initial Capex", 1230, 1100, 1750, 2850, ""],
        ["", "", "", "", "", ""],
        ["OPEX — Annual (×10 years)", "", "", "", "", ""],
        ["  Power / electricity", 610, 0, 0, 0, "Solar in orbit = free electricity"],
        ["  Solar array reserve (annual)", 0, 80, 120, 180, "Degradation replacement"],
        ["  Staff & operations", 100, 50, 80, 120, "Remote ops vs. on-site"],
        ["  Hardware refresh (2 cycles)", 700, 1400, 2500, 4500, "Relaunch cost dominates"],
        ["  Cooling operations", 40, 10, 30, 50, "Passive radiators vs. active"],
        ["  Network & connectivity", 30, 30, 50, 80, "Ground stations + inter-sat links"],
        ["  Insurance", 20, 20, 40, 80, "Space debris risk premium"],
        ["  Repairs & parts", 30, 30, 60, 100, "No on-site repair in orbit"],
        ["  Radiation failure reserve", 0, 30, 60, 100, "SEU mitigation"],
        ["  Decommissioning (end of life)", 5, 50, 100, 200, "Deorbit cost"],
        ["Total 10-Year Opex", 1535, 1700, 3040, 5410, ""],
        ["", "", "", "", "", ""],
        ["TOTAL 10-YEAR TCO", 2765, 2800, 4790, 8260, ""],
        ["TCO per MW per year ($M)", 2.8, 2.8, 4.8, 8.3, ""],
        ["Orbital premium vs. terrestrial", "—", "+1%", "+73%", "+199%", ""],
    ]
    for i, tc in enumerate(tco):
        is_t = "Total" in str(tc[0]) and tc[1] != ""
        write_row(ws, r, tc, bold=is_t, is_total=("TOTAL 10" in str(tc[0])), alt=(i % 2 == 1 and not is_t)); r += 1

    r += 2
    ws.cell(row=r, column=1, value="Key insight: Orbital compute only reaches cost parity with terrestrial in the most optimistic Starship scenario (full reuse at $2-3M/flight).").font = NOTE_FONT
    r += 1
    ws.cell(row=r, column=1, value="The primary advantage is bypassing terrestrial constraints: power grid bottlenecks (42-month interconnect queues), land/permitting (40-month build), and geography.").font = NOTE_FONT
    r += 1
    ws.cell(row=r, column=1, value="Orbital TCO is dominated by hardware replacement costs — inability to physically service equipment makes refresh cycles 3-5x more expensive.").font = NOTE_FONT


# ═══════════════════════════════════════════════════════════════════════
#  SHEET 8 — SUPPLY CHAIN & COMPONENT PRICING
# ═══════════════════════════════════════════════════════════════════════
def build_supply_chain(wb):
    ws = wb.create_sheet("Supply Chain & Components")
    ws.sheet_properties.tabColor = "FFC000"
    ncols = 8
    set_col_widths(ws, [40, 14, 14, 14, 14, 14, 14, 30])

    r = 1
    write_title(ws, r, "Supply Chain — Component Pricing & Vendor Landscape", ncols); r += 2

    write_section(ws, r, "1. COMPONENT PRICE INDEX (2020 = 100)", ncols); r += 1
    write_headers(ws, r, ["Component", "2018", "2020", "2022", "2024", "2026E", "2030E", "Trend"]); r += 1

    price_idx = [
        ["GPU / AI Accelerators", 40, 100, 170, 310, 240, 140, "Peak 2024, declining"],
        ["CPU Servers", 105, 100, 96, 92, 88, 80, "Steady decline"],
        ["Networking (Switches/Optics)", 95, 100, 115, 145, 142, 110, "Elevated, normalizing"],
        ["HV Transformers & Switchgear", 88, 100, 125, 185, 190, 138, "Supply-constrained"],
        ["UPS & Power Distribution", 92, 100, 118, 145, 148, 120, "Demand-driven"],
        ["Backup Generators", 94, 100, 122, 142, 145, 120, "Caterpillar/Cummins backlog"],
        ["Cooling (Chillers/CRAC)", 96, 100, 110, 122, 120, 103, "Stable"],
        ["Liquid Cooling (DLC)", 180, 100, 82, 68, 56, 42, "Rapid cost reduction"],
        ["Fiber & Cabling", 97, 100, 108, 120, 118, 104, "Corning/CommScope supply"],
        ["Construction Labor", 90, 100, 118, 140, 152, 138, "Persistent shortage"],
        ["Steel & Building Materials", 88, 100, 142, 125, 118, 108, "Post-COVID normalization"],
        ["Land & Real Estate", 85, 100, 125, 160, 178, 190, "Structural inflation"],
    ]
    for i, pi in enumerate(price_idx):
        write_row(ws, r, pi, alt=(i % 2 == 1)); r += 1

    r += 1
    write_section(ws, r, "2. DC SUPPLY CHAIN TOTAL ADDRESSABLE MARKET ($B)", ncols); r += 1
    write_headers(ws, r, ["Segment", "2020", "2022", "2024", "2025", "2026E", "2030E", "Key Vendors"]); r += 1

    supply_tam = [
        ["Servers & GPUs", 55, 78, 140, 195, 260, 475, "NVIDIA, AMD, Dell, HPE"],
        ["Networking Equipment", 15, 19, 30, 40, 52, 102, "Arista, Cisco, Broadcom"],
        ["Power Equipment", 10, 14, 24, 33, 44, 90, "Vertiv, Schneider, Eaton"],
        ["Cooling Systems", 6, 8, 15, 21, 29, 64, "Vertiv, Trane, CoolIT"],
        ["Construction & Engineering", 8, 11, 20, 28, 37, 68, "Fluor, DPR, Turner"],
        ["Real Estate / Colo Leasing", 12, 16, 23, 28, 35, 74, "Equinix, DLR, QTS"],
        ["DC Software & Mgmt (DCIM)", 3, 4, 6, 8, 10, 22, "Schneider, Vertiv, Nlyte"],
        ["Fiber & Connectivity", 5, 7, 10, 13, 17, 37, "Corning, CommScope, Lumen"],
        ["TOTAL", 114, 157, 268, 366, 484, 932, ""],
    ]
    for i, st in enumerate(supply_tam):
        write_row(ws, r, st, bold=("TOTAL" in str(st[0])), is_total=("TOTAL" in str(st[0])), alt=(i % 2 == 1)); r += 1

    r += 1
    write_section(ws, r, "3. HBM SUPPLIER MARKET SHARE (2026E)", ncols); r += 1
    write_headers(ws, r, ["Supplier", "Share", "Notes", "", "", "", "", ""]); r += 1

    hbm = [
        ["SK Hynix", "50–55%", "Technology leader, primary NVIDIA supplier for HBM3e/HBM4", "", "", "", "", ""],
        ["Samsung", "30–35%", "Ramping HBM4 mass production (early 2026)", "", "", "", "", ""],
        ["Micron", "15–20%", "Competitive on HBM3e, HBM4 expected late 2026", "", "", "", "", ""],
    ]
    for i, hb in enumerate(hbm):
        write_row(ws, r, hb, alt=(i % 2 == 1)); r += 1

    r += 1
    write_section(ws, r, "4. CAPEX FLOW — WHERE EACH $1 OF DC CAPEX GOES", ncols); r += 1
    write_headers(ws, r, ["Segment", "2020", "2024", "2028E", "2030E", "", "", ""]); r += 1

    capex_flow = [
        ["Servers & GPUs", "35%", "42%", "45%", "44%", "", "", ""],
        ["Networking Equipment", "10%", "9%", "9%", "9%", "", "", ""],
        ["Power Infrastructure", "15%", "14%", "13%", "12%", "", "", ""],
        ["Cooling Systems", "8%", "8%", "8%", "7%", "", "", ""],
        ["Building & Construction", "12%", "10%", "9%", "9%", "", "", ""],
        ["Land & Real Estate", "5%", "5%", "4%", "4%", "", "", ""],
        ["Fiber & Connectivity", "5%", "4%", "4%", "4%", "", "", ""],
        ["Software, Security & Other", "3%", "3%", "3%", "3%", "", "", ""],
        ["Design, Engineering & PM", "4%", "3%", "3%", "3%", "", "", ""],
        ["Contingency", "3%", "2%", "2%", "5%", "", "", ""],
    ]
    for i, cf in enumerate(capex_flow):
        write_row(ws, r, cf, alt=(i % 2 == 1)); r += 1


# ═══════════════════════════════════════════════════════════════════════
#  SHEET 9 — SOURCES & METHODOLOGY
# ═══════════════════════════════════════════════════════════════════════
def build_sources(wb):
    ws = wb.create_sheet("Sources & Methodology")
    ws.sheet_properties.tabColor = "808080"
    ws.column_dimensions['A'].width = 100

    r = 1
    write_title(ws, r, "Sources & Methodology", 1); r += 2

    notes = [
        "DATA SOURCES (from repository branches):",
        "",
        "1. SpaceX / xAI Financial Model  (branch: cursor/spacex-xai-financial-model-f0a5)",
        "   - Launch vehicle pricing, cost per kg, Starship margin trajectory",
        "   - SpaceX 3-statement model (2020–2028), competitive positioning",
        "   - Starlink satellite economics (manufacturing cost, lifespan, per-sat revenue)",
        "   - Sources: Bloomberg, WSJ, Reuters, Quilty Analytics, Morgan Stanley, Payload Space",
        "",
        "2. Datacenter Infrastructure Model  (branch: cursor/cloud-datacenter-model-0202)",
        "   - Hyperscaler capex (AWS, Azure, GCP, Meta, Oracle, Apple) — 2018–2030E",
        "   - GW capacity, server counts, datacenter counts",
        "   - Bill of materials ($/MW), component price indices",
        "   - Power supply & generation, build timelines, capex-to-revenue mapping",
        "   - Neocloud vendors (CoreWeave, Lambda, Crusoe, etc.)",
        "   - Sources: 10-K filings, 4Q25 earnings, industry reports",
        "",
        "3. 1 GW Datacenter Memory Content  (branch: cursor/datacenter-memory-content-1gw-63b2)",
        "   - NVIDIA GB300 NVL72 rack specifications",
        "   - Memory semiconductor content: HBM3e, DRAM, NAND (training vs. inference)",
        "   - Networking vs. memory spend analysis",
        "   - Vera Rubin forward look (HBM4, 2026–2027)",
        "   - Sources: NVIDIA specs, Supermicro datasheets, Citi/Goldman Sachs estimates",
        "",
        "4. Datacenter Vendor Metrics  (branch: cursor/datacenter-vendor-metrics-1337)",
        "   - 10 colo/DC vendors: DLR, Equinix, QTS, CyrusOne, Vantage, Aligned, STACK, Compass, NTT, CoreWeave",
        "   - Capacity (MW), AI share, top customers, GW buildout pipeline",
        "   - Sources: press releases, 10-K filings, IPO S-1, industry reports",
        "",
        "5. Energy Costs & DC Economics  (branch: cursor/energy-costs-data-centre-economics-ab1c)",
        "   - Power pricing, PUE benchmarks, renewable energy mix",
        "",
        "ORBITAL DATA CENTER ESTIMATES:",
        "",
        "Orbital-specific data points (solar array sizing, radiator area, radiation effects,",
        "deorbit costs, space insurance) are derived from published aerospace industry references",
        "and are marked as estimates throughout the workbook. These figures are illustrative",
        "and represent early-stage feasibility analysis, not precise engineering specifications.",
        "",
        "Key orbital assumptions:",
        "  - LEO altitude: 500–600 km (below Van Allen belt peak, above atmospheric drag)",
        "  - Solar flux: ~1,361 W/m² (AM0); array efficiency ~30%; BOL output ~200 W/m²",
        "  - Thermal rejection: radiator panels with emissivity ~0.85; assumes eclipse cycling",
        "  - Radiation environment: ~10 krad/yr total ionizing dose at 500 km, 51.6° inclination",
        "  - Satellite lifespan: 5–7 years (based on Starlink v2 design life)",
        "  - Starship cost: range of $2M–$50M per launch (early vs. mature reuse)",
        "",
        "METHODOLOGY:",
        "",
        "This workbook aggregates data from five separate research artifacts in this repository,",
        "cross-referencing and reconciling where data overlaps (e.g., hyperscaler capex figures",
        "appear in both the SpaceX and DC infrastructure models with consistent values).",
        "",
        "Orbital cluster sizing is derived by mapping terrestrial rack specifications (GB300 NVL72)",
        "against Starship payload constraints (mass and volume). TCO comparison uses a 10-year",
        "horizon with 2 hardware refresh cycles, consistent with terrestrial GPU refresh cadences.",
        "",
        "All orbital cost estimates are presented as ranges (low/base/high) to reflect the",
        "significant uncertainty in space-based compute economics at this stage of development.",
        "",
        f"Model compiled: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')} UTC",
        "Disclaimer: For research/educational purposes only. Not investment advice.",
    ]

    for note in notes:
        ws.cell(row=r, column=1, value=note)
        if note.endswith(":") or note.startswith("ORBITAL") or note.startswith("METHODOLOGY") or note.startswith("DATA"):
            ws.cell(row=r, column=1).font = SECTION_FONT
        elif note.startswith("  "):
            ws.cell(row=r, column=1).font = DATA_FONT
        elif note.startswith("Model compiled") or note.startswith("Disclaimer"):
            ws.cell(row=r, column=1).font = NOTE_FONT
        else:
            ws.cell(row=r, column=1).font = DATA_FONT
        r += 1


# ═══════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_executive_summary(wb)
    build_launch_economics(wb)
    build_terrestrial_comparison(wb)
    build_payload_specs(wb)
    build_cluster_size(wb)
    build_maintenance_costs(wb)
    build_tco_comparison(wb)
    build_supply_chain(wb)
    build_sources(wb)

    for ws in wb:
        ws.freeze_panes = "B3"

    output = "Orbital_Datacenter_Model.xlsx"
    wb.save(output)
    print(f"Saved: {output}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
