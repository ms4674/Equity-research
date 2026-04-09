import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_align = Alignment(vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


def style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def style_rows(ws, num_rows, num_cols):
    for row in range(2, num_rows + 2):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = cell_align
            if row % 2 == 0:
                cell.fill = alt_fill


# ── Sheet 1: Top 20 Consumer Gen AI Apps ──
ws1 = wb.active
ws1.title = "Consumer Gen AI Apps"
ws1.sheet_properties.tabColor = "2F5496"

consumer_headers = ["Rank", "App Name", "Company", "Category", "Monthly Active Users (M)", "Est. Downloads (2025, M)", "Key Use Case"]
ws1.append(consumer_headers)

consumer_data = [
    [1,  "ChatGPT",        "OpenAI",       "Chatbot / Assistant",    297.3, 770,  "Writing, coding, research, general Q&A"],
    [2,  "Doubao",          "ByteDance",    "Chatbot / Assistant",    99.6,  210,  "Conversational AI (China market leader)"],
    [3,  "DeepSeek",        "DeepSeek AI",  "Chatbot / Reasoning",   38.6,  85,   "Reasoning, research, coding"],
    [4,  "Grok",            "xAI",          "Chatbot / Assistant",    23.2,  60,   "Real-time info, social media integration (X)"],
    [5,  "Perplexity",      "Perplexity AI","AI Search",              12.6,  45,   "Research, answer engine with citations"],
    [6,  "Remini",          "Bending Spoons","Photo Enhancement",     10.1,  120,  "AI photo enhancement & restoration"],
    [7,  "Claude",          "Anthropic",    "Chatbot / Assistant",    5.4,   25,   "Writing, analysis, coding, safety-focused"],
    [8,  "Character.ai",    "Character.AI", "AI Characters",          5.3,   40,   "Role-play, entertainment, companion chatbots"],
    [9,  "Yandex Translate","Yandex",       "Translation",            5.1,   30,   "AI-powered language translation"],
    [10, "ChatNow",         "ChatNow Inc.", "Chatbot / Assistant",    5.0,   20,   "Chat companion, writing assistant"],
    [11, "AI Mirror",       "AI Mirror",    "Photo / Selfie",         4.6,   35,   "AI selfies, avatars, photo generation"],
    [12, "Kimi",            "Moonshot AI",  "Chatbot / Reasoning",    4.4,   18,   "Long-context chat, document analysis (China)"],
    [13, "Facetune",        "Bending Spoons","Photo Editing",         4.0,   55,   "AI-powered portrait & selfie editing"],
    [14, "Glam AI",         "Glam AI",      "Beauty / Photo",         3.3,   15,   "AI beauty filters and photo editing"],
    [15, "PictureThis",     "Glority LLC",  "Plant Identification",   3.3,   28,   "AI plant & flower identification"],
    [16, "Genie",           "Genie AI",     "Chatbot / Assistant",    2.7,   12,   "AI assistant, homework help, writing"],
    [17, "KLING AI",        "Kuaishou",     "Video Generation",       2.3,   10,   "AI video & image generation"],
    [18, "CHAI",            "Chai Research","AI Characters",          2.3,   18,   "Conversational AI characters, entertainment"],
    [19, "Chat & Ask AI",   "Vulcan Labs",  "Chatbot / Assistant",    1.8,   8,    "General AI assistant, Q&A"],
    [20, "Viggle AI",       "Viggle",       "Video / Meme",           1.5,   6,    "AI meme & video generation"],
]

for row_data in consumer_data:
    ws1.append(row_data)

style_header(ws1, len(consumer_headers))
style_rows(ws1, len(consumer_data), len(consumer_headers))
ws1.column_dimensions["A"].width = 6
ws1.column_dimensions["B"].width = 20
ws1.column_dimensions["C"].width = 18
ws1.column_dimensions["D"].width = 22
ws1.column_dimensions["E"].width = 26
ws1.column_dimensions["F"].width = 28
ws1.column_dimensions["G"].width = 45

# ── Sheet 2: Top 20 Enterprise Gen AI Apps ──
ws2 = wb.create_sheet("Enterprise Gen AI Apps")
ws2.sheet_properties.tabColor = "548235"

ent_headers = ["Rank", "App / Tool", "Company", "Category", "Enterprise Adoption (%)", "Pricing (per user/mo)", "Primary Enterprise Use Case"]
ws2.append(ent_headers)

ent_data = [
    [1,  "ChatGPT Enterprise / Team",  "OpenAI",          "Chatbot / Productivity",     "62%",  "$25-60",    "Writing, coding, research, data analysis"],
    [2,  "Microsoft 365 Copilot",       "Microsoft",       "Productivity Suite AI",      "57%",  "$30-39",    "Document drafting, email, spreadsheets, presentations"],
    [3,  "GitHub Copilot",              "Microsoft/GitHub", "Code Assistant",             "48%",  "$19-39",    "Code generation, completion, review, documentation"],
    [4,  "Google Gemini for Workspace", "Google",          "Productivity Suite AI",      "38%",  "$48-60",    "Docs, Sheets, Gmail, Meet AI assistance"],
    [5,  "Salesforce Einstein / Agentforce", "Salesforce", "CRM AI",                    "35%",  "$50-125",   "Sales automation, customer insights, service agents"],
    [6,  "Claude for Business",         "Anthropic",       "Chatbot / Productivity",     "30%",  "$25-30",    "Analysis, writing, coding, compliance-sensitive tasks"],
    [7,  "AWS Bedrock",                 "Amazon",          "AI Platform / Gateway",      "28%",  "Usage-based","Multi-model access, RAG, enterprise AI orchestration"],
    [8,  "Azure OpenAI Service",        "Microsoft",       "AI Platform / Gateway",      "27%",  "Usage-based","GPT model hosting with enterprise security & compliance"],
    [9,  "Adobe Firefly",               "Adobe",           "Creative / Design AI",       "24%",  "$4.99-83",  "Image generation, design, creative content for marketing"],
    [10, "Jasper AI",                   "Jasper",          "Marketing Content",          "22%",  "$49-125",   "Marketing copy, blog posts, ad content generation"],
    [11, "Perplexity Enterprise Pro",   "Perplexity AI",   "AI Search / Research",       "18%",  "$40-50",    "Enterprise research, competitive intelligence, citations"],
    [12, "IBM Watsonx.ai",              "IBM",             "AI Platform",                "17%",  "Usage-based","Enterprise AI deployment, fine-tuning, governance"],
    [13, "Cursor",                      "Anysphere",       "Code Assistant",             "15%",  "$20-40",    "AI-native code editor, multi-file editing"],
    [14, "Midjourney",                  "Midjourney",      "Image Generation",           "14%",  "$10-60",    "Creative design, concept art, marketing visuals"],
    [15, "Runway ML",                   "Runway",          "Video Generation",           "12%",  "$12-76",    "AI video editing, motion graphics, VFX"],
    [16, "ElevenLabs",                  "ElevenLabs",      "Voice / Audio AI",           "11%",  "$5-99",     "Voice cloning, text-to-speech, audio content"],
    [17, "Notion AI",                   "Notion",          "Productivity / Knowledge",   "10%",  "$8-10 add-on","Writing assistance, summarization, knowledge management"],
    [18, "Grammarly AI",               "Grammarly",       "Writing Assistant",          "10%",  "$15-25",    "Writing enhancement, tone adjustment, enterprise comms"],
    [19, "Leonardo AI",                 "Leonardo.AI",     "Image Generation",           "8%",   "$12-48",    "Asset generation, game art, product imagery"],
    [20, "Stability AI (Stable Diffusion)", "Stability AI","Image Generation Platform",  "7%",   "Usage-based","Open-source image generation, fine-tuning, enterprise API"],
]

for row_data in ent_data:
    ws2.append(row_data)

style_header(ws2, len(ent_headers))
style_rows(ws2, len(ent_data), len(ent_headers))
ws2.column_dimensions["A"].width = 6
ws2.column_dimensions["B"].width = 34
ws2.column_dimensions["C"].width = 20
ws2.column_dimensions["D"].width = 26
ws2.column_dimensions["E"].width = 24
ws2.column_dimensions["F"].width = 22
ws2.column_dimensions["G"].width = 52

# ── Sheet 3: Top 20 LLM APIs for Enterprise ──
ws3 = wb.create_sheet("Enterprise LLM APIs")
ws3.sheet_properties.tabColor = "BF8F00"

api_headers = [
    "Rank", "Provider", "Top Model(s)", "Top Quality Score",
    "# Models Available", "Input Price ($/MTok)", "Output Price ($/MTok)",
    "Max Context Window", "Key Enterprise Strengths"
]
ws3.append(api_headers)

api_data = [
    [1,  "OpenAI",       "GPT-5.4 Pro, GPT-4.1",          94, 62,  "$1.25-$2.50",  "$8-$10",     "128K-1M",   "Largest ecosystem, highest throughput (10K RPM), broadest tool support"],
    [2,  "Anthropic",    "Claude Opus 4.6, Sonnet 4.5",    92, 13,  "$3-$5",        "$15-$25",    "200K",      "Best safety/compliance, top SWE-bench (80.9%), regulated industries"],
    [3,  "Google",       "Gemini 3 Pro, Flash Preview",    90, 32,  "$0.04-$1.25",  "$3.75-$10",  "1M-2M",     "Largest context window, multimodal leader, Vertex AI integration"],
    [4,  "xAI",          "Grok 4.1 Fast",                  87, 10,  "$0.50-$2",     "$5-$15",     "128K",      "Real-time data access, fast inference, X/Twitter integration"],
    [5,  "NVIDIA",       "Nemotron, NIM microservices",    85, 11,  "Usage-based",  "Usage-based", "128K",      "On-prem deployment, GPU-optimized inference, NIM containers"],
    [6,  "ByteDance",    "Doubao Pro",                     85, 5,   "$0.30-$0.80",  "$1-$3",      "128K",      "Low-cost, strong multilingual (CJK), high-volume consumer scale"],
    [7,  "Alibaba/Qwen", "Qwen 3, Qwen-Max",              85, 51,  "$0.15-$1",     "$0.60-$5",   "128K-1M",   "Most models (51), open-weight options, competitive pricing"],
    [8,  "Moonshot AI",  "Kimi-K2",                        85, 4,   "$0.50-$1",     "$2-$5",      "200K",      "Long-context specialist, strong reasoning, competitive China market"],
    [9,  "Perplexity",   "Sonar Pro, Sonar Reasoning",     85, 5,   "$1-$3",        "$5-$15",     "128K",      "Built-in web search, citation-grounded outputs, research focus"],
    [10, "MiniMax",      "MiniMax-01",                     83, 9,   "$0.20-$0.70",  "$1.10-$3",   "1M-4M",     "Ultra-long context (4M tokens), competitive Chinese AI market"],
    [11, "Arcee AI",     "Arcee Agent, SuperNova",         82, 8,   "$0.15-$0.50",  "$0.50-$2",   "128K",      "Domain-specific fine-tuning, enterprise model merging"],
    [12, "Inception",    "Mercury, Diffusion LLM",         81, 3,   "$0.25-$1",     "$1-$4",      "128K",      "Novel diffusion-based LLM architecture, UAE-backed"],
    [13, "Mistral AI",   "Mistral Large, Small 3.2",       78, 25,  "$0.04-$2",     "$0.12-$6",   "128K",      "European sovereignty, open-weight models, cost-efficient"],
    [14, "Amazon",       "Nova Pro, Nova Lite (Bedrock)",   78, 5,   "$0.06-$0.80",  "$0.24-$3.20","300K",      "Native AWS integration, Bedrock orchestration, enterprise security"],
    [15, "DeepSeek",     "DeepSeek R1, V3.2",              78, 11,  "$0.14-$0.55",  "$0.28-$2.50","128K",      "Ultra-low pricing, strong reasoning, open-source models"],
    [16, "Meta/Llama",   "Llama 4 Maverick, Scout",        77, 15,  "Free (open)",  "Free (open)","128K-10M",  "Open-weight leader, self-hosted flexibility, no API lock-in"],
    [17, "Allen AI",     "OLMo 2, Tülu 3",                75, 7,   "Free (open)",  "Free (open)","128K",      "Fully open-source (data + weights), research-grade transparency"],
    [18, "Baidu",        "ERNIE 4.5, ERNIE X1",           75, 5,   "$0.30-$1",     "$1-$4",      "128K",      "Chinese enterprise leader, integrated with Baidu Cloud ecosystem"],
    [19, "Cohere",       "Command A, Command R+",          60, 4,   "$0.50-$2.50",  "$2-$10",     "128K",      "Enterprise RAG specialist, multilingual embeddings, data privacy"],
    [20, "Together AI",  "Hosted Llama / Mixtral / Qwen",  "-", 100,  "$0.10-$3",    "$0.10-$10",  "128K",      "Model marketplace, fine-tuning platform, cost-efficient inference"],
]

for row_data in api_data:
    ws3.append(row_data)

style_header(ws3, len(api_headers))
style_rows(ws3, len(api_data), len(api_headers))
ws3.column_dimensions["A"].width = 6
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 32
ws3.column_dimensions["D"].width = 18
ws3.column_dimensions["E"].width = 18
ws3.column_dimensions["F"].width = 20
ws3.column_dimensions["G"].width = 22
ws3.column_dimensions["H"].width = 18
ws3.column_dimensions["I"].width = 60

# Freeze top row on each sheet
for ws in [ws1, ws2, ws3]:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

output_path = "/workspace/Top_GenAI_Apps_and_LLM_APIs_2026.xlsx"
wb.save(output_path)
print(f"Spreadsheet saved to {output_path}")
