"""Generate an Excel comparison of AI-compute deals.

Compares the two SpaceX (xAI) compute-rental deals - with Anthropic and with
Google - against TeraWulf's data-center lease with Anthropic.

Run:  python generate_deal_comparison.py
Output:  AI_Compute_Deals_Comparison.xlsx

All figures are sourced from public reporting / regulatory filings (see the
"Sources" sheet). Figures are approximate and reflect reporting as of
July 2026.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "AI_Compute_Deals_Comparison.xlsx"

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
ATTR_FONT = Font(name="Calibri", size=11, bold=True, color="1F2937")
CELL_FONT = Font(name="Calibri", size=11, color="1F2937")

TITLE_FILL = PatternFill("solid", fgColor="1F3864")
DEAL_A_FILL = PatternFill("solid", fgColor="2E75B6")   # SpaceX-Anthropic
DEAL_B_FILL = PatternFill("solid", fgColor="548235")   # SpaceX-Google
DEAL_C_FILL = PatternFill("solid", fgColor="BF8F00")   # TeraWulf-Anthropic
ATTR_FILL = PatternFill("solid", fgColor="D9E1F2")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TOP_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ---------------------------------------------------------------------------
# Data
# ---------------------------------------------------------------------------
DEALS = ["SpaceX (xAI) - Anthropic", "SpaceX (xAI) - Google", "TeraWulf - Anthropic"]

# Each row: (attribute, value_dealA, value_dealB, value_dealC)
ROWS = [
    ("Infrastructure provider",
     "SpaceX (compute from xAI, merged into SpaceX in Feb 2026)",
     "SpaceX (compute from xAI, merged into SpaceX in Feb 2026)",
     "TeraWulf Inc. (via subsidiary landlord Raylan Data LLC)"),
    ("Customer / counterparty",
     "Anthropic PBC",
     "Google (Google Cloud)",
     "Anthropic PBC"),
    ("Announced",
     "Late May 2026",
     "June 5, 2026 (SEC/IPO registration filing)",
     "July 6, 2026 (8-K filing)"),
    ("Deal type",
     "Compute-capacity rental (GPU lease)",
     "Compute-capacity rental (GPU lease)",
     "Real-estate / data-center lease (powered shell + IT load)"),
    ("What is provided",
     "All available compute at the Colossus 1 data center",
     "~110,000 NVIDIA GPUs plus CPUs, memory & related components",
     "~401 MW of critical IT load at a purpose-built AI campus"),
    ("Location / facility",
     "Colossus 1, near Memphis, Tennessee",
     "SpaceX data center (facility not specified; Colossus 2 reserved for xAI)",
     "Justified Data Campus, Hawesville, Kentucky (former Century Aluminum site)"),
    ("Monthly payment",
     "$1.25 billion / month",
     "$920 million / month",
     "~$79 million / month average (implied by ~$19B over 20 yrs)"),
    ("Total / contracted value",
     "~$45 billion over ~3 years",
     "~$30 billion over the contract life (32 months)",
     "~$19 billion of contracted revenue over the initial term"),
    ("Term / duration",
     "~3 years (through May 2029)",
     "~32 months (Oct 2026 - June 2029)",
     "20 years, plus options to extend up to 10 more (two 5-yr renewals)"),
    ("Start date",
     "Ramp at reduced fees May-June 2026; full rate thereafter",
     "Reduced-fee ramp through Sept 2026; full rate from Oct 2026",
     "Phased delivery beginning 2H 2027; full 401 MW by early 2028"),
    ("End date",
     "May 2029",
     "June 2029",
     "~20 years after each phase is delivered (~2047-2048)"),
    ("Ramp-up terms",
     "Reduced fees during May & June 2026 ramp period",
     "Reduced fee during ramp-up through September 2026",
     "Multi-phase buildout; rent begins as each premises is delivered"),
    ("Termination / cancellation",
     "Either party on 90 days' notice after Dec 31, 2026",
     "Either party on 90 days' notice after Dec 31, 2026",
     "Long-term lease (no short-term walk-away); backed by investment-grade credit"),
    ("Hardware / IP ownership",
     "SpaceX owns and operates the hardware",
     "SpaceX owns hardware; Google retains IP for its content, models & data",
     "TeraWulf owns/operates the facility; Anthropic is the tenant"),
    ("Strategic rationale",
     "Anthropic was compute-constrained; raised usage limits the day of the deal",
     "Short-term 'bridge capacity' for surging Gemini Enterprise demand",
     "Anthropic securing dedicated, long-duration purpose-built compute"),
    ("Notable context",
     "Part of SpaceX pre-IPO story; combined w/ Google deal ~$2.17B/mo (~$26B/yr)",
     "Google is a longtime SpaceX investor; deal disclosed in IPO registration",
     "TeraWulf pivoting from bitcoin mining; also sold Abernathy JV stake to Fluidstack"),
]

SOURCES = [
    ("SpaceX - Anthropic",
     "Anthropic to pay ~$1.25B/mo through May 2029 for all Colossus 1 compute (~$45B).",
     "TechCrunch / CNBC / Engadget / Yahoo Finance, May-June 2026"),
    ("SpaceX - Google",
     "Google to pay $920M/mo Oct 2026-June 2029 for ~110,000 NVIDIA GPUs (~$30B).",
     "TechCrunch, CNBC, The Verge, TNW, France24, Yahoo Finance, June 5, 2026"),
    ("TeraWulf - Anthropic",
     "20-year lease, ~401 MW at Justified Data Campus, ~$19B contracted revenue.",
     "TeraWulf press release / SEC 8-K (Ex 99.1) / Reuters / TechTimes, July 6, 2026"),
    ("SpaceX combined",
     "Two disclosed compute deals ~ $26B/yr (~$2.17B/mo); aggregate value > $70B.",
     "Reuters / TNW / Yahoo Finance, June 2026"),
]


def style_title(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30


def build_comparison(ws):
    ncols = 1 + len(DEALS)
    style_title(ws, "AI Compute Deals - Side-by-Side Comparison", ncols)

    header_row = 3
    headers = ["Attribute"] + DEALS
    fills = [ATTR_FILL, DEAL_A_FILL, DEAL_B_FILL, DEAL_C_FILL]
    for col, (text, fill) in enumerate(zip(headers, fills), start=1):
        c = ws.cell(row=header_row, column=col, value=text)
        c.font = ATTR_FONT if col == 1 else HEADER_FONT
        c.fill = fill
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[header_row].height = 32

    for i, row in enumerate(ROWS):
        r = header_row + 1 + i
        for col, value in enumerate(row, start=1):
            c = ws.cell(row=r, column=col, value=value)
            c.border = BORDER
            c.alignment = TOP_WRAP
            if col == 1:
                c.font = ATTR_FONT
                c.fill = ATTR_FILL
            else:
                c.font = CELL_FONT
                if i % 2 == 1:
                    c.fill = ALT_FILL

    ws.column_dimensions["A"].width = 26
    for col in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 42
    ws.freeze_panes = "B4"


def build_key_metrics(ws):
    style_title(ws, "Key Metrics at a Glance", 5)
    headers = ["Deal", "Monthly payment", "Total value", "Term", "Type"]
    hr = 3
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=hr, column=col, value=text)
        c.font = HEADER_FONT
        c.fill = TITLE_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[hr].height = 28

    data = [
        ("SpaceX (xAI) - Anthropic", "$1.25B / month", "~$45B", "~3 yrs (to May 2029)", "GPU compute rental"),
        ("SpaceX (xAI) - Google", "$920M / month", "~$30B", "32 mo (Oct 26-Jun 29)", "GPU compute rental"),
        ("TeraWulf - Anthropic", "~$79M / month (avg)", "~$19B", "20 yrs (+10 opt.)", "Data-center lease"),
    ]
    fills = [DEAL_A_FILL, DEAL_B_FILL, DEAL_C_FILL]
    for i, rowvals in enumerate(data):
        r = hr + 1 + i
        for col, value in enumerate(rowvals, start=1):
            c = ws.cell(row=r, column=col, value=value)
            c.border = BORDER
            c.alignment = TOP_WRAP if col == 1 else CENTER
            c.font = CELL_FONT
            if col == 1:
                c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                c.fill = fills[i]
    ws.column_dimensions["A"].width = 28
    for col, w in zip("BCDE", (18, 12, 22, 20)):
        ws.column_dimensions[col].width = w


def build_sources(ws):
    style_title(ws, "Notes & Sources", 3)
    headers = ["Deal", "Key terms", "Reporting sources"]
    hr = 3
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=hr, column=col, value=text)
        c.font = HEADER_FONT
        c.fill = TITLE_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[hr].height = 28

    for i, rowvals in enumerate(SOURCES):
        r = hr + 1 + i
        for col, value in enumerate(rowvals, start=1):
            c = ws.cell(row=r, column=col, value=value)
            c.border = BORDER
            c.alignment = TOP_WRAP
            c.font = CELL_FONT
            if i % 2 == 1:
                c.fill = ALT_FILL

    note_r = hr + len(SOURCES) + 2
    note = ("Disclaimer: Figures are approximate and compiled from public news "
            "reporting and regulatory filings as of July 6, 2026. Monthly figure "
            "for the TeraWulf lease is an average implied by ~$19B of contracted "
            "revenue over a 20-year term; actual rent is phased and begins on "
            "delivery of each premises.")
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=3)
    c = ws.cell(row=note_r, column=1, value=note)
    c.font = Font(name="Calibri", size=10, italic=True, color="595959")
    c.alignment = TOP_WRAP
    ws.row_dimensions[note_r].height = 70

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 46


def main():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Deal Comparison"
    build_comparison(ws1)

    build_key_metrics(wb.create_sheet("Key Metrics"))
    build_sources(wb.create_sheet("Sources"))

    wb.save(OUTPUT_FILE)
    print(f"Wrote {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
