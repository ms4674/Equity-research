"""Generate an Excel workbook comparing Forward Deployed Engineer (FDE) roles
across Palantir, OpenAI, Anthropic, Microsoft, and Accenture.

Data compiled from public job postings and market commentary (mid-2026).
Compensation figures are approximate market ranges and vary by level/location.

Run:  python generate_fde_comparison.py
Output:  FDE_Comparison.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "FDE_Comparison.xlsx"

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------
NAVY = "1F3864"
BLUE = "2E5496"
LIGHT = "D9E1F2"
LIGHTER = "EAF0FA"
WHITE = "FFFFFF"
GREY = "808080"

TITLE_FONT = Font(name="Calibri", size=18, bold=True, color=WHITE)
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color=WHITE)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
LABEL_FONT = Font(name="Calibri", size=11, bold=True, color=NAVY)
BODY_FONT = Font(name="Calibri", size=10, color="000000")
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color=GREY)

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

TOP_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def header_fill():
    return PatternFill("solid", fgColor=BLUE)


def title_fill():
    return PatternFill("solid", fgColor=NAVY)


def zebra(i):
    return PatternFill("solid", fgColor=LIGHTER if i % 2 else WHITE)


# ---------------------------------------------------------------------------
# Data
# ---------------------------------------------------------------------------
COMPANIES = ["Palantir", "OpenAI", "Anthropic", "Microsoft", "Accenture"]

# Each attribute maps to a list of values (one per company, same order as COMPANIES)
ATTRIBUTES = [
    ("Role title(s)", [
        "Forward Deployed Software Engineer (FDSE); internally \"Delta\"",
        "Forward Deployed Engineer (FDE) / Forward Deployed Software Engineer (FDSWE)",
        "Forward Deployed Engineer (FDE), Applied AI team (a.k.a. Applied AI Engineer)",
        "Forward Deployed Engineer (FDE) / Senior FDE",
        "Forward Deployed Engineer / Forward Deployed AI Engineer",
    ]),
    ("Origin / maturity", [
        "Pioneered the role (~2000s); the original gold standard, largest by history",
        "Function formalized early 2025; largest dedicated FDE team among pure AI labs",
        "Established under Applied AI team; scaling across US/EU/UK",
        "Newer AI-era FDE org built around Copilot / Frontier platform (2025-2026)",
        "Microsoft-Accenture FDE practice launched Mar 2026; also standalone FDE roles",
    ]),
    ("Core mission", [
        "Embed with customers to configure Palantir platforms (Foundry/Gotham/AIP) to solve their hardest data problems",
        "Turn frontier-model research into production systems inside strategic enterprise customers",
        "Embed with strategic customers to drive transformational Claude adoption in production",
        "Embed with customers to move agentic AI (Copilot) from pilot to production",
        "Embed in client accounts to close the 'AI last mile' and ship production-grade AI-native delivery",
    ]),
    ("What they build", [
        "Custom workflows, web apps, data pipelines on Foundry/Gotham/AIP",
        "Full-stack systems, RAG pipelines, agent workflows, evals on OpenAI APIs",
        "Production apps on Claude; MCP servers, sub-agents, agent skills",
        "Agentic Copilot solutions, data pipelines, custom apps, reference architectures",
        "Agentic workflows, RAG pipelines, vector/semantic search across MS/Snowflake/Palantir/etc.",
    ]),
    ("Underlying platform / tech", [
        "Foundry, Gotham, Apollo, AIP; Python, Java, TypeScript",
        "OpenAI API, GPT models, ChatGPT Enterprise; Python, JavaScript",
        "Claude, Model Context Protocol (MCP); Python, TypeScript",
        "Azure, Copilot, Frontier Suite, agent frameworks; Azure AI stack",
        "Multi-vendor: OpenAI, Claude, Vertex, Microsoft, Snowflake, SAP, Palantir; LangGraph/CrewAI/AutoGen",
    ]),
    ("Feedback loop to product", [
        "Field insights feed core product/platform teams",
        "Field feedback shapes Research & Product model roadmaps (eval-driven)",
        "Codify repeatable patterns back to Product & Engineering",
        "Field insights influence AI platform roadmaps & reference architectures",
        "Bring 'field intelligence' back into Tech Reinvention; codify reusable patterns",
    ]),
    ("Typical experience required", [
        "Strong CS fundamentals; new-grad to senior; FAANG-level coding bar",
        "5+ years engineering / technical deployment incl. customer-facing work",
        "3-4+ years customer-facing + production LLM experience",
        "Senior; solution architecture + hands-on coding experience",
        "4+ years production engineering; agentic AI + cloud-native delivery",
    ]),
    ("AI/LLM depth expected", [
        "Growing (AIP); traditionally data/platform-centric more than LLM-native",
        "High: RAG, evals, agents, prompt engineering; must have shipped LLM systems",
        "Very high: frontier-model-native, agents, MCP, evals",
        "High: agentic Copilot, RAG, agent identity/permissions, human-in-the-loop",
        "High: agentic orchestration, RAG, MLOps/LLMOps, multi-provider abstraction",
    ]),
    ("Customer-facing intensity", [
        "Very high - startup-CTO-like ownership, exec stakeholder management",
        "High - primary technical POC for the account",
        "High - white-glove deployment + long-term relationship building",
        "High - bridge between customer devs and Microsoft engineering",
        "Very high - technical owner on the ground inside client accounts",
    ]),
    ("Typical travel", [
        "Substantial; ~25-50% on-site (incl. secure/government facilities)",
        "Up to ~50%; embed on-site for days/weeks",
        "~25% (some postings 25-50%), varies by location",
        "Flexible: on-site, virtual, or hybrid",
        "Varies by account; hybrid (often ~3 days/week in office) + client travel",
    ]),
    ("Primary customer sectors", [
        "Government, defense, intelligence, healthcare, finance, manufacturing",
        "Fortune 500 across healthcare, financial services, legal, technology",
        "Regulated enterprises: financial services, healthcare/life sciences, IT; federal variant",
        "Enterprise & commercial across industries",
        "Broad enterprise across all industries (consulting client base)",
    ]),
    ("Base salary (US, approx.)", [
        "~$130K-$255K (Levels.fyi median ~$215K)",
        "~$162K-$325K",
        "~$200K-$300K",
        "Competitive senior-SWE band (not publicly fixed)",
        "Varies by level/location; ~$120K-$300K+ band",
    ]),
    ("Total comp (US, approx.)", [
        "~$170K-$415K (median ~$215K); travel stipends possible",
        "~$350K-$550K+ (mid-to-senior; equity/PPUs)",
        "~$350K-$550K+ (benchmarked near OpenAI; equity heavy)",
        "Competitive big-tech TC incl. stock",
        "~$190K-$420K depending on level/specialization",
    ]),
    ("Comp structure notes", [
        "Public-company RSUs; travel-related stipends for deployed life",
        "Base + significant equity (PPUs); pays at/above top of FDE market",
        "Base + equity (~60-70% of package at frontier labs)",
        "Base + RSUs + bonus (public company)",
        "Consulting comp: base + bonus + utilization/level-driven; less equity",
    ]),
    ("Business model for the role", [
        "Revenue generator: often billed to client on top of software license",
        "Drives adoption/expansion of high-value API & enterprise contracts",
        "Drives Claude enterprise adoption & expansion",
        "Drives Azure/Copilot platform consumption & adoption",
        "Billable services delivery; scales via thousands of engineers",
    ]),
    ("Scale / headcount signal", [
        "Largest FDE population historically (defense-tech volume leader)",
        "Largest dedicated FDE team among pure AI labs",
        "Smaller, senior, selective team; scaling",
        "Growing FDE community within Microsoft",
        "Thousands of AI-skilled engineers (esp. via MS partnership)",
    ]),
    ("Biggest differentiator", [
        "The original; deepest deployment culture & autonomy; owns messy data problems",
        "Frontier-model access + tightest research feedback loop",
        "Claude + MCP-native; safety-forward; senior + selective",
        "Hyperscaler distribution + enterprise trust/security integration",
        "Industry breadth + change management + massive delivery scale",
    ]),
]

# Pros / cons per company for the analysis sheet
PROS_CONS = {
    "Palantir": (
        "Original FDE culture; extreme ownership; strong brand; deep gov/defense access; clear CTO-track growth.",
        "Heavy travel & deployed lifestyle; less LLM-native historically; secretive/high-stakes environments.",
    ),
    "OpenAI": (
        "Frontier models; top-of-market comp; direct research feedback loop; huge enterprise demand.",
        "High technical bar (must have shipped LLM systems); up to ~50% travel; fast-moving/ambiguous.",
    ),
    "Anthropic": (
        "Frontier Claude + MCP; safety-forward mission; senior, selective; strong comp with equity upside.",
        "Fewer openings; very high bar; equity valuation risk (private); ~25%+ travel.",
    ),
    "Microsoft": (
        "Hyperscaler scale & distribution; enterprise trust/security; flexible on-site/hybrid; public-company stability.",
        "Role newer/less defined than Palantir; tied to Azure/Copilot stack; large-org processes.",
    ),
    "Accenture": (
        "Industry breadth; massive delivery scale; multi-vendor exposure; strong MS partnership; many entry paths.",
        "Consulting comp (less equity, utilization pressure); billable model; outcome ownership varies by account.",
    ),
}

# Data sources
SOURCES = [
    ("Palantir FDSE job posting (Lever)", "jobs.lever.co/palantir"),
    ("Palantir blog - A day in the life of an FDSE", "blog.palantir.com"),
    ("OpenAI Careers - Forward Deployed Engineer / FDSWE (SF)", "openai.com/careers"),
    ("Paraform - OpenAI FDE Guide (2026)", "paraform.com"),
    ("Anthropic Applied AI FDE postings (Greenhouse via aggregators)", "underprompt.com / fde10x.com"),
    ("Microsoft FDE postings (Flexa / Redmond Senior FDE)", "flexa.careers / crawljobs.com"),
    ("Accenture Newsroom - MS FDE practice launch (Mar 2026)", "newsroom.accenture.com"),
    ("Accenture Careers - FDE / Forward Deployed AI Engineer", "accenture.com/careers / builtin.com"),
    ("Levels.fyi & FDE market commentary (2026)", "levels.fyi / fde.academy"),
]


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
def style_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_comparison_sheet(wb):
    ws = wb.active
    ws.title = "Comparison"
    ws.sheet_view.showGridLines = False

    ncols = 1 + len(COMPANIES)  # attribute label + 5 companies

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value="Forward Deployed Engineer (FDE) - Company Comparison")
    t.font = TITLE_FONT
    t.fill = title_fill()
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34

    # Subtitle row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(row=2, column=1,
                value="Compiled from public job postings & market commentary (mid-2026). "
                      "Compensation figures are approximate and vary by level/location.")
    s.font = SUBTITLE_FONT
    s.fill = title_fill()
    s.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    # Header row (row 3)
    hr = 3
    c = ws.cell(row=hr, column=1, value="Attribute")
    c.font = HEADER_FONT
    c.fill = header_fill()
    c.alignment = CENTER_LEFT
    c.border = BORDER
    for j, comp in enumerate(COMPANIES, start=2):
        c = ws.cell(row=hr, column=j, value=comp)
        c.font = HEADER_FONT
        c.fill = header_fill()
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[hr].height = 22

    # Data rows
    r = hr + 1
    for idx, (label, values) in enumerate(ATTRIBUTES):
        lab = ws.cell(row=r, column=1, value=label)
        lab.font = LABEL_FONT
        lab.fill = PatternFill("solid", fgColor=LIGHT)
        lab.alignment = TOP_WRAP
        lab.border = BORDER
        for j, val in enumerate(values, start=2):
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = BODY_FONT
            cell.alignment = TOP_WRAP
            cell.fill = zebra(idx)
            cell.border = BORDER
        r += 1

    style_col_widths(ws, [26, 34, 34, 34, 34, 34])
    ws.freeze_panes = "B4"

    # Footer note
    note = ws.cell(row=r + 1, column=1,
                   value="Note: 'FDE' role definitions differ by company. Palantir and OpenAI/Anthropic "
                         "run product-owned FDE teams; Accenture runs a services/delivery FDE practice "
                         "(often paired with Microsoft). Figures are indicative, not offers.")
    note.font = NOTE_FONT
    note.alignment = TOP_WRAP
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 2, end_column=ncols)


def build_analysis_sheet(wb):
    ws = wb.create_sheet("Pros & Cons")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1, value="FDE Roles - Strengths & Trade-offs")
    t.font = TITLE_FONT
    t.fill = title_fill()
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    hr = 2
    for j, head in enumerate(["Company", "Strengths / Pros", "Trade-offs / Cons"], start=1):
        c = ws.cell(row=hr, column=j, value=head)
        c.font = HEADER_FONT
        c.fill = header_fill()
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[hr].height = 22

    r = hr + 1
    for i, comp in enumerate(COMPANIES):
        pros, cons = PROS_CONS[comp]
        cc = ws.cell(row=r, column=1, value=comp)
        cc.font = LABEL_FONT
        cc.fill = PatternFill("solid", fgColor=LIGHT)
        cc.alignment = TOP_WRAP
        cc.border = BORDER
        p = ws.cell(row=r, column=2, value=pros)
        p.font = BODY_FONT
        p.alignment = TOP_WRAP
        p.fill = zebra(i)
        p.border = BORDER
        n = ws.cell(row=r, column=3, value=cons)
        n.font = BODY_FONT
        n.alignment = TOP_WRAP
        n.fill = zebra(i)
        n.border = BORDER
        ws.row_dimensions[r].height = 58
        r += 1

    style_col_widths(ws, [16, 60, 60])


def build_comp_sheet(wb):
    ws = wb.create_sheet("Compensation")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value="FDE Compensation Snapshot (US, approx. - mid-2026)")
    t.font = TITLE_FONT
    t.fill = title_fill()
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    hr = 2
    for j, head in enumerate(["Company", "Base salary (US)", "Total comp (US)", "Notes"], start=1):
        c = ws.cell(row=hr, column=j, value=head)
        c.font = HEADER_FONT
        c.fill = header_fill()
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[hr].height = 22

    rows = [
        ["Palantir", "~$130K-$255K", "~$170K-$415K (median ~$215K)", "Public RSUs; travel stipends"],
        ["OpenAI", "~$162K-$325K", "~$350K-$550K+", "Equity/PPUs; top of market"],
        ["Anthropic", "~$200K-$300K", "~$350K-$550K+", "Equity ~60-70% of package"],
        ["Microsoft", "Senior-SWE band", "Competitive big-tech TC", "Base + RSUs + bonus"],
        ["Accenture", "~$120K-$300K+", "~$190K-$420K", "Consulting comp; less equity"],
    ]
    r = hr + 1
    for i, row in enumerate(rows):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = LABEL_FONT if j == 1 else BODY_FONT
            cell.fill = PatternFill("solid", fgColor=LIGHT) if j == 1 else zebra(i)
            cell.alignment = TOP_WRAP if j == 4 else CENTER_LEFT
            cell.border = BORDER
        ws.row_dimensions[r].height = 30
        r += 1

    note = ws.cell(row=r + 1, column=1,
                   value="Ranges are indicative market figures aggregated from job postings, Levels.fyi, "
                         "and 2026 compensation commentary. Actual offers vary by level, location, and equity valuation.")
    note.font = NOTE_FONT
    note.alignment = TOP_WRAP
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 2, end_column=4)

    style_col_widths(ws, [16, 22, 30, 40])


def build_sources_sheet(wb):
    ws = wb.create_sheet("Sources")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:B1")
    t = ws.cell(row=1, column=1, value="Sources & References")
    t.font = TITLE_FONT
    t.fill = title_fill()
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    hr = 2
    for j, head in enumerate(["Reference", "Domain"], start=1):
        c = ws.cell(row=hr, column=j, value=head)
        c.font = HEADER_FONT
        c.fill = header_fill()
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[hr].height = 22

    r = hr + 1
    for i, (ref, dom) in enumerate(SOURCES):
        a = ws.cell(row=r, column=1, value=ref)
        a.font = BODY_FONT
        a.alignment = TOP_WRAP
        a.fill = zebra(i)
        a.border = BORDER
        b = ws.cell(row=r, column=2, value=dom)
        b.font = BODY_FONT
        b.alignment = TOP_WRAP
        b.fill = zebra(i)
        b.border = BORDER
        r += 1

    disclaimer = ws.cell(row=r + 1, column=1,
                         value="Disclaimer: Data compiled mid-2026 from public sources for informational "
                               "comparison only. Role scope, comp, and travel vary by team, level, and geography.")
    disclaimer.font = NOTE_FONT
    disclaimer.alignment = TOP_WRAP
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 2, end_column=2)

    style_col_widths(ws, [58, 40])


def main():
    wb = Workbook()
    build_comparison_sheet(wb)
    build_analysis_sheet(wb)
    build_comp_sheet(wb)
    build_sources_sheet(wb)
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT} with sheets: {[ws.title for ws in wb.worksheets]}")


if __name__ == "__main__":
    main()
