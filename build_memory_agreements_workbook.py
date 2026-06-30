"""
Build a granular Excel workbook cataloguing global memory (DRAM / HBM / NAND)
supply & pricing agreements for hyperscalers and neoclouds.

Run:  python3 build_memory_agreements_workbook.py
Output: Memory_Agreements_Hyperscalers_Neoclouds.xlsx

All figures are sourced from public reporting (Reuters, Bloomberg, TrendForce,
company press releases, earnings calls, trade press). This is a research compile
for equity-research use; figures are approximate and several agreements are
non-binding LOIs/MOUs or remain in negotiation. See the "Sources" sheet.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = "Memory_Agreements_Hyperscalers_Neoclouds.xlsx"

# ----------------------------------------------------------------------------- styling
NAVY = "1F3864"
BLUE = "2E5496"
LIGHT = "D9E1F2"
LIGHT2 = "EAF0FA"
AMBER = "FFF2CC"
GREEN = "E2EFDA"
RED = "FCE4D6"
GREY = "F2F2F2"

HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
SUB_FONT = Font(name="Calibri", bold=False, color="FFFFFF", size=10, italic=True)
SECTION_FONT = Font(name="Calibri", bold=True, color=NAVY, size=12)
BOLD = Font(name="Calibri", bold=True, size=10)
NORM = Font(name="Calibri", size=10)

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_C = Alignment(wrap_text=True, vertical="center", horizontal="center")
CENTER = Alignment(vertical="center", horizontal="center")


def hdr_fill(color=BLUE):
    return PatternFill("solid", fgColor=color)


def style_header_row(ws, row, ncols, color=BLUE):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = hdr_fill(color)
        cell.alignment = WRAP_C
        cell.border = BORDER


def title_block(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=title)
    t.font = TITLE_FONT
    t.fill = hdr_fill(NAVY)
    t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = SUB_FONT
    s.fill = hdr_fill(NAVY)
    s.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[2].height = 18


def write_table(ws, start_row, headers, rows, widths, hdr_color=BLUE,
                zebra=True, table_name=None):
    ncols = len(headers)
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=h)
    style_header_row(ws, start_row, ncols, hdr_color)
    r = start_row + 1
    for i, row in enumerate(rows):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = NORM
            cell.alignment = WRAP
            cell.border = BORDER
            if zebra and i % 2 == 1:
                cell.fill = hdr_fill(GREY)
        r += 1
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    if table_name:
        ref = f"A{start_row}:{get_column_letter(ncols)}{r-1}"
        tab = Table(displayName=table_name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        # only add if no merged conflict; safe here
        try:
            ws.add_table(tab)
        except Exception:
            pass
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return r


wb = Workbook()

# =========================================================================== 1. COVER
ws = wb.active
ws.title = "Cover"
ws.sheet_view.showGridLines = False
for col in "ABCDEFGH":
    ws.column_dimensions[col].width = 16

ws.merge_cells("A1:H1")
c = ws["A1"]
c.value = "GLOBAL MEMORY SUPPLY & PRICING AGREEMENTS"
c.font = Font(bold=True, size=20, color=NAVY)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:H2")
c = ws["A2"]
c.value = "Hyperscalers & Neoclouds  |  DRAM · HBM · NAND/SSD  |  Equity-Research Compile"
c.font = Font(bold=True, size=12, color=BLUE)
ws.row_dimensions[2].height = 22

ws.merge_cells("A3:H3")
ws["A3"].value = "Data as of: late June 2026  |  Compiled from public reporting (see Sources tab)"
ws["A3"].font = Font(italic=True, size=10, color="595959")

intro = [
    "",
    "PURPOSE",
    "A granular catalogue of the memory supply agreements reshaping the AI build-out: who is buying memory "
    "(hyperscalers, AI labs, GPU vendors, neoclouds), who is supplying it (SK hynix, Samsung, Micron, Kioxia, "
    "SanDisk, CXMT), and the commercial terms (volume, value, duration, prepayments, pricing mechanism).",
    "",
    "THE STRUCTURAL SHIFT (2025 -> 2026)",
    "The memory market has flipped from a cyclical, quarterly spot/contract model to a 'reserve capacity' model "
    "built on multi-year Long-Term Agreements (LTAs / Strategic Customer Agreements). Key features now standard:",
    "   - Duration: 3-5 years (common formats '3yr fixed + 2yr option', '2yr fixed + 3yr option')",
    "   - Pricing: price-band mechanics — a FLOOR (downside protection for supplier) and a CEILING (upside cap for buyer)",
    "   - Prepayments: 10-30% of contract value upfront (vs <5% historically); forfeited as penalty if buyer under-takes",
    "   - Structure: take-or-pay (buyer pays for committed volume regardless of usage)",
    "   - Some Big Tech (Nvidia, Google, Amazon) are offering to fund supplier fabs/equipment to lock allocation",
    "",
    "WHY IT MATTERS",
    "HBM, server DDR5 and enterprise SSD are effectively sold out for 2026 (HBM for ~3 years). DRAM contract prices "
    "rose ~90-95% QoQ in Q1'26 and a further ~58-63% in Q2'26; NAND ~+70-75% QoQ in Q2'26. The supply deficit is the "
    "widest since 2011. ~60-70% of server DDR5 volume is already locked under LTAs.",
    "",
    "WORKBOOK CONTENTS",
    "   1. Cover (this sheet)",
    "   2. Agreements Master — granular, one row per agreement (the core data set)",
    "   3. Neocloud Exposure — neocloud memory exposure (mostly indirect, via GPU allocation)",
    "   4. Pricing Benchmarks — DRAM/HBM/NAND contract & spot pricing, QoQ moves",
    "   5. HBM Generations & Cost — per-stack / per-GB economics by generation",
    "   6. Supplier Allocation — LTA coverage, capacity & 'sold-out' status by supplier",
    "   7. Hyperscaler Capex — context on AI infra spend driving demand",
    "   8. Sources — full citation list",
    "",
    "CAVEATS",
    "Many headline agreements are non-binding LOIs/MOUs or remain in negotiation; values are analyst estimates and "
    "vary with the memory cycle. Specific contract pricing is almost never disclosed — per-unit figures are market "
    "benchmarks, not contract prices. Treat all numbers as directional. Verify against primary filings before use.",
]
r = 5
for line in intro:
    cell = ws.cell(row=r, column=1, value=line)
    if line in ("PURPOSE", "THE STRUCTURAL SHIFT (2025 -> 2026)", "WHY IT MATTERS",
                "WORKBOOK CONTENTS", "CAVEATS"):
        cell.font = SECTION_FONT
    else:
        cell.font = Font(size=10, color="333333")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if len(line) > 90:
        ws.row_dimensions[r].height = 30
    r += 1

# =========================================================================== 2. AGREEMENTS MASTER
ws = wb.create_sheet("Agreements Master")
ws.sheet_view.showGridLines = False
title_block(ws,
            "AGREEMENTS MASTER — Memory Supply & Pricing Agreements",
            "One row per agreement. Status key: Signed / MOU-LOI (non-binding) / In negotiation. Values are public estimates.",
            16)

headers = [
    "ID", "Supplier", "Buyer / Customer", "Buyer Type", "Region",
    "Memory Product", "Generation / Spec", "Status", "Announced / Date",
    "Term", "Volume / Capacity", "Contract Value (est.)", "Prepayment",
    "Pricing Mechanism", "End-Use / Platform", "Notes & Source",
]

# Buyer Type categories: Hyperscaler / AI Lab / GPU-Accelerator Vendor / Neocloud / OEM
A = [
    # --- HBM / GPU-vendor anchored ---
    ["A01", "SK hynix", "NVIDIA", "GPU/Accelerator Vendor", "Korea/US",
     "HBM", "HBM4 (+ HBM3E)", "Signed (multi-yr co-dev)", "2025-2026",
     "Multi-year", "~60-70% of HBM4 volume for Vera Rubin", "Undisclosed",
     "n/a (co-dev)", "Price + supply security prioritized over spot",
     "NVIDIA Vera Rubin / Rubin Ultra", "Multi-year co-development pact; SK hynix the lead HBM4 supplier. Reuters/TheNextWeb 2026. Reportedly sought ~50-70% price premium vs HBM3E."],

    ["A02", "Samsung Electronics", "NVIDIA", "GPU/Accelerator Vendor", "Korea/US",
     "HBM", "HBM4 (1c DRAM, 4nm base die, 13 Gbps)", "Qualified / supplying", "2026 (mass prod. Feb 2026)",
     "Multi-year (alloc.)", "~25-30% of HBM4 Vera Rubin volume", "Undisclosed",
     "Annual->multi-year transition", "Vera Rubin platform", "Samsung first to mass-produce HBM4 (Feb 2026); cleared for Vera Rubin. 2026 HBM allocation sold out. Also makes Groq LPU on Samsung 4nm foundry. Korea Herald 2026."],

    ["A03", "Micron", "NVIDIA", "GPU/Accelerator Vendor", "US",
     "HBM + SSD + SOCAMM2", "HBM4 36GB 12-high (>2.8 TB/s)", "Signed (volume shipping)", "Mar 2026 (HBM4 vol. ship)",
     "Multi-year", "Part of 2026 HBM sold-out allocation", "Within ~$22B SCA deposits",
     "Floor + ceiling (see A07)", "Vera Rubin platform", "Micron only supplier qualified for full HBM4 + PCIe Gen6 SSD + SOCAMM2 suite for Rubin. HBM4 ramping ~2x faster than HBM3E. Finexus/CryptoBriefing 2026."],

    ["A04", "Samsung Electronics", "AMD", "GPU/Accelerator Vendor", "Korea/US",
     "HBM4 + DDR5", "HBM4 (13 Gbps, 3.3 TB/s) + DDR5", "Signed (MOU)", "2026-03-18",
     "Multi-year (strategic)", "Primary HBM4 supplier for MI455X", "Undisclosed",
     "Strategic alignment", "AMD Instinct MI455X GPU; EPYC 'Venice'; Helios rack", "MOU: Samsung primary HBM4 supplier for MI455X + DDR5 for 6th-Gen EPYC. Foundry/packaging discussions too. Samsung/AMD newsroom 2026-03-18. (Samsung also supplied HBM3E for MI350X/MI355.)"],

    # --- OpenAI / Stargate ---
    ["A05", "SK hynix", "OpenAI", "AI Lab", "Korea/US/Global",
     "HBM (+ DRAM wafers)", "HBM / advanced DRAM", "MOU/LOI (non-binding)", "2025-10-01",
     "Multi-year (to 2029)", "Up to 900k DRAM wafers/mo (combined w/ Samsung)", "~$70-71B combined (4-yr est.)",
     "TBD (to be negotiated)", "Stargate AI data centers", "LOI signed Seoul Oct 2025; production system for up to 900k wafers/mo (>2x current HBM industry capacity). Scaled back after Stargate Abilene expansion abandoned (Mar 2026). Reuters/SK 2025."],

    ["A06", "Samsung Electronics", "OpenAI", "AI Lab", "Korea/US/Global",
     "DRAM wafers + HBM", "Commodity DDR5 + HBM", "MOU/LOI (non-binding)", "2025-10-01",
     "Multi-year (to 2029)", "Up to 900k DRAM wafers/mo (combined w/ SK hynix)", "~$70-71B combined (4-yr est.)",
     "TBD", "Stargate AI data centers (incl. Korea)", "Samsung supplies undiced wafers; Samsung SDS partnered to build/operate Stargate data centers. ~40% of global DRAM output if fully realized. Bloomberg/Reuters 2025."],

    # --- Micron framework ---
    ["A07", "Micron", "4 large + 3 medium customers (CSP/data-center, consumer, auto)", "Hyperscaler (multiple)", "US/Global",
     "DRAM + HBM + NAND", "HBM3E/HBM4, DDR5, eSSD", "Signed (16 SCAs)", "2026 (FQ3'26 call, Jun 2026)",
     "5 yrs (2026-2030); 3 yrs auto", "~20% of DRAM bits; ~1/3 of NAND volume", "~$100B min. revenue (14 of 16 SCAs)",
     "~$22B upfront customer deposits", "Floor (margin > any prior cycle) + ceiling (~Q2'26 mkt price)", "Take-or-pay. When fully executed, ~50%+ of Micron revenue under SCAs; ~40% at fixed/ceiling price. TechTimes/Karsane 2026."],

    # --- SK hynix LTAs ---
    ["A08", "SK hynix", "Microsoft", "Hyperscaler", "US/Korea",
     "DRAM", "DDR5 (server)", "In final negotiation", "2026 (Q2)",
     "3 yrs (from 2026)", "Tens of trillions of won (volume)", "Tens of billions USD",
     "10-30% prepayment", "Min. price floor over term", "Azure AI infrastructure", "Multi-year DDR5 LTA in final coordination. TrendForce/Hankyung/Green Economy 2026."],

    ["A09", "SK hynix", "Google (Alphabet)", "Hyperscaler", "US/Korea",
     "DRAM (+ HBM)", "General-purpose DRAM; HBM3E", "In negotiation", "2026 (H1)",
     "Up to 5 yrs (+2 yr ext.)", "General-purpose DRAM volume", "Undisclosed",
     "10-30% prepayment", "Floor pricing; ext. tied to next-gen HBM", "Google AI infra / TPU", "SK hynix is Google's primary HBM3E supplier; 5yr commodity-DRAM LTA + 2yr extension contingent on next-gen HBM supply. TrendForce/Aju 2026."],

    # --- Samsung LTAs (negotiation) ---
    ["A10", "Samsung Electronics", "Microsoft", "Hyperscaler", "US/Korea",
     "DRAM (+ HBM, eSSD)", "DDR5 / multi-product", "In late-stage negotiation", "2026 (Q1-Q2)",
     "3-5 yrs", "Multi-product volume", ">$10B prepayment (reported, MSFT)",
     "10-30% prepayment", "Floor + ceiling band", "Azure AI infrastructure", "Samsung shifting all new contracts to >=3yr LTAs. MSFT prepayment reported >$10B, adjusted vs shortfall. Korea JoongAng/TrendForce 2026."],

    ["A11", "Samsung Electronics", "Google (Alphabet)", "Hyperscaler", "US/Korea",
     "DRAM (+ HBM, eSSD)", "DDR5 / multi-product", "In late-stage negotiation", "2026 (Q1-Q2)",
     "3-5 yrs", "Multi-product volume", "10-30% of value",
     "10-30% prepayment", "Floor + ceiling band", "Google AI infra / TPU", "Among AMD/Microsoft/Google trio Samsung expected to secure 3yr commitments with. TrendForce/Aju 2026."],

    # --- Samsung Tesla (foundry context) ---
    ["A12", "Samsung Electronics (Foundry)", "Tesla", "AI/Compute (foundry)", "US (Taylor, TX)",
     "Logic (foundry, not memory)", "2nm AI chips (AI5/AI6)", "Signed", "2025 (reported)",
     "Multi-year", "Foundry capacity (Taylor fab)", "$16.5B",
     "n/a", "n/a", "Tesla next-gen AI chips", "CONTEXT ROW (foundry, not memory): $16.5B chip-supply deal; production at Taylor TX from 2027. Reuters/Korea Herald 2026. Included to show Samsung customer linkage."],

    # --- China domestic ---
    ["A13", "CXMT (ChangXin)", "Tencent", "Hyperscaler", "China",
     "DRAM (server)", "DDR5 (server)", "Signed (reported)", "2026-06-29",
     "3-5 yrs", ">20B yuan worth of server DRAM", "~$2.94-3.0B (~20B yuan)",
     "Price bands + prepayments", "LTA price band", "Tencent cloud/AI servers", "Largest domestic Chinese DRAM LTA; HBM inclusion unconfirmed. Reuters via Chosunbiz 2026-06-29. CXMT ~300k wpm capacity (Hefei x2, Beijing)."],

    ["A14", "CXMT (ChangXin)", "Alibaba Cloud / ByteDance / Xiaomi / Lenovo", "Hyperscaler / OEM", "China",
     "DRAM (server)", "DDR5", "In discussion", "2026",
     "Multi-year (3-5 yrs)", "TBD", "TBD",
     "Price bands + prepayments", "TBD", "China cloud / AI / devices", "CXMT in talks with additional Chinese internet majors (per IPO prospectus customers). DDR5 yields still trail Korean/US peers. Reuters/Digg 2026."],

    # --- NAND / SSD ---
    ["A15", "Kioxia", "Hyperscalers (multiple)", "Hyperscaler (multiple)", "Japan/Global",
     "NAND / enterprise SSD", "218-layer; QLC", "In negotiation (multi-yr)", "2026 (Q1-Q2)",
     "Through 2028-2029", "NAND volume allocation", "Undisclosed",
     "Min. volume + pre-negotiated price", "Seller's-market LTA", "AI server / enterprise storage", "Kioxia president confirmed NAND seller's market; LTAs through 2028-2029. eSSD prices +~80% in Q1'26. DailyAlpha/TrendForce 2026."],

    ["A16", "Samsung / SK hynix(Solidigm) / Micron / SanDisk", "Hyperscalers (CSPs)", "Hyperscaler (multiple)", "Global",
     "Enterprise SSD (NAND)", "QLC (128TB class), V9/236L", "Signing LTAs", "2026 (H1)",
     "Multi-year / multi-quarter", "~20% of NAND volume to LTA (2026)", "Undisclosed",
     "Pre-negotiated price + volume", "Allocation by LTA", "AI training datasets / inference / cold data", "Industry-wide eSSD LTA wave; QLC (e.g., SanDisk 128TB 'Stargate' SSD) ramping. eSSD +~80% QoQ Q1'26; NAND +70-75% Q2'26. TrendForce 2026."],

    # --- Big tech fab funding ---
    ["A17", "SK hynix", "NVIDIA / Google / Amazon (proposals)", "Hyperscaler / GPU Vendor", "Global",
     "DRAM / HBM (capacity)", "Capacity/equipment funding", "Proposals (under review)", "2026 (May)",
     "Tied to multi-yr LTAs", "Fab lines + advanced equipment", "Tens of trillions of won (proposed)",
     "~30% down payments", "Floor + ceiling", "Securing future HBM/DRAM allocation", "Big Tech offering to fund SK hynix fabs/equipment in exchange for allocation; SK hynix cautious on single-customer line risk. Seoul Economic Daily 2026-05-08."],
]

r = write_table(ws, 4, headers, A,
                widths=[5, 16, 22, 16, 12, 16, 20, 16, 14, 14, 22, 18, 16, 22, 22, 46],
                hdr_color=BLUE)

# color status column (col 8) by status keyword
for row in range(5, r):
    sval = (ws.cell(row=row, column=8).value or "").lower()
    if "signed" in sval:
        ws.cell(row=row, column=8).fill = hdr_fill(GREEN)
    elif "mou" in sval or "loi" in sval:
        ws.cell(row=row, column=8).fill = hdr_fill(AMBER)
    elif "negotiation" in sval or "proposal" in sval or "discussion" in sval or "qualified" in sval:
        ws.cell(row=row, column=8).fill = hdr_fill(RED)
ws.row_dimensions[4].height = 30

# =========================================================================== 3. NEOCLOUD EXPOSURE
ws = wb.create_sheet("Neocloud Exposure")
ws.sheet_view.showGridLines = False
title_block(ws,
            "NEOCLOUD MEMORY EXPOSURE",
            "Neoclouds rarely buy memory directly — exposure is INDIRECT via Nvidia GPU systems (embedded HBM) + power/backlog. Memory cost flows through GPU BoM.",
            10)

nheaders = ["Neocloud", "Tier", "HQ / Region", "Anchor Customers",
            "Contracted Backlog (RPO)", "Contracted Power", "GPU Fleet / Orders",
            "NVIDIA Relationship", "Memory Sourcing Model", "Notes & Source"]
N = [
    ["CoreWeave", "Giant", "US", "Microsoft, OpenAI, Meta, Anthropic",
     "$99.4B (Q1'26; from $66.8B YE'25)", "3.5 GW contracted (>1 GW energized)",
     "Large Nvidia GB200/GB300; Vera Rubin pipeline", "Strategic shareholder + priority supply",
     "Indirect: HBM embedded in Nvidia systems; priority allocation", "FY25 rev ~$5B (fastest cloud to that mark). GPU-collateralized ABS ($8.5B IG). ModulEdge/Frankk 2026."],
    ["Nebius", "Giant", "Netherlands/EU", "Microsoft, Meta, multiple AI labs",
     "~$50B cumulative (through Q1'26)", "3+ GW (800MW-1GW connected by YE'26)",
     "~30k GPUs (H200); +22k Blackwell/Ultra in 2026", "$2B strategic equity (PIPE) + priority",
     "Indirect via Nvidia allocation", "$17.4-19.4B 5yr Microsoft deal; ~$3B Meta deal; $7-9B ARR target YE'26. SEC filings 2025-26."],
    ["Crusoe", "Giant", "US", "OpenAI (Stargate Abilene), others",
     "n/a (private)", "Stargate Abilene campus builder", "Nvidia GB200+ at Stargate",
     "Nvidia partner / allocation", "Indirect; HBM via Nvidia systems at Stargate", "~$10B valuation; built OpenAI Stargate Abilene campus (expansion later restructured). LinkedIn/Synergy 2025-26."],
    ["Lambda", "Giant", "US", "Enterprises, AI labs, Microsoft (reported)",
     "n/a (private)", "Expanding", "Nvidia H100/H200/Blackwell",
     "Nvidia partner", "Indirect via Nvidia allocation", "Raised >$1.5B late 2025; eyeing IPO. ModulEdge 2026."],
    ["Nscale", "Emerging/Giant", "UK/Europe", "Microsoft",
     "Large (Microsoft Portugal deal)", "Expanding (Portugal)", "66,000 Rubin GPUs (Portugal, for Microsoft)",
     "Nvidia partner", "Indirect; Rubin = HBM4-heavy", "$14.6B valuation (Mar 2026). Deploying Rubin in Portugal for Microsoft. ModulEdge 2026."],
    ["Long tail (Fluidstack, Together AI, Applied Digital, IREN, Nscale peers, Core42, etc.)",
     "Emerging / Brokers", "Global", "Varied AI labs & enterprises",
     "Smaller", "Varied", "Mixed Nvidia / AMD",
     "Mostly arms-length", "Indirect; least allocation leverage", "Without direct strategic ties, face higher cost & supply uncertainty in a sold-out HBM market. Synergy/SemiAnalysis 2025-26."],
]
r = write_table(ws, 4, nheaders, N,
                widths=[26, 16, 16, 26, 24, 24, 26, 26, 28, 44], hdr_color="548235")
ws.row_dimensions[4].height = 30

# note box
note_r = r + 1
ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=10)
c = ws.cell(row=note_r, column=1,
            value="KEY POINT: Neoclouds are memory price-TAKERS, not direct contract counterparties. Their HBM exposure is bundled inside Nvidia GPU systems "
                  "(e.g., a Vera Rubin NVL72 rack carries ~20.7 TB of HBM4). Rising HBM4 prices ($500-600+/stack; Bernstein sees $53/GB by 2027) inflate Nvidia "
                  "system prices, which flow into neocloud capex and ultimately GPU rental rates. Allocation priority (via Nvidia equity/partnership) is the real lever.")
c.font = Font(italic=True, size=10, color=NAVY)
c.fill = hdr_fill(AMBER)
c.alignment = WRAP
c.border = BORDER
ws.row_dimensions[note_r].height = 60

# =========================================================================== 4. PRICING BENCHMARKS
ws = wb.create_sheet("Pricing Benchmarks")
ws.sheet_view.showGridLines = False
title_block(ws,
            "MEMORY PRICING BENCHMARKS",
            "Market reference prices (TrendForce, mid-2026). These are BENCHMARKS, not contract prices — actual LTA pricing is undisclosed and sits within floor/ceiling bands.",
            6)

# 4a contract QoQ moves
ws.cell(row=4, column=1, value="A) Contract Price Moves (QoQ %) — the 2025-2026 surge").font = SECTION_FONT
ph = ["Segment", "Q4 2025 QoQ", "Q1 2026 QoQ", "Q2 2026 QoQ", "Supply backdrop", "Source"]
P = [
    ["DRAM (conventional/contract)", "+18-23% (server)", "+90-95%", "+58-63%", "Capacity diverted to HBM; deficit widest since 2011", "TrendForce/Khan Capital"],
    ["PC DRAM", "rising", "+100%+", "rising, supply-starved", "Crowded out by server/HBM", "NAND Research/TrendForce"],
    ["Server DRAM (DDR5)", "+18-23%", "~+90%", "elevated, LTA-driven", "60-70% locked via LTAs", "TrendForce"],
    ["NAND flash (overall)", "rising", "+50-60%", "+70-75%", "AI server + enterprise SSD demand", "TrendForce/NAND Research"],
    ["Enterprise SSD", "rising", "+53-58% (~80% in qtr)", "shortage-driven, no relief", "Inventories at historic lows", "TrendForce"],
    ["Client SSD", "rising", "+40%", "sharp increase", "Spillover from enterprise", "NAND Research"],
    ["HBM (contract)", "annual mechanism", "lagged spot (below DDR5/wafer)", "2027 talks: multiples higher", "Sold out 2026; ~3yrs for HBM", "TrendForce"],
]
r = write_table(ws, 5, ph, P, widths=[30, 18, 22, 22, 38, 24], hdr_color=BLUE)

# 4b spot/contract unit prices
r2 = r + 2
ws.cell(row=r2, column=1, value="B) Reference Unit Prices (mid-2026, TrendForce session/contract data)").font = SECTION_FONT
uh = ["Item", "Type", "Price (USD)", "Unit", "Note"]
U = [
    ["DDR5 16Gb (2Gx8) 4800/5600", "DRAM chip contract", 46.83, "per chip (session avg)", "Range $31.5-$61.0"],
    ["DDR5 16Gb (2Gx8) eTT", "DRAM chip", 23.50, "per chip", "Untested/grain"],
    ["DDR4 16Gb (2Gx8) 3200", "DRAM chip contract", 73.74, "per chip", "Legacy phase-out premium"],
    ["DDR4 8Gb (1Gx8) 3200", "DRAM chip", 36.00, "per chip", ""],
    ["DDR5 UDIMM 16GB 4800/5600", "Module (spot)", 208.50, "per module", "Weekly avg"],
    ["DDR5 RDIMM 32GB 4800/5600", "Server module (spot)", 1335.00, "per module", "Range $1,250-$1,600; +4.7% w/w"],
    ["DDR4 UDIMM 16GB 3200", "Module (spot)", 149.58, "per module", ""],
    ["HBM2e (16GB, 8-Hi)", "HBM stack", 120.00, "per stack", "$7.50/GB"],
    ["HBM3 (24GB, 12-Hi)", "HBM stack", 200.00, "per stack", "$8.33/GB"],
    ["HBM3E (36GB, 12-Hi)", "HBM stack", 300.00, "per stack", "$8.33/GB; ~20-26 wk lead time"],
    ["HBM4 (48GB, 16-Hi) est.", "HBM stack", 500.00, "per stack", "$10.42/GB; some quotes $550-600+"],
]
rr = write_table(ws, r2 + 1, uh, U, widths=[34, 22, 16, 24, 28], hdr_color="548235")
for row in range(r2 + 2, rr):
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    ws.cell(row=row, column=3).alignment = CENTER

# =========================================================================== 5. HBM GENERATIONS & COST
ws = wb.create_sheet("HBM Generations & Cost")
ws.sheet_view.showGridLines = False
title_block(ws,
            "HBM GENERATIONS — SPEC & COST ECONOMICS",
            "Per-stack / per-GB economics by generation, and Nvidia per-GPU / per-rack memory cost build (analyst estimates).",
            8)

gh = ["Generation", "Capacity/Stack", "Bandwidth/Stack", "Layers", "Cost/Stack (USD)", "Cost/GB (USD)", "Lead Use", "Note"]
G = [
    ["HBM2e", "16 GB", "460 GB/s", "8-Hi", 120, 7.50, "Prior-gen accelerators", "Legacy"],
    ["HBM3", "24 GB", "819 GB/s", "12-Hi", 200, 8.33, "H100", "SK hynix first-mover"],
    ["HBM3E", "36 GB", "1.18 TB/s", "12-Hi", 300, 8.33, "H200 / B200", "~2/3 of 2026 HBM shipments"],
    ["HBM4 (est.)", "48 GB", ">2 TB/s", "16-Hi", 500, 10.42, "Vera Rubin / MI455X", "Quotes $550-600+; +50-70% premium vs HBM3E"],
]
r = write_table(ws, 4, gh, G, widths=[16, 16, 16, 10, 18, 16, 22, 34], hdr_color=BLUE)
for row in range(5, r):
    ws.cell(row=row, column=5).number_format = '#,##0'
    ws.cell(row=row, column=6).number_format = '#,##0.00'

r2 = r + 2
ws.cell(row=r2, column=1, value="Nvidia per-GPU / per-rack memory economics (analyst estimates)").font = SECTION_FONT
xh = ["Platform / Config", "HBM Content", "Memory Cost (est.)", "Source / Note"]
X = [
    ["H100 (per GPU)", "80 GB HBM3 @ 3.4 TB/s", "—", "Reference baseline"],
    ["B200 (per GPU)", "192 GB HBM3E (8 stacks) @ 8.0 TB/s", "~$2,400 memory alone", "Memory > logic die cost. Silicon Analysts 2026"],
    ["Vera Rubin (per GPU)", "288 GB HBM4 (6+ 12-Hi stacks) @ up to 22 TB/s", "~$3,000-3,600", "6+ stacks x $500-600. AI CERTs 2026"],
    ["Rubin NVL72 rack", "20.7 TB HBM4 @ 1.6 PB/s total", "~$2.0M (MS) / up to ~$3.2M (Bernstein)", "Rack ASP $7.8M (MS) to $9.1M (Bernstein); HBM4 to ~$53/GB by 2027"],
]
rr = write_table(ws, r2 + 1, xh, X, widths=[26, 40, 30, 40], hdr_color="548235")

# =========================================================================== 6. SUPPLIER ALLOCATION
ws = wb.create_sheet("Supplier Allocation")
ws.sheet_view.showGridLines = False
title_block(ws,
            "SUPPLIER CAPACITY, ALLOCATION & SOLD-OUT STATUS",
            "How the three (plus challengers) are allocating capacity under the LTA regime. HBM wafer share of total DRAM input rising ~18%/22%/30% by YE 2025/26/27.",
            7)

sh = ["Supplier", "DRAM LTA coverage (2027E)", "HBM/DRAM market position", "Sold-out status",
      "Capacity actions", "Key customers", "Notes & Source"]
S = [
    ["SK hynix", "~18% of DRAM bits", "#1 DRAM rev (H1'25); ~50-55% HBM share",
     "DRAM/NAND/HBM sold out thru 2026; HBM sold out ~3 yrs", "Yongin cluster fab (Phase 1 early 2027); ~620k DRAM wpm w/ HBM conversion",
     "NVIDIA, OpenAI, Microsoft, Google", "HBM-first allocation; raising down payments to ~30%. Asiae/Sedaily 2026."],
    ["Samsung Electronics", "~30% of DRAM bits", "Top-3 DRAM; ~35-40% HBM share",
     "2026 HBM allocation sold out", "Ramp HBM toward ~250k wpm (+47%); record W110tr ($73.5B) 2026 capex; Taylor TX fab",
     "NVIDIA, AMD, (MSFT/Google neg.), Tesla(foundry)", "First to mass-produce HBM4 (Feb 2026). Moving all new contracts to >=3yr LTAs. Korea Herald 2026."],
    ["Micron", "~20% of DRAM bits", "#3 DRAM; ~5-10% HBM (rising fast)",
     "HBM sold out thru 2026; booked thru 2027, demand into 2028", "$100B Clay, NY complex (vol. ~FY2028); 16 SCAs signed",
     "NVIDIA + 4 large/3 medium CSP/auto", "Only full HBM4+SSD+SOCAMM2 supplier for Rubin. ~$22B deposits; ~$100B min rev. TechTimes 2026."],
    ["Kioxia (+ SanDisk alliance)", "NAND-focused", "Top NAND; with SanDisk ~37% of NAND w/ YMTC",
     "Multi-yr NAND LTAs thru 2028-2029", "218-layer ramp; QLC high-capacity (incl. 128TB class)",
     "North American hyperscalers, server OEMs", "Confirmed NAND seller's market. eSSD rev: Kioxia ~$2.22B, SanDisk ~$1.47B Q1'26. TrendForce/DailyAlpha 2026."],
    ["CXMT (China)", "Domestic China", "Largest China DRAM; DDR5 yields lag",
     "Ramping; new Shanghai DRAM fab", "~300k wpm (Hefei x2 + Beijing); IPO to fund expansion (~29.5B yuan)",
     "Tencent, Alibaba, ByteDance, Lenovo, Xiaomi", "$3B Tencent LTA (Jun 2026). Q1'26 rev +700% YoY. Domestic-substitution play vs Korean/US commodity DRAM. Reuters 2026."],
]
r = write_table(ws, 4, sh, S, widths=[24, 22, 28, 30, 38, 28, 44], hdr_color=BLUE)
ws.row_dimensions[4].height = 30

r2 = r + 2
ws.cell(row=r2, column=1, value="LTA structure 'new normal' (industry-wide)").font = SECTION_FONT
facts = [
    "Duration: 3-5 years; formats '3yr fixed + 2yr option' or '2yr fixed + 3yr option' (UBS).",
    "Prepayments: 10-30% of contract value (SK hynix setting ~30% down); historically <5%.",
    "Pricing: floor (margin protection) + ceiling (~current market) price bands; part fixed slightly below market, remainder floats.",
    "Take-or-pay; upfront payment forfeited as penalty if buyer under-takes committed volume.",
    "~60-70% of server DDR5 volume already secured via LTAs; ~20-30% of total DRAM bits and ~20% of NAND moving to LTAs (2026-27).",
    "Reserved for top-tier CSPs (MSFT, Google, Amazon, Meta, Alibaba, ByteDance) + GPU vendors (Nvidia, AMD); mid-tier pushed to spot at premium.",
    "Big Tech (Nvidia, Google, Amazon) offering to fund supplier fabs/equipment to strengthen allocation locks.",
]
rf = r2 + 1
for f in facts:
    cell = ws.cell(row=rf, column=1, value="•  " + f)
    cell.font = NORM
    cell.alignment = WRAP
    ws.merge_cells(start_row=rf, start_column=1, end_row=rf, end_column=7)
    rf += 1

# =========================================================================== 7. HYPERSCALER CAPEX
ws = wb.create_sheet("Hyperscaler Capex")
ws.sheet_view.showGridLines = False
title_block(ws,
            "HYPERSCALER AI CAPEX — DEMAND CONTEXT",
            "The capex wave driving memory demand. Memory is now a top input constraint; ~$650B aggregate hyperscaler infra spend expected in 2026 (+~80% YoY).",
            5)
ch = ["Player", "2026 Capex / Spend (reported)", "Memory-relevant commitments", "Notes & Source"]
C = [
    ["Aggregate hyperscalers", "~$650B infra (2026E, +~80% YoY)", "Driving DRAM/HBM/NAND shortage", "TechBriefly 2026"],
    ["Microsoft", "Azure capex pulled to ~$190B (2026)", "Samsung & SK hynix DDR5 LTAs (neg.); $10B+ to CoreWeave; $17.4B 5yr Nebius", "Frankk/TrendForce 2026"],
    ["Google (Alphabet)", "Large (custom TPU + infra)", "SK hynix HBM3E primary; 5yr DRAM LTA (neg.); offering fab funding", "Sedaily/TrendForce 2026"],
    ["Amazon (AWS)", "Large (Trainium + infra)", "Among CSPs locking LTAs; offering fab funding to SK hynix", "TrendForce/Sedaily 2026"],
    ["Meta", "6 GW GPU deployment (Helios/MI455X)", "AMD Helios (Samsung DDR5/HBM4); Nebius ~$3B; CoreWeave", "HotHardware/Frankk 2026"],
    ["OpenAI", "Stargate ($500B initiative, restructured)", "Samsung+SK hynix LOIs (up to 900k wafers/mo); Nvidia $100B; Crusoe build", "Reuters/SK 2025-26"],
    ["Alibaba / ByteDance / Tencent", "Large (China AI build-out)", "CXMT domestic DRAM LTAs (Tencent $3B signed; others in talks)", "Reuters 2026"],
]
r = write_table(ws, 4, ch, C, widths=[26, 30, 44, 38], hdr_color="548235")

# =========================================================================== 8. SOURCES
ws = wb.create_sheet("Sources")
ws.sheet_view.showGridLines = False
title_block(ws, "SOURCES", "Public reporting used to compile this workbook (accessed June 2026).", 4)
oh = ["#", "Source / Outlet", "Topic", "Reference"]
O = [
    [1, "Reuters (2025-10-01)", "OpenAI Stargate – Samsung & SK hynix LOIs, 900k wafers/mo", "reuters.com/business/media-telecom/samsung-sk-hynix-supply-memory-chips-openais-stargate-project-2025-10-01"],
    [2, "SK Group newsroom (2025-10-01)", "SK hynix–OpenAI LOI, HBM for Stargate", "eng.sk.com/news/sk-group-partners-with-openai"],
    [3, "Notebookcheck (2025-10-30)", "SK hynix sold out DRAM/NAND/HBM thru 2026", "notebookcheck.net"],
    [4, "TheNextWeb (2026)", "Nvidia–SK hynix multi-year HBM4 deal; 60-70% Vera Rubin", "thenextweb.com/news/nvidia-locks-in-sk-hynix"],
    [5, "Khan Capital (2026)", "Nvidia–SK hynix deal; DRAM +90-95% / +58-63% QoQ; HBM4 mid-$500", "khancapitals.com/nvidia-sk-hynix-memory-deal"],
    [6, "Seoul Economic Daily (2026-04-23)", "SK hynix HBM sold out ~3 yrs; LTA requests", "en.sedaily.com"],
    [7, "TechTimes (2026-06-25)", "Micron 16 SCAs, ~$100B min rev, floor/ceiling, $22B deposits", "techtimes.com/articles/319032"],
    [8, "Karsane / Micron FQ3'26", "Micron HBM sold out 2026; ~$22B deposits", "karsane.com"],
    [9, "Finexus / CryptoBriefing (2026)", "Micron HBM4 36GB 12-high; full Rubin suite", "api.finexus.net ; cryptobriefing.com"],
    [10, "TrendForce (2026-04-09)", "Samsung & SK hynix reset to 3-5yr LTAs", "trendforce.com/news/2026/04/09"],
    [11, "TrendForce (2026-05-04)", "LTAs deepen; premiums; SK hynix MSFT 3yr DDR5 / Google 5yr DRAM", "trendforce.com/news/2026/05/04"],
    [12, "TrendForce (2026-06-02)", "HBM 2027 contract prices to surge multiples", "trendforce.com/presscenter/news/20260602-13074"],
    [13, "TrendForce DRAM price page (2026)", "DDR5/DDR4 contract & spot unit prices", "trendforce.com/price/dram/dram_contract"],
    [14, "TrendForce (2025-10-29)", "DDR5 contract profitability to surpass HBM3E", "trendforce.com/presscenter/news/20251029-12758"],
    [15, "TrendForce (2026-06-11)", "Enterprise SSD +~80% QoQ; top-5 $18.46B 1Q26", "trendforce.com/presscenter/news/20260611-13092"],
    [16, "Samsung / AMD newsroom (2026-03-18)", "Samsung primary HBM4 for AMD MI455X + DDR5 EPYC", "news.samsung.com/global/samsung-and-amd ; amd.com/en/newsroom"],
    [17, "Bloomberg (2026-03-18)", "Samsung weighs 3-5yr deals; Samsung-AMD HBM4/DDR5", "bloomberg.com"],
    [18, "Korea Herald (2026)", "Samsung W110tr capex; HBM4 mass prod.; Tesla $16.5B foundry", "koreaherald.com/article/10699099"],
    [19, "Seoul Economic Daily (2026-05-08)", "Big Tech (Nvidia/Google/Amazon) offer to fund SK hynix fabs", "en.sedaily.com/society/2026/05/08"],
    [20, "Asia Business Daily (2026-04-01)", "Advance payments / LTAs; 3-5yr CSP, 3yr GPU vendors", "asiae.co.kr/en/article/2026040109351026768"],
    [21, "MK / UBS (2026)", "LTA formats; coverage 30%/20%/18% (Samsung/Micron/SK hynix)", "mk.co.kr/en/business/12059087"],
    [22, "Reuters via Chosunbiz / Goldsea (2026-06-29)", "CXMT–Tencent ~$3B server-DRAM LTA", "biz.chosun.com ; goldsea.com"],
    [23, "DailyAlpha / NAND Research (2026)", "Kioxia NAND seller's market; LTAs thru 2028-29; QoQ table", "dailyalpha.us ; nand-research.com"],
    [24, "Silicon Analysts (2026)", "HBM per-stack / per-GB cost; B200 memory $2,400", "siliconanalysts.com/data/hbm-pricing"],
    [25, "AI CERTs / NextWaves (2026)", "HBM4 $500-600+/stack; Rubin 6+ stacks; per-GPU $3,000-3,600", "aicerts.ai ; nextwavesinsight.com"],
    [26, "WCCFTech / Bernstein & Morgan Stanley (2026)", "Rubin NVL72 rack $7.8M-9.1M; HBM4 ~$53/GB 2027", "wccftech.com"],
    [27, "ModulEdge / Frankk / Synergy (2025-26)", "Neocloud tiers, backlog, power, Nvidia ties", "moduledge.com/blog/neocloud ; research.frankk.site"],
    [28, "TechBriefly / BigGo / Korea JoongAng (2026)", "MSFT/Google 3yr DRAM deals; prepayments; ~$650B capex", "techbriefly.com ; finance.biggo.com"],
]
r = write_table(ws, 4, oh, O, widths=[5, 34, 50, 58], hdr_color=NAVY)
for row in range(5, r):
    ws.cell(row=row, column=4).font = Font(size=9, color="1155CC")

# disclaimer
dr = r + 1
ws.merge_cells(start_row=dr, start_column=1, end_row=dr, end_column=4)
c = ws.cell(row=dr, column=1,
            value="DISCLAIMER: Compiled from secondary public sources for equity-research reference. Many agreements are non-binding "
                  "(LOI/MOU) or in negotiation; values are analyst estimates that move with the memory cycle. Specific contract pricing is not "
                  "publicly disclosed — unit prices shown are market benchmarks. Not investment advice. Verify against primary filings before use.")
c.font = Font(italic=True, size=9, color="808080")
c.alignment = WRAP
ws.row_dimensions[dr].height = 54

wb.save(OUT)
print("Saved", OUT)
print("Sheets:", wb.sheetnames)
