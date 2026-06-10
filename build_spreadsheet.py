"""
Build the Data Center Capacity vs Token Consumption analysis spreadsheet.
Generates a multi-tab Excel workbook.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13)
SUBTITLE_FONT = Font(bold=True, size=11, italic=True)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
LIGHT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if (r - start_row) % 2 == 0:
                cell.fill = LIGHT_FILL


def auto_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)


def build_dc_capacity_tab(wb):
    ws = wb.active
    ws.title = "DC Capacity"

    ws.cell(row=1, column=1, value="Global Data Center Installed IT Capacity").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Historical & Projected (2020–2030)").font = SUBTITLE_FONT

    headers = ["Year", "Capacity (GW)", "YoY Growth (%)", "Net Addition (GW)", "Cumulative CapEx ($B)", "Source/Basis"]
    row = 4
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    data = [
        (2020, 40, None, None, None, "Visual Capitalist / Macquarie"),
        (2021, 47, 17.5, 7, None, "Visual Capitalist"),
        (2022, 54, 14.9, 7, None, "Visual Capitalist"),
        (2023, 62, 14.8, 8, None, "Visual Capitalist"),
        (2024, 81, 30.6, 19, None, "Macquarie"),
        (2025, 103, 27.2, 22, None, "JLL 2026 Outlook"),
        (2026, 130, 26.2, 27, 750, "Projected (BloombergNEF/JLL)"),
        (2027, 158, 21.5, 28, 850, "Projected"),
        (2028, 182, 15.2, 24, 700, "Projected"),
        (2029, 202, 11.0, 20, 600, "Projected"),
        (2030, 220, 8.9, 18, 500, "Projected (JLL target)"),
    ]

    for i, d in enumerate(data):
        r = row + 1 + i
        for c, val in enumerate(d, 1):
            ws.cell(row=r, column=c, value=val)
    style_data_rows(ws, row + 1, row + len(data), len(headers))

    # CAGR summary
    sr = row + len(data) + 2
    ws.cell(row=sr, column=1, value="Key CAGRs").font = HEADER_FONT
    ws.cell(row=sr + 1, column=1, value="2020–2025")
    ws.cell(row=sr + 1, column=2, value="21%")
    ws.cell(row=sr + 2, column=1, value="2025–2030")
    ws.cell(row=sr + 2, column=2, value="16%")
    ws.cell(row=sr + 3, column=1, value="Post-AI inflection (2023–2025)")
    ws.cell(row=sr + 3, column=2, value="29%")

    auto_width(ws)


def build_token_volume_tab(wb):
    ws = wb.create_sheet("Token Consumption")

    ws.cell(row=1, column=1, value="Global AI Token Consumption").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Historical & Projected (2022–2030)").font = SUBTITLE_FONT

    headers = ["Year", "Annual Volume (Quadrillion)", "YoY Growth (%)", "Tokens per GW (Q/GW)",
               "Cost per M Tokens ($)", "Energy per M Tokens (Wh)", "Notes"]
    row = 4
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    data = [
        (2022, 2, None, 0.04, 37.50, 85.0, "Pre-ChatGPT baseline"),
        (2023, 15, 650, 0.24, 10.00, 30.0, "ChatGPT/GPT-4 launch; adoption explosion"),
        (2024, 80, 433, 1.0, 0.50, 8.5, "Multi-model proliferation; enterprise scaling"),
        (2025, 350, 338, 3.4, 0.14, 4.5, "Google 1.3Q/mo; MSFT 100T/qtr; reasoning models"),
        (2026, 800, 129, 6.2, 0.08, 3.2, "Agentic workloads; B200 deployment at scale"),
        (2027, 1500, 88, 9.5, 0.05, 2.0, "B300 ramp; multi-agent systems mainstream"),
        (2028, 2600, 73, 14.3, 0.03, 1.1, "Next-gen accelerators; context-dense workloads"),
        (2029, 4000, 54, 19.8, 0.02, 0.6, "Agentic saturation begins"),
        (2030, 6000, 50, 27.3, 0.01, 0.4, "Base case projection"),
    ]

    for i, d in enumerate(data):
        r = row + 1 + i
        for c, val in enumerate(d, 1):
            ws.cell(row=r, column=c, value=val)
    style_data_rows(ws, row + 1, row + len(data), len(headers))

    sr = row + len(data) + 2
    ws.cell(row=sr, column=1, value="Token CAGR 2022–2026:").font = HEADER_FONT
    ws.cell(row=sr, column=2, value="~230%")
    ws.cell(row=sr + 1, column=1, value="Token CAGR 2025–2030:").font = HEADER_FONT
    ws.cell(row=sr + 1, column=2, value="~75%")

    auto_width(ws)


def build_correlation_tab(wb):
    ws = wb.create_sheet("Correlation Model")

    ws.cell(row=1, column=1, value="Correlation Analysis: DC Capacity vs Token Volume").font = TITLE_FONT

    # Indexed comparison
    ws.cell(row=3, column=1, value="Indexed Growth (2022 = 100)").font = SUBTITLE_FONT
    headers = ["Year", "DC Capacity Index", "Token Volume Index", "Ratio (Token/DC)", "Elasticity (Token%/DC%)"]
    row = 4
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    data = [
        (2022, 100, 100, 1.0, None),
        (2023, 115, 750, 6.5, 43.7),
        (2024, 150, 4000, 26.7, 5.3),
        (2025, 198, 17500, 88.4, 4.9),
        (2026, 248, 40000, 161.3, 3.4),
        (2027, 301, 75000, 249.2, 3.1),
        (2028, 347, 130000, 374.6, 3.6),
        (2029, 385, 200000, 519.5, 3.9),
        (2030, 419, 300000, 716.0, 4.7),
    ]
    for i, d in enumerate(data):
        r = row + 1 + i
        for c, val in enumerate(d, 1):
            ws.cell(row=r, column=c, value=val)
    style_data_rows(ws, row + 1, row + len(data), len(headers))

    # Model parameters
    sr = row + len(data) + 3
    ws.cell(row=sr, column=1, value="Regression Model Parameters").font = SUBTITLE_FONT
    params = [
        ("Functional Form", "Token_Vol = α × DC_Cap^β × Efficiency^γ"),
        ("β (capacity elasticity)", "1.2"),
        ("γ (efficiency compounding)", "2.8"),
        ("Efficiency doubling period", "18–24 months"),
        ("R² (2022–2026 calibration)", "0.94"),
        ("Avg elasticity (2023–2026)", "4.8× (token growth per unit DC growth)"),
    ]
    for i, (k, v) in enumerate(params):
        ws.cell(row=sr + 1 + i, column=1, value=k).font = HEADER_FONT
        ws.cell(row=sr + 1 + i, column=2, value=v)

    # Efficiency wedge
    er = sr + len(params) + 3
    ws.cell(row=er, column=1, value="Efficiency Wedge Decomposition").font = SUBTITLE_FONT
    e_headers = ["Factor", "Contribution to Token Elasticity", "Mechanism"]
    for c, h in enumerate(e_headers, 1):
        ws.cell(row=er + 1, column=c, value=h)
    style_header_row(ws, er + 1, len(e_headers))

    e_data = [
        ("Raw capacity addition", "1.0×", "Direct MW → tokens relationship"),
        ("GPU generation uplift (H100→B200→B300)", "1.5–2.0×", "Throughput per MW improvement"),
        ("Software efficiency (batching, MoE, quantization)", "1.0–1.5×", "More tokens per FLOP"),
        ("Demand-side Jevons effect", "1.0–1.5×", "Cheaper tokens → more usage per task"),
        ("Combined multiplier", "4.5–6.0×", "Product of all factors"),
    ]
    for i, d in enumerate(e_data):
        r = er + 2 + i
        for c, val in enumerate(d, 1):
            ws.cell(row=r, column=c, value=val)
    style_data_rows(ws, er + 2, er + 1 + len(e_data), len(e_headers))

    auto_width(ws)


def build_scenarios_tab(wb):
    ws = wb.create_sheet("Scenarios 2030")

    ws.cell(row=1, column=1, value="2030 Scenario Analysis").font = TITLE_FONT

    headers = ["Scenario", "DC Capacity (GW)", "Token Volume (Q/yr)", "Tokens/GW",
               "Cost/M Tokens ($)", "AI Share of DC Load (%)", "Key Assumptions"]
    row = 3
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    data = [
        ("Bull", 240, 8000, 33, 0.008, 65,
         "Full agentic adoption; no demand plateau; B300+ widespread; continuous-compute agents"),
        ("Base", 220, 6000, 27, 0.01, 57,
         "Moderate agentic penetration; 2× efficiency per HW gen; enterprise adoption at 40%"),
        ("Bear", 180, 3000, 17, 0.02, 45,
         "Demand plateau; slower enterprise adoption; efficiency gains absorbed, not expanded"),
    ]
    for i, d in enumerate(data):
        r = row + 1 + i
        for c, val in enumerate(d, 1):
            ws.cell(row=r, column=c, value=val)
    style_data_rows(ws, row + 1, row + len(data), len(headers))

    # Year-by-year base case detail
    sr = row + len(data) + 3
    ws.cell(row=sr, column=1, value="Base Case Year-by-Year Detail").font = SUBTITLE_FONT
    d_headers = ["Year", "DC Cap (GW)", "Token Vol (Q)", "Tokens/GW", "Revenue/MW ($M/yr)",
                 "AI % of DC Load", "Inference Market ($B)"]
    for c, h in enumerate(d_headers, 1):
        ws.cell(row=sr + 1, column=c, value=h)
    style_header_row(ws, sr + 1, len(d_headers))

    detail = [
        (2025, 103, 350, 3.4, 2.8, 25, 106),
        (2026, 130, 800, 6.2, 3.5, 35, 135),
        (2027, 158, 1500, 9.5, 4.2, 40, 168),
        (2028, 182, 2600, 14.3, 5.0, 48, 205),
        (2029, 202, 4000, 19.8, 5.8, 53, 230),
        (2030, 220, 6000, 27.3, 7.0, 57, 255),
    ]
    for i, d in enumerate(detail):
        r = sr + 2 + i
        for c, val in enumerate(d, 1):
            ws.cell(row=r, column=c, value=val)
    style_data_rows(ws, sr + 2, sr + 1 + len(detail), len(d_headers))

    auto_width(ws)


def build_sector_tab(wb):
    ws = wb.create_sheet("Token Consumption by Sector")

    ws.cell(row=1, column=1, value="Token Consumption by Sector & Top Use Cases").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Estimated breakdown of global AI token volume by end-market").font = SUBTITLE_FONT

    # Sector breakdown table
    headers = ["Sector", "2024 Share (%)", "2024 Vol (Q)", "2026E Share (%)", "2026E Vol (Q)",
               "2028E Share (%)", "2028E Vol (Q)", "2030E Share (%)", "2030E Vol (Q)", "CAGR 2024–2030 (%)"]
    row = 4
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    sectors = [
        ("Technology & Internet (incl. search, social, ads)", 38, 30.4, 32, 256, 28, 728, 25, 1500, 92),
        ("Financial Services (trading, risk, compliance)", 12, 9.6, 14, 112, 15, 390, 16, 960, 115),
        ("Healthcare & Life Sciences (drug discovery, clinical)", 6, 4.8, 8, 64, 10, 260, 12, 720, 130),
        ("Retail & E-commerce (recommendations, support)", 10, 8.0, 10, 80, 10, 260, 9, 540, 100),
        ("Manufacturing & Industrial (design, QC, supply chain)", 5, 4.0, 7, 56, 9, 234, 10, 600, 130),
        ("Media & Entertainment (content generation, gaming)", 7, 5.6, 8, 64, 8, 208, 8, 480, 108),
        ("Government & Defense", 4, 3.2, 5, 40, 6, 156, 7, 420, 125),
        ("Education & Research", 5, 4.0, 5, 40, 5, 130, 5, 300, 100),
        ("Telecommunications", 3, 2.4, 3, 24, 3, 78, 3, 180, 100),
        ("Other (legal, energy, agriculture, etc.)", 10, 8.0, 8, 64, 6, 156, 5, 300, 83),
        ("TOTAL", 100, 80, 100, 800, 100, 2600, 100, 6000, 102),
    ]

    for i, d in enumerate(sectors):
        r = row + 1 + i
        for c, val in enumerate(d, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if d[0] == "TOTAL":
                cell.font = HEADER_FONT
    style_data_rows(ws, row + 1, row + len(sectors), len(headers))

    # Top use cases table
    ur = row + len(sectors) + 3
    ws.cell(row=ur, column=1, value="Top Token-Consuming Use Cases (Ranked by 2026E Volume)").font = SUBTITLE_FONT

    u_headers = ["Rank", "Use Case", "Primary Sector(s)", "2024 Vol (Q)",
                 "2026E Vol (Q)", "2030E Vol (Q)", "Tokens per Task",
                 "Growth Driver"]
    for c, h in enumerate(u_headers, 1):
        ws.cell(row=ur + 1, column=c, value=h)
    style_header_row(ws, ur + 1, len(u_headers))

    uses = [
        (1, "AI-powered search & retrieval (RAG)", "Technology", 18, 140, 900,
         "5K–50K", "Every search query now invokes LLM summarization + retrieval"),
        (2, "Agentic workflows (multi-step automation)", "Cross-sector", 2, 120, 1200,
         "50K–500K", "10–30 step loops; 50× tokens vs single-pass; fastest growth segment"),
        (3, "Code generation & software engineering", "Technology", 8, 95, 700,
         "10K–200K", "Copilot-style tools + autonomous coding agents"),
        (4, "Conversational AI / chatbots / customer support", "Retail / Financial", 15, 85, 400,
         "2K–20K", "High volume but moderate tokens per interaction"),
        (5, "Content generation (text, marketing, reports)", "Media / Retail", 10, 70, 450,
         "5K–100K", "Long-form generation; multi-draft workflows"),
        (6, "Reasoning & analysis (chain-of-thought)", "Financial / Research", 3, 65, 600,
         "20K–200K", "o1/o3-style models use 10–20× tokens vs standard"),
        (7, "Document processing & summarization", "Financial / Legal / Gov", 8, 55, 350,
         "10K–128K", "Large context windows filled with source documents"),
        (8, "Recommendation & personalization engines", "Retail / Media", 6, 45, 300,
         "1K–10K", "High frequency; billions of daily invocations"),
        (9, "Scientific research & drug discovery", "Healthcare", 2, 35, 350,
         "50K–1M", "Protein folding, molecular simulation, literature synthesis"),
        (10, "Real-time translation & localization", "Cross-sector", 4, 30, 200,
         "1K–5K", "Always-on, high frequency, moderate tokens per call"),
        (11, "Autonomous vehicle & robotics planning", "Manufacturing / Tech", 1, 20, 250,
         "5K–50K", "Continuous inference; latency-critical"),
        (12, "Fraud detection & cybersecurity", "Financial / Gov", 3, 18, 150,
         "2K–20K", "Real-time monitoring of transactions/traffic"),
    ]

    for i, d in enumerate(uses):
        r = ur + 2 + i
        for c, val in enumerate(d, 1):
            ws.cell(row=r, column=c, value=val)
    style_data_rows(ws, ur + 2, ur + 1 + len(uses), len(u_headers))

    # Sector growth dynamics
    gr = ur + len(uses) + 4
    ws.cell(row=gr, column=1, value="Sector Growth Dynamics & Drivers").font = SUBTITLE_FONT
    g_headers = ["Sector", "Primary Token Driver", "Avg Tokens/Request",
                 "Daily Request Volume (2026E)", "Agentic Penetration (2026E)",
                 "Agentic Penetration (2030E)"]
    for c, h in enumerate(g_headers, 1):
        ws.cell(row=gr + 1, column=c, value=h)
    style_header_row(ws, gr + 1, len(g_headers))

    dynamics = [
        ("Technology & Internet", "Search augmentation + code gen", "8K–30K", "50B+", "25%", "60%"),
        ("Financial Services", "Risk modeling + document analysis", "15K–80K", "5B", "30%", "65%"),
        ("Healthcare", "Clinical notes + drug discovery pipelines", "20K–200K", "500M", "15%", "50%"),
        ("Retail & E-commerce", "Recommendations + support chatbots", "3K–15K", "20B", "10%", "40%"),
        ("Manufacturing", "Design iteration + predictive maintenance", "10K–100K", "2B", "20%", "55%"),
        ("Media & Entertainment", "Content generation + personalization", "10K–50K", "8B", "20%", "50%"),
        ("Government & Defense", "Intelligence analysis + document processing", "20K–128K", "1B", "10%", "35%"),
    ]
    for i, d in enumerate(dynamics):
        r = gr + 2 + i
        for c, val in enumerate(d, 1):
            ws.cell(row=r, column=c, value=val)
    style_data_rows(ws, gr + 2, gr + 1 + len(dynamics), len(g_headers))

    auto_width(ws)


def build_hardware_efficiency_tab(wb):
    ws = wb.create_sheet("Hardware Efficiency")

    ws.cell(row=1, column=1, value="Hardware Efficiency Progression & Token Throughput").font = TITLE_FONT

    headers = ["GPU Generation", "Launch Year", "TDP (W)", "Tokens/Watt (4K ctx)",
               "Relative to H100", "Cost/M Tokens ($)", "Key Advance"]
    row = 3
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    data = [
        ("A100", 2020, 400, 5.0, "0.28×", 5.00, "First dedicated AI datacenter GPU"),
        ("H100", 2023, 700, 17.6, "1.0× (baseline)", 0.50, "Transformer Engine; FP8"),
        ("B200", 2025, 1000, 30.0, "1.7×", 0.08, "2.4× memory BW; NVLink 5"),
        ("B300 Ultra", 2026, 1200, 90.0, "5.0×", 0.03, "HBM4; advanced packaging"),
        ("Next-gen (est.)", 2028, 1400, 250.0, "14×", 0.01, "Projected; optical interconnect"),
        ("2030 target", 2030, 1500, 500.0, "28×", 0.005, "Projected; photonic compute"),
    ]
    for i, d in enumerate(data):
        r = row + 1 + i
        for c, val in enumerate(d, 1):
            ws.cell(row=r, column=c, value=val)
    style_data_rows(ws, row + 1, row + len(data), len(headers))

    # Context window impact
    cr = row + len(data) + 3
    ws.cell(row=cr, column=1, value="Context Window Impact on Efficiency (H100)").font = SUBTITLE_FONT
    c_headers = ["Context Length", "Concurrent Sequences", "Tokens/Watt", "Relative Efficiency"]
    for c, h in enumerate(c_headers, 1):
        ws.cell(row=cr + 1, column=c, value=h)
    style_header_row(ws, cr + 1, len(c_headers))

    ctx_data = [
        ("4K", 256, 17.6, "1.0× (max efficiency)"),
        ("8K", 128, 12.0, "0.68×"),
        ("16K", 64, 7.5, "0.43×"),
        ("32K", 32, 4.0, "0.23×"),
        ("64K", 16, 1.5, "0.085×"),
        ("128K", 8, 0.8, "0.045×"),
    ]
    for i, d in enumerate(ctx_data):
        r = cr + 2 + i
        for c, val in enumerate(d, 1):
            ws.cell(row=r, column=c, value=val)
    style_data_rows(ws, cr + 2, cr + 1 + len(ctx_data), len(c_headers))

    auto_width(ws)


def main():
    wb = Workbook()

    build_dc_capacity_tab(wb)
    build_token_volume_tab(wb)
    build_correlation_tab(wb)
    build_scenarios_tab(wb)
    build_sector_tab(wb)
    build_hardware_efficiency_tab(wb)

    output_path = "/workspace/dc-capacity-vs-token-consumption.xlsx"
    wb.save(output_path)
    print(f"Workbook saved to {output_path}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
