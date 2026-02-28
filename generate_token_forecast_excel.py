import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=11)
SECTION_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=12)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_row(ws, row, cols, fill=None, font=None):
    f = fill or HEADER_FILL
    fn = font or HEADER_FONT
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = fn
        cell.fill = f
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN


def auto_width(ws, cols, min_w=14, max_w=45):
    for c in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = min(
            max(min_w, max(len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, ws.max_row + 1)) + 4),
            max_w,
        )


# ── Sheet 1: China AI Token Consumption Timeline ──
ws1 = wb.active
ws1.title = "China Token Growth"

headers1 = ["Period", "Daily Token Consumption", "Growth vs Baseline", "Key Milestone", "Source"]
for c, h in enumerate(headers1, 1):
    ws1.cell(row=1, column=c, value=h)
style_header_row(ws1, 1, len(headers1))

data1 = [
    ["Early 2024", "100 billion", "Baseline", "Starting point for China AI token tracking", "Robonomics Substack"],
    ["June 2025", "30+ trillion", "~300x", "China National Data Bureau official figure", "Robonomics Substack / National Data Bureau"],
    ["H1 2025", "10.19 trillion (enterprise avg)", "~102x", "Enterprise-level large model usage (Sullivan report)", "Sullivan / Futunn"],
    ["H2 2025", "37 trillion (enterprise avg)", "~370x", "263% growth from H1 2025; Qwen leads at 32.1% share", "Sullivan / Futunn"],
    ["Dec 2025", "50+ trillion (Doubao alone)", "~500x (single platform)", "ByteDance Doubao reaches #1 in China MaaS market", "AIBase / 36Kr"],
    ["Jan 2026", "63 trillion (Volcano Engine)", "~630x (single platform)", "Volcano Engine 10x YoY growth", "Robonomics Substack"],
    ["Feb 2026", "~180 trillion (total mainstream)", "~1,800x", "China overtakes US rivals in global token usage", "Robonomics Substack / CGTN"],
]
for i, row in enumerate(data1, 2):
    for c, val in enumerate(row, 1):
        ws1.cell(row=i, column=c, value=val)

style_data_rows(ws1, 2, len(data1) + 1, len(headers1))
auto_width(ws1, len(headers1))

# ── Sheet 2: Global Platform Token Comparison ──
ws2 = wb.create_sheet("Global Platform Comparison")

headers2 = [
    "Platform / Provider", "Country", "Daily Tokens (Estimate)", "Monthly Tokens (Estimate)",
    "Primary Use Cases", "Market Position", "Period", "Source",
]
for c, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=c, value=h)
style_header_row(ws2, 1, len(headers2))

data2 = [
    ["Google (Gemini + all services)", "US", "~43.3 trillion", "~1.3 quadrillion", "Search, Gmail, YouTube, Workspace", "#1 total volume globally", "H2 2025", "Adam Holter / Tom Tunguz"],
    ["OpenAI (ChatGPT + API)", "US", "~8.6 trillion", "~260 trillion", "Consumer AI, Enterprise API, Coding", "#1 consumer AI; 6B tokens/min", "H2 2025", "PYMNTS / Tom Tunguz"],
    ["China Total (all mainstream models)", "China", "~180 trillion", "~5.4 quadrillion (est.)", "Enterprise, Consumer, Video Gen", "Surpassed US in aggregate", "Feb 2026", "Robonomics Substack / CGTN"],
    ["ByteDance (Doubao / Volcano Engine)", "China", "50-63 trillion", "~1.5-1.9 quadrillion (est.)", "Consumer AI, Enterprise, Video", "#1 China, #3 global MaaS", "Dec 2025 - Jan 2026", "AIBase / 36Kr"],
    ["Alibaba Cloud (Qwen)", "China", "~5 trillion (external)", "~150 trillion (est.)", "Enterprise AI, Cloud services", "32.1% enterprise market share in China", "H2 2025", "Futunn / Sullivan"],
    ["Alibaba Cloud (Qwen) - 2026 target", "China", "15-20T ext. / 100T internal", "TBD", "Enterprise + internal business", "Expanding aggressively", "2026 target", "Robonomics Substack"],
    ["OpenRouter (aggregator)", "Global", "1+ trillion", "~30+ trillion", "Developer routing to 300+ models", "5M+ developers", "Late 2025", "a16z / OpenRouter"],
    ["Anthropic (Claude)", "US", "N/A (enterprise-focused)", "N/A", "Programming (80%+ share), Enterprise", "#1 enterprise token consumption", "2025", "PYMNTS"],
    ["Microsoft Foundry", "US", "~57 billion", "~1.7 trillion", "Enterprise applications", "Growing; Azure AI integration", "Apr 2025", "Tom Tunguz"],
]
for i, row in enumerate(data2, 2):
    for c, val in enumerate(row, 1):
        ws2.cell(row=i, column=c, value=val)

style_data_rows(ws2, 2, len(data2) + 1, len(headers2))
auto_width(ws2, len(headers2))

# ── Sheet 3: Token Consumption Use-Case Breakdown ──
ws3 = wb.create_sheet("Use-Case Breakdown")

headers3 = ["Use Case", "Share of Token Usage", "Key Models / Platforms", "Token Intensity", "Trend", "Source"]
for c, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=c, value=h)
style_header_row(ws3, 1, len(headers3))

data3 = [
    ["Programming / Coding", "50%+ of usage (surged from 11%)", "Claude (60%+ share), GPT, DeepSeek", "Moderate per query, high frequency", "Rapidly increasing; agentic workflows", "a16z / OpenRouter Study"],
    ["Roleplay / Creative Writing", "~52% of open-source token usage", "DeepSeek (60%), open-source models", "High per session", "Stable; dominant in open-source segment", "a16z / OpenRouter Study"],
    ["Video Generation", "Emerging major driver", "Seedance 2.0, Sora 2.0", "~350,000 tokens per 10-sec 1080p video; animated projects: hundreds of millions", "Explosive growth expected", "Robonomics Substack"],
    ["Enterprise AI / Business Apps", "37T tokens/day in China alone", "Qwen (32.1% share), Doubao, GPT", "Varies by application", "263% growth H1-to-H2 2025 in China", "Sullivan / Futunn"],
    ["Science & Technology", "Growing segment", "GPT, Gemini, Claude", "High (complex reasoning)", "Increasing with reasoning models", "a16z / OpenRouter Study"],
    ["Search / General Knowledge", "Embedded in platform usage", "Google Gemini, Perplexity", "Low-moderate per query, very high volume", "Driving Google's quadrillion-scale volume", "Adam Holter"],
    ["Agentic / Multi-step Reasoning", "50%+ of all tokens processed", "Reasoning-optimized models", "Very high (extended multi-step workflows)", "Fastest growing category", "a16z / OpenRouter Study"],
]
for i, row in enumerate(data3, 2):
    for c, val in enumerate(row, 1):
        ws3.cell(row=i, column=c, value=val)

style_data_rows(ws3, 2, len(data3) + 1, len(headers3))
auto_width(ws3, len(headers3))

# ── Sheet 4: AI API Pricing Per 1M Tokens (2026) ──
ws4 = wb.create_sheet("API Pricing (2026)")

headers4 = [
    "Provider", "Model", "Input Price (per 1M tokens)", "Output Price (per 1M tokens)",
    "Cost Tier", "Notable Advantage", "Source",
]
for c, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=c, value=h)
style_header_row(ws4, 1, len(headers4))

data4 = [
    ["OpenAI", "GPT-5", "$2.00", "$8.00", "Mid-range", "Latest frontier model", "Dev.to / DevTk.AI"],
    ["OpenAI", "GPT-4.1", "$2.00", "$8.00", "Mid-range", "Optimized for coding", "Dev.to / DevTk.AI"],
    ["OpenAI", "GPT-4o", "$2.50", "$10.00", "Mid-range", "Multimodal flagship", "Dev.to / DevTk.AI"],
    ["Anthropic", "Claude Opus 4.6", "$5.00", "$25.00", "Premium", "Highest quality reasoning", "Dev.to / DevTk.AI"],
    ["Anthropic", "Claude Sonnet 4.6", "$3.00", "$15.00", "Mid-high", "Best price/performance for coding", "Dev.to / DevTk.AI"],
    ["Anthropic", "Claude Haiku 4.5", "$1.00", "$5.00", "Budget", "Fast, cost-effective", "Dev.to / DevTk.AI"],
    ["Google", "Gemini 3.1 Pro", "$1.25", "$10.00", "Mid-range", "Strong multimodal", "Dev.to / DevTk.AI"],
    ["Google", "Gemini 2.5 Flash", "$0.30", "$2.50", "Budget", "Very low cost, fast", "Dev.to / DevTk.AI"],
    ["DeepSeek", "DeepSeek V3.2", "$0.27", "N/A", "Ultra-budget", "Significant cost advantage", "Dev.to / DevTk.AI"],
    ["Alibaba", "Qwen 3.5", "Open-weight (self-host)", "Open-weight (self-host)", "Free / Infra cost", "60% cheaper than predecessor", "VentureBeat"],
]
for i, row in enumerate(data4, 2):
    for c, val in enumerate(row, 1):
        ws4.cell(row=i, column=c, value=val)

style_data_rows(ws4, 2, len(data4) + 1, len(headers4))
auto_width(ws4, len(headers4))

# ── Sheet 5: Market Forecasts & Implications ──
ws5 = wb.create_sheet("Market Forecasts")

headers5 = ["Metric / Forecast", "Value", "Time Frame", "Growth Rate", "Key Implication", "Source"]
for c, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=c, value=h)
style_header_row(ws5, 1, len(headers5))

data5 = [
    ["Global AI Spending", "$2.02 - $2.5 trillion", "2026", "36-44% YoY", "Transition from experimental to industrialized AI", "Gartner / Wedbush"],
    ["Enterprise GenAI Spending", "$37 billion", "2025", "320% YoY (from $11.5B in 2024)", "Jevons' Paradox: lower costs drive higher total spend", "Artur Markus analysis"],
    ["Cumulative Token Spending", ">$1 trillion projected", "2027-2028", "30% YoY enterprise growth", "Token economy becomes major economic force", "AI2.work / OpenAI"],
    ["Hardware share of AI spend", "$1.13 trillion", "2026", "Part of $2T total", "GenAI smartphones: $393B; AI servers: $330B", "Gartner"],
    ["GPU Cloud Price Decline", "$1.49-$3.90/hr (H100)", "2025-2026", "Down from $7-8/hr", "AWS cut 44% in Jun 2025; enables scale", "Introl"],
    ["Per-Token Cost Drop", "~1,000x decline", "2023-2025", "Exponential decrease", "Despite drops, total spend surges (Jevons' Paradox)", "Artur Markus analysis"],
    ["China Daily Token Consumption", "~180 trillion", "Feb 2026", "1,800x since early 2024", "China leads globally in aggregate AI token usage", "Robonomics Substack"],
    ["Google Monthly Token Volume", "1.3 quadrillion", "H2 2025", "~2.7x in 3 months (May-Aug 2025)", "Largest single-company token processor", "Adam Holter"],
    ["Video Gen Token Demand", "350K tokens / 10-sec video", "2025-2026", "Major emerging driver", "Could 10x+ total consumption when widely adopted", "Robonomics Substack"],
    ["Open-Source Model Share", "~30% of all AI usage", "Late 2025", "Growing; Chinese OSS peaked at 30%", "'Good enough but cheaper' threatens pricing power", "a16z / OpenRouter"],
    ["Agentic Inference Share", "50%+ of tokens processed", "Late 2025", "Fastest growing", "Multi-step reasoning workflows drive consumption", "a16z / OpenRouter"],
]
for i, row in enumerate(data5, 2):
    for c, val in enumerate(row, 1):
        ws5.cell(row=i, column=c, value=val)

style_data_rows(ws5, 2, len(data5) + 1, len(headers5))
auto_width(ws5, len(headers5))

# ── Sheet 6: Key Open Questions (from Robonomics) ──
ws6 = wb.create_sheet("Key Open Questions")

headers6 = ["Question / Caveat", "Context", "Potential Implication", "Source"]
for c, h in enumerate(headers6, 1):
    ws6.cell(row=1, column=c, value=h)
style_header_row(ws6, 1, len(headers6))

data6 = [
    [
        "Like-for-like comparison issues",
        "Lower-quality models may require more retries and longer prompts, inflating raw token counts",
        "China's higher token numbers may partly reflect efficiency differences rather than genuine usage advantage",
        "Robonomics Substack",
    ],
    [
        "Structural population differences",
        "China's larger population and historically weaker search alternatives could naturally drive heavier AI usage",
        "Per-capita or per-GDP-normalized token consumption may tell a different story",
        "Robonomics Substack",
    ],
    [
        "'Good enough but cheaper' model viability",
        "If cheaper models prove sufficient for most tasks, pricing power for frontier models erodes",
        "Major consequences for model hierarchy, VC returns, and infrastructure investment thesis",
        "Robonomics Substack / a16z Study",
    ],
    [
        "Token throughput as the right metric",
        "Raw token counts may not capture quality, efficiency, or economic value of AI usage",
        "Need complementary metrics: revenue per token, task completion rate, user satisfaction",
        "Robonomics Substack",
    ],
    [
        "Jevons' Paradox in AI",
        "Per-token costs dropped 1,000x but total spending surged 320%",
        "Cost reductions will continue to drive exponential demand growth, not savings",
        "Artur Markus / Introl",
    ],
]
for i, row in enumerate(data6, 2):
    for c, val in enumerate(row, 1):
        ws6.cell(row=i, column=c, value=val)

style_data_rows(ws6, 2, len(data6) + 1, len(headers6))
auto_width(ws6, len(headers6))

# ── Sheet 7: Sources ──
ws7 = wb.create_sheet("Sources")

headers7 = ["Source", "Title / Description", "URL / Reference", "Data Points Used"]
for c, h in enumerate(headers7, 1):
    ws7.cell(row=1, column=c, value=h)
style_header_row(ws7, 1, len(headers7))

data7 = [
    ["Robonomics Substack (FD)", "Token Tracker & Implications", "https://robonomics.substack.com/p/token-tracker-and-implications", "China token growth, video gen tokens, platform comparisons, open questions"],
    ["a16z / OpenRouter", "State of AI: 100 Trillion Token Study", "https://a16z.com/state-of-ai/", "Model usage breakdown, programming share, agentic inference trends"],
    ["Sullivan / Futunn", "Enterprise Large Model Daily Usage Report", "https://news.futunn.com/en/post/69126161", "China enterprise token usage (37T/day), Qwen market share (32.1%)"],
    ["AIBase / 36Kr", "Doubao Large Model Achievements", "https://news.aibase.com/news/23814", "ByteDance Doubao 50T daily tokens, 417x growth, MaaS ranking"],
    ["Adam Holter", "Google Processes 1.3 Quadrillion AI Tokens Monthly", "https://adam.holter.com/google-now-processes-1-3-quadrillion-ai-tokens-each-month/", "Google token volume, comparison with OpenAI"],
    ["PYMNTS", "OpenAI vs Google Consumer AI Token Consumption", "https://www.pymnts.com/artificial-intelligence-2/2025/openai-bests-google-in-race-for-consumer-ai-token-consumption/", "OpenAI 6B tokens/min, consumer AI leadership"],
    ["Tom Tunguz", "Beyond a Trillion: The Token Race", "https://www.tomtunguz.com/trillion-token-race/", "Google 32.7T daily, Microsoft Foundry data"],
    ["CGTN", "Chinese AI Models Overtake US Rivals", "https://news.cgtn.com/news/2026-02-28/Chinese-AI-models-overtake-U-S-rivals-in-global-token-usage", "China vs US aggregate token comparison"],
    ["Gartner / Wedbush", "AI Supercycle / AI Spending Forecasts", "Various (markets.financialcontent.com, investor.wedbush.com)", "Global AI spending $2-2.5T forecast for 2026"],
    ["Dev.to / DevTk.AI", "AI API Pricing Comparison 2026", "https://dev.to/lemondata_dev/ai-api-pricing-comparison-2026", "Per-token pricing for GPT, Claude, Gemini, DeepSeek"],
    ["Artur Markus", "The Inference Cost Paradox", "https://www.arturmarkus.com/the-inference-cost-paradox", "Enterprise GenAI spend surge 320%, Jevons' Paradox analysis"],
    ["VentureBeat", "Alibaba Qwen 3.5 Analysis", "https://www.venturebeat.com/technology/alibabas-qwen-3-5", "Qwen 3.5 performance, 60% cost reduction"],
]
for i, row in enumerate(data7, 2):
    for c, val in enumerate(row, 1):
        ws7.cell(row=i, column=c, value=val)

style_data_rows(ws7, 2, len(data7) + 1, len(headers7))
auto_width(ws7, len(headers7))

output_path = "/workspace/AI_Token_Forecast_Data_Aggregation.xlsx"
wb.save(output_path)
print(f"Excel file saved to: {output_path}")
