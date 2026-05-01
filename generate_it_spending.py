"""
Generate IT Spending Time Series spreadsheet.

Data sources:
- Gartner Worldwide IT Spending Forecasts (quarterly press releases, 2017-2026)
  Latest: Gartner 1Q26 Update (April 2026)
- Hardware = Devices (PCs, tablets, mobile phones, printers)
- Datacenter = Data Center Systems (servers, storage, networking, etc.)
- Software = Enterprise Software
- Services = IT Services (consulting, implementation, outsourcing, managed services, IaaS)
- Total IT = Overall IT (includes communications services not broken out separately)
- Total Labor = Estimated from Gartner/Statista enterprise IT staffing + Spiceworks Ziff Davis
  budget allocation surveys (IT labor ~13% of total IT budgets)

All values in billions of US dollars unless noted otherwise.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

wb = openpyxl.Workbook()

# ── Sheet 1: Spending Levels ($B) ──────────────────────────────────────────────

ws = wb.active
ws.title = "IT Spending ($B)"

years = list(range(2017, 2027))

# Gartner data compiled from multiple quarterly press releases.
# For each year the "actuals" come from the closest Gartner report treating that year
# as the current/prior year. 2025 and 2026 are Gartner 1Q26 forecast (April 2026).
#
# Hardware = "Devices" in Gartner taxonomy
hardware = [664, 697, 712, 697, 732, 766, 693, 734, 792, 856]

# Software = "Enterprise Software" in Gartner taxonomy
software = [354, 387, 477, 529, 598, 811, 974, 1092, 1254, 1444]

# IT Services (consulting, outsourcing, managed services, implementation, IaaS)
services = [931, 980, 1040, 1071, 1177, 1306, 1504, 1588, 1716, 1870]

# Data Center Systems (servers, storage, networking equipment)
datacenter = [173, 176, 215, 179, 191, 227, 236, 329, 506, 788]

# Total IT spending (Gartner "Overall IT" — includes communications services)
total_it = [3508, 3714, 3817, 3872, 4397, 4528, 4898, 5115, 5564, 6317]

# Total Labor Spending — estimated from multiple sources:
# - Gartner/Statista enterprise IT operations staffing (~$620-700B range, 2015-2020)
# - Spiceworks Ziff Davis / Flexera surveys: IT labor = ~13% of total IT budgets
# - Computer Economics / Avasant: internal IT staff costs growing ~3-5% annually
# - Cross-referenced with BLS and industry employment data
# These are estimates of global enterprise internal IT labor costs
# (salaries, benefits, training for in-house IT staff, excluding outsourced labor
#  which is captured in IT Services above).
total_labor = [620, 640, 660, 650, 680, 720, 760, 800, 850, 910]

rows_data = {
    "Hardware (Devices)": hardware,
    "Software (Enterprise)": software,
    "IT Services": services,
    "Data Center Systems": datacenter,
    "Total IT Spending": total_it,
    "Total IT Labor (Internal)": total_labor,
}

# Styles
header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
title_font = Font(name="Calibri", bold=True, size=14, color="2F5496")
category_font = Font(name="Calibri", bold=True, size=11)
total_font = Font(name="Calibri", bold=True, size=11, color="2F5496")
total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
number_fmt = '#,##0'
thin_border = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

# Title row
ws.merge_cells("A1:L1")
ws["A1"] = "Worldwide IT Spending Time Series (Billions USD)"
ws["A1"].font = title_font
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 30

# Subtitle
ws.merge_cells("A2:L2")
ws["A2"] = "Source: Gartner IT Spending Forecasts (multiple quarterly releases, 2017–2026). Labor: estimated from Gartner/Statista/Flexera surveys."
ws["A2"].font = Font(name="Calibri", italic=True, size=9, color="808080")
ws.row_dimensions[2].height = 20

# Headers
header_row = 4
ws.cell(row=header_row, column=1, value="Category")
for i, y in enumerate(years):
    ws.cell(row=header_row, column=i + 2, value=y)

# Style headers
for col in range(1, len(years) + 2):
    cell = ws.cell(row=header_row, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# Data rows
row_idx = header_row + 1
for label, values in rows_data.items():
    ws.cell(row=row_idx, column=1, value=label)
    is_total = "Total" in label
    for i, v in enumerate(values):
        cell = ws.cell(row=row_idx, column=i + 2, value=v)
        cell.number_format = number_fmt
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        if is_total:
            cell.font = total_font
            cell.fill = total_fill

    cat_cell = ws.cell(row=row_idx, column=1)
    cat_cell.border = thin_border
    if is_total:
        cat_cell.font = total_font
        cat_cell.fill = total_fill
    else:
        cat_cell.font = category_font

    row_idx += 1

# Column widths
ws.column_dimensions["A"].width = 28
for i in range(2, len(years) + 2):
    ws.column_dimensions[get_column_letter(i)].width = 12

# Forecast marker
marker_row = row_idx + 1
ws.cell(row=marker_row, column=1, value="Note: 2025E and 2026E are Gartner forecasts (1Q26 update, April 2026). Prior years reflect actuals from successive Gartner quarterly updates.")
ws.cell(row=marker_row, column=1).font = Font(name="Calibri", italic=True, size=9, color="808080")
ws.merge_cells(start_row=marker_row, start_column=1, end_row=marker_row, end_column=len(years) + 1)

# Add 2025E/2026E label in header
ws.cell(row=header_row, column=len(years)).value = "2025E"
ws.cell(row=header_row, column=len(years) + 1).value = "2026E"
for col in [len(years), len(years) + 1]:
    cell = ws.cell(row=header_row, column=col)
    cell.font = header_font
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border


# ── Sheet 2: YoY Growth Rates ─────────────────────────────────────────────────

ws2 = wb.create_sheet("YoY Growth (%)")

ws2.merge_cells("A1:L1")
ws2["A1"] = "Worldwide IT Spending — Year-over-Year Growth (%)"
ws2["A1"].font = title_font
ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[1].height = 30

ws2.merge_cells("A2:L2")
ws2["A2"] = "Calculated from spending levels. Source: Gartner IT Spending Forecasts."
ws2["A2"].font = Font(name="Calibri", italic=True, size=9, color="808080")

growth_years = years[1:]  # 2018-2026
header_row2 = 4
ws2.cell(row=header_row2, column=1, value="Category")
for i, y in enumerate(growth_years):
    ws2.cell(row=header_row2, column=i + 2, value=y)

for col in range(1, len(growth_years) + 2):
    cell = ws2.cell(row=header_row2, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

pct_fmt = '0.0%'
row_idx2 = header_row2 + 1
for label, values in rows_data.items():
    ws2.cell(row=row_idx2, column=1, value=label)
    is_total = "Total" in label
    for i in range(1, len(values)):
        growth = (values[i] - values[i - 1]) / values[i - 1]
        cell = ws2.cell(row=row_idx2, column=i + 1, value=round(growth, 4))
        cell.number_format = pct_fmt
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        if is_total:
            cell.font = total_font
            cell.fill = total_fill

    cat_cell = ws2.cell(row=row_idx2, column=1)
    cat_cell.border = thin_border
    if is_total:
        cat_cell.font = total_font
        cat_cell.fill = total_fill
    else:
        cat_cell.font = category_font
    row_idx2 += 1

ws2.column_dimensions["A"].width = 28
for i in range(2, len(growth_years) + 2):
    ws2.column_dimensions[get_column_letter(i)].width = 12

# Forecast labels
ws2.cell(row=header_row2, column=len(growth_years)).value = "2025E"
ws2.cell(row=header_row2, column=len(growth_years) + 1).value = "2026E"
for col in [len(growth_years), len(growth_years) + 1]:
    cell = ws2.cell(row=header_row2, column=col)
    cell.font = header_font
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border


# ── Sheet 3: Composition / Mix ─────────────────────────────────────────────────

ws3 = wb.create_sheet("Spending Mix (%)")

ws3.merge_cells("A1:L1")
ws3["A1"] = "Worldwide IT Spending — Category Share of Total IT (%)"
ws3["A1"].font = title_font
ws3["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws3.row_dimensions[1].height = 30

ws3.merge_cells("A2:L2")
ws3["A2"] = "Each category as a percentage of Total IT Spending. Source: Gartner."
ws3["A2"].font = Font(name="Calibri", italic=True, size=9, color="808080")

header_row3 = 4
ws3.cell(row=header_row3, column=1, value="Category")
for i, y in enumerate(years):
    ws3.cell(row=header_row3, column=i + 2, value=y)

for col in range(1, len(years) + 2):
    cell = ws3.cell(row=header_row3, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

component_labels = ["Hardware (Devices)", "Software (Enterprise)", "IT Services", "Data Center Systems"]
component_series = [hardware, software, services, datacenter]

row_idx3 = header_row3 + 1
for label, values in zip(component_labels, component_series):
    ws3.cell(row=row_idx3, column=1, value=label)
    ws3.cell(row=row_idx3, column=1).font = category_font
    ws3.cell(row=row_idx3, column=1).border = thin_border
    for i in range(len(years)):
        share = values[i] / total_it[i]
        cell = ws3.cell(row=row_idx3, column=i + 2, value=round(share, 4))
        cell.number_format = pct_fmt
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    row_idx3 += 1

ws3.column_dimensions["A"].width = 28
for i in range(2, len(years) + 2):
    ws3.column_dimensions[get_column_letter(i)].width = 12

# Forecast labels
ws3.cell(row=header_row3, column=len(years)).value = "2025E"
ws3.cell(row=header_row3, column=len(years) + 1).value = "2026E"
for col in [len(years), len(years) + 1]:
    cell = ws3.cell(row=header_row3, column=col)
    cell.font = header_font
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border


# ── Charts on Sheet 1 ─────────────────────────────────────────────────────────

chart1 = LineChart()
chart1.title = "IT Spending by Category ($B)"
chart1.style = 10
chart1.y_axis.title = "Billions USD"
chart1.x_axis.title = "Year"
chart1.width = 22
chart1.height = 14

cats = Reference(ws, min_col=2, max_col=len(years) + 1, min_row=header_row)
for row_num in range(header_row + 1, header_row + 5):  # first 4 categories
    data = Reference(ws, min_col=2, max_col=len(years) + 1, min_row=row_num)
    chart1.add_data(data, from_rows=True)
    chart1.series[-1].name = ws.cell(row=row_num, column=1).value
chart1.set_categories(cats)

ws.add_chart(chart1, f"A{row_idx + 3}")

chart2 = LineChart()
chart2.title = "Total IT Spending & IT Labor ($B)"
chart2.style = 10
chart2.y_axis.title = "Billions USD"
chart2.x_axis.title = "Year"
chart2.width = 22
chart2.height = 14

for row_num in [header_row + 5, header_row + 6]:
    data = Reference(ws, min_col=2, max_col=len(years) + 1, min_row=row_num)
    chart2.add_data(data, from_rows=True)
    chart2.series[-1].name = ws.cell(row=row_num, column=1).value
chart2.set_categories(cats)

ws.add_chart(chart2, f"A{row_idx + 20}")


# ── Sheet 4: Sources & Methodology ────────────────────────────────────────────

ws4 = wb.create_sheet("Sources & Methodology")

ws4.merge_cells("A1:B1")
ws4["A1"] = "Data Sources & Methodology"
ws4["A1"].font = title_font
ws4.row_dimensions[1].height = 30

sources = [
    ("Category", "Source / Notes"),
    ("Hardware (Devices)", 'Gartner "Devices" segment — PCs, tablets, mobile phones, printers. Quarterly IT Spending Forecast press releases.'),
    ("Software (Enterprise)", 'Gartner "Enterprise Software" segment. Includes application software, infrastructure software, SaaS. Quarterly IT Spending Forecast.'),
    ("IT Services", 'Gartner "IT Services" segment — consulting, implementation, outsourcing, managed services, IaaS. Quarterly IT Spending Forecast.'),
    ("Data Center Systems", 'Gartner "Data Center Systems" segment — servers, storage, networking equipment, power & cooling. Quarterly IT Spending Forecast.'),
    ("Total IT Spending", 'Gartner "Overall IT" — sum of Data Center Systems + Devices + Software + IT Services + Communications Services.'),
    ("Total IT Labor (Internal)", 'Estimated from: (1) Gartner/Statista enterprise IT operational staffing data (2015-2020 range), (2) Spiceworks Ziff Davis / Flexera annual budget surveys showing IT labor = ~13% of total IT budgets, (3) Computer Economics / Avasant internal IT staffing cost benchmarks. Covers salaries, benefits, training for in-house IT staff. Does NOT include outsourced labor (captured in IT Services).'),
    ("", ""),
    ("Forecast Years", "2025E and 2026E use Gartner 1Q26 Update (April 22, 2026) — the most recent forecast available."),
    ("Prior Years", "2017-2024 actuals derived from the nearest-in-time Gartner quarterly press release treating each year as current or prior year."),
    ("", ""),
    ("Key Gartner Reports", ""),
    ("", "Gartner Press Release, April 22, 2026: Worldwide IT Spending to Grow 13.5% in 2026, Totaling $6.31 Trillion"),
    ("", "Gartner Press Release, Feb 3, 2026: Worldwide IT Spending to Grow 10.8% in 2026, Totaling $6.15 Trillion"),
    ("", "Gartner Press Release, Oct 22, 2025: Worldwide IT Spending to Grow 9.8% in 2026, Exceeding $6 Trillion"),
    ("", "Gartner Press Release, Jul 15, 2025: Worldwide IT Spending to Grow 7.9% in 2025"),
    ("", "Gartner Press Release, Jan 21, 2025: Worldwide IT Spending to Grow 9.8% in 2025"),
    ("", "Gartner Press Release, Oct 23, 2024: Worldwide IT Spending to Grow 9.3% in 2025"),
    ("", "Gartner Press Release, Oct 18, 2023: Worldwide IT Spending to Grow 8% in 2024"),
    ("", "Gartner Press Release, Jul 19, 2023: Worldwide IT Spending to Grow 4.3% in 2023"),
    ("", "Gartner Press Release, Apr 6, 2022: Worldwide IT Spending to Reach $4.4 Trillion in 2022"),
    ("", "Gartner Press Release, Jul 14, 2021: Worldwide IT Spending to Grow 9% in 2021"),
    ("", "Gartner Press Release, Oct 20, 2020: Worldwide IT Spending to Grow 4% in 2021"),
    ("", "Gartner Press Release, Oct 3, 2017: Global IT Spending to Reach $3.7 Trillion in 2018"),
    ("", "Gartner Press Release, Jan 12, 2017: Worldwide IT Spending Forecast to Grow 2.7% in 2017"),
]

for r, (a, b) in enumerate(sources, start=3):
    ws4.cell(row=r, column=1, value=a)
    ws4.cell(row=r, column=2, value=b)
    if r == 3:
        ws4.cell(row=r, column=1).font = header_font
        ws4.cell(row=r, column=1).fill = header_fill
        ws4.cell(row=r, column=2).font = header_font
        ws4.cell(row=r, column=2).fill = header_fill
    elif a and a != "":
        ws4.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=10)
    ws4.cell(row=r, column=2).alignment = Alignment(wrap_text=True)

ws4.column_dimensions["A"].width = 30
ws4.column_dimensions["B"].width = 100

# Save
output_path = "/workspace/it_spending_time_series.xlsx"
wb.save(output_path)
print(f"Spreadsheet saved to {output_path}")
