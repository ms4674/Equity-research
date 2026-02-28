import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
subheader_font = Font(bold=True, size=11)
ai_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
wrap_alignment = Alignment(wrap_text=True, vertical="top")
center_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border


def style_data_rows(ws, start_row, end_row, max_col):
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = wrap_alignment


# ============================================================
# SHEET 1: Overview & Key Metrics
# ============================================================
ws1 = wb.active
ws1.title = "Overview & Key Metrics"

headers = [
    "Vendor",
    "Ownership / Ticker",
    "Revenue (2025, $B)",
    "CapEx / Capital Raised ($B)",
    "Development Backlog ($B)",
    "Total Datacenters",
    "Global Footprint (Metros/Countries)",
    "Total Portfolio Capacity (MW)",
    "Capacity Live (MW est.)",
    "Capacity Under Construction (MW est.)",
    "Total Capacity (GW)",
    "Key Markets",
]

for col, h in enumerate(headers, 1):
    ws1.cell(row=1, column=col, value=h)
style_header_row(ws1, 1, len(headers))

data = [
    [
        "Digital Realty (DLR)",
        "Public REIT (NYSE: DLR)",
        "~5.8",
        "~3.5 (CapEx)",
        "~7.0",
        "300+",
        "50+ metros / 25+ countries",
        "2,345",
        "~1,500",
        "~845",
        "~2.3",
        "Virginia, Texas, Chicago, Silicon Valley, Amsterdam, Frankfurt, Singapore",
    ],
    [
        "Equinix (EQIX)",
        "Public REIT (NASDAQ: EQIX)",
        "~9.2",
        "~3.8 (CapEx) + $15B xScale JV",
        "~6.0 (xScale TCV)",
        "270+",
        "72+ metros / 33 countries",
        "~3,000+",
        "~1,800",
        "~1,200",
        "~3.0+",
        "Virginia, Silicon Valley, NY Metro, London, Amsterdam, Frankfurt, Singapore, Tokyo",
    ],
    [
        "QTS Realty (Blackstone)",
        "Private (Blackstone)",
        "N/A (private)",
        "~5.1 (debt financing 2025)",
        "~25.0 (pipeline)",
        "30+",
        "15 states (US) + Europe",
        "4,752",
        "~1,500",
        "~3,250",
        "~4.8",
        "Virginia (PW Digital Gateway), Texas, Georgia, Arizona, Ohio",
    ],
    [
        "CyrusOne",
        "Private (KKR)",
        "N/A (private)",
        "~21.7 (total financing)",
        "~12.0",
        "50+",
        "9 states (US) + Europe",
        "789 (disclosed) + pipeline",
        "~500",
        "~600+",
        "~0.8 (disclosed)",
        "Texas (DFW), Chicago, Virginia, Arizona, Ohio",
    ],
    [
        "Vantage Data Centers",
        "Private",
        "N/A (private)",
        "~28.8 (Frontier $25B + NV1 $3B + EU $0.8B)",
        "~28.0+",
        "36 campuses",
        "12 projects / 8 states (US) + EMEA",
        "4,187 (US) / ~2,900 global",
        "~1,200",
        "~2,900+",
        "~4.2 (US) / ~2.9 (global)",
        "Texas (Frontier 1.4GW), Nevada, Virginia, EMEA",
    ],
    [
        "Aligned Data Centers",
        "Private (AIP/MGX/BlackRock GIP, ~$40B EV)",
        "N/A (private)",
        "~12.0 ($5B equity + $7B debt)",
        "~12.0+",
        "78 DCs / 50 campuses",
        "Americas (US + LatAm via ODATA)",
        "5,000+",
        "~1,500",
        "~3,500+",
        "~5.0+",
        "NoVA, Illinois, Ohio, Phoenix, Dallas, Salt Lake City, Brazil, Mexico",
    ],
    [
        "STACK Infrastructure",
        "Private",
        "N/A (private)",
        "~4.0 (2025) / $20B+ total since 2019",
        "~10.0+",
        "Multiple campuses",
        "22 key markets globally",
        "4,000+ (built/dev) + 6,000+ (planned)",
        "~1,500",
        "~2,500+",
        "~4.0+ (built/dev)",
        "Virginia (Stafford 1GW+), Dallas (220MW), North Texas (3GW planned)",
    ],
    [
        "Compass Datacenters",
        "Private",
        "N/A (private)",
        "~10.3 (Lauderdale $10B + Red Oak $0.3B)",
        "~10.0+",
        "Multiple campuses",
        "US (TX, MS, VA, others)",
        "~500+ (operational) + pipeline",
        "~200",
        "~300+",
        "~0.5+",
        "Red Oak TX (180MW), Lauderdale County MS",
    ],
    [
        "NTT Global Data Centers",
        "NTT Group subsidiary",
        "N/A (subsidiary)",
        "~10.0 (through 2027)",
        "~5.0+",
        "150+",
        "20+ countries",
        "2,000+",
        "~1,200",
        "~800+",
        "~2.0+",
        "Tokyo, Chicago, Dallas, Phoenix, Virginia, Johor Bahru (290MW), Frankfurt",
    ],
    [
        "CoreWeave",
        "Public (NASDAQ: CRWV, IPO 2025)",
        "~3.0 (est. 2025)",
        "~7.0+ (est. 2025 CapEx, doubling in 2026)",
        "~15.0+ (contracted backlog)",
        "28+",
        "US primarily",
        "~1,500+ (GPU-focused)",
        "~800",
        "~700+",
        "~1.5+",
        "Multiple US locations (GPU cloud infrastructure)",
    ],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws1.cell(row=row_idx, column=col_idx, value=val)

style_data_rows(ws1, 2, len(data) + 1, len(headers))

col_widths = [25, 30, 18, 32, 22, 18, 30, 25, 22, 28, 18, 55]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SHEET 2: AI vs Non-AI Capacity
# ============================================================
ws2 = wb.create_sheet("AI vs Non-AI Capacity")

headers2 = [
    "Vendor",
    "Total Capacity (MW)",
    "AI Capacity (MW est.)",
    "AI % of Total",
    "Non-AI Capacity (MW est.)",
    "Non-AI % of Total",
    "AI Capacity Indicators",
    "AI Infrastructure Features",
]

for col, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=col, value=h)
style_header_row(ws2, 1, len(headers2))

ai_data = [
    [
        "Digital Realty (DLR)",
        "2,345",
        "~700",
        "~30%",
        "~1,645",
        "~70%",
        "30% of MW signed in recent quarters are AI-related; growing AI share of new signings",
        "High-density deployments, interconnection fabric, Oracle AI partnership, Microsoft Azure ExpressRoute",
    ],
    [
        "Equinix (EQIX)",
        "~3,000+",
        "~1,200",
        "~40%",
        "~1,800",
        "~60%",
        "60% of largest Q4 2025 deals driven by AI; xScale program primarily AI-focused",
        "xScale hyperscale DCs ($23B+ program), liquid cooling, Equinix Fabric for AI interconnection, 365 MW xScale contracted",
    ],
    [
        "QTS Realty (Blackstone)",
        "4,752",
        "~2,400",
        "~50%",
        "~2,350",
        "~50%",
        "Major AI-focused expansion; rack densities >50kW, custom 100kW deployments for AI accelerator clusters",
        "AI-optimized halls, Lumen fiber backbone connectivity, NVIDIA H100/B200 GPU support, Federal AI infrastructure",
    ],
    [
        "CyrusOne",
        "789+",
        "~400",
        "~50%",
        "~389",
        "~50%",
        "AI-era strategy: 'power + land + interconnect'; all new campuses AI-ready",
        "High-density AI racks, Calpine behind-the-meter power (400MW TX), Eolian battery storage (200MW TX)",
    ],
    [
        "Vantage Data Centers",
        "4,187 (US)",
        "~2,500",
        "~60%",
        "~1,687",
        "~40%",
        "Frontier campus entirely AI-focused (1.4GW); ultra-high density 250kW+ racks",
        "Liquid cooling for next-gen GPUs, 720W/sq ft density, NV1 first two buildings fully leased for AI",
    ],
    [
        "Aligned Data Centers",
        "5,000+",
        "~3,000",
        "~60%",
        "~2,000",
        "~40%",
        "Gigascale AI campuses; 100-300 kW/rack density; NVIDIA Blackwell support",
        "Patented Delta3 & DeltaFlow cooling, Lambda AI partnership, air+liquid cooling hybrid",
    ],
    [
        "STACK Infrastructure",
        "4,000+",
        "~2,000",
        "~50%",
        "~2,000",
        "~50%",
        "All major campuses AI-ready; 3GW TX campus for AI workloads",
        "Flexible cooling (air+liquid), adaptable configurations, massive power availability",
    ],
    [
        "Compass Datacenters",
        "500+",
        "~200",
        "~40%",
        "~300",
        "~60%",
        "Single-tenant hyperscale focus; AI-driven enterprise clients",
        "Schneider Electric AI-driven predictive maintenance, up to 50MW per building scalability",
    ],
    [
        "NTT Global Data Centers",
        "2,000+",
        "~600",
        "~30%",
        "~1,400",
        "~70%",
        "200+ MW of AI workloads enabled in 2024; 30+ kW/rack AI halls",
        "Advanced liquid cooling, AI workload halls, Johor Bahru 290MW AI campus",
    ],
    [
        "CoreWeave",
        "~1,500+",
        "~1,500",
        "~100%",
        "~0",
        "~0%",
        "Pure-play AI/GPU cloud provider; 100% AI-focused infrastructure",
        "NVIDIA GPU clusters (H100, B200), purpose-built for AI training/inference, Microsoft primary customer",
    ],
]

for row_idx, row_data in enumerate(ai_data, 2):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        if col_idx in (3, 4):
            cell.fill = ai_fill

style_data_rows(ws2, 2, len(ai_data) + 1, len(headers2))

col_widths2 = [25, 20, 22, 15, 25, 18, 65, 70]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SHEET 3: Top Customers
# ============================================================
ws3 = wb.create_sheet("Top Customers")

headers3 = [
    "Vendor",
    "Top Customer 1",
    "Top Customer 2",
    "Top Customer 3",
    "Top Customer 4",
    "Top Customer 5",
    "Customer Notes",
]

for col, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=col, value=h)
style_header_row(ws3, 1, len(headers3))

customer_data = [
    [
        "Digital Realty (DLR)",
        "Oracle",
        "Meta (est. unnamed Fortune 50)",
        "Microsoft",
        "IBM",
        "Salesforce",
        "Oracle is largest named customer by revenue; unnamed Fortune 50 software co is largest overall; 15 Azure ExpressRoute on-ramps; Oracle AI Solution Centers partnership",
    ],
    [
        "Equinix (EQIX)",
        "AWS",
        "Microsoft Azure",
        "Google Cloud",
        "Salesforce",
        "Oracle",
        "Hyperscalers are primary xScale customers; Salesforce deepened partnership with Equinix Fabric Cloud Router across 14 countries; ~10,000+ enterprise customers globally",
    ],
    [
        "QTS Realty (Blackstone)",
        "AWS",
        "Microsoft Azure",
        "Google Cloud",
        "Meta",
        "U.S. Federal Government",
        "Serves top 3 hyperscalers; Federal AI infrastructure for government missions; hyperscalers committing $250B+ to DCs through 2026",
    ],
    [
        "CyrusOne",
        "Microsoft",
        "AWS",
        "Google",
        "Enterprise customers",
        "Financial services firms",
        "Hyperscale-focused with significant enterprise base; Texas campuses targeting AI hyperscalers",
    ],
    [
        "Vantage Data Centers",
        "Hyperscale Cloud Provider A",
        "Hyperscale Cloud Provider B",
        "AI Company (undisclosed)",
        "Enterprise (undisclosed)",
        "—",
        "NV1 Nevada first two buildings fully pre-leased; Frontier 1.4GW campus driven by AI customer demand; $820M EU securitization for hyperscale AI clients",
    ],
    [
        "Aligned Data Centers",
        "Lambda (AI Developer Cloud)",
        "Hyperscale Cloud Providers",
        "Enterprise AI Workloads",
        "—",
        "—",
        "Lambda deploys AI infrastructure in DFW-04; serves hyperscalers and enterprise AI customers; acquired by AIP/MGX/BlackRock",
    ],
    [
        "STACK Infrastructure",
        "Hyperscale Cloud Providers",
        "Enterprise Customers",
        "AI Companies",
        "—",
        "—",
        "Single-tenant hyperscale leases; Stafford VA campus (1GW) targeting major cloud providers",
    ],
    [
        "Compass Datacenters",
        "Single-Tenant Hyperscaler (undisclosed)",
        "Enterprise Clients",
        "—",
        "—",
        "—",
        "Exclusively builds for single-tenant hyperscale clients; each facility designed for one customer",
    ],
    [
        "NTT Global Data Centers",
        "Major Cloud Providers",
        "Enterprise Customers",
        "Japanese Enterprises",
        "Government",
        "Financial Services",
        "130MW+ hyperscale agreements across Chicago, Dallas, Phoenix, Virginia; global enterprise and government customers",
    ],
    [
        "CoreWeave",
        "Microsoft (~60%+ of revenue)",
        "AI Startups & Labs",
        "Enterprise AI Teams",
        "Research Institutions",
        "—",
        "Microsoft is dominant customer (~60%+ of revenue per IPO filing); remaining performance obligations ~$15B+; GPU cloud for AI training/inference",
    ],
]

for row_idx, row_data in enumerate(customer_data, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws3.cell(row=row_idx, column=col_idx, value=val)

style_data_rows(ws3, 2, len(customer_data) + 1, len(headers3))

col_widths3 = [25, 28, 28, 25, 25, 25, 75]
for i, w in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SHEET 4: GW Capacity Detail (Live vs Under Construction)
# ============================================================
ws4 = wb.create_sheet("GW Capacity Detail")

headers4 = [
    "Vendor",
    "Total Portfolio Capacity (GW)",
    "Capacity Live / Operational (GW est.)",
    "Capacity Under Construction (GW est.)",
    "Planned / Pipeline (GW est.)",
    "Key Projects Under Construction",
    "Expected Delivery Timeline",
]

for col, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=col, value=h)
style_header_row(ws4, 1, len(headers4))

gw_data = [
    [
        "Digital Realty (DLR)",
        "~2.3",
        "~1.5",
        "~0.8",
        "~3.0+ (development rights)",
        "Multiple hyperscale builds in Virginia, Texas, Chicago; international expansion in Europe and Asia",
        "Rolling delivery through 2026-2028",
    ],
    [
        "Equinix (EQIX)",
        "~3.0+",
        "~1.8",
        "~1.2",
        "~3.0+ (development rights, target 2x by 2029)",
        "xScale campuses across US, Europe, Asia; 16 new projects in 14 metros opened 2025; $23B+ xScale program",
        "On track to double capacity by 2029",
    ],
    [
        "QTS Realty (Blackstone)",
        "~4.8",
        "~1.5",
        "~3.3",
        "~5.0+ (pipeline grown from $1B to $25B)",
        "Prince William Digital Gateway (1GW+), 16 new campuses across AZ, GA, OH, TX, VA, OR",
        "Phased delivery 2025-2028+",
    ],
    [
        "CyrusOne",
        "~0.8 (disclosed)",
        "~0.5",
        "~0.6",
        "~2.0+ (financed)",
        "DFW10 Calpine campus (400MW TX), DFW7 Eolian campus (200MW Fort Worth), Chicago DC (18MW)",
        "DFW10: Q4 2026; DFW7: 2026; Chicago: 2025",
    ],
    [
        "Vantage Data Centers",
        "~4.2 (US) / ~2.9 (global)",
        "~1.2",
        "~2.9+",
        "~2.0+",
        "Frontier TX (1.4GW, $25B), NV1 Nevada (224MW, $3B), EMEA expansion",
        "Frontier H2 2026; NV1 Q2 2026 (first bldg); EMEA ongoing",
    ],
    [
        "Aligned Data Centers",
        "~5.0+",
        "~1.5",
        "~3.5+",
        "~2.0+",
        "Gigascale AI campuses across NoVA, IL, OH, Phoenix, Dallas; LatAm expansion",
        "Phased delivery 2025-2028",
    ],
    [
        "STACK Infrastructure",
        "~4.0+ (built/dev)",
        "~1.5",
        "~2.5+",
        "~6.0+ (planned)",
        "Stafford VA (1GW+, 19 DCs), South Dallas (220MW), North TX (3GW transmission-scale)",
        "Dallas: mid-2026; Stafford: phased; North TX: Q1 2027+",
    ],
    [
        "Compass Datacenters",
        "~0.5+",
        "~0.2",
        "~0.3+",
        "~2.0+ (Lauderdale County)",
        "Red Oak TX (180MW, 5 bldgs), Lauderdale County MS ($10B, 8 DCs)",
        "Red Oak: 2024-2025; Lauderdale: 8-year phased",
    ],
    [
        "NTT Global Data Centers",
        "~2.0+",
        "~1.2",
        "~0.8+",
        "~1.0+ (secured growth)",
        "Chicago, Dallas, Phoenix, Virginia campuses (130MW+), Johor Bahru Malaysia (290MW), Berlin, Frankfurt",
        "Ongoing through 2027",
    ],
    [
        "CoreWeave",
        "~1.5+",
        "~0.8",
        "~0.7+",
        "~2.0+ (contracted backlog)",
        "Multiple US GPU data center facilities; rapid build-out for AI training clusters",
        "Aggressive delivery 2025-2026; CapEx doubling 2026",
    ],
]

for row_idx, row_data in enumerate(gw_data, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws4.cell(row=row_idx, column=col_idx, value=val)

style_data_rows(ws4, 2, len(gw_data) + 1, len(headers4))

col_widths4 = [25, 25, 30, 30, 28, 75, 40]
for i, w in enumerate(col_widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SHEET 5: Datacenter Footprint
# ============================================================
ws5 = wb.create_sheet("Datacenter Footprint")

headers5 = [
    "Vendor",
    "Number of Datacenters / Campuses",
    "Gross Floor Area (sq ft est.)",
    "Geographic Presence",
    "Key Regions / Markets",
    "Recent Expansion (2025)",
]

for col, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=col, value=h)
style_header_row(ws5, 1, len(headers5))

footprint_data = [
    [
        "Digital Realty (DLR)",
        "300+ datacenters",
        "~40M+ sq ft",
        "25+ countries across Americas, EMEA, APAC",
        "Virginia, Texas, Chicago, Silicon Valley, NY Metro, Amsterdam, Frankfurt, London, Singapore, Hong Kong, Tokyo",
        "Continued global expansion; AI-focused builds in core US markets; Oracle & Microsoft partnerships deepened",
    ],
    [
        "Equinix (EQIX)",
        "270+ IBX + xScale datacenters",
        "~55M+ sq ft",
        "33 countries, 72+ metros across Americas, EMEA, APAC",
        "Virginia, Silicon Valley, NY Metro, Los Angeles, London, Amsterdam, Frankfurt, Paris, Singapore, Tokyo, Sydney",
        "16 new projects in 14 metros; 23,250 retail cabinets + 90MW xScale delivered; xScale US entry",
    ],
    [
        "QTS Realty (Blackstone)",
        "30+ facilities",
        "~9M+ sq ft",
        "15 US states + Europe",
        "Virginia (PW Digital Gateway), Georgia (Atlanta), Texas (Dallas, San Antonio), Arizona, Ohio, Oregon",
        "16 new campuses under development; capacity grown 9x under Blackstone; $25B+ pipeline",
    ],
    [
        "CyrusOne",
        "50+ datacenters",
        "~5M+ sq ft",
        "9 US states + Europe (Frankfurt, London, Amsterdam, Dublin)",
        "Texas (DFW), Chicago, Virginia, Arizona, Ohio, New York Metro, North Carolina",
        "DFW10 (400MW), DFW7 (200MW), Chicago DC topped out; $21.7B in financing secured",
    ],
    [
        "Vantage Data Centers",
        "36 global campuses",
        "~8M+ sq ft (incl. planned)",
        "US (8 states) + EMEA (multiple countries)",
        "Texas (Frontier 1,200 acres), Nevada (Reno), Virginia, Silicon Valley, Phoenix, EMEA markets",
        "Frontier TX ($25B, 1.4GW, 3.7M sq ft), NV1 Nevada ($3B, 224MW, 1M+ sq ft), $820M EU raise",
    ],
    [
        "Aligned Data Centers",
        "78 DCs across 50 campuses",
        "~10M+ sq ft",
        "Americas (US + LatAm via ODATA: Brazil, Mexico, Chile)",
        "NoVA, Illinois, Maryland, Ohio, Salt Lake City, Phoenix, Dallas, Sao Paulo, Mexico City",
        "$12B+ capital raise; acquired by AIP/MGX/BlackRock GIP at $40B EV; 5GW+ planned capacity",
    ],
    [
        "STACK Infrastructure",
        "Multiple campuses in 22 markets",
        "~15M+ sq ft (incl. planned)",
        "Global across 22 key markets",
        "Virginia (Stafford 500 acres), Dallas, North Texas, International markets",
        "Stafford 1GW+ campus, South Dallas 220MW, 3GW North TX planned; $4B financing in 2025",
    ],
    [
        "Compass Datacenters",
        "Multiple campuses (single-tenant)",
        "~3M+ sq ft",
        "US primarily",
        "Red Oak TX, Lauderdale County MS, Virginia, Other US markets",
        "$10B Lauderdale County investment (8 DCs); Red Oak 180MW completed",
    ],
    [
        "NTT Global Data Centers",
        "150+ datacenters",
        "~12M+ sq ft",
        "20+ countries across Americas, EMEA, APAC",
        "Tokyo, Chicago, Dallas, Phoenix, Virginia, Hillsboro OR, Johor Bahru, Frankfurt, Berlin, Jakarta, Mumbai",
        "370+ MW delivered in 2024; 10 new facilities; $10B investment through 2027; Johor Bahru 290MW campus",
    ],
    [
        "CoreWeave",
        "28+ facilities",
        "~2M+ sq ft",
        "US primarily",
        "Multiple US locations for GPU cloud infrastructure",
        "Rapid facility buildout; IPO in 2025 (NASDAQ: CRWV); CapEx doubling in 2026",
    ],
]

for row_idx, row_data in enumerate(footprint_data, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws5.cell(row=row_idx, column=col_idx, value=val)

style_data_rows(ws5, 2, len(footprint_data) + 1, len(headers5))

col_widths5 = [25, 30, 22, 45, 70, 70]
for i, w in enumerate(col_widths5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SHEET 6: Sources & Notes
# ============================================================
ws6 = wb.create_sheet("Sources & Notes")

ws6.cell(row=1, column=1, value="Sources & Methodology Notes")
ws6.cell(row=1, column=1).font = Font(bold=True, size=14)

notes = [
    "",
    "Data Sources:",
    "- Public company filings (10-K, 10-Q, earnings calls) for Digital Realty and Equinix",
    "- IPO S-1 filings for CoreWeave",
    "- Press releases and corporate announcements for private companies (QTS, CyrusOne, Vantage, Aligned, STACK, Compass, NTT)",
    "- Industry reports: U.S. Data Center Colocation Market Report 2025-2030 (ResearchAndMarkets)",
    "- U.S. Colocation Data Center Portfolio Analysis Report 2025 (GlobeNewsWire)",
    "- Market data aggregated from michaelbommarito.com/wiki/datacenters",
    "- News sources: DataCenterFrontier, Bisnow, Reuters, CNBC, CRN, VentureBeat, CoStar",
    "",
    "Important Notes:",
    "- All figures are estimates based on publicly available data as of Q4 2025 / early 2026",
    "- Private company financials (revenue, exact CapEx) are not publicly disclosed; capital raised / financing amounts are used as proxies",
    "- AI vs Non-AI capacity split is estimated based on company disclosures about AI share of new signings, AI-specific facility builds, and industry analyst estimates",
    "- MW/GW capacity figures may include operational, under construction, and secured/planned capacity depending on source",
    "- 'Capacity Live' and 'Under Construction' are estimates based on total disclosed capacity minus known pipeline/planned projects",
    "- Digital Realty and Equinix MW figures may undercount total capacity as they include different definitions (colocation MW vs total IT load)",
    "- QTS disclosed capacity of 4,752 MW includes projects across all stages; Vantage 4,187 MW is US-disclosed only",
    "- CoreWeave is a GPU cloud provider, not traditional colocation; included for completeness as AI infrastructure vendor",
    "- Customer information is based on public partnerships, press releases, and analyst reports; actual customer lists are confidential",
    "",
    "Abbreviations:",
    "- MW = Megawatts (of IT critical load capacity)",
    "- GW = Gigawatts (1 GW = 1,000 MW)",
    "- CapEx = Capital Expenditure",
    "- DLR = Digital Realty Trust",
    "- EQIX = Equinix",
    "- EV = Enterprise Value",
    "- TCV = Total Contract Value",
    "- CMBS = Commercial Mortgage-Backed Securities",
    "- NoVA = Northern Virginia",
    "- DFW = Dallas-Fort Worth",
    "- EMEA = Europe, Middle East, Africa",
    "- APAC = Asia-Pacific",
    "",
    "Report compiled: February 2026",
]

for i, note in enumerate(notes, 2):
    ws6.cell(row=i, column=1, value=note)
    if note.endswith(":") and note != "":
        ws6.cell(row=i, column=1).font = Font(bold=True, size=11)

ws6.column_dimensions["A"].width = 120

output_path = "/workspace/datacenter_vendor_metrics.xlsx"
wb.save(output_path)
print(f"Excel file saved to: {output_path}")
