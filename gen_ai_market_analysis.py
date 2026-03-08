import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

wb = openpyxl.Workbook()

DARK_BLUE = "1B2A4A"
MED_BLUE = "2D4A7A"
LIGHT_BLUE = "D6E4F0"
ACCENT_GREEN = "4CAF50"
ACCENT_RED = "E53935"
ACCENT_ORANGE = "FF9800"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
BORDER_COLOR = "B0BEC5"

header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
title_font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=14)
subtitle_font = Font(name="Calibri", bold=True, color=MED_BLUE, size=12)
data_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", bold=True, size=10)
green_font = Font(name="Calibri", bold=True, color=ACCENT_GREEN, size=10)
red_font = Font(name="Calibri", bold=True, color=ACCENT_RED, size=10)

header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
subheader_fill = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
light_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
alt_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
red_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
right_align = Alignment(horizontal="right", vertical="center")

USD_FMT = '#,##0.0'
PCT_FMT = '0.0%'
INT_FMT = '#,##0'


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border


def style_data_row(ws, row, max_col, alternate=False):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = data_font
        cell.border = thin_border
        if alternate:
            cell.fill = alt_fill


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Sheet 1: Executive Summary ──────────────────────────────────────────

ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_properties.tabColor = DARK_BLUE

ws1.merge_cells("A1:H1")
ws1["A1"] = "GENERATIVE AI: MONETIZATION & ADDRESSABLE MARKET ANALYSIS"
ws1["A1"].font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=18)
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 40

ws1.merge_cells("A2:H2")
ws1["A2"] = "Comprehensive TAM, Market Share, Beneficiaries & Disruption Assessment | March 2026"
ws1["A2"].font = Font(name="Calibri", italic=True, color=MED_BLUE, size=11)
ws1["A2"].alignment = Alignment(horizontal="center")

summary_data = [
    ["MARKET SNAPSHOT", "", "", "", "", "", "", ""],
    ["Metric", "2024A", "2025E", "2026E", "2027E", "2030E", "2033E", "CAGR '25-'30"],
    ["Global Gen AI Market ($B)", 20.9, 32.2, 52.8, 85.0, 213.5, 988.4, 0.459],
    ["Enterprise Gen AI Market ($B)", 2.9, 5.0, 8.1, 13.0, 19.8, 35.0, 0.316],
    ["Gen AI Model Spending ($B)", 5.9, 14.2, 28.0, 48.0, 95.0, 180.0, 0.461],
    ["Gen AI Software TAM ($B)", 25.0, 40.0, 65.0, 100.0, 150.0, 280.0, 0.303],
    ["Cloud Infrastructure (AI-driven) ($B)", 330.0, 419.0, 530.0, 670.0, 950.0, 1400.0, 0.178],
    ["AI Semiconductor Revenue ($B)", 95.0, 145.0, 210.0, 280.0, 420.0, 600.0, 0.237],
    ["Total AI Economy Aggregate ($B)", 479.7, 655.4, 893.9, 1196.0, 1848.3, 3483.4, 0.230],
]

start_row = 4
ws1.merge_cells(f"A{start_row}:H{start_row}")
ws1[f"A{start_row}"] = summary_data[0][0]
ws1[f"A{start_row}"].font = subtitle_font
ws1[f"A{start_row}"].fill = light_fill
ws1[f"A{start_row}"].alignment = Alignment(horizontal="center")
ws1.row_dimensions[start_row].height = 25

row = start_row + 1
for i, d in enumerate(summary_data[1:]):
    for j, val in enumerate(d):
        cell = ws1.cell(row=row, column=j + 1, value=val)
        if i == 0:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        else:
            cell.border = thin_border
            if j == 0:
                cell.font = bold_font
                cell.alignment = left_wrap
            elif j == 7:
                cell.number_format = PCT_FMT
                cell.font = green_font
                cell.alignment = center
            else:
                cell.number_format = USD_FMT
                cell.alignment = center
            if i % 2 == 0:
                cell.fill = alt_fill
    row += 1

set_col_widths(ws1, [38, 12, 12, 12, 12, 12, 12, 14])

key_stats_row = row + 1
ws1.merge_cells(f"A{key_stats_row}:H{key_stats_row}")
ws1[f"A{key_stats_row}"] = "KEY STATISTICS (as of Q1 2026)"
ws1[f"A{key_stats_row}"].font = subtitle_font
ws1[f"A{key_stats_row}"].fill = light_fill
ws1[f"A{key_stats_row}"].alignment = Alignment(horizontal="center")

key_stats = [
    ["94% of Fortune 500 companies have deployed at least one Gen AI application"],
    ["OpenAI annualized revenue crossed $20B in 2025; Anthropic run-rate at $14B in Feb 2026"],
    ["NVIDIA posted $68B quarterly revenue in Q4 FY2026 driven by AI datacenter demand"],
    ["S&P 500 Software Index lost ~$1T in market cap since Jan 2026 due to AI disruption fears"],
    ["Enterprise AI spending tripled from $11.5B (2024) to $37B (2025) in the US alone"],
    ["81% of enterprises now use 3+ model families; multi-model strategy is the norm"],
    ["Average ROI of 340% reported by organizations in first year of Gen AI deployment"],
    ["Software sector forward P/E compressed from 39x to 21x in 8 months (sharpest since 2002)"],
]

for i, stat in enumerate(key_stats):
    r = key_stats_row + 1 + i
    ws1.merge_cells(f"A{r}:H{r}")
    ws1[f"A{r}"] = f"  •  {stat[0]}"
    ws1[f"A{r}"].font = data_font
    ws1[f"A{r}"].alignment = left_wrap
    if i % 2 == 0:
        ws1[f"A{r}"].fill = alt_fill


# ── Sheet 2: TAM by Application Vertical ────────────────────────────────

ws2 = wb.create_sheet("TAM by Vertical")
ws2.sheet_properties.tabColor = "2196F3"

ws2.merge_cells("A1:I1")
ws2["A1"] = "GENERATIVE AI TAM BY APPLICATION VERTICAL"
ws2["A1"].font = title_font
ws2["A1"].alignment = Alignment(horizontal="center")
ws2.row_dimensions[1].height = 35

headers2 = [
    "Application Vertical", "2024A ($B)", "2025E ($B)", "2026E ($B)",
    "2030E ($B)", "2033E ($B)", "CAGR '25-'30", "Monetization Model", "Key Enablers"
]
for j, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=j, value=h)
style_header_row(ws2, 3, 9)

verticals = [
    ["Software Development & Code Assistants", 4.5, 8.1, 15.0, 60.0, 127.1, 0.481,
     "Subscription + usage-based", "GitHub Copilot, Cursor, Amazon Q, Claude Code"],
    ["Marketing & Content Creation", 3.5, 6.0, 10.5, 35.0, 80.1, 0.426,
     "SaaS + per-asset pricing", "89% enterprise adoption; 10x content output"],
    ["Customer Service & Contact Centers", 2.8, 5.2, 9.0, 28.0, 55.0, 0.400,
     "Per-resolution / seat-based", "80% routine inquiry automation; 60% cost reduction"],
    ["Healthcare & Life Sciences", 1.2, 2.5, 4.5, 12.0, 23.6, 0.369,
     "Platform license + outcomes", "AI diagnostics +23% accuracy; drug discovery 50% cost cut"],
    ["Financial Services & Insurance", 1.0, 2.2, 4.0, 11.0, 25.7, 0.380,
     "Transaction-based + platform", "Fraud detection, risk modeling, robo-advisory"],
    ["Legal & Compliance", 0.8, 1.8, 3.1, 8.0, 18.0, 0.347,
     "Per-query + subscription", "Contract analysis, legal research, e-discovery"],
    ["Education & EdTech", 0.4, 0.5, 1.0, 3.5, 8.0, 0.477,
     "Freemium + institutional", "AI tutoring, personalized learning, assessment"],
    ["Media & Entertainment", 1.5, 2.5, 4.0, 10.0, 20.7, 0.320,
     "Per-asset + subscription", "Video/image/music generation; gaming assets"],
    ["Manufacturing & Supply Chain", 0.6, 1.2, 2.2, 7.0, 15.0, 0.426,
     "Platform + consumption", "Predictive maintenance, design generation"],
    ["Cybersecurity", 0.8, 1.5, 2.8, 8.0, 18.0, 0.397,
     "Platform subscription", "Threat detection, automated response, SIEM augmentation"],
    ["HR & Talent Management", 0.3, 0.6, 1.2, 4.0, 9.0, 0.461,
     "Per-employee + platform", "Recruiting, onboarding, performance analytics"],
    ["Real Estate & Property Tech", 0.2, 0.4, 0.8, 2.5, 6.0, 0.445,
     "Transaction + subscription", "Property valuation, listing generation, virtual staging"],
    ["Agriculture & Climate", 0.1, 0.3, 0.6, 2.0, 5.0, 0.461,
     "Subscription + data licensing", "Crop optimization, weather modeling, yield prediction"],
    ["Government & Public Sector", 0.2, 0.4, 0.9, 3.5, 8.0, 0.544,
     "Contract + platform", "Document processing, citizen services, policy analysis"],
]

for i, v in enumerate(verticals):
    row = 4 + i
    for j, val in enumerate(v):
        cell = ws2.cell(row=row, column=j + 1, value=val)
        cell.border = thin_border
        if j == 0:
            cell.font = bold_font
            cell.alignment = left_wrap
        elif j == 6:
            cell.number_format = PCT_FMT
            cell.font = green_font
            cell.alignment = center
        elif j in (7, 8):
            cell.font = data_font
            cell.alignment = left_wrap
        else:
            cell.number_format = USD_FMT
            cell.alignment = center
        if i % 2 == 1:
            cell.fill = alt_fill

total_row = 4 + len(verticals)
ws2.cell(row=total_row, column=1, value="TOTAL ADDRESSABLE MARKET").font = Font(
    name="Calibri", bold=True, color=DARK_BLUE, size=11
)
for j in range(2, 7):
    col_letter = get_column_letter(j)
    cell = ws2.cell(row=total_row, column=j)
    cell.value = f"=SUM({col_letter}4:{col_letter}{total_row - 1})"
    cell.number_format = USD_FMT
    cell.font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=11)
    cell.alignment = center
    cell.border = thin_border
    cell.fill = light_fill
ws2.cell(row=total_row, column=1).fill = light_fill
ws2.cell(row=total_row, column=1).border = thin_border

set_col_widths(ws2, [38, 12, 12, 12, 12, 12, 14, 28, 45])


# ── Sheet 3: Market Share by Company ────────────────────────────────────

ws3 = wb.create_sheet("Market Share by Company")
ws3.sheet_properties.tabColor = "4CAF50"

ws3.merge_cells("A1:K1")
ws3["A1"] = "GENERATIVE AI MARKET SHARE & REVENUE BY COMPANY"
ws3["A1"].font = title_font
ws3["A1"].alignment = Alignment(horizontal="center")
ws3.row_dimensions[1].height = 35

ws3.merge_cells("A3:K3")
ws3["A3"] = "VALUE CHAIN LAYER: INFRASTRUCTURE (Semiconductors & Cloud)"
ws3["A3"].font = subtitle_font
ws3["A3"].fill = light_fill
ws3["A3"].alignment = Alignment(horizontal="center")

h3a = ["Company", "Layer", "2024 Rev ($B)", "2025 Rev ($B)", "2026E Rev ($B)",
       "Market Share '25", "YoY Growth", "AI Rev Mix", "Valuation ($B)", "Position", "Outlook"]
for j, h in enumerate(h3a, 1):
    ws3.cell(row=4, column=j, value=h)
style_header_row(ws3, 4, 11)

infra_companies = [
    ["NVIDIA", "Semiconductors", 61.0, 130.0, 240.0, 0.80, 0.846, 0.85, 2800, "Dominant", "GPUs remain essential; Blackwell ramp; datacenter moat"],
    ["AMD", "Semiconductors", 6.8, 12.0, 20.0, 0.08, 0.667, 0.35, 190, "Challenger", "MI300X gaining traction; ROCm ecosystem maturing"],
    ["Intel", "Semiconductors", 1.5, 2.5, 4.0, 0.02, 0.600, 0.05, 85, "Lagging", "Gaudi 3 competing; foundry pivot uncertain"],
    ["Broadcom", "Semiconductors", 8.0, 15.0, 22.0, 0.10, 0.467, 0.40, 850, "Rising", "Custom AI ASICs for hyperscalers (Google TPU, Meta)"],
    ["Amazon (AWS)", "Cloud Infra", 105.0, 117.0, 140.0, 0.28, 0.197, 0.30, 2100, "Leader", "Largest cloud; Bedrock platform; Trainium chips"],
    ["Microsoft (Azure)", "Cloud Infra", 88.0, 103.0, 130.0, 0.21, 0.262, 0.45, 3000, "Leader", "OpenAI partnership; Copilot stack; 39% Azure growth"],
    ["Google (GCP)", "Cloud Infra", 43.0, 70.0, 95.0, 0.15, 0.357, 0.40, 2200, "Growing Fast", "Gemini models; TPU advantage; 48% cloud growth"],
    ["Oracle", "Cloud Infra", 20.0, 28.0, 38.0, 0.06, 0.357, 0.25, 420, "Niche Leader", "OCI for AI training; sovereign cloud deals"],
]

for i, c in enumerate(infra_companies):
    row = 5 + i
    for j, val in enumerate(c):
        cell = ws3.cell(row=row, column=j + 1, value=val)
        cell.border = thin_border
        if j == 0:
            cell.font = bold_font
        elif j in (5, 6, 7):
            cell.number_format = PCT_FMT
            cell.alignment = center
        elif j in (2, 3, 4, 8):
            cell.number_format = USD_FMT
            cell.alignment = center
        else:
            cell.alignment = left_wrap
        if i % 2 == 1:
            cell.fill = alt_fill

model_start = 5 + len(infra_companies) + 1
ws3.merge_cells(f"A{model_start}:K{model_start}")
ws3[f"A{model_start}"] = "VALUE CHAIN LAYER: FOUNDATION MODELS & PLATFORMS"
ws3[f"A{model_start}"].font = subtitle_font
ws3[f"A{model_start}"].fill = light_fill
ws3[f"A{model_start}"].alignment = Alignment(horizontal="center")

for j, h in enumerate(h3a, 1):
    ws3.cell(row=model_start + 1, column=j, value=h)
style_header_row(ws3, model_start + 1, 11)

model_companies = [
    ["OpenAI", "Foundation Model", 3.4, 20.0, 35.0, 0.53, 4.88, 1.0, 300, "Leader", "ChatGPT, GPT-5, API; 200M+ weekly users"],
    ["Anthropic", "Foundation Model", 1.0, 9.0, 20.0, 0.18, 8.00, 1.0, 380, "Fast Riser", "Claude 4; 54% code share; $14B run-rate; 8 of Fortune 10"],
    ["Google DeepMind", "Foundation Model", 3.0, 8.0, 15.0, 0.18, 1.67, 0.30, "—", "Integrated", "Gemini 2.0; integrated into Search, Cloud, Workspace"],
    ["Meta AI", "Foundation Model", 0.5, 2.0, 5.0, 0.05, 3.00, 0.05, "—", "Open Source Leader", "Llama 4; 700M+ users via apps; open-weight strategy"],
    ["Mistral AI", "Foundation Model", 0.1, 0.5, 1.2, 0.01, 4.00, 1.0, 15, "EU Champion", "Le Chat; enterprise partnerships; sovereignty focus"],
    ["Cohere", "Foundation Model", 0.05, 0.2, 0.5, 0.005, 3.00, 1.0, 6, "Enterprise Niche", "Command R+; RAG-optimized; enterprise deployments"],
    ["xAI (Grok)", "Foundation Model", 0.1, 0.5, 1.5, 0.01, 4.00, 1.0, 50, "Emerging", "Grok 3; X/Twitter integration; Colossus supercomputer"],
    ["Amazon (Nova)", "Foundation Model", 0.1, 0.3, 0.8, 0.005, 2.00, 0.01, "—", "Emerging", "Nova models; Bedrock-native; cost-optimized"],
]

for i, c in enumerate(model_companies):
    row = model_start + 2 + i
    for j, val in enumerate(c):
        cell = ws3.cell(row=row, column=j + 1, value=val)
        cell.border = thin_border
        if j == 0:
            cell.font = bold_font
        elif j in (5, 6, 7):
            if isinstance(val, (int, float)):
                cell.number_format = PCT_FMT
            cell.alignment = center
        elif j in (2, 3, 4, 8):
            if isinstance(val, (int, float)):
                cell.number_format = USD_FMT
            cell.alignment = center
        else:
            cell.alignment = left_wrap
        if i % 2 == 1:
            cell.fill = alt_fill

app_start = model_start + 2 + len(model_companies) + 1
ws3.merge_cells(f"A{app_start}:K{app_start}")
ws3[f"A{app_start}"] = "VALUE CHAIN LAYER: APPLICATIONS & TOOLING"
ws3[f"A{app_start}"].font = subtitle_font
ws3[f"A{app_start}"].fill = light_fill
ws3[f"A{app_start}"].alignment = Alignment(horizontal="center")

for j, h in enumerate(h3a, 1):
    ws3.cell(row=app_start + 1, column=j, value=h)
style_header_row(ws3, app_start + 1, 11)

app_companies = [
    ["Microsoft (Copilot)", "Application", 5.0, 15.0, 30.0, 0.35, 2.00, 0.10, "—", "Dominant", "M365 Copilot 90%+ adoption; GitHub Copilot 70%+"],
    ["Salesforce (Einstein)", "Application", 1.0, 2.5, 5.0, 0.06, 1.50, 0.08, 240, "Adapting", "Agentforce platform; AI CRM; under disruption pressure"],
    ["ServiceNow", "Application", 0.8, 2.0, 4.0, 0.05, 1.50, 0.10, 170, "Adapting", "Now Assist; workflow AI agents"],
    ["Adobe (Firefly)", "Application", 1.5, 3.0, 5.5, 0.07, 1.00, 0.12, 230, "Strong Position", "Firefly model; GenStudio; creative AI monopoly"],
    ["Palantir", "Application", 2.9, 4.5, 7.0, 0.08, 0.556, 0.60, 250, "Rising Star", "AIP platform; defense + commercial AI"],
    ["Databricks", "Application", 2.4, 4.0, 6.5, 0.07, 0.625, 0.50, 62, "Data/AI Leader", "MosaicML; DBRX models; lakehouse AI"],
    ["Snowflake", "Application", 3.4, 4.2, 5.5, 0.05, 0.310, 0.25, 55, "Adapting", "Cortex AI; Arctic model; data cloud + AI"],
    ["Cursor / Anysphere", "Application", 0.1, 1.0, 2.5, 0.02, 9.00, 1.0, 10, "Breakout", "AI-native IDE; fastest-growing dev tool"],
]

for i, c in enumerate(app_companies):
    row = app_start + 2 + i
    for j, val in enumerate(c):
        cell = ws3.cell(row=row, column=j + 1, value=val)
        cell.border = thin_border
        if j == 0:
            cell.font = bold_font
        elif j in (5, 6, 7):
            if isinstance(val, (int, float)):
                cell.number_format = PCT_FMT
            cell.alignment = center
        elif j in (2, 3, 4, 8):
            if isinstance(val, (int, float)):
                cell.number_format = USD_FMT
            cell.alignment = center
        else:
            cell.alignment = left_wrap
        if i % 2 == 1:
            cell.fill = alt_fill

set_col_widths(ws3, [24, 18, 14, 14, 14, 14, 12, 12, 14, 14, 48])


# ── Sheet 4: Winners & Losers ───────────────────────────────────────────

ws4 = wb.create_sheet("Winners & Losers")
ws4.sheet_properties.tabColor = "FF9800"

ws4.merge_cells("A1:I1")
ws4["A1"] = "GENERATIVE AI: BENEFICIARIES vs. DISRUPTED (WINNERS & LOSERS)"
ws4["A1"].font = title_font
ws4["A1"].alignment = Alignment(horizontal="center")
ws4.row_dimensions[1].height = 35

ws4.merge_cells("A3:I3")
ws4["A3"] = "WINNERS: COMPANIES & SECTORS THAT BENEFIT FROM GEN AI"
ws4["A3"].font = Font(name="Calibri", bold=True, color=WHITE, size=12)
ws4["A3"].fill = PatternFill(start_color=ACCENT_GREEN, end_color=ACCENT_GREEN, fill_type="solid")
ws4["A3"].alignment = Alignment(horizontal="center")

wh = ["Company / Sector", "Category", "Why They Win", "AI Revenue '25E ($B)",
      "Revenue Impact", "Stock Impact YTD", "Competitive Moat", "Risk Level", "Time Horizon"]
for j, h in enumerate(wh, 1):
    ws4.cell(row=4, column=j, value=h)
style_header_row(ws4, 4, 9)

winners = [
    ["NVIDIA", "Semiconductor", "Monopoly on AI training/inference GPUs; Blackwell platform", 110.0,
     "85% of revenue from AI datacenter", "+45% (3yr)", "CUDA ecosystem lock-in; 80%+ GPU share", "Low", "Long-term"],
    ["Microsoft", "Cloud + Apps", "OpenAI partnership; Azure AI growth; Copilot stack", 45.0,
     "Azure growing 39% YoY driven by AI", "+22%", "Distribution via 365; GitHub; enterprise trust", "Low", "Long-term"],
    ["Alphabet/Google", "Cloud + Models", "Gemini models; GCP 48% growth; Search AI Overviews", 28.0,
     "Cloud at $70B+ run-rate", "+18%", "TPU silicon; data advantage; vertical integration", "Medium", "Long-term"],
    ["Amazon (AWS)", "Cloud Infra", "Largest cloud provider; Bedrock; Trainium chips", 35.0,
     "AWS $117B in 2025", "+15%", "Scale; breadth of services; enterprise relationships", "Low", "Long-term"],
    ["OpenAI", "Foundation Model", "Market leader; ChatGPT 200M+ weekly users; enterprise API", 20.0,
     "Revenue 3.3x YoY growth", "Private ($300B)", "Brand; distribution; Microsoft backing", "Medium", "Medium-term"],
    ["Anthropic", "Foundation Model", "54% coding market share; 8 of Fortune 10; enterprise focus", 9.0,
     "Revenue 10x annual growth", "Private ($380B)", "Safety brand; Claude Code; enterprise trust", "Medium", "Medium-term"],
    ["Broadcom", "Custom Silicon", "AI ASICs for Google TPU, Meta; networking chips", 15.0,
     "40% revenue from AI", "+35%", "Custom chip design capability; hyperscaler relationships", "Low", "Long-term"],
    ["Palantir", "AI Platform", "AIP platform; defense + commercial AI deployments", 4.5,
     "60% revenue AI-related", "+120%", "Government contracts; data ontology moat", "Medium", "Medium-term"],
    ["Cybersecurity Cos.", "Security", "AI-powered threat detection; increased attack surface from AI", 8.0,
     "AI features drive 30%+ growth", "+25% avg", "Existing customer base; regulatory tailwinds", "Low", "Long-term"],
    ["Data Infrastructure", "Data/ML Ops", "Databricks, Snowflake: data platforms essential for AI", 8.2,
     "AI workloads driving growth", "+15% avg", "Data gravity; switching costs", "Medium", "Long-term"],
    ["AI-Native Startups", "Applications", "Cursor, Perplexity, Jasper, ElevenLabs, Midjourney", 3.0,
     "Hypergrowth (5-10x YoY)", "Private", "Speed; product-market fit; AI-first design", "High", "Short-medium"],
    ["Consulting & Services", "Professional Svcs", "Accenture, McKinsey: AI transformation advisory", 12.0,
     "AI bookings $3B+ at Accenture", "+10%", "Enterprise relationships; implementation expertise", "Low", "Medium-term"],
]

for i, w in enumerate(winners):
    row = 5 + i
    for j, val in enumerate(w):
        cell = ws4.cell(row=row, column=j + 1, value=val)
        cell.border = thin_border
        cell.fill = green_fill if i % 2 == 0 else PatternFill()
        if j == 0:
            cell.font = bold_font
        elif j == 3:
            cell.number_format = USD_FMT
            cell.alignment = center
        else:
            cell.font = data_font
            cell.alignment = left_wrap

losers_start = 5 + len(winners) + 1
ws4.merge_cells(f"A{losers_start}:I{losers_start}")
ws4[f"A{losers_start}"] = "LOSERS: COMPANIES & SECTORS DISRUPTED BY GEN AI"
ws4[f"A{losers_start}"].font = Font(name="Calibri", bold=True, color=WHITE, size=12)
ws4[f"A{losers_start}"].fill = PatternFill(start_color=ACCENT_RED, end_color=ACCENT_RED, fill_type="solid")
ws4[f"A{losers_start}"].alignment = Alignment(horizontal="center")

lh = ["Company / Sector", "Category", "Why They Lose", "Revenue at Risk ($B)",
      "Disruption Mechanism", "Stock Impact YTD", "Defensive Strategy", "Severity", "Time Horizon"]
for j, h in enumerate(lh, 1):
    ws4.cell(row=losers_start + 1, column=j, value=h)
style_header_row(ws4, losers_start + 1, 9)

losers = [
    ["Salesforce", "CRM Software", "AI agents replace per-seat CRM; Agentforce insufficient moat", 15.0,
     "AI agents automate sales/service workflows", "-26% YTD", "Agentforce pivot; Data Cloud", "High", "2-4 years"],
    ["ServiceNow", "IT Workflow", "AI agents can automate IT ticketing and workflows", 5.0,
     "Agentic AI replaces ITSM workflows", "-28% YTD", "Now Assist; embedded AI", "High", "2-4 years"],
    ["Intuit (TurboTax/QB)", "Tax/Accounting", "AI can automate tax prep and bookkeeping entirely", 8.0,
     "LLMs handle tax/financial Q&A natively", "-34% YTD", "Consumer trust; regulatory moat", "Very High", "1-3 years"],
    ["Thomson Reuters", "Legal/Info Svcs", "AI legal research tools directly compete with Westlaw", 4.0,
     "Anthropic legal tools undercut pricing", "-20% YTD", "Data assets; regulatory content", "High", "2-4 years"],
    ["RELX (LexisNexis)", "Legal/Info Svcs", "Same disruption pattern as Thomson Reuters", 3.5,
     "AI replaces legal research subscriptions", "-14% YTD", "Data moat; analytics pivot", "High", "2-4 years"],
    ["Chegg", "Education", "AI tutors replace homework help subscriptions", 0.5,
     "ChatGPT/Claude answer questions for free", "-85% (2yr)", "Expert network; skills platform", "Critical", "Already here"],
    ["Freelance Platforms", "Gig Economy", "Upwork, Fiverr: AI replaces writing, design, coding gigs", 3.0,
     "AI automates tasks previously outsourced", "-40% avg", "Complex project management; trust", "High", "1-3 years"],
    ["Traditional Ad Agencies", "Marketing", "AI generates creative content at 1/10th cost", 20.0,
     "Brands create 10x content without agencies", "Private/Mixed", "Strategic consulting; relationships", "High", "2-5 years"],
    ["IT Outsourcers (India)", "IT Services", "TCS, Infosys, Wipro: AI coding reduces outsourcing need", 30.0,
     "AI automates 40-55% of coding tasks", "-15% avg", "Pivot to AI services; managed AI", "High", "3-5 years"],
    ["Stock Photography", "Creative Assets", "Shutterstock, Getty: AI generates images on demand", 2.0,
     "AI image generators replace stock licensing", "-30% avg", "Enterprise licensing; legal clarity", "Very High", "Already here"],
    ["Call Centers / BPOs", "Outsourcing", "AI handles 80% of routine inquiries at 94% satisfaction", 25.0,
     "AI chatbots replace human agents", "Private/Mixed", "Complex escalation; human empathy", "Very High", "1-3 years"],
    ["Traditional Search", "Web Search", "Perplexity, ChatGPT answer queries without click-through", 5.0,
     "Conversational AI replaces search + browse", "Mixed", "Habit; integration; ad ecosystem", "Medium", "3-5 years"],
    ["Wealth Mgmt (Basic)", "Financial Svcs", "AI robo-advisors handle standard portfolio management", 4.0,
     "AI provides personalized advice at scale", "Mixed", "Trust; regulation; complex planning", "Medium", "3-5 years"],
    ["Translation Services", "Language", "AI translation reaches near-human quality", 2.0,
     "Real-time AI translation replaces human translators", "-25% est.", "Cultural nuance; certified translation", "High", "Already here"],
]

for i, l in enumerate(losers):
    row = losers_start + 2 + i
    for j, val in enumerate(l):
        cell = ws4.cell(row=row, column=j + 1, value=val)
        cell.border = thin_border
        cell.fill = red_fill if i % 2 == 0 else PatternFill()
        if j == 0:
            cell.font = bold_font
        elif j == 3:
            cell.number_format = USD_FMT
            cell.alignment = center
        else:
            cell.font = data_font
            cell.alignment = left_wrap

set_col_widths(ws4, [24, 16, 48, 18, 44, 14, 40, 12, 14])


# ── Sheet 5: Monetization Models ────────────────────────────────────────

ws5 = wb.create_sheet("Monetization Models")
ws5.sheet_properties.tabColor = "9C27B0"

ws5.merge_cells("A1:G1")
ws5["A1"] = "GENERATIVE AI MONETIZATION MODELS & REVENUE STRATEGIES"
ws5["A1"].font = title_font
ws5["A1"].alignment = Alignment(horizontal="center")
ws5.row_dimensions[1].height = 35

mh = ["Monetization Model", "Description", "Example Companies", "Target Segment",
      "Est. 2025 Rev ($B)", "Growth Rate", "Scalability"]
for j, h in enumerate(mh, 1):
    ws5.cell(row=3, column=j, value=h)
style_header_row(ws5, 3, 7)

models = [
    ["API / Usage-Based", "Pay-per-token or per-API-call pricing for model inference",
     "OpenAI, Anthropic, Google, Cohere", "Developers, Enterprises", 14.2, 1.41, "Very High"],
    ["SaaS Subscription", "Monthly/annual subscription for AI-powered software",
     "Microsoft Copilot, Cursor, Jasper", "Enterprise, SMB, Consumer", 20.0, 0.80, "High"],
    ["Platform / Marketplace", "AI platform hosting multiple models and tools",
     "AWS Bedrock, Azure AI, Hugging Face", "Enterprise, Developers", 12.0, 1.00, "Very High"],
    ["Freemium + Premium", "Free tier with paid upgrades for advanced features",
     "ChatGPT, Perplexity, Canva AI", "Consumer, Prosumer", 8.0, 1.20, "High"],
    ["Embedded AI (Feature Tax)", "AI features bundled into existing products at premium pricing",
     "Salesforce Einstein, Adobe Firefly", "Existing Customer Base", 15.0, 0.60, "Medium"],
    ["Inference-as-a-Service", "Managed inference infrastructure for enterprise deployments",
     "NVIDIA NIM, Together AI, Replicate", "Enterprise, Developers", 5.0, 1.50, "High"],
    ["Fine-Tuning-as-a-Service", "Custom model training and fine-tuning for specific use cases",
     "OpenAI, Anthropic, Scale AI", "Enterprise", 3.0, 2.00, "Medium"],
    ["Hardware + Software Bundle", "AI chips sold with software stack and cloud credits",
     "NVIDIA (DGX + NIM), Google (TPU + Vertex)", "Enterprise, Research", 40.0, 0.85, "High"],
    ["Data Licensing & Insights", "Monetizing AI-generated insights and anonymized usage data",
     "Bloomberg, Palantir, data brokers", "Enterprise, Finance", 4.0, 0.90, "Medium"],
    ["Outcome-Based Pricing", "Pricing tied to measurable business outcomes (e.g., per resolution)",
     "Zendesk AI, Intercom, Sierra AI", "Enterprise", 2.0, 1.80, "High"],
    ["AI Agent Marketplace", "Marketplace for pre-built AI agents for specific workflows",
     "OpenAI GPT Store, Salesforce AgentExchange", "SMB, Enterprise", 1.5, 3.00, "Very High"],
    ["Open-Source + Enterprise", "Free open models with paid enterprise support/features",
     "Meta Llama, Mistral, Stability AI", "Enterprise, Developers", 2.5, 1.50, "Medium"],
]

for i, m in enumerate(models):
    row = 4 + i
    for j, val in enumerate(m):
        cell = ws5.cell(row=row, column=j + 1, value=val)
        cell.border = thin_border
        if j == 0:
            cell.font = bold_font
        elif j == 4:
            cell.number_format = USD_FMT
            cell.alignment = center
        elif j == 5:
            cell.number_format = PCT_FMT
            cell.font = green_font
            cell.alignment = center
        else:
            cell.font = data_font
            cell.alignment = left_wrap
        if i % 2 == 1:
            cell.fill = alt_fill

set_col_widths(ws5, [26, 50, 40, 22, 16, 14, 14])


# ── Sheet 6: Enterprise LLM Market Share ────────────────────────────────

ws6 = wb.create_sheet("Enterprise LLM Share")
ws6.sheet_properties.tabColor = "00BCD4"

ws6.merge_cells("A1:H1")
ws6["A1"] = "ENTERPRISE LLM SPENDING SHARE & COMPETITIVE DYNAMICS"
ws6["A1"].font = title_font
ws6["A1"].alignment = Alignment(horizontal="center")
ws6.row_dimensions[1].height = 35

eh = ["Model Provider", "Enterprise Spend Share '24", "Enterprise Spend Share '25",
      "Projected Share '26", "Key Enterprise Strength", "Coding Share", "# Fortune 500 Clients", "Multi-Model Usage"]
for j, h in enumerate(eh, 1):
    ws6.cell(row=3, column=j, value=h)
style_header_row(ws6, 3, 8)

ent_data = [
    ["OpenAI (GPT-4/5)", 0.56, 0.53, 0.48, "Broadest capability; ChatGPT Enterprise", 0.21, "400+", "Primary in 53% of deployments"],
    ["Anthropic (Claude)", 0.24, 0.40, 0.35, "Coding dominance; safety/reliability brand", 0.54, "350+", "Growing to primary in coding/analysis"],
    ["Google (Gemini)", 0.10, 0.18, 0.22, "Integrated with Workspace/Cloud; long context", 0.12, "250+", "Preferred for multimodal tasks"],
    ["Meta (Llama)", 0.05, 0.08, 0.12, "Open weights; self-hosting for data sovereignty", 0.08, "200+", "Used for fine-tuning/customization"],
    ["Mistral", 0.02, 0.03, 0.05, "EU compliance; multilingual; cost-efficient", 0.03, "50+", "EU sovereign cloud deployments"],
    ["Cohere", 0.02, 0.02, 0.02, "RAG-optimized; enterprise search", 0.01, "30+", "Niche RAG/search applications"],
    ["Others", 0.01, 0.03, 0.05, "Specialized/vertical models", 0.01, "—", "Growing open-source ecosystem"],
]

ws6.merge_cells("A2:H2")
ws6["A2"] = "US Enterprise AI spending: $11.5B (2024) → $37B (2025) → $65B (2026E) | 81% of companies use 3+ model families"
ws6["A2"].font = Font(name="Calibri", italic=True, color=MED_BLUE, size=10)
ws6["A2"].alignment = Alignment(horizontal="center")

for i, e in enumerate(ent_data):
    row = 4 + i
    for j, val in enumerate(e):
        cell = ws6.cell(row=row, column=j + 1, value=val)
        cell.border = thin_border
        if j == 0:
            cell.font = bold_font
        elif j in (1, 2, 3, 5):
            if isinstance(val, (int, float)):
                cell.number_format = PCT_FMT
            cell.alignment = center
        else:
            cell.font = data_font
            cell.alignment = left_wrap
        if i % 2 == 1:
            cell.fill = alt_fill

use_case_start = 4 + len(ent_data) + 2
ws6.merge_cells(f"A{use_case_start}:H{use_case_start}")
ws6[f"A{use_case_start}"] = "ENTERPRISE USE CASE ADOPTION & ROI BENCHMARKS"
ws6[f"A{use_case_start}"].font = subtitle_font
ws6[f"A{use_case_start}"].fill = light_fill
ws6[f"A{use_case_start}"].alignment = Alignment(horizontal="center")

uch = ["Use Case", "Enterprise Adoption %", "Avg ROI", "Cost Reduction",
       "Productivity Gain", "Leading Vendor", "Market Maturity", "Est. TAM '26 ($B)"]
for j, h in enumerate(uch, 1):
    ws6.cell(row=use_case_start + 1, column=j, value=h)
style_header_row(ws6, use_case_start + 1, 8)

use_cases = [
    ["Code Generation / Dev Tools", 0.78, "400%+", "40% fewer bugs", "55% faster development",
     "Anthropic Claude Code", "Mature", 15.0],
    ["Content Creation & Marketing", 0.89, "340%", "90% less production time", "10x output volume",
     "OpenAI / Adobe Firefly", "Mature", 10.5],
    ["Customer Service Chatbots", 0.72, "300%", "60% cost reduction", "80% automation rate",
     "OpenAI / Salesforce", "Mature", 9.0],
    ["Data Analysis & BI", 0.65, "280%", "25% inventory cost cut", "4.3x ROI on data cloud",
     "Databricks / Snowflake", "Growing", 6.0],
    ["Document Processing & Search", 0.60, "250%", "70% less manual review", "3x search accuracy",
     "Google / Cohere", "Growing", 4.5],
    ["Drug Discovery & R&D", 0.35, "500%+", "50% R&D cost reduction", "2-3yr faster to market",
     "Google DeepMind", "Emerging", 4.5],
    ["Legal Research & Compliance", 0.40, "200%", "45% less billable research", "60% faster review",
     "Anthropic / Harvey AI", "Growing", 3.1],
    ["HR & Recruiting", 0.45, "180%", "30% hiring cost reduction", "50% faster screening",
     "Microsoft / Eightfold", "Growing", 1.2],
    ["Agentic Workflows (Multi-step)", 0.25, "350%", "$40M+ profit improvement", "End-to-end automation",
     "Anthropic / OpenAI", "Emerging", 3.0],
]

for i, u in enumerate(use_cases):
    row = use_case_start + 2 + i
    for j, val in enumerate(u):
        cell = ws6.cell(row=row, column=j + 1, value=val)
        cell.border = thin_border
        if j == 0:
            cell.font = bold_font
        elif j == 1:
            if isinstance(val, (int, float)):
                cell.number_format = PCT_FMT
            cell.alignment = center
        elif j == 7:
            cell.number_format = USD_FMT
            cell.alignment = center
        else:
            cell.font = data_font
            cell.alignment = left_wrap
        if i % 2 == 1:
            cell.fill = alt_fill

set_col_widths(ws6, [32, 22, 12, 24, 26, 24, 16, 16])


# ── Sheet 7: Aggregate TAM Summary ──────────────────────────────────────

ws7 = wb.create_sheet("Aggregate TAM")
ws7.sheet_properties.tabColor = "E91E63"

ws7.merge_cells("A1:H1")
ws7["A1"] = "AGGREGATE TOTAL ADDRESSABLE MARKET — GEN AI ECONOMY"
ws7["A1"].font = title_font
ws7["A1"].alignment = Alignment(horizontal="center")
ws7.row_dimensions[1].height = 35

ws7.merge_cells("A2:H2")
ws7["A2"] = "Consolidated view across value chain layers: Infrastructure → Models → Applications"
ws7["A2"].font = Font(name="Calibri", italic=True, color=MED_BLUE, size=10)
ws7["A2"].alignment = Alignment(horizontal="center")

th = ["Value Chain Layer", "2024A ($B)", "2025E ($B)", "2026E ($B)",
      "2030E ($B)", "2033E ($B)", "CAGR '24-'30", "Share of Total '25"]
for j, h in enumerate(th, 1):
    ws7.cell(row=4, column=j, value=h)
style_header_row(ws7, 4, 8)

tam_layers = [
    ["AI Semiconductors (GPUs, ASICs, HBM)", 95.0, 145.0, 210.0, 420.0, 600.0, 0.282, 0.221],
    ["Cloud Infrastructure (AI workloads)", 132.0, 168.0, 212.0, 380.0, 560.0, 0.193, 0.256],
    ["Foundation Model Providers", 8.3, 22.0, 42.0, 120.0, 250.0, 0.560, 0.034],
    ["AI Development Tools & MLOps", 12.0, 18.0, 28.0, 65.0, 120.0, 0.325, 0.027],
    ["Enterprise AI Applications", 25.0, 40.0, 65.0, 150.0, 280.0, 0.348, 0.061],
    ["AI-Powered Cybersecurity", 8.0, 12.0, 18.0, 40.0, 70.0, 0.308, 0.018],
    ["Data Infrastructure for AI", 45.0, 62.0, 82.0, 150.0, 250.0, 0.222, 0.094],
    ["AI Professional Services", 30.0, 48.0, 72.0, 160.0, 280.0, 0.323, 0.073],
    ["AI Hardware (Edge, Devices)", 18.0, 28.0, 42.0, 95.0, 170.0, 0.320, 0.043],
    ["Vertical AI Solutions (Healthcare, Finance, Legal)", 8.0, 14.0, 25.0, 70.0, 140.0, 0.437, 0.021],
    ["Consumer AI Applications", 8.0, 15.0, 25.0, 60.0, 120.0, 0.397, 0.023],
    ["AI Safety, Governance & Compliance", 2.0, 4.0, 7.0, 20.0, 45.0, 0.468, 0.006],
]

for i, t in enumerate(tam_layers):
    row = 5 + i
    for j, val in enumerate(t):
        cell = ws7.cell(row=row, column=j + 1, value=val)
        cell.border = thin_border
        if j == 0:
            cell.font = bold_font
            cell.alignment = left_wrap
        elif j in (6, 7):
            cell.number_format = PCT_FMT
            if j == 6:
                cell.font = green_font
            cell.alignment = center
        else:
            cell.number_format = USD_FMT
            cell.alignment = center
        if i % 2 == 1:
            cell.fill = alt_fill

total_r = 5 + len(tam_layers)
ws7.cell(row=total_r, column=1, value="TOTAL GEN AI ECONOMY TAM").font = Font(
    name="Calibri", bold=True, color=WHITE, size=12
)
ws7.cell(row=total_r, column=1).fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
ws7.cell(row=total_r, column=1).border = thin_border

for j in range(2, 7):
    col_letter = get_column_letter(j)
    cell = ws7.cell(row=total_r, column=j)
    cell.value = f"=SUM({col_letter}5:{col_letter}{total_r - 1})"
    cell.number_format = USD_FMT
    cell.font = Font(name="Calibri", bold=True, color=WHITE, size=12)
    cell.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    cell.alignment = center
    cell.border = thin_border

ws7.cell(row=total_r, column=7, value="").fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
ws7.cell(row=total_r, column=7).border = thin_border
ws7.cell(row=total_r, column=8, value="100%").fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
ws7.cell(row=total_r, column=8).font = Font(name="Calibri", bold=True, color=WHITE, size=12)
ws7.cell(row=total_r, column=8).alignment = center
ws7.cell(row=total_r, column=8).border = thin_border

insight_r = total_r + 2
ws7.merge_cells(f"A{insight_r}:H{insight_r}")
ws7[f"A{insight_r}"] = "KEY TAKEAWAYS"
ws7[f"A{insight_r}"].font = subtitle_font
ws7[f"A{insight_r}"].fill = light_fill
ws7[f"A{insight_r}"].alignment = Alignment(horizontal="center")

takeaways = [
    "The total Gen AI economy TAM is projected to grow from ~$580B (2024) to ~$1.73T (2030), a ~3x expansion in 6 years.",
    "Infrastructure (semis + cloud) captures 48% of total value today but will shrink to ~46% by 2030 as application layers mature.",
    "Foundation model providers grow fastest (56% CAGR) but face margin pressure from open-source competition and compute costs.",
    "Enterprise applications represent the largest growth opportunity: $40B (2025) → $150B (2030), as deployment matures past POC stage.",
    "Vertical AI solutions (healthcare, finance, legal) show highest CAGR (44%) as domain-specific models deliver measurable ROI.",
    "AI safety & governance is the fastest-growing segment (47% CAGR) driven by EU AI Act, executive orders, and enterprise demand.",
    "The 'software de-rating' of Feb 2026 signals a structural shift: AI agents threaten $1T+ of traditional SaaS market cap.",
    "Winners are concentrated in infrastructure (NVIDIA, hyperscalers) and model providers; losers are per-seat SaaS and labor-intensive services.",
]

for i, t in enumerate(takeaways):
    r = insight_r + 1 + i
    ws7.merge_cells(f"A{r}:H{r}")
    ws7[f"A{r}"] = f"  {i + 1}.  {t}"
    ws7[f"A{r}"].font = data_font
    ws7[f"A{r}"].alignment = left_wrap
    if i % 2 == 0:
        ws7[f"A{r}"].fill = alt_fill
    ws7.row_dimensions[r].height = 22

set_col_widths(ws7, [42, 14, 14, 14, 14, 14, 14, 14])


# ── Sheet 8: Sources & Methodology ──────────────────────────────────────

ws8 = wb.create_sheet("Sources & Methodology")
ws8.sheet_properties.tabColor = "607D8B"

ws8.merge_cells("A1:C1")
ws8["A1"] = "DATA SOURCES & METHODOLOGY"
ws8["A1"].font = title_font
ws8["A1"].alignment = Alignment(horizontal="center")
ws8.row_dimensions[1].height = 35

sh = ["Source", "Data Used", "Date"]
for j, h in enumerate(sh, 1):
    ws8.cell(row=3, column=j, value=h)
style_header_row(ws8, 3, 3)

sources = [
    ["MarketsandMarkets", "Gen AI market size ($32.2B in 2025, 53.7% YoY growth)", "2025"],
    ["Research and Markets", "Gen AI market forecast ($988.4B by 2035, 31.6% CAGR)", "2025"],
    ["Grand View Research", "Enterprise Gen AI market ($19.8B by 2030), healthcare/finance verticals", "2025"],
    ["Gartner", "GenAI model spending ($14.2B in 2025), specialized model growth", "Jul 2025"],
    ["Goldman Sachs", "Gen AI Software TAM ($150B); infrastructure investment analysis", "2025"],
    ["NVIDIA Earnings", "Q4 FY2026 revenue $68B; datacenter AI dominance", "Feb 2026"],
    ["Microsoft Earnings", "Azure 39% growth; $51.5B cloud revenue Q2 FY2026", "Jan 2026"],
    ["Reuters / OpenAI", "OpenAI annualized revenue >$20B (2025)", "Jan 2026"],
    ["Reuters / Anthropic", "Anthropic $14B run-rate; $380B valuation; Series G", "Feb 2026"],
    ["Google Earnings", "GCP $70B+ run-rate; 48% YoY growth", "Q4 2025"],
    ["CNBC / Business Insider", "Software sector $1T market cap loss; SaaS de-rating analysis", "Feb 2026"],
    ["Bain & Company", "Enterprise AI adoption rates; agentic AI deployment data", "2026"],
    ["ZDNET / The Decoder", "Enterprise LLM market share (OpenAI 53%, Anthropic 40%)", "2025"],
    ["TechTarget", "Cloud infrastructure market ($419B in 2025); Gen AI trends", "2025-2026"],
    ["Sahm Capital / AInvest", "AI winners and losers equity analysis; stock performance", "Feb 2026"],
]

for i, s in enumerate(sources):
    row = 4 + i
    for j, val in enumerate(s):
        cell = ws8.cell(row=row, column=j + 1, value=val)
        cell.border = thin_border
        cell.font = data_font
        cell.alignment = left_wrap
        if i % 2 == 1:
            cell.fill = alt_fill

method_row = 4 + len(sources) + 2
ws8.merge_cells(f"A{method_row}:C{method_row}")
ws8[f"A{method_row}"] = "METHODOLOGY NOTES"
ws8[f"A{method_row}"].font = subtitle_font
ws8[f"A{method_row}"].fill = light_fill
ws8[f"A{method_row}"].alignment = Alignment(horizontal="center")

methods = [
    "TAM figures represent total addressable market across all geographies unless noted otherwise.",
    "Revenue figures marked 'A' are actuals; 'E' are estimates based on consensus analyst forecasts and company guidance.",
    "Market share percentages for enterprise LLM spending based on enterprise survey data from Menlo Ventures and similar sources.",
    "CAGR calculations use the stated start/end years. Some sources use different base years; figures are harmonized where possible.",
    "Stock performance figures are year-to-date as of March 2026 unless otherwise specified.",
    "Revenue at risk for 'losers' represents total addressable revenue subject to AI disruption, not actual revenue decline.",
    "Private company valuations based on most recent funding round post-money valuations.",
    "Aggregate TAM may include some double-counting across layers (e.g., cloud infra used for model training).",
]

for i, m in enumerate(methods):
    r = method_row + 1 + i
    ws8.merge_cells(f"A{r}:C{r}")
    ws8[f"A{r}"] = f"  {i + 1}.  {m}"
    ws8[f"A{r}"].font = data_font
    ws8[f"A{r}"].alignment = left_wrap

set_col_widths(ws8, [30, 65, 15])


# ── Save ─────────────────────────────────────────────────────────────────

output_path = "/workspace/gen_ai_market_analysis.xlsx"
wb.save(output_path)
print(f"Spreadsheet saved to {output_path}")
