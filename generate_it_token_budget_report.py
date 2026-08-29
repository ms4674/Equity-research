"""Generate an Excel report benchmarking enterprise IT budgets and AI token
consumption by company size, including token spend as % of the IT budget.

All benchmark figures were compiled in August 2026 from public sources
(Deloitte, Gartner via AIStackHub, Ramp AI Index, VendorBenchmark, Presenc AI,
ITBudgetCalculator, iternal.ai, Alvarez & Marsal). See the "Sources & Notes"
sheet for links and caveats.

Usage: python3 generate_it_token_budget_report.py
Output: it_budget_token_consumption_by_company_size.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT = "it_budget_token_consumption_by_company_size.xlsx"

# ---------------------------------------------------------------- styling ---
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
SUBHEADER_FILL = PatternFill("solid", fgColor="2E5496")
ACCENT_FILL = PatternFill("solid", fgColor="D6E4F0")
KEY_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
NOTE_FONT = Font(italic=True, size=9, color="595959")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

MONEY = '"$"#,##0'
MONEY_M = '"$"#,##0,,"M"'
PCT1 = "0.0%"
PCT2 = "0.00%"
NUM = "#,##0"


def style_header_row(ws, row, first_col, last_col, fill=HEADER_FILL):
    for c in range(first_col, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def style_body(ws, first_row, last_row, first_col, last_col):
    for r in range(first_row, last_row + 1):
        for c in range(first_col, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if cell.alignment.wrap_text is not True:
                cell.alignment = WRAP


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# ---------------------------------------------------------------- Summary ---
ws = wb.active
ws.title = "Summary"

ws["A1"] = "IT Budget vs. AI Token Consumption by Company Size (2026 benchmarks)"
ws["A1"].font = TITLE_FONT
ws["A2"] = ("Modeled from public 2025–2026 benchmarks. Yellow column = token/LLM API spend "
            "as % of total IT budget. All computed cells are live formulas — edit the "
            "assumptions to re-run the model.")
ws["A2"].font = NOTE_FONT

headers = [
    "Company size segment",
    "Employees (range)",
    "Representative headcount",
    "Representative annual revenue",
    "IT budget (% of revenue, median)",
    "Estimated annual IT budget",
    "Total AI spend ($/employee/month)",
    "Estimated annual AI spend (tools + usage)",
    "Token / LLM API spend ($/employee/month)",
    "Estimated annual token spend",
    "Token spend as % of IT budget",
    "Total AI spend as % of IT budget",
    "Est. token consumption (tokens/month)",
]
HDR_ROW = 4
for c, h in enumerate(headers, start=1):
    ws.cell(row=HDR_ROW, column=c, value=h)
style_header_row(ws, HDR_ROW, 1, len(headers))

# segment, employee range, headcount, revenue, IT % of revenue,
# AI PEPM, token PEPM, token volume label
rows = [
    ("Small business", "10–49", 30, 6_000_000, 0.069, 55, 12, "1M – 3M"),
    ("SMB", "50–249", 150, 30_000_000, 0.055, 70, 10, "10M – 30M"),
    ("Mid-market", "250–999", 600, 120_000_000, 0.070, 145, 9, "25M – 75M"),
    ("Enterprise", "1,000–4,999", 2500, 750_000_000, 0.061, 103, 8, "250M – 750M"),
    ("Large enterprise", "5,000+", 20000, 6_000_000_000, 0.048, 103, 8, "2.5B – 7.5B"),
]

r = HDR_ROW
for seg, emp_range, headcount, revenue, it_pct, ai_pepm, tok_pepm, tok_vol in rows:
    r += 1
    ws.cell(row=r, column=1, value=seg).font = BOLD
    ws.cell(row=r, column=2, value=emp_range)
    ws.cell(row=r, column=3, value=headcount).number_format = NUM
    ws.cell(row=r, column=4, value=revenue).number_format = MONEY
    ws.cell(row=r, column=5, value=it_pct).number_format = PCT1
    ws.cell(row=r, column=6, value=f"=D{r}*E{r}").number_format = MONEY
    ws.cell(row=r, column=7, value=ai_pepm).number_format = MONEY
    ws.cell(row=r, column=8, value=f"=C{r}*G{r}*12").number_format = MONEY
    ws.cell(row=r, column=9, value=tok_pepm).number_format = MONEY
    ws.cell(row=r, column=10, value=f"=C{r}*I{r}*12").number_format = MONEY
    kcell = ws.cell(row=r, column=11, value=f"=J{r}/F{r}")
    kcell.number_format = PCT2
    kcell.fill = KEY_FILL
    kcell.font = BOLD
    ws.cell(row=r, column=12, value=f"=H{r}/F{r}").number_format = PCT1
    ws.cell(row=r, column=13, value=tok_vol)
LAST = r
style_body(ws, HDR_ROW + 1, LAST, 1, len(headers))

notes = [
    "How to read: 'Token / LLM API spend' is metered usage-based consumption (API tokens) only. 'Total AI spend' also includes "
    "AI/copilot seat licenses, inference infrastructure, fine-tuning and program costs (tools + usage, excl. AI headcount).",
    "Key finding: token consumption is still a small slice of IT budgets today (~0.5–1.1% across segments; GenAI tools incl. "
    "licenses ≈ 1.7% of IT budget per VendorBenchmark 2026), but it is the fastest-growing line: token usage grew 1,001% "
    "Jan 2025 → Apr 2026 (Ramp) and planning guidance assumes 50–100% annual spend growth.",
    "Total AI spend runs ~5–12% of IT budget in this model; broader industry surveys that include AI infrastructure, "
    "implementation and talent put AI at 12–18% of the average IT budget in 2026 (up from 11% in 2024).",
]
r = LAST + 2
for n in notes:
    ws.cell(row=r, column=1, value=n).font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
    ws.cell(row=r, column=1).alignment = WRAP
    ws.row_dimensions[r].height = 28
    r += 1

set_widths(ws, [18, 11, 13, 15, 13, 14, 13, 14, 13, 13, 12, 12, 14])
ws.freeze_panes = "B5"
ws.row_dimensions[HDR_ROW].height = 45

# ------------------------------------------------- IT Budget Benchmarks -----
ws = wb.create_sheet("IT Budget Benchmarks")
ws["A1"] = "IT Budget Benchmarks by Company Size"
ws["A1"].font = TITLE_FONT

ws["A3"] = "View 1 — By headcount (ITBudgetCalculator / Deloitte-based, 2026)"
ws["A3"].font = BOLD
hdr = ["Company size", "Employees", "IT spend (% of revenue)", "Notes"]
for c, h in enumerate(hdr, start=1):
    ws.cell(row=4, column=c, value=h)
style_header_row(ws, 4, 1, 4, SUBHEADER_FILL)
data = [
    ("Startup", "1–50", "6.9% (avg)", "Highest ratio: fixed SaaS/security costs over little revenue"),
    ("SMB", "51–250", "4–7%", "First dedicated IT hires or MSP contract; security catch-up"),
    ("Mid-market", "251–1,000", "6–8%", "IT governance and FinOps forming; digital transformation spend"),
    ("Enterprise", "1,001–5,000", "4–6%", "Enterprise platforms and compliance tooling at scale"),
    ("Large enterprise", "5,000+", "3.7% (avg)", "Economies of scale on infrastructure and licensing"),
    ("All companies (cross-industry avg)", "—", "5.49%", "Deloitte, most recent measured cross-industry figure"),
]
r = 4
for row in data:
    r += 1
    for c, v in enumerate(row, start=1):
        ws.cell(row=r, column=c, value=v)
style_body(ws, 5, r, 1, 4)

r += 2
ws.cell(row=r, column=1, value="View 2 — By revenue (VendorBenchmark 2026, publicly traded and private companies)").font = BOLD
r += 1
hdr2_row = r
hdr = ["Company size", "Annual revenue", "IT spend range (% of revenue)", "Median"]
for c, h in enumerate(hdr, start=1):
    ws.cell(row=r, column=c, value=h)
style_header_row(ws, r, 1, 4, SUBHEADER_FILL)
data = [
    ("Small business", "$5M – $50M", "8.5% – 15%", "11.2%"),
    ("Small-to-medium business", "$50M – $500M", "6.5% – 9.2%", "7.8%"),
    ("Mid-market enterprise", "$500M – $5B", "5.2% – 7.1%", "6.1%"),
    ("Large enterprise", "$5B+", "4.1% – 5.8%", "4.8%"),
]
for row in data:
    r += 1
    for c, v in enumerate(row, start=1):
        ws.cell(row=r, column=c, value=v)
style_body(ws, hdr2_row + 1, r, 1, 4)

r += 2
ws.cell(row=r, column=1,
        value="The two views differ because size taxonomies differ across sources (headcount vs. revenue) and industry mix "
              "dominates: financial services runs 8–12% of revenue, technology/SaaS 12–18%, manufacturing 2–4%, "
              "traditional retail 2–3.5%. The inverse size relationship reflects economies of scale.").font = NOTE_FONT
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
ws.cell(row=r, column=1).alignment = WRAP
ws.row_dimensions[r].height = 42
set_widths(ws, [32, 16, 24, 52])

# ------------------------------------------- AI & Token Spend Benchmarks ----
ws = wb.create_sheet("AI & Token Spend Benchmarks")
ws["A1"] = "AI / Token Spend Benchmarks (2026)"
ws["A1"].font = TITLE_FONT

sections = [
    ("Median total AI spend by company size (AIStackHub 2026 operator survey)",
     ["Segment", "Revenue band", "Median annual AI spend", "Notes"],
     [
         ("SMB", "< $5M", "$18,400", "Top quartile spends 5–10× the median in every tier"),
         ("Mid-market", "$5M – $50M", "$127,000", ""),
         ("Enterprise", "$50M+", "$1,400,000", "Financial services median $2.1M; tech/SaaS $1.6M"),
     ]),
    ("AI spend per employee per year (Presenc AI, May 2026)",
     ["Segment", "Employees", "Annual AI spend / employee", "Notes"],
     [
         ("Solopreneur / micro", "< 10", "$200 – $600", "ChatGPT Plus + Claude + Cursor + Perplexity is roughly the floor"),
         ("Small business", "10–49", "$500 – $1,000", "Heavy reliance on free-tier consumer AI"),
         ("SMB", "50–249", "$900 – $1,400", "Mix of flagship tools and free tiers"),
         ("Mid-market", "250–999", "$1,500 – $2,000", "Higher per-capita than enterprise (less procurement leverage)"),
         ("Enterprise", "1,000+", "~$1,240 (avg)", "Financial services $3,200; tech $2,400; public sector $400"),
     ]),
    ("Token / LLM API consumption spend (metered usage only)",
     ["Metric", "Value", "Range / percentiles", "Source"],
     [
         ("Enterprise API consumption, per employee/month", "$8 (median)", "$2 – $24", "VendorBenchmark GenAI cohort 2026"),
         ("Token spend share of GenAI budget", "22–38%", "Copilot licenses are 38–52%", "VendorBenchmark GenAI cohort 2026"),
         ("Median company token spend / month", "$2,246", "Avg $140,842 (power-law skew)", "Ramp AI Index, Apr 2026"),
         ("Token spend percentiles / month", "p75: $14,843", "p90: $73,030 · p95: $211,409 · p99: $831,338", "Ramp AI Index, Apr 2026"),
         ("Median token spend per employee/month", "$11.38", "Top 10%: $611 · Top 1%: $7,450", "Ramp AI Index, Jun 2026"),
         ("Median PEPM by model-usage depth", "$46 (overall)", "4–10 models: $28 · 11–25: $130 · 26+: $442", "Ramp AI Index, Apr 2026"),
         ("Typical Fortune 500 annual AI cost", "$30M+", "Agents use 30–60× more tokens/task than chat", "Alvarez & Marsal, 2026"),
     ]),
    ("Token consumption volume (document/agent workloads, iternal.ai 2026)",
     ["Scenario", "Volume", "Tokens / month", "Notes"],
     [
         ("Small business (invoices)", "500 docs/mo", "~1.25M – 2.75M", ""),
         ("Mid-market (mixed docs)", "5,000 docs/mo", "~25M – 75M", ""),
         ("Enterprise (high volume)", "50,000 docs/mo", "~250M – 750M", ""),
         ("Large enterprise (batch)", "500,000 docs/mo", "~2.5B – 7.5B", "Enterprise support ops can reach 1.5–2.5B/mo"),
     ]),
    ("AI as % of IT budget",
     ["Metric", "Value", "Trend", "Source"],
     [
         ("Average IT budget share allocated to AI", "18% (2026)", "Up from 11% in 2024", "AIStackHub / Gartner-derived"),
         ("Mid-market AI share of IT budget", "12–18%", "Fastest-growing IT line item, passed cybersecurity", "Deloitte 2025 via aisavvy.io"),
         ("Enterprise AI share of IT budget (common range)", "5–15%", "High performers commit 20%+ of digital budgets", "CIO surveys (Gartner, Deloitte, PwC) 2025–26"),
         ("GenAI & LLM tools share of total IT budget", "1.7%", "+156% YoY — fastest-growing category", "VendorBenchmark IT budget allocation 2026"),
         ("Token usage growth", "+1,001%", "Jan 2025 → Apr 2026; spend grew 497% (prices fell)", "Ramp AI Index"),
         ("Planning assumption for AI spend growth", "50–100% / year", "Goldman Sachs: global tokens ×24 by 2030", "Ramp / Goldman Sachs Research"),
     ]),
]

r = 3
for title, hdr, data in sections:
    ws.cell(row=r, column=1, value=title).font = BOLD
    r += 1
    for c, h in enumerate(hdr, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, 1, len(hdr), SUBHEADER_FILL)
    first = r + 1
    for row in data:
        r += 1
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
    style_body(ws, first, r, 1, len(hdr))
    r += 2

set_widths(ws, [44, 20, 34, 46])

# ------------------------------------------------------- Sources & Notes ----
ws = wb.create_sheet("Sources & Notes")
ws["A1"] = "Methodology, Caveats and Sources"
ws["A1"].font = TITLE_FONT

method = [
    ("Methodology",
     "The Summary sheet is a model, not a survey: for each company-size segment a representative headcount and revenue are "
     "chosen, the median IT-budget-as-%-of-revenue benchmark is applied to estimate the IT budget, and per-employee-per-month "
     "(PEPM) benchmarks are applied to estimate total AI spend and token/API spend. Token % of IT budget = estimated annual "
     "token spend ÷ estimated annual IT budget. All computed cells are formulas, so assumptions can be edited."),
    ("Caveat — definitions vary",
     "'Token consumption' here means metered, usage-based LLM API spend. Some sources report total AI budgets (licenses + "
     "usage + infrastructure + talent + implementation), which are 3–10× larger than token spend alone. Size taxonomies also "
     "differ across sources (headcount vs. revenue bands)."),
    ("Caveat — extreme variance",
     "AI spend follows a power-law distribution: Ramp's April 2026 data shows a median of $2,246/month vs. an average of "
     "$140,842/month. Top-quartile spenders run 5–10× the median in every size tier. Industry matters more than size: "
     "financial services and tech/SaaS spend 2–3× the cross-industry average on both IT and AI."),
    ("Caveat — fast-moving data",
     "Token usage grew 1,001% between Jan 2025 and Apr 2026 while per-token prices fell ~80% year over year. Vendors moved "
     "enterprise contracts from flat seats to consumption pricing during 2025–26, so these benchmarks age quickly. Figures "
     "compiled August 2026."),
]
r = 3
for title, body in method:
    ws.cell(row=r, column=1, value=title).font = BOLD
    ws.cell(row=r, column=2, value=body).alignment = WRAP
    ws.row_dimensions[r].height = 68
    r += 1

r += 1
ws.cell(row=r, column=1, value="Sources").font = BOLD
r += 1
src_hdr_row = r
for c, h in enumerate(["Source", "What it provided", "URL"], start=1):
    ws.cell(row=r, column=c, value=h)
style_header_row(ws, r, 1, 3, SUBHEADER_FILL)

sources = [
    ("Ramp AI Index (Apr–Jun 2026)", "Company-level token spend distribution, PEPM medians/percentiles, usage growth",
     "https://ramp.com/blog/ai-token-cost-for-businesses"),
    ("VendorBenchmark — Enterprise GenAI Cost Benchmark 2026", "GenAI PEPM cost breakdown; API consumption $8 PEPM median; token share of GenAI budget",
     "https://vendorbenchmark.com/guides/enterprise-genai-cost-benchmark-2026"),
    ("VendorBenchmark — IT Spend as % of Revenue 2026", "IT budget benchmarks by revenue band and industry",
     "https://vendorbenchmark.com/blog/it-spend-percentage-revenue-2026-benchmark"),
    ("VendorBenchmark — IT Budget Allocation 2026", "GenAI & LLM tools = 1.7% of total IT budget, +156% YoY",
     "https://vendorbenchmark.com/blog/it-budget-allocation-benchmark-breakdown"),
    ("ITBudgetCalculator (Deloitte 2022 baseline)", "IT budget % of revenue by headcount segment; 5.49% cross-industry average",
     "https://itbudgetcalculator.com/average-it-budget"),
    ("AIStackHub — AI Spending by Industry 2026", "Median AI spend by size tier ($18K / $127K / $1.4M); AI = 18% of IT budget",
     "https://aistackhub.ai/ai-spending-by-industry"),
    ("Presenc AI — AI Spend per Employee 2026", "Per-employee AI spend by company size and industry",
     "https://presenc.ai/research/ai-spend-per-employee-2026"),
    ("iternal.ai — Token Usage Guide 2026", "Token volume per month by workload scale",
     "https://iternal.ai/token-usage-guide"),
    ("aisavvy.io — 2026 CFO Benchmarks", "Mid-market AI budget = 2–5% of revenue; AI = 12–18% of IT budget",
     "https://aisavvy.io/insights/how-much-should-we-budget-for-ai-this-year"),
    ("Alvarez & Marsal — End of the AI Flat-Rate Era", "Fortune 500 $30M+ annual AI cost; agentic token multipliers; pricing shift",
     "https://alvarezandmarsal-crg.com/insight/the-end-of-the-ai-flat-rate-era/"),
    ("Alice Labs — AI Automation ROI Benchmark 2026", "Enterprise AI = 5–15% of IT budget (CIO survey synthesis)",
     "https://alicelabs.ai/reports/ai-automation-roi-benchmark-2026"),
]
for name, what, url in sources:
    r += 1
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=what)
    cell = ws.cell(row=r, column=3, value=url)
    cell.hyperlink = url
    cell.font = Font(color="0563C1", underline="single")
style_body(ws, src_hdr_row + 1, r, 1, 3)
set_widths(ws, [42, 62, 60])

wb.save(OUTPUT)
print(f"Wrote {OUTPUT}")
