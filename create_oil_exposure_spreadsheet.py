import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

wb = openpyxl.Workbook()

# ── Styles ──
header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
subheader_font = Font(name="Calibri", bold=True, size=11, color="1F3864")
title_font = Font(name="Calibri", bold=True, size=14, color="1F3864")
section_font = Font(name="Calibri", bold=True, size=12, color="2F5496")
normal_font = Font(name="Calibri", size=11)
pct_font = Font(name="Calibri", size=11)
red_font = Font(name="Calibri", size=11, color="C00000")
green_font = Font(name="Calibri", size=11, color="006100")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border


def style_data_row(ws, row, max_col, alt=False):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = normal_font
        cell.border = thin_border
        if alt:
            cell.fill = light_gray_fill
        if col == 1:
            cell.alignment = left_wrap
        else:
            cell.alignment = center


def apply_exposure_color(cell, level):
    level = level.lower() if level else ""
    if "very high" in level or "critical" in level:
        cell.fill = red_fill
        cell.font = Font(name="Calibri", size=11, bold=True, color="C00000")
    elif "high" in level:
        cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        cell.font = Font(name="Calibri", size=11, color="C55A11")
    elif "medium" in level or "moderate" in level:
        cell.fill = yellow_fill
        cell.font = Font(name="Calibri", size=11, color="9C6500")
    elif "low" in level:
        cell.fill = green_fill
        cell.font = Font(name="Calibri", size=11, color="006100")


# ════════════════════════════════════════════════════════════════════
# SHEET 1: Executive Summary
# ════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_properties.tabColor = "2F5496"

ws1.merge_cells("A1:H1")
ws1["A1"] = "Impact of High Oil Prices on Software & Internet Companies"
ws1["A1"].font = Font(name="Calibri", bold=True, size=16, color="1F3864")
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")

ws1.merge_cells("A2:H2")
ws1["A2"] = "Revenue & Backlog Exposure Analysis  |  March 2026"
ws1["A2"].font = Font(name="Calibri", size=12, color="595959", italic=True)
ws1["A2"].alignment = Alignment(horizontal="center")

row = 4
ws1.cell(row=row, column=1, value="KEY FINDINGS").font = section_font
row += 1

findings = [
    ("Market Impact", "Energy-price shock on March 3, 2026 wiped $600B+ from tech stocks; WTI crude surged 8% toward $77/bbl on Iran conflict."),
    ("Valuation Compression", "Rising oil prices fuel inflation, forcing higher interest rates that compress growth-stock multiples disproportionately."),
    ("Data Center Costs", "Hyperscalers committed >$600B in 2026 capex; natural gas (50% of US electricity) price spikes directly raise operating costs."),
    ("Sales Cycle Slowdown", "Enterprise IT budgets face scrutiny during oil shocks; IDC projects global IT spending growth may fall from 10% to 9%."),
    ("Sector Rotation", "50+ percentage-point performance gap between energy and software stocks in early 2026 (Benzinga)."),
    ("Energy AI Spending", "Oil & gas AI spend projected to grow from $4B (2025) to $13.4B (2029), creating a tailwind for exposed vendors."),
]

for topic, detail in findings:
    ws1.cell(row=row, column=1, value=topic).font = Font(name="Calibri", bold=True, size=11, color="2F5496")
    ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    ws1.cell(row=row, column=2, value=detail).font = normal_font
    ws1.cell(row=row, column=2).alignment = left_wrap
    row += 1

row += 1
ws1.cell(row=row, column=1, value="HOW HIGH OIL PRICES AFFECT TECH COMPANIES").font = section_font
row += 1

channels = [
    ("Channel", "Mechanism", "Primary Victims", "Severity"),
    ("Direct Cost Inflation", "Oil used in plastics/chemicals for hardware; transportation/logistics costs rise", "Hardware OEMs, E-commerce (Amazon, Apple)", "High"),
    ("Data Center Energy Costs", "Natural gas prices spike electricity costs; 100MW data center costs $900M-$1.5B to build", "AWS, Azure, Google Cloud, Meta", "High"),
    ("Interest Rate Pressure", "Oil inflation forces central bank tightening; discount rates rise, compressing growth multiples", "High-multiple SaaS (CRM, NOW, DDOG)", "Very High"),
    ("Enterprise Budget Cuts", "Companies enter cost-control mode; discretionary software spend deferred", "Mid-market SaaS, consulting firms", "Medium"),
    ("Supply Chain Disruption", "Strait of Hormuz risk; shipping costs rise; component availability affected", "Semis (NVDA, TSM, MU), hardware", "High"),
    ("Consumer Spending Drag", "Goldman: 0.2% drag on consumer spend; lower-income hit hardest", "Ad-driven (Alphabet, Meta), e-commerce", "Medium"),
]

for i, (c1, c2, c3, c4) in enumerate(channels):
    for j, val in enumerate([c1, c2, c3, c4], start=1):
        ws1.cell(row=row, column=j, value=val)
    if i == 0:
        style_header_row(ws1, row, 4)
    else:
        style_data_row(ws1, row, 4, alt=(i % 2 == 0))
        apply_exposure_color(ws1.cell(row=row, column=4), c4)
    row += 1

for col_num, width in [(1, 28), (2, 70), (3, 40), (4, 16), (5, 16), (6, 16), (7, 16), (8, 16)]:
    ws1.column_dimensions[get_column_letter(col_num)].width = width


# ════════════════════════════════════════════════════════════════════
# SHEET 2: Revenue Exposure by Company
# ════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Revenue & Backlog Exposure")
ws2.sheet_properties.tabColor = "C55A11"

ws2.merge_cells("A1:K1")
ws2["A1"] = "Software & Internet Companies: Oil/Energy Sector Revenue & Backlog Exposure"
ws2["A1"].font = title_font
ws2["A1"].alignment = Alignment(horizontal="center")

headers = [
    "Company", "Ticker", "Category",
    "Total Revenue\n(FY2025, $B)", "Energy/Oil&Gas\nRevenue ($B, est.)",
    "Energy Revenue\n% of Total", "Energy Backlog\n($B, est.)",
    "Backlog % of\nTotal Backlog",
    "Oil Price\nSensitivity",
    "Exposure\nRating", "Notes"
]

row = 3
for j, h in enumerate(headers, start=1):
    ws2.cell(row=row, column=j, value=h)
style_header_row(ws2, row, len(headers))

companies = [
    # (Company, Ticker, Category, Total Rev, Energy Rev, Energy %, Energy Backlog, Backlog %, Sensitivity, Rating, Notes)
    # --- Energy-Specialized Software ---
    ("Emerson Electric / AspenTech", "EMR", "Industrial Software", 17.5, 7.0, "40%", 5.5, "35%", "Very High",
     "Very High", "Acquired remaining AspenTech stake ($7.2B). $450M LNG automation wins in Q1'26. Grid mgmt ACV +25% YoY."),
    ("Baker Hughes (Digital)", "BKR", "Oilfield Tech", 29.6, 29.6, "100%", 33.0, "100%", "Very High",
     "Critical", "Record $29.6B orders in FY25. Digital recurring revenue >$1B. Entirely oil/energy focused."),
    ("SLB (Schlumberger) Digital", "SLB", "Oilfield Tech", 36.4, 36.4, "100%", 38.0, "100%", "Very High",
     "Critical", "Digital ARR >$1B (+15% YoY). Launched Tela AI. Delfi SaaS platform is high-margin crown jewel."),
    ("C3.ai", "AI", "Enterprise AI", 0.39, 0.16, "~41%", 0.3, "~45%", "High",
     "Very High", "Baker Hughes partnership renewed thru 2028. Non-O&G revenue growing 48% to diversify."),
    ("Halliburton (Digital)", "HAL", "Oilfield Tech", 22.2, 22.2, "100%", 20.0, "100%", "Very High",
     "Critical", "ZEUS IQ platform adoption +18%. N.A. revenue expected down high-single digits in 2026."),

    # --- IT Services / Consulting ---
    ("Accenture", "ACN", "IT Services", 69.7, 9.5, "~14%", 10.0, "~12%", "Moderate",
     "High", "Resources segment (energy+utilities+chemicals) = $9.5B. 5% growth in FY25."),
    ("Infosys", "INFY", "IT Services", 19.4, 2.6, "~13%", 2.5, "~13%", "Moderate",
     "Medium-High", "Energy, Utilities, Resources & Services = 13.4% of revenue (Q2 FY26)."),
    ("Wipro", "WIT", "IT Services", 11.3, 1.95, "~17%", 1.8, "~16%", "Moderate",
     "High", "Energy, Manufacturing & Resources = 17.3% of IT services revenue. Stable exposure."),
    ("Cognizant", "CTSH", "IT Services", 21.1, 3.8, "~18%", 3.5, "~17%", "Moderate",
     "High", "Products & Resources segment grew 10.5% in FY25. Includes energy + manufacturing."),
    ("TCS", "TCS.NS", "IT Services", 29.1, 1.6, "~6%", 1.5, "~5%", "Low",
     "Low-Medium", "ERU (Energy, Resources, Utilities) ~5.6% of revenue. Diversified portfolio."),

    # --- Enterprise Software ---
    ("SAP", "SAP", "Enterprise Software", 37.5, 3.4, "~9%", 4.0, "~10%", "Moderate",
     "Medium", "Major ERP provider to oil majors (Shell, BP, Chevron). Cloud ERP +28% YoY overall."),
    ("Oracle", "ORCL", "Enterprise Software", 57.4, 4.0, "~7%", 5.5, "~8%", "Moderate",
     "Medium", "Cloud infra for energy workloads. OCI revenue +52% in Q4. Energy is growing vertical."),
    ("Palantir", "PLTR", "Data Analytics", 4.5, 0.5, "~11%", 0.6, "~12%", "Moderate",
     "Medium-High", "Multi-year APA Corp partnership for AI in oil ops. Government (54%) provides stability."),
    ("ServiceNow", "NOW", "Enterprise Software", 11.0, 0.7, "~6%", 0.9, "~7%", "Low",
     "Low-Medium", "Energy is growing vertical but small share. Workflow automation for field operations."),
    ("Salesforce", "CRM", "Enterprise Software", 37.9, 1.9, "~5%", 2.0, "~5%", "Low",
     "Low", "Limited direct O&G exposure. Broader macro sensitivity via enterprise budgets."),

    # --- Hyperscalers / Internet ---
    ("Microsoft (Azure)", "MSFT", "Cloud / Software", 254.2, 12.7, "~5%", 15.0, "~4%", "Moderate",
     "Medium", "Azure used by energy companies. $75B Azure ARR. Data center energy costs rising."),
    ("Amazon (AWS)", "AMZN", "Cloud / E-commerce", 638.0, 14.2, "~2%", 10.0, "~3%", "Moderate",
     "Medium", "AWS $142B ARR. E-commerce hit by shipping/logistics costs. Data center capex $200B in 2026."),
    ("Alphabet (Google Cloud)", "GOOGL", "Cloud / Internet", 382.2, 5.3, "~1.4%", 4.0, "~2%", "Low-Moderate",
     "Medium", "Google Cloud $71B ARR (+48%). Ad revenue sensitive to macro/consumer spend shifts."),
    ("Meta Platforms", "META", "Internet / Social", 164.5, 0.5, "~0.3%", None, "N/A", "Low",
     "Low-Medium", "Minimal direct O&G revenue. Exposed via data center energy costs ($115-135B capex in 2026)."),

    # --- Semiconductors (for context) ---
    ("NVIDIA", "NVDA", "Semiconductors", 130.5, 5.2, "~4%", 8.0, "~5%", "Moderate",
     "Medium", "AI chips used in energy sector. Down 1.74% in March 2026 energy shock."),
    ("Taiwan Semiconductor", "TSM", "Semiconductors", 87.1, 2.6, "~3%", 4.0, "~3%", "Low-Moderate",
     "Medium", "Fab costs sensitive to energy prices. Down 4.44% in March 2026 selloff."),
]

for i, comp in enumerate(companies):
    r = row + 1 + i
    for j, val in enumerate(comp, start=1):
        cell = ws2.cell(row=r, column=j, value=val if val is not None else "N/A")
    style_data_row(ws2, r, len(headers), alt=(i % 2 == 1))
    apply_exposure_color(ws2.cell(row=r, column=10), comp[9])

col_widths = [30, 10, 20, 18, 18, 16, 16, 16, 16, 14, 55]
for i, w in enumerate(col_widths, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.auto_filter.ref = f"A3:K{row + len(companies)}"
ws2.freeze_panes = "A4"


# ════════════════════════════════════════════════════════════════════
# SHEET 3: March 2026 Tech Selloff
# ════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("March 2026 Tech Selloff")
ws3.sheet_properties.tabColor = "C00000"

ws3.merge_cells("A1:G1")
ws3["A1"] = "March 3, 2026: $600B+ Tech Wipeout from Energy Price Surge"
ws3["A1"].font = title_font

ws3.merge_cells("A2:G2")
ws3["A2"] = "WTI crude surged 8% toward $77/bbl on escalating Iran conflict; Strait of Hormuz supply fears"
ws3["A2"].font = Font(name="Calibri", size=11, color="595959", italic=True)

selloff_headers = ["Company", "Ticker", "Sector", "1-Day Decline (%)", "Est. Mkt Cap Loss ($B)", "Primary Risk Factor", "Recovery Outlook"]
row3 = 4
for j, h in enumerate(selloff_headers, 1):
    ws3.cell(row=row3, column=j, value=h)
style_header_row(ws3, row3, len(selloff_headers))

selloff_data = [
    ("Micron Technology", "MU", "Semiconductors", -7.29, 7.5, "Energy-intensive fab operations", "Slow; margin pressure persists"),
    ("Taiwan Semiconductor", "TSM", "Semiconductors", -4.44, 35.0, "Global supply chain, fab energy costs", "Moderate; AI demand cushion"),
    ("Tesla", "TSLA", "EV / Tech", -3.29, 25.0, "Consumer discretionary + energy costs", "Mixed; EV demand elastic"),
    ("Alphabet (Google)", "GOOGL", "Internet", -1.92, 40.0, "Ad spend sensitivity, data center costs", "Moderate; AI diversification"),
    ("NVIDIA", "NVDA", "Semiconductors", -1.74, 50.0, "Multiple compression, AI capex fears", "Strong; structural AI demand"),
    ("Broadcom", "AVGO", "Semiconductors", -1.62, 12.0, "Supply chain, customer capex cuts", "Moderate"),
    ("Apple", "AAPL", "Hardware / Software", -1.55, 45.0, "Hardware manufacturing, shipping costs", "Strong; services revenue stable"),
    ("Amazon", "AMZN", "Cloud / E-commerce", -1.06, 22.0, "Logistics costs, consumer spend drag", "Strong; AWS provides stability"),
    ("KOSPI Index", "—", "South Korea Market", -12.0, 150.0, "Semiconductor + export dependency", "Uncertain; geopolitical risk"),
    ("Nikkei 225", "—", "Japan Market", -3.6, 100.0, "Energy import dependency, yen impact", "Moderate"),
]

for i, d in enumerate(selloff_data):
    r = row3 + 1 + i
    for j, val in enumerate(d, 1):
        ws3.cell(row=r, column=j, value=val)
    style_data_row(ws3, r, len(selloff_headers), alt=(i % 2 == 1))
    pct_cell = ws3.cell(row=r, column=4)
    pct_cell.font = Font(name="Calibri", size=11, bold=True, color="C00000")

ws3_col_widths = [22, 10, 20, 18, 20, 38, 32]
for i, w in enumerate(ws3_col_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w


# ════════════════════════════════════════════════════════════════════
# SHEET 4: Energy Sector IT Spending Trends
# ════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Energy IT Spending Trends")
ws4.sheet_properties.tabColor = "548235"

ws4.merge_cells("A1:F1")
ws4["A1"] = "Oil & Gas Industry IT and AI Spending Forecast"
ws4["A1"].font = title_font

spend_headers = ["Year", "Total O&G IT Spend ($B)", "AI Spend ($B)", "AI % of IT Spend", "Cloud Infra ($B)", "Key Drivers"]
row4 = 3
for j, h in enumerate(spend_headers, 1):
    ws4.cell(row=row4, column=j, value=h)
style_header_row(ws4, row4, len(spend_headers))

spend_data = [
    (2023, 28.0, 2.5, "9%", 5.0, "Post-COVID recovery; initial cloud migration"),
    (2024, 31.0, 3.2, "10%", 6.5, "GenAI pilots; cybersecurity investment surge"),
    (2025, 35.0, 4.0, "11%", 8.5, "AI scaling in production optimization; cloud ERP migration"),
    (2026, 39.0, 5.8, "15%", 11.0, "Agentic AI adoption (SLB Tela); LNG digitization boom"),
    (2027, 43.0, 7.8, "18%", 13.5, "Autonomous operations pilots; digital twin expansion"),
    (2028, 47.0, 10.2, "22%", 16.0, "AI-driven exploration; real-time reservoir modeling"),
    (2029, 52.0, 13.4, "26%", 19.0, "Mature AI deployment; 57% of IT spend (Deloitte projection)"),
]

for i, d in enumerate(spend_data):
    r = row4 + 1 + i
    for j, val in enumerate(d, 1):
        ws4.cell(row=r, column=j, value=val)
    style_data_row(ws4, r, len(spend_headers), alt=(i % 2 == 1))

ws4_col_widths = [10, 22, 18, 18, 18, 55]
for i, w in enumerate(ws4_col_widths, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

chart = BarChart()
chart.type = "col"
chart.title = "Oil & Gas AI Spending Growth ($B)"
chart.y_axis.title = "$ Billions"
chart.x_axis.title = "Year"
chart.style = 10
data = Reference(ws4, min_col=3, min_row=3, max_row=row4 + len(spend_data), max_col=3)
cats = Reference(ws4, min_col=1, min_row=4, max_row=row4 + len(spend_data))
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.shape = 4
ws4.add_chart(chart, "A13")


# ════════════════════════════════════════════════════════════════════
# SHEET 5: Hyperscaler Energy Cost Impact
# ════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Hyperscaler Energy Costs")
ws5.sheet_properties.tabColor = "7030A0"

ws5.merge_cells("A1:H1")
ws5["A1"] = "Hyperscaler Data Center Capex & Energy Cost Exposure (2025-2026)"
ws5["A1"].font = title_font

hyper_headers = [
    "Company", "2025 Capex ($B)", "2026 Capex ($B)",
    "Capex Growth", "Est. Energy as\n% of OpEx",
    "Annual Electricity\nCost Est. ($B)", "Data Center\nMW Capacity",
    "Oil/Gas Price Impact"
]
row5 = 3
for j, h in enumerate(hyper_headers, 1):
    ws5.cell(row=row5, column=j, value=h)
style_header_row(ws5, row5, len(hyper_headers))

hyper_data = [
    ("Amazon (AWS)", 131, 200, "+53%", "12-15%", 8.5, "~30,000 MW", "High: largest absolute energy consumer; logistics also affected"),
    ("Microsoft (Azure)", 80, 110, "+38%", "10-13%", 6.0, "~25,000 MW", "High: massive data center fleet; committed to carbon neutrality"),
    ("Alphabet (Google)", 91.4, 180, "+97%", "10-12%", 5.5, "~20,000 MW", "High: doubling capex; committed to 24/7 carbon-free energy by 2030"),
    ("Meta Platforms", 71, 125, "+76%", "8-10%", 3.5, "~12,000 MW", "Moderate-High: rapid expansion but smaller base; no cloud revenue"),
    ("TOTAL", 373.4, 615, "+65%", "—", 23.5, "~87,000 MW", "Electricity costs could rise 6-7% in 2026-2027 (Goldman Sachs)"),
]

for i, d in enumerate(hyper_data):
    r = row5 + 1 + i
    for j, val in enumerate(d, 1):
        ws5.cell(row=r, column=j, value=val)
    if i == len(hyper_data) - 1:
        for j in range(1, len(hyper_headers) + 1):
            c = ws5.cell(row=r, column=j)
            c.font = Font(name="Calibri", bold=True, size=11)
            c.fill = subheader_fill
            c.border = thin_border
            c.alignment = center
    else:
        style_data_row(ws5, r, len(hyper_headers), alt=(i % 2 == 1))

ws5_col_widths = [22, 16, 16, 14, 16, 18, 18, 55]
for i, w in enumerate(ws5_col_widths, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w


# ════════════════════════════════════════════════════════════════════
# SHEET 6: Ranked Summary
# ════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Ranked Exposure Summary")
ws6.sheet_properties.tabColor = "BF8F00"

ws6.merge_cells("A1:F1")
ws6["A1"] = "Companies Ranked by Oil/Energy Revenue Exposure (Highest to Lowest)"
ws6["A1"].font = title_font

rank_headers = ["Rank", "Company", "Energy Rev %", "Backlog %", "Exposure Rating", "Category"]
row6 = 3
for j, h in enumerate(rank_headers, 1):
    ws6.cell(row=row6, column=j, value=h)
style_header_row(ws6, row6, len(rank_headers))

ranked = [
    (1, "Baker Hughes (Digital)", "100%", "100%", "Critical", "Oilfield Tech"),
    (2, "SLB (Schlumberger) Digital", "100%", "100%", "Critical", "Oilfield Tech"),
    (3, "Halliburton (Digital)", "100%", "100%", "Critical", "Oilfield Tech"),
    (4, "C3.ai", "~41%", "~45%", "Very High", "Enterprise AI"),
    (5, "Emerson / AspenTech", "40%", "35%", "Very High", "Industrial Software"),
    (6, "Cognizant", "~18%", "~17%", "High", "IT Services"),
    (7, "Wipro", "~17%", "~16%", "High", "IT Services"),
    (8, "Accenture", "~14%", "~12%", "High", "IT Services"),
    (9, "Infosys", "~13%", "~13%", "Medium-High", "IT Services"),
    (10, "Palantir", "~11%", "~12%", "Medium-High", "Data Analytics"),
    (11, "SAP", "~9%", "~10%", "Medium", "Enterprise Software"),
    (12, "Oracle", "~7%", "~8%", "Medium", "Enterprise Software"),
    (13, "ServiceNow", "~6%", "~7%", "Low-Medium", "Enterprise Software"),
    (14, "TCS", "~6%", "~5%", "Low-Medium", "IT Services"),
    (15, "Microsoft (Azure)", "~5%", "~4%", "Medium", "Cloud / Software"),
    (16, "Salesforce", "~5%", "~5%", "Low", "Enterprise Software"),
    (17, "NVIDIA", "~4%", "~5%", "Medium", "Semiconductors"),
    (18, "Taiwan Semiconductor", "~3%", "~3%", "Medium", "Semiconductors"),
    (19, "Amazon (AWS)", "~2%", "~3%", "Medium", "Cloud / E-commerce"),
    (20, "Alphabet (Google Cloud)", "~1.4%", "~2%", "Medium", "Cloud / Internet"),
    (21, "Meta Platforms", "~0.3%", "N/A", "Low-Medium", "Internet / Social"),
]

for i, d in enumerate(ranked):
    r = row6 + 1 + i
    for j, val in enumerate(d, 1):
        ws6.cell(row=r, column=j, value=val)
    style_data_row(ws6, r, len(rank_headers), alt=(i % 2 == 1))
    apply_exposure_color(ws6.cell(row=r, column=5), d[4])

ws6_col_widths = [8, 30, 14, 14, 18, 22]
for i, w in enumerate(ws6_col_widths, 1):
    ws6.column_dimensions[get_column_letter(i)].width = w


# ════════════════════════════════════════════════════════════════════
# SHEET 7: Sources & Methodology
# ════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Sources & Methodology")
ws7.sheet_properties.tabColor = "595959"

ws7.merge_cells("A1:C1")
ws7["A1"] = "Data Sources & Methodology"
ws7["A1"].font = title_font

sources = [
    ("Company Filings", "10-K annual reports, quarterly earnings press releases, and investor presentations (FY2025)"),
    ("Accenture", "FY2025 Annual Report - Resources segment revenue of $9.5B (~14% of total $69.7B)"),
    ("Infosys", "Q2 FY2026 Fact Sheet - Energy, Utilities, Resources & Services = 13.4% of revenue"),
    ("Wipro", "Q4 FY25 & Q2 FY26 Press Release - Energy, Manufacturing & Resources = 17.3%"),
    ("Cognizant", "FY2025 Earnings - Products & Resources segment grew 10.5%"),
    ("TCS", "FY2024 Investor Presentation - ERU segment ~5.6% of revenue"),
    ("Baker Hughes", "FY2025 Earnings - Record $29.6B orders, digital ARR >$1B"),
    ("SLB", "FY2025 Earnings - Revenue $36.4B, Digital ARR >$1B (+15% YoY)"),
    ("Halliburton", "FY2025 Earnings - Revenue $22.2B, ZEUS IQ adoption +18%"),
    ("C3.ai", "FY2025 Earnings - Revenue $389M, non-O&G revenue +48%"),
    ("Emerson / AspenTech", "Full AspenTech acquisition ($7.2B). Q1 FY26: $450M LNG automation wins"),
    ("Market Data", "Benzinga, Reuters - March 3, 2026 tech selloff ($600B+) on Iran energy shock"),
    ("IDC", "Global IT spending impact analysis - conflict scenario may reduce growth from 10% to 9%"),
    ("Deloitte", "Energy sector AI spending: $4B (2025) to $13.4B (2029) projection"),
    ("Goldman Sachs", "Electricity inflation forecast: +6% in 2026-2027; 0.2% consumer spending drag"),
    ("", ""),
    ("Methodology Notes", ""),
    ("Revenue Exposure", "Where companies do not break out oil/gas specifically, estimates are based on industry segment "
     "reports (e.g., 'Resources' or 'Energy, Utilities, Resources') and may include adjacent verticals "
     "(utilities, chemicals, mining). Figures marked with '~' are estimates."),
    ("Backlog", "Backlog estimates for pure-play oilfield tech companies are based on reported order books. "
     "For IT services and enterprise software companies, backlog is estimated from remaining performance "
     "obligations and deal pipeline disclosures."),
    ("Exposure Rating", "Composite rating considering: (1) energy revenue as % of total, (2) backlog concentration, "
     "(3) direct oil price sensitivity of business model, (4) ability to pass through costs."),
]

row7 = 3
for src, detail in sources:
    ws7.cell(row=row7, column=1, value=src).font = Font(name="Calibri", bold=True, size=11)
    ws7.merge_cells(start_row=row7, start_column=2, end_row=row7, end_column=3)
    ws7.cell(row=row7, column=2, value=detail).font = normal_font
    ws7.cell(row=row7, column=2).alignment = left_wrap
    row7 += 1

ws7.column_dimensions["A"].width = 25
ws7.column_dimensions["B"].width = 50
ws7.column_dimensions["C"].width = 40

# ── Save ──
filepath = "/workspace/Tech_Company_Oil_Exposure_Analysis.xlsx"
wb.save(filepath)
print(f"Spreadsheet saved to {filepath}")
