"""
Build an Excel workbook: Parametric vs Memory Scaling for LLMs
================================================================
Covers three memory regimes that drive LLM inference/serving economics:

  1. Parametric memory  -> weights (grows with parameter count / precision)
  2. KV cache memory    -> grows with context length x batch x model depth/width
  3. Decode economics   -> memory-bandwidth-bound token generation

The workbook is built with *live* Excel formulas that reference an
Assumptions tab, so an analyst can flip precision, batch size, GPU, or
overhead assumptions and watch every downstream number recompute.

Run:  python build_llm_memory_scaling.py
Out:  LLM_Parametric_vs_Memory_Scaling.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.axis import ChartLines

OUT = "LLM_Parametric_vs_Memory_Scaling.xlsx"

# ----------------------------------------------------------------------------
# Palette / styles
# ----------------------------------------------------------------------------
NAVY = "1F3864"
BLUE = "2E5496"
LTBLUE = "D6E0F0"
GREY = "F2F2F2"
ACCENT = "C55A11"
GREEN = "375623"
LTGREEN = "E2EFDA"
WHITE = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_title(ws, cell, text, size=15):
    ws[cell] = text
    ws[cell].font = Font(bold=True, size=size, color=NAVY)

def style_subtitle(ws, cell, text, size=10):
    ws[cell] = text
    ws[cell].font = Font(italic=True, size=size, color="595959")

def hdr(cell, fill=BLUE, color=WHITE, wrap=True):
    cell.font = Font(bold=True, color=color, size=10)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = border

def cellfmt(cell, num_fmt=None, bold=False, fill=None, align="right", color=None):
    if num_fmt:
        cell.number_format = num_fmt
    cell.font = Font(bold=bold, color=color or "000000", size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = border
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)

INPUT_FILL = "FFF2CC"   # yellow = editable input
CALC_FILL = None

NUM0 = "#,##0"
NUM1 = "#,##0.0"
NUM2 = "#,##0.00"
GB = '#,##0.0" GB"'
GB2 = '#,##0.00" GB"'
TB = '#,##0.0" TB"'
KB = '#,##0.0" KB"'
PCT = "0.0%"
TPS = '#,##0" tok/s"'
XNUM = '#,##0.0"x"'

wb = Workbook()

# ============================================================================
# 1) ASSUMPTIONS
# ============================================================================
a = wb.active
a.title = "Assumptions"
a.sheet_view.showGridLines = False
style_title(a, "B2", "LLM Parametric vs Memory Scaling — Assumptions")
style_subtitle(a, "B3", "Yellow cells are editable inputs. Every other tab references these values via live formulas.")

# --- Precision / bytes-per-element -----------------------------------------
a["B5"] = "Numeric precision (bytes per element)"
a["B5"].font = Font(bold=True, size=11, color=NAVY)
prec_hdr = ["Precision", "Bytes/elem", "Notes"]
for j, h in enumerate(prec_hdr):
    c = a.cell(row=6, column=2 + j, value=h)
    hdr(c, fill=NAVY)
prec_rows = [
    ("FP32", 4, "Full precision (training / reference)"),
    ("FP16 / BF16", 2, "Standard inference precision"),
    ("FP8", 1, "Hopper/Blackwell native; frontier inference"),
    ("INT4 / NF4", 0.5, "Aggressive weight-only quantization"),
]
prec_start = 7
for i, (p, b, note) in enumerate(prec_rows):
    r = prec_start + i
    cellfmt(a.cell(row=r, column=2, value=p), align="left")
    cellfmt(a.cell(row=r, column=3, value=b), NUM2, fill=INPUT_FILL)
    cellfmt(a.cell(row=r, column=4, value=note), align="left")
# named references
BYTES_FP32 = "Assumptions!$C$7"
BYTES_FP16 = "Assumptions!$C$8"
BYTES_FP8 = "Assumptions!$C$9"
BYTES_INT4 = "Assumptions!$C$10"

# --- Serving inputs ---------------------------------------------------------
a["B13"] = "Default serving assumptions"
a["B13"].font = Font(bold=True, size=11, color=NAVY)
serve = [
    ("Weight precision bytes/elem", 2, "Default = FP16/BF16", "D14"),
    ("KV cache precision bytes/elem", 2, "Default = FP16 KV", "D15"),
    ("Batch size (concurrent seqs)", 1, "Concurrent sequences served", "D16"),
    ("Weights overhead factor", 1.15, "CUDA ctx, fragmentation, buffers", "D17"),
    ("Activation / runtime overhead", 1.10, "Applied on top of weights+KV", "D18"),
    ("Selected GPU (row # below)", 2, "Pick GPU for Decode tab (1-5)", "D19"),
]
for i, (label, val, note, ref) in enumerate(serve):
    r = 14 + i
    cellfmt(a.cell(row=r, column=2, value=label), align="left")
    c = a.cell(row=r, column=4, value=val)
    cellfmt(c, NUM2, fill=INPUT_FILL)
    cellfmt(a.cell(row=r, column=5, value=note), align="left")
KV_BYTES = "Assumptions!$D$15"
BATCH = "Assumptions!$D$16"
W_OH = "Assumptions!$D$17"
ACT_OH = "Assumptions!$D$18"
W_BYTES = "Assumptions!$D$14"
GPU_SEL = "Assumptions!$D$19"

# --- Accelerator specs ------------------------------------------------------
a["B22"] = "Accelerator specifications"
a["B22"].font = Font(bold=True, size=11, color=NAVY)
gpu_hdr = ["#", "Accelerator", "HBM (GB)", "Mem BW (TB/s)", "Notes"]
for j, h in enumerate(gpu_hdr):
    hdr(a.cell(row=23, column=2 + j, value=h), fill=NAVY)
gpus = [
    (1, "NVIDIA A100 80GB", 80, 2.04, "Ampere, HBM2e"),
    (2, "NVIDIA H100 SXM", 80, 3.35, "Hopper, HBM3"),
    (3, "NVIDIA H200", 141, 4.80, "Hopper, HBM3e"),
    (4, "NVIDIA B200", 192, 8.00, "Blackwell, HBM3e"),
    (5, "AMD MI300X", 192, 5.30, "CDNA3, HBM3"),
]
gpu_start = 24
for i, (n, name, mem, bw, note) in enumerate(gpus):
    r = gpu_start + i
    cellfmt(a.cell(row=r, column=2, value=n), NUM0, align="center")
    cellfmt(a.cell(row=r, column=3, value=name), align="left")
    cellfmt(a.cell(row=r, column=4, value=mem), NUM0, fill=INPUT_FILL)
    cellfmt(a.cell(row=r, column=5, value=bw), NUM2, fill=INPUT_FILL)
    cellfmt(a.cell(row=r, column=6, value=note), align="left")
GPU_LAST = gpu_start + len(gpus) - 1  # =28
# Selected GPU pull-through (INDEX on selection)
a["B31"] = "Selected GPU (from row # above)"
a["B31"].font = Font(bold=True, size=11, color=ACCENT)
cellfmt(a.cell(row=32, column=2, value="Name"), align="left")
c = a.cell(row=32, column=4,
           value=f"=INDEX($C${gpu_start}:$C${GPU_LAST},{GPU_SEL})")
cellfmt(c, align="left", bold=True, fill=LTGREEN)
cellfmt(a.cell(row=33, column=2, value="HBM (GB)"), align="left")
cellfmt(a.cell(row=33, column=4,
               value=f"=INDEX($D${gpu_start}:$D${GPU_LAST},{GPU_SEL})"),
        NUM0, bold=True, fill=LTGREEN)
cellfmt(a.cell(row=34, column=2, value="Mem BW (TB/s)"), align="left")
cellfmt(a.cell(row=34, column=4,
               value=f"=INDEX($E${gpu_start}:$E${GPU_LAST},{GPU_SEL})"),
        NUM2, bold=True, fill=LTGREEN)
SEL_HBM = "Assumptions!$D$33"
SEL_BW = "Assumptions!$D$34"

a.column_dimensions["A"].width = 2
a.column_dimensions["B"].width = 30
a.column_dimensions["C"].width = 20
a.column_dimensions["D"].width = 14
a.column_dimensions["E"].width = 16
a.column_dimensions["F"].width = 34

# ============================================================================
# 2) MODELS (architecture reference)
# ============================================================================
m = wb.create_sheet("Models")
m.sheet_view.showGridLines = False
style_title(m, "B2", "Model Architectures")
style_subtitle(m, "B3", "Public / reported configurations. KV cache scales with (layers x kv_heads x head_dim); weights scale with total parameters.")

model_hdr = ["Model", "Params (B)", "Layers", "Hidden", "Attn heads",
             "KV heads", "Head dim", "Vocab", "Max ctx (K)", "Attn type"]
hrow = 5
for j, h in enumerate(model_hdr):
    hdr(m.cell(row=hrow, column=2 + j, value=h), fill=BLUE)

# name, params(B), layers, hidden, attn_heads, kv_heads, head_dim, vocab, maxctx_k, attn
models = [
    ("Llama-3.2-1B",      1.24,   16, 2048,  32,  8, 64,  128256, 128, "GQA"),
    ("Llama-3.2-3B",      3.21,   28, 3072,  24,  8, 128, 128256, 128, "GQA"),
    ("Llama-3.1-8B",      8.03,   32, 4096,  32,  8, 128, 128256, 128, "GQA"),
    ("Llama-2-7B",        6.74,   32, 4096,  32, 32, 128, 32000,    4, "MHA"),
    ("Mistral-7B v0.3",   7.25,   32, 4096,  32,  8, 128, 32768,   32, "GQA"),
    ("Qwen2.5-7B",        7.62,   28, 3584,  28,  4, 128, 152064, 128, "GQA"),
    ("Gemma-2-9B",        9.24,   42, 3584,  16,  8, 256, 256128,    8, "GQA"),
    ("Gemma-2-27B",      27.2,    46, 4608,  32, 16, 128, 256128,    8, "GQA"),
    ("Llama-2-70B",      68.9,    80, 8192,  64,  8, 128, 32000,     4, "GQA"),
    ("Qwen2.5-72B",      72.7,    80, 8192,  64,  8, 128, 152064, 128, "GQA"),
    ("Llama-3.1-70B",    70.6,    80, 8192,  64,  8, 128, 128256, 128, "GQA"),
    ("GPT-3 175B",      175.0,    96, 12288, 96, 96, 128, 50257,     2, "MHA"),
    ("Llama-3.1-405B",  405.0,   126, 16384,128,  8, 128, 128256, 128, "GQA"),
]
mstart = hrow + 1
for i, row in enumerate(models):
    r = mstart + i
    name, params, layers, hidden, ah, kvh, hd, vocab, ctx, attn = row
    fill = GREY if i % 2 else WHITE
    cellfmt(m.cell(row=r, column=2, value=name), align="left", fill=fill, bold=True)
    cellfmt(m.cell(row=r, column=3, value=params), NUM1, fill=fill)
    cellfmt(m.cell(row=r, column=4, value=layers), NUM0, fill=fill)
    cellfmt(m.cell(row=r, column=5, value=hidden), NUM0, fill=fill)
    cellfmt(m.cell(row=r, column=6, value=ah), NUM0, fill=fill)
    cellfmt(m.cell(row=r, column=7, value=kvh), NUM0, fill=fill)
    cellfmt(m.cell(row=r, column=8, value=hd), NUM0, fill=fill)
    cellfmt(m.cell(row=r, column=9, value=vocab), NUM0, fill=fill)
    cellfmt(m.cell(row=r, column=10, value=ctx), NUM0, fill=fill)
    cellfmt(m.cell(row=r, column=11, value=attn), align="center", fill=fill)
mend = mstart + len(models) - 1
NMODELS = len(models)

widths = {"A": 2, "B": 18, "C": 11, "D": 8, "E": 9, "F": 11,
          "G": 10, "H": 10, "I": 10, "J": 11, "K": 10}
for col, w in widths.items():
    m.column_dimensions[col].width = w

# convenience refs into Models sheet
def M(col, r):  # absolute-column ref within Models
    return f"Models!${col}${r}"

# ============================================================================
# 3) WEIGHTS MEMORY (parametric memory)
# ============================================================================
w = wb.create_sheet("Weights Memory")
w.sheet_view.showGridLines = False
style_title(w, "B2", "Parametric (Weights) Memory")
style_subtitle(w, "B3", "Weights memory = params x bytes/elem x overhead. Overhead applied per Assumptions. GB = 1e9 bytes.")

whdr = ["Model", "Params (B)", "FP16 (GB)", "FP8 (GB)", "INT4 (GB)",
        "FP16 + overhead", "# H200 (141GB)", "# B200 (192GB)"]
whrow = 5
for j, h in enumerate(whdr):
    hdr(w.cell(row=whrow, column=2 + j, value=h), fill=BLUE)

wstart = whrow + 1
for i in range(NMODELS):
    r = wstart + i
    mr = mstart + i
    fill = GREY if i % 2 else WHITE
    cellfmt(w.cell(row=r, column=2, value=f"={M('B', mr)}"), align="left", fill=fill, bold=True)
    cellfmt(w.cell(row=r, column=3, value=f"={M('C', mr)}"), NUM1, fill=fill)
    # params (B) * 1e9 * bytes / 1e9 GB  == params * bytes
    cellfmt(w.cell(row=r, column=4, value=f"=C{r}*{BYTES_FP16}"), GB, fill=fill)
    cellfmt(w.cell(row=r, column=5, value=f"=C{r}*{BYTES_FP8}"), GB, fill=fill)
    cellfmt(w.cell(row=r, column=6, value=f"=C{r}*{BYTES_INT4}"), GB, fill=fill)
    cellfmt(w.cell(row=r, column=7, value=f"=D{r}*{W_OH}"), GB, fill=fill, bold=True)
    cellfmt(w.cell(row=r, column=8, value=f"=CEILING(G{r}/141,1)"), NUM0, fill=fill)
    cellfmt(w.cell(row=r, column=9, value=f"=CEILING(G{r}/192,1)"), NUM0, fill=fill)
wend = wstart + NMODELS - 1

for col, wd in {"A": 2, "B": 18, "C": 11, "D": 12, "E": 12, "F": 12,
                "G": 16, "H": 15, "I": 15}.items():
    w.column_dimensions[col].width = wd

# ============================================================================
# 4) KV CACHE SCALING (memory vs context length)
# ============================================================================
kv = wb.create_sheet("KV Cache Scaling")
kv.sheet_view.showGridLines = False
style_title(kv, "B2", "KV Cache Memory Scaling")
style_subtitle(kv, "B3", "KV bytes = 2 x layers x kv_heads x head_dim x seq_len x batch x kv_bytes. Grows LINEARLY with context & batch.")
style_subtitle(kv, "B4", "Per-token column is batch-independent; matrix below uses batch size from Assumptions.")

# per-token KV size then a matrix across context lengths
ctx_lengths = [1024, 2048, 4096, 8192, 16384, 32768, 65536, 131072]
kvhdr = ["Model", "KV / token (KB)"] + [f"{c//1024}K ctx (GB)" for c in ctx_lengths]
khrow = 6
for j, h in enumerate(kvhdr):
    hdr(kv.cell(row=khrow, column=2 + j, value=h), fill=GREEN)

kstart = khrow + 1
for i in range(NMODELS):
    r = kstart + i
    mr = mstart + i
    fill = LTGREEN if i % 2 else WHITE
    cellfmt(kv.cell(row=r, column=2, value=f"={M('B', mr)}"), align="left", fill=fill, bold=True)
    # KV per token (bytes) = 2 * layers(D) * kv_heads(G) * head_dim(H) * kv_bytes ; /1024 -> KB
    per_tok = (f"=2*{M('D', mr)}*{M('G', mr)}*{M('H', mr)}*{KV_BYTES}/1024")
    cellfmt(kv.cell(row=r, column=3, value=per_tok), NUM1, fill=fill)
    for j, cl in enumerate(ctx_lengths):
        col = 4 + j
        # GB = KVperToken_KB * 1024 (bytes) * seq * batch / 1e9
        f = f"=$C{r}*1024*{cl}*{BATCH}/1000000000"
        cellfmt(kv.cell(row=r, column=col, value=f), GB2, fill=fill)
kend = kstart + NMODELS - 1

kv.column_dimensions["A"].width = 2
kv.column_dimensions["B"].width = 18
kv.column_dimensions["C"].width = 15
for j in range(len(ctx_lengths)):
    kv.column_dimensions[get_column_letter(4 + j)].width = 13

# note row
note_r = kend + 2
kv.cell(row=note_r, column=2,
        value="Note: MHA models (Llama-2-7B, GPT-3) carry far larger KV per token than GQA models of similar size — the KV cache, not weights, dominates long-context memory.")
kv.cell(row=note_r, column=2).font = Font(italic=True, size=9, color="595959")

# ============================================================================
# 5) DECODE ECONOMICS (bandwidth-bound generation)
# ============================================================================
d = wb.create_sheet("Decode Economics")
d.sheet_view.showGridLines = False
style_title(d, "B2", "Decode Economics (Memory-Bandwidth Bound)")
style_subtitle(d, "B3", "Autoregressive decode reads all weights + KV each step. Step time ≈ (weight_bytes + kv_bytes) / mem_BW.")

# selected context length input
d["B5"] = "Context length for decode (tokens):"
d["B5"].font = Font(bold=True, size=10)
cellfmt(d.cell(row=5, column=6, value=8192), NUM0, fill=INPUT_FILL)
CTX_SEL = "'Decode Economics'!$F$5"
d["B6"] = "Selected GPU:"
d["B6"].font = Font(bold=True, size=10)
cellfmt(d.cell(row=6, column=6, value=f"={SEL_HBM}"), NUM0, fill=LTGREEN)  # placeholder trigger
d.cell(row=6, column=6, value=f"={SEL_BW}")  # show BW
d.cell(row=6, column=3, value=f"={SEL_HBM}&\" GB HBM @ \"&{SEL_BW}&\" TB/s\"")
d.cell(row=6, column=3).font = Font(bold=True, color=ACCENT, size=10)
d.cell(row=6, column=6).number_format = TB

dhdr = ["Model", "Weights FP16 (GB)", "KV @ ctx (GB)", "Total mem (GB)",
        "Fits on GPU?", "Decode step (ms)", "Throughput (tok/s)", "Bottleneck"]
dhrow = 8
for j, h in enumerate(dhdr):
    hdr(d.cell(row=dhrow, column=2 + j, value=h), fill=ACCENT)

dstart = dhrow + 1
for i in range(NMODELS):
    r = dstart + i
    mr = mstart + i
    wr = wstart + i
    kr = kstart + i
    fill = GREY if i % 2 else WHITE
    cellfmt(d.cell(row=r, column=2, value=f"={M('B', mr)}"), align="left", fill=fill, bold=True)
    # weights FP16 GB from Weights sheet col D
    cellfmt(d.cell(row=r, column=3, value=f"='Weights Memory'!D{wr}"), GB, fill=fill)
    # KV at selected ctx: per-token KB * 1024 * ctx * batch / 1e9
    cellfmt(d.cell(row=r, column=4,
                   value=f"='KV Cache Scaling'!$C{kr}*1024*{CTX_SEL}*{BATCH}/1000000000"),
            GB2, fill=fill)
    # total incl overhead
    cellfmt(d.cell(row=r, column=5, value=f"=(C{r}+D{r})*{ACT_OH}"), GB, fill=fill, bold=True)
    # fits?
    cellfmt(d.cell(row=r, column=6,
                   value=f'=IF(E{r}<={SEL_HBM},"Yes","No — "&CEILING(E{r}/{SEL_HBM},1)&"x GPU")'),
            align="center", fill=fill)
    # decode step ms = (weights+kv) bytes / BW(bytes/s) * 1000. GB=1e9, TB/s=1e12
    #   time_s = (GB*1e9) / (BW*1e12) = GB/(BW*1000); ms = GB/BW
    cellfmt(d.cell(row=r, column=7, value=f"=(C{r}+D{r})/{SEL_BW}"), NUM2, fill=fill)
    # throughput single stream tok/s = 1000/step_ms
    cellfmt(d.cell(row=r, column=8, value=f"=1000/G{r}"), TPS, fill=fill, bold=True)
    cellfmt(d.cell(row=r, column=9,
                   value=f'=IF(D{r}>C{r},"KV-cache","Weights")'),
            align="center", fill=fill)
dend = dstart + NMODELS - 1

for col, wd in {"A": 2, "B": 18, "C": 16, "D": 14, "E": 14, "F": 16,
                "G": 15, "H": 17, "I": 12}.items():
    d.column_dimensions[col].width = wd

dn = dend + 2
d.cell(row=dn, column=2,
       value="Single-stream, memory-bandwidth-bound estimate (ignores compute, kernel efficiency, and batching amortization of weights). Real throughput is lower; batching raises aggregate tok/s by amortizing weight reads across sequences.")
d.cell(row=dn, column=2).font = Font(italic=True, size=9, color="595959")

# ============================================================================
# 6) PARAMETRIC SCALING (weights memory vs param count, log-log)
# ============================================================================
p = wb.create_sheet("Parametric Scaling")
p.sheet_view.showGridLines = False
style_title(p, "B2", "Parametric Memory Scaling Curve")
style_subtitle(p, "B3", "Weights memory grows linearly with parameter count; the multiplier is set purely by precision (bytes/elem).")

phdr = ["Params (B)", "FP32 (GB)", "FP16 (GB)", "FP8 (GB)", "INT4 (GB)",
        "# H100 (80GB) FP16", "# B200 (192GB) FP16"]
phrow = 5
for j, h in enumerate(phdr):
    hdr(p.cell(row=phrow, column=2 + j, value=h), fill=NAVY)

param_series = [0.5, 1, 3, 7, 13, 30, 70, 130, 175, 405, 671, 1000, 2000]
pstart = phrow + 1
for i, pv in enumerate(param_series):
    r = pstart + i
    fill = GREY if i % 2 else WHITE
    cellfmt(p.cell(row=r, column=2, value=pv), NUM1, fill=fill)
    cellfmt(p.cell(row=r, column=3, value=f"=B{r}*{BYTES_FP32}"), NUM0, fill=fill)
    cellfmt(p.cell(row=r, column=4, value=f"=B{r}*{BYTES_FP16}"), NUM0, fill=fill)
    cellfmt(p.cell(row=r, column=5, value=f"=B{r}*{BYTES_FP8}"), NUM0, fill=fill)
    cellfmt(p.cell(row=r, column=6, value=f"=B{r}*{BYTES_INT4}"), NUM0, fill=fill)
    cellfmt(p.cell(row=r, column=7, value=f"=CEILING(D{r}*{W_OH}/80,1)"), NUM0, fill=fill)
    cellfmt(p.cell(row=r, column=8, value=f"=CEILING(D{r}*{W_OH}/192,1)"), NUM0, fill=fill)
pend = pstart + len(param_series) - 1

for col, wd in {"A": 2, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12,
                "G": 18, "H": 18}.items():
    p.column_dimensions[col].width = wd

# ---- chart: params vs memory (weights) ----
chart = LineChart()
chart.title = "Weights Memory vs Parameter Count"
chart.style = 12
chart.y_axis.title = "Weights memory (GB)"
chart.x_axis.title = "Parameters (B)"
chart.height = 10
chart.width = 20
data = Reference(p, min_col=3, max_col=6, min_row=phrow, max_row=pend)
cats = Reference(p, min_col=2, min_row=pstart, max_row=pend)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.x_axis.delete = False
chart.y_axis.delete = False
p.add_chart(chart, "J5")

# ---- chart: KV cache vs context ----
kchart = LineChart()
kchart.title = "KV Cache vs Context Length (batch=1)"
kchart.style = 13
kchart.y_axis.title = "KV cache (GB)"
kchart.x_axis.title = "Context length"
kchart.height = 10
kchart.width = 20
# transpose: use a few representative models. Build a small helper table on this sheet.
# helper table at columns J.. below chart
hb_row = 26
p.cell(row=hb_row, column=10, value="KV cache by context (GB) — helper for chart")
p.cell(row=hb_row, column=10).font = Font(bold=True, size=10, color=GREEN)
sel_models = [("Llama-3.1-8B", 2), ("Llama-3.1-70B", 10), ("Llama-3.1-405B", 12),
              ("GPT-3 175B (MHA)", 11), ("Llama-2-7B (MHA)", 3)]
# header row: context lengths
p.cell(row=hb_row + 1, column=10, value="Context")
hdr(p.cell(row=hb_row + 1, column=10), fill=GREEN)
for j, model_label in enumerate([sm[0] for sm in sel_models]):
    hdr(p.cell(row=hb_row + 1, column=11 + j, value=model_label), fill=GREEN)
for ii, cl in enumerate(ctx_lengths):
    rr = hb_row + 2 + ii
    cellfmt(p.cell(row=rr, column=10, value=f"{cl//1024}K"), align="center")
    for j, (label, midx) in enumerate(sel_models):
        # midx is 1-based position in models list
        kr = kstart + (midx - 1)
        # reference the KV matrix column for this ctx (col 4+ii)
        srccol = get_column_letter(4 + ii)
        cellfmt(p.cell(row=rr, column=11 + j,
                       value=f"='KV Cache Scaling'!{srccol}{kr}"), NUM2)
hb_end = hb_row + 1 + len(ctx_lengths)
kdata = Reference(p, min_col=11, max_col=11 + len(sel_models) - 1,
                  min_row=hb_row + 1, max_row=hb_end)
kcats = Reference(p, min_col=10, min_row=hb_row + 2, max_row=hb_end)
kchart.add_data(kdata, titles_from_data=True)
kchart.set_categories(kcats)
kchart.x_axis.delete = False
kchart.y_axis.delete = False
p.add_chart(kchart, "J27" if False else "R5")

for j in range(len(sel_models)):
    p.column_dimensions[get_column_letter(11 + j)].width = 16
p.column_dimensions["J"].width = 12

# ============================================================================
# 7) README / methodology
# ============================================================================
rd = wb.create_sheet("README")
rd.sheet_view.showGridLines = False
style_title(rd, "B2", "README — Methodology & Formulas")
lines = [
    ("", ""),
    ("Purpose", "Quantify how LLM memory footprint scales along two axes: parametric (weights) and KV cache, and how both drive decode-time economics."),
    ("", ""),
    ("Three memory regimes", ""),
    ("  1. Parametric (weights)", "Fixed per model. Memory = params x bytes/elem. Scales linearly with param count; precision sets the slope (FP16=2B, FP8=1B, INT4=0.5B)."),
    ("  2. KV cache", "Grows at run time. Bytes = 2 x layers x kv_heads x head_dim x seq_len x batch x kv_bytes. Linear in context length AND batch size."),
    ("  3. Decode", "Autoregressive generation reads weights + KV cache from HBM every token → memory-bandwidth bound. Step time ≈ (weights+KV)/BW."),
    ("", ""),
    ("Key formulas", ""),
    ("  Weights (GB)", "params_B x bytes_per_elem   (since params_B x 1e9 x bytes / 1e9 = params_B x bytes)"),
    ("  KV / token (KB)", "2 x layers x kv_heads x head_dim x kv_bytes / 1024"),
    ("  KV total (GB)", "(KV_per_token_KB x 1024) x seq_len x batch / 1e9"),
    ("  Decode step (ms)", "(weights_GB + KV_GB) / mem_BW_TBps          [GB/1e9 ÷ TB/1e12 x1000 = GB/TBps]"),
    ("  Throughput (tok/s)", "1000 / decode_step_ms   (single stream, BW-bound upper bound)"),
    ("", ""),
    ("Why GQA matters", "Grouped-Query Attention slashes kv_heads (e.g. 64→8), cutting KV cache 8x vs multi-head attention. This is why modern long-context models are viable."),
    ("", ""),
    ("Caveats", "Estimates ignore compute-bound prefill, kernel/MFU inefficiency, PagedAttention fragmentation, and batching amortization. Treat throughput as an optimistic ceiling."),
    ("", ""),
    ("How to use", "Edit yellow cells on Assumptions (precision, batch, overhead, GPU) and F5 on Decode Economics (context). All tabs recompute live."),
    ("", ""),
    ("Tabs", "Assumptions · Models · Weights Memory · KV Cache Scaling · Decode Economics · Parametric Scaling (charts) · README"),
]
r = 4
for label, text in lines:
    if label:
        rd.cell(row=r, column=2, value=label).font = Font(bold=True, size=10, color=NAVY)
    if text:
        c = rd.cell(row=r, column=3, value=text)
        c.font = Font(size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
rd.column_dimensions["A"].width = 2
rd.column_dimensions["B"].width = 22
rd.column_dimensions["C"].width = 110

# order sheets: README first
wb.move_sheet("README", -(len(wb.sheetnames) - 1))

# freeze header panes
for sh, cell in [("Models", "B6"), ("Weights Memory", "B6"),
                 ("KV Cache Scaling", "B7"), ("Decode Economics", "B9"),
                 ("Parametric Scaling", "B6")]:
    wb[sh].freeze_panes = cell

wb.save(OUT)
print(f"Wrote {OUT} with sheets: {wb.sheetnames}")
