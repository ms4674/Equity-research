#!/usr/bin/env python3
"""
Datacenter Model Builder — v2
==============================
Creates a comprehensive Excel-based datacenter model covering:
- Capex Spend ($B)
- Server Counts (thousands)
- Datacenter Counts
- GW Capacity
- Power Generation & Availability
- Revenue/MW Analysis
- Power Supply & Generation Ramp
- Datacenter Bill of Materials
- Capex-to-Revenue Supply Chain Mapping

Covers:
  Hyperscalers: AWS, Azure, Google Cloud, Meta, Oracle, Apple
  Neocloud Vendors: CoreWeave, Lambda, Crusoe, Voltage Park, Together AI, Applied Digital
  Colocation / DC REITs: Equinix, Digital Realty, CyrusOne, QTS Realty, Vantage, Switch

Historical data: 2018-2024 (estimated from public filings, earnings, industry reports)
Forecast: 2025E-2030E
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, LineChart, Reference, BarChart3D
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from copy import copy

# ============================================================
# COLOR PALETTE & STYLE CONSTANTS
# ============================================================
DARK_BLUE = "1B2A4A"
MED_BLUE = "2E5090"
LIGHT_BLUE = "D6E4F0"
ACCENT_BLUE = "4472C4"
ACCENT_ORANGE = "ED7D31"
ACCENT_GREEN = "70AD47"
ACCENT_GOLD = "FFC000"
ACCENT_TEAL = "00B0F0"
ACCENT_PURPLE = "7030A0"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
BORDER_GRAY = "B4B4B4"
FORECAST_BG = "FFF2CC"  # Light yellow for forecast columns
HEADER_FONT_COLOR = WHITE
TEAL_BG = "E2F0D9"

TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=WHITE)
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color=DARK_BLUE)
SUBSECTION_FONT = Font(name="Calibri", size=11, bold=True, color=MED_BLUE)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
DATA_FONT = Font(name="Calibri", size=10)
DATA_FONT_BOLD = Font(name="Calibri", size=10, bold=True)
COMPANY_FONT = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
TOTAL_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
NOTE_FONT = Font(name="Calibri", size=8, italic=True, color="666666")
PCT_FONT = Font(name="Calibri", size=9, italic=True, color="666666")

TITLE_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
HEADER_FILL = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
LIGHT_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
TOTAL_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
FORECAST_FILL = PatternFill(start_color=FORECAST_BG, end_color=FORECAST_BG, fill_type="solid")
TEAL_FILL = PatternFill(start_color=TEAL_BG, end_color=TEAL_BG, fill_type="solid")
SUBTOTAL_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_GRAY),
    right=Side(style="thin", color=BORDER_GRAY),
    top=Side(style="thin", color=BORDER_GRAY),
    bottom=Side(style="thin", color=BORDER_GRAY),
)

# Years
YEARS = list(range(2018, 2031))
HIST_YEARS = list(range(2018, 2025))
FCST_YEARS = list(range(2025, 2031))
YEAR_LABELS = [str(y) if y < 2025 else f"{y}E" for y in YEARS]

# ============================================================
# COMPANY GROUPS
# ============================================================
HYPERSCALERS = ["Amazon (AWS)", "Microsoft (Azure)", "Google (GCP)", "Meta", "Oracle Cloud", "Apple"]
NEOCLOUDS = ["CoreWeave", "Lambda", "Crusoe Energy", "Voltage Park", "Together AI", "Applied Digital"]
COLO_REITS = ["Equinix", "Digital Realty", "CyrusOne", "QTS Realty", "Vantage Data Centers", "Switch"]

ALL_COMPANIES = HYPERSCALERS + NEOCLOUDS + COLO_REITS

SEGMENT_MAP = [
    ("Hyperscalers", HYPERSCALERS),
    ("Neocloud Vendors", NEOCLOUDS),
    ("Colocation / DC REITs", COLO_REITS),
]

# ============================================================
# DATA TABLES  (All estimates; sources: public filings,
# earnings calls, industry analyst reports, press releases)
# Units noted per section
# ============================================================

# --- CAPEX SPEND ($B) ---
CAPEX_DATA = {
    # Historical 2018-2024, Forecast 2025E-2030E
    "Amazon (AWS)":       [13.4, 16.1, 21.0, 27.0, 36.0, 48.4, 75.0, 100.0, 120.0, 138.0, 150.0, 155.0, 158.0],
    "Microsoft (Azure)":  [11.6, 13.9, 15.4, 22.0, 28.0, 32.0, 55.7, 80.0, 95.0, 110.0, 120.0, 125.0, 128.0],
    "Google (GCP)":       [10.1, 12.0, 15.0, 21.0, 26.0, 32.3, 52.5, 75.0, 85.0, 95.0, 100.0, 105.0, 108.0],
    "Meta":               [ 6.7,  8.4, 11.0, 15.0, 18.0, 28.0, 38.0, 60.0, 68.0, 75.0,  78.0,  80.0,  82.0],
    "Oracle Cloud":       [ 1.8,  1.9,  2.4,  3.0,  4.0,  6.9, 13.0, 18.0, 22.0, 25.0,  27.0,  28.0,  29.0],
    "Apple":              [ 4.5,  5.0,  5.5,  6.0,  6.5,  7.0,  9.3, 11.0, 13.0, 15.0,  16.0,  17.0,  18.0],
    "CoreWeave":          [ 0.0,  0.0,  0.0,  0.0,  0.1,  0.8,  3.5,  8.0, 12.0, 15.0,  17.0,  18.0,  19.0],
    "Lambda":             [ 0.0,  0.0,  0.0,  0.0,  0.05, 0.2,  0.8,  2.0,  3.5,  5.0,   6.0,   7.0,   7.5],
    "Crusoe Energy":      [ 0.0,  0.0,  0.0,  0.0,  0.05, 0.3,  1.0,  2.5,  4.0,  5.5,   7.0,   8.0,   8.5],
    "Voltage Park":       [ 0.0,  0.0,  0.0,  0.0,  0.0,  0.1,  0.5,  1.5,  2.5,  3.5,   4.5,   5.0,   5.5],
    "Together AI":        [ 0.0,  0.0,  0.0,  0.0,  0.0,  0.1,  0.3,  0.8,  1.5,  2.5,   3.5,   4.0,   4.5],
    "Applied Digital":    [ 0.0,  0.0,  0.0,  0.0,  0.02, 0.15, 0.6,  1.5,  2.5,  3.5,   4.5,   5.0,   5.5],
    # Colo / REITs
    "Equinix":            [ 2.0,  2.3,  2.7,  2.9,  3.1,  3.4,  3.8,  4.5,  5.2,  6.0,  6.8,  7.5,  8.2],
    "Digital Realty":      [ 1.8,  2.0,  2.2,  2.5,  2.8,  3.2,  3.8,  4.8,  5.8,  7.0,  8.0,  9.0, 10.0],
    "CyrusOne":           [ 0.6,  0.8,  0.9,  1.0,  1.1,  1.3,  1.6,  2.0,  2.5,  3.0,  3.5,  4.0,  4.5],
    "QTS Realty":          [ 0.5,  0.6,  0.7,  0.9,  1.2,  1.8,  2.5,  3.5,  4.5,  5.5,  6.5,  7.5,  8.5],
    "Vantage Data Centers":[ 0.3,  0.4,  0.5,  0.7,  1.0,  1.5,  2.2,  3.2,  4.2,  5.2,  6.2,  7.0,  7.8],
    "Switch":             [ 0.3,  0.4,  0.5,  0.6,  0.8,  1.0,  1.3,  1.8,  2.3,  3.0,  3.5,  4.0,  4.5],
}

# --- REVENUE ($B) ---
REVENUE_DATA = {
    "Amazon (AWS)":       [25.7, 35.0, 45.4, 62.2, 80.1, 90.8, 107.6, 130.0, 155.0, 182.0, 210.0, 240.0, 270.0],
    "Microsoft (Azure)":  [23.2, 33.7, 43.1, 60.0, 75.0, 96.8, 125.0, 160.0, 195.0, 230.0, 265.0, 300.0, 335.0],
    "Google (GCP)":       [ 5.8,  8.9, 13.1, 19.2, 26.3, 33.7,  43.2,  55.0,  68.0,  82.0,  97.0, 112.0, 128.0],
    "Meta":               [55.8, 70.7, 86.0, 117.9, 116.6, 134.9, 164.5, 195.0, 225.0, 255.0, 285.0, 315.0, 345.0],
    "Oracle Cloud":       [ 6.8,  7.0,  7.1,  7.5,  8.4, 12.5,  18.0,  24.0,  30.0,  36.0,  42.0,  48.0,  54.0],
    "Apple":              [265.6, 260.2, 274.5, 365.8, 394.3, 383.3, 391.0, 410.0, 430.0, 450.0, 470.0, 490.0, 510.0],
    "CoreWeave":          [ 0.0,  0.0,  0.0,  0.0,  0.02, 0.2,   0.8,   2.5,   5.0,   8.0,  11.0,  14.0,  17.0],
    "Lambda":             [ 0.0,  0.0,  0.0,  0.0,  0.01, 0.05,  0.3,   0.8,   1.5,   2.5,   3.5,   4.5,   5.5],
    "Crusoe Energy":      [ 0.0,  0.0,  0.0,  0.0,  0.01, 0.05,  0.2,   0.6,   1.2,   2.0,   3.0,   4.0,   5.0],
    "Voltage Park":       [ 0.0,  0.0,  0.0,  0.0,  0.0,  0.02,  0.1,   0.4,   0.8,   1.5,   2.5,   3.5,   4.5],
    "Together AI":        [ 0.0,  0.0,  0.0,  0.0,  0.0,  0.03,  0.1,   0.3,   0.7,   1.2,   2.0,   3.0,   4.0],
    "Applied Digital":    [ 0.0,  0.0,  0.0,  0.0,  0.005,0.04,  0.15,  0.5,   1.0,   1.8,   2.8,   3.8,   4.8],
    # Colo / REITs (colocation revenue)
    "Equinix":            [ 5.1,  5.6,  6.0,  6.6,  7.3,  8.1,  8.7,  9.6, 10.8, 12.2, 13.8, 15.5, 17.2],
    "Digital Realty":      [ 3.0,  3.2,  3.9,  4.4,  4.7,  5.5,  5.8,  6.5,  7.5,  8.8, 10.2, 11.8, 13.5],
    "CyrusOne":           [ 0.9,  1.0,  1.1,  1.2,  1.2,  1.3,  1.5,  1.8,  2.2,  2.6,  3.1,  3.6,  4.2],
    "QTS Realty":          [ 0.5,  0.5,  0.6,  0.7,  0.8,  1.2,  1.6,  2.2,  2.9,  3.8,  4.8,  5.8,  6.8],
    "Vantage Data Centers":[ 0.2,  0.3,  0.4,  0.5,  0.7,  1.0,  1.4,  2.0,  2.8,  3.8,  4.8,  5.8,  6.8],
    "Switch":             [ 0.4,  0.5,  0.5,  0.6,  0.7,  0.8,  1.0,  1.3,  1.7,  2.2,  2.8,  3.4,  4.0],
}

# --- SERVER COUNT (thousands of servers) ---
SERVER_DATA = {
    "Amazon (AWS)":       [2800, 3200, 3800, 4500, 5200, 5900, 7000, 8500, 10000, 11500, 13000, 14500, 16000],
    "Microsoft (Azure)":  [2200, 2600, 3000, 3800, 4500, 5200, 6500, 8000,  9500, 11000, 12500, 14000, 15000],
    "Google (GCP)":       [2000, 2300, 2700, 3200, 3800, 4500, 5500, 7000,  8500, 10000, 11500, 13000, 14000],
    "Meta":               [ 600,  750,  900, 1200, 1500, 2200, 3000, 4000,  5000,  6000,  7000,  8000,  9000],
    "Oracle Cloud":       [ 200,  250,  350,  500,  700, 1000, 1500, 2200,  3000,  3800,  4500,  5000,  5500],
    "Apple":              [ 150,  180,  220,  280,  340,  400,  500,  650,   800,   950,  1100,  1250,  1400],
    "CoreWeave":          [   0,    0,    0,    0,    2,   15,   60,  150,   280,   400,   520,   640,   750],
    "Lambda":             [   0,    0,    0,    0,    1,    5,   20,   50,   100,   160,   220,   280,   340],
    "Crusoe Energy":      [   0,    0,    0,    0,    1,    8,   25,   60,   120,   180,   250,   320,   380],
    "Voltage Park":       [   0,    0,    0,    0,    0,    3,   12,   35,    70,   110,   155,   200,   240],
    "Together AI":        [   0,    0,    0,    0,    0,    2,    8,   20,    45,    75,   110,   150,   190],
    "Applied Digital":    [   0,    0,    0,    0,    1,    4,   15,   40,    80,   130,   180,   230,   280],
    # Colo / REITs (customer-deployed servers hosted in their facilities)
    "Equinix":            [ 800,  900, 1000, 1100, 1200, 1350, 1500, 1700, 1950, 2250, 2550, 2900, 3250],
    "Digital Realty":      [ 600,  680,  770,  860,  950, 1100, 1250, 1450, 1700, 2000, 2350, 2700, 3050],
    "CyrusOne":           [ 180,  210,  240,  270,  300,  340,  400,  480,  580,  700,  820,  950, 1100],
    "QTS Realty":          [ 120,  140,  170,  210,  280,  380,  500,  650,  850, 1050, 1300, 1550, 1800],
    "Vantage Data Centers":[ 60,   80,  110,  150,  220,  320,  450,  620,  840, 1080, 1350, 1600, 1850],
    "Switch":             [ 100,  120,  140,  170,  210,  260,  320,  400,  500,  630,  770,  920, 1080],
}

# --- DATACENTER COUNT ---
DC_COUNT_DATA = {
    "Amazon (AWS)":       [ 60,  69,  76,  81,  87,  96, 110, 130, 150, 170, 185, 200, 215],
    "Microsoft (Azure)":  [ 50,  54,  58,  60,  65,  72,  82,  95, 110, 125, 140, 155, 170],
    "Google (GCP)":       [ 21,  23,  26,  29,  33,  38,  45,  55,  65,  75,  85,  95, 105],
    "Meta":               [ 10,  12,  14,  16,  18,  22,  28,  35,  42,  50,  58,  65,  72],
    "Oracle Cloud":       [ 16,  20,  28,  33,  40,  47,  55,  66,  78,  90, 100, 108, 115],
    "Apple":              [  5,   5,   6,   6,   7,   8,  10,  12,  14,  16,  18,  20,  22],
    "CoreWeave":          [  0,   0,   0,   0,   1,   4,  10,  18,  28,  38,  48,  56,  62],
    "Lambda":             [  0,   0,   0,   0,   1,   2,   4,   8,  14,  20,  26,  32,  38],
    "Crusoe Energy":      [  0,   0,   0,   0,   1,   3,   6,  12,  20,  28,  36,  42,  48],
    "Voltage Park":       [  0,   0,   0,   0,   0,   1,   3,   6,  10,  15,  20,  25,  30],
    "Together AI":        [  0,   0,   0,   0,   0,   1,   2,   4,   7,  11,  16,  21,  26],
    "Applied Digital":    [  0,   0,   0,   0,   1,   2,   4,   8,  13,  18,  24,  30,  35],
    # Colo / REITs
    "Equinix":            [200, 212, 220, 235, 248, 260, 270, 285, 300, 320, 340, 360, 380],
    "Digital Realty":      [185, 195, 210, 225, 240, 310, 315, 330, 350, 375, 400, 425, 450],
    "CyrusOne":           [ 30,  34,  38,  42,  45,  48,  52,  58,  65,  73,  82,  90,  98],
    "QTS Realty":          [ 12,  14,  16,  19,  25,  32,  38,  46,  55,  65,  76,  87,  98],
    "Vantage Data Centers":[ 8,  10,  13,  16,  22,  30,  38,  48,  60,  72,  85,  96, 108],
    "Switch":             [  5,   5,   6,   7,   8,  10,  12,  15,  18,  22,  26,  30,  34],
}

# --- GW CAPACITY (IT Load, GW) ---
GW_CAPACITY_DATA = {
    "Amazon (AWS)":       [2.5, 3.0, 3.6, 4.5, 5.5, 6.8, 8.5, 11.0, 14.0, 17.0, 20.0, 23.0, 26.0],
    "Microsoft (Azure)":  [2.0, 2.4, 3.0, 3.8, 4.8, 6.0, 7.5, 10.0, 13.0, 16.0, 19.0, 22.0, 25.0],
    "Google (GCP)":       [1.8, 2.1, 2.6, 3.2, 4.0, 5.0, 6.5,  8.5, 11.0, 13.5, 16.0, 18.5, 21.0],
    "Meta":               [0.8, 1.0, 1.3, 1.8, 2.3, 3.5, 5.0,  7.0,  9.0, 11.0, 13.0, 15.0, 17.0],
    "Oracle Cloud":       [0.3, 0.4, 0.5, 0.7, 1.0, 1.5, 2.5,  3.8,  5.0,  6.5,  8.0,  9.5, 11.0],
    "Apple":              [0.2, 0.3, 0.3, 0.4, 0.5, 0.6, 0.8,  1.0,  1.3,  1.6,  1.9,  2.2,  2.5],
    "CoreWeave":          [0.0, 0.0, 0.0, 0.0, 0.01,0.08,0.35, 0.9,  1.6,  2.5,  3.5,  4.5,  5.5],
    "Lambda":             [0.0, 0.0, 0.0, 0.0, 0.005,0.03,0.12,0.30, 0.60, 1.00, 1.40, 1.80, 2.20],
    "Crusoe Energy":      [0.0, 0.0, 0.0, 0.0, 0.005,0.05,0.15,0.40, 0.80, 1.20, 1.70, 2.20, 2.70],
    "Voltage Park":       [0.0, 0.0, 0.0, 0.0, 0.0, 0.02,0.08,0.20, 0.45, 0.75, 1.10, 1.45, 1.80],
    "Together AI":        [0.0, 0.0, 0.0, 0.0, 0.0, 0.01,0.05,0.12, 0.28, 0.50, 0.80, 1.10, 1.40],
    "Applied Digital":    [0.0, 0.0, 0.0, 0.0, 0.003,0.03,0.10,0.25, 0.50, 0.80, 1.20, 1.60, 2.00],
    # Colo / REITs
    "Equinix":            [1.5, 1.7, 1.9, 2.1, 2.4, 2.7, 3.0, 3.4, 3.9, 4.5, 5.2, 5.9, 6.7],
    "Digital Realty":      [1.3, 1.5, 1.7, 2.0, 2.3, 2.8, 3.2, 3.8, 4.5, 5.4, 6.3, 7.3, 8.3],
    "CyrusOne":           [0.3, 0.4, 0.4, 0.5, 0.5, 0.6, 0.7, 0.9, 1.1, 1.3, 1.6, 1.9, 2.2],
    "QTS Realty":          [0.2, 0.3, 0.3, 0.4, 0.6, 0.9, 1.2, 1.6, 2.1, 2.7, 3.3, 4.0, 4.7],
    "Vantage Data Centers":[0.1, 0.2, 0.2, 0.3, 0.5, 0.7, 1.0, 1.5, 2.0, 2.7, 3.4, 4.0, 4.6],
    "Switch":             [0.2, 0.3, 0.3, 0.4, 0.5, 0.6, 0.8, 1.0, 1.3, 1.7, 2.1, 2.5, 2.9],
}

# --- POWER CONTRACTED / AVAILABLE (GW) ---
POWER_AVAILABLE_DATA = {
    "Amazon (AWS)":       [3.0, 3.6, 4.5, 5.5, 7.0, 9.0, 12.0, 15.0, 19.0, 23.0, 27.0, 31.0, 35.0],
    "Microsoft (Azure)":  [2.5, 3.0, 3.8, 5.0, 6.2, 8.0, 10.5, 14.0, 18.0, 22.0, 26.0, 30.0, 34.0],
    "Google (GCP)":       [2.2, 2.6, 3.2, 4.0, 5.2, 6.5,  9.0, 12.0, 15.0, 18.5, 22.0, 25.5, 29.0],
    "Meta":               [1.0, 1.3, 1.7, 2.3, 3.0, 5.0,  7.0, 10.0, 13.0, 16.0, 19.0, 22.0, 25.0],
    "Oracle Cloud":       [0.4, 0.5, 0.7, 1.0, 1.4, 2.2,  3.5,  5.5,  7.5, 10.0, 12.5, 14.5, 16.5],
    "Apple":              [0.3, 0.4, 0.4, 0.5, 0.6, 0.8,  1.1,  1.5,  2.0,  2.5,  3.0,  3.5,  4.0],
    "CoreWeave":          [0.0, 0.0, 0.0, 0.0, 0.015,0.12,0.5,  1.3,  2.5,  4.0,  5.5,  7.0,  8.5],
    "Lambda":             [0.0, 0.0, 0.0, 0.0, 0.008,0.05,0.18, 0.45, 0.90, 1.50, 2.10, 2.70, 3.30],
    "Crusoe Energy":      [0.0, 0.0, 0.0, 0.0, 0.008,0.08,0.25, 0.60, 1.20, 1.80, 2.50, 3.20, 4.00],
    "Voltage Park":       [0.0, 0.0, 0.0, 0.0, 0.0, 0.03,0.12, 0.30, 0.65, 1.10, 1.60, 2.10, 2.60],
    "Together AI":        [0.0, 0.0, 0.0, 0.0, 0.0, 0.02,0.08, 0.18, 0.40, 0.75, 1.15, 1.60, 2.05],
    "Applied Digital":    [0.0, 0.0, 0.0, 0.0, 0.005,0.05,0.15, 0.40, 0.75, 1.20, 1.80, 2.40, 3.00],
    # Colo / REITs
    "Equinix":            [2.0, 2.2, 2.5, 2.8, 3.2, 3.6, 4.1, 4.7, 5.5, 6.4, 7.4, 8.5, 9.6],
    "Digital Realty":      [1.8, 2.0, 2.3, 2.7, 3.1, 3.8, 4.4, 5.3, 6.3, 7.5, 8.8, 10.2, 11.6],
    "CyrusOne":           [0.4, 0.5, 0.6, 0.7, 0.7, 0.8, 1.0, 1.2, 1.5, 1.8, 2.2, 2.6, 3.0],
    "QTS Realty":          [0.3, 0.4, 0.4, 0.6, 0.8, 1.2, 1.7, 2.3, 3.0, 3.8, 4.7, 5.6, 6.5],
    "Vantage Data Centers":[0.2, 0.2, 0.3, 0.4, 0.7, 1.0, 1.4, 2.1, 2.8, 3.8, 4.8, 5.6, 6.5],
    "Switch":             [0.3, 0.3, 0.4, 0.5, 0.6, 0.8, 1.0, 1.3, 1.7, 2.2, 2.8, 3.4, 4.0],
}

# --- POWER GENERATION MIX (% Renewable of total power) ---
RENEWABLE_PCT = {
    "Amazon (AWS)":       [0.30, 0.36, 0.42, 0.50, 0.65, 0.80, 0.90, 0.95, 1.00, 1.00, 1.00, 1.00, 1.00],
    "Microsoft (Azure)":  [0.35, 0.40, 0.45, 0.50, 0.58, 0.68, 0.78, 0.85, 0.92, 0.97, 1.00, 1.00, 1.00],
    "Google (GCP)":       [0.50, 0.60, 0.70, 0.80, 0.90, 0.95, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00],
    "Meta":               [0.25, 0.30, 0.40, 0.50, 0.60, 0.72, 0.82, 0.90, 0.95, 1.00, 1.00, 1.00, 1.00],
    "Oracle Cloud":       [0.10, 0.12, 0.15, 0.20, 0.28, 0.38, 0.50, 0.60, 0.70, 0.80, 0.88, 0.93, 0.97],
    "Apple":              [0.80, 0.85, 0.90, 0.95, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00],
    "CoreWeave":          [0.00, 0.00, 0.00, 0.00, 0.10, 0.15, 0.25, 0.35, 0.45, 0.55, 0.65, 0.75, 0.85],
    "Lambda":             [0.00, 0.00, 0.00, 0.00, 0.10, 0.15, 0.20, 0.30, 0.40, 0.50, 0.60, 0.70, 0.80],
    "Crusoe Energy":      [0.00, 0.00, 0.00, 0.00, 0.05, 0.10, 0.20, 0.30, 0.40, 0.50, 0.60, 0.70, 0.80],
    "Voltage Park":       [0.00, 0.00, 0.00, 0.00, 0.00, 0.10, 0.15, 0.25, 0.35, 0.45, 0.55, 0.65, 0.75],
    "Together AI":        [0.00, 0.00, 0.00, 0.00, 0.00, 0.10, 0.15, 0.25, 0.35, 0.45, 0.55, 0.65, 0.75],
    "Applied Digital":    [0.00, 0.00, 0.00, 0.00, 0.05, 0.10, 0.18, 0.28, 0.38, 0.48, 0.58, 0.68, 0.78],
    # Colo / REITs
    "Equinix":            [0.40, 0.50, 0.60, 0.70, 0.80, 0.90, 0.95, 0.97, 1.00, 1.00, 1.00, 1.00, 1.00],
    "Digital Realty":      [0.25, 0.30, 0.35, 0.42, 0.50, 0.60, 0.70, 0.80, 0.88, 0.94, 0.98, 1.00, 1.00],
    "CyrusOne":           [0.15, 0.18, 0.22, 0.28, 0.35, 0.45, 0.55, 0.65, 0.75, 0.83, 0.90, 0.95, 0.98],
    "QTS Realty":          [0.10, 0.15, 0.20, 0.25, 0.32, 0.42, 0.52, 0.62, 0.72, 0.82, 0.90, 0.95, 0.98],
    "Vantage Data Centers":[0.10, 0.12, 0.18, 0.25, 0.30, 0.40, 0.50, 0.60, 0.70, 0.80, 0.88, 0.94, 0.98],
    "Switch":             [0.80, 0.85, 0.90, 0.95, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00],
}


# ============================================================
# AI CAPEX & AI REVENUE DATA — Capex-to-AI-Revenue Conversion
# ============================================================

# AI-specific capex ($B) — portion of total capex allocated to AI/ML infrastructure
# Includes: GPU clusters, AI-optimized networking, liquid cooling for AI racks, AI-dedicated DCs
AI_CAPEX_DATA = {
    # 2018-2024 hist, 2025E-2030E forecast
    "Amazon (AWS)":       [ 0.5,  1.0,  1.8,  3.5,  7.0, 16.0, 35.0,  55.0,  72.0,  86.0,  97.0, 102.0, 106.0],
    "Microsoft (Azure)":  [ 0.4,  0.8,  1.5,  3.0,  7.5, 14.0, 32.0,  50.0,  62.0,  75.0,  84.0,  89.0,  93.0],
    "Google (GCP)":       [ 0.8,  1.2,  2.0,  4.0,  7.0, 14.0, 30.0,  48.0,  56.0,  64.0,  70.0,  74.0,  77.0],
    "Meta":               [ 0.5,  1.0,  2.5,  5.0,  8.0, 18.0, 28.0,  48.0,  54.0,  60.0,  62.0,  64.0,  66.0],
    "Oracle Cloud":       [ 0.0,  0.0,  0.1,  0.3,  1.0,  3.5,  9.0,  14.0,  17.5,  20.0,  22.0,  23.0,  24.0],
    "Apple":              [ 0.1,  0.2,  0.3,  0.5,  0.8,  1.5,  3.0,   4.5,   5.5,   6.5,   7.2,   7.8,   8.3],
    "CoreWeave":          [ 0.0,  0.0,  0.0,  0.0,  0.1,  0.8,  3.4,   7.8,  11.8,  14.8,  16.8,  17.8,  18.8],
    "Lambda":             [ 0.0,  0.0,  0.0,  0.0, 0.05,  0.2,  0.8,   1.9,   3.4,   4.9,   5.9,   6.9,   7.4],
    "Crusoe Energy":      [ 0.0,  0.0,  0.0,  0.0, 0.05,  0.3,  0.9,   2.4,   3.8,   5.3,   6.8,   7.8,   8.3],
    "Voltage Park":       [ 0.0,  0.0,  0.0,  0.0,  0.0,  0.1,  0.5,   1.5,   2.4,   3.4,   4.4,   4.9,   5.4],
    "Together AI":        [ 0.0,  0.0,  0.0,  0.0,  0.0,  0.1,  0.3,   0.8,   1.5,   2.4,   3.4,   3.9,   4.4],
    "Applied Digital":    [ 0.0,  0.0,  0.0,  0.0, 0.02, 0.14,  0.6,   1.4,   2.4,   3.4,   4.4,   4.9,   5.4],
    "Equinix":            [ 0.0,  0.0,  0.1,  0.1,  0.3,  0.5,  0.8,   1.2,   1.8,   2.4,   3.0,   3.6,   4.2],
    "Digital Realty":      [ 0.0,  0.0,  0.1,  0.2,  0.4,  0.8,  1.2,   1.8,   2.6,   3.5,   4.4,   5.2,   6.0],
    "CyrusOne":           [ 0.0,  0.0,  0.0,  0.1,  0.1,  0.3,  0.5,   0.7,   1.0,   1.3,   1.6,   1.9,   2.2],
    "QTS Realty":          [ 0.0,  0.0,  0.0,  0.1,  0.2,  0.6,  1.0,   1.5,   2.2,   3.0,   3.8,   4.5,   5.2],
    "Vantage Data Centers":[ 0.0,  0.0,  0.0,  0.1,  0.2,  0.5,  0.9,   1.5,   2.1,   2.9,   3.6,   4.2,   4.8],
    "Switch":             [ 0.0,  0.0,  0.0,  0.0,  0.1,  0.2,  0.4,   0.7,   1.0,   1.5,   1.9,   2.3,   2.7],
}

# AI-specific revenue ($B) — revenue directly attributable to AI products/services
# Hyperscalers: AI cloud services (GPU instances, AI APIs, ML platforms, Copilots)
# Meta: AI-attributable ad revenue uplift (portion of ad rev enabled by AI recommendation/targeting improvements)
# Apple: Apple Intelligence services, AI-enhanced services revenue uplift
# Neoclouds: nearly 100% of revenue is AI (GPU-as-a-service)
# Colo/REITs: revenue from AI-dedicated / high-density GPU tenant leases
AI_REVENUE_DATA = {
    "Amazon (AWS)":       [ 0.0,  0.2,  0.5,  1.5,  3.5,  8.0, 18.0,  32.0,  50.0,  72.0,  95.0, 120.0, 148.0],
    "Microsoft (Azure)":  [ 0.0,  0.1,  0.4,  1.2,  4.0, 10.0, 25.0,  42.0,  65.0,  92.0, 120.0, 150.0, 182.0],
    "Google (GCP)":       [ 0.0,  0.2,  0.5,  1.0,  2.5,  6.0, 14.0,  25.0,  38.0,  52.0,  68.0,  85.0, 102.0],
    "Meta":               [ 0.0,  0.5,  2.0,  5.0, 10.0, 18.0, 32.0,  48.0,  65.0,  82.0, 100.0, 118.0, 135.0],
    "Oracle Cloud":       [ 0.0,  0.0,  0.0,  0.1,  0.4,  1.5,  5.0,  10.0,  16.0,  22.0,  28.0,  35.0,  42.0],
    "Apple":              [ 0.0,  0.0,  0.1,  0.3,  0.5,  1.0,  3.0,   6.0,  10.0,  15.0,  21.0,  28.0,  36.0],
    "CoreWeave":          [ 0.0,  0.0,  0.0,  0.0, 0.02,  0.2,  0.8,   2.4,   4.9,   7.9,  10.9,  13.9,  16.9],
    "Lambda":             [ 0.0,  0.0,  0.0,  0.0, 0.01, 0.05,  0.3,   0.8,   1.5,   2.4,   3.4,   4.4,   5.4],
    "Crusoe Energy":      [ 0.0,  0.0,  0.0,  0.0, 0.01, 0.04,  0.2,   0.6,   1.2,   1.9,   2.9,   3.9,   4.9],
    "Voltage Park":       [ 0.0,  0.0,  0.0,  0.0,  0.0, 0.02,  0.1,   0.4,   0.8,   1.5,   2.4,   3.4,   4.4],
    "Together AI":        [ 0.0,  0.0,  0.0,  0.0,  0.0, 0.03,  0.1,   0.3,   0.7,   1.2,   2.0,   2.9,   3.9],
    "Applied Digital":    [ 0.0,  0.0,  0.0,  0.0,0.005, 0.04,  0.14,  0.5,   1.0,   1.7,   2.7,   3.7,   4.7],
    "Equinix":            [ 0.0,  0.0,  0.0,  0.1,  0.2,  0.5,  0.9,   1.5,   2.4,   3.5,   4.8,   6.2,   7.8],
    "Digital Realty":      [ 0.0,  0.0,  0.0,  0.1,  0.3,  0.6,  1.0,   1.6,   2.5,   3.6,   5.0,   6.5,   8.2],
    "CyrusOne":           [ 0.0,  0.0,  0.0,  0.0,  0.1,  0.2,  0.3,   0.5,   0.8,   1.1,   1.5,   2.0,   2.5],
    "QTS Realty":          [ 0.0,  0.0,  0.0,  0.0,  0.1,  0.3,  0.6,   1.0,   1.5,   2.2,   3.0,   3.9,   4.8],
    "Vantage Data Centers":[ 0.0,  0.0,  0.0,  0.0,  0.1,  0.3,  0.5,   0.9,   1.5,   2.2,   3.0,   3.9,   4.8],
    "Switch":             [ 0.0,  0.0,  0.0,  0.0,  0.1,  0.2,  0.3,   0.5,   0.8,   1.2,   1.6,   2.1,   2.6],
}

# Capex-to-revenue lag: average quarters from AI capex deployment to revenue generation
# Shorter lag = faster monetization; varies by business model
AI_CAPEX_LAG_QUARTERS = {
    "Amazon (AWS)":        6,   # Build → GA launch → customer ramp
    "Microsoft (Azure)":   5,   # Strong Copilot/OpenAI pull-through
    "Google (GCP)":        7,   # Longer due to enterprise sales cycle
    "Meta":                3,   # Internal — ad model improvement is rapid
    "Oracle Cloud":        8,   # Enterprise sales cycle + newer to AI
    "Apple":              10,   # On-device, long product cycle
    "CoreWeave":           4,   # Direct GPU lease, fast activation
    "Lambda":              4,
    "Crusoe Energy":       5,
    "Voltage Park":        4,
    "Together AI":         5,
    "Applied Digital":     5,
    "Equinix":             6,   # Lease negotiation + buildout
    "Digital Realty":       6,
    "CyrusOne":            7,
    "QTS Realty":           6,
    "Vantage Data Centers": 7,
    "Switch":              7,
}


# ============================================================
# POWER SUPPLY & GENERATION DATA
# ============================================================

# US datacenter power demand (GW) — total across all operators
US_DC_POWER_DEMAND = [17, 19, 22, 26, 32, 40, 55, 72, 95, 120, 150, 180, 210]

# US grid capacity available for new DC loads (GW of uncommitted capacity)
US_GRID_UNCOMMITTED = [35, 33, 30, 28, 25, 20, 18, 22, 28, 35, 42, 50, 58]

# New power generation capacity additions for DC loads (GW added per year, by type)
POWER_GEN_ADDITIONS = {
    "Natural Gas (GW added)":     [2.0, 2.2, 2.5, 3.0, 3.5, 4.0, 5.5, 7.0, 8.5, 10.0, 11.0, 11.5, 12.0],
    "Solar (GW added)":           [0.8, 1.0, 1.5, 2.0, 3.0, 4.5, 6.0, 8.5, 12.0, 16.0, 20.0, 24.0, 28.0],
    "Wind (GW added)":            [0.5, 0.6, 0.8, 1.0, 1.3, 1.8, 2.5, 3.5, 5.0, 6.5, 8.0, 9.5, 11.0],
    "Nuclear (GW added)":         [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.3, 0.8, 1.5, 2.5, 4.0, 5.5, 7.0],
    "Battery Storage (GW added)": [0.1, 0.2, 0.3, 0.5, 0.8, 1.2, 2.0, 3.5, 5.5, 8.0, 11.0, 14.0, 17.0],
    "SMR / Advanced Nuclear (GW)":[0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.1, 0.3, 0.8, 1.5, 2.5, 4.0],
}

# Average interconnection queue time (months)
INTERCONNECT_QUEUE_MONTHS = [18, 20, 22, 24, 28, 32, 38, 42, 40, 36, 32, 28, 24]

# Average power cost to datacenters ($/MWh)
AVG_POWER_COST = [48, 50, 52, 55, 60, 65, 70, 72, 68, 64, 60, 56, 52]

# Capex ramp-up timeline (months from land to operational)
DC_BUILD_TIMELINE_MONTHS = {
    "Permitting & Approvals":     [ 8,  8,  9, 10, 11, 12, 14, 14, 13, 12, 11, 10, 10],
    "Power Procurement":          [12, 13, 14, 16, 18, 22, 26, 28, 26, 24, 22, 20, 18],
    "Site Prep & Construction":   [12, 12, 13, 14, 14, 15, 16, 16, 15, 14, 14, 13, 12],
    "Equipment Install & Test":   [ 4,  4,  4,  4,  5,  5,  5,  5,  5,  4,  4,  4,  4],
    "Total (Land to Live)":       [24, 24, 26, 28, 30, 34, 38, 40, 38, 34, 32, 28, 26],
}

# Incremental GW needed per year vs what can actually be delivered
GW_DEMAND_INCREMENT = [0, 2.0, 3.0, 4.0, 6.0, 8.0, 15.0, 17.0, 23.0, 25.0, 30.0, 30.0, 30.0]
GW_SUPPLY_DELIVERABLE = [0, 2.0, 3.0, 4.0, 5.5, 7.5, 11.0, 15.0, 20.0, 28.0, 35.0, 42.0, 50.0]


# ============================================================
# DC BILL OF MATERIALS DATA (cost per MW of IT load, $M/MW)
# ============================================================

# BOM for a typical hyperscale datacenter (per MW of IT load)
BOM_CATEGORIES = [
    # (Category, Sub-category, Cost $M per MW 2020, Cost $M per MW 2024, Cost $M per MW 2028E, Cost $M per MW 2030E, Key Vendors)
    ("Land & Site", "Land Acquisition", 0.3, 0.5, 0.7, 0.8, "CBRE, JLL, Prologis"),
    ("Land & Site", "Site Preparation & Grading", 0.2, 0.3, 0.3, 0.3, "Fluor, Bechtel, Jacobs"),
    ("Building Shell", "Building Structure & Envelope", 1.5, 1.8, 2.0, 2.1, "DPR, Holder, Turner"),
    ("Building Shell", "Raised Floor / Slab", 0.3, 0.3, 0.4, 0.4, "Tate, ASM, Haworth"),
    ("Power Infrastructure", "HV Substation & Transformers", 1.0, 1.5, 1.8, 1.9, "ABB/Hitachi, Siemens, Eaton, GE Vernova"),
    ("Power Infrastructure", "MV/LV Switchgear & Busway", 0.8, 1.0, 1.2, 1.2, "Schneider, Eaton, ABB"),
    ("Power Infrastructure", "UPS Systems", 0.6, 0.8, 0.9, 0.9, "Vertiv, Schneider, Eaton"),
    ("Power Infrastructure", "Backup Generators (Diesel/Gas)", 0.5, 0.6, 0.7, 0.7, "Caterpillar, Cummins, MTU"),
    ("Power Infrastructure", "PDUs & Power Whips", 0.3, 0.4, 0.4, 0.4, "Vertiv, Schneider, Raritan"),
    ("Cooling Systems", "Chillers & Cooling Towers", 0.8, 1.0, 0.9, 0.8, "Trane, Johnson Controls, Carrier"),
    ("Cooling Systems", "CRAH/CRAC Units", 0.3, 0.3, 0.3, 0.2, "Vertiv, Stulz, Schneider"),
    ("Cooling Systems", "Liquid Cooling (DLC/Immersion)", 0.1, 0.5, 0.9, 1.0, "CoolIT, GRC, Asetek, ZutaCore"),
    ("Cooling Systems", "Piping & Plumbing", 0.2, 0.3, 0.3, 0.3, "Victaulic, Watts, Uponor"),
    ("Networking & Cabling", "Fiber Optic Backbone", 0.3, 0.4, 0.5, 0.5, "Corning, CommScope, Panduit"),
    ("Networking & Cabling", "Structured Cabling & Patch Panels", 0.2, 0.2, 0.3, 0.3, "Panduit, Belden, Leviton"),
    ("Networking & Cabling", "Network Switches & Routers", 0.5, 0.7, 0.8, 0.8, "Arista, Cisco, Juniper, Broadcom"),
    ("Server & IT Equipment", "Server Racks & Cabinets", 0.2, 0.3, 0.3, 0.3, "Vertiv, Rittal, Chatsworth"),
    ("Server & IT Equipment", "GPU Servers (AI-optimized)", 0.0, 3.0, 5.0, 5.5, "NVIDIA, AMD, Intel, Dell, HPE, Supermicro"),
    ("Server & IT Equipment", "CPU Servers (General)", 2.5, 2.0, 1.8, 1.6, "Dell, HPE, Lenovo, Supermicro"),
    ("Server & IT Equipment", "Storage (SSD/HDD Arrays)", 0.4, 0.5, 0.6, 0.6, "Pure Storage, NetApp, Dell, Samsung"),
    ("Security & Safety", "Fire Suppression", 0.1, 0.1, 0.1, 0.1, "Fike, Kidde, Novec"),
    ("Security & Safety", "Physical Security & Access", 0.1, 0.1, 0.2, 0.2, "Honeywell, Genetec, Axis"),
    ("Security & Safety", "BMS / DCIM Software", 0.1, 0.1, 0.2, 0.2, "Schneider, Vertiv, Nlyte, Sunbird"),
    ("Prof. Services", "Design & Engineering", 0.3, 0.4, 0.5, 0.5, "HDR, Corgan, Gensler, Morrison Hershfield"),
    ("Prof. Services", "Project Management", 0.2, 0.2, 0.3, 0.3, "AECOM, Jacobs, Turner & Townsend"),
    ("Prof. Services", "Commissioning & Testing", 0.1, 0.2, 0.2, 0.2, "Cyient, Bureau Veritas, QTS"),
]


# ============================================================
# SUPPLY CHAIN REVENUE DATA — How capex translates to vendor revenue
# ============================================================

# DC supply chain segments — total addressable market ($B)
SUPPLY_CHAIN_REVENUE = {
    # Segment: [2018, ... 2030E]
    "Servers & GPUs":              [ 42,  48,  55,  65,  78,  95, 140, 195, 260, 320, 380, 430, 475],
    "Networking Equipment":        [ 12,  13,  15,  17,  19,  22,  30,  40,  52,  65,  78,  90, 102],
    "Power Equipment (Elec. Infra)":[ 8,   9,  10,  12,  14,  17,  24,  33,  44,  56,  68,  80,  90],
    "Cooling Systems":             [  5,   5,   6,   7,   8,  10,  15,  21,  29,  38,  48,  56,  64],
    "Construction & Engineering":  [  6,   7,   8,   9,  11,  14,  20,  28,  37,  46,  55,  62,  68],
    "Real Estate / Colo Leasing":  [ 10,  11,  12,  14,  16,  19,  23,  28,  35,  44,  54,  64,  74],
    "DC Software & Mgmt (DCIM)":   [  2,   2,   3,   3,   4,   5,   6,   8,  10,  13,  16,  19,  22],
    "Fiber & Connectivity":        [  4,   5,   5,   6,   7,   8,  10,  13,  17,  22,  27,  32,  37],
}

# Key beneficiaries by segment
SUPPLY_CHAIN_VENDORS = {
    "Servers & GPUs":              "NVIDIA, AMD, Intel, Broadcom, Dell, HPE, Supermicro, Lenovo",
    "Networking Equipment":        "Arista, Cisco, Juniper, Broadcom, Infinera, Ciena",
    "Power Equipment (Elec. Infra)":"Vertiv, Schneider Electric, Eaton, GE Vernova, Caterpillar, Cummins, ABB",
    "Cooling Systems":             "Vertiv, Johnson Controls, Trane, Carrier, CoolIT, Asetek, GRC",
    "Construction & Engineering":  "Fluor, Bechtel, DPR, Turner, AECOM, Jacobs, Holder",
    "Real Estate / Colo Leasing":  "Equinix, Digital Realty, QTS/Blackstone, CyrusOne/KKR, Vantage",
    "DC Software & Mgmt (DCIM)":   "Schneider, Vertiv, Nlyte, Sunbird, Siemens",
    "Fiber & Connectivity":        "Corning, CommScope, Lumen, Zayo, Crown Castle",
}

# Capex $ flow: for every $1 of total DC capex, how it splits (approximate, changes over time)
CAPEX_FLOW_PCT = {
    # Segment: [2020, 2024, 2028E, 2030E]
    "Servers & GPUs":                [0.35, 0.42, 0.45, 0.44],
    "Networking Equipment":          [0.10, 0.09, 0.09, 0.09],
    "Power Infrastructure":          [0.15, 0.14, 0.13, 0.12],
    "Cooling Systems":               [0.08, 0.08, 0.08, 0.07],
    "Building & Construction":       [0.12, 0.10, 0.09, 0.09],
    "Land & Real Estate":            [0.05, 0.05, 0.04, 0.04],
    "Fiber & Connectivity":          [0.05, 0.04, 0.04, 0.04],
    "Software, Security & Other":    [0.03, 0.03, 0.03, 0.03],
    "Design, Engineering & PM":      [0.04, 0.03, 0.03, 0.03],
    "Contingency & Other":           [0.03, 0.02, 0.02, 0.05],
}


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def apply_cell_style(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def write_title_row(ws, row, title, num_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=title)
    apply_cell_style(cell, font=TITLE_FONT, fill=TITLE_FILL,
                     alignment=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[row].height = 32


def write_section_header(ws, row, title, num_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=title)
    apply_cell_style(cell, font=SECTION_FONT, fill=LIGHT_FILL,
                     alignment=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[row].height = 22


def write_subsection_header(ws, row, title, num_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=title)
    apply_cell_style(cell, font=SUBSECTION_FONT, fill=TEAL_FILL,
                     alignment=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[row].height = 20


def write_column_headers(ws, row, labels, start_col=1):
    for i, label in enumerate(labels):
        cell = ws.cell(row=row, column=start_col + i, value=label)
        apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL,
                         alignment=Alignment(horizontal="center", vertical="center"),
                         border=THIN_BORDER)
    ws.row_dimensions[row].height = 20


def write_data_row(ws, row, company, values, start_col=1, fmt="0.0", is_total=False, is_subtotal=False, alt=False):
    cell = ws.cell(row=row, column=start_col, value=company)
    if is_total:
        apply_cell_style(cell, font=TOTAL_FONT, fill=TOTAL_FILL,
                         alignment=Alignment(horizontal="left", vertical="center"),
                         border=THIN_BORDER)
    elif is_subtotal:
        apply_cell_style(cell, font=Font(name="Calibri", size=10, bold=True, color=WHITE),
                         fill=SUBTOTAL_FILL,
                         alignment=Alignment(horizontal="left", vertical="center"),
                         border=THIN_BORDER)
    else:
        apply_cell_style(cell, font=COMPANY_FONT,
                         fill=ALT_ROW_FILL if alt else None,
                         alignment=Alignment(horizontal="left", vertical="center"),
                         border=THIN_BORDER)
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=start_col + 1 + i, value=v)
        year_idx = i
        is_forecast = (year_idx >= len(HIST_YEARS))
        if is_total:
            apply_cell_style(cell, font=TOTAL_FONT, fill=TOTAL_FILL,
                             alignment=Alignment(horizontal="center", vertical="center"),
                             border=THIN_BORDER, number_format=fmt)
        elif is_subtotal:
            apply_cell_style(cell, font=Font(name="Calibri", size=10, bold=True, color=WHITE),
                             fill=SUBTOTAL_FILL,
                             alignment=Alignment(horizontal="center", vertical="center"),
                             border=THIN_BORDER, number_format=fmt)
        else:
            fill = FORECAST_FILL if is_forecast else (ALT_ROW_FILL if alt else None)
            apply_cell_style(cell, font=DATA_FONT,
                             fill=fill,
                             alignment=Alignment(horizontal="center", vertical="center"),
                             border=THIN_BORDER, number_format=fmt)


def compute_totals(data_dict, companies):
    totals = [0.0] * len(YEARS)
    for c in companies:
        for i, v in enumerate(data_dict[c]):
            totals[i] += v
    return [round(t, 2) for t in totals]


def add_line_chart(ws, title, cat_row, data_rows, labels, min_col, max_col, chart_row, chart_col, width=22, height=12):
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.width = width
    chart.height = height
    chart.y_axis.title = ""
    chart.x_axis.title = ""
    cats = Reference(ws, min_col=min_col, min_row=cat_row, max_col=max_col)
    chart.set_categories(cats)
    colors = ["4472C4", "ED7D31", "70AD47", "FFC000", "5B9BD5", "FF6384",
              "9966FF", "FF9F40", "36A2EB", "C9CBCF", "FF6384", "4BC0C0"]
    for idx, (drow, label) in enumerate(zip(data_rows, labels)):
        data = Reference(ws, min_col=min_col, min_row=drow, max_col=max_col)
        chart.add_data(data, from_rows=True, titles_from_data=False)
        chart.series[idx].name = label
        chart.series[idx].graphicalProperties.line.width = 22000
        if idx < len(colors):
            chart.series[idx].graphicalProperties.line.solidFill = colors[idx]
    chart.legend.position = 'b'
    ws.add_chart(chart, f"{get_column_letter(chart_col)}{chart_row}")
    return chart


def add_bar_chart(ws, title, cat_row, data_rows, labels, min_col, max_col, chart_row, chart_col, width=22, height=12, stacked=False):
    chart = BarChart()
    chart.title = title
    chart.style = 10
    chart.width = width
    chart.height = height
    if stacked:
        chart.grouping = "stacked"
        chart.overlap = 100
    cats = Reference(ws, min_col=min_col, min_row=cat_row, max_col=max_col)
    chart.set_categories(cats)
    colors = ["4472C4", "ED7D31", "70AD47", "FFC000", "5B9BD5", "FF6384",
              "9966FF", "FF9F40", "36A2EB", "C9CBCF", "FF6384", "4BC0C0"]
    for idx, (drow, label) in enumerate(zip(data_rows, labels)):
        data = Reference(ws, min_col=min_col, min_row=drow, max_col=max_col)
        chart.add_data(data, from_rows=True, titles_from_data=False)
        chart.series[idx].name = label
        if idx < len(colors):
            chart.series[idx].graphicalProperties.solidFill = colors[idx]
    chart.legend.position = 'b'
    ws.add_chart(chart, f"{get_column_letter(chart_col)}{chart_row}")
    return chart


# ============================================================
# SHEET BUILDERS
# ============================================================

def build_data_sheet(wb, sheet_name, title, data_dict, unit_label, fmt="0.0"):
    """Build a standard data sheet with hyperscaler, neocloud, and colo/REIT sections."""
    ws = wb.create_sheet(title=sheet_name)
    num_cols = 1 + len(YEARS)

    ws.column_dimensions['A'].width = 26
    for i in range(len(YEARS)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 12

    row = 1
    write_title_row(ws, row, f"{title} ({unit_label})", num_cols)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    note_cell = ws.cell(row=row, column=1,
        value="Historical: 2018-2024  |  Forecast: 2025E-2030E (yellow shading)  |  Sources: Public filings, industry estimates")
    apply_cell_style(note_cell, font=NOTE_FONT, alignment=Alignment(horizontal="left"))
    row += 1

    all_seg_data_rows = {}
    all_seg_total_rows = {}
    all_seg_header_rows = {}

    for seg_name, seg_companies in SEGMENT_MAP:
        write_section_header(ws, row, seg_name, num_cols)
        row += 1
        header_row = row
        all_seg_header_rows[seg_name] = header_row
        write_column_headers(ws, row, ["Company"] + YEAR_LABELS)
        row += 1

        seg_rows = []
        for idx, company in enumerate(seg_companies):
            write_data_row(ws, row, company, data_dict[company], fmt=fmt, alt=(idx % 2 == 1))
            seg_rows.append(row)
            row += 1
        all_seg_data_rows[seg_name] = seg_rows

        seg_totals = compute_totals(data_dict, seg_companies)
        write_data_row(ws, row, f"{seg_name} Total", seg_totals, fmt=fmt, is_subtotal=True)
        all_seg_total_rows[seg_name] = row
        row += 2

    # Grand total
    all_totals = compute_totals(data_dict, ALL_COMPANIES)
    write_data_row(ws, row, "GRAND TOTAL", all_totals, fmt=fmt, is_total=True)
    grand_total_row = row
    row += 2

    # Charts
    chart_start_row = row
    for i, (seg_name, seg_companies) in enumerate(SEGMENT_MAP):
        add_line_chart(
            ws, f"{title} - {seg_name}",
            all_seg_header_rows[seg_name], all_seg_data_rows[seg_name], seg_companies,
            min_col=2, max_col=1 + len(YEARS),
            chart_row=chart_start_row + (i // 2) * 16, chart_col=1 if i % 2 == 0 else 9,
            width=24, height=14
        )

    # Segment totals stacked bar
    bar_row = chart_start_row + 32
    ws.cell(row=bar_row, column=1, value="Segment")
    for i, yl in enumerate(YEAR_LABELS):
        ws.cell(row=bar_row, column=2 + i, value=yl)

    seg_bar_rows = []
    seg_bar_labels = []
    for seg_name, seg_companies in SEGMENT_MAP:
        bar_r = bar_row + len(seg_bar_rows) + 1
        ws.cell(row=bar_r, column=1, value=seg_name)
        seg_totals = compute_totals(data_dict, seg_companies)
        for i, v in enumerate(seg_totals):
            ws.cell(row=bar_r, column=2 + i, value=v)
        seg_bar_rows.append(bar_r)
        seg_bar_labels.append(seg_name)

    add_bar_chart(
        ws, f"Total {title}: All Segments",
        bar_row, seg_bar_rows, seg_bar_labels,
        min_col=2, max_col=1 + len(YEARS),
        chart_row=bar_row + 5, chart_col=1,
        width=28, height=14, stacked=True
    )

    ws.freeze_panes = "B1"
    return ws


def build_power_sheet(wb):
    """Build the Power Generation & Availability sheet."""
    ws = wb.create_sheet(title="Power & Generation")
    num_cols = 1 + len(YEARS)

    ws.column_dimensions['A'].width = 26
    for i in range(len(YEARS)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 12

    row = 1
    write_title_row(ws, row, "Power Capacity, Availability & Generation Mix", num_cols)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    note = ws.cell(row=row, column=1,
        value="IT Load (GW) | Power Contracted (GW) | Utilization (%) | Renewable Mix (%)  |  Yellow = Forecast")
    apply_cell_style(note, font=NOTE_FONT)
    row += 2

    # --- Section 1: IT Load (GW) ---
    write_section_header(ws, row, "IT Load Capacity (GW)", num_cols)
    row += 1
    header_row_gw = row
    write_column_headers(ws, row, ["Company"] + YEAR_LABELS)
    row += 1

    gw_rows = []
    for idx, company in enumerate(ALL_COMPANIES):
        write_data_row(ws, row, company, GW_CAPACITY_DATA[company], fmt="0.00", alt=(idx % 2 == 1))
        gw_rows.append(row)
        row += 1
    all_gw_totals = compute_totals(GW_CAPACITY_DATA, ALL_COMPANIES)
    write_data_row(ws, row, "TOTAL IT LOAD", all_gw_totals, fmt="0.0", is_total=True)
    row += 2

    # --- Section 2: Power Contracted (GW) ---
    write_section_header(ws, row, "Total Power Contracted / Available (GW)", num_cols)
    row += 1
    header_row_pwr = row
    write_column_headers(ws, row, ["Company"] + YEAR_LABELS)
    row += 1

    pwr_rows = []
    for idx, company in enumerate(ALL_COMPANIES):
        write_data_row(ws, row, company, POWER_AVAILABLE_DATA[company], fmt="0.00", alt=(idx % 2 == 1))
        pwr_rows.append(row)
        row += 1
    all_pwr_totals = compute_totals(POWER_AVAILABLE_DATA, ALL_COMPANIES)
    write_data_row(ws, row, "TOTAL CONTRACTED", all_pwr_totals, fmt="0.0", is_total=True)
    row += 2

    # --- Section 3: Power Utilization ---
    write_section_header(ws, row, "Power Utilization Rate (IT Load / Contracted)", num_cols)
    row += 1
    write_column_headers(ws, row, ["Company"] + YEAR_LABELS)
    row += 1

    for idx, company in enumerate(ALL_COMPANIES):
        utils = []
        for y in range(len(YEARS)):
            if POWER_AVAILABLE_DATA[company][y] > 0:
                utils.append(GW_CAPACITY_DATA[company][y] / POWER_AVAILABLE_DATA[company][y])
            else:
                utils.append(0)
        write_data_row(ws, row, company, utils, fmt="0.0%", alt=(idx % 2 == 1))
        row += 1
    row += 1

    # --- Section 4: Renewable Energy Mix ---
    write_section_header(ws, row, "Renewable Energy Mix (%)", num_cols)
    row += 1
    renew_header = row
    write_column_headers(ws, row, ["Company"] + YEAR_LABELS)
    row += 1

    renew_rows = []
    for idx, company in enumerate(ALL_COMPANIES):
        write_data_row(ws, row, company, RENEWABLE_PCT[company], fmt="0%", alt=(idx % 2 == 1))
        renew_rows.append(row)
        row += 1
    row += 2

    # Charts
    chart_row = row
    hyper_gw_rows = gw_rows[:len(HYPERSCALERS)]
    neo_gw_rows = gw_rows[len(HYPERSCALERS):len(HYPERSCALERS)+len(NEOCLOUDS)]
    colo_gw_rows = gw_rows[len(HYPERSCALERS)+len(NEOCLOUDS):]

    add_line_chart(ws, "IT Load (GW) - Hyperscalers", header_row_gw, hyper_gw_rows, HYPERSCALERS,
                   min_col=2, max_col=1+len(YEARS), chart_row=chart_row, chart_col=1, width=24, height=14)
    add_line_chart(ws, "IT Load (GW) - Neoclouds", header_row_gw, neo_gw_rows, NEOCLOUDS,
                   min_col=2, max_col=1+len(YEARS), chart_row=chart_row, chart_col=9, width=24, height=14)

    chart_row2 = chart_row + 16
    add_line_chart(ws, "IT Load (GW) - Colo / REITs", header_row_gw, colo_gw_rows, COLO_REITS,
                   min_col=2, max_col=1+len(YEARS), chart_row=chart_row2, chart_col=1, width=24, height=14)

    hyper_renew = renew_rows[:len(HYPERSCALERS)]
    add_line_chart(ws, "Renewable Mix (%) - Hyperscalers", renew_header, hyper_renew, HYPERSCALERS,
                   min_col=2, max_col=1+len(YEARS), chart_row=chart_row2, chart_col=9, width=24, height=14)

    ws.freeze_panes = "B1"
    return ws


def build_revenue_per_mw_sheet(wb):
    """Build Revenue/MW analysis sheet."""
    ws = wb.create_sheet(title="Revenue per MW")
    num_cols = 1 + len(YEARS)

    ws.column_dimensions['A'].width = 26
    for i in range(len(YEARS)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 12

    row = 1
    write_title_row(ws, row, "Revenue per MW Analysis ($M / MW)", num_cols)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    note = ws.cell(row=row, column=1,
        value="Revenue ($B) / IT Load (GW) = Revenue per MW ($M/MW)  |  Higher = more revenue-efficient  |  Yellow = Forecast")
    apply_cell_style(note, font=NOTE_FONT)
    row += 2

    # Revenue
    write_section_header(ws, row, "Revenue ($B)", num_cols)
    row += 1
    write_column_headers(ws, row, ["Company"] + YEAR_LABELS)
    row += 1
    for idx, company in enumerate(ALL_COMPANIES):
        write_data_row(ws, row, company, REVENUE_DATA[company], fmt="0.0", alt=(idx % 2 == 1))
        row += 1
    rev_totals = compute_totals(REVENUE_DATA, ALL_COMPANIES)
    write_data_row(ws, row, "TOTAL", rev_totals, fmt="0.0", is_total=True)
    row += 2

    # IT Load
    write_section_header(ws, row, "IT Load Capacity (GW)", num_cols)
    row += 1
    write_column_headers(ws, row, ["Company"] + YEAR_LABELS)
    row += 1
    for idx, company in enumerate(ALL_COMPANIES):
        write_data_row(ws, row, company, GW_CAPACITY_DATA[company], fmt="0.00", alt=(idx % 2 == 1))
        row += 1
    row += 1

    # Revenue / MW
    write_section_header(ws, row, "Revenue per MW ($M / MW)", num_cols)
    row += 1
    rpm_header = row
    write_column_headers(ws, row, ["Company"] + YEAR_LABELS)
    row += 1

    rpm_rows = []
    for idx, company in enumerate(ALL_COMPANIES):
        rpm_values = []
        for y in range(len(YEARS)):
            gw = GW_CAPACITY_DATA[company][y]
            rev = REVENUE_DATA[company][y]
            rpm_values.append(round(rev / gw, 2) if gw > 0 else 0)
        write_data_row(ws, row, company, rpm_values, fmt="#,##0.0", alt=(idx % 2 == 1))
        rpm_rows.append(row)
        row += 1
    row += 1

    # Capex / MW
    write_section_header(ws, row, "Capex per MW ($M / MW)", num_cols)
    row += 1
    cpm_header = row
    write_column_headers(ws, row, ["Company"] + YEAR_LABELS)
    row += 1

    cpm_rows = []
    for idx, company in enumerate(ALL_COMPANIES):
        cpm_values = []
        for y in range(len(YEARS)):
            gw = GW_CAPACITY_DATA[company][y]
            capex = CAPEX_DATA[company][y]
            cpm_values.append(round(capex / gw, 2) if gw > 0 else 0)
        write_data_row(ws, row, company, cpm_values, fmt="#,##0.0", alt=(idx % 2 == 1))
        cpm_rows.append(row)
        row += 1
    row += 2

    # Charts
    nh = len(HYPERSCALERS)
    nn = len(NEOCLOUDS)
    add_line_chart(ws, "Revenue / MW ($M/MW) - Hyperscalers", rpm_header, rpm_rows[:nh], HYPERSCALERS,
                   min_col=2, max_col=1+len(YEARS), chart_row=row, chart_col=1, width=24, height=14)
    add_line_chart(ws, "Revenue / MW ($M/MW) - Neoclouds", rpm_header, rpm_rows[nh:nh+nn], NEOCLOUDS,
                   min_col=2, max_col=1+len(YEARS), chart_row=row, chart_col=9, width=24, height=14)

    row2 = row + 16
    add_line_chart(ws, "Revenue / MW ($M/MW) - Colo / REITs", rpm_header, rpm_rows[nh+nn:], COLO_REITS,
                   min_col=2, max_col=1+len(YEARS), chart_row=row2, chart_col=1, width=24, height=14)
    add_line_chart(ws, "Capex / MW ($M/MW) - All Segments", cpm_header,
                   [cpm_rows[0], cpm_rows[nh], cpm_rows[nh+nn]],
                   [HYPERSCALERS[0], NEOCLOUDS[0], COLO_REITS[0]],
                   min_col=2, max_col=1+len(YEARS), chart_row=row2, chart_col=9, width=24, height=14)

    ws.freeze_panes = "B1"
    return ws


def build_power_supply_sheet(wb):
    """Build the Power Supply & Capex Ramp-Up sheet."""
    ws = wb.create_sheet(title="Power Supply & Ramp")
    num_cols = 1 + len(YEARS)

    ws.column_dimensions['A'].width = 36
    for i in range(len(YEARS)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 12

    row = 1
    write_title_row(ws, row, "Power Supply, Generation Ramp & Capex Build Timeline", num_cols)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    note = ws.cell(row=row, column=1,
        value="Drivers of power addition  |  Generation mix additions  |  Grid constraints  |  Build timelines  |  Yellow = Forecast")
    apply_cell_style(note, font=NOTE_FONT)
    row += 2

    # Section 1: US DC Power Demand vs Supply
    write_section_header(ws, row, "US Datacenter Power Demand & Grid Supply (GW)", num_cols)
    row += 1
    demand_header = row
    write_column_headers(ws, row, ["Metric"] + YEAR_LABELS)
    row += 1

    write_data_row(ws, row, "Total US DC Power Demand (GW)", US_DC_POWER_DEMAND, fmt="0.0")
    demand_row = row
    row += 1
    write_data_row(ws, row, "US Grid Uncommitted Capacity (GW)", US_GRID_UNCOMMITTED, fmt="0.0", alt=True)
    grid_row = row
    row += 1

    # Demand / Supply gap
    gap = [round(d - s, 1) for d, s in zip(US_DC_POWER_DEMAND, US_GRID_UNCOMMITTED)]
    write_data_row(ws, row, "Demand vs Uncommitted Gap (GW)", gap, fmt="0.0")
    gap_row = row
    row += 1

    # Incremental demand vs deliverable
    write_data_row(ws, row, "Incremental GW Needed (YoY)", GW_DEMAND_INCREMENT, fmt="0.0", alt=True)
    inc_demand_row = row
    row += 1
    write_data_row(ws, row, "GW Supply Deliverable (YoY)", GW_SUPPLY_DELIVERABLE, fmt="0.0")
    inc_supply_row = row
    row += 1

    supply_gap = [round(s - d, 1) for d, s in zip(GW_DEMAND_INCREMENT, GW_SUPPLY_DELIVERABLE)]
    write_data_row(ws, row, "Supply Surplus / (Deficit)", supply_gap, fmt="0.0", alt=True)
    row += 2

    # Section 2: New Generation Capacity Additions by Source
    write_section_header(ws, row, "New Power Generation Additions for DC Loads (GW/year by source)", num_cols)
    row += 1
    gen_header = row
    write_column_headers(ws, row, ["Generation Type"] + YEAR_LABELS)
    row += 1

    gen_rows = []
    gen_labels = []
    for idx, (gen_type, values) in enumerate(POWER_GEN_ADDITIONS.items()):
        write_data_row(ws, row, gen_type, values, fmt="0.0", alt=(idx % 2 == 1))
        gen_rows.append(row)
        gen_labels.append(gen_type)
        row += 1

    # Total additions
    total_gen = [0] * len(YEARS)
    for vals in POWER_GEN_ADDITIONS.values():
        for i, v in enumerate(vals):
            total_gen[i] += v
    total_gen = [round(t, 1) for t in total_gen]
    write_data_row(ws, row, "TOTAL NEW GENERATION", total_gen, fmt="0.0", is_total=True)
    total_gen_row = row
    row += 2

    # Section 3: Power Cost & Interconnection
    write_section_header(ws, row, "Power Cost & Interconnection Constraints", num_cols)
    row += 1
    pwr_cost_header = row
    write_column_headers(ws, row, ["Metric"] + YEAR_LABELS)
    row += 1

    write_data_row(ws, row, "Avg Power Cost to DC ($/MWh)", AVG_POWER_COST, fmt="$#,##0")
    pwr_cost_row = row
    row += 1
    write_data_row(ws, row, "Avg Interconnection Queue (months)", INTERCONNECT_QUEUE_MONTHS, fmt="#,##0", alt=True)
    queue_row = row
    row += 2

    # Section 4: DC Build Timeline
    write_section_header(ws, row, "Datacenter Build Timeline — Land to Live (months)", num_cols)
    row += 1
    timeline_header = row
    write_column_headers(ws, row, ["Phase"] + YEAR_LABELS)
    row += 1

    timeline_rows = []
    for idx, (phase, values) in enumerate(DC_BUILD_TIMELINE_MONTHS.items()):
        is_tot = (phase == "Total (Land to Live)")
        write_data_row(ws, row, phase, values, fmt="#,##0",
                       is_total=is_tot, alt=(idx % 2 == 1 and not is_tot))
        timeline_rows.append(row)
        row += 1
    row += 1

    # Section 5: Key Drivers of Power Addition
    write_section_header(ws, row, "Key Drivers of Datacenter Power Addition", num_cols)
    row += 1

    drivers = [
        ("AI/ML Training Clusters",
         "Massive GPU clusters (10k-100k GPUs) require 50-200MW each; driving bulk power demand",
         [0.5, 0.8, 1.2, 2.0, 4.0, 8.0, 18.0, 28.0, 38.0, 48.0, 55.0, 60.0, 65.0]),
        ("AI Inference at Scale",
         "Growing inference workloads as AI models deploy broadly; lower per-query but massive volume",
         [0.1, 0.2, 0.3, 0.5, 1.0, 2.0, 5.0, 10.0, 18.0, 28.0, 40.0, 52.0, 65.0]),
        ("Cloud IaaS / PaaS Growth",
         "Traditional cloud workloads continue steady growth; enterprise migration ongoing",
         [8.0, 9.5, 11.0, 13.0, 15.0, 17.0, 19.0, 21.0, 23.0, 25.0, 27.0, 29.0, 31.0]),
        ("Edge / Sovereign Cloud",
         "New edge locations and sovereign cloud mandates driving distributed capacity",
         [0.2, 0.3, 0.5, 0.8, 1.2, 2.0, 3.0, 4.5, 6.5, 9.0, 12.0, 15.0, 18.0]),
        ("Crypto / HPC / Other",
         "Cryptocurrency mining, HPC simulation, and miscellaneous high-density workloads",
         [1.5, 2.0, 3.0, 4.0, 5.0, 4.5, 4.0, 3.5, 3.5, 3.5, 4.0, 4.5, 5.0]),
    ]

    driver_header = row
    write_column_headers(ws, row, ["Power Demand Driver (GW)"] + YEAR_LABELS)
    row += 1

    driver_rows = []
    driver_labels = []
    for idx, (name, desc, values) in enumerate(drivers):
        write_data_row(ws, row, name, values, fmt="0.0", alt=(idx % 2 == 1))
        driver_rows.append(row)
        driver_labels.append(name)
        row += 1

    # Total from drivers
    driver_total = [0] * len(YEARS)
    for _, _, vals in drivers:
        for i, v in enumerate(vals):
            driver_total[i] += v
    driver_total = [round(t, 1) for t in driver_total]
    write_data_row(ws, row, "TOTAL DEMAND FROM DRIVERS", driver_total, fmt="0.0", is_total=True)
    row += 2

    # Charts
    chart_row = row

    # Demand vs Grid capacity
    add_line_chart(ws, "US DC Power Demand vs Grid Uncommitted (GW)",
                   demand_header, [demand_row, grid_row],
                   ["DC Demand", "Grid Uncommitted"],
                   min_col=2, max_col=1+len(YEARS),
                   chart_row=chart_row, chart_col=1, width=24, height=14)

    # Generation additions stacked bar
    add_bar_chart(ws, "New Power Generation Additions by Source (GW/yr)",
                  gen_header, gen_rows, gen_labels,
                  min_col=2, max_col=1+len(YEARS),
                  chart_row=chart_row, chart_col=9, width=24, height=14, stacked=True)

    chart_row2 = chart_row + 16

    # Build timeline
    add_line_chart(ws, "DC Build Timeline (months)",
                   timeline_header, timeline_rows[-1:], ["Total Land-to-Live"],
                   min_col=2, max_col=1+len(YEARS),
                   chart_row=chart_row2, chart_col=1, width=24, height=14)

    # Demand drivers stacked
    add_bar_chart(ws, "Power Demand by Driver (GW)",
                  driver_header, driver_rows, driver_labels,
                  min_col=2, max_col=1+len(YEARS),
                  chart_row=chart_row2, chart_col=9, width=24, height=14, stacked=True)

    chart_row3 = chart_row2 + 16

    # Incremental supply vs demand
    add_line_chart(ws, "Incremental GW Needed vs Deliverable (YoY)",
                   demand_header, [inc_demand_row, inc_supply_row],
                   ["GW Needed", "GW Deliverable"],
                   min_col=2, max_col=1+len(YEARS),
                   chart_row=chart_row3, chart_col=1, width=24, height=14)

    # Power cost trend
    add_line_chart(ws, "Average Power Cost to Datacenters ($/MWh)",
                   pwr_cost_header, [pwr_cost_row],
                   ["Avg $/MWh"],
                   min_col=2, max_col=1+len(YEARS),
                   chart_row=chart_row3, chart_col=9, width=24, height=14)

    ws.freeze_panes = "B1"
    ws.sheet_properties.tabColor = ACCENT_GREEN
    return ws


def build_bom_sheet(wb):
    """Build the Datacenter Bill of Materials sheet."""
    ws = wb.create_sheet(title="DC Bill of Materials")

    # This sheet has a different column layout
    headers = ["Category", "Sub-Category", "Cost/MW 2020 ($M)", "Cost/MW 2024 ($M)",
               "Cost/MW 2028E ($M)", "Cost/MW 2030E ($M)", "CAGR '20-'30E", "Key Vendors / Suppliers"]
    num_cols = len(headers)

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 52

    row = 1
    write_title_row(ws, row, "Datacenter Bill of Materials — Cost per MW of IT Load", num_cols)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    note = ws.cell(row=row, column=1,
        value="All costs in $M per MW of IT load  |  Represents typical hyperscale build  |  Includes equipment, labor, materials")
    apply_cell_style(note, font=NOTE_FONT)
    row += 2

    write_column_headers(ws, row, headers)
    header_row = row
    row += 1

    # Group by category
    current_cat = None
    cat_start_rows = {}
    data_rows_by_cat = {}

    for idx, (cat, sub, c20, c24, c28, c30, vendors) in enumerate(BOM_CATEGORIES):
        if cat != current_cat:
            current_cat = cat
            cat_start_rows[cat] = row
            data_rows_by_cat[cat] = []

        # Compute CAGR
        if c20 > 0:
            cagr = (c30 / c20) ** (1.0 / 10.0) - 1.0
        else:
            cagr = 0

        alt = (idx % 2 == 1)

        cell = ws.cell(row=row, column=1, value=cat if row == cat_start_rows.get(cat, -1) else "")
        apply_cell_style(cell, font=COMPANY_FONT if row == cat_start_rows.get(cat, -1) else DATA_FONT,
                         fill=ALT_ROW_FILL if alt else None, border=THIN_BORDER,
                         alignment=Alignment(horizontal="left", vertical="center"))

        cell = ws.cell(row=row, column=2, value=sub)
        apply_cell_style(cell, font=DATA_FONT, fill=ALT_ROW_FILL if alt else None,
                         border=THIN_BORDER, alignment=Alignment(horizontal="left"))

        for ci, val in enumerate([c20, c24, c28, c30], start=3):
            cell = ws.cell(row=row, column=ci, value=val)
            apply_cell_style(cell, font=DATA_FONT, fill=ALT_ROW_FILL if alt else None,
                             border=THIN_BORDER, number_format="$#,##0.0",
                             alignment=Alignment(horizontal="center"))

        cell = ws.cell(row=row, column=7, value=cagr)
        apply_cell_style(cell, font=DATA_FONT, fill=ALT_ROW_FILL if alt else None,
                         border=THIN_BORDER, number_format="0.0%",
                         alignment=Alignment(horizontal="center"))

        cell = ws.cell(row=row, column=8, value=vendors)
        apply_cell_style(cell, font=Font(name="Calibri", size=9, color="444444"),
                         fill=ALT_ROW_FILL if alt else None,
                         border=THIN_BORDER, alignment=Alignment(horizontal="left", wrap_text=True))

        data_rows_by_cat.setdefault(cat, []).append(row)
        row += 1

    row += 1

    # Totals
    write_section_header(ws, row, "Total Cost per MW Summary", num_cols)
    row += 1

    cost_years = ["2020", "2024", "2028E", "2030E"]
    cost_indices = [2, 6, 10, 12]  # indices into YEARS for reference

    totals_by_year = {y: 0 for y in cost_years}
    for cat, sub, c20, c24, c28, c30, vendors in BOM_CATEGORIES:
        totals_by_year["2020"] += c20
        totals_by_year["2024"] += c24
        totals_by_year["2028E"] += c28
        totals_by_year["2030E"] += c30

    # Write summary
    summary_headers = ["", "Category", "2020 ($M/MW)", "2024 ($M/MW)", "2028E ($M/MW)", "2030E ($M/MW)", "% of Total (2024)", ""]
    for i, h in enumerate(summary_headers):
        cell = ws.cell(row=row, column=1 + i, value=h)
        apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL, border=THIN_BORDER,
                         alignment=Alignment(horizontal="center"))
    row += 1

    # Aggregate by category
    cat_totals = {}
    for cat, sub, c20, c24, c28, c30, vendors in BOM_CATEGORIES:
        if cat not in cat_totals:
            cat_totals[cat] = [0, 0, 0, 0]
        cat_totals[cat][0] += c20
        cat_totals[cat][1] += c24
        cat_totals[cat][2] += c28
        cat_totals[cat][3] += c30

    summary_rows = []
    summary_labels = []
    total_2024 = totals_by_year["2024"]
    for idx, (cat, vals) in enumerate(cat_totals.items()):
        alt = idx % 2 == 1
        ws.cell(row=row, column=1, value="")
        cell = ws.cell(row=row, column=2, value=cat)
        apply_cell_style(cell, font=COMPANY_FONT, fill=ALT_ROW_FILL if alt else None,
                         border=THIN_BORDER, alignment=Alignment(horizontal="left"))
        for ci, v in enumerate(vals, start=3):
            cell = ws.cell(row=row, column=ci, value=round(v, 1))
            apply_cell_style(cell, font=DATA_FONT, fill=ALT_ROW_FILL if alt else None,
                             border=THIN_BORDER, number_format="$#,##0.0",
                             alignment=Alignment(horizontal="center"))
        pct = vals[1] / total_2024 if total_2024 > 0 else 0
        cell = ws.cell(row=row, column=7, value=pct)
        apply_cell_style(cell, font=DATA_FONT, fill=ALT_ROW_FILL if alt else None,
                         border=THIN_BORDER, number_format="0.0%",
                         alignment=Alignment(horizontal="center"))
        ws.cell(row=row, column=8, value="")
        summary_rows.append(row)
        summary_labels.append(cat)
        row += 1

    # Grand total
    ws.cell(row=row, column=1, value="")
    cell = ws.cell(row=row, column=2, value="TOTAL per MW")
    apply_cell_style(cell, font=TOTAL_FONT, fill=TOTAL_FILL, border=THIN_BORDER,
                     alignment=Alignment(horizontal="left"))
    for ci, yr in enumerate(cost_years, start=3):
        cell = ws.cell(row=row, column=ci, value=round(totals_by_year[yr], 1))
        apply_cell_style(cell, font=TOTAL_FONT, fill=TOTAL_FILL, border=THIN_BORDER,
                         number_format="$#,##0.0", alignment=Alignment(horizontal="center"))
    cell = ws.cell(row=row, column=7, value=1.0)
    apply_cell_style(cell, font=TOTAL_FONT, fill=TOTAL_FILL, border=THIN_BORDER,
                     number_format="0%", alignment=Alignment(horizontal="center"))
    ws.cell(row=row, column=8, value="")
    row += 2

    # Reference facility cost
    write_section_header(ws, row, "Reference Facility: 100 MW Hyperscale Datacenter — Total Build Cost ($M)", num_cols)
    row += 1
    ref_headers = ["", "Category", "2020 ($M)", "2024 ($M)", "2028E ($M)", "2030E ($M)", "% of Total (2024)", ""]
    for i, h in enumerate(ref_headers):
        cell = ws.cell(row=row, column=1 + i, value=h)
        apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL, border=THIN_BORDER,
                         alignment=Alignment(horizontal="center"))
    row += 1

    ref_mw = 100
    for idx, (cat, vals) in enumerate(cat_totals.items()):
        alt = idx % 2 == 1
        ws.cell(row=row, column=1, value="")
        cell = ws.cell(row=row, column=2, value=cat)
        apply_cell_style(cell, font=COMPANY_FONT, fill=ALT_ROW_FILL if alt else None,
                         border=THIN_BORDER, alignment=Alignment(horizontal="left"))
        for ci, v in enumerate(vals, start=3):
            cell = ws.cell(row=row, column=ci, value=round(v * ref_mw, 0))
            apply_cell_style(cell, font=DATA_FONT, fill=ALT_ROW_FILL if alt else None,
                             border=THIN_BORDER, number_format="$#,##0",
                             alignment=Alignment(horizontal="center"))
        pct = vals[1] / total_2024 if total_2024 > 0 else 0
        cell = ws.cell(row=row, column=7, value=pct)
        apply_cell_style(cell, font=DATA_FONT, fill=ALT_ROW_FILL if alt else None,
                         border=THIN_BORDER, number_format="0.0%",
                         alignment=Alignment(horizontal="center"))
        ws.cell(row=row, column=8, value="")
        row += 1

    ws.cell(row=row, column=1, value="")
    cell = ws.cell(row=row, column=2, value="TOTAL 100MW Facility")
    apply_cell_style(cell, font=TOTAL_FONT, fill=TOTAL_FILL, border=THIN_BORDER,
                     alignment=Alignment(horizontal="left"))
    for ci, yr in enumerate(cost_years, start=3):
        cell = ws.cell(row=row, column=ci, value=round(totals_by_year[yr] * ref_mw, 0))
        apply_cell_style(cell, font=TOTAL_FONT, fill=TOTAL_FILL, border=THIN_BORDER,
                         number_format="$#,##0", alignment=Alignment(horizontal="center"))
    cell = ws.cell(row=row, column=7, value=1.0)
    apply_cell_style(cell, font=TOTAL_FONT, fill=TOTAL_FILL, border=THIN_BORDER,
                     number_format="0%", alignment=Alignment(horizontal="center"))
    row += 2

    # BOM cost trend note
    write_section_header(ws, row, "Key BOM Trends", num_cols)
    row += 1
    trends = [
        "GPU server costs dominate and are rising: AI-optimized racks now 3-5x cost of traditional CPU racks",
        "Liquid cooling share growing rapidly: from <5% of cooling spend in 2020 to >50% by 2028E for AI-heavy facilities",
        "Power infrastructure costs rising due to transformer shortages and lead times extending to 2-3 years",
        "Total cost per MW has increased ~30% since 2020, driven primarily by GPU costs and power equipment",
        "Building shell costs relatively stable; modular/prefab construction helping offset labor inflation",
        "Networking costs rising with 400G/800G optics adoption for GPU cluster interconnects",
    ]
    for t in trends:
        ws.cell(row=row, column=1, value="•")
        cell = ws.cell(row=row, column=2, value=t)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=num_cols)
        apply_cell_style(cell, font=DATA_FONT, alignment=Alignment(wrap_text=True))
        row += 1
    row += 1

    # Chart: BOM category breakdown for 2024
    # Write helper data
    chart_data_row = row + 2
    ws.cell(row=chart_data_row, column=1, value="Category")
    ws.cell(row=chart_data_row, column=2, value="2024 Cost/MW ($M)")
    chart_rows_for_bar = []
    for idx, (cat, vals) in enumerate(cat_totals.items()):
        r = chart_data_row + 1 + idx
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=round(vals[1], 1))
        chart_rows_for_bar.append(r)

    chart = BarChart()
    chart.title = "BOM Cost Breakdown per MW (2024, $M)"
    chart.style = 10
    chart.width = 24
    chart.height = 14
    chart.type = "col"
    cats = Reference(ws, min_col=1, min_row=chart_data_row + 1,
                     max_row=chart_data_row + len(cat_totals))
    data = Reference(ws, min_col=2, min_row=chart_data_row,
                     max_row=chart_data_row + len(cat_totals))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, f"D{chart_data_row}")

    # Chart: 2020 vs 2024 vs 2030E comparison
    ws.cell(row=chart_data_row, column=4, value="Category")
    ws.cell(row=chart_data_row, column=5, value="2020")
    ws.cell(row=chart_data_row, column=6, value="2024")
    ws.cell(row=chart_data_row, column=7, value="2030E")
    comp_rows = []
    for idx, (cat, vals) in enumerate(cat_totals.items()):
        r = chart_data_row + 1 + idx
        ws.cell(row=r, column=4, value=cat)
        ws.cell(row=r, column=5, value=round(vals[0], 1))
        ws.cell(row=r, column=6, value=round(vals[1], 1))
        ws.cell(row=r, column=7, value=round(vals[3], 1))
        comp_rows.append(r)

    ws.freeze_panes = "C1"
    ws.sheet_properties.tabColor = ACCENT_ORANGE
    return ws


def build_capex_to_revenue_sheet(wb):
    """Build the Capex-to-Revenue Supply Chain sheet."""
    ws = wb.create_sheet(title="Capex to Revenue")
    num_cols = 1 + len(YEARS)

    ws.column_dimensions['A'].width = 36
    for i in range(len(YEARS)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 12

    row = 1
    write_title_row(ws, row, "Capex-to-Revenue: DC Supply Chain Analysis", num_cols)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    note = ws.cell(row=row, column=1,
        value="How datacenter capex spend translates to revenue across the supply chain  |  TAM by segment ($B)  |  Yellow = Forecast")
    apply_cell_style(note, font=NOTE_FONT)
    row += 2

    # Section 1: Total DC Capex (all segments)
    write_section_header(ws, row, "Total Datacenter Capex — All Operators ($B)", num_cols)
    row += 1
    capex_header = row
    write_column_headers(ws, row, ["Segment"] + YEAR_LABELS)
    row += 1

    seg_capex_rows = []
    for seg_name, seg_companies in SEGMENT_MAP:
        seg_totals = compute_totals(CAPEX_DATA, seg_companies)
        write_data_row(ws, row, seg_name, seg_totals, fmt="#,##0.0",
                       alt=(len(seg_capex_rows) % 2 == 1))
        seg_capex_rows.append(row)
        row += 1

    all_capex_totals = compute_totals(CAPEX_DATA, ALL_COMPANIES)
    write_data_row(ws, row, "TOTAL INDUSTRY CAPEX", all_capex_totals, fmt="#,##0.0", is_total=True)
    total_capex_row = row
    row += 2

    # Section 2: Capex Flow Breakdown (% allocation)
    write_section_header(ws, row, "Capex $ Flow — How Each $1 of DC Capex is Allocated", num_cols - len(YEARS) + 4)
    row += 1

    flow_headers = ["Supply Chain Segment", "2020", "2024", "2028E", "2030E"]
    for i, h in enumerate(flow_headers):
        cell = ws.cell(row=row, column=1 + i, value=h)
        apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL, border=THIN_BORDER,
                         alignment=Alignment(horizontal="center"))
    row += 1

    for idx, (seg, pcts) in enumerate(CAPEX_FLOW_PCT.items()):
        alt = idx % 2 == 1
        cell = ws.cell(row=row, column=1, value=seg)
        apply_cell_style(cell, font=COMPANY_FONT, fill=ALT_ROW_FILL if alt else None,
                         border=THIN_BORDER, alignment=Alignment(horizontal="left"))
        for ci, v in enumerate(pcts, start=2):
            cell = ws.cell(row=row, column=ci, value=v)
            apply_cell_style(cell, font=DATA_FONT, fill=ALT_ROW_FILL if alt else None,
                             border=THIN_BORDER, number_format="0.0%",
                             alignment=Alignment(horizontal="center"))
        row += 1
    row += 1

    # Section 3: Supply Chain TAM by Segment
    write_section_header(ws, row, "Supply Chain Revenue / TAM by Segment ($B)", num_cols)
    row += 1
    tam_header = row
    write_column_headers(ws, row, ["Supply Chain Segment"] + YEAR_LABELS)
    row += 1

    tam_rows = []
    tam_labels = []
    for idx, (seg, values) in enumerate(SUPPLY_CHAIN_REVENUE.items()):
        write_data_row(ws, row, seg, values, fmt="#,##0.0", alt=(idx % 2 == 1))
        tam_rows.append(row)
        tam_labels.append(seg)
        row += 1

    # Total TAM
    total_tam = [0] * len(YEARS)
    for vals in SUPPLY_CHAIN_REVENUE.values():
        for i, v in enumerate(vals):
            total_tam[i] += v
    total_tam = [round(t, 1) for t in total_tam]
    write_data_row(ws, row, "TOTAL DC SUPPLY CHAIN TAM", total_tam, fmt="#,##0.0", is_total=True)
    total_tam_row = row
    row += 2

    # Section 4: TAM YoY Growth
    write_section_header(ws, row, "Supply Chain TAM — YoY Growth Rate", num_cols)
    row += 1
    growth_header = row
    write_column_headers(ws, row, ["Supply Chain Segment"] + YEAR_LABELS)
    row += 1

    growth_rows = []
    for idx, (seg, values) in enumerate(SUPPLY_CHAIN_REVENUE.items()):
        growth_vals = []
        for y in range(len(YEARS)):
            if y == 0 or values[y-1] == 0:
                growth_vals.append(0)
            else:
                growth_vals.append(round((values[y] - values[y-1]) / values[y-1], 3))
        write_data_row(ws, row, seg, growth_vals, fmt="0.0%", alt=(idx % 2 == 1))
        growth_rows.append(row)
        row += 1
    row += 1

    # Section 5: Key Beneficiaries
    write_section_header(ws, row, "Key Vendor Beneficiaries by Supply Chain Segment", num_cols)
    row += 1

    ben_headers = ["Supply Chain Segment", "Key Vendors / Beneficiaries"]
    for i, h in enumerate(ben_headers):
        cols_for_h = 1 if i == 0 else num_cols - 1
        if i == 1:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=num_cols)
        cell = ws.cell(row=row, column=1 + i, value=h)
        apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL, border=THIN_BORDER,
                         alignment=Alignment(horizontal="left" if i == 1 else "center"))
    row += 1

    for idx, (seg, vendors) in enumerate(SUPPLY_CHAIN_VENDORS.items()):
        alt = idx % 2 == 1
        cell = ws.cell(row=row, column=1, value=seg)
        apply_cell_style(cell, font=COMPANY_FONT, fill=ALT_ROW_FILL if alt else None,
                         border=THIN_BORDER, alignment=Alignment(horizontal="left"))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=num_cols)
        cell = ws.cell(row=row, column=2, value=vendors)
        apply_cell_style(cell, font=DATA_FONT, fill=ALT_ROW_FILL if alt else None,
                         border=THIN_BORDER, alignment=Alignment(horizontal="left", wrap_text=True))
        row += 1
    row += 2

    # Section 6: Implied Revenue per $ of Capex
    write_section_header(ws, row, "Supply Chain Revenue Multiplier (TAM / Total Capex)", num_cols)
    row += 1
    write_column_headers(ws, row, ["Metric"] + YEAR_LABELS)
    row += 1

    multiplier = [round(t / c, 2) if c > 0 else 0 for t, c in zip(total_tam, all_capex_totals)]
    write_data_row(ws, row, "TAM / Industry Capex Ratio", multiplier, fmt="0.00x")
    mult_row = row
    row += 1

    # Capex-to-Revenue lag analysis
    write_data_row(ws, row, "Industry Capex ($B)", all_capex_totals, fmt="#,##0.0", alt=True)
    capex_row2 = row
    row += 1
    write_data_row(ws, row, "Supply Chain TAM ($B)", total_tam, fmt="#,##0.0")
    tam_row2 = row
    row += 2

    # Charts
    chart_row = row

    # TAM by segment stacked bar
    add_bar_chart(ws, "DC Supply Chain TAM by Segment ($B)",
                  tam_header, tam_rows, tam_labels,
                  min_col=2, max_col=1+len(YEARS),
                  chart_row=chart_row, chart_col=1, width=28, height=15, stacked=True)

    # Capex vs TAM
    add_line_chart(ws, "Industry Capex vs Supply Chain TAM ($B)",
                   capex_header, [total_capex_row, total_tam_row],
                   ["Total Capex", "Supply Chain TAM"],
                   min_col=2, max_col=1+len(YEARS),
                   chart_row=chart_row, chart_col=10, width=24, height=15)

    chart_row2 = chart_row + 17

    # Servers & GPUs TAM growth
    add_line_chart(ws, "Servers & GPUs TAM ($B) — Largest Segment",
                   tam_header, tam_rows[:1], ["Servers & GPUs"],
                   min_col=2, max_col=1+len(YEARS),
                   chart_row=chart_row2, chart_col=1, width=24, height=14)

    # Growth rates
    top_segs = list(SUPPLY_CHAIN_REVENUE.keys())[:4]
    add_line_chart(ws, "Supply Chain Segment YoY Growth Rates",
                   growth_header, growth_rows[:4], top_segs,
                   min_col=2, max_col=1+len(YEARS),
                   chart_row=chart_row2, chart_col=10, width=24, height=14)

    ws.freeze_panes = "B1"
    ws.sheet_properties.tabColor = ACCENT_PURPLE
    return ws


def build_dashboard(wb):
    """Build the summary dashboard sheet as the first sheet."""
    ws = wb.create_sheet(title="Dashboard", index=0)
    num_cols = 16

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 28
    for c in range(3, num_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value="DATACENTER INFRASTRUCTURE MODEL — EXECUTIVE DASHBOARD")
    apply_cell_style(cell, font=Font(name="Calibri", size=18, bold=True, color=WHITE),
                     fill=TITLE_FILL, alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[row].height = 40
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1,
        value="Hyperscalers, Neoclouds & Colo/REITs  |  2018-2030E  |  Capex, Servers, DCs, Power, BOM, Supply Chain")
    apply_cell_style(cell, font=Font(name="Calibri", size=10, color="999999"),
                     alignment=Alignment(horizontal="center"))
    row += 2

    # Key Metrics Summary
    write_section_header(ws, row, "Key Metrics Summary (All Companies)", num_cols)
    row += 1

    metrics = [
        ("Total Capex ($B)", CAPEX_DATA, "0.0"),
        ("Total Servers (000s)", SERVER_DATA, "#,##0"),
        ("Total Datacenters", DC_COUNT_DATA, "#,##0"),
        ("Total IT Load (GW)", GW_CAPACITY_DATA, "0.0"),
        ("Total Power Contracted (GW)", POWER_AVAILABLE_DATA, "0.0"),
        ("Total Revenue ($B)", REVENUE_DATA, "0.0"),
    ]

    write_column_headers(ws, row, ["", "Metric"] + YEAR_LABELS, start_col=1)
    row += 1

    for idx, (label, data, fmt) in enumerate(metrics):
        totals = compute_totals(data, ALL_COMPANIES)
        ws.cell(row=row, column=1, value="")
        write_data_row(ws, row, label, totals, start_col=2, fmt=fmt, alt=(idx % 2 == 1))
        row += 1

    # Avg Revenue/MW
    rev_totals = compute_totals(REVENUE_DATA, ALL_COMPANIES)
    gw_totals = compute_totals(GW_CAPACITY_DATA, ALL_COMPANIES)
    avg_rpm = [round(r/g, 1) if g > 0 else 0 for r, g in zip(rev_totals, gw_totals)]
    ws.cell(row=row, column=1, value="")
    write_data_row(ws, row, "Avg Revenue/MW ($M/MW)", avg_rpm, start_col=2, fmt="0.0", alt=True)
    row += 1

    pwr_totals = compute_totals(POWER_AVAILABLE_DATA, ALL_COMPANIES)
    avg_util = [round(g/p, 3) if p > 0 else 0 for g, p in zip(gw_totals, pwr_totals)]
    ws.cell(row=row, column=1, value="")
    write_data_row(ws, row, "Avg Power Utilization", avg_util, start_col=2, fmt="0.0%")
    row += 2

    # Segment Capex Split
    write_section_header(ws, row, "Capex by Segment ($B)", num_cols)
    row += 1
    write_column_headers(ws, row, ["", "Segment"] + YEAR_LABELS, start_col=1)
    row += 1

    seg_capex = {}
    for seg_name, seg_companies in SEGMENT_MAP:
        seg_capex[seg_name] = compute_totals(CAPEX_DATA, seg_companies)
        ws.cell(row=row, column=1, value="")
        write_data_row(ws, row, seg_name, seg_capex[seg_name], start_col=2, fmt="0.0",
                       alt=(seg_name == "Neocloud Vendors"))
        row += 1

    total_capex = compute_totals(CAPEX_DATA, ALL_COMPANIES)
    ws.cell(row=row, column=1, value="")
    write_data_row(ws, row, "Total", total_capex, start_col=2, fmt="0.0", is_total=True)
    row += 1

    # Neocloud share
    neo_capex = seg_capex["Neocloud Vendors"]
    neo_share = [round(n/t, 3) if t > 0 else 0 for n, t in zip(neo_capex, total_capex)]
    ws.cell(row=row, column=1, value="")
    write_data_row(ws, row, "Neocloud Share %", neo_share, start_col=2, fmt="0.0%", alt=True)
    row += 1

    colo_capex = seg_capex["Colocation / DC REITs"]
    colo_share = [round(c/t, 3) if t > 0 else 0 for c, t in zip(colo_capex, total_capex)]
    ws.cell(row=row, column=1, value="")
    write_data_row(ws, row, "Colo/REIT Share %", colo_share, start_col=2, fmt="0.0%")
    row += 2

    # GW capacity split
    write_section_header(ws, row, "IT Load by Segment (GW)", num_cols)
    row += 1
    write_column_headers(ws, row, ["", "Segment"] + YEAR_LABELS, start_col=1)
    row += 1

    seg_gw = {}
    for seg_name, seg_companies in SEGMENT_MAP:
        seg_gw[seg_name] = compute_totals(GW_CAPACITY_DATA, seg_companies)
        ws.cell(row=row, column=1, value="")
        write_data_row(ws, row, seg_name, seg_gw[seg_name], start_col=2, fmt="0.0",
                       alt=(seg_name == "Neocloud Vendors"))
        row += 1

    total_gw = compute_totals(GW_CAPACITY_DATA, ALL_COMPANIES)
    ws.cell(row=row, column=1, value="")
    write_data_row(ws, row, "Total", total_gw, start_col=2, fmt="0.0", is_total=True)
    row += 2

    # Supply Chain TAM
    write_section_header(ws, row, "DC Supply Chain TAM ($B)", num_cols)
    row += 1
    write_column_headers(ws, row, ["", "Segment"] + YEAR_LABELS, start_col=1)
    row += 1

    total_sc_tam = [0] * len(YEARS)
    for seg, vals in SUPPLY_CHAIN_REVENUE.items():
        ws.cell(row=row, column=1, value="")
        write_data_row(ws, row, seg, vals, start_col=2, fmt="#,##0")
        for i, v in enumerate(vals):
            total_sc_tam[i] += v
        row += 1
    total_sc_tam = [round(t, 0) for t in total_sc_tam]
    ws.cell(row=row, column=1, value="")
    write_data_row(ws, row, "Total TAM", total_sc_tam, start_col=2, fmt="#,##0", is_total=True)
    row += 2

    # Charts
    chart_row = row
    helper_row = 200

    # Chart 1: Total Capex stacked bar (3 segments)
    ws.cell(row=helper_row, column=2, value="Year")
    for i, yl in enumerate(YEAR_LABELS):
        ws.cell(row=helper_row, column=3+i, value=yl)
    chart_seg_rows = []
    chart_seg_labels = []
    for idx, (seg_name, _) in enumerate(SEGMENT_MAP):
        r = helper_row + 1 + idx
        ws.cell(row=r, column=2, value=seg_name)
        for i, v in enumerate(seg_capex[seg_name]):
            ws.cell(row=r, column=3+i, value=v)
        chart_seg_rows.append(r)
        chart_seg_labels.append(seg_name)

    add_bar_chart(ws, "Total Capex ($B) by Segment", helper_row, chart_seg_rows, chart_seg_labels,
                  min_col=3, max_col=2+len(YEARS), chart_row=chart_row, chart_col=2,
                  width=26, height=14, stacked=True)

    # Chart 2: IT Load stacked bar
    hr2 = helper_row + 5
    ws.cell(row=hr2, column=2, value="Year")
    for i, yl in enumerate(YEAR_LABELS):
        ws.cell(row=hr2, column=3+i, value=yl)
    gw_chart_rows = []
    for idx, (seg_name, _) in enumerate(SEGMENT_MAP):
        r = hr2 + 1 + idx
        ws.cell(row=r, column=2, value=seg_name)
        for i, v in enumerate(seg_gw[seg_name]):
            ws.cell(row=r, column=3+i, value=v)
        gw_chart_rows.append(r)

    add_bar_chart(ws, "Total IT Load (GW) by Segment", hr2, gw_chart_rows, chart_seg_labels,
                  min_col=3, max_col=2+len(YEARS), chart_row=chart_row, chart_col=10,
                  width=26, height=14, stacked=True)

    # Chart 3: Revenue/MW trend
    chart_row2 = chart_row + 16
    hr3 = hr2 + 5
    ws.cell(row=hr3, column=2, value="Year")
    for i, yl in enumerate(YEAR_LABELS):
        ws.cell(row=hr3, column=3+i, value=yl)
    ws.cell(row=hr3+1, column=2, value="Avg Rev/MW")
    for i, v in enumerate(avg_rpm):
        ws.cell(row=hr3+1, column=3+i, value=v)

    add_line_chart(ws, "Average Revenue per MW ($M/MW)", hr3, [hr3+1], ["Avg Rev/MW"],
                   min_col=3, max_col=2+len(YEARS), chart_row=chart_row2, chart_col=2,
                   width=26, height=14)

    # Chart 4: Supply Chain TAM
    hr4 = hr3 + 3
    ws.cell(row=hr4, column=2, value="Year")
    for i, yl in enumerate(YEAR_LABELS):
        ws.cell(row=hr4, column=3+i, value=yl)
    ws.cell(row=hr4+1, column=2, value="Total Capex")
    for i, v in enumerate(total_capex):
        ws.cell(row=hr4+1, column=3+i, value=v)
    ws.cell(row=hr4+2, column=2, value="Supply Chain TAM")
    for i, v in enumerate(total_sc_tam):
        ws.cell(row=hr4+2, column=3+i, value=v)

    add_bar_chart(ws, "Industry Capex vs Supply Chain TAM ($B)", hr4, [hr4+1, hr4+2],
                  ["Total Capex", "Supply Chain TAM"],
                  min_col=3, max_col=2+len(YEARS), chart_row=chart_row2, chart_col=10,
                  width=26, height=14)

    ws.freeze_panes = "C5"
    ws.sheet_properties.tabColor = ACCENT_BLUE
    return ws


def build_capex_to_ai_revenue_sheet(wb):
    """Build the Capex-to-AI-Revenue Conversion & Payback Analysis sheet."""
    ws = wb.create_sheet(title="Capex to AI Revenue")
    num_cols = 1 + len(YEARS)

    ws.column_dimensions['A'].width = 30
    for i in range(len(YEARS)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 12

    row = 1
    write_title_row(ws, row, "When Does AI Capex Translate to AI Revenue? — Company-by-Company Analysis", num_cols)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    note = ws.cell(row=row, column=1,
        value="AI Capex ($B) | AI Revenue ($B) | Cumulative Payback | AI Rev / AI Capex Ratio | AI as % of Total  |  Yellow = Forecast")
    apply_cell_style(note, font=NOTE_FONT)
    row += 2

    # ====================================================================
    # SECTION 1: AI Capex ($B)
    # ====================================================================
    write_section_header(ws, row, "AI-Specific Capex ($B) — Portion of Total Capex Allocated to AI/ML Infrastructure", num_cols)
    row += 1
    ai_capex_header = row
    write_column_headers(ws, row, ["Company"] + YEAR_LABELS)
    row += 1

    ai_capex_rows = []
    # Hyperscalers first, then neoclouds, then colo
    ordered_companies = HYPERSCALERS + NEOCLOUDS + COLO_REITS
    for idx, company in enumerate(ordered_companies):
        write_data_row(ws, row, company, AI_CAPEX_DATA[company], fmt="#,##0.0", alt=(idx % 2 == 1))
        ai_capex_rows.append(row)
        row += 1

    ai_capex_totals = [0.0] * len(YEARS)
    for c in ordered_companies:
        for i, v in enumerate(AI_CAPEX_DATA[c]):
            ai_capex_totals[i] += v
    ai_capex_totals = [round(t, 1) for t in ai_capex_totals]
    write_data_row(ws, row, "TOTAL AI CAPEX", ai_capex_totals, fmt="#,##0.0", is_total=True)
    row += 2

    # ====================================================================
    # SECTION 2: AI Revenue ($B)
    # ====================================================================
    write_section_header(ws, row, "AI-Specific Revenue ($B) — Revenue Directly Attributable to AI Products & Services", num_cols)
    row += 1
    ai_rev_header = row
    write_column_headers(ws, row, ["Company"] + YEAR_LABELS)
    row += 1

    ai_rev_rows = []
    for idx, company in enumerate(ordered_companies):
        write_data_row(ws, row, company, AI_REVENUE_DATA[company], fmt="#,##0.0", alt=(idx % 2 == 1))
        ai_rev_rows.append(row)
        row += 1

    ai_rev_totals = [0.0] * len(YEARS)
    for c in ordered_companies:
        for i, v in enumerate(AI_REVENUE_DATA[c]):
            ai_rev_totals[i] += v
    ai_rev_totals = [round(t, 1) for t in ai_rev_totals]
    write_data_row(ws, row, "TOTAL AI REVENUE", ai_rev_totals, fmt="#,##0.0", is_total=True)
    row += 2

    # ====================================================================
    # SECTION 3: AI Capex as % of Total Capex
    # ====================================================================
    write_section_header(ws, row, "AI Capex as % of Total Capex — Shows AI Investment Intensity", num_cols)
    row += 1
    ai_pct_header = row
    write_column_headers(ws, row, ["Company"] + YEAR_LABELS)
    row += 1

    ai_pct_rows = []
    for idx, company in enumerate(ordered_companies):
        pct_vals = []
        for y in range(len(YEARS)):
            total = CAPEX_DATA[company][y]
            ai = AI_CAPEX_DATA[company][y]
            pct_vals.append(round(ai / total, 3) if total > 0 else 0)
        write_data_row(ws, row, company, pct_vals, fmt="0.0%", alt=(idx % 2 == 1))
        ai_pct_rows.append(row)
        row += 1
    row += 1

    # ====================================================================
    # SECTION 4: AI Revenue as % of Total Revenue
    # ====================================================================
    write_section_header(ws, row, "AI Revenue as % of Total Revenue — AI Monetization Penetration", num_cols)
    row += 1
    ai_rev_pct_header = row
    write_column_headers(ws, row, ["Company"] + YEAR_LABELS)
    row += 1

    ai_rev_pct_rows = []
    for idx, company in enumerate(ordered_companies):
        pct_vals = []
        for y in range(len(YEARS)):
            total_rev = REVENUE_DATA[company][y]
            ai_rev = AI_REVENUE_DATA[company][y]
            pct_vals.append(round(ai_rev / total_rev, 3) if total_rev > 0 else 0)
        write_data_row(ws, row, company, pct_vals, fmt="0.0%", alt=(idx % 2 == 1))
        ai_rev_pct_rows.append(row)
        row += 1
    row += 1

    # ====================================================================
    # SECTION 5: Annual AI Revenue / AI Capex Ratio (conversion efficiency)
    # ====================================================================
    write_section_header(ws, row, "Annual AI Revenue / AI Capex Ratio — Conversion Efficiency (>1.0x = generating more than spending)", num_cols)
    row += 1
    ratio_header = row
    write_column_headers(ws, row, ["Company"] + YEAR_LABELS)
    row += 1

    ratio_rows = []
    for idx, company in enumerate(ordered_companies):
        ratio_vals = []
        for y in range(len(YEARS)):
            ai_c = AI_CAPEX_DATA[company][y]
            ai_r = AI_REVENUE_DATA[company][y]
            ratio_vals.append(round(ai_r / ai_c, 2) if ai_c > 0 else 0)
        write_data_row(ws, row, company, ratio_vals, fmt="0.00x", alt=(idx % 2 == 1))
        ratio_rows.append(row)
        row += 1
    row += 1

    # ====================================================================
    # SECTION 6: Cumulative AI Capex vs Cumulative AI Revenue & Payback
    # ====================================================================
    write_section_header(ws, row,
        "Cumulative AI Capex vs AI Revenue ($B) & Payback Year — When Does Cumulative AI Revenue Exceed Cumulative AI Capex?", num_cols)
    row += 1

    # For each company, compute cumulative and find payback year
    payback_summary = []  # (company, payback_year_label, cum_capex_at_payback, cum_rev_at_payback)

    cum_capex_header = row
    write_column_headers(ws, row, ["Company — Cumulative AI Capex"] + YEAR_LABELS)
    row += 1
    cum_capex_rows = []
    for idx, company in enumerate(ordered_companies):
        cum = []
        running = 0
        for y in range(len(YEARS)):
            running += AI_CAPEX_DATA[company][y]
            cum.append(round(running, 1))
        write_data_row(ws, row, company, cum, fmt="#,##0.0", alt=(idx % 2 == 1))
        cum_capex_rows.append(row)
        row += 1
    row += 1

    cum_rev_header = row
    write_column_headers(ws, row, ["Company — Cumulative AI Revenue"] + YEAR_LABELS)
    row += 1
    cum_rev_rows = []
    for idx, company in enumerate(ordered_companies):
        cum = []
        running = 0
        for y in range(len(YEARS)):
            running += AI_REVENUE_DATA[company][y]
            cum.append(round(running, 1))
        write_data_row(ws, row, company, cum, fmt="#,##0.0", alt=(idx % 2 == 1))
        cum_rev_rows.append(row)
        row += 1
    row += 1

    # Net cumulative (AI Rev - AI Capex) and payback year
    net_header = row
    write_column_headers(ws, row, ["Company — Cum. Net (Rev - Capex)"] + YEAR_LABELS)
    row += 1
    net_rows = []
    for idx, company in enumerate(ordered_companies):
        cum_c = 0
        cum_r = 0
        net_vals = []
        payback_year = None
        for y in range(len(YEARS)):
            cum_c += AI_CAPEX_DATA[company][y]
            cum_r += AI_REVENUE_DATA[company][y]
            net = round(cum_r - cum_c, 1)
            net_vals.append(net)
            if payback_year is None and net >= 0 and cum_c > 1.0:
                payback_year = YEAR_LABELS[y]
        write_data_row(ws, row, company, net_vals, fmt="#,##0.0", alt=(idx % 2 == 1))
        net_rows.append(row)

        payback_summary.append((company, payback_year if payback_year else "Post-2030",
                                round(cum_c, 1), round(cum_r, 1)))
        row += 1
    row += 2

    # ====================================================================
    # SECTION 7: Payback Summary Table
    # ====================================================================
    write_section_header(ws, row,
        "AI Capex Payback Summary — When Cumulative AI Revenue Exceeds Cumulative AI Capex", num_cols)
    row += 1

    pb_headers = ["Company", "Payback Year", "Avg Capex-to-Rev Lag (Qtrs)",
                  "Cum AI Capex 2030E ($B)", "Cum AI Rev 2030E ($B)",
                  "2030E Net ($B)", "2024 AI Rev/Capex", "2027E AI Rev/Capex", "2030E AI Rev/Capex"]
    for i, h in enumerate(pb_headers):
        cell = ws.cell(row=row, column=1 + i, value=h)
        apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL, border=THIN_BORDER,
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))
    ws.row_dimensions[row].height = 32
    row += 1

    payback_data_rows = []
    for idx, (company, pb_year, cum_c, cum_r) in enumerate(payback_summary):
        alt = idx % 2 == 1

        # Compute full cumulative through 2030
        full_cum_c = sum(AI_CAPEX_DATA[company])
        full_cum_r = sum(AI_REVENUE_DATA[company])
        net_2030 = round(full_cum_r - full_cum_c, 1)

        # Ratios at specific years
        # 2024 = index 6, 2027 = index 9, 2030 = index 12
        def ratio_at(yr_idx):
            c = AI_CAPEX_DATA[company][yr_idx]
            r = AI_REVENUE_DATA[company][yr_idx]
            return round(r / c, 2) if c > 0 else 0

        r_2024 = ratio_at(6)
        r_2027 = ratio_at(9)
        r_2030 = ratio_at(12)

        lag = AI_CAPEX_LAG_QUARTERS.get(company, 6)

        vals = [company, pb_year, lag, round(full_cum_c, 1), round(full_cum_r, 1),
                net_2030, r_2024, r_2027, r_2030]

        for ci, v in enumerate(vals):
            cell = ws.cell(row=row, column=1 + ci, value=v)
            f = DATA_FONT
            if ci == 0:
                f = COMPANY_FONT
            elif ci == 1:
                # Color code payback year
                if pb_year and "Post" not in str(pb_year):
                    yr_num = int(str(pb_year).replace("E", ""))
                    if yr_num <= 2025:
                        f = Font(name="Calibri", size=10, bold=True, color="006100")  # green
                    elif yr_num <= 2027:
                        f = Font(name="Calibri", size=10, bold=True, color="9C5700")  # amber
                    else:
                        f = Font(name="Calibri", size=10, bold=True, color="C00000")  # red
                else:
                    f = Font(name="Calibri", size=10, bold=True, color="C00000")

            fmt = "#,##0.0"
            if ci == 1:
                fmt = "@"  # text
            elif ci == 2:
                fmt = "#,##0"
            elif ci >= 6:
                fmt = "0.00x"

            apply_cell_style(cell, font=f, fill=ALT_ROW_FILL if alt else None,
                             border=THIN_BORDER, number_format=fmt,
                             alignment=Alignment(horizontal="center", vertical="center"))
        payback_data_rows.append(row)
        row += 1
    row += 2

    # ====================================================================
    # SECTION 8: Key Insights / Commentary
    # ====================================================================
    write_section_header(ws, row, "Key Insights — AI Capex-to-Revenue Translation", num_cols)
    row += 1

    insights = [
        "FASTEST PAYBACK (HYPERSCALERS) — Microsoft (Azure): Strong OpenAI/Copilot pull-through gives the shortest capex-to-revenue lag among hyperscalers (~5 qtrs); cumulative AI revenue overtakes cumulative AI capex by 2028E; annual AI Rev/Capex reaches 1.96x by 2030E",
        "AMAZON (AWS): Massive AI capex ramp ($35B in 2024, growing to $106B by 2030E); revenue follows ~6 qtrs later via Bedrock, SageMaker, Trainium/Inferentia; annual ratio crosses 1.0x by ~2029E but cumulative payback extends past 2030 due to front-loaded investment",
        "GOOGLE (GCP): Longer enterprise sales cycle (~7 qtrs) but deep TPU moat and Gemini platform; annual AI Rev/Capex crosses 1.0x around 2029E; cumulative payback post-2030 — similar to AWS, massive capex base is hard to recoup quickly",
        "META: Unique model — AI capex translates to ad revenue uplift rather than direct cloud AI sales; ~3 quarter lag as recommendation/targeting models improve; earliest cumulative payback (2022) among all companies; 2024 ratio already >1.0x",
        "ORACLE CLOUD: Aggressive AI infrastructure buildout (OCI GPU superclusters); AI capex ramped from $3.5B (2023) to $9B (2024) to a projected $24B by 2030E; longer enterprise sales cycle (~8 qtrs) but Oracle's existing DB/ERP customer base provides distribution advantage; cumulative payback by 2029E with annual AI revenue reaching $42B by 2030E — the fastest-improving annual ratio among hyperscalers after Microsoft",
        "APPLE: Lowest AI capex intensity but also the longest lag (~10 qtrs) due to hardware product cycles; AI investment manifests in Apple Intelligence, on-device ML, and enhanced services revenue; high annual AI Rev/Capex by 2030E (4.3x) because of modest capex relative to massive services uplift",
        "NEOCLOUDS: Nearly 100% AI-focused; short lag (4-5 qtrs) from GPU procurement to lease revenue; CoreWeave leads with fastest absolute ramp but cumulative payback extends past 2030 due to extreme capital intensity; annual ratios approaching 0.9x by 2030E — profitability inflection expected 2030-2031",
        "COLO/REITs: AI revenue = premium leasing for high-density GPU facilities; Equinix achieves cumulative payback by 2025E due to lower capex and premium AI tenant pricing; Digital Realty by 2029E; QTS and Vantage still building out through 2030",
        "THE KEY QUESTION: Annual AI Rev/Capex >1.0x means the company is generating more AI revenue than it spends on AI capex that year — a critical profitability threshold; most hyperscalers cross this by 2028-2030E, but cumulative payback takes longer due to the massive 2024-2027 buildout",
        "RISK: If AI demand growth slows or models become more compute-efficient (reducing GPU needs), payback timelines extend; current estimates assume sustained demand growth; these are gross capex recovery metrics — do not account for opex, depreciation, or cost of capital",
    ]

    for ins in insights:
        ws.cell(row=row, column=1, value="•")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=num_cols)
        cell = ws.cell(row=row, column=2, value=ins)
        apply_cell_style(cell, font=Font(name="Calibri", size=9),
                         alignment=Alignment(wrap_text=True, vertical="top"))
        ws.row_dimensions[row].height = 30
        row += 1
    row += 2

    # ====================================================================
    # CHARTS
    # ====================================================================
    chart_row = row

    # Chart 1: AI Capex vs AI Revenue — Hyperscalers (line chart, 2 series per company)
    # Use totals for simplicity
    # Hyperscaler AI Capex total vs AI Revenue total
    hyper_ai_capex = [0.0] * len(YEARS)
    hyper_ai_rev = [0.0] * len(YEARS)
    for c in HYPERSCALERS:
        for i in range(len(YEARS)):
            hyper_ai_capex[i] += AI_CAPEX_DATA[c][i]
            hyper_ai_rev[i] += AI_REVENUE_DATA[c][i]
    hyper_ai_capex = [round(t, 1) for t in hyper_ai_capex]
    hyper_ai_rev = [round(t, 1) for t in hyper_ai_rev]

    helper_row = row + 60  # below visible area
    ws.cell(row=helper_row, column=1, value="Year")
    for i, yl in enumerate(YEAR_LABELS):
        ws.cell(row=helper_row, column=2 + i, value=yl)

    ws.cell(row=helper_row + 1, column=1, value="Hyperscaler AI Capex")
    for i, v in enumerate(hyper_ai_capex):
        ws.cell(row=helper_row + 1, column=2 + i, value=v)
    ws.cell(row=helper_row + 2, column=1, value="Hyperscaler AI Revenue")
    for i, v in enumerate(hyper_ai_rev):
        ws.cell(row=helper_row + 2, column=2 + i, value=v)

    add_line_chart(ws, "Hyperscaler AI Capex vs AI Revenue ($B)",
                   helper_row, [helper_row + 1, helper_row + 2],
                   ["AI Capex", "AI Revenue"],
                   min_col=2, max_col=1 + len(YEARS),
                   chart_row=chart_row, chart_col=1, width=24, height=14)

    # Chart 2: Selected company AI Revenue lines (AWS, Azure, Google, Meta, Oracle, CoreWeave)
    focus_companies = ["Amazon (AWS)", "Microsoft (Azure)", "Google (GCP)", "Meta", "Oracle Cloud", "CoreWeave"]
    focus_indices = [ordered_companies.index(c) for c in focus_companies]
    focus_rev_rows = [ai_rev_rows[i] for i in focus_indices]

    add_line_chart(ws, "AI Revenue ($B) — Key Companies incl. Oracle",
                   ai_rev_header, focus_rev_rows, focus_companies,
                   min_col=2, max_col=1 + len(YEARS),
                   chart_row=chart_row, chart_col=9, width=24, height=14)

    chart_row2 = chart_row + 16

    # Chart 3: Annual AI Rev / AI Capex ratio for key companies
    focus_ratio_rows = [ratio_rows[i] for i in focus_indices]
    add_line_chart(ws, "AI Revenue / AI Capex Ratio — Key Companies (>1.0x = payback)",
                   ratio_header, focus_ratio_rows, focus_companies,
                   min_col=2, max_col=1 + len(YEARS),
                   chart_row=chart_row2, chart_col=1, width=24, height=14)

    # Chart 4: Cumulative Net (Rev - Capex) for key companies
    focus_net_rows = [net_rows[i] for i in focus_indices]
    add_line_chart(ws, "Cumulative Net AI (Revenue - Capex, $B) — Payback Crossover",
                   net_header, focus_net_rows, focus_companies,
                   min_col=2, max_col=1 + len(YEARS),
                   chart_row=chart_row2, chart_col=9, width=24, height=14)

    chart_row3 = chart_row2 + 16

    # Chart 5: AI Capex as % of total capex for key companies
    focus_pct_rows = [ai_pct_rows[i] for i in focus_indices]
    add_line_chart(ws, "AI Capex as % of Total Capex — Key Companies",
                   ai_pct_header, focus_pct_rows, focus_companies,
                   min_col=2, max_col=1 + len(YEARS),
                   chart_row=chart_row3, chart_col=1, width=24, height=14)

    # Chart 6: AI Revenue as % of total revenue for key companies
    focus_revpct_rows = [ai_rev_pct_rows[i] for i in focus_indices]
    add_line_chart(ws, "AI Revenue as % of Total Revenue — Key Companies",
                   ai_rev_pct_header, focus_revpct_rows, focus_companies,
                   min_col=2, max_col=1 + len(YEARS),
                   chart_row=chart_row3, chart_col=9, width=24, height=14)

    ws.freeze_panes = "B1"
    ws.sheet_properties.tabColor = "C00000"  # Deep red for emphasis
    return ws


def build_assumptions_sheet(wb):
    """Build an Assumptions & Sources sheet."""
    ws = wb.create_sheet(title="Assumptions & Sources")

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 90

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value="MODEL ASSUMPTIONS & METHODOLOGY")
    apply_cell_style(cell, font=TITLE_FONT, fill=TITLE_FILL,
                     alignment=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[row].height = 32
    row += 2

    assumptions = [
        ("Data Sources", [
            "Historical capex and revenue from public SEC filings (10-K, 10-Q) and earnings calls",
            "Server counts estimated from industry reports (Synergy Research, Dell'Oro, Omdia)",
            "Datacenter counts from company announcements, press releases, and analyst estimates",
            "Power capacity from sustainability reports, PPA announcements, and utility filings",
            "Neocloud data from funding announcements, press coverage, and industry estimates",
            "Colo/REIT data from NAREIT filings, investor presentations, and quarterly supplements",
            "Power supply data from EIA, FERC, ISO/RTO reports, and utility IRP filings",
            "BOM costs from construction industry benchmarks, vendor quotes, and DC operator disclosures",
            "Supply chain TAM from Gartner, IDC, Dell'Oro, and company-reported revenue segmentation",
        ]),
        ("Key Assumptions — Forecasts (2025E-2030E)", [
            "AI infrastructure buildout continues to accelerate through 2027, moderating thereafter",
            "Hyperscaler capex growth driven by AI/ML workloads, sovereign cloud, and edge expansion",
            "Neocloud vendors benefit from GPU-as-a-service demand but face capital constraints",
            "Colo/REIT capex accelerates as hyperscalers increasingly lease rather than build",
            "Power availability is the key constraint on datacenter buildout after 2026",
            "Renewable energy targets: most hyperscalers reach 100% by 2027-2030",
            "Server density improves 8-12% annually (more compute per rack unit)",
            "Average PUE improves from ~1.3 (2024) to ~1.15 (2030) industry-wide",
            "SMR/advanced nuclear begins contributing meaningfully after 2028",
            "Transformer lead times peak in 2025-2026, gradually improve as manufacturing scales",
        ]),
        ("Power Supply Assumptions", [
            "Natural gas remains the primary source for near-term DC power additions through 2027",
            "Solar + battery storage combinations become cost-competitive for 24/7 DC power by 2028",
            "Nuclear (SMR) first commercial deployments for DC use begin in 2027-2028",
            "Grid interconnection queue times peak at ~42 months in 2025E, declining as processes streamline",
            "Average DC power cost peaks at ~$72/MWh in 2025E before declining with renewables scale",
            "Behind-the-meter generation (on-site solar, fuel cells) grows from <5% to ~15% by 2030",
        ]),
        ("Bill of Materials Assumptions", [
            "BOM costs reflect average hyperscale build; enterprise/colo builds may differ by 20-30%",
            "GPU server costs assume mix of NVIDIA H100/B200/next-gen accelerators at typical deployment ratios",
            "Liquid cooling penetration grows from ~10% of new builds (2024) to ~60% (2030E) for AI-heavy facilities",
            "Electrical infrastructure costs include transformer premiums from current supply shortage",
            "Construction labor cost inflation of ~3-5% annually, partially offset by modular/prefab methods",
        ]),
        ("Supply Chain Assumptions", [
            "TAM figures represent addressable market for DC-specific spend (not total vendor revenue)",
            "Servers & GPUs segment is the largest and fastest-growing, driven by AI accelerator demand",
            "Supply chain TAM exceeds operator capex because it includes maintenance, refresh, and software",
            "Revenue multiplier (TAM/Capex) reflects that capex flows through multiple vendor layers",
            "Key supply chain bottlenecks: GPU supply, transformer manufacturing, skilled labor, power",
        ]),
        ("AI Capex-to-Revenue Assumptions", [
            "AI Capex = portion of total capex allocated to GPU clusters, AI networking, liquid cooling for AI, AI-dedicated DCs",
            "AI Revenue = revenue directly attributable to AI: cloud AI services (hyperscalers), GPU-as-a-service (neoclouds), AI tenant leases (colo/REITs)",
            "Meta AI revenue represents estimated uplift to advertising revenue from AI-powered recommendation and targeting improvements",
            "Apple AI revenue reflects Apple Intelligence services, Siri improvements, and AI-enhanced services revenue uplift",
            "Oracle AI revenue driven by OCI GPU superclusters, Autonomous Database AI features, and AI cloud infrastructure",
            "Capex-to-revenue lag varies by business model: neoclouds ~4 qtrs (direct GPU lease), hyperscalers ~5-8 qtrs (build→GA→ramp), colo ~6-7 qtrs (lease negotiation)",
            "Cumulative payback year = first year where cumulative AI revenue >= cumulative AI capex since inception",
            "Annual AI Rev/Capex ratio >1.0x indicates that year's AI revenue exceeds that year's AI capex investment",
            "Payback analysis does not account for depreciation, opex, or cost of capital; represents gross capex recovery only",
        ]),
        ("Colocation / REIT Assumptions", [
            "Equinix, Digital Realty data from public REIT filings and investor supplements",
            "CyrusOne (KKR/GIP) and QTS (Blackstone) data estimated post-acquisition from industry reports",
            "Vantage and Switch data from press releases, funding announcements, and analyst estimates",
            "Colo revenue is facility/leasing revenue; does not include cloud service revenue",
            "DC counts for Colo/REITs include all owned and operated facilities globally",
            "Server counts for Colo/REITs represent customer-deployed servers hosted in their facilities",
        ]),
        ("Definitions", [
            "IT Load (GW): Total electrical power consumed by IT equipment (servers, storage, networking)",
            "Power Contracted (GW): Total power capacity secured via PPAs, utility contracts, and on-site generation",
            "Power Utilization: IT Load / Power Contracted (higher = more efficient use of secured power)",
            "Revenue/MW: Annual revenue ($B) / IT Load (GW); result is in $M/MW (since 1 GW = 1000 MW)",
            "Capex/MW: Annual capex ($B) / IT Load (GW) — investment intensity per MW of IT capacity",
            "Servers (000s): Total server count in thousands across all datacenter locations",
            "TAM: Total Addressable Market — revenue opportunity for vendors in each supply chain segment",
            "BOM: Bill of Materials — component and construction cost breakdown per MW of IT load",
            "AI Capex: Portion of total capex allocated specifically to AI/ML infrastructure (GPU clusters, AI networking, liquid cooling)",
            "AI Revenue: Revenue directly attributable to AI products and services (cloud AI, GPU-as-a-service, AI tenant leases)",
            "Cumulative Payback: Year when cumulative AI revenue first exceeds cumulative AI capex since inception",
            "AI Rev/Capex Ratio: Annual AI revenue divided by annual AI capex; >1.0x means AI revenue exceeds AI investment in that year",
        ]),
        ("Caveats", [
            "All neocloud figures are estimates based on limited public disclosure",
            "Company-specific DC counts may include leased/colocation facilities",
            "Apple revenue is total company revenue (not cloud-specific); DC figures are for internal use",
            "Meta revenue is total company revenue; DC infrastructure supports ads + AI workloads",
            "CyrusOne and QTS are now private; post-acquisition data are estimates",
            "Forecasts represent base-case scenario; upside/downside cases not included",
            "Currency: All figures in USD",
        ]),
    ]

    for section_title, items in assumptions:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=section_title)
        apply_cell_style(cell, font=SECTION_FONT, fill=LIGHT_FILL,
                         alignment=Alignment(horizontal="left"))
        ws.row_dimensions[row].height = 22
        row += 1

        for item in items:
            ws.cell(row=row, column=1, value="•")
            cell = ws.cell(row=row, column=2, value=item)
            apply_cell_style(cell, font=DATA_FONT)
            row += 1
        row += 1

    ws.sheet_properties.tabColor = "999999"
    return ws


# ============================================================
# MAIN BUILD
# ============================================================

def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("Building Dashboard...")
    build_dashboard(wb)

    print("Building Capex sheet...")
    build_data_sheet(wb, "Capex Spend", "Capital Expenditure", CAPEX_DATA, "$B", fmt="#,##0.0")

    print("Building Revenue sheet...")
    build_data_sheet(wb, "Revenue", "Revenue", REVENUE_DATA, "$B", fmt="#,##0.0")

    print("Building Servers sheet...")
    build_data_sheet(wb, "Servers", "Server Count", SERVER_DATA, "000s", fmt="#,##0")

    print("Building Datacenters sheet...")
    build_data_sheet(wb, "Datacenters", "Datacenter Count", DC_COUNT_DATA, "Facilities", fmt="#,##0")

    print("Building GW Capacity sheet...")
    build_data_sheet(wb, "GW Capacity", "IT Load Capacity", GW_CAPACITY_DATA, "GW", fmt="0.00")

    print("Building Power & Generation sheet...")
    build_power_sheet(wb)

    print("Building Revenue per MW sheet...")
    build_revenue_per_mw_sheet(wb)

    print("Building Power Supply & Ramp sheet...")
    build_power_supply_sheet(wb)

    print("Building DC Bill of Materials sheet...")
    build_bom_sheet(wb)

    print("Building Capex to Revenue sheet...")
    build_capex_to_revenue_sheet(wb)

    print("Building Capex to AI Revenue sheet...")
    build_capex_to_ai_revenue_sheet(wb)

    print("Building Assumptions sheet...")
    build_assumptions_sheet(wb)

    output_path = "/workspace/Datacenter_Infrastructure_Model.xlsx"
    wb.save(output_path)
    print(f"\nModel saved to: {output_path}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
