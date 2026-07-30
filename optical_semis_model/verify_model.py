"""Verify the LibreOffice-recalculated workbook: no formula errors, sane values."""

import openpyxl

wb = openpyxl.load_workbook("recalc/Optical_Semiconductor_Supply_Chain_Model.xlsx", data_only=True)

ERRS = ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#N/A", "#NULL!", "#NUM!", "Err:")
problems = []
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and any(e in c.value for e in ERRS):
                problems.append(f"{ws.title}!{c.coordinate}: {c.value!r}")

print(f"Formula errors found: {len(problems)}")
for p in problems[:20]:
    print("  ", p)

mm = wb["Market Model"]


def find_row(ws, text, col=1):
    for row in ws.iter_rows(min_col=col, max_col=col):
        c = row[0]
        if isinstance(c.value, str) and text in c.value:
            return c.row
    return None


checks = []

r = find_row(mm, "Total datacom optical ports")
checks.append(("MM total ports 2024A (M)", mm.cell(r, 4).value, 20, 35))
r = find_row(mm, "Mix check")
for col in range(3, 9):
    v = mm.cell(r, col).value
    checks.append((f"MM mix check col{col} == 1", v, 0.999, 1.001))
r = find_row(mm, "Total datacom transceiver market")
checks.append(("MM transceiver mkt 2024A ($M)", mm.cell(r, 4).value, 8000, 25000))
checks.append(("MM transceiver mkt 2028E ($M)", mm.cell(r, 8).value, 30000, 90000))
r = find_row(mm, "Net pluggable transceiver market")
plug28 = mm.cell(r, 8).value
checks.append(("MM net pluggable 2028E ($M)", plug28, 30000, 90000))
r = find_row(mm, "Total TAM")
checks.append(("MM total TAM 2028E ($M)", mm.cell(r, 8).value, 50000, 120000))
r = find_row(mm, "AI/datacom pluggables CAGR")
checks.append(("MM AI CAGR '25-'28", mm.cell(r, 3).value, 0.10, 0.60))
r = find_row(mm, "Fab optics CAGR")
checks.append(("MM fab CAGR '25-'28", mm.cell(r, 3).value, 0.02, 0.20))

cp = wb["Comps"]
r = find_row(cp, "LITE")
checks.append(("Comps LITE EV ($M)", cp.cell(r, 8).value, 40000, 70000))
checks.append(("Comps LITE EV/S (x)", cp.cell(r, 15).value, 5, 40))
r = find_row(cp, "ASML")
checks.append(("Comps ASML rev TTM ($M, EUR conv)", cp.cell(r, 9).value, 30000, 55000))
r = find_row(cp, "Median - Universe", col=2)
checks.append(("Comps universe median EV/S", cp.cell(r, 15).value, 2, 30))
checks.append(("Comps universe median GM", cp.cell(r, 11).value, 0.2, 0.8))

ex = wb["Company Exposure"]
r = find_row(ex, "300308.SZ")
checks.append(("Expo Innolight TTM rev ($M)", ex.cell(r, 4).value, 3000, 15000))
checks.append(("Expo Innolight 2028E rev ($M)", ex.cell(r, 14).value, 5000, 40000))
checks.append(("Expo Innolight implied CAGR", ex.cell(r, 15).value, 0.05, 0.8))
checks.append(("Expo Innolight EV/28E rev", ex.cell(r, 17).value, 1, 40))

ts = wb["Thematic Scores"]
r = find_row(ts, "COHR")
comp_col = 4 + 8
checks.append(("Scores COHR composite", ts.cell(r, comp_col).value, 0, 5))
checks.append(("Scores COHR rank int", ts.cell(r, comp_col + 1).value, 1, 30))
r = find_row(ts, "Weight ->", col=3)
checks.append(("Scores weights sum", ts.cell(r, 12).value, 0.999, 1.001))

sn = wb["Sensitivity"]
r = find_row(sn, "GRID 1")
center = sn.cell(r + 4, 5).value  # attach 3.5 row, ASP $1000 col
checks.append(("Sens grid1 (3.5x, $1000) $B", center, 30, 90))
r2 = find_row(sn, "GRID 2")
center2 = sn.cell(r2 + 4, 4).value  # 18M units, 52% mix
checks.append(("Sens grid2 (18M, 52%) $B", center2, 30, 90))
# grid2 model-case cell should be close to gross market/1000
gross = None
rg = find_row(mm, "Total datacom transceiver market")
gross = mm.cell(rg, 8).value / 1000
checks.append(("Sens grid2 model-case vs MM gross", center2 / gross if gross else None, 0.98, 1.02))

n_fail = 0
for name, val, lo, hi in checks:
    ok = isinstance(val, (int, float)) and lo <= val <= hi
    if not ok:
        n_fail += 1
    print(f"{'OK  ' if ok else 'FAIL'} {name}: {val}")

print(f"\n{len(checks) - n_fail}/{len(checks)} checks passed; {len(problems)} formula errors")
