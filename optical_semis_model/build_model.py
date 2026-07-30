"""Build the Optical x Semiconductor Supply Chain Excel model.

Reads market_data.json (produced by pull_data.py) and writes
Optical_Semiconductor_Supply_Chain_Model.xlsx with 8 tabs:

  1. Read Me            - purpose, tab guide, conventions
  2. Universe           - 30-name universe with segmentation & qualitative detail
  3. Supply Chain Map   - value-chain stage mapping (datacenter + fab-equipment chains)
  4. Market Model       - driver-based AI/datacom optics + fab-optics TAM model (formulas)
  5. Comps              - live valuation / financial comps with segment medians (formulas)
  6. Company Exposure   - revenue exposure model & 2028E scenario per name (formulas)
  7. Thematic Scores    - weighted thematic scoring matrix (formulas)
  8. Sensitivity        - 2028E market sensitivity grids (formulas)

Convention: blue font = input/assumption a user can flex; black = formula or hard data.
"""

import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# Styles
# ----------------------------------------------------------------------------
NAVY = "1F3864"
LIGHT_BLUE = "D9E2F3"
BAND_BLUE = "BDD7EE"
GREY = "F2F2F2"
INPUT_BLUE = "0070C0"

F_TITLE = Font(size=16, bold=True, color=NAVY)
F_SUB = Font(size=10, italic=True, color="595959")
F_HDR = Font(bold=True, color="FFFFFF", size=10)
F_BAND = Font(bold=True, color=NAVY, size=10)
F_SECTION = Font(bold=True, color=NAVY, size=11)
F_LABEL = Font(size=10)
F_LABEL_B = Font(size=10, bold=True)
F_INPUT = Font(size=10, color=INPUT_BLUE)
F_INPUT_B = Font(size=10, color=INPUT_BLUE, bold=True)
F_CALC = Font(size=10)
F_CALC_B = Font(size=10, bold=True)

FILL_HDR = PatternFill("solid", start_color=NAVY)
FILL_BAND = PatternFill("solid", start_color=BAND_BLUE)
FILL_TOTAL = PatternFill("solid", start_color=LIGHT_BLUE)
FILL_GREY = PatternFill("solid", start_color=GREY)

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_C = Alignment(wrap_text=True, vertical="center", horizontal="center")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

# Number formats
NF_M = "#,##0"          # $ millions
NF_B = "#,##0.0"        # $ billions
NF_PCT = "0.0%"
NF_PCT0 = "0%"
NF_MULT = '0.0"x"'
NF_PX = "#,##0.00"
NF_UNITS = "#,##0.0"
NF_USD = "$#,##0"
NF_SCORE = "0.0"

YEARS = ["2023A", "2024A", "2025E", "2026E", "2027E", "2028E"]

# ----------------------------------------------------------------------------
# Universe definition (ordered by segment)
# ----------------------------------------------------------------------------
SEG_TRX = "Datacom Transceivers & Optical Components"
SEG_CHIP = "Optical Semiconductors (DSP / PHY / TIA / PIC)"
SEG_LITHO = "Lithography & Light-Source Optics"
SEG_INSP = "Optical Inspection & Metrology"
SEG_MATL = "Substrates, Materials & Lasers"
SEG_SYS = "Optical Systems & EMS"

SEG_SHORT = {
    SEG_TRX: "Transceivers",
    SEG_CHIP: "Optical Semis",
    SEG_LITHO: "Litho Optics",
    SEG_INSP: "Inspection",
    SEG_MATL: "Materials",
    SEG_SYS: "Systems/EMS",
}

# exposure = (% AI/datacom optics, % semi-fab optics, % telecom optics) of revenue (estimates, inputs)
# adj = company-specific share-gain/(loss) adj to AI-bucket growth, in pts (input)
# scores = [AI DC optics, 1.6T/3.2T leverage, CPO/SiPh, EUV/litho, Inspection, Telecom recovery,
#           B/S & profitability, Risk (5 = low risk)]
UNIVERSE = [
    dict(t="COHR", name="Coherent Corp.", ctry="US", seg=SEG_TRX,
         sub="Vertically integrated photonics (InP lasers -> transceivers)",
         role="Makes 800G/1.6T datacom transceivers, EML/CW lasers for SiPh, and optics/lasers used in semi cap equipment; also SiC materials.",
         products="800G/1.6T transceivers, EML & CW lasers, VCSELs, optical circuit switches, fab lasers",
         customers="Nvidia ecosystem, hyperscalers (MSFT/META/GOOG/AWS), telecom OEMs, semi cap OEMs",
         thesis="One of two western vertically-integrated AI optics plays; indium-phosphide capacity is the scarce asset.",
         exp=(45, 5, 15), adj=0.0, scores=[5, 5, 4, 1, 1, 3, 3, 3]),
    dict(t="LITE", name="Lumentum Holdings", ctry="US", seg=SEG_TRX,
         sub="Lasers & photonics; datacom components -> modules pivot",
         role="EMLs and CW lasers that power 800G+/1.6T modules (incl. merchant sales to Chinese module makers); expanding own module + CPO laser sources.",
         products="EML/CW/DFB lasers, 800G/1.6T modules, ROADMs/WSS, 3D-sensing VCSELs",
         customers="Chinese transceiver makers, hyperscalers, Ciena/telecom OEMs, Apple (3D sensing)",
         thesis="Laser-chip bottleneck owner: sells picks-and-shovels to every 1.6T module vendor; external laser source play for CPO.",
         exp=(55, 5, 25), adj=1.0, scores=[5, 5, 5, 1, 1, 4, 3, 3]),
    dict(t="AAOI", name="Applied Optoelectronics", ctry="US", seg=SEG_TRX,
         sub="Datacenter transceivers (vertically integrated, US/TW/CN fabs)",
         role="400G/800G transceivers to hyperscalers; 100G lasers; CATV amplifiers. Re-shoring US capacity narrative.",
         products="400G/800G/1.6T transceivers, DFB/EML lasers, CATV/HFC gear",
         customers="Microsoft, Amazon (historical), CATV MSOs",
         thesis="High-beta share-recovery story on 800G qualification at US hyperscalers; execution & dilution risk.",
         exp=(85, 0, 5), adj=-2.0, scores=[4, 4, 2, 0, 0, 1, 1, 2]),
    dict(t="FN", name="Fabrinet", ctry="Thailand", seg=SEG_TRX,
         sub="Precision optical EMS (manufactures others' optics)",
         role="Contract manufacturer of optical transceivers/engines - notably Nvidia's 800G/1.6T OSFP modules and coherent modules for Ciena/others.",
         products="Optical module assembly & test, optical engines, sensors, automotive optics",
         customers="Nvidia (largest), Cisco/Acacia, Lumentum, Ciena, Infinera/Nokia",
         thesis="Cleanest 'volume' play on AI optics units with fab-lite risk; Nvidia concentration is the swing factor.",
         exp=(65, 0, 25), adj=0.0, scores=[5, 5, 3, 0, 0, 3, 4, 2]),
    dict(t="300308.SZ", name="Zhongji Innolight", ctry="China", seg=SEG_TRX,
         sub="#1 global datacom transceiver vendor by share",
         role="Largest supplier of 800G (and first-wave 1.6T) modules to US hyperscalers and Nvidia ecosystem.",
         products="800G/1.6T optical modules, SiPh modules, coherent modules",
         customers="Nvidia, Google, Meta, Amazon, Microsoft",
         thesis="Volume leader of the AI optics cycle; geopolitical/tariff risk against dominant share and cost position.",
         exp=(90, 0, 5), adj=1.0, scores=[5, 5, 4, 0, 0, 2, 4, 2]),
    dict(t="300502.SZ", name="Eoptolink", ctry="China", seg=SEG_TRX,
         sub="Fast-follower transceiver vendor",
         role="800G/1.6T modules incl. silicon photonics variants; gaining share at US and Chinese cloud customers.",
         products="800G/1.6T modules (EML & SiPh), 400G, AOCs",
         customers="Nvidia ecosystem, hyperscalers, Chinese cloud (Alibaba/Tencent/ByteDance)",
         thesis="Fastest-growing of the Chinese module trio into the 1.6T generation; similar geopolitical overhang.",
         exp=(92, 0, 4), adj=2.0, scores=[5, 5, 4, 0, 0, 1, 4, 2]),
    dict(t="300394.SZ", name="Suzhou TFC Optical", ctry="China", seg=SEG_TRX,
         sub="Passive optical components & optical engines",
         role="Fiber-array units (FAU), lenses, AWGs and passive assemblies that go inside transceivers and CPO/optical-engine designs.",
         products="FAUs, MT/MPO assemblies, lenses, AWG, WDM components, optical engines",
         customers="Innolight/Eoptolink & global module makers, SiPh/CPO programs",
         thesis="Content-per-port winner as SiPh and CPO raise passive/fiber-coupling content; supplier to everyone.",
         exp=(80, 0, 12), adj=1.0, scores=[4, 5, 5, 0, 0, 2, 4, 3]),
    dict(t="002281.SZ", name="Accelink Technologies", ctry="China", seg=SEG_TRX,
         sub="State-linked Chinese optics incumbent (telecom-heavy)",
         role="China domestic optical components/modules champion across telecom + datacom; localization beneficiary.",
         products="Telecom lasers/detectors, 400G/800G datacom modules, amplifiers",
         customers="Huawei, ZTE, China carriers, Chinese cloud",
         thesis="Play on China network capex + domestic substitution; less exposed to US AI capex than the private-sector trio.",
         exp=(45, 0, 45), adj=-1.0, scores=[3, 3, 2, 0, 0, 4, 3, 3]),

    dict(t="MRVL", name="Marvell Technology", ctry="US", seg=SEG_CHIP,
         sub="PAM4 optical DSP leader + custom AI silicon",
         role="Dominant merchant PAM4 DSP franchise (inherited from Inphi) inside 800G/1.6T modules; coherent DSPs (COLORZ/ZR); custom XPUs.",
         products="PAM4 DSPs, 800ZR/coherent DSPs, TIAs/drivers, custom AI ASICs, switches",
         customers="All major module makers, hyperscalers (custom silicon)",
         thesis="Every pluggable needs a DSP - Marvell taxes the whole module ecosystem; LPO/LRO and Broadcom are the share risks.",
         exp=(20, 0, 5), adj=0.0, scores=[5, 5, 4, 0, 0, 3, 4, 3]),
    dict(t="AVGO", name="Broadcom", ctry="US", seg=SEG_CHIP,
         sub="Switch silicon + DSP + first-mover CPO",
         role="Tomahawk switch ASICs set the port cadence that drives optics demand; sells optical DSPs, VCSELs/EMLs; shipping co-packaged optics (Bailly/TH6).",
         products="Tomahawk/Jericho switch ASICs, PAM4 DSPs, CPO optical engines, custom XPUs",
         customers="Hyperscalers, Arista/OEMs, module makers",
         thesis="Controls the switch roadmap that paces every optics upgrade; best-positioned if CPO displaces pluggables (small % of revenue today).",
         exp=(5, 0, 1), adj=0.0, scores=[4, 4, 5, 0, 0, 2, 5, 4]),
    dict(t="CRDO", name="Credo Technology", ctry="US", seg=SEG_CHIP,
         sub="SerDes house: AECs, optical DSPs, LRO",
         role="Active electrical cables (AEC) for in-rack + optical DSPs incl. low-power LRO variants that compete on power/cost in AI clusters.",
         products="AECs, PAM4 optical DSPs, LRO/linear optics DSPs, retimers, chiplets",
         customers="Hyperscalers (MSFT/AMZN/xAI et al.), module partners",
         thesis="Power-efficiency angle on AI interconnect; AECs cannibalize short optics but optical DSP line ties it to the module chain.",
         exp=(25, 0, 0), adj=2.0, scores=[4, 4, 3, 0, 0, 1, 4, 2]),
    dict(t="MTSI", name="MACOM Technology", ctry="US", seg=SEG_CHIP,
         sub="Analog/photonic semis: TIAs, drivers, lasers",
         role="High-performance analog front-ends (TIA/driver) and laser/photodiode chips for 800G+/1.6T and linear-drive (LPO) optics.",
         products="TIAs, modulator drivers, CDRs, lasers/PDs, RF/microwave (defense)",
         customers="Module makers, telecom OEMs, defense primes",
         thesis="Linear-drive optics needs better analog, not DSPs - MACOM wins either architecture; diversified by defense/industrial.",
         exp=(35, 5, 10), adj=0.5, scores=[4, 4, 4, 0, 0, 3, 4, 4]),
    dict(t="SMTC", name="Semtech", ctry="US", seg=SEG_CHIP,
         sub="FiberEdge TIAs/drivers + CopperEdge",
         role="TIA/driver content in 800G/1.6T modules (FiberEdge); CopperEdge linear redrivers for ACCs; LoRa/IoT rest of business.",
         products="FiberEdge TIAs & drivers, CopperEdge, LoRa, protection ICs",
         customers="Module makers, hyperscaler cable programs, IoT",
         thesis="Under-appreciated analog content per optical port; data-center now the growth engine after LoRa/IoT reset.",
         exp=(25, 0, 5), adj=0.5, scores=[4, 4, 3, 0, 0, 2, 2, 3]),
    dict(t="ALAB", name="Astera Labs", ctry="US", seg=SEG_CHIP,
         sub="Connectivity silicon (PCIe/CXL/Ethernet retimers -> optical)",
         role="Retimers/smart cables inside AI servers; scale-up fabric silicon (Scorpio) with a roadmap into optical interconnect (PCIe-over-optics).",
         products="Aries retimers, Taurus AECs, Leo CXL, Scorpio fabric switches",
         customers="Nvidia-based server OEMs, hyperscalers, AMD platforms",
         thesis="Adjacent today (mostly electrical), but the designated optical on-ramp is PCIe/scale-up over optics - watch Aries optical variants.",
         exp=(5, 0, 0), adj=3.0, scores=[3, 3, 3, 0, 0, 0, 5, 3]),
    dict(t="POET", name="POET Technologies", ctry="Canada", seg=SEG_CHIP,
         sub="Optical interposer / optical engines (speculative)",
         role="Wafer-level optical interposer platform for 800G/1.6T optical engines and external light sources; pre-scale revenue.",
         products="Optical interposers, 800G/1.6T optical engines, light-source modules",
         customers="Module partners (Foxconn Interconnect, Mitsubishi et al.), design wins ramping",
         thesis="Optionality bet on chiplet-style optics packaging; binary execution risk, watch dilution.",
         exp=(90, 0, 5), adj=5.0, scores=[3, 4, 5, 0, 0, 0, 1, 1]),

    dict(t="ASML", name="ASML Holding", ctry="Netherlands", seg=SEG_LITHO,
         sub="EUV/DUV lithography monopoly (optics from Zeiss SMT)",
         role="The single most important optical system in semis: EUV scanners (13.5nm light, Zeiss mirrors, Cymer LPP source). High-NA ramping.",
         products="EUV (NXE/EXE High-NA), DUV immersion, metrology/inspection (YieldStar)",
         customers="TSMC, Samsung, Intel, SK Hynix, Micron",
         thesis="Effective monopoly on the optics that define leading-edge logic/DRAM; AI capex flows through to litho intensity.",
         exp=(0, 100, 0), adj=0.0, scores=[1, 0, 0, 5, 3, 0, 5, 4]),
    dict(t="7731.T", name="Nikon", ctry="Japan", seg=SEG_LITHO,
         sub="ArF/KrF immersion litho #2; in-house optics",
         role="Only non-ASML immersion DUV supplier; precision optics/measurement; exposure tools for packaging/panel.",
         products="ArF-i/KrF scanners, digital-litho for advanced packaging, optical components, cameras",
         customers="Intel (historic), memory fabs, OSATs, panel makers",
         thesis="Deep value on DUV + advanced-packaging lithography optionality; imaging funds the semi business.",
         exp=(0, 30, 0), adj=0.0, scores=[0, 0, 0, 3, 1, 0, 2, 3]),
    dict(t="7751.T", name="Canon", ctry="Japan", seg=SEG_LITHO,
         sub="KrF/i-line litho + nanoimprint (NIL) challenger",
         role="Volume KrF/i-line tools for mature nodes & packaging; nanoimprint litho as a disruptive (non-optical projection) alternative.",
         products="KrF/i-line steppers, FPA packaging litho, nanoimprint systems, imaging/printing",
         customers="Mature-node fabs, memory (NIL pilot), OSATs",
         thesis="Mature-node capex + NIL optionality inside a printing/imaging conglomerate - diluted but real semi-optics exposure.",
         exp=(0, 15, 0), adj=0.0, scores=[0, 0, 0, 2, 1, 0, 3, 4]),
    dict(t="6925.T", name="Ushio", ctry="Japan", seg=SEG_LITHO,
         sub="Industrial light sources; EUV mask-inspection source",
         role="Excimer/UV lamps and lasers across litho/anneal/cure steps; light source for actinic EUV mask inspection (Lasertec ecosystem).",
         products="Excimer lamps/lasers, UV cure/exposure sources, EPL/EUV-related sources",
         customers="Semi cap OEMs (incl. Lasertec), panel, life science",
         thesis="Hidden light-source content across many fab steps; levered to EUV mask-inspection buildout.",
         exp=(0, 40, 0), adj=0.0, scores=[0, 0, 0, 4, 3, 0, 3, 4]),
    dict(t="7741.T", name="Hoya", ctry="Japan", seg=SEG_LITHO,
         sub="EUV mask blanks near-monopoly; photomasks",
         role="Dominant supplier of EUV mask blanks (multilayer-coated low-defect substrates) and large-share DUV blanks/photomasks.",
         products="EUV/DUV mask blanks, photomasks, glass substrates, healthcare optics",
         customers="TSMC/Samsung/Intel mask shops, merchant mask makers",
         thesis="Every EUV layer starts on a Hoya blank; healthcare (~60%) stabilizes the cyclicality.",
         exp=(0, 30, 0), adj=0.0, scores=[0, 0, 0, 5, 2, 0, 5, 4]),

    dict(t="KLAC", name="KLA Corp.", ctry="US", seg=SEG_INSP,
         sub="Process control leader (optical inspection & metrology)",
         role="Broadband-plasma & laser-scanning optical wafer inspection, overlay/film metrology, reticle inspection - yield gatekeeper for leading edge.",
         products="BBP inspection (39xx), laser scanning, overlay, mask inspection, packaging inspection",
         customers="TSMC, Samsung, Intel, memory, OSATs",
         thesis="Process-control intensity rises with EUV layers and advanced packaging (HBM) - KLA is the toll booth.",
         exp=(0, 95, 0), adj=0.0, scores=[1, 0, 0, 4, 5, 0, 5, 4]),
    dict(t="6920.T", name="Lasertec", ctry="Japan", seg=SEG_INSP,
         sub="Actinic (EUV-light) mask inspection monopoly",
         role="Only supplier of actinic patterned-mask inspection (ACTIS) using 13.5nm light; EUV blank inspection (ABICS).",
         products="ACTIS A300+, ABICS, MATRICS, wafer/photomask inspection",
         customers="TSMC, Samsung, Intel, Hoya/mask shops",
         thesis="Pure monopoly on EUV mask actinic inspection; order lumpiness + short-seller history = volatility around a real franchise.",
         exp=(0, 100, 0), adj=0.0, scores=[0, 0, 0, 5, 5, 0, 4, 3]),
    dict(t="ONTO", name="Onto Innovation", ctry="US", seg=SEG_INSP,
         sub="Metrology + advanced-packaging inspection",
         role="Optical film/OCD metrology and the leading inspection franchise for advanced packaging/HBM (Dragonfly).",
         products="Dragonfly G3 (packaging inspection), Atlas OCD, IMPULSE, lithography for panel-level packaging",
         customers="TSMC (CoWoS), HBM makers (SK Hynix/Micron/Samsung), OSATs",
         thesis="The most direct optics play on CoWoS/HBM inspection intensity - AI packaging is its growth engine.",
         exp=(0, 95, 0), adj=1.0, scores=[2, 0, 1, 3, 5, 0, 4, 3]),
    dict(t="CAMT", name="Camtek", ctry="Israel", seg=SEG_INSP,
         sub="Advanced-packaging optical inspection & metrology",
         role="2D/3D optical inspection of HBM stacks, interposers, chiplets - direct beneficiary of AI packaging capacity adds.",
         products="Eagle G5/Hawk inspection, metrology for bumps/TSVs/hybrid bonding",
         customers="HBM makers, OSATs (ASE/Amkor), CoWoS chain",
         thesis="Small-cap pure play on the same HBM/CoWoS inspection theme as Onto; high growth, customer-concentrated.",
         exp=(0, 100, 0), adj=1.0, scores=[2, 0, 1, 2, 5, 0, 4, 3]),
    dict(t="NVMI", name="Nova Ltd.", ctry="Israel", seg=SEG_INSP,
         sub="Dimensional & materials metrology (optical + X-ray)",
         role="Optical CD/film metrology embedded in process tools plus stand-alone; gains content with gate-all-around & advanced packaging.",
         products="Nova Prism (OCD), Ancolt/Metrion, chemical & X-ray metrology",
         customers="TSMC, Samsung, Intel, memory",
         thesis="Metrology intensity compounder at leading edge; consistently outgrows WFE.",
         exp=(0, 100, 0), adj=0.0, scores=[1, 0, 0, 4, 5, 0, 5, 4]),

    dict(t="AXTI", name="AXT Inc.", ctry="US", seg=SEG_MATL,
         sub="InP / GaAs / Ge compound-semi substrates",
         role="Indium-phosphide substrates are the starting wafer for EML/CW lasers and photodetectors in every high-speed optical link.",
         products="InP, GaAs, Ge substrates; raw materials JVs",
         customers="Laser/PD chip makers (Lumentum/Coherent/Landmark etc.), China photonics chain",
         thesis="Upstream-most listed play on InP tightness from 1.6T lasers & CPO light sources; China-heavy footprint is the risk.",
         exp=(45, 10, 15), adj=1.0, scores=[4, 4, 4, 0, 0, 2, 1, 1]),
    dict(t="GLW", name="Corning", ctry="US", seg=SEG_MATL,
         sub="Optical fiber/cable & specialty glass",
         role="Fiber, cable and connectivity for hyperscale/AI datacenters (enterprise segment growing fast); EUV/optics glass; display.",
         products="Optical fiber & cable, datacenter connectivity, Gorilla glass, life-science optics",
         customers="Hyperscalers, carriers, Apple, display makers",
         thesis="AI datacenters need 2-5x the fiber of classic clouds; optical is now Corning's growth engine (Springboard plan).",
         exp=(35, 5, 15), adj=0.0, scores=[4, 3, 3, 1, 0, 3, 4, 4]),
    dict(t="IPGP", name="IPG Photonics", ctry="US", seg=SEG_MATL,
         sub="Fiber lasers (industrial; semi-adjacent)",
         role="Fiber lasers used in wafer/panel processing (anneal, dicing, drilling, marking) and micromachining for electronics & packaging.",
         products="High-power fiber lasers, pulsed lasers for micro-processing, telecom components (legacy)",
         customers="Industrial/EV, electronics & semi-packaging toolmakers",
         thesis="Peripheral but real: laser processing content in advanced packaging; mostly an industrial cycle play with net cash.",
         exp=(5, 10, 5), adj=0.0, scores=[1, 0, 0, 1, 1, 1, 4, 4]),

    dict(t="CIEN", name="Ciena", ctry="US", seg=SEG_SYS,
         sub="Coherent optical networking systems & pluggables",
         role="WaveLogic coherent DSP/modems interconnect datacenters (DCI/metro); 800ZR coherent pluggables carry AI traffic between sites.",
         products="WaveLogic 6 (1.6T coherent), Waveserver DCI, RLS line systems, coherent pluggables",
         customers="Hyperscalers (>50% of orders), carriers (AT&T/Verizon), cable",
         thesis="Inter-datacenter leg of AI networking: cloud now the majority customer; coherent-lite/800ZR expands the served market.",
         exp=(40, 0, 55), adj=0.0, scores=[4, 3, 2, 0, 0, 5, 4, 4]),
    dict(t="CLS", name="Celestica", ctry="Canada", seg=SEG_SYS,
         sub="Hyperscaler networking hardware (ODM/EMS)",
         role="Designs/builds 800G (and 1.6T-ready) Ethernet switches and AI server/storage systems for hyperscalers - each switch port pulls optics.",
         products="800G switches (HPS), AI compute/storage systems, industrial EMS",
         customers="Google, Meta, AWS + other hyperscalers",
         thesis="Switch-box winner of AI networking: optics-adjacent volume compounding, re-rated from EMS to ODM multiple.",
         exp=(60, 0, 5), adj=0.0, scores=[5, 4, 3, 0, 0, 1, 4, 3]),
]

THEMES = [
    ("AI DC Optics Cycle", 0.25),
    ("1.6T/3.2T Leverage", 0.15),
    ("CPO / SiPh Position", 0.15),
    ("EUV / Adv Litho", 0.15),
    ("Inspection / Metrology", 0.10),
    ("Telecom / DCI Recovery", 0.05),
    ("B/S & Profitability", 0.10),
    ("Risk Score (5 = low risk)", 0.05),
]

SEG_ORDER = [SEG_TRX, SEG_CHIP, SEG_LITHO, SEG_INSP, SEG_MATL, SEG_SYS]


# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------
def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def hdr(ws, row, col, text, width_span=1):
    c = ws.cell(row=row, column=col, value=text)
    c.font = F_HDR
    c.fill = FILL_HDR
    c.alignment = WRAP_C
    c.border = BORDER
    for extra in range(1, width_span):
        cc = ws.cell(row=row, column=col + extra)
        cc.fill = FILL_HDR
        cc.border = BORDER
    return c


def put(ws, row, col, value, font=F_LABEL, nf=None, align=None, fill=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if nf:
        c.number_format = nf
    if align:
        c.alignment = align
    if fill:
        c.fill = fill
    if border:
        c.border = BORDER
    return c


def title_block(ws, title, subtitle, as_of):
    ws.cell(row=1, column=1, value=title).font = F_TITLE
    ws.cell(row=2, column=1, value=subtitle + f"   |   Market data as of {as_of} (Yahoo Finance)").font = F_SUB


def pct(x):
    return None if x is None else float(x)


def musd(x, fx=1.0):
    """Convert a raw currency amount to $ millions USD."""
    return None if x is None else float(x) * fx / 1e6


# ----------------------------------------------------------------------------
# Sheet builders
# ----------------------------------------------------------------------------
def build_readme(wb, as_of):
    ws = wb.active
    ws.title = "Read Me"
    ws.sheet_properties.tabColor = NAVY
    set_widths(ws, [28, 110])
    title_block(ws, "Optical Names x Semiconductor Supply Chain - Equity Model",
                "30-company universe, live comps, driver-based TAM model, exposure & thematic scoring", as_of)

    r = 4
    put(ws, r, 1, "PURPOSE", F_SECTION, border=False)
    r += 1
    put(ws, r, 1, "", border=False)
    ws.cell(row=r, column=2, value=(
        "Maps the listed optics/photonics ecosystem to the semiconductor supply chain along two axes: "
        "(1) optics that move data BETWEEN semiconductors - AI/datacenter transceivers, optical DSPs, CPO - and "
        "(2) optics that MAKE semiconductors - lithography light/optics, mask blanks, optical inspection & metrology. "
        "Includes a driver-based market model (accelerator units x attach x speed mix x ASP), live comps, "
        "per-company revenue-exposure scenarios to 2028E, thematic scoring and sensitivities.")).font = F_LABEL
    ws.cell(row=r, column=2).alignment = WRAP
    ws.row_dimensions[r].height = 60

    r += 2
    put(ws, r, 1, "TAB GUIDE", F_SECTION, border=False)
    r += 1
    guide = [
        ("Universe", "30 names in 6 segments: who they are, role in the semi chain, key products/customers, one-line thesis."),
        ("Supply Chain Map", "Stage-by-stage value chain - AI datacenter interconnect chain, fab-equipment optics chain, cross-cutting enablers - with universe tickers and key private players per stage."),
        ("Market Model", "2023A-2028E driver model: AI accelerator units x optics attach x speed mix (400G/800G/1.6T) x ASPs -> datacom optics TAM; CPO displacement; telecom optics; WFE-derived litho & inspection optics TAM. Blue cells are yours to flex."),
        ("Comps", "Live valuation/financial comps in USD with segment medians: growth, margins, EV/S, EV/EBITDA, P/E, FCF yield, 1-yr return."),
        ("Company Exposure", "Estimated revenue split (AI/datacom, fab-optics, telecom, other) per name; grows each bucket at Market Model CAGRs (+/- share shift) to a 2028E revenue scenario; implied CAGR and EV/2028E sales."),
        ("Thematic Scores", "0-5 scores across 8 themes, weighted composite (weights adjustable) and rank."),
        ("Sensitivity", "2028E datacom optics market grids: attach ratio x 1.6T ASP, and accelerator units x 1.6T mix."),
    ]
    hdr(ws, r, 1, "Tab"); hdr(ws, r, 2, "What it contains")
    r += 1
    for name, desc in guide:
        put(ws, r, 1, name, F_LABEL_B)
        c = put(ws, r, 2, desc, F_LABEL, align=WRAP)
        ws.row_dimensions[r].height = 30
        r += 1

    r += 1
    put(ws, r, 1, "CONVENTIONS", F_SECTION, border=False)
    r += 1
    conv = [
        ("Blue font", "Input / assumption - flex these; everything downstream recalculates."),
        ("Black font", "Formula or hard market data (pulled from Yahoo Finance)."),
        ("$ figures", "USD millions unless marked $B; non-USD names converted at spot FX on the as-of date."),
        ("Exposure %s", "Analyst estimates of revenue mix - directional, not company-reported segment data."),
        ("NM / blank", "Not meaningful (negative denominator) or data unavailable."),
    ]
    for k, v in conv:
        put(ws, r, 1, k, F_INPUT_B if k == "Blue font" else F_LABEL_B)
        put(ws, r, 2, v, F_LABEL, align=WRAP)
        r += 1

    r += 1
    put(ws, r, 1, "DISCLAIMER", F_SECTION, border=False)
    r += 1
    ws.cell(row=r, column=2, value=(
        "For research/educational use. Market data via Yahoo Finance and may contain errors; exposure estimates and "
        "forecasts are analyst assumptions, not investment advice. Verify before use.")).font = F_SUB
    ws.cell(row=r, column=2).alignment = WRAP
    return ws


def build_universe(wb):
    ws = wb.create_sheet("Universe")
    ws.sheet_properties.tabColor = "2E74B5"
    widths = [11, 22, 10, 16, 34, 52, 44, 40, 52]
    set_widths(ws, widths)
    cols = ["Ticker", "Company", "Country", "Segment", "Sub-segment / niche",
            "Role in the semiconductor supply chain", "Key products",
            "Key customers / end markets", "Thesis in one line"]
    for i, cname in enumerate(cols, start=1):
        hdr(ws, 1, i, cname)
    ws.freeze_panes = "C2"

    r = 2
    for seg in SEG_ORDER:
        put(ws, r, 1, seg, F_BAND, fill=FILL_BAND)
        for cidx in range(2, len(cols) + 1):
            put(ws, r, cidx, "", fill=FILL_BAND)
        r += 1
        for co in [c for c in UNIVERSE if c["seg"] == seg]:
            vals = [co["t"], co["name"], co["ctry"], SEG_SHORT[seg], co["sub"],
                    co["role"], co["products"], co["customers"], co["thesis"]]
            for i, v in enumerate(vals, start=1):
                f = F_LABEL_B if i <= 2 else F_LABEL
                put(ws, r, i, v, f, align=WRAP)
            ws.row_dimensions[r].height = 42
            r += 1
    return ws


def build_supply_chain(wb):
    ws = wb.create_sheet("Supply Chain Map")
    ws.sheet_properties.tabColor = "2E74B5"
    set_widths(ws, [4, 30, 62, 34, 34, 52])
    cols = ["#", "Value-chain stage", "What it is / why it matters",
            "Universe names (tickers)", "Key private / non-universe players", "Bottlenecks & notes"]

    sections = [
        ("A. AI DATACENTER OPTICAL INTERCONNECT CHAIN  (optics that move data between semiconductors)", [
            ("III-V & SiPh substrates",
             "InP substrates are the starting wafer for EML/CW lasers & photodetectors; SOI wafers for silicon photonics.",
             "AXTI",
             "Sumitomo Electric, JX Metals, IQE, Soitec (SOI)",
             "InP 4\"->6\" transition is the gating capacity item for 1.6T lasers and CPO light sources."),
            ("Laser & modulator chips (EML / CW / VCSEL)",
             "The light engines: EMLs for 200G/lane pluggables, high-power CW lasers feeding silicon photonics and CPO.",
             "LITE, COHR, AVGO, MTSI",
             "Mitsubishi Electric, Sumitomo Electric, Landmark, SEDI",
             "200G/lane EML yield is the industry-wide constraint; CW laser reliability governs CPO adoption."),
            ("Photonic ICs / silicon photonics",
             "Integrates modulators/detectors/waveguides on silicon; the platform for 1.6T DR/FR modules and optical engines.",
             "AVGO, MRVL, COHR, POET",
             "GlobalFoundries, Tower Semiconductor, TSMC (COUPE), imec",
             "Foundry SiPh PDK maturity determines who can play; laser attach/packaging is the yield battleground."),
            ("Analog front-end (TIA / driver / CDR)",
             "Amplifies photodiode current and drives modulators; linear-drive (LPO/LRO) architectures raise analog content.",
             "SMTC, MTSI",
             "Analog Devices, in-house at Broadcom/Marvell",
             "LPO shifts BOM value from DSP to analog - watch architecture wins at 1.6T."),
            ("PAM4 DSP / optical PHY",
             "Retimes and equalizes the electrical-optical interface; the highest-value chip inside a pluggable module.",
             "MRVL, AVGO, CRDO",
             "In-house hyperscaler efforts",
             "Duopoly-plus-one (MRVL/AVGO/CRDO); 3nm DSP tapeouts pace the 1.6T ramp."),
            ("Passive optics & fiber coupling",
             "FAUs, lenses, AWG mux/demux, MPO connectors - content per port rises sharply with SiPh and CPO.",
             "300394.SZ, FN, GLW",
             "Browave, Senko, US Conec",
             "Fiber-array alignment capacity is a quiet bottleneck for CPO scale-up."),
            ("Transceiver module design & assembly",
             "Turns chips+optics into 800G/1.6T pluggables; the most visible P&L beneficiary of AI optics capex.",
             "300308.SZ, 300502.SZ, COHR, LITE, AAOI, 002281.SZ",
             "Hisense Broadband, HG Genuine, Source Photonics",
             "Chinese trio holds ~60% share of hyperscaler pluggables; tariffs/export rules are the key swing factor."),
            ("Optical EMS / contract manufacturing",
             "Precision optical assembly & test at scale for chip vendors and systems OEMs (incl. Nvidia optics).",
             "FN, CLS",
             "Venture Corp, Benchmark",
             "Fabrinet is the de-facto standard for high-mix optical assembly; capacity additions track AI demand."),
            ("Switching, DCI & coherent systems",
             "Switch boxes whose port counts pull optics demand; coherent systems/pluggables link datacenters together.",
             "CLS, CIEN",
             "Arista, Nvidia (Spectrum-X/Quantum), Cisco, Nokia(Infinera)",
             "Switch ASIC cadence (51.2T->102.4T) sets the optics upgrade clock; 800ZR blurs telecom/datacom."),
            ("End demand: AI accelerators & hyperscalers",
             "GPU/XPU cluster architectures determine attach ratios (scale-out, and now scale-up going optical).",
             "-",
             "Nvidia, AMD, Google TPU, AWS Trainium, Microsoft, Meta",
             "Scale-up networks moving from copper to optics is the biggest upside optionality in the model."),
        ]),
        ("B. SEMICONDUCTOR FAB OPTICS CHAIN  (optics that make semiconductors)", [
            ("Lithography light sources",
             "DUV excimer lasers (KrF/ArF) and EUV LPP sources (CO2 laser + tin plasma) that generate the exposure light.",
             "ASML (Cymer), 6925.T",
             "Gigaphoton (Komatsu)",
             "EUV source power (>500W) gates High-NA throughput; excimer duopoly Cymer/Gigaphoton."),
            ("Projection & illumination optics",
             "The most precise optics made: EUV mirror sets (sub-nm figure error) and ArF immersion lens columns.",
             "7731.T, 7751.T (in-house)",
             "Zeiss SMT (ASML's optics partner, private)",
             "Zeiss SMT is a sole-source choke point for EUV mirrors - unlisted, accessed via ASML."),
            ("Photomask blanks, pellicles & photomasks",
             "EUV blanks (40+ multilayer coatings, near-zero defects), DUV blanks, pellicles protecting masks in the scanner.",
             "7741.T",
             "AGC, Mitsui Chemicals (pellicles), Toppan/DNP (masks)",
             "Hoya ~near-monopoly in EUV blanks; blank defectivity feeds straight into Lasertec/KLA inspection demand."),
            ("Lithography systems",
             "Scanner/stepper integration - the optical system that prints every transistor.",
             "ASML, 7731.T, 7751.T",
             "SMEE (China, embargoed at leading edge)",
             "ASML monopoly at EUV; DUV export controls reshape China demand for Nikon/Canon mature tools."),
            ("Mask & blank inspection (actinic)",
             "Inspecting EUV masks with EUV light itself - the only way to see printable phase defects.",
             "6920.T, KLAC",
             "-",
             "Lasertec ACTIS is sole-source actinic patterned-mask inspection; KLA competes at DUV/e-beam."),
            ("Wafer inspection & optical metrology",
             "Broadband-plasma/laser inspection, OCD & overlay metrology - yield control for EUV nodes, GAA and HBM.",
             "KLAC, ONTO, NVMI, CAMT",
             "Applied Materials (PDC), Hitachi High-Tech (e-beam)",
             "Process-control intensity rises with EUV layer count and advanced packaging - structural share of WFE grows."),
            ("Laser processing in fab & packaging",
             "Annealing, wafer dicing/grooving, via drilling, marking - laser steps throughout front & back end.",
             "IPGP, COHR",
             "DISCO, Hamamatsu, Amada/Via Mechanics",
             "Advanced packaging (CoWoS/HBM) adds laser drilling/dicing content per wafer."),
        ]),
        ("C. CROSS-CUTTING ENABLERS", [
            ("Optical fiber, cable & connectivity",
             "Intra/inter-datacenter fiber plant; AI clusters need 2-5x the fiber of classic cloud halls.",
             "GLW",
             "Prysmian, CommScope, Amphenol (connectivity)",
             "Pre-terminated high-density fiber (MPO trunks) is a hidden AI capex line-item."),
            ("Co-packaged optics (CPO) ecosystem",
             "Optical engines mounted beside the switch/XPU die; replaces pluggables at highest speeds.",
             "AVGO, MRVL, COHR, LITE, 300394.SZ, POET",
             "Nvidia (Quantum-X/Spectrum-X photonics), TSMC COUPE",
             "CPO shifts value: less pluggable assembly, more laser sources (external light), fiber coupling and foundry SiPh."),
            ("Optical test & measurement",
             "Module/link test at 200G-per-lane; test intensity per port rises with speed.",
             "-",
             "EXFO, Viavi, Keysight, Yokogawa",
             "Test capacity often gates module shipment ramps in tight quarters."),
        ]),
    ]

    r = 1
    ws.cell(row=r, column=1, value="Optical value chains through the semiconductor ecosystem").font = F_TITLE
    r += 2
    for sec_title, stages in sections:
        put(ws, r, 1, sec_title, F_SECTION, border=False)
        r += 1
        for i, cname in enumerate(cols, start=1):
            hdr(ws, r, i, cname)
        r += 1
        for n, (stage, what, uni, priv, note) in enumerate(stages, start=1):
            put(ws, r, 1, n, F_LABEL, align=CENTER)
            put(ws, r, 2, stage, F_LABEL_B, align=WRAP)
            put(ws, r, 3, what, F_LABEL, align=WRAP)
            put(ws, r, 4, uni, F_CALC_B, align=WRAP)
            put(ws, r, 5, priv, F_LABEL, align=WRAP)
            put(ws, r, 6, note, F_LABEL, align=WRAP)
            ws.row_dimensions[r].height = 40
            r += 1
        r += 1
    return ws


def build_market_model(wb):
    """Driver model. Years 2023A-2028E in columns C:H. Col I holds side assumptions (YoY deltas)."""
    ws = wb.create_sheet("Market Model")
    ws.sheet_properties.tabColor = "C00000"
    set_widths(ws, [52, 12, 11, 11, 11, 11, 11, 11, 12, 60])
    refs = {}

    ws.cell(row=1, column=1, value="AI / Datacom + Fab-Equipment Optics: Driver-Based Market Model").font = F_TITLE
    ws.cell(row=2, column=1, value="Blue = input assumption (flex freely). All $ in millions unless marked $B.").font = F_SUB

    r = 4
    hdr(ws, r, 1, "Driver / line item"); hdr(ws, r, 2, "Units")
    for i, y in enumerate(YEARS):
        hdr(ws, r, 3 + i, y)
    hdr(ws, r, 9, "YoY assump."); hdr(ws, r, 10, "Notes")
    ws.freeze_panes = "C5"
    r += 1

    def label(row, text, units, note="", bold=False):
        put(ws, row, 1, text, F_LABEL_B if bold else F_LABEL, align=LEFT)
        put(ws, row, 2, units, F_LABEL, align=CENTER)
        put(ws, row, 10, note, F_SUB, align=WRAP)

    def inputs_row(row, vals, nf=NF_UNITS):
        for i, v in enumerate(vals):
            put(ws, row, 3 + i, v, F_INPUT, nf=nf, align=CENTER)

    def formula_row(row, formulas, nf=NF_UNITS, bold=False, fill=None):
        for i, f in enumerate(formulas):
            put(ws, row, 3 + i, f, F_CALC_B if bold else F_CALC, nf=nf, align=CENTER, fill=fill)

    def yoy_row(row, src_row, label_text="   YoY growth"):
        label(row, label_text, "%")
        put(ws, row, 3, "", align=CENTER)
        for i in range(1, 6):
            col = get_column_letter(3 + i)
            prev = get_column_letter(2 + i)
            put(ws, row, 3 + i, f"=IFERROR({col}{src_row}/{prev}{src_row}-1,\"\")",
                F_CALC, nf=NF_PCT, align=CENTER)

    # --- Section 1: AI / datacom ports ------------------------------------
    put(ws, r, 1, "1. AI / DATACOM OPTICAL TRANSCEIVER DEMAND", F_SECTION, border=False); r += 1
    refs["accel"] = r
    label(r, "AI accelerator shipments (GPU + custom XPU)", "M units",
          "Nvidia + AMD + hyperscaler ASICs; the primary demand driver.")
    inputs_row(r, [3.9, 6.5, 9.5, 12.5, 15.5, 18.0]); r += 1
    yoy_row(r, refs["accel"]); r += 1
    refs["attach"] = r
    label(r, "Optical ports per accelerator (attach ratio)", "ports/unit",
          "Scale-out today; rises as scale-up fabrics go optical. Key swing input.")
    inputs_row(r, [2.0, 2.4, 2.8, 3.0, 3.3, 3.5]); r += 1
    refs["ai_ports"] = r
    label(r, "AI-driven optical ports", "M units", "Formula: accelerators x attach ratio", bold=True)
    formula_row(r, [f"=({get_column_letter(3+i)}{refs['accel']}*{get_column_letter(3+i)}{refs['attach']})"
                    for i in range(6)], bold=True); r += 1
    refs["nonai_ports"] = r
    label(r, "Non-AI datacom ports (classic cloud/enterprise)", "M units",
          "2023 base is input; grows at rate in col I.")
    put(ws, r, 3, 10.0, F_INPUT, nf=NF_UNITS, align=CENTER)
    for i in range(1, 6):
        put(ws, r, 3 + i, f"={get_column_letter(2+i)}{r}*(1+$I${r})", F_CALC, nf=NF_UNITS, align=CENTER)
    put(ws, r, 9, 0.08, F_INPUT, nf=NF_PCT0, align=CENTER); r += 1
    refs["tot_ports"] = r
    label(r, "Total datacom optical ports", "M units", "", bold=True)
    formula_row(r, [f"={get_column_letter(3+i)}{refs['ai_ports']}+{get_column_letter(3+i)}{refs['nonai_ports']}"
                    for i in range(6)], bold=True, fill=FILL_TOTAL); r += 2

    # --- Section 2: speed mix ----------------------------------------------
    put(ws, r, 1, "2. PORT SPEED MIX (% of total ports)", F_SECTION, border=False); r += 1
    refs["mix400"] = r
    label(r, "<= 400G", "% mix", "Legacy speeds fade but never fully disappear.")
    inputs_row(r, [0.70, 0.45, 0.30, 0.20, 0.12, 0.08], nf=NF_PCT0); r += 1
    refs["mix800"] = r
    label(r, "800G", "% mix", "Workhorse speed of the 2024-2027 build.")
    inputs_row(r, [0.30, 0.55, 0.60, 0.55, 0.48, 0.40], nf=NF_PCT0); r += 1
    refs["mix16"] = r
    label(r, "1.6T", "% mix", "Ramps with GB300/Rubin-class systems and 102.4T switching.")
    inputs_row(r, [0.00, 0.00, 0.10, 0.25, 0.40, 0.52], nf=NF_PCT0); r += 1
    refs["mixchk"] = r
    label(r, "   Mix check (must = 100%)", "%")
    formula_row(r, [f"=SUM({get_column_letter(3+i)}{refs['mix400']}:{get_column_letter(3+i)}{refs['mix16']})"
                    for i in range(6)], nf=NF_PCT0); r += 2

    # --- Section 3: ASPs ----------------------------------------------------
    put(ws, r, 1, "3. ASPs ($ per port; first year input, then declines at rate in col I)", F_SECTION, border=False); r += 1
    asp_rows = {}
    for key, lbl, first_val, first_col, decline, note in [
        ("asp400", "<= 400G ASP", 350, 0, -0.15, "Mature product; steady price erosion."),
        ("asp800", "800G ASP", 950, 0, -0.12, "EML-based OSFP; SiPh variants cheaper."),
        ("asp16", "1.6T ASP", 1700, 2, -0.18, "Enters 2025; fastest erosion as SiPh scales."),
    ]:
        asp_rows[key] = r
        refs[key] = r
        label(r, lbl, "$/port", note)
        for i in range(6):
            if i < first_col:
                put(ws, r, 3 + i, "", align=CENTER)
            elif i == first_col:
                put(ws, r, 3 + i, first_val, F_INPUT, nf=NF_USD, align=CENTER)
            else:
                put(ws, r, 3 + i, f"={get_column_letter(2+i)}{r}*(1+$I${r})", F_CALC, nf=NF_USD, align=CENTER)
        put(ws, r, 9, decline, F_INPUT, nf=NF_PCT0, align=CENTER)
        r += 1
    r += 1

    # --- Section 4: revenue --------------------------------------------------
    put(ws, r, 1, "4. DATACOM OPTICAL TRANSCEIVER MARKET ($M)", F_SECTION, border=False); r += 1
    rev_rows = {}
    for key, mixkey, aspkey, lbl in [
        ("rev400", "mix400", "asp400", "<= 400G revenue"),
        ("rev800", "mix800", "asp800", "800G revenue"),
        ("rev16", "mix16", "asp16", "1.6T revenue"),
    ]:
        rev_rows[key] = r
        refs[key] = r
        label(r, lbl, "$M", "Formula: total ports x mix x ASP  (M units x $ = $M)")
        formula_row(r, [f"={get_column_letter(3+i)}{refs['tot_ports']}*{get_column_letter(3+i)}{refs[mixkey]}"
                        f"*N({get_column_letter(3+i)}{refs[aspkey]})" for i in range(6)], nf=NF_M)
        r += 1
    refs["rev_tot"] = r
    label(r, "Total datacom transceiver market", "$M", "", bold=True)
    formula_row(r, [f"=SUM({get_column_letter(3+i)}{rev_rows['rev400']}:{get_column_letter(3+i)}{rev_rows['rev16']})"
                    for i in range(6)], nf=NF_M, bold=True, fill=FILL_TOTAL); r += 1
    yoy_row(r, refs["rev_tot"]); r += 2

    # --- Section 5: CPO -------------------------------------------------------
    put(ws, r, 1, "5. CO-PACKAGED OPTICS (CPO) DISPLACEMENT", F_SECTION, border=False); r += 1
    refs["cpo_pen"] = r
    label(r, "CPO share of 1.6T ports", "% of 1.6T",
          "Nvidia Quantum-X/Spectrum-X photonics + Broadcom Bailly ramp.")
    inputs_row(r, [0.0, 0.0, 0.0, 0.02, 0.06, 0.12], nf=NF_PCT0); r += 1
    refs["cpo_disp"] = r
    label(r, "Pluggable revenue displaced by CPO", "$M", "Formula: 1.6T revenue x CPO share")
    formula_row(r, [f"={get_column_letter(3+i)}{refs['rev16']}*{get_column_letter(3+i)}{refs['cpo_pen']}"
                    for i in range(6)], nf=NF_M); r += 1
    refs["rev_plug"] = r
    label(r, "Net pluggable transceiver market", "$M",
          "What the module makers (Comps: Transceivers segment) actually ship.", bold=True)
    formula_row(r, [f"={get_column_letter(3+i)}{refs['rev_tot']}-{get_column_letter(3+i)}{refs['cpo_disp']}"
                    for i in range(6)], nf=NF_M, bold=True, fill=FILL_TOTAL); r += 2

    # --- Section 6: telecom ----------------------------------------------------
    put(ws, r, 1, "6. TELECOM OPTICS (components + modules)", F_SECTION, border=False); r += 1
    refs["tel_g"] = r
    label(r, "Telecom optics growth", "% YoY", "2024 inventory digestion, then 400ZR/800ZR-led recovery.")
    put(ws, r, 3, "", align=CENTER)
    for i, g in enumerate([-0.08, 0.04, 0.07, 0.06, 0.05], start=1):
        put(ws, r, 3 + i, g, F_INPUT, nf=NF_PCT0, align=CENTER)
    r += 1
    refs["tel_rev"] = r
    label(r, "Telecom optics market", "$M", "2023 base input.", bold=True)
    put(ws, r, 3, 14000, F_INPUT, nf=NF_M, align=CENTER)
    for i in range(1, 6):
        put(ws, r, 3 + i, f"={get_column_letter(2+i)}{r}*(1+{get_column_letter(3+i)}{refs['tel_g']})",
            F_CALC_B, nf=NF_M, align=CENTER)
    r += 2

    # --- Section 7: fab optics --------------------------------------------------
    put(ws, r, 1, "7. SEMICONDUCTOR FAB OPTICS (equipment chain, $B)", F_SECTION, border=False); r += 1
    refs["wfe"] = r
    label(r, "Wafer fab equipment (WFE) spend", "$B", "AI/leading-edge logic + HBM led cycle.")
    inputs_row(r, [95, 105, 118, 130, 140, 150], nf=NF_B); r += 1
    refs["litho_sh"] = r
    label(r, "Lithography share of WFE", "%", "EUV mix pushes litho share up over time.")
    inputs_row(r, [0.22, 0.23, 0.23, 0.24, 0.24, 0.24], nf=NF_PCT0); r += 1
    refs["litho_mkt"] = r
    label(r, "Lithography systems market", "$B")
    formula_row(r, [f"={get_column_letter(3+i)}{refs['wfe']}*{get_column_letter(3+i)}{refs['litho_sh']}"
                    for i in range(6)], nf=NF_B); r += 1
    refs["litho_opt_sh"] = r
    label(r, "Optics + light-source content of litho tools", "%",
          "Zeiss optics + Cymer source ~1/3 of an EUV tool's value.")
    inputs_row(r, [0.33, 0.34, 0.34, 0.35, 0.35, 0.35], nf=NF_PCT0); r += 1
    refs["litho_opt"] = r
    label(r, "Litho optics & source content", "$B", "", bold=True)
    formula_row(r, [f"={get_column_letter(3+i)}{refs['litho_mkt']}*{get_column_letter(3+i)}{refs['litho_opt_sh']}"
                    for i in range(6)], nf=NF_B, bold=True); r += 1
    refs["pc_sh"] = r
    label(r, "Process control share of WFE", "%", "Rises with EUV layers & advanced packaging.")
    inputs_row(r, [0.105, 0.11, 0.11, 0.115, 0.115, 0.12], nf="0.0%"); r += 1
    refs["pc_mkt"] = r
    label(r, "Process control market", "$B")
    formula_row(r, [f"={get_column_letter(3+i)}{refs['wfe']}*{get_column_letter(3+i)}{refs['pc_sh']}"
                    for i in range(6)], nf=NF_B); r += 1
    refs["pc_opt_sh"] = r
    label(r, "Optical share of process control", "%", "Optical inspection/OCD vs e-beam & X-ray.")
    inputs_row(r, [0.75] * 6, nf=NF_PCT0); r += 1
    refs["insp_opt"] = r
    label(r, "Optical inspection & metrology market", "$B", "", bold=True)
    formula_row(r, [f"={get_column_letter(3+i)}{refs['pc_mkt']}*{get_column_letter(3+i)}{refs['pc_opt_sh']}"
                    for i in range(6)], nf=NF_B, bold=True); r += 1
    refs["fab_opt"] = r
    label(r, "Total fab optics TAM", "$B", "", bold=True)
    formula_row(r, [f"={get_column_letter(3+i)}{refs['litho_opt']}+{get_column_letter(3+i)}{refs['insp_opt']}"
                    for i in range(6)], nf=NF_B, bold=True, fill=FILL_TOTAL); r += 2

    # --- Section 8: total ----------------------------------------------------------
    put(ws, r, 1, "8. TOTAL OPTICAL x SEMICONDUCTOR TAM", F_SECTION, border=False); r += 1
    refs["tam"] = r
    label(r, "Total TAM (pluggables + telecom + fab optics)", "$M", "", bold=True)
    formula_row(r, [f"={get_column_letter(3+i)}{refs['rev_plug']}+{get_column_letter(3+i)}{refs['tel_rev']}"
                    f"+{get_column_letter(3+i)}{refs['fab_opt']}*1000" for i in range(6)],
                nf=NF_M, bold=True, fill=FILL_TOTAL); r += 1
    yoy_row(r, refs["tam"]); r += 2

    # CAGR block used by Company Exposure
    put(ws, r, 1, "MARKET CAGRs (2025E -> 2028E) - feed the Company Exposure tab", F_SECTION, border=False); r += 1
    cagr_specs = [
        ("cagr_ai", "AI/datacom pluggables CAGR '25-'28", refs["rev_plug"]),
        ("cagr_fab", "Fab optics CAGR '25-'28", refs["fab_opt"]),
        ("cagr_tel", "Telecom optics CAGR '25-'28", refs["tel_rev"]),
    ]
    for key, lbl, src in cagr_specs:
        refs[key] = r
        label(r, lbl, "%")
        put(ws, r, 3, f"=(H{src}/E{src})^(1/3)-1", F_CALC_B, nf=NF_PCT, align=CENTER)
        r += 1

    return ws, refs


def build_comps(wb, data):
    ws = wb.create_sheet("Comps")
    ws.sheet_properties.tabColor = "538135"
    headers = ["Ticker", "Company", "Segment", "Ccy", "Price (local)", "Mkt Cap $M", "Net Debt $M",
               "EV $M", "Rev TTM $M", "Rev Grw YoY", "Gross Mgn", "EBITDA Mgn", "Op Mgn",
               "EBITDA $M", "EV / Sales", "EV / EBITDA", "P/E (fwd)", "FCF $M", "FCF Yield",
               "1Y Return", "% off 52w High", "Beta"]
    set_widths(ws, [11, 24, 13, 6, 11, 11, 10, 11, 10, 9, 9, 9, 9, 10, 9, 10, 9, 10, 9, 9, 10, 7])
    ws.cell(row=1, column=1, value="Valuation & Financial Comps (USD, converted at spot FX)").font = F_TITLE
    ws.cell(row=2, column=1,
            value="Values pulled from Yahoo Finance; EV, multiples, yields and medians are live formulas.").font = F_SUB
    hr = 4
    for i, h in enumerate(headers, start=1):
        hdr(ws, hr, i, h)
    ws.freeze_panes = "C5"

    tickers = data["tickers"]
    comps_row = {}
    r = hr + 1
    seg_data_rows = {}

    for seg in SEG_ORDER:
        put(ws, r, 1, seg, F_BAND, fill=FILL_BAND)
        for cidx in range(2, len(headers) + 1):
            put(ws, r, cidx, "", fill=FILL_BAND)
        r += 1
        first = r
        for co in [c for c in UNIVERSE if c["seg"] == seg]:
            d = tickers.get(co["t"], {})
            fx = d.get("fxToUsd") or 1.0
            ffx = d.get("finFxToUsd") or fx
            price = d.get("currentPrice")
            mcap = musd(d.get("marketCap"), fx)
            debt = d.get("totalDebt") or 0
            cash = d.get("totalCash") or 0
            netdebt = musd(debt - cash, ffx)
            rev = musd(d.get("totalRevenue"), ffx)
            ebitda = musd(d.get("ebitda"), ffx)
            fcf = musd(d.get("freeCashflow"), ffx)
            hi = d.get("fiftyTwoWeekHigh")
            off_hi = (price / hi - 1) if (price and hi) else None

            comps_row[co["t"]] = r
            put(ws, r, 1, co["t"], F_LABEL_B)
            put(ws, r, 2, co["name"], F_LABEL)
            put(ws, r, 3, SEG_SHORT[seg], F_LABEL)
            put(ws, r, 4, d.get("currency") or "", F_LABEL, align=CENTER)
            put(ws, r, 5, price, F_CALC, nf=NF_PX)
            put(ws, r, 6, mcap, F_CALC, nf=NF_M)
            put(ws, r, 7, netdebt, F_CALC, nf=NF_M)
            put(ws, r, 8, f"=IF(F{r}=\"\",\"\",F{r}+N(G{r}))", F_CALC, nf=NF_M)
            put(ws, r, 9, rev, F_CALC, nf=NF_M)
            put(ws, r, 10, pct(d.get("revenueGrowth")), F_CALC, nf=NF_PCT)
            put(ws, r, 11, pct(d.get("grossMargins")), F_CALC, nf=NF_PCT)
            put(ws, r, 12, pct(d.get("ebitdaMargins")), F_CALC, nf=NF_PCT)
            put(ws, r, 13, pct(d.get("operatingMargins")), F_CALC, nf=NF_PCT)
            put(ws, r, 14, ebitda, F_CALC, nf=NF_M)
            put(ws, r, 15, f"=IFERROR(IF(I{r}>0,H{r}/I{r},\"NM\"),\"NM\")", F_CALC, nf=NF_MULT)
            put(ws, r, 16, f"=IFERROR(IF(N{r}>0,H{r}/N{r},\"NM\"),\"NM\")", F_CALC, nf=NF_MULT)
            fpe = d.get("forwardPE")
            put(ws, r, 17, fpe if (fpe and fpe > 0) else "NM", F_CALC, nf=NF_MULT)
            put(ws, r, 18, fcf, F_CALC, nf=NF_M)
            put(ws, r, 19, f"=IFERROR(IF(AND(R{r}<>\"\",F{r}>0),R{r}/F{r},\"\"),\"\")", F_CALC, nf=NF_PCT)
            put(ws, r, 20, pct(d.get("oneYearReturn")), F_CALC, nf=NF_PCT)
            put(ws, r, 21, off_hi, F_CALC, nf=NF_PCT)
            put(ws, r, 22, d.get("beta"), F_CALC, nf="0.00")
            r += 1
        last = r - 1
        seg_data_rows[seg] = (first, last)
        # segment median row
        put(ws, r, 1, "", fill=FILL_GREY)
        put(ws, r, 2, f"Median - {SEG_SHORT[seg]}", F_CALC_B, fill=FILL_GREY)
        for cidx, nf in [(10, NF_PCT), (11, NF_PCT), (12, NF_PCT), (13, NF_PCT),
                         (15, NF_MULT), (16, NF_MULT), (17, NF_MULT), (19, NF_PCT), (20, NF_PCT)]:
            col = get_column_letter(cidx)
            put(ws, r, cidx, f"=IFERROR(MEDIAN({col}{first}:{col}{last}),\"\")",
                F_CALC_B, nf=nf, fill=FILL_GREY)
        for cidx in [3, 4, 5, 6, 7, 8, 9, 14, 18, 21, 22]:
            put(ws, r, cidx, "", fill=FILL_GREY)
        r += 1

    # overall median
    all_first = hr + 1
    all_last = r - 1
    put(ws, r, 1, "", fill=FILL_TOTAL)
    put(ws, r, 2, "Median - Universe", F_CALC_B, fill=FILL_TOTAL)
    for cidx, nf in [(10, NF_PCT), (11, NF_PCT), (12, NF_PCT), (13, NF_PCT),
                     (15, NF_MULT), (16, NF_MULT), (17, NF_MULT), (19, NF_PCT), (20, NF_PCT)]:
        col = get_column_letter(cidx)
        # AGGREGATE-free approach: MEDIAN ignores text/blank cells, so range over all rows is fine
        put(ws, r, cidx, f"=IFERROR(MEDIAN({col}{all_first}:{col}{all_last}),\"\")",
            F_CALC_B, nf=nf, fill=FILL_TOTAL)
    for cidx in [3, 4, 5, 6, 7, 8, 9, 14, 18, 21, 22]:
        put(ws, r, cidx, "", fill=FILL_TOTAL)

    return ws, comps_row


def build_exposure(wb, comps_row, mm_refs):
    ws = wb.create_sheet("Company Exposure")
    ws.sheet_properties.tabColor = "538135"
    set_widths(ws, [11, 24, 13, 11, 9, 9, 9, 9, 11, 11, 11, 11, 10, 11, 11, 11, 10])
    ws.cell(row=1, column=1, value="Revenue Exposure Model - who actually monetizes each optical TAM").font = F_TITLE
    ws.cell(row=2, column=1, value=(
        "Exposure %s are analyst estimates (blue - flex them). Each bucket grows at the Market Model "
        "'25-'28 CAGR, +/- a company share-shift adjustment, compounded over the horizon below.")).font = F_SUB

    put(ws, 3, 1, "Years TTM -> CY2028E:", F_LABEL_B, border=False)
    put(ws, 3, 2, 2.5, F_INPUT_B, nf="0.0", align=CENTER)
    HY = "$B$3"

    headers = ["Ticker", "Company", "Segment", "Rev TTM $M", "% AI/ DC opt", "% Fab optics", "% Telecom",
               "% Other", "AI/DC Rev $M", "Fab Rev $M", "Telecom Rev $M", "Other Rev $M",
               "Share adj (pts)", "2028E Rev $M", "Implied CAGR", "EV $M", "EV / 28E Rev"]
    hr = 5
    for i, h in enumerate(headers, start=1):
        hdr(ws, hr, i, h)
    ws.freeze_panes = "C6"

    mm = "'Market Model'"
    cagr_ai = f"{mm}!$C${mm_refs['cagr_ai']}"
    cagr_fab = f"{mm}!$C${mm_refs['cagr_fab']}"
    cagr_tel = f"{mm}!$C${mm_refs['cagr_tel']}"

    put(ws, 3, 4, "Other-bucket growth:", F_LABEL_B, border=False)
    put(ws, 3, 6, 0.03, F_INPUT_B, nf=NF_PCT0, align=CENTER)
    OTH_G = "$F$3"

    r = hr + 1
    first = r
    for seg in SEG_ORDER:
        for co in [c for c in UNIVERSE if c["seg"] == seg]:
            cr = comps_row[co["t"]]
            ai, fab, tel = co["exp"]
            put(ws, r, 1, co["t"], F_LABEL_B)
            put(ws, r, 2, co["name"], F_LABEL)
            put(ws, r, 3, SEG_SHORT[seg], F_LABEL)
            put(ws, r, 4, f"=Comps!I{cr}", F_CALC, nf=NF_M)
            put(ws, r, 5, ai / 100, F_INPUT, nf=NF_PCT0, align=CENTER)
            put(ws, r, 6, fab / 100, F_INPUT, nf=NF_PCT0, align=CENTER)
            put(ws, r, 7, tel / 100, F_INPUT, nf=NF_PCT0, align=CENTER)
            put(ws, r, 8, f"=1-E{r}-F{r}-G{r}", F_CALC, nf=NF_PCT0, align=CENTER)
            put(ws, r, 9, f"=IFERROR($D{r}*E{r},\"\")", F_CALC, nf=NF_M)
            put(ws, r, 10, f"=IFERROR($D{r}*F{r},\"\")", F_CALC, nf=NF_M)
            put(ws, r, 11, f"=IFERROR($D{r}*G{r},\"\")", F_CALC, nf=NF_M)
            put(ws, r, 12, f"=IFERROR($D{r}*H{r},\"\")", F_CALC, nf=NF_M)
            put(ws, r, 13, co["adj"] / 100, F_INPUT, nf="0.0%", align=CENTER)
            put(ws, r, 14,
                (f"=IFERROR(I{r}*(1+{cagr_ai}+M{r})^{HY}"
                 f"+J{r}*(1+{cagr_fab})^{HY}"
                 f"+K{r}*(1+{cagr_tel})^{HY}"
                 f"+L{r}*(1+{OTH_G})^{HY},\"\")"),
                F_CALC_B, nf=NF_M)
            put(ws, r, 15, f"=IFERROR((N{r}/D{r})^(1/{HY})-1,\"\")", F_CALC, nf=NF_PCT)
            put(ws, r, 16, f"=Comps!H{cr}", F_CALC, nf=NF_M)
            put(ws, r, 17, f"=IFERROR(IF(N{r}>0,P{r}/N{r},\"NM\"),\"NM\")", F_CALC_B, nf=NF_MULT)
            r += 1
    last = r - 1
    put(ws, r, 2, "Median", F_CALC_B, fill=FILL_TOTAL)
    for cidx, nf in [(15, NF_PCT), (17, NF_MULT)]:
        col = get_column_letter(cidx)
        put(ws, r, cidx, f"=IFERROR(MEDIAN({col}{first}:{col}{last}),\"\")", F_CALC_B, nf=nf, fill=FILL_TOTAL)
    for cidx in [1, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 16]:
        put(ws, r, cidx, "", fill=FILL_TOTAL)
    return ws


def build_scores(wb):
    ws = wb.create_sheet("Thematic Scores")
    ws.sheet_properties.tabColor = "BF8F00"
    set_widths(ws, [11, 24, 13] + [11] * len(THEMES) + [12, 7])
    ws.cell(row=1, column=1, value="Thematic Scoring Matrix (0 = none, 5 = highest exposure/quality)").font = F_TITLE
    ws.cell(row=2, column=1,
            value="Scores and weights are inputs (blue). Composite = SUMPRODUCT(weights, scores).").font = F_SUB

    wr = 4  # weights row
    put(ws, wr, 3, "Weight ->", F_LABEL_B, align=CENTER, border=False)
    for j, (_, w) in enumerate(THEMES):
        put(ws, wr, 4 + j, w, F_INPUT_B, nf=NF_PCT0, align=CENTER)
    wsum_col = 4 + len(THEMES)
    put(ws, wr, wsum_col, f"=SUM(D{wr}:{get_column_letter(3+len(THEMES))}{wr})", F_CALC_B, nf=NF_PCT0, align=CENTER)
    put(ws, wr, wsum_col + 1, "(=100%)", F_SUB, border=False)

    hr = 5
    hdr(ws, hr, 1, "Ticker"); hdr(ws, hr, 2, "Company"); hdr(ws, hr, 3, "Segment")
    for j, (tname, _) in enumerate(THEMES):
        hdr(ws, hr, 4 + j, tname)
    hdr(ws, hr, wsum_col, "Composite")
    hdr(ws, hr, wsum_col + 1, "Rank")
    ws.freeze_panes = "D6"

    r = hr + 1
    first = r
    w_first = get_column_letter(4)
    w_last = get_column_letter(3 + len(THEMES))
    for seg in SEG_ORDER:
        for co in [c for c in UNIVERSE if c["seg"] == seg]:
            put(ws, r, 1, co["t"], F_LABEL_B)
            put(ws, r, 2, co["name"], F_LABEL)
            put(ws, r, 3, SEG_SHORT[seg], F_LABEL)
            for j, s in enumerate(co["scores"]):
                put(ws, r, 4 + j, s, F_INPUT, nf="0", align=CENTER)
            comp_col = get_column_letter(wsum_col)
            put(ws, r, wsum_col,
                f"=SUMPRODUCT({w_first}${wr}:{w_last}${wr},{w_first}{r}:{w_last}{r})",
                F_CALC_B, nf=NF_SCORE, align=CENTER)
            r += 1
    last = r - 1
    comp_col = get_column_letter(wsum_col)
    for rr in range(first, last + 1):
        put(ws, rr, wsum_col + 1,
            f"=RANK({comp_col}{rr},{comp_col}${first}:{comp_col}${last})", F_CALC, nf="0", align=CENTER)
    return ws


def build_sensitivity(wb, mm_refs):
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_properties.tabColor = "BF8F00"
    set_widths(ws, [34, 11, 11, 11, 11, 11, 11, 11, 11])
    ws.cell(row=1, column=1, value="2028E Datacom Optical Transceiver Market Sensitivities ($B)").font = F_TITLE
    ws.cell(row=2, column=1, value=(
        "Grids recompute 2028E gross transceiver revenue from Market Model 2028E cells, flexing two "
        "drivers at a time and holding everything else at model values.")).font = F_SUB

    mm = "'Market Model'"
    H = lambda key: f"{mm}!$H${mm_refs[key]}"  # 2028E column

    # --- Grid 1: attach ratio x 1.6T ASP -----------------------------------
    r = 4
    put(ws, r, 1, "GRID 1: Optics attach ratio (rows)  x  1.6T ASP (columns)", F_SECTION, border=False)
    r += 1
    asps = [700, 800, 900, 1000, 1100, 1200, 1300]
    attaches = [2.5, 3.0, 3.5, 4.0, 4.5]
    hdr(ws, r, 1, "2028E market $B"); 
    for j, a in enumerate(asps):
        put(ws, r, 2 + j, a, F_INPUT_B, nf=NF_USD, align=CENTER, fill=FILL_GREY)
    hdr_row = r
    r += 1
    for i, att in enumerate(attaches):
        put(ws, r, 1, att, F_INPUT_B, nf=NF_UNITS, align=CENTER, fill=FILL_GREY)
        for j in range(len(asps)):
            col = get_column_letter(2 + j)
            f = (f"=({H('accel')}*$A{r}+{H('nonai_ports')})*"
                 f"({H('mix400')}*{H('asp400')}+{H('mix800')}*{H('asp800')}"
                 f"+{H('mix16')}*{col}${hdr_row})/1000")
            put(ws, r, 2 + j, f, F_CALC, nf=NF_B, align=CENTER)
        r += 1
    r += 1
    put(ws, r, 1, "Model case: attach and ASP per Market Model 2028E inputs; grid holds speed mix, other ASPs and non-AI ports constant.",
        F_SUB, border=False)
    r += 2

    # --- Grid 2: accelerator units x 1.6T mix --------------------------------
    put(ws, r, 1, "GRID 2: AI accelerator shipments, M units (rows)  x  1.6T port mix (columns)", F_SECTION, border=False)
    r += 1
    mixes = [0.30, 0.40, 0.52, 0.60, 0.70]
    units = [12, 15, 18, 21, 24]
    hdr(ws, r, 1, "2028E market $B")
    for j, m in enumerate(mixes):
        put(ws, r, 2 + j, m, F_INPUT_B, nf=NF_PCT0, align=CENTER, fill=FILL_GREY)
    hdr2 = r
    r += 1
    for i, u in enumerate(units):
        put(ws, r, 1, u, F_INPUT_B, nf=NF_UNITS, align=CENTER, fill=FILL_GREY)
        for j in range(len(mixes)):
            col = get_column_letter(2 + j)
            # 800G mix backfills: mix800 = 1 - mix400 - mix1.6T
            f = (f"=($A{r}*{H('attach')}+{H('nonai_ports')})*"
                 f"({H('mix400')}*{H('asp400')}"
                 f"+(1-{H('mix400')}-{col}${hdr2})*{H('asp800')}"
                 f"+{col}${hdr2}*{H('asp16')})/1000")
            put(ws, r, 2 + j, f, F_CALC, nf=NF_B, align=CENTER)
        r += 1
    r += 1
    put(ws, r, 1, "800G mix absorbs whatever 1.6T does not take (<=400G mix held at model value).", F_SUB, border=False)
    return ws


# ----------------------------------------------------------------------------
def main():
    with open("market_data.json") as f:
        data = json.load(f)
    as_of = data["as_of"]

    wb = Workbook()
    build_readme(wb, as_of)
    build_universe(wb)
    build_supply_chain(wb)
    _, mm_refs = build_market_model(wb)
    _, comps_row = build_comps(wb, data)
    build_exposure(wb, comps_row, mm_refs)
    build_scores(wb)
    build_sensitivity(wb, mm_refs)

    out = "Optical_Semiconductor_Supply_Chain_Model.xlsx"
    wb.save(out)
    print(f"Wrote {out}")
    print("Sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()
