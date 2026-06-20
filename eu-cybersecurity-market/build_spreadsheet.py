#!/usr/bin/env python3
"""Aggregate EU cybersecurity market research into a multi-sheet spreadsheet.

This script encodes figures gathered from public analyst reports, vendor
financial disclosures and EU/ENISA policy documents (see the "Sources" sheet
for the full citation list) and writes:

  - EU_Cybersecurity_Market.xlsx  (multi-sheet workbook, primary deliverable)
  - data/*.csv                    (one CSV per sheet, for git-friendly diffs)

All currency figures keep the unit used by the original source (mostly USD bn
for market-sizing analysts, EUR for EU policy/vendor figures). The unit is
always stated in the relevant column or note so values are never ambiguous.

Re-run with:  python3 build_spreadsheet.py
"""

from __future__ import annotations

import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(HERE, "data")
XLSX_PATH = os.path.join(HERE, "EU_Cybersecurity_Market.xlsx")

# ---------------------------------------------------------------------------
# Sheet data. Each entry: sheet name -> (intro_rows, header, list-of-rows)
# intro_rows are free-text lines placed above the table.
# ---------------------------------------------------------------------------

OVERVIEW = {
    "title": "Overview & Key Takeaways",
    "intro": [
        "EU / European Cybersecurity Market — Aggregated Research",
        "Compiled: June 2026. Currency units are stated per column (USD bn or EUR).",
        "'Europe' vs 'EU-27': analyst reports usually cover wider 'Europe' (incl. UK, "
        "Switzerland, sometimes Russia). EU-specific policy figures are flagged as such.",
        "",
        "KEY TAKEAWAYS:",
        "1. Europe cybersecurity market ~USD 50-82 bn in 2024-25; consensus ~USD 83-121 bn by 2030.",
        "2. Growth (CAGR) 8-14% depending on analyst scope; driven by NIS2, DORA, Cyber Resilience Act.",
        "3. Network security is the largest security-type segment; endpoint security is fastest growing.",
        "4. Firewall/network-security is led by US/Israeli vendors: Palo Alto, Fortinet, Cisco, Check Point.",
        "5. Europe's own enterprise-firewall segment was ~USD 3.5 bn in 2024 (~27% of the global firewall market).",
        "6. Largest EU-headquartered cyber players: Thales (FR), Bitdefender (RO), ESET (SK), plus UK Sophos/Darktrace.",
        "7. Sovereignty funding gap is large: Digital Europe earmarks only ~EUR 1.4 bn for cybersecurity (2021-27);",
        "   the 'EuroStack' proposal calls for EUR 300 bn over 10 years for the full sovereign tech stack.",
        "8. EU cyber skills shortage reached ~299,000 professionals in 2024 (wider Europe ~424,000).",
    ],
    "header": ["Metric", "Value", "Year", "Scope", "Source"],
    "rows": [
        ["Europe market size (low estimate)", "USD 50.22 bn", "2024", "Europe", "MarketsandMarkets"],
        ["Europe market size (high estimate)", "USD 81.81 bn", "2025", "Europe", "IMARC Group"],
        ["Consensus 2030 market size range", "USD 83-121 bn", "2030", "Europe", "MnM / IMARC"],
        ["Typical forecast CAGR range", "8.2% - 14.2%", "2025-31", "Europe", "Multiple analysts"],
        ["Europe share of global cyber market", "~25.6%", "2025", "Europe", "Fortune Business Insights"],
        ["Largest country market", "Germany (~22%)", "2025", "Europe", "IMARC Group"],
        ["Europe enterprise firewall market", "USD 3.54 bn", "2024", "Europe", "Mordor Intelligence"],
        ["Network-security global leader share", "28.4% (Palo Alto)", "2024", "Global", "Omdia"],
        ["EU cybersecurity skills shortage", "~299,000 people", "2024", "EU-27", "ENISA NIS Investments 2025"],
        ["Digital Europe cybersecurity budget", "EUR 1.372 bn", "2021-27", "EU-27", "EUR-Lex / EC"],
        ["EuroStack sovereign-tech proposal", "EUR 300 bn / 10y", "2025+", "EU (proposal)", "Bertelsmann Stiftung / CEPS"],
    ],
}

MARKET_SIZE = {
    "title": "EU/Europe Cybersecurity Market Size & Forecast",
    "intro": [
        "Analyst estimates differ by scope (countries included), definition (HW/SW/services) "
        "and methodology. Range is shown rather than a single number on purpose.",
        "Unit: USD billion unless noted.",
    ],
    "header": ["Source", "Base year", "Base size (USD bn)", "Forecast year",
               "Forecast size (USD bn)", "CAGR", "Notes"],
    "rows": [
        ["MarketsandMarkets", 2024, 50.22, 2030, 83.14, "8.7% (2025-30)",
         "Drivers: NIS2, Cyber Resilience Act; 2025 base 54.77"],
        ["Mordor Intelligence", 2025, 63.12, 2031, 115.66, "10.62% (2026-31)",
         "2026 size 69.82; NL fastest country CAGR 12.1%"],
        ["IMARC Group", 2025, 81.81, 2034, 165.73, "8.16% (2026-34)",
         "2030 anchor 121.10; services 54% of mkt; Germany 22%"],
        ["The Insight Partners", 2024, 68.6, 2031, 170.6, "14.2% (2025-31)",
         "Network security largest type; solutions largest component"],
        ["Fortune Business Insights", 2025, 55.98, 2026, 63.11, "n/a (Europe slice)",
         "Europe = 25.6% of global demand in 2025"],
        ["Dimension Market Research", "n/a", "n/a", 2035, 132.8, "n/a",
         "Cloud deployment largest share by 2026"],
    ],
}

COUNTRIES = {
    "title": "Key Country Markets",
    "intro": [
        "Country-level figures drawn from several analysts; treat as indicative, not "
        "directly comparable (different base years / definitions). Unit: USD bn / %.",
    ],
    "header": ["Country", "Indicative market size", "Year", "Note", "Source"],
    "rows": [
        ["Germany", "~22% of Europe; ~USD 11.36 bn", "2025/26", "Largest single market", "IMARC / Fortune BI"],
        ["United Kingdom", "+USD 11.55 bn added", "2026", "Largest growth contributor", "Fortune Business Insights"],
        ["France", "USD 7.97 bn (2026) -> 14.58 bn (2030)", "2026-30", "EUR 1.9 bn public quantum programme", "Fortune BI / Mordor"],
        ["Netherlands", "Highest country CAGR 12.1%", "to 2031", "Amsterdam internet exchange hub", "Mordor Intelligence"],
        ["Nordics", "Premium spend per capita", "2024-31", "High digitization + awareness", "Mordor Intelligence"],
    ],
}

SEGMENTS_TYPE = {
    "title": "Segments by Security Type",
    "intro": [
        "Security-type view of the market. Network security is consistently the largest "
        "type; endpoint is the fastest growing. Shares are qualitative where analysts do "
        "not publish exact splits.",
    ],
    "header": ["Security type", "Relative position", "Trend", "Source"],
    "rows": [
        ["Network security (incl. firewall, VPN, SASE)", "Largest type", "Steady; SASE-driven", "The Insight Partners / MRFR"],
        ["Endpoint security", "Mid", "Fastest growing (remote work, device sprawl)", "Market Research Future"],
        ["Cloud security", "Growing fast", "Fastest-growing deployment; ~62% cloud by 2026", "Dimension Market Research"],
        ["Application security", "Mid", "Rising with API/web exposure", "The Insight Partners"],
        ["Infrastructure / other security", "Smaller", "Stable", "The Insight Partners"],
    ],
}

SEGMENTS_SOLUTION = {
    "title": "Segments by Solution Type & Component",
    "intro": [
        "Solution-type and component view. Analysts disagree on the single largest "
        "solution (IAM vs SIEM) — both are cited as leaders; firewall & VPN is a core "
        "sub-segment. Component split: Solutions ~61-68% vs Services ~32-39%/54% (varies).",
    ],
    "header": ["Solution / component", "Relative position", "Note", "Source"],
    "rows": [
        ["IAM (Identity & Access Management)", "Largest solution (some analysts)", "Authentication/authorization demand", "Market Research Future"],
        ["Log management & SIEM", "Largest solution (other analysts)", "NIS2/CRA audit & monitoring", "MarketsandMarkets"],
        ["Firewall & VPN", "Core network sub-segment", "NGFW + SASE convergence", "MarketsandMarkets"],
        ["Antivirus / Antimalware / Endpoint", "Significant", "Endpoint protection", "MarketsandMarkets"],
        ["Data Loss Prevention (DLP) / encryption", "Significant", "GDPR-driven", "Market Research Future"],
        ["Component: Solutions", "~61-68% of market", "2026 estimate", "Fortune BI / Dimension MR"],
        ["Component: Services (managed + professional)", "~32-54% (analyst-dependent)", "Managed services fastest-growing", "IMARC / Dimension MR"],
        ["Deployment: Cloud vs On-premise", "Cloud rising to ~62%; on-prem 58% (IMARC)", "Methodology differs", "Dimension MR / IMARC"],
    ],
}

FIREWALL = {
    "title": "Firewall / Network Security Detail",
    "intro": [
        "Firewall is the headline network-security sub-segment requested. Figures cover "
        "the Europe enterprise firewall market plus global network-security context.",
        "Unit: USD bn / %.",
    ],
    "header": ["Metric", "Value", "Year", "Scope", "Source"],
    "rows": [
        ["Europe enterprise firewall market size", "USD 3.54 bn", "2024", "Europe", "Mordor Intelligence"],
        ["Europe share of global enterprise firewall market", "~27%", "2024", "Europe", "Industryresearch.co"],
        ["Global network-security market growth (Q4)", "+5.1% YoY", "Q4 2024", "Global", "Omdia"],
        ["Global network-security full-year growth", "+3.1%", "2024", "Global", "Omdia"],
        ["European enterprises deploying firewalls (compliance)", ">70%", "2024", "Europe", "Industryresearch.co"],
        ["Encrypted traffic share of enterprise networks", ">80%", "2024", "Europe", "Industryresearch.co"],
        ["Market structure", "Semi-consolidated; top 4 hold double-digit shares each", "2024", "Europe", "Omdia / Mordor"],
    ],
}

VENDORS_NETWORK = {
    "title": "Top Vendors — Network Security / Firewall",
    "intro": [
        "Network-security / firewall vendor shares. Omdia figures are global totals "
        "(closest published proxy; Europe mix is similar). Enterprise-firewall-only "
        "shares from Industryresearch.co. EU IDC report covers same leaders (no public %).",
    ],
    "header": ["Vendor", "HQ country", "Network-security share 2024 (global, Omdia)",
               "Enterprise firewall share (Industryresearch.co)", "Note"],
    "rows": [
        ["Palo Alto Networks", "USA", "28.4% (leader)", "~23%", "Platformization, NGFW + SASE"],
        ["Fortinet", "USA", "Double-digit (#2)", "~19%", "FortiGate; secure networking + SASE"],
        ["Cisco Systems", "USA", "Double-digit (#3)", "Top tier", "Boosted by Splunk acquisition"],
        ["Check Point", "Israel", "Double-digit (#4)", "Top tier", "OEMs Wiz for cloud-native"],
        ["Zscaler", "USA", "Rising (SSE/SASE)", "n/a", "Cloud-native SSE; covered by EU IDC"],
        ["Others (Juniper, Sophos, WatchGuard, SonicWall, Barracuda, Huawei, Stormshield)",
         "Mixed", "Remainder", "Remainder", "Stormshield (FR) is the notable EU NGFW player"],
    ],
}

VENDORS_GLOBAL = {
    "title": "Top Vendors — Global Cybersecurity (context)",
    "intro": [
        "Whole-of-cybersecurity vendor shares (global, Q2 2024) for context — these are "
        "the platforms competing for European spend. Top 12 vendors = 53.2% of spend.",
        "Source: Omdia / Futurum (reported by Security MEA), Q2 2024.",
    ],
    "header": ["Vendor", "HQ country", "Market share Q2 2024", "Revenue growth YoY", "Note"],
    "rows": [
        ["Palo Alto Networks", "USA", "9.7%", "+11.2%", "Leader; Prisma/Cortex/SASE"],
        ["Fortinet", "USA", "7.0%", "+11.1%", "SecOps + SASE pivot"],
        ["Microsoft", "USA", "6.1%", "+18.6%", "Defender/Sentinel/Entra via E5"],
        ["Cisco", "USA", "5.8%", "+5.3%", "Excl. Splunk"],
        ["CrowdStrike", "USA", "4.5%", "+33.2%", "Falcon platform; fastest grower"],
        ["Okta", "USA", "3.6%", "+16.6%", "Identity"],
        ["Check Point", "Israel", "3.5%", "+6.7%", "Network security"],
        ["Zscaler", "USA", "3.0%", "+31.3%", "SSE/SASE"],
        ["Symantec (Broadcom)", "USA", "2.7%", "+3.2%", "Legacy enterprise"],
        ["IBM", "USA", "2.5%", "+2.5%", "QRadar sold to Palo Alto"],
        ["Trellix", "USA", "2.4%", "+2.1%", "McAfee/FireEye merger"],
        ["Splunk (Cisco)", "USA", "2.4%", "+13.5%", "Security analytics"],
        ["Others", "Mixed", "46.8%", "+6.9%", "Long tail incl. EU vendors"],
    ],
}

VENDORS_EU = {
    "title": "Leading EU / European-HQ Vendors",
    "intro": [
        "European-headquartered cybersecurity vendors and their latest disclosed revenue. "
        "Most relevant for 'sovereignty' since these are EU-controlled suppliers. UK/CH "
        "firms flagged (European but outside EU-27). Unit: EUR.",
    ],
    "header": ["Vendor", "HQ country", "EU-27?", "Latest revenue", "FY", "Focus area"],
    "rows": [
        ["Thales", "France", "Yes", "Group EUR 20.6 bn; Cyber & Digital EUR 4.024 bn", "2024", "Encryption, data security (Imperva), identity, critical infra"],
        ["Atos / Eviden", "France", "Yes", "Part of Atos group (restructuring)", "2024", "Managed security, consulting, SOC"],
        ["Bitdefender", "Romania", "Yes", "Private (est. ~USD 600m+)", "2024", "Endpoint, consumer + enterprise"],
        ["ESET", "Slovakia", "Yes", "EUR 691 m", "2024", "Endpoint, MDR; +9% YoY"],
        ["WithSecure", "Finland", "Yes", "Group EUR 147.4 m", "2024", "Cloud protection, MDR (B2B)"],
        ["F-Secure", "Finland", "Yes", "Consumer security (split from WithSecure)", "2024", "Consumer security"],
        ["Stormshield", "France", "Yes", "Subsidiary of Airbus Defence", "2024", "NGFW, sovereign firewall"],
        ["Telefonica Tech", "Spain", "Yes", "Part of Telefonica", "2024", "Managed security services"],
        ["Orange Cyberdefense", "France", "Yes", "Part of Orange", "2024", "MSSP, threat intel"],
        ["CY4GATE", "Italy", "Yes", "Listed (small-cap)", "2024", "Cyber intelligence, decision intel"],
        ["Sophos", "UK", "No (UK)", "Private (Thoma Bravo)", "2024", "Endpoint, firewall, MDR"],
        ["Darktrace", "UK", "No (UK)", "Private (Thoma Bravo, ~USD 800m ARR)", "2024", "AI-driven detection & response"],
        ["NCC Group", "UK", "No (UK)", "Listed", "2024", "Security consulting/testing"],
        ["Kudelski Security", "Switzerland", "No (CH)", "Part of Kudelski", "2024", "MSSP, advisory"],
    ],
}

SOVEREIGNTY = {
    "title": "EU Cyber Sovereignty — Investment Needs",
    "intro": [
        "How much investment is needed for EU cybersecurity sovereignty? There is no single "
        "official 'cyber sovereignty' number; the table aggregates (a) committed EU funding, "
        "(b) broader sovereign-tech investment proposals that include cybersecurity, and "
        "(c) the talent/skills gap that any sovereignty push must close.",
        "Unit: EUR unless noted.",
    ],
    "header": ["Item", "Amount", "Period", "Type", "Source"],
    "rows": [
        ["Digital Europe Programme — total budget", "EUR 7.59 bn (amended ~8.2 bn)", "2021-2027", "Committed (EU)", "European Commission / EUR-Lex"],
        ["Digital Europe — 'Cybersecurity & Trust' objective", "EUR 1.372 bn", "2021-2027", "Committed (EU)", "EUR-Lex (Reg 2021/694, amended 2025/38)"],
        ["Digital Europe — 2023-24 cybersecurity work programme", "EUR 375 m", "2023-2024", "Committed (EU)", "ECCC / European Commission"],
        ["ECCC multiannual work programmes (all digital)", "EUR 1.284 bn", "2023-2024", "Committed (EU)", "European Commission"],
        ["EuroStack — European Sovereign Technology Fund", "EUR 300 bn", "10 years", "Proposal", "Bertelsmann Stiftung / CEPS / EuroStack"],
        ["EuroStack — initial demonstrator tranche", "EUR 10 bn", "Initial", "Proposal", "Bertelsmann Stiftung / CEPS"],
        ["EuroStack — alternative full-cost estimate", "> EUR 5 trillion", "n/a", "Critique", "Chamber of Progress (via Politico)"],
        ["InvestAI — AI hardware mobilization (related stack layer)", "EUR 200 bn", "2025+", "Mobilization target", "European Commission (via Politico)"],
        ["AI gigafactory sites investment (related)", "Up to EUR 2 bn", "2024+", "Committed", "European Commission (via Politico)"],
        ["Draghi report — overall EU investment gap (all sectors)", "~EUR 800 bn / year", "Annual", "Diagnosis (not cyber-only)", "Draghi Competitiveness Report 2024"],
    ],
}

SKILLS = {
    "title": "Skills Gap & Spending Behaviour",
    "intro": [
        "Sovereignty is constrained by people, not just money. ENISA tracks the EU cyber "
        "workforce gap and how organisations spend. Closing this gap is a core part of any "
        "sovereignty investment case (EU Cybersecurity Skills Academy).",
    ],
    "header": ["Metric", "Value", "Year", "Scope", "Source"],
    "rows": [
        ["Cyber skills shortage", "~299,000 professionals (+9% YoY)", "2024", "EU-27", "ENISA NIS Investments 2025"],
        ["Cyber skills shortage (wider Europe)", "~424,000 professionals", "2024", "Europe", "ENISA NIS Investments 2025"],
        ["Cyber skills shortage (global)", "~4.8 million professionals", "2024", "Global", "ENISA NIS Investments 2025"],
        ["Cyber spend as share of IT budget", "~9% (median EUR 1.5 m)", "2024-25", "EU-27", "ENISA NIS Investments 2025"],
        ["Orgs struggling to attract talent", "76%", "2025", "EU-27", "ENISA NIS Investments 2025"],
        ["Orgs struggling to retain talent", "71%", "2025", "EU-27", "ENISA NIS Investments 2025"],
        ["IDC: EU cyber investment growth", "~+12%", "2024", "Europe", "IDC (via IMARC)"],
        ["EU policy response", "Cybersecurity Skills Academy; ECSF; CyberHEAD", "2023+", "EU-27", "ENISA / European Commission"],
    ],
}

SOURCES = {
    "title": "Sources",
    "intro": [
        "Full source list. Market-research vendor figures are publicly cited estimates and "
        "vary by methodology; EU/ENISA figures are from official documents. Always verify "
        "against the original report before using in a transaction or investment memo.",
    ],
    "header": ["#", "Source", "Used for", "URL"],
    "rows": [
        [1, "MarketsandMarkets — Europe Cybersecurity Market", "Market size, segments, EU vendors",
         "https://www.marketsandmarkets.com/Market-Reports/europe-cybersecurity-market-156644743.html"],
        [2, "Mordor Intelligence — Europe Cybersecurity / Enterprise Firewall", "Market size, firewall, countries",
         "https://www.mordorintelligence.com/industry-reports/europe-cybersecurity-market"],
        [3, "IMARC Group — Europe Cybersecurity Market", "Market size, country & component split",
         "https://www.imarcgroup.com/europe-cybersecurity-market"],
        [4, "The Insight Partners — Europe Cybersecurity Market", "Market size, type/component split",
         "https://www.theinsightpartners.com/reports/europe-cybersecurity-market"],
        [5, "Fortune Business Insights — Cybersecurity Market", "Europe share, country values",
         "https://www.fortunebusinessinsights.com/industry-reports/cyber-security-market-101165"],
        [6, "Dimension Market Research — Europe Cybersecurity", "Deployment & component split",
         "https://dimensionmarketresearch.com/report/europe-cybersecurity-market/"],
        [7, "Market Research Future — Europe Cybersecurity", "Solution & security-type leaders",
         "https://www.marketresearchfuture.com/reports/europe-cybersecurity-market-46043"],
        [8, "Omdia (via Informa) — Network Security Market Q4 2024", "Network-security vendor shares",
         "https://omdia.tech.informa.com/pr/2025/mar/network-security-market-grows-5point1-percent-yearoveryear-in-q4-2024-omdia-reports"],
        [9, "Industryresearch.co — Enterprise Firewall Market", "Firewall shares, Europe %",
         "https://www.industryresearch.co/market-reports/enterprise-firewall-market-308811"],
        [10, "IDC — European Network Security Market Shares 2024", "EU network-security vendors",
         "https://my.idc.com/getdoc.jsp?containerId=EUR152829325"],
        [11, "Security MEA / Omdia — Cybersecurity spending Q2 2024", "Global vendor shares",
         "https://securitymea.com/2024/09/26/platform-adoption-boosts-cybersecurity-spending-to-us21-1-billion-in-q2-2024/"],
        [12, "Thales — 2024 Full-Year Results", "Thales revenue & Cyber & Digital segment",
         "https://www.thalesgroup.com/en/group/investors"],
        [13, "ESET — Annual Report 2024", "ESET revenue",
         "https://www.eset.com/us/about/newsroom/company/eset-2024-annual-report-profit-and-revenue-growth-continues-rd-investment-delivers-strong-returns/"],
        [14, "WithSecure — Restated 2024 financials", "WithSecure revenue",
         "https://www.globenewswire.com/news-release/2025/04/22/3065294/0/en/WithSecure-publishes-restated-2024-financial-information-for-segments.html"],
        [15, "European Commission — Digital Europe Programme", "Sovereignty funding (total)",
         "https://commission.europa.eu/funding-and-tenders/find-funding/eu-funding-programmes/digital-europe-programme_en"],
        [16, "EUR-Lex — Regulation (EU) 2021/694 (amended 2025/38)", "Cybersecurity & Trust budget",
         "https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=legissum:4526703"],
        [17, "ECCC / EC — EUR 1.3 bn Digital Europe work programmes", "Cyber work-programme funding",
         "https://ec.europa.eu/newsroom/ECCC/items/781679/"],
        [18, "Bertelsmann Stiftung — EuroStack (2025)", "EUR 300 bn sovereign-tech proposal",
         "https://www.bertelsmann-stiftung.de/fileadmin/files/user_upload/EuroStack__2025_final.pdf"],
        [19, "CEPS — A bold proposal to build the EuroStack", "Sovereign Technology Fund",
         "https://www.ceps.eu/a-bold-proposal-to-build-the-eurostack-because-doing-nothing-isnt-an-option-anymore/"],
        [20, "Politico — Push for EuroStack", "Investment-size debate, InvestAI",
         "https://www.politico.eu/article/push-for-eurostack-as-eu-us-tech-tensions-grow/"],
        [21, "ENISA — NIS Investments 2025 / What's Driving Cyber Investments", "Skills gap, spend levels",
         "https://www.enisa.europa.eu/news/whats-driving-cybersecurity-investments-and-where-lie-the-challenges"],
        [22, "ENISA — 2024 Report on State of Cybersecurity in the Union", "Skills Academy, policy",
         "https://www.enisa.europa.eu/publications/2024-report-on-the-state-of-the-cybersecurity-in-the-union"],
    ],
}

SHEETS = [
    ("Overview", OVERVIEW),
    ("Market Size", MARKET_SIZE),
    ("Country Markets", COUNTRIES),
    ("Segments by Type", SEGMENTS_TYPE),
    ("Segments by Solution", SEGMENTS_SOLUTION),
    ("Firewall Detail", FIREWALL),
    ("Vendors - Network", VENDORS_NETWORK),
    ("Vendors - Global", VENDORS_GLOBAL),
    ("Vendors - EU", VENDORS_EU),
    ("Sovereignty Investment", SOVEREIGNTY),
    ("Skills & Spending", SKILLS),
    ("Sources", SOURCES),
]

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

TITLE_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FILL = PatternFill("solid", fgColor="2E5496")
INTRO_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FONT = Font(bold=True, color="FFFFFF", size=14)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_sheet(ws, spec):
    row = 1
    ncols = len(spec["header"])

    ws.cell(row=row, column=1, value=spec["title"])
    ws.cell(row=row, column=1).font = TITLE_FONT
    ws.cell(row=row, column=1).fill = TITLE_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 22
    row += 2

    for line in spec.get("intro", []):
        ws.cell(row=row, column=1, value=line)
        ws.cell(row=row, column=1).alignment = WRAP
        if line:
            ws.cell(row=row, column=1).fill = INTRO_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        row += 1
    row += 1

    header_row = row
    for c, name in enumerate(spec["header"], start=1):
        cell = ws.cell(row=row, column=c, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        cell.border = BORDER
    row += 1

    for data_row in spec["rows"]:
        for c, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=c, value=value)
            cell.alignment = WRAP
            cell.border = BORDER
        row += 1

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    widths = [28] + [22] * (ncols - 1)
    if spec is SOURCES:
        widths = [4, 42, 30, 60]
    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = widths[c - 1] if c - 1 < len(widths) else 22


def write_csv(name, spec):
    safe = name.replace(" ", "_").replace("-", "").replace("__", "_")
    path = os.path.join(DATA_DIR, f"{safe}.csv")
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(spec["header"])
        for r in spec["rows"]:
            w.writerow(r)
    return path


def main():
    os.makedirs(DATA_DIR, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    for name, spec in SHEETS:
        ws = wb.create_sheet(title=name[:31])
        write_sheet(ws, spec)
        write_csv(name, spec)
    wb.save(XLSX_PATH)
    print(f"Wrote workbook: {XLSX_PATH}")
    print(f"Wrote {len(SHEETS)} CSVs to: {DATA_DIR}")


if __name__ == "__main__":
    main()
