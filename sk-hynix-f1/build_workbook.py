"""Build an Excel workbook aggregating the key tables from the SK hynix Form F-1.

Source: SK hynix Inc. Form F-1 (US SEC, filed June 24, 2026) — the foreign-private-issuer
equivalent of an S-1, for a Nasdaq ADS listing under ticker "SKHY".

All monetary figures are in billions of Korean Won (W) unless otherwise noted, IFRS,
consolidated. USD convenience rate per prospectus: W1,523.5 = US$1.00 (Mar 31, 2026).

Run:  python build_workbook.py
Output:  SK_hynix_F1_key_tables.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

OUTFILE = "SK_hynix_F1_key_tables.xlsx"

# ---- Shared styles -------------------------------------------------------
NAVY = "1F3864"
BLUE = "2E5496"
LIGHT = "D9E1F2"
GREY = "F2F2F2"
GREEN = "375623"

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
HDR_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
LABEL_FONT = Font(name="Calibri", size=11, bold=False)
BOLD = Font(name="Calibri", size=11, bold=True)
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="595959")

TITLE_FILL = PatternFill("solid", fgColor=NAVY)
HDR_FILL = PatternFill("solid", fgColor=BLUE)
LIGHT_FILL = PatternFill("solid", fgColor=LIGHT)
TOTAL_FILL = PatternFill("solid", fgColor=GREY)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

NUM = "#,##0"
NUM1 = "#,##0.0"
PCT = "0.0%"

PERIODS = ["Q1 2026", "Q1 2025", "FY 2025", "FY 2024", "FY 2023"]


def title_row(ws, text, ncols, subtitle=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26
    start = 2
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        s = ws.cell(row=2, column=1, value=subtitle)
        s.font = NOTE_FONT
        s.alignment = LEFT
        start = 3
    return start


def header(ws, row, headers, first_width=42):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.column_dimensions["A"].width = first_width
    for j in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 13
    return row + 1


def write_rows(ws, start, rows, numfmt=NUM, total_labels=None, pct_rows=None):
    """rows: list of (label, [values...]). Values may be numbers, formula strings, or None."""
    total_labels = total_labels or set()
    pct_rows = pct_rows or set()
    r = start
    for label, vals in rows:
        is_total = label in total_labels
        is_pct = label in pct_rows
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = BOLD if is_total else LABEL_FONT
        lc.alignment = LEFT
        lc.border = BORDER
        if is_total:
            lc.fill = TOTAL_FILL
        for j, v in enumerate(vals, start=2):
            c = ws.cell(row=r, column=j, value=v)
            c.alignment = RIGHT
            c.border = BORDER
            c.number_format = PCT if is_pct else numfmt
            if is_total:
                c.font = BOLD
                c.fill = TOTAL_FILL
        r += 1
    return r


def note(ws, row, ncols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = NOTE_FONT
    c.alignment = LEFT
    return row + 1


wb = Workbook()

# =========================================================================
# 1. OVERVIEW
# =========================================================================
ws = wb.active
ws.title = "Overview"
ws.sheet_view.showGridLines = False
r = title_row(ws, "SK hynix Inc. — Form F-1 (Nasdaq ADS Listing): Key Tables", 2)
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 78

facts = [
    ("Issuer", "SK hynix Inc. (Republic of Korea)"),
    ("Filing", "Form F-1 (foreign-issuer equivalent of an S-1), U.S. SEC"),
    ("Filing date", "June 24, 2026"),
    ("Security offered", "American Depositary Shares (ADSs); 10 ADSs = 1 common share"),
    ("Proposed Nasdaq ticker", "SKHY (Nasdaq Global Select Market)"),
    ("Existing listings", "KOSPI: 000660; Luxembourg GDR: HYNSE"),
    ("Max new shares", "17,790,000 (~2.50% of 712,702,365 shares outstanding)"),
    ("Underwriters", "BofA Securities, Citigroup, Goldman Sachs (Asia), J.P. Morgan"),
    ("Depositary", "Citibank, N.A."),
    ("Lock-up", "90 days"),
    ("Reported target raise", "Up to ~W45.45 trillion (~US$29.4bn); ADR ~255,500 won; Nasdaq debut targeted Jul 10, 2026 (subject to bookbuilding)"),
    ("Use of proceeds", "Korea facility capex (Yongin Fab 1; Cheongju P&T7) + EUV scanners (~W11.9tn)"),
    ("Controlling shareholder", "SK square (20.50%); part of SK Group"),
    ("Accounting / currency", "IFRS, consolidated; W in billions; USD rate W1,523.5 = US$1.00 (Mar 31, 2026)"),
    ("Market position (Q1 2026, IDC)", "DRAM incl. HBM #2 (29.1%); HBM #1 (56.4%); NAND #2 (18.5%)"),
    ("Employees", "47,639 consolidated (35,321 Korea; 11,333 China; 591 U.S.)"),
]
for label, val in facts:
    lc = ws.cell(row=r, column=1, value=label)
    lc.font = BOLD
    lc.alignment = LEFT
    lc.fill = LIGHT_FILL
    lc.border = BORDER
    vc = ws.cell(row=r, column=2, value=val)
    vc.alignment = LEFT
    vc.border = BORDER
    r += 1
r += 1
r = note(ws, r, 2, "Tabs: Income Statement | Balance Sheet | Cash Flow | Adj. EBITDA | "
         "Revenue by Product | Revenue by Region | Use of Proceeds | Dividends | Shareholders | KPIs & Ratios")
note(ws, r, 2, "Source: SK hynix Inc. Form F-1, filed June 24, 2026, plus FY2025 results & June 24, 2026 Korean filing press reporting.")

# =========================================================================
# 2. INCOME STATEMENT
# =========================================================================
ws = wb.create_sheet("Income Statement")
ws.sheet_view.showGridLines = False
r = title_row(ws, "Consolidated Statements of Comprehensive Income (Loss)",
              6, "In billions of Won (IFRS, consolidated). Operating profit derived = Gross profit - SG&A - R&D.")
hr = r
r = header(ws, r, ["W (billions)"] + PERIODS)
data_start = r
# values per period: Q1'26, Q1'25, FY25, FY24, FY23
is_rows = [
    ("Revenue", [52576, 17639, 97147, 66193, 32766]),
    ("Cost of sales", [10897, 7537, 38456, 34365, 33299]),
    ("Gross profit (loss)", [41679, 10102, 58691, 31828, -533]),
    ("Selling & administrative expenses", [1618, 1190, 5019, 3924, 3446]),
    ("Research & development expenses", [2451, 1472, 6466, 4436, 3751]),
]
r = write_rows(ws, r, is_rows)
# Operating profit (formula): gross - sga - r&d
op_row = r
ws.cell(row=r, column=1, value="Operating profit (loss) *").font = BOLD
ws.cell(row=r, column=1).fill = TOTAL_FILL
ws.cell(row=r, column=1).border = BORDER
ws.cell(row=r, column=1).alignment = LEFT
gp = data_start + 2
sga = data_start + 3
rd = data_start + 4
for j in range(2, 7):
    col = get_column_letter(j)
    c = ws.cell(row=r, column=j, value=f"={col}{gp}-{col}{sga}-{col}{rd}")
    c.number_format = NUM
    c.font = BOLD
    c.fill = TOTAL_FILL
    c.alignment = RIGHT
    c.border = BORDER
r += 1
is_rows2 = [
    ("Finance income", [17056, 2687, 16373, 4855, 2262]),
    ("Finance expenses", [3023, 765, 12505, 5708, 6093]),
    ("Share of profit (loss) of equity-accounted investees", [-27, -41, -565, -38, 15]),
    ("Other income", [15, 79, 333, 1477, 624]),
    ("Other expenses", [15, 102, 378, 167, 735]),
    ("Profit (loss) before income tax", [51617, 9299, 50466, 23885, -11658]),
    ("Income tax expense (benefit)", [11271, 1191, 7518, 4088, -2520]),
    ("Profit (loss) for the period", [40346, 8108, 42948, 19797, -9138]),
]
r = write_rows(ws, r, is_rows2,
               total_labels={"Profit (loss) before income tax", "Profit (loss) for the period"})
rev_row = data_start
np_row = r - 1
# Margins block
r += 1
ws.cell(row=r, column=1, value="Margins").font = BOLD
r += 1
for lbl, numr in [("Gross margin", gp), ("Operating margin", op_row), ("Net margin", np_row)]:
    lc = ws.cell(row=r, column=1, value=lbl)
    lc.font = LABEL_FONT
    lc.border = BORDER
    lc.alignment = LEFT
    for j in range(2, 7):
        col = get_column_letter(j)
        c = ws.cell(row=r, column=j, value=f"={col}{numr}/{col}{rev_row}")
        c.number_format = PCT
        c.alignment = RIGHT
        c.border = BORDER
    r += 1
r += 1
note(ws, r, 6, "* Operating profit is not a separate line in the F-1 summary table; it is computed as "
     "Gross profit - SG&A - R&D and matches SK hynix's reported figures (e.g., FY2025 W47,206bn).")
ws.freeze_panes = "B" + str(data_start)

# =========================================================================
# 3. BALANCE SHEET
# =========================================================================
ws = wb.create_sheet("Balance Sheet")
ws.sheet_view.showGridLines = False
bs_periods = ["Mar 31, 2026", "Dec 31, 2025", "Dec 31, 2024", "Dec 31, 2023"]
r = title_row(ws, "Consolidated Statements of Financial Position", 5,
              "In billions of Won (IFRS, consolidated).")
r = header(ws, r, ["W (billions)"] + bs_periods)
bs_rows = [
    ("Cash and cash equivalents", [21167, 14924, 11205, 7587]),
    ("Short-term financial instruments", [18220, 14680, 2382, 473]),
    ("Short-term investment assets", [14943, 5339, 569, 861]),
    ("Trade receivables, net", [33808, 18199, 13019, 6600]),
    ("Inventories, net", [15974, 14289, 13314, 13481]),
    ("Other current assets", [2394, 2027, 1790, 1466]),
    ("Total current assets", [106506, 69458, 42279, 30468]),
    ("Investments in associates and JV", [1356, 1321, 1941, 1367]),
    ("Long-term investment assets", [20658, 14547, 4041, 4106]),
    ("Loans and other receivables, net", [424, 420, 444, 475]),
    ("Property, plant and equipment, net", [82052, 77503, 60157, 52705]),
    ("Right-of-use assets, net", [2354, 2336, 2487, 2695]),
    ("Intangible assets, net", [4051, 4049, 4019, 3835]),
    ("Deferred tax assets", [1832, 3660, 2812, 2989]),
    ("Other non-current assets", [3596, 2812, 1675, 1690]),
    ("Total non-current assets", [116323, 106650, 77576, 69862]),
    ("Total assets", [222829, 176108, 119855, 100330]),
    ("Trade payables", [2798, 2848, 2277, 1846]),
    ("Other payables", [7903, 6434, 6967, 3293]),
    ("Other non-trade payables (current)", [6135, 6283, 3984, 1689]),
    ("Borrowings (current)", [5891, 8162, 5252, 9857]),
    ("Other financial liabilities", [1598, 4914, 1742, 1479]),
    ("Current tax liabilities", [14580, 7024, 3084, 44]),
    ("Lease liabilities (current)", [526, 547, 588, 631]),
    ("Other current liabilities", [1270, 1167, 1071, 2169]),
    ("Total current liabilities", [40701, 37379, 24965, 21008]),
    ("Long-term other payables", [381, 375, 477, 3144]),
    ("Other non-trade payables (non-current)", [21, 20, 52, 97]),
    ("Borrowings (non-current)", [13427, 14086, 17431, 19611]),
    ("Deferred tax liabilities", [285, 248, 218, 114]),
    ("Lease liabilities (non-current)", [1988, 1963, 2180, 2398]),
    ("Other non-current liabilities", [1646, 1370, 616, 455]),
    ("Total non-current liabilities", [17748, 18062, 20974, 25819]),
    ("Total liabilities", [58449, 55441, 45940, 46826]),
    ("Capital stock", [3658, 3658, 3658, 3658]),
    ("Capital surplus", [8510, 8954, 4487, 4373]),
    ("Other equity", [-368, -1349, -2192, -2269]),
    ("Accumulated other comprehensive income", [3745, 2677, 2532, 1014]),
    ("Retained earnings", [148746, 106577, 65418, 46729]),
    ("Equity attributable to owners of the parent", [164291, 120516, 73903, 53504]),
    ("Non-controlling interests", [89, 151, 12, -1]),
    ("Total equity", [164380, 120667, 73916, 53504]),
    ("Total liabilities and equity", [222829, 176108, 119855, 100330]),
]
totals = {"Total current assets", "Total non-current assets", "Total assets",
          "Total current liabilities", "Total non-current liabilities", "Total liabilities",
          "Equity attributable to owners of the parent", "Total equity",
          "Total liabilities and equity"}
data_start = r
r = write_rows(ws, r, bs_rows, total_labels=totals)
ws.freeze_panes = "B" + str(data_start)

# =========================================================================
# 4. CASH FLOW
# =========================================================================
ws = wb.create_sheet("Cash Flow")
ws.sheet_view.showGridLines = False
r = title_row(ws, "Consolidated Statements of Cash Flows (selected)", 6,
              "In billions of Won (IFRS, consolidated).")
r = header(ws, r, ["W (billions)"] + PERIODS)
cf_rows = [
    ("Capital expenditures (PP&E acquisitions)", [7657, 6284, 27519, 15946, 8325]),
    ("Net cash provided by operating activities", [26330, 9024, 53373, 29796, 4278]),
    ("Net cash used in investing activities", [-17635, -8218, -48054, -18005, -7335]),
    ("Net cash provided by (used in) financing activities", [-2951, 509, -1445, -8704, 5697]),
    ("Net increase in cash and cash equivalents", [6243, 1353, 3719, 3618, 2610]),
]
data_start = r
r = write_rows(ws, r, cf_rows, total_labels={"Net increase in cash and cash equivalents"})
r += 1
note(ws, r, 6, "Capital expenditures represent cash outflows for acquisition of PP&E (included within investing activities).")
ws.freeze_panes = "B" + str(data_start)

# =========================================================================
# 5. ADJUSTED EBITDA
# =========================================================================
ws = wb.create_sheet("Adj. EBITDA")
ws.sheet_view.showGridLines = False
r = title_row(ws, "Reconciliation of Profit (Loss) to Adjusted EBITDA (non-IFRS)", 6,
              "In billions of Won.")
r = header(ws, r, ["W (billions)"] + PERIODS)
eb_rows = [
    ("Profit (loss) for the period", [40346, 8108, 42948, 19797, -9138]),
    ("ADD: Income tax expense (benefit)", [11271, 1191, 7518, 4088, -2520]),
    ("SUBTRACT: Finance income", [-17056, -2687, -16373, -4855, -2262]),
    ("ADD: Finance expenses", [3023, 765, 12505, 5708, 6093]),
    ("SUBTRACT: Share of profit (loss) of equity-accounted investees", [27, 41, 565, 38, -15]),
    ("SUBTRACT: Other income", [-15, -79, -333, -1477, -624]),
    ("ADD: Other expenses", [15, 102, 378, 167, 735]),
    ("ADD: Depreciation and amortization", [3726, 3334, 13890, 12545, 13619]),
    ("Adjusted EBITDA", [41336, 10774, 61096, 36012, 5889]),
]
data_start = r
r = write_rows(ws, r, eb_rows, total_labels={"Adjusted EBITDA"})
r += 1
note(ws, r, 6, "Signs reflect the add/subtract direction in the reconciliation. Adjusted EBITDA is a non-IFRS measure.")
ws.freeze_panes = "B" + str(data_start)

# =========================================================================
# 6. REVENUE BY PRODUCT
# =========================================================================
ws = wb.create_sheet("Revenue by Product")
ws.sheet_view.showGridLines = False
r = title_row(ws, "Revenue by Principal Product Category", 11, "In billions of Won and % of total.")
# two-tier header
hdr1 = ["W (billions)"]
for p in PERIODS:
    hdr1 += [p, "%"]
r = header(ws, r, hdr1)
data_start = r
prod = [
    ("DRAM", [40659, 0.773, 14037, 0.796, 74904, 0.771, 44732, 0.676, 20769, 0.634]),
    ("NAND Flash", [11574, 0.220, 3229, 0.183, 20690, 0.213, 19274, 0.291, 9653, 0.295]),
    ("Other products", [343, 0.007, 373, 0.021, 1552, 0.016, 2187, 0.033, 2344, 0.072]),
    ("Total revenue", [52576, 1.0, 17639, 1.0, 97147, 1.0, 66193, 1.0, 32766, 1.0]),
]
for label, vals in prod:
    is_total = label == "Total revenue"
    lc = ws.cell(row=r, column=1, value=label)
    lc.font = BOLD if is_total else LABEL_FONT
    lc.alignment = LEFT
    lc.border = BORDER
    if is_total:
        lc.fill = TOTAL_FILL
    for j, v in enumerate(vals, start=2):
        c = ws.cell(row=r, column=j, value=v)
        c.alignment = RIGHT
        c.border = BORDER
        c.number_format = PCT if (j % 2 == 1) else NUM
        if is_total:
            c.font = BOLD
            c.fill = TOTAL_FILL
    r += 1
for j in range(2, 12):
    ws.column_dimensions[get_column_letter(j)].width = 11
ws.freeze_panes = "B" + str(data_start)

# =========================================================================
# 7. REVENUE BY REGION
# =========================================================================
ws = wb.create_sheet("Revenue by Region")
ws.sheet_view.showGridLines = False
r = title_row(ws, "Revenue by Region (by location of sales entity)", 11, "In billions of Won and % of total.")
r = header(ws, r, hdr1)
data_start = r
region = [
    ("United States", [33999, 0.647, 12795, 0.725, 66885, 0.688, 41961, 0.634, 15390, 0.470]),
    ("China", [12797, 0.243, 2694, 0.153, 19136, 0.197, 15534, 0.235, 10110, 0.309]),
    ("Asia (ex-China/Korea)", [4473, 0.085, 1258, 0.071, 7216, 0.074, 5381, 0.081, 4297, 0.131]),
    ("Europe", [1128, 0.021, 449, 0.025, 1977, 0.020, 1413, 0.021, 935, 0.029]),
    ("Korea", [179, 0.003, 443, 0.025, 1932, 0.020, 1904, 0.029, 2034, 0.062]),
    ("Total revenue", [52576, 1.0, 17639, 1.0, 97147, 1.0, 66193, 1.0, 32766, 1.0]),
]
for label, vals in region:
    is_total = label == "Total revenue"
    lc = ws.cell(row=r, column=1, value=label)
    lc.font = BOLD if is_total else LABEL_FONT
    lc.alignment = LEFT
    lc.border = BORDER
    if is_total:
        lc.fill = TOTAL_FILL
    for j, v in enumerate(vals, start=2):
        c = ws.cell(row=r, column=j, value=v)
        c.alignment = RIGHT
        c.border = BORDER
        c.number_format = PCT if (j % 2 == 1) else NUM
        if is_total:
            c.font = BOLD
            c.fill = TOTAL_FILL
    r += 1
for j in range(2, 12):
    ws.column_dimensions[get_column_letter(j)].width = 11
ws.freeze_panes = "B" + str(data_start)

# =========================================================================
# 8. USE OF PROCEEDS
# =========================================================================
ws = wb.create_sheet("Use of Proceeds")
ws.sheet_view.showGridLines = False
r = title_row(ws, "Use of Proceeds — Planned Investment Projects", 9,
              "In trillions of Won. Additional planned investment by year (as of May 31, 2026).")
hdrs = ["Project", "Cleanroom open", "Total cost", "Previously invested",
        "Additional planned", "2026", "2027", "2028", "2029", "2030"]
r = header(ws, r, hdrs, first_width=34)
data_start = r
uop = [
    ("Fab 1 at the Yongin complex, Korea", "End of 2030", 31.0, 4.4, 26.6, 7.4, 10.1, 6.6, 2.5, 0.0),
    ("P&T7 advanced packaging plant, Cheongju", "End of 2030", 19.0, 0.1, 18.9, 0.5, 2.1, 2.7, 5.8, 7.8),
    ("Total", "", 50.0, 4.5, 45.5, 7.9, 12.2, 9.3, 8.3, 7.8),
]
for row_vals in uop:
    is_total = row_vals[0] == "Total"
    for j, v in enumerate(row_vals, start=1):
        c = ws.cell(row=r, column=j, value=v)
        c.border = BORDER
        if j == 1:
            c.alignment = LEFT
            c.font = BOLD if is_total else LABEL_FONT
        elif j == 2:
            c.alignment = CENTER
        else:
            c.alignment = RIGHT
            c.number_format = NUM1
        if is_total:
            c.fill = TOTAL_FILL
            c.font = BOLD
    r += 1
for j in range(3, 11):
    ws.column_dimensions[get_column_letter(j)].width = 13
ws.column_dimensions["B"].width = 14
r += 1
r = note(ws, r, 10, "Net proceeds intended for (i) W45.5tn of Korean facility capex (above) and "
         "(ii) ~W11.9tn of EUV scanner purchases (delivery by Dec 2027). Shortfalls funded by operating cash flow and debt.")
note(ws, r, 10, "Figures are estimates subject to change with market conditions, specifications and exchange rates.")

# =========================================================================
# 9. DIVIDENDS
# =========================================================================
ws = wb.create_sheet("Dividends")
ws.sheet_view.showGridLines = False
r = title_row(ws, "Dividends Paid (Quarterly & Annual)", 4,
              "Dividend per share in Won; total in billions of Won.")
r = header(ws, r, ["Dividend type / period", "Per share (W)", "Total (W bn)", "Shares entitled"], first_width=46)
data_start = r
div = [
    ("Quarterly — period ended Mar 31, 2023", 300, 206, 688059197),
    ("Quarterly — period ended Jun 30, 2023", 300, 206, 688090311),
    ("Quarterly — period ended Sep 30, 2023", 300, 206, 688116189),
    ("Annual — year ended Dec 31, 2023", 300, 206, 688138649),
    ("Quarterly — period ended Mar 31, 2024", 300, 207, 688614914),
    ("Quarterly — period ended Jun 30, 2024", 300, 207, 688617645),
    ("Quarterly — period ended Sep 30, 2024", 300, 207, 689038731),
    ("Annual — year ended Dec 31, 2024", 1304, 900, 690344530),
    ("Quarterly — period ended Mar 31, 2025", 375, 259, 690412123),
    ("Quarterly — period ended Jun 30, 2025", 375, 259, 690455268),
    ("Quarterly — period ended Sep 30, 2025", 375, 263, 701684263),
    ("Annual — year ended Dec 31, 2025", 1875, 1328, 708113147),
    ("Quarterly — period ended Mar 31, 2026", 375, 267, 711073295),
]
for label, ps, tot, sh in div:
    lc = ws.cell(row=r, column=1, value=label)
    lc.alignment = LEFT
    lc.border = BORDER
    lc.font = LABEL_FONT
    for j, v, fmt in [(2, ps, NUM), (3, tot, NUM), (4, sh, NUM)]:
        c = ws.cell(row=r, column=j, value=v)
        c.alignment = RIGHT
        c.border = BORDER
        c.number_format = fmt
    r += 1
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 16
r += 1
note(ws, r, 4, "FY2025-2027 policy: W1,500/share annually (4 x W375 quarterly), with potential additional returns on excess cash.")
ws.freeze_panes = "B" + str(data_start)

# =========================================================================
# 10. SHAREHOLDERS
# =========================================================================
ws = wb.create_sheet("Shareholders")
ws.sheet_view.showGridLines = False
r = title_row(ws, "Principal Shareholders (as of the F-1 date)", 3)
r = header(ws, r, ["Shareholder", "Common shares", "Ownership %"], first_width=42)
data_start = r
sh_rows = [
    ("SK square (SK Group)", 146100000, 0.2050),
    ("National Pension Service", 57439774, 0.0806),
    ("BlackRock Inc.", 36407157, 0.0511),
    ("Capital Research and Management Company", 25149374, 0.0353),
    ("Others", 445979195, 0.6258),
    ("Treasury shares", 1626865, 0.0023),
    ("Total issued and outstanding", 712702365, 1.0),
]
for label, n, pct in sh_rows:
    is_total = label.startswith("Total")
    lc = ws.cell(row=r, column=1, value=label)
    lc.alignment = LEFT
    lc.border = BORDER
    lc.font = BOLD if is_total else LABEL_FONT
    if is_total:
        lc.fill = TOTAL_FILL
    c2 = ws.cell(row=r, column=2, value=n)
    c2.number_format = NUM
    c2.alignment = RIGHT
    c2.border = BORDER
    c3 = ws.cell(row=r, column=3, value=pct)
    c3.number_format = PCT
    c3.alignment = RIGHT
    c3.border = BORDER
    if is_total:
        c2.font = BOLD; c3.font = BOLD
        c2.fill = TOTAL_FILL; c3.fill = TOTAL_FILL
    r += 1
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 14
ws.freeze_panes = "B" + str(data_start)

# =========================================================================
# 11. KPIs & RATIOS
# =========================================================================
ws = wb.create_sheet("KPIs & Ratios")
ws.sheet_view.showGridLines = False
r = title_row(ws, "Key Performance Indicators & Ratios", 6,
              "Derived from the financial statements. Margins on revenue; balance-sheet ratios at period-end.")
r = header(ws, r, ["Metric"] + PERIODS)
data_start = r
kpi = [
    ("Revenue growth (YoY)", [1.981, None, 0.468, 1.020, None], PCT),
    ("Gross margin", [0.793, 0.573, 0.604, 0.481, -0.016], PCT),
    ("Operating margin", [0.715, 0.422, 0.486, 0.355, -0.236], PCT),
    ("Net margin", [0.767, 0.460, 0.442, 0.299, -0.279], PCT),
    ("Adjusted EBITDA margin", [0.786, 0.611, 0.629, 0.544, 0.180], PCT),
    ("Capex / revenue", [0.146, 0.356, 0.283, 0.241, 0.254], PCT),
    ("DRAM % of revenue", [0.773, 0.796, 0.771, 0.676, 0.634], PCT),
    ("U.S. % of revenue", [0.647, 0.725, 0.688, 0.634, 0.470], PCT),
]
for label, vals, fmt in kpi:
    lc = ws.cell(row=r, column=1, value=label)
    lc.alignment = LEFT
    lc.border = BORDER
    lc.font = LABEL_FONT
    for j, v in enumerate(vals, start=2):
        c = ws.cell(row=r, column=j, value=v)
        c.alignment = RIGHT
        c.border = BORDER
        c.number_format = fmt
    r += 1
r += 1
# balance sheet ratios (period-end, 4 cols)
ws.cell(row=r, column=1, value="Balance-sheet ratios").font = BOLD
r += 1
bs_kpi_periods = ["Mar 31, 2026", "Dec 31, 2025", "Dec 31, 2024", "Dec 31, 2023"]
for j, p in enumerate(bs_kpi_periods, start=2):
    c = ws.cell(row=r, column=j, value=p)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = BORDER
ws.cell(row=r, column=1, value="Metric").font = HDR_FONT
ws.cell(row=r, column=1).fill = HDR_FILL
ws.cell(row=r, column=1).border = BORDER
r += 1
bs_kpi = [
    ("Current ratio (x)", [2.62, 1.86, 1.69, 1.45], NUM1),
    ("Total liabilities / equity (x)", [0.36, 0.46, 0.62, 0.88], NUM1),
    ("Total borrowings (W bn)", [19318, 22248, 22683, 29468], NUM),
    ("Cash & equivalents (W bn)", [21167, 14924, 11205, 7587], NUM),
    ("Equity / total assets", [0.738, 0.685, 0.617, 0.533], PCT),
]
for label, vals, fmt in bs_kpi:
    lc = ws.cell(row=r, column=1, value=label)
    lc.alignment = LEFT
    lc.border = BORDER
    lc.font = LABEL_FONT
    for j, v in enumerate(vals, start=2):
        c = ws.cell(row=r, column=j, value=v)
        c.alignment = RIGHT
        c.border = BORDER
        c.number_format = fmt
    r += 1
r += 1
note(ws, r, 6, "Total borrowings = current + non-current borrowings. YoY growth omitted where the prior comparable period (Q1 2024 / FY2022) is not in the filing.")
ws.freeze_panes = "B" + str(data_start)

wb.save(OUTFILE)
print("Wrote", OUTFILE, "with sheets:", wb.sheetnames)
