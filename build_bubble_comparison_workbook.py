#!/usr/bin/env python3
"""Builds an Excel workbook comparing the 1999 dot-com bubble and the 2008
global financial crisis across capex, semiconductor demand and component
pricing.

Output: Tech_Bubble_Comparison_1999_vs_2008.xlsx

All data points were aggregated from public sources (WSTS/SIA, Gartner/
Dataquest, SEMI, DRAMeXchange/TrendForce, FRBSF, Richmond Fed, TIA, BEA/FRED);
see the "Sources & Notes" sheet for the full citation list.
"""

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

NAVY = "1F3864"
LIGHT_GRAY = "D6DCE5"
DOTCOM_FILL = "FBE5D6"   # light orange
DOTCOM_ACCENT = "C55A11"
GFC_FILL = "DEEBF7"      # light blue
GFC_ACCENT = "2E75B6"

TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=NAVY)
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="595959")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color=NAVY)
BOLD = Font(name="Calibri", size=11, bold=True)
NORMAL = Font(name="Calibri", size=11)
EST_FONT = Font(name="Calibri", size=11, italic=True, color="595959")
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="595959")

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SECTION_FILL = PatternFill("solid", fgColor=LIGHT_GRAY)
DC_FILL = PatternFill("solid", fgColor=DOTCOM_FILL)
GF_FILL = PatternFill("solid", fgColor=GFC_FILL)

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

FMT_BIL = '#,##0.0'
FMT_PCT = '0.0%'
FMT_USD = '$#,##0.00'
FMT_UNITS = '#,##0'


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title_block(ws, title, subtitle, ncols):
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.cell(row=2, column=1, value=subtitle).font = SUBTITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)


def header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers, start=start_col):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BOX


def section(ws, row, text, ncols):
    c = ws.cell(row=row, column=1, value=text)
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    for i in range(1, ncols + 1):
        ws.cell(row=row, column=i).fill = SECTION_FILL
        ws.cell(row=row, column=i).border = BOX


def put(ws, row, col, value, fmt=None, font=None, fill=None, wrap=False,
        border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or NORMAL
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if wrap:
        c.alignment = WRAP
    if border:
        c.border = BOX
    return c


wb = Workbook()

# ===========================================================================
# Sheet 1: Overview
# ===========================================================================
ws = wb.active
ws.title = "Overview"
set_widths(ws, [3, 30, 105])
ws.sheet_view.showGridLines = False

put(ws, 1, 2, "1999 vs 2008: Two Bubbles Compared", font=TITLE_FONT,
    border=False)
put(ws, 2, 2, "Capex, semiconductor demand and component pricing through the "
    "dot-com bust (2000-02) and the Global Financial Crisis (2008-09)",
    font=SUBTITLE_FONT, border=False)

r = 4
put(ws, r, 2, "Purpose", font=SECTION_FONT, border=False)
put(ws, r + 1, 2, "This workbook aggregates public industry data to compare "
    "how the two downturns hit technology capital spending, semiconductor "
    "demand and component (memory) pricing. The 'dot-com' era refers to the "
    "TMT bubble that peaked in March 2000; the '2008' era refers to the "
    "credit/housing bubble whose bust culminated in the September 2008 "
    "Lehman collapse.", font=NORMAL, border=False, wrap=True)
ws.merge_cells(start_row=r + 1, start_column=2, end_row=r + 3, end_column=3)

r = 9
put(ws, r, 2, "Sheet guide", font=SECTION_FONT, border=False)
guide = [
    ("Bubble Comparison", "Side-by-side summary of the two episodes across "
     "every metric in the workbook."),
    ("Semiconductor Demand", "Annual worldwide semiconductor sales 1995-2011 "
     "(WSTS/SIA), DRAM industry revenue, PC and mobile-phone unit shipments."),
    ("Capex", "Semiconductor industry capital spending (Gartner), fab "
     "equipment billings (SEMI), US telecom carrier capex (the dot-com era's "
     "signature excess) and US business fixed investment (BEA)."),
    ("Component Pricing", "DRAM and NAND price timelines for both busts: "
     "128Mb SDRAM in 2000-01 and DDR2 1Gb / NAND in 2007-09."),
    ("Market Context", "Equity index peaks and troughs framing each bubble."),
    ("Sources & Notes", "Full citation list, footnote key and caveats."),
]
for i, (name, desc) in enumerate(guide):
    put(ws, r + 1 + i, 2, name, font=BOLD)
    put(ws, r + 1 + i, 3, desc, wrap=True)

r = 17
put(ws, r, 2, "Headline findings", font=SECTION_FONT, border=False)
findings = [
    "Depth of demand shock: the dot-com bust was ~3x deeper for the chip "
    "industry. Worldwide semiconductor sales fell 32.0% in 2001 (the worst "
    "year in industry history), versus a cumulative -11.5% across 2008-09 "
    "(-2.8% then -9.0%).",
    "Capex was the bubble in 1999-2000: US telecom carriers alone went from "
    "$47B (1995) to $121B (2000) of annual capex before collapsing to $49B "
    "(2002), and semiconductor capex jumped +84% in 2000 alone. In 2008 the "
    "excess sat outside tech (housing/credit); tech capex was cut in a "
    "defensive cash-preservation move, not an unwind of its own excess.",
    "Speed and shape: 2008-09 was a sharper but shorter shock. SEMI fab "
    "equipment billings fell 63% over two years (2007-09) then rebounded "
    "+148% in 2010; after 2001 (-41%), spending stayed depressed for ~3 "
    "years. Semiconductor sales took 4 years to regain their 2000 peak but "
    "only 3 years to regain the 2007 peak.",
    "Component pricing collapsed below cost in both busts: 128Mb SDRAM fell "
    "~92% in 12 months (about $18.40 to about $1.50, mid-2000 to Aug-2001); "
    "DDR2 1Gb fell ~75% inside 2008 ($2.29 in May to a record $0.58 low), "
    "dropping below cash cost in September 2008.",
    "Demand character: 2001 saw the first-ever annual declines in both PC "
    "(-4.6%) and mobile-phone (-3.2%) shipments; in 2008 full-year PC units "
    "still grew +10.9% before the demand stop in Q4-08 (US PC shipments "
    "-10% YoY), showing the shock arrived late and from outside tech.",
]
for i, f in enumerate(findings):
    put(ws, r + 1 + i * 3, 2, chr(8226) + " " + f, wrap=True, border=False)
    ws.merge_cells(start_row=r + 1 + i * 3, start_column=2,
                   end_row=r + 3 + i * 3, end_column=3)

put(ws, r + 16, 2, "All dollar values are nominal US$. Figures marked (e) are "
    "estimated/derived from stated growth rates; (f) were forecasts at the "
    "time; '~' denotes approximate values. See Sources & Notes.",
    font=NOTE_FONT, border=False, wrap=True)
ws.merge_cells(start_row=r + 16, start_column=2, end_row=r + 17, end_column=3)

# ===========================================================================
# Sheet 2: Bubble Comparison
# ===========================================================================
ws = wb.create_sheet("Bubble Comparison")
set_widths(ws, [42, 44, 44, 52])
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A5"

title_block(ws, "Bubble Comparison: 1999 Dot-com vs 2008 GFC",
            "Side-by-side summary. Sources and footnote key on the "
            "'Sources & Notes' sheet.", 4)

header_row(ws, 4, ["Metric", "Dot-com / TMT bubble (1999-2002)",
                   "Credit bubble / GFC (2007-2009)", "Notes"])
ws.cell(row=4, column=2).fill = PatternFill("solid", fgColor=DOTCOM_ACCENT)
ws.cell(row=4, column=3).fill = PatternFill("solid", fgColor=GFC_ACCENT)

rows = [
    ("SECTION", "Character of the bubble", None, None),
    ("Nature of the excess",
     "Overinvestment inside tech itself: internet, telecom networks and "
     "semiconductor capacity, funded by equity mania and cheap debt",
     "Excess sat in housing and credit markets; tech entered 2008 with "
     "relatively healthy balance sheets and was hit by a demand shock",
     "Key structural difference driving the recovery shapes"),
    ("Equity market peak",
     "NASDAQ 5,048.62 on 10-Mar-2000",
     "S&P 500 1,565.15 on 9-Oct-2007",
     "See Market Context sheet"),
    ("Equity market trough",
     "NASDAQ 1,114.11 on 9-Oct-2002 (-77.9%)",
     "S&P 500 676.53 on 9-Mar-2009 (-56.8%)",
     "NASDAQ took ~15 years to regain its 2000 peak; S&P 500 regained its "
     "2007 peak in ~5.5 years"),

    ("SECTION", "Semiconductor demand", None, None),
    ("WW semiconductor sales - pre-bust peak",
     "$204.4B (2000, +36.8% YoY)", "$255.6B (2007, +3.2% YoY)",
     "WSTS/SIA. Note the 2000 peak was itself bubble-inflated (+37% in one "
     "year); 2007 growth was modest"),
    ("WW semiconductor sales - trough",
     "$139.0B (2001)", "$226.3B (2009)",
     "WSTS/SIA"),
    ("Peak-to-trough decline", -0.320, -0.115,
     "2001 = worst single-year decline in industry history (-32.0%); "
     "2008-09 = -2.8% then -9.0%"),
    ("Years to regain prior peak sales",
     "4 (2000 peak first exceeded in 2004: $213.0B)",
     "3 (2007 peak exceeded in 2010: record $298.3B, +31.8%)",
     "V-shaped recovery after 2009 vs long U after 2001"),
    ("DRAM industry revenue",
     "$31.5B (2000) to $14.0B (2001): -55.5%, worst year ever for DRAM",
     "~$31.5B (2007) to ~$23.6B (2008): ~-25%",
     "Dataquest/Gartner"),
    ("PC unit shipments",
     "2001: 128M units, -4.6% - first annual decline since 1985 "
     "(US: -11.1%)",
     "2008: 302.2M units, +10.9% for the full year, but Q4-08 just +1.1% "
     "and US Q4-08 -10%",
     "Gartner/Dataquest. Demand collapse in 2008 was concentrated in Q4"),
    ("Mobile phone unit sales",
     "2001: 399.6M, -3.2% - first annual decline ever (2000: 412.7M)",
     "2009: 1,211M, -0.9% (2008: 1,222M, +6.0%; Q4-08: -4.6% YoY)",
     "Gartner"),

    ("SECTION", "Capital expenditure", None, None),
    ("Semiconductor industry capex - boom",
     "+84% in 2000 alone, to ~$63.9B (e)",
     "Flat pre-bust: $56.3B in 2007 (+0.6% YoY); memory over-spending in "
     "2006-07 had already created oversupply",
     "Gartner/Dataquest"),
    ("Semiconductor industry capex - bust",
     "-26% in 2001 (to $47.3B), -41% in 2002 (to $27.8B): -56% cumulative",
     "Equipment capex -31.7% in 2008 (to $30.7B) and -45.8% in 2009 (to "
     "$16.6B): -63% cumulative; memory capex -54% in 2009 alone",
     "Gartner. 2008-09 cuts were faster and deeper but from a lower base"),
    ("Fab equipment billings (SEMI)",
     "$47.7B (2000) to $28.1B (2001): -41%",
     "$42.8B (2007) to $29.5B (2008) to $15.9B (2009): -63% over two years, "
     "lowest since 1994",
     "SEMI WWSEMS - same source both eras"),
    ("Equipment spending recovery",
     "Slow: spending remained depressed through 2002-03 ($27.8B / $28.9B "
     "total capex); no strong rebound until 2004 (+28%)",
     "V-shaped: SEMI billings +148% in 2010 to $39.5B",
     "IC Insights: 2010 was the only post-downturn rebound where year 1 was "
     "stronger than year 2"),
    ("Telecom / adjacent-sector capex",
     "US telecom carrier capex $47B (1995) to $121B (2000) to $49B (2002): "
     "-60% bust after a +28%/yr boom; ~$444B spent 1996-2001",
     "No equivalent tech-adjacent capex bubble; the overbuild was in "
     "residential construction and structured credit",
     "FRBSF, TIA/CSFB, LA Times. Telecom overbuild was the defining excess "
     "of the 1999 bubble"),
    ("US real business fixed investment (BEA)",
     "-2.2% (2001), -6.9% (2002) - shallow but 2 years long",
     "+0.6% (2008), -14.5% (2009) - deepest single-year cut of the postwar "
     "era",
     "Real private nonresidential fixed investment, annual % change"),

    ("SECTION", "Component pricing", None, None),
    ("Benchmark DRAM price collapse",
     "128Mb SDRAM: ~$18.40 (mid-2000) to ~$1.50 spot (Aug-2001): ~-92% in "
     "12 months; 64Mb fell below $1.00 (Jul-2001)",
     "DDR2 1Gb: $2.29 (6-May-2008) to $1.13 (17-Oct) to record $0.58 low "
     "(Nov/Dec-2008): ~-75% inside 2008",
     "Dataquest; DRAMeXchange. See Component Pricing sheet for timelines"),
    ("Prices vs production cost",
     "Spot below production cost for most makers by mid-2001 (break-even "
     "~$3.00 on 128Mb); top-5 Japanese DRAM makers lost a combined ~$455M "
     "in 2001",
     "Below ~$1.50 cash cost from 6-Sep-2008; at the $0.58 low, price was "
     "at bare materials cost (~$0.60)",
     "Both busts forced selling below cost; consolidation followed (exit of "
     "Japanese DRAM makers post-2001; Qimonda bankruptcy Jan-2009)"),
    ("Supply response",
     "Muted and slow - most makers refused output cuts (only Toshiba cut, "
     "30%); glut persisted into 2002",
     "Fast: Taiwanese makers cut ~30% of output in Q4-08, Hynix shut 8-inch "
     "fabs, NAND makers cut wafer starts ~10% QoQ in 1Q09",
     "DRAMeXchange"),
    ("NAND flash",
     "n/a (market still nascent)",
     "Contract prices -30%+ QoQ in 4Q08; 8Gb fell from $1.50 to $0.95 "
     "during Dec-2008; capex budgets cut and 200mm tools retired",
     "DRAMeXchange"),

    ("SECTION", "Bottom line", None, None),
    ("Summary judgment",
     "A tech-centred bubble: capex and demand inside the sector were the "
     "excess, so the bust was deeper for semis (-32% sales) and the "
     "recovery long and U-shaped",
     "An external demand shock: violent but brief - deeper capex/pricing "
     "cuts at the trough, but a V-shaped snap-back with record sales by "
     "2010",
     "Severity for tech: 1999 bubble much worse on demand; 2008 sharper on "
     "capex velocity and pricing, far faster to heal"),
]

r = 5
for row_def in rows:
    if row_def[0] == "SECTION":
        section(ws, r, row_def[1], 4)
        r += 1
        continue
    metric, dc, gf, note = row_def
    put(ws, r, 1, metric, font=BOLD, wrap=True)
    c_dc = put(ws, r, 2, dc, wrap=True, fill=DC_FILL)
    c_gf = put(ws, r, 3, gf, wrap=True, fill=GF_FILL)
    if isinstance(dc, float):
        c_dc.number_format = FMT_PCT
        c_dc.font = BOLD
    if isinstance(gf, float):
        c_gf.number_format = FMT_PCT
        c_gf.font = BOLD
    put(ws, r, 4, note, font=NOTE_FONT, wrap=True)
    ws.row_dimensions[r].height = 46
    r += 1

# ===========================================================================
# Sheet 3: Semiconductor Demand
# ===========================================================================
ws = wb.create_sheet("Semiconductor Demand")
set_widths(ws, [8, 15, 9, 14, 13, 9, 14, 9, 62])
ws.freeze_panes = "A5"
title_block(ws, "Semiconductor Demand, 1995-2011",
            "Worldwide semiconductor sales (WSTS/SIA), DRAM revenue "
            "(Dataquest/Gartner), PC shipments and mobile phone sales "
            "(Gartner). Nominal US$.", 9)

header_row(ws, 4, ["Year", "WW semi sales ($B)", "YoY %", "DRAM revenue ($B)",
                   "PC shipments (M)", "PC YoY %", "Mobile phones (M)",
                   "Phone YoY %", "Notes"])

# year: (semi_sales, dram_rev, pc_units, phone_units, note)
demand = {
    1995: (144.4, 41.0, None, None,
           "DRAM peak of the mid-90s cycle (~$41B, Dataquest)"),
    1996: (132.0, None, None, None, ""),
    1997: (137.2, None, None, None, ""),
    1998: (125.6, 14.0, None, None,
           "Asian financial crisis; DRAM revenue ~$14B"),
    1999: (149.4, None, None, None, "Dot-com boom accelerates"),
    2000: (204.4, 31.5, 134.0, 412.7,
           "Bubble peak: semi sales +36.8%, the strongest year of the era"),
    2001: (139.0, 14.0, 128.0, 399.6,
           "Worst year in industry history: sales -32%; first-ever declines "
           "in PC (-4.6%, first since 1985) and phone (-3.2%) units; DRAM "
           "revenue -55.5%"),
    2002: (140.7, None, None, None, "Flat; recovery not yet visible"),
    2003: (166.4, None, None, None, ""),
    2004: (213.0, None, None, None,
           "First year above the 2000 peak - 4 years to recover"),
    2005: (227.5, None, None, None, ""),
    2006: (247.7, 34.0, None, None,
           "Memory capex race inflates DRAM revenue to ~$34B"),
    2007: (255.6, 31.5, 272.5, 1153.0,
           "Pre-GFC peak; growth already modest (+3.2%); DRAM oversupply "
           "begins (512Mb below $1 in Nov-07). PC units implied from "
           "Gartner's 2008 +10.9%"),
    2008: (248.6, 23.6, 302.2, 1222.0,
           "Lehman collapse in Sept; full-year PCs +10.9% but Q4-08 +1.1% "
           "(US -10%); Q4-08 phones -4.6% YoY; DRAM revenue ~-25%"),
    2009: (226.3, 22.9, None, 1211.0,
           "Trough: semi sales -9.0%; phones -0.9% (first decline since "
           "2001)"),
    2010: (298.3, None, None, None,
           "Record sales, +31.8% - regains 2007 peak in 3 years"),
    2011: (299.5, None, None, None, ""),
}

r = 5
first_data_row = r
year_rows = {}
for year in range(1995, 2012):
    sales, dram, pcs, phones, note = demand[year]
    year_rows[year] = r
    fill = DC_FILL if 1999 <= year <= 2002 else (
        GF_FILL if 2007 <= year <= 2009 else None)
    put(ws, r, 1, year, fmt='0', font=BOLD, fill=fill)
    put(ws, r, 2, sales, fmt=FMT_BIL)
    if year > 1995:
        put(ws, r, 3, f"=B{r}/B{r-1}-1", fmt=FMT_PCT)
    else:
        put(ws, r, 3, None)
    put(ws, r, 4, dram, fmt=FMT_BIL, font=EST_FONT if dram else NORMAL)
    put(ws, r, 5, pcs, fmt=FMT_BIL)
    prev_pc = demand[year - 1][2] if year > 1995 else None
    put(ws, r, 6, f"=E{r}/E{r-1}-1" if (pcs and prev_pc) else None,
        fmt=FMT_PCT)
    put(ws, r, 7, phones, fmt=FMT_BIL)
    prev_ph = demand[year - 1][3] if year > 1995 else None
    put(ws, r, 8, f"=G{r}/G{r-1}-1" if (phones and prev_ph) else None,
        fmt=FMT_PCT)
    put(ws, r, 9, note, font=NOTE_FONT, wrap=True)
    r += 1
last_data_row = r - 1

r += 1
section(ws, r, "Drawdown comparison (worldwide semiconductor sales)", 9)
r += 1
header_row(ws, r, ["", "Dot-com bust", "GFC"], start_col=1)
dd_rows = [
    ("Peak year / sales ($B)", f"=B{year_rows[2000]}", f"=B{year_rows[2007]}",
     FMT_BIL),
    ("Trough year / sales ($B)", f"=B{year_rows[2001]}",
     f"=B{year_rows[2009]}", FMT_BIL),
    ("Peak-to-trough decline",
     f"=B{year_rows[2001]}/B{year_rows[2000]}-1",
     f"=B{year_rows[2009]}/B{year_rows[2007]}-1", FMT_PCT),
    ("Years to regain prior peak", 4, 3, FMT_UNITS),
]
labels2 = ["Peak (2000 / 2007)", "Trough (2001 / 2009)",
           "Peak-to-trough decline", "Years to regain prior peak"]
for i, (label, dc, gf, fmt) in enumerate(dd_rows):
    rr = r + 1 + i
    put(ws, rr, 1, labels2[i], font=BOLD)
    put(ws, rr, 2, dc, fmt=fmt, fill=DC_FILL)
    put(ws, rr, 3, gf, fmt=fmt, fill=GF_FILL)

chart = LineChart()
chart.title = "Worldwide semiconductor sales ($B), 1995-2011"
chart.style = 12
chart.y_axis.title = "$B"
chart.x_axis.title = "Year"
chart.height = 9
chart.width = 22
data = Reference(ws, min_col=2, min_row=4, max_row=last_data_row)
cats = Reference(ws, min_col=1, min_row=first_data_row, max_row=last_data_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.legend = None
ws.add_chart(chart, f"B{r + 7}")

# ===========================================================================
# Sheet 4: Capex
# ===========================================================================
ws = wb.create_sheet("Capex")
set_widths(ws, [8, 17, 9, 17, 16, 10, 16, 15, 58])
ws.freeze_panes = "A5"
title_block(ws, "Capital Expenditure, 1995-2011",
            "Semiconductor capex (Gartner/Dataquest), fab equipment "
            "(Gartner and SEMI WWSEMS), US telecom carrier capex (FRBSF) "
            "and US business fixed investment (BEA). Nominal US$B unless "
            "noted. (e)=derived from stated growth rates, (f)=forecast at "
            "the time.", 9)

header_row(ws, 4, ["Year", "Semi capex, total ($B, Gartner)", "YoY %",
                   "Semi capital equipment ($B, Gartner)",
                   "Fab equipment billings ($B, SEMI)", "SEMI YoY %",
                   "US telecom carrier capex ($B)",
                   "US real nonres. fixed investment YoY",
                   "Notes"])

# year: (gartner_capex, capex_is_est, gartner_equip, equip_is_est, semi,
#        telecom, bea_pct, note)
capex = {
    1995: (None, False, None, False, None, 47.0, 0.097,
           "Telecom capex baseline before the Telecom Act of 1996"),
    1996: (None, False, None, False, None, None, 0.091,
           "Telecom Act 1996 kicks off a +28%/yr carrier capex boom "
           "(TIA/CSFB)"),
    1997: (None, False, None, False, None, None, 0.108, ""),
    1998: (None, False, None, False, None, None, 0.109, ""),
    1999: (34.7, True, None, False, None, None, 0.099,
           "(e) implied by Gartner's '+84% in 2000'"),
    2000: (63.9, True, 39.9, False, 47.7, 121.0, 0.093,
           "Capex mania peak: semi capex +84% YoY; telecom capex peaks at "
           "$121B (~$444B cumulative 1996-2001, LA Times)"),
    2001: (47.3, False, 27.9, False, 28.1, None, -0.022,
           "Semi capex -26%; SEMI billings -41%; telecom collapse begins "
           "(CLEC bankruptcies)"),
    2002: (27.8, False, None, False, None, 49.0, -0.069,
           "Semi capex -41% (year 2 of cuts); telecom capex -60% from peak; "
           "WorldCom bankruptcy"),
    2003: (28.9, False, 21.7, False, None, None, 0.025,
           "Spending still at the bottom, 3 years after the peak"),
    2004: (37.0, True, 29.5, True, None, None, 0.056,
           "(f) Gartner Dec-2003 forecast: +28% rebound"),
    2005: (None, False, None, False, None, None, 0.077, ""),
    2006: (56.0, True, None, False, 40.5, None, 0.080,
           "(e) implied by Gartner's 2007 +0.6%; memory capex race"),
    2007: (56.3, False, 44.9, True, 42.8, None, 0.069,
           "Pre-GFC peak, but growth only +0.6%: no capex mania this time. "
           "(e) equipment implied by Gartner's 2008 -31.7%"),
    2008: (None, False, 30.7, False, 29.5, None, 0.006,
           "Equipment capex -31.7% (Gartner final); SEMI billings -31%"),
    2009: (None, False, 16.6, False, 15.9, None, -0.145,
           "Equipment capex -45.8%; memory capex -54%; SEMI billings lowest "
           "since 1994; US business investment worst postwar year"),
    2010: (None, False, None, False, 39.5, None, 0.045,
           "V-shaped rebound: SEMI billings +148%"),
    2011: (None, False, None, False, None, None, 0.087, ""),
}

r = 5
first_data_row = r
cap_rows = {}
for year in range(1995, 2012):
    gx, gx_e, ge, ge_e, semi, tel, bea, note = capex[year]
    cap_rows[year] = r
    fill = DC_FILL if 1999 <= year <= 2002 else (
        GF_FILL if 2007 <= year <= 2009 else None)
    put(ws, r, 1, year, fmt='0', font=BOLD, fill=fill)
    put(ws, r, 2, gx, fmt=FMT_BIL, font=EST_FONT if gx_e else NORMAL)
    prev = capex[year - 1][0] if year > 1995 else None
    put(ws, r, 3, f"=B{r}/B{r-1}-1" if (gx and prev) else None, fmt=FMT_PCT)
    put(ws, r, 4, ge, fmt=FMT_BIL, font=EST_FONT if ge_e else NORMAL)
    put(ws, r, 5, semi, fmt=FMT_BIL)
    prev_s = capex[year - 1][4] if year > 1995 else None
    put(ws, r, 6, f"=E{r}/E{r-1}-1" if (semi and prev_s) else None,
        fmt=FMT_PCT)
    put(ws, r, 7, tel, fmt=FMT_BIL)
    put(ws, r, 8, bea, fmt=FMT_PCT)
    put(ws, r, 9, note, font=NOTE_FONT, wrap=True)
    r += 1
last_data_row = r - 1

r += 1
section(ws, r, "Peak-to-trough capex declines", 9)
r += 1
header_row(ws, r, ["Series", "Dot-com bust", "GFC"], start_col=1)
pt_rows = [
    ("Semi capex, total (Gartner)",
     f"=B{cap_rows[2002]}/B{cap_rows[2000]}-1",
     "n/a (equipment series used for GFC)"),
    ("Semi capital equipment (Gartner)",
     f"=D{cap_rows[2001]}/D{cap_rows[2000]}-1",
     f"=D{cap_rows[2009]}/D{cap_rows[2007]}-1"),
    ("Fab equipment billings (SEMI)",
     f"=E{cap_rows[2001]}/E{cap_rows[2000]}-1",
     f"=E{cap_rows[2009]}/E{cap_rows[2007]}-1"),
    ("US telecom carrier capex (FRBSF)",
     f"=G{cap_rows[2002]}/G{cap_rows[2000]}-1",
     "n/a (no equivalent bubble)"),
    ("US real business fixed investment (BEA)",
     "-8.9% cumulative (2001-02)", "-14.0% cumulative (2008-09)"),
]
for i, (label, dc, gf) in enumerate(pt_rows):
    rr = r + 1 + i
    put(ws, rr, 1, label, font=BOLD)
    cdc = put(ws, rr, 2, dc, fill=DC_FILL)
    cgf = put(ws, rr, 3, gf, fill=GF_FILL)
    if isinstance(dc, str) and dc.startswith("="):
        cdc.number_format = FMT_PCT
    if isinstance(gf, str) and gf.startswith("="):
        cgf.number_format = FMT_PCT

chart = BarChart()
chart.type = "col"
chart.title = "Fab equipment billings ($B, SEMI WWSEMS)"
chart.style = 10
chart.y_axis.title = "$B"
chart.height = 9
chart.width = 22
data = Reference(ws, min_col=5, min_row=4, max_row=last_data_row)
cats = Reference(ws, min_col=1, min_row=first_data_row, max_row=last_data_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.legend = None
ws.add_chart(chart, f"B{r + 8}")

# ===========================================================================
# Sheet 5: Component Pricing
# ===========================================================================
ws = wb.create_sheet("Component Pricing")
set_widths(ws, [14, 30, 13, 15, 70])
ws.freeze_panes = "A4"
title_block(ws, "Component Pricing Through Both Busts",
            "Memory is the industry's price barometer: commodity DRAM/NAND "
            "prices collapsed below production cost in both downturns. Spot "
            "prices unless noted. Sources: Dataquest/Gartner (2000-01), "
            "DRAMeXchange/TrendForce (2007-09).", 5)

r = 4
section(ws, r, "Dot-com bust: SDRAM price collapse, 2000-2001", 5)
r += 1
header_row(ws, r, ["Date", "Item", "Price (US$)", "Index (peak=100)",
                   "Comment"])
dc_prices = [
    ("Mid-2000", "128Mb SDRAM (avg selling price)", 18.40,
     "Cycle peak. PC demand strong, supply tight"),
    ("Dec-2000", "64Mb SDRAM (spot)", 3.80,
     "Slide already underway as PC demand slows"),
    ("Jun-2001", "128Mb SDRAM (spot)", 2.00,
     "Below the ~$3.00 production cost of most makers; contract price also "
     "under $3 (Dataquest)"),
    ("Jul-2001", "64Mb SDRAM (spot)", 0.92,
     "Below $1 - approaching inflation-adjusted 1985 crash levels "
     "(Electronics Weekly)"),
    ("Aug-2001", "128Mb SDRAM (spot)", 1.50,
     "~-92% in 12 months (vs $18.40 average a year earlier)"),
]
r += 1
dc_price_start = r
for date, item, price, comment in dc_prices:
    put(ws, r, 1, date, fill=DC_FILL)
    put(ws, r, 2, item, wrap=True)
    put(ws, r, 3, price, fmt=FMT_USD, font=BOLD)
    if "128Mb" in item:
        put(ws, r, 4, f"=C{r}/$C${dc_price_start}*100", fmt='0.0')
    else:
        put(ws, r, 4, None)
    put(ws, r, 5, comment, font=NOTE_FONT, wrap=True)
    r += 1
dc_extra = [
    "DRAM prices fell ~80% in the 12 months to June 2001 (Dataquest); 2001 "
    "DRAM revenue fell 55.5% to $14B - the worst year in DRAM history.",
    "A 128MB memory module that cost ~$120 in 2000 sold for under $20 by "
    "mid-2001.",
    "Losses: top-5 Japanese DRAM makers ~-$455M (2001, Merrill est.); "
    "Infineon -$500M in a single quarter; Micron -$300M. Only Toshiba cut "
    "output (30%) - the glut persisted into 2002.",
]
for note in dc_extra:
    put(ws, r, 1, chr(8226) + " " + note, font=NOTE_FONT, wrap=True,
        border=False)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1

r += 1
section(ws, r, "GFC: DDR2 / NAND price collapse, 2007-2009", 5)
r += 1
header_row(ws, r, ["Date", "Item", "Price (US$)", "Index (peak=100)",
                   "Comment"])
gf_prices = [
    ("Nov-2007", "DDR2 512Mb 667MHz (spot)", 0.98,
     "Falls below $1.00 - near cash cost; DRAM oversupply predates the "
     "financial crisis"),
    ("6-May-2008", "DDR2 1Gb 667MHz (spot)", 2.29,
     "2008 high, after Q2 restocking lifted contract prices ~25%"),
    ("6-Sep-2008", "DDR2 1Gb 667MHz (spot)", 1.50,
     "Falls below the average cash cost of most DRAM makers"),
    ("30-Sep-2008", "DDR2 1Gb 667MHz (spot)", 1.28,
     "-7% in a single day; makers dumping inventory at $1.00"),
    ("17-Oct-2008", "DDR2 1Gb 667MHz (spot)", 1.13,
     "-50% from the May high; production now cash-flow negative"),
    ("Nov/Dec-2008", "DDR2 1Gb eTT (spot, record low)", 0.58,
     "-75% from May; at bare materials cost (~$0.60). Qimonda filed for "
     "bankruptcy Jan-2009"),
]
r += 1
gf_price_start = r
for i, (date, item, price, comment) in enumerate(gf_prices):
    put(ws, r, 1, date, fill=GF_FILL)
    put(ws, r, 2, item, wrap=True)
    put(ws, r, 3, price, fmt=FMT_USD, font=BOLD)
    if "1Gb" in item:
        put(ws, r, 4, f"=C{r}/$C${gf_price_start + 1}*100", fmt='0.0')
    else:
        put(ws, r, 4, None)
    put(ws, r, 5, comment, font=NOTE_FONT, wrap=True)
    r += 1
gf_extra = [
    "Contract DRAM fell -34% QoQ in 3Q08 and -37% QoQ in 4Q08; by Dec-2008 "
    "a 1GB DDR2 module averaged ~$8.50 and 2GB ~$18 on contract.",
    "NAND flash: contract prices fell >30% QoQ in 4Q08; 8Gb slid from $1.50 "
    "to $0.95 during December (-24% in the month); 16Gb fell ~10% in early "
    "Dec alone.",
    "Supply response was much faster than 2001: Taiwanese DRAM makers cut "
    "~30% of output in Q4-08, Hynix closed 8-inch fabs and cut 12-inch "
    "output 20%, NAND makers cut wafer starts ~10% QoQ in 1Q09 and slashed "
    "2009 capex budgets.",
]
for note in gf_extra:
    put(ws, r, 1, chr(8226) + " " + note, font=NOTE_FONT, wrap=True,
        border=False)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1

r += 1
section(ws, r, "Pricing collapse compared", 5)
r += 1
header_row(ws, r, ["", "Dot-com (128Mb SDRAM)", "GFC (DDR2 1Gb)"],
           start_col=1)
cmp_rows = [
    ("Peak price", 18.40, 2.29, FMT_USD),
    ("Trough price", 1.50, 0.58, FMT_USD),
    ("Decline", "=B%d/B%d-1", "=C%d/C%d-1", FMT_PCT),
    ("Time from peak to trough", "~12 months", "~7 months", None),
    ("Below cash/production cost?", "Yes (break-even ~$3.00)",
     "Yes (cash cost ~$1.50; floor = materials ~$0.60)", None),
]
peak_r, trough_r = r + 1, r + 2
for i, (label, dc, gf, fmt) in enumerate(cmp_rows):
    rr = r + 1 + i
    put(ws, rr, 1, label, font=BOLD)
    if label == "Decline":
        dc = dc % (trough_r, peak_r)
        gf = gf % (trough_r, peak_r)
    put(ws, rr, 2, dc, fmt=fmt, fill=DC_FILL)
    put(ws, rr, 3, gf, fmt=fmt, fill=GF_FILL)

chart = LineChart()
chart.title = "DDR2 1Gb spot price through 2008 (US$)"
chart.style = 12
chart.y_axis.title = "US$"
chart.height = 8
chart.width = 16
data = Reference(ws, min_col=3, min_row=gf_price_start + 1,
                 max_row=gf_price_start + len(gf_prices) - 1)
cats = Reference(ws, min_col=1, min_row=gf_price_start + 1,
                 max_row=gf_price_start + len(gf_prices) - 1)
chart.add_data(data, titles_from_data=False)
chart.set_categories(cats)
chart.legend = None
ws.add_chart(chart, f"A{r + 8}")

chart2 = LineChart()
chart2.title = "128Mb SDRAM price, mid-2000 to Aug-2001 (US$)"
chart2.style = 12
chart2.y_axis.title = "US$"
chart2.height = 8
chart2.width = 16
# 128Mb rows only: peak, Jun-2001, Aug-2001
mini_row = r + 25
put(ws, mini_row, 1, "128Mb SDRAM series used for chart:", font=NOTE_FONT,
    border=False)
pts = [("Mid-2000", 18.40), ("Jun-2001", 2.00), ("Aug-2001", 1.50)]
for i, (d, p) in enumerate(pts):
    put(ws, mini_row + 1 + i, 1, d, font=NOTE_FONT)
    put(ws, mini_row + 1 + i, 2, p, fmt=FMT_USD, font=NOTE_FONT)
data = Reference(ws, min_col=2, min_row=mini_row + 1, max_row=mini_row + 3)
cats = Reference(ws, min_col=1, min_row=mini_row + 1, max_row=mini_row + 3)
chart2.add_data(data, titles_from_data=False)
chart2.set_categories(cats)
chart2.legend = None
ws.add_chart(chart2, f"D{r + 8}")

# ===========================================================================
# Sheet 6: Market Context
# ===========================================================================
ws = wb.create_sheet("Market Context")
set_widths(ws, [34, 14, 13, 14, 13, 11, 46])
ws.sheet_view.showGridLines = False
title_block(ws, "Equity Market Context",
            "Index peaks and troughs framing each bubble. SOX values are "
            "approximate closing levels.", 7)

header_row(ws, 4, ["Index", "Peak date", "Peak", "Trough date", "Trough",
                   "Decline", "Note"])
idx = [
    ("SECTION", "Dot-com bust", None, None, None, None),
    ("NASDAQ Composite", "10-Mar-2000", 5048.62, "9-Oct-2002", 1114.11,
     "Did not regain its 2000 peak until 2015 (~15 years)"),
    ("S&P 500", "24-Mar-2000", 1527.46, "9-Oct-2002", 776.76,
     "Broad market fell about half as much as tech"),
    ("PHLX Semiconductor (SOX) ~", "Mar-2000", 1362.0, "Oct-2002", 214.0,
     "Chip stocks fell harder than the NASDAQ itself"),
    ("NASDAQ Telecom index ~", "Mar-2000", 1230.0, "May-2003", 136.0,
     "The single most bombed-out sector of the bubble (Richmond Fed)"),
    ("SECTION", "GFC", None, None, None, None),
    ("S&P 500", "9-Oct-2007", 1565.15, "9-Mar-2009", 676.53,
     "Regained its 2007 peak in ~5.5 years (2013)"),
    ("NASDAQ Composite", "31-Oct-2007", 2859.12, "9-Mar-2009", 1268.64,
     "Tech fell with the market, not more than it"),
    ("PHLX Semiconductor (SOX) ~", "Jul-2007", 549.0, "Nov-2008", 167.0,
     "Bottomed ~4 months before the broad market"),
]
r = 5
for row_def in idx:
    if row_def[0] == "SECTION":
        section(ws, r, row_def[1], 7)
        fill = DC_FILL if "Dot-com" in row_def[1] else GF_FILL
        r += 1
        continue
    name, pd_, pv, td, tv, note = row_def
    put(ws, r, 1, name, font=BOLD, fill=fill)
    put(ws, r, 2, pd_)
    put(ws, r, 3, pv, fmt='#,##0.00')
    put(ws, r, 4, td)
    put(ws, r, 5, tv, fmt='#,##0.00')
    put(ws, r, 6, f"=E{r}/C{r}-1", fmt=FMT_PCT, font=BOLD)
    put(ws, r, 7, note, font=NOTE_FONT, wrap=True)
    r += 1

# ===========================================================================
# Sheet 7: Sources & Notes
# ===========================================================================
ws = wb.create_sheet("Sources & Notes")
set_widths(ws, [5, 46, 95])
ws.sheet_view.showGridLines = False
title_block(ws, "Sources & Notes", "All figures nominal US$ unless noted. "
            "Compiled Aug-2026 from the public sources below.", 3)

r = 4
section(ws, r, "Footnote key", 3)
keys = [
    ("(e)", "Estimated / derived from growth rates stated by the source "
     "(e.g. 2000 semiconductor capex of ~$63.9B is implied by Gartner's "
     "statement that 2001's $47.3B was a 26% decline)."),
    ("(f)", "A forecast published at the time rather than a final actual."),
    ("~", "Approximate value."),
]
for i, (k, v) in enumerate(keys):
    put(ws, r + 1 + i, 1, k, font=BOLD)
    put(ws, r + 1 + i, 2, v, wrap=True)
    ws.merge_cells(start_row=r + 1 + i, start_column=2,
                   end_row=r + 1 + i, end_column=3)

r += 5
section(ws, r, "Caveats", 3)
caveats = [
    "Different sources measure different baskets: Gartner 'capital "
    "spending' covers total semiconductor industry capex (fabs, land, "
    "buildings, equipment); Gartner 'capital equipment' and SEMI 'equipment "
    "billings' cover manufacturing tools only, with slightly different "
    "category and reporter coverage - so levels differ across columns even "
    "for the same year.",
    "Memory spot prices are volatile daily prints from Asian spot markets "
    "(DRAMeXchange); contract prices paid by large PC OEMs moved later and "
    "less violently. Both are quoted where available.",
    "DRAM revenue figures for 2006-2009 are approximate Gartner/Dataquest "
    "values as reported in the trade press.",
    "US telecom carrier capex (FRBSF) covers publicly traded telecom "
    "service companies; the TIA/CSFB and LA Times figures use broader "
    "industry definitions, hence larger cumulative totals.",
    "Values are nominal (not inflation-adjusted). Between 2000 and 2008 the "
    "US CPI rose ~25%, which slightly overstates the 2008-era dollar "
    "figures relative to the dot-com era.",
]
for i, c in enumerate(caveats):
    put(ws, r + 1 + i, 1, chr(8226), font=BOLD)
    put(ws, r + 1 + i, 2, c, wrap=True, font=NORMAL)
    ws.merge_cells(start_row=r + 1 + i, start_column=2,
                   end_row=r + 1 + i, end_column=3)
    ws.row_dimensions[r + 1 + i].height = 42

r += 7
section(ws, r, "Sources", 3)
header_row(ws, r + 1, ["#", "Source", "Data used"])
sources = [
    ("SIA / WSTS market data, Factbooks and press releases "
     "(semiconductors.org; wsts.org)",
     "Annual worldwide semiconductor sales 1995-2011, incl. 1999 $149.4B, "
     "2000 $204.4B, 2009 $226.3B, 2010 record $298.3B (+31.8%)"),
    ("Gartner Dataquest via Lightwave Online, 'Gartner Dataquest forecasts "
     "decline in worldwide semiconductor spending' (Jul-2001)",
     "2001 semi capex $47.3B (-26%); equipment $39.9B (2000) to $27.9B "
     "(2001); capex +84% in 2000"),
    ("Gartner press releases 2002-2007 (compilation at "
     "kirklindstrom.blogspot.com)",
     "Semi capex 2002 $27.8B, 2003 $28.9B, 2004 forecast $37B (+28%); "
     "equipment 2003 $21.7B, 2004f $29.5B; 2007 capex $56.3B (+0.6%)"),
    ("Gartner final results via Semiconductor Digest (Apr-2009) and "
     "PresseBox (Apr-2010)",
     "Equipment capex 2008 $30.7B (-31.7%) and 2009 $16.6B (-45.8%); "
     "memory capex -54% in 2009; WFE -47%"),
    ("IC Insights, McClean Report bulletin (Feb-2016)",
     "Capex double-digit downturn periods (1997-98, 2001-02, 2008-09) and "
     "the >=45% rebound pattern; 2010 rebound anomaly"),
    ("SEMI WWSEMS via Semiconductor Digest (Mar-2002, Mar-2008) and "
     "Electronics Weekly (Dec-2009)",
     "Equipment billings: 2000 $47.7B, 2001 $28.1B (-41%), 2006 $40.47B, "
     "2007 $42.77B, 2008 $29.5B (-31%), 2009 $15.9B (-46%, lowest since "
     "1994), 2010 $39.5B (+148%)"),
    ("FRBSF Economic Review 2004, 'The Boom and Bust in Information "
     "Technology Investment' (Doms)",
     "US publicly traded telecom carrier capex: $47B (1995), $121B peak "
     "(2000), $49B (2002)"),
    ("TIA FCC filing 'Investment, Capital Spending and Service Quality in "
     "U.S. Telecommunications Networks' (Nov-2002)",
     "Carrier capex +28%/yr 1996-2000 vs +10%/yr revenues; CSFB estimate "
     "of 2002 capex -48% vs 2001"),
    ("Los Angeles Times, 'Too Much, Too Soon for Telecom' (30-Jun-2002)",
     "~$444B cumulative telecom capex 1996-2001; ~$300B sector debt"),
    ("Federal Reserve Bank of Richmond Economic Quarterly (Fall-2003), "
     "'Boom and Bust in Telecommunications' (Wolman)",
     "Real communications-equipment investment $62B (1996) to $135B "
     "(Q4-2000) to $93B (Q4-2001); NASDAQ telecom index 198 to 1,230 to "
     "136"),
    ("BEA via FRED, series A008RL1A225NBEA",
     "US real private nonresidential fixed investment, annual % change "
     "1995-2011 (2001 -2.2%, 2002 -6.9%, 2008 +0.6%, 2009 -14.5%)"),
    ("Dataquest via ZDNet/HPCwire, 'Memory makers hit rock bottom' "
     "(Aug-2001)",
     "128Mb SDRAM ~$18.40 (mid-2000) to ~$1.50 spot (Aug-2001)"),
    ("Electronics Weekly, 'Memory prices hit rock bottom as 64Mbit DRAM "
     "sinks below $1' (Jul-2001)",
     "64Mb at $0.92; 128Mb <$2 on Asian spot; break-even ~$3; maker losses "
     "(Merrill, Infineon, Micron); only Toshiba cut output; DRAM market "
     "$40B (1996) / $14B (1998) / $31.5B (2000) / $14B (2001e)"),
    ("EDN / Gartner Dataquest, 'DRAM sales to fall 55.5%...' (Jun-2001)",
     "2001 DRAM revenue -55.5% ($31.5B to $14B), worst year ever; prices "
     "-80% in 12 months; 128MB module $120 to <$20; 1995 DRAM peak ~$40B+"),
    ("CNET, 'Have memory prices hit bottom?' (2001)",
     "64Mb spot $2.25 and 128Mb $4.38 (spring 2001); channel inventory 20 "
     "to 10 weeks"),
    ("DRAMeXchange / TrendForce weekly research (Oct-2008 to Jan-2009)",
     "DDR2 1Gb: $2.29 (6-May-08), below ~$1.50 cash cost (6-Sep), $1.28 "
     "(30-Sep, -7% in a day), $1.13 (17-Oct), record $0.58 low; DDR2 512Mb "
     "<$1.00 (Nov-07); contract -34% QoQ 3Q08 / -37% QoQ 4Q08; module "
     "prices; NAND -30% QoQ 4Q08, 8Gb $1.50 to $0.95; output cuts "
     "(Taiwan ~30%, Hynix 8-inch closures); 2008 DRAM industry review"),
    ("Gartner Dataquest PC data via EDN/Computerworld (Jan-2002) and "
     "Gartner press (Jan-2009)",
     "2001 PCs 128M (-4.6%, first decline since 1985; US -11.1%); 2008 PCs "
     "302.2M (+10.9%); Q4-08 78.1M (+1.1%, worst since 2002; US -10%)"),
    ("Gartner mobile phone data via phys.org (Mar-2009), Mobile World Live "
     "(Feb-2010) and Gartner Dataquest (Mar-2002)",
     "2000 412.7M / 2001 399.6M (-3.2%, first annual decline ever); 2008 "
     "1.22B (+6%), Q4-08 -4.6%; 2009 1.211B (-0.9%)"),
    ("Exchange data (NASDAQ, S&P, PHLX) - standard closing levels",
     "Index peaks/troughs on the Market Context sheet; SOX and NASDAQ "
     "Telecom levels approximate"),
]
for i, (src, used) in enumerate(sources):
    rr = r + 2 + i
    put(ws, rr, 1, i + 1, font=BOLD)
    put(ws, rr, 2, src, wrap=True)
    put(ws, rr, 3, used, wrap=True, font=NOTE_FONT)
    ws.row_dimensions[rr].height = 42

out = "Tech_Bubble_Comparison_1999_vs_2008.xlsx"
wb.save(out)
print(f"Wrote {out}")
