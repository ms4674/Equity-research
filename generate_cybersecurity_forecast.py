import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from copy import copy

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ACCENT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
GROWTH_POSITIVE_FONT = Font(name="Calibri", color="006100", size=11)
GROWTH_NEGATIVE_FONT = Font(name="Calibri", color="9C0006", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="2F5496")
NORMAL_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
PCT_FORMAT = "0.0%"
USD_FORMAT = '#,##0.0'
USD_FORMAT_INT = '#,##0'
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

YEARS = list(range(2019, 2031))
HISTORICAL_END = 2024
FORECAST_LABEL = "2025E–2030E"

SEGMENTS = {
    "SIEM": {
        "full_name": "Security Information & Event Management (SIEM)",
        "market_size": {
            2019: 3.6, 2020: 3.9, 2021: 4.4, 2022: 5.0, 2023: 5.7,
            2024: 6.4, 2025: 7.2, 2026: 8.1, 2027: 9.2, 2028: 10.4,
            2029: 11.8, 2030: 13.4,
        },
        "companies": {
            "Splunk (Cisco)":    {2019: 0.90, 2020: 1.01, 2021: 1.14, 2022: 1.25, 2023: 1.37, 2024: 1.47, 2025: 1.58, 2026: 1.70, 2027: 1.84, 2028: 1.99, 2029: 2.12, 2030: 2.28},
            "Microsoft":        {2019: 0.18, 2020: 0.27, 2021: 0.44, 2022: 0.65, 2023: 0.86, 2024: 1.09, 2025: 1.37, 2026: 1.62, 2027: 1.93, 2028: 2.29, 2029: 2.71, 2030: 3.22},
            "IBM":              {2019: 0.54, 2020: 0.55, 2021: 0.55, 2022: 0.55, 2023: 0.54, 2024: 0.54, 2025: 0.54, 2026: 0.53, 2027: 0.53, 2028: 0.52, 2029: 0.52, 2030: 0.51},
            "Palo Alto (XSIAM)":{2019: 0.00, 2020: 0.00, 2021: 0.00, 2022: 0.05, 2023: 0.17, 2024: 0.38, 2025: 0.58, 2026: 0.81, 2027: 1.01, 2028: 1.25, 2029: 1.53, 2030: 1.87},
            "Securonix":        {2019: 0.11, 2020: 0.14, 2021: 0.18, 2022: 0.22, 2023: 0.26, 2024: 0.29, 2025: 0.32, 2026: 0.36, 2027: 0.41, 2028: 0.47, 2029: 0.53, 2030: 0.60},
            "Exabeam":          {2019: 0.07, 2020: 0.10, 2021: 0.13, 2022: 0.16, 2023: 0.18, 2024: 0.19, 2025: 0.22, 2026: 0.24, 2027: 0.28, 2028: 0.31, 2029: 0.35, 2030: 0.40},
            "LogRhythm":        {2019: 0.14, 2020: 0.14, 2021: 0.14, 2022: 0.13, 2023: 0.12, 2024: 0.11, 2025: 0.11, 2026: 0.10, 2027: 0.10, 2028: 0.09, 2029: 0.09, 2030: 0.08},
        },
    },
    "Vulnerability Assessment": {
        "full_name": "Vulnerability Assessment & Management",
        "market_size": {
            2019: 9.0, 2020: 9.8, 2021: 10.8, 2022: 11.9, 2023: 13.2,
            2024: 14.9, 2025: 16.1, 2026: 17.5, 2027: 19.0, 2028: 20.6,
            2029: 22.3, 2030: 24.1,
        },
        "companies": {
            "Tenable":          {2019: 0.35, 2020: 0.44, 2021: 0.54, 2022: 0.68, 2023: 0.79, 2024: 0.90, 2025: 1.01, 2026: 1.13, 2027: 1.27, 2028: 1.42, 2029: 1.58, 2030: 1.76},
            "Qualys":           {2019: 0.32, 2020: 0.36, 2021: 0.41, 2022: 0.49, 2023: 0.55, 2024: 0.62, 2025: 0.69, 2026: 0.77, 2027: 0.86, 2028: 0.96, 2029: 1.07, 2030: 1.19},
            "Rapid7":           {2019: 0.33, 2020: 0.41, 2021: 0.54, 2022: 0.68, 2023: 0.81, 2024: 0.88, 2025: 0.96, 2026: 1.05, 2027: 1.14, 2028: 1.24, 2029: 1.35, 2030: 1.47},
            "Microsoft":        {2019: 0.10, 2020: 0.15, 2021: 0.22, 2022: 0.33, 2023: 0.46, 2024: 0.60, 2025: 0.77, 2026: 0.96, 2027: 1.14, 2028: 1.34, 2029: 1.56, 2030: 1.81},
            "CrowdStrike":      {2019: 0.05, 2020: 0.08, 2021: 0.14, 2022: 0.24, 2023: 0.36, 2024: 0.52, 2025: 0.66, 2026: 0.79, 2027: 0.95, 2028: 1.14, 2029: 1.37, 2030: 1.64},
            "Ivanti":           {2019: 0.18, 2020: 0.20, 2021: 0.23, 2022: 0.25, 2023: 0.28, 2024: 0.30, 2025: 0.32, 2026: 0.35, 2027: 0.38, 2028: 0.41, 2029: 0.45, 2030: 0.48},
        },
    },
    "Identity (IAM)": {
        "full_name": "Identity & Access Management (IAM)",
        "market_size": {
            2019: 10.5, 2020: 12.0, 2021: 14.0, 2022: 16.5, 2023: 19.5,
            2024: 22.9, 2025: 26.5, 2026: 29.5, 2027: 32.0, 2028: 34.3,
            2029: 37.0, 2030: 40.0,
        },
        "companies": {
            "Microsoft":        {2019: 2.10, 2020: 2.64, 2021: 3.36, 2022: 4.29, 2023: 5.46, 2024: 6.87, 2025: 8.22, 2026: 9.44, 2027: 10.56, 2028: 11.66, 2029: 12.95, 2030: 14.40},
            "Okta":             {2019: 0.47, 2020: 0.59, 2021: 0.84, 2022: 1.32, 2023: 1.86, 2024: 2.52, 2025: 2.92, 2026: 3.25, 2027: 3.52, 2028: 3.77, 2029: 4.07, 2030: 4.40},
            "CyberArk":         {2019: 0.39, 2020: 0.42, 2021: 0.47, 2022: 0.56, 2023: 0.72, 2024: 0.94, 2025: 1.19, 2026: 1.48, 2027: 1.73, 2028: 1.89, 2029: 2.04, 2030: 2.20},
            "SailPoint":        {2019: 0.28, 2020: 0.32, 2021: 0.38, 2022: 0.45, 2023: 0.55, 2024: 0.64, 2025: 0.74, 2026: 0.83, 2027: 0.93, 2028: 1.03, 2029: 1.11, 2030: 1.20},
            "Ping Identity":    {2019: 0.24, 2020: 0.26, 2021: 0.30, 2022: 0.35, 2023: 0.42, 2024: 0.50, 2025: 0.58, 2026: 0.65, 2027: 0.74, 2028: 0.82, 2029: 0.89, 2030: 0.96},
            "IBM":              {2019: 0.63, 2020: 0.66, 2021: 0.70, 2022: 0.74, 2023: 0.78, 2024: 0.82, 2025: 0.85, 2026: 0.88, 2027: 0.90, 2028: 0.93, 2029: 0.96, 2030: 1.00},
            "Oracle":           {2019: 0.53, 2020: 0.54, 2021: 0.56, 2022: 0.59, 2023: 0.63, 2024: 0.69, 2025: 0.74, 2026: 0.80, 2027: 0.86, 2028: 0.93, 2029: 1.00, 2030: 1.08},
        },
    },
    "Endpoint Security": {
        "full_name": "Endpoint Security (EPP / EDR / XDR)",
        "market_size": {
            2019: 12.8, 2020: 13.5, 2021: 14.6, 2022: 16.1, 2023: 17.8,
            2024: 18.4, 2025: 19.8, 2026: 21.3, 2027: 22.9, 2028: 24.7,
            2029: 26.6, 2030: 28.6,
        },
        "companies": {
            "CrowdStrike":      {2019: 0.46, 2020: 0.68, 2021: 1.07, 2022: 1.73, 2023: 2.45, 2024: 3.06, 2025: 3.52, 2026: 3.94, 2027: 4.35, 2028: 4.69, 2029: 5.05, 2030: 5.43},
            "Microsoft":        {2019: 1.15, 2020: 1.49, 2021: 1.90, 2022: 2.42, 2023: 3.02, 2024: 3.50, 2025: 3.96, 2026: 4.47, 2027: 5.04, 2028: 5.68, 2029: 6.38, 2030: 7.15},
            "Palo Alto Networks":{2019: 0.38, 2020: 0.54, 2021: 0.73, 2022: 0.97, 2023: 1.25, 2024: 1.47, 2025: 1.68, 2026: 1.92, 2027: 2.18, 2028: 2.47, 2029: 2.79, 2030: 3.15},
            "SentinelOne":      {2019: 0.05, 2020: 0.10, 2021: 0.20, 2022: 0.42, 2023: 0.62, 2024: 0.74, 2025: 0.85, 2026: 0.96, 2027: 1.07, 2028: 1.19, 2029: 1.33, 2030: 1.49},
            "Trellix (McAfee)": {2019: 1.41, 2020: 1.35, 2021: 1.31, 2022: 1.21, 2023: 1.12, 2024: 1.01, 2025: 0.95, 2026: 0.89, 2027: 0.82, 2028: 0.77, 2029: 0.72, 2030: 0.66},
            "Trend Micro":      {2019: 0.90, 2020: 0.90, 2021: 0.92, 2022: 0.93, 2023: 0.93, 2024: 0.92, 2025: 0.93, 2026: 0.94, 2027: 0.96, 2028: 0.99, 2029: 1.01, 2030: 1.03},
            "Broadcom (Symantec)":{2019: 1.28, 2020: 1.15, 2021: 1.05, 2022: 0.97, 2023: 0.89, 2024: 0.83, 2025: 0.77, 2026: 0.72, 2027: 0.69, 2028: 0.64, 2029: 0.61, 2030: 0.57},
        },
    },
    "Network Security": {
        "full_name": "Network Security (Firewall / IDS/IPS / UTM)",
        "market_size": {
            2019: 15.5, 2020: 16.6, 2021: 18.3, 2022: 20.5, 2023: 23.8,
            2024: 26.6, 2025: 28.7, 2026: 31.0, 2027: 33.5, 2028: 36.2,
            2029: 39.1, 2030: 42.3,
        },
        "companies": {
            "Palo Alto Networks":{2019: 2.79, 2020: 3.07, 2021: 3.66, 2022: 4.31, 2023: 5.24, 2024: 5.99, 2025: 6.60, 2026: 7.28, 2027: 8.04, 2028: 8.69, 2029: 9.39, 2030: 10.15},
            "Fortinet":         {2019: 2.02, 2020: 2.32, 2021: 2.93, 2022: 3.59, 2023: 4.21, 2024: 4.71, 2025: 5.17, 2026: 5.58, 2027: 6.03, 2028: 6.51, 2029: 7.04, 2030: 7.61},
            "Cisco":            {2019: 2.17, 2020: 2.22, 2021: 2.38, 2022: 2.56, 2023: 2.85, 2024: 2.93, 2025: 3.07, 2026: 3.22, 2027: 3.35, 2028: 3.44, 2029: 3.52, 2030: 3.59},
            "Check Point":      {2019: 1.86, 2020: 1.94, 2021: 2.06, 2022: 2.15, 2023: 2.14, 2024: 2.13, 2025: 2.15, 2026: 2.17, 2027: 2.18, 2028: 2.18, 2029: 2.18, 2030: 2.15},
            "Juniper Networks":  {2019: 0.62, 2020: 0.63, 2021: 0.64, 2022: 0.72, 2023: 0.81, 2024: 0.85, 2025: 0.89, 2026: 0.93, 2027: 0.97, 2028: 1.01, 2029: 1.05, 2030: 1.10},
            "Zscaler":          {2019: 0.19, 2020: 0.31, 2021: 0.51, 2022: 0.82, 2023: 1.14, 2024: 1.46, 2025: 1.72, 2026: 2.01, 2027: 2.35, 2028: 2.75, 2029: 3.21, 2030: 3.76},
        },
    },
    "SASE": {
        "full_name": "Secure Access Service Edge (SASE)",
        "market_size": {
            2019: 1.2, 2020: 2.0, 2021: 3.4, 2022: 5.6, 2023: 8.4,
            2024: 9.6, 2025: 10.9, 2026: 12.4, 2027: 14.1, 2028: 16.0,
            2029: 18.2, 2030: 20.7,
        },
        "companies": {
            "Zscaler":          {2019: 0.22, 2020: 0.39, 2021: 0.67, 2022: 1.18, 2023: 1.85, 2024: 2.02, 2025: 2.29, 2026: 2.60, 2027: 2.96, 2028: 3.36, 2029: 3.82, 2030: 4.35},
            "Cisco":            {2019: 0.14, 2020: 0.26, 2021: 0.48, 2022: 0.84, 2023: 1.34, 2024: 1.54, 2025: 1.74, 2026: 1.98, 2027: 2.26, 2028: 2.56, 2029: 2.91, 2030: 3.31},
            "Palo Alto Networks":{2019: 0.08, 2020: 0.18, 2021: 0.37, 2022: 0.73, 2023: 1.18, 2024: 1.34, 2025: 1.53, 2026: 1.74, 2027: 1.97, 2028: 2.24, 2029: 2.55, 2030: 2.90},
            "Broadcom":         {2019: 0.06, 2020: 0.12, 2021: 0.24, 2022: 0.45, 2023: 0.76, 2024: 0.86, 2025: 0.98, 2026: 1.12, 2027: 1.27, 2028: 1.44, 2029: 1.64, 2030: 1.86},
            "Fortinet":         {2019: 0.05, 2020: 0.10, 2021: 0.20, 2022: 0.39, 2023: 0.59, 2024: 0.67, 2025: 0.76, 2026: 0.87, 2027: 0.99, 2028: 1.12, 2029: 1.27, 2030: 1.45},
            "Netskope":         {2019: 0.04, 2020: 0.08, 2021: 0.17, 2022: 0.34, 2023: 0.50, 2024: 0.58, 2025: 0.65, 2026: 0.74, 2027: 0.85, 2028: 0.96, 2029: 1.09, 2030: 1.24},
            "Cloudflare":       {2019: 0.02, 2020: 0.04, 2021: 0.10, 2022: 0.22, 2023: 0.34, 2024: 0.43, 2025: 0.54, 2026: 0.68, 2027: 0.85, 2028: 1.06, 2029: 1.33, 2030: 1.66},
        },
    },
}


##############################################################################
# LLM-IMPACT SCENARIO
# For each segment we define:
#   "llm_impact_pct"  – dict {year: pct} additive market-size adjustment
#                       (positive = LLM tailwind enlarges TAM,
#                        negative = LLM replaces spend / cannibalises)
#   "assumptions"     – list of plain-text assumption strings
#   "winners"         – companies that gain disproportionate share from LLM
#   "losers"          – companies that lose share
#   "winner_boost"    – fractional revenue uplift applied to winners (on top
#                       of the base-case)
#   "loser_drag"      – fractional revenue drag applied to losers
##############################################################################

LLM_IMPACT = {
    "SIEM": {
        "llm_impact_pct": {
            2019: 0, 2020: 0, 2021: 0, 2022: 0, 2023: 0, 2024: 0,
            2025: 0.03, 2026: 0.06, 2027: 0.10, 2028: 0.14, 2029: 0.18, 2030: 0.22,
        },
        "assumptions": [
            "LLM-based search transforms SIEM from query-driven log analysis to natural-language investigation, expanding TAM by attracting non-expert users.",
            "Tier-1 SOC alert triage automated 40-60% by LLM copilots (Microsoft Security Copilot, Palo Alto XSIAM), reducing per-alert cost but increasing detection volume.",
            "Vendors embed LLM features into premium tiers, driving ARPU uplift of 10-15% on new bookings from 2026+.",
            "Legacy rule-based SIEM vendors (LogRhythm, IBM QRadar) lose share faster as LLM-native platforms (Microsoft Sentinel, XSIAM) offer superior natural-language querying.",
            "Net effect: TAM expands because LLM raises detection efficacy, driving demand, but spend per analyst-hour drops — net positive for market overall.",
        ],
        "winners": ["Microsoft", "Palo Alto (XSIAM)"],
        "losers": ["IBM", "LogRhythm", "Exabeam"],
        "winner_boost": {2025: 0.05, 2026: 0.10, 2027: 0.15, 2028: 0.18, 2029: 0.20, 2030: 0.22},
        "loser_drag":   {2025: -0.03, 2026: -0.06, 2027: -0.10, 2028: -0.14, 2029: -0.18, 2030: -0.22},
    },
    "Vulnerability Assessment": {
        "llm_impact_pct": {
            2019: 0, 2020: 0, 2021: 0, 2022: 0, 2023: 0, 2024: 0,
            2025: 0.04, 2026: 0.08, 2027: 0.13, 2028: 0.18, 2029: 0.24, 2030: 0.30,
        },
        "assumptions": [
            "LLM code review automates ~40% of static analysis tasks (Checkmarx, Snyk) — but simultaneously 40%+ of AI-generated code ships with vulnerabilities, dramatically expanding scan demand.",
            "Autonomous LLM pen-testing agents (e.g., MAPTA) achieve 77% success rate at <$0.10/test, creating new low-cost VM consumption tier.",
            "Biggest TAM expansion among all segments: more code = more vulnerabilities = more scanning, offsetting any per-scan price compression.",
            "Platform vendors integrating LLM-driven prioritization (CrowdStrike, Microsoft Defender) gain share vs. point-solution scanners.",
            "Tenable and Qualys defend share by embedding LLM prioritisation into exposure management; Rapid7 risks share loss without comparable LLM investment.",
        ],
        "winners": ["Microsoft", "CrowdStrike"],
        "losers": ["Ivanti"],
        "winner_boost": {2025: 0.06, 2026: 0.12, 2027: 0.18, 2028: 0.22, 2029: 0.26, 2030: 0.30},
        "loser_drag":   {2025: -0.04, 2026: -0.08, 2027: -0.12, 2028: -0.16, 2029: -0.20, 2030: -0.24},
    },
    "Identity (IAM)": {
        "llm_impact_pct": {
            2019: 0, 2020: 0, 2021: 0, 2022: 0, 2023: 0, 2024: 0,
            2025: 0.03, 2026: 0.07, 2027: 0.12, 2028: 0.17, 2029: 0.22, 2030: 0.28,
        },
        "assumptions": [
            "AI agents require machine-to-machine identity at scale, creating a new 'agent identity' sub-segment that expands IAM TAM 15-25% by 2030.",
            "LLMs enable natural-language policy authoring (NIST research), lowering barrier for Zero Trust adoption — accelerating IAM adoption in mid-market.",
            "Privileged access management (CyberArk) sees uplift: LLM agents handling sensitive operations need robust secret vaulting and session recording.",
            "Behavioral biometrics enhanced by LLM anomaly detection raise ARPU for adaptive authentication vendors.",
            "Microsoft benefits most from Entra ID + Copilot integration; Okta and CyberArk well-positioned for workforce & privileged-agent identity.",
        ],
        "winners": ["Microsoft", "CyberArk", "Okta"],
        "losers": ["IBM", "Oracle"],
        "winner_boost": {2025: 0.04, 2026: 0.08, 2027: 0.12, 2028: 0.16, 2029: 0.19, 2030: 0.22},
        "loser_drag":   {2025: -0.02, 2026: -0.04, 2027: -0.06, 2028: -0.08, 2029: -0.10, 2030: -0.12},
    },
    "Endpoint Security": {
        "llm_impact_pct": {
            2019: 0, 2020: 0, 2021: 0, 2022: 0, 2023: 0, 2024: 0,
            2025: 0.02, 2026: 0.05, 2027: 0.08, 2028: 0.12, 2029: 0.15, 2030: 0.18,
        },
        "assumptions": [
            "LLM-powered malware generation increases sophistication of attacks, driving demand for next-gen EDR/XDR with LLM-based behavioral analysis.",
            "LLM copilots embedded in EDR platforms accelerate investigation (CrowdStrike Charlotte AI, Microsoft Copilot for Defender) — premium tier pricing.",
            "Legacy signature-based products accelerate decline; LLM-native XDR platforms capture displaced share.",
            "Moderate TAM uplift: endpoints already well-penetrated; LLM impact is more about vendor shift than market expansion.",
            "AI-generated phishing/malware raises baseline threat level, sustaining demand growth above pre-AI trajectory.",
        ],
        "winners": ["CrowdStrike", "Microsoft", "Palo Alto Networks", "SentinelOne"],
        "losers": ["Trellix (McAfee)", "Broadcom (Symantec)"],
        "winner_boost": {2025: 0.03, 2026: 0.06, 2027: 0.10, 2028: 0.13, 2029: 0.15, 2030: 0.17},
        "loser_drag":   {2025: -0.04, 2026: -0.08, 2027: -0.12, 2028: -0.16, 2029: -0.20, 2030: -0.24},
    },
    "Network Security": {
        "llm_impact_pct": {
            2019: 0, 2020: 0, 2021: 0, 2022: 0, 2023: 0, 2024: 0,
            2025: 0.02, 2026: 0.04, 2027: 0.06, 2028: 0.09, 2029: 0.12, 2030: 0.15,
        },
        "assumptions": [
            "LLM-based search has least direct impact on network security hardware (firewalls, IDS/IPS), which is driven by appliance refresh cycles.",
            "AI-assisted policy management (natural-language firewall rule generation) improves operational efficiency but does not significantly change spend.",
            "LLM-powered DDoS detection and encrypted traffic analysis create incremental software-layer revenue for cloud-native vendors (Zscaler, Cloudflare).",
            "Traditional appliance vendors (Check Point, Juniper) see minimal LLM uplift; cloud-native vendors disproportionately benefit.",
            "Net effect: modest TAM uplift from AI-enhanced threat detection features; market growth primarily structural (cloud migration, Zero Trust).",
        ],
        "winners": ["Zscaler", "Palo Alto Networks"],
        "losers": ["Check Point", "Juniper Networks"],
        "winner_boost": {2025: 0.03, 2026: 0.06, 2027: 0.09, 2028: 0.12, 2029: 0.14, 2030: 0.16},
        "loser_drag":   {2025: -0.02, 2026: -0.04, 2027: -0.06, 2028: -0.08, 2029: -0.10, 2030: -0.12},
    },
    "SASE": {
        "llm_impact_pct": {
            2019: 0, 2020: 0, 2021: 0, 2022: 0, 2023: 0, 2024: 0,
            2025: 0.04, 2026: 0.08, 2027: 0.13, 2028: 0.18, 2029: 0.24, 2030: 0.30,
        },
        "assumptions": [
            "LLM-based search and AI agents massively increase API-to-API and agent-to-service traffic, driving SSE/SASE inspection volume growth.",
            "Natural-language policy authoring simplifies SASE deployment, accelerating mid-market adoption and reducing implementation friction.",
            "AI agent traffic requires real-time DLP and content inspection — strong tailwind for SSE components (CASB, SWG, ZTNA).",
            "Single-vendor SASE platforms with embedded AI analytics gain share faster (Zscaler, Palo Alto Prisma).",
            "Cloudflare benefits disproportionately: AI inference traffic traverses its edge network, creating bundled security + AI delivery revenue.",
        ],
        "winners": ["Zscaler", "Palo Alto Networks", "Cloudflare"],
        "losers": ["Broadcom"],
        "winner_boost": {2025: 0.05, 2026: 0.10, 2027: 0.15, 2028: 0.20, 2029: 0.24, 2030: 0.28},
        "loser_drag":   {2025: -0.03, 2026: -0.06, 2027: -0.09, 2028: -0.12, 2029: -0.15, 2030: -0.18},
    },
}


def compute_llm_adjusted(segment_key):
    """Return (adjusted_market_size, adjusted_companies) dicts for a segment."""
    base = SEGMENTS[segment_key]
    impact = LLM_IMPACT[segment_key]
    adj_market = {}
    for yr in YEARS:
        adj_market[yr] = round(base["market_size"][yr] * (1 + impact["llm_impact_pct"][yr]), 2)

    adj_companies = {}
    for company, rev in base["companies"].items():
        adj_companies[company] = {}
        for yr in YEARS:
            if yr <= HISTORICAL_END:
                adj_companies[company][yr] = rev[yr]
                continue
            base_rev = rev[yr]
            if company in impact["winners"]:
                boost = impact["winner_boost"].get(yr, 0)
                adj_companies[company][yr] = round(base_rev * (1 + boost), 2)
            elif company in impact["losers"]:
                drag = impact["loser_drag"].get(yr, 0)
                adj_companies[company][yr] = round(base_rev * (1 + drag), 2)
            else:
                mkt_mult = 1 + impact["llm_impact_pct"].get(yr, 0)
                adj_companies[company][yr] = round(base_rev * mkt_mult, 2)
    return adj_market, adj_companies


def apply_cell_style(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def style_range(ws, row, col_start, col_end, **kwargs):
    for c in range(col_start, col_end + 1):
        apply_cell_style(ws.cell(row=row, column=c), **kwargs)


def write_segment_sheet(wb, segment_key, data):
    ws = wb.create_sheet(title=segment_key)

    ws.cell(row=1, column=1, value=data["full_name"])
    apply_cell_style(ws.cell(row=1, column=1), font=TITLE_FONT)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(YEARS) + 2)

    ws.cell(row=2, column=1, value="All figures in USD Billions. Sources: IDC, Gartner, Dell'Oro, company filings. 2025-2030 are estimates.")
    apply_cell_style(ws.cell(row=2, column=1), font=Font(name="Calibri", italic=True, size=10, color="666666"))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(YEARS) + 2)

    # --- SECTION 1: Total Market Size ---
    row = 4
    ws.cell(row=row, column=1, value="Total Market Size ($B)")
    apply_cell_style(ws.cell(row=row, column=1), font=HEADER_FONT, fill=HEADER_FILL, alignment=Alignment(horizontal="left"))
    for i, yr in enumerate(YEARS):
        cell = ws.cell(row=row, column=i + 2, value=yr)
        apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL, alignment=Alignment(horizontal="center"))
    style_range(ws, row, 1, len(YEARS) + 1, border=THIN_BORDER)

    row = 5
    ws.cell(row=row, column=1, value="Market Size")
    apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, border=THIN_BORDER)
    for i, yr in enumerate(YEARS):
        cell = ws.cell(row=row, column=i + 2, value=data["market_size"][yr])
        fmt = USD_FORMAT
        apply_cell_style(cell, font=NORMAL_FONT, border=THIN_BORDER, number_format=fmt, alignment=Alignment(horizontal="center"))
        if yr > HISTORICAL_END:
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    row = 6
    ws.cell(row=row, column=1, value="YoY Growth (%)")
    apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, border=THIN_BORDER)
    for i, yr in enumerate(YEARS):
        if i == 0:
            ws.cell(row=row, column=i + 2, value="—")
            apply_cell_style(ws.cell(row=row, column=i + 2), font=NORMAL_FONT, border=THIN_BORDER, alignment=Alignment(horizontal="center"))
            continue
        prev = data["market_size"][YEARS[i - 1]]
        cur = data["market_size"][yr]
        growth = (cur - prev) / prev
        cell = ws.cell(row=row, column=i + 2, value=growth)
        gfont = GROWTH_POSITIVE_FONT if growth >= 0 else GROWTH_NEGATIVE_FONT
        apply_cell_style(cell, font=gfont, border=THIN_BORDER, number_format=PCT_FORMAT, alignment=Alignment(horizontal="center"))
        if yr > HISTORICAL_END:
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # --- SECTION 2: Company Revenue ---
    row = 8
    ws.cell(row=row, column=1, value="Company Revenue ($B)")
    apply_cell_style(ws.cell(row=row, column=1), font=HEADER_FONT, fill=SUBHEADER_FILL, alignment=Alignment(horizontal="left"))
    for i, yr in enumerate(YEARS):
        cell = ws.cell(row=row, column=i + 2, value=yr)
        apply_cell_style(cell, font=HEADER_FONT, fill=SUBHEADER_FILL, alignment=Alignment(horizontal="center"))
    style_range(ws, row, 1, len(YEARS) + 1, border=THIN_BORDER)

    company_start_row = row + 1
    for ci, (company, rev_data) in enumerate(data["companies"].items()):
        row = company_start_row + ci
        ws.cell(row=row, column=1, value=company)
        f = ACCENT_FILL if ci % 2 == 0 else None
        apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, fill=f, border=THIN_BORDER)
        for i, yr in enumerate(YEARS):
            cell = ws.cell(row=row, column=i + 2, value=rev_data[yr])
            apply_cell_style(cell, font=NORMAL_FONT, fill=f, border=THIN_BORDER, number_format=USD_FORMAT, alignment=Alignment(horizontal="center"))
            if yr > HISTORICAL_END:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Others row
    row += 1
    ws.cell(row=row, column=1, value="Others")
    apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, border=THIN_BORDER)
    for i, yr in enumerate(YEARS):
        company_total = sum(rev[yr] for rev in data["companies"].values())
        others = data["market_size"][yr] - company_total
        cell = ws.cell(row=row, column=i + 2, value=round(others, 2))
        apply_cell_style(cell, font=NORMAL_FONT, border=THIN_BORDER, number_format=USD_FORMAT, alignment=Alignment(horizontal="center"))
        if yr > HISTORICAL_END:
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # --- SECTION 3: Company Market Share ---
    row += 2
    share_header_row = row
    ws.cell(row=row, column=1, value="Market Share (%)")
    apply_cell_style(ws.cell(row=row, column=1), font=HEADER_FONT, fill=HEADER_FILL, alignment=Alignment(horizontal="left"))
    for i, yr in enumerate(YEARS):
        cell = ws.cell(row=row, column=i + 2, value=yr)
        apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL, alignment=Alignment(horizontal="center"))
    style_range(ws, row, 1, len(YEARS) + 1, border=THIN_BORDER)

    for ci, (company, rev_data) in enumerate(data["companies"].items()):
        row = share_header_row + 1 + ci
        ws.cell(row=row, column=1, value=company)
        f = ACCENT_FILL if ci % 2 == 0 else None
        apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, fill=f, border=THIN_BORDER)
        for i, yr in enumerate(YEARS):
            share = rev_data[yr] / data["market_size"][yr]
            cell = ws.cell(row=row, column=i + 2, value=share)
            apply_cell_style(cell, font=NORMAL_FONT, fill=f, border=THIN_BORDER, number_format=PCT_FORMAT, alignment=Alignment(horizontal="center"))
            if yr > HISTORICAL_END:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Others share
    row += 1
    ws.cell(row=row, column=1, value="Others")
    apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, border=THIN_BORDER)
    for i, yr in enumerate(YEARS):
        company_total = sum(rev[yr] for rev in data["companies"].values())
        others_share = (data["market_size"][yr] - company_total) / data["market_size"][yr]
        cell = ws.cell(row=row, column=i + 2, value=others_share)
        apply_cell_style(cell, font=NORMAL_FONT, border=THIN_BORDER, number_format=PCT_FORMAT, alignment=Alignment(horizontal="center"))
        if yr > HISTORICAL_END:
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # --- SECTION 4: Company Revenue Growth ---
    row += 2
    growth_header_row = row
    ws.cell(row=row, column=1, value="Company Revenue YoY Growth (%)")
    apply_cell_style(ws.cell(row=row, column=1), font=HEADER_FONT, fill=SUBHEADER_FILL, alignment=Alignment(horizontal="left"))
    for i, yr in enumerate(YEARS):
        cell = ws.cell(row=row, column=i + 2, value=yr)
        apply_cell_style(cell, font=HEADER_FONT, fill=SUBHEADER_FILL, alignment=Alignment(horizontal="center"))
    style_range(ws, row, 1, len(YEARS) + 1, border=THIN_BORDER)

    for ci, (company, rev_data) in enumerate(data["companies"].items()):
        row = growth_header_row + 1 + ci
        ws.cell(row=row, column=1, value=company)
        f = ACCENT_FILL if ci % 2 == 0 else None
        apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, fill=f, border=THIN_BORDER)
        for i, yr in enumerate(YEARS):
            if i == 0:
                ws.cell(row=row, column=i + 2, value="—")
                apply_cell_style(ws.cell(row=row, column=i + 2), font=NORMAL_FONT, fill=f, border=THIN_BORDER, alignment=Alignment(horizontal="center"))
                continue
            prev_val = rev_data[YEARS[i - 1]]
            cur_val = rev_data[yr]
            if prev_val == 0:
                ws.cell(row=row, column=i + 2, value="N/A")
                apply_cell_style(ws.cell(row=row, column=i + 2), font=NORMAL_FONT, fill=f, border=THIN_BORDER, alignment=Alignment(horizontal="center"))
                continue
            growth = (cur_val - prev_val) / prev_val
            cell = ws.cell(row=row, column=i + 2, value=growth)
            gfont = Font(name="Calibri", color="006100", size=11) if growth >= 0 else Font(name="Calibri", color="9C0006", size=11)
            apply_cell_style(cell, font=gfont, fill=f, border=THIN_BORDER, number_format=PCT_FORMAT, alignment=Alignment(horizontal="center"))
            if yr > HISTORICAL_END:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # --- Add chart: Market Size bar + line ---
    chart_row = row + 3
    chart = BarChart()
    chart.type = "col"
    chart.title = f"{segment_key} Market Size ($B)"
    chart.y_axis.title = "USD Billions"
    chart.x_axis.title = "Year"
    chart.style = 10
    chart.width = 28
    chart.height = 14

    cats = Reference(ws, min_col=2, max_col=len(YEARS) + 1, min_row=4)
    vals = Reference(ws, min_col=2, max_col=len(YEARS) + 1, min_row=5)
    chart.add_data(vals, from_rows=True, titles_from_data=False)
    chart.set_categories(cats)
    chart.series[0].title = openpyxl.chart.series.SeriesLabel(v="Market Size")
    chart.series[0].graphicalProperties.solidFill = "4472C4"

    ws.add_chart(chart, f"A{chart_row}")

    # Column widths
    ws.column_dimensions["A"].width = 28
    for i in range(len(YEARS)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 12

    ws.sheet_properties.tabColor = "2F5496"


def write_summary_sheet(wb):
    ws = wb.create_sheet(title="Summary", index=0)

    ws.cell(row=1, column=1, value="Cybersecurity Market Segments — Historical & Forecast Overview")
    apply_cell_style(ws.cell(row=1, column=1), font=TITLE_FONT)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(YEARS) + 2)

    ws.cell(row=2, column=1, value="All figures in USD Billions. Yellow-highlighted columns are forecasts (2025E–2030E).")
    apply_cell_style(ws.cell(row=2, column=1), font=Font(name="Calibri", italic=True, size=10, color="666666"))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(YEARS) + 2)

    # --- Header row ---
    row = 4
    ws.cell(row=row, column=1, value="Segment")
    apply_cell_style(ws.cell(row=row, column=1), font=HEADER_FONT, fill=HEADER_FILL, alignment=Alignment(horizontal="left"))
    for i, yr in enumerate(YEARS):
        cell = ws.cell(row=row, column=i + 2, value=yr)
        apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL, alignment=Alignment(horizontal="center"))
    cagr_col = len(YEARS) + 2
    ws.cell(row=row, column=cagr_col, value="CAGR\n'19-'30")
    apply_cell_style(ws.cell(row=row, column=cagr_col), font=HEADER_FONT, fill=HEADER_FILL, alignment=Alignment(horizontal="center", wrap_text=True))
    style_range(ws, row, 1, cagr_col, border=THIN_BORDER)

    # --- Market size rows ---
    seg_order = ["SIEM", "Vulnerability Assessment", "Identity (IAM)", "Endpoint Security", "Network Security", "SASE"]
    total_by_year = {yr: 0 for yr in YEARS}

    for si, seg_key in enumerate(seg_order):
        row = 5 + si
        d = SEGMENTS[seg_key]
        f = ACCENT_FILL if si % 2 == 0 else None
        ws.cell(row=row, column=1, value=seg_key)
        apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, fill=f, border=THIN_BORDER)
        for i, yr in enumerate(YEARS):
            val = d["market_size"][yr]
            total_by_year[yr] += val
            cell = ws.cell(row=row, column=i + 2, value=val)
            apply_cell_style(cell, font=NORMAL_FONT, fill=f, border=THIN_BORDER, number_format=USD_FORMAT, alignment=Alignment(horizontal="center"))
            if yr > HISTORICAL_END:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        cagr = (d["market_size"][2030] / d["market_size"][2019]) ** (1 / 11) - 1
        cell = ws.cell(row=row, column=cagr_col, value=cagr)
        apply_cell_style(cell, font=BOLD_FONT, fill=f, border=THIN_BORDER, number_format=PCT_FORMAT, alignment=Alignment(horizontal="center"))

    # Total row
    row = 5 + len(seg_order)
    ws.cell(row=row, column=1, value="TOTAL")
    apply_cell_style(ws.cell(row=row, column=1), font=Font(name="Calibri", bold=True, size=11, color="FFFFFF"), fill=PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid"), border=THIN_BORDER)
    for i, yr in enumerate(YEARS):
        cell = ws.cell(row=row, column=i + 2, value=round(total_by_year[yr], 1))
        apply_cell_style(cell, font=Font(name="Calibri", bold=True, size=11, color="FFFFFF"),
                         fill=PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid"),
                         border=THIN_BORDER, number_format=USD_FORMAT, alignment=Alignment(horizontal="center"))
    total_cagr = (total_by_year[2030] / total_by_year[2019]) ** (1 / 11) - 1
    cell = ws.cell(row=row, column=cagr_col, value=total_cagr)
    apply_cell_style(cell, font=Font(name="Calibri", bold=True, size=11, color="FFFFFF"),
                     fill=PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid"),
                     border=THIN_BORDER, number_format=PCT_FORMAT, alignment=Alignment(horizontal="center"))

    # --- Growth rates section ---
    row += 2
    growth_section_start = row
    ws.cell(row=row, column=1, value="YoY Growth Rate (%)")
    apply_cell_style(ws.cell(row=row, column=1), font=HEADER_FONT, fill=SUBHEADER_FILL, alignment=Alignment(horizontal="left"))
    for i, yr in enumerate(YEARS):
        cell = ws.cell(row=row, column=i + 2, value=yr)
        apply_cell_style(cell, font=HEADER_FONT, fill=SUBHEADER_FILL, alignment=Alignment(horizontal="center"))
    style_range(ws, row, 1, len(YEARS) + 1, border=THIN_BORDER)

    for si, seg_key in enumerate(seg_order):
        row = growth_section_start + 1 + si
        d = SEGMENTS[seg_key]
        f = ACCENT_FILL if si % 2 == 0 else None
        ws.cell(row=row, column=1, value=seg_key)
        apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, fill=f, border=THIN_BORDER)
        for i, yr in enumerate(YEARS):
            if i == 0:
                ws.cell(row=row, column=i + 2, value="—")
                apply_cell_style(ws.cell(row=row, column=i + 2), font=NORMAL_FONT, fill=f, border=THIN_BORDER, alignment=Alignment(horizontal="center"))
                continue
            prev = d["market_size"][YEARS[i - 1]]
            cur = d["market_size"][yr]
            growth = (cur - prev) / prev
            cell = ws.cell(row=row, column=i + 2, value=growth)
            gfont = GROWTH_POSITIVE_FONT if growth >= 0 else GROWTH_NEGATIVE_FONT
            apply_cell_style(cell, font=gfont, fill=f, border=THIN_BORDER, number_format=PCT_FORMAT, alignment=Alignment(horizontal="center"))
            if yr > HISTORICAL_END:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # --- Chart ---
    chart_row = row + 3
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.title = "Cybersecurity Segments — Market Size ($B)"
    chart.y_axis.title = "USD Billions"
    chart.x_axis.title = "Year"
    chart.style = 10
    chart.width = 30
    chart.height = 16

    cats = Reference(ws, min_col=2, max_col=len(YEARS) + 1, min_row=4)
    colors = ["4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5", "70AD47"]
    for si, seg_key in enumerate(seg_order):
        vals = Reference(ws, min_col=2, max_col=len(YEARS) + 1, min_row=5 + si)
        chart.add_data(vals, from_rows=True, titles_from_data=False)
        chart.set_categories(cats)
        chart.series[si].title = openpyxl.chart.series.SeriesLabel(v=seg_key)
        chart.series[si].graphicalProperties.solidFill = colors[si % len(colors)]

    ws.add_chart(chart, f"A{chart_row}")

    # Column widths
    ws.column_dimensions["A"].width = 28
    for i in range(len(YEARS)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 12
    ws.column_dimensions[get_column_letter(cagr_col)].width = 12

    ws.sheet_properties.tabColor = "1F3864"


##############################################################################
# LLM IMPACT — ASSUMPTIONS TAB
##############################################################################

LLM_GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LLM_RED_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
LLM_HEADER_FILL = PatternFill(start_color="7B2D8E", end_color="7B2D8E", fill_type="solid")
LLM_SUBHEADER_FILL = PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")
DELTA_POS_FONT = Font(name="Calibri", color="006100", size=11)
DELTA_NEG_FONT = Font(name="Calibri", color="9C0006", size=11)


def write_llm_assumptions_sheet(wb):
    ws = wb.create_sheet(title="LLM Assumptions")

    ws.cell(row=1, column=1, value="How LLM-Based Search Impacts Cybersecurity Segments")
    apply_cell_style(ws.cell(row=1, column=1), font=Font(name="Calibri", bold=True, size=16, color="7B2D8E"))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    ws.cell(row=2, column=1, value=(
        "This tab summarises the qualitative and quantitative assumptions behind the LLM-adjusted scenario. "
        "LLM-based search, AI copilots, and autonomous agents reshape buyer behaviour, vendor differentiation, "
        "and total addressable market for every cybersecurity segment."
    ))
    apply_cell_style(ws.cell(row=2, column=1), font=Font(name="Calibri", italic=True, size=10, color="666666"))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    seg_order = ["SIEM", "Vulnerability Assessment", "Identity (IAM)", "Endpoint Security", "Network Security", "SASE"]
    row = 4

    for seg_key in seg_order:
        impact = LLM_IMPACT[seg_key]

        # Segment header
        ws.cell(row=row, column=1, value=seg_key)
        for c in range(1, 9):
            apply_cell_style(ws.cell(row=row, column=c), font=HEADER_FONT, fill=LLM_HEADER_FILL, border=THIN_BORDER)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1

        # TAM impact multiplier row
        ws.cell(row=row, column=1, value="TAM Impact Multiplier")
        apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, border=THIN_BORDER)
        forecast_years = [yr for yr in YEARS if yr > HISTORICAL_END]
        for i, yr in enumerate(forecast_years):
            ws.cell(row=row, column=2 + i, value=yr)
            apply_cell_style(ws.cell(row=row, column=2 + i), font=BOLD_FONT, fill=ACCENT_FILL, border=THIN_BORDER, alignment=Alignment(horizontal="center"))
        row += 1
        ws.cell(row=row, column=1, value="Market Size Adjustment")
        apply_cell_style(ws.cell(row=row, column=1), font=NORMAL_FONT, border=THIN_BORDER)
        for i, yr in enumerate(forecast_years):
            val = impact["llm_impact_pct"][yr]
            cell = ws.cell(row=row, column=2 + i, value=val)
            apply_cell_style(cell, font=DELTA_POS_FONT if val >= 0 else DELTA_NEG_FONT,
                             border=THIN_BORDER, number_format="+0.0%;-0.0%", alignment=Alignment(horizontal="center"))
        row += 1

        # Winners & Losers
        ws.cell(row=row, column=1, value="Winners (Share Gainers)")
        apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, fill=LLM_GREEN_FILL, border=THIN_BORDER)
        ws.cell(row=row, column=2, value=", ".join(impact["winners"]))
        apply_cell_style(ws.cell(row=row, column=2), font=NORMAL_FONT, fill=LLM_GREEN_FILL, border=THIN_BORDER)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        row += 1

        ws.cell(row=row, column=1, value="Losers (Share Donors)")
        apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, fill=LLM_RED_FILL, border=THIN_BORDER)
        ws.cell(row=row, column=2, value=", ".join(impact["losers"]))
        apply_cell_style(ws.cell(row=row, column=2), font=NORMAL_FONT, fill=LLM_RED_FILL, border=THIN_BORDER)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        row += 1

        # Winner/Loser revenue adjustment
        ws.cell(row=row, column=1, value="Winner Revenue Boost")
        apply_cell_style(ws.cell(row=row, column=1), font=NORMAL_FONT, border=THIN_BORDER)
        for i, yr in enumerate(forecast_years):
            val = impact["winner_boost"].get(yr, 0)
            cell = ws.cell(row=row, column=2 + i, value=val)
            apply_cell_style(cell, font=DELTA_POS_FONT, border=THIN_BORDER, number_format="+0.0%", alignment=Alignment(horizontal="center"))
        row += 1

        ws.cell(row=row, column=1, value="Loser Revenue Drag")
        apply_cell_style(ws.cell(row=row, column=1), font=NORMAL_FONT, border=THIN_BORDER)
        for i, yr in enumerate(forecast_years):
            val = impact["loser_drag"].get(yr, 0)
            cell = ws.cell(row=row, column=2 + i, value=val)
            apply_cell_style(cell, font=DELTA_NEG_FONT, border=THIN_BORDER, number_format="+0.0%;-0.0%", alignment=Alignment(horizontal="center"))
        row += 1

        # Assumptions text
        ws.cell(row=row, column=1, value="Key Assumptions")
        apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, border=THIN_BORDER, alignment=Alignment(vertical="top"))
        for ai, assumption in enumerate(impact["assumptions"]):
            ws.cell(row=row + ai, column=2, value=f"{ai+1}. {assumption}")
            apply_cell_style(ws.cell(row=row + ai, column=2), font=Font(name="Calibri", size=10), border=THIN_BORDER,
                             alignment=Alignment(wrap_text=True, vertical="top"))
            ws.merge_cells(start_row=row + ai, start_column=2, end_row=row + ai, end_column=8)
        row += len(impact["assumptions"]) + 1

    ws.column_dimensions["A"].width = 26
    for c in range(2, 9):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.sheet_properties.tabColor = "7B2D8E"


##############################################################################
# LLM IMPACT — ADJUSTED FORECAST TAB (per-segment)
##############################################################################

def write_llm_adjusted_segment_sheet(wb, segment_key):
    data = SEGMENTS[segment_key]
    impact = LLM_IMPACT[segment_key]
    adj_market, adj_companies = compute_llm_adjusted(segment_key)

    title = f"{segment_key} LLM"
    if len(title) > 31:
        title = title[:31]
    ws = wb.create_sheet(title=title)

    ws.cell(row=1, column=1, value=f"{data['full_name']} — LLM-Adjusted Scenario")
    apply_cell_style(ws.cell(row=1, column=1), font=Font(name="Calibri", bold=True, size=14, color="7B2D8E"))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(YEARS) + 2)

    ws.cell(row=2, column=1, value="Shows base-case vs. LLM-adjusted market size and company revenue. Purple-highlighted columns are LLM-adjusted forecasts.")
    apply_cell_style(ws.cell(row=2, column=1), font=Font(name="Calibri", italic=True, size=10, color="666666"))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(YEARS) + 2)

    PURPLE_FILL = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")

    # --- SECTION 1: Market Size Comparison ---
    row = 4
    ws.cell(row=row, column=1, value="Market Size Comparison ($B)")
    apply_cell_style(ws.cell(row=row, column=1), font=HEADER_FONT, fill=LLM_HEADER_FILL, alignment=Alignment(horizontal="left"))
    for i, yr in enumerate(YEARS):
        cell = ws.cell(row=row, column=i + 2, value=yr)
        apply_cell_style(cell, font=HEADER_FONT, fill=LLM_HEADER_FILL, alignment=Alignment(horizontal="center"))
    style_range(ws, row, 1, len(YEARS) + 1, border=THIN_BORDER)

    # Base case
    row = 5
    ws.cell(row=row, column=1, value="Base Case")
    apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, border=THIN_BORDER)
    for i, yr in enumerate(YEARS):
        cell = ws.cell(row=row, column=i + 2, value=data["market_size"][yr])
        apply_cell_style(cell, font=NORMAL_FONT, border=THIN_BORDER, number_format=USD_FORMAT, alignment=Alignment(horizontal="center"))

    # LLM Adjusted
    row = 6
    ws.cell(row=row, column=1, value="LLM-Adjusted")
    apply_cell_style(ws.cell(row=row, column=1), font=Font(name="Calibri", bold=True, color="7B2D8E"), border=THIN_BORDER)
    for i, yr in enumerate(YEARS):
        cell = ws.cell(row=row, column=i + 2, value=adj_market[yr])
        f = PURPLE_FILL if yr > HISTORICAL_END else None
        apply_cell_style(cell, font=Font(name="Calibri", color="7B2D8E"), fill=f, border=THIN_BORDER, number_format=USD_FORMAT, alignment=Alignment(horizontal="center"))

    # Delta
    row = 7
    ws.cell(row=row, column=1, value="Delta ($B)")
    apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, border=THIN_BORDER)
    for i, yr in enumerate(YEARS):
        delta = round(adj_market[yr] - data["market_size"][yr], 2)
        cell = ws.cell(row=row, column=i + 2, value=delta)
        dfont = DELTA_POS_FONT if delta >= 0 else DELTA_NEG_FONT
        apply_cell_style(cell, font=dfont, border=THIN_BORDER, number_format='+#,##0.0;-#,##0.0', alignment=Alignment(horizontal="center"))

    # Delta %
    row = 8
    ws.cell(row=row, column=1, value="Delta (%)")
    apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, border=THIN_BORDER)
    for i, yr in enumerate(YEARS):
        pct = impact["llm_impact_pct"][yr]
        cell = ws.cell(row=row, column=i + 2, value=pct)
        dfont = DELTA_POS_FONT if pct >= 0 else DELTA_NEG_FONT
        apply_cell_style(cell, font=dfont, border=THIN_BORDER, number_format='+0.0%;-0.0%', alignment=Alignment(horizontal="center"))

    # LLM-adjusted YoY growth
    row = 9
    ws.cell(row=row, column=1, value="LLM-Adj YoY Growth (%)")
    apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, border=THIN_BORDER)
    for i, yr in enumerate(YEARS):
        if i == 0:
            ws.cell(row=row, column=i + 2, value="—")
            apply_cell_style(ws.cell(row=row, column=i + 2), font=NORMAL_FONT, border=THIN_BORDER, alignment=Alignment(horizontal="center"))
            continue
        prev = adj_market[YEARS[i - 1]]
        cur = adj_market[yr]
        g = (cur - prev) / prev
        cell = ws.cell(row=row, column=i + 2, value=g)
        gfont = GROWTH_POSITIVE_FONT if g >= 0 else GROWTH_NEGATIVE_FONT
        apply_cell_style(cell, font=gfont, border=THIN_BORDER, number_format=PCT_FORMAT, alignment=Alignment(horizontal="center"))

    # --- SECTION 2: Company Revenue — LLM Adjusted ---
    row = 11
    ws.cell(row=row, column=1, value="Company Revenue — LLM Adjusted ($B)")
    apply_cell_style(ws.cell(row=row, column=1), font=HEADER_FONT, fill=LLM_SUBHEADER_FILL, alignment=Alignment(horizontal="left"))
    for i, yr in enumerate(YEARS):
        cell = ws.cell(row=row, column=i + 2, value=yr)
        apply_cell_style(cell, font=HEADER_FONT, fill=LLM_SUBHEADER_FILL, alignment=Alignment(horizontal="center"))
    style_range(ws, row, 1, len(YEARS) + 1, border=THIN_BORDER)

    company_start_row = row + 1
    companies_list = list(data["companies"].keys())
    for ci, company in enumerate(companies_list):
        row = company_start_row + ci
        is_winner = company in impact["winners"]
        is_loser = company in impact["losers"]
        label = company
        if is_winner:
            label = f"{company}  [WINNER]"
        elif is_loser:
            label = f"{company}  [LOSER]"
        ws.cell(row=row, column=1, value=label)
        if is_winner:
            bg = LLM_GREEN_FILL
        elif is_loser:
            bg = LLM_RED_FILL
        else:
            bg = ACCENT_FILL if ci % 2 == 0 else None
        apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, fill=bg, border=THIN_BORDER)
        for i, yr in enumerate(YEARS):
            cell = ws.cell(row=row, column=i + 2, value=adj_companies[company][yr])
            cfill = bg
            if yr > HISTORICAL_END and not is_winner and not is_loser:
                cfill = PURPLE_FILL
            apply_cell_style(cell, font=NORMAL_FONT, fill=cfill, border=THIN_BORDER, number_format=USD_FORMAT, alignment=Alignment(horizontal="center"))

    # --- SECTION 3: Revenue Delta vs Base (company level) ---
    row = company_start_row + len(companies_list) + 1
    delta_header_row = row
    ws.cell(row=row, column=1, value="Revenue Delta vs Base ($B)")
    apply_cell_style(ws.cell(row=row, column=1), font=HEADER_FONT, fill=LLM_HEADER_FILL, alignment=Alignment(horizontal="left"))
    for i, yr in enumerate(YEARS):
        cell = ws.cell(row=row, column=i + 2, value=yr)
        apply_cell_style(cell, font=HEADER_FONT, fill=LLM_HEADER_FILL, alignment=Alignment(horizontal="center"))
    style_range(ws, row, 1, len(YEARS) + 1, border=THIN_BORDER)

    for ci, company in enumerate(companies_list):
        row = delta_header_row + 1 + ci
        is_winner = company in impact["winners"]
        is_loser = company in impact["losers"]
        ws.cell(row=row, column=1, value=company)
        if is_winner:
            bg = LLM_GREEN_FILL
        elif is_loser:
            bg = LLM_RED_FILL
        else:
            bg = None
        apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, fill=bg, border=THIN_BORDER)
        for i, yr in enumerate(YEARS):
            delta = round(adj_companies[company][yr] - data["companies"][company][yr], 2)
            cell = ws.cell(row=row, column=i + 2, value=delta)
            dfont = DELTA_POS_FONT if delta >= 0 else DELTA_NEG_FONT
            apply_cell_style(cell, font=dfont, fill=bg, border=THIN_BORDER, number_format='+#,##0.00;-#,##0.00', alignment=Alignment(horizontal="center"))

    # --- SECTION 4: LLM-Adjusted Market Share ---
    row = delta_header_row + len(companies_list) + 2
    share_header_row = row
    ws.cell(row=row, column=1, value="LLM-Adjusted Market Share (%)")
    apply_cell_style(ws.cell(row=row, column=1), font=HEADER_FONT, fill=LLM_SUBHEADER_FILL, alignment=Alignment(horizontal="left"))
    for i, yr in enumerate(YEARS):
        cell = ws.cell(row=row, column=i + 2, value=yr)
        apply_cell_style(cell, font=HEADER_FONT, fill=LLM_SUBHEADER_FILL, alignment=Alignment(horizontal="center"))
    style_range(ws, row, 1, len(YEARS) + 1, border=THIN_BORDER)

    for ci, company in enumerate(companies_list):
        row = share_header_row + 1 + ci
        is_winner = company in impact["winners"]
        is_loser = company in impact["losers"]
        ws.cell(row=row, column=1, value=company)
        if is_winner:
            bg = LLM_GREEN_FILL
        elif is_loser:
            bg = LLM_RED_FILL
        else:
            bg = ACCENT_FILL if ci % 2 == 0 else None
        apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, fill=bg, border=THIN_BORDER)
        for i, yr in enumerate(YEARS):
            share = adj_companies[company][yr] / adj_market[yr] if adj_market[yr] else 0
            cell = ws.cell(row=row, column=i + 2, value=share)
            apply_cell_style(cell, font=NORMAL_FONT, fill=bg, border=THIN_BORDER, number_format=PCT_FORMAT, alignment=Alignment(horizontal="center"))

    # Column widths
    ws.column_dimensions["A"].width = 30
    for i in range(len(YEARS)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 12
    ws.sheet_properties.tabColor = "9C27B0"


##############################################################################
# LLM IMPACT — SUMMARY TAB
##############################################################################

def write_llm_summary_sheet(wb):
    ws = wb.create_sheet(title="LLM Summary")

    ws.cell(row=1, column=1, value="LLM-Based Search Impact — Segment-Level Summary")
    apply_cell_style(ws.cell(row=1, column=1), font=Font(name="Calibri", bold=True, size=16, color="7B2D8E"))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(YEARS) + 3)

    ws.cell(row=2, column=1, value=(
        "Base-case vs. LLM-adjusted total market size by segment. "
        "Delta shows incremental TAM created (or destroyed) by LLM-based search, AI copilots, and autonomous agents."
    ))
    apply_cell_style(ws.cell(row=2, column=1), font=Font(name="Calibri", italic=True, size=10, color="666666"))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(YEARS) + 3)

    seg_order = ["SIEM", "Vulnerability Assessment", "Identity (IAM)", "Endpoint Security", "Network Security", "SASE"]
    PURPLE_FILL = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")

    # ---- BASE CASE ----
    row = 4
    ws.cell(row=row, column=1, value="Base Case ($B)")
    apply_cell_style(ws.cell(row=row, column=1), font=HEADER_FONT, fill=HEADER_FILL, alignment=Alignment(horizontal="left"))
    cagr_col = len(YEARS) + 2
    for i, yr in enumerate(YEARS):
        cell = ws.cell(row=row, column=i + 2, value=yr)
        apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL, alignment=Alignment(horizontal="center"))
    ws.cell(row=row, column=cagr_col, value="CAGR\n'24-'30")
    apply_cell_style(ws.cell(row=row, column=cagr_col), font=HEADER_FONT, fill=HEADER_FILL, alignment=Alignment(horizontal="center", wrap_text=True))
    style_range(ws, row, 1, cagr_col, border=THIN_BORDER)

    base_total = {yr: 0 for yr in YEARS}
    for si, seg in enumerate(seg_order):
        row = 5 + si
        d = SEGMENTS[seg]
        f = ACCENT_FILL if si % 2 == 0 else None
        ws.cell(row=row, column=1, value=seg)
        apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, fill=f, border=THIN_BORDER)
        for i, yr in enumerate(YEARS):
            val = d["market_size"][yr]
            base_total[yr] += val
            cell = ws.cell(row=row, column=i + 2, value=val)
            apply_cell_style(cell, font=NORMAL_FONT, fill=f, border=THIN_BORDER, number_format=USD_FORMAT, alignment=Alignment(horizontal="center"))
        cagr = (d["market_size"][2030] / d["market_size"][2024]) ** (1 / 6) - 1
        cell = ws.cell(row=row, column=cagr_col, value=cagr)
        apply_cell_style(cell, font=BOLD_FONT, fill=f, border=THIN_BORDER, number_format=PCT_FORMAT, alignment=Alignment(horizontal="center"))

    row = 5 + len(seg_order)
    ws.cell(row=row, column=1, value="TOTAL (Base)")
    dark_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    white_bold = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    apply_cell_style(ws.cell(row=row, column=1), font=white_bold, fill=dark_fill, border=THIN_BORDER)
    for i, yr in enumerate(YEARS):
        cell = ws.cell(row=row, column=i + 2, value=round(base_total[yr], 1))
        apply_cell_style(cell, font=white_bold, fill=dark_fill, border=THIN_BORDER, number_format=USD_FORMAT, alignment=Alignment(horizontal="center"))
    base_cagr = (base_total[2030] / base_total[2024]) ** (1 / 6) - 1
    cell = ws.cell(row=row, column=cagr_col, value=base_cagr)
    apply_cell_style(cell, font=white_bold, fill=dark_fill, border=THIN_BORDER, number_format=PCT_FORMAT, alignment=Alignment(horizontal="center"))

    # ---- LLM ADJUSTED ----
    row += 2
    adj_section_start = row
    ws.cell(row=row, column=1, value="LLM-Adjusted ($B)")
    apply_cell_style(ws.cell(row=row, column=1), font=HEADER_FONT, fill=LLM_HEADER_FILL, alignment=Alignment(horizontal="left"))
    for i, yr in enumerate(YEARS):
        cell = ws.cell(row=row, column=i + 2, value=yr)
        apply_cell_style(cell, font=HEADER_FONT, fill=LLM_HEADER_FILL, alignment=Alignment(horizontal="center"))
    ws.cell(row=row, column=cagr_col, value="CAGR\n'24-'30")
    apply_cell_style(ws.cell(row=row, column=cagr_col), font=HEADER_FONT, fill=LLM_HEADER_FILL, alignment=Alignment(horizontal="center", wrap_text=True))
    style_range(ws, row, 1, cagr_col, border=THIN_BORDER)

    adj_total = {yr: 0 for yr in YEARS}
    for si, seg in enumerate(seg_order):
        row = adj_section_start + 1 + si
        adj_mkt, _ = compute_llm_adjusted(seg)
        f = ACCENT_FILL if si % 2 == 0 else None
        ws.cell(row=row, column=1, value=seg)
        apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, fill=f, border=THIN_BORDER)
        for i, yr in enumerate(YEARS):
            val = adj_mkt[yr]
            adj_total[yr] += val
            cfill = PURPLE_FILL if yr > HISTORICAL_END else f
            cell = ws.cell(row=row, column=i + 2, value=val)
            apply_cell_style(cell, font=NORMAL_FONT, fill=cfill, border=THIN_BORDER, number_format=USD_FORMAT, alignment=Alignment(horizontal="center"))
        cagr = (adj_mkt[2030] / adj_mkt[2024]) ** (1 / 6) - 1
        cell = ws.cell(row=row, column=cagr_col, value=cagr)
        apply_cell_style(cell, font=BOLD_FONT, fill=f, border=THIN_BORDER, number_format=PCT_FORMAT, alignment=Alignment(horizontal="center"))

    row = adj_section_start + 1 + len(seg_order)
    purple_dark = PatternFill(start_color="4A148C", end_color="4A148C", fill_type="solid")
    ws.cell(row=row, column=1, value="TOTAL (LLM-Adj)")
    apply_cell_style(ws.cell(row=row, column=1), font=white_bold, fill=purple_dark, border=THIN_BORDER)
    for i, yr in enumerate(YEARS):
        cell = ws.cell(row=row, column=i + 2, value=round(adj_total[yr], 1))
        apply_cell_style(cell, font=white_bold, fill=purple_dark, border=THIN_BORDER, number_format=USD_FORMAT, alignment=Alignment(horizontal="center"))
    adj_cagr = (adj_total[2030] / adj_total[2024]) ** (1 / 6) - 1
    cell = ws.cell(row=row, column=cagr_col, value=adj_cagr)
    apply_cell_style(cell, font=white_bold, fill=purple_dark, border=THIN_BORDER, number_format=PCT_FORMAT, alignment=Alignment(horizontal="center"))

    # ---- DELTA ----
    row += 2
    delta_section_start = row
    ws.cell(row=row, column=1, value="Delta: LLM Uplift ($B)")
    apply_cell_style(ws.cell(row=row, column=1), font=HEADER_FONT, fill=LLM_SUBHEADER_FILL, alignment=Alignment(horizontal="left"))
    for i, yr in enumerate(YEARS):
        cell = ws.cell(row=row, column=i + 2, value=yr)
        apply_cell_style(cell, font=HEADER_FONT, fill=LLM_SUBHEADER_FILL, alignment=Alignment(horizontal="center"))
    ws.cell(row=row, column=cagr_col, value="2030E\nUplift %")
    apply_cell_style(ws.cell(row=row, column=cagr_col), font=HEADER_FONT, fill=LLM_SUBHEADER_FILL, alignment=Alignment(horizontal="center", wrap_text=True))
    style_range(ws, row, 1, cagr_col, border=THIN_BORDER)

    for si, seg in enumerate(seg_order):
        row = delta_section_start + 1 + si
        adj_mkt, _ = compute_llm_adjusted(seg)
        base_mkt = SEGMENTS[seg]["market_size"]
        f = ACCENT_FILL if si % 2 == 0 else None
        ws.cell(row=row, column=1, value=seg)
        apply_cell_style(ws.cell(row=row, column=1), font=BOLD_FONT, fill=f, border=THIN_BORDER)
        for i, yr in enumerate(YEARS):
            delta = round(adj_mkt[yr] - base_mkt[yr], 2)
            cell = ws.cell(row=row, column=i + 2, value=delta)
            dfont = DELTA_POS_FONT if delta >= 0 else DELTA_NEG_FONT
            apply_cell_style(cell, font=dfont, fill=f, border=THIN_BORDER, number_format='+#,##0.0;-#,##0.0', alignment=Alignment(horizontal="center"))
        uplift_pct = (adj_mkt[2030] - base_mkt[2030]) / base_mkt[2030]
        cell = ws.cell(row=row, column=cagr_col, value=uplift_pct)
        apply_cell_style(cell, font=DELTA_POS_FONT if uplift_pct >= 0 else DELTA_NEG_FONT,
                         fill=f, border=THIN_BORDER, number_format='+0.0%;-0.0%', alignment=Alignment(horizontal="center"))

    row = delta_section_start + 1 + len(seg_order)
    ws.cell(row=row, column=1, value="TOTAL Delta")
    apply_cell_style(ws.cell(row=row, column=1), font=white_bold, fill=purple_dark, border=THIN_BORDER)
    for i, yr in enumerate(YEARS):
        delta = round(adj_total[yr] - base_total[yr], 1)
        cell = ws.cell(row=row, column=i + 2, value=delta)
        apply_cell_style(cell, font=white_bold, fill=purple_dark, border=THIN_BORDER, number_format='+#,##0.0;-#,##0.0', alignment=Alignment(horizontal="center"))
    total_uplift = (adj_total[2030] - base_total[2030]) / base_total[2030]
    cell = ws.cell(row=row, column=cagr_col, value=total_uplift)
    apply_cell_style(cell, font=white_bold, fill=purple_dark, border=THIN_BORDER, number_format='+0.0%;-0.0%', alignment=Alignment(horizontal="center"))

    # ---- CHART ----
    chart_row = row + 3
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "2030E Market Size: Base vs. LLM-Adjusted ($B)"
    chart.y_axis.title = "USD Billions"
    chart.style = 10
    chart.width = 28
    chart.height = 14

    yr_2030_col = YEARS.index(2030) + 2
    base_vals = Reference(ws, min_col=yr_2030_col, max_col=yr_2030_col, min_row=5, max_row=5 + len(seg_order) - 1)
    adj_vals = Reference(ws, min_col=yr_2030_col, max_col=yr_2030_col, min_row=adj_section_start + 1, max_row=adj_section_start + len(seg_order))
    cats = Reference(ws, min_col=1, min_row=5, max_row=5 + len(seg_order) - 1)

    chart.add_data(base_vals, titles_from_data=False)
    chart.add_data(adj_vals, titles_from_data=False)
    chart.set_categories(cats)
    chart.series[0].title = openpyxl.chart.series.SeriesLabel(v="Base Case 2030E")
    chart.series[0].graphicalProperties.solidFill = "4472C4"
    chart.series[1].title = openpyxl.chart.series.SeriesLabel(v="LLM-Adjusted 2030E")
    chart.series[1].graphicalProperties.solidFill = "9C27B0"

    ws.add_chart(chart, f"A{chart_row}")

    # Column widths
    ws.column_dimensions["A"].width = 28
    for i in range(len(YEARS)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 12
    ws.column_dimensions[get_column_letter(cagr_col)].width = 12
    ws.sheet_properties.tabColor = "7B2D8E"


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_summary_sheet(wb)

    seg_order = ["SIEM", "Vulnerability Assessment", "Identity (IAM)", "Endpoint Security", "Network Security", "SASE"]
    for seg_key in seg_order:
        write_segment_sheet(wb, seg_key, SEGMENTS[seg_key])

    write_llm_assumptions_sheet(wb)
    write_llm_summary_sheet(wb)
    for seg_key in seg_order:
        write_llm_adjusted_segment_sheet(wb, seg_key)

    output = "cybersecurity_segments_forecast.xlsx"
    wb.save(output)
    print(f"Workbook saved to {output}")


if __name__ == "__main__":
    main()
