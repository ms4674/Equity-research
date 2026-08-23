#!/usr/bin/env python3
"""Build the agentic AI revenue & token usage workbook and CSV mirrors.

Regenerate with:  python3 scripts/build_spreadsheet.py
Outputs:
  - agentic_ai_revenue_token_usage_2026.xlsx
  - data/offerings_revenue.csv
  - data/sector_aggregates.csv
  - data/token_usage_metrics.csv
  - data/sources.csv
"""

import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

AS_OF = "August 23, 2026"

# ---------------------------------------------------------------------------
# Sources
# ---------------------------------------------------------------------------
SOURCES = [
    (1, "Axis Intelligence", "Cursor AI Statistics 2026: Revenue, Users, Valuation & Market Data", "https://axis-intelligence.com/cursor-ai-statistics/"),
    (2, "Value Add VC", "How Does Cursor Make Money: Subscriptions, Token Pricing, Business Model", "https://valueaddvc.com/blog/how-does-cursor-make-money-subscriptions-token-pricing-and-the-business-model-breakdown"),
    (3, "TickerTrends", "Claude Code Reaches $15.1B in Tracked ARR as Codex Climbs to $8.8B (Aug 10, 2026)", "https://blog.tickertrends.io/p/claude-code-vs-openai-codex-arr"),
    (4, "Enterprise DNA", "Claude Code Hits $2.5B ARR in Just 13 Months (Feb 2026 Series G disclosure)", "https://enterprisedna.co/resources/news/claude-code-25-billion-arr-anthropic-2026/"),
    (5, "FourWeekMBA", "Claude Code Scaled from $1B to $8B Annualized Run-Rate in Six Months (Sacra data)", "https://fourweekmba.com/ai-claude-code-run-rate-agentic-layer/"),
    (6, "Forkast / Bloomberg", "Anthropic Q2 2026 Revenue ($11.5B) Overtook OpenAI; $65B run-rate end of July 2026", "https://forkast.news/anthropics-q2-revenue-overtook-openai-for-the-first-time-and-reached-its-first-positive-operating-income/"),
    (7, "Axis Intelligence", "GitHub Copilot Statistics 2026: Users, Revenue, Market Share", "https://axis-intelligence.com/github-copilot-statistics/"),
    (8, "Value Add VC", "How Does Cognition Make Money: Devin Pricing, Windsurf, and $492M ARR", "https://valueaddvc.com/blog/how-does-cognition-make-money-devin-pricing-windsurf-enterprise-and-the-492m-arr-breakdown"),
    (9, "Sacra", "Sierra revenue, valuation & funding ($200M ARR May 2026)", "https://sacra.com/c/sierra/"),
    (10, "Value Add VC", "How Does Sierra AI Make Money: Outcome-Based Pricing ($15.8B valuation)", "https://valueaddvc.com/blog/how-does-sierra-ai-make-money-outcome-based-pricing-enterprise-agents-and-the-business-model-breakdown"),
    (11, "AgentMarketCap", "Vertical Agent Revenue Ranked (Apr 2026): Harvey, Agentforce, Sierra, Genspark", "https://agentmarketcap.ai/blog/2026/04/05/vertical-agent-revenue-ranked-harvey-salesforce-iqvia-domain-specific-agents"),
    (12, "AgentMarketCap", "The Verticalization Premium: Harvey and Sierra valuation multiples (Apr 2026)", "https://agentmarketcap.ai/blog/2026/04/07/vertical-vs-horizontal-ai-agent-valuations-2026"),
    (13, "Sacra", "Abridge revenue, valuation & funding ($100M ARR May 2025)", "https://sacra.com/c/abridge/"),
    (14, "ARR Club", "OpenEvidence ARR, Revenue Growth & Milestones ($300M ARR Jul 2026)", "https://www.arr.club/openevidence"),
    (15, "Information Matters", "Agentic AI in Healthcare 2026 (sector sizing $4.3-7.1B; central ~$5.5B)", "https://informationmatters.net/wp-content/uploads/2026/05/agentic-ai-healthcare-2026-v13.pdf"),
    (16, "Value Add VC", "Best AI Tools for Healthcare in 2026 (Abridge, Hippocratic, Ambience pricing)", "https://valueaddvc.com/blog/best-ai-tools-for-healthcare-in-2026-options-ranked-by-pricing-and-compliance"),
    (17, "CX Today", "Agentforce Crosses $1B ARR; 28.6T tokens processed in Q1 FY2027", "https://www.cxtoday.com/contact-center/agentforce-hits-1-billion-arr-ai-agents-customer-service/"),
    (18, "Customer Experience Magazine", "Salesforce Q1 FY2027: Agentforce ARR $1.2B, +205% YoY", "https://cxm.world/customer-experience/salesforces-ai-agents-now-handle-double-the-volume-of-its-human-ones/"),
    (19, "TURION.AI", "Enterprise Agent Platforms: Salesforce vs ServiceNow vs Microsoft (June 2026 pricing)", "https://turion.ai/blog/enterprise-agent-platforms-comparison-june-2026/"),
    (20, "AgentMarketCap", "Salesforce Agentforce vs Microsoft Copilot: Enterprise Agent Platform War (Apr 2026)", "https://agentmarketcap.ai/blog/2026/04/13/salesforce-agentforce-vs-microsoft-copilot-enterprise-agent-platform-war-2026"),
    (21, "ChatForest", "Google Is Processing 3.2 Quadrillion Tokens a Month (I/O 2026)", "https://chatforest.com/builders-log/google-3-2-quadrillion-tokens-ai-scale-economics/"),
    (22, "Trending Topics", "Open-Weight Models from China Are Capturing a Growing Share of AI Usage (router token volumes)", "https://www.trendingtopics.eu/open-weight-models-from-china-are-capturing-a-growing-share-of-ai-usage/"),
    (23, "Beth Kindig / I/O Fund", "AI Token Demand Is Shattering Forecasts (~370T tokens/day industry-wide)", "https://beth-kindig.medium.com/ai-token-demand-is-shattering-forecasts-ec8831df6c99"),
    (24, "Vercel", "DeepSeek Overtakes Google on Volume; Cost per Token Falls (AI Gateway, July 2026)", "https://vercel.com/blog/deepseek-overtakes-google-on-volume-cost-per-token-falls"),
    (25, "OpenRouter", "DeepSeek V4 Is Earning Agentic Token Share (agentic ~15x tokens per request)", "https://openrouter.ai/blog/insights/deepseek-v4-adoption/"),
    (26, "OpenRouter", "State of AI 2025: 100T Token LLM Usage Study", "https://openrouter.ai/state-of-ai"),
    (27, "OpenRouter", "App & Agent Rankings (token volume by app)", "https://openrouter.ai/apps"),
    (28, "Fabryka", "OpenRouter Apps - Demand Research (app-level token share)", "https://ort.fabryka.ai/apps.html"),
    (29, "Bloomberg", "OpenAI's Revenue Run Rate Tops $40 Billion Ahead of IPO (Aug 13, 2026)", "https://www.bloomberg.com/news/articles/2026-08-13/openai-s-revenue-run-rate-tops-40-billion-ahead-of-ipo"),
    (30, "Sacra", "OpenAI revenue, valuation & funding ($40B annualized July 2026)", "https://sacra.com/c/openai/"),
    (31, "CNBC", "OpenAI CFO Friar: enterprise now majority of revenue (Aug 14, 2026)", "https://www.cnbc.com/2026/08/14/openai-cfo-friar-tells-investors-that-enterprise-bigger-than-consumer.html"),
    (32, "Karsane", "OpenAI Says Agents Top 40% of Enterprise Revenue (~$25B annualized, Feb 2026)", "https://karsane.com/article/openai-commercial-expansion-agents-40-percent-revenue"),
    (33, "Contrary Research", "Rogo Business Breakdown ($2M 2024 -> $15M+ 2025; 50%+ QoQ growth mid-2026)", "https://research.contrary.com/company/rogo"),
    (34, "THE DAILY BRIEF (beri.net)", "Rogo's $160M Series D at ~$2B; Hebbia and AlphaSense comparables (Apr 2026)", "https://www.beri.net/article/rogo-160m-series-d-investment-banking-ai-agents-2026"),
    (35, "TechCrunch", "Glean's top line crosses $300M ARR (May 28, 2026)", "https://techcrunch.com/2026/05/28/gleans-top-line-crosses-300m-as-ai-budget-cutting-becomes-its-major-selling-point/"),
    (36, "Microsoft", "FY2026 Q4 Earnings Call (30M+ M365 Copilot seats; Foundry token run-rates; GitHub Copilot)", "https://www.microsoft.com/en-us/investor/events/fy-2026/earnings-fy-2026-q4"),
    (37, "LicenseQ", "Microsoft FY26 Q4 Results Explained (Copilot seats, Agent 365, Foundry)", "https://licenseq.com/microsoft-fy26-q4-results-explained/"),
    (38, "Presenc AI", "Medical AI Tools Landscape 2026 (ambient scribes ~$600M 2025 revenue)", "https://presenc.ai/research/medical-ai-tools-landscape-2026"),
    (39, "AI Funding Tracker", "Top AI Agent Startups 2026 (Funding & Valuation)", "https://aifundingtracker.com/top-ai-agent-startups/"),
    (40, "THE DAILY BRIEF (beri.net)", "Why SpaceX Paid $60B for Cursor: $4B ARR (June 2026)", "https://www.beri.net/article/spacex-cursor-60b-acquisition-2026"),
]

# ---------------------------------------------------------------------------
# Offerings: one row per agentic offering
# Fields: sector, company, offering, description, revenue_basis, revenue_musd,
#         revenue_asof, trajectory, valuation_busd, valuation_note, pricing,
#         scale_metric, token_metric, confidence, in_aggregate, sources
# ---------------------------------------------------------------------------
OFFERINGS = [
    # --- Coding & software development ---
    dict(sector="Coding & software development", company="Anthropic", offering="Claude Code",
         description="Agentic CLI coding agent: plans, edits, tests and iterates across a codebase",
         revenue_basis="Tracked ARR (3rd-party estimate)", revenue_musd=15120, revenue_asof="Aug 10, 2026",
         trajectory="$1B Nov 2025 -> $2.5B Feb 2026 (company-disclosed, Series G) -> ~$8B May 2026 (Sacra) -> $15.1B Aug 2026 (TickerTrends)",
         valuation_busd=380, valuation_note="Anthropic post-money, Feb 2026 Series G ($30B raise)",
         pricing="Subscription (Pro/Max) + API usage",
         scale_metric="~21.9% of Anthropic tracked ARR; avg developer ~20 hrs/week in tool; ~4% of global GitHub commits (SemiAnalysis)",
         token_metric="6.5-9.6T tokens/month routed via OpenRouter alone (app rankings snapshots)",
         confidence="Tracked estimate (disclosed $2.5B Feb 2026)", in_aggregate="Yes", sources="3, 4, 5, 27, 28"),
    dict(sector="Coding & software development", company="OpenAI", offering="Codex",
         description="Agentic coding agent (CLI/IDE/cloud) for autonomous software tasks",
         revenue_basis="Tracked ARR (3rd-party estimate)", revenue_musd=8830, revenue_asof="Aug 10, 2026",
         trajectory="Near zero start of 2026 -> $5.9B Jul 6, 2026 -> $8.8B Aug 10, 2026; ~44% of OpenAI enterprise ARR (central case)",
         valuation_busd=None, valuation_note="Part of OpenAI (IPO planned)",
         pricing="Bundled with ChatGPT plans + API usage",
         scale_metric="5M+ weekly active users (Feb 2026); 20M WAU across OpenAI coding/work products (Aug 2026)",
         token_metric="~688B tokens/month via OpenRouter (87% on OpenAI models)",
         confidence="Tracked estimate", in_aggregate="Yes", sources="3, 30, 32, 28"),
    dict(sector="Coding & software development", company="Anysphere (SpaceX acq. pending)", offering="Cursor",
         description="AI-native code editor with agentic coding workflows",
         revenue_basis="ARR", revenue_musd=4000, revenue_asof="May-Jun 2026",
         trajectory="$100M Jan 2025 -> $500M Jun 2025 -> $1B Nov 2025 -> $2B Feb 2026 (Bloomberg) -> $4B May/Jun 2026 (Sacra)",
         valuation_busd=60, valuation_note="SpaceX all-stock acquisition agreed Jun 16, 2026 (close Q3 2026); last private mark $29.3B Series D Nov 2025",
         pricing="$20-$200/mo subscriptions + enterprise seats; ~20% margin on model API costs; ~65% enterprise (~$2.6B)",
         scale_metric="1M+ paying customers; used in 64% of Fortune 500",
         token_metric=None,
         confidence="Company-disclosed to $2B; analyst (Sacra) at $4B", in_aggregate="Yes", sources="1, 2, 40"),
    dict(sector="Coding & software development", company="Microsoft (GitHub)", offering="GitHub Copilot",
         description="AI pair programmer evolving into agentic coding (agent mode, usage-based billing)",
         revenue_basis="ARR (analyst estimate)", revenue_musd=1000, revenue_asof="Q2 FY26 (Jan 2026)",
         trajectory="Estimate range $900M-$1.1B (Axis); FT floor 'at least $550M'; Copilot revenue +60% QoQ in Q4 FY26 after usage-based billing launch",
         valuation_busd=None, valuation_note="Microsoft subsidiary (not separately valued)",
         pricing="$10-$39/user/mo tiers + usage-based billing (introduced Q4 FY26)",
         scale_metric="4.7M paid subscribers (Jan 2026, +75% YoY); 50M users (Jul 2026); 90% of Fortune 100",
         token_metric=None,
         confidence="Analyst estimate (Microsoft does not disclose)", in_aggregate="Yes", sources="7, 36, 37"),
    dict(sector="Coding & software development", company="Replit", offering="Replit Agent",
         description="Cloud IDE with autonomous app-building agent",
         revenue_basis="ARR (analyst estimate)", revenue_musd=525, revenue_asof="Mid-2026",
         trajectory="Targeting $1B ARR by end of 2026",
         valuation_busd=9, valuation_note="2026 private round",
         pricing="Effort-based agent pricing (usage) + subscriptions",
         scale_metric=None, token_metric=None,
         confidence="Analyst estimate", in_aggregate="Yes", sources="8"),
    dict(sector="Coding & software development", company="Cognition", offering="Devin + Windsurf",
         description="Autonomous software engineer (Devin) + agentic IDE (Windsurf, acquired 2025)",
         revenue_basis="Annualized revenue (Sacra estimate)", revenue_musd=492, revenue_asof="May 2026",
         trajectory="$37M (2025) -> $492M May 2026 (~13x); targeting $1B by year-end",
         valuation_busd=10.2, valuation_note="2026 private mark",
         pricing="ACU (agent compute unit) credits from $20 + multi-seat enterprise licenses",
         scale_metric=None, token_metric=None,
         confidence="Analyst estimate (Sacra)", in_aggregate="Yes", sources="8, 11"),
    dict(sector="Coding & software development", company="Lovable", offering="Lovable",
         description="Vibe-coding / text-to-app agent platform",
         revenue_basis="ARR (reported)", revenue_musd=400, revenue_asof="Feb 2026",
         trajectory="$100M ARR within ~12 months of launch -> $400M Feb 2026",
         valuation_busd=None, valuation_note=None,
         pricing="Subscription + usage",
         scale_metric=None, token_metric=None,
         confidence="Press-reported", in_aggregate="Yes", sources="8, 39"),

    # --- Customer experience & support ---
    dict(sector="Customer experience & support", company="Sierra", offering="Sierra Agents",
         description="Enterprise customer-service agents across chat, voice, email (Bret Taylor / Clay Bavor)",
         revenue_basis="ARR (Sacra estimate)", revenue_musd=200, revenue_asof="May 2026",
         trajectory="$26M end-2024 -> $100M Nov 2025 (21 months from launch) -> $150M Feb 2026 -> $200M May 2026",
         valuation_busd=15.8, valuation_note="May 2026 Series E ($950M)",
         pricing="Outcome-based (~$1.50 per resolved interaction) + volume-based for routine flows",
         scale_metric="Customers: SoFi, Ramp, Brex, Discord, Rivian, ADT, Cigna, SiriusXM; 50%+ of customers >$1B revenue",
         token_metric=None,
         confidence="Analyst estimate (Sacra)", in_aggregate="Yes", sources="9, 10, 39"),
    dict(sector="Customer experience & support", company="Decagon", offering="Decagon AI Agents",
         description="AI customer-support agents for internet-native and consumer businesses",
         revenue_basis="ARR (reported)", revenue_musd=100, revenue_asof="2026 (crossed $100M)",
         trajectory="~$6M ARR late 2024 -> $100M+ 2026",
         valuation_busd=None, valuation_note=None,
         pricing="Per-conversation pricing",
         scale_metric="Customers: Duolingo, Notion, Webflow",
         token_metric=None,
         confidence="Press-reported", in_aggregate="Yes", sources="9, 12, 39"),
    dict(sector="Customer experience & support", company="Wonderful AI", offering="Wonderful",
         description="Customer-service agents focused on non-English markets",
         revenue_basis="Not disclosed", revenue_musd=None, revenue_asof=None,
         trajectory="Revenue 'growing'; $286M raised in 13 months",
         valuation_busd=2, valuation_note="2026 private mark",
         pricing=None, scale_metric=None, token_metric=None,
         confidence="Not disclosed", in_aggregate="No (revenue n/d)", sources="11"),

    # --- Legal ---
    dict(sector="Legal", company="Harvey", offering="Harvey",
         description="Legal AI agents: research, contract review, document analysis, drafting",
         revenue_basis="ARR (reported)", revenue_musd=300, revenue_asof="Mid-2026",
         trajectory="$50M end-2024 -> $190M Jan 2026 (~4x YoY) -> ~$300M mid-2026",
         valuation_busd=11, valuation_note="Mar 2026 raise ($200M, GIC/Sequoia); up from $8B late 2025",
         pricing="Per-seat + enterprise contracts (7-figure ACVs)",
         scale_metric="100K+ lawyers; 1,300+ organizations",
         token_metric=None,
         confidence="Press-reported", in_aggregate="Yes", sources="11, 12, 39"),

    # --- Healthcare ---
    dict(sector="Healthcare", company="OpenEvidence", offering="OpenEvidence",
         description="AI clinical decision support / evidence search for clinicians",
         revenue_basis="ARR (reported)", revenue_musd=300, revenue_asof="Jul 21, 2026",
         trajectory="$50M 2025 -> $150M ~Dec 2025 -> $300M Jul 2026 (7.1 months to double)",
         valuation_busd=12, valuation_note="Jan 2026 Series D ($250M)",
         pricing="Free to clinicians; monetization via sponsors/enterprise",
         scale_metric="860K clinician users; used by ~40% of US physicians; 16.5M clinical consultations/month (2025)",
         token_metric=None,
         confidence="Press-reported", in_aggregate="Yes", sources="14"),
    dict(sector="Healthcare", company="Abridge", offering="Abridge",
         description="Ambient clinical documentation (AI scribe) with deep Epic integration",
         revenue_basis="ARR (Sacra estimate)", revenue_musd=100, revenue_asof="May 2025",
         trajectory="$60M end-2024 -> $100M May 2025; contracted ARR $117M Q1 2025",
         valuation_busd=5.3, valuation_note="Jun 2025 Series E ($300M, a16z) + $316M extension Apr 2026",
         pricing="~$208/mo (~$2,500/clinician/yr enterprise)",
         scale_metric="~30% ambient-scribe market share; 250+ health systems (Kaiser, Mayo, Hopkins, Duke)",
         token_metric=None,
         confidence="Analyst estimate (Sacra)", in_aggregate="Yes", sources="13, 15, 16, 38"),
    dict(sector="Healthcare", company="Hippocratic AI", offering="Hippocratic Agents",
         description="Patient-facing healthcare agents (pre-op, discharge, chronic-care outreach)",
         revenue_basis="Annual revenue (reported)", revenue_musd=76, revenue_asof="2026",
         trajectory="115M+ patient interactions completed",
         valuation_busd=3.5, valuation_note="Late 2025 Series C ($126M)",
         pricing="Custom enterprise (per-agent-hour model)",
         scale_metric="50+ health-system, payor and pharma partners across 6 countries",
         token_metric=None,
         confidence="Press-reported", in_aggregate="Yes", sources="16, 39"),
    dict(sector="Healthcare", company="Ambience Healthcare", offering="Ambience",
         description="Ambient documentation + coding for multi-specialty health systems",
         revenue_basis="ARR (analyst estimate)", revenue_musd=30, revenue_asof="May 2025",
         trajectory="Unicorn ($1.25B) in 2025",
         valuation_busd=1.25, valuation_note="2025 round",
         pricing="Custom, est. $250+/clinician/mo",
         scale_metric=None, token_metric=None,
         confidence="Analyst estimate", in_aggregate="Yes", sources="15, 16, 38"),

    # --- Financial services ---
    dict(sector="Financial services", company="Rogo", offering="Rogo (Felix)",
         description="Agentic AI analyst for investment banking: models, comps, memos, pitchbooks, CIMs",
         revenue_basis="ARR (analyst-inferred range $50-100M; midpoint shown)", revenue_musd=75, revenue_asof="Mid-2026",
         trajectory="$2M 2024 -> $15M+ 2025 (disclosed); >50% QoQ ARR growth entering mid-2026",
         valuation_busd=2, valuation_note="Apr 2026 Series D ($160M, Kleiner Perkins)",
         pricing="Enterprise seats (single-tenant deployments)",
         scale_metric="35K+ professionals; 300+ institutions (Lazard, Jefferies, Moelis, Rothschild, Nomura, Baird); 50K+ queries/day",
         token_metric=None,
         confidence="Analyst-inferred (ARR not disclosed)", in_aggregate="Yes", sources="33, 34"),
    dict(sector="Financial services", company="Hebbia", offering="Hebbia Matrix",
         description="Multi-document agentic analysis for finance, legal, PE diligence",
         revenue_basis="ARR (reported)", revenue_musd=24.6, revenue_asof="Late 2025",
         trajectory=None,
         valuation_busd=0.7, valuation_note="2024 Series B (a16z); $161M total funding",
         pricing="Enterprise contracts",
         scale_metric=None, token_metric=None,
         confidence="Press-reported", in_aggregate="Yes", sources="34"),
    dict(sector="Financial services", company="AlphaSense", offering="AlphaSense (agentic research)",
         description="Market-intelligence search adding agentic research workflows (partially agentic)",
         revenue_basis="ARR (reported)", revenue_musd=200, revenue_asof="2026",
         trajectory=None,
         valuation_busd=4, valuation_note="2026 private mark",
         pricing="Enterprise subscriptions",
         scale_metric="10,000+ content sources",
         token_metric=None,
         confidence="Press-reported; partially agentic (flagged)", in_aggregate="Yes", sources="34"),

    # --- Enterprise platform agents ---
    dict(sector="Enterprise platform agents", company="Salesforce", offering="Agentforce",
         description="CRM-native agent platform (service, sales, IT) with Atlas reasoning engine",
         revenue_basis="ARR (company-disclosed)", revenue_musd=1200, revenue_asof="Q1 FY2027 (qtr ended Apr 2026)",
         trajectory="$800M Q4 FY2026 (+169% YoY) -> $1.2B Q1 FY2027 (+205% YoY); total AI+Data ARR $3.4B",
         valuation_busd=None, valuation_note="Public (NYSE: CRM)",
         pricing="$2/conversation + Flex Credits ($500 per 100K credits)",
         scale_metric="18,500+ customers (9,500+ paid); 4M autonomous inquiries on Salesforce's own support in 15 months",
         token_metric="28.6T tokens processed in Q1 FY2027 (+152% QoQ); 3.8B agentic work units delivered",
         confidence="Company-disclosed", in_aggregate="Yes", sources="17, 18, 19, 20"),
    dict(sector="Enterprise platform agents", company="ServiceNow", offering="Now Assist / AI Agents",
         description="Workflow-native agents for ITSM/ESM (Agent Fabric, AI Control Tower)",
         revenue_basis="ACV (reported)", revenue_musd=1000, revenue_asof="Mid-2026",
         trajectory="AI portfolio anchored by Now Assist reached $1B ACV",
         valuation_busd=None, valuation_note="Public (NYSE: NOW)",
         pricing="Bundled in Pro Plus / Enterprise Plus tiers (20-40% uplift at renewal) + message packs",
         scale_metric="Gartner #1 for building/managing AI agents (2025)",
         token_metric=None,
         confidence="Press-reported", in_aggregate="Yes", sources="19, 20"),
    dict(sector="Enterprise platform agents", company="Microsoft", offering="M365 Copilot + Agent 365 + Copilot Studio",
         description="Productivity copilot evolving to autonomous 'Autopilots'; agent governance via Agent 365",
         revenue_basis="Not disclosed (excluded from aggregates)", revenue_musd=None, revenue_asof="Q4 FY26 (Jun 2026)",
         trajectory="Seats: 15M Q2 -> 20M Q3 -> 30M+ Q4 FY26; at $30/user/mo list, implies up to ~$10.8B annualized before discounts (not disclosed)",
         valuation_busd=None, valuation_note="Public (NASDAQ: MSFT)",
         pricing="$30-$99/user/mo seats; Copilot Studio $200/mo per 25K credits or $0.01/credit PAYG",
         scale_metric="30M+ paid M365 Copilot seats; Agent 365: ~40M registered agents within 2 months of launch",
         token_metric="Azure AI Foundry: 100K customers; customers at 1T-token annualized run-rate up 4x YoY",
         confidence="Seats disclosed; revenue not disclosed", in_aggregate="No (revenue n/d)", sources="36, 37, 19, 20"),
    dict(sector="Enterprise platform agents", company="Glean", offering="Glean Assistant + Agents",
         description="Enterprise knowledge graph powering search and agentic workflows",
         revenue_basis="ARR / annualized run-rate (company-disclosed)", revenue_musd=300, revenue_asof="May 2026",
         trajectory="$100M Feb 2025 -> $300M May 2026 (3x in 15 months)",
         valuation_busd=7.2, valuation_note="Jun 2025 Series F ($150M)",
         pricing="Consumption-based + hybrid (fixed per-active-user + usage)",
         scale_metric="Customers: Databricks, Reddit, Pinterest, Samsung",
         token_metric="Sells token-cost reduction: context graph cuts tokens consumed by connected AI",
         confidence="Company-disclosed", in_aggregate="Yes", sources="35"),

    # --- General-purpose & consumer agents ---
    dict(sector="General-purpose & consumer agents", company="Perplexity", offering="Comet (agentic browser) + Perplexity",
         description="Agentic browser and answer engine with multi-agent orchestration (company-wide figure)",
         revenue_basis="Annualized revenue (reported, company-wide)", revenue_musd=500, revenue_asof="Apr 2026",
         trajectory=None,
         valuation_busd=None, valuation_note=None,
         pricing="Subscriptions (Pro/Max) + enterprise",
         scale_metric=None, token_metric=None,
         confidence="Press-reported; company-wide (not agent-only)", in_aggregate="Yes", sources="30"),
    dict(sector="General-purpose & consumer agents", company="Genspark", offering="Genspark Super Agent",
         description="All-in-one agentic workspace for search/productivity using multiple frontier models",
         revenue_basis="ARR (reported)", revenue_musd=100, revenue_asof="2026",
         trajectory="$50M ARR in first 5 months",
         valuation_busd=1.6, valuation_note="2026 round",
         pricing="Subscription + credits",
         scale_metric=None, token_metric=None,
         confidence="Press-reported", in_aggregate="Yes", sources="11, 39"),
    dict(sector="General-purpose & consumer agents", company="Manus (Butterfly Effect)", offering="Manus",
         description="Consumer general-purpose agent: browser control, research, integrations",
         revenue_basis="Annualized revenue (press estimate)", revenue_musd=90, revenue_asof="Aug 2025",
         trajectory="Launched Mar 2025",
         valuation_busd=None, valuation_note=None,
         pricing="Credit-based subscriptions",
         scale_metric=None, token_metric=None,
         confidence="Press estimate (dated)", in_aggregate="Yes", sources="30"),
    dict(sector="General-purpose & consumer agents", company="OpenAI", offering="ChatGPT Agent / Agents platform (memo)",
         description="Deep research + computer use + AgentKit; memo row to avoid double-counting Codex",
         revenue_basis="Memo: agents >40% of enterprise revenue (Feb 2026); enterprise >50% of $40B run-rate (Aug 2026) implies ~$8B+ agent revenue incl. Codex",
         revenue_musd=None, revenue_asof="Aug 2026",
         trajectory="ChatGPT for Work: 7M+ seats (Feb 2026, +40% in 2 months); paying business users ~9M; internal forecast ~$29B agent revenue by 2029",
         valuation_busd=None, valuation_note="IPO planned",
         pricing="Seats + usage; per-task fees emerging",
         scale_metric=None, token_metric=None,
         confidence="Derived (memo only)", in_aggregate="No (memo; overlaps Codex)", sources="31, 32"),

    # --- Model providers (memo) ---
    dict(sector="Model providers (memo - not summed)", company="OpenAI", offering="Company total",
         description="Full company run-rate; agentic products (Codex, agents) are the fastest-growing segment",
         revenue_basis="Annualized run-rate (Bloomberg/CNBC-confirmed)", revenue_musd=40000, revenue_asof="Aug 2026",
         trajectory="~$20B end-2025 -> $40B+ Aug 2026 (+20% MoM in July); enterprise >50% of mix; ads ~$1B run-rate",
         valuation_busd=None, valuation_note="IPO planned",
         pricing="Subscriptions, API usage, ads",
         scale_metric="800M+ weekly ChatGPT users; 20M WAU coding/work products",
         token_metric="API: >6B tokens/min (Oct 2025 DevDay) = ~260T tokens/month, excl. ChatGPT",
         confidence="Confirmed by Bloomberg/CNBC", in_aggregate="No (memo)", sources="29, 30, 31"),
    dict(sector="Model providers (memo - not summed)", company="Anthropic", offering="Company total",
         description="Full company run-rate; Claude Code is ~22% of tracked ARR; enterprise API 80-85% of mix",
         revenue_basis="Annualized run-rate (Bloomberg via internal docs)", revenue_musd=65000, revenue_asof="End of July 2026",
         trajectory="$14B Feb 2026 -> $65B Jul 2026; Q2 2026 revenue $11.5B (> OpenAI's quarter, first time); first positive operating income",
         valuation_busd=380, valuation_note="Feb 2026 Series G post-money",
         pricing="API usage, subscriptions",
         scale_metric="34.4% of B2B market vs OpenAI 32.3% (Jul 2026)",
         token_metric="Token throughput not disclosed; Barclays sees Anthropic leading enterprise inference tokens",
         confidence="Bloomberg-reported", in_aggregate="No (memo)", sources="6"),
    dict(sector="Model providers (memo - not summed)", company="Google", offering="Gemini (platform total)",
         description="Agentic revenue not broken out; included for token-scale context",
         revenue_basis="Not broken out", revenue_musd=None, revenue_asof="May 2026",
         trajectory=None,
         valuation_busd=None, valuation_note="Public (NASDAQ: GOOGL)",
         pricing="API usage, subscriptions, ads",
         scale_metric="8.5M developers building with Gemini monthly",
         token_metric="3.2 quadrillion tokens/month across products (May 2026); Gemini models ~22B tokens/min (~1 quadrillion/month) Q2 2026",
         confidence="Company-disclosed (tokens)", in_aggregate="No (memo)", sources="21, 22, 23"),
]

# ---------------------------------------------------------------------------
# Token usage metrics
# ---------------------------------------------------------------------------
TOKEN_METRICS = [
    ("Industry-wide", "Estimated global token processing", "~370 trillion/day (~135 quadrillion/yr)", "Mid-2026",
     "I/O Fund estimate; already 2.4x Dell's 2028 forecast of 57 quadrillion/yr", "23"),
    ("Google (all products)", "Monthly tokens processed", "3.2 quadrillion/month", "May 2026 (I/O)",
     "vs ~480T May 2025 (~7x YoY) and 9.7T May 2024 (~330x in 2 years); includes Search, Workspace, Gemini apps", "21"),
    ("Google Gemini Developer API", "API token throughput", "19B tokens/min (~800T/month)", "May 2026",
     "375+ Google Cloud customers each processed >1T tokens in trailing 12 months", "21, 22"),
    ("Google Gemini models", "Model token throughput", "22B tokens/min (~1 quadrillion/month)", "Q2 2026 earnings",
     "+120% in six months", "23"),
    ("OpenAI API", "API token throughput", ">6B tokens/min (~260T/month)", "Oct 2025 (DevDay)",
     "API only; excludes ChatGPT consumer traffic (800M+ weekly users)", "22"),
    ("Anthropic", "Token throughput", "Not disclosed", "2026",
     "Barclays analysts see Anthropic leading enterprise inference tokens; 65% of Vercel AI Gateway spend on 30% of volume", "22, 24"),
    ("Microsoft Azure AI Foundry", "Customers at 1T-token annualized run-rate", "Up 4x YoY; 100K Foundry customers", "Q4 FY26 (Jun 2026)",
     "Foundry revenue more than doubled YoY; 11,000+ models offered; Copilot workload throughput up 4x during FY26", "36, 37"),
    ("Microsoft Agent 365", "Registered agents", "~40M agents", "Q4 FY26 (2 months post-launch)",
     "Across tens of thousands of companies; agent governance layer", "36, 37"),
    ("Salesforce Agentforce", "Tokens processed", "28.6T in Q1 FY2027 (+152% QoQ)", "Qtr ended Apr 2026",
     "3.8B agentic work units delivered (+111%)", "17, 18"),
    ("OpenRouter (router)", "Total routed tokens", ">20T/week (~80-90T/month)", "Mid-2026",
     "State of AI study covered 100T+ tokens over ~1 year", "22, 26"),
    ("OpenRouter (router)", "Agentic token multiplier", "~15x tokens per request vs human chat", "2026",
     "Agentic share exploded starting Feb 2026; traffic segmented agentic/mixed/human by 7-signal composite", "25"),
    ("OpenRouter (router)", "Reasoning-model share of tokens", ">50% of routed tokens", "Late 2025",
     "Up from negligible in early 2025; driven by agent-style workflows", "26"),
    ("OpenRouter (router)", "Chinese-origin open-weight share", "~46% of identified token volume", "Mid-2026",
     "Up from <2% a year earlier; DeepSeek alone ~17.6%; agentic workloads drove DeepSeek V4 to ~20% token share", "3 (context), 25"),
    ("OpenRouter apps (snapshot)", "Top agent apps by tokens", "Hermes Agent 27.0T (45.9% of tracked app tokens); Kilo Code 6.9T; Claude Code 6.5T; OpenClaw 3.5T; Cline 2.7T; Codex 0.69T", "Mid-2026",
     "Coding/CLI agents dominate router token volume; app defaults are kingmakers", "27, 28"),
    ("Vercel AI Gateway", "Total routed tokens", "Tens of trillions/month", "Jul 2026",
     "Anthropic: 65% of gateway spend on 30% of token volume at 4.4x average price; >80% of coding-agent spend; open-weight models >1/3 of volume; avg price/token -13.6% MoM", "24"),
    ("Coding agents (research)", "Token intensity vs chat", "~1,000x tokens vs a chat request (third-party researchers); multi-agent systems up to 15x (Anthropic estimate)", "2026",
     "Explains why agentic offerings dominate token demand", "23"),
    ("Claude Code", "Output footprint", "~4% of all global public GitHub commits (SemiAnalysis); projected 20%+ of daily commits by end-2026", "Feb 2026",
     "Average developer ~20 hrs/week in tool", "4 (via orbilontech), 5"),
]

# ---------------------------------------------------------------------------
# Helpers / styling
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
SECTION_FONT = Font(bold=True, size=11, color="1F3864")
BODY_FONT = Font(size=10)
MEMO_FILL = PatternFill("solid", fgColor="F2F2F2")
BAND_FILL = PatternFill("solid", fgColor="DDEBF7")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BORDER


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
wb = Workbook()

# --- Read Me ---
ws = wb.active
ws.title = "Read Me"
set_widths(ws, [120])
readme_lines = [
    ("Agentic AI Offerings: Revenue & Token Usage Across Sectors", TITLE_FONT),
    (f"Compiled: {AS_OF}. All figures are annualized USD unless noted.", BODY_FONT),
    ("", None),
    ("SHEETS", SECTION_FONT),
    ("  Offerings - one row per agentic offering: revenue, growth trajectory, valuation, pricing model, scale and token metrics, sources.", BODY_FONT),
    ("  Sector Aggregates - revenue summed by sector (live SUMIFS over the Offerings sheet; memo rows excluded).", BODY_FONT),
    ("  Token Usage - platform- and ecosystem-level token throughput metrics relevant to agentic workloads.", BODY_FONT),
    ("  Sources - numbered source list; the 'Sources' column in other sheets references these numbers.", BODY_FONT),
    ("", None),
    ("DEFINITIONS", SECTION_FONT),
    ("  ARR - annual recurring revenue as disclosed or estimated; in AI, often used interchangeably with annualized run-rate.", BODY_FONT),
    ("  Annualized run-rate - most recent period revenue x 12 (or x4); volatile for fast-growing companies.", BODY_FONT),
    ("  Tracked ARR - third-party estimate (e.g., TickerTrends) anchored to a company disclosure and extrapolated from usage signals.", BODY_FONT),
    ("  ACV - annual contract value (ServiceNow reports its AI portfolio this way).", BODY_FONT),
    ("", None),
    ("KEY CAVEATS", SECTION_FONT),
    ("  1. Double counting: model-provider totals (OpenAI, Anthropic, Google) overlap with application rows (Codex, Claude Code) and with", BODY_FONT),
    ("     startups that resell frontier models (Cursor, Replit, Lovable pay API costs to providers). Memo rows are excluded from aggregates.", BODY_FONT),
    ("  2. Estimate quality varies: figures range from company-disclosed (Agentforce, Glean) to analyst-tracked (Claude Code, Codex) to", BODY_FONT),
    ("     analyst-inferred (Rogo). See the Confidence column.", BODY_FONT),
    ("  3. Microsoft 365 Copilot revenue is not disclosed; 30M+ seats at $30/user/mo list implies up to ~$10.8B annualized, but actual", BODY_FONT),
    ("     discounted revenue is unknown, so it is excluded from sector aggregates (a large downward bias on the Enterprise platform sector).", BODY_FONT),
    ("  4. Token metrics are heterogeneous: platform-wide (Google), API-only (OpenAI), router-visible (OpenRouter/Vercel see only traffic", BODY_FONT),
    ("     routed through them - a small single-digit % of global volume). Do not sum across rows.", BODY_FONT),
    ("  5. Some figures are stale (Abridge/Ambience May 2025; Manus Aug 2025); dates are given per row.", BODY_FONT),
    ("  6. Run-rates in this market can double in a quarter; treat point-in-time numbers accordingly.", BODY_FONT),
]
for i, (text, font) in enumerate(readme_lines, start=1):
    cell = ws.cell(row=i, column=1, value=text)
    if font:
        cell.font = font

# --- Offerings ---
ws = wb.create_sheet("Offerings")
headers = ["Sector", "Company", "Offering", "What it does", "Revenue basis",
           "Annualized revenue ($M)", "Revenue as-of", "Growth trajectory",
           "Valuation ($B)", "Valuation note", "Pricing model",
           "Scale / adoption metrics", "Token / usage metrics",
           "Confidence", "In sector aggregate?", "Sources (#)"]
ws.append(headers)
style_header_row(ws, 1, len(headers))
for off in OFFERINGS:
    ws.append([
        off["sector"], off["company"], off["offering"], off["description"],
        off["revenue_basis"], off["revenue_musd"], off["revenue_asof"],
        off["trajectory"], off["valuation_busd"], off["valuation_note"],
        off["pricing"], off["scale_metric"], off["token_metric"],
        off["confidence"], off["in_aggregate"], off["sources"],
    ])
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
    is_memo = "memo" in str(row[0].value).lower()
    for cell in row:
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = BORDER
        if is_memo:
            cell.fill = MEMO_FILL
    row[5].number_format = "#,##0.0"
    row[8].number_format = "#,##0.0"
set_widths(ws, [24, 18, 22, 34, 26, 12, 14, 42, 10, 26, 30, 38, 38, 22, 14, 12])
ws.freeze_panes = "D2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

# --- Sector Aggregates ---
ws = wb.create_sheet("Sector Aggregates")
agg_headers = ["Sector", "# offerings tracked", "# with revenue in aggregate",
               "Aggregate annualized revenue ($M)", "Share of tracked total", "Notes"]
ws.append(agg_headers)
style_header_row(ws, 1, len(agg_headers))

sector_notes = {
    "Coding & software development": "Largest agentic sector by revenue; Claude Code + Codex figures are third-party tracked estimates. Cursor acquired by SpaceX for $60B (Jun 2026).",
    "Customer experience & support": "Excludes Agentforce (classified under Enterprise platforms) and Wonderful AI (revenue n/d). Category commands the highest revenue multiples (~127x ARR avg).",
    "Legal": "Harvey dominates; 7-figure enterprise contracts at large law firms.",
    "Healthcare": "Sub-segment sizing (Information Matters, May 2026): scribes $300-500M; RCM automation $0.8-1.2B; payer prior-auth $200-400M; sector central case ~$5.5B incl. drug-discovery value.",
    "Financial services": "Rogo ARR not disclosed - $75M shown is midpoint of analyst-inferred $50-100M. AlphaSense only partially agentic.",
    "Enterprise platform agents": "Excludes Microsoft M365 Copilot (revenue n/d; 30M+ seats imply up to ~$10.8B at list price) - materially understated.",
    "General-purpose & consumer agents": "Perplexity figure is company-wide; Manus figure dated Aug 2025. OpenAI ChatGPT Agent excluded (memo; overlaps Codex).",
}
sectors_in_order = []
for off in OFFERINGS:
    if off["sector"] not in sectors_in_order and "memo" not in off["sector"].lower():
        sectors_in_order.append(off["sector"])

r = 2
for sector in sectors_in_order:
    n_tracked = sum(1 for o in OFFERINGS if o["sector"] == sector)
    ws.cell(row=r, column=1, value=sector)
    ws.cell(row=r, column=2, value=n_tracked)
    ws.cell(row=r, column=3,
            value=f'=COUNTIFS(Offerings!A:A,A{r},Offerings!O:O,"Yes",Offerings!F:F,">0")')
    ws.cell(row=r, column=4,
            value=f'=SUMIFS(Offerings!F:F,Offerings!A:A,A{r},Offerings!O:O,"Yes")')
    ws.cell(row=r, column=5, value=f"=D{r}/D${len(sectors_in_order) + 2}")
    ws.cell(row=r, column=6, value=sector_notes.get(sector, ""))
    r += 1
total_row = r
ws.cell(row=total_row, column=1, value="TOTAL (tracked, ex-memo rows)")
ws.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row - 1})")
ws.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row - 1})")
ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{total_row - 1})")
ws.cell(row=total_row, column=5, value=1)
memo_row = total_row + 2
ws.cell(row=memo_row, column=1, value="Memo: model-provider company totals (not summed above)")
ws.cell(row=memo_row, column=4,
        value='=SUMIFS(Offerings!F:F,Offerings!A:A,"Model providers (memo - not summed)")')
ws.cell(row=memo_row, column=6,
        value="OpenAI $40B (Aug 2026) + Anthropic $65B (Jul 2026) run-rates; agentic products are the top growth driver for both.")
for row in ws.iter_rows(min_row=2, max_row=memo_row, max_col=len(agg_headers)):
    for cell in row:
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = BORDER
    row[3].number_format = "#,##0"
    row[4].number_format = "0.0%"
for c in range(1, len(agg_headers) + 1):
    ws.cell(row=total_row, column=c).font = Font(bold=True, size=10)
    ws.cell(row=total_row, column=c).fill = BAND_FILL
    ws.cell(row=memo_row, column=c).fill = MEMO_FILL
set_widths(ws, [34, 12, 14, 18, 12, 80])
ws.freeze_panes = "A2"

# --- Token Usage ---
ws = wb.create_sheet("Token Usage")
tok_headers = ["Entity / platform", "Metric", "Value", "As-of", "Context / notes", "Sources (#)"]
ws.append(tok_headers)
style_header_row(ws, 1, len(tok_headers))
for row_data in TOKEN_METRICS:
    ws.append(list(row_data))
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(tok_headers)):
    for cell in row:
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = BORDER
set_widths(ws, [26, 32, 44, 18, 62, 12])
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:F{ws.max_row}"

# --- Sources ---
ws = wb.create_sheet("Sources")
src_headers = ["#", "Publisher", "Title / description", "URL"]
ws.append(src_headers)
style_header_row(ws, 1, len(src_headers))
for num, pub, title, url in SOURCES:
    ws.append([num, pub, title, url])
    ws.cell(row=ws.max_row, column=4).hyperlink = url
    ws.cell(row=ws.max_row, column=4).font = Font(size=10, color="0563C1", underline="single")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(src_headers)):
    for cell in row[:3]:
        cell.font = BODY_FONT
    for cell in row:
        cell.alignment = WRAP
        cell.border = BORDER
set_widths(ws, [6, 26, 70, 80])
ws.freeze_panes = "A2"

os.makedirs("data", exist_ok=True)
wb.save("agentic_ai_revenue_token_usage_2026.xlsx")

# ---------------------------------------------------------------------------
# CSV mirrors
# ---------------------------------------------------------------------------
with open("data/offerings_revenue.csv", "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["sector", "company", "offering", "description", "revenue_basis",
                "annualized_revenue_usd_m", "revenue_as_of", "growth_trajectory",
                "valuation_usd_b", "valuation_note", "pricing_model",
                "scale_metrics", "token_metrics", "confidence",
                "in_sector_aggregate", "source_ids"])
    for o in OFFERINGS:
        w.writerow([o["sector"], o["company"], o["offering"], o["description"],
                    o["revenue_basis"], o["revenue_musd"], o["revenue_asof"],
                    o["trajectory"], o["valuation_busd"], o["valuation_note"],
                    o["pricing"], o["scale_metric"], o["token_metric"],
                    o["confidence"], o["in_aggregate"], o["sources"]])

with open("data/sector_aggregates.csv", "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["sector", "offerings_tracked", "offerings_in_aggregate",
                "aggregate_annualized_revenue_usd_m", "share_of_tracked_total", "notes"])
    totals = {}
    counts = {}
    tracked = {}
    for o in OFFERINGS:
        s = o["sector"]
        if "memo" in s.lower():
            continue
        tracked[s] = tracked.get(s, 0) + 1
        if o["in_aggregate"] == "Yes" and o["revenue_musd"]:
            totals[s] = totals.get(s, 0) + o["revenue_musd"]
            counts[s] = counts.get(s, 0) + 1
    grand = sum(totals.values())
    for s in sectors_in_order:
        w.writerow([s, tracked.get(s, 0), counts.get(s, 0), round(totals.get(s, 0), 1),
                    f"{totals.get(s, 0) / grand:.1%}", sector_notes.get(s, "")])
    w.writerow(["TOTAL (tracked, ex-memo rows)", sum(tracked.values()),
                sum(counts.values()), round(grand, 1), "100.0%", ""])

with open("data/token_usage_metrics.csv", "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["entity_platform", "metric", "value", "as_of", "context_notes", "source_ids"])
    for row_data in TOKEN_METRICS:
        w.writerow(list(row_data))

with open("data/sources.csv", "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["id", "publisher", "title", "url"])
    for num, pub, title, url in SOURCES:
        w.writerow([num, pub, title, url])

print("Wrote agentic_ai_revenue_token_usage_2026.xlsx and 4 CSVs under data/")
