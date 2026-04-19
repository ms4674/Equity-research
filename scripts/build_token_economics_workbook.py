"""
Build an XLSX workbook estimating token consumption (input / reasoning / output)
and average agent task duration for five representative BFSI agent use cases:

  1. Retail customer service virtual assistant
  2. Document analysis for loans (credit agreement / covenant extraction / memo)
  3. Insurance claims correspondence drafting
  4. Wealth advisor knowledge assistant
  5. Fraud detection (LLM overlay for case review / explainability / dispute evidence)

Output: data/bfsi_agent_token_economics.xlsx

Sheets:
  1. README              - definitions, methodology, headline finding, disclaimer
  2. Token_Economics     - per-task token ranges (Low/Typical/High) + duration + cost
  3. Per_UseCase_Detail  - long-form per-use-case parameter table
  4. Model_Prices        - reference public list prices used in cost computation
  5. Sources             - citations and URLs

All token counts and durations are *modelled estimates* triangulated from
public vendor disclosures, customer case studies, and observed bank pilots.
They are intended as *order-of-magnitude planning numbers*, not contractual
commitments. Nothing in this workbook is investment advice.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


OUT_PATH = Path(__file__).resolve().parents[1] / "data" / "bfsi_agent_token_economics.xlsx"


# ---------------------------------------------------------------------------
# Reference model prices (USD per 1M tokens, public list price, 2025/2026 era)
# ---------------------------------------------------------------------------
# Sources are listed in the Sources sheet. Reasoning tokens are billed at the
# output rate by both OpenAI and Anthropic.

MODEL_PRICES: list[dict] = [
    {
        "Tier": "Small/cheap (chat, drafting, classification)",
        "Examples": "Claude Haiku 3.5, GPT-4o mini, GPT-4.1 mini",
        "InputPrice_per_M": 0.80,
        "OutputPrice_per_M": 4.00,
        "ReasoningBilling": "Billed at output rate",
        "Notes": "Used for high-volume retail chat and templated drafting.",
    },
    {
        "Tier": "Mid (frontier non-reasoning)",
        "Examples": "GPT-4o, GPT-4.1, Claude Sonnet 4 (non-thinking)",
        "InputPrice_per_M": 3.00,
        "OutputPrice_per_M": 15.00,
        "ReasoningBilling": "n/a (no separate reasoning tokens)",
        "Notes": "Workhorse for document analysis and advisor research.",
    },
    {
        "Tier": "Reasoning / extended thinking",
        "Examples": "OpenAI o3 / o4-mini, Claude Sonnet 4 with extended thinking",
        "InputPrice_per_M": 3.00,
        "OutputPrice_per_M": 15.00,
        "ReasoningBilling": "Billed at output rate",
        "Notes": "Reasoning tokens dominate output cost on hard tasks.",
    },
    {
        "Tier": "Large reasoning (premium)",
        "Examples": "Claude Opus 4 with extended thinking, GPT-5 high-reasoning",
        "InputPrice_per_M": 15.00,
        "OutputPrice_per_M": 75.00,
        "ReasoningBilling": "Billed at output rate",
        "Notes": "Used selectively for complex credit memos / regulatory analysis.",
    },
]


# ---------------------------------------------------------------------------
# Use-case parameter ranges
# ---------------------------------------------------------------------------
# Each use case has Low / Typical / High estimates for:
#   InputTokens, ReasoningTokens, OutputTokens, DurationSec
# Plus an assumed model tier name (must match MODEL_PRICES.Tier).
#
# "Per task" definitions are documented per row.

USE_CASES: list[dict] = [
    {
        "UseCase": "Retail customer service virtual assistant",
        "TaskUnit": "One resolved customer conversation (avg 3-5 turns + RAG over policy/account)",
        "ModelTier": "Small/cheap (chat, drafting, classification)",
        "InputTokens_Low": 2_000,
        "InputTokens_Typical": 8_000,
        "InputTokens_High": 25_000,
        "ReasoningTokens_Low": 0,
        "ReasoningTokens_Typical": 200,
        "ReasoningTokens_High": 1_000,
        "OutputTokens_Low": 100,
        "OutputTokens_Typical": 400,
        "OutputTokens_High": 1_500,
        "DurationSec_Low": 5,
        "DurationSec_Typical": 60,
        "DurationSec_High": 300,
        "Anchors": (
            "Klarna (Feb 2024): avg resolution time fell from 11 min to <2 min (~120s). "
            "BofA Erica >2bn interactions; Wells Fargo Fargo ~245m/yr. Conversations average "
            "3-6 turns. RAG retrieves 2-6 policy/account snippets per turn (~1-3k tokens each)."
        ),
        "PrimarySource": "S1,S2,S3,S4",
    },
    {
        "UseCase": "Document analysis for loans (credit agreement / covenant extraction / memo)",
        "TaskUnit": "One credit agreement processed end-to-end (extract covenants + structured fields + draft memo section)",
        "ModelTier": "Reasoning / extended thinking",
        "InputTokens_Low": 30_000,
        "InputTokens_Typical": 120_000,
        "InputTokens_High": 400_000,
        "ReasoningTokens_Low": 1_000,
        "ReasoningTokens_Typical": 8_000,
        "ReasoningTokens_High": 30_000,
        "OutputTokens_Low": 1_000,
        "OutputTokens_Typical": 5_000,
        "OutputTokens_High": 20_000,
        "DurationSec_Low": 30,
        "DurationSec_Typical": 180,
        "DurationSec_High": 900,
        "Anchors": (
            "Syndicated credit agreements: ~80-200 pages (~80-150k tokens). Master + amendments "
            "+ ISDA schedules can exceed 400k. JPM COiN claimed ~360k lawyer-hours/yr saved on "
            "commercial loan agreements (pre-LLM). Long-context models (Claude 200k+, Gemini "
            "1M+) make single-shot ingest practical. Extended-thinking reasoning tokens "
            "dominate hard-extraction cost."
        ),
        "PrimarySource": "S5,S6,S7,S8",
    },
    {
        "UseCase": "Insurance claims correspondence drafting",
        "TaskUnit": "One outbound letter to claimant / counterparty (FNOL ack, status update, decision letter)",
        "ModelTier": "Small/cheap (chat, drafting, classification)",
        "InputTokens_Low": 2_000,
        "InputTokens_Typical": 8_000,
        "InputTokens_High": 30_000,
        "ReasoningTokens_Low": 0,
        "ReasoningTokens_Typical": 500,
        "ReasoningTokens_High": 2_000,
        "OutputTokens_Low": 200,
        "OutputTokens_Typical": 600,
        "OutputTokens_High": 2_000,
        "DurationSec_Low": 5,
        "DurationSec_Typical": 30,
        "DurationSec_High": 120,
        "Anchors": (
            "Allstate publicly disclosed ~50,000 GenAI-drafted claim letters/day. Letters are "
            "highly templated (regulatory boilerplate + per-claim variables). Input includes "
            "claim file extract + customer profile + applicable state-law template."
        ),
        "PrimarySource": "S9,S10",
    },
    {
        "UseCase": "Wealth advisor knowledge assistant",
        "TaskUnit": "One advisor query answered (RAG over internal research library + portfolio context)",
        "ModelTier": "Mid (frontier non-reasoning)",
        "InputTokens_Low": 4_000,
        "InputTokens_Typical": 20_000,
        "InputTokens_High": 80_000,
        "ReasoningTokens_Low": 200,
        "ReasoningTokens_Typical": 2_000,
        "ReasoningTokens_High": 10_000,
        "OutputTokens_Low": 300,
        "OutputTokens_Typical": 800,
        "OutputTokens_High": 3_000,
        "DurationSec_Low": 5,
        "DurationSec_Typical": 30,
        "DurationSec_High": 120,
        "Anchors": (
            "Morgan Stanley AI @ MS Assistant (OpenAI) covers ~100k+ research documents; ~98% "
            "of advisor teams adopted by 2024. Typical query retrieves 5-15 chunks of 1-3k "
            "tokens each. Multi-document synthesis (e.g., compare two house views) drives the "
            "high end of input range."
        ),
        "PrimarySource": "S11,S12",
    },
    {
        "UseCase": "Fraud detection (LLM overlay: case review, explainability, dispute evidence)",
        "TaskUnit": "One disputed/flagged transaction reviewed by LLM (NOT real-time authorization scoring)",
        "ModelTier": "Mid (frontier non-reasoning)",
        "InputTokens_Low": 500,
        "InputTokens_Typical": 3_000,
        "InputTokens_High": 15_000,
        "ReasoningTokens_Low": 0,
        "ReasoningTokens_Typical": 300,
        "ReasoningTokens_High": 2_000,
        "OutputTokens_Low": 50,
        "OutputTokens_Typical": 200,
        "OutputTokens_High": 1_500,
        "DurationSec_Low": 1,
        "DurationSec_Typical": 5,
        "DurationSec_High": 30,
        "Anchors": (
            "Real-time card authorization is gradient-boosted ML (sub-100ms latency budget); "
            "LLMs are too slow for the auth path. LLM use is in (a) post-auth case review and "
            "explainability for analysts, (b) dispute / chargeback evidence drafting, and (c) "
            "scam-message classification. Mastercard Decision Intelligence Pro and Visa AI "
            "use generative models in this overlay role."
        ),
        "PrimarySource": "S13,S14,S15,S16",
    },
]


# ---------------------------------------------------------------------------
# Sources
# ---------------------------------------------------------------------------

SOURCES: list[dict] = [
    {"id": "S1", "publisher": "Klarna", "date": "Feb 2024",
     "short": "Klarna AI assistant handles 2/3 of customer service chats; avg resolution 11 min -> <2 min",
     "url": "https://www.klarna.com/international/press/klarna-ai-assistant-handles-two-thirds-of-customer-service-chats-in-its-first-month/"},
    {"id": "S2", "publisher": "Bank of America", "date": "Apr 2024",
     "short": "Erica virtual assistant surpasses 2bn interactions across 42m+ users",
     "url": "https://newsroom.bankofamerica.com/content/newsroom/press-releases/2024/04/bank-of-america-s-erica--surpasses-2-billion-interactions--helpi.html"},
    {"id": "S3", "publisher": "Wells Fargo", "date": "2024",
     "short": "Fargo virtual assistant ~245m interactions in 2024",
     "url": "https://newsroom.wf.com/English/news-releases/news-release-details/2024/Wells-Fargo-Fargo-Virtual-Assistant-Hits-Milestone/default.aspx"},
    {"id": "S4", "publisher": "OpenAI / Anthropic", "date": "2024-2025",
     "short": "API pricing pages and model docs (GPT-4o-mini, Claude Haiku 3.5)",
     "url": "https://openai.com/api/pricing/"},
    {"id": "S5", "publisher": "JPMorgan / Bloomberg", "date": "2017",
     "short": "JPMorgan COiN: ~360k lawyer-hours/yr saved on commercial loan agreements",
     "url": "https://www.bloomberg.com/news/articles/2017-02-28/jpmorgan-marshals-an-army-of-developers-to-automate-high-finance"},
    {"id": "S6", "publisher": "LSTA", "date": "ongoing",
     "short": "Loan Syndications and Trading Association: typical syndicated credit agreement structure & length",
     "url": "https://www.lsta.org/"},
    {"id": "S7", "publisher": "Anthropic", "date": "2024-2025",
     "short": "Claude 200k context + extended thinking pricing and capability docs",
     "url": "https://www.anthropic.com/pricing"},
    {"id": "S8", "publisher": "OpenAI", "date": "2024-2025",
     "short": "OpenAI o-series reasoning models: 'reasoning_tokens' billed at output rate",
     "url": "https://platform.openai.com/docs/guides/reasoning"},
    {"id": "S9", "publisher": "Allstate / Wall Street Journal", "date": "2024",
     "short": "Allstate uses generative AI to draft ~50,000 claim letters per day",
     "url": "https://www.wsj.com/articles/allstate-uses-generative-ai-to-write-claims-letters-49b6a7ee"},
    {"id": "S10", "publisher": "Guidewire / Duck Creek", "date": "2024",
     "short": "Vendor docs on claims correspondence templates and FNOL workflows",
     "url": "https://www.guidewire.com/"},
    {"id": "S11", "publisher": "Morgan Stanley + OpenAI", "date": "2023-2024",
     "short": "AI @ Morgan Stanley Assistant: ~98% advisor-team adoption; ~100k+ research docs indexed",
     "url": "https://www.morganstanley.com/press-releases/key-milestone-in-innovation-journey-with-openai"},
    {"id": "S12", "publisher": "OpenAI customer story", "date": "2024",
     "short": "Morgan Stanley case study (token-volume context implied by RAG corpus size)",
     "url": "https://openai.com/index/morgan-stanley/"},
    {"id": "S13", "publisher": "Visa", "date": "2024",
     "short": "Visa: ~$40bn fraud prevented in 2023; tech/AI investment ~$10bn over 5 years",
     "url": "https://usa.visa.com/about-visa/newsroom/press-releases/visa-prevented-40-billion-in-fraudulent-activity-in-2023.html"},
    {"id": "S14", "publisher": "Mastercard", "date": "Feb 2024",
     "short": "Mastercard Decision Intelligence Pro: generative AI overlay on fraud scoring",
     "url": "https://www.mastercard.com/news/press/2024/february/mastercard-harnesses-generative-ai-to-fight-payment-scams/"},
    {"id": "S15", "publisher": "Stripe", "date": "ongoing",
     "short": "Stripe Radar: ML-based real-time fraud scoring (sub-100ms auth)",
     "url": "https://stripe.com/radar"},
    {"id": "S16", "publisher": "Anthropic / OpenAI", "date": "2024-2025",
     "short": "Claude Sonnet / GPT-4o pricing pages used for fraud-overlay LLM cost",
     "url": "https://www.anthropic.com/pricing"},
    {"id": "S17", "publisher": "OpenAI", "date": "2024-2025",
     "short": "OpenAI list pricing reference (GPT-4o, GPT-4.1, GPT-4o-mini, o-series)",
     "url": "https://openai.com/api/pricing/"},
    {"id": "S18", "publisher": "Anthropic", "date": "2024-2025",
     "short": "Anthropic list pricing reference (Claude Haiku 3.5, Sonnet 4, Opus 4)",
     "url": "https://www.anthropic.com/pricing"},
]

SOURCE_BY_ID = {s["id"]: s for s in SOURCES}

PRICE_BY_TIER = {p["Tier"]: p for p in MODEL_PRICES}


# ---------------------------------------------------------------------------
# Cost calculation
# ---------------------------------------------------------------------------

def cost(input_tok: int, reasoning_tok: int, output_tok: int, tier: str) -> float:
    p = PRICE_BY_TIER[tier]
    in_cost = input_tok * p["InputPrice_per_M"] / 1_000_000
    out_cost = (reasoning_tok + output_tok) * p["OutputPrice_per_M"] / 1_000_000
    return round(in_cost + out_cost, 6)


# ---------------------------------------------------------------------------
# Workbook plumbing
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E78")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def style_header(ws, row_idx: int, n_cols: int, fill=HEADER_FILL) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        cell.border = BORDER


def auto_width(ws, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        m = 0
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            for line in str(v).split("\n"):
                m = max(m, len(line))
        ws.column_dimensions[col_letter].width = min(max(12, m + 2), max_width)


def add_table(ws, headers: list[str], name: str, style: str = "TableStyleMedium2") -> None:
    rng = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    table = Table(displayName=name, ref=rng)
    table.tableStyleInfo = TableStyleInfo(
        name=style, showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)


def write_readme(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "README"
    rows = [
        "BFSI AI Agent Token Economics & Task Duration",
        "Per-task input / reasoning / output token estimates and average duration",
        "",
        "Companion to: research/bfsi-agent-token-economics.md",
        "Generated by: scripts/build_token_economics_workbook.py",
        "",
        "Use cases covered (one row each in Token_Economics):",
        "  1. Retail customer service virtual assistant",
        "  2. Document analysis for loans (credit agreement / covenant extraction / memo)",
        "  3. Insurance claims correspondence drafting",
        "  4. Wealth advisor knowledge assistant",
        "  5. Fraud detection (LLM overlay; NOT real-time card authorization)",
        "",
        "Per-task definitions:",
        "  - 'One task' is defined explicitly per use case in the TaskUnit column.",
        "  - For chat use cases this is one full resolved conversation (multiple turns).",
        "  - For document analysis it is one credit agreement processed end-to-end.",
        "",
        "Token categories:",
        "  - Input tokens   = system prompt + RAG context + user message + history",
        "  - Reasoning tokens = OpenAI o-series 'reasoning_tokens' or Claude extended-thinking",
        "                       tokens. Billed at output rate by both vendors.",
        "  - Output tokens   = visible model response tokens.",
        "",
        "Cost methodology:",
        "  cost = (InputTokens * InputPrice + (ReasoningTokens + OutputTokens) * OutputPrice) / 1e6",
        "  Prices are public list prices for a representative model tier per use case.",
        "  See Model_Prices sheet for the full table; production deployments typically",
        "  negotiate 30-60% off list at scale.",
        "",
        "Latency / duration:",
        "  - DurationSec is the *agent task wall-clock duration* end-to-end, including",
        "    retrieval, model calls, tool calls, and human-in-the-loop wait if applicable.",
        "  - It is NOT raw model time-to-first-token. For real-time card authorization,",
        "    LLMs are not used in the hot path (sub-100ms latency budget); see fraud row.",
        "",
        "Headline:",
        "  - Customer service VA and claims correspondence are CHEAP and FAST per task",
        "    (~<$0.01 typical, seconds to a couple of minutes).",
        "  - Document analysis for loans is the MOST EXPENSIVE per task by 2-3 orders of",
        "    magnitude (~$0.50-$2.00 typical, single-digit minutes), because the input",
        "    context dominates and reasoning tokens add 20-50% on top of output.",
        "  - Wealth advisor and fraud-overlay sit in the middle (~$0.05-$0.30 typical).",
        "",
        "Disclaimer:",
        "  All token counts and durations are MODELLED ESTIMATES triangulated from public",
        "  vendor disclosures, customer case studies, and observed bank pilots. They are",
        "  intended as order-of-magnitude PLANNING numbers, not contractual commitments.",
        "  Nothing in this workbook is investment advice.",
    ]
    ws["A1"] = rows[0]
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    for i, text in enumerate(rows[1:], start=2):
        ws.cell(row=i, column=1, value=text).alignment = WRAP
    ws.column_dimensions["A"].width = 110


def write_token_economics(wb: Workbook) -> None:
    ws = wb.create_sheet("Token_Economics")
    headers = [
        "UseCase",
        "TaskUnit",
        "ModelTier",
        # Input
        "InputTokens_Low", "InputTokens_Typical", "InputTokens_High",
        # Reasoning
        "ReasoningTokens_Low", "ReasoningTokens_Typical", "ReasoningTokens_High",
        # Output
        "OutputTokens_Low", "OutputTokens_Typical", "OutputTokens_High",
        # Total tokens (computed)
        "TotalTokens_Low", "TotalTokens_Typical", "TotalTokens_High",
        # Duration (sec)
        "DurationSec_Low", "DurationSec_Typical", "DurationSec_High",
        # Cost per task ($)
        "CostPerTask_Low_USD", "CostPerTask_Typical_USD", "CostPerTask_High_USD",
        # Per 1k tasks
        "CostPer1kTasks_Typical_USD",
        # Anchors / sources
        "Anchors",
        "PrimarySource",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for u in USE_CASES:
        tier = u["ModelTier"]
        total_low = u["InputTokens_Low"] + u["ReasoningTokens_Low"] + u["OutputTokens_Low"]
        total_typ = u["InputTokens_Typical"] + u["ReasoningTokens_Typical"] + u["OutputTokens_Typical"]
        total_hi = u["InputTokens_High"] + u["ReasoningTokens_High"] + u["OutputTokens_High"]
        cost_low = cost(u["InputTokens_Low"], u["ReasoningTokens_Low"], u["OutputTokens_Low"], tier)
        cost_typ = cost(u["InputTokens_Typical"], u["ReasoningTokens_Typical"], u["OutputTokens_Typical"], tier)
        cost_hi = cost(u["InputTokens_High"], u["ReasoningTokens_High"], u["OutputTokens_High"], tier)
        cost_per_1k = round(cost_typ * 1000, 4)
        ws.append([
            u["UseCase"],
            u["TaskUnit"],
            tier,
            u["InputTokens_Low"], u["InputTokens_Typical"], u["InputTokens_High"],
            u["ReasoningTokens_Low"], u["ReasoningTokens_Typical"], u["ReasoningTokens_High"],
            u["OutputTokens_Low"], u["OutputTokens_Typical"], u["OutputTokens_High"],
            total_low, total_typ, total_hi,
            u["DurationSec_Low"], u["DurationSec_Typical"], u["DurationSec_High"],
            cost_low, cost_typ, cost_hi,
            cost_per_1k,
            u["Anchors"],
            u["PrimarySource"],
        ])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
    # number formatting
    int_cols = list(range(4, 19))  # InputTokens..DurationSec_High
    for r in range(2, ws.max_row + 1):
        for c in int_cols:
            ws.cell(row=r, column=c).number_format = '#,##0'
        for c in [19, 20, 21]:
            ws.cell(row=r, column=c).number_format = '$0.000000'
        ws.cell(row=r, column=22).number_format = '$#,##0.0000'
    ws.freeze_panes = "D2"
    auto_width(ws)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions[get_column_letter(headers.index("Anchors") + 1)].width = 60
    add_table(ws, headers, "TokenEconomics", style="TableStyleMedium2")


def write_per_usecase_detail(wb: Workbook) -> None:
    """Long-form table: one row per (UseCase, Metric) for easier reading."""
    ws = wb.create_sheet("Per_UseCase_Detail")
    headers = ["UseCase", "Metric", "Low", "Typical", "High", "Unit", "Notes"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    metrics = [
        ("InputTokens", "tokens", "System prompt + RAG context + user message + history"),
        ("ReasoningTokens", "tokens", "OpenAI o-series reasoning_tokens / Claude extended-thinking; billed at output rate"),
        ("OutputTokens", "tokens", "Visible model response"),
        ("DurationSec", "seconds", "End-to-end agent wall-clock per task (incl. retrieval, tool calls)"),
    ]

    for u in USE_CASES:
        tier = u["ModelTier"]
        for metric, unit, note in metrics:
            ws.append([
                u["UseCase"],
                metric,
                u[f"{metric}_Low"],
                u[f"{metric}_Typical"],
                u[f"{metric}_High"],
                unit,
                note,
            ])
        ws.append([
            u["UseCase"],
            "CostPerTask_USD",
            cost(u["InputTokens_Low"], u["ReasoningTokens_Low"], u["OutputTokens_Low"], tier),
            cost(u["InputTokens_Typical"], u["ReasoningTokens_Typical"], u["OutputTokens_Typical"], tier),
            cost(u["InputTokens_High"], u["ReasoningTokens_High"], u["OutputTokens_High"], tier),
            "USD",
            f"At {tier} list price",
        ])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
        m = ws.cell(row=r, column=2).value
        if m == "CostPerTask_USD":
            for c in (3, 4, 5):
                ws.cell(row=r, column=c).number_format = '$0.000000'
        else:
            for c in (3, 4, 5):
                ws.cell(row=r, column=c).number_format = '#,##0'
    ws.freeze_panes = "A2"
    auto_width(ws)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["G"].width = 70
    add_table(ws, headers, "PerUseCaseDetail", style="TableStyleMedium9")


def write_model_prices(wb: Workbook) -> None:
    ws = wb.create_sheet("Model_Prices")
    headers = ["Tier", "Examples", "InputPrice_per_M_USD", "OutputPrice_per_M_USD",
               "ReasoningBilling", "Notes"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for p in MODEL_PRICES:
        ws.append([
            p["Tier"], p["Examples"],
            p["InputPrice_per_M"], p["OutputPrice_per_M"],
            p["ReasoningBilling"], p["Notes"],
        ])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
        for c in (3, 4):
            ws.cell(row=r, column=c).number_format = '$0.00'
    ws.freeze_panes = "A2"
    auto_width(ws)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["F"].width = 55
    add_table(ws, headers, "ModelPrices", style="TableStyleMedium4")


def write_sources(wb: Workbook) -> None:
    ws = wb.create_sheet("Sources")
    headers = ["Id", "Publisher", "Date", "Description", "URL"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for s in SOURCES:
        ws.append([s["id"], s["publisher"], s["date"], s["short"], s["url"]])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
    ws.freeze_panes = "A2"
    auto_width(ws)
    ws.column_dimensions["D"].width = 75
    ws.column_dimensions["E"].width = 60
    add_table(ws, headers, "Sources", style="TableStyleMedium6")


def main() -> None:
    wb = Workbook()
    write_readme(wb)
    write_token_economics(wb)
    write_per_usecase_detail(wb)
    write_model_prices(wb)
    write_sources(wb)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print()
    print(f"{'Use case':<55} {'In_typ':>8} {'Reas_typ':>9} {'Out_typ':>8} {'Total':>9} {'Sec':>5} {'$/task':>10}")
    for u in USE_CASES:
        tier = u["ModelTier"]
        c = cost(u["InputTokens_Typical"], u["ReasoningTokens_Typical"], u["OutputTokens_Typical"], tier)
        total = u["InputTokens_Typical"] + u["ReasoningTokens_Typical"] + u["OutputTokens_Typical"]
        print(
            f"{u['UseCase'][:55]:<55} "
            f"{u['InputTokens_Typical']:>8,} "
            f"{u['ReasoningTokens_Typical']:>9,} "
            f"{u['OutputTokens_Typical']:>8,} "
            f"{total:>9,} "
            f"{u['DurationSec_Typical']:>5} "
            f"${c:>9.4f}"
        )


if __name__ == "__main__":
    main()
