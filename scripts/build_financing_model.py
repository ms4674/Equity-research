"""
Build an Excel model of AI-infrastructure financing (equity + debt) through 2030
for the major hyperscalers and neocloud / AI-native infrastructure companies.

The workbook breaks total debt financing into mutually-exclusive instrument
buckets (corporate bonds, bank & syndicated loans, private credit, insurance-linked
lending, securitization / ABS, convertibles, vendor & equipment-lease financing)
plus a memo line for the off-balance-sheet SPV / JV portion that those instruments
fund.

Figures are research estimates for the cumulative 2025-2030 buildout. They blend
disclosed transactions (2023-early 2026) with forward projections and are intended
for analytical/illustrative use, not as audited figures. Sources are listed on the
"Key Deals & Sources" sheet and in the methodology notes.

Run:  python scripts/build_financing_model.py
Output: AI_Infrastructure_Financing_2025-2030.xlsx  (repo root)
"""

from __future__ import annotations

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ---------------------------------------------------------------------------
# Palette
# ---------------------------------------------------------------------------
NAVY = "1F2A44"
BLUE = "2E5A9C"
LIGHT_BLUE = "D6E1F2"
TEAL = "1B6E6E"
LIGHT_TEAL = "D2E8E8"
AMBER = "B5760A"
GREY = "5A5A5A"
LIGHT_GREY = "EEEEEE"
WHITE = "FFFFFF"
TOTAL_FILL = "FCEFC7"
HYPER_FILL = "E8EEF7"
NEO_FILL = "E6F2EF"
COLO_FILL = "F4E9DA"
SOV_FILL = "EDE4F4"

THIN = Side(style="thin", color="BFBFBF")
MED = Side(style="medium", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP_BORDER = Border(top=MED)

# ---------------------------------------------------------------------------
# DEBT INSTRUMENT BUCKETS  (mutually exclusive -> sum to total debt)
# ---------------------------------------------------------------------------
DEBT_COLS = [
    "Corporate bonds (public IG/HY)",
    "Bank & syndicated loans",
    "Alternate private credit",
    "Insurance-linked lending",
    "Securitization / ABS / CMBS",
    "Convertibles / equity-linked",
    "Vendor & equipment-lease financing",
]

# Segment display order + row fill colors
SEGMENTS = [
    ("Hyperscaler", HYPER_FILL),
    ("Neocloud / AI-native", NEO_FILL),
    ("Colocation / data-center REIT", COLO_FILL),
    ("Sovereign cloud", SOV_FILL),
]

# ---------------------------------------------------------------------------
# DATA  (cumulative 2025-2030, US$ billions)
#   each entry: segment, equity/internal funding, {debt bucket: $B}, SPV memo
# ---------------------------------------------------------------------------
# fmt: off
COMPANIES = [
    # ---------------- HYPERSCALERS ----------------
    ("Microsoft",  "Hyperscaler", 750,
        {"Corporate bonds (public IG/HY)": 130, "Bank & syndicated loans": 25,
         "Alternate private credit": 30, "Insurance-linked lending": 20,
         "Securitization / ABS / CMBS": 10, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 35}, 70),
    ("Amazon (AWS)", "Hyperscaler", 950,
        {"Corporate bonds (public IG/HY)": 200, "Bank & syndicated loans": 25,
         "Alternate private credit": 35, "Insurance-linked lending": 20,
         "Securitization / ABS / CMBS": 10, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 30}, 50),
    ("Alphabet (Google)", "Hyperscaler", 931,
        {"Corporate bonds (public IG/HY)": 160, "Bank & syndicated loans": 15,
         "Alternate private credit": 25, "Insurance-linked lending": 12,
         "Securitization / ABS / CMBS": 8, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 10}, 30),
    ("Meta Platforms", "Hyperscaler", 487,
        {"Corporate bonds (public IG/HY)": 150, "Bank & syndicated loans": 15,
         "Alternate private credit": 110, "Insurance-linked lending": 70,
         "Securitization / ABS / CMBS": 35, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 20}, 180),
    ("Oracle (OCI)", "Hyperscaler", 175,
        {"Corporate bonds (public IG/HY)": 120, "Bank & syndicated loans": 30,
         "Alternate private credit": 35, "Insurance-linked lending": 15,
         "Securitization / ABS / CMBS": 5, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 10}, 45),
    # ---------------- NEOCLOUDS / AI-NATIVE INFRA ----------------
    ("CoreWeave", "Neocloud / AI-native", 25,
        {"Corporate bonds (public IG/HY)": 10, "Bank & syndicated loans": 50,
         "Alternate private credit": 30, "Insurance-linked lending": 12,
         "Securitization / ABS / CMBS": 8, "Convertibles / equity-linked": 10,
         "Vendor & equipment-lease financing": 0}, 95),
    ("SpaceX / xAI (Colossus)", "Neocloud / AI-native", 140,
        {"Corporate bonds (public IG/HY)": 15, "Bank & syndicated loans": 25,
         "Alternate private credit": 30, "Insurance-linked lending": 5,
         "Securitization / ABS / CMBS": 3, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 27}, 70),
    ("Nebius", "Neocloud / AI-native", 18,
        {"Corporate bonds (public IG/HY)": 2, "Bank & syndicated loans": 12,
         "Alternate private credit": 8, "Insurance-linked lending": 3,
         "Securitization / ABS / CMBS": 2, "Convertibles / equity-linked": 6,
         "Vendor & equipment-lease financing": 2}, 10),
    ("IREN", "Neocloud / AI-native", 10,
        {"Corporate bonds (public IG/HY)": 3, "Bank & syndicated loans": 8,
         "Alternate private credit": 8, "Insurance-linked lending": 3,
         "Securitization / ABS / CMBS": 3, "Convertibles / equity-linked": 5,
         "Vendor & equipment-lease financing": 0}, 8),
    ("Crusoe", "Neocloud / AI-native", 8,
        {"Corporate bonds (public IG/HY)": 2, "Bank & syndicated loans": 6,
         "Alternate private credit": 9, "Insurance-linked lending": 3,
         "Securitization / ABS / CMBS": 2, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 3}, 12),
    ("Lambda", "Neocloud / AI-native", 8,
        {"Corporate bonds (public IG/HY)": 1, "Bank & syndicated loans": 8,
         "Alternate private credit": 5, "Insurance-linked lending": 2,
         "Securitization / ABS / CMBS": 1, "Convertibles / equity-linked": 1,
         "Vendor & equipment-lease financing": 0}, 10),
    ("Fluidstack", "Neocloud / AI-native", 4,
        {"Corporate bonds (public IG/HY)": 1, "Bank & syndicated loans": 6,
         "Alternate private credit": 8, "Insurance-linked lending": 2,
         "Securitization / ABS / CMBS": 1, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 2}, 14),
    ("Nscale", "Neocloud / AI-native", 5,
        {"Corporate bonds (public IG/HY)": 1, "Bank & syndicated loans": 6,
         "Alternate private credit": 5, "Insurance-linked lending": 1,
         "Securitization / ABS / CMBS": 1, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 1}, 8),
    ("Other neoclouds & AI-native DCs", "Neocloud / AI-native", 20,
        {"Corporate bonds (public IG/HY)": 8, "Bank & syndicated loans": 18,
         "Alternate private credit": 18, "Insurance-linked lending": 6,
         "Securitization / ABS / CMBS": 6, "Convertibles / equity-linked": 2,
         "Vendor & equipment-lease financing": 2}, 25),
    # ---------------- COLOCATION / WHOLESALE DATA-CENTER REITs ----------------
    # (own powered shells / campuses leased to hyperscalers & neoclouds; heavy
    #  users of ABS/CMBS securitization, infra equity, green bonds & private credit)
    ("Equinix", "Colocation / data-center REIT", 25,
        {"Corporate bonds (public IG/HY)": 18, "Bank & syndicated loans": 6,
         "Alternate private credit": 3, "Insurance-linked lending": 2,
         "Securitization / ABS / CMBS": 14, "Convertibles / equity-linked": 1,
         "Vendor & equipment-lease financing": 1}, 20),
    ("Digital Realty", "Colocation / data-center REIT", 20,
        {"Corporate bonds (public IG/HY)": 16, "Bank & syndicated loans": 5,
         "Alternate private credit": 3, "Insurance-linked lending": 2,
         "Securitization / ABS / CMBS": 12, "Convertibles / equity-linked": 1,
         "Vendor & equipment-lease financing": 1}, 18),
    ("QTS (Blackstone)", "Colocation / data-center REIT", 18,
        {"Corporate bonds (public IG/HY)": 2, "Bank & syndicated loans": 8,
         "Alternate private credit": 8, "Insurance-linked lending": 4,
         "Securitization / ABS / CMBS": 19, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 1}, 35),
    ("Vantage Data Centers", "Colocation / data-center REIT", 15,
        {"Corporate bonds (public IG/HY)": 3, "Bank & syndicated loans": 8,
         "Alternate private credit": 7, "Insurance-linked lending": 3,
         "Securitization / ABS / CMBS": 13, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 1}, 28),
    ("Aligned Data Centers (MGX/BlackRock)", "Colocation / data-center REIT", 12,
        {"Corporate bonds (public IG/HY)": 1, "Bank & syndicated loans": 6,
         "Alternate private credit": 7, "Insurance-linked lending": 3,
         "Securitization / ABS / CMBS": 10, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 1}, 22),
    ("CyrusOne (KKR/GIP)", "Colocation / data-center REIT", 12,
        {"Corporate bonds (public IG/HY)": 2, "Bank & syndicated loans": 6,
         "Alternate private credit": 6, "Insurance-linked lending": 3,
         "Securitization / ABS / CMBS": 10, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 1}, 22),
    ("Switch (DigitalBridge)", "Colocation / data-center REIT", 8,
        {"Corporate bonds (public IG/HY)": 1, "Bank & syndicated loans": 4,
         "Alternate private credit": 5, "Insurance-linked lending": 2,
         "Securitization / ABS / CMBS": 9, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 1}, 16),
    ("Stack Infrastructure", "Colocation / data-center REIT", 8,
        {"Corporate bonds (public IG/HY)": 1, "Bank & syndicated loans": 4,
         "Alternate private credit": 4, "Insurance-linked lending": 2,
         "Securitization / ABS / CMBS": 6, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 0}, 12),
    ("Other colo & powered-shell owners", "Colocation / data-center REIT", 30,
        {"Corporate bonds (public IG/HY)": 6, "Bank & syndicated loans": 14,
         "Alternate private credit": 12, "Insurance-linked lending": 6,
         "Securitization / ABS / CMBS": 15, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 2}, 30),
    # ---------------- SOVEREIGN CLOUDS (state / SWF-funded national AI infra) ----------------
    ("HUMAIN (Saudi PIF)", "Sovereign cloud", 80,
        {"Corporate bonds (public IG/HY)": 4, "Bank & syndicated loans": 8,
         "Alternate private credit": 4, "Insurance-linked lending": 1,
         "Securitization / ABS / CMBS": 0, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 3}, 10),
    ("MGX / Mubadala (UAE)", "Sovereign cloud", 30,
        {"Corporate bonds (public IG/HY)": 1, "Bank & syndicated loans": 3,
         "Alternate private credit": 4, "Insurance-linked lending": 1,
         "Securitization / ABS / CMBS": 0, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 1}, 8),
    ("Stargate UAE / G42 / Khazna", "Sovereign cloud", 22,
        {"Corporate bonds (public IG/HY)": 1, "Bank & syndicated loans": 3,
         "Alternate private credit": 2, "Insurance-linked lending": 0,
         "Securitization / ABS / CMBS": 0, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 2}, 6),
    ("Qatar (QIA / Qai + Brookfield JV)", "Sovereign cloud", 14,
        {"Corporate bonds (public IG/HY)": 1, "Bank & syndicated loans": 2,
         "Alternate private credit": 2, "Insurance-linked lending": 1,
         "Securitization / ABS / CMBS": 0, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 0}, 4),
    ("EU / France (Mistral, sovereign cloud)", "Sovereign cloud", 12,
        {"Corporate bonds (public IG/HY)": 1, "Bank & syndicated loans": 2,
         "Alternate private credit": 1, "Insurance-linked lending": 0,
         "Securitization / ABS / CMBS": 0, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 1}, 3),
    ("India (IndiaAI Mission)", "Sovereign cloud", 8,
        {"Corporate bonds (public IG/HY)": 0, "Bank & syndicated loans": 1,
         "Alternate private credit": 0, "Insurance-linked lending": 0,
         "Securitization / ABS / CMBS": 0, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 1}, 1),
    ("Other sovereign AI (Japan, Korea, Singapore, etc.)", "Sovereign cloud", 25,
        {"Corporate bonds (public IG/HY)": 2, "Bank & syndicated loans": 3,
         "Alternate private credit": 2, "Insurance-linked lending": 0,
         "Securitization / ABS / CMBS": 0, "Convertibles / equity-linked": 0,
         "Vendor & equipment-lease financing": 1}, 4),
]

# The 3rd field of each COMPANIES entry is TOTAL equity / internal funding
# (operating cash flow self-funded + external equity). The two dicts below split
# it and add a cumulative operating-cash-flow (CFO) estimate so we can derive
# free cash flow and the external-funding gap. (US$bn, cumulative 2025-2030.)
#   operating cash flow self-funded into AI capex = total equity - external equity
EXTERNAL_EQUITY = {
    # Hyperscalers self-fund from CFO; Oracle uses ATM + mandatory convertible preferred.
    "Microsoft": 0, "Amazon (AWS)": 0, "Alphabet (Google)": 0, "Meta Platforms": 0,
    "Oracle (OCI)": 30,
    # Neoclouds: VC / IPO / strategic (NVIDIA) / SPV equity dominate.
    "CoreWeave": 22, "SpaceX / xAI (Colossus)": 90, "Nebius": 16, "IREN": 7,
    "Crusoe": 7, "Lambda": 7, "Fluidstack": 4, "Nscale": 5,
    "Other neoclouds & AI-native DCs": 16,
    # Colocation: sponsor / infra equity + REIT equity issuance.
    "Equinix": 10, "Digital Realty": 8, "QTS (Blackstone)": 16, "Vantage Data Centers": 13,
    "Aligned Data Centers (MGX/BlackRock)": 10, "CyrusOne (KKR/GIP)": 10,
    "Switch (DigitalBridge)": 7, "Stack Infrastructure": 7,
    "Other colo & powered-shell owners": 22,
    # Sovereign: essentially all sovereign-wealth / state equity.
    "HUMAIN (Saudi PIF)": 80, "MGX / Mubadala (UAE)": 30, "Stargate UAE / G42 / Khazna": 22,
    "Qatar (QIA / Qai + Brookfield JV)": 14, "EU / France (Mistral, sovereign cloud)": 12,
    "India (IndiaAI Mission)": 8, "Other sovereign AI (Japan, Korea, Singapore, etc.)": 25,
}

# Cumulative operating cash flow (CFO) generated 2025-2030 (US$bn, est.). For
# hyperscalers CFO >> self-funded capex because much CFO goes to dividends/buybacks;
# for others CFO is set = self-funded capex (they reinvest ~all internal cash).
CFO_TOTAL = {
    "Microsoft": 1300, "Amazon (AWS)": 1100, "Alphabet (Google)": 1350,
    "Meta Platforms": 1000, "Oracle (OCI)": 210,
    "CoreWeave": 3, "SpaceX / xAI (Colossus)": 50, "Nebius": 2, "IREN": 3,
    "Crusoe": 1, "Lambda": 1, "Fluidstack": 0, "Nscale": 0,
    "Other neoclouds & AI-native DCs": 4,
    "Equinix": 15, "Digital Realty": 12, "QTS (Blackstone)": 2, "Vantage Data Centers": 2,
    "Aligned Data Centers (MGX/BlackRock)": 2, "CyrusOne (KKR/GIP)": 2,
    "Switch (DigitalBridge)": 1, "Stack Infrastructure": 1,
    "Other colo & powered-shell owners": 8,
    "HUMAIN (Saudi PIF)": 0, "MGX / Mubadala (UAE)": 0, "Stargate UAE / G42 / Khazna": 0,
    "Qatar (QIA / Qai + Brookfield JV)": 0, "EU / France (Mistral, sovereign cloud)": 0,
    "India (IndiaAI Mission)": 0, "Other sovereign AI (Japan, Korea, Singapore, etc.)": 0,
}


def ocf_self(name, total_equity):
    """Operating cash flow self-funded into AI capex = total equity - external equity."""
    return total_equity - EXTERNAL_EQUITY[name]

CAPEX_YEARS = [2025, 2026, 2027, 2028, 2029, 2030]

# Annual AI-infrastructure capex (US$B). Hyperscalers have explicit annual paths;
# every other company's cumulative capex (= its total financing) is distributed
# across the years using a segment ramp profile (see build_capex_rows()).
HYPER_CAPEX = {
    "Microsoft":          [95, 140, 165, 185, 200, 215],
    "Amazon (AWS)":       [100, 200, 225, 240, 250, 255],
    "Alphabet (Google)":  [91, 180, 200, 215, 230, 245],
    "Meta Platforms":     [72, 125, 150, 165, 180, 195],
    "Oracle (OCI)":       [25, 50, 65, 75, 85, 90],
}
# Ramp weights (share of cumulative spend per year, 2025-2030) by segment.
SEGMENT_RAMP = {
    "Neocloud / AI-native":         [0.10, 0.16, 0.19, 0.18, 0.18, 0.19],
    "Colocation / data-center REIT": [0.12, 0.15, 0.17, 0.18, 0.19, 0.19],
    "Sovereign cloud":              [0.08, 0.14, 0.18, 0.20, 0.20, 0.20],
}


def _distribute(total, weights):
    """Split an integer `total` across years per `weights`, summing exactly to total."""
    raw = [total * w for w in weights]
    floored = [int(x) for x in raw]
    rem = total - sum(floored)
    order = sorted(range(len(weights)), key=lambda i: raw[i] - floored[i], reverse=True)
    for i in range(rem):
        floored[order[i % len(order)]] += 1
    return floored


def build_capex_rows():
    """Return ordered [(segment, [(name, [annual...]), ...]), ...] for every company."""
    out = []
    for seg, _fill in SEGMENTS:
        rows = []
        members = [c for c in COMPANIES if c[1] == seg]
        members.sort(key=lambda c: c[2] + total_debt(c[3]), reverse=True)
        for name, s, equity, debt, spv in members:
            cumulative = equity + total_debt(debt)
            if name in HYPER_CAPEX:
                annual = HYPER_CAPEX[name]
            else:
                annual = _distribute(cumulative, SEGMENT_RAMP[seg])
            rows.append((name, annual))
        out.append((seg, rows))
    return out

# Notable transactions feeding the estimates
DEALS = [
    ("Meta Platforms", "Hyperion JV (Beignet Investor LLC) — Blue Owl 80% / Meta 20%; $27.3bn A+ private bond led by PIMCO (~$18bn) & BlackRock; off-B/S SPV w/ residual-value guarantee", "2025", "$27.3bn debt", "SPV / private credit / insurance", "Meta IR; FT; S&P"),
    ("Meta Platforms", "$30bn multi-tranche IG corporate bond (largest 2025 IG deal), maturities to 2063", "2025", "$30bn", "Corporate bonds", "BofA; MUFG"),
    ("Amazon (AWS)", "~$54bn USD + €14.5bn (~$16.8bn) IG bonds (near-record), ~4x oversubscribed", "2026", "~$71bn", "Corporate bonds", "CNA; Reuters"),
    ("Amazon (AWS)", "$15bn IG bond", "2025", "$15bn", "Corporate bonds", "CNA"),
    ("Alphabet (Google)", "~$32bn multi-currency bond incl. rare 100-yr tranche; plus $17.5bn Nov 2025", "2025-26", "~$50bn", "Corporate bonds", "ConstructConnect; CNA"),
    ("Oracle (OCI)", "$25bn 8-part IG bond (Feb 2026) within $45-50bn CY26 debt+equity plan; $20bn ATM + mandatory convertible preferred (equity)", "2026", "$25bn debt", "Corporate bonds + equity", "Oracle IR; IFR"),
    ("Oracle (OCI)", "$18bn 6-part IG bond (Sep 2025)", "2025", "$18bn", "Corporate bonds", "Reuters/CNA"),
    ("Oracle (OCI)", "Project-level SPV financing: ~$16.3bn Michigan campus; further Texas / Wisconsin deals", "2025-26", "~$25bn+", "SPV / project finance", "Press reports"),
    ("Microsoft", "Finance-lease driven build (~$108bn uncommenced finance leases as of 9/30/24; commence FY25-FY30)", "2024-30", "$100bn+ leases", "Vendor / lease / SPV", "MSFT 10-Q; CNBC"),
    ("CoreWeave", "DDTL 4.0 — $8.5bn delayed-draw term loan; first IG-rated (Moody's A3 / DBRS A-low) GPU-backed financing; SOFR+225 / ~5.9%", "2026", "$8.5bn", "Bank / GPU-backed loan", "CoreWeave IR"),
    ("CoreWeave", "$7.5bn debt facility led by Blackstone & Magnetar (largest private debt deal at the time)", "2024", "$7.5bn", "Private credit / loan", "Blackstone"),
    ("CoreWeave", "$2.7bn 1.75% convertible notes; ~$30bn total debt by early 2026; ~$28bn debt+equity raised in 12 mos", "2025-26", "$2.7bn conv.", "Convertibles", "thedig; CoreWeave"),
    ("CoreWeave", "$2bn strategic equity from NVIDIA; IPO Mar 2025 raised ~$1.5bn", "2025-26", "$3.5bn equity", "Equity", "CNBC; CoreWeave"),
    ("SpaceX / xAI", "SpaceX acquired xAI effective Feb 2, 2026; Colossus I & II now SpaceX-owned (vertically-integrated AI platform; Terafab chip JV w/ Tesla & Intel)", "2026", "merger", "Equity / M&A", "SpaceX S-1"),
    ("SpaceX / xAI", "SpaceX AI capex $12.7bn in 2025 (~61% of total) + $7.7bn in Q1 2026; AI buildout funded by Starlink FCF, private raises, planned IPO", "2025-26", "$20bn+ capex", "Internal / equity", "SpaceX S-1; The Verge"),
    ("SpaceX / xAI", "Anthropic compute contract: $1.25bn/month (~$45bn) through May 2029 for Colossus I & II — anchor revenue underwriting infra financing", "2026-29", "~$45bn", "Take-or-pay (demand)", "SpaceX S-1; The Verge"),
    ("SpaceX / xAI", "xAI $10bn raise ($5bn secured notes/term loans incl. Apollo $5bn facility + $5bn equity) via Morgan Stanley", "2025", "$10bn", "Loans + equity", "CNBC"),
    ("SpaceX / xAI", "Colossus 2 SPV (Valor Equity): ~$12.5bn debt + ~$7.5bn equity (NVIDIA up to $2bn) to buy & lease GPUs", "2025-26", "~$20bn", "SPV / vendor lease", "The Information; Yahoo"),
    ("Nebius", "$3bn+ convertibles + bank lines; lighter balance sheet; Yandex spin-off cash ~$2.5bn; NVIDIA ~$2bn equity; Microsoft ~$7bn prepay", "2025-26", "$3bn+", "Convertibles", "frankk research"),
    ("Lambda", "$500m GPU-backed facility from Macquarie", "2025", "$0.5bn", "Bank / GPU-backed loan", "theaiinsider"),
    ("Nscale", "$1.4bn term loan", "2026", "$1.4bn", "Bank loan", "theaiinsider"),
    ("Fluidstack", "Facility backed by ~$6.7bn Google-contracted revenue (Google lease backstop)", "2026", "~$6.7bn", "SPV / private credit", "theaiinsider"),
    ("Applied Digital", "$1.59bn high-yield bond (7%) for CoreWeave-leased campus (Polaris Forge 1)", "2026", "$1.59bn", "Securitization / HY bond", "TNW"),
    ("Aligned Data Centers", "$12bn capital raise ($5bn+ new equity + $7bn debt commitments) led by Macquarie (early 2025); later acquired by MGX/BlackRock consortium (~$40bn)", "2025", "$12bn / $40bn", "Infra equity + debt", "SFA; NYU DRI"),
    ("QTS (Blackstone)", "BX 2025-VOLT SASB CMBS $3.5bn (~10 QTS DCs); + Phoenix ABS (QTS Issuer 2025-1)", "2025", "$3.5bn+", "Securitization (CMBS/ABS)", "CREFC; SFA"),
    ("Switch (DigitalBridge)", "$2.4bn SASB CMBS (SWCH 2025-DATA) + $1.1bn ABS — green bonds", "2024-25", "$3.5bn", "Securitization (CMBS/ABS)", "SFA; CREFC"),
    ("Vantage Data Centers", "First public data-center ABS (2024); ongoing multi-billion securitization & infra-equity program", "2024-26", "multi-$bn", "Securitization / equity", "Dentons"),
    ("Colocation sector", "US data-center ABS+CMBS issuance ~$26bn in 2025 (~10x 2020); ~$57bn since 2021; Morgan Stanley ~$130bn securitized net issuance 2026-28; ~$150bn permanent financing need 2026-27", "2025-28", "~$130bn", "Securitization (ABS/CMBS)", "Impax; CREFC; Morgan Stanley"),
    ("HUMAIN (Saudi PIF)", "PIF national AI champion (announced 13 May 2025): ~$100bn across 11 data centers / 2.2 GW; ~600k NVIDIA GPUs; ALLAM Arabic LLM", "2025-30", "~$100bn", "Sovereign equity", "Presenc AI; NYU DRI"),
    ("Stargate UAE / G42", "$30bn+ 5GW Abu Dhabi campus (Khazna/G42 w/ OpenAI, Oracle, NVIDIA, Cisco, SoftBank); Phase 1 200MW Q3 2026; Microsoft ~$1.5bn equity", "2025-30", "~$30bn+", "Sovereign equity + JV", "Khaleej Times; dcpulse"),
    ("MGX / Mubadala (UAE)", "Anchored Stargate; co-led ~$40bn Aligned Data Centers acquisition w/ BlackRock; AIP (BlackRock-Microsoft) targets up to $100bn global AI infra", "2025-26", "~$40bn+", "Sovereign equity / infra", "NYU DRI"),
    ("Qatar (QIA / Qai)", "~$20bn AI-infrastructure JV with Brookfield (late 2025); earlier positions in xAI & Databricks", "2025", "~$20bn", "Sovereign equity + JV", "NYU DRI"),
    ("Sector-wide", "HY/unrated AI-infra issuers raised ~$107bn funded+committed debt by ~May 2026 ($68.7bn neoclouds / $38.7bn AI-native DCs)", "to 2026", "~$107bn", "All debt", "Octus"),
    ("Sector-wide", "Insurance-linked platforms deployed ~$180bn into private credit in 2025 (~25% of global private credit AUM)", "2025", "~$180bn", "Insurance / private credit", "McKinsey; ABF Journal"),
    ("Sector-wide", "Goldman Sachs: ~$5.3tn AI + data-center capex 2025-2030; Morgan Stanley: ~$800bn private credit for AI DCs 2025-28", "2025-30", "$5.3tn capex", "Context", "Goldman; Morgan Stanley"),
]

# ---------------------------------------------------------------------------
# PRIVATE CREDIT TRANSACTIONS (AI infrastructure) — transaction log.
# fields: (period, borrower / project, segment, lead lenders / arrangers,
#          amount $bn [None for context/framework rows], structure / instrument,
#          pricing, tenor / maturity, collateral / notes, source)
# Amounts are the DEBT / private-credit portion as reported ('~' = approx).
# This log spans private placements (144A), SPV/project debt, GPU-backed DDTLs and
# direct lending; it overlaps the 'Alternate private credit', 'Insurance-linked' and
# parts of the 'Securitization (144A)' buckets in the aggregated model.
# ---------------------------------------------------------------------------
PRIVATE_CREDIT = [
    ("Oct 2025", "Meta — Hyperion (Beignet Investor LLC), Louisiana", "Hyperscaler",
     "PIMCO (~$18bn), BlackRock, Apollo (debt); Blue Owl (equity); MS adviser", 27.3,
     "Off-B/S SPV; 144A senior secured bond", "6.581% / T+225bp", "Due May 2049 (~23.6y)",
     "2.1GW campus; Meta lease + residual-value guarantee (up to ~$28bn); A+ (S&P)", "IFR; FT; Quinn Emanuel"),
    ("2026", "Anthropic — Google TPU chip SPV", "Neocloud / AI-native",
     "Apollo (Atlas SP $0.8bn equity; Athene), Blackstone; Broadcom RVG", 35.0,
     "Chip-leasing SPV; 3 tranches; ~half syndicated", "A1 T+100bp; A2 5.75%; B 8.5%",
     "~5y", "Buys Google TPUs leased to Anthropic; Broadcom residual-value support", "privatedebtnews; Build.inc"),
    ("2025-26", "xAI — Colossus 2 chip SPV (Memphis/Southaven)", "Neocloud / AI-native",
     "Valor Equity (lead), Apollo, Diameter (debt); NVIDIA up to $2bn equity", 12.5,
     "SPV; debt + ~$7.5bn equity to buy & lease GPUs", "~10.5%", "5y (~2.5y avg life)",
     "Buys NVIDIA GPUs leased to xAI; chips as collateral", "FT; WSJ; AAE filing"),
    ("Jun-Jul 2025", "xAI — secured notes / term loan (part of $10bn raise)", "Neocloud / AI-native",
     "Apollo, Diameter; Morgan Stanley arranger", 5.0,
     "Secured notes + term loan", "up to ~12.5%", "~3-5y",
     "Lenders can seize/operate Colossus on default", "CNBC; ciphertalk"),
    ("Mar 2026", "CoreWeave — DDTL 4.0 (Compute Acq. Co. VIII)", "Neocloud / AI-native",
     "Blackstone Credit & Insurance (anchor); MUFG, MS, GS, JPM", 8.5,
     "Non-recourse GPU-backed delayed-draw term loan", "SOFR+225 / ~5.9% fixed", "Mar 2032",
     "First IG-rated (A3 / A low) HPC-collateralized loan", "CoreWeave IR"),
    ("May 2024", "CoreWeave — $7.5bn debt facility", "Neocloud / AI-native",
     "Blackstone, Magnetar (co-leads), Coatue, Carlyle, CDPQ, BlackRock, Eldridge", 7.5,
     "GPU + customer-contract secured facility", "~10%+ (at issue)", "n/a",
     "Largest private debt deal at the time", "Blackstone"),
    ("Aug 2023", "CoreWeave — $2.3bn debt facility", "Neocloud / AI-native",
     "Blackstone, Magnetar", 2.3, "GPU-backed facility", "high", "n/a",
     "Early template for GPU-collateralized debt", "Blackstone"),
    ("Jul 2025", "CoreWeave — OpenAI contract financing", "Neocloud / AI-native",
     "Private lenders (SPV)", 2.6, "SPV debt vs. $11.9bn OpenAI contract", "n/a", "n/a",
     "Funds CoreWeave's OpenAI compute obligations", "FT"),
    ("2025", "Oracle/OpenAI — Abilene, TX (Stargate; Crusoe developer)", "Hyperscaler",
     "Blue Owl + JPMorgan (incl. Crusoe, Primary Digital Infrastructure)", 10.0,
     "Off-B/S SPV (~$13bn incl. equity)", "n/a", "n/a",
     "SPV owns OpenAI Abilene facility", "FT"),
    ("2025-26", "Oracle — Texas + Wisconsin data centers", "Hyperscaler",
     "Private credit consortium", 38.0, "SPV debt package", "n/a", "n/a",
     "Two-site off-B/S financing", "FT"),
    ("2025-26", "Oracle — New Mexico site", "Hyperscaler",
     "Private credit lenders", 18.0, "SPV / project loan", "n/a", "n/a",
     "Single-site project loan", "FT"),
    ("2025", "Oracle — Michigan campus", "Hyperscaler",
     "Project-finance lenders", 16.3, "Project-level SPV financing", "n/a", "n/a",
     "OCI capacity build", "Press reports"),
    ("2026", "Fluidstack — Google-backed facility", "Neocloud / AI-native",
     "Private credit (Google lease backstop)", 6.7, "SPV backed by contracted revenue", "n/a", "n/a",
     "~$6.7bn Google-contracted revenue backstop", "theaiinsider"),
    ("2024", "Applied Digital — Macquarie financing", "Colocation / data-center REIT",
     "Macquarie", 5.0, "Perpetual/structured HPC financing", "n/a", "n/a",
     "Powered-shell build for CoreWeave (Polaris Forge)", "press"),
    ("Early 2025", "Aligned Data Centers — debt commitments", "Colocation / data-center REIT",
     "Macquarie + global investors", 7.0, "Debt within $12bn raise (+$5bn equity)", "n/a", "n/a",
     "Later acquired by MGX/BlackRock (~$40bn)", "SFA; NYU DRI"),
    ("2026", "Nscale — term loan", "Neocloud / AI-native",
     "Private credit", 1.4, "Term loan", "n/a", "n/a", "GPU-cloud build", "theaiinsider"),
    ("2025", "Lambda — GPU-backed facility", "Neocloud / AI-native",
     "Macquarie", 0.5, "GPU-backed facility", "n/a", "n/a", "GPU-cloud build", "theaiinsider"),
    # ---- context / frameworks (not summed into the itemized total) ----
    ("2024", "KKR + Energy Capital Partners — AI infrastructure partnership", "Context / market",
     "KKR, Energy Capital Partners", None, "Strategic partnership/commitment", "n/a", "n/a",
     "~$50bn to accelerate AI infrastructure", "Business Times"),
    ("Late 2025", "Qatar (QIA / Qai) + Brookfield — AI infrastructure JV", "Context / market",
     "QIA / Qai, Brookfield", None, "Joint venture / infra credit", "n/a", "n/a",
     "~$20bn JV", "NYU DRI"),
    ("2025-28", "Broadcom + Apollo + Blackstone — AI compute platform", "Context / market",
     "Apollo, Blackstone, Broadcom", None, "Platform framework", "n/a", "through 2028",
     ">20 GW of compute to be deployed", "privatedebtnews"),
    ("from Nov 2025", "Sector — 144A private placements (data centers)", "Context / market",
     "Multiple (Blue Owl, PIMCO, etc.)", None, "144A private placement bonds", "n/a", "n/a",
     ">$40bn by Applied Digital, CoreWeave, Hut 8, Related since Nov 2025", "Bisnow"),
    ("2025-28", "Market — projected private credit for AI data centers", "Context / market",
     "Industry-wide", None, "Projection", "n/a", "n/a",
     "Morgan Stanley: ~$800bn over next 2 years; >$200bn already outstanding", "Morgan Stanley; Quinn Emanuel"),
]
# fmt: on


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------
def title_cell(ws, cell, text, size=14):
    ws[cell] = text
    ws[cell].font = Font(bold=True, size=size, color=NAVY)


def header_row(ws, row, start_col, labels, fill=BLUE, color=WHITE, wrap=True):
    for i, label in enumerate(labels):
        c = ws.cell(row=row, column=start_col + i, value=label)
        c.font = Font(bold=True, color=color, size=10)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        c.border = BORDER


def num_fmt(cell, fmt="#,##0"):
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right")


def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# Sheet 1: Cover & Methodology
# ---------------------------------------------------------------------------
def build_cover(wb):
    ws = wb.active
    ws.title = "Cover & Methodology"
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 3, "B": 100})

    ws.merge_cells("B2:B2")
    title_cell(ws, "B2", "AI Infrastructure Financing — Hyperscalers & Neoclouds", 18)
    ws["B3"] = "Equity & debt funding to build AI infrastructure, cumulative 2025-2030"
    ws["B3"].font = Font(size=12, italic=True, color=GREY)
    ws["B4"] = "All figures in US$ billions unless noted."
    ws["B4"].font = Font(size=10, color=GREY)

    blocks = [
        ("What this workbook contains", BLUE, [
            "• Summary (by Company): equity / internal funding vs. total debt, with % debt-funded.",
            "• Debt by Type: total debt split into mutually-exclusive instrument buckets.",
            "• Cash-Flow Bridge: operating cash flow -> free cash flow -> external funding gap.",
            "• Capex Projections: annual AI-infrastructure capex 2025-2030 by company.",
            "• Private Credit Deals: transaction log of private-credit / SPV / 144A debt for AI infra.",
            "• Key Deals & Sources: notable disclosed transactions underpinning the estimates.",
        ]),
        ("Scope & definitions", TEAL, [
            "• Hyperscalers: Microsoft, Amazon (AWS), Alphabet (Google), Meta, Oracle (OCI).",
            "• Neoclouds / AI-native infra: GPU-cloud & AI-native data-center builders (CoreWeave,",
            "   SpaceX/xAI, Nebius, IREN, Crusoe, Lambda, Fluidstack, Nscale, + an 'Other' aggregate).",
            "• SpaceX acquired xAI (effective 2 Feb 2026), so Colossus I & II are now SpaceX-owned; the two",
            "   are shown as one consolidated 'SpaceX / xAI' line to avoid double-counting the same clusters.",
            "• Colocation / data-center REITs: wholesale & retail colo / powered-shell owners (Equinix,",
            "   Digital Realty, QTS, Vantage, Aligned, CyrusOne, Switch, Stack, + 'Other') that build campuses",
            "   and lease them to hyperscalers / neoclouds; heavy users of ABS/CMBS securitization & infra equity.",
            "• Sovereign clouds: state / sovereign-wealth-funded national AI infrastructure (HUMAIN-Saudi PIF,",
            "   MGX/Mubadala, Stargate UAE/G42/Khazna, Qatar QIA, EU/France, India, + 'Other'); mostly equity-funded.",
            "• Funding sources now split three ways:  Capex = Operating cash flow (self-funded) +",
            "   External equity + Total debt.",
            "   - Operating cash flow (self-funded): internally-generated cash (CFO) reinvested into AI capex.",
            "   - External equity: IPOs, secondaries, strategic stakes (e.g. NVIDIA), ATM programs, sovereign-",
            "     wealth equity and mandatory convertible PREFERRED (treated as equity).",
            "   - Total debt: external borrowing, split by instrument (see below).",
            "• The 'Cash-Flow Bridge' sheet separates OPERATING CASH FLOW (CFO) and FREE CASH FLOW from the",
            "   equity portion:  CFO - dividends/buybacks/other = self-funded capex; FREE CASH FLOW = CFO -",
            "   AI capex (negative for most builders, which is what drives external equity + debt issuance).",
        ]),
        ("Debt instrument buckets (mutually exclusive — they sum to total debt)", AMBER, [
            "1. Corporate bonds (public IG/HY) — on-balance-sheet senior notes sold in public markets.",
            "2. Bank & syndicated loans — term loans, revolvers, delayed-draw term loans (incl. GPU-backed DDTLs).",
            "3. Alternate private credit — direct lending / private placements by credit funds (Blackstone,",
            "   Apollo, Blue Owl, PIMCO, Magnetar, etc.), including privately-placed SPV/project debt.",
            "4. Insurance-linked lending — debt funded off insurer/annuity balance sheets (Athene/Apollo,",
            "   Global Atlantic/KKR, MetLife, Brookfield, etc.). NOTE: large overlap with private credit &",
            "   bonds; shown separately to size the insurance channel — see double-counting note below.",
            "5. Securitization / ABS / CMBS — secured data-center bonds & GPU-backed asset-backed securities.",
            "6. Convertibles / equity-linked debt — convertible notes (excludes mandatory convert. PREFERRED).",
            "7. Vendor & equipment-lease financing — chip/equipment leases & sale-leasebacks (e.g. NVIDIA",
            "   chip-lease SPVs, finance leases). Microsoft's large finance-lease book sits largely here.",
            "Memo: 'Off-B/S SPV / JV (memo)' flags the portion routed through special-purpose vehicles /",
            "   joint ventures (e.g. Meta-Hyperion, xAI Colossus 2, Oracle project SPVs). It OVERLAPS the",
            "   buckets above and is therefore NOT added into the debt total.",
        ]),
        ("Methodology & caveats", GREY, [
            "• Figures are RESEARCH ESTIMATES for the cumulative 2025-2030 buildout. They blend disclosed",
            "  transactions (2023-early 2026) with forward projections; they are illustrative, not audited.",
            "• Aggregate cross-check: total financing ≈ $5.25tn, consistent with Goldman Sachs' ~$5.3tn",
            "  2025-2030 AI + data-center capex estimate; aggregate debt ≈ $1.8tn, consistent with Morgan",
            "  Stanley / BofA projections of $1-2tn of debt financing for the buildout.",
            "• CROSS-SEGMENT OVERLAP: hyperscalers LEASE much colocation capacity (so colo financing partly",
            "  funds the same capacity hyperscalers report as off-B/S leases), and sovereign-wealth funds",
            "  (MGX, PIF, QIA) also INVEST in neoclouds/labs and colo (e.g. MGX/BlackRock bought Aligned; PIF",
            "  put ~$3bn into xAI). The all-segment 'grand total' is therefore a GROSS figure of capital raised",
            "  across distinct balance sheets and is not strictly additive; treat it as an upper-bound view.",
            "  Hyperscaler + Neocloud financing (~$5.4tn) ties to Goldman's ~$5.3tn big-tech capex; adding",
            "  third-party colo + sovereign brings the gross total toward broader ~$6-7tn AI-infra estimates.",
            "• Insurance-linked lending double counts with private credit/bonds at the capital-source level;",
            "  to avoid inflating totals, each dollar of debt is classified ONCE by its primary instrument and",
            "  the insurance bucket captures only directly insurer-originated / insurance-balance-sheet debt.",
            "• Company splits reflect observed behavior: Microsoft/Alphabet/Amazon are FCF-heavy with",
            "  selective bonds; Meta & Oracle lean more on debt & SPVs; neoclouds are debt-heavy (asset-backed).",
            "• Built " + "with openpyxl. See Key Deals & Sources sheet for citations.",
        ]),
    ]
    r = 6
    for heading, color, lines in blocks:
        c = ws.cell(row=r, column=2, value=heading)
        c.font = Font(bold=True, size=12, color=WHITE)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[r].height = 20
        r += 1
        for line in lines:
            cc = ws.cell(row=r, column=2, value=line)
            cc.font = Font(size=10, color="333333")
            cc.alignment = Alignment(wrap_text=False, indent=1)
            r += 1
        r += 1

    disclaimer = ("Disclaimer: For analytical/illustrative purposes only. Not investment advice and not a "
                  "representation of any company's actual financing. Verify against primary filings before use.")
    c = ws.cell(row=r + 1, column=2, value=disclaimer)
    c.font = Font(size=9, italic=True, color="999999")
    c.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r + 1].height = 30


# ---------------------------------------------------------------------------
# Sheet 2: Summary (by Company)
# ---------------------------------------------------------------------------
def total_debt(debt: dict) -> int:
    return sum(debt.values())


def build_summary(wb):
    ws = wb.create_sheet("Summary (by Company)")
    ws.sheet_view.showGridLines = False
    title_cell(ws, "A1", "Summary — Funding Sources, Cumulative 2025-2030 (US$bn)")
    ws["A2"] = ("Funding = Operating cash flow (self-funded) + External equity + Total debt. "
                "Sorted within segment by total financing. % Debt = total debt / total financing.")
    ws["A2"].font = Font(italic=True, size=9, color=GREY)

    # columns: 1 Company 2 Segment 3 OCF 4 ExtEquity 5 Debt 6 Total 7 %Debt 8 SPV memo
    cols = ["Company", "Segment", "Operating cash\nflow (self-funded)", "External\nequity",
            "Total debt\nfinancing", "Total\nfinancing", "% Debt-\nfunded", "Off-B/S SPV/JV\n(memo)"]
    NCOL = len(cols)
    hdr = 4
    header_row(ws, hdr, 1, cols)
    ws.row_dimensions[hdr].height = 32

    r = hdr + 1
    seg_rows = {s[0]: [] for s in SEGMENTS}

    def write_company(name, seg, equity, debt, spv, fill):
        nonlocal r
        td = total_debt(debt)
        ocf = ocf_self(name, equity)
        ext = EXTERNAL_EQUITY[name]
        tot = equity + td
        ws.cell(row=r, column=1, value=name).font = Font(bold=False, size=10)
        ws.cell(row=r, column=2, value=seg).font = Font(size=9, color=GREY)
        num_fmt(ws.cell(row=r, column=3, value=ocf))
        ws.cell(row=r, column=3).font = Font(color=TEAL)
        num_fmt(ws.cell(row=r, column=4, value=ext))
        ws.cell(row=r, column=4).font = Font(color="2E7D32")
        num_fmt(ws.cell(row=r, column=5, value=td))
        cc = ws.cell(row=r, column=6, value=f"=C{r}+D{r}+E{r}")
        num_fmt(cc)
        cc.font = Font(bold=True)
        pct = ws.cell(row=r, column=7, value=f"=E{r}/F{r}")
        pct.number_format = "0%"
        pct.alignment = Alignment(horizontal="right")
        num_fmt(ws.cell(row=r, column=8, value=spv))
        ws.cell(row=r, column=8).font = Font(italic=True, color=GREY)
        for col in range(1, NCOL + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = BORDER
            if cell.fill.patternType is None:
                cell.fill = PatternFill("solid", fgColor=fill)
        seg_rows[seg].append(r)
        r += 1

    sum_cols = [3, 4, 5, 6, 8]  # columns to SUM in subtotal/grand-total
    for seg, fill in SEGMENTS:
        sect = ws.cell(row=r, column=1, value=seg.upper())
        sect.font = Font(bold=True, color=WHITE, size=10)
        for col in range(1, NCOL + 1):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=NAVY)
        r += 1
        members = [c for c in COMPANIES if c[1] == seg]
        members.sort(key=lambda c: c[2] + total_debt(c[3]), reverse=True)
        for name, s, equity, debt, spv in members:
            write_company(name, s, equity, debt, spv, fill)
        # subtotal
        sub = ws.cell(row=r, column=1, value=f"{seg} subtotal")
        sub.font = Font(bold=True, size=10, color=NAVY)
        rows = seg_rows[seg]
        for col in sum_cols:
            letter = get_column_letter(col)
            cc = ws.cell(row=r, column=col, value=f"=SUM({letter}{rows[0]}:{letter}{rows[-1]})")
            num_fmt(cc)
            cc.font = Font(bold=True, color=NAVY)
        pc = ws.cell(row=r, column=7, value=f"=E{r}/F{r}")
        pc.number_format = "0%"
        pc.font = Font(bold=True, color=NAVY)
        pc.alignment = Alignment(horizontal="right")
        for col in range(1, NCOL + 1):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=LIGHT_GREY)
            ws.cell(row=r, column=col).border = BORDER
        r += 1

    # grand total
    gt = ws.cell(row=r, column=1, value="GRAND TOTAL — all companies")
    gt.font = Font(bold=True, size=11, color=WHITE)
    comp_rows = [rr for lst in seg_rows.values() for rr in lst]
    for col in sum_cols:
        letter = get_column_letter(col)
        ref = ",".join(f"{letter}{rr}" for rr in comp_rows)
        cc = ws.cell(row=r, column=col, value=f"=SUM({ref})")
        num_fmt(cc)
        cc.font = Font(bold=True, color=WHITE)
    pc = ws.cell(row=r, column=7, value=f"=E{r}/F{r}")
    pc.number_format = "0%"
    pc.font = Font(bold=True, color=WHITE)
    pc.alignment = Alignment(horizontal="right")
    for col in range(1, NCOL + 1):
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=BLUE)
        ws.cell(row=r, column=col).border = BORDER
    grand_row = r

    set_widths(ws, {"A": 32, "B": 19, "C": 15, "D": 11, "E": 13, "F": 12, "G": 10, "H": 14})
    ws.freeze_panes = "A5"

    # Chart: OCF / external equity / debt by company (stacked)
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = "Funding Mix by Company (2025-2030, $bn)"
    chart.height = 9
    chart.width = 28
    hx = 10  # helper table column J
    labels = ["Company", "Op. cash flow", "External equity", "Debt"]
    for i, lab in enumerate(labels):
        ws.cell(row=hdr, column=hx + i, value=lab).font = Font(bold=True, size=9, color=GREY)
    hr = hdr + 1
    for name, s, equity, debt, spv in COMPANIES:
        ws.cell(row=hr, column=hx, value=name)
        ws.cell(row=hr, column=hx + 1, value=ocf_self(name, equity)).number_format = "#,##0"
        ws.cell(row=hr, column=hx + 2, value=EXTERNAL_EQUITY[name]).number_format = "#,##0"
        ws.cell(row=hr, column=hx + 3, value=total_debt(debt)).number_format = "#,##0"
        hr += 1
    data = Reference(ws, min_col=hx + 1, max_col=hx + 3, min_row=hdr, max_row=hr - 1)
    cats = Reference(ws, min_col=hx, max_col=hx, min_row=hdr + 1, max_row=hr - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.y_axis.title = "US$bn"
    ws.add_chart(chart, f"A{grand_row + 3}")
    for row in ws.iter_rows(min_row=hdr + 1, max_row=hr - 1, min_col=hx, max_col=hx + 3):
        for cell in row:
            cell.font = Font(size=8, color="BBBBBB")


# ---------------------------------------------------------------------------
# Sheet 3: Debt by Type (the main breakdown)
# ---------------------------------------------------------------------------
def build_debt_breakdown(wb):
    ws = wb.create_sheet("Debt by Type")
    ws.sheet_view.showGridLines = False
    title_cell(ws, "A1", "Debt Financing by Instrument Type — Cumulative 2025-2030 (US$bn)")
    ws["A2"] = ("Instrument buckets are mutually exclusive and sum to 'Total debt'. "
                "'Off-B/S SPV/JV' is a memo (overlaps the buckets; not added).")
    ws["A2"].font = Font(italic=True, size=9, color=GREY)

    cols = ["Company", "Segment"] + DEBT_COLS + ["Total debt", "Off-B/S SPV/JV\n(memo)",
                                                 "Op. cash flow\n(self-funded)", "External\nequity"]
    hdr = 4
    header_row(ws, hdr, 1, cols)
    ws.row_dimensions[hdr].height = 46

    n_debt = len(DEBT_COLS)
    first_debt_col = 3
    last_debt_col = first_debt_col + n_debt - 1
    total_col = last_debt_col + 1            # Total debt
    spv_col = total_col + 1                   # SPV memo
    ocf_col = spv_col + 1                     # Operating cash flow (self-funded)
    ext_col = ocf_col + 1                     # External equity
    equity_col = ext_col                      # last column (alias for layout loops)

    r = hdr + 1
    seg_rows = {s[0]: [] for s in SEGMENTS}

    for seg, fill in SEGMENTS:
        sect = ws.cell(row=r, column=1, value=seg.upper())
        sect.font = Font(bold=True, color=WHITE, size=10)
        for col in range(1, equity_col + 1):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=NAVY)
        r += 1
        members = [c for c in COMPANIES if c[1] == seg]
        members.sort(key=lambda c: total_debt(c[3]), reverse=True)
        for name, s, equity, debt, spv in members:
            ws.cell(row=r, column=1, value=name).font = Font(size=10)
            ws.cell(row=r, column=2, value=s).font = Font(size=9, color=GREY)
            for i, key in enumerate(DEBT_COLS):
                cc = ws.cell(row=r, column=first_debt_col + i, value=debt[key])
                num_fmt(cc)
            tcell = ws.cell(row=r, column=total_col,
                            value=f"=SUM({get_column_letter(first_debt_col)}{r}:{get_column_letter(last_debt_col)}{r})")
            num_fmt(tcell)
            tcell.font = Font(bold=True)
            num_fmt(ws.cell(row=r, column=spv_col, value=spv))
            ws.cell(row=r, column=spv_col).font = Font(italic=True, color=GREY)
            num_fmt(ws.cell(row=r, column=ocf_col, value=ocf_self(name, equity)))
            ws.cell(row=r, column=ocf_col).font = Font(color=TEAL)
            num_fmt(ws.cell(row=r, column=ext_col, value=EXTERNAL_EQUITY[name]))
            ws.cell(row=r, column=ext_col).font = Font(color="2E7D32")
            for col in range(1, equity_col + 1):
                cell = ws.cell(row=r, column=col)
                cell.border = BORDER
                if cell.fill.patternType is None:
                    cell.fill = PatternFill("solid", fgColor=fill)
            seg_rows[seg].append(r)
            r += 1
        # subtotal
        ws.cell(row=r, column=1, value=f"{seg} subtotal").font = Font(bold=True, color=NAVY)
        rows = seg_rows[seg]
        for col in range(first_debt_col, equity_col + 1):
            letter = get_column_letter(col)
            cc = ws.cell(row=r, column=col, value=f"=SUM({letter}{rows[0]}:{letter}{rows[-1]})")
            num_fmt(cc)
            cc.font = Font(bold=True, color=NAVY)
        for col in range(1, equity_col + 1):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=LIGHT_GREY)
            ws.cell(row=r, column=col).border = BORDER
        r += 1

    # grand total
    ws.cell(row=r, column=1, value="GRAND TOTAL").font = Font(bold=True, color=WHITE, size=11)
    comp_rows = [rr for lst in seg_rows.values() for rr in lst]
    for col in range(first_debt_col, equity_col + 1):
        letter = get_column_letter(col)
        ref = ",".join(f"{letter}{rr}" for rr in comp_rows)
        cc = ws.cell(row=r, column=col, value=f"=SUM({ref})")
        num_fmt(cc)
        cc.font = Font(bold=True, color=WHITE)
    for col in range(1, equity_col + 1):
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=BLUE)
        ws.cell(row=r, column=col).border = BORDER
    grand_row = r

    # % of total debt row
    r += 1
    ws.cell(row=r, column=1, value="% of total debt").font = Font(bold=True, italic=True, color=AMBER)
    for col in range(first_debt_col, total_col):
        letter = get_column_letter(col)
        cc = ws.cell(row=r, column=col,
                     value=f"={letter}{grand_row}/${get_column_letter(total_col)}${grand_row}")
        cc.number_format = "0%"
        cc.font = Font(italic=True, color=AMBER)
        cc.alignment = Alignment(horizontal="right")
    cc = ws.cell(row=r, column=total_col, value=f"={get_column_letter(total_col)}{grand_row}/{get_column_letter(total_col)}{grand_row}")
    cc.number_format = "0%"
    cc.font = Font(italic=True, color=AMBER, bold=True)
    cc.alignment = Alignment(horizontal="right")

    widths = {"A": 30, "B": 19}
    for i in range(n_debt):
        widths[get_column_letter(first_debt_col + i)] = 15
    widths[get_column_letter(total_col)] = 11
    widths[get_column_letter(spv_col)] = 13
    widths[get_column_letter(ocf_col)] = 14
    widths[get_column_letter(ext_col)] = 11
    set_widths(ws, widths)
    ws.freeze_panes = "C5"

    # Chart: debt mix by instrument (grand total)
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Aggregate Debt Mix by Instrument (2025-2030, $bn)"
    chart.height = 8
    chart.width = 22
    data = Reference(ws, min_col=first_debt_col, max_col=last_debt_col,
                     min_row=grand_row, max_row=grand_row)
    cats = Reference(ws, min_col=first_debt_col, max_col=last_debt_col, min_row=hdr, max_row=hdr)
    chart.add_data(data, from_rows=True, titles_from_data=False)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, f"A{r + 3}")


# ---------------------------------------------------------------------------
# Sheet 4: Cash-Flow Bridge (operating cash flow -> free cash flow -> funding gap)
# ---------------------------------------------------------------------------
def build_cashflow_bridge(wb):
    ws = wb.create_sheet("Cash-Flow Bridge")
    ws.sheet_view.showGridLines = False
    title_cell(ws, "A1", "Cash-Flow Bridge — Operating Cash Flow to Funding Gap, 2025-2030 (US$bn)")
    ws["A2"] = ("Operating cash flow (CFO) less shareholder returns/other = self-funded capex; the "
                "remaining AI capex is met by external equity + debt. Free cash flow = CFO - AI capex.")
    ws["A2"].font = Font(italic=True, size=9, color=GREY)

    cols = ["Company", "Segment", "Operating cash\nflow (CFO)",
            "Less: dividends,\nbuybacks & other", "= Op. cash flow\nself-funded",
            "External\nequity", "Total debt", "= AI-infra capex\nfunded",
            "Memo: Free cash\nflow (CFO - capex)"]
    NCOL = len(cols)
    hdr = 4
    header_row(ws, hdr, 1, cols)
    ws.row_dimensions[hdr].height = 34

    r = hdr + 1
    seg_rows = {s[0]: [] for s in SEGMENTS}
    sum_cols = [3, 4, 5, 6, 7, 8, 9]

    def style_row(row, fill, border=True):
        for col in range(1, NCOL + 1):
            cell = ws.cell(row=row, column=col)
            if border:
                cell.border = BORDER
            if cell.fill.patternType is None:
                cell.fill = PatternFill("solid", fgColor=fill)

    for seg, fill in SEGMENTS:
        ws.cell(row=r, column=1, value=seg.upper()).font = Font(bold=True, color=WHITE, size=10)
        for col in range(1, NCOL + 1):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=NAVY)
        r += 1
        members = [c for c in COMPANIES if c[1] == seg]
        members.sort(key=lambda c: c[2] + total_debt(c[3]), reverse=True)
        for name, s, equity, debt, spv in members:
            cfo = CFO_TOTAL[name]
            ocf = ocf_self(name, equity)
            ext = EXTERNAL_EQUITY[name]
            td = total_debt(debt)
            ws.cell(row=r, column=1, value=name).font = Font(size=10)
            ws.cell(row=r, column=2, value=s).font = Font(size=9, color=GREY)
            num_fmt(ws.cell(row=r, column=3, value=cfo))
            num_fmt(ws.cell(row=r, column=4, value=f"=C{r}-E{r}"))
            num_fmt(ws.cell(row=r, column=5, value=ocf))
            ws.cell(row=r, column=5).font = Font(color=TEAL)
            num_fmt(ws.cell(row=r, column=6, value=ext))
            ws.cell(row=r, column=6).font = Font(color="2E7D32")
            num_fmt(ws.cell(row=r, column=7, value=td))
            cc = ws.cell(row=r, column=8, value=f"=E{r}+F{r}+G{r}")
            num_fmt(cc); cc.font = Font(bold=True)
            fcf = ws.cell(row=r, column=9, value=f"=C{r}-H{r}")
            num_fmt(fcf)
            fcf.font = Font(color="B00020")  # red-ish; negatives shown in parens via format
            fcf.number_format = "#,##0;(#,##0)"
            style_row(r, fill)
            seg_rows[seg].append(r)
            r += 1
        # subtotal
        ws.cell(row=r, column=1, value=f"{seg} subtotal").font = Font(bold=True, color=NAVY)
        rows = seg_rows[seg]
        for col in sum_cols:
            letter = get_column_letter(col)
            cc = ws.cell(row=r, column=col, value=f"=SUM({letter}{rows[0]}:{letter}{rows[-1]})")
            num_fmt(cc); cc.font = Font(bold=True, color=NAVY)
            if col == 9:
                cc.number_format = "#,##0;(#,##0)"
        style_row(r, LIGHT_GREY)
        r += 1

    # grand total
    ws.cell(row=r, column=1, value="GRAND TOTAL").font = Font(bold=True, color=WHITE, size=11)
    comp_rows = [rr for lst in seg_rows.values() for rr in lst]
    for col in sum_cols:
        letter = get_column_letter(col)
        ref = ",".join(f"{letter}{rr}" for rr in comp_rows)
        cc = ws.cell(row=r, column=col, value=f"=SUM({ref})")
        num_fmt(cc); cc.font = Font(bold=True, color=WHITE)
        if col == 9:
            cc.number_format = "#,##0;(#,##0)"
    for col in range(1, NCOL + 1):
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=BLUE)
        ws.cell(row=r, column=col).border = BORDER

    set_widths(ws, {"A": 32, "B": 19, "C": 13, "D": 14, "E": 14, "F": 11,
                    "G": 11, "H": 14, "I": 15})
    ws.freeze_panes = "C5"

    note = ("Note: hyperscaler CFO greatly exceeds self-funded capex because much CFO is paid out as "
            "dividends/buybacks — which is why even FCF-positive names still issue debt. Neoclouds, "
            "colo and sovereigns generate little/no operating cash flow yet, so capex is met almost "
            "entirely by external equity + debt (deeply negative free cash flow during the buildout).")
    nr = r + 2
    ws.cell(row=nr, column=1, value=note).font = Font(size=9, italic=True, color=GREY)
    ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=NCOL)
    ws.cell(row=nr, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[nr].height = 42


# ---------------------------------------------------------------------------
# Sheet 5: Capex Projections
# ---------------------------------------------------------------------------
def build_capex(wb):
    ws = wb.create_sheet("Capex Projections")
    ws.sheet_view.showGridLines = False
    title_cell(ws, "A1", "AI-Infrastructure Capex Projections, 2025-2030 (US$bn)")
    ws["A2"] = ("Estimated annual capital expenditure directed at AI infrastructure (data centers, GPUs, "
                "networking, power). Cumulative ≈ total financing on the Summary sheet.")
    ws["A2"].font = Font(italic=True, size=9, color=GREY)

    cols = ["Company", "Segment"] + [str(y) for y in CAPEX_YEARS] + ["Cumulative\n2025-30"]
    NCOL = len(cols)
    yr0 = 3                       # first year column (C)
    yrN = yr0 + len(CAPEX_YEARS) - 1
    cum_col = yrN + 1
    hdr = 4
    header_row(ws, hdr, 1, cols)
    ws.row_dimensions[hdr].height = 28

    r = hdr + 1
    seg_rows = {s[0]: [] for s in SEGMENTS}
    subtotal_rows = []

    for seg, fill in SEGMENTS:
        ws.cell(row=r, column=1, value=seg.upper()).font = Font(bold=True, color=WHITE, size=10)
        for col in range(1, NCOL + 1):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=NAVY)
        r += 1
        for name, annual in dict(build_capex_rows())[seg]:
            ws.cell(row=r, column=1, value=name).font = Font(size=10)
            ws.cell(row=r, column=2, value=seg).font = Font(size=9, color=GREY)
            for i, v in enumerate(annual):
                num_fmt(ws.cell(row=r, column=yr0 + i, value=v))
            cc = ws.cell(row=r, column=cum_col,
                         value=f"=SUM({get_column_letter(yr0)}{r}:{get_column_letter(yrN)}{r})")
            num_fmt(cc); cc.font = Font(bold=True)
            for col in range(1, NCOL + 1):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=fill)
                ws.cell(row=r, column=col).border = BORDER
            seg_rows[seg].append(r)
            r += 1
        # segment subtotal
        ws.cell(row=r, column=1, value=f"{seg} subtotal").font = Font(bold=True, color=NAVY)
        rows = seg_rows[seg]
        for col in range(yr0, cum_col + 1):
            letter = get_column_letter(col)
            cc = ws.cell(row=r, column=col, value=f"=SUM({letter}{rows[0]}:{letter}{rows[-1]})")
            num_fmt(cc); cc.font = Font(bold=True, color=NAVY)
        for col in range(1, NCOL + 1):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=LIGHT_GREY)
            ws.cell(row=r, column=col).border = BORDER
        subtotal_rows.append((seg, r))
        r += 1

    # grand total
    ws.cell(row=r, column=1, value="TOTAL — all companies").font = Font(bold=True, color=WHITE, size=11)
    comp_rows = [rr for lst in seg_rows.values() for rr in lst]
    for col in range(yr0, cum_col + 1):
        letter = get_column_letter(col)
        ref = ",".join(f"{letter}{rr}" for rr in comp_rows)
        cc = ws.cell(row=r, column=col, value=f"=SUM({ref})")
        num_fmt(cc); cc.font = Font(bold=True, color=WHITE)
    for col in range(1, NCOL + 1):
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=BLUE)
        ws.cell(row=r, column=col).border = BORDER
    total_row = r

    widths = {"A": 34, "B": 19, get_column_letter(cum_col): 13}
    for col in range(yr0, yrN + 1):
        widths[get_column_letter(col)] = 9
    set_widths(ws, widths)
    ws.freeze_panes = "C5"

    # Chart: annual capex by SEGMENT (subtotals) — cleaner than 30+ company series
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = "Annual AI-Infrastructure Capex by Segment ($bn)"
    chart.height = 9
    chart.width = 24
    for seg, srow in subtotal_rows:
        sdata = Reference(ws, min_col=yr0, max_col=yrN, min_row=srow, max_row=srow)
        series_ref = Reference(ws, min_col=1, max_col=1, min_row=srow, max_row=srow)
        chart.add_data(sdata, from_rows=True)
    # set series names + categories
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.chart.data_source import StrRef
    for i, (seg, srow) in enumerate(subtotal_rows):
        chart.series[i].tx = SeriesLabel(strRef=StrRef(f"'{ws.title}'!$A${srow}"))
    chart.set_categories(Reference(ws, min_col=yr0, max_col=yrN, min_row=hdr, max_row=hdr))
    chart.y_axis.title = "US$bn"
    ws.add_chart(chart, f"A{total_row + 3}")


# ---------------------------------------------------------------------------
# Sheet 6: Private Credit Transactions
# ---------------------------------------------------------------------------
def build_private_credit(wb):
    ws = wb.create_sheet("Private Credit Deals")
    ws.sheet_view.showGridLines = False
    title_cell(ws, "A1", "Private Credit Transactions — AI Infrastructure (US$bn)")
    ws["A2"] = ("Transaction log of private-credit / SPV / private-placement (144A) debt for AI infra. "
                "Itemized deals are summed; 'Context / market' rows (frameworks, projections) are not.")
    ws["A2"].font = Font(italic=True, size=9, color=GREY)

    cols = ["Period", "Borrower / Project", "Segment", "Lead lenders / arrangers",
            "Debt amount\n($bn)", "Structure / instrument", "Pricing",
            "Tenor /\nmaturity", "Collateral / notes", "Source(s)"]
    NCOL = len(cols)
    hdr = 4
    header_row(ws, hdr, 1, cols, fill=TEAL)
    ws.row_dimensions[hdr].height = 30

    seg_color = {
        "Hyperscaler": HYPER_FILL, "Neocloud / AI-native": NEO_FILL,
        "Colocation / data-center REIT": COLO_FILL, "Sovereign cloud": SOV_FILL,
        "Context / market": "F0F0F0",
    }

    deals = list(PRIVATE_CREDIT)
    itemized = [d for d in deals if d[4] is not None]
    context = [d for d in deals if d[4] is None]
    itemized.sort(key=lambda d: d[4], reverse=True)

    r = hdr + 1
    item_rows = []

    def write_deal(d, dim=False):
        nonlocal r
        period, borrower, seg, lenders, amt, structure, pricing, tenor, notes, src = d
        ws.cell(row=r, column=1, value=period).font = Font(size=9)
        ws.cell(row=r, column=2, value=borrower).font = Font(size=9, bold=not dim)
        ws.cell(row=r, column=3, value=seg).font = Font(size=8, color=GREY)
        ws.cell(row=r, column=4, value=lenders).font = Font(size=9)
        if amt is not None:
            num_fmt(ws.cell(row=r, column=5, value=amt), "#,##0.0")
            ws.cell(row=r, column=5).font = Font(size=9, bold=True)
        else:
            c = ws.cell(row=r, column=5, value="—")
            c.alignment = Alignment(horizontal="right")
            c.font = Font(size=9, color=GREY)
        ws.cell(row=r, column=6, value=structure).font = Font(size=9)
        ws.cell(row=r, column=7, value=pricing).font = Font(size=9)
        ws.cell(row=r, column=8, value=tenor).font = Font(size=9)
        ws.cell(row=r, column=9, value=notes).font = Font(size=9)
        ws.cell(row=r, column=10, value=src).font = Font(size=9, color=BLUE)
        fill = seg_color.get(seg, WHITE)
        for col in range(1, NCOL + 1):
            cell = ws.cell(row=r, column=col)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(wrap_text=True, vertical="top",
                                       horizontal="right" if col == 5 else "left")
            cell.border = BORDER
        ws.row_dimensions[r].height = 30
        item_rows.append(r)
        r += 1

    for d in itemized:
        write_deal(d)

    # itemized total
    ws.cell(row=r, column=2, value="Total — itemized private credit deals").font = Font(bold=True, color=NAVY)
    tot = ws.cell(row=r, column=5, value=f"=SUM(E{item_rows[0]}:E{item_rows[-1]})")
    num_fmt(tot, "#,##0.0"); tot.font = Font(bold=True, color=NAVY)
    for col in range(1, NCOL + 1):
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=TOTAL_FILL)
        ws.cell(row=r, column=col).border = BORDER
    r += 2

    # context section
    ws.cell(row=r, column=1, value="CONTEXT / FRAMEWORKS / PROJECTIONS (not summed)").font = \
        Font(bold=True, color=WHITE, size=10)
    for col in range(1, NCOL + 1):
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=NAVY)
    r += 1
    for d in context:
        write_deal(d, dim=True)

    set_widths(ws, {"A": 11, "B": 30, "C": 16, "D": 30, "E": 11, "F": 22,
                    "G": 14, "H": 13, "I": 38, "J": 18})
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# Sheet 7: Key Deals & Sources
# ---------------------------------------------------------------------------
def build_deals(wb):
    ws = wb.create_sheet("Key Deals & Sources")
    ws.sheet_view.showGridLines = False
    title_cell(ws, "A1", "Key Disclosed Transactions & Sources")
    ws["A2"] = ("Representative deals underpinning the estimates. Amounts as reported; '~' denotes "
                "approximate/announced figures. Not exhaustive.")
    ws["A2"].font = Font(italic=True, size=9, color=GREY)

    cols = ["Company / Group", "Transaction", "Period", "Amount", "Primary instrument", "Source(s)"]
    hdr = 4
    header_row(ws, hdr, 1, cols)
    ws.row_dimensions[hdr].height = 18

    r = hdr + 1
    for i, (co, desc, period, amt, instr, src) in enumerate(DEALS):
        ws.cell(row=r, column=1, value=co).font = Font(size=9, bold=True)
        ws.cell(row=r, column=2, value=desc).font = Font(size=9)
        ws.cell(row=r, column=3, value=period).font = Font(size=9)
        ws.cell(row=r, column=4, value=amt).font = Font(size=9)
        ws.cell(row=r, column=5, value=instr).font = Font(size=9)
        ws.cell(row=r, column=6, value=src).font = Font(size=9, color=BLUE)
        fill = WHITE if i % 2 == 0 else LIGHT_GREY
        for col in range(1, 7):
            cell = ws.cell(row=r, column=col)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDER
        ws.row_dimensions[r].height = 30
        r += 1

    set_widths(ws, {"A": 20, "B": 64, "C": 10, "D": 14, "E": 22, "F": 22})
    ws.freeze_panes = "A5"

    # source list footer
    r += 1
    ws.cell(row=r, column=1, value="Selected sources:").font = Font(bold=True, size=10, color=NAVY)
    r += 1
    sources = [
        "Goldman Sachs Research — ~$5.3tn AI + data-center capex 2025-2030; private markets in DC financing.",
        "Morgan Stanley — ~$2tn capex 2025-28 with >$1tn debt; ~$800bn private credit for AI DCs.",
        "Bank of America / BofA Securities — hyperscaler bond issuance ($121bn 2025; ~$175bn 2026E).",
        "Moody's Ratings — hyperscaler capex (~$785bn 2026, ~$1tn 2027); $662bn off-B/S DC leases.",
        "McKinsey — ~$2.7tn US (5.2tn global) data-center capex by 2030; insurance-linked private credit.",
        "Octus — HY/unrated AI-infra debt (~$107bn to ~May 2026); ~14GW unfunded ≈ $344bn need.",
        "Company disclosures & press: Meta IR / FT / S&P (Hyperion); Oracle IR / IFR / Reuters; CoreWeave IR;",
        "  CNBC / WSJ / The Information (xAI); Blackstone; Apollo Academy; MUFG; CreditSights; CNA.",
    ]
    for s in sources:
        ws.cell(row=r, column=1, value=s).font = Font(size=9, color="444444")
        r += 1


# ---------------------------------------------------------------------------
def main():
    wb = Workbook()
    # force Excel/Sheets to recalculate all formulas when the file is opened
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    build_cover(wb)
    build_summary(wb)
    build_debt_breakdown(wb)
    build_cashflow_bridge(wb)
    build_capex(wb)
    build_private_credit(wb)
    build_deals(wb)

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       "AI_Infrastructure_Financing_2025-2030.xlsx")
    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
