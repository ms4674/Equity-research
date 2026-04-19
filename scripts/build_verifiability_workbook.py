"""
Build an XLSX workbook comparing the verifiability of AI-agent tasks in
BFSI (banking, payments, finance) against general software-engineering /
coding agents.

Output: data/bfsi_vs_coding_agent_verifiability.xlsx

Sheets:
  1. README                - methodology, rubric, headline finding, disclaimer
  2. Verifiability_Rubric  - the 6-dimension scoring framework
  3. BFSI_Tasks            - finance-domain agent tasks scored on verifiability
  4. Coding_Tasks          - general coding-agent tasks scored on verifiability
  5. SideBySide            - aggregate comparison + ranked combined view
  6. Benchmarks            - public benchmarks / oracles referenced per domain
  7. Sources               - bibliography of cited public sources

Verifiability is defined here as: the degree to which an agent's output for a
task can be checked, after the fact, by an automated or near-automated oracle
that returns a clear pass / fail signal usable for evaluation, RL training, or
production gating.

Nothing in this workbook is investment advice.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


OUT_PATH = Path(__file__).resolve().parents[1] / "data" / "bfsi_vs_coding_agent_verifiability.xlsx"


# ---------------------------------------------------------------------------
# Verifiability rubric
# ---------------------------------------------------------------------------
# Six dimensions, each scored 1 (low) - 5 (high). Composite is a weighted
# average. Higher composite = more verifiable.

RUBRIC: list[dict] = [
    {
        "Dimension": "GroundTruthAvailability",
        "Weight": 0.25,
        "What_it_measures": (
            "Does a single, agreed reference answer exist (or can it be cheaply produced)?"
        ),
        "High_score_example": "Unit-test passes / fails; a recomputed P&L matches GL to the cent.",
        "Low_score_example": "Open-ended advice, narrative SAR, customer-empathy email.",
    },
    {
        "Dimension": "AutomatedOracleStrength",
        "Weight": 0.25,
        "What_it_measures": (
            "Can the check be run by a fast, deterministic program (compiler, test suite, "
            "schema validator, recompute) without a human?"
        ),
        "High_score_example": "pytest exit code, mypy clean, JSON-schema validation, SQL row diff.",
        "Low_score_example": "LLM-as-judge over subjective tone or 'reasonableness'.",
    },
    {
        "Dimension": "OutputDeterminism",
        "Weight": 0.15,
        "What_it_measures": (
            "Is the correct output well-defined and stable, or is there a wide manifold of "
            "acceptable answers?"
        ),
        "High_score_example": "Reconciliation result, regex output, parsed amount.",
        "Low_score_example": "Marketing copy, equity-research narrative.",
    },
    {
        "Dimension": "MeasurementLatency",
        "Weight": 0.10,
        "What_it_measures": (
            "How quickly can a verdict be produced after the agent acts? Lower latency = "
            "tighter feedback loop for evals, RL, gating."
        ),
        "High_score_example": "Test suite in seconds; transaction posts immediately.",
        "Low_score_example": "Credit decision quality only verifiable months later via defaults.",
    },
    {
        "Dimension": "RegulatoryAuditability",
        "Weight": 0.10,
        "What_it_measures": (
            "Can the verification artifact serve as evidence to a regulator, auditor, or "
            "model-risk-management committee?"
        ),
        "High_score_example": "Reproducible recomputation of regulatory ratio (LCR, RWA).",
        "Low_score_example": "Subjective summary 'looks reasonable to a reviewer'.",
    },
    {
        "Dimension": "RewardSignalCleanliness",
        "Weight": 0.15,
        "What_it_measures": (
            "Suitability of the verification signal as a training reward (sparse vs dense, "
            "low-noise, low reward-hacking surface)."
        ),
        "High_score_example": "Compile + unit tests + lint; re-execution of SQL matches expected.",
        "Low_score_example": "Human satisfaction surveys; downstream business KPI.",
    },
]

WEIGHTS = {r["Dimension"]: r["Weight"] for r in RUBRIC}


# ---------------------------------------------------------------------------
# BFSI agent tasks
# ---------------------------------------------------------------------------
# Scores are 1 (low) - 5 (high) on each rubric dimension.

BFSI_TASKS: list[dict] = [
    {
        "Task": "Trade reconciliation (front-to-back match against custodian / clearing file)",
        "Category": "Capital markets ops",
        "ExampleOracle": "Deterministic key-by-key compare of expected vs received ledger; diff produces pass/fail",
        "GroundTruthAvailability": 5,
        "AutomatedOracleStrength": 5,
        "OutputDeterminism": 5,
        "MeasurementLatency": 5,
        "RegulatoryAuditability": 5,
        "RewardSignalCleanliness": 5,
        "Notes": "Effectively a unit-test analog for finance; very high-quality reward signal.",
    },
    {
        "Task": "Payments rail formatting & validation (ISO 20022, SWIFT MT/MX, NACHA, SEPA)",
        "Category": "Payments",
        "ExampleOracle": "Schema validation + scheme-specific rule engine (e.g., XSD + business rules)",
        "GroundTruthAvailability": 5,
        "AutomatedOracleStrength": 5,
        "OutputDeterminism": 5,
        "MeasurementLatency": 5,
        "RegulatoryAuditability": 4,
        "RewardSignalCleanliness": 5,
        "Notes": "Schema + ruleset is effectively a compiler for payment messages.",
    },
    {
        "Task": "Recompute regulatory ratios (LCR, NSFR, RWA, leverage ratio)",
        "Category": "Risk & compliance",
        "ExampleOracle": "Re-derive ratio from book-of-record inputs; compare to reported within tolerance",
        "GroundTruthAvailability": 5,
        "AutomatedOracleStrength": 4,
        "OutputDeterminism": 5,
        "MeasurementLatency": 4,
        "RegulatoryAuditability": 5,
        "RewardSignalCleanliness": 4,
        "Notes": "Very high regulatory auditability; oracle complexity comes from data pipelines.",
    },
    {
        "Task": "Bank statement / GL reconciliation (entries to confirmations)",
        "Category": "Finance ops",
        "ExampleOracle": "Set match of (date, amount, counterparty) tuples; residual must net to zero",
        "GroundTruthAvailability": 5,
        "AutomatedOracleStrength": 5,
        "OutputDeterminism": 5,
        "MeasurementLatency": 5,
        "RegulatoryAuditability": 5,
        "RewardSignalCleanliness": 5,
        "Notes": "Classic verifiable task; reward is binary at the row level.",
    },
    {
        "Task": "Sanctions / PEP screening hit adjudication",
        "Category": "Risk & compliance",
        "ExampleOracle": "Gold-standard labelled hits + name-matching ground truth from compliance team",
        "GroundTruthAvailability": 4,
        "AutomatedOracleStrength": 4,
        "OutputDeterminism": 4,
        "MeasurementLatency": 4,
        "RegulatoryAuditability": 5,
        "RewardSignalCleanliness": 4,
        "Notes": "Has gold-standard labels; some judgment in fuzzy match thresholds.",
    },
    {
        "Task": "AML transaction monitoring alert disposition",
        "Category": "Risk & compliance",
        "ExampleOracle": "L2/L3 reviewer labels + downstream SAR-filed outcome",
        "GroundTruthAvailability": 3,
        "AutomatedOracleStrength": 2,
        "OutputDeterminism": 3,
        "MeasurementLatency": 2,
        "RegulatoryAuditability": 4,
        "RewardSignalCleanliness": 2,
        "Notes": "Truth often only known months later; class imbalance brutal (TPR ~1-3%).",
    },
    {
        "Task": "SAR / narrative drafting (Suspicious Activity Report)",
        "Category": "Risk & compliance",
        "ExampleOracle": "Compliance reviewer scoring + structured-field completeness checker",
        "GroundTruthAvailability": 2,
        "AutomatedOracleStrength": 2,
        "OutputDeterminism": 2,
        "MeasurementLatency": 2,
        "RegulatoryAuditability": 3,
        "RewardSignalCleanliness": 2,
        "Notes": "Narrative quality is judgmental; structured fields can be auto-checked.",
    },
    {
        "Task": "KYC document extraction (passport, utility bill, corporate registry)",
        "Category": "Onboarding",
        "ExampleOracle": "Schema validation + cross-check vs. authoritative registries (e.g., GLEIF, secretary of state)",
        "GroundTruthAvailability": 4,
        "AutomatedOracleStrength": 4,
        "OutputDeterminism": 5,
        "MeasurementLatency": 5,
        "RegulatoryAuditability": 4,
        "RewardSignalCleanliness": 4,
        "Notes": "Highly structured outputs verifiable against registry; OCR errors are local.",
    },
    {
        "Task": "Credit-agreement covenant extraction (loan docs, ISDA schedules)",
        "Category": "Lending / capital markets",
        "ExampleOracle": "Span-level F1 vs. analyst-labelled covenants; structured field validators",
        "GroundTruthAvailability": 3,
        "AutomatedOracleStrength": 3,
        "OutputDeterminism": 3,
        "MeasurementLatency": 3,
        "RegulatoryAuditability": 3,
        "RewardSignalCleanliness": 3,
        "Notes": "Has labelled corpora; cleaner than free-text but legal nuance reduces determinism.",
    },
    {
        "Task": "Fraud / transaction risk scoring (decision: approve / step-up / decline)",
        "Category": "Payments / fraud",
        "ExampleOracle": "Backtested chargeback / confirmed-fraud labels with delayed labels",
        "GroundTruthAvailability": 4,
        "AutomatedOracleStrength": 4,
        "OutputDeterminism": 4,
        "MeasurementLatency": 2,
        "RegulatoryAuditability": 3,
        "RewardSignalCleanliness": 3,
        "Notes": "Strong long-run signal; latency 30-120 days; biased by suppression of declined tx.",
    },
    {
        "Task": "Equity research / earnings note synthesis",
        "Category": "Capital markets",
        "ExampleOracle": "LLM-as-judge vs analyst rubric; downstream price-move correlation (very noisy)",
        "GroundTruthAvailability": 1,
        "AutomatedOracleStrength": 1,
        "OutputDeterminism": 1,
        "MeasurementLatency": 1,
        "RegulatoryAuditability": 2,
        "RewardSignalCleanliness": 1,
        "Notes": "Effectively unverifiable in the strict sense; relies on human ratings.",
    },
    {
        "Task": "Customer-service chat resolution (retail bank)",
        "Category": "Retail banking",
        "ExampleOracle": "Containment, repeat-contact rate, post-chat CSAT, policy-compliance checker",
        "GroundTruthAvailability": 3,
        "AutomatedOracleStrength": 3,
        "OutputDeterminism": 2,
        "MeasurementLatency": 4,
        "RegulatoryAuditability": 3,
        "RewardSignalCleanliness": 3,
        "Notes": "Mixed: hard policy violations are auto-detectable; tone/empathy are not.",
    },
    {
        "Task": "Wealth advisor research summarization",
        "Category": "Wealth & advisory",
        "ExampleOracle": "Source-grounding checker (every claim cites a source) + advisor rubric",
        "GroundTruthAvailability": 2,
        "AutomatedOracleStrength": 2,
        "OutputDeterminism": 2,
        "MeasurementLatency": 3,
        "RegulatoryAuditability": 3,
        "RewardSignalCleanliness": 2,
        "Notes": "Citation/grounding is verifiable; synthesis quality is not.",
    },
    {
        "Task": "Regulatory rule mapping (e.g., new OCC bulletin -> internal control library)",
        "Category": "Risk & compliance",
        "ExampleOracle": "SME-labelled mapping + downstream control-test pass rates",
        "GroundTruthAvailability": 3,
        "AutomatedOracleStrength": 2,
        "OutputDeterminism": 3,
        "MeasurementLatency": 2,
        "RegulatoryAuditability": 4,
        "RewardSignalCleanliness": 2,
        "Notes": "Auditable artifact, but truth set evolves with regulation.",
    },
    {
        "Task": "Insurance claims FNOL coding & subrogation routing",
        "Category": "Insurance",
        "ExampleOracle": "Closed-claim outcome labels + structured-field validators",
        "GroundTruthAvailability": 4,
        "AutomatedOracleStrength": 3,
        "OutputDeterminism": 3,
        "MeasurementLatency": 3,
        "RegulatoryAuditability": 3,
        "RewardSignalCleanliness": 3,
        "Notes": "Structured codes verifiable; narrative letters less so.",
    },
    {
        "Task": "Algo / autonomous order routing decision",
        "Category": "Trading",
        "ExampleOracle": "Best-ex TCA framework, IS / VWAP slippage vs benchmark",
        "GroundTruthAvailability": 4,
        "AutomatedOracleStrength": 4,
        "OutputDeterminism": 3,
        "MeasurementLatency": 4,
        "RegulatoryAuditability": 4,
        "RewardSignalCleanliness": 3,
        "Notes": "Slippage vs benchmark is a strong signal but counterfactual hard.",
    },
    {
        "Task": "Disputes / chargeback evidence packet generation",
        "Category": "Payments",
        "ExampleOracle": "Issuer/scheme decision (won/lost) + completeness checker against scheme rule",
        "GroundTruthAvailability": 4,
        "AutomatedOracleStrength": 4,
        "OutputDeterminism": 3,
        "MeasurementLatency": 3,
        "RegulatoryAuditability": 4,
        "RewardSignalCleanliness": 4,
        "Notes": "Win/lose is a clean binary outcome; latency is days/weeks.",
    },
]


# ---------------------------------------------------------------------------
# General coding-agent tasks
# ---------------------------------------------------------------------------

CODING_TASKS: list[dict] = [
    {
        "Task": "Compile / typecheck a codebase",
        "Category": "Build",
        "ExampleOracle": "Compiler / tsc / mypy exit code",
        "GroundTruthAvailability": 5,
        "AutomatedOracleStrength": 5,
        "OutputDeterminism": 5,
        "MeasurementLatency": 5,
        "RegulatoryAuditability": 3,
        "RewardSignalCleanliness": 5,
        "Notes": "Canonical verifiable signal; basis of CI gating.",
    },
    {
        "Task": "Run unit tests",
        "Category": "Test",
        "ExampleOracle": "pytest / jest / go test exit code + counts",
        "GroundTruthAvailability": 5,
        "AutomatedOracleStrength": 5,
        "OutputDeterminism": 5,
        "MeasurementLatency": 5,
        "RegulatoryAuditability": 3,
        "RewardSignalCleanliness": 5,
        "Notes": "Underlies SWE-bench, MBPP, HumanEval; the prototype verifiable task.",
    },
    {
        "Task": "Resolve a real GitHub issue (SWE-bench style)",
        "Category": "Bug fix",
        "ExampleOracle": "Hidden test suite from the actual PR that fixed the issue",
        "GroundTruthAvailability": 5,
        "AutomatedOracleStrength": 5,
        "OutputDeterminism": 4,
        "MeasurementLatency": 5,
        "RegulatoryAuditability": 2,
        "RewardSignalCleanliness": 5,
        "Notes": "Multiple correct patches possible, but tests give clean reward.",
    },
    {
        "Task": "Pass a coding-interview problem (HumanEval / MBPP / LeetCode)",
        "Category": "Algorithms",
        "ExampleOracle": "Reference test cases + complexity asserts",
        "GroundTruthAvailability": 5,
        "AutomatedOracleStrength": 5,
        "OutputDeterminism": 5,
        "MeasurementLatency": 5,
        "RegulatoryAuditability": 2,
        "RewardSignalCleanliness": 5,
        "Notes": "Pure verifiable; foundation of code-RL pipelines.",
    },
    {
        "Task": "Translate code between languages (Python -> TypeScript)",
        "Category": "Refactoring",
        "ExampleOracle": "Original test suite executed against translated code",
        "GroundTruthAvailability": 4,
        "AutomatedOracleStrength": 4,
        "OutputDeterminism": 4,
        "MeasurementLatency": 5,
        "RegulatoryAuditability": 2,
        "RewardSignalCleanliness": 4,
        "Notes": "Tests give clean signal if portable; runtime semantics differences add noise.",
    },
    {
        "Task": "Add type annotations / mypy clean a codebase",
        "Category": "Refactoring",
        "ExampleOracle": "mypy --strict + tests still pass",
        "GroundTruthAvailability": 5,
        "AutomatedOracleStrength": 5,
        "OutputDeterminism": 4,
        "MeasurementLatency": 5,
        "RegulatoryAuditability": 2,
        "RewardSignalCleanliness": 5,
        "Notes": "Clean two-stage gate (static + dynamic).",
    },
    {
        "Task": "Refactor for performance (no behavior change)",
        "Category": "Performance",
        "ExampleOracle": "Tests pass + benchmark improves by >= X% under repeated runs",
        "GroundTruthAvailability": 4,
        "AutomatedOracleStrength": 4,
        "OutputDeterminism": 3,
        "MeasurementLatency": 4,
        "RegulatoryAuditability": 2,
        "RewardSignalCleanliness": 4,
        "Notes": "Microbenchmark variance is the main noise source.",
    },
    {
        "Task": "Implement an HTTP/gRPC API to a spec",
        "Category": "Backend",
        "ExampleOracle": "Conformance suite + OpenAPI/proto contract tests",
        "GroundTruthAvailability": 5,
        "AutomatedOracleStrength": 5,
        "OutputDeterminism": 4,
        "MeasurementLatency": 5,
        "RegulatoryAuditability": 3,
        "RewardSignalCleanliness": 5,
        "Notes": "Schema-driven; very clean reward.",
    },
    {
        "Task": "SQL query authoring against a known schema (Spider / BIRD)",
        "Category": "Data",
        "ExampleOracle": "Execute query, compare result set to gold",
        "GroundTruthAvailability": 5,
        "AutomatedOracleStrength": 5,
        "OutputDeterminism": 4,
        "MeasurementLatency": 5,
        "RegulatoryAuditability": 3,
        "RewardSignalCleanliness": 5,
        "Notes": "Multiple SQL strings can produce identical results; that's fine for the oracle.",
    },
    {
        "Task": "Infrastructure-as-code change (Terraform plan)",
        "Category": "DevOps",
        "ExampleOracle": "terraform plan diff + policy-as-code (OPA / Sentinel) checks",
        "GroundTruthAvailability": 4,
        "AutomatedOracleStrength": 5,
        "OutputDeterminism": 4,
        "MeasurementLatency": 5,
        "RegulatoryAuditability": 4,
        "RewardSignalCleanliness": 5,
        "Notes": "Static plan + policy; high regulatory auditability via policy-as-code.",
    },
    {
        "Task": "Security vulnerability fix (CVE remediation)",
        "Category": "Security",
        "ExampleOracle": "SAST/DAST scanner + targeted exploit test + regression suite",
        "GroundTruthAvailability": 4,
        "AutomatedOracleStrength": 4,
        "OutputDeterminism": 3,
        "MeasurementLatency": 4,
        "RegulatoryAuditability": 4,
        "RewardSignalCleanliness": 4,
        "Notes": "Scanners noisy; exploit tests are gold but expensive to author.",
    },
    {
        "Task": "Frontend UI implementation from a Figma design",
        "Category": "Frontend",
        "ExampleOracle": "Visual regression tests + accessibility (axe) + e2e tests",
        "GroundTruthAvailability": 3,
        "AutomatedOracleStrength": 3,
        "OutputDeterminism": 2,
        "MeasurementLatency": 4,
        "RegulatoryAuditability": 2,
        "RewardSignalCleanliness": 3,
        "Notes": "Pixel-level diff is brittle; semantic match harder.",
    },
    {
        "Task": "Write developer documentation for a module",
        "Category": "Docs",
        "ExampleOracle": "Doctest execution + reviewer rubric (LLM-as-judge)",
        "GroundTruthAvailability": 2,
        "AutomatedOracleStrength": 2,
        "OutputDeterminism": 2,
        "MeasurementLatency": 3,
        "RegulatoryAuditability": 1,
        "RewardSignalCleanliness": 2,
        "Notes": "Doctest snippet is verifiable; prose quality is not.",
    },
    {
        "Task": "System design / architecture proposal",
        "Category": "Design",
        "ExampleOracle": "Senior-engineer rubric (judge model)",
        "GroundTruthAvailability": 1,
        "AutomatedOracleStrength": 1,
        "OutputDeterminism": 1,
        "MeasurementLatency": 1,
        "RegulatoryAuditability": 1,
        "RewardSignalCleanliness": 1,
        "Notes": "Effectively unverifiable in the strict sense.",
    },
    {
        "Task": "Code review / PR comments",
        "Category": "Review",
        "ExampleOracle": "Overlap with human reviewer comments + bug-catch rate on seeded defects",
        "GroundTruthAvailability": 3,
        "AutomatedOracleStrength": 3,
        "OutputDeterminism": 2,
        "MeasurementLatency": 4,
        "RegulatoryAuditability": 2,
        "RewardSignalCleanliness": 3,
        "Notes": "Synthetic-defect injection gives a clean reward subset.",
    },
    {
        "Task": "Database migration script (with rollback)",
        "Category": "Data",
        "ExampleOracle": "Apply on shadow DB + integrity checks + reversibility test",
        "GroundTruthAvailability": 5,
        "AutomatedOracleStrength": 5,
        "OutputDeterminism": 4,
        "MeasurementLatency": 4,
        "RegulatoryAuditability": 4,
        "RewardSignalCleanliness": 5,
        "Notes": "Clean shadow-environment evaluation.",
    },
    {
        "Task": "Browser-use / computer-use task (e.g., file an expense)",
        "Category": "Agentic",
        "ExampleOracle": "Goal predicate over final UI/system state (WebArena/SWE-Lancer style)",
        "GroundTruthAvailability": 4,
        "AutomatedOracleStrength": 4,
        "OutputDeterminism": 3,
        "MeasurementLatency": 4,
        "RegulatoryAuditability": 2,
        "RewardSignalCleanliness": 3,
        "Notes": "Goal predicates are gold but brittle to UI drift.",
    },
]


# ---------------------------------------------------------------------------
# Public benchmarks / oracles to anchor each domain
# ---------------------------------------------------------------------------

BENCHMARKS: list[dict] = [
    {
        "Domain": "General coding",
        "Benchmark": "SWE-bench / SWE-bench Verified",
        "Oracle_type": "Hidden unit-test suite from the resolving PR",
        "Verifiability_class": "Strong, fully automated",
        "SourceId": "B1",
    },
    {
        "Domain": "General coding",
        "Benchmark": "HumanEval / MBPP",
        "Oracle_type": "Reference test cases over function outputs",
        "Verifiability_class": "Strong, fully automated",
        "SourceId": "B2",
    },
    {
        "Domain": "General coding",
        "Benchmark": "LiveCodeBench",
        "Oracle_type": "Held-out competitive-programming test cases (contamination-resistant)",
        "Verifiability_class": "Strong, fully automated",
        "SourceId": "B3",
    },
    {
        "Domain": "General coding",
        "Benchmark": "Spider / BIRD (text-to-SQL)",
        "Oracle_type": "Execution-equivalence of result set vs gold",
        "Verifiability_class": "Strong, fully automated",
        "SourceId": "B4",
    },
    {
        "Domain": "Agentic / web",
        "Benchmark": "WebArena / VisualWebArena / OSWorld",
        "Oracle_type": "Goal predicates over final environment state",
        "Verifiability_class": "Medium-strong; brittle to UI drift",
        "SourceId": "B5",
    },
    {
        "Domain": "Agentic / SWE",
        "Benchmark": "SWE-Lancer / SWE-Gym",
        "Oracle_type": "Tests + economic outcome of freelance task",
        "Verifiability_class": "Strong",
        "SourceId": "B6",
    },
    {
        "Domain": "Finance reasoning",
        "Benchmark": "FinanceBench (Patronus AI)",
        "Oracle_type": "Open-book QA against 10-K/10-Q with reference answers",
        "Verifiability_class": "Medium (human-graded ref answers)",
        "SourceId": "B7",
    },
    {
        "Domain": "Finance reasoning",
        "Benchmark": "FinQA / TAT-QA / ConvFinQA",
        "Oracle_type": "Numerical answer + program-of-thought match",
        "Verifiability_class": "Medium-strong (numeric)",
        "SourceId": "B8",
    },
    {
        "Domain": "Finance reasoning",
        "Benchmark": "BloombergGPT FinNLP suite (sentiment, NER, headlines)",
        "Oracle_type": "Labelled gold for classification / NER",
        "Verifiability_class": "Medium (label-noise risk)",
        "SourceId": "B9",
    },
    {
        "Domain": "Finance reasoning",
        "Benchmark": "PIXIU / FinBen",
        "Oracle_type": "Labelled gold across 24+ financial NLP tasks",
        "Verifiability_class": "Medium",
        "SourceId": "B10",
    },
    {
        "Domain": "Finance ops",
        "Benchmark": "(internal) Reconciliation / ratio-recomputation harnesses",
        "Oracle_type": "Recompute against book-of-record; row-level diff",
        "Verifiability_class": "Strong (when data lineage exists)",
        "SourceId": "B11",
    },
    {
        "Domain": "Risk / compliance",
        "Benchmark": "(internal) AML alert disposition + delayed SAR/STR outcome",
        "Oracle_type": "Long-horizon labelled outcomes",
        "Verifiability_class": "Weak (latency, class imbalance)",
        "SourceId": "B12",
    },
]


# ---------------------------------------------------------------------------
# Sources / bibliography
# ---------------------------------------------------------------------------

SOURCES: list[dict] = [
    {"id": "B1", "publisher": "Princeton / Anthropic / OpenAI", "date": "2023-2024",
     "short": "SWE-bench: Can Language Models Resolve Real-World GitHub Issues? (Jimenez et al.) + SWE-bench Verified",
     "url": "https://www.swebench.com/"},
    {"id": "B2", "publisher": "OpenAI / Google", "date": "2021",
     "short": "HumanEval (Chen et al., Codex paper) and MBPP (Austin et al.)",
     "url": "https://github.com/openai/human-eval"},
    {"id": "B3", "publisher": "UC Berkeley et al.", "date": "2024",
     "short": "LiveCodeBench: holistic and contamination-free evaluation of LLMs for code",
     "url": "https://livecodebench.github.io/"},
    {"id": "B4", "publisher": "Yale / Salesforce / academia", "date": "2018-2023",
     "short": "Spider, BIRD: large-scale text-to-SQL benchmarks with execution-based evaluation",
     "url": "https://yale-lily.github.io/spider"},
    {"id": "B5", "publisher": "CMU / OSU / academia", "date": "2023-2024",
     "short": "WebArena, VisualWebArena, OSWorld: realistic web/desktop agent benchmarks with goal predicates",
     "url": "https://webarena.dev/"},
    {"id": "B6", "publisher": "OpenAI / academia", "date": "2024-2025",
     "short": "SWE-Lancer / SWE-Gym: economic and engineering agent benchmarks",
     "url": "https://github.com/openai/SWELancer-Benchmark"},
    {"id": "B7", "publisher": "Patronus AI", "date": "2023",
     "short": "FinanceBench: a new benchmark for financial question answering",
     "url": "https://www.patronus.ai/announcements/patronus-ai-launches-financebench"},
    {"id": "B8", "publisher": "Academic (EMNLP/ACL)", "date": "2021-2022",
     "short": "FinQA, TAT-QA, ConvFinQA: numerical reasoning over financial documents",
     "url": "https://github.com/czyssrs/FinQA"},
    {"id": "B9", "publisher": "Bloomberg", "date": "2023",
     "short": "BloombergGPT: A Large Language Model for Finance (Wu et al.)",
     "url": "https://www.bloomberg.com/company/press/bloomberggpt-50-billion-parameter-llm-tuned-finance/"},
    {"id": "B10", "publisher": "Academic (PIXIU / FinBen consortium)", "date": "2023-2024",
     "short": "PIXIU / FinBen: comprehensive Chinese & English financial NLP benchmarks",
     "url": "https://github.com/The-FinAI/PIXIU"},
    {"id": "B11", "publisher": "Industry practice", "date": "ongoing",
     "short": "Reconciliation and reg-ratio recomputation harnesses (BCBS 239 lineage practice)",
     "url": "https://www.bis.org/publ/bcbs239.htm"},
    {"id": "B12", "publisher": "FATF / FinCEN / industry", "date": "ongoing",
     "short": "AML alert outcomes - delayed-label, class-imbalanced verification",
     "url": "https://www.fatf-gafi.org/"},
    {"id": "S1", "publisher": "OpenAI", "date": "2024-2025",
     "short": "OpenAI o-series and verifier-based RL: 'Learning to Reason with LLMs'",
     "url": "https://openai.com/index/learning-to-reason-with-llms/"},
    {"id": "S2", "publisher": "Anthropic", "date": "2024-2025",
     "short": "Claude 3.5/3.7 Sonnet system cards: emphasis on agentic coding and verifiable RL",
     "url": "https://www.anthropic.com/news"},
    {"id": "S3", "publisher": "BIS / Basel Committee", "date": "2013 (still current)",
     "short": "BCBS 239: Principles for effective risk data aggregation and risk reporting",
     "url": "https://www.bis.org/publ/bcbs239.htm"},
    {"id": "S4", "publisher": "Federal Reserve / OCC", "date": "2011 (still current)",
     "short": "SR 11-7: Guidance on Model Risk Management",
     "url": "https://www.federalreserve.gov/supervisionreg/srletters/sr1107.htm"},
    {"id": "S5", "publisher": "European Commission", "date": "2024",
     "short": "EU AI Act - high-risk obligations include credit scoring (Annex III)",
     "url": "https://artificialintelligenceact.eu/"},
    {"id": "S6", "publisher": "ISO / SWIFT / Nacha", "date": "ongoing",
     "short": "ISO 20022, SWIFT MT/MX, Nacha and SEPA scheme rulebooks (deterministic schemas)",
     "url": "https://www.iso20022.org/"},
]

SOURCE_BY_ID = {s["id"]: s for s in SOURCES}


# ---------------------------------------------------------------------------
# Scoring
# ---------------------------------------------------------------------------

def composite(task: dict) -> float:
    return round(sum(task[d] * w for d, w in WEIGHTS.items()), 2)


def annotate(tasks: list[dict], domain_label: str) -> None:
    for t in tasks:
        t["Domain"] = domain_label
        t["CompositeScore"] = composite(t)
    tasks.sort(key=lambda t: t["CompositeScore"], reverse=True)
    for i, t in enumerate(tasks, start=1):
        t["Rank"] = i


# ---------------------------------------------------------------------------
# Workbook plumbing
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E78")
WRAP = Alignment(wrap_text=True, vertical="top")
BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def style_header(ws, row_idx: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        cell.border = BORDER


def auto_width(ws, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        m = 0
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            for line in str(v).split("\n"):
                m = max(m, len(line))
        ws.column_dimensions[col_letter].width = min(max(12, m + 2), max_width)


def add_table(ws, headers: list[str], name: str, style: str = "TableStyleMedium2") -> None:
    rng = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    table = Table(displayName=name, ref=rng)
    table.tableStyleInfo = TableStyleInfo(
        name=style, showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)


def write_readme(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "README"
    rows = [
        "Verifiable Tasks: BFSI AI Agents vs General Coding Agents",
        "Side-by-side analysis of how cleanly agent outputs can be checked",
        "",
        "Companion to: research/verifiable-tasks-bfsi-vs-coding.md",
        "Generated by: scripts/build_verifiability_workbook.py",
        "",
        "Definition of verifiability:",
        "  The degree to which an agent's output can be checked, after the fact, by an",
        "  automated or near-automated oracle that returns a clear pass / fail signal",
        "  usable for evaluation, RL training, or production gating.",
        "",
        "Sheets:",
        "  1. README                - this page",
        "  2. Verifiability_Rubric  - the 6-dimension scoring framework",
        "  3. BFSI_Tasks            - finance-domain agent tasks scored",
        "  4. Coding_Tasks          - general coding-agent tasks scored",
        "  5. SideBySide            - aggregate comparison + ranked combined view",
        "  6. Benchmarks            - public benchmarks per domain",
        "  7. Sources               - bibliography",
        "",
        "Scoring rubric (1=low, 5=high; weights in Verifiability_Rubric sheet):",
        "  GroundTruthAvailability  - Does a single, agreed reference exist?           weight 0.25",
        "  AutomatedOracleStrength  - Can the check be run by a fast deterministic prog? 0.25",
        "  OutputDeterminism        - Is the correct output well-defined?               0.15",
        "  MeasurementLatency       - How quickly can a verdict be produced?            0.10",
        "  RegulatoryAuditability   - Can the artifact serve as audit evidence?         0.10",
        "  RewardSignalCleanliness  - Suitability as a training reward                  0.15",
        "",
        "Headline finding:",
        "  General coding tasks have, on average, materially higher verifiability than BFSI",
        "  agent tasks - because compilers, type-checkers, and unit-test suites are an",
        "  almost-perfect oracle. BFSI verifiability is bimodal: ops/payments/reconciliation",
        "  tasks score as high as coding (deterministic schemas, recomputable ledgers) while",
        "  narrative-heavy tasks (SAR drafting, equity research, advisor synthesis) score",
        "  the lowest of any task in this workbook. The practical implication for RL and",
        "  evaluation is that the highest-leverage BFSI agent investments are the ones",
        "  closest to 'finance unit tests': reconciliations, ISO 20022/SWIFT formatting,",
        "  KYC extraction, regulatory-ratio recomputation, and chargeback-evidence",
        "  generation.",
        "",
        "Disclaimer: educational / discussion only. Not investment advice.",
    ]
    ws["A1"] = rows[0]
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    for i, text in enumerate(rows[1:], start=2):
        ws.cell(row=i, column=1, value=text).alignment = WRAP
    ws.column_dimensions["A"].width = 110


def write_rubric(wb: Workbook) -> None:
    ws = wb.create_sheet("Verifiability_Rubric")
    headers = ["Dimension", "Weight", "What_it_measures", "High_score_example", "Low_score_example"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for r in RUBRIC:
        ws.append([r[h] for h in headers])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
    ws.freeze_panes = "A2"
    auto_width(ws)
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 50
    add_table(ws, headers, "Rubric", style="TableStyleMedium4")


def write_task_sheet(wb: Workbook, sheet_name: str, tasks: list[dict], table_name: str,
                     style: str) -> None:
    ws = wb.create_sheet(sheet_name)
    headers = [
        "Rank", "Domain", "Category", "Task", "ExampleOracle",
        "GroundTruthAvailability", "AutomatedOracleStrength", "OutputDeterminism",
        "MeasurementLatency", "RegulatoryAuditability", "RewardSignalCleanliness",
        "CompositeScore", "Notes",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for t in tasks:
        ws.append([t.get(h, "") for h in headers])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
    ws.freeze_panes = "A2"
    auto_width(ws)
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 55
    ws.column_dimensions["M"].width = 55
    add_table(ws, headers, table_name, style=style)


def write_sidebyside(wb: Workbook, bfsi: list[dict], coding: list[dict]) -> None:
    ws = wb.create_sheet("SideBySide")

    def stats(tasks: list[dict]) -> dict:
        n = len(tasks)
        agg = {"N": n}
        for d in WEIGHTS:
            vals = [t[d] for t in tasks]
            agg[f"avg_{d}"] = round(sum(vals) / n, 2)
        agg["avg_Composite"] = round(sum(t["CompositeScore"] for t in tasks) / n, 2)
        agg["min_Composite"] = round(min(t["CompositeScore"] for t in tasks), 2)
        agg["max_Composite"] = round(max(t["CompositeScore"] for t in tasks), 2)
        return agg

    headers = ["Metric", "BFSI agents", "Coding agents", "Delta (BFSI - Coding)"]
    bfsi_stats = stats(bfsi)
    coding_stats = stats(coding)
    ws.append(headers)
    style_header(ws, 1, len(headers))
    rows = [
        ("Number of tasks scored", bfsi_stats["N"], coding_stats["N"]),
        ("Avg GroundTruthAvailability", bfsi_stats["avg_GroundTruthAvailability"], coding_stats["avg_GroundTruthAvailability"]),
        ("Avg AutomatedOracleStrength", bfsi_stats["avg_AutomatedOracleStrength"], coding_stats["avg_AutomatedOracleStrength"]),
        ("Avg OutputDeterminism", bfsi_stats["avg_OutputDeterminism"], coding_stats["avg_OutputDeterminism"]),
        ("Avg MeasurementLatency", bfsi_stats["avg_MeasurementLatency"], coding_stats["avg_MeasurementLatency"]),
        ("Avg RegulatoryAuditability", bfsi_stats["avg_RegulatoryAuditability"], coding_stats["avg_RegulatoryAuditability"]),
        ("Avg RewardSignalCleanliness", bfsi_stats["avg_RewardSignalCleanliness"], coding_stats["avg_RewardSignalCleanliness"]),
        ("Avg CompositeScore", bfsi_stats["avg_Composite"], coding_stats["avg_Composite"]),
        ("Min CompositeScore", bfsi_stats["min_Composite"], coding_stats["min_Composite"]),
        ("Max CompositeScore", bfsi_stats["max_Composite"], coding_stats["max_Composite"]),
    ]
    for label, bv, cv in rows:
        delta = round(bv - cv, 2) if isinstance(bv, (int, float)) and isinstance(cv, (int, float)) else ""
        ws.append([label, bv, cv, delta])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
    ws.freeze_panes = "A2"
    auto_width(ws)
    add_table(ws, headers, "AggregateCompare", style="TableStyleMedium2")

    blank_row = ws.max_row + 2
    ws.cell(row=blank_row, column=1, value="Combined ranking (all tasks across both domains)").font = TITLE_FONT
    ws.merge_cells(start_row=blank_row, start_column=1, end_row=blank_row, end_column=4)

    combined = sorted(bfsi + coding, key=lambda t: t["CompositeScore"], reverse=True)
    headers2 = ["CombinedRank", "Domain", "Task", "CompositeScore"]
    header_row = blank_row + 1
    for ci, h in enumerate(headers2, start=1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        cell.border = BORDER
    for i, t in enumerate(combined, start=1):
        r = header_row + i
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=t["Domain"])
        ws.cell(row=r, column=3, value=t["Task"])
        ws.cell(row=r, column=4, value=t["CompositeScore"])
        for c in range(1, 5):
            ws.cell(row=r, column=c).alignment = WRAP
            ws.cell(row=r, column=c).border = BORDER
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 18


def write_benchmarks(wb: Workbook) -> None:
    ws = wb.create_sheet("Benchmarks")
    headers = ["Domain", "Benchmark", "Oracle_type", "Verifiability_class", "SourceId", "URL"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for b in BENCHMARKS:
        url = SOURCE_BY_ID[b["SourceId"]]["url"]
        ws.append([b["Domain"], b["Benchmark"], b["Oracle_type"], b["Verifiability_class"],
                   b["SourceId"], url])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
    ws.freeze_panes = "A2"
    auto_width(ws)
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["F"].width = 55
    add_table(ws, headers, "Benchmarks", style="TableStyleMedium6")


def write_sources(wb: Workbook) -> None:
    ws = wb.create_sheet("Sources")
    headers = ["Id", "Publisher", "Date", "Description", "URL"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for s in SOURCES:
        ws.append([s["id"], s["publisher"], s["date"], s["short"], s["url"]])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
    ws.freeze_panes = "A2"
    auto_width(ws)
    ws.column_dimensions["D"].width = 75
    ws.column_dimensions["E"].width = 60
    add_table(ws, headers, "Sources", style="TableStyleMedium9")


def main() -> None:
    annotate(BFSI_TASKS, "BFSI")
    annotate(CODING_TASKS, "Coding")

    wb = Workbook()
    write_readme(wb)
    write_rubric(wb)
    write_task_sheet(wb, "BFSI_Tasks", BFSI_TASKS, "BFSITasks", "TableStyleMedium2")
    write_task_sheet(wb, "Coding_Tasks", CODING_TASKS, "CodingTasks", "TableStyleMedium3")
    write_sidebyside(wb, BFSI_TASKS, CODING_TASKS)
    write_benchmarks(wb)
    write_sources(wb)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print()

    def show(title, tasks):
        print(title)
        for t in tasks[:5]:
            print(f"  #{t['Rank']:>2}  {t['CompositeScore']:.2f}  {t['Task']}")
        print(f"  ... ({len(tasks)} total)")
        print()

    show("BFSI - top 5 by verifiability:", BFSI_TASKS)
    show("Coding - top 5 by verifiability:", CODING_TASKS)
    avg_b = sum(t["CompositeScore"] for t in BFSI_TASKS) / len(BFSI_TASKS)
    avg_c = sum(t["CompositeScore"] for t in CODING_TASKS) / len(CODING_TASKS)
    print(f"Avg verifiability composite: BFSI={avg_b:.2f}  Coding={avg_c:.2f}  Delta={avg_b - avg_c:+.2f}")


if __name__ == "__main__":
    main()
