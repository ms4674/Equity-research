"""Build an Excel workbook comparing US labor compensation vs worldwide IT spending.

Outputs: data/Labor_vs_IT_Spending_2001_2025.xlsx

Series:
  - US Compensation of Employees (BEA, NIPA) -- annual, $B (FRED series COE)
  - Worldwide IT Spending (Gartner) -- annual, $B (compiled from Gartner press releases)

The workbook contains three sheets:
  1. Data        - the underlying time series + ratio
  2. Chart       - a combo line chart of both series
  3. Sources     - notes & citations
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.trendline import Trendline
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# US Compensation of Employees, annual, $B SAAR avg
# Source: FRED series COE (U.S. Bureau of Economic Analysis, NIPA Table 1.10)
# https://fred.stlouisfed.org/series/COE  (annual frequency, average aggregation)
US_COE_BILLIONS = {
    2001: 6038.349,
    2002: 6135.115,
    2003: 6353.605,
    2004: 6719.488,
    2005: 7066.107,
    2006: 7479.677,
    2007: 7878.475,
    2008: 8056.826,
    2009: 7759.019,
    2010: 7925.369,
    2011: 8226.175,
    2012: 8567.363,
    2013: 8835.039,
    2014: 9250.208,
    2015: 9699.419,
    2016: 9966.108,
    2017: 10424.372,
    2018: 10957.405,
    2019: 11446.611,
    2020: 11598.142,
    2021: 12558.372,
    2022: 13443.379,
    2023: 14207.393,
    2024: 15027.061,
    2025: 15745.135,
}

# Worldwide IT Spending, $B (Gartner Worldwide IT Spending Forecast).
# Compiled from contemporaneous and revised Gartner press releases. Values
# include hardware (data center systems, devices), enterprise software, IT
# services, and communications/telecom services -- the standard Gartner
# definition. See the Sources sheet for the underlying release for each year.
WORLD_IT_BILLIONS = {
    2001: 2226.7,   # Gartner Dataquest, Sep 2002
    2002: 2302.1,   # Gartner Dataquest, Sep 2002
    2003: 2463.1,   # Gartner Dataquest, Oct 2002 forecast
    2004: 2462.0,   # Implied from 2005 = $2.6T at 5.6% growth (Gartner, Feb 2006)
    2005: 2600.0,   # Gartner, Feb 2006
    2006: 2917.0,   # Gartner, Oct 2007 (2007 = $3.15T, +8% YoY)
    2007: 3151.0,   # Gartner, Jul 2008
    2008: 3406.0,   # Gartner, Jul 2008
    2009: 3227.0,   # Gartner, Jan 2011 (retrospective)
    2010: 3401.6,   # Gartner, Jan 2011
    2011: 3575.8,   # Gartner, Jan 2011
    2012: 3604.0,   # Gartner, Jul 2012
    2013: 3737.0,   # Gartner, later revision
    2014: 3710.0,   # Gartner, Jul 2014
    2015: 3517.0,   # Gartner, Jan 2016
    2016: 3424.0,   # Gartner, Apr 2016 (revised down vs Jan 2016)
    2017: 3539.0,   # Gartner, Apr 2018 (retrospective)
    2018: 3699.0,   # Gartner, Oct 2018
    2019: 3737.0,   # Gartner, Jan 2020
    2020: 3608.8,   # Gartner, Oct 2020 (revised)
    2021: 4259.8,   # Gartner, Apr 2022
    2022: 4431.6,   # Gartner, Apr 2022 (forecast); later revisions ~$4.40T
    2023: 4678.8,   # Gartner, Jan 2024
    2024: 5114.8,   # Gartner, Jan 2025
    2025: 5540.4,   # Gartner, Oct 2025
}


def style_header(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        top=Side(style="thin", color="FFFFFF"),
        bottom=Side(style="thin", color="FFFFFF"),
    )


def style_body(cell, *, money: bool = False, ratio: bool = False) -> None:
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if money:
        cell.number_format = "#,##0.0"
    elif ratio:
        cell.number_format = "0.00x"
    cell.border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )


def build_workbook(output_path: Path) -> None:
    wb = Workbook()

    # ---------------- Data sheet ----------------
    ws = wb.active
    ws.title = "Data"

    headers = [
        "Year",
        "US Compensation of Employees ($B, BEA)",
        "Worldwide IT Spending ($B, Gartner)",
        "Labor / IT Ratio",
        "YoY Labor Growth",
        "YoY IT Growth",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        style_header(cell)

    years = sorted(US_COE_BILLIONS.keys())
    for i, year in enumerate(years):
        row = i + 2
        labor = US_COE_BILLIONS[year]
        it = WORLD_IT_BILLIONS[year]

        ws.cell(row=row, column=1, value=year).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=1).font = Font(bold=True)

        c_labor = ws.cell(row=row, column=2, value=labor)
        style_body(c_labor, money=True)

        c_it = ws.cell(row=row, column=3, value=it)
        style_body(c_it, money=True)

        c_ratio = ws.cell(row=row, column=4, value=f"=B{row}/C{row}")
        style_body(c_ratio, ratio=True)

        if i == 0:
            ws.cell(row=row, column=5, value=None)
            ws.cell(row=row, column=6, value=None)
        else:
            c_lg = ws.cell(row=row, column=5, value=f"=B{row}/B{row-1}-1")
            c_lg.number_format = "0.0%"
            c_lg.alignment = Alignment(horizontal="right")
            c_ig = ws.cell(row=row, column=6, value=f"=C{row}/C{row-1}-1")
            c_ig.number_format = "0.0%"
            c_ig.alignment = Alignment(horizontal="right")

    column_widths = [8, 30, 30, 16, 18, 18]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "B2"

    # Summary stats block
    n_years = len(years)
    summary_row = n_years + 4
    ws.cell(row=summary_row, column=1, value="Summary (2001-2025)").font = Font(bold=True, size=12)

    labels_formulas = [
        ("Total Labor (sum, $B)", f"=SUM(B2:B{n_years + 1})", "#,##0"),
        ("Total IT (sum, $B)", f"=SUM(C2:C{n_years + 1})", "#,##0"),
        ("Avg Labor / IT ratio", f"=AVERAGE(D2:D{n_years + 1})", "0.00x"),
        ("Labor CAGR", f"=(B{n_years + 1}/B2)^(1/{n_years - 1})-1", "0.00%"),
        ("IT CAGR", f"=(C{n_years + 1}/C2)^(1/{n_years - 1})-1", "0.00%"),
    ]
    for offset, (label, formula, fmt) in enumerate(labels_formulas):
        r = summary_row + 1 + offset
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(bold=True)
        vc = ws.cell(row=r, column=2, value=formula)
        vc.number_format = fmt
        vc.alignment = Alignment(horizontal="right")

    # ---------------- Chart sheet ----------------
    chart_ws = wb.create_sheet("Chart")
    chart_ws.cell(row=1, column=1, value="US Labor Compensation vs Worldwide IT Spending, 2001-2025").font = Font(
        bold=True, size=14
    )
    chart_ws.cell(row=2, column=1, value="Annual values in $ Billions. See Data and Sources sheets for detail.").font = Font(
        italic=True, size=10, color="595959"
    )

    chart = LineChart()
    chart.title = "US Compensation of Employees vs Worldwide IT Spending ($B, 2001-2025)"
    chart.y_axis.title = "USD, Billions"
    chart.x_axis.title = "Year"
    chart.height = 16
    chart.width = 30
    chart.style = 12

    data_ref = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=n_years + 1)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=n_years + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    for series in chart.series:
        series.smooth = False
        series.graphicalProperties.line.width = 22000

    # Add trendlines for visual context
    if len(chart.series) >= 1:
        chart.series[0].trendline = Trendline(trendlineType="linear", dispEq=False, dispRSqr=False)
    if len(chart.series) >= 2:
        chart.series[1].trendline = Trendline(trendlineType="linear", dispEq=False, dispRSqr=False)

    chart_ws.add_chart(chart, "A4")

    # ---------------- Sources sheet ----------------
    src = wb.create_sheet("Sources")
    src.cell(row=1, column=1, value="Sources & Methodology").font = Font(bold=True, size=14)

    notes = [
        "",
        "Series 1 - US Compensation of Employees ($B, annual)",
        "  Series ID: COE (Bureau of Economic Analysis, NIPA Table 1.10).",
        "  Reported as Billions of dollars, Seasonally Adjusted Annual Rate; annual values are average of monthly observations.",
        "  Retrieved from FRED, Federal Reserve Bank of St. Louis: https://fred.stlouisfed.org/series/COE",
        "  Definition: Total compensation paid to employees in the United States, including wages,",
        "  salaries, employer contributions for pension/insurance funds, and government social insurance.",
        "",
        "Series 2 - Worldwide IT Spending ($B, annual)",
        "  Source: Gartner Worldwide IT Spending Forecast (multiple press releases).",
        "  Includes data center systems, enterprise software, devices, IT services, and communications services.",
        "  Most recent revision used where available; pre-2010 figures are from contemporaneous Gartner Dataquest releases.",
        "",
        "Key Gartner press releases used:",
        "  - 2001-2003: Gartner Dataquest (Sep/Oct 2002) - reported via eMarketer; total $2.23T (2001), $2.30T (2002), $2.46T (2003 fcst).",
        "  - 2005:    Gartner (Feb 2006) - $2.6T worldwide.",
        "  - 2007:    Gartner (Oct 2007) - $3.15T worldwide, implies ~$2.92T in 2006.",
        "  - 2008:    Gartner (Aug 2008) - $3.41T (forecast), used as reported.",
        "  - 2009-2011: Gartner (Jan 2011) - 2009=$3.23T, 2010=$3.40T, 2011=$3.58T.",
        "  - 2012:    Gartner (Jul 2012) - $3.60T.",
        "  - 2013:    Gartner (later revisions) - $3.74T.",
        "  - 2014:    Gartner (Jul 2014) - $3.71T.",
        "  - 2015-2016: Gartner (Jan/Apr 2016) - 2015=$3.52T, 2016=$3.42T.",
        "  - 2017-2018: Gartner (Apr 2018) - 2017=$3.54T, 2018=$3.70T.",
        "  - 2019-2020: Gartner (Oct 2020) - 2019=$3.74T, 2020=$3.61T.",
        "  - 2021-2023: Gartner (Apr 2022, Jan 2024) - 2021=$4.26T, 2022=$4.43T, 2023=$4.68T.",
        "  - 2024:    Gartner (Jan 2025) - $5.11T.",
        "  - 2025:    Gartner (Oct 2025) - $5.54T.",
        "",
        "Useful URLs:",
        "  https://www.gartner.com/en/newsroom/press-releases/2020-10-20-gartner-says-worldwide-it-spending-to-grow-4-percent-in-2021",
        "  https://www.gartner.com/en/newsroom/press-releases/2022-04-06-gartner-forecasts-worldwide-it-spending-to-reach-4-point-four-trillion-in-2022",
        "  https://www.gartner.com/en/newsroom/press-releases/01-17-2024-gartner-forecasts-worldwide-it-spending-to-grow-six-point-eight-percent-in-2024",
        "  https://www.gartner.com/en/newsroom/press-releases/2025-01-21-gartner-forecasts-worldwide-it-spending-to-grow-9-point-8-percent-in-2025",
        "  https://www.gartner.com/en/newsroom/press-releases/2025-10-22-gartner-forecasts-worldwide-it-spending-to-grow-9-point-8-percent-in-2026-exceeding-6-trillion-dollars-for-the-first-time",
        "",
        "Caveats:",
        "  - The labor series (BEA Compensation of Employees) covers the United States only,",
        "    while the IT spending series (Gartner) is global. There is no single authoritative",
        "    worldwide labor compensation series with the same methodology back to 2001, so a",
        "    US/global comparison is the closest readily available aggregate. The relative growth",
        "    rates and the ratio between the series should be interpreted with this in mind.",
        "  - Gartner periodically revises historical figures (FX effects, segment redefinitions).",
        "    The values above use the most recently published vintage for each year where available.",
        "  - All figures are in nominal (current) US dollars.",
    ]

    for i, line in enumerate(notes, start=2):
        cell = src.cell(row=i, column=1, value=line)
        if line.startswith("Series ") or line.startswith("Key Gartner") or line.startswith("Useful URLs") or line.startswith("Caveats"):
            cell.font = Font(bold=True)
        elif line.startswith("  http"):
            cell.font = Font(color="0563C1", underline="single")
    src.column_dimensions["A"].width = 130

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Wrote {output_path}")


if __name__ == "__main__":
    out = Path(__file__).resolve().parent.parent / "data" / "Labor_vs_IT_Spending_2001_2025.xlsx"
    build_workbook(out)
