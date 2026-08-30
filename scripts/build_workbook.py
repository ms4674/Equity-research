#!/usr/bin/env python3
"""Build the blended token pricing Excel workbook.

Reads the price-history dataset in data/token_price_history.py and produces:
  * Blended_Token_Pricing_Frontier_and_Open_Weight_Models.xlsx (5 sheets + charts)
  * data/token_price_history.csv (flat export of the same dataset)

Run from the repo root:  python3 scripts/build_workbook.py
"""

import csv
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT / "data"))

from token_price_history import FIELDS, RECORDS, blended  # noqa: E402

AS_OF = date(2026, 8, 29)
XLSX_PATH = REPO_ROOT / "Blended_Token_Pricing_Frontier_and_Open_Weight_Models.xlsx"
CSV_PATH = REPO_ROOT / "data" / "token_price_history.csv"

PRICE_FMT = '"$"#,##0.00###'
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BLEND_FILL = PatternFill("solid", fgColor="DDEBF7")
TITLE_FONT = Font(bold=True, size=16, color="1F3864")
SUB_FONT = Font(bold=True, size=12, color="1F3864")

# --------------------------------------------------------------------------
# Lineages used for the carry-forward quarterly series. Each lineage is the
# succession of models that represented that vendor tier over time.
# --------------------------------------------------------------------------
FRONTIER_FLAGSHIP = {
    "OpenAI - GPT flagship": ["GPT-4", "GPT-4 Turbo", "GPT-4o", "GPT-5", "GPT-5.1", "GPT-5.2", "GPT-5.4", "GPT-5.5", "GPT-5.6 Sol"],
    "Anthropic - Opus line": ["Claude 2", "Claude 2.1", "Claude 3 Opus", "Claude Opus 4", "Claude Opus 4.1", "Claude Opus 4.5", "Claude Opus 4.6", "Claude Opus 4.7", "Claude Opus 4.8", "Claude Opus 5"],
    "Google - Gemini Pro": ["Gemini 1.5 Pro", "Gemini 2.5 Pro", "Gemini 3 Pro", "Gemini 3.1 Pro"],
    "xAI - Grok flagship": ["Grok Beta", "Grok 2", "Grok 3", "Grok 4", "Grok 4.20", "Grok 4.3", "Grok 4.5", "Grok 4.6"],
}
OPEN_FLAGSHIP = {
    "Meta - Llama flagship": ["Llama 2 70B", "Llama 3 70B", "Llama 3.1 405B", "Llama 4 Maverick"],
    "DeepSeek - V-series": ["DeepSeek V2", "DeepSeek V3", "DeepSeek V3.1", "DeepSeek V3.2", "DeepSeek V4 Pro"],
    "Alibaba - open Qwen": ["Qwen3 235B-A22B", "Qwen3.5 397B"],
    "Moonshot - Kimi": ["Kimi K2", "Kimi K2 Thinking", "Kimi K2.5", "Kimi K2.6", "Kimi K3"],
    "Zhipu - GLM": ["GLM-4.5", "GLM-4.6", "GLM-4.7", "GLM-5", "GLM-5.1", "GLM-5.2", "GLM-5.3"],
    "Mistral - open flagship": ["Mistral Large 2", "Mistral Large 3"],
}
BUDGET_TIER = {
    "OpenAI - budget": ["GPT-3.5 Turbo", "GPT-4o mini", "GPT-5 mini", "GPT-5.4 mini", "GPT-5.6 Luna"],
    "Anthropic - Instant/Haiku": ["Claude Instant", "Claude Instant 1.2", "Claude 3 Haiku", "Claude 3.5 Haiku", "Claude Haiku 4.5"],
    "Google - Flash": ["Gemini 1.5 Flash", "Gemini 2.0 Flash", "Gemini 2.5 Flash", "Gemini 3 Flash", "Gemini 3.5 Flash"],
}

# --------------------------------------------------------------------------
# Estimated average monthly LLM token consumption per company, by company
# size segment (millions of tokens per company per month). MODELED ESTIMATES,
# not vendor-reported data - calibrated to the anchors documented on the
# 'Token Usage by Size' sheet (Ramp AI Index, Deloitte via OpenRouter,
# OpenAI Enterprise Signals, VendorBenchmark; see Sources sheet).
# --------------------------------------------------------------------------
USAGE_COLUMNS = [
    "Small business (1-99 emp.)",
    "Mid-market (100-999 emp.)",
    "Enterprise (1,000-9,999 emp.)",
    "Large enterprise (10,000+ emp.)",
    "AI-native startup (memo)",
    "All companies - median (memo)",
]
# quarter -> tokens in millions per company per month
USAGE_SERIES = [
    ("2023 Q1", 1, 15, 90, 500, 60, 2),
    ("2023 Q2", 1.5, 22, 140, 800, 120, 3),
    ("2023 Q3", 2.2, 32, 210, 1300, 230, 4.5),
    ("2023 Q4", 3.2, 48, 320, 2000, 420, 6.5),
    ("2024 Q1", 4.5, 70, 480, 3100, 750, 9),
    ("2024 Q2", 6.5, 100, 720, 4800, 1300, 13),
    ("2024 Q3", 9, 145, 1050, 7200, 2200, 18),
    ("2024 Q4", 13, 210, 1600, 11000, 3700, 26),
    ("2025 Q1", 18, 300, 2400, 17000, 6000, 36),
    ("2025 Q2", 28, 470, 3800, 27000, 10000, 56),
    ("2025 Q3", 45, 750, 6200, 44000, 16500, 90),
    ("2025 Q4", 75, 1250, 10500, 74000, 27000, 150),
    ("2026 Q1", 130, 2150, 18000, 128000, 44000, 260),
    ("2026 Q2", 200, 3300, 28000, 200000, 68000, 400),
    ("2026 Q3", 280, 4600, 39000, 280000, 95000, 560),
]


def parse_date(s):
    return datetime.strptime(s, "%Y-%m-%d").date()


def quarter_label(d):
    return f"{d.year} Q{(d.month - 1) // 3 + 1}"


def quarter_ends(first, last):
    """All quarter-end dates from the quarter containing `first` to the one containing `last`."""
    out = []
    y, q = first.year, (first.month - 1) // 3 + 1
    while (y, q) <= (last.year, (last.month - 1) // 3 + 1):
        m = q * 3
        end_day = {3: 31, 6: 30, 9: 30, 12: 31}[m]
        out.append(date(y, m, end_day))
        q += 1
        if q == 5:
            y, q = y + 1, 1
    return out


def records_sorted():
    recs = [dict(zip(FIELDS, r)) for r in RECORDS]
    for r in recs:
        r["date_obj"] = parse_date(r["date"])
        r["blended"] = blended(r["input_usd_per_m"], r["output_usd_per_m"])
    recs.sort(key=lambda r: (r["date_obj"], r["developer"], r["model"]))
    return recs


def lineage_value_at(recs, models, when):
    """Blended price of the lineage's newest model as of `when` (carry-forward)."""
    launches = {}
    for r in recs:
        if r["model"] in models and r["date_obj"] <= when:
            launches.setdefault(r["model"], r["date_obj"])
    if not launches:
        return None
    current = max(launches, key=lambda m: launches[m])
    events = [r for r in recs if r["model"] == current and r["date_obj"] <= when]
    return events[-1]["blended"] if events else None


def style_header(ws, row, n_cols, height=28):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = height


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# --------------------------------------------------------------------------
def sheet_readme(wb):
    ws = wb.active
    ws.title = "README"
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 118])

    lines = [
        ("T", "Blended Token Pricing - Frontier & Open-Weight Models, Global"),
        ("", f"Compiled {AS_OF.isoformat()}. Time series of API list prices per 1 million tokens, 2020-2026."),
        ("", ""),
        ("S", "What 'blended price' means"),
        ("", "Blended $/1M tokens = (3 x input price + 1 x output price) / 4."),
        ("", "This is the standard 3:1 input:output usage weighting used by industry trackers (Artificial Analysis, ModelPriceWatch, AIMultiple), reflecting that typical workloads consume roughly three input tokens for every output token."),
        ("", ""),
        ("S", "Scope"),
        ("", "- Frontier (proprietary): closed API models from OpenAI, Anthropic, Google, xAI, Meta (Muse Spark), Amazon, Cohere, Mistral (commercial line), and Alibaba's Max line."),
        ("", "- Open-weight: models with downloadable weights (Llama, DeepSeek, Qwen open line, Kimi, GLM, MiniMax, Mistral open line, gpt-oss). Priced at the developer's first-party API rate where one exists; otherwise at a reference hosted rate (noted in the Basis column), since open weights themselves are free."),
        ("", "- Geography: USA (OpenAI, Anthropic, Google, xAI, Meta, Amazon), China (DeepSeek, Alibaba, Moonshot, Zhipu, MiniMax), France (Mistral), Canada (Cohere)."),
        ("", ""),
        ("S", "Sheets"),
        ("", "- Price History: long-format time series - one row per price event (launch, cut, increase) per model. ~160 observations."),
        ("", "- Current Snapshot: the latest standard list price for every model as of the compile date, ranked by blended price."),
        ("", "- Quarterly Series: carry-forward blended price of each vendor's flagship / budget lineage at each quarter end (the price of the newest model in that lineage at that date). Feeds the charts."),
        ("", "- Token Usage by Size: quarterly time series of estimated average monthly token consumption per company, by company-size segment (modeled estimates calibrated to Ramp AI Index, Deloitte, OpenAI Enterprise Signals, and VendorBenchmark anchors; methodology on the sheet)."),
        ("", "- Charts: blended price trajectories on a log scale - frontier flagships, open-weight flagships, and budget tiers."),
        ("", "- Sources: pricing pages, launch announcements, and trackers used."),
        ("", ""),
        ("S", "Methodology notes & caveats"),
        ("", "1. Prices are standard pay-as-you-go list rates for the base (short) context tier. Long-context surcharges, cached-input rates, batch (-50%), priority/fast modes, and free tiers are excluded (noted where material)."),
        ("", "2. Reasoning models bill hidden reasoning tokens at the output rate; the blended figure understates their effective cost per visible token."),
        ("", "3. Chinese CNY list prices converted at the rates published by the trackers used (~6.75-7.1 CNY/USD); conversion noted in the Basis column."),
        ("", "4. Date precision: 'approx' marks launch dates estimated to the month from secondary sources; 'exact' dates come from vendor announcements."),
        ("", "5. Promotional prices are recorded as events and flagged (e.g. GPT-5.6 Sol's $4/$20 promo through Nov 2026; DeepSeek's off-peak windows)."),
        ("", "6. Open-weight hosted rates vary 2-4x across hosts; the Basis column names the reference host. Third-party host rates for the same model differ from the first-party rate shown."),
        ("", "7. Tokenizers differ across vendors (e.g. Claude Sonnet 5 produces ~30% more tokens for the same text), so $/token is not a perfect cross-vendor unit of work."),
        ("", ""),
        ("S", "Headline findings"),
        ("", "- Frontier flagship blended price fell from $37.50 (GPT-4, Mar 2023) to a floor of ~$3.44 (GPT-5, Aug 2025), then REBOUNDED to $8-11+ in 2026 as reasoning-heavy flagships (GPT-5.5/5.6 at $5/$30, Claude Fable 5 at $10/$50) reset the ceiling."),
        ("", "- The budget tier collapsed >99%: $2.00 blended (GPT-3.5 Turbo, Mar 2023) to $0.20-0.45 (GPT-5.6 Luna, Gemini Flash-Lite class, 2026)."),
        ("", "- Open-weight flagships remain 3-10x cheaper than closed flagships, but their launch prices are rising (Kimi K3 at $6.00 blended is the most expensive open-weight launch to date)."),
        ("", "- China's open-weight labs (DeepSeek, Qwen, Kimi, GLM) anchor the global price floor; DeepSeek V4 Flash off-peak reaches $0.33 blended."),
    ]
    row = 1
    for kind, text in lines:
        cell = ws.cell(row=row, column=2, value=text)
        if kind == "T":
            cell.font = TITLE_FONT
            ws.row_dimensions[row].height = 24
        elif kind == "S":
            cell.font = SUB_FONT
            ws.row_dimensions[row].height = 18
        else:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if len(text) > 110:
                ws.row_dimensions[row].height = 30
        row += 1


def sheet_history(wb, recs):
    ws = wb.create_sheet("Price History")
    headers = ["Effective date", "Quarter", "Developer", "Country", "Model", "Class", "Tier",
               "Event", "Input $/1M", "Output $/1M", "Blended $/1M (3:1)", "Pricing basis",
               "Date precision", "Notes"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for i, r in enumerate(recs, start=2):
        ws.cell(row=i, column=1, value=r["date_obj"]).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=2, value=quarter_label(r["date_obj"]))
        ws.cell(row=i, column=3, value=r["developer"])
        ws.cell(row=i, column=4, value=r["country"])
        ws.cell(row=i, column=5, value=r["model"])
        ws.cell(row=i, column=6, value=r["model_class"])
        ws.cell(row=i, column=7, value=r["tier"])
        ws.cell(row=i, column=8, value=r["event"])
        ws.cell(row=i, column=9, value=r["input_usd_per_m"]).number_format = PRICE_FMT
        ws.cell(row=i, column=10, value=r["output_usd_per_m"]).number_format = PRICE_FMT
        b = ws.cell(row=i, column=11, value=f"=(3*I{i}+J{i})/4")
        b.number_format = PRICE_FMT
        b.fill = BLEND_FILL
        b.font = Font(bold=True)
        ws.cell(row=i, column=12, value=r["basis"])
        ws.cell(row=i, column=13, value=r["date_precision"])
        ws.cell(row=i, column=14, value=r["notes"])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{len(recs) + 1}"
    set_widths(ws, [12, 9, 13, 9, 24, 20, 17, 15, 11, 11, 15, 30, 12, 58])


def sheet_snapshot(wb, recs):
    ws = wb.create_sheet("Current Snapshot")
    latest = {}
    for r in recs:
        if r["date_obj"] <= AS_OF:
            key = (r["developer"], r["model"])
            if key not in latest or r["date_obj"] > latest[key]["date_obj"]:
                latest[key] = r
    rows = sorted(latest.values(), key=lambda r: r["blended"])

    ws.cell(row=1, column=1, value=f"Latest standard list price per model, as of {AS_OF.isoformat()} - ranked cheapest to most expensive by blended price").font = SUB_FONT
    headers = ["Rank", "Developer", "Country", "Model", "Class", "Tier", "Input $/1M",
               "Output $/1M", "Blended $/1M (3:1)", "Price as of", "Pricing basis", "Notes"]
    ws.append(headers)
    style_header(ws, 2, len(headers))
    for i, r in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=i - 2)
        ws.cell(row=i, column=2, value=r["developer"])
        ws.cell(row=i, column=3, value=r["country"])
        ws.cell(row=i, column=4, value=r["model"])
        ws.cell(row=i, column=5, value=r["model_class"])
        ws.cell(row=i, column=6, value=r["tier"])
        ws.cell(row=i, column=7, value=r["input_usd_per_m"]).number_format = PRICE_FMT
        ws.cell(row=i, column=8, value=r["output_usd_per_m"]).number_format = PRICE_FMT
        b = ws.cell(row=i, column=9, value=f"=(3*G{i}+H{i})/4")
        b.number_format = PRICE_FMT
        b.fill = BLEND_FILL
        b.font = Font(bold=True)
        ws.cell(row=i, column=10, value=r["date_obj"]).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=11, value=r["basis"])
        ws.cell(row=i, column=12, value=r["notes"])
    n = len(rows)
    ws.conditional_formatting.add(
        f"I3:I{n + 2}",
        DataBarRule(start_type="min", end_type="max", color="2E75B6", showValue=True),
    )
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:L{n + 2}"
    set_widths(ws, [6, 13, 9, 24, 20, 17, 11, 11, 15, 12, 30, 58])
    return n


def sheet_quarterly(wb, recs):
    ws = wb.create_sheet("Quarterly Series")
    first = min(r["date_obj"] for r in recs if r["date_obj"] >= date(2023, 1, 1))
    qends = quarter_ends(first, AS_OF)

    groups = [("FRONTIER FLAGSHIPS (proprietary)", FRONTIER_FLAGSHIP),
              ("OPEN-WEIGHT FLAGSHIPS", OPEN_FLAGSHIP),
              ("BUDGET / SMALL TIER (proprietary)", BUDGET_TIER)]

    ws.cell(row=1, column=1, value="Blended $/1M tokens (3:1) of each lineage's newest model at quarter end - carry-forward series").font = SUB_FONT

    # Group banner row (row 2) and lineage header row (row 3).
    col = 2
    group_spans = []
    lineages = []
    for gname, gmap in groups:
        start = col
        for lname, models in gmap.items():
            lineages.append((lname, models))
            col += 1
        group_spans.append((gname, start, col - 1))
    ncols = col - 1

    for gname, c0, c1 in group_spans:
        ws.merge_cells(start_row=2, start_column=c0, end_row=2, end_column=c1)
        cell = ws.cell(row=2, column=c0, value=gname)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        fill = PatternFill("solid", fgColor={0: "C00000", 1: "2E7D32", 2: "7F6000"}[group_spans.index((gname, c0, c1))])
        for c in range(c0, c1 + 1):
            ws.cell(row=2, column=c).fill = fill

    ws.cell(row=3, column=1, value="Quarter")
    for j, (lname, _) in enumerate(lineages, start=2):
        ws.cell(row=3, column=j, value=lname)
    style_header(ws, 3, ncols, height=42)

    for i, qe in enumerate(qends, start=4):
        ws.cell(row=i, column=1, value=quarter_label(qe)).font = Font(bold=True)
        for j, (_, models) in enumerate(lineages, start=2):
            v = lineage_value_at(recs, models, qe)
            if v is not None:
                ws.cell(row=i, column=j, value=round(v, 4)).number_format = PRICE_FMT

    note_row = 4 + len(qends) + 1
    ws.cell(row=note_row, column=1,
            value="Note: each column tracks the succession of models in one vendor lineage (see scripts/build_workbook.py for the mapping). "
                  "A value is the blended list price of the newest lineage model as of that quarter end, including any price cuts in effect. "
                  f"{quarter_label(AS_OF)} is partial (through {AS_OF.isoformat()}).").alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=min(ncols, 10))
    ws.row_dimensions[note_row].height = 42
    ws.freeze_panes = "B4"
    set_widths(ws, [10] + [15] * (ncols - 1))
    return len(qends), lineages, group_spans


def sheet_usage(wb):
    ws = wb.create_sheet("Token Usage by Size")
    ws.cell(row=1, column=1, value="Estimated average monthly LLM token consumption per company, by company size - millions of tokens per company per month").font = SUB_FONT
    ws.cell(row=2, column=1, value="MODELED ESTIMATES calibrated to published anchors (see notes below) - not vendor-reported data. Includes API and subscription-mediated usage.").font = Font(italic=True, color="C00000")

    headers = ["Quarter"] + USAGE_COLUMNS
    ws.append([])  # row 3 spacer consumed by append offset below
    for j, h in enumerate(headers, start=1):
        ws.cell(row=4, column=j, value=h)
    style_header(ws, 4, len(headers), height=42)

    first_data_row = 5
    for i, row in enumerate(USAGE_SERIES, start=first_data_row):
        ws.cell(row=i, column=1, value=row[0]).font = Font(bold=True)
        for j, v in enumerate(row[1:], start=2):
            c = ws.cell(row=i, column=j, value=v)
            c.number_format = "#,##0.0" if v < 10 else "#,##0"
    last_data_row = first_data_row + len(USAGE_SERIES) - 1

    n = len(USAGE_SERIES)
    ws.conditional_formatting.add(
        f"B{first_data_row}:E{last_data_row}",
        DataBarRule(start_type="min", end_type="max", color="70AD47", showValue=True),
    )

    notes = [
        "How this series was built:",
        "1. Anchor (2026 Q2): Ramp AI Index (Apr-Jun 2026, 70,000+ US businesses) - median company token spend $2,246/month at an observed effective rate of $0.72 per 1M tokens "
        "implies ~3.1B tokens/month for the median Ramp-tracked (mid-market, tech-forward) company; average spend $140,842/month implies ~200B tokens for the large-enterprise average.",
        "2. Growth path: Ramp reports token usage among businesses with connected AI grew 1,001% (~11x) from Jan 2025 to Apr 2026; each segment's series is backcast at that pace for 2025-2026 "
        "and at slower adoption-era rates for 2023-2024 (ChatGPT API launched Mar 2023).",
        "3. Cross-checks: Deloitte 2026 (cited by OpenRouter): 67% of enterprises consume >1B tokens/month - consistent with the Enterprise column crossing 1B in 2024 and reaching ~28-39B in 2026. "
        "VendorBenchmark 2026: enterprise LLM API spend of $4-22 per employee/month supports the Enterprise and Large-enterprise levels. "
        "OpenAI Enterprise Signals (Jun 2026): top-decile 'frontier firms' generate 8.3x the output tokens per active user of typical firms - reflected in the AI-native startup memo column.",
        "4. Segment averages are means; token usage is power-law distributed within every segment (Ramp: median $2,246 vs average $140,842 monthly spend), so a segment's typical company sits "
        "well below its average. The 'All companies - median' memo column is the better 'typical firm' line.",
        "5. Usage includes tokens consumed through subscription products (ChatGPT/Claude/Copilot seats) valued at effective token rates, not only direct API billing.",
        "Caveats: no vendor publishes tokens-by-company-size; treat levels as order-of-magnitude estimates. Growth rates and 2026 anchors are well-sourced; 2023-2024 levels are extrapolations. "
        "Aggregate context: OpenRouter alone routed ~100T tokens/month (May 2026, 5x in 6 months); Google reported >3.2 quadrillion tokens/month across products (I/O 2026); Meta burned ~70T/month (Feb 2026, SemiAnalysis).",
    ]
    note_row = last_data_row + 2
    for k, text in enumerate(notes):
        cell = ws.cell(row=note_row + k, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=note_row + k, start_column=1, end_row=note_row + k, end_column=7)
        ws.row_dimensions[note_row + k].height = 15 if len(text) < 120 else 45
    ws.cell(row=note_row, column=1).font = Font(bold=True)

    chart = LineChart()
    chart.title = "Avg monthly tokens per company by size (millions, log scale)"
    chart.style = 12
    chart.height = 12
    chart.width = 26
    chart.y_axis.title = "Millions of tokens per company per month"
    chart.x_axis.title = "Quarter"
    chart.y_axis.scaling.logBase = 10
    chart.y_axis.majorGridlines = None
    for c in range(2, 2 + len(USAGE_COLUMNS)):
        ref = Reference(ws, min_col=c, min_row=4, max_row=last_data_row)
        chart.add_data(ref, titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=first_data_row, max_row=last_data_row))
    for s in chart.series:
        s.marker = Marker(symbol="circle", size=5)
        s.smooth = False
    ws.add_chart(chart, "I4")

    ws.freeze_panes = "B5"
    set_widths(ws, [10, 15, 15, 16, 17, 15, 16])


def sheet_charts(wb, n_quarters, group_spans):
    ws = wb.create_sheet("Charts")
    ws.sheet_view.showGridLines = False
    src = wb["Quarterly Series"]
    first_data_row, last_data_row = 4, 3 + n_quarters
    cats = Reference(src, min_col=1, min_row=first_data_row, max_row=last_data_row)

    titles = {
        "FRONTIER FLAGSHIPS (proprietary)": "Frontier flagship models - blended $/1M tokens (log scale)",
        "OPEN-WEIGHT FLAGSHIPS": "Open-weight flagship models - blended $/1M tokens (log scale)",
        "BUDGET / SMALL TIER (proprietary)": "Budget / small tier - blended $/1M tokens (log scale)",
    }
    anchor_row = 2
    for gname, c0, c1 in group_spans:
        chart = LineChart()
        chart.title = titles[gname]
        chart.style = 12
        chart.height = 11
        chart.width = 30
        chart.y_axis.title = "$ per 1M tokens (blended 3:1)"
        chart.x_axis.title = "Quarter"
        chart.y_axis.scaling.logBase = 10
        chart.y_axis.majorGridlines = None
        for c in range(c0, c1 + 1):
            ref = Reference(src, min_col=c, min_row=3, max_row=last_data_row)
            chart.add_data(ref, titles_from_data=True)
        chart.set_categories(cats)
        for s in chart.series:
            s.marker = Marker(symbol="circle", size=5)
            s.smooth = False
        ws.add_chart(chart, f"B{anchor_row}")
        anchor_row += 24
    ws.cell(row=anchor_row, column=2,
            value="Y axes are logarithmic. Gaps before a lineage's first launch are blank. Source: Quarterly Series sheet.")


def sheet_sources(wb):
    ws = wb.create_sheet("Sources")
    headers = ["Source", "Used for", "URL", "Accessed"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    sources = [
        ("OpenAI - pricing page & model docs", "GPT / o-series / gpt-oss current prices, GPT-5.6 tiers", "https://developers.openai.com/api/docs/pricing", "2026-08-29"),
        ("OpenAI - Introducing GPT-5.4", "GPT-5.2 / 5.4 / Pro launch prices, launch dates", "https://openai.com/index/introducing-gpt-5-4/", "2026-08-29"),
        ("OpenAI - Introducing GPT-5.5", "GPT-5.5 / 5.5 Pro launch prices", "https://openai.com/index/introducing-gpt-5-5/", "2026-08-29"),
        ("Anthropic - Claude platform pricing", "All current + retired Claude model prices", "https://platform.claude.com/docs/en/about-claude/pricing", "2026-08-29"),
        ("Anthropic - Claude release notes & model overview", "Claude launch dates (Sonnet 5, Opus 5, Fable 5, Opus 4.x)", "https://platform.claude.com/docs/en/release-notes/overview", "2026-08-29"),
        ("Google - Gemini API pricing (via metacto/eesel summaries)", "Gemini 2.5 / 3 / 3.1 / 3.5 prices", "https://www.metacto.com/blogs/the-true-cost-of-google-gemini-a-guide-to-api-pricing-and-integration", "2026-08-29"),
        ("xAI - developer pricing & release notes", "Grok 4.20-4.6 prices and context tiers", "https://docs.x.ai/developers/pricing", "2026-08-29"),
        ("DeepSeek API docs (via packet.ai summary)", "V4 Flash/Pro peak & off-peak pricing, 2026-08-16 restructure", "https://packet.ai/blog/open-source-llm-api", "2026-08-29"),
        ("LLM Abacus - Chinese LLM API pricing comparison", "DeepSeek / Qwen / Kimi / GLM CNY list prices + USD conversions", "https://www.llmabacus.com/en/chinese-llm-api-pricing", "2026-08-29"),
        ("Tencent Cloud TokenHub model prices", "Kimi K2.5-K3, GLM 5.x, Qwen 3.5/3.6, MiniMax M3 CNY prices", "https://cloud.tencent.cn/document/product/1823/130055", "2026-08-29"),
        ("Alibaba Model Studio pricing", "Qwen 3.7/3.8 Max prices and snapshot dates", "https://help.aliyun.com/zh/model-studio/model-pricing", "2026-08-29"),
        ("Layer3 Labs - AI model pricing table (verified 2026-07-31)", "Cross-vendor current snapshot incl. Mistral, MiniMax, Amazon, Cohere", "https://www.layer3labs.io/ai-model-pricing", "2026-08-29"),
        ("BenchLM - OpenAI API pricing registry", "GPT-5.6 launch/cut dates and rates", "https://benchlm.ai/openai/api-pricing", "2026-08-29"),
        ("AIMultiple - LLM pricing (blended launch-price series)", "3:1 blending convention; open vs closed launch trends; Muse Spark, Kimi K3", "https://aimultiple.com/llm-pricing", "2026-08-29"),
        ("AI Cost Check / aipricing.guru / techjack", "Llama hosted-rate ranges across providers", "https://aicostcheck.com/provider/meta", "2026-08-29"),
        ("hidekazu-konishi.com - OpenAI GPT model release timeline", "OpenAI launch dates 2023-2026", "https://hidekazu-konishi.com/entry/openai_gpt_model_release_timeline.html", "2026-08-29"),
        ("arXiv 2603.28576 - Tiered Super-Moore's Law", "Historical context: 600x price decline, tier half-lives", "https://arxiv.org/abs/2603.28576", "2026-08-29"),
        ("Ramp - AI token cost benchmarks (AI Index, Apr-Jun 2026)", "Token Usage sheet anchors: $0.72/1M effective rate, median/average company spend, 1,001% usage growth Jan 2025-Apr 2026, PEPM benchmarks", "https://ramp.com/blog/ai-token-cost-for-businesses", "2026-08-30"),
        ("OpenAI - Enterprise Signals", "Token Usage sheet: frontier firms generate 8.3x output tokens per active user vs typical firms (Jun 2026)", "https://openai.com/signals/enterprise-data/", "2026-08-30"),
        ("TechStartups - OpenRouter Series B (Deloitte 2026 study cited)", "Token Usage sheet: 67% of enterprises consume >1B tokens/month; OpenRouter ~100T tokens/month, 5x in 6 months", "https://techstartups.com/2026/05/26/openrouter-raises-113m-as-ai-token-usage-surges-to-100-trillion-monthly/", "2026-08-30"),
        ("VendorBenchmark - Enterprise GenAI cost benchmark 2026", "Token Usage sheet: enterprise LLM API spend $4-22 per employee/month; segment PEPM ranges", "https://vendorbenchmark.com/guides/enterprise-genai-cost-benchmark-2026", "2026-08-30"),
        ("SemiAnalysis - TokenBudgeting newsletter", "Token Usage sheet: Meta ~70T tokens/month (Feb 2026); enterprise spend percentiles", "https://newsletter.semianalysis.com/p/tokenbudgeting-our-conversations", "2026-08-30"),
        ("Trending Topics EU - open-weight share of AI usage", "Aggregate token volumes: Google >3.2 quadrillion tokens/month (I/O 2026), OpenAI API ~260T/month (Oct 2025)", "https://www.trendingtopics.eu/open-weight-models-from-china-are-capturing-a-growing-share-of-ai-usage/", "2026-08-30"),
        ("Vendor announcements 2023-2025 (compiled)", "Historical launch prices: GPT-4/Turbo/4o, Claude 2/3/4, Gemini 1.5-2.5, Grok 2-4, Llama 2-4, DeepSeek V2-V3.2, Kimi K2, GLM-4.5+", "various (see notes column in Price History)", "2026-08-29"),
    ]
    for i, s in enumerate(sources, start=2):
        for j, v in enumerate(s, start=1):
            ws.cell(row=i, column=j, value=v)
    ws.freeze_panes = "A2"
    set_widths(ws, [52, 62, 72, 12])


def write_csv(recs):
    with open(CSV_PATH, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(FIELDS + ["blended_usd_per_m_3to1"])
        for r in recs:
            w.writerow([r[k] for k in FIELDS] + [round(r["blended"], 4)])


def main():
    recs = records_sorted()
    wb = Workbook()
    sheet_readme(wb)
    sheet_history(wb, recs)
    n_models = sheet_snapshot(wb, recs)
    n_quarters, lineages, group_spans = sheet_quarterly(wb, recs)
    sheet_usage(wb)
    sheet_charts(wb, n_quarters, group_spans)
    sheet_sources(wb)
    wb.save(XLSX_PATH)
    write_csv(recs)
    print(f"Wrote {XLSX_PATH.name}: {len(recs)} price events, {n_models} models, "
          f"{n_quarters} quarters, {len(lineages)} lineage series")
    print(f"Wrote {CSV_PATH.relative_to(REPO_ROOT)}")


if __name__ == "__main__":
    main()
