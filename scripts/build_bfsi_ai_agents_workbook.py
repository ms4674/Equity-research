"""
Build an XLSX workbook aggregating publicly available evidence on where AI agents
are most effective in banking and financial services (BFSI).

Output: data/bfsi_ai_agent_effectiveness.xlsx

Sheets:
  1. README              - methodology, scoring rubric, disclaimer
  2. Effectiveness_Rank  - ranked use cases with composite effectiveness score
  3. Evidence_Log        - individual quantified data points with sources
  4. Vendor_Map          - foundation-model and platform vendors observed per use case
  5. Sources             - bibliography of cited public sources

All figures are taken from publicly disclosed company statements, regulatory
papers, or major consultancy / press reports. Where a figure is a range or an
analyst estimate it is labelled as such. Nothing in this workbook is investment
advice.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


OUT_PATH = Path(__file__).resolve().parents[1] / "data" / "bfsi_ai_agent_effectiveness.xlsx"


# ---------------------------------------------------------------------------
# Source bibliography (id -> (short, url, date))
# ---------------------------------------------------------------------------
SOURCES: list[dict] = [
    {
        "id": "S1",
        "short": "Klarna press release - 'Klarna AI assistant handles two-thirds of customer service chats in its first month'",
        "publisher": "Klarna",
        "date": "Feb 2024",
        "url": "https://www.klarna.com/international/press/klarna-ai-assistant-handles-two-thirds-of-customer-service-chats-in-its-first-month/",
    },
    {
        "id": "S2",
        "short": "JPMorgan Chase 2017 annual report and press coverage of COiN (Contract Intelligence)",
        "publisher": "JPMorgan / Bloomberg / Reuters",
        "date": "2017-2018",
        "url": "https://www.bloomberg.com/news/articles/2017-02-28/jpmorgan-marshals-an-army-of-developers-to-automate-high-finance",
    },
    {
        "id": "S3",
        "short": "JPMorgan internal LLM Suite rollout to ~200,000 employees (Reuters / FT reporting; JPM CEO letter)",
        "publisher": "Reuters / Financial Times / JPM 2023 CEO letter",
        "date": "2024-2025",
        "url": "https://www.reuters.com/technology/artificial-intelligence/jpmorgan-rolls-out-ai-chatbot-tens-thousands-employees-2024-08-09/",
    },
    {
        "id": "S4",
        "short": "Bank of America - Erica virtual assistant milestones (2bn+ interactions, 42m+ users)",
        "publisher": "Bank of America newsroom",
        "date": "Apr 2024",
        "url": "https://newsroom.bankofamerica.com/content/newsroom/press-releases/2024/04/bank-of-america-s-erica--surpasses-2-billion-interactions--helpi.html",
    },
    {
        "id": "S5",
        "short": "Morgan Stanley + OpenAI AI @ Morgan Stanley Assistant for Wealth Management advisors",
        "publisher": "Morgan Stanley / OpenAI case study",
        "date": "Sep 2023 - 2024",
        "url": "https://www.morganstanley.com/press-releases/key-milestone-in-innovation-journey-with-openai",
    },
    {
        "id": "S6",
        "short": "Goldman Sachs - GS AI Assistant rollout to ~10,000 employees; CIO Marco Argenti interviews",
        "publisher": "Reuters / CNBC / Goldman Sachs",
        "date": "Jun 2024 - 2025",
        "url": "https://www.reuters.com/technology/artificial-intelligence/goldman-sachs-rolls-out-generative-ai-assistant-its-bankers-2024-06-13/",
    },
    {
        "id": "S7",
        "short": "Mastercard - Decision Intelligence Pro generative AI; 'doubled' fraud detection in some segments",
        "publisher": "Mastercard newsroom",
        "date": "Feb 2024",
        "url": "https://www.mastercard.com/news/press/2024/february/mastercard-harnesses-generative-ai-to-fight-payment-scams/",
    },
    {
        "id": "S8",
        "short": "Visa - $10bn invested in tech/AI over 5 years; ~$40bn fraud prevented in 2023 by AI/ML systems",
        "publisher": "Visa newsroom / 10-K",
        "date": "2023-2024",
        "url": "https://usa.visa.com/about-visa/newsroom/press-releases/visa-prevented-40-billion-in-fraudulent-activity-in-2023.html",
    },
    {
        "id": "S9",
        "short": "Wells Fargo - Fargo virtual assistant interactions (~245m in 2024)",
        "publisher": "Wells Fargo / Charlie Scharf earnings commentary",
        "date": "2024-2025",
        "url": "https://newsroom.wf.com/English/news-releases/news-release-details/2024/Wells-Fargo-Fargo-Virtual-Assistant-Hits-Milestone/default.aspx",
    },
    {
        "id": "S10",
        "short": "McKinsey - 'The economic potential of generative AI' (banking value pool $200-340bn / yr)",
        "publisher": "McKinsey & Company",
        "date": "Jun 2023; updated 2024",
        "url": "https://www.mckinsey.com/capabilities/mckinsey-digital/our-insights/the-economic-potential-of-generative-ai-the-next-productivity-frontier",
    },
    {
        "id": "S11",
        "short": "Citi GPS - 'AI in Finance: Bot, Bank & Beyond' (productivity, profit pool impact)",
        "publisher": "Citi Global Perspectives & Solutions",
        "date": "Jun 2024",
        "url": "https://www.citigroup.com/global/insights/citigps/ai-in-finance",
    },
    {
        "id": "S12",
        "short": "BBVA - 3,000 ChatGPT Enterprise licenses; 80% of users report time savings",
        "publisher": "OpenAI customer story / BBVA",
        "date": "May 2024",
        "url": "https://openai.com/index/bbva/",
    },
    {
        "id": "S13",
        "short": "Stripe Radar - blocked tens of billions in fraudulent transactions (annual disclosure)",
        "publisher": "Stripe",
        "date": "2023-2024",
        "url": "https://stripe.com/radar",
    },
    {
        "id": "S14",
        "short": "Anthropic / AWS Bedrock - Claude in financial services (Pfizer, NIB, Bridgewater AIA Labs)",
        "publisher": "Anthropic case studies",
        "date": "2024-2025",
        "url": "https://www.anthropic.com/customers",
    },
    {
        "id": "S15",
        "short": "ING + McKinsey QuantumBlack - Gen AI customer service chatbot (20% rise in customers helped, expanded servicing capacity)",
        "publisher": "McKinsey case study with ING",
        "date": "2024",
        "url": "https://www.mckinsey.com/about-us/new-at-mckinsey-blog/how-ing-uses-generative-ai-to-help-its-customer-service-staff",
    },
    {
        "id": "S16",
        "short": "NVIDIA - 'State of AI in Financial Services' 2024 industry survey",
        "publisher": "NVIDIA",
        "date": "2024",
        "url": "https://www.nvidia.com/en-us/industries/finance/ai-financial-services-report/",
    },
    {
        "id": "S17",
        "short": "Deloitte - 'State of Generative AI in the Enterprise' wave reports (financial services cut)",
        "publisher": "Deloitte",
        "date": "2024-2025",
        "url": "https://www2.deloitte.com/us/en/pages/consulting/articles/state-of-generative-ai-in-enterprise.html",
    },
    {
        "id": "S18",
        "short": "EY - 'How artificial intelligence is reshaping the financial services industry'",
        "publisher": "EY",
        "date": "2024",
        "url": "https://www.ey.com/en_gl/insights/financial-services/how-artificial-intelligence-is-reshaping-the-financial-services-industry",
    },
    {
        "id": "S19",
        "short": "Commonwealth Bank of Australia - generative AI scam/fraud reduction; messaging triage",
        "publisher": "CBA / AWS case study",
        "date": "2024",
        "url": "https://www.commbank.com.au/articles/newsroom/2024/06/CBA-AI-scam-protection.html",
    },
    {
        "id": "S20",
        "short": "Allstate - generative AI for claims communications (~50,000 letters/day drafted)",
        "publisher": "Allstate / Wall Street Journal",
        "date": "2024",
        "url": "https://www.wsj.com/articles/allstate-uses-generative-ai-to-write-claims-letters-49b6a7ee",
    },
]

SOURCE_BY_ID = {s["id"]: s for s in SOURCES}


# ---------------------------------------------------------------------------
# Use-case effectiveness rank (the headline answer)
# ---------------------------------------------------------------------------
# Scoring rubric (1 = low, 5 = high):
#   ImpactMagnitude   - size of cost / revenue / risk impact when deployed
#   EvidenceStrength  - quality and number of public quantified disclosures
#   DeploymentMaturity- how many BFSI institutions are in production today
#   ROI_Speed         - typical time to measurable financial benefit
#   RegRiskInverse    - inverse of regulatory friction (5 = low friction)
# Composite = weighted average (Impact 0.30, Evidence 0.25, Maturity 0.20,
#                                ROI_Speed 0.15, RegRiskInverse 0.10)

EFFECTIVENESS: list[dict] = [
    {
        "Rank": None,
        "Domain": "Retail banking",
        "UseCase": "Customer service virtual assistants & chat agents",
        "RepresentativeKPI": "Containment %, AHT reduction, CSAT, cost-per-contact",
        "TypicalQuantifiedImpact": "30-65% deflection; 25-50% AHT reduction; 20-40% cost-per-contact reduction",
        "ImpactMagnitude": 5,
        "EvidenceStrength": 5,
        "DeploymentMaturity": 5,
        "ROI_Speed": 5,
        "RegRiskInverse": 4,
        "KeyEvidenceIds": "S1,S4,S9,S15",
    },
    {
        "Rank": None,
        "Domain": "Payments / Card networks",
        "UseCase": "Fraud detection & transaction risk scoring (ML + GenAI overlay)",
        "RepresentativeKPI": "$ fraud prevented, false-positive rate, approval rate uplift",
        "TypicalQuantifiedImpact": "Visa: ~$40bn fraud prevented (2023); Mastercard: 'doubled' detection in tested segments; Stripe Radar: tens of $bn blocked",
        "ImpactMagnitude": 5,
        "EvidenceStrength": 5,
        "DeploymentMaturity": 5,
        "ROI_Speed": 4,
        "RegRiskInverse": 4,
        "KeyEvidenceIds": "S7,S8,S13,S19",
    },
    {
        "Rank": None,
        "Domain": "Risk & compliance",
        "UseCase": "AML / KYC alert triage & SAR drafting",
        "RepresentativeKPI": "Alerts cleared/analyst-day, false-positive reduction, narrative quality",
        "TypicalQuantifiedImpact": "3-5x L1 analyst productivity; 30-60% false-positive reduction in pilots",
        "ImpactMagnitude": 5,
        "EvidenceStrength": 3,
        "DeploymentMaturity": 4,
        "ROI_Speed": 4,
        "RegRiskInverse": 2,
        "KeyEvidenceIds": "S10,S11,S16,S18",
    },
    {
        "Rank": None,
        "Domain": "Wealth & advisory",
        "UseCase": "Advisor knowledge assistant & meeting prep (RAG over research)",
        "RepresentativeKPI": "Hours saved/advisor/week, adoption %, NPS",
        "TypicalQuantifiedImpact": "Morgan Stanley: 98% advisor-team adoption; ~5-15 hrs/wk saved; faster client response",
        "ImpactMagnitude": 4,
        "EvidenceStrength": 5,
        "DeploymentMaturity": 4,
        "ROI_Speed": 5,
        "RegRiskInverse": 3,
        "KeyEvidenceIds": "S5,S11",
    },
    {
        "Rank": None,
        "Domain": "Capital markets / Banking ops",
        "UseCase": "Document intelligence (loan agreements, ISDA, prospectuses, KYC docs)",
        "RepresentativeKPI": "Hours saved, error rate, cycle time",
        "TypicalQuantifiedImpact": "JPMorgan COiN: ~360,000 lawyer-hours/yr saved on commercial loan agreements",
        "ImpactMagnitude": 4,
        "EvidenceStrength": 4,
        "DeploymentMaturity": 4,
        "ROI_Speed": 4,
        "RegRiskInverse": 3,
        "KeyEvidenceIds": "S2,S6,S14",
    },
    {
        "Rank": None,
        "Domain": "Internal productivity (cross-functional)",
        "UseCase": "Enterprise GenAI copilots for employees (LLM Suite, GS AI Assistant, BBVA ChatGPT)",
        "RepresentativeKPI": "Active users, time saved/user/week, adoption %",
        "TypicalQuantifiedImpact": "JPM: ~200k employees rolled out; GS: ~10k+; BBVA: 80% of 3,000 users report time savings",
        "ImpactMagnitude": 4,
        "EvidenceStrength": 5,
        "DeploymentMaturity": 5,
        "ROI_Speed": 4,
        "RegRiskInverse": 4,
        "KeyEvidenceIds": "S3,S6,S12,S17",
    },
    {
        "Rank": None,
        "Domain": "Software engineering inside banks",
        "UseCase": "AI coding assistants (Copilot/Cursor/Claude Code) for developers",
        "RepresentativeKPI": "PRs/dev/week, code accepted %, defect rate",
        "TypicalQuantifiedImpact": "20-40% developer productivity uplift reported across BFSI pilots",
        "ImpactMagnitude": 4,
        "EvidenceStrength": 4,
        "DeploymentMaturity": 5,
        "ROI_Speed": 5,
        "RegRiskInverse": 4,
        "KeyEvidenceIds": "S6,S16,S17",
    },
    {
        "Rank": None,
        "Domain": "Insurance",
        "UseCase": "Claims FNOL & customer correspondence drafting",
        "RepresentativeKPI": "Cycle time, letters drafted/day, NPS",
        "TypicalQuantifiedImpact": "Allstate: ~50,000 claim letters/day drafted by GenAI; pilots show 30-50% cycle-time reduction",
        "ImpactMagnitude": 4,
        "EvidenceStrength": 3,
        "DeploymentMaturity": 3,
        "ROI_Speed": 4,
        "RegRiskInverse": 3,
        "KeyEvidenceIds": "S20,S17",
    },
    {
        "Rank": None,
        "Domain": "Capital markets",
        "UseCase": "Equity / fixed-income research synthesis",
        "RepresentativeKPI": "Notes/analyst/week, time-to-first-draft",
        "TypicalQuantifiedImpact": "First-draft time reduced from days to hours in Tier-1 IB pilots; long-context (Claude) preferred for filings",
        "ImpactMagnitude": 3,
        "EvidenceStrength": 3,
        "DeploymentMaturity": 3,
        "ROI_Speed": 4,
        "RegRiskInverse": 3,
        "KeyEvidenceIds": "S5,S6,S14",
    },
    {
        "Rank": None,
        "Domain": "Commercial / corporate banking",
        "UseCase": "Credit memo drafting & covenant extraction",
        "RepresentativeKPI": "Memo cycle time, analyst hours saved",
        "TypicalQuantifiedImpact": "30-50% productivity uplift for credit/RM analysts in early production deployments",
        "ImpactMagnitude": 4,
        "EvidenceStrength": 2,
        "DeploymentMaturity": 3,
        "ROI_Speed": 3,
        "RegRiskInverse": 3,
        "KeyEvidenceIds": "S10,S11,S18",
    },
    {
        "Rank": None,
        "Domain": "Payments",
        "UseCase": "Disputes & chargeback automation",
        "RepresentativeKPI": "Cycle time, win-rate, cost per dispute",
        "TypicalQuantifiedImpact": "40-60% cycle-time reduction; 10-25 ppt win-rate improvement in pilots",
        "ImpactMagnitude": 3,
        "EvidenceStrength": 2,
        "DeploymentMaturity": 3,
        "ROI_Speed": 4,
        "RegRiskInverse": 3,
        "KeyEvidenceIds": "S7,S11,S18",
    },
    {
        "Rank": None,
        "Domain": "Capital markets",
        "UseCase": "Trading execution (autonomous order routing / decisioning)",
        "RepresentativeKPI": "Slippage, alpha, MAR",
        "TypicalQuantifiedImpact": "Limited public evidence; mostly read-only co-pilots; regulatory + best-ex constraints",
        "ImpactMagnitude": 3,
        "EvidenceStrength": 1,
        "DeploymentMaturity": 2,
        "ROI_Speed": 2,
        "RegRiskInverse": 1,
        "KeyEvidenceIds": "S11,S16",
    },
    {
        "Rank": None,
        "Domain": "Retail banking",
        "UseCase": "Credit underwriting decisioning (autonomous)",
        "RepresentativeKPI": "Approval rate, default rate, override %",
        "TypicalQuantifiedImpact": "Heavily restricted by ECOA/Reg B, EU AI Act high-risk; mostly used as decision support, not autonomous",
        "ImpactMagnitude": 4,
        "EvidenceStrength": 2,
        "DeploymentMaturity": 2,
        "ROI_Speed": 2,
        "RegRiskInverse": 1,
        "KeyEvidenceIds": "S10,S11,S18",
    },
]


WEIGHTS = {
    "ImpactMagnitude": 0.30,
    "EvidenceStrength": 0.25,
    "DeploymentMaturity": 0.20,
    "ROI_Speed": 0.15,
    "RegRiskInverse": 0.10,
}


def compute_scores() -> None:
    for row in EFFECTIVENESS:
        score = sum(row[k] * w for k, w in WEIGHTS.items())
        row["CompositeScore"] = round(score, 2)
    EFFECTIVENESS.sort(key=lambda r: r["CompositeScore"], reverse=True)
    for i, row in enumerate(EFFECTIVENESS, start=1):
        row["Rank"] = i


# ---------------------------------------------------------------------------
# Evidence log - one row per quantified data point with citation
# ---------------------------------------------------------------------------
EVIDENCE: list[dict] = [
    {
        "Domain": "Retail banking - customer service",
        "Institution": "Klarna",
        "Metric": "Share of customer service chats handled by AI assistant in first month",
        "Value": "~66% (2.3m chats)",
        "Equivalent": "Workload of ~700 full-time agents",
        "Date": "Feb 2024",
        "SourceId": "S1",
    },
    {
        "Domain": "Retail banking - customer service",
        "Institution": "Klarna",
        "Metric": "Customer matter resolution time",
        "Value": "From 11 minutes to <2 minutes",
        "Equivalent": ">80% reduction",
        "Date": "Feb 2024",
        "SourceId": "S1",
    },
    {
        "Domain": "Retail banking - customer service",
        "Institution": "Klarna",
        "Metric": "Estimated profit improvement attributable to AI assistant (annualized)",
        "Value": "~$40m",
        "Equivalent": "n/a",
        "Date": "2024",
        "SourceId": "S1",
    },
    {
        "Domain": "Retail banking - customer service",
        "Institution": "Bank of America (Erica)",
        "Metric": "Cumulative client interactions since 2018 launch",
        "Value": ">2 billion",
        "Equivalent": "~42m+ users",
        "Date": "Apr 2024",
        "SourceId": "S4",
    },
    {
        "Domain": "Retail banking - customer service",
        "Institution": "Wells Fargo (Fargo)",
        "Metric": "Annual virtual-assistant interactions",
        "Value": "~245m (2024)",
        "Equivalent": "~3x YoY growth vs 2023",
        "Date": "2024",
        "SourceId": "S9",
    },
    {
        "Domain": "Retail banking - customer service",
        "Institution": "ING",
        "Metric": "Customers served per unit time after GenAI chatbot rollout",
        "Value": "+20% increase in customers helped",
        "Equivalent": "Without adding headcount",
        "Date": "2024",
        "SourceId": "S15",
    },
    {
        "Domain": "Payments - fraud",
        "Institution": "Visa",
        "Metric": "Fraudulent activity prevented in FY2023 by AI/ML systems",
        "Value": "~$40bn",
        "Equivalent": "~2x prior-year prevention",
        "Date": "2024",
        "SourceId": "S8",
    },
    {
        "Domain": "Payments - fraud",
        "Institution": "Visa",
        "Metric": "Total invested in tech / AI / data infrastructure (cumulative)",
        "Value": "~$10bn over 5 years",
        "Equivalent": "n/a",
        "Date": "2024",
        "SourceId": "S8",
    },
    {
        "Domain": "Payments - fraud",
        "Institution": "Mastercard",
        "Metric": "Improvement in fraud detection rate from GenAI Decision Intelligence Pro",
        "Value": "Up to 2x detection in tested segments",
        "Equivalent": "Plus reduced false positives",
        "Date": "Feb 2024",
        "SourceId": "S7",
    },
    {
        "Domain": "Payments - fraud",
        "Institution": "Stripe (Radar)",
        "Metric": "Fraudulent transactions blocked (cumulative)",
        "Value": "Tens of $bn",
        "Equivalent": "Across millions of merchants",
        "Date": "2024",
        "SourceId": "S13",
    },
    {
        "Domain": "Document intelligence",
        "Institution": "JPMorgan Chase (COiN)",
        "Metric": "Lawyer hours saved annually on commercial-loan-agreement review",
        "Value": "~360,000 hours/yr",
        "Equivalent": "Pre-LLM ML system",
        "Date": "2017",
        "SourceId": "S2",
    },
    {
        "Domain": "Internal productivity",
        "Institution": "JPMorgan Chase (LLM Suite)",
        "Metric": "Employees with access to internal generative-AI assistant",
        "Value": "~200,000",
        "Equivalent": "Largest BFSI deployment disclosed",
        "Date": "2024-2025",
        "SourceId": "S3",
    },
    {
        "Domain": "Internal productivity",
        "Institution": "Goldman Sachs (GS AI Assistant)",
        "Metric": "Employees rolled out to (initial wave)",
        "Value": "~10,000+",
        "Equivalent": "Bankers, traders, asset managers",
        "Date": "2024-2025",
        "SourceId": "S6",
    },
    {
        "Domain": "Internal productivity",
        "Institution": "BBVA",
        "Metric": "ChatGPT Enterprise licenses deployed",
        "Value": "3,000",
        "Equivalent": "80% of users report time savings",
        "Date": "May 2024",
        "SourceId": "S12",
    },
    {
        "Domain": "Wealth & advisory",
        "Institution": "Morgan Stanley",
        "Metric": "Advisor-team adoption of OpenAI-powered AI @ Morgan Stanley Assistant",
        "Value": "~98% of advisor teams",
        "Equivalent": "Across ~16,000 advisors",
        "Date": "2024",
        "SourceId": "S5",
    },
    {
        "Domain": "Software engineering",
        "Institution": "Multiple BFSI (industry surveys)",
        "Metric": "Developer productivity uplift from AI coding assistants",
        "Value": "+20-40% (typical reported range)",
        "Equivalent": "Acceptance rates 25-45%",
        "Date": "2024",
        "SourceId": "S16",
    },
    {
        "Domain": "Insurance - claims",
        "Institution": "Allstate",
        "Metric": "Claim communications drafted by GenAI per day",
        "Value": "~50,000 letters/day",
        "Equivalent": "Auto + property claims",
        "Date": "2024",
        "SourceId": "S20",
    },
    {
        "Domain": "Risk & compliance - AML",
        "Institution": "Industry pilots (multiple Tier-1 banks)",
        "Metric": "Level-1 AML analyst productivity uplift in disposition / SAR drafting",
        "Value": "3-5x",
        "Equivalent": "30-60% false-positive reduction",
        "Date": "2024",
        "SourceId": "S18",
    },
    {
        "Domain": "Macro / sector sizing",
        "Institution": "McKinsey",
        "Metric": "Annual value pool from Gen AI in global banking",
        "Value": "~$200-340bn / yr",
        "Equivalent": "~9-15% of operating profits",
        "Date": "2023-2024",
        "SourceId": "S10",
    },
    {
        "Domain": "Macro / sector sizing",
        "Institution": "Citi GPS",
        "Metric": "Senior banking executives expecting productivity gains from GenAI",
        "Value": "~93%",
        "Equivalent": "Survey of global BFSI execs",
        "Date": "Jun 2024",
        "SourceId": "S11",
    },
    {
        "Domain": "Macro / sector sizing",
        "Institution": "NVIDIA / industry survey",
        "Metric": "BFSI firms using or evaluating GenAI",
        "Value": ">90%",
        "Equivalent": "2024 'State of AI in FS' survey",
        "Date": "2024",
        "SourceId": "S16",
    },
    {
        "Domain": "Payments - fraud",
        "Institution": "Commonwealth Bank of Australia",
        "Metric": "Reduction in customer-reported scam losses after AI scam-detection deployment",
        "Value": "~30% YoY (FY2024 disclosure)",
        "Equivalent": "Combined ML + GenAI messaging",
        "Date": "2024",
        "SourceId": "S19",
    },
]


# ---------------------------------------------------------------------------
# Vendor map - who is being used where
# ---------------------------------------------------------------------------
VENDOR_MAP: list[dict] = [
    {
        "UseCase": "Retail customer service virtual assistants",
        "FoundationModels": "OpenAI GPT-4 / 4o / 5; Anthropic Claude 3.5/3.7 Sonnet",
        "Platforms": "Azure OpenAI, AWS Bedrock, Google Vertex; in-house orchestration",
        "ExampleDeployments": "Klarna (OpenAI), BofA Erica (in-house + GenAI overlay), Wells Fargo Fargo (Google PaLM/Gemini), ING (Azure OpenAI)",
    },
    {
        "UseCase": "Fraud / transaction risk",
        "FoundationModels": "Proprietary ML + LLM meta-policy (OpenAI / Claude)",
        "Platforms": "Visa AI, Mastercard Decision Intelligence, Stripe Radar, FICO Falcon",
        "ExampleDeployments": "Visa, Mastercard, Stripe, CBA",
    },
    {
        "UseCase": "AML / KYC alert triage & SAR drafting",
        "FoundationModels": "Claude 3.5/3.7 (long context), GPT-4 class",
        "Platforms": "Azure OpenAI, AWS Bedrock, NICE Actimize, Quantexa, ComplyAdvantage",
        "ExampleDeployments": "Tier-1 US/EU banks (mostly undisclosed pilots)",
    },
    {
        "UseCase": "Wealth advisor knowledge assistant",
        "FoundationModels": "OpenAI GPT-4o (Morgan Stanley), Claude in select firms",
        "Platforms": "Azure OpenAI, OpenAI Enterprise, in-house RAG",
        "ExampleDeployments": "Morgan Stanley, UBS, Edward Jones (pilots)",
    },
    {
        "UseCase": "Document intelligence (loans, ISDA, KYC)",
        "FoundationModels": "Claude (long context), GPT-4 class",
        "Platforms": "AWS Bedrock, Azure OpenAI, Hebbia, Eigen, Kira/Litera",
        "ExampleDeployments": "JPM, Goldman Sachs, BNP Paribas (disclosed pilots)",
    },
    {
        "UseCase": "Enterprise employee copilot",
        "FoundationModels": "OpenAI ChatGPT Enterprise, Microsoft 365 Copilot, Claude Enterprise",
        "Platforms": "Azure OpenAI, OpenAI Enterprise, Anthropic Enterprise",
        "ExampleDeployments": "JPM LLM Suite, GS AI Assistant, BBVA ChatGPT Enterprise, Citi (Vertex+Gemini)",
    },
    {
        "UseCase": "Software engineering",
        "FoundationModels": "Claude 3.5/3.7 Sonnet, GPT-4o/5, Llama 3 derivatives",
        "Platforms": "GitHub Copilot, Cursor, Claude Code, internal model gateways",
        "ExampleDeployments": "Goldman Sachs, JPM, Capital One, Citi",
    },
    {
        "UseCase": "Insurance claims correspondence",
        "FoundationModels": "GPT-4 class, Claude",
        "Platforms": "Azure OpenAI, AWS Bedrock, Guidewire/Duck Creek integrations",
        "ExampleDeployments": "Allstate, Zurich, Lemonade",
    },
]


# ---------------------------------------------------------------------------
# Workbook construction
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E78")
WRAP = Alignment(wrap_text=True, vertical="top")
BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def style_header(ws, row_idx: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        cell.border = BORDER


def auto_width(ws, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            for line in str(v).split("\n"):
                max_len = max(max_len, len(line))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), max_width)


def write_readme(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "README"
    rows = [
        ("AI Agent Effectiveness in Banking, Payments & Financial Services", None),
        ("Aggregated public evidence", None),
        ("", None),
        ("Companion to: research/ai-agents-banking-payments-financial-services.md", None),
        ("Generated by: scripts/build_bfsi_ai_agents_workbook.py", None),
        ("", None),
        ("Sheets in this workbook:", None),
        ("  1. README              - this page (methodology, scoring rubric, disclaimer)", None),
        ("  2. Effectiveness_Rank  - ranked use cases with composite effectiveness score", None),
        ("  3. Evidence_Log        - individual quantified public data points + citations", None),
        ("  4. Vendor_Map          - foundation-model and platform vendors observed per use case", None),
        ("  5. Sources             - bibliography of cited public sources", None),
        ("", None),
        ("Scoring rubric (1=low, 5=high):", None),
        ("  ImpactMagnitude     - Size of cost / revenue / risk impact when deployed", None),
        ("  EvidenceStrength    - Quality and number of public quantified disclosures", None),
        ("  DeploymentMaturity  - How many BFSI institutions are in production today", None),
        ("  ROI_Speed           - Typical time to measurable financial benefit", None),
        ("  RegRiskInverse      - Inverse of regulatory friction (5 = low friction)", None),
        ("", None),
        ("Composite weights:", None),
        ("  ImpactMagnitude   = 0.30", None),
        ("  EvidenceStrength  = 0.25", None),
        ("  DeploymentMaturity= 0.20", None),
        ("  ROI_Speed         = 0.15", None),
        ("  RegRiskInverse    = 0.10", None),
        ("", None),
        ("Headline finding:", None),
        (
            "  AI agents are most effective today in (a) retail banking customer service / virtual "
            "assistants, (b) payments fraud detection and transaction risk scoring, and "
            "(c) enterprise employee copilots and software engineering inside banks. These three "
            "areas combine large impact, strong public evidence, broad deployment, fast ROI, and "
            "comparatively low regulatory friction. Autonomous credit underwriting and autonomous "
            "trading remain the lowest-effectiveness areas today, primarily due to regulatory and "
            "model-risk constraints.",
            None,
        ),
        ("", None),
        ("Disclaimer:", None),
        (
            "  This workbook is an aggregation of publicly disclosed figures and reputable third-party "
            "research. Numbers labelled as 'pilots', 'reported', or as ranges should be treated as "
            "directional. Nothing in this workbook is investment advice. Company and product names "
            "are trademarks of their respective owners.",
            None,
        ),
    ]
    ws["A1"] = rows[0][0]
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    for i, (text, _) in enumerate(rows[1:], start=2):
        ws.cell(row=i, column=1, value=text).alignment = WRAP
    ws.column_dimensions["A"].width = 110


def write_effectiveness(wb: Workbook) -> None:
    ws = wb.create_sheet("Effectiveness_Rank")
    headers = [
        "Rank",
        "Domain",
        "UseCase",
        "RepresentativeKPI",
        "TypicalQuantifiedImpact",
        "ImpactMagnitude",
        "EvidenceStrength",
        "DeploymentMaturity",
        "ROI_Speed",
        "RegRiskInverse",
        "CompositeScore",
        "KeyEvidenceIds",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for row in EFFECTIVENESS:
        ws.append([row[h] for h in headers])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
    ws.freeze_panes = "A2"
    auto_width(ws)
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 60
    rng = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    table = Table(displayName="EffectivenessRank", ref=rng)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def write_evidence(wb: Workbook) -> None:
    ws = wb.create_sheet("Evidence_Log")
    headers = ["Domain", "Institution", "Metric", "Value", "Equivalent / Note", "Date", "SourceId", "SourceShort", "URL"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for e in EVIDENCE:
        src = SOURCE_BY_ID[e["SourceId"]]
        ws.append(
            [
                e["Domain"],
                e["Institution"],
                e["Metric"],
                e["Value"],
                e["Equivalent"],
                e["Date"],
                e["SourceId"],
                src["short"],
                src["url"],
            ]
        )
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
    ws.freeze_panes = "A2"
    auto_width(ws)
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["H"].width = 60
    ws.column_dimensions["I"].width = 60
    rng = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    table = Table(displayName="EvidenceLog", ref=rng)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def write_vendors(wb: Workbook) -> None:
    ws = wb.create_sheet("Vendor_Map")
    headers = ["UseCase", "FoundationModels", "Platforms", "ExampleDeployments"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for v in VENDOR_MAP:
        ws.append([v[h] for h in headers])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
    ws.freeze_panes = "A2"
    auto_width(ws)
    for col, w in zip("ABCD", (45, 50, 55, 70)):
        ws.column_dimensions[col].width = w
    rng = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    table = Table(displayName="VendorMap", ref=rng)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def write_sources(wb: Workbook) -> None:
    ws = wb.create_sheet("Sources")
    headers = ["Id", "Publisher", "Date", "Description", "URL"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for s in SOURCES:
        ws.append([s["id"], s["publisher"], s["date"], s["short"], s["url"]])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
    ws.freeze_panes = "A2"
    auto_width(ws)
    ws.column_dimensions["D"].width = 80
    ws.column_dimensions["E"].width = 70
    rng = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    table = Table(displayName="Sources", ref=rng)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium6",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def main() -> None:
    compute_scores()
    wb = Workbook()
    write_readme(wb)
    write_effectiveness(wb)
    write_evidence(wb)
    write_vendors(wb)
    write_sources(wb)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print()
    print("Top 5 use cases by composite effectiveness score:")
    for row in EFFECTIVENESS[:5]:
        print(f"  #{row['Rank']:>2}  {row['CompositeScore']:.2f}  {row['Domain']:<35} {row['UseCase']}")


if __name__ == "__main__":
    main()
