#!/usr/bin/env python3
"""
Generate an Excel workbook with monthly total token counts (training vs inference)
for OpenAI, Anthropic, Google (Gemini), Meta, xAI, and Cursor.

Training tokens are allocated to the month(s) in which each model was trained.
Inference tokens are estimated monthly volumes based on publicly available data
points and interpolated where necessary.

Sources and methodology are documented in the "Methodology" sheet.
"""

import os
import math
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "..", "data", "token_volume_monthly.xlsx")

# ---------------------------------------------------------------------------
# Color palette
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
COMPANY_FILLS = {
    "OpenAI": PatternFill("solid", fgColor="D5F5E3"),
    "Anthropic": PatternFill("solid", fgColor="FDEBD0"),
    "Google": PatternFill("solid", fgColor="D6EAF8"),
    "Meta": PatternFill("solid", fgColor="D4E6F1"),
    "xAI": PatternFill("solid", fgColor="E8DAEF"),
    "Cursor": PatternFill("solid", fgColor="FADBD8"),
}
TOTAL_FILL = PatternFill("solid", fgColor="F2F3F4")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
NUM_FMT = '#,##0'
NUM_FMT_T = '#,##0.0,,"T"'


def _months(start_year, start_month, end_year, end_month):
    """Yield (year, month) tuples from start to end inclusive."""
    y, m = start_year, start_month
    while (y, m) <= (end_year, end_month):
        yield y, m
        m += 1
        if m > 12:
            m = 1
            y += 1


# ---------------------------------------------------------------------------
# Training token data: (company, model, training_start, training_end, total_tokens)
#
# training_start/end are (year, month). Tokens are spread evenly across months.
# Where exact training windows are unknown, we estimate based on release date
# and known GPU-hours / typical training durations.
# ---------------------------------------------------------------------------
TRAINING_RUNS = [
    # OpenAI
    ("OpenAI", "GPT-1",           (2018, 1), (2018, 5),    40_000_000_000),       # ~40B tokens (small model, est.)
    ("OpenAI", "GPT-2",           (2018, 8), (2019, 1),    10_000_000_000),       # ~10B tokens (40GB WebText)
    ("OpenAI", "GPT-3",           (2020, 1), (2020, 5),    300_000_000_000),      # 300B tokens (confirmed)
    ("OpenAI", "GPT-3.5/ChatGPT", (2022, 6), (2022, 10),   300_000_000_000),      # Fine-tuned on GPT-3 data + RLHF
    ("OpenAI", "GPT-4",           (2022, 8), (2023, 2),    13_000_000_000_000),   # ~13T tokens (estimated)
    ("OpenAI", "GPT-4o",          (2023, 10), (2024, 4),   13_000_000_000_000),   # ~13T tokens (similar to GPT-4)
    ("OpenAI", "GPT-4.1",         (2024, 8), (2025, 3),    15_000_000_000_000),   # ~15T tokens (est.)
    ("OpenAI", "GPT-5/5.4",       (2025, 6), (2026, 2),    20_000_000_000_000),   # ~20T tokens (est.)

    # Anthropic
    ("Anthropic", "Claude 1",        (2022, 9), (2023, 2),   500_000_000_000),      # ~500B tokens (est.)
    ("Anthropic", "Claude 2",        (2023, 2), (2023, 6),   2_000_000_000_000),    # ~2T tokens (est.)
    ("Anthropic", "Claude 3 family", (2023, 8), (2024, 2),   8_000_000_000_000),    # ~8T tokens (est.)
    ("Anthropic", "Claude 3.5",      (2024, 1), (2024, 5),   10_000_000_000_000),   # ~10T tokens (est.)
    ("Anthropic", "Claude 4/4.5",    (2024, 10), (2025, 4),  15_000_000_000_000),   # ~15T tokens (est.)
    ("Anthropic", "Claude 4.6",      (2025, 6), (2025, 12),  20_000_000_000_000),   # ~20T tokens (est.)

    # Google (Gemini)
    ("Google", "Gemini 1.0",       (2023, 4), (2023, 11),  8_000_000_000_000),    # ~8T tokens (est.)
    ("Google", "Gemini 1.5 Pro",   (2023, 10), (2024, 1),  10_000_000_000_000),   # ~10T tokens (est.)
    ("Google", "Gemini 2.0",       (2024, 4), (2024, 11),  15_000_000_000_000),   # ~15T tokens (est.)
    ("Google", "Gemini 2.5 Pro",   (2024, 10), (2025, 2),  15_000_000_000_000),   # ~15T tokens (est.)
    ("Google", "Gemini 3",         (2025, 3), (2025, 11),  20_000_000_000_000),   # ~20T tokens (est.)

    # Meta
    ("Meta", "LLaMA 1",      (2022, 8), (2023, 1),   1_400_000_000_000),   # 1.4T tokens (confirmed)
    ("Meta", "LLaMA 2",      (2023, 1), (2023, 6),   2_000_000_000_000),   # 2T tokens (confirmed)
    ("Meta", "LLaMA 3",      (2023, 10), (2024, 3),  15_000_000_000_000),  # 15T tokens (confirmed)
    ("Meta", "LLaMA 3.1",    (2024, 1), (2024, 6),   15_000_000_000_000),  # 15T tokens (confirmed)
    ("Meta", "LLaMA 4",      (2024, 8), (2025, 3),   30_000_000_000_000),  # 30T+ tokens (confirmed)

    # xAI
    ("xAI", "Grok 1",   (2023, 5), (2023, 10),   500_000_000_000),       # ~500B tokens (est.)
    ("xAI", "Grok 1.5", (2023, 10), (2024, 2),   2_000_000_000_000),     # ~2T tokens (est.)
    ("xAI", "Grok 2",   (2024, 2), (2024, 7),    5_000_000_000_000),     # ~5T tokens (est.)
    ("xAI", "Grok 3",   (2024, 6), (2025, 1),    10_000_000_000_000),    # ~10T tokens (est.)
    ("xAI", "Grok 4",   (2025, 3), (2025, 10),   15_000_000_000_000),    # ~15T tokens (est.)
    ("xAI", "Grok 5",   (2025, 10), (2026, 3),   20_000_000_000_000),    # ~20T tokens (est., in progress)

    # Cursor
    ("Cursor", "cursor-small",  (2024, 2), (2024, 5),   200_000_000_000),     # ~200B tokens (est.)
    ("Cursor", "Composer 1.5",  (2025, 2), (2025, 8),   500_000_000_000),     # ~500B tokens (est.)
    ("Cursor", "Composer 2",    (2025, 9), (2026, 2),   1_000_000_000_000),   # ~1T tokens (est.)
]

# ---------------------------------------------------------------------------
# Inference token data: (company, month_year, tokens_per_month)
# Based on public disclosures and interpolated between known data points.
# ---------------------------------------------------------------------------

def _build_inference_series():
    """Build monthly inference token estimates for each company."""
    data = {}  # (company, year, month) -> tokens

    def _interp(company, points):
        """
        points: list of ((year, month), tokens_per_month) sorted by date.
        Interpolate exponentially between known data points.
        """
        for i in range(len(points) - 1):
            (y1, m1), v1 = points[i]
            (y2, m2), v2 = points[i + 1]
            total_months = (y2 - y1) * 12 + (m2 - m1)
            if total_months <= 0:
                data[(company, y1, m1)] = v1
                continue
            for j, (y, m) in enumerate(_months(y1, m1, y2, m2)):
                if total_months == 0:
                    frac = 0
                else:
                    frac = j / total_months
                if v1 > 0 and v2 > 0:
                    val = v1 * math.exp(frac * math.log(v2 / v1))
                else:
                    val = v1 + frac * (v2 - v1)
                data[(company, y, m)] = int(val)
        # last point
        (yl, ml), vl = points[-1]
        data[(company, yl, ml)] = vl

    # OpenAI inference (tokens/month)
    # Data points: Q4 2022 ChatGPT launch ~trivial, grew to ~260T/month by mid-2025
    _interp("OpenAI", [
        ((2022, 12), 500_000_000_000),            # ChatGPT launch, ~500B/mo
        ((2023, 3),  2_000_000_000_000),           # Rapid growth, ~2T/mo
        ((2023, 6),  5_000_000_000_000),           # ~5T/mo
        ((2023, 9),  8_000_000_000_000),           # ~8T/mo
        ((2023, 11), 15_000_000_000_000),          # GPT-4 Turbo launch, ~15T/mo
        ((2024, 1),  20_000_000_000_000),          # ~20T/mo
        ((2024, 3),  30_000_000_000_000),          # ~30T/mo (Q1 2024 ~1.5T total tokens = heavy API)
        ((2024, 6),  50_000_000_000_000),          # GPT-4o driving adoption, ~50T/mo
        ((2024, 9),  80_000_000_000_000),          # ~80T/mo
        ((2024, 12), 120_000_000_000_000),         # ~120T/mo
        ((2025, 3),  160_000_000_000_000),         # ~160T/mo
        ((2025, 6),  200_000_000_000_000),         # ~200T/mo
        ((2025, 9),  260_000_000_000_000),         # ~260T/mo (confirmed estimate)
        ((2025, 12), 300_000_000_000_000),         # ~300T/mo
        ((2026, 3),  350_000_000_000_000),         # ~350T/mo (GPT-5.4 driving growth)
    ])

    # Google (Gemini) inference
    # Data points: Apr 2024 ~9.7T/mo, May 2025 480T, Jun 2025 950T, Oct 2025 1.3Q
    _interp("Google", [
        ((2023, 12), 500_000_000_000),             # Gemini 1.0 launch, ~500B/mo
        ((2024, 3),  3_000_000_000_000),            # ~3T/mo
        ((2024, 4),  9_700_000_000_000),            # 9.7T/mo (confirmed)
        ((2024, 6),  15_000_000_000_000),           # ~15T/mo
        ((2024, 9),  30_000_000_000_000),           # ~30T/mo
        ((2024, 12), 80_000_000_000_000),           # ~80T/mo
        ((2025, 3),  200_000_000_000_000),          # ~200T/mo
        ((2025, 5),  480_000_000_000_000),          # 480T/mo (confirmed)
        ((2025, 6),  950_000_000_000_000),          # 950T/mo (confirmed)
        ((2025, 7),  980_000_000_000_000),          # 980T/mo (confirmed)
        ((2025, 10), 1_300_000_000_000_000),        # 1.3Q/mo (confirmed)
        ((2025, 12), 1_400_000_000_000_000),        # ~1.4Q/mo
        ((2026, 3),  1_600_000_000_000_000),        # ~1.6Q/mo (est.)
    ])

    # Anthropic inference
    # Revenue-based estimation: Q1 2024 $36M → Q3 2025 $1B
    # Avg price ~$5/Mtok → tokens ~ revenue / $5 * 1M
    _interp("Anthropic", [
        ((2023, 3),  100_000_000_000),              # Claude 1 launch, ~100B/mo
        ((2023, 7),  500_000_000_000),              # Claude 2 launch, ~500B/mo
        ((2023, 12), 1_500_000_000_000),            # ~1.5T/mo
        ((2024, 3),  3_000_000_000_000),            # Q1 rev $36M → ~3T/mo est.
        ((2024, 6),  8_000_000_000_000),            # Q2 rev $106M → ~8T/mo
        ((2024, 9),  15_000_000_000_000),           # Q3 rev $184M → ~15T/mo
        ((2024, 12), 20_000_000_000_000),           # Q4 rev $228M → ~20T/mo
        ((2025, 3),  35_000_000_000_000),           # Q1 rev $375M → ~35T/mo
        ((2025, 6),  60_000_000_000_000),           # Q2 rev $625M → ~60T/mo
        ((2025, 9),  100_000_000_000_000),          # Q3 rev $1B → ~100T/mo
        ((2025, 12), 130_000_000_000_000),          # ~130T/mo
        ((2026, 3),  170_000_000_000_000),          # ~170T/mo (Claude 4.6 + $14B ARR)
    ])

    # Meta (LLaMA ecosystem — self-hosted + cloud partners)
    # 600M MAU by end 2024 + massive cloud partner token growth
    _interp("Meta", [
        ((2023, 7),  100_000_000_000),              # LLaMA 2 launch, ~100B/mo
        ((2023, 12), 500_000_000_000),              # ~500B/mo
        ((2024, 1),  1_000_000_000_000),            # ~1T/mo
        ((2024, 4),  5_000_000_000_000),            # LLaMA 3 launch, ~5T/mo
        ((2024, 7),  10_000_000_000_000),           # 10x growth Jan-Jul → ~10T/mo
        ((2024, 9),  15_000_000_000_000),           # 50% MoM growth in Sep
        ((2024, 12), 30_000_000_000_000),           # 600M MAU, ~30T/mo
        ((2025, 3),  50_000_000_000_000),           # ~50T/mo
        ((2025, 6),  80_000_000_000_000),           # ~80T/mo
        ((2025, 9),  120_000_000_000_000),          # ~120T/mo
        ((2025, 12), 150_000_000_000_000),          # ~150T/mo
        ((2026, 3),  200_000_000_000_000),          # ~200T/mo
    ])

    # xAI (Grok)
    # 64M MAU by Sep 2025, 1B monthly queries earlier, 17.8% US share Jan 2026
    _interp("xAI", [
        ((2023, 11), 20_000_000_000),               # Grok 1 launch, ~20B/mo
        ((2024, 3),  100_000_000_000),              # Grok 1.5, ~100B/mo
        ((2024, 6),  300_000_000_000),              # ~300B/mo
        ((2024, 8),  500_000_000_000),              # Grok 2 launch
        ((2024, 12), 2_000_000_000_000),            # ~2T/mo
        ((2025, 2),  5_000_000_000_000),            # Grok 3 launch, ~5T/mo
        ((2025, 6),  15_000_000_000_000),           # ~15T/mo
        ((2025, 9),  30_000_000_000_000),           # 64M MAU, ~30T/mo
        ((2025, 12), 50_000_000_000_000),           # ~50T/mo (17.8% US share)
        ((2026, 3),  70_000_000_000_000),           # ~70T/mo
    ])

    # Cursor
    # 1M+ DAU, $2B ARR, avg user ~250M tokens/year
    # ~5M MAU → ~100B tokens/mo in 2025, growing rapidly
    _interp("Cursor", [
        ((2024, 3),  5_000_000_000),                # Early days, ~5B/mo
        ((2024, 6),  10_000_000_000),               # cursor-small launch, ~10B/mo
        ((2024, 9),  30_000_000_000),               # ~30B/mo
        ((2024, 12), 80_000_000_000),               # ~80B/mo
        ((2025, 1),  100_000_000_000),              # $100M ARR milestone
        ((2025, 4),  200_000_000_000),              # $300M ARR
        ((2025, 6),  350_000_000_000),              # $500M ARR
        ((2025, 9),  600_000_000_000),              # ~600B/mo
        ((2025, 11), 1_000_000_000_000),            # $1B ARR
        ((2026, 2),  2_000_000_000_000),            # $2B ARR, ~2T/mo
        ((2026, 3),  2_500_000_000_000),            # ~2.5T/mo
    ])

    return data


def _build_training_series():
    """Spread training tokens evenly across the training months."""
    data = {}
    for company, model, (y1, m1), (y2, m2), total_tokens in TRAINING_RUNS:
        months_list = list(_months(y1, m1, y2, m2))
        per_month = total_tokens / len(months_list)
        for y, m in months_list:
            key = (company, y, m)
            data[key] = data.get(key, 0) + int(per_month)
    return data


def main():
    companies = ["OpenAI", "Anthropic", "Google", "Meta", "xAI", "Cursor"]
    all_months = list(_months(2018, 1, 2026, 3))

    training = _build_training_series()
    inference = _build_inference_series()

    wb = openpyxl.Workbook()

    # ====================================================================
    # Sheet 1: Monthly data – all companies, training vs inference
    # ====================================================================
    ws = wb.active
    ws.title = "Monthly Token Volume"
    ws.sheet_properties.tabColor = "1F4E79"

    # Build headers
    headers = ["Month"]
    for c in companies:
        headers.extend([f"{c} Training", f"{c} Inference", f"{c} Total"])
    headers.extend(["All Training", "All Inference", "Grand Total"])

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Freeze top row
    ws.freeze_panes = "B2"

    # Fill data
    for row_idx, (y, m) in enumerate(all_months, 2):
        month_str = f"{y}-{m:02d}"
        ws.cell(row=row_idx, column=1, value=month_str).border = THIN_BORDER
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")

        all_train = 0
        all_infer = 0
        col = 2
        for c in companies:
            t = training.get((c, y, m), 0)
            i = inference.get((c, y, m), 0)
            total = t + i
            all_train += t
            all_infer += i

            for val in [t, i, total]:
                cell = ws.cell(row=row_idx, column=col, value=val if val > 0 else 0)
                cell.number_format = NUM_FMT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="right")
                col += 1

        for val in [all_train, all_infer, all_train + all_infer]:
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.number_format = NUM_FMT
            cell.border = THIN_BORDER
            cell.fill = TOTAL_FILL
            cell.alignment = Alignment(horizontal="right")
            cell.font = Font(bold=True)
            col += 1

    # Column widths
    ws.column_dimensions["A"].width = 10
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    # ====================================================================
    # Sheet 2: Per-company summary sheets
    # ====================================================================
    for c in companies:
        ws2 = wb.create_sheet(title=c)
        ws2.sheet_properties.tabColor = {
            "OpenAI": "10A37F", "Anthropic": "D97706", "Google": "4285F4",
            "Meta": "0668E1", "xAI": "9333EA", "Cursor": "F14E32",
        }[c]

        for col_idx, h in enumerate(["Month", "Training Tokens", "Inference Tokens", "Total Tokens"], 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        ws2.freeze_panes = "B2"

        data_start_row = None
        row = 2
        for y, m in all_months:
            t = training.get((c, y, m), 0)
            i = inference.get((c, y, m), 0)
            if t == 0 and i == 0:
                continue
            if data_start_row is None:
                data_start_row = row
            ws2.cell(row=row, column=1, value=f"{y}-{m:02d}").border = THIN_BORDER
            for col_idx, val in enumerate([t, i, t + i], 2):
                cell = ws2.cell(row=row, column=col_idx, value=val)
                cell.number_format = NUM_FMT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="right")
            row += 1

        ws2.column_dimensions["A"].width = 10
        ws2.column_dimensions["B"].width = 22
        ws2.column_dimensions["C"].width = 22
        ws2.column_dimensions["D"].width = 22

        # Add a chart per company
        if data_start_row and row > data_start_row + 1:
            chart = LineChart()
            chart.title = f"{c} — Monthly Token Volume"
            chart.y_axis.title = "Tokens"
            chart.x_axis.title = "Month"
            chart.style = 10
            chart.width = 30
            chart.height = 15

            cats = Reference(ws2, min_col=1, min_row=data_start_row, max_row=row - 1)
            for col_idx, label in [(2, "Training"), (3, "Inference")]:
                vals = Reference(ws2, min_col=col_idx, min_row=data_start_row - 1, max_row=row - 1)
                chart.add_data(vals, titles_from_data=True)
            chart.set_categories(cats)
            ws2.add_chart(chart, f"F2")

    # ====================================================================
    # Sheet 3: Training runs reference
    # ====================================================================
    ws3 = wb.create_sheet(title="Training Runs")
    ws3.sheet_properties.tabColor = "5B2C6F"
    train_headers = ["Company", "Model", "Training Start", "Training End", "Total Training Tokens", "Source / Notes"]
    for col_idx, h in enumerate(train_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    ws3.freeze_panes = "B2"

    notes_map = {
        "GPT-1": "Estimated; small model",
        "GPT-2": "~40GB WebText (~10B tokens)",
        "GPT-3": "Confirmed: 300B tokens (Brown et al. 2020)",
        "GPT-3.5/ChatGPT": "Fine-tuned GPT-3 + RLHF; est. similar data volume",
        "GPT-4": "Estimated ~13T tokens (multiple sources)",
        "GPT-4o": "Estimated similar to GPT-4",
        "GPT-4.1": "Estimated ~15T tokens",
        "GPT-5/5.4": "Estimated ~20T tokens",
        "Claude 1": "Estimated ~500B tokens",
        "Claude 2": "Estimated ~2T tokens",
        "Claude 3 family": "Estimated ~8T tokens",
        "Claude 3.5": "Estimated ~10T tokens",
        "Claude 4/4.5": "Estimated ~15T tokens",
        "Claude 4.6": "Estimated ~20T tokens",
        "Gemini 1.0": "Estimated ~8T tokens",
        "Gemini 1.5 Pro": "Estimated ~10T tokens",
        "Gemini 2.0": "Estimated ~15T tokens",
        "Gemini 2.5 Pro": "Estimated ~15T tokens",
        "Gemini 3": "Estimated ~20T tokens",
        "LLaMA 1": "Confirmed: 1.4T tokens (Meta)",
        "LLaMA 2": "Confirmed: 2T tokens (Meta)",
        "LLaMA 3": "Confirmed: 15T tokens (Meta)",
        "LLaMA 3.1": "Confirmed: 15T tokens (Meta)",
        "LLaMA 4": "Confirmed: 30T+ tokens (Meta)",
        "Grok 1": "Estimated ~500B tokens",
        "Grok 1.5": "Estimated ~2T tokens",
        "Grok 2": "Estimated ~5T tokens",
        "Grok 3": "Estimated ~10T tokens; 200M GPU-hrs on 100K H100s",
        "Grok 4": "Estimated ~15T tokens",
        "Grok 5": "Estimated ~20T tokens; training in progress on Colossus 2",
        "cursor-small": "Estimated ~200B tokens; code-specialized",
        "Composer 1.5": "Estimated ~500B tokens",
        "Composer 2": "Estimated ~1T tokens",
    }

    for row_idx, (company, model, (y1, m1), (y2, m2), total) in enumerate(TRAINING_RUNS, 2):
        ws3.cell(row=row_idx, column=1, value=company).border = THIN_BORDER
        ws3.cell(row=row_idx, column=2, value=model).border = THIN_BORDER
        ws3.cell(row=row_idx, column=3, value=f"{y1}-{m1:02d}").border = THIN_BORDER
        ws3.cell(row=row_idx, column=4, value=f"{y2}-{m2:02d}").border = THIN_BORDER
        cell = ws3.cell(row=row_idx, column=5, value=total)
        cell.number_format = NUM_FMT
        cell.border = THIN_BORDER
        ws3.cell(row=row_idx, column=6, value=notes_map.get(model, "Estimated")).border = THIN_BORDER

    for col_idx, w in enumerate([12, 20, 14, 14, 24, 50], 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = w

    # ====================================================================
    # Sheet 4: Methodology
    # ====================================================================
    ws4 = wb.create_sheet(title="Methodology")
    ws4.sheet_properties.tabColor = "7D3C98"

    methodology_text = [
        ("Methodology & Sources", True),
        ("", False),
        ("This workbook contains monthly estimates of total token counts for six AI companies,", False),
        ("split between training tokens (tokens consumed during model training) and inference", False),
        ("tokens (tokens processed in production serving user requests).", False),
        ("", False),
        ("TRAINING TOKENS", True),
        ("Training token counts represent the total dataset size consumed during pre-training.", False),
        ("Where official figures are available (GPT-3: 300B, LLaMA 1: 1.4T, LLaMA 2: 2T,", False),
        ("LLaMA 3/3.1: 15T, LLaMA 4: 30T+, GPT-4: ~13T est.), those are used.", False),
        ("For models without disclosed training data sizes (Claude, Gemini, Grok, Cursor),", False),
        ("estimates are based on parameter counts, training compute (GPU-hours), and", False),
        ("industry scaling trends from Epoch AI and similar sources.", False),
        ("Training tokens are allocated evenly across estimated training months.", False),
        ("", False),
        ("INFERENCE TOKENS", True),
        ("Monthly inference volumes are estimated using multiple data points:", False),
        ("- Google: 9.7T/mo (Apr 2024), 480T (May 2025), 950T (Jun 2025),", False),
        ("  980T (Jul 2025), 1.3Q (Oct 2025) — from Google earnings/blog posts", False),
        ("- OpenAI: ~260T/mo (mid-2025) — industry estimates (Adam Holter)", False),
        ("- Anthropic: Revenue-based estimation ($36M Q1 2024 → $1B Q3 2025)", False),
        ("  at ~$5/Mtok average implies token volumes", False),
        ("- Meta: Token volume doubled May-Jul 2024, 10x Jan-Jul 2024 (Meta blog)", False),
        ("  600M MAU Meta AI by end 2024", False),
        ("- xAI: 64M MAU Sep 2025, 17.8% US share Jan 2026, 1B monthly queries", False),
        ("- Cursor: 1M+ DAU, $2B ARR Feb 2026, avg ~250M tokens/user/year", False),
        ("Between known data points, exponential interpolation is used.", False),
        ("", False),
        ("IMPORTANT CAVEATS", True),
        ("- Training tokens are TOTAL tokens consumed during training, not monthly new tokens.", False),
        ("  They are spread across training months for time-series presentation.", False),
        ("- Inference estimates for Anthropic, Meta, xAI, and Cursor involve significant", False),
        ("  uncertainty as these companies do not publicly disclose token volumes.", False),
        ("- Google's 1.3Q tokens/mo includes all AI-powered products (Search, Gmail,", False),
        ("  YouTube, Workspace), not just Gemini API.", False),
        ("- Cursor's inference tokens include all underlying model providers' tokens", False),
        ("  consumed on behalf of Cursor users (Claude, GPT, Gemini, etc.).", False),
        ("", False),
        ("DATA SOURCES", True),
        ("- Epoch AI (epoch.ai) — training compute and data trends", False),
        ("- Stealth Cloud AI Training Data Volume Tracker", False),
        ("- Google earnings calls and developer blog posts (Q4 2025)", False),
        ("- Adam Holter analysis of industry token volumes", False),
        ("- Anthropic quarterly revenue reports (Business of Apps)", False),
        ("- Meta AI blog posts on LLaMA adoption", False),
        ("- xAI announcements and Grokipedia", False),
        ("- Cursor revenue/usage reports (TechCrunch, GetPanto)", False),
        ("- OpenAI API documentation and OpenAI blog", False),
        ("", False),
        (f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M UTC')}", False),
    ]

    for row_idx, (text, is_bold) in enumerate(methodology_text, 1):
        cell = ws4.cell(row=row_idx, column=1, value=text)
        if is_bold:
            cell.font = Font(bold=True, size=12)
        else:
            cell.font = Font(size=11)

    ws4.column_dimensions["A"].width = 90

    # ====================================================================
    # Save
    # ====================================================================
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
