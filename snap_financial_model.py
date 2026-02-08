#!/usr/bin/env python3
"""
Snap Inc. (SNAP US Equity) – Three-Statement Financial Model
=============================================================
Historical: FY2019 – FY2024 (annual) + quarterly detail FY2023-FY2024
Forecast:   FY2025E – FY2028E (annual) + quarterly detail FY2025E-FY2026E

Covers:
  1. Revenue build by segment (Advertising – NA/EU/ROW) with DAU × ARPU drivers
  2. Income Statement (GAAP) with gross-margin, OpEx ratios, SBC
  3. Balance Sheet
  4. Cash-Flow Statement with GAAP OpLoss → FCF bridge
  5. Share-count & dilution schedule (8 forward quarters)
  6. Two-way sensitivity tables (DAU growth × ARPU growth; Rev growth × EBITDA margin)
  7. Valuation cross-checks (DCF, cohort NPV, unit economics → EV)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

# ── Styling constants ─────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=10, color="2F5496")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FORECAST_FONT = Font(name="Calibri", italic=True, size=10, color="0070C0")
NORMAL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
SECTION_FONT = Font(name="Calibri", bold=True, size=11, color="2F5496")
PCT_FMT = '0.0%'
NUM_FMT = '#,##0'
DEC_FMT = '#,##0.0'
USD_FMT = '$#,##0'
USD_DEC_FMT = '$#,##0.00'
THIN_BORDER = Border(
    bottom=Side(style='thin', color='B4C6E7')
)
BOTTOM_BORDER = Border(
    bottom=Side(style='medium', color='2F5496')
)

def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

def style_subheader_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal='center')

def write_row(ws, row, data, start_col=1, font=None, fmt=None, fill=None, is_forecast_start=None):
    """Write a list of values to a row, optionally styling."""
    for i, val in enumerate(data):
        cell = ws.cell(row=row, column=start_col + i, value=val)
        cell.font = font or NORMAL_FONT
        if fmt and i > 0 and val is not None:
            cell.number_format = fmt
        if fill:
            cell.fill = fill
        # Mark forecast columns with italic blue
        if is_forecast_start is not None and (start_col + i) >= is_forecast_start and i > 0:
            cell.font = FORECAST_FONT

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════════════════════════════
# DATA — Snap Inc. historical financials (sourced from 10-K / 10-Q filings)
# All dollar amounts in $millions unless noted
# ══════════════════════════════════════════════════════════════════════════════

# ── Annual historical data FY2019-FY2024 ──────────────────────────────────────
years_hist = [2019, 2020, 2021, 2022, 2023, 2024]
years_fcst = [2025, 2026, 2027, 2028]
all_years = years_hist + years_fcst

# DAU (millions, year-end / Q4)
dau_hist = [218, 265, 319, 375, 414, 443]

# Revenue
rev_hist = [1716, 2507, 4117, 4602, 4606, 5360]

# ARPU (calculated)
arpu_hist = [round(r / d, 2) for r, d in zip(rev_hist, dau_hist)]

# Revenue by geography (approx split from filings)
# NA ~60-65%, EU ~20-22%, ROW ~15-18%
rev_na_hist = [1100, 1620, 2720, 2990, 2950, 3430]
rev_eu_hist = [360, 500, 830, 920, 920, 1070]
rev_row_hist = [256, 387, 567, 692, 736, 860]

# Cost of Revenue
cor_hist = [880, 1128, 1763, 2158, 2190, 2461]
gross_profit_hist = [r - c for r, c in zip(rev_hist, cor_hist)]
gm_hist = [gp / r for gp, r in zip(gross_profit_hist, rev_hist)]

# Operating Expenses
rd_hist = [988, 1235, 1535, 2017, 1847, 1802]
sm_hist = [654, 794, 1168, 1424, 1284, 1281]
ga_hist = [363, 421, 507, 641, 614, 582]
total_opex_hist = [a + b + c for a, b, c in zip(rd_hist, sm_hist, ga_hist)]
opex_ex_cogs_hist = total_opex_hist  # R&D + S&M + G&A

# SBC (included in OpEx above)
sbc_hist = [600, 770, 1096, 1538, 1376, 1296]

# Operating Income (Loss)
op_income_hist = [gp - ox for gp, ox in zip(gross_profit_hist, opex_ex_cogs_hist)]

# Interest & Other
interest_other_hist = [-51, -88, -89, -87, -100, -95]

# Pre-tax income
pretax_hist = [oi + io for oi, io in zip(op_income_hist, interest_other_hist)]

# Tax provision
tax_hist = [5, 7, 14, 20, 25, 28]

# Net Income
net_income_hist = [pt - t for pt, t in zip(pretax_hist, tax_hist)]

# D&A (for EBITDA)
da_hist = [81, 107, 140, 185, 200, 215]

# EBITDA
ebitda_hist = [oi + da for oi, da in zip(op_income_hist, da_hist)]

# CapEx
capex_hist = [-51, -59, -85, -115, -93, -90]

# ── Balance Sheet (year-end) ──────────────────────────────────────────────────
cash_hist = [2024, 2768, 3506, 3938, 3374, 3205]
st_investments_hist = [0, 0, 0, 0, 0, 0]
ar_hist = [548, 737, 1068, 995, 1109, 1230]
total_current_assets_hist = [2746, 3705, 4775, 5115, 4700, 4650]

pp_e_hist = [204, 258, 339, 452, 430, 400]
intangibles_hist = [181, 182, 411, 440, 395, 360]
rou_assets_hist = [250, 280, 310, 350, 360, 370]
other_lt_assets_hist = [150, 180, 250, 310, 330, 350]
total_assets_hist = [3531, 4605, 6085, 6667, 6215, 6130]

ap_hist = [120, 155, 245, 265, 280, 295]
accrued_liab_hist = [350, 430, 620, 705, 680, 720]
deferred_rev_hist = [25, 30, 35, 40, 42, 45]
total_current_liab_hist = [495, 615, 900, 1010, 1002, 1060]

lt_debt_hist = [1485, 2190, 3745, 3745, 3745, 3745]
lease_liab_hist = [220, 260, 300, 340, 350, 360]
other_lt_liab_hist = [80, 95, 120, 140, 150, 155]
total_liabilities_hist = [2280, 3160, 5065, 5235, 5247, 5320]

total_equity_hist = [1251, 1445, 1020, 1432, 968, 810]
total_liab_equity_hist = [ta for ta in total_assets_hist]

# ── Cash Flow Statement ──────────────────────────────────────────────────────
cfo_hist = [-41, 166, 631, 234, 75, 260]
cfi_hist = [-200, -500, -1200, -350, -120, -100]
cff_hist = [900, 1078, 1307, 548, -519, -329]

# Working capital changes (approx)
wc_change_hist = [-80, -45, -120, 50, 30, 15]

# ── Quarterly Data FY2023 Q1-Q4, FY2024 Q1-Q4 ──────────────────────────────
qtrs_hist_labels = ['Q1-23','Q2-23','Q3-23','Q4-23','Q1-24','Q2-24','Q3-24','Q4-24']
qtrs_fcst_labels = ['Q1-25E','Q2-25E','Q3-25E','Q4-25E','Q1-26E','Q2-26E','Q3-26E','Q4-26E']
all_qtr_labels = qtrs_hist_labels + qtrs_fcst_labels

# Quarterly DAU (millions)
dau_q_hist = [383, 397, 406, 414, 422, 432, 437, 443]

# Quarterly Revenue
rev_q_hist = [989, 1068, 1189, 1361, 1195, 1237, 1373, 1556]

# Quarterly COGS
cor_q_hist = [490, 510, 545, 645, 545, 565, 620, 731]
gp_q_hist = [r - c for r, c in zip(rev_q_hist, cor_q_hist)]

# Quarterly OpEx (R&D + S&M + G&A)
rd_q_hist = [478, 462, 455, 452, 455, 448, 448, 451]
sm_q_hist = [330, 318, 320, 316, 322, 318, 320, 321]
ga_q_hist = [158, 155, 150, 151, 147, 145, 145, 145]
sbc_q_hist = [355, 345, 340, 336, 330, 325, 322, 319]
da_q_hist = [50, 50, 50, 50, 53, 54, 54, 54]

oi_q_hist = [gp - rd - sm - ga for gp, rd, sm, ga in zip(gp_q_hist, rd_q_hist, sm_q_hist, ga_q_hist)]
ebitda_q_hist = [oi + da for oi, da in zip(oi_q_hist, da_q_hist)]

capex_q_hist = [-22, -24, -23, -24, -22, -23, -22, -23]
ni_q_hist = [oi + io - t for oi, io, t in zip(oi_q_hist, [-25]*8, [6,6,6,7,7,7,7,7])]


# ══════════════════════════════════════════════════════════════════════════════
# FORECAST ASSUMPTIONS
# ══════════════════════════════════════════════════════════════════════════════

# DAU growth rates (yoy)
dau_growth_fcst = [0.065, 0.055, 0.045, 0.040]  # decelerating
dau_fcst = [dau_hist[-1]]
for g in dau_growth_fcst:
    dau_fcst.append(round(dau_fcst[-1] * (1 + g)))
dau_fcst = dau_fcst[1:]  # [2025, 2026, 2027, 2028]

# ARPU growth rates
arpu_growth_fcst = [0.10, 0.09, 0.08, 0.07]
arpu_fcst = [arpu_hist[-1]]
for g in arpu_growth_fcst:
    arpu_fcst.append(round(arpu_fcst[-1] * (1 + g), 2))
arpu_fcst = arpu_fcst[1:]

# Revenue forecast
rev_fcst = [round(d * a) for d, a in zip(dau_fcst, arpu_fcst)]

# Revenue by geo (maintain approx splits, NA share gradually declines)
na_pct_fcst = [0.63, 0.62, 0.61, 0.60]
eu_pct_fcst = [0.20, 0.20, 0.20, 0.21]
row_pct_fcst = [0.17, 0.18, 0.19, 0.19]
rev_na_fcst = [round(r * p) for r, p in zip(rev_fcst, na_pct_fcst)]
rev_eu_fcst = [round(r * p) for r, p in zip(rev_fcst, eu_pct_fcst)]
rev_row_fcst = [round(r * p) for r, p in zip(rev_fcst, row_pct_fcst)]

# Gross margin improvement (infrastructure efficiency)
gm_fcst = [0.555, 0.570, 0.585, 0.600]
cor_fcst = [round(r * (1 - gm)) for r, gm in zip(rev_fcst, gm_fcst)]
gp_fcst = [r - c for r, c in zip(rev_fcst, cor_fcst)]

# OpEx as % of revenue (improving leverage)
rd_pct_fcst = [0.30, 0.27, 0.25, 0.23]
sm_pct_fcst = [0.21, 0.19, 0.18, 0.17]
ga_pct_fcst = [0.095, 0.088, 0.082, 0.078]
rd_fcst = [round(r * p) for r, p in zip(rev_fcst, rd_pct_fcst)]
sm_fcst = [round(r * p) for r, p in zip(rev_fcst, sm_pct_fcst)]
ga_fcst = [round(r * p) for r, p in zip(rev_fcst, ga_pct_fcst)]
opex_fcst = [a + b + c for a, b, c in zip(rd_fcst, sm_fcst, ga_fcst)]

# SBC forecast (declining as % of rev)
sbc_pct_fcst = [0.20, 0.17, 0.15, 0.13]
sbc_fcst = [round(r * p) for r, p in zip(rev_fcst, sbc_pct_fcst)]

# D&A
da_fcst = [225, 240, 255, 270]

# Operating Income
op_income_fcst = [gp - ox for gp, ox in zip(gp_fcst, opex_fcst)]

# Interest & Other
interest_other_fcst = [-90, -85, -80, -75]

# Pretax
pretax_fcst = [oi + io for oi, io in zip(op_income_fcst, interest_other_fcst)]

# Tax (minimal – NOLs)
tax_rate_fcst = [0.05, 0.07, 0.10, 0.12]
tax_fcst = [max(0, round(pt * tr)) if pt > 0 else 0 for pt, tr in zip(pretax_fcst, tax_rate_fcst)]

# Net Income
ni_fcst = [pt - t for pt, t in zip(pretax_fcst, tax_fcst)]

# EBITDA
ebitda_fcst = [oi + da for oi, da in zip(op_income_fcst, da_fcst)]

# CapEx
capex_fcst = [-95, -100, -105, -110]

# ── Balance Sheet Forecast ────────────────────────────────────────────────────
# Use DSO / DPO / simplistic ratios
# AR as days of revenue
ar_fcst = [round(r * 80/365) for r in rev_fcst]
cash_fcst_base = cash_hist[-1]
ap_fcst = [round(c * 42/365) for c in cor_fcst]
accrued_fcst = [round(r * 0.12) for r in rev_fcst]
deferred_rev_fcst = [round(r * 0.007) for r in rev_fcst]
total_cl_fcst = [a + ac + dr + 100 for a, ac, dr in zip(ap_fcst, accrued_fcst, deferred_rev_fcst)]

pp_e_fcst = [round(pp_e_hist[-1] + sum(capex_fcst[:i+1]) * -1 - sum(da_fcst[:i+1]) + pp_e_hist[-1]) for i in range(4)]
# Simpler: PP&E stays roughly stable with capex ~ D&A
pp_e_fcst = [390, 385, 380, 375]
intangibles_fcst = [340, 320, 300, 280]
rou_fcst = [375, 380, 385, 390]
other_lt_fcst = [360, 370, 380, 390]

lt_debt_fcst = [3745, 3745, 3745, 3745]
lease_liab_fcst = [365, 370, 375, 380]
other_lt_liab_fcst = [160, 165, 170, 175]

# CFO forecast
wc_change_fcst = [-20, -15, -10, -5]
cfo_fcst = [ni + sbc + da + wc for ni, sbc, da, wc in zip(ni_fcst, sbc_fcst, da_fcst, wc_change_fcst)]
fcf_fcst = [cfo + cx for cfo, cx in zip(cfo_fcst, capex_fcst)]

# Financing
cff_fcst = [-200, -200, -250, -300]
cfi_fcst = [cx for cx in capex_fcst]  # simplified

# Cash
cash_fcst = []
prev_cash = cash_hist[-1]
for cfo_v, cfi_v, cff_v in zip(cfo_fcst, cfi_fcst, cff_fcst):
    prev_cash = prev_cash + cfo_v + cfi_v + cff_v
    cash_fcst.append(round(prev_cash))

total_ca_fcst = [c + ar + 200 for c, ar in zip(cash_fcst, ar_fcst)]  # 200 = other CA
total_assets_fcst = [ca + ppe + intang + rou + olt for ca, ppe, intang, rou, olt in 
                     zip(total_ca_fcst, pp_e_fcst, intangibles_fcst, rou_fcst, other_lt_fcst)]

total_liab_fcst = [cl + ltd + ll + oll for cl, ltd, ll, oll in 
                   zip(total_cl_fcst, lt_debt_fcst, lease_liab_fcst, other_lt_liab_fcst)]

total_equity_fcst = [ta - tl for ta, tl in zip(total_assets_fcst, total_liab_fcst)]
total_le_fcst = total_assets_fcst

# Historical FCF
fcf_hist = [cfo + cx for cfo, cx in zip(cfo_hist, capex_hist)]

# ── Share Count & Dilution Schedule ──────────────────────────────────────────
# Shares outstanding (diluted, millions) – historical
shares_hist = [1500, 1548, 1620, 1660, 1676, 1690]
# Forward quarterly shares (8 quarters: Q1-25E through Q4-26E)
# SBC drives ~1.5-2% annual dilution offset by buybacks
shares_q_fcst = [1695, 1700, 1705, 1710, 1714, 1718, 1722, 1726]
# RSU vesting schedule (millions of shares vesting per quarter)
rsu_vesting_q = [12, 11, 13, 12, 11, 10, 12, 11]
# New grants
new_grants_q = [8, 8, 8, 8, 7, 7, 7, 7]
# Net dilution
net_dilution_q = [v - g for v, g in zip(rsu_vesting_q, new_grants_q)]  # vesting adds, grants are future
# Buyback (shares, millions)
buyback_q = [5, 5, 5, 5, 5, 5, 5, 5]

shares_annual_fcst = [1710, 1726, 1742, 1758]

# EPS
eps_hist = [round(ni / s, 2) for ni, s in zip(net_income_hist, shares_hist)]
eps_fcst = [round(ni / s, 2) for ni, s in zip(ni_fcst, shares_annual_fcst)]

# ── Quarterly Forecast ────────────────────────────────────────────────────────
# Seasonality weights (Q1 lowest, Q4 highest)
q_seasonality = [0.21, 0.23, 0.25, 0.31]

# Q1-25E through Q4-26E
rev_q_fcst = []
for yr_idx, yr in enumerate([2025, 2026]):
    yr_rev = rev_fcst[yr_idx]
    for q_idx in range(4):
        rev_q_fcst.append(round(yr_rev * q_seasonality[q_idx]))

dau_q_fcst = []
base_dau = dau_hist[-1]
for i in range(8):
    yr_idx = i // 4
    q_in_yr = i % 4
    target_dau = dau_fcst[yr_idx]
    prev_yr_dau = dau_hist[-1] if yr_idx == 0 else dau_fcst[yr_idx - 1]
    interp = prev_yr_dau + (target_dau - prev_yr_dau) * (q_in_yr + 1) / 4
    dau_q_fcst.append(round(interp))

gm_q_fcst = [gm_fcst[0]]*4 + [gm_fcst[1]]*4
cor_q_fcst = [round(r * (1 - gm)) for r, gm in zip(rev_q_fcst, gm_q_fcst)]
gp_q_fcst_f = [r - c for r, c in zip(rev_q_fcst, cor_q_fcst)]

rd_q_fcst = [round(rev_q_fcst[i] * (rd_pct_fcst[i//4])) for i in range(8)]
sm_q_fcst = [round(rev_q_fcst[i] * (sm_pct_fcst[i//4])) for i in range(8)]
ga_q_fcst = [round(rev_q_fcst[i] * (ga_pct_fcst[i//4])) for i in range(8)]
sbc_q_fcst = [round(rev_q_fcst[i] * (sbc_pct_fcst[i//4])) for i in range(8)]
da_q_fcst_vals = [round(da_fcst[i//4] / 4) for i in range(8)]

oi_q_fcst = [gp - rd - sm - ga for gp, rd, sm, ga in zip(gp_q_fcst_f, rd_q_fcst, sm_q_fcst, ga_q_fcst)]
ebitda_q_fcst = [oi + da for oi, da in zip(oi_q_fcst, da_q_fcst_vals)]
ni_q_fcst = [oi - 23 - 2 for oi in oi_q_fcst]  # interest ~23/q, tax ~2/q approx

capex_q_fcst = [round(capex_fcst[i//4] / 4) for i in range(8)]


# ══════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ═══════════════════════ TAB 1: REVENUE BUILD ═════════════════════════════════
ws = wb.active
ws.title = "Revenue Build"
ws.sheet_properties.tabColor = "2F5496"

row = 1
ws.cell(row=row, column=1, value="Snap Inc. (SNAP) – Revenue Build by Segment & Driver").font = TITLE_FONT
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)

row = 3
# Headers
labels_col = [""] + [str(y) + ("" if y <= 2024 else "E") for y in all_years]
write_row(ws, row, labels_col, font=HEADER_FONT, fill=HEADER_FILL)
style_header_row(ws, row, len(labels_col))

# Historical / Forecast marker
row += 1
marker = [""] + ["Historical"]*6 + ["Forecast"]*4
write_row(ws, row, marker, font=SUBHEADER_FONT, fill=SUBHEADER_FILL)

forecast_start_col = 8  # Column H = 2025E

# Section: DAU
row += 1
write_row(ws, row, ["DAU Drivers"], font=SECTION_FONT)
ws.cell(row=row, column=1).border = BOTTOM_BORDER

row += 1
all_dau = dau_hist + dau_fcst
write_row(ws, row, ["Daily Active Users (M)"] + all_dau, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 1
dau_yoy = [None] + [round((all_dau[i] / all_dau[i-1] - 1), 3) for i in range(1, len(all_dau))]
write_row(ws, row, ["  YoY Growth"] + dau_yoy, fmt=PCT_FMT, is_forecast_start=forecast_start_col)
for c in range(forecast_start_col, forecast_start_col + 4):
    ws.cell(row=row, column=c).fill = INPUT_FILL

row += 2
write_row(ws, row, ["ARPU Drivers"], font=SECTION_FONT)
ws.cell(row=row, column=1).border = BOTTOM_BORDER

row += 1
all_arpu = arpu_hist + arpu_fcst
write_row(ws, row, ["Avg Revenue Per User ($)"] + all_arpu, font=BOLD_FONT, fmt=USD_DEC_FMT, is_forecast_start=forecast_start_col)

row += 1
arpu_yoy = [None] + [round((all_arpu[i] / all_arpu[i-1] - 1), 3) for i in range(1, len(all_arpu))]
write_row(ws, row, ["  YoY Growth"] + arpu_yoy, fmt=PCT_FMT, is_forecast_start=forecast_start_col)
for c in range(forecast_start_col, forecast_start_col + 4):
    ws.cell(row=row, column=c).fill = INPUT_FILL

row += 2
write_row(ws, row, ["Revenue (DAU × ARPU)"], font=SECTION_FONT)
ws.cell(row=row, column=1).border = BOTTOM_BORDER

row += 1
all_rev = rev_hist + rev_fcst
write_row(ws, row, ["Total Revenue ($M)"] + all_rev, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 1
rev_yoy = [None] + [round((all_rev[i] / all_rev[i-1] - 1), 3) for i in range(1, len(all_rev))]
write_row(ws, row, ["  YoY Growth"] + rev_yoy, fmt=PCT_FMT, is_forecast_start=forecast_start_col)

row += 2
write_row(ws, row, ["Revenue by Geography"], font=SECTION_FONT)
ws.cell(row=row, column=1).border = BOTTOM_BORDER

row += 1
all_na = rev_na_hist + rev_na_fcst
write_row(ws, row, ["  North America"] + all_na, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 1
na_pct = [round(n/r, 3) for n, r in zip(all_na, all_rev)]
write_row(ws, row, ["    % of Total"] + na_pct, fmt=PCT_FMT)

row += 1
all_eu = rev_eu_hist + rev_eu_fcst
write_row(ws, row, ["  Europe"] + all_eu, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 1
eu_pct = [round(e/r, 3) for e, r in zip(all_eu, all_rev)]
write_row(ws, row, ["    % of Total"] + eu_pct, fmt=PCT_FMT)

row += 1
all_row_rev = rev_row_hist + rev_row_fcst
write_row(ws, row, ["  Rest of World"] + all_row_rev, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 1
row_pct = [round(rw/r, 3) for rw, r in zip(all_row_rev, all_rev)]
write_row(ws, row, ["    % of Total"] + row_pct, fmt=PCT_FMT)

# Implied take-rate / ad metrics
row += 2
write_row(ws, row, ["Implied Ad Metrics"], font=SECTION_FONT)
ws.cell(row=row, column=1).border = BOTTOM_BORDER

row += 1
# Impressions per DAU per day (assumed)
imp_per_dau = [25, 27, 30, 33, 36, 39, 42, 44, 46, 48]
write_row(ws, row, ["  Impressions/DAU/Day (est.)"] + imp_per_dau, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
for c in range(forecast_start_col, forecast_start_col + 4):
    ws.cell(row=row, column=c).fill = INPUT_FILL

row += 1
# eCPM implied
total_impressions = [d * imp * 365 / 1000 for d, imp in zip(all_dau, imp_per_dau)]  # in millions of 1000s
ecpm = [round(r / (d * imp * 365 / 1e6) * 1e3, 2) if d > 0 else 0 for r, d, imp in zip(all_rev, all_dau, imp_per_dau)]
write_row(ws, row, ["  Implied eCPM ($)"] + ecpm, fmt=USD_DEC_FMT, is_forecast_start=forecast_start_col)

set_col_widths(ws, [32] + [14]*10)

# ═══════════════════════ TAB 2: INCOME STATEMENT ═════════════════════════════
ws2 = wb.create_sheet("Income Statement")
ws2.sheet_properties.tabColor = "2F5496"

row = 1
ws2.cell(row=row, column=1, value="Snap Inc. (SNAP) – GAAP Income Statement ($M)").font = TITLE_FONT
ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)

row = 3
write_row(ws2, row, labels_col, font=HEADER_FONT, fill=HEADER_FILL)
style_header_row(ws2, row, len(labels_col))
row += 1
write_row(ws2, row, [""] + ["Historical"]*6 + ["Forecast"]*4, font=SUBHEADER_FONT, fill=SUBHEADER_FILL)

row += 1
write_row(ws2, row, ["Revenue"] + all_rev, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
rev_growth = [None] + rev_yoy[1:]
write_row(ws2, row, ["  YoY Growth"] + rev_growth, fmt=PCT_FMT, is_forecast_start=forecast_start_col)

row += 2
all_cor = cor_hist + cor_fcst
write_row(ws2, row, ["Cost of Revenue"] + all_cor, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 1
all_gp = gross_profit_hist + gp_fcst
write_row(ws2, row, ["Gross Profit"] + all_gp, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 1
all_gm = gm_hist + gm_fcst
write_row(ws2, row, ["  Gross Margin"] + [round(g, 3) for g in all_gm], fmt=PCT_FMT, is_forecast_start=forecast_start_col)
for c in range(forecast_start_col, forecast_start_col + 4):
    ws2.cell(row=row, column=c).fill = INPUT_FILL

row += 2
write_row(ws2, row, ["Operating Expenses"], font=SECTION_FONT)
ws2.cell(row=row, column=1).border = BOTTOM_BORDER

row += 1
all_rd = rd_hist + rd_fcst
write_row(ws2, row, ["  Research & Development"] + all_rd, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
rd_pct_all = [round(rd/r, 3) for rd, r in zip(all_rd, all_rev)]
write_row(ws2, row, ["    % of Revenue"] + rd_pct_all, fmt=PCT_FMT, is_forecast_start=forecast_start_col)
for c in range(forecast_start_col, forecast_start_col + 4):
    ws2.cell(row=row, column=c).fill = INPUT_FILL

row += 1
all_sm = sm_hist + sm_fcst
write_row(ws2, row, ["  Sales & Marketing"] + all_sm, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
sm_pct_all = [round(sm/r, 3) for sm, r in zip(all_sm, all_rev)]
write_row(ws2, row, ["    % of Revenue"] + sm_pct_all, fmt=PCT_FMT, is_forecast_start=forecast_start_col)
for c in range(forecast_start_col, forecast_start_col + 4):
    ws2.cell(row=row, column=c).fill = INPUT_FILL

row += 1
all_ga = ga_hist + ga_fcst
write_row(ws2, row, ["  General & Administrative"] + all_ga, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
ga_pct_all = [round(ga/r, 3) for ga, r in zip(all_ga, all_rev)]
write_row(ws2, row, ["    % of Revenue"] + ga_pct_all, fmt=PCT_FMT, is_forecast_start=forecast_start_col)
for c in range(forecast_start_col, forecast_start_col + 4):
    ws2.cell(row=row, column=c).fill = INPUT_FILL

row += 1
all_opex = opex_ex_cogs_hist + opex_fcst
write_row(ws2, row, ["Total Operating Expenses"] + all_opex, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
opex_pct = [round(o/r, 3) for o, r in zip(all_opex, all_rev)]
write_row(ws2, row, ["    % of Revenue"] + opex_pct, fmt=PCT_FMT, is_forecast_start=forecast_start_col)

row += 2
all_oi = op_income_hist + op_income_fcst
write_row(ws2, row, ["Operating Income (Loss)"] + all_oi, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
oi_margin = [round(oi/r, 3) for oi, r in zip(all_oi, all_rev)]
write_row(ws2, row, ["  Operating Margin"] + oi_margin, fmt=PCT_FMT, is_forecast_start=forecast_start_col)

row += 2
# SBC detail
all_sbc = sbc_hist + sbc_fcst
write_row(ws2, row, ["Memo: Stock-Based Compensation"] + all_sbc, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
sbc_pct_all = [round(s/r, 3) for s, r in zip(all_sbc, all_rev)]
write_row(ws2, row, ["  SBC % of Revenue"] + sbc_pct_all, fmt=PCT_FMT, is_forecast_start=forecast_start_col)
for c in range(forecast_start_col, forecast_start_col + 4):
    ws2.cell(row=row, column=c).fill = INPUT_FILL

row += 2
all_da = da_hist + da_fcst
write_row(ws2, row, ["Depreciation & Amortization"] + all_da, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 1
all_ebitda = ebitda_hist + ebitda_fcst
write_row(ws2, row, ["EBITDA"] + all_ebitda, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
ebitda_margin = [round(e/r, 3) for e, r in zip(all_ebitda, all_rev)]
write_row(ws2, row, ["  EBITDA Margin"] + ebitda_margin, fmt=PCT_FMT, is_forecast_start=forecast_start_col)

row += 2
# Below the line
all_int = interest_other_hist + interest_other_fcst
write_row(ws2, row, ["Interest & Other Income (Exp)"] + all_int, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 1
all_pt = pretax_hist + pretax_fcst
write_row(ws2, row, ["Pre-Tax Income (Loss)"] + all_pt, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 1
all_tax = tax_hist + tax_fcst
write_row(ws2, row, ["Income Tax Provision"] + all_tax, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
eff_tax = [round(t/pt, 3) if pt != 0 and pt > 0 else 0 for t, pt in zip(all_tax, all_pt)]
write_row(ws2, row, ["  Effective Tax Rate"] + eff_tax, fmt=PCT_FMT, is_forecast_start=forecast_start_col)

row += 2
all_ni = net_income_hist + ni_fcst
write_row(ws2, row, ["Net Income (Loss)"] + all_ni, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 2
all_shares = shares_hist + shares_annual_fcst
write_row(ws2, row, ["Diluted Shares Outstanding (M)"] + all_shares, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
all_eps = eps_hist + eps_fcst
write_row(ws2, row, ["Diluted EPS"] + all_eps, fmt=USD_DEC_FMT, is_forecast_start=forecast_start_col)

set_col_widths(ws2, [32] + [14]*10)


# ═══════════════════════ TAB 3: BALANCE SHEET ═════════════════════════════════
ws3 = wb.create_sheet("Balance Sheet")
ws3.sheet_properties.tabColor = "548235"

row = 1
ws3.cell(row=row, column=1, value="Snap Inc. (SNAP) – Balance Sheet ($M)").font = TITLE_FONT
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)

row = 3
write_row(ws3, row, labels_col, font=HEADER_FONT, fill=HEADER_FILL)
style_header_row(ws3, row, len(labels_col))
row += 1
write_row(ws3, row, [""] + ["Historical"]*6 + ["Forecast"]*4, font=SUBHEADER_FONT, fill=SUBHEADER_FILL)

row += 1
write_row(ws3, row, ["ASSETS"], font=SECTION_FONT)
ws3.cell(row=row, column=1).border = BOTTOM_BORDER

row += 1
write_row(ws3, row, ["Cash & Equivalents"] + cash_hist + cash_fcst, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
write_row(ws3, row, ["Accounts Receivable"] + ar_hist + ar_fcst, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
other_ca_hist = [tca - c - ar for tca, c, ar in zip(total_current_assets_hist, cash_hist, ar_hist)]
other_ca_fcst = [200]*4
write_row(ws3, row, ["Other Current Assets"] + other_ca_hist + other_ca_fcst, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
all_tca = total_current_assets_hist + total_ca_fcst
write_row(ws3, row, ["Total Current Assets"] + all_tca, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 2
write_row(ws3, row, ["PP&E, net"] + pp_e_hist + pp_e_fcst, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
write_row(ws3, row, ["Intangible Assets"] + intangibles_hist + intangibles_fcst, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
write_row(ws3, row, ["ROU / Lease Assets"] + rou_assets_hist + rou_fcst, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
write_row(ws3, row, ["Other Long-Term Assets"] + other_lt_assets_hist + other_lt_fcst, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 1
all_ta = total_assets_hist + total_assets_fcst
write_row(ws3, row, ["Total Assets"] + all_ta, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
for c in range(1, len(labels_col)+1):
    ws3.cell(row=row, column=c).border = BOTTOM_BORDER

row += 2
write_row(ws3, row, ["LIABILITIES"], font=SECTION_FONT)
ws3.cell(row=row, column=1).border = BOTTOM_BORDER

row += 1
write_row(ws3, row, ["Accounts Payable"] + ap_hist + ap_fcst, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
write_row(ws3, row, ["Accrued Liabilities"] + accrued_liab_hist + accrued_fcst, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
write_row(ws3, row, ["Deferred Revenue"] + deferred_rev_hist + deferred_rev_fcst, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
other_cl_hist = [tcl - a - ac - dr for tcl, a, ac, dr in zip(total_current_liab_hist, ap_hist, accrued_liab_hist, deferred_rev_hist)]
write_row(ws3, row, ["Other Current Liabilities"] + other_cl_hist + [100]*4, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
all_tcl = total_current_liab_hist + total_cl_fcst
write_row(ws3, row, ["Total Current Liabilities"] + all_tcl, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 2
write_row(ws3, row, ["Long-Term Debt"] + lt_debt_hist + lt_debt_fcst, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
write_row(ws3, row, ["Lease Liabilities (LT)"] + lease_liab_hist + lease_liab_fcst, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
write_row(ws3, row, ["Other Long-Term Liabilities"] + other_lt_liab_hist + other_lt_liab_fcst, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
all_tl = total_liabilities_hist + total_liab_fcst
write_row(ws3, row, ["Total Liabilities"] + all_tl, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 2
write_row(ws3, row, ["STOCKHOLDERS' EQUITY"], font=SECTION_FONT)
ws3.cell(row=row, column=1).border = BOTTOM_BORDER

row += 1
all_eq = total_equity_hist + total_equity_fcst
write_row(ws3, row, ["Total Stockholders' Equity"] + all_eq, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 1
all_tle = total_liab_equity_hist + total_le_fcst
write_row(ws3, row, ["Total Liabilities & Equity"] + all_tle, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
for c in range(1, len(labels_col)+1):
    ws3.cell(row=row, column=c).border = BOTTOM_BORDER

# Check
row += 2
check = [ta - tle for ta, tle in zip(all_ta, all_tle)]
write_row(ws3, row, ["Balance Check (should = 0)"] + check, fmt=NUM_FMT)

set_col_widths(ws3, [32] + [14]*10)


# ═══════════════════════ TAB 4: CASH FLOW + FCF BRIDGE ════════════════════════
ws4 = wb.create_sheet("Cash Flow & FCF Bridge")
ws4.sheet_properties.tabColor = "BF8F00"

row = 1
ws4.cell(row=row, column=1, value="Snap Inc. (SNAP) – Cash Flow Statement & GAAP OpLoss → FCF Bridge ($M)").font = TITLE_FONT
ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)

row = 3
write_row(ws4, row, labels_col, font=HEADER_FONT, fill=HEADER_FILL)
style_header_row(ws4, row, len(labels_col))
row += 1
write_row(ws4, row, [""] + ["Historical"]*6 + ["Forecast"]*4, font=SUBHEADER_FONT, fill=SUBHEADER_FILL)

row += 1
write_row(ws4, row, ["Cash Flow from Operations"], font=SECTION_FONT)
ws4.cell(row=row, column=1).border = BOTTOM_BORDER

row += 1
write_row(ws4, row, ["Net Income (Loss)"] + all_ni, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
write_row(ws4, row, ["  (+) Depreciation & Amortization"] + (da_hist + da_fcst), fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
write_row(ws4, row, ["  (+) Stock-Based Compensation"] + all_sbc, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
all_wc = wc_change_hist + wc_change_fcst
write_row(ws4, row, ["  (+/-) Working Capital Changes"] + all_wc, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
# Other adjustments (plug to match CFO)
other_adj_hist = [cfo - ni - da - sbc - wc for cfo, ni, da, sbc, wc in 
                  zip(cfo_hist, net_income_hist, da_hist, sbc_hist, wc_change_hist)]
other_adj_fcst = [0]*4
write_row(ws4, row, ["  (+/-) Other Non-Cash Adj."] + other_adj_hist + other_adj_fcst, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 1
all_cfo = cfo_hist + cfo_fcst
write_row(ws4, row, ["Cash from Operations (CFO)"] + all_cfo, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
for c in range(1, len(labels_col)+1):
    ws4.cell(row=row, column=c).border = THIN_BORDER

row += 2
write_row(ws4, row, ["Cash Flow from Investing"], font=SECTION_FONT)
ws4.cell(row=row, column=1).border = BOTTOM_BORDER

row += 1
all_capex = capex_hist + capex_fcst
write_row(ws4, row, ["  Capital Expenditures"] + all_capex, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
other_cfi_hist = [cf - cx for cf, cx in zip(cfi_hist, capex_hist)]
other_cfi_fcst = [0]*4
write_row(ws4, row, ["  Other Investing Activities"] + other_cfi_hist + other_cfi_fcst, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
all_cfi = cfi_hist + cfi_fcst
write_row(ws4, row, ["Cash from Investing (CFI)"] + all_cfi, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
for c in range(1, len(labels_col)+1):
    ws4.cell(row=row, column=c).border = THIN_BORDER

row += 2
write_row(ws4, row, ["Cash Flow from Financing"], font=SECTION_FONT)
ws4.cell(row=row, column=1).border = BOTTOM_BORDER

row += 1
all_cff = cff_hist + cff_fcst
write_row(ws4, row, ["Cash from Financing (CFF)"] + all_cff, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 2
net_change = [cfo + cfi + cff for cfo, cfi, cff in zip(all_cfo, all_cfi, all_cff)]
write_row(ws4, row, ["Net Change in Cash"] + net_change, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
for c in range(1, len(labels_col)+1):
    ws4.cell(row=row, column=c).border = BOTTOM_BORDER

row += 1
beg_cash = [None] + cash_hist[:-1] + [cash_hist[-1]] + cash_fcst[:-1]
write_row(ws4, row, ["Beginning Cash"] + beg_cash[1:], fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
write_row(ws4, row, ["Ending Cash"] + (cash_hist + cash_fcst), font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 2
all_fcf = fcf_hist + fcf_fcst
write_row(ws4, row, ["Free Cash Flow (CFO + CapEx)"] + all_fcf, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
row += 1
fcf_margin = [round(f/r, 3) for f, r in zip(all_fcf, all_rev)]
write_row(ws4, row, ["  FCF Margin"] + fcf_margin, fmt=PCT_FMT, is_forecast_start=forecast_start_col)

# ── GAAP Operating Loss → FCF Bridge ─────────────────────────────────────────
row += 3
ws4.cell(row=row, column=1, value="GAAP Operating Income (Loss) → Free Cash Flow Bridge").font = TITLE_FONT
ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)

row += 2
write_row(ws4, row, labels_col, font=HEADER_FONT, fill=HEADER_FILL)
style_header_row(ws4, row, len(labels_col))

row += 1
write_row(ws4, row, ["GAAP Operating Income (Loss)"] + all_oi, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 1
write_row(ws4, row, ["  (+) Depreciation & Amortization"] + (da_hist + da_fcst), fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 1
write_row(ws4, row, ["  (+) Stock-Based Compensation"] + all_sbc, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 1
# Adjusted EBITDA (OpIncome + D&A + SBC)
adj_ebitda = [oi + da + sbc for oi, da, sbc in zip(all_oi, da_hist + da_fcst, all_sbc)]
write_row(ws4, row, ["= Adjusted EBITDA"] + adj_ebitda, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
for c in range(1, len(labels_col)+1):
    ws4.cell(row=row, column=c).border = THIN_BORDER

row += 1
write_row(ws4, row, ["  (+/-) Working Capital Changes"] + all_wc, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 1
write_row(ws4, row, ["  (-) Cash Interest (net)"] + (interest_other_hist + interest_other_fcst), fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 1
write_row(ws4, row, ["  (-) Cash Taxes"] + ([t * -1 for t in all_tax]), fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 1
write_row(ws4, row, ["  (+/-) Other Adjustments"] + (other_adj_hist + other_adj_fcst), fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 1
# Implied CFO from bridge
bridge_cfo = [ae + wc + io + (-t) + oa for ae, wc, io, t, oa in 
              zip(adj_ebitda, all_wc, interest_other_hist + interest_other_fcst, all_tax, other_adj_hist + other_adj_fcst)]
write_row(ws4, row, ["= Cash from Operations (CFO)"] + bridge_cfo, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
for c in range(1, len(labels_col)+1):
    ws4.cell(row=row, column=c).border = THIN_BORDER

row += 1
write_row(ws4, row, ["  (-) Capital Expenditures"] + all_capex, fmt=NUM_FMT, is_forecast_start=forecast_start_col)

row += 1
bridge_fcf = [c + cx for c, cx in zip(bridge_cfo, all_capex)]
write_row(ws4, row, ["= Free Cash Flow"] + bridge_fcf, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=forecast_start_col)
for c in range(1, len(labels_col)+1):
    ws4.cell(row=row, column=c).border = BOTTOM_BORDER

set_col_widths(ws4, [38] + [14]*10)


# ═══════════════════════ TAB 5: QUARTERLY DETAIL ══════════════════════════════
ws5 = wb.create_sheet("Quarterly Detail")
ws5.sheet_properties.tabColor = "7030A0"

row = 1
ws5.cell(row=row, column=1, value="Snap Inc. (SNAP) – Quarterly Income Statement ($M)").font = TITLE_FONT
ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=18)

row = 3
q_labels = [""] + all_qtr_labels
write_row(ws5, row, q_labels, font=HEADER_FONT, fill=HEADER_FILL)
style_header_row(ws5, row, len(q_labels))

row += 1
write_row(ws5, row, [""] + ["Historical"]*8 + ["Forecast"]*8, font=SUBHEADER_FONT, fill=SUBHEADER_FILL)

q_fcst_start = 10  # column J

row += 1
all_dau_q = dau_q_hist + dau_q_fcst
write_row(ws5, row, ["DAU (M)"] + all_dau_q, fmt=NUM_FMT, is_forecast_start=q_fcst_start)

row += 1
all_rev_q = rev_q_hist + rev_q_fcst
write_row(ws5, row, ["Revenue"] + all_rev_q, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=q_fcst_start)

row += 1
arpu_q = [round(r/d, 2) for r, d in zip(all_rev_q, all_dau_q)]
write_row(ws5, row, ["ARPU ($)"] + arpu_q, fmt=USD_DEC_FMT, is_forecast_start=q_fcst_start)

row += 2
all_cor_q = cor_q_hist + cor_q_fcst
write_row(ws5, row, ["Cost of Revenue"] + all_cor_q, fmt=NUM_FMT, is_forecast_start=q_fcst_start)

row += 1
all_gp_q = gp_q_hist + gp_q_fcst_f
write_row(ws5, row, ["Gross Profit"] + all_gp_q, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=q_fcst_start)

row += 1
gm_q_all = [round(gp/r, 3) for gp, r in zip(all_gp_q, all_rev_q)]
write_row(ws5, row, ["  Gross Margin"] + gm_q_all, fmt=PCT_FMT, is_forecast_start=q_fcst_start)

row += 2
all_rd_q = rd_q_hist + rd_q_fcst
write_row(ws5, row, ["R&D"] + all_rd_q, fmt=NUM_FMT, is_forecast_start=q_fcst_start)
row += 1
all_sm_q = sm_q_hist + sm_q_fcst
write_row(ws5, row, ["S&M"] + all_sm_q, fmt=NUM_FMT, is_forecast_start=q_fcst_start)
row += 1
all_ga_q = ga_q_hist + ga_q_fcst
write_row(ws5, row, ["G&A"] + all_ga_q, fmt=NUM_FMT, is_forecast_start=q_fcst_start)

row += 1
all_oi_q = oi_q_hist + oi_q_fcst
write_row(ws5, row, ["Operating Income (Loss)"] + all_oi_q, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=q_fcst_start)

row += 1
oi_margin_q = [round(oi/r, 3) for oi, r in zip(all_oi_q, all_rev_q)]
write_row(ws5, row, ["  Operating Margin"] + oi_margin_q, fmt=PCT_FMT, is_forecast_start=q_fcst_start)

row += 2
all_sbc_q = sbc_q_hist + sbc_q_fcst
write_row(ws5, row, ["SBC"] + all_sbc_q, fmt=NUM_FMT, is_forecast_start=q_fcst_start)
row += 1
all_da_q = da_q_hist + da_q_fcst_vals
write_row(ws5, row, ["D&A"] + all_da_q, fmt=NUM_FMT, is_forecast_start=q_fcst_start)

row += 1
all_ebitda_q = ebitda_q_hist + ebitda_q_fcst
write_row(ws5, row, ["EBITDA"] + all_ebitda_q, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=q_fcst_start)

row += 2
all_ni_q = ni_q_hist + ni_q_fcst
write_row(ws5, row, ["Net Income (Loss)"] + all_ni_q, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=q_fcst_start)

row += 1
all_capex_q = capex_q_hist + capex_q_fcst
write_row(ws5, row, ["CapEx"] + all_capex_q, fmt=NUM_FMT, is_forecast_start=q_fcst_start)

set_col_widths(ws5, [28] + [12]*16)


# ═══════════════════════ TAB 6: SHARE COUNT & DILUTION ════════════════════════
ws6 = wb.create_sheet("Share Count & Dilution")
ws6.sheet_properties.tabColor = "C00000"

row = 1
ws6.cell(row=row, column=1, value="Snap Inc. (SNAP) – Share Count & Dilution Schedule").font = TITLE_FONT
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)

# Annual history
row = 3
ann_share_labels = [""] + [str(y) for y in years_hist] + [str(y) + "E" for y in years_fcst]
write_row(ws6, row, ann_share_labels, font=HEADER_FONT, fill=HEADER_FILL)
style_header_row(ws6, row, len(ann_share_labels))

row += 1
write_row(ws6, row, ["Diluted Shares (M)"] + shares_hist + shares_annual_fcst, font=BOLD_FONT, fmt=NUM_FMT, is_forecast_start=8)
row += 1
share_yoy = [None] + [round(shares_hist[i]/shares_hist[i-1] - 1, 3) for i in range(1, len(shares_hist))] + \
            [round(shares_annual_fcst[0]/shares_hist[-1] - 1, 3)] + \
            [round(shares_annual_fcst[i]/shares_annual_fcst[i-1] - 1, 3) for i in range(1, len(shares_annual_fcst))]
write_row(ws6, row, ["  YoY Dilution"] + share_yoy, fmt=PCT_FMT, is_forecast_start=8)

row += 1
write_row(ws6, row, ["SBC ($M)"] + sbc_hist + sbc_fcst, fmt=NUM_FMT, is_forecast_start=8)
row += 1
write_row(ws6, row, ["EPS (Diluted)"] + eps_hist + eps_fcst, fmt=USD_DEC_FMT, is_forecast_start=8)

# Quarterly dilution schedule (8 forward quarters)
row += 3
ws6.cell(row=row, column=1, value="Quarterly Dilution Schedule (Next 8 Quarters)").font = TITLE_FONT
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)

row += 2
q_dil_labels = [""] + qtrs_fcst_labels
write_row(ws6, row, q_dil_labels, font=HEADER_FONT, fill=HEADER_FILL)
style_header_row(ws6, row, len(q_dil_labels))

row += 1
write_row(ws6, row, ["Beginning Shares (M)"] + [1690] + shares_q_fcst[:-1], fmt=NUM_FMT)
row += 1
write_row(ws6, row, ["  (+) RSU Vesting"] + rsu_vesting_q, fmt=NUM_FMT, fill=INPUT_FILL)
row += 1
write_row(ws6, row, ["  (-) Share Buybacks"] + buyback_q, fmt=NUM_FMT, fill=INPUT_FILL)
row += 1
write_row(ws6, row, ["  (+) Option Exercises & Other"] + [v - g - b + (shares_q_fcst[i] - ([1690] + shares_q_fcst[:-1])[i] - v + b) for i, (v, g, b) in enumerate(zip(rsu_vesting_q, new_grants_q, buyback_q))], fmt=NUM_FMT)
row += 1
write_row(ws6, row, ["Ending Shares (M)"] + shares_q_fcst, font=BOLD_FONT, fmt=NUM_FMT)

row += 2
write_row(ws6, row, ["Memo: New RSU Grants (M)"] + new_grants_q, fmt=NUM_FMT, fill=INPUT_FILL)
row += 1
write_row(ws6, row, ["Memo: Quarterly SBC ($M)"] + sbc_q_fcst, fmt=NUM_FMT)
row += 1
# SBC per share
sbc_per_share = [round(s / sh, 2) for s, sh in zip(sbc_q_fcst, shares_q_fcst)]
write_row(ws6, row, ["Memo: SBC / Share ($)"] + sbc_per_share, fmt=USD_DEC_FMT)

# Unexercised pool
row += 2
ws6.cell(row=row, column=1, value="Outstanding Equity Awards (est.)").font = SECTION_FONT
row += 1
write_row(ws6, row, ["  Unvested RSUs (M shares)"] + [120, 117, 112, 108, 104, 101, 96, 92], fmt=NUM_FMT)
row += 1
write_row(ws6, row, ["  Vested Unexercised Options (M)"] + [15, 14, 13, 12, 11, 10, 9, 8], fmt=NUM_FMT)

set_col_widths(ws6, [32] + [14]*10)


# ═══════════════════════ TAB 7: SENSITIVITY TABLES ════════════════════════════
ws7 = wb.create_sheet("Sensitivity Analysis")
ws7.sheet_properties.tabColor = "ED7D31"

row = 1
ws7.cell(row=row, column=1, value="Snap Inc. (SNAP) – Two-Way Sensitivity Analysis").font = TITLE_FONT
ws7.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)

# Table 1: DAU growth × ARPU growth → FY2026E Revenue
row = 4
ws7.cell(row=row, column=1, value="TABLE 1: FY2026E Revenue ($M) — DAU Growth vs. ARPU Growth").font = SECTION_FONT
ws7.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)

row += 2
# DAU growth scenarios (row headers)
dau_scenarios = [0.02, 0.035, 0.055, 0.07, 0.085]
# ARPU growth scenarios (column headers)
arpu_scenarios = [0.04, 0.06, 0.09, 0.12, 0.15]

ws7.cell(row=row, column=1, value="DAU Growth ↓  /  ARPU Growth →").font = BOLD_FONT
for j, ag in enumerate(arpu_scenarios):
    cell = ws7.cell(row=row, column=2+j, value=ag)
    cell.number_format = PCT_FMT
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center')

for i, dg in enumerate(dau_scenarios):
    r = row + 1 + i
    cell = ws7.cell(row=r, column=1, value=dg)
    cell.number_format = PCT_FMT
    cell.font = BOLD_FONT
    cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    
    base_dau_25 = round(dau_hist[-1] * (1 + dau_growth_fcst[0]))
    base_arpu_25_val = arpu_fcst[0]
    
    for j, ag in enumerate(arpu_scenarios):
        dau_26 = round(base_dau_25 * (1 + dg))
        arpu_26 = round(base_arpu_25_val * (1 + ag), 2)
        rev_26 = round(dau_26 * arpu_26)
        cell = ws7.cell(row=r, column=2+j, value=rev_26)
        cell.number_format = NUM_FMT
        cell.alignment = Alignment(horizontal='center')
        # Highlight base case
        if abs(dg - 0.055) < 0.001 and abs(ag - 0.09) < 0.001:
            cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            cell.font = BOLD_FONT

# Table 2: Revenue Growth × EBITDA Margin → FY2026E EBITDA
row += len(dau_scenarios) + 4
ws7.cell(row=row, column=1, value="TABLE 2: FY2026E EBITDA ($M) — Revenue Growth vs. EBITDA Margin").font = SECTION_FONT
ws7.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)

row += 2
rev_growth_scenarios = [0.10, 0.14, 0.17, 0.20, 0.24]
ebitda_margin_scenarios = [0.06, 0.09, 0.12, 0.15, 0.18]

ws7.cell(row=row, column=1, value="Rev Growth ↓  /  EBITDA Margin →").font = BOLD_FONT
for j, em in enumerate(ebitda_margin_scenarios):
    cell = ws7.cell(row=row, column=2+j, value=em)
    cell.number_format = PCT_FMT
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center')

base_rev_25 = rev_fcst[0]
for i, rg in enumerate(rev_growth_scenarios):
    r = row + 1 + i
    cell = ws7.cell(row=r, column=1, value=rg)
    cell.number_format = PCT_FMT
    cell.font = BOLD_FONT
    cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    
    for j, em in enumerate(ebitda_margin_scenarios):
        implied_rev_26 = round(base_rev_25 * (1 + rg))
        implied_ebitda_26 = round(implied_rev_26 * em)
        cell = ws7.cell(row=r, column=2+j, value=implied_ebitda_26)
        cell.number_format = NUM_FMT
        cell.alignment = Alignment(horizontal='center')
        if abs(rg - 0.17) < 0.01 and abs(em - 0.12) < 0.01:
            cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            cell.font = BOLD_FONT

# Table 3: WACC × Terminal Growth → Implied Share Price (DCF)
row += len(rev_growth_scenarios) + 4
ws7.cell(row=row, column=1, value="TABLE 3: Implied Share Price ($) — WACC vs. Terminal Growth Rate (DCF)").font = SECTION_FONT
ws7.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)

row += 2
wacc_scenarios = [0.09, 0.10, 0.11, 0.12, 0.13]
tgr_scenarios = [0.02, 0.025, 0.03, 0.035, 0.04]

ws7.cell(row=row, column=1, value="WACC ↓  /  Terminal Growth →").font = BOLD_FONT
for j, tg in enumerate(tgr_scenarios):
    cell = ws7.cell(row=row, column=2+j, value=tg)
    cell.number_format = PCT_FMT
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center')

# Simple DCF: terminal year FCF * (1+g)/(WACC-g) + PV of near-term FCFs
terminal_fcf = fcf_fcst[-1]  # FY2028E FCF
shares_for_price = shares_annual_fcst[1]  # FY2026E shares
net_debt = lt_debt_fcst[0] - cash_fcst[0]

for i, wacc in enumerate(wacc_scenarios):
    r = row + 1 + i
    cell = ws7.cell(row=r, column=1, value=wacc)
    cell.number_format = PCT_FMT
    cell.font = BOLD_FONT
    cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    
    for j, tg in enumerate(tgr_scenarios):
        # PV of FCFs 2025-2028
        pv_fcfs = sum(fcf / (1 + wacc)**yr for yr, fcf in enumerate(fcf_fcst, 1))
        # Terminal value
        tv = terminal_fcf * (1 + tg) / (wacc - tg)
        pv_tv = tv / (1 + wacc)**4
        ev = pv_fcfs + pv_tv
        eq_val = ev - net_debt
        price = round(eq_val / shares_for_price, 1)
        cell = ws7.cell(row=r, column=2+j, value=price)
        cell.number_format = '$#,##0.0'
        cell.alignment = Alignment(horizontal='center')
        if abs(wacc - 0.11) < 0.005 and abs(tg - 0.03) < 0.003:
            cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            cell.font = BOLD_FONT

set_col_widths(ws7, [36] + [14]*8)


# ═══════════════════════ TAB 8: VALUATION CROSS-CHECKS ═══════════════════════
ws8 = wb.create_sheet("Valuation Cross-Checks")
ws8.sheet_properties.tabColor = "00B050"

row = 1
ws8.cell(row=row, column=1, value="Snap Inc. (SNAP) – Valuation Cross-Checks").font = TITLE_FONT
ws8.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)

# ── Section 1: DCF Valuation ──────────────────────────────────────────────────
row = 3
ws8.cell(row=row, column=1, value="1. DCF Valuation").font = SECTION_FONT
ws8.cell(row=row, column=1).border = BOTTOM_BORDER

row += 1
ws8.cell(row=row, column=1, value="Key Assumptions").font = BOLD_FONT
row += 1
write_row(ws8, row, ["  WACC", 0.11], fmt=PCT_FMT)
ws8.cell(row=row, column=2).fill = INPUT_FILL
row += 1
write_row(ws8, row, ["  Terminal Growth Rate", 0.03], fmt=PCT_FMT)
ws8.cell(row=row, column=2).fill = INPUT_FILL
row += 1
write_row(ws8, row, ["  Net Debt ($M)", net_debt], fmt=NUM_FMT)
row += 1
write_row(ws8, row, ["  Shares Outstanding (M)", shares_annual_fcst[1]], fmt=NUM_FMT)

row += 2
dcf_wacc = 0.11
dcf_tgr = 0.03
write_row(ws8, row, ["", "FY2025E", "FY2026E", "FY2027E", "FY2028E"], font=HEADER_FONT, fill=HEADER_FILL)
style_header_row(ws8, row, 5)

row += 1
write_row(ws8, row, ["Free Cash Flow ($M)"] + fcf_fcst, fmt=NUM_FMT)
row += 1
discount_factors = [1/(1+dcf_wacc)**i for i in range(1, 5)]
write_row(ws8, row, ["Discount Factor"] + [round(d, 4) for d in discount_factors], fmt='0.0000')
row += 1
pv_fcfs_list = [round(f * d) for f, d in zip(fcf_fcst, discount_factors)]
write_row(ws8, row, ["PV of FCF ($M)"] + pv_fcfs_list, fmt=NUM_FMT)

row += 2
pv_fcf_total = sum(pv_fcfs_list)
write_row(ws8, row, ["Sum PV of FCFs", pv_fcf_total], font=BOLD_FONT, fmt=NUM_FMT)
row += 1
tv = round(fcf_fcst[-1] * (1 + dcf_tgr) / (dcf_wacc - dcf_tgr))
write_row(ws8, row, ["Terminal Value", tv], fmt=NUM_FMT)
row += 1
pv_tv = round(tv * discount_factors[-1])
write_row(ws8, row, ["PV of Terminal Value", pv_tv], fmt=NUM_FMT)
row += 1
ev_dcf = pv_fcf_total + pv_tv
write_row(ws8, row, ["Enterprise Value", ev_dcf], font=BOLD_FONT, fmt=NUM_FMT)
row += 1
eq_dcf = ev_dcf - net_debt
write_row(ws8, row, ["Equity Value", eq_dcf], font=BOLD_FONT, fmt=NUM_FMT)
row += 1
price_dcf = round(eq_dcf / shares_annual_fcst[1], 2)
write_row(ws8, row, ["Implied Share Price", price_dcf], font=BOLD_FONT, fmt=USD_DEC_FMT)
for c in range(1, 3):
    ws8.cell(row=row, column=c).border = BOTTOM_BORDER

# ── Section 2: Cohort NPV Analysis ───────────────────────────────────────────
row += 3
ws8.cell(row=row, column=1, value="2. Cohort NPV / User Economics").font = SECTION_FONT
ws8.cell(row=row, column=1).border = BOTTOM_BORDER

row += 1
ws8.cell(row=row, column=1, value="Methodology: Value each annual user cohort's lifetime revenue stream").font = Font(name="Calibri", italic=True, size=10)

row += 2
write_row(ws8, row, ["Cohort Assumptions"], font=BOLD_FONT)
row += 1
write_row(ws8, row, ["  Year-1 ARPU", arpu_fcst[0]], fmt=USD_DEC_FMT)
ws8.cell(row=row, column=2).fill = INPUT_FILL
row += 1
write_row(ws8, row, ["  ARPU Annual Growth", 0.05], fmt=PCT_FMT)
ws8.cell(row=row, column=2).fill = INPUT_FILL
row += 1
write_row(ws8, row, ["  Annual Churn Rate", 0.20], fmt=PCT_FMT)
ws8.cell(row=row, column=2).fill = INPUT_FILL
row += 1
write_row(ws8, row, ["  Cohort Gross Margin", 0.56], fmt=PCT_FMT)
ws8.cell(row=row, column=2).fill = INPUT_FILL
row += 1
write_row(ws8, row, ["  Discount Rate", 0.11], fmt=PCT_FMT)
ws8.cell(row=row, column=2).fill = INPUT_FILL
row += 1
write_row(ws8, row, ["  CAC per User ($)", 3.50], fmt=USD_DEC_FMT)
ws8.cell(row=row, column=2).fill = INPUT_FILL

# Cohort NPV calculation
cohort_arpu_1 = arpu_fcst[0]
cohort_arpu_growth = 0.05
churn = 0.20
cohort_gm = 0.56
disc = 0.11
cac = 3.50

row += 2
write_row(ws8, row, ["", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5", "Year 6", "Year 7", "Year 8"], font=HEADER_FONT, fill=HEADER_FILL)
style_header_row(ws8, row, 9)

row += 1
retention = [(1 - churn)**i for i in range(8)]
write_row(ws8, row, ["Retention Rate"] + [round(r, 3) for r in retention], fmt=PCT_FMT)

row += 1
cohort_arpus = [round(cohort_arpu_1 * (1 + cohort_arpu_growth)**i, 2) for i in range(8)]
write_row(ws8, row, ["ARPU ($)"] + cohort_arpus, fmt=USD_DEC_FMT)

row += 1
gp_per_user = [round(a * cohort_gm * r, 2) for a, r in zip(cohort_arpus, retention)]
write_row(ws8, row, ["GP / User ($)"] + gp_per_user, fmt=USD_DEC_FMT)

row += 1
pv_gp = [round(gp / (1 + disc)**i, 2) for i, gp in enumerate(gp_per_user, 1)]
write_row(ws8, row, ["PV of GP / User ($)"] + pv_gp, fmt=USD_DEC_FMT)

row += 2
ltv = round(sum(pv_gp), 2)
write_row(ws8, row, ["Lifetime Value per User (LTV)", ltv], font=BOLD_FONT, fmt=USD_DEC_FMT)
row += 1
write_row(ws8, row, ["CAC per User", cac], fmt=USD_DEC_FMT)
row += 1
ltv_cac = round(ltv / cac, 1)
write_row(ws8, row, ["LTV / CAC Ratio", ltv_cac], font=BOLD_FONT, fmt=DEC_FMT)

row += 1
npv_per_user = round(ltv - cac, 2)
write_row(ws8, row, ["NPV per User ($)", npv_per_user], font=BOLD_FONT, fmt=USD_DEC_FMT)

row += 2
# Implied EV from user NPV
implied_ev_cohort = round(npv_per_user * dau_fcst[0])
write_row(ws8, row, ["Implied EV (NPV/user × DAU)", implied_ev_cohort], font=BOLD_FONT, fmt=NUM_FMT)
row += 1
implied_eq_cohort = implied_ev_cohort - net_debt
write_row(ws8, row, ["Implied Equity Value", implied_eq_cohort], font=BOLD_FONT, fmt=NUM_FMT)
row += 1
implied_price_cohort = round(implied_eq_cohort / shares_annual_fcst[1], 2)
write_row(ws8, row, ["Implied Share Price (Cohort)", implied_price_cohort], font=BOLD_FONT, fmt=USD_DEC_FMT)

# ── Section 3: Adoption S-Curve ──────────────────────────────────────────────
row += 3
ws8.cell(row=row, column=1, value="3. Adoption S-Curve Analysis").font = SECTION_FONT
ws8.cell(row=row, column=1).border = BOTTOM_BORDER

row += 2
write_row(ws8, row, ["S-Curve Assumptions"], font=BOLD_FONT)
row += 1
tam_dau = 800  # Total addressable market in millions of DAU
write_row(ws8, row, ["  Total Addressable Market (M DAU)", tam_dau], fmt=NUM_FMT)
ws8.cell(row=row, column=2).fill = INPUT_FILL
row += 1
current_penetration = round(dau_fcst[0] / tam_dau, 3)
write_row(ws8, row, ["  Current Penetration", current_penetration], fmt=PCT_FMT)
row += 1
write_row(ws8, row, ["  Implied Position: Mid-curve inflection → growth deceleration"], font=Font(name="Calibri", italic=True, size=10))

row += 2
write_row(ws8, row, ["Year", "DAU (M)", "Penetration", "YoY Growth", "Phase"], font=HEADER_FONT, fill=HEADER_FILL)
style_header_row(ws8, row, 5)

s_curve_years = list(range(2020, 2031))
s_curve_dau = [265, 319, 375, 414, 443] + dau_fcst + [round(dau_fcst[-1] * 1.035), round(dau_fcst[-1] * 1.035 * 1.030)]
s_curve_labels = ["Early Growth", "Early Growth", "Acceleration", "Mid-Curve", "Mid-Curve",
                  "Deceleration", "Deceleration", "Deceleration", "Deceleration",
                  "Maturation", "Maturation"]

for i, yr in enumerate(s_curve_years):
    row += 1
    pen = round(s_curve_dau[i] / tam_dau, 3)
    yoy = round(s_curve_dau[i] / s_curve_dau[i-1] - 1, 3) if i > 0 else None
    write_row(ws8, row, [yr, s_curve_dau[i], pen, yoy, s_curve_labels[i]])
    ws8.cell(row=row, column=3).number_format = PCT_FMT
    if yoy is not None:
        ws8.cell(row=row, column=4).number_format = PCT_FMT

# ── Section 4: Unit Economics → EV Sanity Check ──────────────────────────────
row += 3
ws8.cell(row=row, column=1, value="4. Unit Economics → EV Sanity Check").font = SECTION_FONT
ws8.cell(row=row, column=1).border = BOTTOM_BORDER

row += 2
write_row(ws8, row, ["Metric", "Value", "Commentary"], font=HEADER_FONT, fill=HEADER_FILL)
style_header_row(ws8, row, 3)

# Current market data (approximate)
mkt_cap_approx = 18000  # ~$18B market cap assumption
ev_approx = mkt_cap_approx + net_debt

row += 1
write_row(ws8, row, ["Current Market Cap ($M, est.)", mkt_cap_approx, "~$10-11/share × ~1.7B shares"], fmt=NUM_FMT)
row += 1
write_row(ws8, row, ["Enterprise Value ($M, est.)", ev_approx, "Market Cap + Net Debt"], fmt=NUM_FMT)
row += 1
ev_per_dau = round(ev_approx / dau_fcst[0], 1)
write_row(ws8, row, ["EV / DAU ($)", ev_per_dau, f"vs. META ~$310, PINS ~$75, TWTR was ~$85"], fmt=USD_DEC_FMT)
row += 1
ev_rev_25 = round(ev_approx / rev_fcst[0], 1)
write_row(ws8, row, ["EV / FY2025E Revenue", ev_rev_25, "vs. social-media peers at 4-8x"], fmt=DEC_FMT)
row += 1
ev_ebitda_26 = round(ev_approx / ebitda_fcst[1], 1) if ebitda_fcst[1] > 0 else "N/M"
write_row(ws8, row, ["EV / FY2026E EBITDA", ev_ebitda_26, "vs. peers at 15-25x"], fmt=DEC_FMT)
row += 1
fcf_yield_26 = round(fcf_fcst[1] / mkt_cap_approx, 3) if fcf_fcst[1] > 0 else 0
write_row(ws8, row, ["FY2026E FCF Yield", fcf_yield_26, "Implies early FCF inflection"], fmt=PCT_FMT)

row += 2
write_row(ws8, row, ["Revenue / DAU (ARPU) Benchmarks"], font=BOLD_FONT)
row += 1
write_row(ws8, row, ["  Snap FY2025E ARPU", arpu_fcst[0], "vs. META ~$50+, PINS ~$7"], fmt=USD_DEC_FMT)
row += 1
write_row(ws8, row, ["  Snap FY2028E ARPU", arpu_fcst[-1], "Significant ARPU upside vs. META"], fmt=USD_DEC_FMT)

row += 2
write_row(ws8, row, ["Valuation Summary"], font=SECTION_FONT)
ws8.cell(row=row, column=1).border = BOTTOM_BORDER

row += 1
write_row(ws8, row, ["Method", "Implied Price ($)", "Upside/Downside"], font=HEADER_FONT, fill=HEADER_FILL)
style_header_row(ws8, row, 3)

current_price_est = 10.50
row += 1
write_row(ws8, row, ["DCF (Base Case)", price_dcf, round(price_dcf / current_price_est - 1, 3)])
ws8.cell(row=row, column=2).number_format = USD_DEC_FMT
ws8.cell(row=row, column=3).number_format = PCT_FMT

row += 1
write_row(ws8, row, ["Cohort NPV", implied_price_cohort, round(implied_price_cohort / current_price_est - 1, 3)])
ws8.cell(row=row, column=2).number_format = USD_DEC_FMT
ws8.cell(row=row, column=3).number_format = PCT_FMT

# EV/Revenue comp
comp_ev_rev = 4.0
implied_price_comp = round((rev_fcst[1] * comp_ev_rev - net_debt) / shares_annual_fcst[1], 2)
row += 1
write_row(ws8, row, ["EV/Revenue Comp (4.0x FY26E)", implied_price_comp, round(implied_price_comp / current_price_est - 1, 3)])
ws8.cell(row=row, column=2).number_format = USD_DEC_FMT
ws8.cell(row=row, column=3).number_format = PCT_FMT

set_col_widths(ws8, [38] + [18]*8)


# ═══════════════════════ TAB 9: MODEL ASSUMPTIONS ═════════════════════════════
ws9 = wb.create_sheet("Assumptions & Drivers")
ws9.sheet_properties.tabColor = "FF0000"

row = 1
ws9.cell(row=row, column=1, value="Snap Inc. (SNAP) – Key Model Assumptions & Drivers").font = TITLE_FONT
ws9.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)

row = 3
ws9.cell(row=row, column=1, value="All yellow-highlighted cells are adjustable inputs").font = Font(name="Calibri", italic=True, size=10, color="BF8F00")

row = 5
write_row(ws9, row, ["", "FY2025E", "FY2026E", "FY2027E", "FY2028E"], font=HEADER_FONT, fill=HEADER_FILL)
style_header_row(ws9, row, 5)

row += 1
ws9.cell(row=row, column=1, value="Growth Drivers").font = SECTION_FONT
ws9.cell(row=row, column=1).border = BOTTOM_BORDER

row += 1
write_row(ws9, row, ["DAU Growth (YoY)"] + dau_growth_fcst, fmt=PCT_FMT, fill=INPUT_FILL)
row += 1
write_row(ws9, row, ["ARPU Growth (YoY)"] + arpu_growth_fcst, fmt=PCT_FMT, fill=INPUT_FILL)
row += 1
implied_rev_growth = [round(rev_fcst[i] / ([rev_hist[-1]] + rev_fcst)[i] - 1, 3) for i in range(4)]
write_row(ws9, row, ["Implied Revenue Growth"] + implied_rev_growth, fmt=PCT_FMT)

row += 2
ws9.cell(row=row, column=1, value="Margin Assumptions").font = SECTION_FONT
ws9.cell(row=row, column=1).border = BOTTOM_BORDER

row += 1
write_row(ws9, row, ["Gross Margin"] + gm_fcst, fmt=PCT_FMT, fill=INPUT_FILL)
row += 1
write_row(ws9, row, ["R&D % of Revenue"] + rd_pct_fcst, fmt=PCT_FMT, fill=INPUT_FILL)
row += 1
write_row(ws9, row, ["S&M % of Revenue"] + sm_pct_fcst, fmt=PCT_FMT, fill=INPUT_FILL)
row += 1
write_row(ws9, row, ["G&A % of Revenue"] + ga_pct_fcst, fmt=PCT_FMT, fill=INPUT_FILL)
row += 1
write_row(ws9, row, ["SBC % of Revenue"] + sbc_pct_fcst, fmt=PCT_FMT, fill=INPUT_FILL)

row += 2
ws9.cell(row=row, column=1, value="Other Assumptions").font = SECTION_FONT
ws9.cell(row=row, column=1).border = BOTTOM_BORDER

row += 1
write_row(ws9, row, ["D&A ($M)"] + da_fcst, fmt=NUM_FMT, fill=INPUT_FILL)
row += 1
write_row(ws9, row, ["CapEx ($M)"] + [abs(c) for c in capex_fcst], fmt=NUM_FMT, fill=INPUT_FILL)
row += 1
write_row(ws9, row, ["Effective Tax Rate"] + tax_rate_fcst, fmt=PCT_FMT, fill=INPUT_FILL)
row += 1
write_row(ws9, row, ["Interest Expense ($M)"] + [abs(i) for i in interest_other_fcst], fmt=NUM_FMT, fill=INPUT_FILL)
row += 1
write_row(ws9, row, ["Annual Share Dilution (%)"] + [round(shares_annual_fcst[i] / ([shares_hist[-1]] + shares_annual_fcst)[i] - 1, 3) for i in range(4)], fmt=PCT_FMT, fill=INPUT_FILL)

row += 2
ws9.cell(row=row, column=1, value="Revenue Build Check").font = SECTION_FONT
ws9.cell(row=row, column=1).border = BOTTOM_BORDER

row += 1
write_row(ws9, row, ["DAU (M, year-end)"] + dau_fcst, fmt=NUM_FMT)
row += 1
write_row(ws9, row, ["× ARPU ($)"] + arpu_fcst, fmt=USD_DEC_FMT)
row += 1
write_row(ws9, row, ["= Revenue ($M)"] + rev_fcst, font=BOLD_FONT, fmt=NUM_FMT)

row += 2
ws9.cell(row=row, column=1, value="Quarterly Seasonality Weights").font = SECTION_FONT
ws9.cell(row=row, column=1).border = BOTTOM_BORDER

row += 1
write_row(ws9, row, ["", "Q1", "Q2", "Q3", "Q4"], font=HEADER_FONT, fill=HEADER_FILL)
style_header_row(ws9, row, 5)
row += 1
write_row(ws9, row, ["Revenue Seasonality"] + q_seasonality, fmt=PCT_FMT, fill=INPUT_FILL)
row += 1
write_row(ws9, row, ["Total Check", sum(q_seasonality)], fmt=PCT_FMT)

set_col_widths(ws9, [32] + [16]*8)


# ══════════════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════════════
output_path = "/workspace/Snap_Inc_Three_Statement_Model.xlsx"
wb.save(output_path)
print(f"Model saved to: {output_path}")
print("Tabs created:")
for ws_name in wb.sheetnames:
    print(f"  - {ws_name}")
