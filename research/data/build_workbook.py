"""Builds the Excel workbook aggregating US government spending to quantum / PQC
beneficiaries, plus a tab detailing the executive orders and technologies touched.

Run: python3 research/data/build_workbook.py
Output: research/data/us-spending-beneficiaries.xlsx
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent / "us-spending-beneficiaries.xlsx"

# ---- shared styles -------------------------------------------------------
NAVY = "1F3864"
BLUE = "2E5496"
LIGHT = "D9E1F2"
ALT = "F2F6FC"
AMBER = "FFF2CC"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SUB_FILL = PatternFill("solid", fgColor=BLUE)
ALT_FILL = PatternFill("solid", fgColor=ALT)
NOTE_FILL = PatternFill("solid", fgColor=AMBER)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color=NAVY, size=15)
SUB_FONT = Font(bold=True, color="FFFFFF", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_C = Alignment(wrap_text=True, vertical="top", horizontal="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_C
        cell.border = BORDER


def write_table(ws, start_row, headers, rows, widths=None, money_cols=None):
    money_cols = money_cols or set()
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=h)
    style_header(ws, start_row, len(headers))
    for i, rowdata in enumerate(rows, start=start_row + 1):
        for j, val in enumerate(rowdata, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = WRAP
            cell.border = BORDER
            if (i - start_row) % 2 == 0:
                cell.fill = ALT_FILL
            if j in money_cols and isinstance(val, (int, float)):
                cell.number_format = '#,##0'
    if widths:
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return start_row + 1 + len(rows)


def title(ws, text, sub=None):
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT
    row = 2
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(italic=True, color="595959", size=10)
        row = 3
    return row + 1


wb = Workbook()

# ========================================================================
# TAB 1 — Overview
# ========================================================================
ws = wb.active
ws.title = "Overview"
r = title(
    ws,
    "US Government Spending → Quantum & PQC/Cybersecurity Beneficiaries",
    "Companion workbook to the equity-research note. Compiled June 23, 2026. Informational only — not investment advice.",
)
overview_rows = [
    ["Tab", "Contents"],
    ["Executive Orders", "EO 14411 & EO 14409 (Jun 22, 2026): scope, mechanisms, deadlines, and the specific technologies each order touches"],
    ["Headline Programs", "Top-level federal programs/vehicles driving spend (CHIPS LOIs, OMB/ONCD PQC est., DARPA QBI, CISA FY26)"],
    ["CHIPS LOIs", "$2.013B Commerce/NIST quantum letters of intent (May 21, 2026), with ticker mapping & equity condition"],
    ["DARPA QBI", "Quantum Benchmarking Initiative Stage B cohort (Nov 6, 2025) + US2QC final phase"],
    ["Company Contracts", "Company-level disclosed federal contracts/awards (IonQ, Rigetti, D-Wave, SandboxAQ, etc.)"],
    ["PQC Market Sizing", "Demand-side estimates for PQC migration and procurement gating"],
    ["Beneficiaries (data)", "Flat, machine-readable line-item table of all disclosed flows"],
]
hdr = overview_rows[0]
body = overview_rows[1:]
nr = write_table(ws, r, hdr, body, widths=[24, 110])
r = nr + 1
ws.cell(row=r, column=1, value="CAVEAT: Headline values frequently exceed obligated (funded) dollars. 'Up to' / 'ceiling' / 'IDIQ' figures are not committed spend. Reconcile against 10-K/10-Q, USASpending.gov and FPDS.")
ws.cell(row=r, column=1).fill = NOTE_FILL
ws.cell(row=r, column=1).alignment = WRAP
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

# ========================================================================
# TAB 2 — Executive Orders & Technologies Touched  (requested)
# ========================================================================
ws = wb.create_sheet("Executive Orders")
r = title(
    ws,
    "Executive Orders Issued & Technologies Touched",
    "Signed June 22, 2026 at the White House. Two companion orders: one demand-pull (quantum innovation), one compliance-mandate (PQC).",
)
eo_headers = [
    "EO number", "Title", "Type", "Sponsoring/lead agencies",
    "Core mechanism", "Key deadlines / milestones", "Technologies touched", "Primary beneficiary categories",
]
eo_rows = [
    [
        "EO 14411",
        "Ushering In the Next Frontier of Quantum Innovation",
        "Demand-pull / R&D acceleration ('offense')",
        "DOE (lead), Commerce, Dept. of War (DoD), DNI/IC, NASA, GSA",
        "Whole-of-government effort to field a science-grade quantum computer (QC-ADDS); advance market commitments (AMCs); expanded counterintelligence (QCPT) for the QIST supply chain",
        "DOE technical specs +90d; DOE delivery/partnership models +180d; science-grade quantum computer target ~2028; next-gen quantum sensors fielded by Sep 30, 2028",
        "Quantum computing (all modalities: trapped-ion, superconducting, neutral-atom, silicon-spin, photonic); quantum sensing; quantum networking/communications; quantum supply chain & workforce",
        "Quantum hardware OEMs, foundries, quantum sensor & networking vendors, national labs, universities",
    ],
    [
        "EO 14409",
        "Securing the Nation Against Advanced Cryptographic Attacks",
        "Compliance mandate / cybersecurity ('defense')",
        "OMB & National Cyber Director (lead), Commerce/NIST, NSA, DHS/CISA, State, FAR Council, DoD, NASA, GSA",
        "Enforceable PQC migration deadlines for federal HVAs/high-impact systems + FAR rule flowing the mandate to covered contractors; NIST/Commerce migration pilot; assistance to critical-infrastructure operators",
        "Agency PQC migration leads +30d; OMB binding guidance & HVA inventories +90d; FAR proposed rule +180d; NIST/Commerce pilot done by Dec 31, 2027; key establishment by Dec 31, 2030; digital signatures by Dec 31, 2031; contractor compliance by end-2030",
        "Post-quantum cryptography (ML-KEM/FIPS 203; ML-DSA/FIPS 204; SLH-DSA/FIPS 205); crypto-agility middleware; cryptographic discovery & inventory (CBOM); PKI / certificate lifecycle; HSMs; FIPS 140-3 validated libraries; network/SASE/VPN/firewall PQC",
        "PQC software vendors, PKI/HSM providers, crypto-discovery firms, SASE/firewall/network security platforms, FIPS validation labs, systems integrators",
    ],
]
nr = write_table(
    ws, r, eo_headers, eo_rows,
    widths=[11, 34, 26, 30, 40, 42, 46, 34],
)
# bump row heights for the two data rows so wrapped text is readable
ws.row_dimensions[r + 1].height = 150
ws.row_dimensions[r + 2].height = 150

r = nr + 1
# Technology -> order mapping mini-table
ws.cell(row=r, column=1, value="Technology-to-order quick map")
ws.cell(row=r, column=1).font = Font(bold=True, color=NAVY, size=12)
r += 1
tech_headers = ["Technology / capability", "EO 14411 (Quantum Innovation)", "EO 14409 (PQC)", "Theme"]
tech_rows = [
    ["Quantum computing hardware", "Direct (QC-ADDS, AMCs)", "Indirect (threat driver)", "Quantum"],
    ["Quantum sensing", "Direct (2028 sensor target)", "—", "Quantum"],
    ["Quantum networking / communications", "Direct", "Indirect (QKD context)", "Quantum"],
    ["Quantum supply chain / foundries / workforce", "Direct (counterintelligence, onshoring)", "—", "Quantum"],
    ["Post-quantum cryptography (PQC algorithms)", "References (migration awareness)", "Direct (FIPS 203/204/205)", "Cybersecurity"],
    ["Crypto-agility middleware", "—", "Direct", "Cybersecurity"],
    ["Cryptographic discovery & inventory (CBOM)", "—", "Direct (+30/+90d inventories)", "Cybersecurity"],
    ["PKI / certificate lifecycle / HSM", "—", "Direct (signatures by 2031)", "Cybersecurity"],
    ["SASE / VPN / firewall / network PQC", "—", "Direct (via FAR procurement)", "Cybersecurity"],
    ["FIPS 140-3 validation services", "—", "Direct (procurement rule)", "Cybersecurity"],
]
write_table(ws, r, tech_headers, tech_rows, widths=[42, 34, 32, 16])

# ========================================================================
# TAB 3 — Headline Programs
# ========================================================================
ws = wb.create_sheet("Headline Programs")
r = title(ws, "Headline Programs Driving the Spend")
hp_headers = ["Program / vehicle", "Sponsor", "Size (USD)", "Size note", "Date", "Mechanism", "Notes"]
hp_rows = [
    ["CHIPS & Science quantum LOIs", "Commerce / NIST", 2013000000, "$2.013B across 9 cos.", "2026-05-21", "Grants + minority non-controlling equity stake", "Largest single federal intervention in quantum to date"],
    ["EO 14411 (Quantum Innovation)", "White House", None, "Directive (no $ attached)", "2026-06-22", "QC-ADDS, AMCs, 2028 targets", "Demand-pull"],
    ["EO 14409 (PQC)", "White House", None, "Directive (no $ attached)", "2026-06-22", "PQC deadlines + FAR mandate", "Market-wide procurement requirement"],
    ["Government-wide PQC migration", "OMB / ONCD", 7100000000, "~$7.1B (2025-2035, 2024$)", "2024 est.", "Annual agency inventories + budget justifications", "Excludes classified / NSS"],
    ["DARPA QBI Stage B", "DARPA (DoW)", 5000000, "$1M-$5M per award; 11 cos.", "2025-11-06", "Staged R&D validation toward utility-scale by 2033", "Not a procurement competition"],
    ["CISA cybersecurity (FY26)", "DHS / CISA", 1367400000, "~$1.367B for FCEB protection", "FY2026", "Appropriated federal civilian cyber defense", "Broader than PQC"],
]
write_table(ws, r, hp_headers, hp_rows, widths=[34, 18, 16, 28, 12, 40, 36], money_cols={3})

# ========================================================================
# TAB 4 — CHIPS LOIs
# ========================================================================
ws = wb.create_sheet("CHIPS LOIs")
r = title(ws, "CHIPS & Science Act — Quantum LOIs ($2.013B, May 21, 2026)",
          "Every recipient grants Commerce a minority, non-controlling equity stake as a condition of funding.")
chips_headers = ["Company", "Ticker", "Ownership", "Planned funding (USD)", "Value type", "Modality / purpose", "Investability"]
chips_rows = [
    ["IBM", "IBM", "public", 1000000000, "planned", "Domestic quantum foundry", "Public mega-cap (immaterial to EPS)"],
    ["GlobalFoundries", "GFS", "public", 375000000, "planned", "Secure multi-modality quantum foundry (~1% govt equity)", "Public"],
    ["D-Wave Quantum", "QBTS", "public", 100000000, "planned", "Annealing + gate-model superconducting (issues $100M stock to Commerce)", "Public pure-play"],
    ["Rigetti Computing", "RGTI", "public", 100000000, "ceiling (up to)", "Next-gen superconducting", "Public pure-play"],
    ["Infleqtion", "INFQ", "public", 100000000, "planned", "Neutral-atom systems integration", "Public (SPAC Feb 2026)"],
    ["Quantinuum", "HON (IPO filed)", "majority Honeywell", 100000000, "planned", "Fault-tolerant trapped-ion (QCCD)", "IPO pending"],
    ["PsiQuantum", "—", "private", 100000000, "planned", "Photonic fault-tolerant", "Not investable"],
    ["Atom Computing", "—", "private", 100000000, "planned", "Neutral-atom", "Not investable"],
    ["Diraq", "—", "private", 38000000, "ceiling (up to)", "Silicon-spin (CMOS)", "Not investable"],
    ["TOTAL", "", "", 2013000000, "", "", ""],
]
endrow = write_table(ws, r, chips_headers, chips_rows, widths=[20, 16, 18, 20, 16, 44, 30], money_cols={4})
# bold total row
for c in range(1, 8):
    ws.cell(row=endrow - 1, column=c).font = Font(bold=True)
r = endrow + 1
ws.cell(row=r, column=1, value="Equity-day reaction (2026-05-21): INFQ +31.4%, QBTS +33%, RGTI +30.6%; non-recipients IONQ +12.3%, QUBT +19%.")
ws.cell(row=r, column=1).fill = NOTE_FILL
ws.cell(row=r, column=1).alignment = WRAP
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)

# ========================================================================
# TAB 5 — DARPA QBI
# ========================================================================
ws = wb.create_sheet("DARPA QBI")
r = title(ws, "DARPA Quantum Benchmarking Initiative — Stage B (Nov 6, 2025)",
          "11 companies advanced; awards $1M-$5M each. Goal: verify utility-scale quantum by 2033.")
qbi_headers = ["Company", "Ticker", "In CHIPS LOI?", "Modality", "Country"]
qbi_rows = [
    ["IBM", "IBM", "Yes ($1B foundry)", "Modular superconducting", "USA"],
    ["IonQ", "IONQ", "No", "Trapped-ion", "USA"],
    ["Quantinuum", "HON (IPO filed)", "Yes ($100M)", "Trapped-ion QCCD", "USA"],
    ["Atom Computing", "private", "Yes ($100M)", "Neutral-atom", "USA"],
    ["Diraq", "private", "Yes ($38M)", "Silicon-spin", "Australia"],
    ["QuEra Computing", "private", "No", "Neutral-atom", "USA"],
    ["Nord Quantique", "private", "No", "Superconducting (bosonic)", "Canada"],
    ["Photonic Inc.", "private", "No", "Optically-linked silicon spin", "Canada"],
    ["Quantum Motion", "private", "No", "Silicon MOS spin", "UK"],
    ["Silicon Quantum Computing", "private", "No", "Precision donor atom in Si", "Australia"],
    ["Xanadu", "private", "No", "Photonic (continuous-variable)", "Canada"],
]
endrow = write_table(ws, r, qbi_headers, qbi_rows, widths=[26, 16, 18, 30, 12])
r = endrow + 1
ws.cell(row=r, column=1, value="US2QC final phase (~QBI Stage C): Microsoft (MSFT), PsiQuantum (private). Stage A non-advancers: Rigetti (RGTI), HPE (HPE), Atlantic Quantum, Oxford Ionics (acquired by IonQ Sep 2025, ~$1.075B).")
ws.cell(row=r, column=1).fill = NOTE_FILL
ws.cell(row=r, column=1).alignment = WRAP
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

# ========================================================================
# TAB 6 — Company Contracts
# ========================================================================
ws = wb.create_sheet("Company Contracts")
r = title(ws, "Company-Level Disclosed Federal Contracts / Awards")
cc_headers = ["Company", "Ticker", "Award / vehicle", "Sponsor", "Headline value (USD)", "Value type", "Obligated / note", "Date", "Theme"]
cc_rows = [
    ["IonQ", "IONQ", "Quantum networking contract", "AFRL", 54500000, "ceiling", "Only ~$11.99M obligated; unfunded FY2026", "2024-09-25", "Quantum"],
    ["IonQ", "IONQ", "AFRL FY22-24 obligations", "AFRL", 51000000, "obligated", "Earmark-driven cumulative", "2024-12-31", "Quantum"],
    ["IonQ", "IONQ", "SHIELD IDIQ (vendor eligibility)", "MDA", 151000000000, "program ceiling (IDIQ)", "Not a direct award", "2026", "Quantum"],
    ["Rigetti", "RGTI", "Superconducting networking (w/ QphoX)", "AFRL", 5800000, "contract (3yr)", "", "2025-09-18", "Quantum"],
    ["Rigetti", "RGTI", "ABAA chip-fab consortium", "AFOSR", 5480000, "award", "Incl. LLNL", "2025-04-28", "Quantum"],
    ["D-Wave", "QBTS", "CHIPS LOI", "Commerce/NIST", 100000000, "planned", "Issues $100M stock to Commerce", "2026-05-21", "Quantum"],
    ["SandboxAQ", "private", "DoW CIO AQtive Guard (5-yr)", "Dept. of War CIO", None, "undisclosed", "ACDI / PQC discovery", "2025-12-10", "Cybersecurity"],
    ["SandboxAQ", "private", "Quantum navigation SBIR", "US Air Force", 1200000, "contract (SBIR)", "Magnetic navigation prototype", "2022", "Cybersecurity"],
    ["Cloudflare", "NET", "PQC SASE platform", "Multiple federal", None, "productized", "Hybrid ML-KEM TLS/IPsec", "2026", "Cybersecurity"],
    ["Palo Alto Networks", "PANW", "NGFW PQC ciphersuites", "Multiple federal", None, "productized", "7+ PQC ciphersuites", "2026", "Cybersecurity"],
    ["SEALSQ", "LAES", "PQC semiconductor", "Multiple", None, "productized", "Secure-element / PQC chips", "2026", "Cybersecurity"],
]
write_table(ws, r, cc_headers, cc_rows, widths=[18, 14, 34, 18, 18, 20, 32, 12, 14], money_cols={5})

# ========================================================================
# TAB 7 — PQC Market Sizing
# ========================================================================
ws = wb.create_sheet("PQC Market Sizing")
r = title(ws, "PQC Migration Market Sizing (Demand Side)")
ms_headers = ["Segment", "Estimate (USD)", "Estimate note", "Source basis"]
ms_rows = [
    ["Government-wide migration (FCEB, excl. NSS)", 7100000000, "~$7.1B 2025-2035 (2024$)", "OMB/ONCD annual agency cost rollup"],
    ["Small agency (each)", 20000000, "$5M-$20M each", "CISA/agency budget justifications"],
    ["Medium agency (each)", 200000000, "$50M-$200M each", "CISA/agency budget justifications"],
    ["Large agency (DoD, VA, Treasury, each)", 2000000000, "$500M-$2B+ each", "CISA/agency budget justifications"],
    ["Classified / NSS (DoD + IC)", None, "Separate, undisclosed", "NSM-10 / CNSA 2.0 (faster track)"],
]
endrow = write_table(ws, r, ms_headers, ms_rows, widths=[42, 18, 24, 40], money_cols={2})
r = endrow + 1
ws.cell(row=r, column=1, value="Procurement gating already live: CISA published PQC-capable product categories effective 2026-01-23; EO 14409 FAR rule extends to all covered contractors by end-2030.")
ws.cell(row=r, column=1).fill = NOTE_FILL
ws.cell(row=r, column=1).alignment = WRAP
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

# ========================================================================
# TAB 8 — Beneficiaries (flat data, from CSV)
# ========================================================================
ws = wb.create_sheet("Beneficiaries (data)")
import csv as _csv

csv_path = Path(__file__).parent / "us-spending-beneficiaries.csv"
with open(csv_path, newline="") as f:
    reader = list(_csv.reader(f))
flat_headers = reader[0]
flat_rows = []
for row in reader[1:]:
    new = []
    for j, v in enumerate(row):
        if flat_headers[j] == "headline_value_usd" and v.strip():
            try:
                new.append(int(v))
            except ValueError:
                new.append(v)
        else:
            new.append(v)
    flat_rows.append(new)
val_col = flat_headers.index("headline_value_usd") + 1
write_table(ws, 1, flat_headers, flat_rows,
            widths=[26, 16, 16, 38, 20, 20, 22, 36, 14, 24, 16], money_cols={val_col})

wb.save(OUT)
print(f"Wrote {OUT} with {len(wb.sheetnames)} tabs: {wb.sheetnames}")
