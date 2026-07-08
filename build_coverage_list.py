"""Generate the Asia-listed Semiconductor & Data Center coverage universe spreadsheet.

Scope: names listed in mainland China (SSE/SZSE/STAR), Hong Kong, Taiwan, South Korea,
Japan, Singapore and other parts of Asia. No US- or Europe-listed names.

Produces:
  - semi_datacenter_coverage.xlsx  (formatted, multi-sheet)
  - semi_datacenter_coverage.csv   (flat, portable)

Market caps are approximate and rounded, as of ~Jul 8, 2026. They are meant for
coverage-scoping purposes only and are NOT investment advice. Refresh from a live
data source before use.
"""

import csv
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

AS_OF = "2026-07-08"

# Columns:
# name, ticker, exchange, subsector, listing region, approx market cap ($B),
# recent IPO/listing, IPO date, notes
COVERAGE = [
    # =========================== MAINLAND CHINA ===========================
    ("CXMT (ChangXin Memory)", "STAR (pending)", "SSE STAR (pending)", "Semis - Memory (DRAM)", "Mainland China", 150.0, "Yes", "2026-H2 (pending)",
     "China's largest DRAM maker; ~$4.3B STAR IPO registration approved Jun-2026 (2nd-largest STAR listing ever); HBM3 ramp; ~8% global DRAM share."),
    ("SMIC", "688981.SS / 0981.HK", "SSE STAR / HKEX", "Semis - Foundry", "Mainland China", 100.0, "No", "2020-07 (STAR) / 2004-03 (HK)",
     "China's largest foundry; advanced-node localization flagship amid export controls."),
    ("Hua Hong Semiconductor", "688347.SS / 1347.HK", "SSE STAR / HKEX", "Semis - Foundry", "Mainland China", 30.0, "No", "2023-08 (STAR) / 2014-10 (HK)",
     "China #2 foundry; specialty/mature nodes; power & analog platforms."),
    ("NAURA Technology", "002371.SZ", "SZSE", "Semis - Equipment (WFE)", "Mainland China", 55.0, "No", "2010-06",
     "China's largest domestic WFE maker; etch/deposition/thermal; localization play."),
    ("AMEC (Advanced Micro-Fabrication)", "688012.SS", "SSE STAR", "Semis - Equipment (WFE)", "Mainland China", 30.0, "No", "2019-07",
     "China etch equipment leader; MOCVD; leading-edge localization."),
    ("Cambricon Technologies", "688256.SS", "SSE STAR", "Semis - AI Accelerators", "Mainland China", 100.0, "No", "2020-07",
     "China AI training/inference chip champion; revenue up >4,000% YoY in 1H25; NVIDIA substitution play."),
    ("Hygon Information Technology", "688041.SS", "SSE STAR", "Semis - CPU/DCU", "Mainland China", 95.0, "No", "2022-08",
     "Domestic x86-compatible server CPUs + DCU accelerators; merging with Sugon (server OEM)."),
    ("Moore Threads", "688795.SS", "SSE STAR", "Semis - GPU", "Mainland China", 43.0, "Yes", "2025-12",
     "'China's NVIDIA'; full-function GPUs (MUSA architecture); raised ~$1.1B, +400% on debut."),
    ("MetaX Integrated Circuits", "688802.SS", "SSE STAR", "Semis - GPU", "Mainland China", 30.0, "Yes", "2025-12",
     "Domestic GPU for AI training/inference (Xiyun C500); founded by ex-AMD execs; ~700% debut pop."),
    ("Montage Technology", "688008.SS", "SSE STAR", "Semis - Memory Interface", "Mainland China", 15.0, "No", "2019-07",
     "DDR5 memory interface / CXL chips for servers; AI memory bandwidth beneficiary."),
    ("GigaDevice Semiconductor", "603986.SS", "SSE", "Semis - Memory (NOR/NAND/DRAM)", "Mainland China", 12.0, "No", "2016-08",
     "NOR flash leader; expanding niche DRAM (CXMT ecosystem partner); MCUs."),
    ("Runze Technology", "300442.SZ", "SZSE ChiNext", "Data Center - Wholesale IDC", "Mainland China", 10.0, "No", "2015 (backdoor 2022)",
     "Pure-play wholesale IDC campuses (Beijing-Langfang, Yangtze Delta); AI/intelligent-computing demand."),
    ("Beijing Sinnet Technology", "300383.SZ", "SZSE ChiNext", "Data Center - IDC/Cloud", "Mainland China", 4.0, "No", "2014-01",
     "Beijing-area IDC + AWS China (Ningxia/Beijing) cloud operator."),
    ("Shanghai AtHub", "603881.SS", "SSE", "Data Center - Wholesale IDC", "Mainland China", 3.0, "No", "2017-01",
     "Wholesale datacenter developer; anchor customer Alibaba; East China AI capacity."),
    # =========================== HONG KONG ===========================
    ("GDS Holdings", "9698.HK (also GDS US ADR)", "HKEX (dual-primary)", "Data Center - Colocation", "Hong Kong", 6.7, "No", "2022-01 (HK) / 2016-11 (ADR)",
     "Carrier-neutral hyperscale colo leader in China; AI-driven wholesale demand; DayOne (international) angle."),
    ("SUNeVision Holdings", "1686.HK", "HKEX", "Data Center - Colocation", "Hong Kong", 2.7, "No", "2000-03",
     "Hong Kong's largest carrier-neutral datacenter operator (iAdvantage); SHKP-backed; MEGA campuses."),
    ("ASMPT", "0522.HK", "HKEX", "Semis - Equipment (Packaging)", "Hong Kong", 11.0, "No", "1989-06",
     "Advanced packaging equipment leader; TCB bonders for HBM/CoWoS; key AI memory enabler."),
    ("Biren Technology", "6082.HK", "HKEX", "Semis - GPU", "Hong Kong", 15.0, "Yes", "2026-01",
     "High-performance GPU designer (BR20X/BR30X); raised ~$717M in Jan-2026 HK IPO, +76% debut."),
    ("Horizon Robotics", "9660.HK", "HKEX", "Semis - Auto AI SoC", "Hong Kong", 8.5, "Yes", "2024-10",
     "Leading China ADAS/AD computing SoC (Journey series); largest HK tech IPO of 2024."),
    ("InnoScience", "2577.HK", "HKEX", "Semis - Power (GaN)", "Hong Kong", 8.0, "Yes", "2024-12",
     "World's largest GaN-on-Si IDM; AI server power / fast-charging adoption."),
    ("Black Sesame International", "2533.HK", "HKEX", "Semis - Auto AI SoC", "Hong Kong", 2.0, "Yes", "2024-08",
     "Auto-grade AI SoCs (Huashan/Wudang); first 18C-chapter HK listing."),
    # =========================== TAIWAN ===========================
    ("TSMC", "2330.TW", "TWSE", "Semis - Foundry", "Taiwan", 1200.0, "No", "1994-09",
     "World's leading-edge foundry; CoWoS/advanced packaging bottleneck for AI accelerators."),
    ("MediaTek", "2454.TW", "TWSE", "Semis - Fabless (SoC)", "Taiwan", 90.0, "No", "2001-07",
     "Mobile/connectivity SoC leader; custom ASIC push into AI datacenter (with Google/others)."),
    ("United Microelectronics (UMC)", "2303.TW", "TWSE", "Semis - Foundry", "Taiwan", 18.0, "No", "1985-07",
     "Mature/specialty foundry; 22/28nm; potential Intel collaboration optionality."),
    ("ASE Technology Holding", "3711.TW", "TWSE", "Semis - OSAT", "Taiwan", 25.0, "No", "2018-04 (holding)",
     "World's largest OSAT; advanced packaging/test capacity for AI chips."),
    ("Hon Hai Precision (Foxconn)", "2317.TW", "TWSE", "Data Center - AI Servers (ODM)", "Taiwan", 100.0, "No", "1991-06",
     "Largest AI server assembler (GB200/GB300 racks); NVIDIA's key systems partner."),
    ("Quanta Computer", "2382.TW", "TWSE", "Data Center - AI Servers (ODM)", "Taiwan", 35.0, "No", "1999-01",
     "Top AI server ODM for US hyperscalers; cloud infrastructure leverage."),
    ("Wiwynn", "6669.TW", "TWSE", "Data Center - AI Servers (ODM)", "Taiwan", 35.0, "No", "2019-03",
     "Hyperscale-focused server pure-play (Wistron spin-off); Meta/Microsoft exposure."),
    ("Delta Electronics", "2308.TW", "TWSE", "Data Center - Power/Cooling", "Taiwan", 60.0, "No", "1988-12",
     "Datacenter power supplies, busbars, liquid cooling; key AI rack power beneficiary."),
    # =========================== SOUTH KOREA ===========================
    ("Samsung Electronics", "005930.KS", "KRX", "Semis - Memory/Foundry (IDM)", "South Korea", 500.0, "No", "1975-06",
     "DRAM/NAND/HBM + foundry + devices; HBM4 catch-up is the key AI catalyst."),
    ("SK hynix", "000660.KS", "KRX", "Semis - Memory (IDM)", "South Korea", 800.0, "No", "1996-12",
     "HBM leader into AI supercycle; primary NVIDIA HBM supplier; DRAM/NAND."),
    ("Hanmi Semiconductor", "042700.KS", "KRX", "Semis - Equipment (Packaging)", "South Korea", 9.0, "No", "2005-07",
     "TC bonders for HBM stacking (SK hynix/Micron supplier); pure-play AI memory equipment."),
    # =========================== JAPAN ===========================
    ("Tokyo Electron", "8035.T", "TSE Prime", "Semis - Equipment (WFE)", "Japan", 150.0, "No", "1980-10",
     "Top-tier WFE (coater/developer near-monopoly, etch, deposition); Japan flagship."),
    ("Advantest", "6857.T", "TSE Prime", "Semis - Equipment (Test)", "Japan", 150.0, "No", "1983-09",
     "Dominant SoC/memory tester; every AI accelerator and HBM stack needs test time."),
    ("Kioxia Holdings", "285A.T", "TSE Prime", "Semis - Memory (NAND)", "Japan", 20.0, "Yes", "2024-12",
     "NAND flash pure-play (ex-Toshiba Memory); recent Tokyo IPO; AI storage demand."),
    ("Lasertec", "6920.T", "TSE Prime", "Semis - Equipment (Inspection)", "Japan", 20.0, "No", "1990-08",
     "Monopoly in EUV mask/blank inspection; leading-edge logic capex proxy."),
    ("Disco Corporation", "6146.T", "TSE Prime", "Semis - Equipment (Packaging)", "Japan", 30.0, "No", "1989-11",
     "Dicing/grinding near-monopoly; HBM wafer thinning; advanced packaging leverage."),
    ("Ibiden", "4062.T", "TSE Prime", "Semis - Substrates", "Japan", 15.0, "No", "1960s",
     "ABF package substrates for AI GPUs (NVIDIA supplier); capacity expansion cycle."),
    ("Fujikura", "5803.T", "TSE Prime", "Data Center - Optical Connectivity", "Japan", 25.0, "No", "1949 (TSE)",
     "Optical fiber/cable and high-density interconnect for AI datacenters; top TOPIX performer."),
    # =========================== SINGAPORE & OTHER ASIA ===========================
    ("Keppel DC REIT", "AJBU.SI", "SGX", "Data Center - REIT", "Singapore", 4.3, "No", "2014-12",
     "Asia's first pure-play datacenter REIT; ~S$5B AUM, 20+ assets across 10 countries."),
    ("NTT DC REIT", "NTDU.SI", "SGX", "Data Center - REIT", "Singapore", 1.0, "Yes", "2025-07",
     "Largest SGX REIT IPO in a decade (US$773M raise); NTT-sponsored global datacenter portfolio."),
    ("YTL Power International", "6742.KL", "Bursa Malaysia", "Data Center - Power/AI Campus", "Malaysia", 10.0, "No", "1997 (listed)",
     "Utility building Johor AI datacenter campus (NVIDIA GPU cloud partnership); power + DC play."),
]

HEADERS = [
    "Company", "Ticker", "Exchange", "Subsector", "Listing Region",
    "Approx. Market Cap ($B)", "Recent IPO/Listing", "IPO/Listing Date", "Notes / Thesis",
]

REGION_ORDER = ["Mainland China", "Hong Kong", "Taiwan", "South Korea", "Japan", "Singapore", "Malaysia"]


def write_csv(path):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        for row in COVERAGE:
            w.writerow(row)


def write_xlsx(path):
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Coverage universe ----
    ws = wb.active
    ws.title = "Coverage Universe"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    ipo_fill = PatternFill("solid", fgColor="E2EFDA")  # highlight recent IPOs
    region_fill = PatternFill("solid", fgColor="D6E4F0")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "Semiconductor & Data Center - Proposed Coverage Universe (Asia-Listed Only)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (f"As of {AS_OF}. Scope: mainland China, Hong Kong, Taiwan, South Korea, Japan, Singapore, "
                "Malaysia listings only - no US or Europe listings. Market caps approximate/rounded for "
                "scoping only (not investment advice). Recent IPOs/listings highlighted in green.")
    ws["A2"].font = Font(italic=True, size=9, color="808080")
    ws.append([])  # row 3 spacer

    header_row_idx = 4
    ws.append(HEADERS)
    for c, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row_idx, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = border

    for region in REGION_ORDER:
        rows = [r for r in COVERAGE if r[4] == region]
        # Region banner row
        ws.append([region.upper()])
        r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(HEADERS))
        cell = ws.cell(row=r, column=1)
        cell.fill = region_fill
        cell.font = Font(bold=True, size=11, color="1F3864")
        cell.border = border
        for row in sorted(rows, key=lambda x: -x[5]):
            ws.append(row)
            r = ws.max_row
            recent = row[6] == "Yes"
            for c in range(1, len(HEADERS) + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=(c == 9))
                if recent:
                    cell.fill = ipo_fill
            ws.cell(row=r, column=6).number_format = "#,##0.0"
            ws.cell(row=r, column=6).alignment = Alignment(horizontal="right", vertical="top")

    ws.freeze_panes = f"A{header_row_idx + 1}"

    widths = [30, 26, 22, 32, 16, 18, 14, 24, 78]
    for i, wdt in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wdt

    # ---- Sheet 2: Summary by region & subsector ----
    ws2 = wb.create_sheet("Summary")

    def agg(keyfunc, keys):
        out = []
        for k in keys:
            rows = [r for r in COVERAGE if keyfunc(r) == k]
            if rows:
                out.append((k, len(rows), round(sum(r[5] for r in rows), 1),
                            sum(1 for r in rows if r[6] == "Yes")))
        return out

    ws2.append(["By Listing Region", "# Names", "Agg. Approx. Mkt Cap ($B)", "# Recent IPOs"])
    for c in range(1, 5):
        cell = ws2.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
    for k, n, cap, ipos in agg(lambda r: r[4], REGION_ORDER):
        ws2.append([k, n, cap, ipos])
        for c in range(1, 5):
            ws2.cell(row=ws2.max_row, column=c).border = border

    total_names = len(COVERAGE)
    total_cap = round(sum(r[5] for r in COVERAGE), 1)
    total_ipos = sum(1 for r in COVERAGE if r[6] == "Yes")
    ws2.append(["TOTAL", total_names, total_cap, total_ipos])
    for c in range(1, 5):
        cell = ws2.cell(row=ws2.max_row, column=c)
        cell.font = Font(bold=True)
        cell.border = border

    ws2.append([])
    start = ws2.max_row + 1
    ws2.append(["By Subsector", "# Names", "Agg. Approx. Mkt Cap ($B)", "# Recent IPOs"])
    for c in range(1, 5):
        cell = ws2.cell(row=start, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
    subsectors = sorted({r[3] for r in COVERAGE})
    for k, n, cap, ipos in agg(lambda r: r[3], subsectors):
        ws2.append([k, n, cap, ipos])
        for c in range(1, 5):
            ws2.cell(row=ws2.max_row, column=c).border = border

    for i, wdt in enumerate([36, 12, 26, 16], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = wdt
    ws2.freeze_panes = "A2"

    wb.save(path)


if __name__ == "__main__":
    write_csv("semi_datacenter_coverage.csv")
    write_xlsx("semi_datacenter_coverage.xlsx")
    print(f"Coverage names: {len(COVERAGE)}")
    print(f"Recent IPOs/listings: {sum(1 for r in COVERAGE if r[6] == 'Yes')}")
    print(f"Names with mkt cap >= $1B: {sum(1 for r in COVERAGE if r[5] >= 1)}")
    for region in REGION_ORDER:
        n = sum(1 for r in COVERAGE if r[4] == region)
        print(f"  {region}: {n}")
    print(f"Generated on {date.today().isoformat()}")
