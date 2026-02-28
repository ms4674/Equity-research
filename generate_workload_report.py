import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

wb = openpyxl.Workbook()

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="2F5496")
DATA_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)
ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
CURRENCY_FMT = '#,##0.0'
PCT_FMT = '0.0%'
INT_FMT = '#,##0'


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if (r - start_row) % 2 == 1:
                cell.fill = ALT_ROW_FILL


def auto_width(ws, max_col, min_width=12, max_width=40):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        longest = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    longest = max(longest, min(len(str(cell.value)) + 4, max_width))
        ws.column_dimensions[letter].width = longest


# ─────────────────────────────────────────────────────────────────────
# SHEET 1: Executive Summary
# ─────────────────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_properties.tabColor = "2F5496"

ws1.merge_cells("A1:F1")
ws1["A1"] = "Hyperscale Cloud Workload Aggregation — Global Overview"
ws1["A1"].font = TITLE_FONT
ws1["A1"].alignment = Alignment(horizontal="left")

ws1.merge_cells("A2:F2")
ws1["A2"] = "Sources: Gartner, IDC, Synergy Research, Flexera, CloudZero, CRN, Omdia, provider earnings reports | Data as of Q4 2025 / early 2026"
ws1["A2"].font = Font(name="Calibri", size=9, italic=True, color="808080")

summary_headers = ["Metric", "Value", "Period", "Source"]
for i, h in enumerate(summary_headers, 1):
    ws1.cell(row=4, column=i, value=h)
style_header_row(ws1, 4, len(summary_headers))

summary_data = [
    ["Global Public Cloud Spending", "$723.4B", "2025 Forecast", "Gartner"],
    ["Global Cloud Infrastructure Revenue", "$419B+", "Full Year 2025", "Synergy / CRN"],
    ["Q4 2025 Cloud Infra Revenue", "$119B", "Q4 2025", "CRN"],
    ["YoY Cloud Infra Growth", "30%", "Q4 2025", "CRN"],
    ["IaaS Market Size", "$211.9B", "2025 Forecast", "Gartner"],
    ["PaaS Market Size", "$208.6B", "2025 Forecast", "Gartner"],
    ["SaaS Market Size", "$299B", "2025 Forecast", "Gartner"],
    ["Total Hyperscale Data Centers", "~1,300", "2025", "Synergy Research"],
    ["Top-3 Provider Market Share", "68%", "Q4 2025", "CRN"],
    ["AI/ML as % of Cloud Spend", "2.44%", "Dec 2025", "CloudZero"],
    ["AI Infra Spending Growth", "105% YoY", "2025-2026", "Various"],
    ["Worldwide AI Spending Projection", "$2.02T", "2026 Forecast", "IDC / Various"],
    ["Hyperscale Cloud Market (CAGR)", "18.8%", "2025-2029", "GlobeNewsWire"],
]

for i, row in enumerate(summary_data, 5):
    for j, val in enumerate(row, 1):
        ws1.cell(row=i, column=j, value=val)
style_data_rows(ws1, 5, 5 + len(summary_data) - 1, len(summary_headers))
auto_width(ws1, len(summary_headers))

# ─────────────────────────────────────────────────────────────────────
# SHEET 2: Provider Overview
# ─────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Provider Overview")
ws2.sheet_properties.tabColor = "4472C4"

ws2.merge_cells("A1:I1")
ws2["A1"] = "Major Hyperscale Cloud Providers — Global Footprint & Financials"
ws2["A1"].font = TITLE_FONT

prov_headers = [
    "Provider", "Parent Company", "Q4 2025 Revenue ($B)",
    "Annualized Run Rate ($B)", "YoY Growth", "Global Regions",
    "Market Share (Q4 2025)", "GPU Spend Share", "Headquarters"
]
for i, h in enumerate(prov_headers, 1):
    ws2.cell(row=3, column=i, value=h)
style_header_row(ws2, 3, len(prov_headers))

prov_data = [
    ["AWS", "Amazon", 35.6, 142, 0.24, 36, "~31%", "8%", "Seattle, USA"],
    ["Microsoft Azure", "Microsoft", 32.9, 131, 0.29, 60, "~24%", "1.5%", "Redmond, USA"],
    ["Google Cloud", "Alphabet", 17.7, 71, 0.48, 42, "~13%", "14%", "Mountain View, USA"],
    ["Alibaba Cloud", "Alibaba Group", 4.3, 17.2, 0.34, 29, "~4%", "N/A", "Hangzhou, China"],
    ["Oracle Cloud (OCI)", "Oracle", 6.7, 26.8, 0.27, 50, "~3%", "N/A", "Austin, USA"],
    ["IBM Cloud", "IBM", 1.5, 6.0, 0.05, 15, "~2%", "N/A", "Armonk, USA"],
    ["Huawei Cloud", "Huawei", 3.0, 12.0, 0.22, 33, "~3%", "N/A", "Shenzhen, China"],
    ["Tencent Cloud", "Tencent", 2.0, 8.0, 0.15, 21, "~2%", "N/A", "Shenzhen, China"],
]

for i, row in enumerate(prov_data, 4):
    for j, val in enumerate(row, 1):
        cell = ws2.cell(row=i, column=j, value=val)
        if j in (3, 4):
            cell.number_format = CURRENCY_FMT
        if j == 5:
            cell.number_format = PCT_FMT

style_data_rows(ws2, 4, 4 + len(prov_data) - 1, len(prov_headers))
auto_width(ws2, len(prov_headers), min_width=14)

# ─────────────────────────────────────────────────────────────────────
# SHEET 3: Workload Types
# ─────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Workload Types")
ws3.sheet_properties.tabColor = "548235"

ws3.merge_cells("A1:H1")
ws3["A1"] = "Cloud Workload Types — Classification & Distribution Across Hyperscale Providers"
ws3["A1"].font = TITLE_FONT

wl_headers = [
    "Workload Category", "Description", "% of Total Cloud Spend (2025)",
    "YoY Growth", "Dominant Providers",
    "Key Services / Products", "Trend", "Notes"
]
for i, h in enumerate(wl_headers, 1):
    ws3.cell(row=3, column=i, value=h)
style_header_row(ws3, 3, len(wl_headers))

wl_data = [
    [
        "Compute (VMs & Bare Metal)", "Virtual machines, dedicated hosts, HPC instances",
        0.47, 0.12, "AWS, Azure, GCP",
        "EC2, Azure VMs, Compute Engine", "Softening",
        "Dropped below 49% by Dec 2025; shift to specialized compute"
    ],
    [
        "Storage", "Object, block, and file storage services",
        0.14, 0.18, "AWS, Azure, GCP",
        "S3, Azure Blob, Cloud Storage", "Steady growth",
        "Growing with data-intensive AI pipelines"
    ],
    [
        "Databases", "Managed relational, NoSQL, and in-memory databases",
        0.11, 0.22, "AWS, Azure, GCP, Oracle",
        "RDS, Cosmos DB, Cloud SQL, Autonomous DB", "Accelerating",
        "Expanding to support AI/ML data layers"
    ],
    [
        "Networking & CDN", "Virtual networks, load balancers, DNS, CDN",
        0.07, 0.15, "AWS, Azure, GCP, Alibaba",
        "VPC, Azure VNet, Cloud CDN", "Steady",
        "Edge and CDN gaining with AI inference at edge"
    ],
    [
        "AI / ML & GenAI", "Training, inference, MLOps, foundation models, GenAI APIs",
        0.0244, 0.85, "AWS, Azure, GCP",
        "Bedrock, Azure OpenAI, Vertex AI", "Rapid acceleration",
        "Reached 2.44% of spend in Dec 2025; triple-digit growth at some providers"
    ],
    [
        "Analytics & Big Data", "Data warehousing, BI, streaming analytics",
        0.06, 0.20, "AWS, GCP, Azure",
        "Redshift, BigQuery, Synapse, Databricks", "Growing",
        "Converging with AI/ML for real-time insights"
    ],
    [
        "Containers & Orchestration", "Managed Kubernetes, container registries, serverless containers",
        0.04, 0.30, "AWS, Azure, GCP",
        "EKS, AKS, GKE, Fargate", "Strong growth",
        "Dominant deployment model for microservices & AI serving"
    ],
    [
        "Serverless / FaaS", "Event-driven functions, managed application backends",
        0.025, 0.25, "AWS, Azure, GCP",
        "Lambda, Azure Functions, Cloud Functions", "Growing",
        "Increasingly used for AI inference endpoints"
    ],
    [
        "Security & Identity", "IAM, encryption, compliance, threat detection",
        0.03, 0.28, "AWS, Azure, GCP",
        "IAM, Security Hub, Sentinel, Chronicle", "Accelerating",
        "77% of orgs cite security as top cloud challenge"
    ],
    [
        "DevOps & CI/CD", "Source control, CI/CD pipelines, IaC, monitoring",
        0.02, 0.22, "AWS, Azure, GCP",
        "CodePipeline, Azure DevOps, Cloud Build", "Steady growth",
        "Projected $9.1B marketplace category by 2030"
    ],
    [
        "IoT & Edge", "Device management, edge compute, IoT analytics",
        0.015, 0.20, "AWS, Azure, GCP",
        "IoT Core, Azure IoT Hub, Cloud IoT", "Emerging",
        "Growing with edge AI inference deployments"
    ],
    [
        "Business Applications (SaaS on Cloud)", "ERP, CRM, HCM, collaboration delivered via cloud",
        0.09, 0.15, "Oracle, Microsoft, Salesforce",
        "Fusion ERP, Dynamics 365, Workspace", "Steady",
        "Projected $9.1B marketplace category by 2030"
    ],
]

for i, row in enumerate(wl_data, 4):
    for j, val in enumerate(row, 1):
        cell = ws3.cell(row=i, column=j, value=val)
        if j == 3:
            cell.number_format = PCT_FMT
        if j == 4:
            cell.number_format = PCT_FMT

style_data_rows(ws3, 4, 4 + len(wl_data) - 1, len(wl_headers))
auto_width(ws3, len(wl_headers), max_width=50)

# ─────────────────────────────────────────────────────────────────────
# SHEET 4: AI Token Consumption
# ─────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("AI Token Consumption")
ws4.sheet_properties.tabColor = "BF8F00"

ws4.merge_cells("A1:I1")
ws4["A1"] = "AI Workload Token Consumption — Model Pricing, Throughput & Enterprise Usage Patterns"
ws4["A1"].font = TITLE_FONT

# Section A: Model Pricing
ws4["A3"] = "A. LLM API Pricing per Million Tokens (February 2026)"
ws4["A3"].font = SUBTITLE_FONT

tok_headers = [
    "Model", "Provider", "Tier", "Input $/1M Tokens",
    "Output $/1M Tokens", "Context Window",
    "Max Output Tokens", "Throughput (tok/min)", "Notes"
]
for i, h in enumerate(tok_headers, 1):
    ws4.cell(row=4, column=i, value=h)
style_header_row(ws4, 4, len(tok_headers))

tok_data = [
    ["GPT-5.2", "OpenAI", "Flagship", 1.75, 14.00, "256K", "16K", "800K", "Latest reasoning model"],
    ["Claude Opus 4.6", "Anthropic", "Flagship", 5.00, 25.00, "200K", "32K", "400K", "Highest quality; extended thinking"],
    ["Gemini 3.1 Pro", "Google", "Flagship", 2.00, 12.00, "2M", "65K", "2M+", "Largest context window"],
    ["Claude Sonnet 4.6", "Anthropic", "Mid-Tier", 3.00, 15.00, "200K", "16K", "400K", "Best mid-tier balance"],
    ["GPT-5", "OpenAI", "Mid-Tier", 1.25, 10.00, "128K", "16K", "800K", "Strong general purpose"],
    ["Gemini 2.5 Flash", "Google", "Mid-Tier", 0.15, 0.60, "1M", "65K", "2M+", "Thinking model; fast"],
    ["GPT-5 Mini", "OpenAI", "Budget", 0.25, 2.00, "128K", "16K", "800K", "Cost-effective at scale"],
    ["Claude Haiku 4.5", "Anthropic", "Budget", 1.00, 5.00, "200K", "8K", "400K", "Fast; lightweight tasks"],
    ["DeepSeek V3", "DeepSeek", "Budget", 0.14, 0.28, "128K", "8K", "N/A", "Open-weights; 10-30x cheaper than GPT-4"],
    ["Gemini 2.0 Flash", "Google", "Budget", 0.10, 0.40, "1M", "8K", "2M+", "Cheapest mainstream model"],
    ["Llama 4 (405B)", "Meta (self-hosted)", "Open Source", 0.00, 0.00, "128K", "8K", "Varies", "Free weights; infrastructure cost only"],
]

for i, row in enumerate(tok_data, 5):
    for j, val in enumerate(row, 1):
        cell = ws4.cell(row=i, column=j, value=val)
        if j in (4, 5):
            cell.number_format = '$#,##0.00'

style_data_rows(ws4, 5, 5 + len(tok_data) - 1, len(tok_headers))

# Section B: Average Token Consumption by Use Case
row_b = 5 + len(tok_data) + 2
ws4.cell(row=row_b, column=1, value="B. Average Token Consumption by AI Workload Type").font = SUBTITLE_FONT

use_headers = [
    "AI Workload Type", "Avg Input Tokens/Request",
    "Avg Output Tokens/Request", "Avg Requests/Day (Enterprise)",
    "Avg Daily Token Consumption (M)", "Est. Monthly Cost (Mid-Tier Model)",
    "Primary Cloud Services", "% of AI Cloud Spend"
]
for i, h in enumerate(use_headers, 1):
    ws4.cell(row=row_b + 1, column=i, value=h)
style_header_row(ws4, row_b + 1, len(use_headers))

use_data = [
    ["Chatbot / Customer Support", 250, 150, 50000, 20.0, "$6,000 – $15,000",
     "Azure OpenAI, Bedrock, Vertex AI", "18%"],
    ["Code Generation & Assistance", 800, 500, 30000, 39.0, "$12,000 – $30,000",
     "GitHub Copilot, CodeWhisperer, Gemini Code", "15%"],
    ["Document Processing & Summarization", 15000, 1000, 10000, 160.0, "$25,000 – $80,000",
     "Azure Doc Intelligence, Textract, Vertex AI", "14%"],
    ["Search & RAG (Retrieval-Augmented Gen.)", 5000, 500, 25000, 137.5, "$20,000 – $60,000",
     "Azure AI Search, Kendra, Vertex Search", "13%"],
    ["Content Generation (Marketing/Creative)", 500, 2000, 5000, 12.5, "$8,000 – $25,000",
     "Azure OpenAI, Bedrock, Vertex AI", "10%"],
    ["Data Analytics & BI Copilots", 3000, 800, 8000, 30.4, "$10,000 – $35,000",
     "Copilot for Power BI, Q in QuickSight", "8%"],
    ["AI Training (Fine-tuning)", 50000, 0, 500, 25.0, "$15,000 – $50,000",
     "SageMaker, Azure ML, Vertex AI", "8%"],
    ["AI Inference (Batch / Offline)", 2000, 500, 100000, 250.0, "$30,000 – $100,000",
     "SageMaker Batch, Azure Batch, Vertex Batch", "7%"],
    ["Image / Video Generation", 200, 100, 15000, 4.5, "$5,000 – $20,000",
     "DALL-E, Imagen, Titan Image", "4%"],
    ["Agent / Agentic AI Workflows", 5000, 3000, 2000, 16.0, "$10,000 – $40,000",
     "LangChain on Cloud, Azure AI Agent Service", "3%"],
]

for i, row in enumerate(use_data, row_b + 2):
    for j, val in enumerate(row, 1):
        cell = ws4.cell(row=i, column=j, value=val)
        if j in (2, 3, 4):
            cell.number_format = INT_FMT
        if j == 5:
            cell.number_format = '#,##0.0'

style_data_rows(ws4, row_b + 2, row_b + 2 + len(use_data) - 1, len(use_headers))

# Section C: Token Cost Trends
row_c = row_b + 2 + len(use_data) + 2
ws4.cell(row=row_c, column=1, value="C. Token Cost Decline Trends (2023 → 2026)").font = SUBTITLE_FONT

trend_headers = ["Year", "GPT-4 Class Input $/1M", "GPT-4 Class Output $/1M",
                 "Median Price Decline Factor", "Inference as % of AI Spend"]
for i, h in enumerate(trend_headers, 1):
    ws4.cell(row=row_c + 1, column=i, value=h)
style_header_row(ws4, row_c + 1, len(trend_headers))

trend_data = [
    [2023, 30.00, 60.00, "1x (baseline)", "35%"],
    [2024, 10.00, 30.00, "~3x decline", "45%"],
    [2025, 2.50, 10.00, "~12x decline", "50%"],
    [2026, 1.25, 10.00, "~24-50x decline", "55-70%"],
]

for i, row in enumerate(trend_data, row_c + 2):
    for j, val in enumerate(row, 1):
        cell = ws4.cell(row=i, column=j, value=val)
        if j in (2, 3):
            cell.number_format = '$#,##0.00'

style_data_rows(ws4, row_c + 2, row_c + 2 + len(trend_data) - 1, len(trend_headers))
auto_width(ws4, len(tok_headers), max_width=45)

# ─────────────────────────────────────────────────────────────────────
# SHEET 5: Monetization by Workload
# ─────────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("Monetization by Workload")
ws5.sheet_properties.tabColor = "C00000"

ws5.merge_cells("A1:J1")
ws5["A1"] = "Monetization by Workload Type — Revenue, Growth & Forecast Across Hyperscale Clouds"
ws5["A1"].font = TITLE_FONT

# Section A: Current Revenue by Workload
ws5["A3"] = "A. Estimated Global Revenue by Workload Category (2025)"
ws5["A3"].font = SUBTITLE_FONT

mon_headers = [
    "Workload Category", "Est. 2024 Revenue ($B)", "Est. 2025 Revenue ($B)",
    "YoY Growth", "% of Total Cloud Revenue",
    "2029 Projected Revenue ($B)", "CAGR (2025-2029)",
    "Key Revenue Drivers", "Margin Profile"
]
for i, h in enumerate(mon_headers, 1):
    ws5.cell(row=4, column=i, value=h)
style_header_row(ws5, 4, len(mon_headers))

mon_data = [
    ["Compute (VMs, HPC)", 165.0, 193.0, 0.17, 0.267,
     310.0, 0.125, "Reserved instances, spot, on-demand", "Moderate (30-40%)"],
    ["Storage & Data Mgmt", 55.0, 65.0, 0.18, 0.090,
     110.0, 0.14, "Volume-based; tiered pricing", "High (40-50%)"],
    ["Databases (Managed)", 45.0, 55.0, 0.22, 0.076,
     100.0, 0.16, "Consumption-based; serverless tiers", "High (45-55%)"],
    ["Networking & CDN", 32.0, 37.0, 0.15, 0.051,
     58.0, 0.12, "Egress fees, CDN bandwidth", "High (50-60%)"],
    ["AI / ML & GenAI", 12.0, 22.0, 0.85, 0.030,
     95.0, 0.44, "GPU compute, API tokens, model hosting", "Emerging (15-35%)"],
    ["Analytics & Big Data", 28.0, 34.0, 0.20, 0.047,
     62.0, 0.16, "Query-based (BigQuery), cluster (Redshift)", "Moderate (35-45%)"],
    ["Containers & K8s", 15.0, 20.0, 0.30, 0.028,
     45.0, 0.22, "Managed K8s, serverless containers", "Moderate (30-40%)"],
    ["Serverless / FaaS", 10.0, 12.5, 0.25, 0.017,
     28.0, 0.22, "Per-invocation, per-duration", "High (50-60%)"],
    ["Security & Identity", 14.0, 18.0, 0.28, 0.025,
     42.0, 0.24, "Consumption + seat-based", "High (50-60%)"],
    ["DevOps & CI/CD", 9.0, 11.0, 0.22, 0.015,
     22.0, 0.19, "Pipeline minutes, artifact storage", "Moderate (35-45%)"],
    ["IoT & Edge Computing", 6.0, 7.5, 0.20, 0.010,
     16.0, 0.21, "Device connections, edge compute", "Emerging (20-30%)"],
    ["Business Apps (SaaS)", 75.0, 87.0, 0.15, 0.120,
     140.0, 0.13, "Per-user, per-org subscription", "High (60-70%)"],
    ["Marketplace / ISV", 30.0, 42.0, 0.40, 0.058,
     163.0, 0.40, "Commission on third-party sales", "Very High (60-80%)"],
    ["Other / Misc Services", 20.0, 22.0, 0.10, 0.030,
     35.0, 0.12, "Various", "Mixed"],
]

for i, row in enumerate(mon_data, 5):
    for j, val in enumerate(row, 1):
        cell = ws5.cell(row=i, column=j, value=val)
        if j in (2, 3, 6):
            cell.number_format = CURRENCY_FMT
        if j in (4, 5, 7):
            cell.number_format = PCT_FMT

style_data_rows(ws5, 5, 5 + len(mon_data) - 1, len(mon_headers))

# Totals row
total_row = 5 + len(mon_data)
ws5.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Calibri", bold=True, size=11)
ws5.cell(row=total_row, column=2, value=sum(r[1] for r in mon_data)).number_format = CURRENCY_FMT
ws5.cell(row=total_row, column=2).font = Font(name="Calibri", bold=True, size=11)
ws5.cell(row=total_row, column=3, value=sum(r[2] for r in mon_data)).number_format = CURRENCY_FMT
ws5.cell(row=total_row, column=3).font = Font(name="Calibri", bold=True, size=11)
for c in range(1, len(mon_headers) + 1):
    ws5.cell(row=total_row, column=c).border = THIN_BORDER

# Section B: Provider-Level Revenue Breakdown
row_b5 = total_row + 3
ws5.cell(row=row_b5, column=1, value="B. Revenue by Provider & Workload (Q4 2025 Annualized, $B)").font = SUBTITLE_FONT

prov_wl_headers = [
    "Workload Category", "AWS", "Azure", "Google Cloud",
    "Oracle Cloud", "Alibaba Cloud", "Others", "Total"
]
for i, h in enumerate(prov_wl_headers, 1):
    ws5.cell(row=row_b5 + 1, column=i, value=h)
style_header_row(ws5, row_b5 + 1, len(prov_wl_headers))

prov_wl_data = [
    ["Compute", 55.0, 38.0, 20.0, 7.0, 5.0, 18.0, 143.0],
    ["Storage", 20.0, 15.0, 8.0, 2.5, 2.0, 7.5, 55.0],
    ["Databases", 18.0, 12.0, 6.0, 6.0, 1.5, 5.5, 49.0],
    ["Networking & CDN", 12.0, 8.0, 5.0, 1.0, 2.0, 5.0, 33.0],
    ["AI/ML & GenAI", 6.5, 8.0, 5.5, 0.5, 1.0, 1.5, 23.0],
    ["Analytics", 8.0, 7.0, 8.0, 1.0, 0.5, 3.5, 28.0],
    ["Containers & Serverless", 7.0, 5.0, 4.0, 0.5, 0.5, 2.0, 19.0],
    ["Security & Identity", 5.0, 6.0, 3.0, 0.5, 0.5, 2.0, 17.0],
    ["Other Services", 10.5, 8.0, 6.5, 3.0, 3.0, 5.0, 36.0],
]

for i, row in enumerate(prov_wl_data, row_b5 + 2):
    for j, val in enumerate(row, 1):
        cell = ws5.cell(row=i, column=j, value=val)
        if j >= 2:
            cell.number_format = CURRENCY_FMT

style_data_rows(ws5, row_b5 + 2, row_b5 + 2 + len(prov_wl_data) - 1, len(prov_wl_headers))

# Section C: AI Workload Monetization Deep-Dive
row_c5 = row_b5 + 2 + len(prov_wl_data) + 2
ws5.cell(row=row_c5, column=1, value="C. AI/GenAI Workload Monetization Breakdown (2025-2026)").font = SUBTITLE_FONT

ai_mon_headers = [
    "AI Revenue Stream", "Est. 2025 Revenue ($B)", "Est. 2026 Revenue ($B)",
    "YoY Growth", "Revenue Model", "Key Products"
]
for i, h in enumerate(ai_mon_headers, 1):
    ws5.cell(row=row_c5 + 1, column=i, value=h)
style_header_row(ws5, row_c5 + 1, len(ai_mon_headers))

ai_mon_data = [
    ["GPU / Accelerator Compute", 8.0, 15.0, 0.88, "Per-hour / reserved instances",
     "EC2 P5, Azure ND, GCP A3 (NVIDIA H100/B200)"],
    ["Foundation Model APIs (Token-Based)", 4.5, 10.0, 1.22, "Per-token (input + output)",
     "Azure OpenAI, Bedrock, Vertex AI"],
    ["ML Platform & MLOps", 3.5, 5.0, 0.43, "Consumption + subscription",
     "SageMaker, Azure ML, Vertex AI"],
    ["AI Application Services", 2.5, 4.5, 0.80, "Per-API-call / tiered",
     "Vision, Speech, Translation, Doc AI"],
    ["Fine-tuning & Custom Models", 1.5, 3.0, 1.00, "Compute-hours + storage",
     "OpenAI Fine-tuning, SageMaker Training"],
    ["Vector Databases & AI Search", 1.0, 2.5, 1.50, "Storage + query-based",
     "Azure AI Search, Kendra, pgvector"],
    ["AI Infrastructure (Chips/Custom Silicon)", 1.0, 2.5, 1.50, "Bundled with compute",
     "AWS Trainium/Inferentia, Google TPU, Azure Maia"],
]

for i, row in enumerate(ai_mon_data, row_c5 + 2):
    for j, val in enumerate(row, 1):
        cell = ws5.cell(row=i, column=j, value=val)
        if j in (2, 3):
            cell.number_format = CURRENCY_FMT
        if j == 4:
            cell.number_format = PCT_FMT

style_data_rows(ws5, row_c5 + 2, row_c5 + 2 + len(ai_mon_data) - 1, len(ai_mon_headers))
auto_width(ws5, len(mon_headers), max_width=45)

# ─────────────────────────────────────────────────────────────────────
# SHEET 6: Regional Breakdown
# ─────────────────────────────────────────────────────────────────────
ws6 = wb.create_sheet("Regional Breakdown")
ws6.sheet_properties.tabColor = "7030A0"

ws6.merge_cells("A1:G1")
ws6["A1"] = "Cloud Workload Distribution by Global Region"
ws6["A1"].font = TITLE_FONT

reg_headers = [
    "Region", "% of Global Cloud Spend", "Leading Providers",
    "Top Workload Types", "Key Growth Driver",
    "Notable Regulations", "2025 Est. Market Size ($B)"
]
for i, h in enumerate(reg_headers, 1):
    ws6.cell(row=3, column=i, value=h)
style_header_row(ws6, 3, len(reg_headers))

reg_data = [
    ["North America", 0.42, "AWS, Azure, GCP",
     "AI/ML, Compute, SaaS", "GenAI adoption, enterprise migration",
     "FedRAMP, SOC 2, HIPAA", 304],
    ["Europe (EU + UK)", 0.22, "Azure, AWS, GCP",
     "Compute, Security, Compliance", "Data sovereignty, AI Act",
     "GDPR, EU AI Act, Schrems II", 159],
    ["Asia-Pacific (excl. China)", 0.14, "AWS, Azure, GCP, Alibaba",
     "Compute, Storage, Analytics", "Digital transformation, fintech",
     "PDPA, APPI, local data residency", 101],
    ["China", 0.12, "Alibaba, Tencent, Huawei",
     "Compute, AI/ML, E-commerce", "AI innovation, domestic cloud",
     "PIPL, CSL, data localization", 87],
    ["Middle East & Africa", 0.04, "Azure, AWS, Oracle",
     "Compute, Storage, Gov Cloud", "Smart city initiatives, Vision 2030",
     "Emerging data protection laws", 29],
    ["Latin America", 0.04, "AWS, Azure, GCP",
     "Compute, SaaS, Databases", "SMB digitization, fintech growth",
     "LGPD (Brazil), local emerging rules", 29],
    ["India (standalone)", 0.02, "AWS, Azure, GCP",
     "Compute, AI/ML, Analytics", "Digital public infra, AI startups",
     "DPDP Act 2023, data localization push", 14],
]

for i, row in enumerate(reg_data, 4):
    for j, val in enumerate(row, 1):
        cell = ws6.cell(row=i, column=j, value=val)
        if j == 2:
            cell.number_format = PCT_FMT
        if j == 7:
            cell.number_format = CURRENCY_FMT

style_data_rows(ws6, 4, 4 + len(reg_data) - 1, len(reg_headers))
auto_width(ws6, len(reg_headers), max_width=45)

# ─────────────────────────────────────────────────────────────────────
# SHEET 7: Cloud Growth — Traditional vs AI Workloads
# ─────────────────────────────────────────────────────────────────────
ws_growth = wb.create_sheet("Growth — Trad. vs AI")
ws_growth.sheet_properties.tabColor = "00B050"

ws_growth.merge_cells("A1:N1")
ws_growth["A1"] = "Hyperscaler Cloud Growth — Traditional Workloads vs AI Workloads"
ws_growth["A1"].font = TITLE_FONT

ws_growth.merge_cells("A2:N2")
ws_growth["A2"] = "Sources: Provider earnings (Q4 2025 / FY26 Q1-Q2), CRN, analyst estimates | Traditional = compute, storage, databases, networking, SaaS; AI = GPU compute, model APIs, ML platforms, AI services"
ws_growth["A2"].font = Font(name="Calibri", size=9, italic=True, color="808080")

# Section A: Quarterly Growth Comparison
ws_growth["A4"] = "A. Revenue Growth Split: Traditional Cloud vs AI Workloads by Provider"
ws_growth["A4"].font = SUBTITLE_FONT

gr_headers = [
    "Provider",
    "Total Cloud Revenue Q4 2025 ($B)",
    "Total YoY Growth",
    "Traditional Cloud Revenue ($B)",
    "Traditional YoY Growth",
    "AI Workload Revenue ($B)",
    "AI YoY Growth",
    "AI as % of Total Revenue",
    "AI Contribution to Growth (pp)",
    "Traditional Contribution to Growth (pp)",
    "2026E Total Growth",
    "2026E AI Growth",
    "2026E Traditional Growth",
    "Notes"
]
for i, h in enumerate(gr_headers, 1):
    ws_growth.cell(row=5, column=i, value=h)
style_header_row(ws_growth, 5, len(gr_headers))

# Data derived from provider earnings, analyst decompositions, and disclosed metrics
gr_data = [
    [
        "AWS", 35.6, 0.24,
        31.5, 0.17, 4.1, 1.10,
        0.115, 8, 16,
        "26-28%", "90-100%", "15-18%",
        "Custom silicon ($10B+ ARR, triple-digit growth); Bedrock multi-billion ARR; traditional reaccelerating on migration wave"
    ],
    [
        "Microsoft Azure", 32.9, 0.29,
        24.7, 0.14, 8.2, 0.95,
        0.249, 15, 14,
        "35-38%", "70-80%", "13-17%",
        "AI contributes ~22-26pp of Azure's 39% growth (CFO Amy Hood); traditional at 13-17%; deliberately constraining 3P for 1P AI"
    ],
    [
        "Google Cloud", 17.7, 0.48,
        13.3, 0.28, 4.4, 1.50,
        0.249, 20, 28,
        "40-45%", "100-120%", "22-26%",
        "Fastest overall growth; Gemini processing 10B+ tokens/min; $240B backlog; traditional also strong (core infra migration)"
    ],
    [
        "Oracle Cloud (OCI)", 8.0, 0.34,
        4.8, 0.11, 3.2, 1.05,
        0.400, 23, 11,
        "50-55%", "80-100%", "10-12%",
        "OCI IaaS up 68% (AI-driven); SaaS at 11%; $523B RPO backlog (438% YoY); multi-cloud OCI for Meta/NVIDIA"
    ],
    [
        "Alibaba Cloud", 4.3, 0.34,
        3.2, 0.12, 1.1, 2.00,
        0.256, 22, 12,
        "30-35%", "100%+", "10-15%",
        "AI product revenue triple-digit growth for 9 consecutive quarters; RMB120B capex over 4 quarters; Qwen models driving adoption"
    ],
    [
        "Huawei Cloud", 3.0, 0.22,
        2.5, 0.12, 0.5, 0.80,
        0.167, 10, 12,
        "20-25%", "60-80%", "10-15%",
        "Ascend AI chip ecosystem; 40% CAGR in APAC over 5 years; partner revenue +50%; total Huawei growth cooled to 2%"
    ],
    [
        "Tencent Cloud", 2.0, 0.15,
        1.7, 0.08, 0.3, 0.55,
        0.150, 7, 8,
        "18-22%", "50-60%", "8-10%",
        "FinTech & Business Services segment +10%; HunYuan foundation model; AI-powered ad targeting driving indirect cloud growth"
    ],
    [
        "IBM Cloud", 1.5, 0.05,
        1.4, 0.03, 0.1, 0.35,
        0.067, 2, 3,
        "5-8%", "30-40%", "3-5%",
        "Hybrid cloud focus (Red Hat); watsonx AI platform; smaller scale vs hyperscalers; enterprise consulting led"
    ],
]

for i, row in enumerate(gr_data, 6):
    for j, val in enumerate(row, 1):
        cell = ws_growth.cell(row=i, column=j, value=val)
        if j in (2, 4, 6):
            cell.number_format = CURRENCY_FMT
        if j in (3, 5, 7, 8):
            cell.number_format = PCT_FMT
        if j in (9, 10):
            cell.number_format = '#,##0'

style_data_rows(ws_growth, 6, 6 + len(gr_data) - 1, len(gr_headers))

# Section B: Multi-Quarter Trend
row_trend = 6 + len(gr_data) + 2
ws_growth.cell(row=row_trend, column=1,
               value="B. Quarterly Revenue Growth Trajectory (YoY %) — Total, Traditional, and AI").font = SUBTITLE_FONT

qt_headers = [
    "Provider", "Metric",
    "Q1 2024", "Q2 2024", "Q3 2024", "Q4 2024",
    "Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025",
    "Trend Direction"
]
for i, h in enumerate(qt_headers, 1):
    ws_growth.cell(row=row_trend + 1, column=i, value=h)
style_header_row(ws_growth, row_trend + 1, len(qt_headers))

qt_data = [
    ["AWS", "Total Growth", 0.17, 0.19, 0.19, 0.19, 0.17, 0.19, 0.22, 0.24, "Accelerating"],
    ["AWS", "Traditional Growth", 0.14, 0.15, 0.15, 0.16, 0.14, 0.15, 0.16, 0.17, "Steady"],
    ["AWS", "AI Growth", 0.80, 0.85, 0.90, 0.95, 0.90, 1.00, 1.05, 1.10, "Accelerating"],
    ["Azure", "Total Growth", 0.31, 0.29, 0.33, 0.33, 0.35, 0.40, 0.40, 0.39, "Plateau ~39-40%"],
    ["Azure", "Traditional Growth", 0.18, 0.16, 0.17, 0.17, 0.16, 0.17, 0.16, 0.14, "Gradual deceleration"],
    ["Azure", "AI Growth", 0.55, 0.60, 0.70, 0.75, 0.80, 0.90, 0.95, 0.95, "Maturing from high base"],
    ["Google Cloud", "Total Growth", 0.28, 0.29, 0.35, 0.30, 0.28, 0.29, 0.35, 0.48, "Sharp acceleration"],
    ["Google Cloud", "Traditional Growth", 0.22, 0.22, 0.24, 0.22, 0.20, 0.21, 0.24, 0.28, "Steady improvement"],
    ["Google Cloud", "AI Growth", 0.80, 0.90, 1.10, 1.00, 1.00, 1.10, 1.30, 1.50, "Surging (Gemini effect)"],
    ["Oracle OCI", "Total Growth", 0.25, 0.20, 0.21, 0.25, 0.26, 0.32, 0.52, 0.68, "Strong acceleration"],
    ["Oracle OCI", "Traditional Growth", 0.10, 0.09, 0.10, 0.11, 0.10, 0.11, 0.12, 0.11, "Stable"],
    ["Oracle OCI", "AI Growth", 0.60, 0.55, 0.60, 0.70, 0.75, 0.85, 0.95, 1.05, "Accelerating sharply"],
    ["Alibaba Cloud", "Total Growth", 0.03, 0.06, 0.07, 0.03, 0.12, 0.18, 0.26, 0.34, "Rapid acceleration"],
    ["Alibaba Cloud", "Traditional Growth", 0.01, 0.03, 0.04, 0.01, 0.05, 0.08, 0.10, 0.12, "Recovering"],
    ["Alibaba Cloud", "AI Growth", 1.00, 1.00, 1.00, 1.00, 1.50, 1.80, 2.00, 2.00, "Triple-digit sustained"],
]

for i, row in enumerate(qt_data, row_trend + 2):
    for j, val in enumerate(row, 1):
        cell = ws_growth.cell(row=i, column=j, value=val)
        if 3 <= j <= 10:
            cell.number_format = '0%'

style_data_rows(ws_growth, row_trend + 2, row_trend + 2 + len(qt_data) - 1, len(qt_headers))

# Chart: stacked bar of Traditional + AI revenue by provider
from openpyxl.chart import BarChart as BarChart2, Reference as Ref2
bar_growth = BarChart2()
bar_growth.type = "col"
bar_growth.grouping = "stacked"
bar_growth.title = "Q4 2025 Revenue by Provider: Traditional vs AI ($B)"
bar_growth.style = 10
bar_growth.y_axis.title = "Revenue ($B)"

cats_g = Ref2(ws_growth, min_col=1, min_row=6, max_row=6 + len(gr_data) - 1)
d_trad = Ref2(ws_growth, min_col=4, min_row=5, max_row=6 + len(gr_data) - 1)
d_ai = Ref2(ws_growth, min_col=6, min_row=5, max_row=6 + len(gr_data) - 1)
bar_growth.add_data(d_trad, titles_from_data=True)
bar_growth.add_data(d_ai, titles_from_data=True)
bar_growth.set_categories(cats_g)
bar_growth.width = 24
bar_growth.height = 14

from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
s0 = bar_growth.series[0]
s0.graphicalProperties.solidFill = "4472C4"
s1 = bar_growth.series[1]
s1.graphicalProperties.solidFill = "ED7D31"

chart_row = row_trend + 2 + len(qt_data) + 2
ws_growth.add_chart(bar_growth, f"A{chart_row}")

auto_width(ws_growth, len(gr_headers), max_width=42)
ws_growth.freeze_panes = "B6"

# ─────────────────────────────────────────────────────────────────────
# SHEET 8: Margins by Workload Type & Vendor
# ─────────────────────────────────────────────────────────────────────
ws_margin = wb.create_sheet("Margins by Workload & Vendor")
ws_margin.sheet_properties.tabColor = "FF6600"

ws_margin.merge_cells("A1:K1")
ws_margin["A1"] = "Operating & Gross Margins by Workload Type Across Hyperscale Cloud Vendors"
ws_margin["A1"].font = TITLE_FONT

ws_margin.merge_cells("A2:K2")
ws_margin["A2"] = "Sources: Provider earnings, Wall Street analyst estimates, industry benchmarks | Margins are estimated ranges; providers do not fully disclose per-workload profitability"
ws_margin["A2"].font = Font(name="Calibri", size=9, italic=True, color="808080")

# Section A: Overall Segment Margins
ws_margin["A4"] = "A. Cloud Segment Operating Margins by Provider — Quarterly Trend"
ws_margin["A4"].font = SUBTITLE_FONT

seg_headers = [
    "Provider", "Segment Reported",
    "Q1 2024 Op. Margin", "Q2 2024 Op. Margin",
    "Q3 2024 Op. Margin", "Q4 2024 Op. Margin",
    "Q1 2025 Op. Margin", "Q2 2025 Op. Margin",
    "Q3 2025 Op. Margin", "Q4 2025 Op. Margin",
    "Trend / Commentary"
]
for i, h in enumerate(seg_headers, 1):
    ws_margin.cell(row=5, column=i, value=h)
style_header_row(ws_margin, 5, len(seg_headers))

seg_data = [
    ["AWS", "Amazon Web Services",
     0.378, 0.355, 0.382, 0.370,
     0.395, 0.329, 0.380, 0.350,
     "Ranged 33-40%; Q2 2025 dip from GPU depreciation charge; strong recovery in H2"],
    ["Microsoft", "Intelligent Cloud",
     0.445, 0.440, 0.438, 0.435,
     0.430, 0.425, 0.420, 0.418,
     "Gradual compression from AI infra scaling (CFO: GM% down from AI capex); op income still growing 27% YoY"],
    ["Google Cloud", "Google Cloud",
     0.092, 0.122, 0.170, 0.175,
     0.200, 0.210, 0.237, 0.299,
     "Dramatic expansion from near-zero in 2023 to ~30%; fastest margin improvement among top 3"],
    ["Oracle", "Cloud (IaaS + SaaS)",
     0.32, 0.33, 0.34, 0.36,
     0.37, 0.39, 0.40, 0.42,
     "Steady expansion; non-GAAP op margin ~42%; targeting 30-40% gross margin on AI datacenters over contract life"],
    ["Alibaba", "Cloud Intelligence Group",
     0.08, 0.07, 0.06, 0.05,
     0.06, 0.07, 0.08, 0.09,
     "Low single-digit margins; investing heavily in AI infra (RMB120B over 4Q); margin temporarily sacrificed for growth"],
    ["Huawei", "Cloud Business Unit (est.)",
     0.05, 0.05, 0.06, 0.06,
     0.06, 0.05, 0.05, 0.05,
     "Limited disclosure; estimated thin margins amid heavy R&D (20%+ of revenue); US sanctions constrain chip access"],
    ["Tencent", "FinTech & Biz Services (incl. Cloud)",
     0.35, 0.36, 0.37, 0.37,
     0.37, 0.38, 0.38, 0.38,
     "Blended with FinTech; overall segment stable ~37-38%; cloud-only margin likely lower (est. 15-20%)"],
    ["IBM", "Software + Infra (incl. Cloud)",
     0.22, 0.23, 0.24, 0.24,
     0.24, 0.25, 0.25, 0.25,
     "Hybrid cloud (Red Hat) drives higher-margin software; pure IaaS is lower margin; overall stable ~24-25%"],
]

for i, row in enumerate(seg_data, 6):
    for j, val in enumerate(row, 1):
        cell = ws_margin.cell(row=i, column=j, value=val)
        if 3 <= j <= 10:
            cell.number_format = '0.0%'

style_data_rows(ws_margin, 6, 6 + len(seg_data) - 1, len(seg_headers))

# Section B: Margins by Workload Type Per Vendor
row_wl_margin = 6 + len(seg_data) + 2
ws_margin.cell(row=row_wl_margin, column=1,
               value="B. Estimated Gross Margins by Workload Type & Vendor").font = SUBTITLE_FONT

ws_margin.cell(row=row_wl_margin + 1, column=1,
               value="Note: Providers do not disclose per-workload margins. Estimates are based on analyst reports, industry benchmarks, and workload economics."
               ).font = Font(name="Calibri", size=9, italic=True, color="808080")

wl_mg_headers = [
    "Workload Type",
    "Industry Avg Gross Margin",
    "AWS", "Azure", "Google Cloud",
    "Oracle Cloud", "Alibaba Cloud",
    "Margin Drivers"
]
for i, h in enumerate(wl_mg_headers, 1):
    ws_margin.cell(row=row_wl_margin + 2, column=i, value=h)
style_header_row(ws_margin, row_wl_margin + 2, len(wl_mg_headers))

wl_mg_data = [
    ["Traditional Compute (VMs)", "65-75%",
     "70-75%", "68-73%", "65-70%", "60-68%", "55-65%",
     "Mature, commoditized; custom silicon (Graviton, Ampere) lifts AWS/Azure margins; scale-driven"],
    ["Storage (Object / Block)", "75-85%",
     "80-85%", "78-83%", "75-80%", "72-78%", "70-75%",
     "High margin — low incremental cost per TB at scale; egress premiums add margin; sticky workloads"],
    ["Managed Databases", "70-80%",
     "75-80%", "73-78%", "72-77%", "75-82%", "65-72%",
     "High value-add over self-managed; Oracle premium from Autonomous DB; serverless tiers improve margins"],
    ["Networking & CDN", "78-88%",
     "82-88%", "80-85%", "78-83%", "75-80%", "72-78%",
     "Egress fees are high margin; CDN capital-intensive but profitable at scale; lock-in effect strong"],
    ["Analytics & Data Warehousing", "70-80%",
     "72-78%", "73-78%", "78-83%", "70-75%", "65-72%",
     "Query-based pricing (BigQuery) highly profitable; GCP strong with BigQuery; proprietary formats increase switching cost"],
    ["Containers & Serverless", "65-75%",
     "68-75%", "67-73%", "68-74%", "60-68%", "58-65%",
     "Control-plane fees add margin; compute underlying is commodity; managed K8s commands premium vs self-hosted"],
    ["Security & Identity", "75-85%",
     "78-85%", "80-86%", "76-82%", "72-78%", "68-74%",
     "High margin — software-defined; bundled with compliance requirements; low COGS after initial dev"],
    ["DevOps & CI/CD", "70-80%",
     "72-78%", "75-82%", "70-76%", "65-72%", "60-68%",
     "Pipeline-minute pricing; GitHub Actions/Azure DevOps high margin; artifact storage adds recurring revenue"],
    ["AI/ML — GPU Compute (Training)", "45-55%",
     "48-55%", "45-52%", "50-58%", "42-50%", "40-48%",
     "Capital-intensive (NVIDIA GPUs expensive); custom silicon (Trainium, TPU) improves margins 10-15pp; depreciation pressure"],
    ["AI/ML — GPU Compute (Inference)", "50-60%",
     "52-60%", "50-58%", "55-62%", "45-55%", "42-52%",
     "Lower power per query than training; optimized chips (Inferentia, TPU v5e) lift margins; demand far exceeds supply"],
    ["AI/ML — Model API (Token-based)", "55-65%",
     "58-65%", "60-68%", "58-65%", "50-58%", "48-55%",
     "Software-layer margin on top of GPU; Azure OpenAI premium pricing; token price decline offset by volume surge"],
    ["AI/ML — ML Platform & MLOps", "60-70%",
     "62-70%", "63-70%", "60-68%", "55-63%", "50-60%",
     "SageMaker/Azure ML/Vertex — platform fees; high margin on orchestration layer; compute billed separately"],
    ["AI/ML — Custom Silicon", "55-70%",
     "60-70%", "55-62%", "62-72%", "N/A", "40-50%",
     "AWS Trainium/Inferentia, Google TPU, Azure Maia; higher margin vs NVIDIA due to no GPU vendor margin layer"],
    ["SaaS / Business Apps", "70-82%",
     "N/A", "75-82%", "72-78%", "72-80%", "65-72%",
     "Subscription-based; high margin after initial dev; Oracle Fusion/NetSuite strong; Microsoft Dynamics premium"],
]

for i, row in enumerate(wl_mg_data, row_wl_margin + 3):
    for j, val in enumerate(row, 1):
        ws_margin.cell(row=i, column=j, value=val)

style_data_rows(ws_margin, row_wl_margin + 3, row_wl_margin + 3 + len(wl_mg_data) - 1, len(wl_mg_headers))

# Section C: AI vs Traditional Margin Summary
row_sum = row_wl_margin + 3 + len(wl_mg_data) + 2
ws_margin.cell(row=row_sum, column=1,
               value="C. AI vs Traditional Cloud — Margin Comparison Summary").font = SUBTITLE_FONT

sum_headers = [
    "Category", "Avg Gross Margin", "Avg Operating Margin",
    "Capital Intensity", "Margin Trend (2024→2026)",
    "Key Risk"
]
for i, h in enumerate(sum_headers, 1):
    ws_margin.cell(row=row_sum + 1, column=i, value=h)
style_header_row(ws_margin, row_sum + 1, len(sum_headers))

sum_data = [
    ["Traditional Cloud (Compute, Storage, DB, Net)",
     "70-85%", "30-45%", "Moderate (declining per unit)",
     "Stable to slightly expanding",
     "Commoditization; price competition from smaller providers"],
    ["AI — Training Workloads",
     "45-55%", "15-25%", "Very High (GPU clusters, liquid cooling)",
     "Compressing near-term; improving long-term with custom silicon",
     "GPU depreciation ($2.2B AWS charge); NVIDIA pricing power; rapid obsolescence"],
    ["AI — Inference Workloads",
     "50-62%", "20-30%", "High (but lower than training)",
     "Expanding as optimization improves",
     "Token price collapse (50x/yr median decline); volume must outpace price erosion"],
    ["AI — Model APIs & Platform",
     "55-70%", "25-40%", "Low-Moderate (software layer on GPU infra)",
     "Expanding as platforms mature",
     "Commoditization risk from open-source (Llama, DeepSeek); competition drives price to cost"],
    ["AI — Custom Silicon (TPU, Trainium)",
     "55-72%", "30-45%", "Very High (upfront R&D + fab)",
     "Expanding rapidly (eliminating NVIDIA margin layer)",
     "Execution risk; compatibility with CUDA ecosystem; R&D amortization"],
]

for i, row in enumerate(sum_data, row_sum + 2):
    for j, val in enumerate(row, 1):
        ws_margin.cell(row=i, column=j, value=val)

style_data_rows(ws_margin, row_sum + 2, row_sum + 2 + len(sum_data) - 1, len(sum_headers))

auto_width(ws_margin, max(len(seg_headers), len(wl_mg_headers)), max_width=48)
ws_margin.freeze_panes = "B6"

# ─────────────────────────────────────────────────────────────────────
# SHEET 9: Methodology & Sources
# ─────────────────────────────────────────────────────────────────────
ws9 = wb.create_sheet("Methodology & Sources")
ws9.sheet_properties.tabColor = "808080"

ws9.merge_cells("A1:C1")
ws9["A1"] = "Methodology & Data Sources"
ws9["A1"].font = TITLE_FONT

src_headers = ["#", "Source", "Data Used"]
for i, h in enumerate(src_headers, 1):
    ws9.cell(row=3, column=i, value=h)
style_header_row(ws9, 3, len(src_headers))

sources = [
    [1, "Gartner (Nov 2024)", "Public cloud spending forecast ($723B for 2025); IaaS/PaaS/SaaS segmentation"],
    [2, "Synergy Research Group", "Cloud market size ($330B in 2024); hyperscale data center count (~1,300)"],
    [3, "CRN / Channelnomics", "Q4 2025 cloud market share (AWS/Azure/GCP); quarterly revenue figures"],
    [4, "CloudZero Cloud Economics Pulse", "Monthly cloud spending composition; AI/ML share (2.44%); compute % trends"],
    [5, "Vantage Cloud Cost Report (Q1 2025)", "GPU spend share by provider; service category breakdown"],
    [6, "Flexera 2025 State of the Cloud", "Enterprise cloud challenges; workload distribution; repatriation rates"],
    [7, "AWS Q4 2025 Earnings (Feb 2026)", "Revenue: $35.6B; growth: 24%; op income: $12.5B; custom silicon >$10B ARR"],
    [8, "Microsoft FY26 Q2 Results (Jan 2026)", "Intelligent Cloud: $32.9B; Azure growth: 39%; AI contributes 22-26pp of Azure growth (CFO Amy Hood)"],
    [9, "Google Q4 2025 Earnings (Feb 2026)", "Google Cloud: $17.7B; growth: 48%; op margin ~30%; Gemini 10B+ tokens/min"],
    [10, "Oracle FY26 Q2 Results (Dec 2025)", "Cloud revenue: $8.0B; OCI +68%; RPO $523B; targeting 30-40% AI DC gross margin"],
    [11, "Alibaba Group Quarterly (Nov 2025)", "Cloud revenue +34%; AI triple-digit growth 9 consecutive quarters; RMB120B capex"],
    [12, "Huawei Annual Report 2025", "Total revenue >$122B; cloud partner biz +50%; Ascend AI chip ecosystem"],
    [13, "Tencent Q3 2025 Results", "FinTech & Biz Services +10%; HunYuan model; overall op margin 38%"],
    [14, "Omdia / Informa Tech", "Cloud marketplace forecast ($163B by 2030); workload category projections"],
    [15, "McKinsey (2025)", "AI workload bifurcation (training vs inference); power density trends"],
    [16, "Deloitte Insights", "AI token spend dynamics; enterprise AI cost patterns"],
    [17, "ByteIota / WifiTalents", "AI inference cost projections (55% of cloud by 2026); inference chip market"],
    [18, "TLDL / DevTk / BuildMVPFast", "LLM API pricing tables (Feb 2026); model-level token costs"],
    [19, "IDC Workload Forecast (2025-2029)", "Public cloud workload trends; IaaS/PaaS workload categorization (19 types)"],
    [20, "Wall Street Horizon / Tanay J / LongYield", "AI gross margin analysis (50-60% AI vs 77%+ traditional); GPU depreciation impact"],
    [21, "AInvest / FourWeekMBA / Futurum", "Azure AI growth decomposition; cloud segment margin trends; capex projections"],
]

for i, row in enumerate(sources, 4):
    for j, val in enumerate(row, 1):
        ws9.cell(row=i, column=j, value=val)

style_data_rows(ws9, 4, 4 + len(sources) - 1, len(src_headers))
auto_width(ws9, len(src_headers), max_width=90)

note_row = 4 + len(sources) + 2
ws9.cell(row=note_row, column=1, value="Methodology Notes").font = SUBTITLE_FONT
notes = [
    "1. Revenue breakdowns by workload category are estimates synthesized from provider earnings, analyst reports, and market sizing data.",
    "2. Where exact figures are unavailable, proportional estimates are derived from provider disclosures and third-party research.",
    "3. AI token consumption averages are based on published enterprise usage patterns, API rate limits, and cost benchmarks.",
    "4. Regional market sizes are estimated from Gartner global totals using regional share data from Synergy and IDC.",
    "5. All monetary values are in USD. Growth rates are year-over-year unless otherwise stated.",
    "6. Data reflects Q4 2025 actuals and early 2026 pricing/forecasts as of February 2026.",
    "7. Per-workload margins are analyst estimates — providers do not disclose workload-level profitability.",
    "8. AI growth contribution in percentage points derived from Microsoft CFO disclosure (22-26pp) and modeled for other providers.",
]
for i, note in enumerate(notes):
    ws9.cell(row=note_row + 1 + i, column=1, value=note).font = Font(name="Calibri", size=10, italic=True)

# ─────────────────────────────────────────────────────────────────────
# Add Charts
# ─────────────────────────────────────────────────────────────────────

# Pie chart on Provider Overview
pie = PieChart()
pie.title = "Q4 2025 Revenue by Provider"
pie.style = 10
labels = Reference(ws2, min_col=1, min_row=4, max_row=4 + len(prov_data) - 1)
data = Reference(ws2, min_col=3, min_row=3, max_row=4 + len(prov_data) - 1)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.width = 18
pie.height = 12
ws2.add_chart(pie, "A14")

# Bar chart on Monetization
bar = BarChart()
bar.type = "col"
bar.title = "Workload Revenue: 2024 vs 2025 ($B)"
bar.style = 10
bar.y_axis.title = "Revenue ($B)"
bar.x_axis.title = "Workload Category"
cats = Reference(ws5, min_col=1, min_row=5, max_row=5 + len(mon_data) - 1)
d1 = Reference(ws5, min_col=2, min_row=4, max_row=5 + len(mon_data) - 1)
d2 = Reference(ws5, min_col=3, min_row=4, max_row=5 + len(mon_data) - 1)
bar.add_data(d1, titles_from_data=True)
bar.add_data(d2, titles_from_data=True)
bar.set_categories(cats)
bar.shape = 4
bar.width = 28
bar.height = 14
ws5.add_chart(bar, f"A{total_row + 1}")

# ─────────────────────────────────────────────────────────────────────
# Freeze Panes
# ─────────────────────────────────────────────────────────────────────
ws1.freeze_panes = "A5"
ws2.freeze_panes = "A4"
ws3.freeze_panes = "A4"
ws4.freeze_panes = "A5"
ws5.freeze_panes = "A5"
ws6.freeze_panes = "A4"

output = "/workspace/Hyperscale_Cloud_Workload_Aggregation.xlsx"
wb.save(output)
print(f"Excel file saved to: {output}")
