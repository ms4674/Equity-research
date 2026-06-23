#!/usr/bin/env python3
"""
Build an aggregated spreadsheet of LLM token-pricing indices and indicators.

This script is the single source of truth for the dataset. It writes:
  - LLM_Token_Pricing_Indices.xlsx  (multi-sheet, formatted workbook)
  - data/indicators.csv             (the main comparison matrix)
  - data/silicon_data_products.csv  (Silicon Data product suite detail)
  - data/methodology.csv            (methodology / construction comparison)
  - data/sources.csv                (reference links)

Run:  python3 build_spreadsheet.py

Data was hand-compiled from public product/documentation pages of each
publisher (see data/sources.csv). Figures such as "models tracked" and current
index values reflect the publishers' public materials as of the as-of date
below and are point-in-time references, not live values. Where a provider
gates data behind a paid API or login, that is noted in the "Access" column.
"""

from __future__ import annotations

import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

AS_OF = "2026-06-23"

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(HERE, "data")
XLSX_PATH = os.path.join(HERE, "LLM_Token_Pricing_Indices.xlsx")

# ---------------------------------------------------------------------------
# Sheet 1: Indicators - the main comparison matrix
# ---------------------------------------------------------------------------

INDICATORS_HEADERS = [
    "Indicator / Index",
    "Publisher",
    "Ticker / ID",
    "Type",
    "What it measures",
    "Unit",
    "Blend ratio (in:out / cache)",
    "Aggregation method",
    "Weighting",
    "Model scope",
    "Coverage",
    "Update frequency",
    "History from",
    "Access",
    "Data source",
    "URL",
]

INDICATORS = [
    [
        "LLM Token Expenditure Index",
        "Silicon Data",
        "SDLLMTK",
        "Index",
        "Usage/expenditure-weighted average realized cost-to-serve of LLM "
        "inference; a gauge of marginal willingness-to-pay (frontier vs "
        "open-weight migration)",
        "USD per 1M tokens (normalized blended rate)",
        "Normalized for input/output mix, context window, batching, reliability",
        "Usage/expenditure-weighted blended reading",
        "Weighted by where usage is concentrated and sustained",
        "Frontier API + open-weight + brokered dedicated-instance + self-hosted",
        "400+ models tracked, 20+ in daily basket, 20+ price/volume sources "
        "(~90%+ of global LLM inference spend)",
        "Daily (every trading day)",
        "2025-12-01 (API min date)",
        "Paid (Portal subscription / Token Index API; 7-day trial)",
        "Proprietary multi-source observations",
        "https://www.silicondata.com/products/silicon-index/llm-token-expenditure-index",
    ],
    [
        "Token Marketpulse",
        "Silicon Data",
        "-",
        "Tracker",
        "Per-model pricing, adoption and momentum across the LLM market "
        "(Cost, Adoption, Momentum panels)",
        "USD per 1M tokens (8 normalized metrics)",
        "Normalized for apples-to-apples comparison",
        "Per-model (not a single index)",
        "n/a (per-model tracker)",
        "Closed, open-weight and open-source models",
        "142+ models",
        "Daily",
        "n/a",
        "Paid (request access)",
        "Proprietary multi-source observations",
        "https://www.silicondata.com/products/token-marketpulse",
    ],
    [
        "Token Pricebook",
        "Silicon Data",
        "-",
        "Dataset / API",
        "Standardized model-level input & output token prices plus model "
        "metadata (params, context length)",
        "USD per 1M tokens (input & output, separate)",
        "Separate input / output (not blended)",
        "Per-model rows",
        "n/a (catalog)",
        "open-source / open-weight / closed-source",
        "Leading developers & models across model-labs, platforms, marketplaces",
        "Daily collection",
        "2025-08 (example rows)",
        "Paid (API)",
        "Multi-source: model-lab, model-platform, marketplace",
        "https://docs.silicondata.com/products/token_pricebook",
    ],
    [
        "AA Intelligence Index (price series)",
        "Artificial Analysis",
        "-",
        "Benchmark + price",
        "Blended price per token reported alongside a composite intelligence "
        "benchmark; plus Cost-per-Task for agentic workloads",
        "USD per 1M tokens; also USD per task",
        "7:2:1 cache:input:output (default; adjustable)",
        "Arithmetic / per-model; intelligence-band aggregates over time",
        "Per-model; banded by Intelligence Index for trend charts",
        "Proprietary + open-weight (frontier focus)",
        "500+ models evaluated",
        "Continuous / per release",
        "Multi-year trend series",
        "Free (web) + paid API/data access",
        "Provider APIs, live measurements",
        "https://artificialanalysis.ai/trends",
    ],
    [
        "LLM Stats Score (with pricing)",
        "LLM-Stats",
        "-",
        "Leaderboard + price",
        "Composite leaderboard blending verified benchmarks, live performance "
        "and per-token pricing",
        "USD per 1M tokens (blended) + composite score",
        "Blended price per 1M tokens",
        "Composite weighting of benchmarks + price",
        "Per-model",
        "Proprietary + open-weight",
        "300-321+ canonical models",
        "Pricing/metadata revalidate hourly; perf 7-day rolling",
        "Continuous",
        "Free (web)",
        "Provider price lists + LLM Stats proxy billing samples",
        "https://llm-stats.com/",
    ],
    [
        "AI Cost Index (Frontier & Budget)",
        "MyTokenTracker",
        "-",
        "Index",
        "Blended price of fixed baskets of models to track token economics "
        "over time (Frontier Index, Budget Index)",
        "USD per 1M tokens",
        "3:1 input:output ((3*in + 1*out)/4)",
        "Equal-weighted average of basket constituents",
        "Fixed, named baskets (so series stays comparable)",
        "Proprietary + open-weight (basket-defined)",
        "Fixed baskets (Frontier / Budget)",
        "Daily",
        "~2026-06 (series grows daily)",
        "Free (web + open-data archive)",
        "LiteLLM open dataset",
        "https://mytokentracker.io/cost-index",
    ],
    [
        "Pulse Inference Token Index",
        "Pulse (pulsebenchmarks.com)",
        "-",
        "Index",
        "Posted per-token prices for open-weight inference at commodity hosts; "
        "records what the market charges and how the distribution moves",
        "USD per 1M tokens (median; also P90, min, dispersion)",
        "3:1 input:output (headline); input-only & output-only as peers",
        "Breadth-weighted median across eligible endpoints",
        "Breadth-weighted; per model/quantization/jurisdiction series",
        "Open-weight only (proprietary APIs excluded by design)",
        "Per-model series; anchor = Llama 3.3 70B FP8 US",
        "Weekly",
        "v1.0 (provisional)",
        "Free (web + /api/indices feed)",
        "Commodity-host endpoint posted prices",
        "https://pulsebenchmarks.com/indices/inference-token-index/",
    ],
    [
        "Global Compute Price Index (GCPI)",
        "GCPI Research",
        "GCPI",
        "Index",
        "Publicly posted synchronous inference API prices, standardized to a "
        "single unit; tracks market price level over time",
        "USD per 1M input tokens (2026-Q2 = 100)",
        "Input price only (single standardized unit)",
        "Weighted geometric mean of constituents",
        "Weights proxy estimated market share (rebalanced quarterly)",
        "Synchronous token-priced inference APIs",
        "8 providers (Tier S, Llama 3.1 8B); reconstructed 2024-Q3+",
        "Quarterly reconstruction",
        "2024-Q3 (reconstructed)",
        "Free (web / research)",
        "Public provider price pages + analyst estimates",
        "https://globalcomputeindex.org/",
    ],
    [
        "AI CPI & Budget Index",
        "AIscending (llm-pricing-index)",
        "-",
        "Index / open dataset",
        "Cost pressure index (weighted avg cost) and Budget Index (efficiency "
        "vs frontier price ratio) across tracked models",
        "USD per 1M tokens (prompt, completion, blended)",
        "75% input / 25% output (blended)",
        "AI CPI: avg over frontier+reasoning; Budget: efficiency+open source",
        "Category baskets (Frontier/Efficiency/Reasoning/Open Source)",
        "Proprietary + open-source",
        "22 models, 4 categories",
        "Monthly (1st of month)",
        "2026-04",
        "Free (open data: JSON/CSV on GitHub)",
        "OpenRouter API (/api/v1/models)",
        "https://github.com/AIscending/llm-pricing-index",
    ],
    [
        "LLMRates pricing tracker",
        "LLMRates",
        "-",
        "Tracker / API",
        "Multi-source verified token pricing with full history and real-time "
        "change feed",
        "USD per 1M tokens (input & output)",
        "Separate input / output",
        "Per-model; 2-source agreement before publishing",
        "Source-reconciled, confidence-scored",
        "Proprietary + open-weight",
        "~2,330 models, 6 sources, <60s latency",
        "Daily to every 6h (per source); <60s change feed",
        "Live (with trust metadata)",
        "Freemium API",
        "OpenAI, Anthropic, Google, OpenRouter, LiteLLM, Hugging Face",
        "https://llmrates.live/",
    ],
    [
        "BenchGecko llm-pricing",
        "BenchGecko",
        "-",
        "Open dataset",
        "Canonical standardized pricing for 300+ models in one JSON file",
        "USD per 1M tokens (input & output)",
        "Separate input / output",
        "Per-model catalog",
        "n/a (catalog)",
        "Proprietary + open-weight",
        "346 models",
        "Weekly",
        "git history",
        "Free (open data: JSON / npm)",
        "OpenRouter API",
        "https://github.com/BenchGecko/llm-pricing",
    ],
    [
        "OpenRouter model pricing",
        "OpenRouter",
        "-",
        "Source feed",
        "Aggregated pass-through per-token pricing across providers; the de-"
        "facto upstream source for many indices",
        "USD per 1M tokens (input & output)",
        "Separate input / output",
        "Per-model pass-through",
        "n/a (catalog)",
        "Proprietary + open-weight",
        "300-400+ models",
        "Continuous",
        "Live",
        "Free (public API)",
        "Provider APIs",
        "https://openrouter.ai/api/v1/models",
    ],
    [
        "LiteLLM model price map",
        "LiteLLM (BerriAI)",
        "-",
        "Open dataset",
        "Open JSON map of per-token prices + context windows used as an "
        "upstream source by several indices",
        "USD per 1M tokens (input & output)",
        "Separate input / output",
        "Per-model catalog",
        "n/a (catalog)",
        "Proprietary + open-weight",
        "Hundreds of models",
        "Continuous (community-maintained)",
        "git history",
        "Free (open data)",
        "Provider docs / community",
        "https://github.com/BerriAI/litellm",
    ],
]

# ---------------------------------------------------------------------------
# Sheet 2: Silicon Data product suite detail
# ---------------------------------------------------------------------------

SILICON_HEADERS = [
    "Product",
    "Ticker / Endpoint",
    "Purpose",
    "Output / Unit",
    "Granularity",
    "Coverage",
    "Frequency",
    "Access",
    "Notes",
    "URL",
]

SILICON = [
    [
        "LLM Token Expenditure Index",
        "SDLLMTK",
        "Headline benchmark for LLM inference token pricing (marginal "
        "willingness-to-pay)",
        "USD per 1M tokens (normalized blended)",
        "Single blended index value",
        "400+ tracked, 20+ daily basket, 20+ sources (~90%+ of spend)",
        "Daily (trading days)",
        "Portal subscription / API; 7-day trial",
        "Despite the name it is an expenditure/usage-weighted price index; "
        "can be read as a 'quality premium' of frontier over open-weight",
        "https://www.silicondata.com/products/silicon-index/llm-token-expenditure-index",
    ],
    [
        "Token Index API",
        "POST /api/token-index/index",
        "Programmatic access to daily index values for a token, version and "
        "date range",
        "JSON: {date: index_value}; missing dates returned as '-1'",
        "Daily values over requested range",
        "token=expenditure (v1)",
        "Daily",
        "Paid API",
        "starting_date must be >= 2025-12-01; ending_date <= server date",
        "https://docs.silicondata.com/api-reference/token_index_api",
    ],
    [
        "Token Marketpulse",
        "-",
        "LLM token market usage & pricing tracker for FinOps / analysts "
        "(Cost, Adoption, Momentum)",
        "8 normalized metrics, USD per 1M tokens",
        "Per-model",
        "142+ models (closed, open-weight, open-source)",
        "Daily",
        "Request access",
        "Built for benchmarking vendor pricing and tracking price/adoption drift",
        "https://www.silicondata.com/products/token-marketpulse",
    ],
    [
        "Token Pricebook",
        "-",
        "Standardized model-level input/output token pricing + model metadata",
        "USD per 1M tokens (input & output) + params/context",
        "Per-model rows with source attribution",
        "Leading developers/models across labs, platforms, marketplaces",
        "Daily collection",
        "Paid API",
        "Fields: developer, model_name, model_type, params, context_length, "
        "price_input, price_output, source_id, source_type",
        "https://docs.silicondata.com/products/token_pricebook",
    ],
    [
        "Silicon Index Dashboard / Portal",
        "-",
        "Daily GPU & LLM market intelligence for institutional use",
        "Index values + 90-day historical pricing",
        "Market-wide",
        "All GPU, CPU & accelerator indices + LLM indices",
        "Daily",
        "Subscription (7-day trial)",
        "Includes PriceIQ predictive modeling, NeoCloud vs Hyperscaler "
        "comparison, SiliconCarbon",
        "https://www.silicondata.com/",
    ],
]

# ---------------------------------------------------------------------------
# Sheet 3: Methodology comparison
# ---------------------------------------------------------------------------

METHOD_HEADERS = [
    "Indicator",
    "Single number or per-model?",
    "Aggregation",
    "Blend ratio",
    "Normalization",
    "Model scope",
    "Weighting basis",
    "Key distinction",
]

METHOD = [
    [
        "Silicon Data SDLLMTK",
        "Single blended index",
        "Usage/expenditure-weighted average",
        "Normalized in/out mix (not a fixed ratio)",
        "Input/output mix, context, batching, reliability -> cost-to-serve",
        "Frontier + open-weight + brokered + self-hosted",
        "Where usage is concentrated & sustained",
        "Measures realized expenditure / willingness-to-pay, not list price",
    ],
    [
        "Artificial Analysis",
        "Per-model (+ banded trends)",
        "Arithmetic blend; cost-per-task",
        "7:2:1 cache:input:output (default)",
        "Provider token counts + cache-hit measurements",
        "Frontier + open-weight",
        "n/a / band aggregates",
        "Pairs price with intelligence; cost-per-task for agents",
    ],
    [
        "MyTokenTracker AI Cost Index",
        "Two indices (Frontier, Budget)",
        "Equal-weighted average",
        "3:1 input:output",
        "Fixed ratio so it measures price, not usage mix",
        "Basket-defined",
        "Equal weight, fixed baskets",
        "Fixed named baskets; reconstructs history at price-change dates",
    ],
    [
        "Pulse Inference Token Index",
        "Family of per-model series (+ anchor)",
        "Breadth-weighted median (P90, min, dispersion)",
        "3:1 input:output (headline)",
        "Per model / quantization / jurisdiction; pinned checkpoints",
        "Open-weight only (commodity hosts)",
        "Breadth-weighted across eligible endpoints",
        "Price-of-posted-pricing, excludes single-vendor proprietary APIs",
    ],
    [
        "GCPI",
        "Single index (2026-Q2=100)",
        "Weighted geometric mean",
        "Input-only (single unit)",
        "Standardized to USD/1M input tokens",
        "Synchronous token-priced APIs",
        "Estimated market share (quarterly rebalance)",
        "Geometric mean to dampen 10x spreads; log-additive changes",
    ],
    [
        "AIscending AI CPI / Budget Index",
        "Two indices",
        "Weighted / category averages",
        "75% input / 25% output",
        "OpenRouter standardized rates",
        "Frontier/Efficiency/Reasoning/Open Source",
        "Category baskets",
        "Fully open data (JSON/CSV); monthly snapshots",
    ],
    [
        "LLM-Stats Score",
        "Composite score + per-model price",
        "Composite weighting",
        "Blended per 1M tokens",
        "Provider lists cross-checked vs proxy billing",
        "Frontier + open-weight",
        "Benchmark + price composite",
        "Quality+price in one rank; hourly price revalidation",
    ],
]

# ---------------------------------------------------------------------------
# Sheet 4: Sources
# ---------------------------------------------------------------------------

SOURCES_HEADERS = ["#", "Publisher", "Resource", "URL"]

SOURCES = [
    ["1", "Silicon Data", "LLM Token Expenditure Index (SDLLMTK) product page",
     "https://www.silicondata.com/products/silicon-index/llm-token-expenditure-index"],
    ["2", "Silicon Data", "Token Index API reference",
     "https://docs.silicondata.com/api-reference/token_index_api"],
    ["3", "Silicon Data", "Token Marketpulse",
     "https://www.silicondata.com/products/token-marketpulse"],
    ["4", "Silicon Data", "Token Pricebook docs",
     "https://docs.silicondata.com/products/token_pricebook"],
    ["5", "Artificial Analysis", "Intelligence Index + price trends",
     "https://artificialanalysis.ai/trends"],
    ["6", "Artificial Analysis", "Intelligence Index v4.1 (cost-per-task)",
     "https://artificialanalysis.ai/articles/artificial-analysis-intelligence-index-v4-1"],
    ["7", "LLM-Stats", "AI leaderboard (score + pricing)",
     "https://llm-stats.com/"],
    ["8", "MyTokenTracker", "AI Cost Index",
     "https://mytokentracker.io/cost-index"],
    ["9", "Pulse", "Inference Token Index + methodology v1.0",
     "https://pulsebenchmarks.com/indices/inference-token-index/"],
    ["10", "GCPI Research", "Global Compute Price Index + methodology",
     "https://globalcomputeindex.org/"],
    ["11", "AIscending", "llm-pricing-index (open dataset)",
     "https://github.com/AIscending/llm-pricing-index"],
    ["12", "LLMRates", "Token pricing tracker / API",
     "https://llmrates.live/"],
    ["13", "BenchGecko", "llm-pricing (open dataset)",
     "https://github.com/BenchGecko/llm-pricing"],
    ["14", "OpenRouter", "Models pricing API",
     "https://openrouter.ai/api/v1/models"],
    ["15", "LiteLLM (BerriAI)", "Model price map",
     "https://github.com/BerriAI/litellm"],
]

SHEETS = {
    "Indicators": (INDICATORS_HEADERS, INDICATORS),
    "Silicon Data Suite": (SILICON_HEADERS, SILICON),
    "Methodology": (METHOD_HEADERS, METHOD),
    "Sources": (SOURCES_HEADERS, SOURCES),
}

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
SUB_FONT = Font(italic=True, size=9, color="595959")
WRAP = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Per-sheet column widths (chars).
WIDTHS = {
    "Indicators": [26, 20, 12, 16, 40, 24, 26, 26, 26, 22, 30, 18, 14, 22, 26, 40],
    "Silicon Data Suite": [26, 24, 38, 30, 24, 34, 14, 22, 40, 40],
    "Methodology": [26, 24, 26, 22, 34, 26, 26, 38],
    "Sources": [5, 22, 44, 50],
}


def style_sheet(ws, headers, rows, widths):
    ws.sheet_view.showGridLines = False

    # Title block (rows 1-2), table starts at row 4.
    ws.cell(row=1, column=1, value="LLM Token Pricing Indices & Indicators").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Aggregated reference - as of {AS_OF}. "
            f"Point-in-time figures from public publisher materials.").font = SUB_FONT

    header_row = 4
    for c, name in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER

    for r, row in enumerate(rows, start=header_row + 1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = WRAP
            cell.border = BORDER
            if (r - header_row) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="EDF1F7")
            # Hyperlink URL-looking cells.
            if isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")

    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.row_dimensions[header_row].height = 30


def build_xlsx():
    wb = Workbook()
    wb.remove(wb.active)
    for name, (headers, rows) in SHEETS.items():
        ws = wb.create_sheet(title=name)
        style_sheet(ws, headers, rows, WIDTHS[name])
    wb.save(XLSX_PATH)
    print(f"Wrote {XLSX_PATH}")


def build_csvs():
    os.makedirs(DATA_DIR, exist_ok=True)
    csv_map = {
        "indicators.csv": (INDICATORS_HEADERS, INDICATORS),
        "silicon_data_products.csv": (SILICON_HEADERS, SILICON),
        "methodology.csv": (METHOD_HEADERS, METHOD),
        "sources.csv": (SOURCES_HEADERS, SOURCES),
    }
    for fname, (headers, rows) in csv_map.items():
        path = os.path.join(DATA_DIR, fname)
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(headers)
            w.writerows(rows)
        print(f"Wrote {path}")


if __name__ == "__main__":
    build_csvs()
    build_xlsx()
    print("Done.")
