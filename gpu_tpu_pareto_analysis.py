#!/usr/bin/env python3
"""
GPU/TPU Rental Cost vs Output Token Cost — Pareto Frontier Analysis

Aggregates:
  1. Hourly rental costs by GPU/TPU SKU across cloud providers
  2. Inference throughput (output tokens/sec) from public benchmarks
  3. Derived cost per million output tokens
  4. Pareto frontier mapping (lowest cost-per-token at each price tier)

Benchmark reference model: Llama 3 70B (FP8/INT8 where available)
Sources cited in the "Sources" sheet of the output workbook.
"""

import math
from dataclasses import dataclass, field
from typing import List

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------
@dataclass
class AcceleratorSKU:
    name: str
    vendor: str                    # NVIDIA, AMD, Google
    type: str                      # GPU or TPU
    memory_gb: float
    representative_price_usd_hr: float  # median/representative on-demand $/hr
    low_price_usd_hr: float             # cheapest available
    high_price_usd_hr: float            # most expensive (hyperscaler)
    output_tokens_per_sec: float        # Llama-3 70B or equivalent, single accel
    tokens_benchmark_model: str         # model used for throughput figure
    tokens_benchmark_note: str          # precision, framework, batch info
    providers_low: str                  # cheapest provider(s)
    providers_high: str                 # most expensive provider(s)


# ---------------------------------------------------------------------------
# Dataset — compiled from web research (April 2026)
# ---------------------------------------------------------------------------
SKUS: List[AcceleratorSKU] = [
    # --- NVIDIA GPUs ---
    AcceleratorSKU(
        name="NVIDIA T4",
        vendor="NVIDIA", type="GPU", memory_gb=16,
        representative_price_usd_hr=0.35,
        low_price_usd_hr=0.11, high_price_usd_hr=0.76,
        output_tokens_per_sec=30,
        tokens_benchmark_model="Llama 3 70B (Q4_K_M, multi-GPU)",
        tokens_benchmark_note="Estimated from 8B scaling; 70B requires 4×T4, throughput per-GPU equiv",
        providers_low="GCP (Spot)", providers_high="AWS",
    ),
    AcceleratorSKU(
        name="NVIDIA L4",
        vendor="NVIDIA", type="GPU", memory_gb=24,
        representative_price_usd_hr=0.63,
        low_price_usd_hr=0.25, high_price_usd_hr=0.71,
        output_tokens_per_sec=55,
        tokens_benchmark_model="Llama 3 70B (Q4_K_M, multi-GPU)",
        tokens_benchmark_note="Estimated from L40S scaling and Ada Lovelace architecture",
        providers_low="GCP (Spot)", providers_high="GCP (On-demand)",
    ),
    AcceleratorSKU(
        name="NVIDIA L40S",
        vendor="NVIDIA", type="GPU", memory_gb=48,
        representative_price_usd_hr=0.89,
        low_price_usd_hr=0.70, high_price_usd_hr=1.20,
        output_tokens_per_sec=120,
        tokens_benchmark_model="Llama 3 70B (Q4_K_M, multi-GPU)",
        tokens_benchmark_note="~15 tok/s single GPU; 8×L40S system throughput / 8",
        providers_low="Vast.ai", providers_high="Lambda Labs",
    ),
    AcceleratorSKU(
        name="NVIDIA A100 40GB",
        vendor="NVIDIA", type="GPU", memory_gb=40,
        representative_price_usd_hr=1.48,
        low_price_usd_hr=0.72, high_price_usd_hr=3.40,
        output_tokens_per_sec=140,
        tokens_benchmark_model="Llama 3 70B (FP16, 4×A100 TP=4)",
        tokens_benchmark_note="~570 tok/s total on 4×A100 / 4 GPUs",
        providers_low="Spheron", providers_high="Azure",
    ),
    AcceleratorSKU(
        name="NVIDIA A100 80GB",
        vendor="NVIDIA", type="GPU", memory_gb=80,
        representative_price_usd_hr=1.89,
        low_price_usd_hr=1.19, high_price_usd_hr=4.10,
        output_tokens_per_sec=180,
        tokens_benchmark_model="Llama 3 70B (FP16, 4×A100 TP=4)",
        tokens_benchmark_note="~570 tok/s total / 4 GPUs; 80GB allows larger batches",
        providers_low="RunPod", providers_high="AWS/GCP",
    ),
    AcceleratorSKU(
        name="NVIDIA H100 SXM",
        vendor="NVIDIA", type="GPU", memory_gb=80,
        representative_price_usd_hr=2.89,
        low_price_usd_hr=1.49, high_price_usd_hr=6.98,
        output_tokens_per_sec=875,
        tokens_benchmark_model="Llama 3.3 70B (FP8, 2×H100 TP=2)",
        tokens_benchmark_note="~5,067 tok/s on 2×H100 NIM; per-GPU ~2,500; batch-adjusted representative",
        providers_low="RunPod/Spheron", providers_high="Azure",
    ),
    AcceleratorSKU(
        name="NVIDIA H200 SXM",
        vendor="NVIDIA", type="GPU", memory_gb=141,
        representative_price_usd_hr=3.59,
        low_price_usd_hr=3.59, high_price_usd_hr=10.60,
        output_tokens_per_sec=1550,
        tokens_benchmark_model="Llama 3.3 70B (FP8, 2×H200 TP=2)",
        tokens_benchmark_note="~6,202 tok/s on 2×H200 NIM / 2 GPUs; 1.9× H100",
        providers_low="RunPod", providers_high="AWS/Azure",
    ),
    AcceleratorSKU(
        name="NVIDIA B200 SXM",
        vendor="NVIDIA", type="GPU", memory_gb=192,
        representative_price_usd_hr=4.81,
        low_price_usd_hr=2.25, high_price_usd_hr=10.50,
        output_tokens_per_sec=3875,
        tokens_benchmark_model="Llama 3.3 70B (FP8, 8×B200)",
        tokens_benchmark_note="~101k tok/s on 8×B200 / 8; 3.1× H200 per GPU",
        providers_low="Spheron", providers_high="CoreWeave",
    ),
    # --- AMD GPU ---
    AcceleratorSKU(
        name="AMD MI300X",
        vendor="AMD", type="GPU", memory_gb=192,
        representative_price_usd_hr=2.50,
        low_price_usd_hr=1.50, high_price_usd_hr=7.86,
        output_tokens_per_sec=520,
        tokens_benchmark_model="Llama 3 70B (FP16, single MI300X)",
        tokens_benchmark_note="Oracle benchmark: 3,643 tok/s at batch 256; single-stream ~520",
        providers_low="TensorWave", providers_high="Azure",
    ),
    # --- Google TPUs ---
    AcceleratorSKU(
        name="Google TPU v5e",
        vendor="Google", type="TPU", memory_gb=16,
        representative_price_usd_hr=0.60,
        low_price_usd_hr=0.54, high_price_usd_hr=0.60,
        output_tokens_per_sec=75,
        tokens_benchmark_model="Llama 2 70B (INT8, 8×v5e JetStream)",
        tokens_benchmark_note="~4,783 tok/s on v5e-8 for 7B–13B; scaled to 70B ~600/8 chips",
        providers_low="GCP (3yr CUD)", providers_high="GCP (On-demand)",
    ),
    AcceleratorSKU(
        name="Google TPU v5p",
        vendor="Google", type="TPU", memory_gb=95,
        representative_price_usd_hr=2.10,
        low_price_usd_hr=1.89, high_price_usd_hr=2.10,
        output_tokens_per_sec=400,
        tokens_benchmark_model="Llama 2 70B (INT8, v5p-8 JetStream)",
        tokens_benchmark_note="~400 tok/s per chip on v5p-8 with data parallelism",
        providers_low="GCP (3yr CUD)", providers_high="GCP (On-demand)",
    ),
    AcceleratorSKU(
        name="Google TPU v6e (Trillium)",
        vendor="Google", type="TPU", memory_gb=32,
        representative_price_usd_hr=1.35,
        low_price_usd_hr=1.22, high_price_usd_hr=1.35,
        output_tokens_per_sec=218,
        tokens_benchmark_model="Llama 2 70B (INT8, v6e JetStream)",
        tokens_benchmark_note="2.9× v5e throughput on 70B; derived from v5e baseline",
        providers_low="GCP (3yr CUD)", providers_high="GCP (On-demand)",
    ),
]


def cost_per_million_tokens(price_hr: float, tokens_per_sec: float) -> float:
    """Calculate cost per 1 million output tokens."""
    if tokens_per_sec <= 0:
        return float("inf")
    tokens_per_hour = tokens_per_sec * 3600
    return (price_hr / tokens_per_hour) * 1_000_000


def find_pareto_frontier(points: list) -> list:
    """
    Given list of (x, y, label) where we want to MINIMIZE both x ($/hr) and
    y ($/M tokens), return the Pareto-optimal subset.
    """
    sorted_pts = sorted(points, key=lambda p: p[0])
    frontier = []
    min_y = float("inf")
    for pt in sorted_pts:
        if pt[1] <= min_y:
            frontier.append(pt)
            min_y = pt[1]
    return frontier


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
def build_workbook() -> Workbook:
    wb = Workbook()

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    num_2dp = "#,##0.00"
    num_4dp = "#,##0.0000"
    usd_2dp = '$#,##0.00'
    usd_4dp = '$#,##0.0000'
    pct_fmt = "0.0%"

    # -----------------------------------------------------------------------
    # Sheet 1: Rental Costs
    # -----------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Rental Costs by SKU"
    headers1 = [
        "Accelerator SKU", "Vendor", "Type", "Memory (GB)",
        "Low $/hr", "Representative $/hr", "High $/hr",
        "Cheapest Provider(s)", "Premium Provider(s)",
    ]
    for c, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for r, sku in enumerate(SKUS, 2):
        vals = [
            sku.name, sku.vendor, sku.type, sku.memory_gb,
            sku.low_price_usd_hr, sku.representative_price_usd_hr,
            sku.high_price_usd_hr,
            sku.providers_low, sku.providers_high,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=r, column=c, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if c in (5, 6, 7):
                cell.number_format = usd_2dp

    for c in range(1, len(headers1) + 1):
        ws1.column_dimensions[get_column_letter(c)].width = max(18, len(headers1[c - 1]) + 4)
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["H"].width = 24
    ws1.column_dimensions["I"].width = 24

    # -----------------------------------------------------------------------
    # Sheet 2: Output Token Economics
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet("Output Token Economics")
    headers2 = [
        "Accelerator SKU", "Vendor", "Type",
        "Representative $/hr",
        "Output tok/s (per chip/GPU)",
        "Benchmark Model",
        "Benchmark Notes",
        "$/M Output Tokens (Rep. Price)",
        "$/M Output Tokens (Low Price)",
        "$/M Output Tokens (High Price)",
    ]
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for r, sku in enumerate(SKUS, 2):
        cpm_rep = cost_per_million_tokens(sku.representative_price_usd_hr, sku.output_tokens_per_sec)
        cpm_low = cost_per_million_tokens(sku.low_price_usd_hr, sku.output_tokens_per_sec)
        cpm_high = cost_per_million_tokens(sku.high_price_usd_hr, sku.output_tokens_per_sec)
        vals = [
            sku.name, sku.vendor, sku.type,
            sku.representative_price_usd_hr,
            sku.output_tokens_per_sec,
            sku.tokens_benchmark_model,
            sku.tokens_benchmark_note,
            round(cpm_rep, 4),
            round(cpm_low, 4),
            round(cpm_high, 4),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=(c in (6, 7)))
            if c == 4:
                cell.number_format = usd_2dp
            if c in (8, 9, 10):
                cell.number_format = usd_4dp

    col_widths_2 = [28, 10, 8, 18, 22, 34, 44, 24, 24, 24]
    for c, w in enumerate(col_widths_2, 1):
        ws2.column_dimensions[get_column_letter(c)].width = w

    # -----------------------------------------------------------------------
    # Sheet 3: Pareto Frontier
    # -----------------------------------------------------------------------
    ws3 = wb.create_sheet("Pareto Frontier")

    points = []
    for sku in SKUS:
        cpm = cost_per_million_tokens(sku.representative_price_usd_hr, sku.output_tokens_per_sec)
        points.append((sku.representative_price_usd_hr, cpm, sku.name, sku.type, sku.vendor))

    frontier = find_pareto_frontier([(p[0], p[1], p[2]) for p in points])
    frontier_names = {f[2] for f in frontier}

    headers3 = [
        "Accelerator SKU", "Vendor", "Type",
        "Representative $/hr",
        "$/M Output Tokens",
        "On Pareto Frontier?",
        "Pareto Rank",
    ]
    for c, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    sorted_points = sorted(points, key=lambda p: p[0])
    pareto_rank = 0
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    for r, pt in enumerate(sorted_points, 2):
        on_frontier = pt[2] in frontier_names
        if on_frontier:
            pareto_rank += 1
        vals = [
            pt[2], pt[4], pt[3],
            pt[0], round(pt[1], 4),
            "YES" if on_frontier else "No",
            pareto_rank if on_frontier else "",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=r, column=c, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if c == 4:
                cell.number_format = usd_2dp
            if c == 5:
                cell.number_format = usd_4dp
            if on_frontier:
                cell.fill = green_fill

    col_widths_3 = [28, 10, 8, 18, 22, 20, 14]
    for c, w in enumerate(col_widths_3, 1):
        ws3.column_dimensions[get_column_letter(c)].width = w

    # -----------------------------------------------------------------------
    # Sheet 4: Sources & Methodology
    # -----------------------------------------------------------------------
    ws4 = wb.create_sheet("Sources & Methodology")
    ws4.column_dimensions["A"].width = 120
    sources = [
        "GPU / TPU Pareto Frontier Analysis — Sources & Methodology",
        "",
        "Data compiled: April 2026",
        "",
        "=== PRICING SOURCES ===",
        "• IntuitionLabs H100 Rental Price Comparison 2026",
        "• Spheron GPU Cloud Pricing Comparison 2026",
        "• GPUCloudList GPU Cloud Price Comparison 2026",
        "• getdeploying.com B200/T4/MI300X Cloud Pricing 2026",
        "• Google Cloud TPU Pricing (cloud.google.com/tpu/pricing)",
        "• RunPod, Lambda Labs, CoreWeave, Jarvislabs published pricing pages",
        "• ByteIota GPU Cloud Pricing 2026 Real AI Inference Cost Analysis",
        "",
        "=== THROUGHPUT BENCHMARK SOURCES ===",
        "• NVIDIA NIM LLMs Benchmarking (docs.nvidia.com) — H100, H200 Llama 3.3 70B",
        "• Trifonova (Feb 2026) 'Benchmarking LLM Inference on B200, H200, H100' — Medium",
        "• Metrum AI 'Llama 4 Maverick on H200 vs B200 using vLLM' — B200 throughput",
        "• dlewis.io 'Evaluating Llama 3.3 70B on H100 and A100' — A100 throughput",
        "• NVIDIA TensorRT-LLM H200 Launch Blog — H200 Llama 2 benchmarks",
        "• Google Cloud Blog 'AI Hypercomputer inference updates' — TPU v5e/v6e JetStream",
        "• EaseCloud 'Maximize LLM Throughput with Google TPU v5p' — TPU v5p benchmarks",
        "• openllmbenchmarks.com L40S token generation speed benchmarks",
        "• Oracle Cloud MI300X LLM serving benchmarks",
        "",
        "=== METHODOLOGY ===",
        "• Reference model: Llama 3 / Llama 2 70B-class (closest available benchmark)",
        "• Throughput: output tokens/second per single accelerator chip/GPU",
        "  - For multi-GPU benchmarks, total throughput divided by GPU count",
        "  - Precision: FP8 where available (H100/H200/B200), FP16/INT8 otherwise",
        "• Cost per million output tokens = ($/hr ÷ (tok/s × 3600)) × 1,000,000",
        "• Pareto frontier: minimize BOTH rental $/hr AND $/M output tokens",
        "  - Sorted by ascending $/hr; a point is Pareto-optimal if no prior point",
        "    has lower $/M output tokens",
        "• Prices are on-demand unless noted; spot/preemptible shown in Low column",
        "• TPU prices are per-chip-hour (Google Cloud pricing convention)",
    ]
    for r, line in enumerate(sources, 1):
        cell = ws4.cell(row=r, column=1, value=line)
        if r == 1:
            cell.font = Font(name="Calibri", bold=True, size=14)
        elif line.startswith("==="):
            cell.font = Font(name="Calibri", bold=True, size=11)

    return wb


# ---------------------------------------------------------------------------
# Generate Pareto frontier chart as PNG
# ---------------------------------------------------------------------------
def generate_pareto_chart():
    fig, ax = plt.subplots(figsize=(14, 9))

    type_markers = {"GPU": "o", "TPU": "s"}
    vendor_colors = {"NVIDIA": "#76B900", "AMD": "#ED1C24", "Google": "#4285F4"}

    all_x, all_y, all_labels = [], [], []

    for sku in SKUS:
        x = sku.representative_price_usd_hr
        y = cost_per_million_tokens(sku.representative_price_usd_hr, sku.output_tokens_per_sec)
        marker = type_markers.get(sku.type, "^")
        color = vendor_colors.get(sku.vendor, "#888888")
        ax.scatter(x, y, s=120, marker=marker, c=color, edgecolors="black",
                   linewidths=0.8, zorder=5)
        all_x.append(x)
        all_y.append(y)
        all_labels.append(sku.name)

    for x, y, label in zip(all_x, all_y, all_labels):
        offset = (8, 8)
        if "B200" in label:
            offset = (8, -14)
        elif "H200" in label:
            offset = (-10, 12)
        elif "v6e" in label:
            offset = (8, -12)
        elif "MI300" in label:
            offset = (-10, -14)
        ax.annotate(label, (x, y), textcoords="offset points", xytext=offset,
                    fontsize=8.5, fontweight="bold", ha="left")

    points = []
    for sku in SKUS:
        cpm = cost_per_million_tokens(sku.representative_price_usd_hr, sku.output_tokens_per_sec)
        points.append((sku.representative_price_usd_hr, cpm, sku.name))

    frontier = find_pareto_frontier(points)
    frontier_x = [f[0] for f in frontier]
    frontier_y = [f[1] for f in frontier]
    ax.plot(frontier_x, frontier_y, color="#FF6600", linewidth=2.5, linestyle="--",
            label="Pareto Frontier", zorder=4, alpha=0.85)

    from matplotlib.lines import Line2D
    legend_elements = [
        Line2D([0], [0], marker="o", color="w", markerfacecolor="#76B900",
               markeredgecolor="black", markersize=10, label="NVIDIA GPU"),
        Line2D([0], [0], marker="o", color="w", markerfacecolor="#ED1C24",
               markeredgecolor="black", markersize=10, label="AMD GPU"),
        Line2D([0], [0], marker="s", color="w", markerfacecolor="#4285F4",
               markeredgecolor="black", markersize=10, label="Google TPU"),
        Line2D([0], [0], color="#FF6600", linewidth=2.5, linestyle="--",
               label="Pareto Frontier"),
    ]
    ax.legend(handles=legend_elements, loc="upper right", fontsize=10,
              framealpha=0.9)

    ax.set_xlabel("Rental Cost ($/hr per GPU/chip)", fontsize=12, fontweight="bold")
    ax.set_ylabel("Cost per Million Output Tokens ($)", fontsize=12, fontweight="bold")
    ax.set_title("GPU & TPU Pareto Frontier: Rental Cost vs Output Token Cost\n"
                 "(Llama 70B-class, April 2026 pricing)",
                 fontsize=14, fontweight="bold")

    ax.grid(True, alpha=0.3, linestyle="-")
    ax.set_xlim(left=0)
    ax.set_ylim(bottom=0)

    fig.tight_layout()
    fig.savefig("/workspace/pareto_frontier_chart.png", dpi=200, bbox_inches="tight")
    plt.close(fig)
    print("Chart saved to pareto_frontier_chart.png")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    wb = build_workbook()
    output_path = "/workspace/gpu_tpu_pareto_frontier.xlsx"
    wb.save(output_path)
    print(f"Workbook saved to {output_path}")

    generate_pareto_chart()

    print("\n=== Pareto Frontier Summary ===")
    points = []
    for sku in SKUS:
        cpm = cost_per_million_tokens(sku.representative_price_usd_hr, sku.output_tokens_per_sec)
        points.append((sku.representative_price_usd_hr, cpm, sku.name))

    frontier = find_pareto_frontier(points)
    print(f"{'SKU':<30} {'$/hr':>8} {'$/M tokens':>12}")
    print("-" * 52)
    for f in frontier:
        print(f"{f[2]:<30} {f[0]:>8.2f} {f[1]:>12.4f}")
