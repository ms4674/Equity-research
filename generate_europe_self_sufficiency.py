import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
subheader_font = Font(bold=True, size=10)
subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
wrap_align = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_align
        cell.border = thin_border


def style_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.alignment = wrap_align
        cell.border = thin_border


def auto_width(ws, min_width=12, max_width=55):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = min_width
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[col_letter].width = max_len + 2


# =============================================================================
# Sheet 1: Executive Summary
# =============================================================================
ws = wb.active
ws.title = "Executive Summary"

summary_data = [
    ["Category", "Key Metric", "Value", "Source / Date"],
    ["EU AI Investment", "InvestAI total mobilisation target", "€200 billion (2027-2030)", "AI Continent Action Plan, Apr 2025"],
    ["EU AI Investment", "AI Gigafactories earmarked", "€20 billion for 5 sites", "AI Continent Action Plan, Apr 2025"],
    ["EU AI Investment", "EuroHPC total investment (2021-2027)", "€10 billion", "EuroHPC JU / EC, 2026"],
    ["EU AI Investment", "Open Source Strategy funding", "€2 billion over 7 years", "Tech Sovereignty Package, Jun 2026"],
    ["EU AI Investment", "Cloud & AI additional investment needed", "€100 billion", "EC Tech Sovereignty Package, Jun 2026"],
    ["France AI", "Total AI investment package", "€109 billion", "Macron AI Action Summit, Feb 2025"],
    ["France AI", "Bpifrance AI allocation", "€10 billion", "Feb 2025"],
    ["France AI", "Nuclear power pledged to AI training", "1 GW by end-2026", "French Government, Feb 2025"],
    ["UK AI", "Sovereign AI Fund", "£500 million", "UK Government, Apr 2026"],
    ["EU Cybersecurity", "Digital Europe Programme (cyber, 2025-27)", "€353 million", "ECCC Work Programme"],
    ["EU Cybersecurity", "SAFE loans for defence/cyber", "Up to €150 billion", "Council of EU, May 2025"],
    ["EU Cybersecurity", "ECIP target fund size", "€1 billion+", "ECSO"],
    ["EU Cybersecurity", "Annual VC gap in EU cybersecurity", "€1.75 billion/year", "ECSO / EIB"],
    ["EU Chips", "Chips Act 2.0 target (global share by 2030)", "20% by value", "EC Digital Decade"],
    ["EU Chips", "Projected actual share by 2030", "11.7%", "EU Court of Auditors, 2025"],
    ["EU Chips", "Total investment mobilised under Chips Act", "€80 billion+", "EC, 2025"],
    ["Cloud Market", "US hyperscaler share of EU cloud market", "~85%", "Synergy Research, Feb 2026"],
    ["Cloud Market", "European provider share of EU cloud market", "~15%", "Synergy Research, 2025-2026"],
    ["Cloud Market", "EU sovereign cloud contract awarded", "€180 million (6 years)", "EC, Apr 2026"],
    ["LLM Market", "Mistral AI ARR (May 2026)", "~$1.0 billion", "Sacra / Presenc AI, 2026"],
    ["LLM Market", "Mistral AI valuation", "~$14-15 billion", "Series C, Q1 2026"],
    ["LLM Market", "Cohere + Aleph Alpha combined valuation", "~$20 billion", "TechCrunch, Apr 2026"],
    ["Cybersecurity Market", "Europe cybersecurity market (2025)", "$63.12 billion", "Mordor Intelligence"],
    ["Cybersecurity Market", "Europe cybersecurity market (2026E)", "$69.82 billion", "Mordor Intelligence"],
    ["Cybersecurity Market", "Europe cybersecurity market (2031E)", "$115.66 billion", "Mordor Intelligence"],
    ["Cybersecurity Market", "CAGR (2026-2031)", "10.62%", "Mordor Intelligence"],
]

for r_idx, row in enumerate(summary_data, 1):
    for c_idx, val in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=val)
    if r_idx == 1:
        style_header(ws, r_idx, len(row))
    else:
        style_row(ws, r_idx, len(row))

auto_width(ws)

# =============================================================================
# Sheet 2: EU Policy & Regulatory Framework
# =============================================================================
ws2 = wb.create_sheet("Policy & Regulation")

policy_data = [
    ["Initiative / Regulation", "Type", "Status", "Key Dates", "Relevance to Self-Sufficiency", "Investment/Budget"],
    ["EU AI Act", "Regulation", "Phased implementation", "Prohibitions: Feb 2025; GPAI: Aug 2025; High-risk: Aug 2026 (deferred to Dec 2027 via Omnibus)", "Sets rules for AI development; GPAI governance; creates demand for compliant European AI", "N/A (compliance costs €150K-€2M for essential entities)"],
    ["AI Continent Action Plan", "Strategy", "Adopted Apr 2025", "Apr 2025 - 2030", "Central framework for EU AI sovereignty; tripling data centre capacity", "€200 billion via InvestAI"],
    ["Cloud and AI Development Act (CADA)", "Proposed Regulation", "Proposed Jun 2026", "Expected adoption 2026-2027", "Streamlines data centre permitting; sovereignty assessment framework; cloud autonomy", "Part of €200B InvestAI"],
    ["Chips Act 2.0", "Proposed Regulation", "Proposed Jun 2026", "Jun 2026 legislative process", "Advanced chip production in EU; reduce strategic semiconductor dependencies", "Builds on €80B+ from Chips Act 1.0"],
    ["EU Open Source Strategy", "Strategy", "Adopted Jun 2026", "2026-2033", "Reduce dependencies across tech stack via open-source alternatives", "€2 billion over 7 years"],
    ["Tech Sovereignty Package", "Legislative Package", "Proposed Jun 2026", "Four interconnected initiatives", "Comprehensive approach: chips → infrastructure → software → cloud/AI", "Combined ~€100B+ additional needed"],
    ["NIS2 Directive", "Directive", "Transposition underway", "Transposition deadline: Oct 2024; Enforcement: 2025-2026", "Mandates cybersecurity for critical infrastructure; drives spending", "Compliance costs: €30K-€2M per entity"],
    ["Cyber Resilience Act (CRA)", "Regulation", "In force (Dec 2024)", "Reporting: Sep 2026; Full: Dec 2027", "Product cybersecurity by design; reduces dependency on non-EU vendors", "N/A"],
    ["DORA", "Regulation", "In force (Jan 2025)", "Fully applicable Jan 2025", "Financial sector digital resilience; ICT third-party risk management", "N/A"],
    ["SAFE (Security Action for Europe)", "Financial Instrument", "In force (May 2025)", "From May 2025", "€150B loans for defence including cyber; supports EU defence industry", "Up to €150 billion in loans"],
    ["ReArm Europe / Readiness 2030", "Strategy", "Active", "Target: 2030", "Unlock €800B+ in defence spending; cyber/AI as key capability area", "€800 billion+ total defence"],
    ["Digital Decade Policy Programme", "Strategy", "Active", "Targets by 2030", "20% semiconductor share; sovereign cloud; digital skills", "Multi-billion across pillars"],
    ["Frontier AI Initiative", "Non-profit Initiative", "Launched Nov 2025", "Q1 2026 establishment", "France-Germany-EC joint frontier AI research; access to giga-compute", "TBD (aims to be best-funded non-profit frontier AI initiative)"],
]

for r_idx, row in enumerate(policy_data, 1):
    for c_idx, val in enumerate(row, 1):
        ws2.cell(row=r_idx, column=c_idx, value=val)
    if r_idx == 1:
        style_header(ws2, r_idx, len(row))
    else:
        style_row(ws2, r_idx, len(row))

auto_width(ws2)

# =============================================================================
# Sheet 3: AI/LLM Infrastructure & Compute
# =============================================================================
ws3 = wb.create_sheet("AI-LLM Infrastructure")

infra_data = [
    ["Facility / Initiative", "Country", "Type", "Compute Capacity", "Status", "Investment", "Key Partners"],
    ["IT4LIA AI Factory", "Italy (Bologna)", "AI Factory", "160+ ExaFLOPS AI inference", "Contract signed Apr 2026", "EuroHPC funded", "CINECA, Dell, NVIDIA, Axelera AI, SiPearl"],
    ["HammerHAI", "Germany", "AI Factory", "15+ ExaFLOPS AI inference", "Contract signed Mar 2026", "EuroHPC funded", "HPE, NVIDIA GB200, Axelera AI, VAST Data"],
    ["JUPITER", "Germany (Jülich)", "Exascale Supercomputer", "First EU exascale system", "Operational", "EuroHPC funded", "Jülich Supercomputing Centre"],
    ["JAIF (Jülich AI Factory)", "Germany", "AI Factory", "Centred around JUPITER", "Selected Dec 2024", "EuroHPC funded", "Jülich Supercomputing Centre"],
    ["Alice Recoque", "France", "AI Factory", "TBD", "Selected Mar 2025", "EuroHPC funded", "French consortium"],
    ["LUMI", "Finland", "Pre-exascale", "~550 PetaFLOPS", "Operational", "EuroHPC funded", "CSC Finland"],
    ["Leonardo", "Italy", "Pre-exascale", "~250 PetaFLOPS", "Operational", "EuroHPC funded", "CINECA"],
    ["LitAI Factory", "Lithuania", "AI Factory", "TBD", "Selected Oct 2025", "EuroHPC funded", "Vilnius University consortium"],
    ["NLAIF", "Netherlands", "AI Factory", "TBD", "Selected Oct 2025", "EuroHPC funded", "Dutch consortium"],
    ["Fluidstack Supercomputer (France)", "France", "Private AI compute", "500,000 GPUs by 2026", "Under construction", "€10 billion", "Fluidstack, French Government"],
    ["Mistral Compute (Essonne)", "France", "Private sovereign cloud", "18,000 NVIDIA Grace Blackwell Superchips, 40MW", "Launching", "Part of Series C", "Mistral AI, NVIDIA"],
    ["AI Gigafactories (5 planned)", "EU-wide", "Next-gen AI compute", "100,000+ advanced AI processors each", "Mandate expanded (Council Reg. 2026/150)", "€20 billion", "EuroHPC JU + Member States"],
    ["13 AI Factory Antennas", "7 EU + 6 partner countries", "Extended access points", "Complementary to AI Factories", "Selected Oct 2025", "EuroHPC funded", "Belgium, Cyprus, Hungary, Ireland, Latvia, Malta, Slovakia + Iceland, Moldova, N. Macedonia, Serbia, Switzerland, UK"],
    ["Total EuroHPC AI Factories", "19 sites across EU", "AI Factory network", "Tripling current AI compute capacity", "Operational/deploying", "€10 billion (2021-2027)", "EuroHPC Joint Undertaking"],
]

for r_idx, row in enumerate(infra_data, 1):
    for c_idx, val in enumerate(row, 1):
        ws3.cell(row=r_idx, column=c_idx, value=val)
    if r_idx == 1:
        style_header(ws3, r_idx, len(row))
    else:
        style_row(ws3, r_idx, len(row))

auto_width(ws3)

# =============================================================================
# Sheet 4: Key European AI/LLM Companies
# =============================================================================
ws4 = wb.create_sheet("European AI-LLM Companies")

ai_companies = [
    ["Company", "HQ", "Focus", "Valuation", "Revenue / ARR", "Key Funding", "Enterprise Customers", "Sovereignty Angle"],
    ["Mistral AI", "Paris, France", "Foundation models (LLMs)", "$14-15 billion", "~$1.0B ARR (May 2026); target $1.1-1.2B for 2026", "€2.8B total; €1.7B Series C led by ASML (11% stake)", "40% of Europe's Fortune 500; BNP Paribas, AXA, Orange, TotalEnergies, Stellantis, CMA CGM, French MoD", "European sovereign AI anchor; on-prem deployments; EU data residency"],
    ["Cohere + Aleph Alpha (merged)", "Toronto + Germany", "Enterprise sovereign AI", "~$20 billion (combined)", "Cohere: ~$240M ARR (2025)", "€500M from Schwarz Group (Series E); total raising TBD", "German public sector, regulated industries, EU institutions", "Transatlantic sovereign AI; STACKIT cloud; data sovereignty guarantee"],
    ["DeepL", "Cologne, Germany", "AI translation", "~$2 billion (2024)", "Est. $100M+ ARR", "Series C: $300M (Jan 2024)", "Enterprise translation for EU institutions, corporates", "European language AI; data processed in EU"],
    ["Stability AI", "London, UK", "Image/video generation", "~$1 billion (restructured)", "Limited", "~$200M total", "Creative industries", "Open-source generative models"],
    ["Poolside AI", "Paris, France", "AI code generation", "$3 billion (Oct 2024)", "Pre-revenue / early", "$500M Series B", "Developer tools", "European-founded code AI"],
    ["Hugging Face", "Paris/New York", "ML platform / open-source hub", "$4.5 billion (Aug 2023)", "Growing; freemium model", "$235M Series D", "Platform for all EU AI developers", "Open-source AI democratisation; EU-founded"],
    ["SiPearl", "France", "Sovereign AI/HPC processors", "Private", "Pre-revenue (chip development)", "€200M+ in public/private funding", "EuroHPC supercomputers", "European processor design for sovereign compute"],
    ["Axelera AI", "Netherlands", "AI inference accelerators", "Private", "Pre-revenue", "€120M+ raised", "Integrated in HammerHAI and IT4LIA", "European AI chip alternative"],
]

for r_idx, row in enumerate(ai_companies, 1):
    for c_idx, val in enumerate(row, 1):
        ws4.cell(row=r_idx, column=c_idx, value=val)
    if r_idx == 1:
        style_header(ws4, r_idx, len(row))
    else:
        style_row(ws4, r_idx, len(row))

auto_width(ws4)

# =============================================================================
# Sheet 5: Cybersecurity Market & Investment
# =============================================================================
ws5 = wb.create_sheet("Cybersecurity Market")

cyber_market = [
    ["Metric", "Value", "Year / Period", "Source", "Notes"],
    ["Europe cybersecurity market size", "$63.12 billion", "2025", "Mordor Intelligence", ""],
    ["Europe cybersecurity market size", "$69.82 billion", "2026E", "Mordor Intelligence", ""],
    ["Europe cybersecurity market size", "$115.66 billion", "2031E", "Mordor Intelligence", "CAGR 10.62% (2026-2031)"],
    ["Europe cybersecurity market size (alt.)", "$54.77 billion", "2025", "MarketsandMarkets", ""],
    ["Europe cybersecurity market size (alt.)", "$83.14 billion", "2030E", "MarketsandMarkets", "CAGR 8.7% (2025-2030)"],
    ["Managed detection & response CAGR", "13.56%", "2026-2031", "Mordor Intelligence", "Fastest growing segment"],
    ["EU annual VC gap in cybersecurity", "€1.75 billion/year", "2024-2026", "ECSO", "Compared to US/Israel"],
    ["EU cybersecurity VC investment", "€814 million", "2021", "EIB", "vs $15B US, $2.5B Israel"],
    ["EU average specialised VC fund size", "3x smaller than US", "2022", "EIB", "Structural scale disadvantage"],
    ["Digital Europe Programme (Cyber)", "€353 million", "2025-2027", "ECCC Work Programme", "Post-quantum, AI security, cross-border hubs"],
    ["- New technologies & AI for cybersecurity", "€127 million", "2025-2027", "ECCC", "Subset of DEP cyber budget"],
    ["- Cyber Solidarity Act implementation", "€111 million", "2025-2027", "ECCC", "Alert system, emergency mechanism"],
    ["- EU resilience improvement", "€106 million", "2025-2027", "ECCC", "Additional resilience actions"],
    ["ECIP target fund size", "€1 billion+", "TBD", "ECSO", "Fund-of-funds for EU cyber scale-ups"],
    ["SAFE (defence/cyber loans)", "Up to €150 billion", "From May 2025", "Council of EU", "Includes cyber as key capability area"],
    ["InvestEU Defence Equity Facility", "€175 million", "Active", "EIB/EC", "VC backing for dual-use/cyber startups"],
    ["ReArm Europe total defence target", "€800 billion+", "By 2030", "EC", "Cyber/AI as 1 of 9 capability coalitions"],
    ["Germany cyber losses (annual)", "~€300 billion", "2024", "German govt estimates", "Driving spending increase"],
    ["ENISA budget increase (projected)", "+81.5%", "By 2028", "EU budget projections", "Mandate expansion for operational cooperation"],
]

for r_idx, row in enumerate(cyber_market, 1):
    for c_idx, val in enumerate(row, 1):
        ws5.cell(row=r_idx, column=c_idx, value=val)
    if r_idx == 1:
        style_header(ws5, r_idx, len(row))
    else:
        style_row(ws5, r_idx, len(row))

auto_width(ws5)

# =============================================================================
# Sheet 6: Key European Cybersecurity Companies
# =============================================================================
ws6 = wb.create_sheet("EU Cybersecurity Companies")

cyber_companies = [
    ["Company", "HQ Country", "Focus Area", "Revenue (Latest)", "Key Metrics", "Sovereignty Relevance"],
    ["Thales (Cyber & Digital)", "France", "Cybersecurity, digital identity, defence", "€4.0B (Cyber & Digital, 2025); of which cybersecurity: €1.455B", "Group: €22.1B sales; €4.5B R&D; 85K employees", "French sovereign cloud (Bleu JV with Google); SecNumCloud certified; EU defence"],
    ["Airbus CyberSecurity", "Germany/France", "SOC/MDR, OT security, defence cyber", "Part of Airbus Defence & Space", "Key EU defence contractor", "NATO/EU institutional cyber; sovereign defence programmes"],
    ["Atos / Eviden", "France", "Managed security, digital workplace, HPC", "Part of Atos group (restructuring)", "Major EU institutional contracts", "EuroHPC operator; EU institutional IT; sovereign cloud services"],
    ["Orange Cyberdefense", "France", "Managed security services", "~€1B+ revenue (2025)", "Largest EU-headquartered MSSP", "European SOC network; NIS2-aligned services"],
    ["WithSecure", "Finland", "Endpoint & cloud security (MSP focus)", "~€135M revenue (2025)", "Positioned as 'European cybersecurity'", "EU data sovereignty; GDPR-native; no US data routing"],
    ["Darktrace", "UK", "AI-driven cyber defence (NDR)", "~$600M ARR (before Thoma Bravo acquisition)", "Pioneered AI for cybersecurity", "UK sovereign; AI self-learning network defence"],
    ["ESET", "Slovakia", "Endpoint protection, threat intelligence", "~$700M revenue", "400M+ users globally", "EU-headquartered alternative to US/Russian AV"],
    ["Bitdefender", "Romania", "Endpoint, XDR, managed detection", "~$700M+ revenue", "Major OEM partnerships", "EU-based; significant R&D in Romania"],
    ["Secunet Security Networks", "Germany", "IT security for government/defence", "~€400M revenue (2024)", "German govt preferred partner", "Classified systems; German sovereignty; BSI certified"],
    ["Rohde & Schwarz Cybersecurity", "Germany", "Network security, encryption", "Part of R&S group", "Defence & govt focused", "German sovereign encryption; NATO-grade"],
    ["Stormshield (Airbus subsidiary)", "France", "Firewalls, endpoint, data security", "~€80M revenue", "EU-certified (ANSSI, BSI, CCN)", "Certified EU-only cyber; no foreign code dependency"],
    ["Deutsche Telekom Security (T-Systems)", "Germany", "Managed security, SOC, consulting", "Part of DT group", "Large enterprise customer base", "German sovereign cloud (Open Telekom Cloud)"],
    ["Capgemini (Cyber)", "France", "Cybersecurity consulting & managed services", "Part of €22B group", "Global delivery but EU-HQ", "Major EU institutional cyber consultancy"],
    ["Sophos", "UK", "Endpoint, firewall, MDR", "~$1B+ revenue", "Thoma Bravo portfolio (private)", "UK-based; significant EU customer base"],
    ["NCC Group", "UK", "Cyber consulting, assurance, managed services", "~£350M revenue", "UK NCSC-aligned", "UK/EU critical infrastructure testing"],
    ["Kaspersky Lab*", "Netherlands (HQ moved)", "Endpoint, threat intelligence", "~$700M+ revenue", "Transparency centres in EU", "*Russian-origin; geopolitical risk; EU data centres"],
    ["CY4GATE", "Italy", "Offensive/defensive cyber, intelligence", "~€80M revenue", "Italian defence contractor", "Italian/EU sovereign; NATO programmes"],
    ["Schwarz Digits / STACKIT", "Germany", "Sovereign cloud + cyber", "Part of Schwarz Group (Lidl)", "Operating sovereign cloud for Cohere/Aleph Alpha", "German-owned sovereign cloud; GDPR-native"],
]

for r_idx, row in enumerate(cyber_companies, 1):
    for c_idx, val in enumerate(row, 1):
        ws6.cell(row=r_idx, column=c_idx, value=val)
    if r_idx == 1:
        style_header(ws6, r_idx, len(row))
    else:
        style_row(ws6, r_idx, len(row))

auto_width(ws6)

# =============================================================================
# Sheet 7: Semiconductor Self-Sufficiency
# =============================================================================
ws7 = wb.create_sheet("Semiconductors")

semi_data = [
    ["Metric / Initiative", "Value / Status", "Target Date", "Source", "Notes"],
    ["EU global chip market share (2020)", "~7%", "Baseline", "EU Court of Auditors", "Starting point for Chips Act"],
    ["EU global chip market share (projected 2030)", "11.7%", "2030", "EU Court of Auditors / EC own forecast", "Based on current investment trajectory"],
    ["EU Digital Decade target", "20% of world's cutting-edge chips by value", "2030", "Digital Decade Policy Programme", "Deemed 'disconnected from reality' by auditors"],
    ["Total investment mobilised (Chips Act 1.0)", "€80 billion+", "Since 2023", "European Commission", "Excludes €34.2B Intel freeze"],
    ["Chips JTI pilot lines investment", "€3.7 billion (EU + national)", "Ongoing", "EP Briefing 2026", "FAMES, NanoIC, APECS, WBG, PIXEurope"],
    ["IPF/OEF projects awarded", "5 projects across EU", "By Mar 2026", "EP Briefing 2026", "Integrated Production Facilities & Open EU Foundries"],
    ["Intel mega-projects (Germany/Poland)", "€34.2 billion", "Frozen/delayed", "Multiple sources", "Single-point-of-failure risk to EU targets"],
    ["Chips Act 2.0 (proposed)", "New regulation", "Jun 2026", "EC Tech Sovereignty Package", "Build capacity in cutting-edge; reduce dependencies"],
    ["Chips Act 2.0 — Semiconductor alliances target", "5 alliances by 2030", "2030", "EP Briefing", "Automotive, energy, defence, healthcare, telecoms, data centres, AI"],
    ["Semicon Coalition declaration", "27 Member States signed", "Sep 2025", "EC", "Calling for Chips Act 2.0"],
    ["EU lacks manufacturing below 10nm", "Full dependency on third countries", "Current", "EP Briefing 2026", "Critical for defence, security, automotive, space, HPC"],
    ["SiPearl (European processor)", "Developing HPC/AI chips", "In development", "SiPearl / EuroHPC", "European-designed processors for sovereign compute"],
    ["Axelera AI (inference accelerators)", "Netherlands-based", "Integrated in EU AI Factories", "EuroHPC JU contracts", "European AI chip company in HammerHAI & IT4LIA"],
    ["ASML (lithography equipment)", "Dutch; global monopoly on EUV", "Operational", "Market data", "EU strategic asset; invested €1.3B in Mistral AI"],
]

for r_idx, row in enumerate(semi_data, 1):
    for c_idx, val in enumerate(row, 1):
        ws7.cell(row=r_idx, column=c_idx, value=val)
    if r_idx == 1:
        style_header(ws7, r_idx, len(row))
    else:
        style_row(ws7, r_idx, len(row))

auto_width(ws7)

# =============================================================================
# Sheet 8: Cloud Market & Sovereignty
# =============================================================================
ws8 = wb.create_sheet("Cloud Sovereignty")

cloud_data = [
    ["Metric / Initiative", "Value", "Date", "Source", "Notes"],
    ["US hyperscaler share of EU cloud (IaaS+PaaS)", "~85%", "Feb 2026", "Synergy Research Group", "AWS 32%, Azure 23%, Google 10%, others 7%"],
    ["European provider share of EU cloud", "~15%", "2025-2026", "Synergy Research / The Register", "Down from 27% in 2017"],
    ["EU cloud market total value", "~€61 billion ($70B)", "2024", "Synergy Research", "6x growth since 2017"],
    ["US quarterly capex in European cloud", "~€10 billion/quarter", "2025", "Synergy Research", "'Impossible hill to climb' for EU challengers"],
    ["OVHcloud market share (largest EU provider)", "<2%", "2025", "Market data", "Revenue: €993M"],
    ["Deutsche Telekom / SAP market share (each)", "~2%", "2025", "Market data", "Among largest EU players"],
    ["EU sovereign cloud contract awarded", "€180 million (6 years)", "Apr 2026", "European Commission", "4 consortia selected"],
    ["- Post Telecom + CleverCloud + OVHcloud", "SEAL-3 level", "Apr 2026", "EC", "Fully European-owned"],
    ["- STACKIT (Schwarz Group)", "SEAL-3 level", "Apr 2026", "EC", "Fully European-owned"],
    ["- Scaleway (Iliad Group)", "SEAL-3 level", "Apr 2026", "EC", "Fully European-owned"],
    ["- Proximus + S3NS (Thales-Google JV) + Mistral", "SEAL-2 level", "Apr 2026", "EC", "US tech operated by EU companies"],
    ["CADA target", "Triple EU data centre capacity in 5-7 years", "2025-2032", "AI Continent Action Plan", "Requires parallel grid infrastructure"],
    ["GAIA-X", "EU sovereign cloud framework", "Ongoing since 2019", "Franco-German initiative", "Criticised for slow progress; US members admitted"],
    ["Bleu (Thales + Google Cloud JV)", "French sovereign cloud", "Operational", "Thales", "SecNumCloud certified"],
    ["STACKIT (Schwarz Digits)", "German sovereign cloud", "Operational", "Schwarz Group", "Backing Cohere/Aleph Alpha sovereign AI"],
    ["92% of Western data stored on US-controlled cloud", "92%", "2025", "Industry estimates", "Structural dependency metric"],
]

for r_idx, row in enumerate(cloud_data, 1):
    for c_idx, val in enumerate(row, 1):
        ws8.cell(row=r_idx, column=c_idx, value=val)
    if r_idx == 1:
        style_header(ws8, r_idx, len(row))
    else:
        style_row(ws8, r_idx, len(row))

auto_width(ws8)

# =============================================================================
# Sheet 9: Gaps, Risks & Challenges
# =============================================================================
ws9 = wb.create_sheet("Gaps & Challenges")

gaps_data = [
    ["Domain", "Gap / Challenge", "Severity", "Detail", "Potential Mitigation"],
    ["LLMs", "Scale disadvantage vs US labs", "High", "Mistral: $1B ARR vs OpenAI $20B, Anthropic $14B; 7-20x gap", "InvestAI + AI Gigafactories; consolidation (Cohere+Aleph Alpha)"],
    ["LLMs", "Talent retention to US", "High", "Top researchers attracted to US compensation/compute access", "Frontier AI Initiative; competitive academic environments"],
    ["LLMs", "Training compute gap", "High", "EU aggregate ~50 PFlops across 13 factories vs single US hyperscale site", "AI Gigafactories (100K+ GPUs each); planned tripling"],
    ["LLMs", "Consumer model adoption", "Medium", "Mistral 1.1B API queries vs OpenAI 10B+; limited consumer products", "Focus on enterprise/sovereign niche instead of consumer"],
    ["Cybersecurity", "VC funding gap vs US/Israel", "High", "EU: €814M (2021) vs US: $15B; annual gap €1.75B", "ECIP fund-of-funds; InvestEU Defence Equity"],
    ["Cybersecurity", "Market dominated by non-EU vendors", "High", "Top 5 in EU market: Cisco, Palo Alto, IBM, Check Point, Fortinet (all non-EU)", "Procurement preferences; CADA sovereignty framework"],
    ["Cybersecurity", "Scale-up failure", "High", "EU cyber companies struggle to reach $1B+ scale; funding cliff after Series A", "ECIP; SAFE loans; defence procurement consolidation"],
    ["Cybersecurity", "NIS2 fragmented implementation", "Medium", "23/27 states missed Oct 2024 deadline; divergent national requirements", "EC infringement procedures; enforcement pressure"],
    ["Cloud", "85% US dependency", "Critical", "AWS+Azure+Google = 70% alone; EU providers declining share since 2017", "CADA mandatory diversification; sovereign cloud contracts"],
    ["Cloud", "Investment asymmetry", "Critical", "US investing €10B/quarter in EU capex; no EU company can match", "Public investment via InvestAI; AI Gigafactories"],
    ["Cloud", "GAIA-X underdelivery", "Medium", "3+ years of development with limited commercial traction", "CADA regulation to operationalise sovereignty"],
    ["Semiconductors", "No sub-10nm manufacturing", "Critical", "100% import dependency for cutting-edge chips (defence, AI, 5G)", "Chips Act 2.0; TSMC/Samsung EU fabs; domestic design (SiPearl)"],
    ["Semiconductors", "20% target unreachable", "High", "Projected 11.7% by 2030 vs 20% goal; Intel delays critical", "Target revision; focus on 'indispensability' over market share"],
    ["Semiconductors", "Energy cost disadvantage", "Medium", "Higher electricity costs than US/Asia for fab operations", "Nuclear expansion (France); grid infrastructure investment"],
    ["Regulation", "AI Act compliance burden on EU startups", "Medium", "High-risk obligations may disadvantage smaller EU players vs US competitors", "Omnibus simplification (May 2026); sandbox provisions"],
    ["Regulation", "Fragmented national transposition (NIS2)", "Medium", "Different deadlines, requirements, enforcement across 27 states", "EC oversight; ENISA coordination"],
    ["Talent", "AI researcher brain drain", "High", "US offers 2-5x compensation; better compute access historically", "Frontier AI Initiative; EuroHPC access; national programmes"],
    ["Energy", "Data centre grid constraints", "High", "Tripling DC capacity requires parallel grid buildout; decades-long permitting", "CADA grid provisions; energy sovereignty roadmap"],
]

for r_idx, row in enumerate(gaps_data, 1):
    for c_idx, val in enumerate(row, 1):
        ws9.cell(row=r_idx, column=c_idx, value=val)
    if r_idx == 1:
        style_header(ws9, r_idx, len(row))
    else:
        style_row(ws9, r_idx, len(row))

auto_width(ws9)

# =============================================================================
# Sheet 10: National AI Strategies
# =============================================================================
ws10 = wb.create_sheet("National Strategies")

national_data = [
    ["Country", "Key Initiative", "Investment Amount", "Timeline", "Focus", "Key Partners / Champions"],
    ["France", "Macron AI Investment Package", "€109 billion (public + private)", "2025-2030", "Data centres, sovereign compute, LLM development, nuclear-powered AI", "Mistral AI, Bpifrance, Brookfield, UAE/MGX, Fluidstack"],
    ["France", "Bpifrance AI allocation", "€10 billion", "By 2029", "AI ecosystem support, startup funding", "Bpifrance"],
    ["France", "1 GW nuclear for AI", "TBD (part of €109B)", "By end-2026", "Decarbonised AI training compute", "EDF, nuclear operators"],
    ["France", "Fluidstack supercomputer", "€10 billion", "500K GPUs by 2026; 1 GW by 2028", "World's largest low-carbon AI supercomputer", "Fluidstack (UK), French government"],
    ["France", "France 2030 (AI portion)", "€2.22 billion over 5 years", "2021-2026", "AI research & technology (€1.5B public + €506M private)", "National research labs, startups"],
    ["Germany", "Chips Act national contributions", "€10 billion+ for Intel Magdeburg (paused)", "Uncertain", "Semiconductor manufacturing", "Intel (paused); TSMC possible"],
    ["Germany", "HammerHAI AI Factory", "EuroHPC funded", "Contract Mar 2026", "15+ ExaFLOPS AI inference", "HPE, NVIDIA, Axelera AI"],
    ["Germany", "JAIF (Jülich)", "EuroHPC funded", "Selected Dec 2024", "Exascale AI research", "Jülich Supercomputing Centre"],
    ["Germany", "Frontier AI Initiative co-lead", "TBD", "Q1 2026 establishment", "Joint frontier AI research with France", "EC, France, private partners"],
    ["UK", "Sovereign AI Fund", "£500 million", "Apr 2026 - 2030", "Equity stakes, compute access, grants for AI startups", "UK Sovereign AI Unit"],
    ["UK", "AI Safety Institute", "£100 million+", "2024-ongoing", "Frontier AI safety research and testing", "UK DSIT"],
    ["Italy", "IT4LIA AI Factory", "EuroHPC + national co-funding", "Contract Apr 2026", "160+ ExaFLOPS AI inference", "CINECA, Dell, NVIDIA, Axelera, SiPearl"],
    ["Netherlands", "NLAIF AI Factory", "EuroHPC funded", "Selected Oct 2025", "Sovereign compute for sensitive data (health, IP)", "Dutch consortium"],
    ["Netherlands", "Axelera AI", "€120M+ private funding", "Ongoing", "European AI inference chip design", "Axelera AI"],
    ["EU-wide", "InvestAI initiative", "€200 billion", "2027-2030", "Pan-EU AI infrastructure mobilisation", "EC, Member States, private sector"],
    ["EU-wide", "EuroHPC JU (total)", "€10 billion", "2021-2027", "19 AI Factories + supercomputers", "32 participating states"],
    ["EU-wide", "AI Gigafactories (5)", "€20 billion", "Mandate expanded 2026", "100K+ processors each; sovereign training capacity", "EuroHPC JU, Member States"],
]

for r_idx, row in enumerate(national_data, 1):
    for c_idx, val in enumerate(row, 1):
        ws10.cell(row=r_idx, column=c_idx, value=val)
    if r_idx == 1:
        style_header(ws10, r_idx, len(row))
    else:
        style_row(ws10, r_idx, len(row))

auto_width(ws10)

# =============================================================================
# Sheet 11: Timeline & Milestones
# =============================================================================
ws11 = wb.create_sheet("Timeline")

timeline_data = [
    ["Date", "Milestone", "Domain", "Significance"],
    ["Aug 2024", "EU AI Act enters into force", "Regulation", "World's first comprehensive AI regulation"],
    ["Oct 2024", "NIS2 transposition deadline (missed by 23 states)", "Cybersecurity", "Fragmented implementation; EC infringement proceedings"],
    ["Dec 2024", "Cyber Resilience Act enters into force", "Cybersecurity", "Product cybersecurity by design"],
    ["Dec 2024", "First 7 AI Factories selected", "AI Infrastructure", "Finland, Germany, Greece, Italy, Luxembourg, Spain, Sweden"],
    ["Jan 2025", "DORA fully applicable", "Cybersecurity", "Financial sector digital resilience"],
    ["Feb 2025", "AI Act prohibitions and literacy apply", "Regulation", "First enforcement phase"],
    ["Feb 2025", "Macron announces €109B AI package", "AI Investment", "France's answer to US Stargate"],
    ["Mar 2025", "6 additional AI Factories selected", "AI Infrastructure", "Austria, Bulgaria, France, Germany, Poland, Slovenia"],
    ["May 2025", "SAFE adopted (€150B defence loans)", "Defence/Cyber", "Loans for cyber and other defence capabilities"],
    ["Aug 2025", "GPAI rules apply under AI Act", "Regulation", "General-purpose AI governance obligations"],
    ["Sep 2025", "Semicon Coalition declaration (27 states)", "Semiconductors", "Unanimous call for Chips Act 2.0"],
    ["Oct 2025", "6 more AI Factories + 13 Antennas selected", "AI Infrastructure", "Total: 19 factories across EU"],
    ["Nov 2025", "Frontier AI Initiative launched", "AI Research", "France-Germany-EC joint non-profit for frontier AI"],
    ["Q1 2026", "Mistral AI Series C ($1.5B at ~$15B)", "LLMs", "Europe's leading LLM crosses $1B ARR"],
    ["Mar 2026", "HammerHAI contract signed (Germany)", "AI Infrastructure", "15+ ExaFLOPS, NVIDIA GB200"],
    ["Apr 2026", "Cohere acquires Aleph Alpha", "LLMs", "Transatlantic sovereign AI entity ($20B)"],
    ["Apr 2026", "EU awards €180M sovereign cloud contract", "Cloud", "First major EU sovereign cloud procurement"],
    ["Apr 2026", "IT4LIA contract signed (Italy)", "AI Infrastructure", "160+ ExaFLOPS, largest EU AI system"],
    ["Apr 2026", "UK launches £500M Sovereign AI Fund", "AI Investment", "UK national AI sovereignty programme"],
    ["May 2026", "Digital Omnibus agreement (AI Act deferral)", "Regulation", "High-risk AI obligations delayed to Dec 2027"],
    ["Jun 2026", "Tech Sovereignty Package adopted", "Policy", "Chips Act 2.0 + CADA + Open Source Strategy + Energy AI Roadmap"],
    ["Sep 2026", "CRA reporting obligations begin", "Cybersecurity", "Mandatory vulnerability/incident reporting for manufacturers"],
    ["Aug 2026", "AI Act majority rules apply", "Regulation", "Transparency, high-risk (Annex III per baseline), sandboxes"],
    ["2027-2030", "InvestAI €200B mobilisation window", "AI Investment", "Core funding period for EU AI sovereignty"],
    ["Dec 2027", "CRA full application", "Cybersecurity", "All product cybersecurity obligations apply"],
    ["Dec 2027", "AI Act high-risk (Annex III) apply (if Omnibus adopted)", "Regulation", "Deferred from Aug 2026"],
    ["2030", "Digital Decade targets deadline", "Policy", "20% chips target; triple DC capacity; digital sovereignty"],
]

for r_idx, row in enumerate(timeline_data, 1):
    for c_idx, val in enumerate(row, 1):
        ws11.cell(row=r_idx, column=c_idx, value=val)
    if r_idx == 1:
        style_header(ws11, r_idx, len(row))
    else:
        style_row(ws11, r_idx, len(row))

auto_width(ws11)

# Save
output_path = "/workspace/europe_self_sufficiency_llms_cybersecurity.xlsx"
wb.save(output_path)
print(f"Excel file saved to: {output_path}")
