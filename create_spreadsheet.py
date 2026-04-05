#!/usr/bin/env python3
"""Generate the Agentic Install Base spreadsheet.

Run:  python3 create_spreadsheet.py
Requires: openpyxl  (pip install openpyxl)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────────────────────────
header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
subheader_font = Font(name="Calibri", bold=True, size=11, color="1F3864")
title_font = Font(name="Calibri", bold=True, size=14, color="1F3864")
section_font = Font(name="Calibri", bold=True, size=11, color="2F5496")
section_fill = PatternFill(start_color="E9EFF7", end_color="E9EFF7", fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
wrap_alignment = Alignment(wrap_text=True, vertical="top")
center_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def style_data(ws, row, cols, is_section=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = thin_border
        cell.alignment = wrap_alignment if c > 1 else Alignment(vertical="top", wrap_text=True)
        if is_section:
            cell.font = section_font
            cell.fill = section_fill


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1 – Summary Comparison
# ═══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Summary Comparison"

ws1.merge_cells("A1:H1")
ws1["A1"] = "Agentic AI Install Base Comparison – April 2026"
ws1["A1"].font = title_font
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 30

headers = [
    "Product / Platform",
    "Category",
    "Parent Company",
    "Total Users / Install Base",
    "Paid / Subscribed Users",
    "Enterprise Adoption",
    "Key Growth Metric",
    "Data As-Of",
]
for c, h in enumerate(headers, 1):
    ws1.cell(row=3, column=c, value=h)
style_header(ws1, 3, len(headers))

data = [
    # ── Agentic Browsers ──
    ["AGENTIC BROWSERS", "", "", "", "", "", "", ""],

    [
        "Dia (The Browser Company)",
        "Agentic Browser",
        "The Browser Company\n(acq. Atlassian, Sep 2025, $610M)",
        "Not publicly disclosed",
        "Free (no paid tier)",
        "Enterprise integration with Jira/Linear planned post-acquisition",
        "Atlassian acquisition at $610M validates product",
        "Q1 2026",
    ],
    [
        "ChatGPT Atlas",
        "Agentic Browser",
        "OpenAI",
        "Part of 900M weekly ChatGPT users; standalone browser downloads not disclosed",
        "Included with ChatGPT Plus/Pro ($20–$200/mo)",
        "Available to 1M+ business customers",
        "macOS launch Oct 2025; Windows & mobile pending",
        "Feb 2026",
    ],
    [
        "Perplexity Comet",
        "Agentic Browser",
        "Perplexity AI",
        "~45M monthly active Perplexity users; Comet-specific downloads not disclosed",
        "Comet is free; Perplexity Pro $20/mo",
        "Growing; 147M monthly web visits",
        "Made free worldwide Oct 2025; 100%+ YoY user growth",
        "Q1 2026",
    ],
    [
        "Microsoft Edge Copilot Mode",
        "Agentic Browser (add-on)",
        "Microsoft",
        "Part of Edge's ~700M+ user base",
        "Bundled with Edge (free); M365 Copilot $30/user/mo",
        "90%+ Fortune 500 use Edge; 15M paid M365 Copilot seats",
        "Incremental approach; 33M active Copilot users across surfaces",
        "Q2 FY2026",
    ],

    # ── Agentic Coding / AI Assistants ──
    ["AGENTIC CODING & AI ASSISTANTS", "", "", "", "", "", "", ""],

    [
        "Claude Code (Anthropic)",
        "Agentic Coding Tool (CLI)",
        "Anthropic",
        "115K+ developers (Jul 2025); 109K+ GitHub stars",
        "Pay-per-use via Anthropic API; included in Claude Pro/Team",
        "69% market share among AI coding tools (Jan 2026 ACTI Index)",
        "195M lines of code/week; 4% of public GitHub commits",
        "Jan 2026",
    ],
    [
        "OpenAI Operator / ChatGPT Agent",
        "Agentic AI Assistant / Browser Agent",
        "OpenAI",
        "Part of 900M weekly ChatGPT users",
        "50M+ paid ChatGPT subscribers; Operator bundled with Plus/Pro/Team",
        "9M+ paying business users; 1M+ business customers",
        "87% web navigation success rate; 350% user growth in 18 months",
        "Feb 2026",
    ],
    [
        "OpenAI Codex",
        "Agentic Coding Agent",
        "OpenAI",
        "2M weekly users",
        "Included with ChatGPT Plus/Pro/Team",
        "4M developers building on OpenAI platform",
        "5x growth in 3 months (to Feb 2026)",
        "Feb 2026",
    ],
    [
        "Microsoft Copilot (M365)",
        "AI Productivity Assistant",
        "Microsoft",
        "33M active users across Copilot surfaces",
        "15M paid M365 Copilot seats (160% YoY growth)",
        "70%+ Fortune 500 evaluating/deployed; 3.3% conversion of 450M M365 base",
        "Daily active users up ~10x YoY; avg conversations/user doubled",
        "Q2 FY2026",
    ],
    [
        "GitHub Copilot",
        "AI Coding Assistant (IDE plugin)",
        "Microsoft / GitHub",
        "~20M total users (incl. trials)",
        "4.7M paid subscribers (75% YoY growth)",
        "~90% Fortune 100 companies",
        "Copilot-authored code: ~30% of new code at adopting orgs",
        "Jan 2026",
    ],
    [
        "Cursor (Anysphere)",
        "AI-Native IDE",
        "Anysphere",
        "1M+ monthly active users",
        "Est. 800K+ paid (based on $2B ARR at $20/mo avg)",
        "70%+ Fortune 500 testing or deploying; enterprise = 60% of revenue",
        "$2B ARR (Mar 2026), doubled from $1B in 3 months",
        "Mar 2026",
    ],
]

row = 4
for d in data:
    is_section = d[1] == "" and d[0].isupper()
    for c, v in enumerate(d, 1):
        ws1.cell(row=row, column=c, value=v)
    style_data(ws1, row, len(headers), is_section=is_section)
    if is_section:
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
    row += 1

col_widths = [32, 26, 28, 32, 30, 34, 38, 14]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
for r in range(4, row):
    ws1.row_dimensions[r].height = 48

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2 – Agentic Browsers Deep-Dive
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Agentic Browsers Detail")

ws2.merge_cells("A1:F1")
ws2["A1"] = "Agentic Browsers – Detailed Comparison"
ws2["A1"].font = title_font
ws2["A1"].alignment = Alignment(horizontal="center")
ws2.row_dimensions[1].height = 28

b_headers = [
    "Browser",
    "Launch Date",
    "Platform Availability",
    "Pricing",
    "Key Capabilities",
    "Install Base / Adoption Indicators",
]
for c, h in enumerate(b_headers, 1):
    ws2.cell(row=3, column=c, value=h)
style_header(ws2, 3, len(b_headers))

browsers = [
    [
        "Dia (The Browser Company)",
        "2025 (public beta)",
        "macOS, Windows (Chromium-based)",
        "Free",
        "Tab conversations, contextual page understanding, research assistance, built-in AI without extensions",
        "Atlassian acquired for $610M (Sep 2025). Arc predecessor had ~1M users. Specific Dia download numbers not disclosed.",
    ],
    [
        "ChatGPT Atlas (OpenAI)",
        "Oct 2025",
        "macOS (Windows & mobile planned)",
        "Included with ChatGPT Plus ($20/mo) and Pro ($200/mo)",
        "Full agentic browsing, deep ChatGPT integration, autonomous web tasks, multi-step research",
        "Part of 900M weekly ChatGPT ecosystem. Browser-specific installs not disclosed. Available to 50M+ paid subscribers.",
    ],
    [
        "Perplexity Comet",
        "Oct 2025 (free worldwide)",
        "macOS, Windows, iOS, Android",
        "Free (Perplexity Pro $20/mo for advanced features)",
        "AI-powered search-first browsing, workflow automation, real-time web answers, answer-engine integration",
        "Millions joined waitlist pre-launch. Parent Perplexity has 45M MAU and 147M monthly visits. Comet-specific numbers not disclosed.",
    ],
    [
        "Microsoft Edge Copilot Mode",
        "2025 (incremental rollout)",
        "Windows, macOS, iOS, Android",
        "Free (bundled with Edge)",
        "AI sidebar, page summarization, content generation, shopping assistant, Copilot chat integration",
        "Edge has ~700M+ users. 33M active Copilot users across surfaces. 15M paid M365 Copilot seats.",
    ],
    [
        "Opera Neon (AI Browser)",
        "2025 (beta)",
        "macOS, Windows",
        "Free",
        "AI-native interface, contextual tab management, multi-model AI integration",
        "Opera has ~380M+ global users. Neon-specific adoption data not publicly available.",
    ],
]

for r_idx, b in enumerate(browsers, 4):
    for c, v in enumerate(b, 1):
        ws2.cell(row=r_idx, column=c, value=v)
    style_data(ws2, r_idx, len(b_headers))
    ws2.row_dimensions[r_idx].height = 72

b_widths = [30, 18, 30, 28, 50, 55]
for i, w in enumerate(b_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3 – Market Context & Growth
# ═══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Market Context")

ws3.merge_cells("A1:D1")
ws3["A1"] = "Market Context & Growth Trends"
ws3["A1"].font = title_font
ws3["A1"].alignment = Alignment(horizontal="center")
ws3.row_dimensions[1].height = 28

m_headers = ["Metric", "Value", "Source / Context", "Date"]
for c, h in enumerate(m_headers, 1):
    ws3.cell(row=3, column=c, value=h)
style_header(ws3, 3, len(m_headers))

market_data = [
    ["OVERALL AGENTIC AI MARKET", "", "", ""],
    ["Global AI Browser Market (2024)", "$4.5 billion", "Market.us", "2024"],
    ["Projected AI Browser Market (2034)", "$76.8 billion", "Market.us (32.8% CAGR)", "2034 projection"],
    ["Enterprise Agentic Browser Adoption", "27.7% of enterprises in production", "Industry surveys", "Q1 2026"],
    ["Traditional Search Volume Decline", "-25% projected", "Gartner forecast", "By 2026"],
    ["AI Code Tools Market (2023)", "$4.86 billion", "Industry reports", "2023"],
    ["AI Code Tools Market (projected 2030)", "$26 billion", "Industry projections", "2030"],
    ["", "", "", ""],
    ["CONSUMER AI PLATFORM TRAFFIC (Monthly)", "", "", ""],
    ["ChatGPT", "4.1 billion monthly visits", "Web analytics", "Feb 2026"],
    ["Bing (Copilot)", "3.0 billion monthly visits", "Web analytics", "Feb 2026"],
    ["Google Gemini", "1.5 billion monthly visits", "Web analytics", "Feb 2026"],
    ["Anthropic (Claude)", "181 million monthly visits", "Web analytics", "Feb 2026"],
    ["Perplexity AI", "147 million monthly visits", "Web analytics", "Feb 2026"],
    ["", "", "", ""],
    ["KEY MILESTONES", "", "", ""],
    ["OpenAI – 900M weekly users", "Milestone reached Feb 2026", "OpenAI announcement", "Feb 2026"],
    ["OpenAI – 50M paid subscribers", "Consumer paying users", "OpenAI announcement", "Feb 2026"],
    ["Microsoft – 15M paid M365 Copilot seats", "160% YoY growth", "Microsoft earnings", "Q2 FY2026"],
    ["Claude Code – 69% coding market share", "ACTI Index survey", "ACTI January 2026 Report", "Jan 2026"],
    ["Cursor – $2B ARR", "Doubled in 3 months from $1B", "TechCrunch", "Mar 2026"],
    ["GitHub Copilot – 4.7M paid subscribers", "75% YoY growth", "GitHub / Microsoft", "Jan 2026"],
    ["Atlassian acquires Dia", "$610M acquisition", "The Browser Company", "Sep 2025"],
]

row = 4
for d in market_data:
    is_section = d[1] == "" and d[0] != ""
    for c, v in enumerate(d, 1):
        ws3.cell(row=row, column=c, value=v)
    style_data(ws3, row, len(m_headers), is_section=is_section)
    if is_section:
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(m_headers))
    row += 1

m_widths = [42, 36, 36, 20]
for i, w in enumerate(m_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4 – Sources & Methodology
# ═══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Sources & Notes")

ws4.merge_cells("A1:C1")
ws4["A1"] = "Sources & Methodology Notes"
ws4["A1"].font = title_font
ws4["A1"].alignment = Alignment(horizontal="center")
ws4.row_dimensions[1].height = 28

s_headers = ["#", "Source", "Notes"]
for c, h in enumerate(s_headers, 1):
    ws4.cell(row=3, column=c, value=h)
style_header(ws4, 3, len(s_headers))

sources = [
    ["1", "OpenAI Official Blog & Announcements", "900M weekly users, 50M subscribers, Operator/Agent details, Codex 2M weekly users (Feb 2026)"],
    ["2", "Microsoft Earnings (Q2 FY2026)", "15M paid M365 Copilot seats, 33M active users, 160% YoY growth"],
    ["3", "ACTI Index – January 2026 Report", "Claude Code 69% market share, developer adoption survey across experience levels"],
    ["4", "ppc.land / Anthropic", "Claude Code 115K developers, 195M lines/week (Jul 2025)"],
    ["5", "GitHub / Microsoft", "4.7M paid Copilot subscribers, ~20M total users, 90% Fortune 100 adoption (Jan 2026)"],
    ["6", "TechCrunch / Anysphere", "Cursor $2B ARR (Mar 2026), 1M+ MAU, 70%+ Fortune 500 evaluation"],
    ["7", "Market.us", "AI browser market sizing: $4.5B (2024) to $76.8B (2034), 32.8% CAGR"],
    ["8", "Gartner", "Traditional search volume decline -25% by 2026 forecast"],
    ["9", "Web Analytics (SimilarWeb / Semrush)", "Monthly visit data for ChatGPT, Bing, Gemini, Claude, Perplexity"],
    ["10", "Business of Apps / DemandSage / Statista", "Perplexity 45M MAU, growth trajectory data"],
    ["11", "The Browser Company / Atlassian", "Dia browser details, $610M acquisition (Sep 2025)"],
    ["12", "Various industry publications", "Opera, Edge Copilot Mode, Atlas, Comet feature comparisons"],
]

for r_idx, s in enumerate(sources, 4):
    for c, v in enumerate(s, 1):
        ws4.cell(row=r_idx, column=c, value=v)
    style_data(ws4, r_idx, len(s_headers))
    ws4.row_dimensions[r_idx].height = 36

notes_row = 4 + len(sources) + 1
ws4.cell(row=notes_row, column=1, value="METHODOLOGY NOTES")
ws4.cell(row=notes_row, column=1).font = section_font
ws4.cell(row=notes_row, column=1).fill = section_fill
ws4.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=3)

notes = [
    "Data compiled as of April 2026 from publicly available sources.",
    "'Install base' for agentic browsers is difficult to determine precisely as most products are new and do not disclose standalone download figures.",
    "User counts may overlap (e.g., ChatGPT Atlas users are a subset of ChatGPT's 900M weekly users).",
    "Market share figures from ACTI Index are based on developer surveys and may not reflect enterprise-only usage.",
    "Revenue figures (e.g., Cursor $2B ARR) are annualized run-rate figures, not trailing twelve-month revenue.",
    "Paid subscriber counts include individual and enterprise/team seats unless otherwise noted.",
]

for i, note in enumerate(notes):
    r = notes_row + 1 + i
    ws4.cell(row=r, column=1, value=note)
    ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws4.cell(row=r, column=1).alignment = wrap_alignment
    ws4.row_dimensions[r].height = 24

s_widths = [6, 40, 80]
for i, w in enumerate(s_widths, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

# ── Save ────────────────────────────────────────────────────────────────────
output_path = "/workspace/agentic_install_base_comparison.xlsx"
wb.save(output_path)
print(f"Spreadsheet saved to {output_path}")
