import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

wb = Workbook()

header_font = Font(bold=True, size=11)
title_font = Font(bold=True, size=14)
subtitle_font = Font(bold=True, size=12)
currency_format = '#,##0.0'
pct_format = '0.0%'
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
openai_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
anthropic_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border


def auto_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)


# =============================================================================
# Sheet 1: Executive Summary
# =============================================================================
ws = wb.active
ws.title = "Executive Summary"

ws.cell(row=1, column=1, value="OpenAI & Anthropic ARR Analysis").font = title_font
ws.cell(row=2, column=1, value="Token Spend by Industry & Customer").font = subtitle_font
ws.cell(row=3, column=1, value="Data as of: June 2026 | Sources: Sacra, Menlo Ventures, Ramp, Presenc AI, VentureBeat, The Information")

row = 5
headers = ["Metric", "OpenAI", "Anthropic", "Notes"]
for c, h in enumerate(headers, 1):
    ws.cell(row=row, column=c, value=h)
style_header_row(ws, row, len(headers))

data = [
    ["Latest ARR (Jun 2026 est.)", "$33B", "$45B", "The Information (May/Jun 2026); annualized run-rate"],
    ["Confirmed ARR (Apr/May 2026)", "$25B", "$30B", "Company-confirmed figures"],
    ["End-2025 ARR", "$20B", "$9B", "CFO-confirmed (OpenAI); Company-confirmed (Anthropic)"],
    ["End-2024 ARR", "$6B", "$1B", ""],
    ["End-2023 ARR", "$2B", "$100M", ""],
    ["YoY Growth Rate (2025→2026)", "~3x", "~10x", "Based on confirmed run-rates"],
    ["Enterprise Revenue Share", "~40%", "~85%", "Enterprise + developer customers"],
    ["Enterprise LLM Market Share (2025)", "27%", "40%", "Menlo Ventures estimate"],
    ["Fortune 500 Penetration", "92%", "8 of Fortune 10", ""],
    ["Customers >$1M/yr spend", "Not disclosed", "1,000+", "Doubled in 2 months (Feb→Apr 2026)"],
    ["Total Business Customers", "1M+", "300K+", ""],
    ["API Throughput", "15B tokens/min", "Not disclosed", "OpenAI Mar 2026"],
    ["Gross Margin", "33%", "Not disclosed", ""],
    ["Projected 2026 Cash Burn", "$27B", "Not disclosed", ""],
]

for i, d in enumerate(data, 1):
    for c, v in enumerate(d, 1):
        cell = ws.cell(row=row + i, column=c, value=v)
        cell.border = thin_border
        if c == 2:
            cell.fill = openai_fill
        elif c == 3:
            cell.fill = anthropic_fill

auto_width(ws)

# =============================================================================
# Sheet 2: Revenue Breakdown by Segment
# =============================================================================
ws2 = wb.create_sheet("Revenue Segments")

ws2.cell(row=1, column=1, value="OpenAI Revenue Breakdown (May 2026 - $25B ARR)").font = subtitle_font
row = 3
headers = ["Revenue Segment", "Est. Share", "Est. ARR ($B)", "Growth Driver"]
for c, h in enumerate(headers, 1):
    ws2.cell(row=row, column=c, value=h)
style_header_row(ws2, row, len(headers))

openai_segments = [
    ["ChatGPT Plus, Pro & Team Subscriptions", 0.48, 12.0, "900M+ weekly users, 50M+ paying subscribers"],
    ["OpenAI API (Developer + Business)", 0.22, 5.5, "4M+ developers, 15B tokens/min throughput"],
    ["ChatGPT Enterprise & Edu", 0.16, 4.0, "9x YoY seat growth, 92% Fortune 500 penetration"],
    ["OpenAI for Business (Custom GPT, Agents)", 0.06, 1.5, "Agentic workflows, custom deployments"],
    ["Microsoft Revenue Share & OEM", 0.05, 1.25, "Capped at $38B through 2030"],
    ["Other (Sora, Ads pilot, etc.)", 0.03, 0.75, "Ads pilot reached $100M+ ARR in 6 weeks"],
    ["TOTAL", 1.00, 25.0, ""],
]

for i, d in enumerate(openai_segments, 1):
    ws2.cell(row=row + i, column=1, value=d[0]).border = thin_border
    cell = ws2.cell(row=row + i, column=2, value=d[1])
    cell.number_format = pct_format
    cell.border = thin_border
    cell = ws2.cell(row=row + i, column=3, value=d[2])
    cell.number_format = currency_format
    cell.border = thin_border
    ws2.cell(row=row + i, column=4, value=d[3]).border = thin_border
    if d[0] == "TOTAL":
        for c in range(1, 5):
            ws2.cell(row=row + i, column=c).font = header_font

row2 = row + len(openai_segments) + 3
ws2.cell(row=row2, column=1, value="Anthropic Revenue Breakdown (Apr 2026 - $30B ARR)").font = subtitle_font
row2 += 2
headers2 = ["Revenue Segment", "Est. Share", "Est. ARR ($B)", "Growth Driver"]
for c, h in enumerate(headers2, 1):
    ws2.cell(row=row2, column=c, value=h)
style_header_row(ws2, row2, len(headers2))

anthropic_segments = [
    ["Enterprise API (direct + cloud partners)", 0.55, 16.5, "1,000+ customers at $1M+/yr; AWS Bedrock, GCP Vertex, Azure Foundry"],
    ["Claude Code", 0.085, 2.5, "$2.5B ARR in 9 months; developers & engineering teams"],
    ["Claude Enterprise Seats (CoWork)", 0.15, 4.5, "Deloitte 470K, Cognizant 350K, Accenture 30K trained"],
    ["Consumer (Claude Pro/Max)", 0.15, 4.5, "Growing but secondary; 15% of revenue"],
    ["Claude Partner Network", 0.035, 1.0, "$100M Anthropic investment; certified architects"],
    ["Other (Claude for Teams, integrations)", 0.03, 1.0, "Salesforce Agentforce, platform integrations"],
    ["TOTAL", 1.00, 30.0, ""],
]

for i, d in enumerate(anthropic_segments, 1):
    ws2.cell(row=row2 + i, column=1, value=d[0]).border = thin_border
    cell = ws2.cell(row=row2 + i, column=2, value=d[1])
    cell.number_format = pct_format
    cell.border = thin_border
    cell = ws2.cell(row=row2 + i, column=3, value=d[2])
    cell.number_format = currency_format
    cell.border = thin_border
    ws2.cell(row=row2 + i, column=4, value=d[3]).border = thin_border
    if d[0] == "TOTAL":
        for c in range(1, 5):
            ws2.cell(row=row2 + i, column=c).font = header_font

auto_width(ws2)

# =============================================================================
# Sheet 3: Token Spend by Industry
# =============================================================================
ws3 = wb.create_sheet("Industry Token Spend")

ws3.cell(row=1, column=1, value="Estimated Token Spend by Industry Vertical").font = title_font
ws3.cell(row=2, column=1, value="Based on: Menlo Ventures ($37B enterprise gen AI spend 2025), Ramp AI Index, TechnologyChecker, OpenAI Enterprise Report")

row = 4
headers = [
    "Industry", "Total Enterprise AI Spend 2025 ($B)",
    "OpenAI Share", "OpenAI Est. ($B)",
    "Anthropic Share", "Anthropic Est. ($B)",
    "AI Adoption Level", "Provider Preference",
    "Key Use Cases"
]
for c, h in enumerate(headers, 1):
    ws3.cell(row=row, column=c, value=h)
style_header_row(ws3, row, len(headers))

industry_data = [
    ["Technology / Software", 8.5, 0.25, 2.13, 0.50, 4.25, "Very High", "Anthropic", "Coding (Claude Code), product development, DevOps automation"],
    ["Financial Services", 6.5, 0.30, 1.95, 0.45, 2.93, "Very High", "Anthropic", "Wealth mgmt assistants, compliance, risk analysis, trading ops"],
    ["Professional Services / Consulting", 5.0, 0.30, 1.50, 0.45, 2.25, "Very High", "Anthropic", "Audit support, tax analysis, strategy docs, client deliverables"],
    ["Healthcare & Life Sciences", 4.0, 0.35, 1.40, 0.30, 1.20, "High (growing 8x)", "Slight OpenAI", "Clinical documentation, drug discovery, ambient scribes, HIPAA workflows"],
    ["Manufacturing & Industrial", 3.0, 0.30, 0.90, 0.30, 0.90, "High (growing 7x)", "Even split", "Supply chain optimization, quality control, predictive maintenance"],
    ["Retail & E-commerce", 2.5, 0.35, 0.88, 0.25, 0.63, "Medium", "Slight OpenAI", "Product descriptions, customer service, demand forecasting"],
    ["Media & Entertainment", 2.0, 0.30, 0.60, 0.35, 0.70, "Medium-High", "Slight Anthropic", "Content generation, personalization, copyright analysis"],
    ["Telecommunications", 1.8, 0.25, 0.45, 0.40, 0.72, "Medium-High", "Anthropic", "Customer service automation, network ops, billing analysis"],
    ["Legal", 1.5, 0.20, 0.30, 0.50, 0.75, "High", "Strong Anthropic", "Contract review, discovery, compliance monitoring, legal research"],
    ["Education", 1.2, 0.40, 0.48, 0.25, 0.30, "Medium", "OpenAI", "Tutoring, content creation, grading, research assistance"],
    ["Government & Defense", 1.0, 0.35, 0.35, 0.20, 0.20, "Medium", "OpenAI", "Document processing, citizen services, policy analysis"],
    ["Transportation & Logistics", 0.8, 0.30, 0.24, 0.30, 0.24, "Medium", "Even split", "Route optimization, fleet management, supplier risk"],
    ["Real Estate", 0.5, 0.35, 0.18, 0.25, 0.13, "Low-Medium", "Slight OpenAI", "Listings, market analysis, property valuations"],
    ["Construction", 0.4, 0.40, 0.16, 0.15, 0.06, "Low", "OpenAI", "Project planning, safety docs, bid estimation"],
    ["Agriculture", 0.3, 0.35, 0.11, 0.15, 0.05, "Low", "OpenAI", "Crop analysis, weather modeling, supply chain"],
    ["Hospitality & Travel", 0.5, 0.35, 0.18, 0.25, 0.13, "Low-Medium", "Slight OpenAI", "Booking assistants, customer service, dynamic pricing"],
]

for i, d in enumerate(industry_data, 1):
    r = row + i
    ws3.cell(row=r, column=1, value=d[0]).border = thin_border
    cell = ws3.cell(row=r, column=2, value=d[1])
    cell.number_format = currency_format
    cell.border = thin_border
    cell = ws3.cell(row=r, column=3, value=d[2])
    cell.number_format = pct_format
    cell.border = thin_border
    cell = ws3.cell(row=r, column=4, value=d[3])
    cell.number_format = currency_format
    cell.border = thin_border
    cell.fill = openai_fill
    cell = ws3.cell(row=r, column=5, value=d[4])
    cell.number_format = pct_format
    cell.border = thin_border
    cell = ws3.cell(row=r, column=6, value=d[5])
    cell.number_format = currency_format
    cell.border = thin_border
    cell.fill = anthropic_fill
    ws3.cell(row=r, column=7, value=d[6]).border = thin_border
    ws3.cell(row=r, column=8, value=d[7]).border = thin_border
    ws3.cell(row=r, column=9, value=d[8]).border = thin_border

total_row = row + len(industry_data) + 1
ws3.cell(row=total_row, column=1, value="TOTAL ESTIMATED").font = header_font
ws3.cell(row=total_row, column=2, value=sum(d[1] for d in industry_data)).number_format = currency_format
ws3.cell(row=total_row, column=4, value=sum(d[3] for d in industry_data)).number_format = currency_format
ws3.cell(row=total_row, column=6, value=sum(d[5] for d in industry_data)).number_format = currency_format
for c in range(1, 10):
    ws3.cell(row=total_row, column=c).font = header_font
    ws3.cell(row=total_row, column=c).border = thin_border

auto_width(ws3)

# =============================================================================
# Sheet 4: Named Enterprise Customers
# =============================================================================
ws4 = wb.create_sheet("Enterprise Customers")

ws4.cell(row=1, column=1, value="Major Enterprise Customers by Provider & Industry").font = title_font
ws4.cell(row=2, column=1, value="Sources: OpenAI announcements, Anthropic partner releases, press reports (2024-2026)")

row = 4
headers = ["Company", "Industry", "Provider", "Deployment Scale", "Use Case", "Est. Annual Spend"]
for c, h in enumerate(headers, 1):
    ws4.cell(row=row, column=c, value=h)
style_header_row(ws4, row, len(headers))

customers = [
    # OpenAI customers
    ["Morgan Stanley", "Financial Services", "OpenAI", "Wealth Mgmt division (15,000+ advisors)", "AI Assistant for advisors, meeting debrief, research retrieval", "$50M-100M+"],
    ["PwC", "Professional Services", "OpenAI", "100,000+ seats (entire global workforce)", "Audit, tax, advisory workflow automation", "$100M+"],
    ["Coca-Cola", "Consumer Goods", "OpenAI", "Multi-market deployment", "Creative content generation, marketing copy, ad workflows", "$20M-50M"],
    ["Stripe", "Financial Technology", "OpenAI", "Developer platform integration", "Documentation search, customer insights, content moderation", "$10M-30M"],
    ["Cisco", "Technology", "OpenAI", "Enterprise-wide", "Internal knowledge mgmt, customer support, product docs", "$20M-50M"],
    ["T-Mobile", "Telecommunications", "OpenAI", "Customer service + internal", "Customer service automation, network troubleshooting", "$15M-40M"],
    ["Target", "Retail", "OpenAI", "Store operations + corporate", "Inventory planning, customer experience, supply chain", "$10M-30M"],
    ["Lowe's", "Retail", "OpenAI", "1,700+ stores", "Mylow Companion in-store assistant for associates", "$15M-40M"],
    ["Amgen", "Pharmaceuticals", "OpenAI", "R&D + operations", "Drug discovery, clinical trial optimization, regulatory docs", "$20M-50M"],
    ["Thermo Fisher Scientific", "Life Sciences", "OpenAI", "Enterprise-wide", "Lab automation, research synthesis, quality control", "$15M-30M"],
    ["Booking.com", "Travel & Hospitality", "OpenAI", "Customer-facing + internal", "Travel planning AI, customer service, personalization", "$20M-50M"],
    ["Commonwealth Bank", "Financial Services", "OpenAI", "Enterprise-wide", "Customer service, fraud detection, compliance", "$10M-30M"],
    ["Shopify", "E-commerce", "OpenAI", "Platform integration", "Agentic Commerce Protocol, merchant tools", "$10M-20M"],
    ["Walmart", "Retail", "OpenAI", "Platform integration", "Agentic shopping experiences, supply chain", "$20M-50M"],
    ["Spotify", "Media/Entertainment", "OpenAI", "Platform integration", "Personalization, content discovery, playlist curation", "$10M-20M"],
    ["Intercom", "Technology (SaaS)", "OpenAI", "Core product (Fin agent)", "Customer service AI agent powering entire product", "$5M-15M"],
    # Anthropic customers
    ["Deloitte", "Professional Services", "Anthropic", "470,000 employees globally (150 countries)", "Audit, consulting, advisory — largest enterprise AI rollout", "$200M+"],
    ["Accenture", "Professional Services", "Anthropic", "30,000 Claude-trained professionals", "Anthropic Business Group — regulated industry consulting", "$100M+"],
    ["Cognizant", "IT Services", "Anthropic", "350,000 employees", "Agent Foundry, engineering platform, financial services vertical", "$100M+"],
    ["Uber", "Technology / Transportation", "Anthropic", "Enterprise-wide", "Operations optimization, customer service, driver tools", "$30M-75M"],
    ["Netflix", "Media / Entertainment", "Anthropic", "Enterprise-wide", "Content analysis, personalization, internal productivity", "$20M-50M"],
    ["IBM", "Technology", "Anthropic", "Internal + Watsonx integration", "HR chatbots, enterprise Watsonx deployments", "$30M-75M"],
    ["Infosys", "IT Services", "Anthropic", "Topaz AI Platform + Center of Excellence", "Telecom, financial services, manufacturing verticals", "$50M-100M"],
    ["TELUS", "Telecommunications", "Anthropic", "57,000 employees", "Customer service, network ops, internal productivity", "$20M-50M"],
    ["Cleary Gottlieb", "Legal", "Anthropic", "Firm-wide", "Agentic legal workflows, complex document synthesis", "$5M-15M"],
    ["Epic Systems", "Healthcare Technology", "Anthropic", "Platform integration", "Clinical documentation, EHR assistance, HIPAA workflows", "$10M-30M"],
    ["United Airlines", "Aviation / Travel", "Anthropic (via Slalom/AWS)", "Customer-facing", "AI-powered flight update customization", "$5M-15M"],
    ["Salesforce", "Technology (CRM)", "Anthropic", "Agentforce platform", "Foundation model for autonomous AI agent platform", "$50M-100M"],
    ["KPMG", "Professional Services", "Anthropic", "Global deployment", "Audit, tax, advisory automation", "$50M-100M"],
    ["Slalom", "Consulting", "Anthropic", "Partner deployment", "Ethical AI deployment, customer implementations", "$10M-30M"],
]

for i, d in enumerate(customers, 1):
    r = row + i
    for c, v in enumerate(d, 1):
        cell = ws4.cell(row=r, column=c, value=v)
        cell.border = thin_border
        if d[2] == "OpenAI":
            if c == 3:
                cell.fill = openai_fill
        elif d[2] == "Anthropic":
            if c == 3:
                cell.fill = anthropic_fill

auto_width(ws4)

# =============================================================================
# Sheet 5: Market Share & Competitive Dynamics
# =============================================================================
ws5 = wb.create_sheet("Market Dynamics")

ws5.cell(row=1, column=1, value="Enterprise LLM Market Share & Competitive Dynamics").font = title_font

row = 3
ws5.cell(row=row, column=1, value="Enterprise LLM Market Share Evolution").font = subtitle_font
row += 1
headers = ["Year", "OpenAI", "Anthropic", "Google", "Others", "Total Enterprise Gen AI Spend"]
for c, h in enumerate(headers, 1):
    ws5.cell(row=row, column=c, value=h)
style_header_row(ws5, row, len(headers))

market_share = [
    [2023, 0.50, 0.05, 0.20, 0.25, "$1.7B"],
    [2024, 0.40, 0.20, 0.20, 0.20, "$11.5B"],
    [2025, 0.27, 0.40, 0.21, 0.12, "$37B"],
    ["2026E", 0.25, 0.45, 0.20, 0.10, "$70B+ (projected)"],
]

for i, d in enumerate(market_share, 1):
    r = row + i
    ws5.cell(row=r, column=1, value=d[0]).border = thin_border
    for c in range(2, 5):
        cell = ws5.cell(row=r, column=c, value=d[c - 1])
        cell.number_format = pct_format
        cell.border = thin_border
    ws5.cell(row=r, column=5, value=d[4]).border = thin_border
    ws5.cell(row=r, column=6, value=d[5]).border = thin_border

row2 = row + len(market_share) + 3
ws5.cell(row=row2, column=1, value="Provider Preference by Industry (Ramp AI Index, Feb 2026)").font = subtitle_font
row2 += 1
headers2 = ["Industry", "AI Adoption Level", "Anthropic Share of AI Spend", "OpenAI Share of AI Spend", "Trend"]
for c, h in enumerate(headers2, 1):
    ws5.cell(row=row2, column=c, value=h)
style_header_row(ws5, row2, len(headers2))

ramp_data = [
    ["Information / Technology", "Highest", 0.70, 0.30, "Strongly Anthropic — widest lead"],
    ["Finance & Insurance", "Very High", 0.65, 0.35, "Anthropic preference — compliance-driven"],
    ["Professional Services", "Very High", 0.65, 0.35, "Anthropic preference — reasoning quality"],
    ["Education", "Medium", 0.45, 0.55, "Slight OpenAI — brand recognition"],
    ["Manufacturing", "Medium", 0.48, 0.52, "Near parity — early adopter phase"],
    ["Wholesale Trade", "Medium", 0.47, 0.53, "Near parity"],
    ["Retail", "Medium", 0.42, 0.58, "Slight OpenAI — Microsoft ecosystem"],
    ["Transportation", "Medium", 0.45, 0.55, "Near parity"],
    ["Real Estate", "Low-Medium", 0.40, 0.60, "OpenAI — brand recognition"],
    ["Healthcare", "Low (growing fast)", 0.38, 0.62, "OpenAI slight edge; Anthropic gaining via HIPAA"],
    ["Construction", "Low", 0.30, 0.70, "OpenAI — fragmented workforce"],
    ["Arts & Entertainment", "Low", 0.45, 0.55, "Near parity"],
    ["Agriculture", "Very Low", 0.25, 0.75, "OpenAI — basic adoption stage"],
    ["Hospitality", "Very Low", 0.30, 0.70, "OpenAI — Microsoft ecosystem on-ramp"],
]

for i, d in enumerate(ramp_data, 1):
    r = row2 + i
    ws5.cell(row=r, column=1, value=d[0]).border = thin_border
    ws5.cell(row=r, column=2, value=d[1]).border = thin_border
    cell = ws5.cell(row=r, column=3, value=d[2])
    cell.number_format = pct_format
    cell.border = thin_border
    cell.fill = anthropic_fill
    cell = ws5.cell(row=r, column=4, value=d[3])
    cell.number_format = pct_format
    cell.border = thin_border
    cell.fill = openai_fill
    ws5.cell(row=r, column=5, value=d[4]).border = thin_border

auto_width(ws5)

# =============================================================================
# Sheet 6: ARR Growth Timeline
# =============================================================================
ws6 = wb.create_sheet("ARR Timeline")

ws6.cell(row=1, column=1, value="ARR Growth Timeline: OpenAI vs Anthropic").font = title_font

row = 3
headers = ["Period", "OpenAI ARR ($B)", "Anthropic ARR ($B)", "OpenAI MoM Growth", "Anthropic MoM Growth"]
for c, h in enumerate(headers, 1):
    ws6.cell(row=row, column=c, value=h)
style_header_row(ws6, row, len(headers))

timeline = [
    ["Jan 2024", 3.0, 0.087, "", ""],
    ["Dec 2024", 6.0, 1.0, "", ""],
    ["Mid-2025", 13.0, 4.5, "", ""],
    ["End-2025", 20.0, 9.0, "", ""],
    ["Jan 2026", 22.0, 11.0, "10%", "22%"],
    ["Feb 2026", 25.0, 14.0, "14%", "27%"],
    ["Mar 2026", 24.0, 19.0, "-4%", "36%"],
    ["Apr 2026", 25.0, 30.0, "4%", "58%"],
    ["May 2026 (est.)", 33.0, 45.0, "32%", "50%"],
]

for i, d in enumerate(timeline, 1):
    r = row + i
    ws6.cell(row=r, column=1, value=d[0]).border = thin_border
    cell = ws6.cell(row=r, column=2, value=d[1])
    cell.number_format = currency_format
    cell.border = thin_border
    cell.fill = openai_fill
    cell = ws6.cell(row=r, column=3, value=d[2])
    cell.number_format = currency_format
    cell.border = thin_border
    cell.fill = anthropic_fill
    ws6.cell(row=r, column=4, value=d[3]).border = thin_border
    ws6.cell(row=r, column=5, value=d[4]).border = thin_border

auto_width(ws6)

# =============================================================================
# Sheet 7: Vertical AI Spend Detail (Menlo Ventures)
# =============================================================================
ws7 = wb.create_sheet("Vertical AI Detail")

ws7.cell(row=1, column=1, value="Enterprise Gen AI Spend Breakdown (Menlo Ventures 2025)").font = title_font
ws7.cell(row=2, column=1, value="Total Enterprise Gen AI Spend: $37B (2025) — up from $11.5B (2024) and $1.7B (2023)")

row = 4
headers = ["Category", "2025 Spend ($B)", "Share of Total", "Key Sub-segments"]
for c, h in enumerate(headers, 1):
    ws7.cell(row=row, column=c, value=h)
style_header_row(ws7, row, len(headers))

spend_categories = [
    ["Model APIs (Infrastructure)", 18.0, 0.49, "Anthropic, OpenAI, Google — direct API consumption"],
    ["Horizontal AI (Copilots)", 8.4, 0.23, "ChatGPT Enterprise, Claude for Work, MS Copilot, Zoom AI"],
    ["Departmental AI", 7.3, 0.20, "Coding ($4.2B), Customer Success ($630M), IT Ops, Marketing"],
    ["Vertical AI", 3.5, 0.09, "Healthcare ($1.5B), Legal ($650M), Creator ($360M), Gov ($350M)"],
    ["TOTAL", 37.0, 1.00, ""],
]

for i, d in enumerate(spend_categories, 1):
    r = row + i
    ws7.cell(row=r, column=1, value=d[0]).border = thin_border
    cell = ws7.cell(row=r, column=2, value=d[1])
    cell.number_format = currency_format
    cell.border = thin_border
    cell = ws7.cell(row=r, column=3, value=d[2])
    cell.number_format = pct_format
    cell.border = thin_border
    ws7.cell(row=r, column=4, value=d[3]).border = thin_border
    if d[0] == "TOTAL":
        for c in range(1, 5):
            ws7.cell(row=r, column=c).font = header_font

row2 = row + len(spend_categories) + 3
ws7.cell(row=row2, column=1, value="Vertical AI Spend by Industry (2025)").font = subtitle_font
row2 += 1
headers2 = ["Industry Vertical", "Spend ($B)", "Share of Vertical AI", "YoY Growth", "Key Driver"]
for c, h in enumerate(headers2, 1):
    ws7.cell(row=row2, column=c, value=h)
style_header_row(ws7, row2, len(headers2))

verticals = [
    ["Healthcare", 1.5, 0.43, "3.3x (from $450M)", "Ambient AI scribes (92% penetration), clinical docs"],
    ["Legal", 0.65, 0.19, "~2.5x", "Contract review, discovery, compliance monitoring"],
    ["Creator Tools", 0.36, 0.10, "~2x", "Content generation, image/video, design workflows"],
    ["Government", 0.35, 0.10, "~2x", "Document processing, citizen services, policy"],
    ["Financial Services (vertical-specific)", 0.30, 0.09, "~2x", "Fraud detection, underwriting, compliance"],
    ["Other Verticals", 0.34, 0.10, "Various", "Education, construction, agriculture, etc."],
    ["TOTAL VERTICAL AI", 3.50, 1.00, "2.9x", ""],
]

for i, d in enumerate(verticals, 1):
    r = row2 + i
    ws7.cell(row=r, column=1, value=d[0]).border = thin_border
    cell = ws7.cell(row=r, column=2, value=d[1])
    cell.number_format = currency_format
    cell.border = thin_border
    cell = ws7.cell(row=r, column=3, value=d[2])
    cell.number_format = pct_format
    cell.border = thin_border
    ws7.cell(row=r, column=4, value=d[3]).border = thin_border
    ws7.cell(row=r, column=5, value=d[4]).border = thin_border
    if "TOTAL" in d[0]:
        for c in range(1, 6):
            ws7.cell(row=r, column=c).font = header_font

auto_width(ws7)

# =============================================================================
# Sheet 8: Pricing & Token Economics
# =============================================================================
ws8 = wb.create_sheet("Token Economics")

ws8.cell(row=1, column=1, value="Token Pricing & Enterprise Economics").font = title_font

row = 3
ws8.cell(row=row, column=1, value="Enterprise Billing Models (2026)").font = subtitle_font
row += 1
headers = ["Dimension", "OpenAI", "Anthropic"]
for c, h in enumerate(headers, 1):
    ws8.cell(row=row, column=c, value=h)
style_header_row(ws8, row, len(headers))

pricing_data = [
    ["Enterprise Seat Fee", "$60/user/month (Enterprise)", "$20/user/month (Claude.ai) or $20/user/month (Claude Code)"],
    ["Billing Model", "Per-seat with included usage tiers", "Per-seat (access only) + per-token usage billing"],
    ["API Pricing (flagship)", "GPT-4o: $2.50/$10 per 1M tokens (in/out)", "Claude Sonnet 4: $3/$15 per 1M tokens (in/out)"],
    ["API Pricing (premium)", "GPT-4.5: $75/$150 per 1M tokens", "Claude Opus 4: $15/$75 per 1M tokens"],
    ["Usage Model Shift", "Bundled allowances for Enterprise seats", "Mandatory monthly spend commitment; all usage metered"],
    ["Revenue per GW Compute", "$12.6B per GW", "$21.4B per GW"],
    ["API Throughput", "15B tokens/minute (Mar 2026)", "Not disclosed"],
    ["Key Pricing Trend", "Microsoft subsidization; aggressive bundling", "Unbundled tokens from seats; pure consumption billing"],
]

for i, d in enumerate(pricing_data, 1):
    r = row + i
    ws8.cell(row=r, column=1, value=d[0]).border = thin_border
    cell = ws8.cell(row=r, column=2, value=d[1])
    cell.border = thin_border
    cell.fill = openai_fill
    cell = ws8.cell(row=r, column=3, value=d[2])
    cell.border = thin_border
    cell.fill = anthropic_fill

auto_width(ws8)

# =============================================================================
# Sheet 9: Spend per Customer Data (Cledara/Ramp)
# =============================================================================
ws9 = wb.create_sheet("Per-Customer Spend")

ws9.cell(row=1, column=1, value="Average Monthly Spend Per Customer (Cledara Data)").font = title_font
ws9.cell(row=2, column=1, value="Source: Cledara SaaS spending data across thousands of companies")

row = 4
headers = ["Month", "OpenAI Avg Spend/Customer", "Anthropic Avg Spend/Customer", "Leader"]
for c, h in enumerate(headers, 1):
    ws9.cell(row=row, column=c, value=h)
style_header_row(ws9, row, len(headers))

spend_data = [
    ["Apr 2025", "$543", "$163", "OpenAI (3.3x)"],
    ["Jul 2025", "$663", "$332", "OpenAI (2.0x)"],
    ["Oct 2025", "$818", "$642", "OpenAI (1.3x)"],
    ["Jan 2026", "$1,050", "$1,049", "Parity (~1:1)"],
    ["Feb 2026", "$1,172", "$1,351", "Anthropic (1.15x)"],
    ["Mar 2026", "$1,014", "$1,548", "Anthropic (1.53x)"],
]

for i, d in enumerate(spend_data, 1):
    r = row + i
    for c, v in enumerate(d, 1):
        cell = ws9.cell(row=r, column=c, value=v)
        cell.border = thin_border
        if c == 2:
            cell.fill = openai_fill
        elif c == 3:
            cell.fill = anthropic_fill

row2 = row + len(spend_data) + 3
ws9.cell(row=row2, column=1, value="Subscription Adoption Rates (Cledara, Mar 2026)").font = subtitle_font
row2 += 1
headers2 = ["Metric", "OpenAI", "Anthropic"]
for c, h in enumerate(headers2, 1):
    ws9.cell(row=row2, column=c, value=h)
style_header_row(ws9, row2, len(headers2))

adoption = [
    ["Companies with active subscription", "56.1%", "51.0%"],
    ["Share of all SaaS transactions", "5.0%", "15.4%"],
    ["Companies using BOTH providers", "36.2%", "36.2%"],
    ["YoY subscription growth", "Flat (6.3% → 5.0%)", "12x (1.2% → 15.4%)"],
    ["Mean transaction size", "$299", "$143"],
    ["Median transaction size", "$45", "$39"],
]

for i, d in enumerate(adoption, 1):
    r = row2 + i
    for c, v in enumerate(d, 1):
        cell = ws9.cell(row=r, column=c, value=v)
        cell.border = thin_border

auto_width(ws9)

# =============================================================================
# Sheet 10: Sources & Methodology
# =============================================================================
ws10 = wb.create_sheet("Sources & Notes")

ws10.cell(row=1, column=1, value="Sources & Methodology").font = title_font

row = 3
sources = [
    ["Source", "Data Used", "Date", "URL"],
    ["Sacra", "OpenAI ARR estimates, revenue breakdown", "Feb 2026", "sacra.com/c/openai/"],
    ["Presenc AI", "OpenAI IPO analysis, revenue line breakdown", "May 2026", "presenc.ai/research/openai-ipo-watch-2026"],
    ["VentureBeat", "Anthropic $30B run-rate announcement", "Apr 2026", "venturebeat.com"],
    ["The Information", "Latest ARR estimates ($33B/$45B)", "May/Jun 2026", "theinformation.com"],
    ["Sherwood News", "Comparative ARR ($45B vs $33B)", "Jun 2026", "sherwood.news"],
    ["Menlo Ventures", "Enterprise gen AI spend $37B, market share", "Dec 2025", "menlovc.com"],
    ["Ramp AI Index", "Industry preference data, spend share", "Feb 2026", "officechai.com (analysis)"],
    ["Cledara", "Per-customer spend comparison", "Mar 2026", "cledara.com/blog"],
    ["OpenAI", "Enterprise AI report, customer announcements", "2025-2026", "openai.com"],
    ["Anthropic", "Compute partnership, pricing changes", "2026", "anthropic.com"],
    ["TechnologyChecker", "Industry distribution of OpenAI adopters", "2026", "technologychecker.io"],
    ["Vercel AI Gateway", "Token volume & spend share by provider", "Apr 2026", "vercel.com/blog"],
    ["HfS Research", "Anthropic IT services penetration", "Apr 2026", "horsesforsources.com"],
    ["Meridian48", "OpenAI 2026 revenue breakdown", "2026", "meridian48.com"],
    ["The Register", "Anthropic enterprise pricing changes", "Apr 2026", "theregister.com"],
]

for i, s in enumerate(sources):
    r = row + i
    for c, v in enumerate(s, 1):
        cell = ws10.cell(row=r, column=c, value=v)
        cell.border = thin_border
        if i == 0:
            cell.font = header_font_white
            cell.fill = header_fill

row3 = row + len(sources) + 2
ws10.cell(row=row3, column=1, value="Methodology Notes").font = subtitle_font
notes = [
    "1. ARR figures are annualized run-rates (monthly revenue × 12 or 4-week revenue × 13), NOT audited GAAP annual revenue.",
    "2. Anthropic reports revenue GROSS (includes cloud partner pass-through); OpenAI reports NET for Azure resale (20% cut only).",
    "3. Industry token spend estimates are derived by cross-referencing: Menlo Ventures total market size, Ramp provider preference data,",
    "   OpenAI's enterprise growth report (industry-level growth rates), and TechnologyChecker adoption distribution.",
    "4. Individual customer spend estimates are illustrative ranges based on seat counts, published pricing, and analyst models.",
    "5. 'Enterprise LLM Market Share' from Menlo Ventures covers model API spend only, not total AI budget.",
    "6. Vercel AI Gateway data covers production developer workloads specifically and may not represent overall enterprise patterns.",
    "7. The crossover where Anthropic surpassed OpenAI in enterprise spend occurred approximately January 2026 (Cledara) / mid-2025 (Menlo).",
]

for i, n in enumerate(notes, 1):
    ws10.cell(row=row3 + i, column=1, value=n)

auto_width(ws10)

# Save
wb.save("/workspace/OpenAI_Anthropic_ARR_Token_Spend_Analysis.xlsx")
print("Excel file saved: /workspace/OpenAI_Anthropic_ARR_Token_Spend_Analysis.xlsx")
