#!/usr/bin/env python3
"""Builds datacenter_buildouts_2026.xlsx — an aggregation of the largest
data center buildouts (single campuses), multi-site programs / contracted
compute pipelines, and hyperscaler capex, with costs and sources.

Data compiled from public reporting as of 27 July 2026.
Run: python3 build_datacenter_spreadsheet.py
"""

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = "datacenter_buildouts_2026.xlsx"
COMPILED = "27 July 2026"

# ---------------------------------------------------------------- styling
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL = PatternFill("solid", fgColor="EAF0F8")
TITLE_FONT = Font(bold=True, size=15, color="1F3864")
SUB_FONT = Font(italic=True, size=10, color="595959")
SECTION_FONT = Font(bold=True, size=12, color="1F3864")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def put_table(ws, top_row, headers, rows, widths, num_cols=()):
    """Write a styled table starting at row `top_row`, return last row."""
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=top_row, column=c, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BORDER
    for r, row in enumerate(rows, top_row + 1):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = WRAP
            cell.border = BORDER
            if (r - top_row) % 2 == 0:
                cell.fill = ALT_FILL
            if c in num_cols and isinstance(v, (int, float)):
                cell.number_format = "#,##0.0" if isinstance(v, float) else "#,##0"
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    return top_row + len(rows)


wb = Workbook()

# ================================================================ overview
ov = wb.active
ov.title = "Overview"
ov.sheet_view.showGridLines = False

ov["B2"] = "Largest Data Center Buildouts — Costs & Pipelines"
ov["B2"].font = TITLE_FONT
ov["B3"] = f"Compiled from public reporting as of {COMPILED}. All dollar figures in US$ billions unless noted."
ov["B3"].font = SUB_FONT

notes = [
    "WHAT'S IN THIS WORKBOOK",
    "  • 'Largest Buildouts' — the biggest single data center campuses worldwide, with status, capacity, cost and power strategy.",
    "  • 'Programs & Pipelines' — multi-site buildout programs, contracted compute deals and financing pipelines behind the campuses.",
    "  • 'Hyperscaler Capex' — 2025 actual vs 2026 guided capital expenditure for Amazon, Microsoft, Alphabet and Meta.",
    "  • 'Sources' — publications used, with URLs and what each supports.",
    "",
    "KEY TAKEAWAYS (as of Jul 2026)",
    "  • Big-4 hyperscaler capex guided to ~$710–725B for 2026, up ~77% from ~$410B in 2025; analysts project $1T+ in 2027.",
    "  • Stargate (OpenAI/Oracle/SoftBank) spans ~11 GW of headline planned US capacity plus Stargate UAE; ~$400B committed vs the $500B envelope.",
    "  • OpenAI's total infrastructure spending plan reached ~$750B through 2030 (WSJ, 22 Jul 2026), incl. its first self-built campus (Project Camellia, GA).",
    "  • Largest single campuses by planned power: Fermi 'Project Matador' claim 11 GW (unproven), Nexus Hubbard potential 7.7 GW,",
    "    Meta Hyperion 5 GW, Stargate UAE 5 GW, Stargate Doña Ana ('Project Jupiter') 4.5 GW.",
    "  • Largest disclosed single-campus costs: Meta Hyperion >$50B, Stargate Abilene ~$47B (incl. GPUs), xAI Colossus 2 ~$35.8B (est., incl. chips).",
    "",
    "CAVEATS — READ BEFORE COMPARING NUMBERS",
    "  • Cost bases are inconsistent across projects: some figures include IT hardware/GPUs (often the majority of cost), others cover only land,",
    "    shell, power and cooling. See the 'Cost basis / notes' column before comparing.",
    "  • Capacity is reported variously as IT (critical) load or total grid draw; grid draw runs ~1.2–1.3x IT load.",
    "  • Figures are company statements, regulatory filings and press estimates — not audited; several projects have already been re-scoped.",
    "  • Announced/pipeline projects (e.g. Project Matador) may never be built as described.",
]
r = 5
for line in notes:
    cell = ov.cell(row=r, column=2, value=line)
    if line and not line.startswith(" "):
        cell.font = SECTION_FONT
    r += 1

# helper table for the chart: single campuses with disclosed cost estimates
chart_rows = [
    ("Meta Hyperion (LA)", 50.0),
    ("Stargate Abilene (TX, incl. GPUs)", 47.0),
    ("xAI Colossus 2 (TN, est. incl. chips)", 35.8),
    ("Stargate UAE (Abu Dhabi)", 30.0),
    ("OpenAI Project Camellia (GA)", 30.0),
    ("Stargate 'Frontier' Shackelford (TX)", 25.0),
    ("Stargate 'Jupiter' Doña Ana (NM, financing)", 21.0),
    ("Meta Prometheus (OH, est.)", 16.0),
    ("Stargate 'Lighthouse' Port Washington (WI)", 15.0),
    ("AWS Project Rainier (IN)", 11.0),
    ("Stargate 'The Barn' Saline Twp (MI)", 10.0),
    ("Meta El Paso (TX)", 10.0),
    ("Homer City Energy Campus (PA)", 10.0),
    ("Microsoft Fairwater Wisconsin", 7.3),
    ("Nexus Hubbard initial phase (TX)", 5.0),
]
anchor = r + 1
ov.cell(row=anchor, column=2, value="Estimated cost — largest single campuses (US$B)").font = SECTION_FONT
hdr_row = anchor + 1
ov.cell(row=hdr_row, column=2, value="Campus")
ov.cell(row=hdr_row, column=3, value="Est. cost (US$B)")
for i, (name, cost) in enumerate(chart_rows, 1):
    ov.cell(row=hdr_row + i, column=2, value=name)
    ov.cell(row=hdr_row + i, column=3, value=cost).number_format = "#,##0.0"

chart = BarChart()
chart.type = "bar"  # horizontal bars
chart.style = 10
chart.title = "Largest single-campus buildouts by estimated cost (US$B)"
chart.y_axis.delete = False
chart.x_axis.delete = False
data = Reference(ov, min_col=3, min_row=hdr_row, max_row=hdr_row + len(chart_rows))
cats = Reference(ov, min_col=2, min_row=hdr_row + 1, max_row=hdr_row + len(chart_rows))
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.legend = None
chart.height = 12
chart.width = 24
ov.add_chart(chart, f"E{anchor}")

ov.column_dimensions["A"].width = 2
ov.column_dimensions["B"].width = 46
ov.column_dimensions["C"].width = 16

# ======================================================= largest buildouts
bo = wb.create_sheet("Largest Buildouts")
headers = [
    "Project / Campus", "Lead company / Developer", "Partners / Anchor tenant",
    "Location", "Country", "Status (Jul 2026)",
    "Capacity now (GW)", "Planned capacity (GW)",
    "Est. cost (US$B)", "Cost basis / notes",
    "Construction start", "Target completion", "Power strategy",
]
# Ordered by estimated cost (desc); not-disclosed costs follow, by capacity.
projects = [
    ("Meta Hyperion", "Meta (JV with Blue Owl Capital)", "Entergy Louisiana (power)",
     "Richland Parish, Louisiana", "USA", "Under construction",
     0, 5.0, 50.0,
     ">$50B announced Jul 2026 (up from $10B at Dec 2024 announcement); $27B initial JV with Blue Owl; ~$29B off-balance-sheet financing ($26B debt + $3B equity). Excludes ~$3B of on-site gas plants funded via Entergy.",
     "2025", "First phase 2028; full build ~2030",
     "10 new Entergy gas plants (>7 GW combined) + up to 1.5 GW solar; Meta pays full energy/water infrastructure cost"),

    ("Stargate Abilene (flagship; 'Project Ludicrous')", "Crusoe (developer) / Lancium (site)", "Oracle (lessee, hardware), OpenAI (tenant), Microsoft (~700 MW lease), Nvidia",
     "Abilene, Texas", "USA", "Partially operational (4 of 8 buildings live)",
     0.3, 1.2, 47.0,
     "~$47B at full buildout INCLUDING ~400k Nvidia Blackwell GPUs; Crusoe construction cost ~$12B excl. chips. 8 buildings, ~4M sq ft. Planned 600 MW expansion cancelled early 2026; capacity re-let to Microsoft/Nvidia.",
     "Jun 2024", "Full 1.2 GW: Q4 2026",
     "On-site natural gas + AEP grid; closed-loop liquid cooling"),

    ("xAI Colossus 2", "xAI (SpaceX)", "Used by xAI/SpaceX, Anthropic, Cursor (per Epoch AI)",
     "Memphis, Tennessee", "USA", "Operational & expanding",
     0.95, 1.5, 35.8,
     "Epoch AI capital-cost estimate INCLUDING chips (~1.1M H100-equivalents; GB200/GB300); build cost excl. chips estimated ~$9B. 946 MW IT power live; further ~400 MW expansion phase in progress.",
     "2025", "Expansion phases through 2026",
     "Gas-fired (mobile generators + power plant) + MLGW grid"),

    ("Stargate UAE", "Khazna Data Centres (G42)", "OpenAI, Oracle, Nvidia, Cisco, SoftBank; Mubadala; South Korea joined",
     "Abu Dhabi", "UAE", "Under construction",
     0, 5.0, 30.0,
     ">$30B per UAE AI minister (Jan 2026), up ~50% from initial ~$20B estimate. 19.2 km² campus (~9x Monaco). 200 MW phase 1 due Q3 2026; 1 GW foundation cluster in ~3-year window.",
     "2025", "200 MW Q3 2026; 1 GW late 2026+; 5 GW full build later",
     "Dedicated generation + grid; GB300 systems on site since May 2026"),

    ("OpenAI Project Camellia", "OpenAI (first self-designed/built/operated site)", "Georgia Power (25-yr PPA)",
     "Effingham County, Georgia", "USA", "Announced / pipeline",
     0, 3.2, 30.0,
     "$20B initial commitment; >$30B at full scale per OpenAI VP of Compute Strategy (announced 22 Jul 2026). 1,400 acres, four buildings. OpenAI pays full infrastructure and electric-service costs.",
     "Announced Jul 2026", "Power phased 2028–2032",
     "Georgia Power grid (3.2 GW, 25-yr agreement) + up to 1 GW flexible demand response"),

    ("Stargate Shackelford 'Frontier'", "Vantage Data Centers", "Oracle (named occupant), OpenAI workloads; VoltaGrid (power)",
     "Shackelford County, Texas", "USA", "Under construction",
     0, 1.4, 25.0,
     ">$25B total investment. 10 buildings, ~3.7M sq ft, 1.4 GW IT capacity (~2.0 GW grid-draw equivalent); >250 kW/rack densities. Part of $38B Vantage borrowing across its two Stargate campuses.",
     "Dec 2025", "First building 2H 2026; full ~Q4 2028",
     "Behind-the-meter gas microgrid (210 VoltaGrid Jenbacher engines, 700 MW), fully off ERCOT grid"),

    ("Stargate Doña Ana 'Project Jupiter'", "STACK Infrastructure + BorderPlex Digital Assets", "Oracle (customer); lenders SMBC/BNP/Goldman/MUFG; Blue Owl equity",
     "Doña Ana County, New Mexico", "USA", "Planning / pre-construction",
     0, 4.5, 21.0,
     "~$18B bank project financing + ~$3B Blue Owl equity arranged; up to $165B industrial revenue bonds authorized by county (30-yr). Largest single Stargate campus by planned capacity. Air-quality permit was the gating item (decision deadline Jul 2026).",
     "Financing arranged; no confirmed start", "~2.2 GW by 2028 (est.); 4.5 GW full build",
     "Two behind-the-fence gas microgrids (no utility grid connection); dedicated Transwestern gas lateral"),

    ("Meta Prometheus", "Meta", "—",
     "New Albany, Ohio", "USA", "Online 2026 (first gigawatt-scale cluster)",
     0, 1.0, 16.0,
     "Meta has not disclosed cost; analyst/Epoch estimates $15.6B–$32.4B including hardware, depending on expansion scope. Low end used here.",
     "2024", "Online 2026",
     "On-site natural gas (~200 MW) + grid"),

    ("Stargate Port Washington 'Lighthouse'", "Vantage Data Centers + Cloverleaf Infrastructure", "OpenAI, Oracle; We Energies (power)",
     "Port Washington, Wisconsin", "USA", "Under construction (early)",
     0, 0.9, 15.0,
     ">$15B investment. 902 MW IT across 4 buildings (~2.5M sq ft); 1.3 GW grid draw; expandable to 3.5 GW on up to 1,900 acres. Requires new $1.4–1.6B ATC transmission corridor (in service late 2027).",
     "First steel Mar 2026", "2028",
     "Utility-integrated: We Energies adding >3 GW new generation (solar/battery/wind + >1.2 GW gas)"),

    ("AWS Project Rainier", "Amazon Web Services", "Anthropic (anchor; trains Claude here)",
     "New Carlisle, Indiana", "USA", "Operational & expanding",
     1.0, 2.25, 11.0,
     "$11B — largest capital investment in Indiana history. 18 of 32 buildings operating (Jul 2026); ~6.5M sq ft at full build. >500k Trainium2 chips at opening (Oct 2025), 1M+ by end 2025. Largest non-Nvidia compute deployment.",
     "2024", "Full 32-building build ~2027–28",
     "Grid (Indiana Michigan Power, 2.25 GW draw) + 878 backup diesel generators"),

    ("Stargate Saline Township 'The Barn'", "Related Digital (Related Companies)", "Oracle subsidiary (contracting customer), OpenAI; DTE Energy (power)",
     "Saline Township, Michigan", "USA", "Under construction (financing unresolved)",
     0, 1.4, 10.0,
     "~$10B estimated project cost; ~$14B debt package described in reporting but lender composition unconfirmed (Blue Owl declined; Blackstone in talks). 1.38 GW contracted utility capacity across 3 buildings (~550k sq ft each).",
     "Site work from Dec 2025", "~2028",
     "Grid (DTE) + project-financed battery storage (1.33 GW of contracts approved Mar 2026)"),

    ("Meta El Paso", "Meta", "—",
     "El Paso, Texas", "USA", "Under construction",
     0, 1.0, 10.0,
     "$10B announced investment; ~1 GW AI data center.",
     "2025", "2028",
     "Grid + on-site generation"),

    ("Homer City Energy Campus", "Homer City Redevelopment (+ Kiewit)", "GE Vernova (turbines)",
     "Homer City, Pennsylvania", "USA", "Site prep / pipeline",
     0, 4.5, 10.0,
     ">$10B initial phases announced Apr 2025 for power infrastructure and site; redevelopment of the largest former coal plant site in PA into a gas-powered data center campus (up to 4.5 GW generation).",
     "2025 (demolition/site prep)", "Power from ~2027",
     "On-site combined-cycle natural gas (GE Vernova 7HA.02 turbines)"),

    ("Microsoft Fairwater Wisconsin", "Microsoft (self-build)", "WE Energies (power)",
     "Mount Pleasant, Wisconsin", "USA", "Partially operational (DC1 live Jun 2026)",
     None, None, 7.3,
     ">$7.3B committed ($3.3B first DC + $4B second, Sep 2025); ~$4.7B local hyperscale construction 2024–28. Capacity not disclosed ('world's most powerful AI datacenter' claim); 15 additional buildings approved Jan 2026.",
     "2023", "DC1 live Jun 2026; DC2 2028",
     "Grid (WE Energies) with pre-paid infrastructure; 250 MW solar match; closed-loop liquid cooling"),

    ("Nexus Hubbard Campus", "Nexus Data Centers", "Anthropic (anchor lease); Google (construction financing); Eagle Point (credit)",
     "Hubbard, Texas", "USA", "Under construction",
     0, 7.7, 5.0,
     "$5B+ initial phase (up-to figure incl. Google-backed financing); first permitted building 491k sq ft / ~600 MW. ~2,900-acre site could ultimately host up to 7.7 GW of behind-the-meter generation.",
     "Broke ground 2025–26", "500–612 MW late 2026; 7.7 GW long-term potential",
     "Behind-the-meter natural gas turbines (direct pipeline access), independent of ERCOT queue"),

    ("Stargate Milam County 'Orion'", "SB Energy (SoftBank)", "OpenAI ($500M each into SB Energy); Ares ($800M pref)",
     "Milam County, Texas", "USA", "Under construction",
     0, 1.2, None,
     "Site cost not disclosed; $1.8B identified equity into SB Energy. Vertically integrated developer model; modular units manufactured at SoftBank's Lordstown plant targeting 7–8 month build cycles.",
     "Oct 2025", "~Q4 2028",
     "Co-located with SB Energy's 900 MWdc Orion Solar Belt + on-site infrastructure"),

    ("Stargate Lordstown", "SoftBank (+ Foxconn JV)", "OpenAI",
     "Lordstown, Ohio", "USA", "Under construction",
     0, 0.3, None,
     "Former Foxconn EV plant acquired for $375M (Aug 2025); dual role as data center (~0.3 GW) and modular data-center-equipment factory. With Milam County, scalable to a combined 1.5 GW by 2027.",
     "Q4 2025", "Operational 2026–27",
     "Grid"),

    ("xAI Colossus 1", "xAI", "—",
     "Memphis, Tennessee", "USA", "Operational",
     0.43, 0.43, None,
     "~425 MW; ~230k GPUs (H100/H200/GB200 era); built in ~122 days in 2024. Cost not formally disclosed.",
     "2024", "Operational since Sep 2024",
     "Gas turbines + MLGW grid"),

    ("Microsoft Fairwater Atlanta (Fairwater 2)", "Microsoft (leased)", "QTS / Blackstone (landlord); Georgia Power",
     "Fayetteville (Atlanta), Georgia", "USA", "Operational (Oct–Nov 2025)",
     0.7, 1.5, None,
     "Cost not disclosed (Microsoft is anchor tenant on long triple-net lease of QTS-built shells; 2 buildings, ~2.17M sq ft with Fairwater 1). >700 MW Microsoft IT capacity by end 2026; QTS campus ultimate load ~1.5 GW across 13 buildings.",
     "2024", ">700 MW by end 2026",
     "Georgia Power grid only — no on-site generation or UPS; linked with other Fairwater sites into one training system"),

    ("Anthropic–Fluidstack sites (TX & NY)", "Fluidstack (developer/operator)", "Anthropic ($50B commitment)",
     "Texas & New York (multiple sites)", "USA", "Under construction",
     0, None, 50.0,
     "$50B program announced 12 Nov 2025 — Anthropic's first major independent buildout; capacity not disclosed (multi-GW). ~800 permanent + 2,400 construction jobs. Listed here as a program; sites not yet individually disclosed.",
     "2025–26", "Facilities online through 2026",
     "Mixed / site-dependent"),

    ("Google Texas new campuses", "Google", "—",
     "Armstrong, Haskell (x2), Pampa & Wilbarger Counties, Texas", "USA", "Announced / early works",
     0, None, 40.0,
     "Within Google's $40B Texas investment through 2027 (also covers Ellis County/Red Oak expansion, energy funds, workforce). Campus-level capacity/cost not yet disclosed; one Haskell site co-located with solar + battery storage.",
     "Announced Nov 2025", "Through 2027",
     "Grid + ~6.2 GW of Texas PPAs; one campus co-located with solar/BESS"),

    ("Fermi America 'Project Matador' (HyperGrid)", "Fermi America", "—",
     "Amarillo, Texas", "USA", "Announced only — no construction or customers",
     0, 11.0, None,
     "Claims an 11 GW energy-and-data campus (nuclear + gas + solar). As of late 2025 reporting, construction had not begun and no customers were secured. Included for completeness; treat as speculative.",
     "—", "—",
     "Proposed nuclear + gas + solar 'HyperGrid'"),

    ("HUMAIN buildout", "HUMAIN (Saudi PIF)", "Nvidia, AMD silicon supply; NEOM",
     "Riyadh / NEOM / Eastern Province", "Saudi Arabia", "Early construction",
     0, 0.5, None,
     "PIF-backed sovereign AI company; ~500 MW near-term build target across three regions with Nvidia and AMD supply pipelines (multi-GW ambition into the 2030s). Costs not disclosed.",
     "2025", "First phases 2026+",
     "Grid + renewables"),
]
last = put_table(bo, 1, headers, projects,
                 widths=[38, 30, 34, 26, 12, 26, 11, 12, 11, 60, 16, 24, 42],
                 num_cols=(7, 8, 9))
bo.freeze_panes = "B2"
bo.auto_filter.ref = f"A1:M{last}"
note = bo.cell(row=last + 2, column=1,
               value="Blank cost/capacity = not disclosed. Costs are NOT directly comparable — some include IT hardware/GPUs, others only land/shell/power (see 'Cost basis / notes'). "
                     f"Capacity mixes IT load and grid draw as reported. Compiled {COMPILED}.")
note.font = SUB_FONT

# ==================================================== programs & pipelines
pp = wb.create_sheet("Programs & Pipelines")
p_headers = ["Program / Deal", "Type", "Parties", "Value (US$B)", "Capacity (GW)",
             "Timeframe", "Status (Jul 2026)", "Notes"]
programs = [
    ("OpenAI infrastructure spending plan", "Corporate pipeline", "OpenAI",
     750.0, "~8 GW fleet target by 2030",
     "Through 2030", "Active",
     "Raised to ~$750B (WSJ, 22 Jul 2026) from ~$600B earlier in 2026; spans Stargate, self-builds (Camellia) and cloud contracts below. OpenAI ended 2025 with ~1.9 GW of compute."),
    ("Meta US infrastructure pledge", "Corporate pipeline", "Meta",
     600.0, ">10 GW total capacity targeted by end 2026",
     "Through 2028", "Active",
     "US data centers and infrastructure pledge; anchored by Hyperion (LA), Prometheus (OH), El Paso (TX) and ~30 existing data centers."),
    ("Stargate program", "Multi-site buildout JV", "OpenAI, Oracle, SoftBank, MGX",
     500.0, "~11 GW headline planned across 7 US sites + UAE",
     "2025–2029", "Active — only ~0.2–0.6 GW operational so far",
     "$500B/4-yr envelope announced at White House Jan 2025 ($100B immediate); ~$400B of committed 3-yr spend by mid-2026. Buildout is proceeding site-by-site with different developers, power models and financing."),
    ("OpenAI–Oracle compute agreement", "Compute contract", "OpenAI, Oracle",
     300.0, "4.5 GW additional (~6 GW total contracted)",
     "~5 years (from Jul 2025)", "Signed",
     ">$300B over ~5 years for additional Stargate capacity; underpins Shackelford, Doña Ana, Port Washington, Saline Twp sites. Oracle RPO hit $553B in Q3 FY26 (+325% YoY)."),
    ("OpenAI–Microsoft Azure commitment", "Compute contract", "OpenAI, Microsoft",
     250.0, "n/d",
     "~6 years (from Oct 2025 restructuring)", "Signed",
     "Up to $250B of contracted Azure spend agreed as part of the Oct 2025 partnership restructuring."),
    ("OpenAI–AWS agreement", "Compute contract", "OpenAI, Amazon Web Services",
     138.0, "n/d",
     "8 years", "Signed (expanded)",
     "Expanded from initial ~$38B (Nov 2025) to $138B over eight years (reported Jul 2026)."),
    ("Anthropic–AWS commitment", "Compute contract + investment", "Anthropic, Amazon",
     100.0, "Up to 5 GW Trainium reserved; ~1 GW online by end 2026",
     "10 years", "Active",
     ">$100B compute commitment; Amazon investing up to $25B more ($5B equity + $20B milestone-gated facility) on top of $8B already invested. Project Rainier (IN) is the flagship site."),
    ("Nvidia–OpenAI letter of intent", "Strategic investment", "Nvidia, OpenAI",
     100.0, "10 GW of Nvidia systems",
     "Announced Sep 2025", "LOI — progressive",
     "Up to $100B invested progressively as each GW of Nvidia systems is deployed."),
    ("Anthropic–Fluidstack program", "Multi-site buildout", "Anthropic, Fluidstack",
     50.0, "n/d (multi-GW)",
     "Announced 12 Nov 2025; online through 2026", "Under construction",
     "Anthropic's first major independent data center buildout — custom sites in Texas and New York."),
    ("Google Texas investment", "Regional program", "Google",
     40.0, "n/d; 6.2 GW of PPAs",
     "Through 2027", "Active",
     "$40B for cloud/AI infrastructure incl. new campuses in Armstrong, Haskell (x2), Pampa, Wilbarger counties + Ellis County expansion; $30M Energy Impact Fund."),
    ("Vantage Stargate financing", "Project debt", "Vantage Data Centers (DigitalBridge)",
     38.0, "Shackelford 1.4 GW IT + Port Washington 0.9 GW IT",
     "2025–2026", "Drawn/committed",
     "$38B borrowed across the two Vantage-built Stargate campuses."),
    ("Stargate UAE / G42 program", "International buildout", "Khazna (G42), OpenAI, Oracle, Nvidia, Cisco, SoftBank",
     30.0, "5 GW campus (1 GW foundation cluster)",
     "Phase 1 Q3 2026; 1 GW late 2026+", "Under construction",
     ">$30B estimate (Jan 2026); template being studied for sovereign builds with Vietnam and India."),
    ("Microsoft UK investment", "Regional program", "Microsoft (+ Nscale)",
     30.0, "n/d",
     "Announced Sep 2025", "Active",
     "$30B UK commitment incl. the country's largest supercomputer with Nscale."),
    ("Amazon–Anthropic investment", "Investment", "Amazon, Anthropic",
     33.0, "—",
     "2023–", "Active",
     "Up to ~$33B total: $8B invested + $5B equity + $20B milestone-gated financing facility tied to compute delivery."),
    ("OpenAI–CoreWeave agreement", "Compute contract", "OpenAI, CoreWeave",
     22.4, "n/d",
     "Through 2029", "Signed",
     "Up to $22.4B of contracted capacity."),
    ("Anthropic–Google Cloud TPU deal", "Compute contract", "Anthropic, Google",
     None, "Up to 1 GW of TPU capacity in 2026",
     "Announced Oct 2025", "Signed",
     "Reported to be worth tens of billions of dollars; value not formally disclosed."),
    ("Meta–AMD chip supply", "Chip supply", "Meta, AMD",
     None, "6 GW",
     "Signed Feb 2026", "Signed",
     "6 GW AMD accelerator deal (per Measured AI reporting); shifted competitive dynamics for Nvidia at Stargate Abilene."),
    ("SoftBank–DigitalBridge acquisition", "M&A", "SoftBank, DigitalBridge",
     3.9, "—",
     "Expected close H2 2026", "Pending",
     "Gives SoftBank control of Vantage Data Centers — consolidating Stargate exposure across SB Energy (Milam), Vantage (Shackelford, Port Washington) and Lordstown."),
    ("Oracle contracted backlog (RPO)", "Backlog indicator", "Oracle",
     553.0, "—",
     "Q3 FY2026 (Mar 2026)", "Reported",
     "Remaining performance obligations $553B, +325% YoY — proxy for contracted AI cloud pipeline. Oracle FY26 capex ~$50B; total borrowings $134.6B (Feb 2026); $50B capital program announced Feb 2026."),
]
last = put_table(pp, 1, p_headers, programs,
                 widths=[36, 24, 40, 13, 34, 26, 26, 78],
                 num_cols=(4,))
pp.freeze_panes = "A2"
pp.auto_filter.ref = f"A1:H{last}"
pp.cell(row=last + 2, column=1,
        value="Values mix committed spend, 'up to' contract ceilings, debt facilities and backlog — see Notes column. Contract values overlap with campus costs on the 'Largest Buildouts' sheet; do not sum across sheets.").font = SUB_FONT

# ========================================================= hyperscaler capex
cx = wb.create_sheet("Hyperscaler Capex")
c_headers = ["Company", "2025 capex (US$B, approx.)", "2026 guidance (US$B)",
             "2026 midpoint (US$B)", "YoY growth", "Primary AI drivers", "Notes"]
capex = [
    ("Amazon", 100, "~200", 200, "~+100%", "AWS, Trainium 2/3, Anthropic",
     "Q1 2026 capex $44.2B (+77% YoY); ~$200B full-year run rate reiterated by CEO Andy Jassy."),
    ("Microsoft", 95, "~190", 190, "~+100%", "Azure, OpenAI, Maia silicon",
     "Guide includes ~$25B attributed to higher component pricing (Q3 FY26 call, 29 Apr 2026). AI business run rate >$37B (+123% YoY)."),
    ("Alphabet (Google)", 85, "180–190", 185, "~+110%", "Google Cloud, TPU v7, Gemini",
     "Raised from $175–185B in Feb 2026. Google Cloud backlog >$460B."),
    ("Meta", 71, "125–145", 135, "~+90%", "Llama, ads ranking, MTIA silicon",
     "Raised both ends by $10B (from $115–135B), citing memory prices and added data center costs. No public cloud to monetize."),
    ("Combined (Big 4)", 410, "~700–725", 710, "~+77%", "—",
     "vs ~$226B in 2024. Roughly $2B/day of infrastructure spend; ~3/4 AI-related. Analysts project $1T+ combined in 2027."),
]
last = put_table(cx, 1, c_headers, capex,
                 widths=[22, 22, 20, 18, 12, 34, 70],
                 num_cols=(2, 4))
cx.freeze_panes = "A2"
cx.cell(row=last + 2, column=1,
        value="Capex covers all capital expenditure (mostly, but not exclusively, AI/data centers): GPUs and custom silicon, data center shells, power and land. Source: Q1 2026 earnings calls (29 Apr 2026) and company disclosures.").font = SUB_FONT

# =================================================================== sources
src = wb.create_sheet("Sources")
s_headers = ["Source", "URL", "Supports", "Accessed"]
sources = [
    ("OpenAI — Five new Stargate sites", "https://openai.com/index/five-new-stargate-sites/",
     "Stargate expansion: Shackelford, Doña Ana, Lordstown, Milam; $300B+/4.5 GW Oracle deal"),
    ("Epoch AI — Stargate: where the US sites stand", "https://epoch.ai/publications/openai-stargate-where-the-us-sites-stand",
     "Site-level capacity/status/timeline table for all 7 US Stargate sites"),
    ("Measured AI — Stargate site-by-site status (Mar 2026)", "https://measuredai.substack.com/p/stargate-site-by-site-status",
     "Abilene $47B/$12B costs; Frontier >$25B; Lighthouse >$15B; The Barn ~$10B; Jupiter $18B+$3B financing; Vantage $38B debt; OpenAI Azure $250B / AWS $138B / CoreWeave $22.4B; Oracle RPO/debt"),
    ("Presenc AI — Stargate status report 2026", "https://presenc.ai/research/stargate-project-status-2026",
     "~7 GW planned, ~$400B committed; Abilene 4-of-8 buildings live; Stargate UAE phase status"),
    ("The Energy Mag — Meta expands Louisiana AI DC (Jul 2026)", "https://backend.theenergymag.com/news/2026-07-13/meta-ai-louisiana-expand/",
     "Hyperion 5 GW / >$50B expansion"),
    ("Silicon Report — Meta Hyperion profile", "https://www.siliconreport.com/meta-hyperion-louisiana-megacampus-profile-84de9440",
     "Hyperion power plan: 10 Entergy gas plants >7 GW; $29B off-balance-sheet financing"),
    ("TechCrunch — OpenAI spending balloons to $750B (22 Jul 2026)", "https://techcrunch.com/2026/07/22/openais-ai-spending-spree-has-ballooned-to-750b/",
     "OpenAI $750B-through-2030 plan; Project Camellia basics"),
    ("Enterprise DNA / ConstructConnect — Project Camellia", "https://enterprisedna.co/resources/news/openai-project-camellia-georgia-data-center-30-billion-2026/",
     "Camellia 3.2 GW, $20B→$30B+, Georgia Power 25-yr PPA, 2028–2032 phasing"),
    ("Epoch AI — Colossus 2 directory entry", "https://epoch.ai/data/ai-data-centers/directory/colossus-2",
     "Colossus 2: 946 MW IT, $35.8B est. capital cost incl. chips, users incl. Anthropic/Cursor"),
    ("Measured AI — AWS New Carlisle / Project Rainier", "https://measuredai.substack.com/p/aws-new-carlisle-data-center-campus",
     "$11B, 32 buildings, 2.25 GW draw, Anthropic–AWS $100B/5 GW commitments"),
    ("Introl — Anthropic $50B Fluidstack partnership", "https://introl.com/blog/anthropic-50b-fluidstack-data-center-december-2025",
     "Anthropic–Fluidstack $50B, TX & NY, jobs figures"),
    ("Microsoft — Made in Wisconsin (Fairwater)", "https://blogs.microsoft.com/on-the-issues/2025/09/18/made-in-wisconsin-the-worlds-most-powerful-ai-datacenter/",
     "Fairwater WI $3.3B + $4B; DCD follow-up: DC1 operational Jun 2026, 15 more buildings approved"),
    ("Measured AI — Microsoft Fairwater Atlanta", "https://measuredai.substack.com/p/microsoft-fairwater-atlanta-data-center",
     "Fairwater Atlanta: QTS lease, >700 MW by end 2026, ~1.5 GW campus load, grid-only design"),
    ("The National / AGBI — Stargate UAE >$30B (Jan 2026)", "https://www.thenationalnews.com/future/technology/2026/01/26/stargate-uae-data-centre-to-cost-more-than-30bn-ai-minister-says/",
     "Stargate UAE 5 GW, >$30B, 19.2 km², Q3 2026 phase 1"),
    ("DCD — Google commits $40B to Texas", "https://www.datacenterdynamics.com/en/news/google-commits-40bn-to-expand-cloud-and-ai-infrastructure-in-texas-through-2027/",
     "Google $40B Texas program, new campuses, 6.2 GW PPAs; Fermi 'Project Matador' 11 GW claim"),
    ("MLQ / Construction Review — Nexus Hubbard campus", "https://mlq.ai/news/nexus-data-centers-files-permit-for-491000-sq-ft-data-center-in-hubbard-texas/",
     "Nexus Hubbard: $5B+ initial, Anthropic lease, Google financing, 7.7 GW potential"),
    ("Value Add VC / Analysis Atlas — 2026 hyperscaler capex", "https://valueaddvc.com/pulse/big-tech-ai-capex-725-billion-2026-guidance",
     "Big-4 2026 capex guidance ($710–725B), per-company breakdown, 2025 baseline"),
    ("AI in Arabia — G42 Stargate UAE GB300s; HUMAIN", "https://aiinarabia.com/news/g42-stargate-uae-200mw-gb300-chips-on-site-news-2026-05-18",
     "Stargate UAE phase-1 progress; HUMAIN 500 MW target across Riyadh/NEOM/Eastern Province"),
]
rows = [(t, u, s, COMPILED) for (t, u, s) in sources]
last = put_table(src, 1, s_headers, rows, widths=[46, 62, 80, 14])
src.freeze_panes = "A2"
src.cell(row=last + 2, column=1,
         value="Homer City Energy Campus (PA) figures are from its April 2025 announcement (Homer City Redevelopment / GE Vernova press coverage). "
               "All other rows trace to the sources above. Figures are press/company statements, not audited.").font = SUB_FONT

wb.save(OUT)
print(f"Wrote {OUT}")
