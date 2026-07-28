#!/usr/bin/env python3
"""Generate llm_architectural_innovations.xlsx.

Aggregates the role of five architectural/training innovations
(MLA, Muon, KDA, Attention Residuals, Latent MoE) across open-weight
and proprietary LLMs, compiled from public papers, model cards and
technical blogs as of July 2026.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- styling

TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F3864")
H2_FONT = Font(name="Calibri", size=12, bold=True, color="1F3864")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10.5)
BOLD_BODY = Font(name="Calibri", size=10.5, bold=True)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
YES_FILL = PatternFill("solid", fgColor="C6EFCE")        # green
VARIANT_FILL = PatternFill("solid", fgColor="E2F0D9")    # light green
PARTIAL_FILL = PatternFill("solid", fgColor="FFF2CC")    # yellow
NO_FILL = PatternFill("solid", fgColor="F8CBAD")         # orange/red
UNKNOWN_FILL = PatternFill("solid", fgColor="D9D9D9")    # gray
OPEN_FILL = PatternFill("solid", fgColor="DDEBF7")
PROP_FILL = PatternFill("solid", fgColor="FCE4D6")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def status_fill(value: str):
    v = value.lower()
    if v.startswith("yes"):
        if "variant" in v or "hybrid" in v or "gated" in v or "originator" in v:
            return VARIANT_FILL
        return YES_FILL
    if v.startswith(("hybrid", "variant")):
        return VARIANT_FILL
    if v.startswith(("partial", "rumored", "related", "validated")):
        return PARTIAL_FILL
    if v.startswith("no") and not v.startswith("not stated"):
        return NO_FILL
    return UNKNOWN_FILL  # Undisclosed / Unknown


def write_header_row(ws, row, headers):
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=text)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = WRAP_CENTER
        c.border = BORDER


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# ================================================================ 1. Overview

ws = wb.active
ws.title = "1. Overview"
set_widths(ws, [4, 30, 110])
ws.sheet_view.showGridLines = False

ws.cell(row=2, column=2, value="Architectural Innovations in LLM Training").font = TITLE_FONT
ws.cell(row=3, column=2, value="MLA · Muon · KDA · Attention Residuals · Latent MoE — adoption across open-weight and proprietary models").font = Font(
    size=11, italic=True, color="595959")
ws.cell(row=4, column=2, value="Compiled July 28, 2026 from public papers, model cards, official tech blogs and framework documentation.").font = Font(
    size=10, color="595959")

overview_rows = [
    ("SCOPE", ""),
    ("What this workbook covers",
     "Five architectural / training innovations that shaped frontier LLM training in 2024–2026: "
     "Multi-head Latent Attention (MLA), the Muon optimizer, Kimi Delta Attention (KDA), Attention Residuals (AttnRes), "
     "and Latent MoE. For each: origin, mechanism, role in training, quantified impact, and which models (open and proprietary) use it."),
    ("Sheet guide",
     "Sheet 2 'Innovations Summary' — one row per innovation with origin, mechanism and impact.  "
     "Sheet 3 'Model Adoption Matrix' — one row per model, one column per innovation, color-coded.  "
     "Sheet 4 'Innovation Deep-Dive' — aspect-by-aspect comparison (problem, mechanism, training role, variants, ecosystem).  "
     "Sheet 5 'Sources' — numbered references with URLs."),
    ("KEY TAKEAWAYS", ""),
    ("1. Efficiency drives architecture",
     "All five innovations attack the same constraints: KV-cache memory (MLA, KDA), optimizer efficiency / step quality (Muon), "
     "signal propagation across depth (AttnRes), and MoE memory-bandwidth + communication cost (Latent MoE). Together they let labs "
     "convert a fixed compute budget into more capability — Kimi K3 claims ~2.5x overall scaling efficiency over K2 from this stack."),
    ("2. Open labs publish, proprietary labs stay silent",
     "Every confirmed adoption in this workbook comes from open-weight labs (DeepSeek, Moonshot, Zhipu, NVIDIA, Prime Intellect). "
     "OpenAI, Anthropic, Google and xAI do not disclose architecture details; their rows are marked 'Undisclosed' with rumor-level "
     "evidence flagged where it exists (e.g. leaked claims that GPT-5.5 uses a new MoE router; analyses suggesting Claude remains dense or lightly-MoE)."),
    ("3. Innovations compound",
     "The techniques are complementary and increasingly co-deployed: Kimi K3 (2.8T params) combines all five at once "
     "(KDA + Gated MLA attention, AttnRes residuals, Stable LatentMoE experts, Per-Head Muon optimizer). DeepSeek-V4 pairs Muon with "
     "its own alternatives (CSA/HCA attention instead of MLA, mHC instead of plain residuals)."),
    ("4. Optimizer shift is real",
     "Muon is the first serious challenger to Adam/AdamW in production: ~2x compute efficiency at compute-optimal scale, "
     "adopted for pretraining by Kimi K2/K3 (1T/2.8T), GLM-4.5/5 (355B/744B), DeepSeek-V4 (1.6T), and integrated into "
     "PyTorch 2.9, DeepSpeed, and NVIDIA Megatron Core."),
    ("LEGEND (Sheet 3 color coding)", ""),
    ("Green — Yes", "Confirmed use, stated in a paper, model card, or official blog."),
    ("Light green — Yes (variant / hybrid / originator)", "Confirmed use of a modified form (e.g. Gated MLA, MuonClip, Muon Split, Stable LatentMoE) or a hybrid layout."),
    ("Yellow — Partial / Rumored / Related", "Used for part of training only (e.g. SFT/RL), rumor-level evidence, or a closely related sibling technique."),
    ("Orange — No", "Explicitly uses a different mechanism (documented)."),
    ("Gray — Undisclosed", "No public information; typical for proprietary frontier models."),
    ("CAVEATS", ""),
    ("Proprietary data quality",
     "Architecture claims about GPT-5.x, Gemini, Claude and Grok rest on leaks, inference-cost analyses and journalism — not primary sources. "
     "Treat every non-green cell in those rows as unverified."),
    ("Snapshot in time",
     "The field moves fast; adoption data reflects publications available as of late July 2026 (Kimi K3 weights released July 27, 2026)."),
]

r = 6
for label, text in overview_rows:
    if text == "":
        c = ws.cell(row=r, column=2, value=label)
        c.font = H2_FONT
        c.fill = SECTION_FILL
        ws.cell(row=r, column=3).fill = SECTION_FILL
        r += 1
        continue
    lc = ws.cell(row=r, column=2, value=label)
    lc.font = BOLD_BODY
    lc.alignment = WRAP_TOP
    tc = ws.cell(row=r, column=3, value=text)
    tc.font = BODY_FONT
    tc.alignment = WRAP_TOP
    est_lines = max(1, len(text) // 105 + 1)
    ws.row_dimensions[r].height = max(16, est_lines * 14)
    r += 1

# ====================================================== 2. Innovations Summary

ws = wb.create_sheet("2. Innovations Summary")
headers = [
    "Innovation", "Abbrev.", "Category", "Introduced by", "Key paper / source", "Date",
    "What it replaces / changes", "Core mechanism", "Role in TRAINING",
    "Role in INFERENCE / serving", "Quantified impact", "Known variants",
    "Open-source implementations", "Adoption — open-weight models", "Adoption — proprietary models",
]
innovations = [
    [
        "Multi-head Latent Attention", "MLA", "Attention / KV-cache compression",
        "DeepSeek", "DeepSeek-V2 (arXiv:2405.04434); DeepSeek-V3 Technical Report", "May 2024",
        "Replaces MHA / GQA as the attention mechanism.",
        "Low-rank joint compression of keys and values into a small shared latent vector (e.g. 512-dim vs 128 heads x 128-dim); "
        "only the latent is cached and it is up-projected (or absorbed into query/output projections) at compute time. Decoupled RoPE path carries position info.",
        "Makes long-context pretraining and RL economically feasible at MoE scale; DeepSeek reports better modeling quality than MHA, "
        "not just parity — so it is a quality AND efficiency choice. Interacts with Muon (per-head structure motivated GLM-5's 'Muon Split' and Kimi's Per-Head Muon).",
        "KV cache shrinks ~93.3% vs MHA (V2 figure); at DeepSeek-V3 dims a 128K context needs ~1/64 the cache of MHA — the difference between "
        "impractical (~hundreds of GB) and servable. FlashMLA kernels reach up to 660 TFLOPS / 3000 GB/s on H800.",
        "~93.3% KV-cache reduction (V2); ~2x+ inference throughput vs MHA baseline; quality >= MHA in DeepSeek ablations.",
        "Gated MLA (Kimi K3); DSA = DeepSeek Sparse Attention built on top of MLA (V3.2, GLM-5); MHA2MLA / TransMLA (post-hoc conversion of GQA models).",
        "FlashMLA (DeepSeek CUDA kernels, open); vLLM / SGLang support; HF Transformers.",
        "Very high: DeepSeek V2/V3/R1/V3.2, Kimi K2/K2.5, Kimi Linear & K3 (Gated MLA in hybrid), GLM-5 (via DSA), Moonlight. "
        "Note: DeepSeek-V4 moved OFF MLA to CSA/HCA.",
        "Undisclosed. No proprietary lab has confirmed MLA use.",
    ],
    [
        "Muon optimizer", "Muon", "Optimizer (matrix-orthogonalized momentum)",
        "Keller Jordan (NanoGPT speedrun); scaled by Moonshot AI",
        "Jordan et al. 2024 (blog); 'Muon is Scalable for LLM Training' (arXiv:2502.16982)", "Dec 2024 / Feb 2025",
        "Replaces AdamW for 2-D weight matrices (embeddings, norms, scalars usually stay on AdamW).",
        "MomentUm Orthogonalized by Newton-Schulz: takes the momentum-averaged gradient of each weight matrix and orthogonalizes it "
        "(approximate matrix sign via 5-step Newton-Schulz iteration) before the update, so every step is well-conditioned across all directions. "
        "Moonshot's scaling recipe adds weight decay + AdamW-matched per-parameter update RMS.",
        "This is the purest 'training' innovation of the five: ~2x compute efficiency vs AdamW at compute-optimal scale (same loss for ~52% of FLOPs), "
        "~half the optimizer state memory (one momentum buffer vs Adam's two), stable bf16 behavior, and works out-of-the-box without re-tuning. "
        "MuonClip's QK-Clip eliminated attention-logit explosions: Kimi K2 pretrained 15.5T tokens at 1T params with zero loss spikes.",
        "None directly (optimizer only affects training), but faster convergence lowers cost per capability.",
        "~2x FLOP efficiency vs AdamW (Moonlight scaling laws); 1T-param production run with zero instabilities (K2); "
        "adopted for 744B (GLM-5) and 1.6T (DeepSeek-V4-Pro) pretraining.",
        "MuonClip / QK-Clip (Kimi K2); Muon Split (GLM-5 — orthogonalizes MLA up-projections per attention head); "
        "Per-Head Muon (Kimi K3); NorMuon; distributed Muon (Moonshot, open-sourced).",
        "PyTorch 2.9 torch.optim.Muon; DeepSpeed (ZeRO 2/3); NVIDIA Megatron Core & NeMo; Moonshot's distributed impl; nanochat.",
        "High and accelerating: Moonlight, Kimi K2 / K2 Thinking / K3, GLM-4.5, GLM-5, DeepSeek-V4 (Pro & Flash), "
        "INTELLECT-3 (SFT/RL), nanoGPT speedrun, nanochat.",
        "Undisclosed for GPT/Gemini/Claude/Grok. NVIDIA/Megatron integration signals broad industry interest.",
    ],
    [
        "Kimi Delta Attention", "KDA", "Linear attention (hybrid layouts)",
        "Moonshot AI (Kimi Team)", "'Kimi Linear: An Expressive, Efficient Attention Architecture' (arXiv:2510.26692)", "Oct 2025",
        "Replaces most full-attention layers; deployed as hybrid (typically 3 KDA layers : 1 full/MLA layer).",
        "Extends Gated DeltaNet with fine-grained channel-wise (diagonal) forget gates: each feature dimension gets its own data-dependent "
        "decay instead of one scalar per head, so the fixed-size RNN state can hold some channels almost indefinitely while decaying others. "
        "Custom chunkwise algorithm uses a specialized Diagonal-Plus-Low-Rank (DPLR) transition for hardware efficiency.",
        "First linear-attention design reported to beat full attention (MLA) under identical training recipes on short-context, long-context AND "
        "RL workloads — removing the historical quality penalty of linear attention. Linear-time sequence mixing makes 1M-token context training "
        "and long-horizon RL rollouts affordable.",
        "Fixed-size state instead of KV cache for 3/4 of layers: up to 75% KV-cache reduction and up to 6x decoding throughput at 1M context "
        "(Kimi K3 reports 6.3x faster decoding).",
        "Kimi Linear 48B/3B: outperforms full MLA across all evaluated tasks at identical recipe; -75% KV cache; 6x TPOT speedup @1M.",
        "Deployed as KDA + MLA hybrid (3:1); K3 pairs KDA with Gated MLA (69 KDA + 24 Gated MLA layers).",
        "KDA kernels in FLA (flash-linear-attention); vLLM integration; Kimi Linear checkpoints (48B) open-sourced.",
        "Kimi Linear (48B, open weights); Kimi K3 (2.8T, open weights) — the production backbone of Moonshot's lineup.",
        "None disclosed.",
    ],
    [
        "Attention Residuals", "AttnRes", "Residual stream / depth-wise connectivity",
        "Moonshot AI (Kimi Team)", "'Attention Residuals' (arXiv:2603.15031); open impl: wdlctc/open-attention-residuals", "Mar 2026",
        "Replaces standard PreNorm additive residual connections.",
        "Each layer applies softmax attention over the outputs of preceding layers (via a learned per-layer pseudo-query), selecting "
        "which earlier representations to aggregate instead of summing everything with fixed unit weights. Block AttnRes groups layers into ~8 blocks, "
        "cutting memory/communication from O(L*d) to O(N*d), making it a practical drop-in at scale.",
        "Fixes PreNorm dilution during training of very deep stacks: hidden-state magnitudes stay bounded and gradient norms distribute uniformly "
        "across depth, improving optimization of 90+ layer models. Scaling-law experiments show consistent gains across model sizes; "
        "Block AttnRes matches a baseline given ~1.25x more compute. Overhead: ~0.03% params, <2% latency.",
        "Minimal serving impact (small extra state per block); primarily a trainability / quality innovation.",
        "Matches baseline trained with ~1.25x more compute; uniform gradient distribution; validated on Kimi-Linear-48B pretrained on 1.4T tokens.",
        "Full AttnRes (attend over all layer outputs) vs Block AttnRes (block-level, N~8). "
        "Related-but-different: DeepSeek-V4's Manifold-Constrained Hyper-Connections (mHC) attack the same residual-stream problem.",
        "wdlctc/open-attention-residuals (open reference implementation).",
        "Kimi K3 (2.8T) — named as one of two backbone innovations; validated on Kimi Linear 48B architecture.",
        "None disclosed.",
    ],
    [
        "Latent MoE", "LatentMoE / l-MoE", "Mixture-of-Experts / sparse FFN",
        "NVIDIA (Nemotron team, Elango et al.)", "'LatentMoE: Toward Optimal Accuracy per FLOP and Parameter in MoE' (arXiv:2601.18089)", "Jan 2026",
        "Restructures the routed-expert path of standard MoE layers.",
        "Projects each token from hidden dim d down to a latent dim l (shared down-projection) BEFORE expert dispatch; routed experts compute "
        "entirely in the latent space; a shared up-projection maps back to d. Router and shared experts stay in dim d. Dispatch bytes and "
        "expert weight bytes both shrink by ~d/l; the savings are reinvested into more experts and higher top-k at the same serving cost.",
        "Enables much higher expert counts / sparsity for the same training and serving budget (combinatorially richer routing). "
        "Kimi K3's 'Stable LatentMoE' (latent dim 3584 vs hidden 7168) makes a 896-expert / 16-active layout trainable, paired with "
        "Quantile Balancing for load balance without a tuned bias hyperparameter.",
        "Cuts MoE all-to-all communication volume and per-token expert weight-loading (memory bandwidth) — the true serving bottlenecks of MoE; "
        "projected up to 3.5x throughput speedup at iso-accuracy vs standard MoE at trillion-param scale.",
        "Standard MoE needs ~350B extra params to match LatentMoE accuracy in NVIDIA's analysis; added projection compute ~9% (vs native Kimi-K2-1T); "
        "K3 activates 16 of 896 experts (~1.8% expert sparsity).",
        "l-MoE_acc (scale top-k, iso-cost, higher accuracy) vs l-MoE_eff (keep top-k, cheaper); Stable LatentMoE (Kimi K3); "
        "Multi-Head LatentMoE + Head Parallel (arXiv:2602.04870, deterministic communication).",
        "kyegomez/Latent-MoE (PyTorch reference); NVIDIA Nemotron stack.",
        "NVIDIA Nemotron-3 Super & Ultra (flagship adoption); Kimi K3 (Stable LatentMoE, 2.8T).",
        "Undisclosed elsewhere; NVIDIA positions it as the template for trillion-scale MoE serving.",
    ],
]

write_header_row(ws, 1, headers)
widths = [22, 12, 20, 22, 30, 12, 26, 60, 60, 45, 42, 42, 32, 45, 32]
set_widths(ws, widths)
for i, row in enumerate(innovations, start=2):
    for j, val in enumerate(row, start=1):
        c = ws.cell(row=i, column=j, value=val)
        c.font = BODY_FONT if j > 1 else BOLD_BODY
        c.alignment = WRAP_TOP
        c.border = BORDER
    ws.row_dimensions[i].height = 170
ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(innovations) + 1}"

# ==================================================== 3. Model Adoption Matrix

ws = wb.create_sheet("3. Model Adoption Matrix")
headers = [
    "Model", "Developer", "Access", "Release", "Total params", "Active params", "Context",
    "MLA", "Muon", "KDA", "AttnRes", "Latent MoE",
    "Other notable architecture / training features", "Evidence / notes",
]

# (model, dev, access, release, total, active, ctx, mla, muon, kda, attnres, lmoe, other, evidence)
models = [
    # --- open weights -------------------------------------------------------
    ("DeepSeek-V2", "DeepSeek", "Open weights", "May 2024", "236B", "21B", "128K",
     "Yes (originator)", "No (AdamW)", "No", "No", "No",
     "DeepSeekMoE fine-grained + shared experts; introduced MLA.",
     "arXiv:2405.04434 — MLA introduced here; 93.3% KV-cache reduction headline result."),
    ("DeepSeek-V3 / R1", "DeepSeek", "Open weights", "Dec 2024 / Jan 2025", "671B", "37B", "128K",
     "Yes", "No (AdamW)", "No", "No", "No",
     "FP8 mixed-precision training; Multi-Token Prediction (MTP); aux-loss-free load balancing; R1 adds RL reasoning on V3-Base.",
     "DeepSeek-V3 Technical Report; R1 inherits MLA unchanged."),
    ("DeepSeek-V3.2-Exp", "DeepSeek", "Open weights", "Sep 2025", "671B", "37B", "128K",
     "Yes (+ DSA sparse attention on top)", "No (AdamW)", "No", "No", "No",
     "DeepSeek Sparse Attention (DSA): token-level top-k sparse attention built on the MLA cache; FlashMLA sparse kernels (640 TFLOPS prefill).",
     "DeepSeek-V3.2-Exp release + FlashMLA repo."),
    ("DeepSeek-V4-Pro", "DeepSeek", "Open weights (MIT)", "Jun–Jul 2026", "1.6T", "49B", "1M",
     "No — replaced by CSA + HCA hybrid", "Yes (AdamW kept for embeddings/norms/gates)", "No", "Related — mHC (Manifold-Constrained Hyper-Connections) replaces plain residuals", "No (DeepSeekMoE retained)",
     "Hybrid attention: Compressed Sparse Attention (FP4 indexer, top-k blocks) + Heavily Compressed Attention (cheap global view) + 128-token sliding window; "
     "Hash-MoE bootstrap in first ~3 layers; Sqrt(Softplus) affinity; FP4 QAT expert weights; MTP; 61 layers.",
     "arXiv:2606.19348 — vs V3.2: 27% of 1M-token inference FLOPs, 10% of KV cache. Notably moved OFF MLA."),
    ("DeepSeek-V4-Flash", "DeepSeek", "Open weights (MIT)", "Jun–Jul 2026", "284B", "13B", "1M",
     "No — SWA / CSA / HCA per-layer mix", "Yes (AdamW for embeddings)", "No", "Related — mHC (4 hyper-connection streams, Sinkhorn-routed)", "No (256 routed + 1 shared expert, top-6)",
     "43-layer all-MoE backbone; per-layer attention 'zoo' via compress_ratios; hash-gate first layers; FP4 experts / FP8 elsewhere.",
     "HF model card + NVIDIA NeMo AutoModel docs."),
    ("Moonlight", "Moonshot AI", "Open weights", "Feb 2025", "16B", "3B", "8K",
     "Yes (DeepSeek-V2-Lite-style arch)", "Yes (originator at scale — the Muon scaling-law paper)", "No", "No", "No",
     "3B/16B MoE trained on 5.7T tokens as the demonstration vehicle for scalable Muon; intermediate checkpoints released.",
     "arXiv:2502.16982 — ~2x FLOP efficiency vs AdamW; matched AdamW baseline with ~52% of FLOPs."),
    ("Kimi K2 / K2 Thinking", "Moonshot AI", "Open weights", "Jul / Nov 2025", "1.04T", "32B", "128K–256K",
     "Yes", "Yes — variant MuonClip (QK-Clip)", "No", "No", "No",
     "Ultra-sparse MoE: 384 experts, 8 active + 1 shared; 64 attention heads; DeepSeek-V3-like layout chosen by scaling-law analysis; 15.5T pretraining tokens.",
     "arXiv:2507.20534 — zero loss spikes across the entire 1T-param run; the flagship proof of Muon in production."),
    ("Kimi Linear", "Moonshot AI", "Open weights", "Oct 2025", "48B", "3B", "1M",
     "Yes — hybrid (1 MLA layer per 3 KDA)", "Not stated in paper", "Yes (originator)", "Validated — AttnRes paper integrates it into this architecture (1.4T-token run)", "No",
     "Proof-of-concept for the KDA:MLA 3:1 hybrid; 5.7T-token checkpoints released; KDA kernel open-sourced in FLA + vLLM.",
     "arXiv:2510.26692 — beats full MLA on all tasks at identical recipe; -75% KV cache; 6x decode throughput @1M."),
    ("Kimi K3", "Moonshot AI", "Open weights (Jul 27, 2026)", "Jul 2026", "2.78T", "104B", "1M",
     "Yes — variant Gated MLA (24 of 93 layers)", "Yes — variant Per-Head Muon", "Yes (69 of 93 attention layers)", "Yes (backbone)", "Yes — Stable LatentMoE (latent dim 3584 vs hidden 7168)",
     "ALL FIVE innovations in one model. 896 experts / 16 active + 2 shared; Quantile Balancing router; SiTU-GLU activation; MXFP4 weights + MXFP8 activations (QAT); "
     "MoonViT-V2 401M vision encoder; first open 3T-class model; ~2.5x scaling efficiency vs K2.",
     "Kimi K3 tech blog + OpenLM.ai spec table; 6.3x faster decoding @1M vs full attention."),
    ("GLM-4.5", "Zhipu AI (Z.ai)", "Open weights (MIT)", "2025", "355B", "32B", "128K",
     "No (GQA with partial RoPE)", "Yes", "No", "No", "No",
     "MoE; also released GLM-4.5-Air (106B).",
     "arXiv:2508.06471 (GLM-4.5 Technical Report) confirms Muon pretraining."),
    ("GLM-5", "Zhipu AI (Z.ai)", "Open weights", "2026", "744B", "40B", "long-context",
     "Yes — via DSA (MLA-based sparse attention)", "Yes — variant Muon Split (per-head orthogonalization of MLA up-projections)", "No", "No", "No",
     "Adopted DeepSeek Sparse Attention; Muon Split closes the MLA-vs-GQA performance gap under Muon.",
     "PyTorch/DeepSpeed blog + GLM-5 disclosures."),
    ("INTELLECT-3", "Prime Intellect", "Open weights", "Late 2025", "106B", "12B", "128K",
     "No (GQA — GLM-4.5-Air base)", "Partial — Muon for SFT + RL only (base was pretrained with Muon by Zhipu)", "No", "No", "No",
     "Decentralized post-training on GLM-4.5-Air-Base.",
     "arXiv:2512.16144."),
    ("Nemotron-3 Super / Ultra", "NVIDIA", "Open weights (rolling out)", "2026", "n/a (flagship scale)", "n/a", "long-context",
     "No (Mamba-Attention hybrid)", "Undisclosed (Muon integrated in Megatron Core; training optimizer not confirmed)", "No", "No", "Yes (originator — flagship adoption)",
     "Mamba-Attention hybrid MoE; LatentMoE scaled to longer horizons and larger sizes per Nemotron-3 white paper.",
     "NVIDIA Nemotron LatentMoE research page."),
    ("Qwen3-235B-A22B", "Alibaba", "Open weights", "2025", "235B", "22B", "128K",
     "No (GQA)", "Undisclosed (optimizer not stated; NVIDIA ran Muon experiments on Qwen3-30B)", "No", "No", "No",
     "Counter-example: frontier open MoE that adopted none of the five (public info).",
     "Qwen3 model cards; NVIDIA Megatron blog for the Muon experiments."),
    ("MiniMax-M1 / M2.5", "MiniMax", "Open weights", "2025–2026", "456B (M1)", "45.9B (M1)", "1M (M1)",
     "No — Lightning Attention (different linear-attention lineage)", "Undisclosed", "No (sibling linear-attention approach)", "No", "No",
     "Hybrid MoE + Lightning Attention; parallel evolution to KDA in the linear-attention space.",
     "MiniMax-M1 release."),
    ("nanochat / NanoGPT speedrun", "Karpathy / Jordan (community)", "Open source", "2024–2025", "<1B", "<1B", "small",
     "No", "Yes (Muon's birthplace)", "No", "No", "No",
     "Muon was created for the NanoGPT speedrun and popularized via nanochat before frontier adoption.",
     "Keller Jordan blog; nanochat repo."),
    # --- proprietary --------------------------------------------------------
    ("GPT-5.5 / GPT-5.6", "OpenAI", "Proprietary", "2026", "Undisclosed", "Undisclosed", "long-context",
     "Undisclosed", "Undisclosed", "Undisclosed", "Undisclosed", "Rumored — leaks claim a new MoE router with ~4x active params of GPT-5.4 at equal inference cost",
     "GPT-5.4 release notes referenced 'sparse activation experiments'; no primary-source architecture disclosure.",
     "May 2026 leak coverage (unverified, rumor-tier)."),
    ("Gemini 2.5 / 3.5 Pro", "Google DeepMind", "Proprietary", "2025–2026", "Undisclosed", "Undisclosed", "1M+",
     "Undisclosed", "Undisclosed", "Undisclosed", "Undisclosed", "Rumored — MoE confirmed for 2.5 (Google architecture overview); soft-routing MoE reported for Ultra; latent variant unknown",
     "3.5 Pro reportedly required a full base-model rebuild (Jul 2026 launch); details unpublished.",
     "Google architecture overview (MoE only); press reporting (unverified)."),
    ("Claude (Opus 4.7 / Fable 5)", "Anthropic", "Proprietary", "2025–2026", "Undisclosed", "Undisclosed", "long-context",
     "Undisclosed", "Undisclosed", "Undisclosed", "Undisclosed", "Rumored — public evidence suggests dense or lightly-MoE (the notable frontier hold-out); unverified 5T/10T MoE claims circulate",
     "Anthropic discloses nothing; inference-cost analyses point to dense-leaning designs prioritizing predictability/interpretability.",
     "Third-party analyses only (unverified, rumor-tier)."),
    ("Grok 4 / 5", "xAI", "Proprietary", "2025–2026", "Undisclosed", "Undisclosed", "long-context",
     "Undisclosed", "Undisclosed", "Undisclosed", "Undisclosed", "Undisclosed",
     "No architecture disclosures.",
     "—"),
]

write_header_row(ws, 1, headers)
set_widths(ws, [24, 16, 18, 14, 12, 12, 12, 26, 30, 22, 30, 34, 60, 45])
for i, row in enumerate(models, start=2):
    for j, val in enumerate(row, start=1):
        c = ws.cell(row=i, column=j, value=val)
        c.border = BORDER
        c.alignment = WRAP_TOP
        if j == 1:
            c.font = BOLD_BODY
        else:
            c.font = BODY_FONT
        if j == 3:  # access column
            c.fill = OPEN_FILL if "open" in val.lower() else PROP_FILL
        if 8 <= j <= 12:  # innovation columns
            c.fill = status_fill(val)
    ws.row_dimensions[i].height = 78
ws.freeze_panes = "B2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(models) + 1}"

# ===================================================== 4. Innovation Deep-Dive

ws = wb.create_sheet("4. Innovation Deep-Dive")
aspects = [
    ("Problem addressed", [
        "KV cache grows linearly with context and heads: at V3 dims a 128K-token MHA cache would be ~hundreds of GB — unservable. "
        "GQA trades quality for cache size.",
        "AdamW treats every parameter as an independent scalar, ignoring the matrix structure of weights; steps are poorly conditioned, "
        "convergence is slower than the geometry allows, and it stores two state buffers.",
        "Full attention is O(T^2) and its KV cache is O(T); at 1M-token contexts both training and decoding become prohibitive. "
        "Prior linear attention lost too much quality to be used.",
        "PreNorm residual streams sum every layer's output with fixed unit weights: hidden-state norms grow unboundedly with depth "
        "and each new layer's contribution is progressively diluted — an optimization problem for 90+ layer models.",
        "MoE serving is bottlenecked by memory bandwidth (loading expert weights per token) and all-to-all dispatch traffic — not FLOPs. "
        "Standard MoE pays both in the full hidden dimension d.",
    ]),
    ("Core mechanism", [
        "Jointly compress K and V into one low-rank latent vector per token (plus a small decoupled RoPE key); cache only the latent; "
        "up-project at attention time (projections can be absorbed for decode).",
        "Momentum SGD on each weight matrix, but the update is orthogonalized via ~5 Newton-Schulz iterations (approximate matrix sign / "
        "nearest semi-orthogonal matrix). Scaling recipe: weight decay + match update RMS to Adam's.",
        "Delta-rule linear attention with channel-wise (diagonal) data-dependent forget gates — each feature dimension has its own decay — "
        "implemented with a hardware-efficient chunkwise DPLR-variant algorithm; fixed-size recurrent state.",
        "Replace 'x + f(x)' accumulation with softmax attention over preceding layer (or block) outputs, weighted by a learned per-layer "
        "pseudo-query; zero-initialized so training starts from the standard-residual behavior.",
        "Shared down-projection d->l before expert dispatch; routed experts and all-to-all run entirely in the latent dim l; shared up-projection "
        "l->d after combine. Router and shared experts stay in d. Savings reinvested in expert count and top-k.",
    ]),
    ("Role in TRAINING specifically", [
        "Cuts activation memory of attention and makes long-context pretraining/RL affordable; DeepSeek reports quality above MHA, so no "
        "capability tax. Its per-head low-rank structure later forced optimizer co-design (Muon Split, Per-Head Muon).",
        "~2x compute efficiency vs AdamW (same loss, ~52% FLOPs) at compute-optimal scale; halves optimizer memory; MuonClip's QK-Clip caps "
        "attention logits — Kimi K2's 15.5T-token, 1T-param run finished with zero loss spikes. Now the default for new frontier open-model pretrains.",
        "Linear-time sequence mixing slashes the cost of long-context pretraining and long-horizon RL (rollouts dominated by decode). "
        "Beats full attention under identical recipes — first linear attention adopted as a production backbone (K3).",
        "Purely a trainability innovation: bounded hidden-state magnitudes, uniform gradient distribution across depth, consistent scaling-law "
        "gains; Block AttnRes ~= baseline given 1.25x more compute, at 0.03% extra params and <2% latency.",
        "Lets a fixed budget train far more experts at higher sparsity (K3: 896 experts, 16 active) — richer routing combinatorics per FLOP. "
        "Kimi's 'Stable LatentMoE' + Quantile Balancing keep extreme sparsity trainable without hand-tuned balancing.",
    ]),
    ("Role in INFERENCE / serving", [
        "~93% smaller KV cache -> long contexts and high concurrency on the same hardware; FlashMLA kernels hit 660 TFLOPS / 3000 GB/s (H800).",
        "None at serving time (training-side only); indirect effect is cheaper capability.",
        "Up to 75% KV-cache reduction and ~6x decode throughput at 1M tokens (K3: 6.3x); most layers keep a fixed-size state regardless of context.",
        "Negligible serving change; small per-block state retained.",
        "Reduces expert weight-loading bytes and all-to-all volume by ~d/l; projected up to 3.5x throughput at iso-accuracy vs standard MoE at "
        "trillion scale (else standard MoE needs ~+350B params to match).",
    ]),
    ("Key variants & follow-ups", [
        "Gated MLA (K3); DSA — sparse attention on the MLA substrate (DeepSeek-V3.2, GLM-5); TransMLA / MHA2MLA post-hoc conversions; "
        "note DeepSeek-V4 superseded MLA in-house with CSA/HCA.",
        "MuonClip (QK-Clip, K2); Muon Split (GLM-5, per-head MLA up-projections); Per-Head Muon (K3); NorMuon (+21.7% vs Adam at 1.1B); "
        "distributed Muon (open-sourced by Moonshot).",
        "KDA is itself the third step of a lineage: DeltaNet -> Gated DeltaNet (scalar gate) -> KDA (channel-wise gate). "
        "Sibling approaches: Lightning Attention (MiniMax), Mamba hybrids (Hunyuan, Nemotron).",
        "Full AttnRes vs Block AttnRes (N~8 blocks); two-phase computation + cache-based pipeline communication for large-scale training. "
        "Parallel line: DeepSeek-V4's mHC (doubly-stochastic mixing of 4 residual streams).",
        "l-MoE_acc (iso-cost, more experts + higher top-k) vs l-MoE_eff (iso-accuracy, cheaper); Stable LatentMoE (K3); "
        "Multi-Head LatentMoE + Head Parallel (deterministic all-to-all).",
    ]),
    ("Framework / ecosystem support", [
        "FlashMLA (open CUDA kernels); vLLM; SGLang; HF Transformers.",
        "PyTorch 2.9 (torch.optim.Muon); DeepSpeed ZeRO 2/3; NVIDIA Megatron Core + NeMo; Moonshot distributed implementation.",
        "FLA (flash-linear-attention) kernels; vLLM; open Kimi Linear checkpoints.",
        "Open reference implementation (wdlctc/open-attention-residuals); described drop-in for Megatron-style pipelines.",
        "Reference PyTorch implementation (kyegomez/Latent-MoE); NVIDIA Nemotron training stack.",
    ]),
    ("Confirmed adopters (open weights)", [
        "DeepSeek V2 / V3 / R1 / V3.2; Kimi K2 / K2.5; Moonlight; Kimi Linear + K3 (Gated MLA in hybrid); GLM-5 (via DSA).",
        "Moonlight; Kimi K2 / K2 Thinking / K3; GLM-4.5 / GLM-5; DeepSeek-V4 Pro & Flash; INTELLECT-3 (SFT/RL); nanoGPT speedrun; nanochat.",
        "Kimi Linear (48B); Kimi K3 (2.8T).",
        "Kimi K3 (2.8T); validated on Kimi-Linear-48B (1.4T tokens).",
        "NVIDIA Nemotron-3 Super & Ultra; Kimi K3 (Stable LatentMoE).",
    ]),
    ("Proprietary adoption (public evidence)", [
        "None confirmed. Frontier proprietary labs do not disclose attention mechanisms.",
        "None confirmed. NVIDIA's Megatron/NeMo integration and PyTorch upstreaming signal broad industry demand beyond open labs.",
        "None confirmed.",
        "None confirmed.",
        "None confirmed beyond NVIDIA's own (open) Nemotron line; MoE itself is the rumored norm at OpenAI/Google (Claude the likely dense hold-out).",
    ]),
    ("Caveats", [
        "Post-hoc conversion of GQA models (TransMLA) recovers efficiency but not the training-time quality benefit. "
        "DeepSeek itself moved on (V4 uses CSA/HCA) — MLA is dominant but no longer the frontier's endpoint.",
        "Needs AdamW alongside for embeddings/norms/scalars; naive Muon destabilizes at scale without weight decay + RMS matching (+ QK-Clip for "
        "attention logits); per-head structure of MLA requires split/per-head variants.",
        "Fixed-size state is lossy compression: exact long-range retrieval/copying degrades without interleaved full-attention layers — "
        "hence hybrid 3:1 layouts, never pure KDA.",
        "Newest and least battle-tested of the five (one production model); Full AttnRes memory is O(L*d) — blocking is required at scale.",
        "Adds ~9% projection compute; benefits contingent on serving being bandwidth/communication-bound; very new (Jan 2026) with two "
        "adopter families so far.",
    ]),
]

col_heads = ["Aspect", "MLA (Multi-head Latent Attention)", "Muon (optimizer)", "KDA (Kimi Delta Attention)",
             "Attention Residuals (AttnRes)", "Latent MoE"]
write_header_row(ws, 1, col_heads)
set_widths(ws, [22, 52, 52, 52, 52, 52])
for i, (aspect, cells) in enumerate(aspects, start=2):
    a = ws.cell(row=i, column=1, value=aspect)
    a.font = BOLD_BODY
    a.fill = SECTION_FILL
    a.alignment = WRAP_TOP
    a.border = BORDER
    for j, text in enumerate(cells, start=2):
        c = ws.cell(row=i, column=j, value=text)
        c.font = BODY_FONT
        c.alignment = WRAP_TOP
        c.border = BORDER
    ws.row_dimensions[i].height = 118
ws.freeze_panes = "B2"

# ================================================================ 5. Sources

ws = wb.create_sheet("5. Sources")
headers = ["#", "Source", "Type", "URL", "Supports"]
sources = [
    ("DeepSeek-V2: A Strong, Economical, and Efficient MoE Language Model", "Paper (arXiv:2405.04434)",
     "https://arxiv.org/abs/2405.04434", "MLA origin; 93.3% KV-cache reduction."),
    ("DeepSeek-V3 Technical Report", "Paper",
     "https://arxiv.org/abs/2412.19437", "MLA at 671B scale; MTP; FP8 training."),
    ("FlashMLA (DeepSeek kernels)", "GitHub",
     "https://github.com/deepseek-ai/FlashMLA", "MLA/DSA kernels; 660 TFLOPS / 3000 GB/s; powers V3 and V3.2."),
    ("Muon: An optimizer for hidden layers in neural networks (Keller Jordan)", "Blog",
     "https://kellerjordan.github.io/posts/muon/", "Muon origin; NanoGPT speedrun."),
    ("Muon is Scalable for LLM Training (Moonshot)", "Paper (arXiv:2502.16982)",
     "https://arxiv.org/abs/2502.16982", "Muon scaling recipe; ~2x efficiency vs AdamW; Moonlight 16B."),
    ("Kimi K2: Open Agentic Intelligence", "Paper (arXiv:2507.20534)",
     "https://arxiv.org/abs/2507.20534", "MuonClip / QK-Clip; MLA at 1T params; zero loss spikes over 15.5T tokens."),
    ("MoonshotAI/Kimi-K2 model card", "GitHub",
     "https://github.com/MoonshotAI/Kimi-K2", "K2 architecture table (MLA, 384 experts, Muon)."),
    ("Kimi Linear: An Expressive, Efficient Attention Architecture", "Paper (arXiv:2510.26692)",
     "https://arxiv.org/abs/2510.26692", "KDA design; 3:1 KDA:MLA hybrid; -75% KV cache; 6x decode @1M."),
    ("MoonshotAI/Kimi-Linear", "GitHub",
     "https://github.com/MoonshotAI/Kimi-Linear", "Open KDA kernels (FLA), vLLM support, 48B checkpoints."),
    ("Attention Residuals", "Paper (arXiv:2603.15031)",
     "https://arxiv.org/abs/2603.15031", "AttnRes / Block AttnRes; PreNorm dilution fix; Kimi Linear 48B validation."),
    ("open-attention-residuals (reference implementation)", "GitHub",
     "https://github.com/wdlctc/open-attention-residuals", "AttnRes implementation details; 0.03% params, <2% latency overhead."),
    ("LatentMoE: Toward Optimal Accuracy per FLOP and Parameter in MoE (NVIDIA)", "Paper (arXiv:2601.18089)",
     "https://arxiv.org/abs/2601.18089", "Latent MoE mechanism; iso-FLOP/iso-param analysis."),
    ("NVIDIA Nemotron LatentMoE research page", "Blog",
     "https://research.nvidia.com/labs/nemotron/LatentMoE/", "Nemotron-3 Super/Ultra adoption; 3.5x projected speedup; +350B-param equivalence."),
    ("Multi-Head LatentMoE and Head Parallel", "Paper (arXiv:2602.04870)",
     "https://arxiv.org/abs/2602.04870", "Latent MoE follow-up; deterministic MoE parallelism."),
    ("Kimi K3 Tech Blog: Open Frontier Intelligence (Moonshot)", "Official blog",
     "https://www.kimi.com/en/blog/kimi-k3", "K3: KDA + AttnRes backbone; Stable LatentMoE (16/896); Per-Head Muon; Quantile Balancing; Gated MLA; SiTU."),
    ("Kimi K3 spec table (OpenLM.ai)", "Spec aggregation",
     "https://openlm.ai/kimi-k3/", "K3: 2.8T/104B; 69 KDA + 24 Gated MLA layers; latent MoE dim 3584; MXFP4/MXFP8."),
    ("DeepSeek-V4: Towards Highly Efficient Million-Token Context Intelligence", "Paper (arXiv:2606.19348)",
     "https://arxiv.org/abs/2606.19348", "V4 Pro/Flash: Muon adoption; CSA+HCA replacing MLA; mHC; Hash-MoE."),
    ("DeepSeek-V4 model docs (Hugging Face Transformers)", "Docs",
     "https://huggingface.co/docs/transformers/main/model_doc/deepseek_v4", "V4 architecture details (hash_moe, Sqrt-Softplus affinity, mHC)."),
    ("Using Muon Optimizer with DeepSpeed (PyTorch blog)", "Blog",
     "https://pytorch.org/blog/using-muon-optimizer-with-deepspeed/", "Muon adoption roundup: K2, GLM-4.5/5 (Muon Split), DeepSeek-V4, nanochat; DeepSpeed support."),
    ("Advancing Emerging Optimizers with NVIDIA Megatron", "NVIDIA blog",
     "https://developer.nvidia.com/blog/advancing-emerging-optimizers-for-accelerated-llm-training-with-nvidia-megatron/",
     "Muon in Megatron Core / NeMo; near-AdamW throughput at scale."),
    ("GLM-4.5 Technical Report", "Paper (arXiv:2508.06471)",
     "https://arxiv.org/abs/2508.06471", "GLM-4.5 Muon pretraining."),
    ("INTELLECT-3 (Prime Intellect)", "Paper (arXiv:2512.16144)",
     "https://arxiv.org/abs/2512.16144", "Muon for SFT/RL on GLM-4.5-Air base."),
    ("MoonshotAI/Moonlight", "GitHub",
     "https://github.com/MoonshotAI/Moonlight", "Moonlight benchmark table; Muon vs AdamW comparison."),
    ("MoE Architecture: GPT, Claude, DeepSeek, Qwen Compared (Digital Applied)", "Analysis (secondary)",
     "https://www.digitalapplied.com/blog/moe-architecture-comparison-gpt-claude-deepseek-qwen",
     "Proprietary landscape: MoE convergence; Claude as dense hold-out. RUMOR-TIER for closed labs."),
    ("GPT-5.5 leak coverage (TrendWatch)", "Rumor (secondary)",
     "https://trends.thicket.sh/gpt-5-5-leak-may-2026", "Rumored GPT-5.5 MoE router. UNVERIFIED."),
    ("Claude Opus 5 parameter claims — evidence review", "Rumor analysis (secondary)",
     "https://aithinkerlab.com/claude-opus-5-trillion-parameters/", "Unverified Claude MoE claims; evidence-tiering. UNVERIFIED."),
    ("Behind Kimi K3: Understanding Kimi Delta Attention (Zhou Yao)", "Analysis (secondary)",
     "https://zhouyaoai.substack.com/p/behind-kimi-k3-understanding-kimi", "KDA lineage (DeltaNet -> GDN -> KDA); K3 3:1 layout."),
]
write_header_row(ws, 1, headers)
set_widths(ws, [5, 60, 24, 62, 70])
for i, (name, typ, url, supports) in enumerate(sources, start=2):
    ws.cell(row=i, column=1, value=i - 1).font = BODY_FONT
    ws.cell(row=i, column=2, value=name).font = BODY_FONT
    ws.cell(row=i, column=3, value=typ).font = BODY_FONT
    u = ws.cell(row=i, column=4, value=url)
    u.font = Font(size=10.5, color="0563C1", underline="single")
    u.hyperlink = url
    ws.cell(row=i, column=5, value=supports).font = BODY_FONT
    for col in range(1, 6):
        ws.cell(row=i, column=col).alignment = WRAP_TOP
        ws.cell(row=i, column=col).border = BORDER
    ws.row_dimensions[i].height = 30
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:E{len(sources) + 1}"

wb.save("/workspace/llm_architectural_innovations.xlsx")
print("Saved llm_architectural_innovations.xlsx")
for sheet in wb.sheetnames:
    print(" -", sheet)
