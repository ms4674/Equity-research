"""
Build a granular Excel workbook aggregating Meta Platforms' total compute
(data center) capacity, split between company-OWNED data centers and LEASED /
colocation capacity.

Primary "compute capacity" proxy = electric power. In the data-center industry
capacity is measured in power (MW / GW), because power is the binding constraint
on how much compute (servers / GPUs) a site can host. Meta does not publish a
single "owned vs leased MW" number, but it DOES publish, in its annual
Environmental Data Index, actual electricity consumption (MWh/yr) for every
owned online data center individually AND a single line for "Leased data center
facilities." That gives the cleanest apples-to-apples owned-vs-leased split of
realized compute, which anchors this model. Nameplate power (MW), square
footage, buildings, investment, and the forward AI GW build-out are layered on
from Meta disclosures and third-party trackers.

Run:  python build_workbook.py
Out:  Meta_Compute_Capacity_Owned_vs_Leased.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

# ----------------------------------------------------------------------------
# Style helpers
# ----------------------------------------------------------------------------
NAVY = "1F3864"
BLUE = "2E5496"
LIGHT = "D9E1F2"
LIGHTER = "EAF0FA"
OWNED_FILL = "C6E0B4"   # green
LEASED_FILL = "FFE699"  # amber
GREY = "F2F2F2"
ACCENT = "8EAADB"

thin = Side(style="thin", color="BFBFBF")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

TITLE_FONT = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
SUB_FONT = Font(name="Calibri", size=11, italic=True, color="FFFFFF")
H_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BOLD = Font(name="Calibri", size=11, bold=True)
BOLD_NAVY = Font(name="Calibri", size=11, bold=True, color=NAVY)
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="595959")
REG = Font(name="Calibri", size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFTTOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

FMT_INT = "#,##0"
FMT_1 = "#,##0.0"
FMT_PCT = "0.0%"
FMT_USD = '"$"#,##0.0,,"M"'  # not used heavily


def hdr(cell, fill=BLUE, font=H_FONT, align=CENTER):
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = font
    cell.alignment = align
    cell.border = border_all


def box(cell, align=LEFT):
    cell.border = border_all
    cell.alignment = align
    if cell.font is None or cell.font.name is None:
        cell.font = REG


def title_block(ws, title, subtitle, span=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1, 1, title)
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    for col in range(1, span + 1):
        ws.cell(1, col).fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[1].height = 32
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    s = ws.cell(2, 1, subtitle)
    s.font = SUB_FONT
    s.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, span + 1):
        ws.cell(2, col).fill = PatternFill("solid", fgColor=BLUE)
    ws.row_dimensions[2].height = 20


wb = Workbook()

# ============================================================================
# DATA
# ============================================================================
YEARS = [2020, 2021, 2022, 2023, 2024]

# Meta Environmental Data Index (2025 report, FY2024) -- electricity consumption
# (MWh) by facility. Owned online data centers reported individually; leased
# facilities as one line; "other" = warehouses/network/small colo <100k MWh.
# rows: name, region, country, [2020..2024 MWh]
OWNED_ENERGY = [
    ("Altoona",         "US-Midwest",   "USA",     980000, 950705, 1043606, 1243306, 1585392),
    ("Prineville",      "US-West",      "USA",     686000, 898409, 982177, 1375321, 1728291),
    ("Sarpy County",    "US-Midwest",   "USA",     519000, 736810, 1007635, 1148091, 1258239),
    ("Stanton Springs", "US-Southeast", "USA",          0, 215279, 636266, 968565, 1184380),
    ("Los Lunas",       "US-Southwest", "USA",     571000, 717932, 929488, 1110100, 1143067),
    ("Eagle Mountain",  "US-West",      "USA",          0, 229946, 504049, 787740, 1115619),
    ("Fort Worth",      "US-Southwest", "USA",     941000, 1014447, 959419, 1029570, 1109004),
    ("Clonee",          "Europe",       "Ireland", 487000, 634648, 668290, 953837, 1076961),
    ("Henrico",         "US-Southeast", "USA",     204000, 515270, 701003, 805061, 948859),
    ("Huntsville",      "US-Southeast", "USA",          0, 85286, 368841, 614198, 865803),
    ("Odense",          "Europe",       "Denmark", 343000, 500863, 517718, 518005, 569374),
    ("Forest City",     "US-Southeast", "USA",     595000, 580842, 492786, 507068, 535555),
    ("New Albany",      "US-Midwest",   "USA",     270000, 511414, 702694, 793063, 521217),
    ("Luleå",           "Europe",       "Sweden",  369000, 306054, 267471, 351931, 468809),
    ("DeKalb",          "US-Midwest",   "USA",          0, 4724, 16934, 138965, 372339),
    ("Gallatin",        "US-Southeast", "USA",          0, 0, 6264, 116520, 359730),
    ("Mesa",            "US-Southwest", "USA",          0, 0, 0, 0, 24657),
    ("Kansas City",     "US-Midwest",   "USA",          0, 0, 0, 0, 22963),
]
# Meta totals lines
LEASED_ENERGY = [795000, 964650, 1105834, 2187020, 3069504]
OTHER_ENERGY = [206000, 249843, 256939, 327073, 102016]
DC_TOTAL_ENERGY = [6966000, 9117122, 11167416, 14975435, 18061781]

# Campus detail (Dgtl Infra / Meta): online yr, buildings, sqft (M), investment ($B),
# third-party nameplate power estimate (MW, usdatamap) -- None if not tracked.
# ownership status all "Owned"
CAMPUS = [
    # name, region, country, online, buildings, sqft_m, invest_b, mw_est, status
    ("Prineville, OR",        "US-West",      "USA",     2011, 11, 4.6, 2.0, 180, "Operational"),
    ("Forest City, NC",       "US-Southeast", "USA",     2012, 3, 1.3, 0.8, 160, "Operational"),
    ("Luleå, Sweden",         "Europe",       "Sweden",  2013, 3, 1.0, 0.8, None, "Operational"),
    ("Altoona, IA",           "US-Midwest",   "USA",     2014, 10, 5.0, 2.5, 200, "Operational"),
    ("Fort Worth, TX",        "US-Southwest", "USA",     2017, 5, 2.6, 1.5, None, "Operational"),
    ("Clonee, Ireland",       "Europe",       "Ireland", 2018, 3, 1.6, 1.5, None, "Operational"),
    ("Los Lunas, NM",         "US-Southwest", "USA",     2019, 8, 3.8, 2.0, 100, "Operational"),
    ("Odense, Denmark",       "Europe",       "Denmark", 2019, 3, 0.9, 1.9, None, "Operational"),
    ("Sarpy County, NE",      "US-Midwest",   "USA",     2019, 9, 4.0, 1.5, 80, "Operational"),
    ("New Albany, OH",        "US-Midwest",   "USA",     2020, 5, 2.5, 1.5, 250, "Operational"),
    ("Henrico County, VA",    "US-Southeast", "USA",     2020, 5, 2.5, 1.0, 200, "Operational"),
    ("Huntsville, AL",        "US-Southeast", "USA",     2021, 6, 3.5, 1.5, 120, "Operational"),
    ("Newton Co (Stanton Springs), GA", "US-Southeast", "USA", 2021, 5, 2.5, 1.0, 200, "Operational"),
    ("Eagle Mountain, UT",    "US-West",      "USA",     2021, 7, 4.5, 1.5, None, "Operational"),
    ("Singapore",             "APAC",         "Singapore", 2022, 1, 1.8, 1.0, None, "Operational"),
    ("DeKalb, IL",            "US-Midwest",   "USA",     2022, 5, 2.4, 1.0, 300, "Operational"),
    ("Gallatin, TN",          "US-Southeast", "USA",     2023, 4, 1.6, 1.0, None, "Operational"),
    ("Mesa, AZ",              "US-Southwest", "USA",     2023, 5, 2.5, 1.0, None, "Operational"),
    ("Temple, TX",            "US-Southwest", "USA",     2024, 1, 0.9, 0.8, None, "Operational"),
    ("Kansas City, MO",       "US-Midwest",   "USA",     2024, 1, 0.7, 0.8, None, "Operational"),
    ("Kuna, ID",              "US-West",      "USA",     2025, 1, 1.0, 0.8, 100, "Operational/Ramping"),
    ("Jeffersonville, IN",    "US-Midwest",   "USA",     2026, 1, 0.7, 0.8, None, "Under construction"),
    ("Rosemount, MN",         "US-Midwest",   "USA",     2026, 1, 0.72, 0.8, None, "Under construction"),
    ("Montgomery, AL",        "US-Southeast", "USA",     2026, 1, 0.72, 0.8, None, "Under construction"),
]

# Forward AI GW build-out (public reporting / SemiAnalysis). ownership model noted.
AI_BUILD = [
    # project, location, capacity_gw, timeline, ownership, notes
    ("Prometheus", "New Albany, Ohio", 1.0, "~1 GW by 2026",
     "Owned self-build + leased + on-site gas",
     "World's first ~1GW-scale AI training cluster; Meta pre-leased more capacity in H2-2024 than any hyperscaler, mostly Ohio; two 200MW behind-the-meter gas plants w/ Williams."),
    ("Hyperion", "Richland Parish, Louisiana", 2.0, "2 GW by 2030 (1.5 GW IT by 2027); scalable to 5 GW",
     "JV-owned (Blue Owl 80% / Meta 20%), leased back to Meta",
     "$27-30B Blue Owl JV finances the campus; fully leased by Meta. >$10B Meta investment; ~4M+ sqft."),
    ("Lebanon (El Paso project)", "Lebanon, Indiana", 1.0, "~1 GW, from 2027",
     "Owned self-build", "~4M sqft, 13 buildings, $10B+."),
    ("Aiken (New Albany SC)", "Aiken, South Carolina", None, "Under construction",
     "Owned self-build", "AI-optimized campus, $800M+."),
    ("El Paso", "El Paso, Texas", None, "Under construction",
     "Owned self-build", "AI-optimized; reporting ~$10B."),
    ("Beaver Dam", "Beaver Dam, Wisconsin", None, "2027",
     "Owned self-build", "700k+ sqft, $1B+, closed-loop liquid cooling."),
    ("Cheyenne", "Cheyenne, Wyoming", None, "Announced/UC",
     "Owned self-build", "715k sqft."),
    ("Tulsa", "Tulsa, Oklahoma", None, "Announced Apr 2026",
     "Owned self-build", "Meta's 32nd data center globally."),
]

# Leased COLOCATION known markets (Dgtl Infra) -- Meta owns the compute here
LEASE_MARKETS = [
    "Ashburn, Virginia", "Manassas, Virginia", "Richmond, Virginia",
    "Aurora, Illinois", "Chicago, Illinois", "Dallas, Texas",
    "Houston, Texas", "Phoenix, Arizona", "Hillsboro, Oregon",
    "Santa Clara, California", "Chennai, India",
]

# NEOCLOUD / third-party CLOUD commitments (compute-as-a-service; provider owns GPUs)
# provider, category, announced, value_$b (None=undisclosed), term, scope, notes
NEOCLOUD_DEALS = [
    ("CoreWeave — initial", "Neocloud (GPU-as-a-service)", "Sep 2025", 14.2,
     "through Dec 14, 2031 (option to 2032)", "NVIDIA GB300 systems",
     "First major Meta neocloud contract; disclosed in CoreWeave SEC filing."),
    ("CoreWeave — expansion", "Neocloud (GPU-as-a-service)", "Apr 2026", 21.0,
     "2027-2032", "Dedicated capacity, multi-site; incl. NVIDIA Vera Rubin",
     "Brings Meta-CoreWeave total to ~$35B; supports large-scale AI inference."),
    ("Nebius", "Neocloud (GPU-as-a-service)", "2025", 12.0,
     "multi-year (+$15B option)", "Fixed dedicated compute",
     "Up to $27B including a $15B option exercisable at Nebius's discretion."),
    ("Google Cloud", "Hyperscaler cloud", "Aug 2025", 10.0,
     "6 years", "Servers, storage, networking for AI",
     "Minimum ~$10B; mainly AI infrastructure scaling."),
    ("Amazon Web Services", "Hyperscaler cloud", "Ongoing", None,
     "Ongoing", "Cloud infrastructure", "Long-standing; value undisclosed."),
    ("Microsoft Azure", "Hyperscaler cloud", "Ongoing", None,
     "Ongoing", "Cloud infrastructure", "Long-standing; value undisclosed."),
]

# COLOCATION rental rates -- by deal tier (CREFC data-center primer, 2026)
# tier, typical size, typical term, rate $/kW/month, note
COLO_RATE_TIER = [
    ("Retail / edge colocation", "< ~250 kW", "1-3 yr", "$200 - $400",
     "Highest unit price; smallest deployments."),
    ("Wholesale colocation", "~250 kW - a few MW", "5-10 yr", "$150 - $250",
     "Mid-size; typical enterprise / AI."),
    ("Hyperscale / build-to-suit", "5 - 40+ MW", "10-15 yr", "$100 - $150",
     "Largest deals, lowest unit rate — Meta's tier."),
]
# COLOCATION rates by metro (CBRE, 250-500 kW requirement)
COLO_RATE_METRO = [
    ("Primary N. America avg (250-500 kW)", "~$196 / kW/mo", "+6.6% YoY", "CBRE H2-2025 record; 4th straight annual rise."),
    ("Chicago", "$200 - $230", "+14.7% YoY", "Highest US primary rate (Q1-2026)."),
    ("Northern Virginia", "$190 - $235", "single-digit", "Largest US market."),
    ("Atlanta", "$160 - $180", "+2% YoY", "Fast-growing; vacancy ~1%."),
    ("Dallas-Ft. Worth", "market rate", "~flat", "Asking rents roughly unchanged YoY."),
    ("Frankfurt", "$235 - $265", "up YoY", "Highest European rate."),
    ("Singapore", "$310 - $475", "up YoY", "Highest APAC rate; supply-constrained."),
    ("10 MW+ deployments", "premium/scarce", "up to +19% YoY", "Large contiguous blocks command premiums."),
]
# COLOCATION utilization / occupancy
COLO_UTIL = [
    ("Primary-market vacancy (CBRE, YE2025)", "1.4%", "Record low."),
    ("Primary-market vacancy (JLL, YE2025)", "1.0%", "2nd straight year; ~99% occupancy."),
    ("Under-construction precommitted (JLL)", "92%", "Binding leases or owner-occupied."),
    ("Under-construction that is leased (JLL)", "~60%", "Remaining ~40% owner-occupied."),
    ("N. America capacity under construction (JLL)", "35+ GW", "Extraordinary by historical standards."),
    ("Net absorption, primary markets 2025 (CBRE)", "2,498 MW", "Record; prior peak 1,810 MW (2024)."),
    ("Occupancy 'yellow / red flag' (CREFC)", "<85% / <75%", "Credit-stress thresholds."),
    ("Power utilization 'yellow / red flag' (CREFC)", "<60% / <50%", "Increasingly the key tracked metric."),
]

# NEOCLOUD GPU rental rates ($/GPU-hour); on-demand unless noted
# sku, coreweave_8x, lambda, hyperscaler_ondemand(str), specialty_reserved3yr(str)
NEO_GPU_RATES = [
    ("A100 80GB", 2.70, 1.29, "$1.50 - $3.00", "—"),
    ("H100 SXM", 6.16, 3.29, "$3.50 - $5.00", "$1.70 - $2.30"),
    ("H200 SXM", 6.30, 3.50, "$4.50 - $7.00", "—"),
    ("B200", 8.60, 4.99, "$5.50+", "—"),
    ("GB200 NVL72", 10.50, None, "n/d", "—"),
    ("GH200", 6.50, 1.49, "n/d", "—"),
]
# NEOCLOUD utilization / economics
NEO_UTIL = [
    ("Debt-financed cluster break-even utilization", "~65-70%", "Little margin of safety (American Compute)."),
    ("1,024x H100 cluster @ 55% vs 85% util", "-$330k vs +$340k /mo", "Illustrates utilization sensitivity."),
    ("CoreWeave Model FLOPs Utilization (MFU)", "50%+", "Claimed vs industry typical 30-45%."),
    ("CoreWeave active power (Q1-2026)", ">1.0 GW", "Plan: ~1.7 GW active in 2026."),
    ("CoreWeave contracted power (Q1-2026)", ">3.5 GW", "Active/contracted ≈ 29% (energization gap)."),
    ("CoreWeave revenue backlog (Mar-2026)", "$99.4B", ">50% est. from Meta / OpenAI / hyperscalers."),
    ("Bare-metal gross margin (pre-depreciation)", "55-65%", "Before heavy GPU depreciation."),
    ("Neocloud market size (2025)", "$25B+", "Q4-2025 $9B, +223% YoY (Synergy Research)."),
    ("Reserved-contract discounts", "1yr 25-40%; 3yr up to ~50%", "Spot 50-80% below on-demand."),
]

# ============================================================================
# SHEET 1 — COVER & KEY TAKEAWAYS
# ============================================================================
ws = wb.active
ws.title = "Cover & Notes"
for col, w in zip("AB", (26, 96)):
    ws.column_dimensions[col].width = w
title_block(ws, "Meta Platforms — Total Compute Capacity: Owned vs. Leased",
            "Data-center footprint aggregated from Meta disclosures and third-party trackers  |  Compiled for equity research",
            span=2)

r = 4
intro = [
    ("What this workbook shows",
     "A granular aggregation of Meta's total data-center (compute) capacity, split between company-OWNED "
     "facilities and LEASED / colocation capacity. In the data-center industry, capacity is measured in "
     "electric power (MW / GW) because power is the binding constraint on how much compute (servers / GPUs) "
     "a site can host."),
    ("The cleanest owned-vs-leased split",
     "Meta does not publish a single 'owned vs leased MW' figure. However, its annual Environmental Data Index "
     "reports ACTUAL electricity consumption (MWh/yr) for every owned online data center individually, plus one "
     "line for 'Leased data center facilities.' That is the most authoritative apples-to-apples split of realized "
     "compute and anchors the 'Summary' tab."),
    ("Three procurement modes — the key distinction",
     "Meta's compute is sourced three ways: (1) OWNED self-build data centers; (2) LEASED COLOCATION — Meta leases "
     "physical space & power (powered shells / wholesale colo) from REITs/operators and installs its OWN servers/GPUs "
     "(measured in MW & sqft; shows up in Meta's leased-facility electricity); and (3) NEOCLOUD / CLOUD — Meta rents "
     "GPU compute-as-a-service, where a THIRD PARTY owns the GPUs & the building (measured in $ multi-year "
     "commitments / GPU-hours; it is opex, not in Meta's owned-electricity reporting). This workbook now separates "
     "colocation from neocloud."),
    ("Headline #1 — physical footprint (FY2024, by electricity consumed)",
     "Owned online data centers: 14.89 TWh (~82%)  |  Leased colocation: 3.07 TWh (~17%)  |  Other DC-related: "
     "0.10 TWh (~1%).  Total data-center electricity = 18.06 TWh. Leased colocation share roughly doubled since "
     "2020-22 (~9-11%) as Meta pre-leased heavily for AI."),
    ("Headline #2 — neocloud / cloud commitments ($, off the electricity split)",
     "CoreWeave ~$35B (initial $14.2B Sep-2025 + $21B expansion Apr-2026, through 2032; GB300 / Vera Rubin) | "
     "Nebius up to $27B ($12B fixed + $15B option) | Google Cloud $10B+ over 6 yrs (Aug-2025) | plus ongoing AWS & "
     "Azure. ~$57B+ disclosed (up to ~$72B with options). These are compute rentals, NOT part of the owned-vs-"
     "colocation electricity split (different unit)."),
    ("Forward-looking AI capacity",
     "Realized power today is dwarfed by the announced AI build-out. Trackers count ~15.8 GW of Meta AI data-center "
     "capacity across ~20 sites (~50% operational). Flagship clusters: Prometheus (Ohio, ~1 GW) and Hyperion "
     "(Louisiana, 2 GW scaling to 5 GW). Hyperion uses a $27-30B Blue Owl JV (Meta ~20% equity) that is leased "
     "back to Meta — an owned/leased hybrid. See 'AI Buildout' tab."),
    ("Tabs",
     "1) Summary — owned vs colocation vs neocloud.  2) Owned - Site Energy — per-site MWh 2020-24.  "
     "3) Owned - Campus Detail — buildings, sqft, investment, nameplate MW est.  "
     "4) Colocation (space & power) — leased-facility energy trend + markets.  "
     "5) Neocloud & Cloud — GPU-as-a-service commitments.  "
     "6) Utilization & Rental Rates — colocation ($/kW/mo, vacancy) & neocloud ($/GPU-hr, utilization).  "
     "7) AI Buildout (GW).  8) Sources & Methodology."),
    ("Important caveats",
     "• Electricity consumption ≠ nameplate MW, but is the best public realized-capacity proxy and is Meta-reported & "
     "third-party assured.  • Colocation and neocloud are measured in DIFFERENT units (MWh vs $) and are not additive.  "
     "• Neocloud $ are multi-year contract values, not annual spend or installed capacity.  • Rental rates & "
     "utilization on tab 6 are market benchmarks (CBRE/JLL/CREFC/neocloud disclosures), not Meta-specific contract "
     "terms, which are confidential.  • GW figures for AI clusters are targets/announcements.  • FY2024 = calendar 2024."),
]
for h, b in intro:
    ws.cell(r, 1, h).font = BOLD_NAVY
    ws.cell(r, 1).alignment = LEFTTOP
    c = ws.cell(r, 2, b)
    c.alignment = LEFTTOP
    c.font = REG
    ws.row_dimensions[r].height = max(30, 15 * (len(b) // 95 + 1))
    r += 1

ws.cell(r + 1, 1, "Units: MWh = megawatt-hours (energy/yr) · MW = megawatts (power) · GW = 1,000 MW · TWh = 1,000,000 MWh · sqft = square feet").font = NOTE_FONT

# ============================================================================
# SHEET 2 — SUMMARY: OWNED vs LEASED (authoritative, energy based)
# ============================================================================
ws = wb.create_sheet("Summary - Owned vs Leased")
title_block(ws, "Summary — Owned vs Leased Compute Capacity",
            "Basis: actual data-center electricity consumption (MWh/yr), Meta 2025 Environmental Data Index (FY2024). Power = capacity proxy.",
            span=8)
for i, w in enumerate([34, 14, 14, 14, 14, 14, 12, 12], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# owned total per year = sum of owned energy
owned_tot = [sum(row[3 + k] for row in OWNED_ENERGY) for k in range(5)]

r = 4
ws.cell(r, 1, "Data-center electricity consumption (MWh/yr)").font = BOLD_NAVY
r += 1
head = ["Category"] + [str(y) for y in YEARS] + ["2024 Share", "CAGR 20-24"]
for c, h in enumerate(head, start=1):
    hdr(ws.cell(r, c, h))
rows_def = [
    ("Owned online data centers", owned_tot, OWNED_FILL),
    ("Leased colocation facilities (Meta owns the compute)", LEASED_ENERGY, LEASED_FILL),
    ("Other DC-related (warehouse/network/small colo)", OTHER_ENERGY, GREY),
]
first_data = r + 1
for name, vals, fill in rows_def:
    r += 1
    c = ws.cell(r, 1, name)
    c.font = BOLD if fill != GREY else REG
    c.fill = PatternFill("solid", fgColor=fill)
    c.border = border_all
    c.alignment = LEFT
    for k, v in enumerate(vals):
        cell = ws.cell(r, 2 + k, v)
        cell.number_format = FMT_INT
        cell.border = border_all
        cell.fill = PatternFill("solid", fgColor=fill)
    # share of 2024 total
    sc = ws.cell(r, 7, f"={get_column_letter(6)}{r}/{get_column_letter(6)}{first_data+3}")
    sc.number_format = FMT_PCT
    sc.border = border_all
    sc.fill = PatternFill("solid", fgColor=fill)
    # CAGR
    cg = ws.cell(r, 8, f'=IFERROR(({get_column_letter(6)}{r}/{get_column_letter(2)}{r})^(1/4)-1,"n/m")')
    cg.number_format = FMT_PCT
    cg.border = border_all
    cg.fill = PatternFill("solid", fgColor=fill)
last_data = r
# total row
r += 1
tc = ws.cell(r, 1, "Total data-center electricity")
tc.font = BOLD
tc.fill = PatternFill("solid", fgColor=LIGHT)
tc.border = border_all
for k in range(5):
    col = get_column_letter(2 + k)
    cell = ws.cell(r, 2 + k, f"=SUM({col}{first_data}:{col}{last_data})")
    cell.number_format = FMT_INT
    cell.font = BOLD
    cell.fill = PatternFill("solid", fgColor=LIGHT)
    cell.border = border_all
sc = ws.cell(r, 7, f"=G{first_data}+G{first_data+1}+G{first_data+2}")
sc.number_format = FMT_PCT
sc.font = BOLD; sc.fill = PatternFill("solid", fgColor=LIGHT); sc.border = border_all
ws.cell(r, 8).fill = PatternFill("solid", fgColor=LIGHT); ws.cell(r, 8).border = border_all
total_row = r
# cross-check row
r += 1
ws.cell(r, 1, "Meta-reported 'Data centers total' (check)").font = NOTE_FONT
for k, v in enumerate(DC_TOTAL_ENERGY):
    cc = ws.cell(r, 2 + k, v); cc.number_format = FMT_INT; cc.font = NOTE_FONT

# Leased share row
r += 2
ws.cell(r, 1, "Leased share of data-center electricity").font = BOLD_NAVY
r += 1
hdr(ws.cell(r, 1, "Metric"))
for k, y in enumerate(YEARS):
    hdr(ws.cell(r, 2 + k, str(y)))
r += 1
ws.cell(r, 1, "Leased % of total").border = border_all
ws.cell(r, 1).alignment = LEFT
for k in range(5):
    col = get_column_letter(2 + k)
    cell = ws.cell(r, 2 + k, f"={col}{first_data+1}/{col}{total_row}")
    cell.number_format = FMT_PCT
    cell.border = border_all
r += 1
ws.cell(r, 1, "Owned % of total").border = border_all
ws.cell(r, 1).alignment = LEFT
for k in range(5):
    col = get_column_letter(2 + k)
    cell = ws.cell(r, 2 + k, f"=({col}{first_data})/{col}{total_row}")
    cell.number_format = FMT_PCT
    cell.border = border_all

# ---- Neocloud / cloud commitments block (separate unit: $) ----
r += 2
ws.cell(r, 1, "Neocloud / third-party cloud commitments ($ multi-year) — separate from the electricity split above").font = BOLD_NAVY
r += 1
for c, h in enumerate(["Counterparty", "Category", "Announced", "Value ($B)", "Term"], start=1):
    hdr(ws.cell(r, c, h))
neo_start = r + 1
for prov, cat, ann, val, term, scope, notes in NEOCLOUD_DEALS:
    r += 1
    ws.cell(r, 1, prov).alignment = LEFT
    ws.cell(r, 2, cat).alignment = LEFT
    ws.cell(r, 3, ann).alignment = CENTER
    vc = ws.cell(r, 4, val if val is not None else "n/d")
    vc.alignment = CENTER
    if isinstance(val, (int, float)):
        vc.number_format = FMT_1
    ws.cell(r, 5, term).alignment = LEFT
    for c in range(1, 6):
        ws.cell(r, c).border = border_all
        ws.cell(r, c).fill = PatternFill("solid", fgColor="FCE4D6")  # light orange = neocloud
neo_end = r
r += 1
ws.cell(r, 1, "Total disclosed neocloud/cloud commitments").font = BOLD
tc = ws.cell(r, 4, f'=SUMIF(D{neo_start}:D{neo_end},"<>n/d")')
tc.number_format = FMT_1; tc.font = BOLD
for c in range(1, 6):
    ws.cell(r, c).border = border_all
    ws.cell(r, c).fill = PatternFill("solid", fgColor="F8CBAD")
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
ws.cell(r, 1, "Taxonomy: OWNED = Meta builds & owns the DC and compute. COLOCATION = Meta leases space/power, owns the "
              "compute (in electricity split above, amber). NEOCLOUD/CLOUD = Meta rents GPU compute-as-a-service; the "
              "provider owns the GPUs & building (this block, orange; measured in $, not MWh — not additive with the "
              "table above). See Neocloud & Cloud and Utilization & Rental Rates tabs.").font = NOTE_FONT
ws.row_dimensions[r].height = 42

# ---- Pie chart (2024 split) ----
pie = PieChart()
pie.title = "FY2024 DC electricity: Owned vs Leased vs Other"
labels = Reference(ws, min_col=1, min_row=first_data, max_row=first_data + 2)
data = Reference(ws, min_col=6, min_row=first_data - 1, max_row=first_data + 2)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.height = 7.5
pie.width = 12
ws.add_chart(pie, "A" + str(r + 3))

# ---- Bar chart (trend) ----
bar = BarChart()
bar.type = "col"
bar.grouping = "stacked"
bar.overlap = 100
bar.title = "DC electricity by ownership (MWh/yr)"
cats = Reference(ws, min_col=2, max_col=6, min_row=first_data - 1, max_row=first_data - 1)
bdata = Reference(ws, min_col=1, max_col=6, min_row=first_data, max_row=first_data + 2)
# Build a clean bar using transposed structure via helper table
# Simpler: chart the three category rows across years
bdata2 = Reference(ws, min_col=2, max_col=6, min_row=first_data, max_row=first_data + 2)
bar.add_data(bdata2, from_rows=True, titles_from_data=False)
bar.set_categories(Reference(ws, min_col=2, max_col=6, min_row=5, max_row=5))
bar.series[0].tx = None
bar.height = 7.5
bar.width = 14
ws.add_chart(bar, "H" + str(r + 3))

ws.cell(r + 2, 1, "Note: 'compute capacity' proxied by realized electricity consumption; see Cover & Sources.").font = NOTE_FONT

# ============================================================================
# SHEET 3 — OWNED: SITE ENERGY DETAIL
# ============================================================================
ws = wb.create_sheet("Owned - Site Energy")
title_block(ws, "Owned Data Centers — Electricity Consumption by Site",
            "MWh/yr, per site, 2020-2024. Source: Meta 2025 Environmental Data Index (FY2024). Sorted by FY2024.",
            span=9)
widths = [30, 15, 12, 13, 13, 13, 13, 13, 12]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
r = 4
head = ["Owned Site", "Region", "Country"] + [str(y) for y in YEARS] + ["'24 % of owned"]
for c, h in enumerate(head, start=1):
    hdr(ws.cell(r, c, h))
data_start = r + 1
for row in OWNED_ENERGY:
    r += 1
    ws.cell(r, 1, row[0]); ws.cell(r, 2, row[1]); ws.cell(r, 3, row[2])
    for k in range(5):
        cell = ws.cell(r, 4 + k, row[3 + k])
        cell.number_format = FMT_INT
    for c in range(1, 10):
        ws.cell(r, c).border = border_all
        if c <= 3:
            ws.cell(r, c).alignment = LEFT
    if r % 2 == 0:
        for c in range(1, 10):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=LIGHTER)
data_end = r
# share col (of owned total 2024)
for rr in range(data_start, data_end + 1):
    cell = ws.cell(rr, 9, f"=H{rr}/SUM($H${data_start}:$H${data_end})")
    cell.number_format = FMT_PCT
    cell.border = border_all
# total row
r += 1
ws.cell(r, 1, "Total owned online data centers").font = BOLD
ws.cell(r, 2, "").font = BOLD
for k in range(5):
    col = get_column_letter(4 + k)
    cell = ws.cell(r, 4 + k, f"=SUM({col}{data_start}:{col}{data_end})")
    cell.number_format = FMT_INT
    cell.font = BOLD
ws.cell(r, 9, f"=SUM(I{data_start}:I{data_end})").number_format = FMT_PCT
for c in range(1, 10):
    ws.cell(r, c).fill = PatternFill("solid", fgColor=OWNED_FILL)
    ws.cell(r, c).border = border_all
    ws.cell(r, c).font = BOLD
# memo: leased + other
r += 2
ws.cell(r, 1, "Memo — Leased colocation facilities").font = REG
for k, v in enumerate(LEASED_ENERGY):
    ws.cell(r, 4 + k, v).number_format = FMT_INT
for c in range(1, 9):
    ws.cell(r, c).fill = PatternFill("solid", fgColor=LEASED_FILL)
r += 1
ws.cell(r, 1, "Memo — Other DC-related facilities").font = REG
for k, v in enumerate(OTHER_ENERGY):
    ws.cell(r, 4 + k, v).number_format = FMT_INT

ws.freeze_panes = "D5"

# ============================================================================
# SHEET 4 — OWNED: CAMPUS DETAIL
# ============================================================================
ws = wb.create_sheet("Owned - Campus Detail")
title_block(ws, "Owned Data Centers — Campus Detail (buildings, area, investment, power)",
            "Sources: Dgtl Infra / Meta (buildings, sqft, investment). Nameplate MW = third-party estimate (usdatamap), incomplete.",
            span=9)
widths = [34, 15, 11, 10, 11, 12, 13, 14, 18]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
r = 4
head = ["Campus", "Region", "Country", "First online", "Buildings",
        "Area (M sqft)", "Investment ($B)", "Nameplate MW (est.)", "Status"]
for c, h in enumerate(head, start=1):
    hdr(ws.cell(r, c, h))
data_start = r + 1
for row in CAMPUS:
    r += 1
    name, region, country, online, bld, sqft, inv, mw, status = row
    vals = [name, region, country, online, bld, sqft, inv, mw if mw is not None else "n/a", status]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(r, c, v)
        cell.border = border_all
        if c in (1, 2, 3, 9):
            cell.alignment = LEFT
        else:
            cell.alignment = CENTER
    ws.cell(r, 5).number_format = FMT_INT
    ws.cell(r, 6).number_format = FMT_1
    ws.cell(r, 7).number_format = FMT_1
    if isinstance(mw, (int, float)):
        ws.cell(r, 8).number_format = FMT_INT
    if r % 2 == 0:
        for c in range(1, 10):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=LIGHTER)
data_end = r
# totals
r += 1
ws.cell(r, 1, "TOTAL (owned campuses)").font = BOLD
ws.cell(r, 5, f"=SUM(E{data_start}:E{data_end})").number_format = FMT_INT
ws.cell(r, 6, f"=SUM(F{data_start}:F{data_end})").number_format = FMT_1
ws.cell(r, 7, f"=SUM(G{data_start}:G{data_end})").number_format = FMT_1
ws.cell(r, 8, f'=SUMIF(H{data_start}:H{data_end},"<>n/a")').number_format = FMT_INT
for c in range(1, 10):
    ws.cell(r, c).fill = PatternFill("solid", fgColor=OWNED_FILL)
    ws.cell(r, c).border = border_all
    ws.cell(r, c).font = BOLD
ws.cell(r + 1, 1, "Note: nameplate MW total is only for the subset of sites the third-party tracker reports and understates true installed power; it excludes the AI build-out (see AI Buildout tab).").font = NOTE_FONT
ws.freeze_panes = "A5"

# ============================================================================
# SHEET 5 — COLOCATION (space & power; Meta owns the compute)
# ============================================================================
ws = wb.create_sheet("Colocation (space & power)")
title_block(ws, "Leased Colocation — Space & Power (Meta owns the compute)",
            "Meta leases powered shells / wholesale colo and installs its own servers. Sources: Meta Env. Data Index; Dgtl Infra; SemiAnalysis.",
            span=6)
for i, w in enumerate([34, 14, 14, 14, 14, 14], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
r = 4
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws.cell(r, 1, "Definition: colocation = Meta rents physical data-center space & power from third parties (REITs/operators) "
              "but owns and operates the servers/GPUs inside. It IS captured in Meta's 'Leased data center facilities' "
              "electricity line. (Contrast with neocloud — see next tab.)").font = NOTE_FONT
ws.row_dimensions[r].height = 42
r += 2
ws.cell(r, 1, "Leased colocation electricity consumption (MWh/yr)").font = BOLD_NAVY
r += 1
hdr(ws.cell(r, 1, "Metric"))
for k, y in enumerate(YEARS):
    hdr(ws.cell(r, 2 + k, str(y)))
r += 1
ws.cell(r, 1, "Leased colocation facilities (MWh)").border = border_all
ws.cell(r, 1).alignment = LEFT
for k, v in enumerate(LEASED_ENERGY):
    cell = ws.cell(r, 2 + k, v); cell.number_format = FMT_INT; cell.border = border_all
    cell.fill = PatternFill("solid", fgColor=LEASED_FILL)
lease_row = r
r += 1
ws.cell(r, 1, "YoY growth").border = border_all
ws.cell(r, 1).alignment = LEFT
ws.cell(r, 2, "n/a").alignment = CENTER; ws.cell(r, 2).border = border_all
for k in range(1, 5):
    col = get_column_letter(2 + k); pcol = get_column_letter(1 + k)
    cell = ws.cell(r, 2 + k, f"={col}{lease_row}/{pcol}{lease_row}-1")
    cell.number_format = FMT_PCT; cell.border = border_all

r += 2
ws.cell(r, 1, "Known colocation markets (Meta)").font = BOLD_NAVY
r += 1
hdr(ws.cell(r, 1, "#"))
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
hdr(ws.cell(r, 2, "Market"))
for i, m in enumerate(LEASE_MARKETS, start=1):
    r += 1
    ws.cell(r, 1, i).border = border_all; ws.cell(r, 1).alignment = CENTER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    c = ws.cell(r, 2, m); c.border = border_all; c.alignment = LEFT

r += 2
ws.cell(r, 1, "Key relationships & context").font = BOLD_NAVY
notes = [
    "Digital Realty (DLR): Meta is its ~11th-largest customer; ~48 leased locations; ~$61.8M annualized base rent (Dgtl Infra).",
    "SemiAnalysis: Meta pre-leased more colocation capacity in H2-2024 than any other hyperscaler, concentrated in Ohio.",
    "Dual-track 'all of the above' strategy: self-build + colocation leasing + neocloud + on-site gas generation.",
    "Hyperion (Louisiana) is financed via a $27-30B Blue Owl JV (Meta ~20% equity) and leased back to Meta — an owned/leased hybrid (see AI Buildout).",
    "Leased colocation electricity nearly quadrupled 2020->2024 (0.80 -> 3.07 TWh); colocation share of DC power rose from ~11% to ~17%.",
    "Market rental rates & utilization/vacancy for colocation are on the 'Utilization & Rental Rates' tab.",
]
for n in notes:
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(r, 1, "• " + n); c.alignment = LEFTTOP; c.font = REG
    ws.row_dimensions[r].height = 30

# ============================================================================
# SHEET 6 — NEOCLOUD & CLOUD (compute-as-a-service; provider owns the GPUs)
# ============================================================================
ws = wb.create_sheet("Neocloud & Cloud")
title_block(ws, "Neocloud / Third-Party Cloud — Compute-as-a-Service",
            "Meta rents GPU capacity; the provider owns the GPUs & building. Measured in $ commitments (opex), not Meta MWh.",
            span=7)
for i, w in enumerate([22, 26, 12, 12, 26, 26, 40], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
r = 4
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
ws.cell(r, 1, "Definition: neocloud/cloud = Meta buys managed GPU compute from a third party that owns the chips and "
              "data centers (CoreWeave, Nebius, Google Cloud, AWS, Azure). This is NOT in Meta's owned/colocation "
              "electricity split; it is a separate, $-denominated procurement channel.").font = NOTE_FONT
ws.row_dimensions[r].height = 42
r += 2
head = ["Counterparty", "Category", "Announced", "Value ($B)", "Term", "Scope / hardware", "Notes"]
for c, h in enumerate(head, start=1):
    hdr(ws.cell(r, c, h))
ns = r + 1
for prov, cat, ann, val, term, scope, notes in NEOCLOUD_DEALS:
    r += 1
    vals = [prov, cat, ann, val if val is not None else "n/d", term, scope, notes]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(r, c, v)
        cell.border = border_all
        cell.alignment = CENTER if c in (3, 4) else LEFTTOP
        cell.font = REG
    if isinstance(val, (int, float)):
        ws.cell(r, 4).number_format = FMT_1
    ws.row_dimensions[r].height = 34
ne = r
r += 1
ws.cell(r, 1, "Total disclosed commitments").font = BOLD
tc = ws.cell(r, 4, f'=SUMIF(D{ns}:D{ne},"<>n/d")'); tc.number_format = FMT_1; tc.font = BOLD
for c in range(1, 8):
    ws.cell(r, c).border = border_all
    ws.cell(r, c).fill = PatternFill("solid", fgColor="F8CBAD")
r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
ws.cell(r, 1, "Context").font = BOLD_NAVY
ncontext = [
    "Meta describes this as a 'portfolio-based approach to infrastructure' — self-build + colocation + neocloud together.",
    "CoreWeave total with Meta ~$35B (through 2032); CoreWeave overall backlog $99.4B (Mar-2026) on >3.5 GW contracted power.",
    "Nebius Meta deal up to $27B: $12B fixed dedicated compute + $15B option (exercisable by Nebius).",
    "Google Cloud: $10B+ / 6 yrs; Meta historically also uses AWS and Azure (values undisclosed).",
    "Neocloud GPU rental rates and utilization economics are on the 'Utilization & Rental Rates' tab.",
]
for n in ncontext:
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    c = ws.cell(r, 1, "• " + n); c.alignment = LEFTTOP; c.font = REG
    ws.row_dimensions[r].height = 28

# ============================================================================
# SHEET 7 — UTILIZATION & RENTAL RATES
# ============================================================================
ws = wb.create_sheet("Utilization & Rental Rates")
title_block(ws, "Utilization & Rental Rates — Colocation vs Neocloud",
            "Market benchmarks (NOT Meta-specific contract terms, which are confidential). Sources: CBRE, JLL, CREFC, neocloud disclosures.",
            span=6)
for i, w in enumerate([40, 20, 18, 44, 12, 12], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

r = 4
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
h = ws.cell(r, 1, "A.  COLOCATION (space & power leasing)")
h.font = Font(bold=True, size=12, color="FFFFFF")
h.fill = PatternFill("solid", fgColor=LEASED_FILL[:6] if False else "BF8F00")
h.alignment = LEFT
r += 1
ws.cell(r, 1, "Rental rates by deal tier ($/kW/month)").font = BOLD_NAVY
r += 1
for c, hh in enumerate(["Tier", "Typical size", "Typical term", "Rate ($/kW/mo)", "", ""], start=1):
    if hh:
        hdr(ws.cell(r, c, hh))
for tier, size, term, rate, note in COLO_RATE_TIER:
    r += 1
    ws.cell(r, 1, tier).alignment = LEFT
    ws.cell(r, 2, size).alignment = CENTER
    ws.cell(r, 3, term).alignment = CENTER
    ws.cell(r, 4, rate).alignment = CENTER
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=4)
    ws.cell(r, 4).font = BOLD
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    ws.cell(r, 5, note).alignment = LEFT; ws.cell(r, 5).font = NOTE_FONT
    for c in range(1, 5):
        ws.cell(r, c).border = border_all
r += 2
ws.cell(r, 1, "Rental rates by market (250-500 kW requirement)").font = BOLD_NAVY
r += 1
for c, hh in enumerate(["Market", "Asking rate", "YoY", "Note"], start=1):
    hdr(ws.cell(r, c, hh))
    if c == 4:
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        hdr(ws.cell(r, 4, "Note"))
for metro, rate, yoy, note in COLO_RATE_METRO:
    r += 1
    ws.cell(r, 1, metro).alignment = LEFT
    ws.cell(r, 2, rate).alignment = CENTER
    ws.cell(r, 3, yoy).alignment = CENTER
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    ws.cell(r, 4, note).alignment = LEFT; ws.cell(r, 4).font = NOTE_FONT
    for c in range(1, 4):
        ws.cell(r, c).border = border_all
    ws.cell(r, 4).border = border_all
r += 2
ws.cell(r, 1, "Utilization / occupancy").font = BOLD_NAVY
r += 1
for c, hh in enumerate(["Metric", "Value", "Note"], start=1):
    hdr(ws.cell(r, c, hh))
    if c == 3:
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        hdr(ws.cell(r, 3, "Note"))
for metric, val, note in COLO_UTIL:
    r += 1
    ws.cell(r, 1, metric).alignment = LEFT
    ws.cell(r, 2, val).alignment = CENTER; ws.cell(r, 2).font = BOLD
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws.cell(r, 3, note).alignment = LEFT; ws.cell(r, 3).font = NOTE_FONT
    for c in range(1, 3):
        ws.cell(r, c).border = border_all
    ws.cell(r, 3).border = border_all

r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
h = ws.cell(r, 1, "B.  NEOCLOUD (GPU compute-as-a-service)")
h.font = Font(bold=True, size=12, color="FFFFFF")
h.fill = PatternFill("solid", fgColor="C55A11")
h.alignment = LEFT
r += 1
ws.cell(r, 1, "GPU rental rates ($/GPU-hour, on-demand unless noted)").font = BOLD_NAVY
r += 1
for c, hh in enumerate(["GPU SKU", "CoreWeave (8x)", "Lambda", "Hyperscaler on-dmd", "Specialty rsv 3yr", ""], start=1):
    if hh:
        hdr(ws.cell(r, c, hh))
for sku, cw, lam, hyp, rsv in NEO_GPU_RATES:
    r += 1
    ws.cell(r, 1, sku).alignment = LEFT
    a = ws.cell(r, 2, cw); a.alignment = CENTER
    if isinstance(cw, (int, float)):
        a.number_format = '"$"#,##0.00'
    b = ws.cell(r, 3, lam if lam is not None else "n/d"); b.alignment = CENTER
    if isinstance(lam, (int, float)):
        b.number_format = '"$"#,##0.00'
    ws.cell(r, 4, hyp).alignment = CENTER
    ws.cell(r, 5, rsv).alignment = CENTER
    for c in range(1, 6):
        ws.cell(r, c).border = border_all
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws.cell(r, 1, "Discounts: reserved 1yr ~25-40% off on-demand; 3yr up to ~50%; spot ~50-80% below on-demand. "
              "CoreWeave bills 8-GPU clusters (premium for guaranteed capacity/orchestration); long-tail H100 ~$1.99-2.50.").font = NOTE_FONT
ws.row_dimensions[r].height = 28
r += 2
ws.cell(r, 1, "Utilization & unit economics").font = BOLD_NAVY
r += 1
for c, hh in enumerate(["Metric", "Value", "Note"], start=1):
    hdr(ws.cell(r, c, hh))
    if c == 3:
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        hdr(ws.cell(r, 3, "Note"))
for metric, val, note in NEO_UTIL:
    r += 1
    ws.cell(r, 1, metric).alignment = LEFT
    ws.cell(r, 2, val).alignment = CENTER; ws.cell(r, 2).font = BOLD
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws.cell(r, 3, note).alignment = LEFT; ws.cell(r, 3).font = NOTE_FONT
    for c in range(1, 3):
        ws.cell(r, c).border = border_all
    ws.cell(r, 3).border = border_all
r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws.cell(r, 1, "Note: Meta's own contracted colocation rents and neocloud rates are confidential; figures above are "
              "market benchmarks used to frame Meta's economics. Owned/self-build effective cost is often quoted at "
              "~$1.40-1.80/GPU-hr for a single GPU at ~70% utilization (before cluster/networking overhead).").font = NOTE_FONT
ws.row_dimensions[r].height = 40

# ============================================================================
# SHEET 8 — AI BUILDOUT (GW)
# ============================================================================
ws = wb.create_sheet("AI Buildout (GW)")
title_block(ws, "Forward AI Capacity Build-out (nameplate GW)",
            "Announced/targeted AI clusters. Sources: Meta, SemiAnalysis, Data Center Frontier, CBRE, AI Data Center Index. Targets, not installed.",
            span=6)
for i, w in enumerate([24, 26, 12, 30, 30, 44], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
r = 4
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws.cell(r, 1, "Tracker view: ~15.8 GW of Meta AI data-center capacity across ~20 sites, ~50% operational, in 5 countries (AI Data Center Index, 2026).").font = BOLD_NAVY
r += 2
head = ["Project", "Location", "Capacity (GW)", "Timeline", "Ownership model", "Notes"]
for c, h in enumerate(head, start=1):
    hdr(ws.cell(r, c, h))
data_start = r + 1
for row in AI_BUILD:
    r += 1
    proj, loc, gw, tl, own, notes = row
    vals = [proj, loc, gw if gw is not None else "n/d", tl, own, notes]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(r, c, v)
        cell.border = border_all
        cell.alignment = CENTER if c == 3 else LEFTTOP
        cell.font = REG
    if isinstance(gw, (int, float)):
        ws.cell(r, 3).number_format = FMT_1
    fill = OWNED_FILL if own.startswith("Owned") else (LEASED_FILL if "JV" in own or "leased" in own else GREY)
    ws.cell(r, 5).fill = PatternFill("solid", fgColor=fill)
    ws.row_dimensions[r].height = 42
data_end = r
r += 1
ws.cell(r, 1, "Sum of quantified GW targets").font = BOLD
ws.cell(r, 3, f'=SUMIF(C{data_start}:C{data_end},"<>n/d")').number_format = FMT_1
ws.cell(r, 3).font = BOLD
for c in range(1, 7):
    ws.cell(r, c).border = border_all
r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws.cell(r, 1, "Legend: green = owned self-build; amber = JV-owned/leased hybrid; grey = ownership TBD/self-build unquantified. "
              "GW figures are targets/announcements and include multi-year phasing.").font = NOTE_FONT

# ============================================================================
# SHEET 9 — SOURCES & METHODOLOGY
# ============================================================================
ws = wb.create_sheet("Sources & Methodology")
for i, w in enumerate([40, 100], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
title_block(ws, "Sources & Methodology", "All figures as-disclosed; compiled from public sources.", span=2)
r = 4
ws.cell(r, 1, "Methodology").font = BOLD_NAVY
r += 1
meth = [
    "Three procurement modes are tracked separately: (1) OWNED self-build; (2) LEASED COLOCATION — Meta leases space/power and owns the compute; (3) NEOCLOUD/CLOUD — Meta rents compute-as-a-service and the provider owns the GPUs. Owned + colocation are comparable in electricity (MWh); neocloud is a distinct $-denominated channel and is NOT additive with the electricity split.",
    "Capacity proxy: Data-center capacity is measured in electric power. Because Meta does not disclose a single owned-vs-leased MW figure, the owned vs colocation split uses ACTUAL electricity consumption (MWh/yr) as the realized-capacity proxy — the one metric Meta reports separately for owned sites and for leased facilities.",
    "Utilization & rental rates (tab 6) are external market benchmarks (CBRE/JLL/CREFC for colocation; neocloud disclosures & third-party price trackers for GPU rates). They frame Meta's economics but are NOT Meta's confidential contract terms.",
    "Owned online data centers are reported individually by Meta (18 sites with material 2024 consumption). 'Leased data center facilities' and 'Other DC-related facilities' are each reported as single lines.",
    "Nameplate MW (Campus Detail) are third-party estimates (usdatamap) available only for a subset of sites; they understate installed power and exclude the AI expansion. Use directionally.",
    "AI build-out GW are announced targets / analyst estimates with multi-year phasing — NOT installed capacity.",
    "Owned campus counts vary by source and date (Meta cites '27-28 owned locations in operation or under construction'; Dgtl Infra lists 24 established campuses; newer sites e.g. Tulsa push counts higher). Reconciliation notes included where relevant.",
]
for m in meth:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c = ws.cell(r, 1, "• " + m); c.alignment = LEFTTOP; c.font = REG
    ws.row_dimensions[r].height = 45
    r += 1
r += 1
ws.cell(r, 1, "Sources").font = BOLD_NAVY
r += 1
hdr(ws.cell(r, 1, "Source")); hdr(ws.cell(r, 2, "URL / detail"))
sources = [
    ("Meta 2025 Environmental Data Index (FY2024)", "https://sustainability.atmeta.com/wp-content/uploads/2025/10/Meta_2025-Environmental-Data-Index.pdf — per-site electricity (MWh), water, GHG; owned vs leased lines."),
    ("Meta 2025 Sustainability Report", "https://www.smartenergydecisions.com/wp-content/uploads/2025/10/Meta_2025-Sustainability-Report_-compressed.pdf — '27 owned DC locations in operation or under construction'; new DC design uses less sqft per unit compute."),
    ("Meta — U.S. Data Centers list (2026)", "https://datacenters.atmeta.com/wp-content/uploads/2026/04/V_Meta-United-States-Data-Centers.pdf — 28 U.S. data centers; Richland Parish >2 GW."),
    ("Dgtl Infra — Meta data center locations", "https://dgtlinfra.com/meta-data-center-locations-facebook/ — 24 campuses, 104 buildings, 53.4M sqft, ~$29.8B; leased markets; Digital Realty relationship."),
    ("SemiAnalysis — Meta Superintelligence (Jul 2025)", "https://newsletter.semianalysis.com/p/meta-superintelligence-leadership-compute-talent-and-data — Prometheus 1GW, Hyperion, pre-leasing, on-site gas."),
    ("Data Center Frontier — Hyperion & Prometheus", "https://www.datacenterfrontier.com/hyperscale/article/55310441 — ownership/lease structure; 2 GW->5 GW; CAPEX."),
    ("AI Data Center Index — Meta operator report", "https://aidatacenterindex.com/reports/operator/meta/ — ~15.8 GW tracked, 20 sites, ~50% operational, 5 countries."),
    ("usdatamap — Meta", "https://usdatamap.com/company/meta — third-party per-site nameplate MW estimates (incomplete)."),
    ("CBRE North America DC Trends H2 2025", "https://www.cbre.com/insights/books/north-america-data-center-trends-h2-2025 — Blue Owl JV; $27B Hyperion; market context."),
    ("Redact.dev — Facebook data center list (May 2026)", "https://redact.dev/blog/full-list-of-known-and-upcoming-facebook-datacenters-as-of-may-2026 — campus sqft/investment; Lebanon ~1GW; Hyperion up to 5GW."),
    ("CoreWeave — Meta $21B expansion (Apr 2026)", "https://coreweave.com/news/coreweave-and-meta-announce-21-billion-expanded-ai-infrastructure-agreement — through Dec 2032; Vera Rubin; brings total to ~$35B."),
    ("Reuters — CoreWeave $14B Meta deal (Sep 2025)", "https://www.reuters.com/technology/coreweave-signs-14-billion-ai-deal-with-meta-bloomberg-news-reports-2025-09-30/ — $14.2B through Dec 14 2031; GB300."),
    ("CNBC — Meta +$21B CoreWeave (Apr 2026)", "https://www.cnbc.com/2026/04/09/meta-commits-to-spending-additional-21-billion-with-coreweave-.html — 2027-2032; portfolio approach."),
    ("CNBC / Reuters / Bloomberg — Meta-Google Cloud (Aug 2025)", "https://www.cnbc.com/2025/08/21/google-scores-six-year-meta-cloud-deal-worth-over-10-billion.html — $10B+ over 6 yrs."),
    ("Alatirok — Neocloud Economics 2026", "https://alatirok.com/neocloud-economics-2026/ — Nebius-Meta up to $27B ($12B + $15B option); CoreWeave $99.4B backlog / 3.5 GW."),
    ("CoreWeave Q1 FY2026 results", "https://www.businesswire.com/news/home/20260507558197 — >1 GW active, >3.5 GW contracted power; Meta $21B; MFU context."),
    ("CBRE — N. America DC Trends H2 2025 & Global 2026", "https://www.cbre.com/insights/books/north-america-data-center-trends-h2-2025 — $196/kW/mo, 1.4% vacancy, metro rates."),
    ("JLL — N. America DC Report YE 2025", "https://www.jll.com/en-us/insights/market-dynamics/north-america-data-centers — 1% vacancy, 92% precommitted, 35+ GW UC, ~60% leased."),
    ("CREFC — Data Center e-primer (2026)", "https://www.crefc.org/common/Uploaded%20files/Learn/DataCenters-eprimer-2026.02.25.pdf — rate tiers by deal type; occupancy/power-utilization flags."),
    ("Mercatus / DeployBase / Verda — GPU cloud pricing", "https://www.mercatus-ai.com/blog/cloud-gpu-pricing — cross-provider $/GPU-hr by SKU (H100/H200/B200/GB200); reserved & spot discounts."),
    ("American Compute / ModulEdge — Neocloud unit economics", "https://www.moduledge.com/blog/neocloud — ~70% break-even utilization; cluster P&L sensitivity; margins."),
]
for name, url in sources:
    r += 1
    a = ws.cell(r, 1, name); a.border = border_all; a.alignment = LEFTTOP; a.font = BOLD
    b = ws.cell(r, 2, url); b.border = border_all; b.alignment = LEFTTOP; b.font = REG
    ws.row_dimensions[r].height = 40
r += 2
ws.cell(r, 1, "Compiled: FY2024 basis. Prepared as an equity-research reference; verify against primary filings before use.").font = NOTE_FONT

# ============================================================================
out = "Meta_Compute_Capacity_Owned_vs_Leased.xlsx"
wb.save(out)
print("Saved", out)
